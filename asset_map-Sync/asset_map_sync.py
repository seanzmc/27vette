#!/usr/bin/env python3
"""
asset_map_sync.py
-----------------
Reconcile the `asset_map` sheet of stingray_master.xlsx against two sources of
truth and refresh image URLs from the WordPress.com media library.

Two sources of truth:
  - {model}_options sheets  -> which rows SHOULD exist (active AND selectable),
                               plus the authoritative `rpo` and `option_name`.
  - WP media library        -> the URL value for each (model, rpo).

asset_map owns the curation columns (image_alt, image_fit, image_position,
notes) and is never rebuilt from scratch -- it is UPSERTED:

  * desired option not yet in asset_map        -> INSERT a row (url filled if a
                                                   matching image exists, else blank)
  * row already present                        -> keep curation, reconcile URL
  * row whose option is no longer active+sel.   -> flag stale (optionally deactivate)

The same run builds (first time, empty sheet) and maintains (every run after).

URL reconciliation per row:
  - existing url live (2xx/3xx)  -> KEEP (trust existing)
  - existing url 404/dead        -> REPLACE with matched image
  - url blank                    -> FILL with matched image
  - bare-RPO file claimed by >1 model -> AMBIGUOUS (left blank, flagged)

Filename -> (model, rpo) parsing:
  - model = single leading letter + hyphen: c stingray, e grand_sport, h z06,
    r zr1, s zr1x, g grand_sport_x. A leading letter with NO hyphen (e.g. hzp)
    is part of the RPO. `imgi_<n>_` scrape prefixes and trailing qualifiers
    (-c, _v1, -23x10-cp) are stripped. One image per option.

Default run is a DRY RUN (no writes; two CSV reports). Pass --apply to write
(workbook backed up to *.bak first). New rows are appended.

Setup:
  pip install requests openpyxl
  export WP_USER="your-wordpress-login-username"
  export WP_APP_PASSWORD="xxxx xxxx xxxx xxxx xxxx xxxx"

Usage:
  python asset_map_sync.py --workbook stingray_master.xlsx                 # dry run (build preview)
  python asset_map_sync.py --workbook stingray_master.xlsx --apply         # build/maintain
  python asset_map_sync.py --workbook ... --since auto --apply             # incremental maintenance
  python asset_map_sync.py --workbook ... --status-col --apply             # also maintain image_status
  python asset_map_sync.py --workbook ... --deactivate-stale --apply       # set active=FALSE on stale rows
"""

import argparse
import os
import re
import csv
import sys
import json
import shutil
from datetime import datetime, timedelta, timezone
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse, unquote

import requests
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------
SITE = "stingraychevroletcorvette.com"
MEDIA_ENDPOINT = f"https://{SITE}/wp-json/wp/v2/media"
PATH_FILTER = "/wp-content/uploads/pictures/27vette/"   # this folder + subfolders
ASSET_SHEET = "asset_map"
TARGET_TYPE = "option"

# Source option sheets -> asset_map model_key. Missing sheets are skipped.
OPTION_SHEETS = {
    "stingray_options": "stingray",
    "grandSport_options": "grand_sport",
    "z06_options": "z06",
    "zr1_options": "zr1",
    "zr1x_options": "zr1x",
    "grandSportX_options": "grand_sport_x",
}

MODEL_PREFIX = {
    "c": "stingray", "e": "grand_sport", "h": "z06",
    "r": "zr1", "s": "zr1x", "g": "grand_sport_x",
}

NEW_ROW_FIT = "cover"
NEW_ROW_POSITION = "center"
NEW_ROW_NOTE = "auto-seeded"

IMGI_RE   = re.compile(r"^imgi_\d+_(.+)$")
PREFIX_RE = re.compile(r"^([cehrsg])-(.+)$")
SPLIT_RE  = re.compile(r"[-_]")
RPO_RE    = re.compile(r"^[0-9a-z]{3}$")


def truthy(v):
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


# --------------------------------------------------------------------------
# Filename parsing
# --------------------------------------------------------------------------
def filename_stem(url):
    base = unquote(os.path.basename(urlparse(url).path))
    return os.path.splitext(base)[0].lower()


def parse_media(url):
    """Return (model_or_None, rpo, is_valid)."""
    s = filename_stem(url)
    m = IMGI_RE.match(s)
    if m:
        s = m.group(1)
    model = None
    m = PREFIX_RE.match(s)
    if m:
        model = MODEL_PREFIX[m.group(1)]
        s = m.group(2)
    rpo = SPLIT_RE.split(s)[0]
    return model, rpo, bool(RPO_RE.match(rpo))


# --------------------------------------------------------------------------
# WordPress media pull
# --------------------------------------------------------------------------
def fetch_media(auth, timeout, modified_after=None):
    urls = []
    page = 1
    while True:
        params = {"per_page": 100, "page": page, "_fields": "source_url", "media_type": "image"}
        if modified_after:
            params.update({"modified_after": modified_after, "orderby": "modified", "order": "desc"})
        r = requests.get(MEDIA_ENDPOINT, params=params, auth=auth, timeout=timeout)
        if r.status_code == 400 and page > 1:
            break
        r.raise_for_status()
        batch = r.json()
        if not batch:
            break
        for item in batch:
            u = item.get("source_url", "")
            if PATH_FILTER in u:
                urls.append(u)
        total_pages = int(r.headers.get("X-WP-TotalPages", page))
        if page >= total_pages:
            break
        page += 1
    return urls


# --------------------------------------------------------------------------
# Incremental cursor (for --since auto)
# --------------------------------------------------------------------------
def state_path(report_dir):
    return os.path.join(report_dir, ".asset_map_sync_state.json")


def read_since_auto(report_dir, cushion_hours=6):
    p = state_path(report_dir)
    if not os.path.exists(p):
        return None
    try:
        ts = json.load(open(p)).get("last_run_utc")
        return (datetime.fromisoformat(ts) - timedelta(hours=cushion_hours)).strftime("%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None


def write_state(report_dir):
    json.dump({"last_run_utc": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec="seconds")},
              open(state_path(report_dir), "w"))


# --------------------------------------------------------------------------
# URL liveness
# --------------------------------------------------------------------------
def url_alive(session, url, timeout):
    try:
        resp = session.head(url, timeout=timeout, allow_redirects=True)
        if resp.status_code == 405:
            resp = session.get(url, timeout=timeout, stream=True, allow_redirects=True)
        return resp.status_code < 400
    except requests.RequestException:
        return False


def check_existing(urls, timeout, workers):
    session = requests.Session()
    out = {}
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for u, ok in ex.map(lambda u: (u, url_alive(session, u, timeout)), urls):
            out[u] = ok
    return out


# --------------------------------------------------------------------------
# Indexing
# --------------------------------------------------------------------------
def build_media_index(media_urls):
    exact, bare, unparseable = defaultdict(list), defaultdict(list), []
    for u in media_urls:
        model, rpo, ok = parse_media(u)
        if not ok:
            unparseable.append(u)
        elif model:
            exact[(model, rpo)].append(u)
        else:
            bare[rpo].append(u)
    return exact, bare, unparseable


def read_option_sheets(wb):
    """Desired set: {(model_key, target_id): {'rpo':.., 'name':..}}."""
    desired = {}
    for sheet, model_key in OPTION_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
        idx = {n: i for i, n in enumerate(hdr)}
        missing = {"option_id", "rpo", "selectable", "active"} - set(idx)
        if missing:
            print(f"  ! {sheet}: missing columns {sorted(missing)}, skipped")
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            oid = row[idx["option_id"]]
            if oid is None or not str(oid).strip():
                continue
            if not (truthy(row[idx["active"]]) and truthy(row[idx["selectable"]])):
                continue
            rpo = str(row[idx["rpo"]] or "").strip().lower()
            name = str(row[idx["option_name"]] or "").strip() if "option_name" in idx else ""
            desired[(model_key, str(oid).strip().lower())] = {"rpo": rpo, "name": name}
    return desired


# --------------------------------------------------------------------------
# Pure reconcile
# --------------------------------------------------------------------------
def reconcile(desired, exact, bare, existing_rows, alive, incremental):
    """
    desired:       {(model, tid): {'rpo','name'}}
    existing_rows: {(model, tid): {'row':int, 'url':str}}  (asset_map option rows)
    alive:         {url: bool}
    Returns: report(list of dicts), url_writes({row->url}), inserts(list of dicts),
             status({row->status} and inserts carry 'status'), used(set of urls)
    """
    models_for_rpo = defaultdict(set)
    for (model, tid), info in desired.items():
        if info["rpo"]:
            models_for_rpo[info["rpo"]].add(model)
    prefixed_models_for_rpo = defaultdict(set)
    for (model, rpo) in exact:
        prefixed_models_for_rpo[rpo].add(model)

    def resolve(model, rpo):
        if not rpo:
            return None, "no-rpo"
        if (model, rpo) in exact:
            return exact[(model, rpo)][0], "prefixed"
        if rpo in bare:
            eligible = models_for_rpo.get(rpo, set()) - prefixed_models_for_rpo.get(rpo, set())
            if eligible == {model}:
                return bare[rpo][0], "bare-unique"
            if model in eligible:
                return None, "bare-ambiguous"
        return None, "none"

    report, url_writes, inserts, status, used = [], {}, [], {}, set()

    def add_report(scope, model, tid, rpo, action, source, existing_url, new_url, st, note=""):
        report.append({"scope": scope, "model_key": model, "target_id": tid, "rpo": rpo,
                        "action": action, "candidate_source": source,
                        "existing_url": existing_url, "new_url": new_url,
                        "image_status": st, "note": note})

    # Drive off the desired set: inserts + updates.
    for (model, tid), info in desired.items():
        rpo = info["rpo"]
        cand, source = resolve(model, rpo)
        if cand:
            used.add(cand)
        note = "" if rpo else "no rpo in option sheet"
        if source == "bare-ambiguous":
            note = f"bare file for '{rpo}' shared across models; add a c/e/h/r/s/g prefix"

        if (model, tid) in existing_rows:
            row = existing_rows[(model, tid)]["row"]
            existing_url = existing_rows[(model, tid)]["url"]
            if existing_url:
                ok = alive.get(existing_url, True)
                if ok:
                    status[row] = "ok"
                    add_report("existing", model, tid, rpo, "keep", source, existing_url, existing_url, "ok")
                elif cand:
                    url_writes[row] = cand
                    status[row] = "ok"
                    add_report("existing", model, tid, rpo, "replace_404", source, existing_url, cand, "ok")
                else:
                    st = "url_dead"
                    act = "dead_no_match_incremental" if incremental else "flag_dead_no_match"
                    status[row] = st
                    add_report("existing", model, tid, rpo, act, source, existing_url, existing_url, st, note)
            else:
                if cand:
                    url_writes[row] = cand
                    status[row] = "ok"
                    add_report("existing", model, tid, rpo, "fill", source, "", cand, "ok")
                elif source == "bare-ambiguous":
                    status[row] = "ambiguous"
                    add_report("existing", model, tid, rpo, "flag_ambiguous", source, "", "", "ambiguous", note)
                elif incremental:
                    add_report("existing", model, tid, rpo, "skip_no_candidate_incremental", source, "", "", "missing")
                else:
                    status[row] = "missing"
                    add_report("existing", model, tid, rpo, "flag_missing", source, "", "", "missing", note)
        else:
            # INSERT a new row for this desired option.
            if cand:
                st, act, url = "ok", "insert_filled", cand
            elif source == "bare-ambiguous":
                st, act, url = "ambiguous", "insert_ambiguous", ""
            else:
                st, act, url = "missing", "insert_blank", ""
            inserts.append({"model": model, "tid": tid, "rpo": rpo, "name": info["name"],
                            "url": url, "status": st})
            add_report("new", model, tid, rpo, act, source, "", url, st, note)

    # Stale: existing option rows whose option is no longer desired.
    for (model, tid), row in existing_rows.items():
        if (model, tid) not in desired:
            status[row["row"]] = "stale_option"
            add_report("stale", model, tid, "", "stale_option", "", row["url"], row["url"],
                       "stale_option", "option no longer active+selectable")

    return report, url_writes, inserts, status, used


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Seed + reconcile asset_map and refresh image URLs.")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--asset-sheet", default=ASSET_SHEET)
    ap.add_argument("--apply", action="store_true", help="Write changes (default dry run)")
    ap.add_argument("--report-dir", default=".")
    ap.add_argument("--timeout", type=float, default=10.0)
    ap.add_argument("--workers", type=int, default=16)
    ap.add_argument("--no-verify-existing", action="store_true", help="Skip 404-checking existing URLs")
    ap.add_argument("--since", default=None, metavar="DATE",
                    help="Incremental: media modified after DATE (YYYY-MM-DD/ISO), or 'auto' for saved cursor")
    ap.add_argument("--status-col", action="store_true", help="Maintain an image_status column in asset_map")
    ap.add_argument("--deactivate-stale", action="store_true", help="Set active=FALSE on stale rows")
    args = ap.parse_args()

    user, pw = os.environ.get("WP_USER"), os.environ.get("WP_APP_PASSWORD")
    if not user or not pw:
        sys.exit("Set WP_USER and WP_APP_PASSWORD environment variables.")
    auth = HTTPBasicAuth(user, pw.replace(" ", ""))

    modified_after = read_since_auto(args.report_dir) if args.since == "auto" else args.since
    incremental = modified_after is not None
    print(f"Pulling media [{'incremental after ' + modified_after if incremental else 'full'}] ...")
    media_urls = fetch_media(auth, args.timeout, modified_after)
    print(f"  {len(media_urls)} images under {PATH_FILTER}")
    exact, bare, unparseable = build_media_index(media_urls)

    wb = load_workbook(args.workbook)
    if args.asset_sheet not in wb.sheetnames:
        sys.exit(f"Sheet '{args.asset_sheet}' not found.")
    ws = wb[args.asset_sheet]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    hidx = {n: i for i, n in enumerate(headers)}
    for col in ("model_key", "target_type", "target_id", "image_url"):
        if col not in hidx:
            sys.exit(f"asset_map missing column: {col}")

    # Optional image_status column.
    status_col = hidx.get("image_status")
    if args.status_col and status_col is None:
        status_col = len(headers)
        ws.cell(1, status_col + 1).value = "image_status"
        headers.append("image_status")

    # Existing option rows.
    existing_rows = {}
    for ri in range(2, ws.max_row + 1):
        if str(ws.cell(ri, hidx["target_type"] + 1).value or "").strip().lower() != TARGET_TYPE:
            continue
        model = str(ws.cell(ri, hidx["model_key"] + 1).value or "").strip().lower()
        tid = str(ws.cell(ri, hidx["target_id"] + 1).value or "").strip().lower()
        if not tid:
            continue
        url = str(ws.cell(ri, hidx["image_url"] + 1).value or "").strip()
        existing_rows[(model, tid)] = {"row": ri, "url": url}

    desired = read_option_sheets(wb)
    print(f"  {len(desired)} active+selectable options across option sheets")
    print(f"  {len(existing_rows)} existing option rows in asset_map")

    # Verify existing URLs (those that survive to be 'kept' or 'replaced').
    if args.no_verify_existing:
        alive = {}
    else:
        urls = sorted({v["url"] for v in existing_rows.values() if v["url"]})
        print(f"Checking {len(urls)} existing URLs ...")
        alive = check_existing(urls, args.timeout, args.workers)
        if args.no_verify_existing:
            alive = {}

    report, url_writes, inserts, status, used = reconcile(
        desired, exact, bare, existing_rows, alive, incremental)

    # Apply URL writes to existing rows (in memory).
    for row, url in url_writes.items():
        ws.cell(row, hidx["image_url"] + 1).value = url
    # Status writes.
    if args.status_col:
        for row, st in status.items():
            ws.cell(row, status_col + 1).value = st
    # Deactivate stale.
    if args.deactivate_stale and "active" in hidx:
        for r in report:
            if r["action"] == "stale_option":
                key = (r["model_key"], r["target_id"])
                if key in existing_rows:
                    ws.cell(existing_rows[key]["row"], hidx["active"] + 1).value = False
    # Append inserts.
    ncols = len(headers)
    for ins in inserts:
        rowvals = [""] * ncols
        rowvals[hidx["model_key"]] = ins["model"]
        rowvals[hidx["target_type"]] = TARGET_TYPE
        rowvals[hidx["target_id"]] = ins["tid"]
        rowvals[hidx["image_url"]] = ins["url"]
        if "image_alt" in hidx:
            rowvals[hidx["image_alt"]] = ins["name"]
        if "image_fit" in hidx:
            rowvals[hidx["image_fit"]] = NEW_ROW_FIT
        if "image_position" in hidx:
            rowvals[hidx["image_position"]] = NEW_ROW_POSITION
        if "active" in hidx:
            rowvals[hidx["active"]] = True
        if "notes" in hidx:
            rowvals[hidx["notes"]] = NEW_ROW_NOTE
        if args.status_col:
            rowvals[status_col] = ins["status"]
        ws.append(rowvals)

    # Unmatched media (only meaningful on a full run).
    unmatched = sorted(set(media_urls) - used - set(unparseable))

    # Reports.
    os.makedirs(args.report_dir, exist_ok=True)
    rep1 = os.path.join(args.report_dir, "asset_map_sync_report.csv")
    with open(rep1, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["scope", "model_key", "target_id", "rpo", "action",
                                          "candidate_source", "existing_url", "new_url",
                                          "image_status", "note"])
        w.writeheader()
        w.writerows(report)
    rep2 = os.path.join(args.report_dir, "asset_map_unmatched_media.csv")
    with open(rep2, "w", newline="") as f:
        w = csv.writer(f)
        reason = "new media in window, no row yet" if incremental else "no desired (model, rpo) for this file"
        w.writerow(["source_url", "parsed_model", "parsed_rpo", "reason"])
        for u in unmatched:
            m, r, _ = parse_media(u)
            w.writerow([u, m or "", r, reason])
        for u in unparseable:
            w.writerow([u, "", "", "filename did not yield a 3-char RPO"])

    # Summary.
    counts = Counter(r["action"] for r in report)
    print("\n=== Summary ===")
    for k in ("keep", "fill", "replace_404", "insert_filled", "insert_blank", "insert_ambiguous",
              "flag_dead_no_match", "dead_no_match_incremental", "flag_ambiguous", "flag_missing",
              "skip_no_candidate_incremental", "stale_option"):
        if counts[k]:
            print(f"  {k:<30} {counts[k]}")
    print(f"  {'unmatched media':<30} {len(unmatched)}")
    print(f"  {'unparseable files':<30} {len(unparseable)}")
    print(f"\nReports: {rep1}\n         {rep2}")
    if incremental:
        print("\nNote: incremental run; coverage flags suppressed for rows whose image was "
              "outside the window. Run a full sync (no --since) periodically.")

    if args.apply:
        backup = args.workbook + ".bak"
        shutil.copy2(args.workbook, backup)
        wb.save(args.workbook)
        if args.since == "auto":
            write_state(args.report_dir)
        print(f"\nAPPLIED: {len(url_writes)} url change(s), {len(inserts)} row insert(s). Backup -> {backup}")
    else:
        print(f"\nDRY RUN -- would write {len(url_writes)} url change(s) and {len(inserts)} new row(s). "
              "Re-run with --apply.")


if __name__ == "__main__":
    main()
