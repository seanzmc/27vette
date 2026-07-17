"""Rebuild a comparison workbook from canonical SQL and reversible lineage."""

from __future__ import annotations

import hashlib
import json
import shutil
import sqlite3
from collections import Counter
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook


class ReverseMappingError(ValueError):
    """Raised when persisted mapping evidence cannot express a SQL value."""


def _quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _jsonable(value: object) -> object:
    if isinstance(value, Mapping):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, (tuple, list, set, frozenset)):
        return [_jsonable(item) for item in value]
    return value


def _json(value: object) -> str:
    return json.dumps(
        _jsonable(value), sort_keys=True, separators=(",", ":"), default=str
    )


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _compiled_lineage_hashes(compiled) -> dict[tuple[str, str], str]:
    hashes = {}
    for table in compiled.tables:
        for row in table.rows:
            key = {column: row.values[column] for column in table.primary_key}
            evidence = {
                "values": row.values,
                "source_sheet": row.source_sheet,
                "source_row": row.source_row,
                "lineage_role": row.lineage_role,
                "mapping_parameters": row.mapping_parameters,
            }
            hashes[(table.name, _json(key))] = hashlib.sha256(
                _json(evidence).encode("utf-8")
            ).hexdigest()
    return hashes


def _validated_import_run(
    conn: sqlite3.Connection,
    source_path: Path,
) -> int:
    run_id_row = conn.execute(
        "SELECT value FROM meta WHERE key='last_import_run_id'"
    ).fetchone()
    if run_id_row is None:
        raise ReverseMappingError("Persisted lineage has no import run identity")
    run_id = int(run_id_row[0])
    run = conn.execute(
        "SELECT workbook_sha256 FROM import_runs WHERE id=?", (run_id,)
    ).fetchone()
    if run is None:
        raise ReverseMappingError("Persisted lineage import run is missing")
    source_sha256 = _file_sha256(source_path)
    trusted_row = conn.execute(
        "SELECT value FROM meta WHERE key='trusted_workbook_sha256'"
    ).fetchone()
    trusted_sha256 = str(trusted_row[0]) if trusted_row else str(
        run["workbook_sha256"]
    )
    if trusted_sha256 != source_sha256:
        raise ReverseMappingError(
            "Source workbook SHA does not match the trusted synchronization base"
        )
    return run_id


def _validate_persisted_lineage(
    conn: sqlite3.Connection,
    compiled,
    run_id: int,
) -> None:
    hashes = _compiled_lineage_hashes(compiled)
    expected = Counter(
        (
            entry.destination_table,
            _json(dict(entry.destination_key)),
            entry.source_sheet,
            entry.source_row,
            hashes[(entry.destination_table, _json(entry.destination_key))],
            entry.mapping_role,
            "mapped",
        )
        for entry in compiled.lineage
    )
    persisted_rows = tuple(
        conn.execute(
            "SELECT sql_table, primary_key_json, source_sheet, source_row, "
            "source_row_hash, lineage_role, transform_status "
            "FROM import_lineage WHERE import_run_id=?",
            (run_id,),
        )
    )
    persisted = Counter(
        (
            row["sql_table"],
            _json(json.loads(row["primary_key_json"])),
            row["source_sheet"],
            row["source_row"],
            row["source_row_hash"],
            row["lineage_role"],
            row["transform_status"],
        )
        for row in persisted_rows
    )
    if persisted != expected:
        raise ReverseMappingError(
            "Persisted import_lineage does not exactly match recompiled lineage"
        )

    lineage_destinations: dict[tuple[str, int], set[tuple[str, str]]] = {}
    for row in persisted_rows:
        lineage_destinations.setdefault(
            (row["source_sheet"], row["source_row"]), set()
        ).add(
            (
                row["sql_table"],
                _json(json.loads(row["primary_key_json"])),
            )
        )
    disposition_destinations = {}
    for row in conn.execute(
        "SELECT source_sheet, source_row, destinations_json "
        "FROM source_row_disposition WHERE import_run_id=? "
        "AND disposition='emitted'",
        (run_id,),
    ):
        destination_set = {
            (
                destination["destination_table"],
                _json(destination["destination_key"]),
            )
            for destination in json.loads(row["destinations_json"])
        }
        disposition_destinations[(row["source_sheet"], row["source_row"])] = (
            destination_set
        )
    if disposition_destinations != lineage_destinations:
        raise ReverseMappingError(
            "Emitted source-row disposition destinations do not exactly match "
            "persisted import_lineage"
        )


def _truthy(value: object) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value != 0)
    return int(str(value or "").strip().lower() in {"1", "true", "yes", "y"})


def _reverse_value(current: object, original: object, reverse: str) -> object:
    if reverse == "restore_exact_model_owner":
        return original
    if reverse == "restore_scope_active_flag":
        return _reverse_value(
            current, original, "restore_original_boolean_from_lineage"
        )
    if reverse == "restore_original_boolean_from_lineage":
        boolean = bool(current)
        if isinstance(original, bool):
            return boolean
        if isinstance(original, (int, float)):
            return int(boolean)
        text = str(original or "")
        if text.isupper():
            return "TRUE" if boolean else "FALSE"
        if text[:1].isupper():
            return "True" if boolean else "False"
        return "true" if boolean else "false"
    if reverse == "restore_original_scope_from_lineage":
        if current is None:
            return original if original in (None, "", "*") else None
        return current
    if reverse == "restore_original_number_from_lineage":
        if isinstance(original, float):
            return float(current) if current is not None else None
        return int(current) if current is not None else None
    if reverse in {
        "restore_target_type_option",
        "restore_target_type_context_choice",
    }:
        return "option" if reverse.endswith("option") else "context_choice"
    return current


def _destination_row(
    conn: sqlite3.Connection,
    table: str,
    key: Mapping[str, object],
) -> Mapping[str, object]:
    if not key:
        raise ReverseMappingError(f"Missing destination key for {table}")
    clauses = " AND ".join(f"{_quote(column)} IS ?" for column in key)
    row = conn.execute(
        f"SELECT * FROM {_quote(table)} WHERE {clauses}", tuple(key.values())
    ).fetchone()
    if row is None:
        raise ReverseMappingError(
            f"Mapped destination row is missing: {table} {dict(key)!r}"
        )
    return dict(row)


def _mapped_value(
    conn: sqlite3.Connection,
    mappings: tuple[Mapping[str, object], ...],
    destinations: tuple[Mapping[str, object], ...],
    original: object,
    expected_rows: Mapping[tuple[str, str], Mapping[str, object]],
) -> object:
    source_sheet = str(mappings[0]["source_sheet"])
    direct_tables = {
        str(row["sql_table"])
        for row in mappings
        if str(row["sql_table"]) == source_sheet
    }
    if direct_tables:
        mappings = tuple(
            row for row in mappings if str(row["sql_table"]) in direct_tables
        )
        destinations = tuple(
            row
            for row in destinations
            if str(row["destination_table"]) in direct_tables
        )
    if (
        original == "*"
        and all(row["source_column"] == "model_key" for row in mappings)
        and all(row["sql_column"] == "model_key" for row in mappings)
    ):
        # The compiler's documented wildcard expansion produces one row per
        # model table; the single source cell remains the wildcard marker.
        return original
    values: list[object] = []
    for destination in destinations:
        table = str(destination["destination_table"])
        table_mappings = tuple(row for row in mappings if row["sql_table"] == table)
        if not table_mappings:
            continue
        sql_row = _destination_row(conn, table, destination["destination_key"])
        key_json = json.dumps(
            destination["destination_key"], sort_keys=True, separators=(",", ":")
        )
        expected_row = expected_rows.get((table, key_json))
        if expected_row is None:
            raise ReverseMappingError(
                f"Source lineage has no compiler baseline for {table} {key_json}"
            )
        changed_mappings = tuple(
            row
            for row in table_mappings
            if sql_row[row["sql_column"]] != expected_row[row["sql_column"]]
        )
        if not changed_mappings:
            values.append(original)
            continue
        table_mappings = changed_mappings
        destination_values = [sql_row[row["sql_column"]] for row in table_mappings]
        non_null = [value for value in destination_values if value is not None]
        if len(set(non_null)) > 1:
            raise ReverseMappingError(
                f"Source field maps to contradictory columns in {table}: {non_null!r}"
            )
        current = non_null[0] if non_null else None
        mapping = next(
            (row for row in table_mappings if sql_row[row["sql_column"]] is not None),
            table_mappings[0],
        )
        parameters = json.loads(str(mapping["transform_parameters_json"] or "{}"))
        values.append(
            _reverse_value(
                current,
                original,
                str(parameters.get("reverse_transform") or "identity"),
            )
        )
    distinct = []
    for value in values:
        if value not in distinct:
            distinct.append(value)
    if not distinct:
        return original
    if len(distinct) != 1:
        raise ReverseMappingError(
            "A shared workbook field has model-specific SQL values and cannot "
            f"be represented by one source cell: {distinct!r}"
        )
    return distinct[0]


def _assert_recompiled_matches_sql(
    conn: sqlite3.Connection,
    compiled,
) -> None:
    """Prove that the rebuilt workbook recreates every canonical SQL row."""

    for table in compiled.tables:
        if not table.rows:
            count = conn.execute(
                f"SELECT COUNT(*) FROM {_quote(table.name)}"
            ).fetchone()[0]
            if count:
                raise ReverseMappingError(
                    f"{table.name} has {count} SQL rows with no workbook reconstruction"
                )
            continue
        columns = tuple(
            sorted({column for row in table.rows for column in row.values})
        )
        expected = {
            tuple(row.values[column] for column in table.primary_key): {
                column: row.values[column] for column in columns
            }
            for row in table.rows
        }
        selected = ", ".join(_quote(column) for column in columns)
        actual_rows = tuple(
            dict(row)
            for row in conn.execute(
                f"SELECT {selected} FROM {_quote(table.name)}"
            )
        )
        actual = {
            tuple(row[column] for column in table.primary_key): row
            for row in actual_rows
        }
        if set(actual) != set(expected):
            missing = sorted(set(expected) - set(actual), key=repr)
            extra = sorted(set(actual) - set(expected), key=repr)
            raise ReverseMappingError(
                f"{table.name} row identities are not reversible; "
                f"missing={missing[:3]!r}, extra={extra[:3]!r}"
            )
        for key in sorted(expected, key=repr):
            for column in columns:
                if actual[key][column] != expected[key][column]:
                    raise ReverseMappingError(
                        f"{table.name} {key!r}.{column} is not reversible: "
                        f"SQL={actual[key][column]!r}, "
                        f"workbook={expected[key][column]!r}"
                    )

    from .catalog import LIVE_MODELS, MODEL_TABLE_ROLES, physical_table

    by_role = {
        (table.model_key, table.role): table
        for table in compiled.tables
        if table.model_key
    }
    expected_registry = {}
    for model_key in LIVE_MODELS:
        for role in MODEL_TABLE_ROLES:
            table = by_role[(model_key, role)]
            source_sheets = tuple(sorted(
                {row.source_sheet for row in table.rows if row.source_sheet}
                or {
                    mapping.source_sheet
                    for mapping in table.schema_mappings
                    if mapping.source_sheet
                }
            ))
            split = any(
                row.lineage_role == "shared_source_split" for row in table.rows
            )
            expected_registry[(model_key, role)] = {
                "model_key": model_key,
                "table_role": role,
                "sql_table": physical_table(model_key, role),
                "source_sheets_json": json.dumps(
                    source_sheets, sort_keys=True, separators=(",", ":")
                ),
                "source_filter": f"model_key={model_key}" if split else "",
                "mapping_type": "split" if split else "exact",
                "active": 1,
            }
    actual_registry = {
        (row["model_key"], row["table_role"]): dict(row)
        for row in conn.execute("SELECT * FROM model_table_registry")
    }
    if actual_registry != expected_registry:
        for key in sorted(set(actual_registry) | set(expected_registry)):
            if actual_registry.get(key) != expected_registry.get(key):
                raise ReverseMappingError(
                    "model_table_registry is not reversible for "
                    f"{key!r}: SQL={actual_registry.get(key)!r}, "
                    f"workbook={expected_registry.get(key)!r}"
                )


def export_comparison_workbook(
    conn: sqlite3.Connection,
    source: Path,
    destination: Path,
) -> Path:
    """Copy ``source`` and rewrite mapped canonical cells from ``conn``.

    Unmanaged and inactive sheets remain byte-content-equivalent at the cell
    level because the source package is copied first. Only rows with exact
    persisted lineage are considered for reconstruction.
    """

    source_path = Path(source)
    destination_path = Path(destination)

    # Recompile the immutable source to obtain the exact canonical baseline.
    # This avoids guessing how a derived destination field related to an
    # original cell: only SQL values that differ from that compiler baseline
    # are reverse-applied.
    from .importer import compile_workbook

    run_id = _validated_import_run(conn, source_path)
    compiled = compile_workbook(source_path)
    _validate_persisted_lineage(conn, compiled, run_id)
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, destination_path)
    expected_rows: dict[tuple[str, str], Mapping[str, object]] = {}
    for table in compiled.tables:
        for row in table.rows:
            key = {column: row.values[column] for column in table.primary_key}
            key_json = json.dumps(key, sort_keys=True, separators=(",", ":"))
            expected_rows[(table.name, key_json)] = dict(row.values)

    mapping_rows = tuple(dict(row) for row in conn.execute("SELECT * FROM schema_mapping"))
    mappings_by_field: dict[tuple[str, str], list[Mapping[str, object]]] = {}
    for row in mapping_rows:
        source_column = str(row["source_column"])
        if source_column.startswith("__"):
            continue
        mappings_by_field.setdefault(
            (str(row["source_sheet"]), source_column), []
        ).append(row)

    workbook = load_workbook(destination_path, data_only=False)
    try:
        dispositions = conn.execute(
            "SELECT source_sheet, source_row, destinations_json, evidence_json "
            "FROM source_row_disposition WHERE import_run_id=? "
            "AND disposition='emitted' "
            "ORDER BY source_sheet, source_row",
            (run_id,),
        )
        for disposition in dispositions:
            sheet_name = str(disposition["source_sheet"])
            if sheet_name not in workbook.sheetnames:
                raise ReverseMappingError(f"Comparison workbook missing {sheet_name!r}")
            worksheet = workbook[sheet_name]
            headers = {
                str(cell.value): cell.column
                for cell in worksheet[1]
                if cell.value is not None and str(cell.value) != ""
            }
            evidence = json.loads(str(disposition["evidence_json"] or "{}"))
            originals = evidence.get("source_values") or {}
            destinations = tuple(json.loads(str(disposition["destinations_json"])))
            for source_column, original in originals.items():
                column = headers.get(source_column)
                mappings = tuple(mappings_by_field.get((sheet_name, source_column), ()))
                if column is None or not mappings:
                    continue
                try:
                    rebuilt = _mapped_value(
                        conn, mappings, destinations, original, expected_rows
                    )
                except ReverseMappingError as error:
                    raise ReverseMappingError(
                        f"{sheet_name}!{source_column} row "
                        f"{disposition['source_row']}: {error}"
                    ) from error
                worksheet.cell(int(disposition["source_row"]), column, rebuilt)
        workbook.save(destination_path)
    except BaseException:
        workbook.close()
        try:
            destination_path.unlink()
        except OSError:
            pass
        raise
    workbook.close()
    try:
        rebuilt = compile_workbook(destination_path)
        _assert_recompiled_matches_sql(conn, rebuilt)
    except BaseException:
        try:
            destination_path.unlink()
        except OSError:
            pass
        raise
    return destination_path
