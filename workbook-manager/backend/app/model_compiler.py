"""Compile registry-bound model sheets into identical relational families."""

from __future__ import annotations

import hashlib
from pathlib import Path
from types import MappingProxyType
from typing import Iterable, Mapping

from openpyxl import load_workbook

from .catalog import LIVE_MODELS, physical_table
from .compile_types import (
    CompiledRow,
    CompiledTable,
    DecisionRequired,
    SchemaMapping,
    WorkbookProfile,
    freeze_mapping,
)


_DIRECT_ROLE_SOURCES: tuple[tuple[str, str], ...] = (
    ("options", "source_option_sheet"),
    ("option_availability", "status_sheet"),
    ("rule_mapping", "rule_mapping_sheet"),
    ("price_rules", "price_rules_sheet"),
    ("rule_groups", "rule_groups_sheet"),
    ("rule_group_members", "rule_group_members_sheet"),
    ("exclusive_groups", "exclusive_groups_sheet"),
    ("exclusive_group_members", "exclusive_group_members_sheet"),
    ("variant_overrides", "variant_option_overrides_sheet"),
)

_HEADERS: Mapping[str, tuple[str, ...]] = MappingProxyType(
    {
        "options": (
            "option_id",
            "rpo",
            "price",
            "option_name",
            "description",
            "detail_raw",
            "section_id",
            "selectable",
            "display_order",
            "active",
            "display_behavior",
        ),
        "option_availability": ("option_id", "variant_id", "status"),
        "rule_mapping": (
            "rule_id",
            "source_id",
            "rule_type",
            "target_id",
            "original_detail_raw",
            "body_style_scope",
            "runtime_action",
            "disabled_reason",
        ),
        "price_rules": (
            "price_rule_id",
            "condition_option_id",
            "price_rule_type",
            "target_option_id",
            "price_value",
            "body_style_scope",
            "trim_level_scope",
            "notes",
        ),
        "rule_groups": (
            "group_id",
            "group_type",
            "source_id",
            "body_style_scope",
            "trim_level_scope",
            "variant_scope",
            "disabled_reason",
            "active",
            "notes",
        ),
        "rule_group_members": (
            "group_id",
            "target_id",
            "display_order",
            "active",
        ),
        "exclusive_groups": (
            "group_id",
            "selection_mode",
            "active",
            "notes",
        ),
        "exclusive_group_members": (
            "group_id",
            "option_id",
            "display_order",
            "active",
        ),
        "variant_overrides": (
            "option_id",
            "variant_id",
            "selectable",
            "display_behavior",
            "section_id",
            "active",
            "note",
        ),
    }
)

_PRIMARY_KEYS: Mapping[str, tuple[str, ...]] = MappingProxyType(
    {
        "options": ("option_id",),
        "option_availability": ("option_id", "variant_id"),
        "rule_mapping": ("rule_id",),
        "price_rules": ("price_rule_id",),
        "rule_groups": ("group_id",),
        "rule_group_members": ("group_id", "target_option_id"),
        "exclusive_groups": ("group_id",),
        "exclusive_group_members": ("group_id", "option_id"),
        "variant_overrides": ("option_id", "variant_id"),
    }
)

_COLUMN_DESTINATIONS: Mapping[str, Mapping[str, tuple[str, ...]]] = MappingProxyType(
    {
        "rule_mapping": MappingProxyType(
            {
                "source_id": ("source_option_id", "source_interior_id"),
                "target_id": ("target_option_id",),
            }
        ),
        "price_rules": MappingProxyType(
            {
                "condition_option_id": (
                    "condition_option_id",
                    "condition_interior_id",
                ),
            }
        ),
        "rule_groups": MappingProxyType({"source_id": ("source_option_id",)}),
        "rule_group_members": MappingProxyType(
            {"target_id": ("target_option_id",)}
        ),
    }
)

_BOOLEAN_COLUMNS = frozenset(
    {
        ("options", "selectable"),
        ("options", "active"),
        ("rule_groups", "active"),
        ("rule_group_members", "active"),
        ("exclusive_groups", "active"),
        ("exclusive_group_members", "active"),
        ("variant_overrides", "selectable"),
        ("variant_overrides", "active"),
    }
)

_SCOPE_COLUMNS = frozenset(
    {
        ("rule_mapping", "body_style_scope"),
        ("price_rules", "body_style_scope"),
        ("price_rules", "trim_level_scope"),
        ("rule_groups", "body_style_scope"),
        ("rule_groups", "trim_level_scope"),
        ("rule_groups", "variant_scope"),
    }
)

_BLANK_TO_ZERO_INTEGER_COLUMNS = frozenset(
    {
        ("options", "price"),
        ("rule_group_members", "display_order"),
    }
)

_REQUIRED_INTEGER_COLUMNS = frozenset(
    {
        ("options", "display_order"),
        ("price_rules", "price_value"),
        ("exclusive_group_members", "display_order"),
    }
)

_BLANK_TO_NULL_TEXT_COLUMNS = frozenset(
    {
        ("options", "display_behavior"),
        ("rule_mapping", "runtime_action"),
        ("variant_overrides", "display_behavior"),
        ("variant_overrides", "section_id"),
    }
)

_BLANK_TO_EMPTY_TEXT_COLUMNS = frozenset(
    {
        ("options", "rpo"),
        ("options", "description"),
        ("options", "detail_raw"),
        ("rule_mapping", "original_detail_raw"),
        ("rule_mapping", "disabled_reason"),
        ("price_rules", "notes"),
        ("rule_groups", "disabled_reason"),
        ("rule_groups", "notes"),
        ("exclusive_groups", "notes"),
        ("variant_overrides", "note"),
    }
)

_REQUIRED_TEXT_COLUMNS = frozenset(
    {
        ("options", "option_id"),
        ("options", "option_name"),
        ("options", "section_id"),
        ("option_availability", "option_id"),
        ("option_availability", "variant_id"),
        ("option_availability", "status"),
        ("rule_mapping", "rule_id"),
        ("rule_mapping", "rule_type"),
        ("price_rules", "price_rule_id"),
        ("price_rules", "price_rule_type"),
        ("price_rules", "target_option_id"),
        ("rule_groups", "group_id"),
        ("rule_groups", "group_type"),
        ("rule_group_members", "group_id"),
        ("exclusive_groups", "group_id"),
        ("exclusive_groups", "selection_mode"),
        ("exclusive_group_members", "group_id"),
        ("exclusive_group_members", "option_id"),
        ("variant_overrides", "option_id"),
        ("variant_overrides", "variant_id"),
    }
)


def _decision(
    code: str,
    message: str,
    *,
    sheet: str = "",
    row: int | None = None,
    column: str = "",
    value: object = None,
) -> DecisionRequired:
    return DecisionRequired(
        code,
        message,
        source_sheet=sheet,
        source_row=row,
        source_column=column,
        value=value,
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as workbook_file:
        for chunk in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _required_text(
    value: object, *, sheet: str, row: int, column: str
) -> str:
    result = _text(value)
    if not result:
        raise _decision(
            "required_value_missing",
            "A required direct model value is blank.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    return result


def _integer(
    value: object,
    *,
    sheet: str,
    row: int,
    column: str,
    nullable: bool = False,
) -> int | None:
    if value is None or _text(value) == "":
        if nullable:
            return None
        raise _decision(
            "integer_value_missing",
            "A required direct model integer is blank.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    try:
        result = int(value)
    except (TypeError, ValueError, OverflowError) as error:
        raise _decision(
            "integer_value_invalid",
            "A direct model integer field is invalid.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        ) from error
    if isinstance(value, float) and value != result:
        raise _decision(
            "integer_value_invalid",
            "A direct model integer field is fractional.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    return result


def _boolean(
    value: object,
    *,
    sheet: str,
    row: int,
    column: str,
    nullable: bool = False,
) -> tuple[bool | None, dict[str, object]]:
    if value is None or (isinstance(value, str) and not value.strip()):
        if nullable:
            return None, {}
        raise _decision(
            "boolean_value_missing",
            "A required direct model boolean is blank.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    if isinstance(value, bool):
        return value, {}
    if value in (0, 1):
        canonical = bool(value)
        return canonical, {
            "original": value,
            "canonical": canonical,
            "transform": "integer_to_boolean",
            "reverse_transform": "restore_original_boolean_from_lineage",
        }
    if isinstance(value, str) and value.strip().lower() in {"true", "false"}:
        canonical = value.strip().lower() == "true"
        return canonical, {
            "original": value,
            "canonical": canonical,
            "transform": "text_to_boolean",
            "reverse_transform": "restore_original_boolean_from_lineage",
        }
    raise _decision(
        "boolean_value_invalid",
        "A direct model boolean field is invalid.",
        sheet=sheet,
        row=row,
        column=column,
        value=value,
    )


def _read_rows(workbook, sheet_name: str, role: str):
    if sheet_name not in workbook.sheetnames:
        raise _decision(
            "registered_source_sheet_missing",
            "A registered direct model source sheet is missing.",
            sheet="model_workbook_sources",
            column="sheet_name",
            value={"role": role, "sheet_name": sheet_name},
        )
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = tuple(next(iterator, ()))
    expected = _HEADERS[role]
    if headers != expected:
        raise _decision(
            "source_role_mismatch",
            "A registered sheet does not match its direct model source role.",
            sheet=sheet_name,
            row=1,
            value={"role": role, "expected": expected, "actual": headers},
        )
    return [
        (row_number, dict(zip(headers, values)))
        for row_number, values in enumerate(iterator, start=2)
        if any(value is not None for value in values)
    ]


def _central_domains(central: Iterable[CompiledTable]):
    tables = {table.name: table for table in central}
    required = {
        "models",
        "body_styles",
        "trim_levels",
        "variants",
        "model_variants",
        "sections",
    }
    if required - set(tables):
        raise _decision(
            "central_domain_missing",
            "Direct model compilation is missing required central domains.",
            value=tuple(sorted(required - set(tables))),
        )
    models = {row.values["model_key"] for row in tables["models"].rows}
    bodies = {row.values["body_style"] for row in tables["body_styles"].rows}
    trims = {row.values["trim_level"] for row in tables["trim_levels"].rows}
    variants = {row.values["variant_id"] for row in tables["variants"].rows}
    model_variants: dict[str, set[str]] = {model: set() for model in LIVE_MODELS}
    for row in tables["model_variants"].rows:
        model_variants.setdefault(str(row.values["model_key"]), set()).add(
            str(row.values["variant_id"])
        )
    sections = {row.values["section_id"] for row in tables["sections"].rows}
    return models, bodies, trims, variants, model_variants, sections


def _ensure_reference(
    value: object,
    domain: set[object],
    *,
    code: str,
    sheet: str,
    row: int,
    column: str,
) -> None:
    if value not in domain:
        raise _decision(
            code,
            "A direct model relationship does not resolve.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )


def _scope(
    value: object,
    domain: set[object],
    *,
    sheet: str,
    row: int,
    column: str,
) -> tuple[str | None, dict[str, object]]:
    original = "" if value is None else str(value)
    trimmed = original.strip()
    if not trimmed or trimmed == "*":
        parameters = {}
        if value is not None:
            transforms: list[str] = []
            if trimmed != original:
                transforms.append("trim")
            transforms.append("asterisk_to_null" if trimmed == "*" else "blank_to_null")
            parameters = {
                "original": original,
                "canonical": None,
                "transform": "_then_".join(transforms),
                "reverse_transform": "restore_original_scope_from_lineage",
            }
        return None, parameters
    normalized = trimmed.lower()
    _ensure_reference(
        normalized,
        domain,
        code="scope_reference_missing",
        sheet=sheet,
        row=row,
        column=column,
    )
    transforms: list[str] = []
    if trimmed != original:
        transforms.append("trim")
    if normalized != trimmed:
        transforms.append("lowercase")
    if not transforms:
        return normalized, {}
    return normalized, {
        "original": original,
        "canonical": normalized,
        "transform": "_then_".join(transforms),
        "reverse_transform": "restore_original_scope_from_lineage",
    }


def _typed_entity_reference(
    entity_id: str,
    option_ids: set[str],
    interior_ids: set[str],
    *,
    sheet: str,
    row: int,
    column: str,
) -> tuple[str | None, str | None]:
    in_options = entity_id in option_ids
    in_interiors = entity_id in interior_ids
    if in_options == in_interiors:
        raise _decision(
            "entity_reference_ambiguous_or_missing",
            "A polymorphic entity must resolve to exactly one model domain.",
            sheet=sheet,
            row=row,
            column=column,
            value=entity_id,
        )
    return (entity_id, None) if in_options else (None, entity_id)


def _compiled_row(
    values: Mapping[str, object],
    sheet: str,
    row: int,
    parameters: Mapping[str, object] | None = None,
    *,
    source_values: Mapping[str, object] | None = None,
    role: str = "",
) -> CompiledRow:
    mapping_parameters = dict(parameters or {})
    if source_values is not None:
        destinations = _COLUMN_DESTINATIONS.get(role, {})
        for source_column in _HEADERS[role]:
            raw = source_values.get(source_column)
            destination_columns = destinations.get(source_column, (source_column,))
            is_alias = (
                len(destination_columns) > 1
                or destination_columns != (source_column,)
            )
            for destination in destination_columns:
                evidence = dict(mapping_parameters.get(destination, {}))
                if "original" in evidence:
                    continue
                canonical = values.get(destination)
                if not is_alias and raw == canonical:
                    continue
                transform, reverse_transform = _schema_contract(
                    role, source_column, destination
                )
                if (
                    not is_alias
                    and isinstance(raw, str)
                    and isinstance(canonical, str)
                    and raw != canonical
                    and raw.strip() == canonical
                ):
                    transform = "trim_text"
                evidence.update(
                    {
                        "original": raw,
                        "canonical": canonical,
                        "transform": str(evidence.get("transform") or transform),
                        "reverse_transform": str(
                            evidence.get("reverse_transform") or reverse_transform
                        ),
                    }
                )
                mapping_parameters[destination] = evidence
    for destination, evidence_value in tuple(mapping_parameters.items()):
        if not isinstance(evidence_value, Mapping):
            continue
        evidence = dict(evidence_value)
        evidence.setdefault("canonical", values.get(destination))
        mapping_parameters[destination] = evidence
    return CompiledRow(
        values=MappingProxyType(dict(values)),
        source_sheet=sheet,
        source_row=row,
        lineage_role="normalized" if any(
            isinstance(value, Mapping) and "original" in value
            for value in mapping_parameters.values()
        ) else "direct",
        mapping_parameters=freeze_mapping(mapping_parameters),
    )


def _schema_contract(
    role: str, source_column: str, destination_column: str
) -> tuple[str, str]:
    destinations = _COLUMN_DESTINATIONS.get(role, {}).get(source_column, ())
    if len(destinations) == 2:
        return (
            "require_nonblank_trim_text_then_typed_entity_reference",
            "coalesce_typed_entity_reference_then_restore_original_text_from_lineage",
        )
    if source_column != destination_column:
        return (
            "require_nonblank_trim_text_then_option_reference",
            "restore_original_text_from_lineage",
        )
    identity = (role, source_column)
    if identity in _BOOLEAN_COLUMNS:
        return (
            "normalize_workbook_boolean",
            "restore_original_boolean_from_lineage",
        )
    if identity in _SCOPE_COLUMNS:
        return (
            "blank_or_asterisk_to_null_else_trim_and_lowercase",
            "restore_original_scope_from_lineage",
        )
    if identity in _BLANK_TO_ZERO_INTEGER_COLUMNS:
        return (
            "blank_to_zero_else_normalize_integer",
            "restore_original_number_from_lineage",
        )
    if identity in _REQUIRED_INTEGER_COLUMNS:
        return (
            "normalize_required_integer",
            "restore_original_number_from_lineage",
        )
    if identity in _BLANK_TO_NULL_TEXT_COLUMNS:
        return (
            "blank_to_null_else_trim_text",
            "restore_original_text_from_lineage",
        )
    if identity in _BLANK_TO_EMPTY_TEXT_COLUMNS:
        return (
            "blank_to_empty_else_trim_text",
            "restore_original_text_from_lineage",
        )
    if identity in _REQUIRED_TEXT_COLUMNS:
        return (
            "require_nonblank_trim_text",
            "restore_original_text_from_lineage",
        )
    raise RuntimeError(f"missing direct schema mapping contract: {identity!r}")


def _schema_mappings(
    model_key: str, role: str, source_sheet: str
) -> tuple[SchemaMapping, ...]:
    destination_table = physical_table(model_key, role)
    destinations = _COLUMN_DESTINATIONS.get(role, {})
    mappings: list[SchemaMapping] = []
    for source_column in _HEADERS[role]:
        for destination_column in destinations.get(source_column, (source_column,)):
            transform, reverse = _schema_contract(
                role, source_column, destination_column
            )
            mappings.append(
                SchemaMapping(
                    source_sheet=source_sheet,
                    source_column=source_column,
                    destination_table=destination_table,
                    destination_column=destination_column,
                    model_key=model_key,
                    transform=transform,
                    reverse_transform=reverse,
                )
            )
    return tuple(mappings)


def _alias_parameters(role: str) -> dict[str, object]:
    parameters: dict[str, object] = {}
    for source_column, destinations in _COLUMN_DESTINATIONS.get(role, {}).items():
        for destination in destinations:
            transform = (
                "typed_entity_reference"
                if len(destinations) == 2
                else "option_reference"
            )
            parameters[destination] = {
                "source_column": source_column,
                "transform": transform,
                "reverse_transform": (
                    "coalesce_typed_entity_reference"
                    if len(destinations) == 2
                    else "restore_original_text_from_lineage"
                ),
            }
    return parameters


def _interior_ids_by_model(workbook, profile: WorkbookProfile) -> dict[str, set[str]]:
    if "model_interior_scope" not in workbook.sheetnames:
        raise _decision(
            "registered_interior_scope_missing",
            "Polymorphic reference typing requires model_interior_scope.",
            sheet="model_interior_scope",
        )
    scope_sheet = workbook["model_interior_scope"]
    iterator = scope_sheet.iter_rows(values_only=True)
    headers = tuple(next(iterator, ()))
    required = {"model_key", "interior_id", "active"}
    if not required <= set(headers):
        raise _decision(
            "source_role_mismatch",
            "model_interior_scope lacks required reference-typing columns.",
            sheet="model_interior_scope",
            row=1,
            value=tuple(sorted(required - set(headers))),
        )
    scoped: dict[str, set[str]] = {model: set() for model in LIVE_MODELS}
    for row_number, values in enumerate(iterator, start=2):
        if not any(value is not None for value in values):
            continue
        row = dict(zip(headers, values))
        model_key = _text(row.get("model_key"))
        if model_key not in scoped:
            continue
        active, _ = _boolean(
            row.get("active"),
            sheet="model_interior_scope",
            row=row_number,
            column="active",
        )
        if active:
            scoped[model_key].add(
                _required_text(
                    row.get("interior_id"),
                    sheet="model_interior_scope",
                    row=row_number,
                    column="interior_id",
                )
            )

    result: dict[str, set[str]] = {}
    for model_key in LIVE_MODELS:
        sources = profile.active_sources.get(model_key, {})
        source_sheet = sources.get("interior_source_sheet")
        if not source_sheet:
            raise _decision(
                "source_role_mismatch",
                "A live model lacks its registered interior source.",
                sheet="model_workbook_sources",
                column="source_role",
                value={"model_key": model_key, "role": "interior_source_sheet"},
            )
        if source_sheet not in workbook.sheetnames:
            raise _decision(
                "registered_source_sheet_missing",
                "A registered interior source sheet is missing.",
                sheet="model_workbook_sources",
                column="sheet_name",
                value={"model_key": model_key, "sheet_name": source_sheet},
            )
        source = workbook[source_sheet]
        source_iterator = source.iter_rows(values_only=True)
        source_headers = tuple(next(source_iterator, ()))
        if "interior_id" not in source_headers:
            raise _decision(
                "source_role_mismatch",
                "A registered interior source lacks interior_id.",
                sheet=source_sheet,
                row=1,
                column="interior_id",
            )
        source_ids = {
            _text(dict(zip(source_headers, values)).get("interior_id"))
            for values in source_iterator
            if any(value is not None for value in values)
        }
        missing = scoped[model_key] - source_ids
        if missing:
            missing_id = sorted(missing)[0]
            raise _decision(
                "entity_reference_ambiguous_or_missing",
                "An active model interior scope does not resolve in its registered source.",
                sheet="model_interior_scope",
                column="interior_id",
                value=missing_id,
            )
        result[model_key] = source_ids & scoped[model_key]
    return result


def _unique_table(
    model_key: str,
    role: str,
    source_sheet: str,
    rows: Iterable[CompiledRow],
) -> CompiledTable:
    compiled_rows = tuple(rows)
    primary_key = _PRIMARY_KEYS[role]
    seen: set[tuple[object, ...]] = set()
    for row in compiled_rows:
        key = tuple(row.values[column] for column in primary_key)
        if key in seen:
            raise _decision(
                "duplicate_direct_key",
                "A direct model source has a duplicate canonical key.",
                sheet=row.source_sheet,
                row=row.source_row,
                value={"role": role, "key": key},
            )
        seen.add(key)
    return CompiledTable(
        name=physical_table(model_key, role),
        primary_key=primary_key,
        rows=compiled_rows,
        model_key=model_key,
        role=role,
        schema_mappings=_schema_mappings(model_key, role, source_sheet),
    )


def compile_direct_model_tables(
    profile: WorkbookProfile,
    workbook_path: Path,
    central: Iterable[CompiledTable],
) -> tuple[CompiledTable, ...]:
    """Compile all nine directly bound roles for every live model."""
    path = Path(workbook_path)
    if path.resolve() != profile.workbook_path.resolve() or _sha256(path) != profile.workbook_sha256:
        raise _decision(
            "workbook_profile_mismatch",
            "The workbook path or content no longer matches its profile.",
            value=str(path),
        )
    blocking = [
        finding
        for finding in profile.findings
        if finding.severity == "error"
        or finding.status in {"decision_required", "contract_mismatch"}
    ]
    if blocking:
        finding = blocking[0]
        raise _decision(
            finding.code,
            finding.message,
            sheet=finding.source_sheet,
            row=finding.source_row,
            column=finding.source_column,
            value=finding.value,
        )
    if profile.active_models != LIVE_MODELS:
        raise _decision(
            "live_model_catalog_mismatch",
            "Direct model compilation requires the canonical live models.",
            value={"expected": LIVE_MODELS, "actual": profile.active_models},
        )

    models, bodies, trims, variants, model_variants, sections = _central_domains(central)
    if models != set(LIVE_MODELS):
        raise _decision(
            "central_model_domain_mismatch",
            "The central model domain differs from the live catalog.",
            value={"expected": LIVE_MODELS, "actual": tuple(sorted(models))},
        )

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        interiors_by_model = _interior_ids_by_model(workbook, profile)
        tables: list[CompiledTable] = []
        for model_key in LIVE_MODELS:
            sources = profile.active_sources.get(model_key)
            if sources is None:
                raise _decision(
                    "source_role_mismatch",
                    "A live model has no active source registry.",
                    sheet="model_workbook_sources",
                    value=model_key,
                )
            source_names: dict[str, str] = {}
            for role, source_role in _DIRECT_ROLE_SOURCES:
                source_sheet = sources.get(source_role)
                if not source_sheet:
                    raise _decision(
                        "source_role_mismatch",
                        "A live model lacks a required direct source binding.",
                        sheet="model_workbook_sources",
                        column="source_role",
                        value={"model_key": model_key, "role": source_role},
                    )
                source_names[role] = source_sheet
            duplicate_sources = {
                source_sheet
                for source_sheet in source_names.values()
                if tuple(source_names.values()).count(source_sheet) > 1
            }
            if duplicate_sources:
                raise _decision(
                    "source_role_mismatch",
                    "One source sheet resolves to multiple direct roles.",
                    sheet="model_workbook_sources",
                    value={
                        "model_key": model_key,
                        "sheets": tuple(sorted(duplicate_sources)),
                    },
                )

            source_rows = {
                role: _read_rows(workbook, source_names[role], role)
                for role, _ in _DIRECT_ROLE_SOURCES
            }
            options_sheet = source_names["options"]
            option_rows: list[CompiledRow] = []
            option_ids: set[str] = set()
            for row_number, row in source_rows["options"]:
                option_id = _required_text(
                    row.get("option_id"),
                    sheet=options_sheet,
                    row=row_number,
                    column="option_id",
                )
                if option_id in option_ids:
                    raise _decision(
                        "duplicate_direct_key",
                        "An options source contains a duplicate option_id.",
                        sheet=options_sheet,
                        row=row_number,
                        column="option_id",
                        value=option_id,
                    )
                option_ids.add(option_id)
                section_id = _required_text(
                    row.get("section_id"),
                    sheet=options_sheet,
                    row=row_number,
                    column="section_id",
                )
                _ensure_reference(
                    section_id,
                    sections,
                    code="section_reference_missing",
                    sheet=options_sheet,
                    row=row_number,
                    column="section_id",
                )
                selectable, selectable_parameter = _boolean(
                    row.get("selectable"),
                    sheet=options_sheet,
                    row=row_number,
                    column="selectable",
                )
                active, active_parameter = _boolean(
                    row.get("active"),
                    sheet=options_sheet,
                    row=row_number,
                    column="active",
                )
                parameters = {
                    key: value
                    for key, value in {
                        "selectable": selectable_parameter,
                        "active": active_parameter,
                    }.items()
                    if value
                }
                raw_price = row.get("price")
                if raw_price is None or _text(raw_price) == "":
                    price = 0
                    parameters["price"] = {
                        "original": raw_price,
                        "canonical": 0,
                        "transform": "blank_to_zero",
                        "reverse_transform": "restore_original_number_from_lineage",
                    }
                else:
                    price = _integer(
                        raw_price,
                        sheet=options_sheet,
                        row=row_number,
                        column="price",
                    )
                option_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "option_id": option_id,
                            "rpo": _text(row.get("rpo")),
                            "price": price,
                            "option_name": _required_text(
                                row.get("option_name"),
                                sheet=options_sheet,
                                row=row_number,
                                column="option_name",
                            ),
                            "description": _text(row.get("description")),
                            "detail_raw": _text(row.get("detail_raw")),
                            "section_id": section_id,
                            "selectable": selectable,
                            "display_order": _integer(
                                row.get("display_order"),
                                sheet=options_sheet,
                                row=row_number,
                                column="display_order",
                            ),
                            "active": active,
                            "display_behavior": _text(row.get("display_behavior")) or None,
                        },
                        options_sheet,
                        row_number,
                        parameters,
                        source_values=row,
                        role="options",
                    )
                )
            tables.append(
                _unique_table(model_key, "options", options_sheet, option_rows)
            )

            ovs_sheet = source_names["option_availability"]
            ovs_rows: list[CompiledRow] = []
            for row_number, row in source_rows["option_availability"]:
                option_id = _required_text(
                    row.get("option_id"),
                    sheet=ovs_sheet,
                    row=row_number,
                    column="option_id",
                )
                _ensure_reference(
                    option_id,
                    option_ids,
                    code="option_reference_missing",
                    sheet=ovs_sheet,
                    row=row_number,
                    column="option_id",
                )
                variant_id = _required_text(
                    row.get("variant_id"),
                    sheet=ovs_sheet,
                    row=row_number,
                    column="variant_id",
                )
                _ensure_reference(
                    variant_id,
                    model_variants[model_key],
                    code="variant_reference_missing",
                    sheet=ovs_sheet,
                    row=row_number,
                    column="variant_id",
                )
                ovs_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "option_id": option_id,
                            "variant_id": variant_id,
                            "status": _required_text(
                                row.get("status"),
                                sheet=ovs_sheet,
                                row=row_number,
                                column="status",
                            ),
                        },
                        ovs_sheet,
                        row_number,
                        source_values=row,
                        role="option_availability",
                    )
                )
            tables.append(
                _unique_table(
                    model_key, "option_availability", ovs_sheet, ovs_rows
                )
            )

            rule_sheet = source_names["rule_mapping"]
            rule_rows: list[CompiledRow] = []
            for row_number, row in source_rows["rule_mapping"]:
                source_id = _required_text(
                    row.get("source_id"),
                    sheet=rule_sheet,
                    row=row_number,
                    column="source_id",
                )
                source_option, source_interior = _typed_entity_reference(
                    source_id,
                    option_ids,
                    interiors_by_model[model_key],
                    sheet=rule_sheet,
                    row=row_number,
                    column="source_id",
                )
                target = _required_text(
                    row.get("target_id"),
                    sheet=rule_sheet,
                    row=row_number,
                    column="target_id",
                )
                _ensure_reference(
                    target,
                    option_ids,
                    code="option_reference_missing",
                    sheet=rule_sheet,
                    row=row_number,
                    column="target_id",
                )
                body_scope, body_parameter = _scope(
                    row.get("body_style_scope"),
                    bodies,
                    sheet=rule_sheet,
                    row=row_number,
                    column="body_style_scope",
                )
                parameters = _alias_parameters("rule_mapping")
                if body_parameter:
                    parameters["body_style_scope"] = body_parameter
                rule_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "rule_id": _required_text(
                                row.get("rule_id"),
                                sheet=rule_sheet,
                                row=row_number,
                                column="rule_id",
                            ),
                            "source_option_id": source_option,
                            "source_interior_id": source_interior,
                            "rule_type": _required_text(
                                row.get("rule_type"),
                                sheet=rule_sheet,
                                row=row_number,
                                column="rule_type",
                            ),
                            "target_option_id": target,
                            "original_detail_raw": _text(row.get("original_detail_raw")),
                            "body_style_scope": body_scope,
                            "trim_level_scope": None,
                            "variant_scope": None,
                            "runtime_action": _text(row.get("runtime_action")) or None,
                            "disabled_reason": _text(row.get("disabled_reason")),
                        },
                        rule_sheet,
                        row_number,
                        parameters,
                        source_values=row,
                        role="rule_mapping",
                    )
                )
            tables.append(
                _unique_table(model_key, "rule_mapping", rule_sheet, rule_rows)
            )

            price_sheet = source_names["price_rules"]
            price_rows: list[CompiledRow] = []
            for row_number, row in source_rows["price_rules"]:
                condition = _required_text(
                    row.get("condition_option_id"),
                    sheet=price_sheet,
                    row=row_number,
                    column="condition_option_id",
                )
                condition_option, condition_interior = _typed_entity_reference(
                    condition,
                    option_ids,
                    interiors_by_model[model_key],
                    sheet=price_sheet,
                    row=row_number,
                    column="condition_option_id",
                )
                target = _required_text(
                    row.get("target_option_id"),
                    sheet=price_sheet,
                    row=row_number,
                    column="target_option_id",
                )
                _ensure_reference(
                    target,
                    option_ids,
                    code="option_reference_missing",
                    sheet=price_sheet,
                    row=row_number,
                    column="target_option_id",
                )
                body_scope, body_parameter = _scope(
                    row.get("body_style_scope"),
                    bodies,
                    sheet=price_sheet,
                    row=row_number,
                    column="body_style_scope",
                )
                trim_scope, trim_parameter = _scope(
                    row.get("trim_level_scope"),
                    trims,
                    sheet=price_sheet,
                    row=row_number,
                    column="trim_level_scope",
                )
                parameters = _alias_parameters("price_rules")
                if body_parameter:
                    parameters["body_style_scope"] = body_parameter
                if trim_parameter:
                    parameters["trim_level_scope"] = trim_parameter
                price_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "price_rule_id": _required_text(
                                row.get("price_rule_id"),
                                sheet=price_sheet,
                                row=row_number,
                                column="price_rule_id",
                            ),
                            "condition_option_id": condition_option,
                            "condition_interior_id": condition_interior,
                            "price_rule_type": _required_text(
                                row.get("price_rule_type"),
                                sheet=price_sheet,
                                row=row_number,
                                column="price_rule_type",
                            ),
                            "target_option_id": target,
                            "price_value": _integer(
                                row.get("price_value"),
                                sheet=price_sheet,
                                row=row_number,
                                column="price_value",
                            ),
                            "body_style_scope": body_scope,
                            "trim_level_scope": trim_scope,
                            "variant_scope": None,
                            "notes": _text(row.get("notes")),
                        },
                        price_sheet,
                        row_number,
                        parameters,
                        source_values=row,
                        role="price_rules",
                    )
                )
            tables.append(
                _unique_table(model_key, "price_rules", price_sheet, price_rows)
            )

            group_sheet = source_names["rule_groups"]
            group_rows: list[CompiledRow] = []
            group_ids: set[str] = set()
            for row_number, row in source_rows["rule_groups"]:
                group_id = _required_text(
                    row.get("group_id"),
                    sheet=group_sheet,
                    row=row_number,
                    column="group_id",
                )
                group_ids.add(group_id)
                source_id = _required_text(
                    row.get("source_id"),
                    sheet=group_sheet,
                    row=row_number,
                    column="source_id",
                )
                _ensure_reference(
                    source_id,
                    option_ids,
                    code="option_reference_missing",
                    sheet=group_sheet,
                    row=row_number,
                    column="source_id",
                )
                body_scope, body_parameter = _scope(
                    row.get("body_style_scope"),
                    bodies,
                    sheet=group_sheet,
                    row=row_number,
                    column="body_style_scope",
                )
                trim_scope, trim_parameter = _scope(
                    row.get("trim_level_scope"),
                    trims,
                    sheet=group_sheet,
                    row=row_number,
                    column="trim_level_scope",
                )
                variant_scope, variant_parameter = _scope(
                    row.get("variant_scope"),
                    model_variants[model_key],
                    sheet=group_sheet,
                    row=row_number,
                    column="variant_scope",
                )
                active, active_parameter = _boolean(
                    row.get("active"),
                    sheet=group_sheet,
                    row=row_number,
                    column="active",
                )
                parameters = _alias_parameters("rule_groups")
                parameters.update(
                    {
                        key: value
                        for key, value in {
                            "body_style_scope": body_parameter,
                            "trim_level_scope": trim_parameter,
                            "variant_scope": variant_parameter,
                            "active": active_parameter,
                        }.items()
                        if value
                    }
                )
                group_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "group_id": group_id,
                            "group_type": _required_text(
                                row.get("group_type"),
                                sheet=group_sheet,
                                row=row_number,
                                column="group_type",
                            ),
                            "source_option_id": source_id,
                            "body_style_scope": body_scope,
                            "trim_level_scope": trim_scope,
                            "variant_scope": variant_scope,
                            "disabled_reason": _text(row.get("disabled_reason")),
                            "active": active,
                            "notes": _text(row.get("notes")),
                        },
                        group_sheet,
                        row_number,
                        parameters,
                        source_values=row,
                        role="rule_groups",
                    )
                )
            tables.append(
                _unique_table(model_key, "rule_groups", group_sheet, group_rows)
            )

            member_sheet = source_names["rule_group_members"]
            member_rows: list[CompiledRow] = []
            for row_number, row in source_rows["rule_group_members"]:
                group_id = _required_text(
                    row.get("group_id"),
                    sheet=member_sheet,
                    row=row_number,
                    column="group_id",
                )
                _ensure_reference(
                    group_id,
                    group_ids,
                    code="rule_group_reference_missing",
                    sheet=member_sheet,
                    row=row_number,
                    column="group_id",
                )
                target = _required_text(
                    row.get("target_id"),
                    sheet=member_sheet,
                    row=row_number,
                    column="target_id",
                )
                _ensure_reference(
                    target,
                    option_ids,
                    code="option_reference_missing",
                    sheet=member_sheet,
                    row=row_number,
                    column="target_id",
                )
                active, active_parameter = _boolean(
                    row.get("active"),
                    sheet=member_sheet,
                    row=row_number,
                    column="active",
                )
                parameters = _alias_parameters("rule_group_members")
                if active_parameter:
                    parameters["active"] = active_parameter
                raw_display_order = row.get("display_order")
                if raw_display_order is None or _text(raw_display_order) == "":
                    display_order = 0
                    parameters["display_order"] = {
                        "original": raw_display_order,
                        "canonical": 0,
                        "transform": "blank_to_zero",
                        "reverse_transform": "restore_original_number_from_lineage",
                    }
                else:
                    display_order = _integer(
                        raw_display_order,
                        sheet=member_sheet,
                        row=row_number,
                        column="display_order",
                    )
                member_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "group_id": group_id,
                            "target_option_id": target,
                            "display_order": display_order,
                            "active": active,
                        },
                        member_sheet,
                        row_number,
                        parameters,
                        source_values=row,
                        role="rule_group_members",
                    )
                )
            tables.append(
                _unique_table(
                    model_key, "rule_group_members", member_sheet, member_rows
                )
            )

            exclusive_sheet = source_names["exclusive_groups"]
            exclusive_rows: list[CompiledRow] = []
            exclusive_ids: set[str] = set()
            for row_number, row in source_rows["exclusive_groups"]:
                group_id = _required_text(
                    row.get("group_id"),
                    sheet=exclusive_sheet,
                    row=row_number,
                    column="group_id",
                )
                exclusive_ids.add(group_id)
                active, active_parameter = _boolean(
                    row.get("active"),
                    sheet=exclusive_sheet,
                    row=row_number,
                    column="active",
                )
                exclusive_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "group_id": group_id,
                            "selection_mode": _required_text(
                                row.get("selection_mode"),
                                sheet=exclusive_sheet,
                                row=row_number,
                                column="selection_mode",
                            ),
                            "active": active,
                            "notes": _text(row.get("notes")),
                        },
                        exclusive_sheet,
                        row_number,
                        {"active": active_parameter} if active_parameter else {},
                        source_values=row,
                        role="exclusive_groups",
                    )
                )
            tables.append(
                _unique_table(
                    model_key, "exclusive_groups", exclusive_sheet, exclusive_rows
                )
            )

            exclusive_member_sheet = source_names["exclusive_group_members"]
            exclusive_member_rows: list[CompiledRow] = []
            for row_number, row in source_rows["exclusive_group_members"]:
                group_id = _required_text(
                    row.get("group_id"),
                    sheet=exclusive_member_sheet,
                    row=row_number,
                    column="group_id",
                )
                _ensure_reference(
                    group_id,
                    exclusive_ids,
                    code="exclusive_group_reference_missing",
                    sheet=exclusive_member_sheet,
                    row=row_number,
                    column="group_id",
                )
                option_id = _required_text(
                    row.get("option_id"),
                    sheet=exclusive_member_sheet,
                    row=row_number,
                    column="option_id",
                )
                _ensure_reference(
                    option_id,
                    option_ids,
                    code="option_reference_missing",
                    sheet=exclusive_member_sheet,
                    row=row_number,
                    column="option_id",
                )
                active, active_parameter = _boolean(
                    row.get("active"),
                    sheet=exclusive_member_sheet,
                    row=row_number,
                    column="active",
                )
                exclusive_member_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "group_id": group_id,
                            "option_id": option_id,
                            "display_order": _integer(
                                row.get("display_order"),
                                sheet=exclusive_member_sheet,
                                row=row_number,
                                column="display_order",
                            ),
                            "active": active,
                        },
                        exclusive_member_sheet,
                        row_number,
                        {"active": active_parameter} if active_parameter else {},
                        source_values=row,
                        role="exclusive_group_members",
                    )
                )
            tables.append(
                _unique_table(
                    model_key,
                    "exclusive_group_members",
                    exclusive_member_sheet,
                    exclusive_member_rows,
                )
            )

            override_sheet = source_names["variant_overrides"]
            override_rows: list[CompiledRow] = []
            for row_number, row in source_rows["variant_overrides"]:
                option_id = _required_text(
                    row.get("option_id"),
                    sheet=override_sheet,
                    row=row_number,
                    column="option_id",
                )
                _ensure_reference(
                    option_id,
                    option_ids,
                    code="option_reference_missing",
                    sheet=override_sheet,
                    row=row_number,
                    column="option_id",
                )
                variant_id = _required_text(
                    row.get("variant_id"),
                    sheet=override_sheet,
                    row=row_number,
                    column="variant_id",
                )
                _ensure_reference(
                    variant_id,
                    model_variants[model_key],
                    code="variant_reference_missing",
                    sheet=override_sheet,
                    row=row_number,
                    column="variant_id",
                )
                section_id = _text(row.get("section_id")) or None
                if section_id is not None:
                    _ensure_reference(
                        section_id,
                        sections,
                        code="section_reference_missing",
                        sheet=override_sheet,
                        row=row_number,
                        column="section_id",
                    )
                selectable, selectable_parameter = _boolean(
                    row.get("selectable"),
                    sheet=override_sheet,
                    row=row_number,
                    column="selectable",
                    nullable=True,
                )
                active, active_parameter = _boolean(
                    row.get("active"),
                    sheet=override_sheet,
                    row=row_number,
                    column="active",
                )
                parameters = {
                    key: value
                    for key, value in {
                        "selectable": selectable_parameter,
                        "active": active_parameter,
                    }.items()
                    if value
                }
                override_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "option_id": option_id,
                            "variant_id": variant_id,
                            "selectable": selectable,
                            "display_behavior": _text(row.get("display_behavior")) or None,
                            "section_id": section_id,
                            "active": active,
                            "note": _text(row.get("note")),
                        },
                        override_sheet,
                        row_number,
                        parameters,
                        source_values=row,
                        role="variant_overrides",
                    )
                )
            tables.append(
                _unique_table(
                    model_key, "variant_overrides", override_sheet, override_rows
                )
            )
        return tuple(tables)
    finally:
        workbook.close()


__all__ = ["compile_direct_model_tables"]
