"""Workbook synchronization and export.

Committed-but-unsynced ``change_history`` rows are translated into an
``editor_ops`` operation batch and handed to the repo's existing gated write
pipeline (``apply_batch`` -> ``save_workbook_safely``), which enforces:
Excel-lock refusal, mtime staleness, batch validation, dry-run on a temp
copy, package + schema validation, automatic backup, and atomic replacement.
This module never writes the workbook through any other path (AGENTS.md §5).

Dry-run is the default; a live write requires ``write=True`` plus the exact
``workbook_mtime_ns`` captured by the dry-run the user reviewed.
"""

from __future__ import annotations

import json
import shutil
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from . import config  # ensures scripts/ on sys.path
from corvette_form_generator import editor_ops  # noqa: E402

from .specs import SPEC_BY_TABLE
from .staging import target_sheet_for


def _now_slug() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def pending_history(conn) -> list[dict]:
    rows = conn.execute(
        "SELECT * FROM change_history WHERE sync_status='pending' "
        "AND status='committed' ORDER BY id").fetchall()
    return [dict(r) for r in rows]


def build_batch(conn, workbook_path: Path) -> dict:
    """editor_ops batch from committed-but-unsynced history rows."""
    items = []
    skipped = []
    for row in pending_history(conn):
        spec = SPEC_BY_TABLE.get(row["entity_type"])
        if spec is None or not spec.editable:
            skipped.append({"history_id": row["id"],
                            "reason": f"no workbook write path for "
                                      f"{row['entity_type']}"})
            continue
        old = json.loads(row["old_json"]) if row["old_json"] else {}
        new = json.loads(row["new_json"]) if row["new_json"] else {}
        sheet = old.get("src_sheet") or target_sheet_for(
            conn, spec, row["model_id"])
        if not sheet:
            skipped.append({"history_id": row["id"],
                            "reason": "target sheet could not be resolved"})
            continue
        key_source = old if row["op"] != "add" else new
        key = {k: str(key_source.get(k, "")) for k in spec.key}
        op: dict = {"action": row["op"], "sheet": sheet, "key": key,
                    "_history_id": row["id"]}
        if row["op"] in ("add", "update"):
            # key columns are immutable on update in editor_ops; they travel
            # in `key` only. Adds carry the full row including keys.
            skip = set(spec.key) if row["op"] == "update" else set()
            op["row"] = {c.header: str(new.get(c.sql_name(), "") or "")
                         for c in spec.columns if c.sql_name() not in skip}
        items.append(op)
    return {
        "workbookMtimeNs": str(workbook_path.stat().st_mtime_ns),
        "items": [{k: v for k, v in op.items() if not k.startswith("_")}
                  for op in items],
        "historyIds": [op["_history_id"] for op in items],
        "skipped": skipped,
    }


def sync_workbook(conn: sqlite3.Connection, workbook_path: Path, *,
                  write: bool = False, confirmed_warnings=(),
                  expected_mtime_ns: str | None = None) -> dict:
    batch = build_batch(conn, workbook_path)
    if not batch["items"]:
        return {"ok": False, "status": "empty", "errors":
                ["no committed changes are pending synchronization"],
                "skipped": batch["skipped"]}
    if write and expected_mtime_ns is not None and \
            str(expected_mtime_ns) != batch["workbookMtimeNs"]:
        return {"ok": False, "status": "stale",
                "errors": ["workbook changed since the reviewed dry-run; "
                           "run dry-run again"], "skipped": batch["skipped"]}
    result = editor_ops.apply_batch(
        workbook_path,
        {"workbookMtimeNs": batch["workbookMtimeNs"], "items": batch["items"]},
        write=write,
        confirmed_warnings=confirmed_warnings,
        source="workbook-manager",
        log_path=config.EDIT_LOG_PATH if write else None,
    )
    result = dict(result)
    result["skipped"] = batch["skipped"]
    result["workbookMtimeNs"] = batch["workbookMtimeNs"]
    result["opCount"] = len(batch["items"])
    if write and result.get("ok"):
        now = datetime.now(timezone.utc).isoformat(timespec="seconds")
        detail = json.dumps({
            "backup": str(result.get("backupPath", "")),
            "synced_at": now,
        })
        conn.executemany(
            "UPDATE change_history SET sync_status='synced', sync_detail=? "
            "WHERE id=?", [(detail, hid) for hid in batch["historyIds"]])
        conn.commit()
    elif write and not result.get("ok"):
        conn.executemany(
            "UPDATE change_history SET sync_status='sync_failed', "
            "sync_detail=? WHERE id=?",
            [(json.dumps({"status": result.get("status"),
                          "errors": result.get("errors", [])}), hid)
             for hid in batch["historyIds"]])
        conn.commit()
        # a failed gated write leaves the workbook untouched (atomic path),
        # but flag loudly for the caller.
    return result


# ── comparison export ────────────────────────────────────────────────

def _coerce_export(raw: str, ctype: str):
    """Best-effort native typing for the comparison artifact (display diffing
    only; the live workbook write path coerces via editor_ops instead)."""
    text = "" if raw is None else str(raw)
    if text == "":
        return ""
    if ctype == "int":
        try:
            return int(text.replace(",", ""))
        except ValueError:
            return text
    if ctype == "bool":
        if text in ("True", "False"):
            return text == "True"
        return text
    return text



def export_comparison_workbook(conn: sqlite3.Connection,
                               workbook_path: Path) -> dict:
    """Regenerate a workbook copy from the database for diffing.

    Starts from a byte copy of the current workbook (preserving unmanaged
    sheets, formatting, and tables), then rewrites the data rows of every
    managed sheet from the normalized tables. Never touches the original.
    """
    from openpyxl import load_workbook

    config.ensure_dirs()
    out_path = config.EXPORT_DIR / f"regenerated-{_now_slug()}.xlsx"
    shutil.copy2(workbook_path, out_path)
    wb = load_workbook(out_path)

    rewritten: dict[str, int] = {}
    for spec in SPEC_BY_TABLE.values():
        sheets: list[tuple[str, str | None]] = []
        if spec.sheet:
            sheets = [(s, None) for s in spec.sheet]
        elif spec.role:
            regs = conn.execute(
                "SELECT model_key, sheet_name FROM sheet_registry WHERE "
                "source_role=?", (spec.role,)).fetchall()
            sheets = [(r["sheet_name"], r["model_key"]) for r in regs]
        for sheet, model in sheets:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = [c.value for c in ws[1]]
            if spec.sheet and len(spec.sheet) > 1:
                where, params = "src_sheet=?", [sheet]
            elif spec.model_scoped:
                where, params = "model_id=?", [model]
            else:
                where, params = "1=1", []
            order = ", ".join(f'"{k}"' for k in ("src_row", *spec.key))
            rows = conn.execute(
                f"SELECT * FROM {spec.table} WHERE {where} "
                f"ORDER BY {order}", params).fetchall()
            header_to_col = {c.header: c for c in spec.columns}
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            for row in rows:
                values = []
                for h in headers:
                    col = header_to_col.get(h)
                    if col is None:
                        values.append("")
                        continue
                    raw = row[col.sql_name()]
                    values.append(_coerce_export(raw, col.ctype))
                ws.append(values)
            rewritten[sheet] = rewritten.get(sheet, 0) + len(rows)
    wb.save(out_path)
    return {"ok": True, "path": str(out_path), "rewritten": rewritten}


def backup_database(conn: sqlite3.Connection, db_path: Path) -> dict:
    config.ensure_dirs()
    out = config.DB_BACKUP_DIR / f"workbook-manager-{_now_slug()}.sqlite3"
    dest = sqlite3.connect(out)
    with dest:
        conn.backup(dest)
    dest.close()
    return {"ok": True, "path": str(out)}
