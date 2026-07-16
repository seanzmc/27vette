"""Compile proven shared workbook sources into model-owned table families."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Iterable, Mapping

from openpyxl import load_workbook

from .catalog import LIVE_MODELS, MODEL_TABLE_ROLES, physical_table
from .compile_types import (
    CompiledRow,
    CompiledTable,
    DecisionRequired,
    Finding,
    SchemaMapping,
    WorkbookProfile,
    freeze_mapping,
)


_SHARED_ROLES = MODEL_TABLE_ROLES[9:]

_PRIMARY_KEYS: Mapping[str, tuple[str, ...]] = MappingProxyType(
    {
        "interiors": ("interior_id",),
        "interior_scope": (
            "interior_id",
            "trim_level",
            "body_style",
            "variant_id",
        ),
        "interior_components": ("interior_id", "rpo", "component_type"),
        "color_overrides": ("interior_id", "option_id"),
        "option_assets": ("option_id",),
        "context_choice_assets": ("context_choice_id",),
        "default_selection_rules": ("rule_id",),
        "runtime_rule_exceptions": ("exception_id",),
    }
)

_INTERIOR_HEADERS = (
    "interior_id",
    "Interior Name",
    "Material",
    "Price",
    "Detail from Disclosure",
    "Color Overrides",
    "Trim",
    "Seat",
    "Interior Code",
    "Suede",
    "Stitch",
    "Two Tone",
    "section_id",
    "active_for_stingray",
    "requires_r6x",
    "included_option_id",
)

_INTERIOR_DESTINATIONS: Mapping[str, str] = MappingProxyType(
    {
        "interior_id": "interior_id",
        "Interior Name": "interior_name",
        "Material": "material",
        "Price": "price",
        "Detail from Disclosure": "detail_from_disclosure",
        "Color Overrides": "color_overrides",
        "Trim": "trim",
        "Seat": "seat",
        "Interior Code": "interior_code",
        "Suede": "suede",
        "Stitch": "stitch",
        "Two Tone": "two_tone",
        "section_id": "section_id",
        "active_for_stingray": "active",
        "requires_r6x": "requires_r6x",
        "included_option_id": "included_option_id",
    }
)

_SHEET_HEADERS: Mapping[str, tuple[str, ...]] = MappingProxyType(
    {
        "model_interior_scope": (
            "model_key",
            "interior_id",
            "trim_level",
            "active",
            "requires_option_id",
            "notes",
            "interior_seat_label",
            "interior_color_family",
            "interior_material_family",
            "interior_variant_label",
            "interior_group_display_order",
            "interior_material_display_order",
            "interior_choice_display_order",
            "interior_hierarchy_levels",
            "interior_parent_group_label",
            "interior_leaf_label",
            "interior_reference_order",
            "grouping_source",
        ),
        "interior_components": (
            "model_key",
            "interior_id",
            "rpo",
            "component_type",
            "label",
            "price_ref_type",
            "price_ref_code",
            "price_trim_scope",
            "display_order",
            "active",
            "notes",
        ),
        "color_overrides": (
            "interior_id",
            "option_id",
            "rule_type",
            "adds_rpo",
        ),
        "asset_map": (
            "model_key",
            "target_type",
            "target_id",
            "image_url",
            "image_alt",
            "image_fit",
            "image_position",
            "hover_image_url",
            "hover_image_alt",
            "hover_image_position",
            "active",
            "notes",
        ),
        "default_selection_rules": (
            "model_key",
            "rule_id",
            "target_option_id",
            "condition_type",
            "condition_id",
            "body_style_scope",
            "trim_level_scope",
            "variant_scope",
            "priority",
            "active",
            "notes",
            "display_behavior",
        ),
        "runtime_rule_exceptions": (
            "model_key",
            "exception_id",
            "source_option_id",
            "target_option_id",
            "exception_type",
            "body_style_scope",
            "trim_level_scope",
            "variant_scope",
            "disabled_reason",
            "active",
            "notes",
        ),
    }
)

_ASSET_FIELDS = (
    "image_url",
    "image_alt",
    "image_fit",
    "image_position",
    "hover_image_url",
    "hover_image_alt",
    "hover_image_position",
)


@dataclass(frozen=True)
class SharedCompilation:
    tables: tuple[CompiledTable, ...]
    mappings: tuple[SchemaMapping, ...]
    findings: tuple[Finding, ...]

    def table(self, name: str) -> CompiledTable:
        return next(table for table in self.tables if table.name == name)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as workbook_file:
        for chunk in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _required_text(
    value: object, *, sheet: str, row: int, column: str
) -> str:
    result = _text(value)
    if not result:
        raise DecisionRequired(
            "shared_required_value_missing",
            "A required shared-source value is blank.",
            source_sheet=sheet,
            source_row=row,
            source_column=column,
            value=value,
        )
    return result


def _boolean(
    value: object, *, sheet: str, row: int, column: str
) -> tuple[bool, dict[str, object]]:
    if isinstance(value, bool):
        return value, {}
    if value in (0, 1):
        canonical = bool(value)
        return canonical, {
            "original": value,
            "canonical": canonical,
            "transform": "integer_to_boolean",
            "reverse_transform": "restore_original_boolean_from_lineage",
        }
    if isinstance(value, str) and value.strip().lower() in {"true", "false"}:
        canonical = value.strip().lower() == "true"
        return canonical, {
            "original": value,
            "canonical": canonical,
            "transform": "text_to_boolean",
            "reverse_transform": "restore_original_boolean_from_lineage",
        }
    raise DecisionRequired(
        "shared_boolean_invalid",
        "A shared-source boolean has no exact interpretation.",
        source_sheet=sheet,
        source_row=row,
        source_column=column,
        value=value,
    )


def _integer(
    value: object,
    *,
    sheet: str,
    row: int,
    column: str,
    nullable: bool = False,
) -> tuple[int | None, dict[str, object]]:
    if value is None or _text(value) == "":
        if nullable:
            return None, {}
        raise DecisionRequired(
            "shared_integer_missing",
            "A required shared-source integer is blank.",
            source_sheet=sheet,
            source_row=row,
            source_column=column,
            value=value,
        )
    try:
        canonical = int(value)
    except (TypeError, ValueError, OverflowError) as error:
        raise DecisionRequired(
            "shared_integer_invalid",
            "A shared-source integer is invalid.",
            source_sheet=sheet,
            source_row=row,
            source_column=column,
            value=value,
        ) from error
    if isinstance(value, float) and value != canonical:
        raise DecisionRequired(
            "shared_integer_invalid",
            "A shared-source integer is fractional.",
            source_sheet=sheet,
            source_row=row,
            source_column=column,
            value=value,
        )
    if value == canonical and not isinstance(value, str):
        return canonical, {}
    return canonical, {
        "original": value,
        "canonical": canonical,
        "transform": "normalize_integer",
        "reverse_transform": "restore_original_number_from_lineage",
    }


def _scope(
    value: object,
    domain: set[str],
    *,
    sheet: str,
    row: int,
    column: str,
) -> tuple[str | None, dict[str, object]]:
    raw = "" if value is None else str(value)
    trimmed = raw.strip()
    if not trimmed or trimmed == "*":
        return None, {
            "original": value,
            "canonical": None,
            "transform": "unrestricted_to_null",
            "reverse_transform": "restore_original_scope_from_lineage",
        }
    canonical = trimmed.lower()
    if canonical not in domain:
        raise DecisionRequired(
            "shared_scope_reference_missing",
            "A restricted shared-source scope has no relational domain row.",
            source_sheet=sheet,
            source_row=row,
            source_column=column,
            value=value,
        )
    if raw == canonical:
        return canonical, {}
    return canonical, {
        "original": value,
        "canonical": canonical,
        "transform": "trim_and_lowercase",
        "reverse_transform": "restore_original_scope_from_lineage",
    }


def _read_rows(workbook, sheet_name: str, expected: tuple[str, ...]):
    if sheet_name not in workbook.sheetnames:
        raise DecisionRequired(
            "registered_source_sheet_missing",
            "A required shared source sheet is missing.",
            source_sheet=sheet_name,
        )
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = tuple(next(iterator, ()))
    if headers != expected:
        raise DecisionRequired(
            "shared_source_role_mismatch",
            "A shared source sheet does not match its canonical role.",
            source_sheet=sheet_name,
            source_row=1,
            value={"expected": expected, "actual": headers},
        )
    return [
        (row_number, dict(zip(headers, values)))
        for row_number, values in enumerate(iterator, start=2)
        if any(value is not None for value in values)
    ]


def _finding(
    code: str,
    message: str,
    *,
    sheet: str,
    row: int,
    column: str = "",
    model_key: str = "",
    value: object = None,
) -> Finding:
    return Finding(
        severity="error",
        status="decision_required",
        code=code,
        message=message,
        source_sheet=sheet,
        source_row=row,
        source_column=column,
        model_key=model_key,
        value=value,
    )


def _row(
    values: Mapping[str, object],
    *,
    sheet: str,
    row: int,
    lineage_role: str,
    parameters: Mapping[str, object] | None = None,
) -> CompiledRow:
    return CompiledRow(
        values=freeze_mapping(values),
        source_sheet=sheet,
        source_row=row,
        lineage_role=lineage_role,
        mapping_parameters=freeze_mapping(parameters or {}),
    )


def _mapping(
    source_sheet: str,
    source_column: str,
    model_key: str,
    role: str,
    destination_column: str,
    *,
    transform: str = "identity",
    reverse_transform: str = "identity",
) -> SchemaMapping:
    return SchemaMapping(
        source_sheet=source_sheet,
        source_column=source_column,
        destination_table=physical_table(model_key, role),
        destination_column=destination_column,
        model_key=model_key,
        transform=transform,
        reverse_transform=reverse_transform,
    )


def _text_parameter(
    raw: object,
    canonical: object,
    *,
    source_column: str | None = None,
    transform: str = "trim_text",
    reverse_transform: str = "restore_original_text_from_lineage",
) -> dict[str, object]:
    parameter: dict[str, object] = {
        "original": raw,
        "canonical": canonical,
        "transform": transform,
        "reverse_transform": reverse_transform,
    }
    if source_column:
        parameter = {"source_column": source_column, **parameter}
    return parameter


def _unique_table(
    model_key: str,
    role: str,
    rows: Iterable[CompiledRow],
    mappings: Iterable[SchemaMapping],
    findings: list[Finding],
) -> CompiledTable:
    unique_rows: list[CompiledRow] = []
    seen: set[tuple[object, ...]] = set()
    primary_key = _PRIMARY_KEYS[role]
    for row in rows:
        key = tuple(row.values[column] for column in primary_key)
        if key in seen:
            findings.append(
                _finding(
                    "duplicate_shared_key",
                    "A shared-source split produced a duplicate canonical key.",
                    sheet=row.source_sheet,
                    row=row.source_row,
                    model_key=model_key,
                    value={"role": role, "key": key},
                )
            )
            continue
        seen.add(key)
        unique_rows.append(row)
    return CompiledTable(
        name=physical_table(model_key, role),
        primary_key=primary_key,
        rows=tuple(unique_rows),
        model_key=model_key,
        role=role,
        schema_mappings=tuple(mappings),
    )


def _central_domains(central: Iterable[CompiledTable]):
    tables = {table.name: table for table in central}
    required = {
        "models",
        "body_styles",
        "trim_levels",
        "model_variants",
        "sections",
        "runtime_context_choices",
    }
    if required - set(tables):
        raise DecisionRequired(
            "central_domain_missing",
            "Shared compilation is missing required central domains.",
            value=tuple(sorted(required - set(tables))),
        )
    return {
        "models": {row.values["model_key"] for row in tables["models"].rows},
        "bodies": {
            row.values["body_style"] for row in tables["body_styles"].rows
        },
        "trims": {
            row.values["trim_level"] for row in tables["trim_levels"].rows
        },
        "variants": {
            model: {
                row.values["variant_id"]
                for row in tables["model_variants"].rows
                if row.values["model_key"] == model
            }
            for model in LIVE_MODELS
        },
        "sections": {
            row.values["section_id"] for row in tables["sections"].rows
        },
        "context_choices": {
            (row.values["model_key"], row.values["context_choice_id"])
            for row in tables["runtime_context_choices"].rows
        },
    }


def _direct_domains(direct: Iterable[CompiledTable]):
    tables = tuple(direct)
    expected = {
        (model, role)
        for model in LIVE_MODELS
        for role in MODEL_TABLE_ROLES[:9]
    }
    actual = {(table.model_key, table.role) for table in tables}
    if actual != expected or len(tables) != len(expected):
        raise DecisionRequired(
            "direct_model_family_incomplete",
            "Shared compilation requires every direct model table exactly once.",
            value={
                "missing": tuple(sorted(expected - actual)),
                "unexpected": tuple(sorted(actual - expected)),
            },
        )
    options: dict[str, set[str]] = {}
    rpos: dict[str, set[str]] = {}
    by_key = {(table.model_key, table.role): table for table in tables}
    for model in LIVE_MODELS:
        rows = by_key[(model, "options")].rows
        options[model] = {str(row.values["option_id"]) for row in rows}
        rpos[model] = {
            str(row.values["rpo"])
            for row in rows
            if str(row.values["rpo"])
        }
    return tables, by_key, options, rpos


def _interior_mappings(model_key: str, source_sheet: str):
    mappings = []
    for source_column in _INTERIOR_HEADERS:
        destination = _INTERIOR_DESTINATIONS[source_column]
        transform = "identity"
        reverse = "identity"
        if source_column != destination:
            transform = "semantic_header_alias"
            reverse = "restore_source_header"
        if source_column == "Price":
            transform = "normalize_required_integer"
            reverse = "restore_original_number_from_lineage"
        elif source_column == "active_for_stingray":
            transform = "legacy_flag_superseded_by_active_model_scope"
            reverse = "restore_original_boolean_from_lineage"
        elif source_column == "requires_r6x":
            transform = "normalize_workbook_boolean"
            reverse = "restore_original_boolean_from_lineage"
        mappings.append(
            _mapping(
                source_sheet,
                source_column,
                model_key,
                "interiors",
                destination,
                transform=transform,
                reverse_transform=reverse,
            )
        )
    mappings.extend(
        (
            _mapping(
                "model_interior_scope",
                "model_key",
                model_key,
                "interiors",
                "model_key",
                transform="prove_registered_source_ownership",
                reverse_transform="restore_exact_model_owner",
            ),
            _mapping(
                "model_interior_scope",
                "active",
                model_key,
                "interiors",
                "active",
                transform="active_scope_membership_to_model_row",
                reverse_transform="restore_scope_active_flag",
            ),
        )
    )
    return tuple(mappings)


def _compile_interiors(
    workbook,
    profile: WorkbookProfile,
    options: Mapping[str, set[str]],
    sections: set[str],
    domains: Mapping[str, object],
    findings: list[Finding],
):
    scope_source_rows = _read_rows(
        workbook,
        "model_interior_scope",
        _SHEET_HEADERS["model_interior_scope"],
    )
    source_sheets = {
        profile.active_sources[model]["interior_source_sheet"]
        for model in LIVE_MODELS
    }
    source_rows = {
        sheet: _read_rows(workbook, sheet, _INTERIOR_HEADERS)
        for sheet in source_sheets
    }
    source_by_sheet_id = {
        sheet: {
            _text(row["interior_id"]): (row_number, row)
            for row_number, row in rows
        }
        for sheet, rows in source_rows.items()
    }

    scope_compiled: dict[str, list[CompiledRow]] = {
        model: [] for model in LIVE_MODELS
    }
    scope_by_model: dict[str, dict[str, tuple[int, Mapping[str, object]]]] = {
        model: {} for model in LIVE_MODELS
    }
    for row_number, raw in scope_source_rows:
        model_key = _text(raw.get("model_key")).lower()
        if model_key not in LIVE_MODELS:
            continue
        try:
            active, active_parameter = _boolean(
                raw.get("active"),
                sheet="model_interior_scope",
                row=row_number,
                column="active",
            )
            if not active:
                continue
            interior_id = _required_text(
                raw.get("interior_id"),
                sheet="model_interior_scope",
                row=row_number,
                column="interior_id",
            )
            source_sheet = profile.active_sources[model_key][
                "interior_source_sheet"
            ]
            if interior_id not in source_by_sheet_id[source_sheet]:
                findings.append(
                    _finding(
                        "shared_reference_missing",
                        "An active interior scope does not resolve in the model's registered interior source.",
                        sheet="model_interior_scope",
                        row=row_number,
                        column="interior_id",
                        model_key=model_key,
                        value=interior_id,
                    )
                )
                continue
            trim_level, trim_parameter = _scope(
                raw.get("trim_level"),
                domains["trims"],
                sheet="model_interior_scope",
                row=row_number,
                column="trim_level",
            )
            requires_option_id = _text(raw.get("requires_option_id")) or None
            if requires_option_id is not None and requires_option_id not in options[model_key]:
                raise DecisionRequired(
                    "shared_option_reference_missing",
                    "An interior scope requirement is not a model option.",
                    source_sheet="model_interior_scope",
                    source_row=row_number,
                    source_column="requires_option_id",
                    value=requires_option_id,
                )
            parameters = {
                key: value
                for key, value in {
                    "trim_level": trim_parameter,
                    "active": active_parameter,
                }.items()
                if value
            }
            values: dict[str, object] = {
                "model_key": model_key,
                "interior_id": interior_id,
                "trim_level": trim_level,
                "body_style": None,
                "variant_id": None,
                "active": active,
                "requires_option_id": requires_option_id,
                "notes": _text(raw.get("notes")),
                "interior_seat_label": _text(raw.get("interior_seat_label")),
                "interior_color_family": _text(raw.get("interior_color_family")),
                "interior_material_family": _text(raw.get("interior_material_family")),
                "interior_variant_label": _text(raw.get("interior_variant_label")),
                "interior_hierarchy_levels": _text(raw.get("interior_hierarchy_levels")),
                "interior_parent_group_label": _text(raw.get("interior_parent_group_label")),
                "interior_leaf_label": _text(raw.get("interior_leaf_label")),
                "grouping_source": _text(raw.get("grouping_source")),
            }
            for column in (
                "interior_group_display_order",
                "interior_material_display_order",
                "interior_choice_display_order",
                "interior_reference_order",
            ):
                value, parameter = _integer(
                    raw.get(column),
                    sheet="model_interior_scope",
                    row=row_number,
                    column=column,
                    nullable=True,
                )
                values[column] = value
                if parameter:
                    parameters[column] = parameter
            scope_compiled[model_key].append(
                _row(
                    values,
                    sheet="model_interior_scope",
                    row=row_number,
                    lineage_role="normalized" if parameters else "direct",
                    parameters=parameters,
                )
            )
            scope_by_model[model_key][interior_id] = (row_number, raw)
        except DecisionRequired as error:
            findings.append(
                _finding(
                    error.code,
                    str(error),
                    sheet=error.source_sheet or "model_interior_scope",
                    row=error.source_row or row_number,
                    column=error.source_column,
                    model_key=model_key,
                    value=error.value,
                )
            )

    interior_compiled: dict[str, list[CompiledRow]] = {
        model: [] for model in LIVE_MODELS
    }
    for source_sheet, rows in source_rows.items():
        for source_row, raw in rows:
            interior_id = _text(raw.get("interior_id"))
            owners = [
                model
                for model in LIVE_MODELS
                if profile.active_sources[model]["interior_source_sheet"]
                == source_sheet
                and interior_id in scope_by_model[model]
            ]
            if not owners:
                findings.append(
                    _finding(
                        "shared_row_owner_unresolved",
                        "An interior source row has no active model scope in a registered owner.",
                        sheet=source_sheet,
                        row=source_row,
                        column="interior_id",
                        value=interior_id,
                    )
                )
                continue
            for model_key in owners:
                scope_row_number, scope_raw = scope_by_model[model_key][interior_id]
                try:
                    section_id = _required_text(
                        raw.get("section_id"),
                        sheet=source_sheet,
                        row=source_row,
                        column="section_id",
                    )
                    if section_id not in sections:
                        raise DecisionRequired(
                            "shared_section_reference_missing",
                            "An interior section does not resolve centrally.",
                            source_sheet=source_sheet,
                            source_row=source_row,
                            source_column="section_id",
                            value=section_id,
                        )
                    price, price_parameter = _integer(
                        raw.get("Price"),
                        sheet=source_sheet,
                        row=source_row,
                        column="Price",
                    )
                    requires_r6x, requires_parameter = _boolean(
                        raw.get("requires_r6x"),
                        sheet=source_sheet,
                        row=source_row,
                        column="requires_r6x",
                    )
                    included = _text(raw.get("included_option_id")) or None
                    if included is not None and included not in options[model_key]:
                        raise DecisionRequired(
                            "shared_option_reference_missing",
                            "An included interior option is not a model option.",
                            source_sheet=source_sheet,
                            source_row=source_row,
                            source_column="included_option_id",
                            value=included,
                        )
                    legacy_active = raw.get("active_for_stingray")
                    parameters: dict[str, object] = {
                        "model_key": {
                            "original": scope_raw.get("model_key"),
                            "canonical": model_key,
                            "transform": "prove_registered_source_and_active_scope_owner",
                            "reverse_transform": "restore_exact_model_owner",
                            "owner_source_sheet": "model_interior_scope",
                            "owner_source_row": scope_row_number,
                        },
                        "active": {
                            "original": scope_raw.get("active"),
                            "canonical": True,
                            "transform": "active_scope_membership_to_model_row",
                            "reverse_transform": "restore_scope_active_flag",
                            "owner_source_sheet": "model_interior_scope",
                            "owner_source_row": scope_row_number,
                            "legacy_active_for_stingray": legacy_active,
                        },
                    }
                    if price_parameter:
                        parameters["price"] = price_parameter
                    if requires_parameter:
                        parameters["requires_r6x"] = requires_parameter
                    values = {
                        "model_key": model_key,
                        "interior_id": interior_id,
                        "interior_name": _text(raw.get("Interior Name")),
                        "material": _text(raw.get("Material")),
                        "price": price,
                        "detail_from_disclosure": _text(
                            raw.get("Detail from Disclosure")
                        ),
                        "color_overrides": _text(raw.get("Color Overrides")),
                        "trim": _text(raw.get("Trim")),
                        "seat": _text(raw.get("Seat")),
                        "interior_code": _text(raw.get("Interior Code")),
                        "suede": _text(raw.get("Suede")),
                        "stitch": _text(raw.get("Stitch")),
                        "two_tone": _text(raw.get("Two Tone")),
                        "section_id": section_id,
                        "requires_r6x": requires_r6x,
                        "included_option_id": included,
                        "active": True,
                    }
                    for source_column, destination in _INTERIOR_DESTINATIONS.items():
                        if source_column == destination or destination == "active":
                            continue
                        transform = "semantic_header_alias"
                        reverse_transform = "restore_source_header"
                        if source_column == "Price":
                            transform = "normalize_required_integer"
                            reverse_transform = (
                                "restore_original_number_from_lineage"
                            )
                        evidence = dict(parameters.get(destination, {}))
                        evidence.setdefault("source_column", source_column)
                        evidence.setdefault("original", raw.get(source_column))
                        evidence.setdefault("canonical", values[destination])
                        evidence.setdefault("transform", transform)
                        evidence.setdefault(
                            "reverse_transform", reverse_transform
                        )
                        parameters[destination] = evidence
                    interior_compiled[model_key].append(
                        _row(
                            values,
                            sheet=source_sheet,
                            row=source_row,
                            lineage_role="shared_source_split",
                            parameters=parameters,
                        )
                    )
                except DecisionRequired as error:
                    findings.append(
                        _finding(
                            error.code,
                            str(error),
                            sheet=error.source_sheet or source_sheet,
                            row=error.source_row or source_row,
                            column=error.source_column,
                            model_key=model_key,
                            value=error.value,
                        )
                    )

    scope_mappings: dict[str, tuple[SchemaMapping, ...]] = {}
    for model in LIVE_MODELS:
        mappings: list[SchemaMapping] = []
        for source_column in _SHEET_HEADERS["model_interior_scope"]:
            destination = source_column
            transform = "identity"
            reverse = "identity"
            if source_column in {
                "trim_level",
            }:
                transform = "blank_or_asterisk_to_null_else_trim_and_lowercase"
                reverse = "restore_original_scope_from_lineage"
            elif source_column == "active":
                transform = "normalize_workbook_boolean"
                reverse = "restore_original_boolean_from_lineage"
            elif source_column in {
                "interior_group_display_order",
                "interior_material_display_order",
                "interior_choice_display_order",
                "interior_reference_order",
            }:
                transform = "blank_to_null_else_normalize_integer"
                reverse = "restore_original_number_from_lineage"
            mappings.append(
                _mapping(
                    "model_interior_scope",
                    source_column,
                    model,
                    "interior_scope",
                    destination,
                    transform=transform,
                    reverse_transform=reverse,
                )
            )
        scope_mappings[model] = tuple(mappings)
    return interior_compiled, scope_compiled, scope_by_model, {
        model: _interior_mappings(
            model, profile.active_sources[model]["interior_source_sheet"]
        )
        for model in LIVE_MODELS
    }, scope_mappings


def _compile_components(
    workbook,
    scope_by_model,
    findings: list[Finding],
):
    rows = _read_rows(
        workbook,
        "interior_components",
        _SHEET_HEADERS["interior_components"],
    )
    compiled = {model: [] for model in LIVE_MODELS}
    for row_number, raw in rows:
        model_key = _text(raw.get("model_key")).lower()
        if model_key not in LIVE_MODELS:
            continue
        try:
            interior_id = _required_text(
                raw.get("interior_id"),
                sheet="interior_components",
                row=row_number,
                column="interior_id",
            )
            if interior_id not in scope_by_model[model_key]:
                raise DecisionRequired(
                    "shared_reference_missing",
                    "An interior component does not resolve to an owned model interior.",
                    source_sheet="interior_components",
                    source_row=row_number,
                    source_column="interior_id",
                    value=interior_id,
                )
            display_order, order_parameter = _integer(
                raw.get("display_order"),
                sheet="interior_components",
                row=row_number,
                column="display_order",
            )
            active, active_parameter = _boolean(
                raw.get("active"),
                sheet="interior_components",
                row=row_number,
                column="active",
            )
            parameters = {
                key: value
                for key, value in {
                    "display_order": order_parameter,
                    "active": active_parameter,
                }.items()
                if value
            }
            compiled[model_key].append(
                _row(
                    {
                        "model_key": model_key,
                        "interior_id": interior_id,
                        "rpo": _required_text(
                            raw.get("rpo"),
                            sheet="interior_components",
                            row=row_number,
                            column="rpo",
                        ),
                        "component_type": _required_text(
                            raw.get("component_type"),
                            sheet="interior_components",
                            row=row_number,
                            column="component_type",
                        ),
                        "label": _required_text(
                            raw.get("label"),
                            sheet="interior_components",
                            row=row_number,
                            column="label",
                        ),
                        "price_ref_type": _text(raw.get("price_ref_type")),
                        "price_ref_code": _text(raw.get("price_ref_code")),
                        "price_trim_scope": _text(raw.get("price_trim_scope")),
                        "display_order": display_order,
                        "active": active,
                        "notes": _text(raw.get("notes")),
                    },
                    sheet="interior_components",
                    row=row_number,
                    lineage_role="normalized" if parameters else "direct",
                    parameters=parameters,
                )
            )
        except DecisionRequired as error:
            findings.append(
                _finding(
                    error.code,
                    str(error),
                    sheet=error.source_sheet or "interior_components",
                    row=error.source_row or row_number,
                    column=error.source_column,
                    model_key=model_key,
                    value=error.value,
                )
            )
    mappings = {
        model: tuple(
            _mapping(
                "interior_components",
                column,
                model,
                "interior_components",
                column,
                transform=(
                    "normalize_required_integer"
                    if column == "display_order"
                    else "normalize_workbook_boolean"
                    if column == "active"
                    else "identity"
                ),
                reverse_transform=(
                    "restore_original_number_from_lineage"
                    if column == "display_order"
                    else "restore_original_boolean_from_lineage"
                    if column == "active"
                    else "identity"
                ),
            )
            for column in _SHEET_HEADERS["interior_components"]
        )
        for model in LIVE_MODELS
    }
    return compiled, mappings


def _compile_color_overrides(
    workbook,
    profile: WorkbookProfile,
    scope_by_model,
    options: Mapping[str, set[str]],
    findings: list[Finding],
):
    color_sheets = {
        profile.active_sources[model]["color_overrides_sheet"]
        for model in LIVE_MODELS
    }
    compiled = {model: [] for model in LIVE_MODELS}
    mappings = {model: [] for model in LIVE_MODELS}
    for sheet in color_sheets:
        rows = _read_rows(workbook, sheet, _SHEET_HEADERS["color_overrides"])
        for row_number, raw in rows:
            interior_id = _text(raw.get("interior_id"))
            option_id = _text(raw.get("option_id"))
            added_option_id = _text(raw.get("adds_rpo"))
            eligible_models = tuple(
                model
                for model in LIVE_MODELS
                if profile.active_sources[model]["color_overrides_sheet"] == sheet
            )
            reference_models = {
                "interior_id": tuple(
                    model
                    for model in eligible_models
                    if interior_id in scope_by_model[model]
                ),
                "option_id": tuple(
                    model for model in eligible_models if option_id in options[model]
                ),
                "adds_rpo": tuple(
                    model
                    for model in eligible_models
                    if added_option_id in options[model]
                ),
            }
            reference_values = {
                "interior_id": interior_id,
                "option_id": option_id,
                "adds_rpo": added_option_id,
            }
            unresolved = [
                column for column, models in reference_models.items() if not models
            ]
            for column in unresolved:
                findings.append(
                    _finding(
                        "shared_color_reference_missing",
                        "A color override reference resolves in no eligible live model.",
                        sheet=sheet,
                        row=row_number,
                        column=column,
                        value=reference_values[column],
                    )
                )
            if unresolved:
                continue
            owner_set = set(eligible_models)
            for models in reference_models.values():
                owner_set.intersection_update(models)
            owners = tuple(model for model in LIVE_MODELS if model in owner_set)
            if not owners:
                findings.append(
                    _finding(
                        "shared_color_owner_conflict",
                        "Color override references resolve separately but share no model owner.",
                        sheet=sheet,
                        row=row_number,
                        value={
                            column: {
                                "value": reference_values[column],
                                "models": models,
                            }
                            for column, models in reference_models.items()
                        },
                    )
                )
                continue
            for model in owners:
                compiled[model].append(
                    _row(
                        {
                            "model_key": model,
                            "interior_id": interior_id,
                            "option_id": option_id,
                            "rule_type": _text(raw.get("rule_type")).lower(),
                            "added_option_id": added_option_id,
                        },
                        sheet=sheet,
                        row=row_number,
                        lineage_role="shared_source_split",
                        parameters={
                            "model_key": {
                                "original": None,
                                "canonical": model,
                                "transform": "resolve_owned_interior_and_option_references",
                                "reverse_transform": "merge_shared_source_row",
                            },
                            "added_option_id": _text_parameter(
                                raw.get("adds_rpo"),
                                added_option_id,
                                source_column="adds_rpo",
                                transform="option_reference",
                            ),
                        },
                    )
                )
    for model in LIVE_MODELS:
        sheet = profile.active_sources[model]["color_overrides_sheet"]
        mappings[model] = [
            _mapping(sheet, "interior_id", model, "color_overrides", "interior_id"),
            _mapping(sheet, "option_id", model, "color_overrides", "option_id"),
            _mapping(sheet, "rule_type", model, "color_overrides", "rule_type", transform="trim_and_lowercase", reverse_transform="restore_original_text_from_lineage"),
            _mapping(sheet, "adds_rpo", model, "color_overrides", "added_option_id", transform="option_reference", reverse_transform="restore_original_text_from_lineage"),
        ]
    return compiled, {model: tuple(rows) for model, rows in mappings.items()}


def _asset_mappings(model: str, role: str):
    mappings = [
        _mapping("asset_map", "model_key", model, role, "model_key"),
    ]
    if role == "option_assets":
        mappings.extend(
            (
                _mapping("asset_map", "target_type", model, role, "option_id", transform="route_option_target", reverse_transform="restore_target_type_option"),
                _mapping("asset_map", "target_id", model, role, "option_id", transform="option_reference", reverse_transform="restore_original_text_from_lineage"),
            )
        )
    else:
        mappings.extend(
            (
                _mapping("asset_map", "target_type", model, role, "context_choice_id", transform="route_context_choice_target", reverse_transform="restore_target_type_context_choice"),
                _mapping("asset_map", "target_id", model, role, "context_choice_id", transform="context_choice_reference", reverse_transform="restore_original_text_from_lineage"),
            )
        )
    mappings.extend(
        _mapping("asset_map", column, model, role, column)
        for column in (*_ASSET_FIELDS, "active", "notes")
    )
    return tuple(mappings)


def _compile_assets(
    workbook,
    options: Mapping[str, set[str]],
    context_choices: set[tuple[str, str]],
    findings: list[Finding],
):
    rows = _read_rows(workbook, "asset_map", _SHEET_HEADERS["asset_map"])
    option_rows: dict[str, dict[str, CompiledRow]] = {
        model: {} for model in LIVE_MODELS
    }
    context_rows: dict[str, dict[str, CompiledRow]] = {
        model: {} for model in LIVE_MODELS
    }

    wildcard_rows = []
    exact_rows = []
    for row_number, raw in rows:
        try:
            active, _ = _boolean(
                raw.get("active"),
                sheet="asset_map",
                row=row_number,
                column="active",
            )
        except DecisionRequired as error:
            findings.append(
                _finding(
                    error.code,
                    str(error),
                    sheet="asset_map",
                    row=row_number,
                    column="active",
                    value=error.value,
                )
            )
            continue
        if not active:
            continue
        if _text(raw.get("model_key")) == "*":
            wildcard_rows.append((row_number, raw))
        else:
            exact_rows.append((row_number, raw))

    first_by_precedence: dict[tuple[str, str, str], int] = {}
    for row_number, raw in (*wildcard_rows, *exact_rows):
        row_model = _text(raw.get("model_key")).lower()
        target_type = _text(raw.get("target_type")).lower()
        target_id = _text(raw.get("target_id"))
        image_url = _text(raw.get("image_url"))
        if row_model == "*" and target_type != "option":
            findings.append(
                _finding(
                    "wildcard_asset_target_unsupported",
                    "Wildcard asset rows are supported only for option targets.",
                    sheet="asset_map",
                    row=row_number,
                    column="target_type",
                    value=target_type,
                )
            )
            continue
        if target_type == "model":
            # Central compilation owns model-card assets.
            continue
        precedence_key = (row_model, target_type, target_id)
        if precedence_key in first_by_precedence:
            findings.append(
                _finding(
                    "duplicate_asset_precedence_key",
                    "Multiple active asset rows occupy the same precedence key.",
                    sheet="asset_map",
                    row=row_number,
                    column="target_id",
                    model_key=row_model,
                    value={
                        "model_key": row_model,
                        "target_type": target_type,
                        "target_id": target_id,
                        "first_source_row": first_by_precedence[precedence_key],
                    },
                )
            )
            continue
        first_by_precedence[precedence_key] = row_number
        if row_model != "*" and row_model not in LIVE_MODELS:
            continue
        if not target_id or not image_url:
            findings.append(
                _finding(
                    "shared_asset_target_invalid",
                    "An active asset row requires a target and image URL.",
                    sheet="asset_map",
                    row=row_number,
                    column="target_id" if not target_id else "image_url",
                    model_key="" if row_model == "*" else row_model,
                    value=target_id or image_url,
                )
            )
            continue
        asset_values = {
            column: _text(raw.get(column)) for column in _ASSET_FIELDS
        }
        if target_type == "option":
            owners = (
                [model for model in LIVE_MODELS if target_id in options[model]]
                if row_model == "*"
                else [row_model] if target_id in options[row_model] else []
            )
            if not owners:
                findings.append(
                    _finding(
                        "shared_row_owner_unresolved",
                        "An option asset does not resolve to an applicable model option.",
                        sheet="asset_map",
                        row=row_number,
                        column="target_id",
                        model_key="" if row_model == "*" else row_model,
                        value=target_id,
                    )
                )
                continue
            for model in owners:
                parameters = {
                    "option_id": _text_parameter(
                        raw.get("target_id"),
                        target_id,
                        source_column="target_id",
                        transform="option_reference",
                    )
                }
                if row_model == "*":
                    parameters["model_key"] = {
                        "original": "*",
                        "canonical": model,
                        "transform": "expand_wildcard_to_matching_option_owner",
                        "reverse_transform": "merge_wildcard_source_row",
                    }
                option_rows[model][target_id] = _row(
                    {
                        "model_key": model,
                        "option_id": target_id,
                        **asset_values,
                        "active": True,
                        "notes": _text(raw.get("notes")),
                    },
                    sheet="asset_map",
                    row=row_number,
                    lineage_role=(
                        "shared_source_split" if row_model == "*" else "direct"
                    ),
                    parameters=parameters,
                )
        elif target_type == "context_choice":
            if row_model == "*":
                raise AssertionError("wildcard context asset passed earlier guard")
            if (row_model, target_id) not in context_choices:
                findings.append(
                    _finding(
                        "shared_context_choice_reference_missing",
                        "A context-choice asset target does not resolve to the model runtime context domain.",
                        sheet="asset_map",
                        row=row_number,
                        column="target_id",
                        model_key=row_model,
                        value=target_id,
                    )
                )
                continue
            context_rows[row_model][target_id] = _row(
                {
                    "model_key": row_model,
                    "context_choice_id": target_id,
                    **asset_values,
                    "active": True,
                    "notes": _text(raw.get("notes")),
                },
                sheet="asset_map",
                row=row_number,
                lineage_role="direct",
                parameters={
                    "context_choice_id": _text_parameter(
                        target_id,
                        target_id,
                        source_column="target_id",
                        transform="context_choice_reference",
                    ),
                },
            )
        else:
            findings.append(
                _finding(
                    "shared_asset_target_type_unknown",
                    "An active asset target type has no canonical destination.",
                    sheet="asset_map",
                    row=row_number,
                    column="target_type",
                    model_key="" if row_model == "*" else row_model,
                    value=target_type,
                )
            )
    return (
        {model: tuple(rows.values()) for model, rows in option_rows.items()},
        {model: tuple(rows.values()) for model, rows in context_rows.items()},
        {model: _asset_mappings(model, "option_assets") for model in LIVE_MODELS},
        {
            model: _asset_mappings(model, "context_choice_assets")
            for model in LIVE_MODELS
        },
    )


def _rule_mappings(model: str, role: str, headers: tuple[str, ...]):
    mappings = []
    for column in headers:
        transform = "identity"
        reverse = "identity"
        if column in {"body_style_scope", "trim_level_scope", "variant_scope"}:
            transform = "blank_or_asterisk_to_null_else_trim_and_lowercase"
            reverse = "restore_original_scope_from_lineage"
        elif column == "priority":
            transform = "normalize_required_integer"
            reverse = "restore_original_number_from_lineage"
        elif column == "active":
            transform = "normalize_workbook_boolean"
            reverse = "restore_original_boolean_from_lineage"
        mappings.append(
            _mapping(
                role,
                column,
                model,
                role,
                column,
                transform=transform,
                reverse_transform=reverse,
            )
        )
    return tuple(mappings)


def _compile_rules(
    workbook,
    role: str,
    options: Mapping[str, set[str]],
    option_rpos: Mapping[str, set[str]],
    domains: Mapping[str, object],
    findings: list[Finding],
):
    rows = _read_rows(workbook, role, _SHEET_HEADERS[role])
    compiled = {model: [] for model in LIVE_MODELS}
    id_column = "rule_id" if role == "default_selection_rules" else "exception_id"
    for row_number, raw in rows:
        model = _text(raw.get("model_key")).lower()
        if model not in LIVE_MODELS:
            continue
        try:
            body_scope, body_parameter = _scope(
                raw.get("body_style_scope"),
                domains["bodies"],
                sheet=role,
                row=row_number,
                column="body_style_scope",
            )
            trim_scope, trim_parameter = _scope(
                raw.get("trim_level_scope"),
                domains["trims"],
                sheet=role,
                row=row_number,
                column="trim_level_scope",
            )
            variant_scope, variant_parameter = _scope(
                raw.get("variant_scope"),
                domains["variants"][model],
                sheet=role,
                row=row_number,
                column="variant_scope",
            )
            active, active_parameter = _boolean(
                raw.get("active"), sheet=role, row=row_number, column="active"
            )
            parameters = {
                key: value
                for key, value in {
                    "body_style_scope": body_parameter,
                    "trim_level_scope": trim_parameter,
                    "variant_scope": variant_parameter,
                    "active": active_parameter,
                }.items()
                if value
            }
            if role == "default_selection_rules":
                target = _required_text(
                    raw.get("target_option_id"),
                    sheet=role,
                    row=row_number,
                    column="target_option_id",
                )
                if target not in options[model]:
                    raise DecisionRequired(
                        "shared_option_reference_missing",
                        "A default target is not an option owned by its exact model.",
                        source_sheet=role,
                        source_row=row_number,
                        source_column="target_option_id",
                        value=target,
                    )
                condition_type = _required_text(
                    raw.get("condition_type"),
                    sheet=role,
                    row=row_number,
                    column="condition_type",
                )
                condition_id = _text(raw.get("condition_id")) or None
                if condition_type == "always":
                    if condition_id is not None:
                        raise DecisionRequired(
                            "shared_default_condition_invalid",
                            "An always default must not name a condition target.",
                            source_sheet=role,
                            source_row=row_number,
                            source_column="condition_id",
                            value=condition_id,
                        )
                elif condition_type == "unless_selected_rpo":
                    if condition_id not in option_rpos[model]:
                        raise DecisionRequired(
                            "shared_default_condition_reference_missing",
                            "A default condition RPO is not owned by its exact model.",
                            source_sheet=role,
                            source_row=row_number,
                            source_column="condition_id",
                            value=condition_id,
                        )
                elif condition_type == "unless_selected_section":
                    if condition_id not in domains["sections"]:
                        raise DecisionRequired(
                            "shared_default_condition_reference_missing",
                            "A default condition section has no central route section.",
                            source_sheet=role,
                            source_row=row_number,
                            source_column="condition_id",
                            value=condition_id,
                        )
                elif condition_type == "when_selected_unless_selected_section":
                    if condition_id not in options[model]:
                        raise DecisionRequired(
                            "shared_default_condition_reference_missing",
                            "A selected-option default condition is not owned by its exact model.",
                            source_sheet=role,
                            source_row=row_number,
                            source_column="condition_id",
                            value=condition_id,
                        )
                else:
                    raise DecisionRequired(
                        "shared_default_condition_type_unknown",
                        "A default condition type has no canonical reference contract.",
                        source_sheet=role,
                        source_row=row_number,
                        source_column="condition_type",
                        value=condition_type,
                    )
                priority, priority_parameter = _integer(
                    raw.get("priority"),
                    sheet=role,
                    row=row_number,
                    column="priority",
                )
                if priority_parameter:
                    parameters["priority"] = priority_parameter
                values = {
                    "model_key": model,
                    "rule_id": _required_text(
                        raw.get("rule_id"),
                        sheet=role,
                        row=row_number,
                        column="rule_id",
                    ),
                    "target_option_id": target,
                    "condition_type": condition_type,
                    "condition_id": condition_id,
                    "body_style_scope": body_scope,
                    "trim_level_scope": trim_scope,
                    "variant_scope": variant_scope,
                    "priority": priority,
                    "active": active,
                    "notes": _text(raw.get("notes")),
                    "display_behavior": _text(raw.get("display_behavior")) or None,
                }
            else:
                source = _required_text(
                    raw.get("source_option_id"),
                    sheet=role,
                    row=row_number,
                    column="source_option_id",
                )
                target = _required_text(
                    raw.get("target_option_id"),
                    sheet=role,
                    row=row_number,
                    column="target_option_id",
                )
                for column, value in (
                    ("source_option_id", source),
                    ("target_option_id", target),
                ):
                    if value not in options[model]:
                        raise DecisionRequired(
                            "shared_option_reference_missing",
                            "A runtime exception option is not owned by its exact model.",
                            source_sheet=role,
                            source_row=row_number,
                            source_column=column,
                            value=value,
                        )
                values = {
                    "model_key": model,
                    "exception_id": _required_text(
                        raw.get("exception_id"),
                        sheet=role,
                        row=row_number,
                        column="exception_id",
                    ),
                    "source_option_id": source,
                    "target_option_id": target,
                    "exception_type": _required_text(
                        raw.get("exception_type"),
                        sheet=role,
                        row=row_number,
                        column="exception_type",
                    ),
                    "body_style_scope": body_scope,
                    "trim_level_scope": trim_scope,
                    "variant_scope": variant_scope,
                    "disabled_reason": _text(raw.get("disabled_reason")),
                    "active": active,
                    "notes": _text(raw.get("notes")),
                }
            compiled[model].append(
                _row(
                    values,
                    sheet=role,
                    row=row_number,
                    lineage_role="normalized" if parameters else "direct",
                    parameters=parameters,
                )
            )
        except DecisionRequired as error:
            findings.append(
                _finding(
                    error.code,
                    str(error),
                    sheet=error.source_sheet or role,
                    row=error.source_row or row_number,
                    column=error.source_column,
                    model_key=model,
                    value=error.value,
                )
            )
    return compiled, {
        model: _rule_mappings(model, role, _SHEET_HEADERS[role])
        for model in LIVE_MODELS
    }


def compile_shared_model_tables(
    profile: WorkbookProfile,
    workbook_path: Path,
    central: Iterable[CompiledTable],
    direct: Iterable[CompiledTable],
) -> SharedCompilation:
    """Complete direct model families with proven shared-source splits."""
    path = Path(workbook_path)
    if (
        path.resolve() != profile.workbook_path.resolve()
        or _sha256(path) != profile.workbook_sha256
    ):
        raise DecisionRequired(
            "workbook_profile_mismatch",
            "The workbook path or content no longer matches its profile.",
            value=str(path),
        )
    if profile.active_models != LIVE_MODELS:
        raise DecisionRequired(
            "live_model_catalog_mismatch",
            "Shared compilation requires the canonical live models.",
            value={"expected": LIVE_MODELS, "actual": profile.active_models},
        )
    central_tables = tuple(central)
    domains = _central_domains(central_tables)
    if domains["models"] != set(LIVE_MODELS):
        raise DecisionRequired(
            "central_model_domain_mismatch",
            "The central model domain differs from the live catalog.",
            value=domains["models"],
        )
    direct_tables, direct_by_key, options, option_rpos = _direct_domains(direct)
    findings: list[Finding] = [
        finding
        for finding in profile.findings
        if finding.code == "inactive_future_row_excluded"
    ]

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        (
            interior_rows,
            scope_rows,
            scope_by_model,
            interior_mappings,
            scope_mappings,
        ) = _compile_interiors(
            workbook,
            profile,
            options,
            domains["sections"],
            domains,
            findings,
        )
        component_rows, component_mappings = _compile_components(
            workbook, scope_by_model, findings
        )
        color_rows, color_mappings = _compile_color_overrides(
            workbook, profile, scope_by_model, options, findings
        )
        (
            option_asset_rows,
            context_asset_rows,
            option_asset_mappings,
            context_asset_mappings,
        ) = _compile_assets(
            workbook, options, domains["context_choices"], findings
        )
        default_rows, default_mappings = _compile_rules(
            workbook,
            "default_selection_rules",
            options,
            option_rpos,
            domains,
            findings,
        )
        exception_rows, exception_mappings = _compile_rules(
            workbook,
            "runtime_rule_exceptions",
            options,
            option_rpos,
            domains,
            findings,
        )
    finally:
        workbook.close()

    rows_by_role = {
        "interiors": interior_rows,
        "interior_scope": scope_rows,
        "interior_components": component_rows,
        "color_overrides": color_rows,
        "option_assets": option_asset_rows,
        "context_choice_assets": context_asset_rows,
        "default_selection_rules": default_rows,
        "runtime_rule_exceptions": exception_rows,
    }
    mappings_by_role = {
        "interiors": interior_mappings,
        "interior_scope": scope_mappings,
        "interior_components": component_mappings,
        "color_overrides": color_mappings,
        "option_assets": option_asset_mappings,
        "context_choice_assets": context_asset_mappings,
        "default_selection_rules": default_mappings,
        "runtime_rule_exceptions": exception_mappings,
    }

    shared_tables: dict[tuple[str, str], CompiledTable] = {}
    for model in LIVE_MODELS:
        for role in _SHARED_ROLES:
            shared_tables[(model, role)] = _unique_table(
                model,
                role,
                rows_by_role[role][model],
                mappings_by_role[role][model],
                findings,
            )

    completed = tuple(
        direct_by_key[(model, role)]
        if role in MODEL_TABLE_ROLES[:9]
        else shared_tables[(model, role)]
        for model in LIVE_MODELS
        for role in MODEL_TABLE_ROLES
    )
    mappings = tuple(
        mapping
        for model in LIVE_MODELS
        for role in _SHARED_ROLES
        for mapping in shared_tables[(model, role)].schema_mappings
    )
    return SharedCompilation(
        tables=completed,
        mappings=mappings,
        findings=tuple(findings),
    )


__all__ = ["SharedCompilation", "compile_shared_model_tables"]
