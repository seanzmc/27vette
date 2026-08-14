"""Isolated primary-runtime-only acceptance helpers for Workbook Manager Pass 4."""

from __future__ import annotations

import json
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from corvette_form_generator.model_configs import discover_generation_model_configs
from corvette_form_generator.output import write_json_output
from corvette_form_generator.registry_promotion import runtime_contract_artifact_path
from corvette_form_generator.runtime_contract import assert_runtime_contract
from corvette_form_generator.source_assembly import assemble_model_source
from corvette_form_generator.workbook import clean, rows_from_sheet, workbook_truthy


_IGNORED_GENERATED_KEYS = {"generated_at", "sourceGeneratedAt", "generatedAt"}


def _normalized_contract(value):
    if isinstance(value, list):
        return [_normalized_contract(item) for item in value]
    if isinstance(value, dict):
        return {
            key: _normalized_contract(value[key])
            for key in sorted(value)
            if key not in _IGNORED_GENERATED_KEYS
        }
    return value


def promoted_runtime_models(workbook_path: Path, repo_root: Path) -> tuple[str, ...]:
    """Return the complete strict runtime-contract promotion set."""
    workbook_path = Path(workbook_path)
    repo_root = Path(repo_root)
    configs = discover_generation_model_configs(workbook_path, root=repo_root)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "model_registry_promotion" not in workbook.sheetnames:
            raise ValueError("model_registry_promotion is missing")
        rows = [
            row
            for row in rows_from_sheet(workbook, "model_registry_promotion")
            if workbook_truthy(row.get("active"))
            and workbook_truthy(row.get("promoted_to_runtime"))
        ]
    finally:
        workbook.close()
    if not rows:
        raise ValueError("model_registry_promotion has no active promoted rows")

    models: list[str] = []
    for row in rows:
        model_key = clean(row.get("model_key")).lower()
        artifact_type = clean(row.get("artifact_type"))
        artifact_path = clean(row.get("artifact_path"))
        if artifact_type != "runtime_contract":
            raise ValueError(
                f"promoted {model_key!r} uses artifact_type {artifact_type!r}; "
                "Pass 4 accepts only 'runtime_contract'"
            )
        if model_key not in configs:
            raise ValueError(f"promoted model {model_key!r} is not generatable")
        expected = runtime_contract_artifact_path(repo_root, model_key).resolve()
        actual = (repo_root / artifact_path).resolve()
        if actual != expected:
            raise ValueError(
                f"promoted {model_key!r} artifact path {artifact_path!r} does not "
                f"resolve to {expected}"
            )
        models.append(model_key)
    if len(models) != len(set(models)):
        raise ValueError("model_registry_promotion contains duplicate promoted model rows")
    return tuple(models)


def generate_contract_snapshot(
    workbook_path: Path,
    output_root: Path,
    model_key: str,
) -> Path:
    """Generate exactly one canonical runtime contract below ``output_root``."""
    workbook_path = Path(workbook_path)
    output_root = Path(output_root)
    configs = discover_generation_model_configs(workbook_path)
    if model_key not in configs:
        raise ValueError(f"model {model_key!r} is not generatable from {workbook_path}")
    config = configs[model_key].with_overrides(
        root=output_root,
        workbook_path=workbook_path,
        output_dir=output_root / "form-output",
        app_dir=output_root / "form-app",
    )
    assembly = assemble_model_source(config)
    contract = assembly.runtime_contract
    assert_runtime_contract(
        contract,
        source=f"Workbook Manager Pass 4 acceptance for {model_key}",
        config=config,
        expected_model_label=config.model_label,
    )
    path = runtime_contract_artifact_path(output_root, model_key)
    write_json_output(path, contract)
    return path


def validate_primary_runtime_parity(
    source_workbook: Path,
    reconstructed_workbook: Path,
    repo_root: Path,
) -> list[dict]:
    """Return blocking findings for source/reconstruction runtime drift."""
    issues: list[dict] = []
    with tempfile.TemporaryDirectory(prefix="wbm-contract-parity-") as tempdir:
        root = Path(tempdir)
        source_input = root / "source-input" / Path(source_workbook).name
        reconstructed_input = root / "reconstruction-input" / Path(source_workbook).name
        source_input.parent.mkdir()
        reconstructed_input.parent.mkdir()
        shutil.copy2(source_workbook, source_input)
        shutil.copy2(reconstructed_workbook, reconstructed_input)
        source_models = promoted_runtime_models(source_input, repo_root)
        reconstructed_models = promoted_runtime_models(reconstructed_input, repo_root)
        if reconstructed_models != source_models:
            return [{
                "severity": "error",
                "category": "promoted_model_set_drift",
                "message": (
                    f"reconstructed promoted models {reconstructed_models!r} do not "
                    f"match source models {source_models!r}"
                ),
            }]
        for model_key in source_models:
            source_contract = generate_contract_snapshot(
                source_input,
                root / "source" / model_key,
                model_key,
            )
            reconstructed_contract = generate_contract_snapshot(
                reconstructed_input,
                root / "reconstruction" / model_key,
                model_key,
            )
            source_payload = _normalized_contract(
                json.loads(source_contract.read_text(encoding="utf-8"))
            )
            reconstructed_payload = _normalized_contract(
                json.loads(reconstructed_contract.read_text(encoding="utf-8"))
            )
            if reconstructed_payload != source_payload:
                issues.append({
                    "severity": "error",
                    "category": "generated_contract_drift",
                    "model_id": model_key,
                    "message": (
                        f"reconstructed {model_key} runtime contract differs "
                        "from the source contract"
                    ),
                })
    return issues
