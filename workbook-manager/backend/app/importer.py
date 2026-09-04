"""Workbook -> SQLite import adapter (openpyxl lives only behind this API).

Import is tolerant: rows with duplicate identifiers or
unresolved references are still ingested (first occurrence wins for
duplicates) and every problem is recorded in ``import_issues`` so nothing
fails silently. Sheets without a normalized table remain workbook-owned and
are classified without copying their cells into a second SQLite store.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sqlite3
import tempfile
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from corvette_form_generator.schema_validation import REQUIRED_SHEETS
from corvette_form_generator.schema_validation import validate_workbook_schema
from corvette_form_generator import editor_ops
from corvette_form_generator.workbook import clean, workbook_truthy
from corvette_form_generator.workbook_package import validate_workbook_package

from . import db as dbmod
from .catalog import (
    RAW_PRESERVED_SHEETS,
    SPEC_BY_TABLE,
    TABLE_SPECS,
    TableSpec,
    RequiredValueError,
    classify_workbook_sheets,
    projection_value,
    reconcile_columns,
)
from .validation import check_references
from corvette_form_generator.workbook_domain.registry import WILDCARD_MODEL_FAMILIES


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sheet_rows(ws) -> tuple[list[str], list[tuple[int, dict]]]:
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        return [], []
    headers = [str(h) if h is not None else "" for h in header]
    rows: list[tuple[int, dict]] = []
    for idx, values in enumerate(it, start=2):
        if values is None:
            continue
        record = {h: _cell_text(v) for h, v in zip(headers, values) if h}
        if all(v == "" for v in record.values()):
            continue
        rows.append((idx, record))
    return headers, rows


def _key_values(spec: TableSpec, rec: dict) -> tuple[str, ...]:
    """Key values from a header-keyed sheet record.

    ``spec.key`` holds SQL names; the record is keyed by literal headers.
    Every pre-2D family had headers equal to their SQL names, so the two were
    interchangeable until PriceRef (``OptionType`` -> ``optiontype``).
    """
    values = []
    for key in spec.key:
        column = spec.column_by_name(key)
        header = column.header if column is not None else key
        values.append(str(rec.get(header, "") or "").strip())
    return tuple(values)


def _active_model_keys(wb) -> set[str]:
    if "model_master" not in wb.sheetnames:
        return set()
    _, rows = _sheet_rows(wb["model_master"])
    return {
        str(r.get("model_key") or "").strip()
        for _, r in rows
        if _truthy(r.get("active", "")) and str(r.get("model_key") or "").strip()
    }


def _registry_rows(wb) -> list[dict]:
    _, rows = _sheet_rows(wb["model_workbook_sources"])
    return [r for _, r in rows if _truthy(r.get("active", ""))]


def _truthy(value: str) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y", "x"}


def workbook_identity(path: Path) -> dict[str, int | str]:
    path = Path(path)
    return {
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "mtime_ns": path.stat().st_mtime_ns,
    }


class Importer:
    def __init__(self, conn: sqlite3.Connection, workbook_path: Path):
        self.conn = conn
        self.workbook_path = Path(workbook_path)
        self.issues: list[dict] = []
        self.row_counts: dict[str, int] = {}
        self.active_models: list[str] = []

    # ── issue reporting ───────────────────────────────────────────────
    def issue(self, severity, category, *, sheet="", src_row=None,
              table="", model="", key="", field="", message=""):
        self.issues.append({
            "severity": severity, "category": category, "sheet": sheet,
            "src_row": src_row, "table_name": table, "model_id": model,
            "entity_key": key, "field": field, "message": message,
        })

    def row_disposition(
        self,
        *,
        sheet: str,
        src_row: int,
        spec: TableSpec,
        physical_key: tuple[str, ...],
        model_context: list[str],
        disposition: str,
        blocking: bool = False,
        field: str = "",
        value: str = "",
        reason_token: str = "",
        contract_impact: str = "",
    ) -> None:
        self.conn.execute(
            "INSERT INTO managed_row_dispositions(sheet, src_row, family, "
            "physical_key, model_context, disposition, blocking, field, value, "
            "reason_token, contract_impact) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (
                sheet,
                src_row,
                spec.editor_family or spec.family,
                json.dumps(list(physical_key), separators=(",", ":")),
                json.dumps(model_context, separators=(",", ":")),
                disposition,
                1 if blocking else 0,
                field,
                value,
                reason_token,
                contract_impact,
            ),
        )

    # ── main entry ────────────────────────────────────────────────────
    def run(self) -> dict:
        digest = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        mtime_ns = str(self.workbook_path.stat().st_mtime_ns)
        wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet_names = set(wb.sheetnames)

        dbmod.clear_imported_data(self.conn)
        self.conn.execute("DELETE FROM sheet_dispositions")
        self.conn.execute("DELETE FROM managed_row_dispositions")

        classifications = classify_workbook_sheets(wb)
        for name in wb.sheetnames:
            classification = classifications[name]
            opaque: tuple[str, ...] = ()
            if classification.spec is not None:
                headers, _rows = _sheet_rows(wb[name])
                opaque = reconcile_columns(classification.spec, headers).opaque
            self.conn.execute(
                "INSERT INTO sheet_dispositions(sheet, disposition, family, "
                "model_context, opaque_columns_json) VALUES(?,?,?,?,?)",
                (
                    name,
                    classification.disposition,
                    classification.family,
                    json.dumps(list(classification.models), separators=(",", ":")),
                    json.dumps(list(opaque), separators=(",", ":")),
                ),
            )

        registry = _registry_rows(wb)
        handled_sheets: set[str] = set()
        # D1 ownership doctrine (registry.models_for_write_targets): a row in a
        # global fixed family, or a wildcard model_key row, can change any
        # model's generated output, so it is owned by every active model.
        # Recorded at import so drafts, ChangeSets, and Apply/Rebuild derive a
        # non-empty affected set from stored ownership rather than UI state.
        self.active_models = sorted(_active_model_keys(wb))

        for spec in TABLE_SPECS:
            if spec.role:
                self._import_model_scoped(wb, spec, registry, handled_sheets,
                                          sheet_names)
            else:
                self._import_fixed(wb, spec, handled_sheets, sheet_names)

        # Preserved and unknown sheets remain workbook-owned. Their sheet-level
        # disposition was recorded above; do not read, copy, or count their rows.
        for name in wb.sheetnames:
            if name in handled_sheets:
                continue
            if name not in RAW_PRESERVED_SHEETS:
                self.issue("warning", "unmanaged_sheet", sheet=name,
                           message=f"sheet {name!r} has no normalized table; "
                                   "left workbook-owned without cell projection")

        expected_managed_rows = 0
        for name, classification in classifications.items():
            if classification.spec is None:
                continue
            _headers, physical_rows = _sheet_rows(wb[name])
            expected_managed_rows += len(physical_rows)
        actual_managed_rows = self.conn.execute(
            "SELECT COUNT(*) c FROM managed_row_dispositions"
        ).fetchone()["c"]
        if actual_managed_rows != expected_managed_rows:
            self.issue(
                "error",
                "incomplete_row_reconciliation",
                message=(
                    f"managed row reconciliation recorded {actual_managed_rows} of "
                    f"{expected_managed_rows} nonblank physical rows"
                ),
            )

        wb.close()

        ref_issues = check_references(self.conn)
        excluded_reference_rows: set[tuple[str, str, int]] = set()
        for iss in ref_issues:
            self.issues.append(iss)
            if iss.get("severity") != "error":
                continue
            table = str(iss.get("table_name") or "")
            sheet = str(iss.get("sheet") or "")
            src_row = int(iss.get("src_row") or 0)
            identity = (table, sheet, src_row)
            if table not in SPEC_BY_TABLE or not sheet or not src_row or \
                    identity in excluded_reference_rows:
                continue
            excluded_reference_rows.add(identity)
            family = SPEC_BY_TABLE[table].editor_family or SPEC_BY_TABLE[table].family
            prior = self.conn.execute(
                "SELECT physical_key, model_context FROM managed_row_dispositions "
                "WHERE sheet=? AND src_row=? AND family=?",
                (sheet, src_row, family),
            ).fetchone()
            field = str(iss.get("field") or "")
            offending_value = ""
            if SPEC_BY_TABLE[table].column_by_name(field) is not None:
                offending = self.conn.execute(
                    f'SELECT "{field}" FROM "{table}" '
                    "WHERE src_sheet=? AND src_row=? LIMIT 1",
                    (sheet, src_row),
                ).fetchone()
                if offending is not None:
                    offending_value = clean(offending[field])
            self.conn.execute(
                "DELETE FROM managed_row_dispositions "
                "WHERE sheet=? AND src_row=? AND family=?",
                (sheet, src_row, family),
            )
            self.conn.execute(
                f'DELETE FROM "{table}" WHERE src_sheet=? AND src_row=?',
                (sheet, src_row),
            )
            self.row_disposition(
                sheet=sheet,
                src_row=src_row,
                spec=SPEC_BY_TABLE[table],
                physical_key=tuple(
                    json.loads(prior["physical_key"]) if prior else
                    [str(iss.get("entity_key") or "")]
                ),
                model_context=list(
                    json.loads(prior["model_context"]) if prior else
                    [str(iss.get("model_id") or "")]
                ),
                disposition="excluded",
                blocking=True,
                field=field,
                value=offending_value,
                reason_token=str(iss.get("category") or "unresolved_reference"),
                contract_impact=str(iss.get("message") or "reference validation failed"),
            )
            self.row_counts[table] = max(0, self.row_counts.get(table, 0) - 1)

        issue_counts = {"error": 0, "warning": 0}
        for iss in self.issues:
            issue_counts[iss["severity"]] = issue_counts.get(iss["severity"], 0) + 1

        cur = self.conn.execute(
            "INSERT INTO import_runs(ts, workbook_path, workbook_mtime_ns, "
            "workbook_sha256, status, row_counts_json, issue_counts_json) "
            "VALUES(?,?,?,?,?,?,?)",
            (_now(), str(self.workbook_path), mtime_ns, digest,
             "imported_with_issues" if issue_counts["error"] else "imported",
             json.dumps(self.row_counts), json.dumps(issue_counts)),
        )
        run_id = cur.lastrowid
        self.conn.executemany(
            "INSERT INTO import_issues(run_id, severity, category, sheet, "
            "src_row, table_name, model_id, entity_key, field, message) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)",
            [(run_id, i["severity"], i["category"], i["sheet"], i["src_row"],
              i["table_name"], i["model_id"], i["entity_key"], i["field"],
              i["message"]) for i in self.issues],
        )
        dbmod.set_meta(self.conn, "workbook_mtime_ns", mtime_ns)
        dbmod.set_meta(self.conn, "workbook_sha256", digest)
        dbmod.set_meta(self.conn, "last_import_run_id", str(run_id))
        self.conn.commit()
        return self.report(run_id)

    def report(self, run_id: int) -> dict:
        run = self.conn.execute(
            "SELECT * FROM import_runs WHERE id=?", (run_id,)).fetchone()
        issues = self.conn.execute(
            "SELECT * FROM import_issues WHERE run_id=? ORDER BY id",
            (run_id,)).fetchall()
        sheet_dispositions = self.conn.execute(
            "SELECT * FROM sheet_dispositions ORDER BY sheet"
        ).fetchall()
        row_dispositions = self.conn.execute(
            "SELECT * FROM managed_row_dispositions ORDER BY sheet, src_row, family"
        ).fetchall()
        return {
            "run": dict(run),
            "issues": [dict(i) for i in issues],
            "sheet_dispositions": [dict(row) for row in sheet_dispositions],
            "managed_row_dispositions": [dict(row) for row in row_dispositions],
        }

    # ── sheet import paths ────────────────────────────────────────────
    def _import_fixed(self, wb, spec: TableSpec, handled: set, names: set):
        total = 0
        for sheet in spec.sheet:
            if sheet not in names:
                if sheet in REQUIRED_SHEETS:
                    self.issue("error", "missing_sheet", sheet=sheet,
                               table=spec.table,
                               message=f"expected sheet {sheet!r} not found")
                continue
            handled.add(sheet)
            total += self._ingest_sheet(
                wb, sheet, spec, model_id=None,
                model_context=None if spec.has_model_key_column else list(self.active_models),
            )
        self.row_counts[spec.table] = total

    def _import_model_scoped(self, wb, spec: TableSpec, registry, handled,
                             names):
        total = 0
        seen_sheets: set[tuple[str, ...]] = set()
        shared_models: dict[str, set[str]] = {}
        if not spec.model_scoped:
            for row in registry:
                if row.get("source_role") == spec.role and row.get("sheet_name"):
                    shared_models.setdefault(row["sheet_name"], set()).add(
                        row.get("model_key", "")
                    )
        for row in registry:
            if row.get("source_role") != spec.role:
                continue
            model = row.get("model_key", "")
            sheet = row.get("sheet_name", "")
            if not model or not sheet:
                continue
            sheet_identity = (model, sheet) if spec.model_scoped else (sheet,)
            if sheet_identity in seen_sheets:
                continue
            seen_sheets.add(sheet_identity)
            if sheet not in names:
                self.issue("error", "missing_sheet", sheet=sheet, model=model,
                           table=spec.table,
                           message=f"registered sheet {sheet!r} for model "
                                   f"{model!r} not found in workbook")
                continue
            # Shared sheets (color_overrides, interiors) are fixed-table specs;
            # a model-scoped spec re-registering a shared sheet is skipped when
            # another table already owns that sheet.
            handled.add(sheet)
            total += self._ingest_sheet(
                wb,
                sheet,
                spec,
                model_id=model if spec.model_scoped else None,
                model_context=sorted(shared_models.get(sheet, ())),
            )
        self.row_counts[spec.table] = total


    def _ingest_sheet(
        self,
        wb,
        sheet: str,
        spec: TableSpec,
        model_id: str | None,
        model_context: list[str] | None = None,
    ) -> int:
        headers, rows = _sheet_rows(wb[sheet])
        reconciliation = reconcile_columns(spec, headers)
        if reconciliation.duplicate:
            self.issue(
                "error",
                "duplicate_headers",
                sheet=sheet,
                table=spec.table,
                model=model_id or "",
                message=f"duplicate managed headers {list(reconciliation.duplicate)}",
            )
        if reconciliation.missing_required:
            self.issue("error", "missing_columns", sheet=sheet,
                       table=spec.table, model=model_id or "",
                       message=f"missing required columns "
                               f"{list(reconciliation.missing_required)}")
        if reconciliation.opaque:
            self.issue("warning", "unmapped_columns", sheet=sheet,
                       table=spec.table, model=model_id or "",
                       message=f"opaque managed-sheet columns "
                               f"{list(reconciliation.opaque)} remain workbook-owned")

        structural_failures = tuple(reconciliation.duplicate) + tuple(
            reconciliation.missing_required
        )
        if structural_failures:
            reason = (
                "duplicate_headers" if reconciliation.duplicate
                else "missing_required_headers"
            )
            for idx, rec in rows:
                keyvals = _key_values(spec, rec)
                row_context = list(model_context or ([model_id] if model_id else []))
                if spec.has_model_key_column:
                    row_model = str(rec.get("model_key") or "").strip()
                    row_context = [row_model] if row_model else []
                self.row_disposition(
                    sheet=sheet,
                    src_row=idx,
                    spec=spec,
                    physical_key=keyvals,
                    model_context=row_context,
                    disposition="excluded",
                    blocking=True,
                    field=",".join(structural_failures),
                    reason_token=reason,
                    contract_impact="managed sheet structure is ambiguous or incomplete",
                )
            return 0

        sql_cols = ["src_sheet", "src_row", "src_family", "physical_key", "model_context"]
        if spec.model_scoped:
            sql_cols.append("model_id")
        sql_cols += [c.sql_name() for c in spec.columns]
        placeholders = ",".join("?" for _ in sql_cols)
        collist = ",".join(f'"{c}"' for c in sql_cols)
        insert = f"INSERT INTO {spec.table} ({collist}) VALUES ({placeholders})"

        seen_keys: set[tuple] = set()
        count = 0
        for idx, rec in rows:
            keyvals = _key_values(spec, rec)
            key_id = (model_id or "",) + keyvals if spec.model_scoped else keyvals
            key_label = "/".join(v for v in keyvals if v) or f"row{idx}"
            row_context = list(model_context or ([model_id] if model_id else []))
            if spec.has_model_key_column:
                row_model = str(rec.get("model_key") or "").strip()
                row_context = [row_model] if row_model else []
                if row_model == "*" and (spec.editor_family in WILDCARD_MODEL_FAMILIES):
                    row_context = list(self.active_models)
            blank_ok = spec.blank_key_columns()
            missing_keys = [key for key, value in zip(spec.key, keyvals)
                            if value == "" and key not in blank_ok]
            if missing_keys:
                self.issue("error", "missing_identifier", sheet=sheet,
                           src_row=idx, table=spec.table,
                           model=model_id or "", key=key_label,
                           field=",".join(missing_keys),
                           message=f"blank key column(s) in {sheet} row {idx}")
                self.row_disposition(
                    sheet=sheet,
                    src_row=idx,
                    spec=spec,
                    physical_key=keyvals,
                    model_context=row_context,
                    disposition="excluded",
                    blocking=True,
                    field=",".join(missing_keys),
                    reason_token="missing_identifier",
                    contract_impact="managed row cannot be identified or reconstructed",
                )
                continue
            if key_id in seen_keys:
                self.issue("error", "duplicate_id", sheet=sheet, src_row=idx,
                           table=spec.table, model=model_id or "",
                           key=key_label,
                           message=f"duplicate key {key_label!r} in {sheet} "
                                   f"row {idx}; first occurrence kept")
                self.row_disposition(
                    sheet=sheet,
                    src_row=idx,
                    spec=spec,
                    physical_key=keyvals,
                    model_context=row_context,
                    disposition="excluded",
                    blocking=True,
                    field=",".join(spec.key),
                    value=key_label,
                    reason_token="duplicate_identifier",
                    contract_impact="physical row identity is ambiguous",
                )
                continue
            seen_keys.add(key_id)
            values = [
                sheet,
                idx,
                spec.editor_family or spec.family,
                json.dumps(list(keyvals), separators=(",", ":")),
                json.dumps(row_context),
            ]
            if spec.model_scoped:
                values.append(model_id)
            projected_values = []
            missing_required: list[str] = []
            for column in spec.columns:
                try:
                    projected_values.append(
                        projection_value(column, rec.get(column.header, ""))
                    )
                except RequiredValueError as exc:
                    self.issue(
                        "error",
                        "missing_required_value",
                        sheet=sheet,
                        src_row=idx,
                        table=spec.table,
                        model=model_id or "",
                        key=key_label,
                        field=column.header,
                        message=str(exc),
                    )
                    missing_required.append(column.header)
                    projected_values.append(None)
            if missing_required:
                self.row_disposition(
                    sheet=sheet,
                    src_row=idx,
                    spec=spec,
                    physical_key=keyvals,
                    model_context=row_context,
                    disposition="excluded",
                    blocking=True,
                    field=",".join(missing_required),
                    reason_token="missing_required_value",
                    contract_impact="managed row cannot satisfy the shared registry contract",
                )
                continue
            values += projected_values
            try:
                self.conn.execute(insert, values)
                count += 1
                self.row_disposition(
                    sheet=sheet,
                    src_row=idx,
                    spec=spec,
                    physical_key=keyvals,
                    model_context=row_context,
                    disposition="imported",
                )
            except sqlite3.IntegrityError as exc:
                # Fixed-sheet tables with global keys (e.g. interiors across
                # two sheets) can still collide; report instead of failing.
                self.issue("error", "duplicate_id", sheet=sheet, src_row=idx,
                           table=spec.table, model=model_id or "",
                           key=key_label, message=str(exc))
                self.row_disposition(
                    sheet=sheet,
                    src_row=idx,
                    spec=spec,
                    physical_key=keyvals,
                    model_context=row_context,
                    disposition="excluded",
                    blocking=True,
                    field=",".join(spec.key),
                    value=key_label,
                    reason_token="sqlite_uniqueness_violation",
                    contract_impact="candidate uniqueness could not be proven",
                )
        return count


def import_workbook(conn: sqlite3.Connection, workbook_path: Path) -> dict:
    return Importer(conn, workbook_path).run()


def latest_report(conn: sqlite3.Connection) -> dict | None:
    row = conn.execute(
        "SELECT id FROM import_runs ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        return None
    imp = Importer(conn, Path("."))
    return imp.report(row["id"])


def _semantic_fingerprint(conn: sqlite3.Connection) -> str:
    payload: dict[str, list] = {}
    for spec in TABLE_SPECS:
        columns = [
            row["name"]
            for row in conn.execute(f'PRAGMA table_info("{spec.table}")')
            if row["name"] != "id"
        ]
        quoted = ", ".join(f'"{column}"' for column in columns)
        values = [
            tuple(row[column] for column in columns)
            for row in conn.execute(f'SELECT {quoted} FROM "{spec.table}"')
        ]
        payload[spec.table] = sorted(values, key=repr)
    for table in ("sheet_dispositions", "managed_row_dispositions"):
        columns = [
            row["name"]
            for row in conn.execute(f'PRAGMA table_info("{table}")')
            if row["name"] != "id"
        ]
        quoted = ", ".join(f'"{column}"' for column in columns)
        values = [
            tuple(row[column] for column in columns)
            for row in conn.execute(f'SELECT {quoted} FROM "{table}"')
        ]
        payload[table] = sorted(values, key=repr)
    encoded = json.dumps(payload, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _projection_overlay_operations(
    source_projection: sqlite3.Connection,
    candidate_projection: sqlite3.Connection,
) -> tuple[list[dict], list[dict]]:
    """Build registry-owned update operations for candidate/source drift."""
    operations: list[dict] = []
    issues: list[dict] = []
    for spec in TABLE_SPECS:
        if not spec.editable:
            continue
        source_rows = [dict(row) for row in source_projection.execute(
            f'SELECT * FROM "{spec.table}"'
        )]
        candidate_rows = [dict(row) for row in candidate_projection.execute(
            f'SELECT * FROM "{spec.table}"'
        )]

        def identity(row: dict) -> tuple:
            return (
                str(row.get("model_id") or ""),
                str(row.get("src_sheet") or ""),
                *(str(row.get(key) or "") for key in spec.key),
            )

        source_by_key = {identity(row): row for row in source_rows}
        candidate_by_key = {identity(row): row for row in candidate_rows}
        if source_by_key.keys() != candidate_by_key.keys():
            issues.append({
                "severity": "error",
                "category": "projection_row_set_drift",
                "table_name": spec.table,
                "message": (
                    f"candidate {spec.table} row identities cannot be overlaid "
                    "as registry-owned field updates"
                ),
            })
            continue

        for row_key in sorted(candidate_by_key, key=repr):
            source_row = source_by_key[row_key]
            candidate_row = candidate_by_key[row_key]
            changed_columns = [
                column
                for column in spec.columns
                if column.sql_name() not in spec.key
                and source_row[column.sql_name()] != candidate_row[column.sql_name()]
            ]
            changed = bool(changed_columns)
            if not changed:
                continue
            operations.append({
                "action": "update",
                "sheet": candidate_row["src_sheet"],
                "_src_row": int(candidate_row["src_row"]),
                "key": {
                    key: str(candidate_row.get(key) or "")
                    for key in spec.key
                },
                "row": {
                    column.header: str(candidate_row.get(column.sql_name()) or "")
                    for column in changed_columns
                },
            })
    return operations, issues


def _anchor_signature(anchor) -> tuple:
    if isinstance(anchor, str):
        return (anchor,)
    return tuple(
        getattr(getattr(anchor, point, None), field, None)
        for point in ("_from", "to")
        for field in ("col", "colOff", "row", "rowOff")
    )


def _workbook_preservation_snapshot(path: Path) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        worksheet_xml_shapes: dict[str, dict] = {}
        with ZipFile(path) as package:
            namespace = {
                "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }
            relationships = ET.fromstring(
                package.read("xl/_rels/workbook.xml.rels")
            )
            target_by_id = {
                relationship.get("Id"): relationship.get("Target")
                for relationship in relationships
            }
            workbook_xml = ET.fromstring(package.read("xl/workbook.xml"))
            for sheet_node in workbook_xml.findall(".//x:sheets/x:sheet", namespace):
                sheet_name = str(sheet_node.get("name"))
                relationship_id = sheet_node.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                target = str(target_by_id[relationship_id]).lstrip("/")
                if not target.startswith("xl/"):
                    target = f"xl/{target}"
                root = ET.fromstring(package.read(target))
                dimension = root.find("x:dimension", namespace)
                worksheet_xml_shapes[sheet_name] = {
                    "dimension_ref": (
                        dimension.get("ref") if dimension is not None else None
                    ),
                    "physical_rows": tuple(
                        row.get("r")
                        for row in root.findall(".//x:sheetData/x:row", namespace)
                    ),
                    "physical_cells": tuple(
                        cell.get("r")
                        for cell in root.findall(
                            ".//x:sheetData/x:row/x:c", namespace
                        )
                    ),
                }
        snapshot: dict = {}
        for sheet in workbook.worksheets:
            cells = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None and not cell.has_style and cell.comment is None \
                            and cell.hyperlink is None:
                        continue
                    cells[cell.coordinate] = (
                        cell.value,
                        cell.data_type,
                        tuple(cell._style) if cell.has_style else None,
                        cell.number_format,
                        cell.comment.text if cell.comment else None,
                        cell.hyperlink.target if cell.hyperlink else None,
                    )
            row_dimensions = {
                index: (
                    dimension.height,
                    dimension.hidden,
                    dimension.outlineLevel,
                    dimension.collapsed,
                    dimension.style_id,
                )
                for index, dimension in sheet.row_dimensions.items()
            }
            column_dimensions = {
                index: (
                    dimension.min,
                    dimension.max,
                    dimension.width,
                    dimension.hidden,
                    dimension.outlineLevel,
                    dimension.collapsed,
                    dimension.bestFit,
                    dimension.style_id,
                )
                for index, dimension in sheet.column_dimensions.items()
            }
            snapshot[sheet.title] = {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "cells": cells,
                "merged": tuple(sorted(str(item) for item in sheet.merged_cells.ranges)),
                "row_dimensions": row_dimensions,
                "column_dimensions": column_dimensions,
                "tables": tuple(sorted(
                    (table.name, table.ref) for table in sheet.tables.values()
                )),
                "freeze_panes": str(sheet.freeze_panes or ""),
                "auto_filter": str(sheet.auto_filter.ref or ""),
                "images": tuple(
                    _anchor_signature(image.anchor)
                    for image in getattr(sheet, "_images", [])
                ),
                "charts": tuple(
                    _anchor_signature(chart.anchor)
                    for chart in getattr(sheet, "_charts", [])
                ),
                **worksheet_xml_shapes[sheet.title],
            }
        return snapshot
    finally:
        workbook.close()


def _preservation_issues(
    before: dict,
    after: dict,
    operations: list[dict],
    workbook_path: Path,
) -> list[dict]:
    allowed_cells: set[tuple[str, str]] = set()
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        for operation in operations:
            sheet = workbook[operation["sheet"]]
            headers = {
                str(cell.value or ""): getattr(cell, "column_letter", "")
                for cell in sheet[1]
            }
            for header in operation["row"]:
                column = headers.get(header)
                if column:
                    allowed_cells.add(
                        (operation["sheet"], f"{column}{operation['_src_row']}")
                    )
    finally:
        workbook.close()

    issues: list[dict] = []
    if before.keys() != after.keys():
        issues.append({
            "severity": "error",
            "category": "reconstruction_sheet_set_drift",
            "message": "reconstruction changed the workbook sheet set",
        })
        return issues
    for sheet_name in before:
        before_sheet = before[sheet_name]
        after_sheet = after[sheet_name]
        for key in (
            "max_row",
            "max_column",
            "merged",
            "row_dimensions",
            "column_dimensions",
            "tables",
            "freeze_panes",
            "auto_filter",
            "images",
            "charts",
            "dimension_ref",
            "physical_rows",
        ):
            if before_sheet[key] != after_sheet[key]:
                issues.append({
                    "severity": "error",
                    "category": "reconstruction_physical_drift",
                    "sheet": sheet_name,
                    "field": key,
                    "message": f"reconstruction changed {sheet_name!r} {key}",
                })
        changed_physical_cells = (
            set(before_sheet["physical_cells"])
            ^ set(after_sheet["physical_cells"])
        )
        unowned_physical_cells = {
            coordinate for coordinate in changed_physical_cells
            if (sheet_name, coordinate) not in allowed_cells
        }
        if unowned_physical_cells:
            issues.append({
                "severity": "error",
                "category": "reconstruction_physical_drift",
                "sheet": sheet_name,
                "field": "physical_cells",
                "value": sorted(unowned_physical_cells),
                "message": (
                    f"reconstruction changed untouched physical cells on "
                    f"{sheet_name!r}"
                ),
            })
        coordinates = before_sheet["cells"].keys() | after_sheet["cells"].keys()
        for coordinate in coordinates:
            if (sheet_name, coordinate) in allowed_cells:
                continue
            if before_sheet["cells"].get(coordinate) != after_sheet["cells"].get(coordinate):
                issues.append({
                    "severity": "error",
                    "category": "reconstruction_cell_drift",
                    "sheet": sheet_name,
                    "field": coordinate,
                    "message": (
                        f"reconstruction changed untouched cell "
                        f"{sheet_name}!{coordinate}"
                    ),
                })
                if len(issues) >= 100:
                    return issues
    return issues


def _validate_reconstruction(
    source: Path,
    candidate: sqlite3.Connection,
    *,
    reconstruction_path: Path | None = None,
) -> tuple[bool, bool, bool, list[dict]]:
    """Copy, overlay through editor_ops, and verify semantic reconstruction."""
    with tempfile.TemporaryDirectory(prefix="wbm-reconstruction-") as tempdir:
        root = Path(tempdir)
        copied_workbook = reconstruction_path or (root / source.name)
        if reconstruction_path is None:
            shutil.copy2(source, copied_workbook)
        if copied_workbook.read_bytes() != source.read_bytes():
            return False, False, False, [{
                "severity": "error",
                "category": "reconstruction_copy_drift",
                "message": "identity reconstruction copy differs from its source",
            }]
        baseline_path = root / "baseline-projection.sqlite3"
        baseline = dbmod.connect(baseline_path)
        try:
            dbmod.init_projection_schema(baseline)
            baseline_report = import_workbook(baseline, copied_workbook)
            baseline_errors = [
                issue for issue in baseline_report["issues"]
                if issue["severity"] == "error"
            ]
            operations, overlay_issues = _projection_overlay_operations(
                baseline, candidate
            )
        finally:
            baseline.close()

        preservation_issues: list[dict] = []
        if operations and not baseline_errors and not overlay_issues:
            before_snapshot = _workbook_preservation_snapshot(source)
            identity = workbook_identity(copied_workbook)
            overlay_result = editor_ops.apply_batch(
                copied_workbook,
                {
                    "workbookMtimeNs": str(identity["mtime_ns"]),
                    "workbookSha256": identity["sha256"],
                    "items": [
                        {
                            key: value
                            for key, value in operation.items()
                            if not key.startswith("_")
                        }
                        for operation in operations
                    ],
                },
                write=True,
                source="workbook-manager-pass4-reconstruction",
                log_path=root / "reconstruction-edits.jsonl",
            )
            if not overlay_result.get("ok"):
                overlay_issues.append({
                    "severity": "error",
                    "category": "reconstruction_overlay_failed",
                    "message": "; ".join(
                        str(error) for error in overlay_result.get("errors", [])
                    ) or str(overlay_result.get("status") or "overlay failed"),
                })
            else:
                preservation_issues = _preservation_issues(
                    before_snapshot,
                    _workbook_preservation_snapshot(copied_workbook),
                    operations,
                    source,
                )

        package_issues = validate_workbook_package(copied_workbook)
        schema_issues = validate_workbook_schema(
            copied_workbook, check_live_contract=False
        )
        schema_errors = [
            issue for issue in schema_issues if getattr(issue, "severity", "error") == "error"
        ]
        verifier_path = root / "semantic-readback.sqlite3"
        verifier = dbmod.connect(verifier_path)
        try:
            dbmod.init_projection_schema(verifier)
            verifier_report = import_workbook(verifier, copied_workbook)
            verifier_errors = [
                issue for issue in verifier_report["issues"]
                if issue["severity"] == "error"
            ]
            semantic_equal = (
                not baseline_errors
                and not overlay_issues
                and not preservation_issues
                and not verifier_errors
                and _semantic_fingerprint(candidate) == _semantic_fingerprint(verifier)
            )
        finally:
            verifier.close()
        issues: list[dict] = [dict(issue) for issue in package_issues]
        issues.extend(
            {
                "severity": getattr(issue, "severity", "error"),
                "category": getattr(issue, "check_id", "schema_validation"),
                "message": getattr(issue, "message", str(issue)),
            }
            for issue in schema_errors
        )
        issues.extend(baseline_errors)
        issues.extend(overlay_issues)
        issues.extend(preservation_issues)
        issues.extend(verifier_errors)
        if not semantic_equal:
            issues.append({
                "severity": "error",
                "category": "semantic_readback_drift",
                "message": "reconstructed workbook semantic readback differs from candidate",
            })
        return (
            not package_issues,
            not schema_errors,
            semantic_equal,
            issues,
        )


def promote_verified_projection(workbook_path: Path, projection_path: Path) -> dict:
    """Build, verify, and atomically promote a same-filesystem projection.

    The caller owns ``STATE_LOCK`` when invoked from the API. This helper never
    opens or mutates the durable manager-state database.
    """
    workbook_path = Path(workbook_path)
    projection_path = Path(projection_path)
    initial_identity = workbook_identity(workbook_path)
    projection_path.parent.mkdir(parents=True, exist_ok=True)
    candidate_path = projection_path.with_name(
        f".wbm-import-candidate-{uuid.uuid4().hex}.sqlite3"
    )
    candidate: sqlite3.Connection | None = None
    report: dict = {
        "status": "blocked",
        "promoted": False,
        "source_sha256": initial_identity["sha256"],
        "source_mtime_ns": initial_identity["mtime_ns"],
        "blocking_findings": 1,
        "package_valid": False,
        "schema_valid": False,
        "semantic_readback_verified": False,
        "issues": [],
    }
    try:
        candidate = dbmod.connect(candidate_path)
        dbmod.init_projection_schema(candidate)
        imported = import_workbook(candidate, workbook_path)
        report["import"] = imported
        blocking = [
            issue for issue in imported["issues"] if issue["severity"] == "error"
        ]
        integrity = candidate.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            blocking.append({
                "severity": "error",
                "category": "projection_integrity",
                "message": str(integrity),
            })
        if blocking:
            report["issues"] = blocking
            report["blocking_findings"] = len(blocking)
            return report
        (
            package_valid,
            schema_valid,
            semantic_equal,
            validation_issues,
        ) = (
            _validate_reconstruction(workbook_path, candidate)
        )
        report.update({
            "package_valid": package_valid,
            "schema_valid": schema_valid,
            "semantic_readback_verified": semantic_equal,
        })
        blocking.extend(
            issue for issue in validation_issues
            if issue.get("severity", "error") == "error"
        )
        report["issues"] = blocking
        report["blocking_findings"] = len(blocking)
        if blocking:
            return report

        final_identity = workbook_identity(workbook_path)
        if final_identity != initial_identity:
            report["status"] = "source_changed"
            report["blocking_findings"] = 1
            report["issues"] = [{
                "severity": "error",
                "category": "source_identity_drift",
                "message": "source workbook SHA-256 or mtime changed during candidate build",
            }]
            return report

        counts = {
            spec.table: candidate.execute(
                f'SELECT COUNT(*) c FROM "{spec.table}"'
            ).fetchone()["c"]
            for spec in TABLE_SPECS
        }
        migration_id = f"import-{uuid.uuid4().hex}"
        dbmod._write_manifest(
            candidate,
            store_kind="projection",
            migration_id=migration_id,
            source_sha256=str(initial_identity["sha256"]),
            source_path=workbook_path,
            archive_path=None,
            table_counts=counts,
        )
        candidate.commit()
        checkpoint = candidate.execute("PRAGMA wal_checkpoint(TRUNCATE)").fetchone()
        if checkpoint[0] != 0 or checkpoint[1] != checkpoint[2]:
            raise RuntimeError(
                "candidate projection WAL could not be fully checkpointed"
            )
        candidate.close()
        candidate = None
        for sidecar in (
            Path(f"{candidate_path}-wal"),
            Path(f"{candidate_path}-shm"),
        ):
            sidecar.unlink(missing_ok=True)
        dbmod._fsync_dir(candidate_path.parent)
        dbmod._fsync_file(candidate_path)

        def verify_promoted(path: Path) -> None:
            reopened = dbmod.connect(path)
            try:
                if reopened.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
                    raise RuntimeError("promoted projection failed integrity check")
                manifest = dbmod.storage_manifest(reopened) or {}
                expected = {
                    "store_kind": "projection",
                    "migration_id": migration_id,
                    "source_sha256": str(initial_identity["sha256"]),
                    "schema_version": dbmod.SCHEMA_VERSION,
                }
                actual = {key: manifest.get(key) for key in expected}
                if actual != expected:
                    raise RuntimeError(
                        f"promoted projection manifest mismatch: {actual!r} != {expected!r}"
                    )
                reopened_report = latest_report(reopened)
                if not reopened_report or reopened_report["run"]["status"] != "imported":
                    raise RuntimeError("promoted projection is missing its verified import run")
            finally:
                reopened.close()

        dbmod._replace_projection(
            candidate_path,
            projection_path,
            verify=verify_promoted,
        )
        report.update({
            "status": "promoted",
            "promoted": True,
            "blocking_findings": 0,
            "issues": [],
        })
        return report
    except dbmod.ProjectionBusyError:
        raise
    except Exception as exc:
        report["issues"] = [{
            "severity": "error",
            "category": "candidate_build_exception",
            "message": f"{type(exc).__name__}: {exc}",
        }]
        report["blocking_findings"] = 1
        return report
    finally:
        if candidate is not None:
            candidate.close()
        candidate_path.unlink(missing_ok=True)
        Path(f"{candidate_path}-wal").unlink(missing_ok=True)
        Path(f"{candidate_path}-shm").unlink(missing_ok=True)
