"""Read-only workbook profiling and explicit source-sheet disposition."""

from __future__ import annotations

import hashlib
from collections import defaultdict
from pathlib import Path
from types import MappingProxyType
from typing import Iterable

from openpyxl import load_workbook

from .catalog import LIVE_MODELS, MODEL_TABLE_ROLES, physical_table
from .compile_types import Finding, SourceSheet, WorkbookProfile


_CONTROL_SHEETS = (
    "model_master",
    "model_registry_promotion",
    "model_workbook_sources",
)

# Fixed workbook metadata sources are explicit. Model-specific source sheet
# names are intentionally absent: those come only from model_workbook_sources.
_CENTRAL_DESTINATIONS: dict[str, tuple[str, ...]] = {
    "model_master": ("models",),
    "model_registry_promotion": ("model_registry_promotion",),
    "model_workbook_sources": ("model_table_registry",),
    "model_variants": ("model_variants",),
    "variant_master": ("body_styles", "trim_levels", "variants"),
    "section_master": ("sections",),
    "context_section_master": ("runtime_context_sections",),
    "section_presentation": ("section_presentation",),
    "runtime_steps": ("runtime_route_keys", "runtime_steps"),
    "context_choice_copy": ("runtime_context_choices",),
    "order_summary_sections": ("runtime_summary_sections",),
    "step_order_summary_map": (
        "runtime_route_keys",
        "runtime_step_summary_map",
    ),
    "default_selection_rules": tuple(
        physical_table(model, "default_selection_rules") for model in LIVE_MODELS
    ),
    "runtime_rule_exceptions": tuple(
        physical_table(model, "runtime_rule_exceptions") for model in LIVE_MODELS
    ),
    "asset_map": (
        "model_assets",
        *(
            physical_table(model, role)
            for model in LIVE_MODELS
            for role in ("option_assets", "context_choice_assets")
        ),
    ),
    "PriceRef": ("price_ref",),
    "model_interior_scope": tuple(
        physical_table(model, "interior_scope") for model in LIVE_MODELS
    ),
    "interior_components": tuple(
        physical_table(model, "interior_components") for model in LIVE_MODELS
    ),
    "rule_phrase_map": ("rule_phrase_map",),
}

_SOURCE_ROLE_TO_TABLE_ROLE = {
    "source_option_sheet": "options",
    "status_sheet": "option_availability",
    "rule_mapping_sheet": "rule_mapping",
    "price_rules_sheet": "price_rules",
    "rule_groups_sheet": "rule_groups",
    "rule_group_members_sheet": "rule_group_members",
    "exclusive_groups_sheet": "exclusive_groups",
    "exclusive_group_members_sheet": "exclusive_group_members",
    "variant_option_overrides_sheet": "variant_overrides",
    "color_overrides_sheet": "color_overrides",
    "interior_source_sheet": "interiors",
}

if not set(_SOURCE_ROLE_TO_TABLE_ROLE.values()) <= set(MODEL_TABLE_ROLES):
    raise RuntimeError("workbook source roles reference an unknown canonical role")


def _truthy(value: object) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y"}
    return value is True or value == 1


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as workbook_file:
        for chunk in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _rows(sheet) -> tuple[tuple[str, ...], list[tuple[int, dict[str, object]]]]:
    iterator = sheet.iter_rows(values_only=True)
    raw_headers = next(iterator, ())
    seen_headers: set[str] = set()
    duplicate_headers: list[str] = []
    for header in raw_headers:
        if not isinstance(header, str) or not header:
            continue
        if header in seen_headers and header not in duplicate_headers:
            duplicate_headers.append(header)
        seen_headers.add(header)
    if duplicate_headers:
        duplicates = ", ".join(repr(header) for header in duplicate_headers)
        noun = "header" if len(duplicate_headers) == 1 else "headers"
        raise ValueError(f"Sheet {sheet.title!r} has duplicate {noun} {duplicates}")
    if not raw_headers or any(not isinstance(header, str) or not header for header in raw_headers):
        raise ValueError(f"Sheet {sheet.title!r} has missing or non-text headers")
    headers = tuple(raw_headers)
    records = [
        (row_number, dict(zip(headers, row)))
        for row_number, row in enumerate(iterator, start=2)
        if any(value is not None for value in row)
    ]
    return headers, records


def _require_columns(
    source_sheet: str,
    headers: Iterable[str],
    required: set[str],
) -> None:
    missing = required - set(headers)
    if missing:
        raise ValueError(
            f"Sheet {source_sheet!r} is missing required columns {sorted(missing)!r}"
        )


def _discover_active_models(model_rows, promotion_rows) -> tuple[str, ...]:
    active_master = {
        str(row["model_key"])
        for _, row in model_rows
        if row.get("model_key") and _truthy(row.get("active"))
    }
    promoted = {
        str(row["model_key"])
        for _, row in promotion_rows
        if row.get("model_key")
        and _truthy(row.get("active"))
        and _truthy(row.get("promoted_to_runtime"))
    }
    return tuple(
        str(row["model_key"])
        for _, row in model_rows
        if row.get("model_key") in active_master & promoted
    )


def _active_source_registry(
    source_rows,
    active_models: tuple[str, ...],
) -> dict[str, dict[str, str]]:
    sources = {model: {} for model in active_models}
    for source_row, row in source_rows:
        model_key = str(row.get("model_key") or "")
        if model_key not in sources or not _truthy(row.get("active")):
            continue
        source_role = str(row.get("source_role") or "")
        sheet_name = str(row.get("sheet_name") or "")
        if not source_role or not sheet_name:
            raise ValueError(
                f"model_workbook_sources row {source_row} has an empty active source binding"
            )
        if source_role in sources[model_key]:
            raise ValueError(
                "model_workbook_sources has duplicate active binding for "
                f"{model_key!r} role {source_role!r} at row {source_row}"
            )
        sources[model_key][source_role] = sheet_name
    return sources


def profile_workbook(path: Path) -> WorkbookProfile:
    """Profile a workbook without mutating it or inferring unknown ownership."""
    workbook_path = Path(path)
    workbook_hash = _sha256(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    findings: list[Finding] = []
    try:
        missing_control = set(_CONTROL_SHEETS) - set(workbook.sheetnames)
        if missing_control:
            raise ValueError(f"Workbook is missing control sheets {sorted(missing_control)!r}")

        sheet_data = {name: _rows(workbook[name]) for name in workbook.sheetnames}
        model_headers, model_rows = sheet_data["model_master"]
        promotion_headers, promotion_rows = sheet_data["model_registry_promotion"]
        source_headers, source_rows = sheet_data["model_workbook_sources"]
        _require_columns("model_master", model_headers, {"model_key", "active"})
        _require_columns(
            "model_registry_promotion",
            promotion_headers,
            {"model_key", "promoted_to_runtime", "active"},
        )
        _require_columns(
            "model_workbook_sources",
            source_headers,
            {"model_key", "source_role", "sheet_name", "active"},
        )

        active_models = _discover_active_models(model_rows, promotion_rows)
        if active_models != LIVE_MODELS:
            findings.append(
                Finding(
                    severity="error",
                    status="contract_mismatch",
                    code="live_model_catalog_mismatch",
                    message="Workbook active/promoted models differ from the canonical catalog.",
                    source_sheet="model_master",
                    value={"expected": LIVE_MODELS, "actual": active_models},
                )
            )

        active_sources = _active_source_registry(source_rows, active_models)
        expected_roles = frozenset(_SOURCE_ROLE_TO_TABLE_ROLE)
        for model_key, sources in active_sources.items():
            if frozenset(sources) != expected_roles:
                findings.append(
                    Finding(
                        severity="error",
                        status="contract_mismatch",
                        code="source_role_contract_mismatch",
                        message="Live model source roles differ from the canonical source-role contract.",
                        source_sheet="model_workbook_sources",
                        model_key=model_key,
                        value={
                            "expected": tuple(sorted(expected_roles)),
                            "actual": tuple(sorted(sources)),
                        },
                    )
                )

        destinations_by_source: dict[str, set[str]] = defaultdict(set)
        unclassified_registered: set[str] = set()
        ownership_conflicts: set[str] = set()
        for source_row, row in source_rows:
            source_role = str(row.get("source_role") or "")
            sheet_name = str(row.get("sheet_name") or "")
            if sheet_name in _CENTRAL_DESTINATIONS:
                ownership_conflicts.add(sheet_name)
                findings.append(
                    Finding(
                        severity="error",
                        status="decision_required",
                        code="source_ownership_conflict",
                        message=(
                            "Workbook source binding conflicts with explicit central "
                            "sheet ownership; correct the registry binding before import."
                        ),
                        source_sheet=sheet_name,
                        source_row=source_row,
                        source_column="sheet_name",
                        model_key=str(row.get("model_key") or ""),
                        value={
                            "source_role": source_role,
                            "central_destinations": _CENTRAL_DESTINATIONS[sheet_name],
                        },
                    )
                )
            if sheet_name and source_role not in _SOURCE_ROLE_TO_TABLE_ROLE:
                unclassified_registered.add(sheet_name)
                findings.append(
                    Finding(
                        severity="error",
                        status="decision_required",
                        code="source_role_unclassified",
                        message="Workbook source role has no canonical table-role mapping.",
                        source_sheet="model_workbook_sources",
                        source_row=source_row,
                        source_column="source_role",
                        model_key=str(row.get("model_key") or ""),
                        value=source_role,
                    )
                )
        for model_key, sources in active_sources.items():
            for source_role, sheet_name in sources.items():
                table_role = _SOURCE_ROLE_TO_TABLE_ROLE.get(source_role)
                if table_role is None:
                    unclassified_registered.add(sheet_name)
                    continue
                destinations_by_source[sheet_name].add(
                    physical_table(model_key, table_role)
                )

        inactive_source_sheets = {
            str(row.get("sheet_name") or "")
            for _, row in source_rows
            if row.get("sheet_name")
            and (
                str(row.get("model_key") or "") not in active_models
                or not _truthy(row.get("active"))
            )
        }

        known_sheet_names = set(workbook.sheetnames)
        for model_key, sources in active_sources.items():
            for source_role, sheet_name in sources.items():
                if sheet_name not in known_sheet_names:
                    findings.append(
                        Finding(
                            severity="error",
                            status="contract_mismatch",
                            code="registered_source_sheet_missing",
                            message="Active workbook source binding names a missing sheet.",
                            source_sheet="model_workbook_sources",
                            source_column="sheet_name",
                            model_key=model_key,
                            value={"role": source_role, "sheet_name": sheet_name},
                        )
                    )

        profiled_sheets: list[SourceSheet] = []
        for sheet_name in workbook.sheetnames:
            headers, rows = sheet_data[sheet_name]
            if sheet_name.startswith("form_"):
                disposition = "generated_artifact_validation"
                destination_tables = ()
                reason = (
                    "Retired generated workbook surface; validate against runtime "
                    "artifacts and never import as source data."
                )
                findings.append(
                    Finding(
                        severity="error",
                        status="contract_mismatch",
                        code="retired_generated_sheet_present",
                        message="Retired generated workbook sheet is present.",
                        source_sheet=sheet_name,
                    )
                )
            elif sheet_name in ownership_conflicts:
                disposition = "decision_required"
                destination_tables = ()
                reason = "Central and registry source ownership conflict."
            elif sheet_name in _CENTRAL_DESTINATIONS:
                destination_tables = _CENTRAL_DESTINATIONS[sheet_name]
                disposition = (
                    "canonical_direct"
                    if len(destination_tables) == 1
                    else "canonical_split"
                )
                reason = "Explicit canonical metadata source mapping."
            elif sheet_name in unclassified_registered:
                disposition = "decision_required"
                destination_tables = ()
                reason = "Registered source role has no approved canonical mapping."
            elif sheet_name in destinations_by_source:
                destination_tables = tuple(sorted(destinations_by_source[sheet_name]))
                disposition = (
                    "canonical_direct"
                    if len(destination_tables) == 1
                    else "canonical_split"
                )
                reason = "Destination ownership derived from active model source bindings."
            elif sheet_name in inactive_source_sheets:
                disposition = "inactive_future_source"
                destination_tables = ()
                reason = "Registered only to inactive or unpromoted future models."
            else:
                disposition = "decision_required"
                destination_tables = ()
                reason = "No approved central, active, inactive, or generated-sheet policy applies."
                findings.append(
                    Finding(
                        severity="error",
                        status="decision_required",
                        code="source_sheet_unclassified",
                        message="Workbook sheet has no approved source ownership mapping.",
                        source_sheet=sheet_name,
                    )
                )
            profiled_sheets.append(
                SourceSheet(
                    source_sheet=sheet_name,
                    disposition=disposition,
                    headers=headers,
                    row_count=len(rows),
                    destination_tables=destination_tables,
                    reason=reason,
                )
            )

        frozen_sources = MappingProxyType(
            {
                model: MappingProxyType(dict(sources))
                for model, sources in active_sources.items()
            }
        )
        return WorkbookProfile(
            workbook_path=workbook_path,
            workbook_sha256=workbook_hash,
            sheets=tuple(profiled_sheets),
            active_models=active_models,
            active_sources=frozen_sources,
            findings=tuple(findings),
        )
    finally:
        workbook.close()
