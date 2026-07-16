"""Compile workbook-owned central metadata into canonical relational rows."""

from __future__ import annotations

import hashlib
from collections import defaultdict
from pathlib import Path
from types import MappingProxyType
from typing import Iterable, Mapping

from openpyxl import load_workbook

from .catalog import LIVE_MODELS
from .compile_types import (
    CompiledRow,
    CompiledTable,
    DecisionRequired,
    WorkbookProfile,
)


_SHEET_HEADERS: Mapping[str, tuple[str, ...]] = MappingProxyType(
    {
        "model_master": (
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "notes",
        ),
        "model_registry_promotion": (
            "model_key",
            "registry_key",
            "promoted_to_runtime",
            "default_model",
            "artifact_path",
            "artifact_type",
            "legacy_alias",
            "active",
            "display_order",
            "notes",
        ),
        "variant_master": (
            "variant_id",
            "model_year",
            "trim_level",
            "body_style",
            "display_name",
            "base_price",
            "display_order",
            "active",
        ),
        "model_variants": (
            "model_key",
            "variant_id",
            "display_order",
            "active",
            "notes",
        ),
        "section_master": (
            "section_id",
            "section_name",
            "selection_mode",
            "is_required",
            "display_order",
            "standard_behavior",
            "step_key",
        ),
        "section_presentation": (
            "model_key",
            "section_id",
            "display_label",
            "step_key",
            "display_behavior",
            "section_display_order",
            "standard_equipment_bucket",
            "standard_equipment_group_type",
            "auto_added_bucket",
            "active",
            "notes",
        ),
        "runtime_steps": (
            "model_key",
            "step_key",
            "step_label",
            "runtime_order",
            "source",
            "active",
            "notes",
        ),
        "context_section_master": (
            "model_key",
            "context_type",
            "section_id",
            "section_name",
            "selection_mode",
            "choice_mode",
            "is_required",
            "standard_behavior",
            "section_display_order",
            "step_key",
            "step_label",
            "active",
            "notes",
        ),
        "context_choice_copy": (
            "model_key",
            "context_type",
            "value",
            "body_style",
            "info_tooltip",
            "active",
            "notes",
        ),
        "order_summary_sections": (
            "model_key",
            "section_key",
            "section_label",
            "display_order",
            "active",
            "notes",
        ),
        "step_order_summary_map": (
            "model_key",
            "step_key",
            "section_key",
            "active",
            "notes",
        ),
        "asset_map": (
            "model_key",
            "target_type",
            "target_id",
            "image_url",
            "image_alt",
            "image_fit",
            "image_position",
            "hover_image_url",
            "hover_image_alt",
            "hover_image_position",
            "active",
            "notes",
        ),
        "PriceRef": ("OptionType", "Trim", "Code", "Price"),
        "rule_phrase_map": (
            "phrase",
            "rule_type",
            "direction",
            "stop_phrases",
            "review_flag_default",
            "active",
            "notes",
        ),
    }
)


_TABLE_PRIMARY_KEYS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("models", ("model_key",)),
    ("model_registry_promotion", ("model_key",)),
    ("body_styles", ("body_style",)),
    ("trim_levels", ("trim_level",)),
    ("variants", ("variant_id",)),
    ("model_variants", ("model_key", "variant_id")),
    ("sections", ("section_id",)),
    ("section_presentation", ("model_key", "section_id")),
    ("runtime_route_keys", ("model_key", "route_key")),
    ("runtime_steps", ("model_key", "step_key")),
    (
        "runtime_context_sections",
        ("model_key", "context_type", "section_id"),
    ),
    ("runtime_context_choices", ("model_key", "context_choice_id")),
    ("runtime_summary_sections", ("model_key", "section_key")),
    (
        "runtime_step_summary_map",
        ("model_key", "step_key"),
    ),
    ("model_assets", ("model_key",)),
    ("price_ref", ("option_type", "trim_level", "code")),
    ("rule_phrase_map", ("phrase",)),
)


def _decision(
    code: str,
    message: str,
    *,
    sheet: str = "",
    row: int | None = None,
    column: str = "",
    value: object = None,
) -> DecisionRequired:
    return DecisionRequired(
        code,
        message,
        source_sheet=sheet,
        source_row=row,
        source_column=column,
        value=value,
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as workbook_file:
        for chunk in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _required_text(
    value: object, *, sheet: str, row: int, column: str
) -> str:
    result = _text(value)
    if not result:
        raise _decision(
            "central_required_value_missing",
            "A required central value is blank.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    return result


def _lower(
    value: object, *, sheet: str, row: int, column: str
) -> tuple[str, dict[str, object]]:
    original = "" if value is None else str(value)
    trimmed = original.strip()
    if not trimmed:
        raise _decision(
            "central_required_value_missing",
            "A required central value is blank.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    normalized = trimmed.lower()
    transforms: list[str] = []
    if trimmed != original:
        transforms.append("trim")
    if normalized != trimmed:
        transforms.append("lowercase")
    parameters = (
        {"original": original, "transform": "_then_".join(transforms)}
        if transforms
        else {}
    )
    return normalized, parameters


def _integer(
    value: object,
    *,
    sheet: str,
    row: int,
    column: str,
    nullable: bool = False,
) -> int | None:
    if value is None or _text(value) == "":
        if nullable:
            return None
        raise _decision(
            "central_required_value_missing",
            "A required central integer is blank.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    try:
        number = int(value)
    except (TypeError, ValueError, OverflowError) as error:
        raise _decision(
            "central_integer_invalid",
            "A central integer field is not an integer.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        ) from error
    if isinstance(value, float) and value != number:
        raise _decision(
            "central_integer_invalid",
            "A central integer field has a fractional value.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )
    return number


def _boolean(value: object, *, sheet: str, row: int, column: str) -> bool:
    if value is True or value == 1:
        return True
    if value is False or value == 0:
        return False
    normalized = _text(value).lower()
    if normalized in {"true", "1", "yes", "y"}:
        return True
    if normalized in {"false", "0", "no", "n"}:
        return False
    raise _decision(
        "central_boolean_invalid",
        "A central boolean field has no exact boolean interpretation.",
        sheet=sheet,
        row=row,
        column=column,
        value=value,
    )


def _freeze_mapping(value: Mapping[str, object]) -> Mapping[str, object]:
    frozen: dict[str, object] = {}
    for key, item in value.items():
        frozen[key] = MappingProxyType(dict(item)) if isinstance(item, dict) else item
    return MappingProxyType(frozen)


def _compiled_row(
    values: Mapping[str, object],
    sheet: str,
    row: int,
    *,
    lineage_role: str = "direct",
    mapping_parameters: Mapping[str, object] | None = None,
) -> CompiledRow:
    return CompiledRow(
        values=MappingProxyType(dict(values)),
        source_sheet=sheet,
        source_row=row,
        lineage_role=lineage_role,
        mapping_parameters=_freeze_mapping(mapping_parameters or {}),
    )


def _derived_mapping_parameter(
    row: CompiledRow,
    source_column: str,
    derived_transform: str,
) -> Mapping[str, object]:
    source_parameter = row.mapping_parameters.get(source_column)
    if isinstance(source_parameter, Mapping):
        original = source_parameter.get("original", row.values[source_column])
        source_transform = str(source_parameter.get("transform") or "identity")
    else:
        original = row.values[source_column]
        source_transform = "identity"
    transform = (
        derived_transform
        if source_transform == "identity"
        else f"{source_transform}_then_{derived_transform}"
    )
    return MappingProxyType({"original": original, "transform": transform})


def _read_rows(workbook, sheet_name: str) -> list[tuple[int, dict[str, object]]]:
    if sheet_name not in workbook.sheetnames:
        raise _decision(
            "central_source_sheet_missing",
            "A required central source sheet is missing.",
            sheet=sheet_name,
        )
    iterator = workbook[sheet_name].iter_rows(values_only=True)
    headers = tuple(next(iterator, ()))
    expected = _SHEET_HEADERS[sheet_name]
    if headers != expected:
        raise _decision(
            "central_source_headers_mismatch",
            "Central source headers differ from the canonical mapping contract.",
            sheet=sheet_name,
            row=1,
            value={"expected": expected, "actual": headers},
        )
    return [
        (row_number, dict(zip(headers, values)))
        for row_number, values in enumerate(iterator, start=2)
        if any(value is not None for value in values)
    ]


def _active_rows(
    rows: Iterable[tuple[int, dict[str, object]]], sheet: str
) -> list[tuple[int, dict[str, object]]]:
    return [
        (row_number, row)
        for row_number, row in rows
        if _boolean(
            row.get("active"),
            sheet=sheet,
            row=row_number,
            column="active",
        )
    ]


def _ensure_reference(
    value: object,
    domain: set[object],
    *,
    code: str,
    sheet: str,
    row: int,
    column: str,
) -> None:
    if value not in domain:
        raise _decision(
            code,
            "A central relationship references no compiled canonical row.",
            sheet=sheet,
            row=row,
            column=column,
            value=value,
        )


def _table(
    name: str, primary_key: tuple[str, ...], rows: Iterable[CompiledRow]
) -> CompiledTable:
    compiled_rows = tuple(rows)
    seen: dict[tuple[object, ...], CompiledRow] = {}
    for row in compiled_rows:
        try:
            key = tuple(row.values[column] for column in primary_key)
        except KeyError as error:
            raise _decision(
                "central_primary_key_missing",
                f"Compiled table {name!r} is missing primary-key column {error.args[0]!r}.",
                sheet=row.source_sheet,
                row=row.source_row,
            ) from error
        if key in seen:
            raise _decision(
                "central_primary_key_duplicate",
                f"Compiled table {name!r} has duplicate canonical key {key!r}.",
                sheet=row.source_sheet,
                row=row.source_row,
                value=key,
            )
        seen[key] = row
    return CompiledTable(name=name, primary_key=primary_key, rows=compiled_rows)


def compile_central_tables(
    profile: WorkbookProfile, workbook_path: Path
) -> tuple[CompiledTable, ...]:
    """Compile central workbook rows and hard-stop on unresolved relationships."""
    path = Path(workbook_path)
    if path.resolve() != profile.workbook_path.resolve() or _sha256(path) != profile.workbook_sha256:
        raise _decision(
            "workbook_profile_mismatch",
            "The workbook path or content no longer matches its read-only profile.",
            value=str(path),
        )
    blocking_findings = [
        finding
        for finding in profile.findings
        if finding.severity == "error"
        or finding.status in {"decision_required", "contract_mismatch"}
    ]
    if blocking_findings:
        finding = blocking_findings[0]
        raise _decision(
            finding.code,
            finding.message,
            sheet=finding.source_sheet,
            row=finding.source_row,
            column=finding.source_column,
            value=finding.value,
        )
    if profile.active_models != LIVE_MODELS:
        raise _decision(
            "live_model_catalog_mismatch",
            "Central compilation requires the exact canonical live-model catalog.",
            value={"expected": LIVE_MODELS, "actual": profile.active_models},
        )

    expected_destinations = {
        table_name for table_name, _ in _TABLE_PRIMARY_KEYS
    }
    profiled_destinations = {
        destination
        for sheet in profile.sheets
        if sheet.disposition in {"canonical_direct", "canonical_split"}
        for destination in sheet.destination_tables
    }
    missing_destinations = expected_destinations - profiled_destinations
    if missing_destinations:
        raise _decision(
            "central_profile_coverage_missing",
            "The workbook profile does not own every central compiler destination.",
            value=tuple(sorted(missing_destinations)),
        )

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        source = {
            sheet_name: _read_rows(workbook, sheet_name)
            for sheet_name in _SHEET_HEADERS
        }

        model_rows: list[CompiledRow] = []
        registry_by_model: dict[str, str] = {}
        expected_variants_by_model: dict[str, int] = {}
        for row_number, row in _active_rows(source["model_master"], "model_master"):
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="model_master",
                row=row_number,
                column="model_key",
            )
            if model_key not in profile.active_models:
                continue
            registry_key = _required_text(
                row.get("registry_key"),
                sheet="model_master",
                row=row_number,
                column="registry_key",
            )
            registry_by_model[model_key] = registry_key
            expected_variant_count = _integer(
                row.get("expected_variant_count"),
                sheet="model_master",
                row=row_number,
                column="expected_variant_count",
            )
            assert expected_variant_count is not None
            expected_variants_by_model[model_key] = expected_variant_count
            parameters = {"model_key": model_parameter} if model_parameter else {}
            model_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "registry_key": registry_key,
                        "model_label": _required_text(
                            row.get("model_label"),
                            sheet="model_master",
                            row=row_number,
                            column="model_label",
                        ),
                        "model_year": _integer(
                            row.get("model_year"),
                            sheet="model_master",
                            row=row_number,
                            column="model_year",
                        ),
                        "dataset_name": _text(row.get("dataset_name")),
                        "export_slug": _text(row.get("export_slug")),
                        "expected_variant_count": expected_variant_count,
                        "default_model": _boolean(
                            row.get("default_model"),
                            sheet="model_master",
                            row=row_number,
                            column="default_model",
                        ),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "model_master",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )
        model_keys = {row.values["model_key"] for row in model_rows}
        if model_keys != set(profile.active_models):
            raise _decision(
                "live_model_rows_missing",
                "Active model rows do not exactly cover the profiled live models.",
                sheet="model_master",
                value={"expected": profile.active_models, "actual": tuple(sorted(model_keys))},
            )

        promotion_rows: list[CompiledRow] = []
        for row_number, row in _active_rows(
            source["model_registry_promotion"], "model_registry_promotion"
        ):
            if not _boolean(
                row.get("promoted_to_runtime"),
                sheet="model_registry_promotion",
                row=row_number,
                column="promoted_to_runtime",
            ):
                continue
            model_key, parameter = _lower(
                row.get("model_key"),
                sheet="model_registry_promotion",
                row=row_number,
                column="model_key",
            )
            _ensure_reference(
                model_key,
                model_keys,
                code="promotion_model_reference_missing",
                sheet="model_registry_promotion",
                row=row_number,
                column="model_key",
            )
            registry_key = _required_text(
                row.get("registry_key"),
                sheet="model_registry_promotion",
                row=row_number,
                column="registry_key",
            )
            if registry_key != registry_by_model[model_key]:
                raise _decision(
                    "promotion_registry_reference_mismatch",
                    "Promotion registry_key differs from its model row.",
                    sheet="model_registry_promotion",
                    row=row_number,
                    column="registry_key",
                    value=registry_key,
                )
            promotion_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "registry_key": registry_key,
                        "promoted_to_runtime": True,
                        "default_model": _boolean(
                            row.get("default_model"),
                            sheet="model_registry_promotion",
                            row=row_number,
                            column="default_model",
                        ),
                        "artifact_path": _required_text(
                            row.get("artifact_path"),
                            sheet="model_registry_promotion",
                            row=row_number,
                            column="artifact_path",
                        ),
                        "artifact_type": _required_text(
                            row.get("artifact_type"),
                            sheet="model_registry_promotion",
                            row=row_number,
                            column="artifact_type",
                        ),
                        "legacy_alias": _text(row.get("legacy_alias")),
                        "active": True,
                        "display_order": _integer(
                            row.get("display_order"),
                            sheet="model_registry_promotion",
                            row=row_number,
                            column="display_order",
                        ),
                        "notes": _text(row.get("notes")),
                    },
                    "model_registry_promotion",
                    row_number,
                    lineage_role="normalized" if parameter else "direct",
                    mapping_parameters={"model_key": parameter} if parameter else {},
                )
            )
        if {row.values["model_key"] for row in promotion_rows} != model_keys:
            raise _decision(
                "live_promotion_rows_missing",
                "Promoted registry rows do not exactly cover the live models.",
                sheet="model_registry_promotion",
            )

        variant_rows: list[CompiledRow] = []
        body_sources: dict[str, tuple[int, str, dict[str, object]]] = {}
        trim_sources: dict[str, tuple[int, str, dict[str, object]]] = {}
        for row_number, row in _active_rows(source["variant_master"], "variant_master"):
            variant_id, variant_parameter = _lower(
                row.get("variant_id"),
                sheet="variant_master",
                row=row_number,
                column="variant_id",
            )
            trim_level, trim_parameter = _lower(
                row.get("trim_level"),
                sheet="variant_master",
                row=row_number,
                column="trim_level",
            )
            body_style, body_parameter = _lower(
                row.get("body_style"),
                sheet="variant_master",
                row=row_number,
                column="body_style",
            )
            parameters = {
                key: value
                for key, value in {
                    "variant_id": variant_parameter,
                    "trim_level": trim_parameter,
                    "body_style": body_parameter,
                }.items()
                if value
            }
            variant_rows.append(
                _compiled_row(
                    {
                        "variant_id": variant_id,
                        "model_year": _integer(
                            row.get("model_year"),
                            sheet="variant_master",
                            row=row_number,
                            column="model_year",
                        ),
                        "trim_level": trim_level,
                        "body_style": body_style,
                        "display_name": _required_text(
                            row.get("display_name"),
                            sheet="variant_master",
                            row=row_number,
                            column="display_name",
                        ),
                        "base_price": _integer(
                            row.get("base_price"),
                            sheet="variant_master",
                            row=row_number,
                            column="base_price",
                        ),
                        "display_order": _integer(
                            row.get("display_order"),
                            sheet="variant_master",
                            row=row_number,
                            column="display_order",
                        ),
                        "active": True,
                    },
                    "variant_master",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )
            body_sources.setdefault(body_style, (row_number, _text(row.get("body_style")), body_parameter))
            trim_sources.setdefault(trim_level, (row_number, _text(row.get("trim_level")), trim_parameter))

        body_rows = [
            _compiled_row(
                {"body_style": body_style},
                "variant_master",
                source_row,
                lineage_role="normalized" if parameter else "direct",
                mapping_parameters={"body_style": parameter} if parameter else {},
            )
            for body_style, (source_row, _original, parameter) in sorted(body_sources.items())
        ]
        trim_rows = [
            _compiled_row(
                {"trim_level": trim_level},
                "variant_master",
                source_row,
                lineage_role="normalized" if parameter else "direct",
                mapping_parameters={"trim_level": parameter} if parameter else {},
            )
            for trim_level, (source_row, _original, parameter) in sorted(trim_sources.items())
        ]
        variant_ids = {row.values["variant_id"] for row in variant_rows}
        body_styles = set(body_sources)
        trim_levels = set(trim_sources)

        model_variant_rows: list[CompiledRow] = []
        variants_by_model: dict[str, set[str]] = defaultdict(set)
        for row_number, row in _active_rows(source["model_variants"], "model_variants"):
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="model_variants",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                raise _decision(
                    "model_variant_model_reference_missing",
                    "An active model_variants row references no live model.",
                    sheet="model_variants",
                    row=row_number,
                    column="model_key",
                    value=model_key,
                )
            variant_id, variant_parameter = _lower(
                row.get("variant_id"),
                sheet="model_variants",
                row=row_number,
                column="variant_id",
            )
            _ensure_reference(
                variant_id,
                variant_ids,
                code="model_variant_reference_missing",
                sheet="model_variants",
                row=row_number,
                column="variant_id",
            )
            parameters = {
                key: value
                for key, value in {
                    "model_key": model_parameter,
                    "variant_id": variant_parameter,
                }.items()
                if value
            }
            model_variant_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "variant_id": variant_id,
                        "display_order": _integer(
                            row.get("display_order"),
                            sheet="model_variants",
                            row=row_number,
                            column="display_order",
                        ),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "model_variants",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )
            variants_by_model[model_key].add(variant_id)
        for model_key in sorted(model_keys):
            actual = len(variants_by_model.get(model_key, set()))
            expected = expected_variants_by_model[model_key]
            if actual != expected:
                raise _decision(
                    "model_variant_count_mismatch",
                    "Active model variant membership does not match the model contract.",
                    sheet="model_variants",
                    value={
                        "model_key": model_key,
                        "expected": expected,
                        "actual": actual,
                    },
                )

        variant_by_id = {row.values["variant_id"]: row.values for row in variant_rows}
        model_body_styles = {
            model: {variant_by_id[variant]["body_style"] for variant in variants}
            for model, variants in variants_by_model.items()
        }
        model_trim_levels = {
            model: {variant_by_id[variant]["trim_level"] for variant in variants}
            for model, variants in variants_by_model.items()
        }

        section_rows: list[CompiledRow] = []
        sections_by_id: dict[str, Mapping[str, object]] = {}
        for row_number, row in source["section_master"]:
            section_id, parameter = _lower(
                row.get("section_id"),
                sheet="section_master",
                row=row_number,
                column="section_id",
            )
            step_key, step_parameter = _lower(
                row.get("step_key"),
                sheet="section_master",
                row=row_number,
                column="step_key",
            )
            parameters = {
                key: value
                for key, value in {
                    "section_id": parameter,
                    "step_key": step_parameter,
                }.items()
                if value
            }
            values = {
                "section_id": section_id,
                "section_name": _required_text(
                    row.get("section_name"),
                    sheet="section_master",
                    row=row_number,
                    column="section_name",
                ),
                "selection_mode": _text(row.get("selection_mode")) or None,
                "is_required": _boolean(
                    row.get("is_required"),
                    sheet="section_master",
                    row=row_number,
                    column="is_required",
                ),
                "display_order": _integer(
                    row.get("display_order"),
                    sheet="section_master",
                    row=row_number,
                    column="display_order",
                    nullable=True,
                ),
                "standard_behavior": _text(row.get("standard_behavior")) or None,
                "step_key": step_key,
            }
            section_rows.append(
                _compiled_row(
                    values,
                    "section_master",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )
            sections_by_id[section_id] = values
        section_ids = set(sections_by_id)
        runtime_step_rows: list[CompiledRow] = []
        runtime_step_keys: set[tuple[str, str]] = set()
        for row_number, row in _active_rows(source["runtime_steps"], "runtime_steps"):
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="runtime_steps",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                continue
            step_key, step_parameter = _lower(
                row.get("step_key"),
                sheet="runtime_steps",
                row=row_number,
                column="step_key",
            )
            parameters = {
                key: value
                for key, value in {
                    "model_key": model_parameter,
                    "step_key": step_parameter,
                }.items()
                if value
            }
            runtime_step_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "step_key": step_key,
                        "step_label": _required_text(
                            row.get("step_label"),
                            sheet="runtime_steps",
                            row=row_number,
                            column="step_label",
                        ),
                        "runtime_order": _integer(
                            row.get("runtime_order"),
                            sheet="runtime_steps",
                            row=row_number,
                            column="runtime_order",
                        ),
                        "source": _required_text(
                            row.get("source"),
                            sheet="runtime_steps",
                            row=row_number,
                            column="source",
                        ),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "runtime_steps",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )
            runtime_step_keys.add((model_key, step_key))

        runtime_route_keys = set(runtime_step_keys)
        for row_number, row in _active_rows(
            source["step_order_summary_map"], "step_order_summary_map"
        ):
            model_key, _ = _lower(
                row.get("model_key"),
                sheet="step_order_summary_map",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                continue
            route_key, _ = _lower(
                row.get("step_key"),
                sheet="step_order_summary_map",
                row=row_number,
                column="step_key",
            )
            runtime_route_keys.add((model_key, route_key))

        section_presentation_rows: list[CompiledRow] = []
        for row_number, row in _active_rows(
            source["section_presentation"], "section_presentation"
        ):
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="section_presentation",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                continue
            section_id, section_parameter = _lower(
                row.get("section_id"),
                sheet="section_presentation",
                row=row_number,
                column="section_id",
            )
            _ensure_reference(
                section_id,
                section_ids,
                code="section_presentation_reference_missing",
                sheet="section_presentation",
                row=row_number,
                column="section_id",
            )
            section = sections_by_id[section_id]
            raw_step = "" if row.get("step_key") is None else str(row["step_key"])
            if raw_step.strip():
                step_key, step_parameter = _lower(
                    row.get("step_key"),
                    sheet="section_presentation",
                    row=row_number,
                    column="step_key",
                )
            else:
                step_key = str(section["step_key"])
                step_parameter = {}
            parameters: dict[str, object] = {
                key: value
                for key, value in {
                    "model_key": model_parameter,
                    "section_id": section_parameter,
                }.items()
                if value
            }
            display_label = _text(row.get("display_label"))
            if not display_label:
                display_label = str(section["section_name"])
                parameters["display_label"] = {
                    "original": row.get("display_label"),
                    "transform": "fallback_from_sections.section_name",
                }
            if not raw_step.strip():
                parameters["step_key"] = {
                    "original": row.get("step_key"),
                    "transform": "fallback_from_sections.step_key",
                }
            elif step_parameter:
                parameters["step_key"] = step_parameter
            display_order = _integer(
                row.get("section_display_order"),
                sheet="section_presentation",
                row=row_number,
                column="section_display_order",
                nullable=True,
            )
            if display_order is None:
                display_order = int(section["display_order"])
                parameters["section_display_order"] = {
                    "original": row.get("section_display_order"),
                    "transform": "fallback_from_sections.display_order",
                }
            section_presentation_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "section_id": section_id,
                        "display_label": display_label,
                        "step_key": step_key,
                        "display_behavior": _text(row.get("display_behavior")) or None,
                        "section_display_order": display_order,
                        "standard_equipment_bucket": _text(row.get("standard_equipment_bucket")),
                        "standard_equipment_group_type": _text(row.get("standard_equipment_group_type")),
                        "auto_added_bucket": _text(row.get("auto_added_bucket")),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "section_presentation",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )

        context_section_rows: list[CompiledRow] = []
        context_keys: set[tuple[str, str]] = set()
        for row_number, row in _active_rows(
            source["context_section_master"], "context_section_master"
        ):
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="context_section_master",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                continue
            context_type, context_parameter = _lower(
                row.get("context_type"),
                sheet="context_section_master",
                row=row_number,
                column="context_type",
            )
            section_id, section_parameter = _lower(
                row.get("section_id"),
                sheet="context_section_master",
                row=row_number,
                column="section_id",
            )
            if section_id in section_ids:
                raise _decision(
                    "context_section_ownership_ambiguous",
                    "A context section collides with an option section identifier.",
                    sheet="context_section_master",
                    row=row_number,
                    column="section_id",
                    value=section_id,
                )
            step_key, step_parameter = _lower(
                row.get("step_key"),
                sheet="context_section_master",
                row=row_number,
                column="step_key",
            )
            _ensure_reference(
                (model_key, step_key),
                runtime_route_keys,
                code="context_runtime_route_reference_missing",
                sheet="context_section_master",
                row=row_number,
                column="step_key",
            )
            parameters = {
                key: value
                for key, value in {
                    "model_key": model_parameter,
                    "context_type": context_parameter,
                    "section_id": section_parameter,
                    "step_key": step_parameter,
                }.items()
                if value
            }
            context_section_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "context_type": context_type,
                        "section_id": section_id,
                        "section_name": _required_text(
                            row.get("section_name"),
                            sheet="context_section_master",
                            row=row_number,
                            column="section_name",
                        ),
                        "selection_mode": _required_text(
                            row.get("selection_mode"),
                            sheet="context_section_master",
                            row=row_number,
                            column="selection_mode",
                        ),
                        "choice_mode": _required_text(
                            row.get("choice_mode"),
                            sheet="context_section_master",
                            row=row_number,
                            column="choice_mode",
                        ),
                        "is_required": _boolean(
                            row.get("is_required"),
                            sheet="context_section_master",
                            row=row_number,
                            column="is_required",
                        ),
                        "standard_behavior": _required_text(
                            row.get("standard_behavior"),
                            sheet="context_section_master",
                            row=row_number,
                            column="standard_behavior",
                        ),
                        "section_display_order": _integer(
                            row.get("section_display_order"),
                            sheet="context_section_master",
                            row=row_number,
                            column="section_display_order",
                        ),
                        "step_key": step_key,
                        "step_label": _required_text(
                            row.get("step_label"),
                            sheet="context_section_master",
                            row=row_number,
                            column="step_label",
                        ),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "context_section_master",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )
            context_keys.add((model_key, context_type))

        # Context choices are derived from active model variants, matching
        # contract.build_model_context_choices.  context_choice_copy supplies
        # wildcard/exact tooltip overlays; it is not the choice inventory.
        normalized_copy_rows: list[dict[str, object]] = []
        for row_number, row in _active_rows(
            source["context_choice_copy"], "context_choice_copy"
        ):
            raw_model = _text(row.get("model_key")) or "*"
            model_key = raw_model.lower()
            if model_key != "*":
                _ensure_reference(
                    model_key,
                    model_keys,
                    code="context_choice_model_reference_missing",
                    sheet="context_choice_copy",
                    row=row_number,
                    column="model_key",
                )
            context_type, _ = _lower(
                row.get("context_type"),
                sheet="context_choice_copy",
                row=row_number,
                column="context_type",
            )
            if context_type not in {"body_style", "trim_level"}:
                raise _decision(
                    "context_choice_type_unresolved",
                    "A context choice copy type has no approved relational domain.",
                    sheet="context_choice_copy",
                    row=row_number,
                    column="context_type",
                    value=context_type,
                )
            value, _ = _lower(
                row.get("value"),
                sheet="context_choice_copy",
                row=row_number,
                column="value",
            )
            raw_body_style = _text(row.get("body_style")) or "*"
            body_style = raw_body_style.lower()
            if body_style != "*":
                _ensure_reference(
                    body_style,
                    body_styles,
                    code="context_choice_body_style_reference_missing",
                    sheet="context_choice_copy",
                    row=row_number,
                    column="body_style",
                )
            candidate_models = profile.active_models if model_key == "*" else (model_key,)
            accepted_models = [
                candidate
                for candidate in candidate_models
                if (candidate, context_type) in context_keys
                and value
                in (
                    model_trim_levels[candidate]
                    if context_type == "trim_level"
                    else model_body_styles[candidate]
                )
                and (
                    body_style == "*"
                    or body_style in model_body_styles[candidate]
                )
            ]
            if not accepted_models:
                raise _decision(
                    "context_choice_owner_unresolved",
                    "A context choice copy row applies to no active model choice.",
                    sheet="context_choice_copy",
                    row=row_number,
                    value={"context_type": context_type, "value": value},
                )
            normalized_copy_rows.append(
                {
                    "source_row": row_number,
                    "model_key": model_key,
                    "context_type": context_type,
                    "value": value,
                    "body_style": body_style,
                    "info_tooltip": _text(row.get("info_tooltip")),
                    "notes": _text(row.get("notes")),
                }
            )

        def _copy_for(
            model_key: str,
            context_type: str,
            value: str,
            body_style: str,
        ) -> dict[str, object] | None:
            best_score = -1
            best: dict[str, object] | None = None
            for copy_row in normalized_copy_rows:
                if copy_row["context_type"] != context_type:
                    continue
                if copy_row["value"] != value.lower():
                    continue
                if copy_row["model_key"] not in {"*", model_key}:
                    continue
                if copy_row["body_style"] not in {"*", body_style}:
                    continue
                score = (
                    2 if copy_row["model_key"] == model_key else 0
                ) + (1 if copy_row["body_style"] == body_style else 0)
                if copy_row["info_tooltip"] and score > best_score:
                    best_score = score
                    best = copy_row
            return best

        context_section_by_key = {
            (row.values["model_key"], row.values["context_type"]): row
            for row in context_section_rows
        }
        variant_row_by_id = {
            row.values["variant_id"]: row for row in variant_rows
        }
        memberships_by_model = {
            model_key: [
                row
                for row in model_variant_rows
                if row.values["model_key"] == model_key
            ]
            for model_key in LIVE_MODELS
        }
        context_choice_rows: list[CompiledRow] = []
        for model_key in LIVE_MODELS:
            memberships = memberships_by_model[model_key]
            ordered_body_styles: list[str] = []
            for membership in memberships:
                body_style = variant_row_by_id[
                    membership.values["variant_id"]
                ].values["body_style"]
                if body_style not in ordered_body_styles:
                    ordered_body_styles.append(body_style)
            body_section = context_section_by_key[(model_key, "body_style")]
            for display_order, body_style in enumerate(
                ordered_body_styles, start=1
            ):
                body_memberships = [
                    membership
                    for membership in memberships
                    if variant_row_by_id[
                        membership.values["variant_id"]
                    ].values["body_style"]
                    == body_style
                ]
                source_membership = body_memberships[0]
                copy_row = _copy_for(
                    model_key, "body_style", body_style, body_style
                )
                context_choice_id = f"body_style__{body_style}"
                parameters: dict[str, object] = {
                    "context_choice_id": {
                        "original": tuple(
                            row.values["variant_id"] for row in body_memberships
                        ),
                        "canonical": context_choice_id,
                        "transform": "derive_body_context_choice_id",
                    },
                    "description": {
                        "original": len(body_memberships),
                        "canonical": f"{len(body_memberships)} trims available",
                        "transform": "describe_body_variant_count",
                    },
                    "section_id": {
                        "source_sheet": body_section.source_sheet,
                        "source_row": body_section.source_row,
                        "original": body_section.values["section_id"],
                        "canonical": body_section.values["section_id"],
                        "transform": "join_context_section",
                    },
                    "step_key": {
                        "source_sheet": body_section.source_sheet,
                        "source_row": body_section.source_row,
                        "original": body_section.values["step_key"],
                        "canonical": body_section.values["step_key"],
                        "transform": "join_context_route",
                    },
                }
                if copy_row is not None:
                    parameters["info_tooltip"] = {
                        "source_sheet": "context_choice_copy",
                        "source_row": copy_row["source_row"],
                        "original": copy_row["info_tooltip"],
                        "canonical": copy_row["info_tooltip"],
                        "transform": "apply_wildcard_exact_tooltip_precedence",
                    }
                context_choice_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "context_choice_id": context_choice_id,
                            "context_type": "body_style",
                            "value": body_style,
                            "label": body_style.title(),
                            "description": f"{len(body_memberships)} trims available",
                            "info_tooltip": (
                                str(copy_row["info_tooltip"])
                                if copy_row is not None
                                else ""
                            ),
                            "section_id": body_section.values["section_id"],
                            "step_key": body_section.values["step_key"],
                            "body_style": body_style,
                            "trim_level": None,
                            "variant_id": None,
                            "base_price": None,
                            "display_order": display_order,
                            "active": True,
                            "notes": (
                                str(copy_row["notes"])
                                if copy_row is not None
                                else ""
                            ),
                        },
                        source_membership.source_sheet,
                        source_membership.source_row,
                        lineage_role="normalized",
                        mapping_parameters=parameters,
                    )
                )

            trim_section = context_section_by_key[(model_key, "trim_level")]
            for membership in memberships:
                variant = variant_row_by_id[membership.values["variant_id"]]
                body_style = str(variant.values["body_style"])
                trim_level = str(variant.values["trim_level"])
                display_value = trim_level.upper()
                context_choice_id = (
                    f"trim_level__{body_style}__{trim_level.lower()}"
                )
                copy_row = _copy_for(
                    model_key,
                    "trim_level",
                    display_value,
                    body_style,
                )
                parameters = {
                    "context_choice_id": {
                        "original": variant.values["variant_id"],
                        "canonical": context_choice_id,
                        "transform": "derive_trim_context_choice_id",
                    },
                    "value": {
                        "original": trim_level,
                        "canonical": display_value,
                        "transform": "uppercase_runtime_trim_label",
                    },
                    "label": {
                        "original": trim_level,
                        "canonical": display_value,
                        "transform": "uppercase_runtime_trim_label",
                    },
                    "variant_id": {
                        "source_sheet": variant.source_sheet,
                        "source_row": variant.source_row,
                        "original": variant.values["variant_id"],
                        "canonical": variant.values["variant_id"],
                        "transform": "join_model_variant",
                    },
                    "section_id": {
                        "source_sheet": trim_section.source_sheet,
                        "source_row": trim_section.source_row,
                        "original": trim_section.values["section_id"],
                        "canonical": trim_section.values["section_id"],
                        "transform": "join_context_section",
                    },
                    "step_key": {
                        "source_sheet": trim_section.source_sheet,
                        "source_row": trim_section.source_row,
                        "original": trim_section.values["step_key"],
                        "canonical": trim_section.values["step_key"],
                        "transform": "join_context_route",
                    },
                }
                if copy_row is not None:
                    parameters["info_tooltip"] = {
                        "source_sheet": "context_choice_copy",
                        "source_row": copy_row["source_row"],
                        "original": copy_row["info_tooltip"],
                        "canonical": copy_row["info_tooltip"],
                        "transform": "apply_wildcard_exact_tooltip_precedence",
                    }
                context_choice_rows.append(
                    _compiled_row(
                        {
                            "model_key": model_key,
                            "context_choice_id": context_choice_id,
                            "context_type": "trim_level",
                            "value": display_value,
                            "label": display_value,
                            "description": variant.values["display_name"],
                            "info_tooltip": (
                                str(copy_row["info_tooltip"])
                                if copy_row is not None
                                else ""
                            ),
                            "section_id": trim_section.values["section_id"],
                            "step_key": trim_section.values["step_key"],
                            "body_style": body_style,
                            "trim_level": trim_level,
                            "variant_id": variant.values["variant_id"],
                            "base_price": variant.values["base_price"],
                            "display_order": variant.values["display_order"],
                            "active": True,
                            "notes": (
                                str(copy_row["notes"])
                                if copy_row is not None
                                else ""
                            ),
                        },
                        membership.source_sheet,
                        membership.source_row,
                        lineage_role="normalized",
                        mapping_parameters=parameters,
                    )
                )

        summary_rows: list[CompiledRow] = []
        summary_keys: set[tuple[str, str]] = set()
        for row_number, row in _active_rows(
            source["order_summary_sections"], "order_summary_sections"
        ):
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="order_summary_sections",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                continue
            section_key, section_parameter = _lower(
                row.get("section_key"),
                sheet="order_summary_sections",
                row=row_number,
                column="section_key",
            )
            parameters = {
                key: value
                for key, value in {
                    "model_key": model_parameter,
                    "section_key": section_parameter,
                }.items()
                if value
            }
            summary_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "section_key": section_key,
                        "section_label": _required_text(
                            row.get("section_label"),
                            sheet="order_summary_sections",
                            row=row_number,
                            column="section_label",
                        ),
                        "display_order": _integer(
                            row.get("display_order"),
                            sheet="order_summary_sections",
                            row=row_number,
                            column="display_order",
                        ),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "order_summary_sections",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )
            summary_keys.add((model_key, section_key))

        step_summary_rows: list[CompiledRow] = []
        step_summary_keys: set[tuple[str, str]] = set()
        for row_number, row in _active_rows(
            source["step_order_summary_map"], "step_order_summary_map"
        ):
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="step_order_summary_map",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                continue
            step_key, step_parameter = _lower(
                row.get("step_key"),
                sheet="step_order_summary_map",
                row=row_number,
                column="step_key",
            )
            section_key, section_parameter = _lower(
                row.get("section_key"),
                sheet="step_order_summary_map",
                row=row_number,
                column="section_key",
            )
            _ensure_reference(
                (model_key, section_key),
                summary_keys,
                code="summary_section_reference_missing",
                sheet="step_order_summary_map",
                row=row_number,
                column="section_key",
            )
            route_identity = (model_key, step_key)
            if route_identity in step_summary_keys:
                raise _decision(
                    "step_summary_route_duplicate",
                    "A model route maps to more than one active summary destination.",
                    sheet="step_order_summary_map",
                    row=row_number,
                    column="step_key",
                    value=route_identity,
                )
            step_summary_keys.add(route_identity)
            parameters = {
                key: value
                for key, value in {
                    "model_key": model_parameter,
                    "step_key": step_parameter,
                    "section_key": section_parameter,
                }.items()
                if value
            }
            step_summary_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "step_key": step_key,
                        "section_key": section_key,
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "step_order_summary_map",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )

        model_asset_rows: list[CompiledRow] = []
        for row_number, row in _active_rows(source["asset_map"], "asset_map"):
            if _text(row.get("target_type")).lower() != "model":
                continue
            model_key, model_parameter = _lower(
                row.get("model_key"),
                sheet="asset_map",
                row=row_number,
                column="model_key",
            )
            if model_key not in model_keys:
                continue
            target_id = _required_text(
                row.get("target_id"),
                sheet="asset_map",
                row=row_number,
                column="target_id",
            )
            if target_id != registry_by_model[model_key]:
                raise _decision(
                    "model_asset_registry_reference_mismatch",
                    "A model asset target differs from the canonical registry key.",
                    sheet="asset_map",
                    row=row_number,
                    column="target_id",
                    value=target_id,
                )
            model_asset_rows.append(
                _compiled_row(
                    {
                        "model_key": model_key,
                        "image_url": _required_text(
                            row.get("image_url"),
                            sheet="asset_map",
                            row=row_number,
                            column="image_url",
                        ),
                        "image_alt": _text(row.get("image_alt")),
                        "image_fit": _text(row.get("image_fit")),
                        "image_position": _text(row.get("image_position")),
                        "hover_image_url": _text(row.get("hover_image_url")),
                        "hover_image_alt": _text(row.get("hover_image_alt")),
                        "hover_image_position": _text(row.get("hover_image_position")),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "asset_map",
                    row_number,
                    lineage_role="normalized" if model_parameter else "direct",
                    mapping_parameters={"model_key": model_parameter} if model_parameter else {},
                )
            )
        if {row.values["model_key"] for row in model_asset_rows} != model_keys:
            raise _decision(
                "live_model_assets_missing",
                "Every live model must have one active model asset.",
                sheet="asset_map",
            )

        price_rows: list[CompiledRow] = []
        for row_number, row in source["PriceRef"]:
            option_type, option_parameter = _lower(
                row.get("OptionType"),
                sheet="PriceRef",
                row=row_number,
                column="OptionType",
            )
            raw_trim = "" if row.get("Trim") is None else str(row["Trim"])
            trim_level: str | None
            trim_parameter: dict[str, object] = {}
            if not raw_trim.strip():
                trim_level = None
                trim_parameter = {
                    "original": row.get("Trim"),
                    "transform": "unrestricted_to_null",
                }
            else:
                trim_level, trim_parameter = _lower(
                    row.get("Trim"),
                    sheet="PriceRef",
                    row=row_number,
                    column="Trim",
                )
            parameters = {
                key: item
                for key, item in {
                    "option_type": option_parameter,
                    "trim_level": trim_parameter,
                }.items()
                if item
            }
            price_rows.append(
                _compiled_row(
                    {
                        "option_type": option_type,
                        "trim_level": trim_level,
                        "code": _required_text(
                            row.get("Code"),
                            sheet="PriceRef",
                            row=row_number,
                            column="Code",
                        ),
                        "price": _integer(
                            row.get("Price"),
                            sheet="PriceRef",
                            row=row_number,
                            column="Price",
                        ),
                    },
                    "PriceRef",
                    row_number,
                    lineage_role="normalized" if parameters else "direct",
                    mapping_parameters=parameters,
                )
            )

        phrase_rows: list[CompiledRow] = []
        for row_number, row in _active_rows(source["rule_phrase_map"], "rule_phrase_map"):
            phrase_rows.append(
                _compiled_row(
                    {
                        "phrase": _required_text(
                            row.get("phrase"),
                            sheet="rule_phrase_map",
                            row=row_number,
                            column="phrase",
                        ),
                        "rule_type": _required_text(
                            row.get("rule_type"),
                            sheet="rule_phrase_map",
                            row=row_number,
                            column="rule_type",
                        ),
                        "direction": _required_text(
                            row.get("direction"),
                            sheet="rule_phrase_map",
                            row=row_number,
                            column="direction",
                        ),
                        "stop_phrases": _text(row.get("stop_phrases")),
                        "review_flag_default": _boolean(
                            row.get("review_flag_default"),
                            sheet="rule_phrase_map",
                            row=row_number,
                            column="review_flag_default",
                        ),
                        "active": True,
                        "notes": _text(row.get("notes")),
                    },
                    "rule_phrase_map",
                    row_number,
                )
            )

        route_rows_by_key: dict[tuple[str, str], CompiledRow] = {}
        for runtime_step in runtime_step_rows:
            model_key = str(runtime_step.values["model_key"])
            step_key = str(runtime_step.values["step_key"])
            route_parameters = dict(runtime_step.mapping_parameters)
            route_parameters.update(
                {
                    "model_key": _derived_mapping_parameter(
                        runtime_step,
                        "model_key",
                        "derived_from_runtime_steps.model_key",
                    ),
                    "route_key": _derived_mapping_parameter(
                        runtime_step,
                        "step_key",
                        "derived_from_runtime_steps.step_key",
                    ),
                }
            )
            route_rows_by_key[(model_key, step_key)] = _compiled_row(
                {
                    "model_key": model_key,
                    "route_key": step_key,
                    "route_kind": "visible_step",
                },
                runtime_step.source_sheet,
                runtime_step.source_row,
                lineage_role=runtime_step.lineage_role,
                mapping_parameters=route_parameters,
            )
        for summary_map in step_summary_rows:
            model_key = str(summary_map.values["model_key"])
            step_key = str(summary_map.values["step_key"])
            route_parameters = dict(summary_map.mapping_parameters)
            route_parameters.update(
                {
                    "model_key": _derived_mapping_parameter(
                        summary_map,
                        "model_key",
                        "derived_from_step_order_summary_map.model_key",
                    ),
                    "route_key": _derived_mapping_parameter(
                        summary_map,
                        "step_key",
                        "derived_from_step_order_summary_map.step_key",
                    ),
                }
            )
            route_rows_by_key.setdefault(
                (model_key, step_key),
                _compiled_row(
                    {
                        "model_key": model_key,
                        "route_key": step_key,
                        "route_kind": "hidden_summary_bucket",
                    },
                    summary_map.source_sheet,
                    summary_map.source_row,
                    lineage_role=summary_map.lineage_role,
                    mapping_parameters=route_parameters,
                ),
            )
        route_rows = [
            route_rows_by_key[key] for key in sorted(route_rows_by_key)
        ]

        rows_by_table: Mapping[str, Iterable[CompiledRow]] = {
            "models": model_rows,
            "model_registry_promotion": promotion_rows,
            "body_styles": body_rows,
            "trim_levels": trim_rows,
            "variants": variant_rows,
            "model_variants": model_variant_rows,
            "sections": section_rows,
            "section_presentation": section_presentation_rows,
            "runtime_route_keys": route_rows,
            "runtime_steps": runtime_step_rows,
            "runtime_context_sections": context_section_rows,
            "runtime_context_choices": context_choice_rows,
            "runtime_summary_sections": summary_rows,
            "runtime_step_summary_map": step_summary_rows,
            "model_assets": model_asset_rows,
            "price_ref": price_rows,
            "rule_phrase_map": phrase_rows,
        }
        return tuple(
            _table(name, primary_key, rows_by_table[name])
            for name, primary_key in _TABLE_PRIMARY_KEYS
        )
    finally:
        workbook.close()


__all__ = [
    "CompiledRow",
    "CompiledTable",
    "DecisionRequired",
    "compile_central_tables",
]
