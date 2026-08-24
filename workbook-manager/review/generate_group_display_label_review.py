#!/usr/bin/env python3
"""Checkpoint 2C: generate the complete group display-label review artifacts.

Spec: docs/superpowers/specs/2026-08-21-workbook-manager-ux-recovery.md §7.3.
One record per existing exclusive/rule group across every generatable model,
assembled from the canonical workbook. Evidence-only: `proposed_display_label`
is left blank (no authoritative label exists), `review_status` starts as
`pending`, and nothing is auto-approved. Ordering is stable by model, group
type, source sheet/row, and group_id. Read-only against the workbook.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts"))

from openpyxl import load_workbook  # noqa: E402

from corvette_form_generator.model_configs import discover_generation_model_configs  # noqa: E402
from corvette_form_generator.rules import load_exclusive_groups, load_rule_groups  # noqa: E402
from corvette_form_generator.workbook import clean, rows_from_optional_sheet  # noqa: E402
from corvette_form_generator.workbook_domain.registry import (  # noqa: E402
    READONLY_SHEET_META,
)

ARTIFACT_SCHEMA_VERSION = "1"
REVIEW_DIR = ROOT / "workbook-manager" / "review"
_HASH_SUFFIX = re.compile(r"_[0-9a-f]{12}$", re.IGNORECASE)

PLACEHOLDER_FALLBACK = "Label pending workbook review"


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _fallback_label(group_id: str) -> str:
    """Mirror the Manager's factual fallback (explorer._group_fallback)."""
    if _HASH_SUFFIX.search(group_id or ""):
        return PLACEHOLDER_FALLBACK
    return (group_id or "Unnamed group").replace("_", " ").strip().title()


def _section_names(wb) -> dict[str, str]:
    """section_id -> section_name from the registry-owned read-only sheet."""
    sheet = READONLY_SHEET_META["sections"]["sheet"]
    names: dict[str, str] = {}
    for row in rows_from_optional_sheet(wb, sheet):
        sid = clean(row.get("section_id"))
        if sid:
            names[sid] = clean(row.get("section_name"))
    return names


def _exclusive_fallback_label(
    group_id: str,
    member_ids: list[str],
    option_sections: dict[str, str],
    section_names: dict[str, str],
) -> str:
    """Mirror the Manager's full pre-label path (explorer._exclusive_group_label).

    `_group_fallback` alone is not what the Manager renders: for a hash-suffixed
    group whose members resolve to exactly one section it shows
    `Exclusive group · <section_name>`. Recording the placeholder instead
    misstates what a reviewer would actually see today.
    """
    if not _HASH_SUFFIX.search(group_id or ""):
        return _fallback_label(group_id)
    resolved = {
        section_names.get(option_sections.get(oid, ""), "")
        for oid in member_ids
    }
    resolved.discard("")
    if len(resolved) == 1:
        return f"Exclusive group \u00b7 {next(iter(resolved))}"
    return _fallback_label(group_id)


def _member_rows(wb, config, sheet_name: str) -> dict[str, list[dict]]:
    """Active member rows keyed by group_id, in sheet order."""
    out: dict[str, list[dict]] = {}
    for row in rows_from_optional_sheet(wb, sheet_name):
        if clean(row.get("active", "True")) != "True":
            continue
        gid = clean(row.get("group_id"))
        if gid:
            out.setdefault(gid, []).append(row)
    return out


def _option_labels(wb, config) -> tuple[dict[str, tuple[str, str]], dict[str, str]]:
    """(option_id -> (rpo, name)), plus (option_id -> section_id)."""
    labels: dict[str, tuple[str, str]] = {}
    sections: dict[str, str] = {}
    for row in rows_from_optional_sheet(wb, config.source_option_sheet):
        oid = clean(row.get("option_id"))
        if not oid:
            continue
        labels[oid] = (clean(row.get("rpo")), clean(row.get("option_name")))
        sid = clean(row.get("section_id"))
        if sid:
            sections[oid] = sid
    return labels, sections


def _fmt_members(members: list[dict], id_column: str, labels: dict) -> tuple[int, str]:
    ids = [clean(m.get(id_column)) for m in members]
    parts = []
    for mid in ids:
        rpo, name = labels.get(mid, ("", ""))
        parts.append(" ".join(p for p in (rpo, name) if p) or mid)
    return len(ids), "; ".join(parts)


def _sheet_and_row(wb, sheet_name: str, group_id: str) -> tuple[str, int]:
    """Physical source sheet and 1-based data row for a group_id."""
    ws = wb[sheet_name]
    headers = [str(c.value or "") for c in ws[1]]
    gid_col = headers.index("group_id")
    for r in range(2, ws.max_row + 1):
        if clean(ws.cell(row=r, column=gid_col + 1).value) == group_id:
            return sheet_name, r
    return sheet_name, 0


def _refuse_reviewed_decision_overwrite() -> None:
    csv_path = REVIEW_DIR / "group-display-label-review.csv"
    if not csv_path.exists():
        return
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    if any(
        (row.get("proposed_display_label") or "").strip()
        or (row.get("review_status") or "").strip().casefold() not in {"", "pending"}
        or (row.get("reviewer_note") or "").strip()
        for row in rows
    ):
        raise RuntimeError(
            "reviewed decisions already exist; use sync_group_display_label_review.py "
            "instead of regenerating the evidence inventory"
        )


def main() -> None:
    _refuse_reviewed_decision_overwrite()
    wb_path = ROOT / "stingray_master.xlsx"
    wb_sha = _sha256(wb_path)
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    try:
        configs = discover_generation_model_configs(wb_path)
        records: list[dict] = []

        for model_key in sorted(configs):
            cfg = configs[model_key]
            option_labels, option_sections = _option_labels(wb, cfg)
            section_names = _section_names(wb)
            excl_members = _member_rows(wb, cfg, cfg.exclusive_group_members_sheet)
            rule_members = _member_rows(wb, cfg, cfg.rule_group_members_sheet)

            for group in load_rule_groups(wb, cfg):
                gid = group["group_id"]
                member_ids = [t for t in group.get("target_ids", []) if t]
                count, member_text = _fmt_members(
                    [{"target_id": t} for t in member_ids], "target_id", option_labels
                )
                src_rpo, src_name = option_labels.get(group.get("source_id", ""), ("", ""))
                src_text = " ".join(p for p in (src_rpo, src_name) if p) or group.get("source_id", "")
                section_ids = sorted({
                    option_sections.get(t, "") for t in member_ids if option_sections.get(t)
                })
                records.append({
                    "model_key": model_key,
                    "group_type": "rule",
                    "group_id": gid,
                    "current_fallback_label": f"{src_text}: {clean(group.get('group_type'))}".strip(": "),
                    "source_sheet_row": "",
                    "active": clean(group.get("active")) == "True",
                    "customer_visible": False,
                    "audience": "manager",
                    "resolved_step_section": "; ".join(section_ids),
                    "selection_behavior": clean(group.get("group_type")),
                    "member_count": len(member_ids),
                    "members_ordered": (
                        f"source {src_text}; targets {member_text}" if src_text else member_text
                    ),
                    "notes": clean(group.get("notes")),
                    "proposed_display_label": "",
                    "review_status": "pending",
                    "reviewer_note": "",
                    "_sheet": cfg.rule_groups_sheet,
                    "_gid": gid,
                })

            for group in load_exclusive_groups(wb, cfg):
                gid = group["group_id"]
                member_ids = [o for o in group.get("option_ids", []) if o]
                count, member_text = _fmt_members(
                    [{"option_id": o} for o in member_ids], "option_id", option_labels
                )
                section_ids = sorted({
                    option_sections.get(o, "") for o in member_ids if option_sections.get(o)
                })
                records.append({
                    "model_key": model_key,
                    "group_type": "exclusive",
                    "group_id": gid,
                    "current_fallback_label": _exclusive_fallback_label(
                        gid, member_ids, option_sections, section_names
                    ),
                    "source_sheet_row": "",
                    "active": clean(group.get("active")) == "True",
                    "customer_visible": True,
                    "audience": "customer",
                    "resolved_step_section": "; ".join(section_ids),
                    "selection_behavior": clean(group.get("selection_mode")),
                    "member_count": len(member_ids),
                    "members_ordered": member_text,
                    "notes": clean(group.get("notes")),
                    "proposed_display_label": "",
                    "review_status": "pending",
                    "reviewer_note": "",
                    "_sheet": cfg.exclusive_groups_sheet,
                    "_gid": gid,
                })

        # Resolve physical source sheet + row after collection (§7.3 evidence).
        for rec in records:
            sheet, rownum = _sheet_and_row(wb, rec["_sheet"], rec["_gid"])
            rec["source_sheet_row"] = f"{sheet}!{rownum}"
            del rec["_sheet"], rec["_gid"]
    finally:
        wb.close()

    # §7.3 ordering: model, group type, source sheet/row, group_id.
    def sort_key(r: dict) -> tuple:
        sheet, _, rownum = r["source_sheet_row"].partition("!")
        return (r["model_key"], r["group_type"], sheet, int(rownum or 0), r["group_id"])

    records.sort(key=sort_key)

    header = [
        "artifact_schema_version", "source_workbook_sha256",
        "model_key", "group_type", "group_id",
        "current_fallback_label", "source_sheet_row",
        "active", "customer_visible", "audience",
        "resolved_step_section", "selection_behavior",
        "member_count", "members_ordered", "notes",
        "proposed_display_label", "review_status", "reviewer_note",
    ]
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)

    csv_path = REVIEW_DIR / "group-display-label-review.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=header, lineterminator="\n")
        writer.writeheader()
        for rec in records:
            row = dict(rec)
            row["artifact_schema_version"] = ARTIFACT_SCHEMA_VERSION
            row["source_workbook_sha256"] = wb_sha
            writer.writerow({k: json.dumps(v, ensure_ascii=False) if isinstance(v, bool) else v
                             for k, v in row.items()})

    payload = {
        "artifact_schema_version": ARTIFACT_SCHEMA_VERSION,
        "source_workbook_sha256": wb_sha,
        "record_count": len(records),
        "records": [
            {**rec,
             "artifact_schema_version": ARTIFACT_SCHEMA_VERSION,
             "source_workbook_sha256": wb_sha}
            for rec in records
        ],
    }
    json_path = REVIEW_DIR / "group-display-label-review.json"
    json_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(f"wrote {csv_path.name} and {json_path.name}: {len(records)} records; "
          f"workbook sha256={wb_sha}")


if __name__ == "__main__":
    main()
