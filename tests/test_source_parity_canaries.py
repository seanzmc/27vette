"""Forced-mismatch canaries for the source-to-contract parity gate.

Spec `docs/superpowers/specs/2026-08-17-fast-layered-validation-suite.md` §4.2:
a parity gate is only worth its runtime if it follows the workbook instead of
matching today's rows. The canaries here make a *valid* authoring edit against a
workbook copy, regenerate, and assert the gate still passes — the failure mode a
green run cannot rule out on its own, because every override row in the tracked
workbook is currently active.

Nothing here touches the canonical workbook or any tracked artifact: the copy,
the regenerated contract, and the snapshot all live under `tmp_path`.
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = REPO_ROOT / "stingray_master.xlsx"
PARITY_GATE = "tests/source-to-contract-parity.test.mjs"
OVERRIDE_SHEET = "stingray_variant_overrides"
MODEL_KEY = "stingray"


def _python() -> str:
    venv = REPO_ROOT / ".venv" / "bin" / "python"
    return str(venv) if venv.exists() else sys.executable


def _deactivate_one_override(workbook_path: Path) -> dict[str, str]:
    """Turn off the first override row that restates section or behavior.

    Returns the authored row, so the test can assert the edit is observable in
    generated output rather than a no-op dressed as a canary.
    """

    wb = load_workbook(workbook_path)
    try:
        ws = wb[OVERRIDE_SHEET]
        headers = [str(ws.cell(1, column).value or "").strip() for column in range(1, ws.max_column + 1)]
        index = {header: position + 1 for position, header in enumerate(headers) if header}
        for column in ("option_id", "variant_id", "section_id", "display_behavior", "active"):
            assert column in index, f"{OVERRIDE_SHEET} has no {column} column"
        for row in range(2, ws.max_row + 1):
            record = {header: str(ws.cell(row, index[header]).value or "").strip() for header in index}
            if not record.get("section_id") and not record.get("display_behavior"):
                continue
            ws.cell(row, index["active"]).value = "False"
            wb.save(workbook_path)
            return record
        pytest.skip(f"{OVERRIDE_SHEET} authors no override that restates section or behavior")
    finally:
        wb.close()


def _generate(workbook_path: Path, output_root: Path) -> None:
    output_root.mkdir(parents=True, exist_ok=True)
    completed = subprocess.run(
        [
            _python(),
            "scripts/generate_form.py",
            "--model",
            MODEL_KEY,
            "--workbook",
            str(workbook_path),
            "--output-root",
            str(output_root),
        ],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
    )
    assert completed.returncode == 0, completed.stderr[-4000:] or completed.stdout[-4000:]


def _build_snapshot(workbook_path: Path, out_path: Path) -> None:
    completed = subprocess.run(
        [_python(), "scripts/build_workbook_truth.py", "--workbook", str(workbook_path), "--out", str(out_path)],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
    )
    assert completed.returncode == 0, completed.stderr[-4000:] or completed.stdout[-4000:]


def _contract_path(snapshot_path: Path) -> str:
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    for row in snapshot["promotions"]["rows"]:
        if row["model_key"] == MODEL_KEY:
            return row["artifact_path"]
    raise AssertionError(f"{MODEL_KEY} is not promoted in the snapshot")


def test_parity_follows_a_deactivated_variant_override(tmp_path: Path) -> None:
    """Deactivating an override is a valid edit, not a parity failure.

    `variant_overrides` carries an `active` column and generation reads the
    sheet through `active_rows`, so a deactivated row restates nothing. A parity
    gate that indexed every override row would keep resolving section and
    display behavior through the dead row and reject correct output.
    """

    # Same basename as the canonical workbook: the contract records the
    # workbook it was built from and the gate checks that provenance, so a
    # renamed copy would fail on the filename rather than on parity.
    workbook = tmp_path / WORKBOOK.name
    shutil.copy2(WORKBOOK, workbook)
    row = _deactivate_one_override(workbook)

    # The gate reads every promoted model's contract, so the candidate root
    # needs all six. Only the Stingray override sheet was edited, so the other
    # five tracked contracts are still this workbook's own output; Stingray is
    # regenerated over its copy.
    output_root = tmp_path / "candidate"
    output_root.mkdir()
    shutil.copytree(REPO_ROOT / "form-output", output_root / "form-output")
    _generate(workbook, output_root)
    snapshot = tmp_path / "workbook-truth.json"
    _build_snapshot(workbook, snapshot)

    # The edit has to bite, or this canary proves nothing. Deactivating the
    # override returns the choice to the option row's own section and display
    # behavior, so the regenerated contract must differ from the tracked one.
    relative = _contract_path(snapshot)
    mutated = json.loads((output_root / relative).read_text(encoding="utf-8"))
    tracked = json.loads((REPO_ROOT / relative).read_text(encoding="utf-8"))
    key = (row["option_id"], row["variant_id"])

    def choice_for(contract: dict) -> dict | None:
        for choice in contract["choices"]:
            if (choice["option_id"], choice["variant_id"]) == key:
                return choice
        return None

    before, after = choice_for(tracked), choice_for(mutated)
    assert before is not None, f"{key} is not in the tracked contract; pick another canary row"
    assert (
        after is None
        or after.get("section_id") != before.get("section_id")
        or after.get("display_behavior") != before.get("display_behavior")
    ), "deactivating the override changed nothing in generated output; the canary is vacuous"

    env = {
        **os.environ,
        "CORVETTE_WORKBOOK_TRUTH": str(snapshot),
        "CORVETTE_CONTRACT_ROOT": str(output_root),
    }
    completed = subprocess.run(
        ["node", "--test", PARITY_GATE], cwd=REPO_ROOT, capture_output=True, text=True, env=env
    )
    assert completed.returncode == 0, completed.stdout[-6000:] or completed.stderr[-6000:]
