"""Checkpoint 5 fixture-helper contract.

Owning spec: docs/superpowers/specs/2026-08-17-fast-layered-validation-suite.md
§6.4 / Checkpoint 5. These tests own clone isolation and compact negative
workbooks. They must not import or promote the canonical workbook; the real
projection/candidate is built only when a Layer 3 owner asks for it.
"""

from __future__ import annotations

import hashlib
import sqlite3
import sys
import tempfile
import unittest
from unittest import mock
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "workbook-manager" / "backend"
SCRIPTS = ROOT / "scripts"
for path in (BACKEND, SCRIPTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from app import db as dbmod  # noqa: E402
from app import importer  # noqa: E402
from workbook_manager_fixtures import (  # noqa: E402
    assert_fixture_unmutated,
    clone_sqlite,
    write_compact_missing_identifier_workbook,
    write_compact_missing_sheet_workbook,
    write_compact_unresolved_reference_workbook,
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _tiny_sqlite(path: Path, marker: str) -> None:
    conn = sqlite3.connect(path)
    try:
        conn.execute("CREATE TABLE marker(value TEXT)")
        conn.execute("INSERT INTO marker(value) VALUES(?)", (marker,))
        conn.commit()
    finally:
        conn.close()


class TestCloneIsolation(unittest.TestCase):
    def test_clone_is_byte_identical_and_mutations_stay_local(self):
        with tempfile.TemporaryDirectory(prefix="wbm-cp5-clone-") as raw:
            root = Path(raw)
            source = root / "source.sqlite3"
            _tiny_sqlite(source, "immutable")
            source_hash = _sha256(source)

            dest = root / "clone.sqlite3"
            returned = clone_sqlite(source, dest)

            self.assertEqual(returned, dest)
            self.assertEqual(_sha256(dest), source_hash)
            dest.write_bytes(dest.read_bytes() + b"dirty")
            self.assertEqual(_sha256(source), source_hash)
            assert_fixture_unmutated(source, source_hash)

class TestCompactNegativeWorkbooks(unittest.TestCase):
    def _empty_projection(self, root: Path) -> Path:
        projection = root / "workbook_projection.sqlite3"
        conn = dbmod.connect(projection)
        try:
            dbmod.init_projection_schema(conn)
            conn.execute(
                "INSERT INTO meta(key, value) VALUES('sentinel', 'prior-projection')"
            )
            conn.commit()
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        finally:
            conn.close()
        return projection

    def test_missing_identifier_blocks_promotion_on_a_compact_workbook(self):
        with tempfile.TemporaryDirectory(prefix="wbm-cp5-missing-id-") as raw:
            root = Path(raw)
            workbook = write_compact_missing_identifier_workbook(
                root / "missing-id.xlsx"
            )
            projection = self._empty_projection(root)
            before = _sha256(projection)

            with mock.patch.object(dbmod, "_replace_projection") as replace:
                result = importer.promote_verified_projection(workbook, projection)

            self.assertFalse(result["promoted"])
            self.assertEqual(result["status"], "blocked")
            self.assertEqual(
                [issue["category"] for issue in result["issues"]],
                ["missing_identifier"],
            )
            replace.assert_not_called()
            self.assertEqual(_sha256(projection), before)
            self.assertFalse(list(root.glob(".wbm-import-candidate-*")))

    def test_unresolved_reference_records_the_offending_value(self):
        with tempfile.TemporaryDirectory(prefix="wbm-cp5-unresolved-") as raw:
            root = Path(raw)
            workbook = write_compact_unresolved_reference_workbook(
                root / "unresolved.xlsx"
            )
            projection = self._empty_projection(root)
            missing_target = "missing_target_for_checkpoint_5"

            result = importer.promote_verified_projection(workbook, projection)

            self.assertFalse(result["promoted"])
            disposition = next(
                row
                for row in result["import"]["managed_row_dispositions"]
                if row["sheet"] == "rule_mapping" and row["src_row"] == 2
            )
            self.assertEqual(disposition["disposition"], "excluded")
            self.assertEqual(disposition["field"], "target_id")
            self.assertEqual(disposition["value"], missing_target)
            self.assertTrue(disposition["blocking"])

    def test_missing_required_sheet_blocks_promotion_on_a_compact_workbook(self):
        with tempfile.TemporaryDirectory(prefix="wbm-cp5-missing-sheet-") as raw:
            root = Path(raw)
            workbook = write_compact_missing_sheet_workbook(
                root / "missing-sheet.xlsx"
            )
            projection = self._empty_projection(root)
            before = _sha256(projection)

            result = importer.promote_verified_projection(workbook, projection)

            self.assertFalse(result["promoted"])
            self.assertEqual(result["status"], "blocked")
            self.assertEqual(_sha256(projection), before)
            self.assertFalse(list(root.glob(".wbm-import-candidate-*")))

    def test_compact_builders_do_not_read_the_canonical_workbook(self):
        canonical = ROOT / "stingray_master.xlsx"
        before = canonical.stat()
        with tempfile.TemporaryDirectory(prefix="wbm-cp5-no-live-") as raw:
            root = Path(raw)
            write_compact_missing_identifier_workbook(root / "a.xlsx")
            write_compact_unresolved_reference_workbook(root / "b.xlsx")
            write_compact_missing_sheet_workbook(root / "c.xlsx")
        after = canonical.stat()
        self.assertEqual(before.st_mtime_ns, after.st_mtime_ns)
        self.assertEqual(before.st_size, after.st_size)


class TestCompactBuildersAreSelfContained(unittest.TestCase):
    def test_missing_identifier_workbook_has_no_live_product_rows(self):
        with tempfile.TemporaryDirectory(prefix="wbm-cp5-shape-") as raw:
            path = write_compact_missing_identifier_workbook(
                Path(raw) / "shape.xlsx"
            )
            workbook = Workbook()
            workbook.close()
            from openpyxl import load_workbook

            loaded = load_workbook(path, read_only=True, data_only=False)
            try:
                self.assertIn("asset_map", loaded.sheetnames)
                self.assertNotIn("stingray_options", loaded.sheetnames)
            finally:
                loaded.close()


if __name__ == "__main__":
    unittest.main()
