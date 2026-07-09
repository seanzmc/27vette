#!/usr/bin/env python3
"""Pass C editor_ops extensions: global sheet families + create_sheet ops.

Fixture workbooks only. Write tests keep schema validation enabled while
mocking a successful schema issue scan for the compact workbook; the live
workbook is never used.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from corvette_form_generator.editor_ops import apply_batch  # noqa: E402
from ingest_wizard_fixtures import build_master_workbook  # noqa: E402


def batch(path: Path, items: list[dict]) -> dict:
    return {"workbookMtimeNs": str(path.stat().st_mtime_ns), "items": items}


class GlobalFamilyOpsTest(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.path = build_master_workbook(Path(self._tmp.name) / "master.xlsx")
        wb = load_workbook(self.path)
        wb["model_workbook_sources"].append(
            ["z06", "interior_source_sheet", "LZ_Interiors", True, "test registration"]
        )
        asset_map = wb.create_sheet("asset_map")
        asset_map.append(
            ["model_key", "target_type", "target_id", "image_url", "image_alt", "image_fit",
             "image_position", "hover_image_url", "hover_image_alt", "hover_image_position", "active", "notes"]
        )
        components = wb.create_sheet("interior_components")
        components.append(
            ["model_key", "interior_id", "rpo", "component_type", "label", "price_ref_type",
             "price_ref_code", "price_trim_scope", "display_order", "active", "notes"]
        )
        wb.save(self.path)
        wb.close()

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def run_batch(self, items: list[dict], *, write: bool = False, confirmed: tuple[str, ...] = ()) -> dict:
        with patch("corvette_form_generator.editor_ops.validate_workbook_schema", return_value=[]):
            return apply_batch(
                self.path,
                batch(self.path, items),
                write=write,
                confirmed_warnings=confirmed,
                run_schema_validation=write,
                log_path=Path(self._tmp.name) / "edit-log.jsonl",
            )

    def test_asset_map_and_interior_components_accept_canonical_operations(self) -> None:
        result = self.run_batch(
            [
                {
                    "action": "add",
                    "sheet": "asset_map",
                    "key": {"model_key": "z06", "target_type": "option", "target_id": "opt_pdb_001"},
                    "row": {
                        "model_key": "z06",
                        "target_type": "option",
                        "target_id": "opt_pdb_001",
                        "image_url": "https://example.test/pdb.png",
                        "active": True,
                    },
                },
                {
                    "action": "add",
                    "sheet": "interior_components",
                    "key": {
                        "model_key": "z06",
                        "interior_id": "1LZ_AQ9_HTA",
                        "rpo": "AQ9",
                        "component_type": "seat",
                    },
                    "row": {
                        "model_key": "z06",
                        "interior_id": "1LZ_AQ9_HTA",
                        "rpo": "AQ9",
                        "component_type": "seat",
                        "label": "GT1 Bucket Seats",
                        "display_order": 1,
                        "active": True,
                    },
                },
            ],
            write=True,
        )

        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path, read_only=True, data_only=True)
        self.assertEqual(wb["asset_map"][2][2].value, "opt_pdb_001")
        self.assertEqual(wb["interior_components"][2][1].value, "1LZ_AQ9_HTA")
        wb.close()

    def test_model_master_add_via_global_family(self) -> None:
        result = self.run_batch(
            [
                {
                    "action": "add",
                    "sheet": "model_master",
                    "key": {"model_key": "grand_sport_x"},
                    "row": {
                        "model_key": "grand_sport_x",
                        "registry_key": "grandSportX",
                        "model_label": "Grand Sport X",
                        "expected_variant_count": 6,
                        "default_model": False,
                        "active": False,
                    },
                }
            ],
            write=True,
        )
        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path, read_only=True)
        rows = list(wb["model_master"].iter_rows(values_only=True))
        wb.close()
        self.assertIn("grand_sport_x", [row[0] for row in rows])

    def test_presentation_family_add(self) -> None:
        result = self.run_batch(
            [
                {
                    "action": "add",
                    "sheet": "runtime_steps",
                    "key": {"model_key": "zr1", "step_key": "paint"},
                    "row": {"model_key": "zr1", "step_key": "paint", "step_label": "Paint", "runtime_order": 1},
                }
            ],
            write=True,
        )
        self.assertTrue(result["ok"], result)

    def test_default_selection_rules_global_family_add(self) -> None:
        result = self.run_batch(
            [
                {
                    "action": "add",
                    "sheet": "default_selection_rules",
                    "key": {"model_key": "z06", "rule_id": "default_pdb"},
                    "row": {
                        "model_key": "z06",
                        "rule_id": "default_pdb",
                        "target_option_id": "opt_pdb_001",
                        "condition_type": "always",
                        "body_style_scope": "*",
                        "trim_level_scope": "*",
                        "variant_scope": "*",
                        "priority": 10,
                        "active": True,
                        "display_behavior": "default_selected",
                    },
                }
            ],
            write=True,
        )
        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path, read_only=True)
        rows = list(wb["default_selection_rules"].iter_rows(values_only=True))
        wb.close()
        self.assertIn("default_pdb", [row[1] for row in rows])

    def test_default_selection_rules_rejects_bad_display_behavior(self) -> None:
        result = self.run_batch(
            [
                {
                    "action": "add",
                    "sheet": "default_selection_rules",
                    "key": {"model_key": "zr1", "rule_id": "bad_default"},
                    "row": {
                        "model_key": "zr1",
                        "rule_id": "bad_default",
                        "target_option_id": "opt_pdb_001",
                        "condition_type": "always",
                        "priority": 10,
                        "active": True,
                        "display_behavior": "default-selected",
                    },
                }
            ]
        )
        self.assertFalse(result["ok"])
        self.assertTrue(any("display_behavior" in error for error in result["errors"]))

    def test_global_family_key_and_type_enforcement(self) -> None:
        result = self.run_batch(
            [
                {
                    "action": "add",
                    "sheet": "model_master",
                    "key": {"model_key": "x"},
                    "row": {"model_key": "x", "expected_variant_count": "not-a-number"},
                }
            ]
        )
        self.assertFalse(result["ok"])
        self.assertTrue(any("expected integer" in e for e in result["errors"]))
        result = self.run_batch(
            [{"action": "update", "sheet": "model_variants", "key": {"model_key": "zr1"}, "row": {"active": True}}]
        )
        self.assertFalse(result["ok"])
        self.assertTrue(any("key must be exactly" in e for e in result["errors"]))

    def test_create_sheet_then_add_rows_in_same_batch(self) -> None:
        items = [
                {
                    "action": "create_sheet",
                    "sheet": "grandSportX_options",
                    "family": "options",
                    "headersFrom": "z06_options",
                },
                # New sheet is not in the model registry yet, so ops carry the
                # created family; a follow-up source-row add registers it.
                {
                    "action": "add",
                    "sheet": "model_workbook_sources",
                    "key": {"model_key": "grand_sport_x", "source_role": "source_option_sheet"},
                    "row": {
                        "model_key": "grand_sport_x",
                        "source_role": "source_option_sheet",
                        "sheet_name": "grandSportX_options",
                        "active": True,
                    },
                },
                {
                    "action": "add",
                    "sheet": "grandSportX_options",
                    "key": {"option_id": "opt_zzz_001"},
                    "row": {
                        "option_id": "opt_zzz_001",
                        "rpo": "ZZZ",
                        "option_name": "Test Option",
                        "section_id": "sec_whee_001",
                        "active": True,
                    },
                },
            ]
        result = self.run_batch(items, write=True)
        self.assertEqual(result["status"], "needs_confirmation", result)
        result = self.run_batch(items, write=True, confirmed=tuple(warning["id"] for warning in result["warnings"]))
        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path, read_only=True)
        self.assertIn("grandSportX_options", wb.sheetnames)
        rows = list(wb["grandSportX_options"].iter_rows(values_only=True))
        wb.close()
        self.assertEqual(rows[0][0], "option_id")
        self.assertEqual(rows[1][0], "opt_zzz_001")

    def test_existing_text_bool_convention_is_preserved_on_write(self) -> None:
        wb = load_workbook(self.path)
        headers = [cell.value for cell in wb["context_section_master"][1]]
        is_required_col = headers.index("is_required") + 1
        active_col = headers.index("active") + 1
        wb["context_section_master"].cell(row=2, column=is_required_col, value="True")
        wb["context_section_master"].cell(row=2, column=active_col, value="True")
        wb.save(self.path)
        wb.close()

        result = self.run_batch(
            [
                {
                    "action": "add",
                    "sheet": "context_section_master",
                    "key": {"model_key": "zr1", "context_type": "body", "section_id": "ctx_body"},
                    "row": {
                        "model_key": "zr1",
                        "context_type": "body",
                        "section_id": "ctx_body",
                        "section_name": "Body",
                        "is_required": True,
                        "active": True,
                    },
                }
            ],
            write=True,
        )

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["boolHygieneResult"]["error_count"], 0)
        wb = load_workbook(self.path, read_only=True, data_only=True)
        rows = list(wb["context_section_master"].iter_rows(values_only=True))
        wb.close()
        added = next(row for row in rows[1:] if row[0] == "zr1")
        self.assertEqual(added[6], "True")
        self.assertEqual(added[11], "True")

    def test_new_sheet_inherits_text_bool_convention_from_template(self) -> None:
        wb = load_workbook(self.path)
        headers = [cell.value for cell in wb["z06_options"][1]]
        selectable_col = headers.index("selectable") + 1
        active_col = headers.index("active") + 1
        for row in range(2, wb["z06_options"].max_row + 1):
            wb["z06_options"].cell(row=row, column=selectable_col, value="True")
            wb["z06_options"].cell(row=row, column=active_col, value="True")
        wb.save(self.path)
        wb.close()

        items = [
            {"action": "create_sheet", "sheet": "grandSportX_options", "family": "options", "headersFrom": "z06_options"},
            {
                "action": "add",
                "sheet": "model_workbook_sources",
                "key": {"model_key": "grand_sport_x", "source_role": "source_option_sheet"},
                "row": {
                    "model_key": "grand_sport_x",
                    "source_role": "source_option_sheet",
                    "sheet_name": "grandSportX_options",
                    "active": True,
                },
            },
            {
                "action": "add",
                "sheet": "grandSportX_options",
                "key": {"option_id": "opt_yyy_001"},
                "row": {
                    "option_id": "opt_yyy_001",
                    "rpo": "YYY",
                    "option_name": "Template Bool Test",
                    "section_id": "sec_whee_001",
                    "selectable": True,
                    "active": True,
                },
            },
        ]
        preview = self.run_batch(items, write=True)
        result = self.run_batch(items, write=True, confirmed=tuple(warning["id"] for warning in preview["warnings"]))

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["boolHygieneResult"]["error_count"], 0)
        wb = load_workbook(self.path, read_only=True, data_only=True)
        rows = list(wb["grandSportX_options"].iter_rows(values_only=True))
        wb.close()
        self.assertEqual(rows[1][7], "True")
        self.assertEqual(rows[1][9], "True")

    def test_pending_registered_existing_sheet_preserves_text_bool_convention(self) -> None:
        wb = load_workbook(self.path)
        ws = wb["zr1_options"]
        source_headers = [cell.value for cell in ws[1]]
        selectable_col = source_headers.index("selectable") + 1
        active_col = source_headers.index("active") + 1
        ws.cell(row=2, column=selectable_col, value="False")
        ws.cell(row=2, column=active_col, value="True")
        wb.save(self.path)
        wb.close()

        items = [
            {
                "action": "update",
                "sheet": "model_workbook_sources",
                "key": {"model_key": "zr1", "source_role": "source_option_sheet"},
                "row": {
                    "sheet_name": "zr1_options",
                    "active": True,
                },
            },
            {
                "action": "add",
                "sheet": "zr1_options",
                "key": {"option_id": "opt_zzz_001"},
                "row": {
                    "option_id": "opt_zzz_001",
                    "rpo": "ZZZ",
                    "option_name": "Pending Registry Bool Test",
                    "section_id": "sec_whee_001",
                    "selectable": False,
                    "active": True,
                },
            },
        ]

        preview = self.run_batch(items, write=True)
        result = self.run_batch(items, write=True, confirmed=tuple(warning["id"] for warning in preview["warnings"]))

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["boolHygieneResult"]["error_count"], 0)
        wb = load_workbook(self.path, read_only=True, data_only=True)
        rows = list(wb["zr1_options"].iter_rows(values_only=True))
        wb.close()
        added = next(row for row in rows[1:] if row[0] == "opt_zzz_001")
        self.assertEqual(added[7], "False")
        self.assertEqual(added[9], "True")

    def test_create_sheet_failures(self) -> None:
        result = self.run_batch(
            [{"action": "create_sheet", "sheet": "z06_options", "family": "options", "headersFrom": "z06_options"}]
        )
        self.assertFalse(result["ok"])
        self.assertTrue(any("already exists" in e for e in result["errors"]))
        result = self.run_batch(
            [{"action": "create_sheet", "sheet": "new_sheet", "family": "nope", "headersFrom": "z06_options"}]
        )
        self.assertFalse(result["ok"])
        result = self.run_batch(
            [{"action": "create_sheet", "sheet": "new_sheet", "family": "options", "headersFrom": "missing"}]
        )
        self.assertFalse(result["ok"])

    def test_dry_run_leaves_file_untouched(self) -> None:
        before = self.path.read_bytes()
        result = self.run_batch(
            [
                {
                    "action": "create_sheet",
                    "sheet": "grandSportX_ovs",
                    "family": "ovs",
                    "headersFrom": "z06_options",
                }
            ]
        )
        # ovs keys aren't in z06_options headers -> template key check fails.
        self.assertFalse(result["ok"])
        self.assertEqual(self.path.read_bytes(), before)


if __name__ == "__main__":
    unittest.main()
