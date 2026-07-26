#!/usr/bin/env python3
"""Contract invariants that must hold for every model once one builder serves all six.

These are requirement-derived, not builder-shaped: they say what a publishable
contract may contain, so they stay meaningful after the Stingray fork is gone.
"""

from __future__ import annotations

import functools
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from corvette_form_generator.model_configs import discover_generation_model_configs
from corvette_form_generator.source_assembly import assemble_model_source

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "stingray_master.xlsx"
MODEL_KEYS = ("stingray", "grand_sport", "grand_sport_x", "z06", "zr1", "zr1x")


@pytest.fixture(scope="module")
def contracts() -> dict[str, dict]:
    with tempfile.TemporaryDirectory() as tmpdir:
        configs = discover_generation_model_configs(WORKBOOK, root=Path(tmpdir))
        return {key: assemble_model_source(configs[key]).runtime_contract for key in MODEL_KEYS}


@functools.cache
def workbook_authored_reasons(model_key: str) -> frozenset[str]:
    """Every disabled_reason string the workbook itself authors for one model."""

    with tempfile.TemporaryDirectory() as tmpdir:
        sheet = discover_generation_model_configs(WORKBOOK, root=Path(tmpdir))[model_key].rule_mapping_sheet
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        rows = workbook[sheet].iter_rows(values_only=True)
        headers = list(next(rows))
        reason = headers.index("disabled_reason")
        return frozenset(str(row[reason]).strip() for row in rows if row[reason])
    finally:
        workbook.close()


def entity_ids(contract: dict) -> set[str]:
    return {row["option_id"] for row in contract["choices"]} | {
        row["interior_id"] for row in contract["interiors"]
    }


@pytest.mark.parametrize("model_key", MODEL_KEYS)
def test_no_rule_references_an_entity_the_contract_does_not_contain(model_key, contracts) -> None:
    """A rule pointing at a deactivated option is dead payload the browser can never fire."""

    contract = contracts[model_key]
    known = entity_ids(contract)
    dangling = [
        (rule["rule_id"], rule["source_id"], rule["target_id"])
        for rule in contract["rules"]
        if rule["source_id"] not in known or rule["target_id"] not in known
    ]

    assert dangling == [], f"{model_key}: {len(dangling)} rules reference absent entities, e.g. {dangling[:3]}"


@pytest.mark.parametrize("model_key", MODEL_KEYS)
def test_no_color_override_references_an_absent_interior(model_key, contracts) -> None:
    contract = contracts[model_key]
    interiors = {row["interior_id"] for row in contract["interiors"]}
    dangling = sorted(
        {row["interior_id"] for row in contract["colorOverrides"] if row["interior_id"] not in interiors}
    )

    assert dangling == [], f"{model_key}: colorOverrides reference interiors outside the model: {dangling}"


@pytest.mark.parametrize("model_key", MODEL_KEYS)
def test_every_section_holds_at_least_one_row(model_key, contracts) -> None:
    """An empty section renders nothing; another model's section is never this model's."""

    contract = contracts[model_key]
    populated = {row["section_id"] for row in contract["choices"]}
    populated |= {row["section_id"] for row in contract["standardEquipment"]}
    populated |= {row["section_id"] for row in contract["contextChoices"]}
    empty = sorted({row["section_id"] for row in contract["sections"]} - populated)

    assert empty == [], f"{model_key}: {len(empty)} sections carry no rows: {empty}"


@pytest.mark.parametrize("model_key", MODEL_KEYS)
def test_display_behavior_is_omitted_rather_than_blank(model_key, contracts) -> None:
    """Least-change decision 2026-07-26: carry the key only when it has a value."""

    blank = [
        row["choice_id"]
        for row in contracts[model_key]["choices"]
        if "display_behavior" in row and not str(row["display_behavior"]).strip()
    ]

    assert blank == [], f"{model_key}: {len(blank)} choices carry a blank display_behavior"


@pytest.mark.parametrize("model_key", MODEL_KEYS)
def test_no_composed_reason_leaks_an_internal_interior_id(model_key, contracts) -> None:
    """Copy the GENERATOR composes names an interior the way the browser does.

    Workbook-authored reasons are exempt: the workbook is the authority, so a bad
    string there is a workbook defect to report, not something to rewrite in code.
    """

    contract = contracts[model_key]
    interior_ids = {row["interior_id"] for row in contract["interiors"]}
    authored = workbook_authored_reasons(model_key)
    leaks = [
        (rule["rule_id"], rule["disabled_reason"])
        for rule in contract["rules"]
        if str(rule.get("disabled_reason", "")) not in authored
        and any(interior_id in str(rule.get("disabled_reason", "")) for interior_id in interior_ids)
    ]

    assert leaks == [], f"{model_key}: {len(leaks)} composed reasons contain a raw interior id, e.g. {leaks[:2]}"


def test_one_builder_serves_every_model() -> None:
    """No model-keyed fork survives in the source-assembly facade."""

    source = (ROOT / "scripts" / "corvette_form_generator" / "source_assembly.py").read_text()

    assert 'config.model_key == "stingray"' not in source
    assert "build_production_source_data" not in source
