#!/usr/bin/env python3
"""Tests for promoted-model runtime metadata fallback guards."""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.runtime_metadata import (  # noqa: E402
    derived_default_selected_display_behavior,
    load_context_sections,
    load_default_selection_display_rules,
    load_default_selection_rules,
    load_order_summary_metadata,
    load_runtime_steps,
)

def _raw_sheet_rows(wb: object, sheet_name: str) -> list[dict[str, object]]:
    """Cells of one sheet, read directly.

    The independent side of a parity assertion (spec §4.2): openpyxl only, no
    `corvette_form_generator` call, so the expected side cannot be produced by
    the code under test.
    """

    ws = wb[sheet_name]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, object]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = {header: value for header, value in zip(headers, raw) if header}
        if any(value is not None for value in record.values()):
            rows.append(record)
    return rows


PROMOTION_HEADERS = [
    "model_key",
    "registry_key",
    "promoted_to_runtime",
    "default_model",
    "artifact_path",
    "artifact_type",
    "legacy_alias",
    "active",
    "display_order",
    "notes",
]
RUNTIME_STEP_HEADERS = ["model_key", "step_key", "step_label", "runtime_order", "source", "active", "notes"]
CONTEXT_SECTION_HEADERS = [
    "model_key",
    "context_type",
    "section_id",
    "section_name",
    "selection_mode",
    "choice_mode",
    "is_required",
    "standard_behavior",
    "section_display_order",
    "step_key",
    "step_label",
    "active",
    "notes",
]
ORDER_SUMMARY_HEADERS = ["model_key", "section_key", "section_label", "display_order", "active", "notes"]
STEP_SUMMARY_HEADERS = ["model_key", "step_key", "section_key", "active", "notes"]
DEFAULT_SELECTION_HEADERS = [
    "model_key",
    "rule_id",
    "target_option_id",
    "condition_type",
    "condition_id",
    "body_style_scope",
    "trim_level_scope",
    "variant_scope",
    "priority",
    "active",
    "display_behavior",
    "notes",
]


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows or []:
        ws.append([row.get(header, None) for header in headers])


def metadata_workbook(
    *,
    promoted: bool,
    runtime_rows: list[dict[str, object]] | None = None,
    context_rows: list[dict[str, object]] | None = None,
) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "model_registry_promotion",
        PROMOTION_HEADERS,
        [
            {
                "model_key": "z06",
                "registry_key": "z06",
                "promoted_to_runtime": promoted,
                "default_model": False,
                "artifact_path": "form-output/runtime/z06-runtime-contract.json",
                "artifact_type": "runtime_contract",
                "active": True,
                "display_order": 3,
            }
        ],
    )
    append_sheet(wb, "runtime_steps", RUNTIME_STEP_HEADERS, runtime_rows or [])
    append_sheet(wb, "context_section_master", CONTEXT_SECTION_HEADERS, context_rows or [])
    append_sheet(wb, "order_summary_sections", ORDER_SUMMARY_HEADERS, [])
    append_sheet(wb, "step_order_summary_map", STEP_SUMMARY_HEADERS, [])
    append_sheet(wb, "default_selection_rules", DEFAULT_SELECTION_HEADERS, [])
    return wb


def default_selection_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "default_selection_rules",
        DEFAULT_SELECTION_HEADERS,
        [
            {
                "model_key": "stingray",
                "rule_id": "default_bc7",
                "target_option_id": "opt_bc7_001",
                "condition_type": "always",
                "body_style_scope": "coupe",
                "trim_level_scope": "*",
                "variant_scope": "*",
                "priority": 40,
                "active": True,
                "display_behavior": "default_selected",
                "notes": "derive display metadata",
            },
            {
                "model_key": "stingray",
                "rule_id": "default_nga",
                "target_option_id": "opt_nga_001",
                "condition_type": "unless_selected_rpo",
                "body_style_scope": "*",
                "trim_level_scope": "*",
                "variant_scope": "*",
                "priority": 20,
                "active": True,
                "display_behavior": "",
                "notes": "runtime default only",
            },
            {
                "model_key": "z06",
                "rule_id": "z06_default_bc7",
                "target_option_id": "opt_bc7_001",
                "condition_type": "always",
                "body_style_scope": "coupe",
                "trim_level_scope": "*",
                "variant_scope": "*",
                "priority": 10,
                "active": True,
                "display_behavior": "default_selected",
                "notes": "other model",
            },
        ],
    )
    return wb


class RuntimeMetadataGuardTests(unittest.TestCase):
    def test_unpromoted_models_cannot_use_a_runtime_step_fallback_either(self) -> None:
        """No Python fallback for any active model — promotion is irrelevant."""

        wb = metadata_workbook(promoted=False)

        with self.assertRaisesRegex(ValueError, "requires active workbook-owned runtime_steps rows"):
            load_runtime_steps(wb, "z06")

    def test_promoted_models_cannot_use_runtime_step_fallback(self) -> None:
        wb = metadata_workbook(promoted=True)

        with self.assertRaisesRegex(ValueError, "requires active workbook-owned runtime_steps rows"):
            load_runtime_steps(wb, "z06")

    def test_incomplete_runtime_steps_fail_against_workbook_referenced_steps(self) -> None:
        """Completeness is measured against the workbook's own section metadata."""

        wb = metadata_workbook(
            promoted=True,
            runtime_rows=[
                {
                    "model_key": "z06",
                    "step_key": "body_style",
                    "step_label": "Body Style",
                    "runtime_order": 1,
                    "source": "test",
                    "active": True,
                }
            ],
            context_rows=[
                {
                    "model_key": "z06",
                    "context_type": "trim_level",
                    "section_id": "sec_context_trim_level",
                    "section_name": "Trim Level",
                    "step_key": "trim_level",
                    "active": True,
                }
            ],
        )

        with self.assertRaisesRegex(ValueError, "missing step_key values: trim_level"):
            load_runtime_steps(wb, "z06")

    def test_promoted_models_cannot_use_context_section_fallback(self) -> None:
        wb = metadata_workbook(promoted=True)

        with self.assertRaisesRegex(ValueError, "requires active workbook-owned context_section_master rows"):
            load_context_sections(wb, "z06")

    def test_promoted_models_cannot_use_browser_order_summary_fallback(self) -> None:
        wb = metadata_workbook(promoted=True)

        with self.assertRaisesRegex(ValueError, "requires workbook-owned order_summary_sections and step_order_summary_map rows"):
            load_order_summary_metadata(wb, "z06")

    def test_default_selection_rules_strip_workbook_only_display_behavior(self) -> None:
        rows = load_default_selection_rules(default_selection_workbook(), "stingray")

        self.assertEqual([row["rule_id"] for row in rows], ["default_nga", "default_bc7"])
        self.assertTrue(all("display_behavior" not in row for row in rows))

    def test_default_selection_display_rules_load_only_workbook_flagged_rows(self) -> None:
        rows = load_default_selection_display_rules(default_selection_workbook(), "stingray")

        self.assertEqual([row["rule_id"] for row in rows], ["default_bc7"])
        self.assertEqual(rows[0]["display_behavior"], "default_selected")

    def test_default_selected_display_derives_from_workbook_flagged_rule(self) -> None:
        choice = {
            "option_id": "opt_bc7_001",
            "status": "standard",
            "selectable": "True",
            "active": "True",
            "body_style": "coupe",
            "trim_level": "1LT",
            "variant_id": "1lt_c07",
        }
        groups = [
            {
                "active": "True",
                "selection_mode": "single_within_group",
                "option_ids": ["opt_bc7_001", "opt_bcs_001"],
            }
        ]

        self.assertTrue(
            derived_default_selected_display_behavior(
                choice,
                "stingray",
                load_default_selection_display_rules(default_selection_workbook(), "stingray"),
                groups,
            )
        )

    def test_default_selection_rule_without_display_behavior_does_not_derive(self) -> None:
        choice = {
            "option_id": "opt_nga_001",
            "status": "standard",
            "selectable": "True",
            "active": "True",
            "body_style": "coupe",
            "trim_level": "1LT",
            "variant_id": "1lt_c07",
        }
        groups = [
            {
                "active": "True",
                "selection_mode": "single_within_group",
                "option_ids": ["opt_nga_001", "opt_nwi_001"],
            }
        ]

        self.assertFalse(
            derived_default_selected_display_behavior(
                choice,
                "stingray",
                load_default_selection_rules(default_selection_workbook(), "stingray"),
                groups,
            )
        )

    def test_runtime_metadata_has_no_default_selected_rule_id_allowlist(self) -> None:
        source = (ROOT / "scripts" / "corvette_form_generator" / "runtime_metadata.py").read_text()

        self.assertNotIn("_DEFAULT_SELECTED_DISPLAY_RULE_IDS_BY_MODEL", source)

    def test_live_workbook_default_selection_display_behavior_rows_are_explicit(self) -> None:
        """Every workbook-authored `default_selected` row reaches the loader, and no other.

        Checkpoint 1 of the fast layered validation suite (spec §9) replaced the
        hardcoded three-row expectation this test carried over a hardcoded
        three-model tuple. The workbook has since authored a fourth valid row
        and promotes six models, so the literal failed for a valid data state
        while the three-model sweep silently covered half the active models.
        The expected side is now a direct read of the same sheet's cells — no
        `runtime_metadata` call — so a valid new row extends the comparison
        instead of breaking it, and a loader that drops, invents, or misroutes a
        row still fails.
        """

        wb = load_workbook(ROOT / "stingray_master.xlsx", read_only=True, data_only=True)
        try:
            model_keys = sorted(
                str(row["model_key"]).strip()
                for row in _raw_sheet_rows(wb, "model_master")
                if str(row.get("active", "")).strip().lower() in {"true", "yes", "1", "y"}
            )
            expected = sorted(
                (
                    str(row["model_key"]).strip(),
                    str(row["rule_id"]).strip(),
                    str(row["display_behavior"]).strip(),
                )
                for row in _raw_sheet_rows(wb, "default_selection_rules")
                if str(row.get("active", "")).strip().lower() in {"true", "yes", "1", "y"}
                and str(row.get("display_behavior", "")).strip() == "default_selected"
                and str(row.get("rule_id", "")).strip()
            )
            actual = sorted(
                (model_key, row["rule_id"], row["display_behavior"])
                for model_key in model_keys
                for row in load_default_selection_display_rules(wb, model_key)
            )
        finally:
            wb.close()

        # Guard the sweep: an empty expected side would make the comparison
        # vacuous, and every model must be swept, not a named subset.
        self.assertGreaterEqual(len(model_keys), 2)
        self.assertTrue(expected, "no active default_selected row in default_selection_rules")
        self.assertEqual(actual, expected)


if __name__ == "__main__":
    unittest.main()
