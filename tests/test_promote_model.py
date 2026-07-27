#!/usr/bin/env python3
"""Tests for model promotion parity with generator discovery metadata."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_configs import REQUIRED_GENERATION_SOURCE_ROLES  # noqa: E402
from corvette_form_generator.workbook import rows_from_sheet  # noqa: E402
import promote_model as promotion_module  # noqa: E402
from promote_model import model_promotion_plan, promote_model, verify_workbook  # noqa: E402


def append_sheet(
    wb: Workbook,
    name: str,
    headers: list[str],
    rows: list[dict[str, object]],
) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def promotion_workbook(*, missing_source_role: str | None = None) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "model_master",
        [
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "notes",
        ],
        [
            {
                "model_key": "zr1",
                "registry_key": "zr1",
                "model_label": "ZR1",
                "model_year": 2027,
                "dataset_name": "ZR1 scaffold",
                "export_slug": "zr1",
                "expected_variant_count": 2,
                "default_model": False,
                "active": False,
            }
        ],
    )
    append_sheet(
        wb,
        "model_registry_promotion",
        [
            "model_key",
            "registry_key",
            "promoted_to_runtime",
            "default_model",
            "artifact_path",
            "artifact_type",
            "legacy_alias",
            "active",
            "display_order",
            "notes",
        ],
        [
            {
                "model_key": "zr1",
                "registry_key": "zr1",
                "promoted_to_runtime": False,
                "default_model": False,
                "active": False,
                "display_order": 4,
            }
        ],
    )
    append_sheet(
        wb,
        "variant_master",
        [
            "variant_id",
            "model_year",
            "trim_level",
            "body_style",
            "display_name",
            "base_price",
            "display_order",
            "active",
        ],
        [
            {
                "variant_id": "1lz_r07",
                "model_year": 2027,
                "trim_level": "1LZ",
                "body_style": "coupe",
                "display_name": "ZR1 Coupe 1LZ",
                "base_price": 197195,
                "display_order": 1,
                "active": False,
            },
            {
                "variant_id": "3lz_r67",
                "model_year": 2027,
                "trim_level": "3LZ",
                "body_style": "convertible",
                "display_name": "ZR1 Convertible 3LZ",
                "base_price": 218195,
                "display_order": 2,
                "active": False,
            },
            {
                "variant_id": "1lz_s07",
                "model_year": 2027,
                "trim_level": "1LZ",
                "body_style": "coupe",
                "display_name": "ZR1X Coupe 1LZ",
                "base_price": 227395,
                "display_order": 3,
                "active": False,
            },
        ],
    )
    append_sheet(
        wb,
        "model_variants",
        ["model_key", "variant_id", "display_order", "active", "notes"],
        [
            {"model_key": "zr1", "variant_id": "1lz_r07", "display_order": 1, "active": False},
            {"model_key": "zr1", "variant_id": "3lz_r67", "display_order": 2, "active": False},
            {"model_key": "zr1x", "variant_id": "1lz_s07", "display_order": 1, "active": False},
        ],
    )
    source_rows = [
        {
            "model_key": "zr1",
            "source_role": role,
            "sheet_name": f"zr1_{role}",
            "active": False,
        }
        for role in REQUIRED_GENERATION_SOURCE_ROLES
        if role != missing_source_role
    ]
    append_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        source_rows,
    )
    return wb


def save_promoted_workbook(path: Path, *, missing_source_role: str | None = None) -> dict[str, object]:
    wb = promotion_workbook(missing_source_role=missing_source_role)
    plan = model_promotion_plan(wb, "zr1")
    promote_model(wb, "zr1", plan)
    wb.save(path)
    wb.close()
    return plan


def test_promotion_activates_exact_target_memberships_and_preserves_unrelated_rows() -> None:
    wb = promotion_workbook()
    plan = model_promotion_plan(wb, "zr1")

    promote_model(wb, "zr1", plan)

    memberships = {
        (row["model_key"], row["variant_id"]): row["active"]
        for row in rows_from_sheet(wb, "model_variants")
    }
    variants = {row["variant_id"]: row["active"] for row in rows_from_sheet(wb, "variant_master")}
    assert memberships == {
        ("zr1", "1lz_r07"): "True",
        ("zr1", "3lz_r67"): "True",
        ("zr1x", "1lz_s07"): "False",
    }
    assert variants == {"1lz_r07": "True", "3lz_r67": "True", "1lz_s07": "False"}
    wb.close()


def test_promotion_activates_every_existing_generation_source_row() -> None:
    wb = promotion_workbook()
    plan = model_promotion_plan(wb, "zr1")

    changes = promote_model(wb, "zr1", plan)

    source_rows = [
        row
        for row in rows_from_sheet(wb, "model_workbook_sources")
        if row["model_key"] == "zr1"
    ]
    assert len(source_rows) == len(REQUIRED_GENERATION_SOURCE_ROLES)
    assert all(row["active"] == "True" for row in source_rows)
    assert sum(change["sheet"] == "model_workbook_sources" for change in changes) == len(source_rows)
    wb.close()


def test_promotion_is_idempotent() -> None:
    wb = promotion_workbook()
    plan = model_promotion_plan(wb, "zr1")

    first_changes = promote_model(wb, "zr1", plan)
    second_changes = promote_model(wb, "zr1", plan)

    assert first_changes
    assert second_changes == []
    wb.close()


def test_verify_workbook_proves_target_memberships_and_generator_discovery(tmp_path: Path) -> None:
    path = tmp_path / "complete-promotion.xlsx"
    plan = save_promoted_workbook(path)

    verification = verify_workbook(path, "zr1", plan)

    assert verification["model_variant_active"] == {"1lz_r07": "True", "3lz_r67": "True"}
    assert verification["discovery"] == {
        "ok": True,
        "model_key": "zr1",
        "expected_variant_ids": ["1lz_r07", "3lz_r67"],
        "discovered_model_keys": ["zr1"],
        "variant_ids": ["1lz_r07", "3lz_r67"],
        "error": None,
    }
    assert verification["failures"] == []


def test_verify_workbook_fails_when_a_target_membership_is_inactive(tmp_path: Path) -> None:
    path = tmp_path / "inactive-membership.xlsx"
    plan = save_promoted_workbook(path)
    wb = load_workbook(path)
    ws = wb["model_variants"]
    ws.cell(3, 4).value = False
    wb.save(path)
    wb.close()

    verification = verify_workbook(path, "zr1", plan)

    assert verification["model_variant_active"]["3lz_r67"] == "False"
    assert any("model_variants zr1/3lz_r67 active expected 'True'" in item for item in verification["failures"])
    assert verification["discovery"]["ok"] is False
    assert "expected 2 active model_variants rows; found 1" in verification["discovery"]["error"]["message"]


def test_verify_workbook_captures_incomplete_source_metadata_as_structured_evidence(tmp_path: Path) -> None:
    path = tmp_path / "incomplete-sources.xlsx"
    plan = save_promoted_workbook(path, missing_source_role="status_sheet")

    verification = verify_workbook(path, "zr1", plan)

    assert verification["discovery"]["ok"] is False
    assert verification["discovery"]["error"]["type"] == "ValueError"
    assert "missing required active model_workbook_sources roles: status_sheet" in verification["discovery"]["error"]["message"]
    assert any("generator discovery failed" in item for item in verification["failures"])


def test_atomic_write_preflight_failure_leaves_canonical_file_unchanged(tmp_path: Path) -> None:
    path = tmp_path / "incomplete-promotion.xlsx"
    wb = promotion_workbook(missing_source_role="status_sheet")
    wb.save(path)
    wb.close()
    before = path.read_bytes()

    execute = getattr(promotion_module, "execute_promotion", None)
    assert callable(execute), "promotion command needs an atomic preflight API"
    result = execute(path, ["zr1"], write=True)

    assert result["ok"] is False
    assert result["status"] == "preflight_failed"
    assert path.read_bytes() == before
    assert "backup_path" not in result


def test_post_save_verification_failure_restores_exact_original_workbook(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "post-save-failure.xlsx"
    wb = promotion_workbook()
    wb.save(path)
    wb.close()
    before = path.read_bytes()
    real_verify = promotion_module.verify_promotions
    call_count = 0

    def fail_second_verification(candidate_path, plans):
        nonlocal call_count
        call_count += 1
        if call_count == 1:
            return real_verify(candidate_path, plans)
        return {"ok": False, "models": {}, "failures": ["forced post-save failure"]}

    monkeypatch.setattr(promotion_module, "verify_promotions", fail_second_verification)

    # The candidate lane is exercised by tests/test_verify_workbook_candidate.py
    # and by test_candidate_lane_gates_promotion below; this fixture workbook is
    # a minimal promotion scaffold, not a generatable one.
    result = promotion_module.execute_promotion(path, ["zr1"], write=True, run_candidate_lane=False)

    assert result["ok"] is False
    assert result["status"] == "post_save_verification_failed_restored"
    assert result["restored"] is True
    assert path.read_bytes() == before


def test_cli_routes_repeated_models_through_one_atomic_execution(tmp_path: Path, monkeypatch, capsys) -> None:
    path = tmp_path / "three-model-promotion.xlsx"
    calls = []

    def execute(workbook_path, model_keys, *, write=False):
        calls.append((Path(workbook_path), model_keys, write))
        return {"ok": True, "status": "validated", "model_keys": model_keys, "write": write}

    monkeypatch.setattr(promotion_module, "execute_promotion", execute)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "promote_model.py",
            "--workbook",
            str(path),
            "--model",
            "grand_sport_x",
            "--model",
            "zr1",
            "--model",
            "zr1x",
        ],
    )

    promotion_module.main()

    assert calls == [(path, ["grand_sport_x", "zr1", "zr1x"], False)]
    assert json.loads(capsys.readouterr().out)["status"] == "validated"


def test_candidate_lane_gates_promotion(tmp_path: Path, monkeypatch) -> None:
    """Spec Pass 3 requirements 1-5: promotion cannot report success without the lane.

    Breaks if `execute_promotion` ever reaches the canonical write while the
    composed candidate lane reports a failure — the exact "validated without
    generating or runtime-testing the candidate" shape §2.1 item 2 recorded.
    """

    path = tmp_path / "candidate-lane-gate.xlsx"
    wb = promotion_workbook()
    wb.save(path)
    wb.close()
    before = path.read_bytes()
    seen: list[dict] = []

    def failing_lane(candidate, *, changed_models=None, **kwargs):
        seen.append({"candidate": Path(candidate), "changed_models": list(changed_models or [])})
        return {"ok": False, "failedStage": "generate_models", "models": {}}

    monkeypatch.setattr(promotion_module, "verify_candidate", failing_lane)

    result = promotion_module.execute_promotion(path, ["zr1"], write=True)

    assert result["ok"] is False
    assert result["status"] == "candidate_lane_failed"
    assert result["candidate_readiness"]["failedStage"] == "generate_models"
    assert path.read_bytes() == before, "canonical workbook was written despite a failing candidate lane"
    assert "backup_path" not in result
    # The lane must see the post-promotion scratch workbook, not the canonical one,
    # and must be told which models the promotion touched.
    assert len(seen) == 1
    assert seen[0]["candidate"] != path
    assert seen[0]["changed_models"] == ["zr1"]
