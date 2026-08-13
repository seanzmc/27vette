#!/usr/bin/env python3
"""Focused tests for the asset_map sync maintenance command."""

from __future__ import annotations

import subprocess
import sys
import json
import csv
from email.message import Message
from pathlib import Path
from urllib.error import HTTPError

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator import asset_map_sync  # noqa: E402


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def make_discovery_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    add_sheet(
        wb,
        "model_registry_promotion",
        ["model_key", "promoted_to_runtime", "active", "display_order"],
        [
            {"model_key": "stingray", "promoted_to_runtime": True, "active": True, "display_order": 1},
            {"model_key": "grand_sport", "promoted_to_runtime": True, "active": True, "display_order": 2},
            {"model_key": "zr1", "promoted_to_runtime": False, "active": False, "display_order": 3},
        ],
    )
    add_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active"],
        [
            {"model_key": "stingray", "source_role": "source_option_sheet", "sheet_name": "stingray_options", "active": True},
            {"model_key": "grand_sport", "source_role": "source_option_sheet", "sheet_name": "grandSport_options", "active": True},
            {"model_key": "zr1", "source_role": "source_option_sheet", "sheet_name": "zr1_options", "active": True},
        ],
    )
    option_headers = ["option_id", "rpo", "option_name", "active", "selectable"]
    add_sheet(
        wb,
        "stingray_options",
        option_headers,
        [
            {"option_id": "opt_gba_001", "rpo": "GBA", "option_name": "Black", "active": True, "selectable": True},
            {"option_id": "opt_hidden_001", "rpo": "ZZZ", "option_name": "Hidden", "active": False, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "grandSport_options",
        option_headers,
        [
            {"option_id": "opt_qe6_001", "rpo": "QE6", "option_name": "Wheel", "active": True, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "zr1_options",
        option_headers,
        [
            {"option_id": "opt_future_001", "rpo": "FUT", "option_name": "Future", "active": True, "selectable": True},
        ],
    )
    return wb


def make_apply_workbook(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    add_sheet(
        wb,
        "model_registry_promotion",
        ["model_key", "promoted_to_runtime", "active", "display_order"],
        [{"model_key": "stingray", "promoted_to_runtime": True, "active": True, "display_order": 1}],
    )
    add_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active"],
        [{"model_key": "stingray", "source_role": "source_option_sheet", "sheet_name": "stingray_options", "active": True}],
    )
    add_sheet(
        wb,
        "stingray_options",
        ["option_id", "rpo", "option_name", "active", "selectable"],
        [{"option_id": "opt_gba_001", "rpo": "GBA", "option_name": "Black", "active": True, "selectable": True}],
    )
    add_sheet(
        wb,
        "asset_map",
        ["model_key", "target_type", "target_id", "image_url", "image_alt", "image_fit", "image_position", "active", "notes"],
        [],
    )
    wb.save(path)
    wb.close()


class FakeJsonResponse:
    def __init__(self, payload: bytes = b"[]", headers: dict[str, str] | None = None) -> None:
        self.payload = payload
        self.headers = headers or {}

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback) -> None:
        return None

    def read(self) -> bytes:
        return self.payload


def make_missing_report_workbook(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    add_sheet(
        wb,
        "model_registry_promotion",
        ["model_key", "promoted_to_runtime", "active", "display_order"],
        [
            {"model_key": "stingray", "promoted_to_runtime": True, "active": True, "display_order": 1},
            {"model_key": "grand_sport", "promoted_to_runtime": True, "active": True, "display_order": 2},
        ],
    )
    add_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active"],
        [
            {"model_key": "stingray", "source_role": "source_option_sheet", "sheet_name": "stingray_options", "active": True},
            {"model_key": "grand_sport", "source_role": "source_option_sheet", "sheet_name": "grandSport_options", "active": True},
        ],
    )
    option_headers = ["option_id", "rpo", "option_name", "section_id", "active", "selectable"]
    add_sheet(
        wb,
        "stingray_options",
        option_headers,
        [
            {"option_id": "opt_gba_001", "rpo": "GBA", "option_name": "Black", "section_id": "sec_paint_001", "active": True, "selectable": True},
            {"option_id": "opt_noimg_001", "rpo": "NIX", "option_name": "No Image", "section_id": "sec_test_001", "active": True, "selectable": True},
            {"option_id": "opt_stx_001", "rpo": "STX", "option_name": "Stripe", "section_id": "sec_stripe_001", "active": True, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "grandSport_options",
        option_headers,
        [
            {"option_id": "opt_gba_002", "rpo": "GBA", "option_name": "Black GS", "section_id": "sec_paint_001", "active": True, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "asset_map",
        ["model_key", "target_type", "target_id", "image_url", "image_alt", "image_fit", "image_position", "active", "notes"],
        [
            {
                "model_key": "stingray",
                "target_type": "option",
                "target_id": "opt_gba_001",
                "image_url": "https://example.test/current-gba.png",
                "image_alt": "Black",
                "image_fit": "cover",
                "image_position": "center",
                "active": True,
                "notes": "existing",
            },
        ],
    )
    wb.save(path)
    wb.close()


def test_open_json_sends_browser_like_user_agent(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, str | None] = {}

    def fake_urlopen(request, timeout):
        captured["accept"] = request.get_header("Accept")
        captured["cache_control"] = request.get_header("Cache-control")
        captured["pragma"] = request.get_header("Pragma")
        captured["user_agent"] = request.get_header("User-agent")
        return FakeJsonResponse(b"[]", {"X-WP-TotalPages": "1"})

    monkeypatch.setattr(asset_map_sync, "urlopen", fake_urlopen)

    payload, headers = asset_map_sync._open_json("https://example.test/wp-json/wp/v2/media", auth_header=None, timeout=1)

    assert payload == []
    assert headers == {"x-wp-totalpages": "1"}
    assert captured["accept"] == "application/json"
    assert captured["cache_control"] == "no-cache, no-store, max-age=0"
    assert captured["pragma"] == "no-cache"
    assert captured["user_agent"]
    assert "Mozilla/5.0" in captured["user_agent"]


def test_open_json_preserves_optional_basic_auth_header(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, str | None] = {}
    monkeypatch.setenv("WP_USER", "media-user")
    monkeypatch.setenv("WP_APP_PASSWORD", "app pass")

    def fake_urlopen(request, timeout):
        captured["authorization"] = request.get_header("Authorization")
        captured["user_agent"] = request.get_header("User-agent")
        return FakeJsonResponse()

    monkeypatch.setattr(asset_map_sync, "urlopen", fake_urlopen)

    asset_map_sync._open_json(
        "https://example.test/wp-json/wp/v2/media",
        auth_header=asset_map_sync._auth_header_from_env(),
        timeout=1,
    )

    assert captured["authorization"]
    assert captured["authorization"].startswith("Basic ")
    assert captured["user_agent"]


def test_fetch_media_403_mentions_media_url_list_fallback(monkeypatch: pytest.MonkeyPatch) -> None:
    def blocked(url, *, auth_header, timeout):
        raise HTTPError(url, 403, "Forbidden", Message(), None)

    monkeypatch.setattr(asset_map_sync, "_open_json", blocked)

    with pytest.raises(Exception) as excinfo:
        asset_map_sync.fetch_media(timeout=1)

    message = str(excinfo.value)
    assert "HTTP 403" in message
    assert "--media-url-list" in message


def test_stable_media_fetch_requires_two_identical_uncached_snapshots(monkeypatch: pytest.MonkeyPatch) -> None:
    snapshots = iter(
        [
            asset_map_sync.MediaSnapshot(["https://example.test/27vette/a.png"], {"https://example.test/27vette/a.png": "1"}),
            asset_map_sync.MediaSnapshot(["https://example.test/27vette/a.png", "https://example.test/27vette/b.png"], {"https://example.test/27vette/a.png": "1", "https://example.test/27vette/b.png": "2"}),
            asset_map_sync.MediaSnapshot(["https://example.test/27vette/b.png", "https://example.test/27vette/a.png"], {"https://example.test/27vette/a.png": "1", "https://example.test/27vette/b.png": "2"}),
        ]
    )
    tokens: list[str | None] = []

    def fake_fetch(timeout, modified_after=None, *, cache_token=None):
        tokens.append(cache_token)
        return next(snapshots)

    monkeypatch.setattr(asset_map_sync, "fetch_media_snapshot", fake_fetch)
    monkeypatch.setattr(asset_map_sync.time, "sleep", lambda _seconds: None)

    snapshot = asset_map_sync.fetch_media_stable(timeout=1, attempts=4)

    assert len(snapshot.urls) == 2
    assert len(tokens) == 3
    assert all(tokens)
    assert len(set(tokens)) == 3


def test_revision_tokens_only_change_after_a_known_modified_time_changes() -> None:
    url = "https://example.test/wp-content/uploads/pictures/27vette/brakes/r-j6d.webp"
    snapshot = asset_map_sync.MediaSnapshot([url], {url: "2026-08-13T12:00:00"})

    baseline_urls, baseline_tokens = asset_map_sync.prepare_revisioned_media_urls(snapshot, {})
    revised_urls, revised_tokens = asset_map_sync.prepare_revisioned_media_urls(
        snapshot,
        {"media_modified": {url: "2026-08-13T11:00:00"}, "revision_tokens": {}},
    )
    repeated_urls, repeated_tokens = asset_map_sync.prepare_revisioned_media_urls(
        snapshot,
        {"media_modified": {url: "2026-08-13T12:00:00"}, "revision_tokens": revised_tokens},
    )

    assert baseline_urls == [url]
    assert baseline_tokens == {}
    assert revised_urls == [f"{url}?asset_rev=20260813T120000"]
    assert revised_tokens == {url: "20260813T120000"}
    assert repeated_urls == revised_urls
    assert repeated_tokens == revised_tokens


def test_parse_media_requires_hyphen_for_model_prefix() -> None:
    assert asset_map_sync.parse_media("https://example.test/imgi_47_379.png") == (None, "379", True)
    assert asset_map_sync.parse_media("https://example.test/h-stx.png") == ("z06", "stx", True)
    assert asset_map_sync.parse_media("https://example.test/hzp.png") == (None, "hzp", True)
    assert asset_map_sync.parse_media("https://example.test/c-qe6_v1.png") == ("stingray", "qe6", True)
    assert asset_map_sync.parse_media("https://example.test/27vette/paint/gba.png") == (None, "gba", True)
    assert asset_map_sync.parse_media("https://example.test/27vette/paint/c-gba.png") == ("stingray", "gba", True)


def test_parse_shared_option_media_accepts_any_valid_multi_model_prefix_and_suffix() -> None:
    assert asset_map_sync.parse_shared_option_media(
        "https://example.test/27vette/brakes/e-g-j6d-o-cmp.webp"
    ) == ("e-g", "j6d", True)
    assert asset_map_sync.parse_shared_option_media(
        "https://example.test/27vette/brakes/h-s-r-j6d-o-cmp.webp"
    ) == ("h-s-r", "j6d", True)
    assert asset_map_sync.parse_shared_option_media(
        "https://example.test/27vette/brakes/e-j6d.webp"
    ) == (None, "", False)
    assert asset_map_sync.parse_shared_option_media(
        "https://example.test/27vette/brakes/h-h-j6d.webp"
    ) == ("h-h", "j6d", False)


def test_duplicate_wordpress_records_for_same_url_are_one_media_candidate() -> None:
    url = "https://example.test/wp-content/uploads/pictures/27vette/baz.jpg"

    inventory = asset_map_sync.build_media_inventory([url, url])

    assert inventory.option_bare == {"baz": [url]}
    assert inventory.unparseable == []


def test_fetch_media_snapshot_collapses_duplicate_attachment_records(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    url = "https://example.test/wp-content/uploads/pictures/27vette/baz.jpg"

    monkeypatch.setattr(
        asset_map_sync,
        "_open_json",
        lambda *_args, **_kwargs: (
            [
                {"source_url": url, "modified": "2026-08-12T10:00:00"},
                {"source_url": url, "modified": "2026-08-13T11:00:00"},
            ],
            {"x-wp-totalpages": "1"},
        ),
    )

    snapshot = asset_map_sync.fetch_media_snapshot(timeout=1)

    assert snapshot.urls == [url]
    assert snapshot.modified_by_url == {url: "2026-08-13T11:00:00"}


def test_parse_model_and_bodystyle_media_names() -> None:
    assert asset_map_sync.parse_model_media("https://example.test/27vette/grandsport.png") == ("grand_sport", "grandSport")
    assert asset_map_sync.parse_bodystyle_media("https://example.test/27vette/c07-1.png") == (
        "stingray",
        "body_style__coupe",
        "image_url",
    )
    assert asset_map_sync.parse_bodystyle_media("https://example.test/27vette/h67-2.png") == (
        "z06",
        "body_style__convertible",
        "hover_image_url",
    )


def test_discovery_uses_promoted_runtime_models_not_inactive_future_sheets() -> None:
    wb = make_discovery_workbook()

    sources = asset_map_sync.discover_promoted_option_sources(wb)
    desired = asset_map_sync.read_option_sheets(wb, sources)

    assert sources == {"stingray": "stingray_options", "grand_sport": "grandSport_options"}
    assert sorted(desired) == [
        ("grand_sport", "option", "opt_qe6_001"),
        ("stingray", "option", "opt_gba_001"),
    ]
    assert ("zr1", "option", "opt_future_001") not in desired


def test_reconcile_uses_bare_media_as_shared_fallback_after_model_prefixed_media() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
        ("grand_sport", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
        ("z06", "option", "opt_stx_001"): {"target_type": "option", "rpo": "stx", "name": "Stripe"},
        ("stingray", "option", "opt_noimg_001"): {"target_type": "option", "rpo": "nix", "name": "No Image"},
    }
    exact, bare, _ = asset_map_sync.build_media_index(
        [
            "https://example.test/27vette/paint/gba.png",
            "https://example.test/27vette/stripes/stx.png",
            "https://example.test/27vette/z06/h-stx.png",
        ]
    )

    report, url_writes, inserts, status, _used = asset_map_sync.reconcile(
        desired,
        exact,
        bare,
        existing_rows={},
        alive={},
        incremental=False,
    )

    actions = {(row["model_key"], row["target_id"]): row["action"] for row in report}
    sources = {(row["model_key"], row["target_id"]): row["candidate_source"] for row in report}
    urls = {(row["model_key"], row["target_id"]): row["new_url"] for row in report}
    assert actions[("stingray", "opt_gba_001")] == "insert_filled"
    assert actions[("grand_sport", "opt_gba_001")] == "insert_filled"
    assert actions[("z06", "opt_stx_001")] == "insert_filled"
    assert actions[("stingray", "opt_noimg_001")] == "flag_missing"
    assert sources[("stingray", "opt_gba_001")] == "bare-shared"
    assert sources[("grand_sport", "opt_gba_001")] == "bare-shared"
    assert sources[("z06", "opt_stx_001")] == "prefixed"
    assert urls[("stingray", "opt_gba_001")] == "https://example.test/27vette/paint/gba.png"
    assert urls[("grand_sport", "opt_gba_001")] == "https://example.test/27vette/paint/gba.png"
    assert urls[("z06", "opt_stx_001")] == "https://example.test/27vette/z06/h-stx.png"
    assert [(row["model"], row["tid"], row["url"], row["target_type"]) for row in inserts] == [
        ("stingray", "opt_gba_001", "https://example.test/27vette/paint/gba.png", "option"),
        ("grand_sport", "opt_gba_001", "https://example.test/27vette/paint/gba.png", "option"),
        ("z06", "opt_stx_001", "https://example.test/27vette/z06/h-stx.png", "option"),
    ]
    assert url_writes == {}
    assert status == {}


def test_reconcile_uses_model_fallback_chain_before_bare_media() -> None:
    desired = {
        ("grand_sport_x", "option", "opt_aaa_001"): {"target_type": "option", "rpo": "aaa", "name": "GSX exact"},
        ("grand_sport_x", "option", "opt_bbb_001"): {"target_type": "option", "rpo": "bbb", "name": "Grand Sport fallback"},
        ("grand_sport_x", "option", "opt_ccc_001"): {"target_type": "option", "rpo": "ccc", "name": "Stingray chain fallback"},
        ("grand_sport", "option", "opt_ddd_001"): {"target_type": "option", "rpo": "ddd", "name": "Stingray fallback"},
        ("zr1", "option", "opt_eee_001"): {"target_type": "option", "rpo": "eee", "name": "ZR1 exact"},
        ("zr1x", "option", "opt_fff_001"): {"target_type": "option", "rpo": "fff", "name": "Z06 fallback"},
        ("z06", "option", "opt_ggg_001"): {"target_type": "option", "rpo": "ggg", "name": "Bare fallback"},
    }
    media = asset_map_sync.build_media_inventory(
        [
            "https://example.test/27vette/g-aaa.png",
            "https://example.test/27vette/e-aaa.png",
            "https://example.test/27vette/aaa.png",
            "https://example.test/27vette/e-bbb.png",
            "https://example.test/27vette/c-bbb.png",
            "https://example.test/27vette/bbb.png",
            "https://example.test/27vette/c-ccc.png",
            "https://example.test/27vette/ccc.png",
            "https://example.test/27vette/c-ddd.png",
            "https://example.test/27vette/ddd.png",
            "https://example.test/27vette/r-eee.png",
            "https://example.test/27vette/h-eee.png",
            "https://example.test/27vette/eee.png",
            "https://example.test/27vette/h-fff.png",
            "https://example.test/27vette/fff.png",
            "https://example.test/27vette/ggg.png",
        ]
    )

    plan = asset_map_sync.reconcile(desired, media, existing_rows={}, alive={}, incremental=False)
    by_target = {row["target_id"]: row for row in plan.report}

    assert by_target["opt_aaa_001"]["candidate_source"] == "prefixed"
    assert by_target["opt_aaa_001"]["new_url"].endswith("/g-aaa.png")
    assert by_target["opt_bbb_001"]["candidate_source"] == "model-fallback:grand_sport"
    assert by_target["opt_bbb_001"]["new_url"].endswith("/e-bbb.png")
    assert by_target["opt_ccc_001"]["candidate_source"] == "model-fallback:stingray"
    assert by_target["opt_ccc_001"]["new_url"].endswith("/c-ccc.png")
    assert by_target["opt_ddd_001"]["candidate_source"] == "model-fallback:stingray"
    assert by_target["opt_ddd_001"]["new_url"].endswith("/c-ddd.png")
    assert by_target["opt_eee_001"]["candidate_source"] == "prefixed"
    assert by_target["opt_eee_001"]["new_url"].endswith("/r-eee.png")
    assert by_target["opt_fff_001"]["candidate_source"] == "model-fallback:z06"
    assert by_target["opt_fff_001"]["new_url"].endswith("/h-fff.png")
    assert by_target["opt_ggg_001"]["candidate_source"] == "bare-shared"
    assert by_target["opt_ggg_001"]["new_url"].endswith("/ggg.png")


def test_reconcile_uses_exact_then_shared_prefix_then_model_fallback_then_bare() -> None:
    desired = {
        ("grand_sport", "option", "opt_aaa_001"): {"target_type": "option", "rpo": "aaa", "name": "GS exact"},
        ("grand_sport", "option", "opt_bbb_001"): {"target_type": "option", "rpo": "bbb", "name": "GS shared"},
        ("grand_sport_x", "option", "opt_aaa_001"): {"target_type": "option", "rpo": "aaa", "name": "GSX exact"},
        ("grand_sport_x", "option", "opt_bbb_001"): {"target_type": "option", "rpo": "bbb", "name": "GSX shared"},
    }
    media = asset_map_sync.build_media_inventory(
        [
            "https://example.test/27vette/e-aaa.png",
            "https://example.test/27vette/g-aaa.png",
            "https://example.test/27vette/e-g-aaa-o-cmp.webp",
            "https://example.test/27vette/e-g-bbb-o-cmp.webp",
            "https://example.test/27vette/c-bbb.png",
            "https://example.test/27vette/bbb.png",
        ]
    )

    plan = asset_map_sync.reconcile(desired, media, existing_rows={}, alive={}, incremental=False)
    by_key = {(row["model_key"], row["target_id"]): row for row in plan.report}

    assert by_key[("grand_sport", "opt_aaa_001")]["candidate_source"] == "prefixed"
    assert by_key[("grand_sport", "opt_aaa_001")]["new_url"].endswith("/e-aaa.png")
    assert by_key[("grand_sport_x", "opt_aaa_001")]["candidate_source"] == "prefixed"
    assert by_key[("grand_sport_x", "opt_aaa_001")]["new_url"].endswith("/g-aaa.png")
    assert by_key[("grand_sport", "opt_bbb_001")]["candidate_source"] == "shared-prefix:e-g"
    assert by_key[("grand_sport", "opt_bbb_001")]["new_url"].endswith("/e-g-bbb-o-cmp.webp")
    assert by_key[("grand_sport_x", "opt_bbb_001")]["candidate_source"] == "shared-prefix:e-g"
    assert by_key[("grand_sport_x", "opt_bbb_001")]["new_url"].endswith("/e-g-bbb-o-cmp.webp")


def test_reconcile_resolves_arbitrary_shared_group_for_every_named_model() -> None:
    desired = {
        (model_key, "option", "opt_j6d_001"): {
            "target_type": "option",
            "rpo": "j6d",
            "name": "Calipers",
        }
        for model_key in ("z06", "zr1", "zr1x")
    }
    media = asset_map_sync.build_media_inventory(
        ["https://example.test/27vette/brakes/h-s-r-j6d-o-cmp.webp"]
    )

    plan = asset_map_sync.reconcile(desired, media, existing_rows={}, alive={}, incremental=False)

    assert {row["model_key"] for row in plan.report} == {"z06", "zr1", "zr1x"}
    assert {row["candidate_source"] for row in plan.report} == {"shared-prefix:h-s-r"}
    assert {row["new_url"] for row in plan.report} == {
        "https://example.test/27vette/brakes/h-s-r-j6d-o-cmp.webp"
    }


def test_reconcile_prefers_narrower_shared_group_over_broader_shared_group() -> None:
    desired = {
        ("zr1", "option", "opt_j6d_001"): {
            "target_type": "option",
            "rpo": "j6d",
            "name": "Calipers",
        },
    }
    media = asset_map_sync.build_media_inventory(
        [
            "https://example.test/27vette/brakes/h-r-j6d.webp",
            "https://example.test/27vette/brakes/h-s-r-j6d.webp",
        ]
    )

    plan = asset_map_sync.reconcile(desired, media, existing_rows={}, alive={}, incremental=False)

    assert plan.report[0]["candidate_source"] == "shared-prefix:h-r"
    assert plan.report[0]["new_url"].endswith("/h-r-j6d.webp")


def test_reconcile_flags_duplicate_shared_prefix_media_without_falling_through() -> None:
    desired = {
        ("grand_sport_x", "option", "opt_j6d_001"): {
            "target_type": "option",
            "rpo": "j6d",
            "name": "Calipers",
        },
    }
    media = asset_map_sync.build_media_inventory(
        [
            "https://example.test/27vette/brakes/e-g-j6d-o-cmp.webp",
            "https://example.test/27vette/alternate/e-g-j6d-o-cmp.webp",
            "https://example.test/27vette/brakes/e-j6d.png",
            "https://example.test/27vette/brakes/j6d.png",
        ]
    )

    plan = asset_map_sync.reconcile(desired, media, existing_rows={}, alive={}, incremental=False)

    assert len(plan.report) == 1
    assert plan.report[0]["action"] == "flag_ambiguous"
    assert plan.report[0]["candidate_source"] == "shared-prefix:e-g:ambiguous"
    assert "multiple e-g shared-prefix files" in plan.report[0]["note"]
    assert plan.url_writes == {}
    assert plan.inserts == []


def test_reconcile_flags_ambiguous_highest_priority_model_candidate_without_falling_through() -> None:
    desired = {
        ("grand_sport_x", "option", "opt_bbb_001"): {"target_type": "option", "rpo": "bbb", "name": "Ambiguous Grand Sport fallback"},
        ("grand_sport_x", "option", "opt_ccc_001"): {"target_type": "option", "rpo": "ccc", "name": "Ambiguous GSX exact"},
    }
    media = asset_map_sync.build_media_inventory(
        [
            "https://example.test/27vette/mech/e-bbb.png",
            "https://example.test/27vette/exhaust/e-bbb.png",
            "https://example.test/27vette/mech/c-bbb.png",
            "https://example.test/27vette/mech/bbb.png",
            "https://example.test/27vette/mech/g-ccc.png",
            "https://example.test/27vette/exhaust/g-ccc.png",
            "https://example.test/27vette/mech/e-ccc.png",
            "https://example.test/27vette/mech/ccc.png",
        ]
    )

    plan = asset_map_sync.reconcile(desired, media, existing_rows={}, alive={}, incremental=False)
    by_target = {row["target_id"]: row for row in plan.report}

    assert len(plan.report) == 2
    assert by_target["opt_bbb_001"]["action"] == "flag_ambiguous"
    assert by_target["opt_bbb_001"]["candidate_source"] == "model-fallback:grand_sport:ambiguous"
    assert "multiple grand_sport-prefixed files" in by_target["opt_bbb_001"]["note"]
    assert by_target["opt_ccc_001"]["action"] == "flag_ambiguous"
    assert by_target["opt_ccc_001"]["candidate_source"] == "prefixed-ambiguous"
    assert "multiple grand_sport_x-prefixed files" in by_target["opt_ccc_001"]["note"]
    assert plan.url_writes == {}
    assert plan.inserts == []


def test_reconcile_flags_duplicate_bare_media_for_same_rpo_as_ambiguous() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
        ("grand_sport", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
    }
    exact, bare, _ = asset_map_sync.build_media_index(
        [
            "https://example.test/27vette/paint/gba.png",
            "https://example.test/27vette/exterior/gba.png",
        ]
    )

    report, url_writes, inserts, status, _used = asset_map_sync.reconcile(
        desired,
        exact,
        bare,
        existing_rows={},
        alive={},
        incremental=False,
    )

    assert {row["action"] for row in report} == {"flag_ambiguous"}
    assert {row["candidate_source"] for row in report} == {"bare-ambiguous"}
    assert all("multiple bare files" in row["note"] for row in report)
    assert url_writes == {}
    assert inserts == []
    assert status == {}


def test_reconcile_replaces_existing_url_when_canonical_media_inventory_differs() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
    }
    media = asset_map_sync.build_media_inventory(["https://example.test/27vette/paint/c-gba-new.png"])

    report, url_writes, inserts, status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows={
            ("stingray", "option", "opt_gba_001"): {
                "row": 12,
                "target_type": "option",
                "url": "https://example.test/27vette/paint/c-gba-old.png",
            }
        },
        alive={"https://example.test/27vette/paint/c-gba-old.png": True},
        incremental=False,
    )

    assert report[0]["action"] == "replace_canonical"
    assert report[0]["new_url"] == "https://example.test/27vette/paint/c-gba-new.png"
    assert url_writes == {(12, "image_url"): "https://example.test/27vette/paint/c-gba-new.png"}
    assert inserts == []
    assert status == {12: "ok"}


def test_reconcile_inserts_context_choice_base_and_hover_media_from_naming_pair() -> None:
    desired = {
        ("stingray", "context_choice", "body_style__coupe"): {
            "target_type": "context_choice",
            "rpo": "",
            "name": "Coupe",
            "source_sheet": "generated_body_style_context",
        }
    }
    media = asset_map_sync.build_media_inventory(
        [
            "https://example.test/27vette/body/c07-1.png",
            "https://example.test/27vette/body/c07-2.png",
        ]
    )

    report, url_writes, inserts, status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows={},
        alive={},
        incremental=False,
    )

    assert report[0]["action"] == "insert_filled"
    assert report[0]["target_type"] == "context_choice"
    assert inserts == [
        {
            "model": "stingray",
            "target_type": "context_choice",
            "tid": "body_style__coupe",
            "rpo": "",
            "name": "Coupe",
            "fields": {
                "image_url": "https://example.test/27vette/body/c07-1.png",
                "hover_image_url": "https://example.test/27vette/body/c07-2.png",
            },
            "url": "https://example.test/27vette/body/c07-1.png",
            "status": "ok",
            "source_sheet": "generated_body_style_context",
        }
    ]
    assert url_writes == {}
    assert status == {}


def test_reconcile_keys_by_target_type_so_same_id_string_across_types_does_not_collide() -> None:
    """A shared model_key + target_id string under two different target_types must stay distinct."""
    desired = {
        ("stingray", "option", "shared_id"): {
            "target_type": "option",
            "rpo": "gba",
            "name": "Black Option",
        },
        ("stingray", "model", "shared_id"): {
            "target_type": "model",
            "rpo": "",
            "name": "Stingray Model Card",
        },
    }
    media = asset_map_sync.MediaInventory(
        option_exact={("stingray", "gba"): ["https://example.test/27vette/paint/gba.png"]},
        option_bare={},
        model={("stingray", "shared_id"): ["https://example.test/27vette/stingray-shared_id.png"]},
        bodystyle={},
        unparseable=[],
    )

    report, url_writes, inserts, status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows={},
        alive={},
        incremental=False,
    )

    by_target_type = {row["target_type"]: row for row in report}
    assert set(by_target_type) == {"option", "model"}
    assert by_target_type["option"]["action"] == "insert_filled"
    assert by_target_type["model"]["action"] == "insert_filled"
    assert by_target_type["option"]["new_url"] == "https://example.test/27vette/paint/gba.png"
    assert by_target_type["model"]["new_url"] == "https://example.test/27vette/stingray-shared_id.png"
    assert {(insert["target_type"], insert["tid"], insert["url"]) for insert in inserts} == {
        ("option", "shared_id", "https://example.test/27vette/paint/gba.png"),
        ("model", "shared_id", "https://example.test/27vette/stingray-shared_id.png"),
    }
    assert url_writes == {}
    assert status == {}


def test_existing_asset_rows_keys_by_target_type_so_same_id_string_across_types_does_not_collide() -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    add_sheet(
        wb,
        "asset_map",
        ["model_key", "target_type", "target_id", "image_url", "active"],
        [
            {"model_key": "stingray", "target_type": "option", "target_id": "shared_id", "image_url": "https://example.test/opt.png", "active": True},
            {"model_key": "stingray", "target_type": "model", "target_id": "shared_id", "image_url": "https://example.test/model.png", "active": True},
        ],
    )
    ws = wb["asset_map"]
    header_index = {header: index for index, header in enumerate(["model_key", "target_type", "target_id", "image_url", "active"])}

    rows = asset_map_sync.existing_asset_rows(ws, header_index)

    assert set(rows) == {("stingray", "option", "shared_id"), ("stingray", "model", "shared_id")}
    assert rows[("stingray", "option", "shared_id")]["url"] == "https://example.test/opt.png"
    assert rows[("stingray", "model", "shared_id")]["url"] == "https://example.test/model.png"


def test_existing_asset_rows_includes_wildcard_rows_under_star_model_key() -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    add_sheet(
        wb,
        "asset_map",
        ["model_key", "target_type", "target_id", "image_url", "active"],
        [
            {"model_key": "*", "target_type": "option", "target_id": "opt_gba_001", "image_url": "https://example.test/shared.png", "active": True},
        ],
    )
    ws = wb["asset_map"]
    header_index = {header: index for index, header in enumerate(["model_key", "target_type", "target_id", "image_url", "active"])}

    rows = asset_map_sync.existing_asset_rows(ws, header_index)

    assert set(rows) == {("*", "option", "opt_gba_001")}
    assert rows[("*", "option", "opt_gba_001")]["url"] == "https://example.test/shared.png"


def test_reconcile_wildcard_covered_target_is_keep_not_insert() -> None:
    """The anti-undo contract: a sync run must not re-insert per-model rows for wildcard-covered targets."""
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
        ("grand_sport", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
    }
    media = asset_map_sync.build_media_inventory(["https://example.test/27vette/paint/gba.png"])
    existing_rows = {
        ("*", "option", "opt_gba_001"): {"row": 2, "target_type": "option", "url": "https://example.test/27vette/paint/gba.png"},
    }

    report, url_writes, inserts, status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive={},
        incremental=False,
    )

    per_model = {row["model_key"]: row for row in report if row["target_id"] == "opt_gba_001"}
    assert set(per_model) == {"stingray", "grand_sport"}
    for row in per_model.values():
        assert row["action"] == "keep"
        assert "wildcard" in row["note"]
        assert row["existing_url"] == "https://example.test/27vette/paint/gba.png"
    assert inserts == []
    assert url_writes == {}
    assert not any(row["action"] == "stale_target" for row in report)


def test_reconcile_wildcard_covered_target_without_media_candidate_is_keep_not_missing() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
    }
    media = asset_map_sync.build_media_inventory([])
    existing_rows = {
        ("*", "option", "opt_gba_001"): {"row": 2, "target_type": "option", "url": "https://example.test/27vette/paint/gba.png"},
    }

    report, url_writes, inserts, _status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive={},
        incremental=False,
    )

    row = next(r for r in report if r["model_key"] == "stingray" and r["target_id"] == "opt_gba_001")
    assert row["action"] == "keep"
    assert inserts == []
    assert url_writes == {}


def test_reconcile_wildcard_url_conflict_is_review_action_with_no_writes() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
    }
    media = asset_map_sync.build_media_inventory(["https://example.test/27vette/paint/c-gba-new.png"])
    existing_rows = {
        ("*", "option", "opt_gba_001"): {"row": 2, "target_type": "option", "url": "https://example.test/27vette/paint/gba-old.png"},
    }

    report, url_writes, inserts, _status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive={},
        incremental=False,
    )

    row = next(r for r in report if r["model_key"] == "stingray" and r["target_id"] == "opt_gba_001")
    assert row["action"] == "wildcard_conflict"
    assert row["existing_url"] == "https://example.test/27vette/paint/gba-old.png"
    assert row["new_url"] == "https://example.test/27vette/paint/c-gba-new.png"
    assert "never edits wildcard rows" in row["note"]
    assert url_writes == {}
    assert inserts == []


def test_complete_reconcile_updates_one_wildcard_row_for_unique_bare_media() -> None:
    desired = {
        ("stingray", "option", "opt_eri_001"): {"target_type": "option", "rpo": "eri", "name": "Battery Protection"},
        ("zr1", "option", "opt_eri_001"): {"target_type": "option", "rpo": "eri", "name": "Battery Protection"},
    }
    new_url = "https://example.test/27vette/lpo/eri-cmp.webp"
    media = asset_map_sync.build_media_inventory([new_url])
    existing_rows = {
        ("*", "option", "opt_eri_001"): {"row": 2, "target_type": "option", "url": "https://example.test/27vette/ext/eri.jpg"},
    }

    report, url_writes, inserts, _status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive={},
        incremental=False,
        update_safe_wildcards=True,
    )

    assert url_writes == {(2, "image_url"): new_url}
    assert inserts == []
    assert {row["action"] for row in report} == {"replace_shared_canonical"}


def test_complete_reconcile_does_not_write_model_specific_candidate_to_wildcard() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
    }
    media = asset_map_sync.build_media_inventory(["https://example.test/27vette/paint/c-gba-new.png"])
    existing_rows = {
        ("*", "option", "opt_gba_001"): {"row": 2, "target_type": "option", "url": "https://example.test/27vette/paint/gba-old.png"},
    }

    report, url_writes, inserts, _status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive={},
        incremental=False,
        update_safe_wildcards=True,
    )

    assert report[0]["action"] == "wildcard_conflict"
    assert url_writes == {}
    assert inserts == []


def test_reconcile_exact_row_takes_precedence_over_wildcard_row() -> None:
    desired = {
        ("z06", "option", "opt_j6d_001"): {"target_type": "option", "rpo": "j6d", "name": "Calipers"},
    }
    media = asset_map_sync.build_media_inventory([])
    existing_rows = {
        ("*", "option", "opt_j6d_001"): {"row": 2, "target_type": "option", "url": "https://example.test/shared.png"},
        ("z06", "option", "opt_j6d_001"): {"row": 3, "target_type": "option", "url": "https://example.test/z06-exact.png"},
    }

    report, url_writes, inserts, _status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive={},
        incremental=False,
    )

    row = next(r for r in report if r["model_key"] == "z06" and r["target_id"] == "opt_j6d_001")
    assert row["existing_url"] == "https://example.test/z06-exact.png"
    assert url_writes == {}
    assert inserts == []


def test_reconcile_wildcard_row_is_stale_only_when_no_model_desires_target() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black"},
    }
    media = asset_map_sync.build_media_inventory([])
    existing_rows = {
        ("*", "option", "opt_gba_001"): {"row": 2, "target_type": "option", "url": "https://example.test/shared.png"},
        ("*", "option", "opt_gone_001"): {"row": 3, "target_type": "option", "url": "https://example.test/gone.png"},
    }

    report, _url_writes, _inserts, status, _used = asset_map_sync.reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive={},
        incremental=False,
    )

    stale = [row for row in report if row["action"] == "stale_target"]
    assert [(row["model_key"], row["target_id"]) for row in stale] == [("*", "opt_gone_001")]
    assert status[3] == "stale_target"
    assert 2 not in status


def test_section_coverage_counts_wildcard_row_as_covered() -> None:
    desired = {
        ("stingray", "option", "opt_gba_001"): {"target_type": "option", "rpo": "gba", "name": "Black", "section_id": "sec_pain_001"},
        ("stingray", "option", "opt_nix_001"): {"target_type": "option", "rpo": "nix", "name": "None", "section_id": "sec_pain_001"},
    }
    existing_rows = {
        ("*", "option", "opt_gba_001"): {"row": 2, "target_type": "option", "url": "https://example.test/shared.png"},
    }

    stats = asset_map_sync.build_section_coverage_stats(desired, existing_rows)

    section = stats["stingray"]["sections"]["sec_pain_001"]
    assert section["total_targets"] == 2
    assert section["covered"] == 1
    assert section["missing"] == 1


def test_apply_uses_injected_safe_save_and_inserts_confident_candidate(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sync.xlsx"
    make_apply_workbook(workbook_path)
    report_dir = tmp_path / "reports"
    fixture = ROOT / "tests" / "fixtures" / "asset-map-sync-media-urls.txt"
    calls: list[tuple[Path, int | None]] = []

    def fake_safe_save(wb, path, *, loaded_mtime_ns=None):
        calls.append((Path(path), loaded_mtime_ns))
        wb.save(path)
        return tmp_path / "backup.xlsx"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=asset_map_sync.read_media_url_list(fixture),
        apply=True,
        verify_existing=False,
        save_fn=fake_safe_save,
    )

    assert result.url_write_count == 0
    assert result.insert_count == 1
    assert calls and calls[0][0] == workbook_path and calls[0][1] is not None
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows = list(wb["asset_map"].iter_rows(min_row=2, values_only=True))
    finally:
        wb.close()
    assert rows[0][:9] == (
        "stingray",
        "option",
        "opt_gba_001",
        "https://example.test/c-gba.png",
        "Black",
        "cover",
        "center",
        True,
        "auto-seeded",
    )
    assert (report_dir / "asset_map_sync_report.csv").exists()
    assert (report_dir / "asset_map_unmatched_media.csv").exists()


def test_apply_with_no_unambiguous_changes_does_not_rewrite_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sync.xlsx"
    make_apply_workbook(workbook_path)
    report_dir = tmp_path / "reports"
    calls: list[Path] = []

    def fake_safe_save(wb, path, *, loaded_mtime_ns=None):
        calls.append(Path(path))
        raise AssertionError("save must not run without workbook changes")

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=[],
        apply=True,
        verify_existing=False,
        save_fn=fake_safe_save,
    )

    assert result.url_write_count == 0
    assert result.insert_count == 0
    assert result.backup_path is None
    assert calls == []


def test_data_cache_version_bump_is_targeted_and_repeatable(tmp_path: Path) -> None:
    index_path = tmp_path / "index.html"
    index_path.write_text('<script src="./data.js?v=26"></script>\n<script src="./app.js?v=27"></script>\n', encoding="utf-8")

    asset_map_sync._bump_data_cache_version(index_path)

    assert index_path.read_text(encoding="utf-8") == (
        '<script src="./data.js?v=27"></script>\n<script src="./app.js?v=27"></script>\n'
    )


def test_report_manifest_records_source_inventory_and_counts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sync.xlsx"
    make_apply_workbook(workbook_path)
    report_dir = tmp_path / "reports"
    fixture = ROOT / "tests" / "fixtures" / "asset-map-sync-media-urls.txt"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=asset_map_sync.read_media_url_list(fixture),
        apply=False,
        verify_existing=False,
        media_source="media-url-list",
        since_argument=None,
        resolved_modified_after=None,
        state_read=False,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    required = {
        "version",
        "generated_at_utc",
        "workbook_path",
        "asset_sheet",
        "apply",
        "media_source",
        "since_argument",
        "resolved_modified_after",
        "incremental",
        "state_path",
        "state_read",
        "state_written",
        "media_url_count",
        "verify_existing",
        "included_sources",
        "action_counts",
        "url_write_count",
        "insert_count",
        "unmatched_count",
        "unparseable_count",
        "report_path",
        "unmatched_path",
    }
    assert required <= set(manifest)
    assert manifest["apply"] is False
    assert manifest["media_source"] == "media-url-list"
    assert manifest["verify_existing"] is False
    assert manifest["included_sources"] == [{"model_key": "stingray", "option_sheet": "stingray_options"}]
    assert manifest["action_counts"] == {"flag_missing": 3, "insert_filled": 1}
    assert manifest["url_write_count"] == 0
    assert manifest["insert_count"] == 1
    assert manifest["media_url_count"] == len(asset_map_sync.read_media_url_list(fixture))
    assert Path(manifest["report_path"]).name == "asset_map_sync_report.csv"
    assert Path(manifest["unmatched_path"]).name == "asset_map_unmatched_media.csv"
    assert result.manifest_path == report_dir / "asset_map_sync_manifest.json"


def test_missing_images_artifact_written_and_manifested(tmp_path: Path) -> None:
    workbook_path = tmp_path / "missing-report.xlsx"
    make_missing_report_workbook(workbook_path)
    report_dir = tmp_path / "reports"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=[
            "https://example.test/gba.png",
            "https://example.test/c-stx.png",
            "https://example.test/abc.png",
        ],
        apply=False,
        verify_existing=False,
        media_source="media-url-list",
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    missing_path = report_dir / "asset_map_missing_images.csv"
    assert result.missing_path == missing_path
    assert Path(manifest["missing_images_path"]) == missing_path

    rows = list(csv.DictReader(missing_path.open(encoding="utf-8")))
    assert manifest["missing_images_count"] == len(rows) == 7
    assert ("opt_noimg_001", "flag_missing") in {(row["target_id"], row["action"]) for row in rows}
    no_image = next(row for row in rows if row["target_id"] == "opt_noimg_001")
    assert no_image["section_id"] == "sec_test_001"
    assert no_image["option_name"] == "No Image"


def test_missing_images_artifact_excludes_keep_insert_and_unmatched(tmp_path: Path) -> None:
    workbook_path = tmp_path / "missing-filter.xlsx"
    make_missing_report_workbook(workbook_path)
    report_dir = tmp_path / "reports"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=[
            "https://example.test/gba.png",
            "https://example.test/c-stx.png",
            "https://example.test/abc.png",
        ],
        apply=False,
        verify_existing=False,
        media_source="media-url-list",
    )

    rows = list(csv.DictReader(result.missing_path.open(encoding="utf-8")))
    assert {row["action"] for row in rows} <= {"flag_missing", "flag_ambiguous", "flag_dead_no_match"}
    assert "opt_gba_001" not in {row["target_id"] for row in rows}
    assert "opt_stx_001" not in {row["target_id"] for row in rows}
    unmatched_rows = list(csv.DictReader(result.unmatched_path.open(encoding="utf-8")))
    assert {row["parsed_rpo"] for row in unmatched_rows} == {"abc"}
    assert "abc" not in {row["rpo"] for row in rows}


def test_dry_run_reports_without_saving_workbook_rows_or_state(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sync.xlsx"
    make_apply_workbook(workbook_path)
    report_dir = tmp_path / "reports"
    fixture = ROOT / "tests" / "fixtures" / "asset-map-sync-media-urls.txt"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=asset_map_sync.read_media_url_list(fixture),
        apply=False,
        verify_existing=False,
        media_source="media-url-list",
        since_argument="auto",
        resolved_modified_after=None,
        state_read=False,
    )

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows = list(wb["asset_map"].iter_rows(min_row=2, values_only=True))
    finally:
        wb.close()
    assert rows == []
    assert result.insert_count == 1
    assert not asset_map_sync.state_path(report_dir).exists()
    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["state_read"] is False
    assert manifest["state_written"] is False


def test_apply_state_written_only_after_safe_save_success(tmp_path: Path) -> None:
    fixture = ROOT / "tests" / "fixtures" / "asset-map-sync-media-urls.txt"

    success_workbook = tmp_path / "success.xlsx"
    make_apply_workbook(success_workbook)
    success_dir = tmp_path / "success-report"

    def fake_safe_save(wb, path, *, loaded_mtime_ns=None):
        wb.save(path)
        return tmp_path / "backup.xlsx"

    result = asset_map_sync.run_sync(
        workbook_path=success_workbook,
        report_dir=success_dir,
        media_urls=asset_map_sync.read_media_url_list(fixture),
        apply=True,
        verify_existing=False,
        save_fn=fake_safe_save,
        since_argument="auto",
        resolved_modified_after="2026-06-26T00:00:00",
        state_read=True,
        media_source="media-url-list",
    )

    assert asset_map_sync.state_path(success_dir).exists()
    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["state_read"] is True
    assert manifest["state_written"] is True

    failed_workbook = tmp_path / "failed.xlsx"
    make_apply_workbook(failed_workbook)
    failed_dir = tmp_path / "failed-report"

    def failing_safe_save(wb, path, *, loaded_mtime_ns=None):
        raise RuntimeError("save failed")

    with pytest.raises(RuntimeError, match="save failed"):
        asset_map_sync.run_sync(
            workbook_path=failed_workbook,
            report_dir=failed_dir,
            media_urls=asset_map_sync.read_media_url_list(fixture),
            apply=True,
            verify_existing=False,
            save_fn=failing_safe_save,
            since_argument="auto",
            resolved_modified_after="2026-06-26T00:00:00",
            state_read=True,
            media_source="media-url-list",
        )

    assert not asset_map_sync.state_path(failed_dir).exists()


def test_cli_help_works_without_requests_dependency() -> None:
    completed = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "sync_asset_map.py"), "--help"],
        cwd=ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )

    assert completed.returncode == 0, completed.stdout
    assert "--media-url-list" in completed.stdout
    assert "--apply" in completed.stdout
    assert "--complete" in completed.stdout
    assert "--status-col" not in completed.stdout
    assert "--deactivate-stale" not in completed.stdout
    assert "--seed-blank-missing" not in completed.stdout


def test_cli_rejects_unsupported_schema_and_lifecycle_flags(tmp_path: Path) -> None:
    for flag in ["--status-col", "--deactivate-stale", "--seed-blank-missing"]:
        completed = subprocess.run(
            [
                sys.executable,
                str(ROOT / "scripts" / "sync_asset_map.py"),
                "--workbook",
                str(tmp_path / "missing.xlsx"),
                "--report-dir",
                str(tmp_path / "reports"),
                "--media-url-list",
                str(ROOT / "tests" / "fixtures" / "asset-map-sync-media-urls.txt"),
                flag,
            ],
            cwd=ROOT,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            check=False,
        )

        assert completed.returncode != 0
        assert flag in completed.stdout


def test_legacy_entrypoint_stays_removed() -> None:
    # The retired asset_map-Sync/asset_map_sync.py wrote the workbook directly;
    # the supported entrypoint is scripts/sync_asset_map.py (docs/asset-map-sync.md).
    assert not (ROOT / "asset_map-Sync").exists()


def make_coverage_workbook(path: Path) -> None:
    """Workbook exercising every coverage-intent classifier rule."""

    wb = Workbook()
    del wb[wb.sheetnames[0]]
    add_sheet(
        wb,
        "model_registry_promotion",
        ["model_key", "promoted_to_runtime", "active", "display_order", "registry_key", "model_label"],
        [
            {"model_key": "stingray", "promoted_to_runtime": True, "active": True, "display_order": 1, "registry_key": "stingray", "model_label": "Stingray"},
            {"model_key": "grand_sport", "promoted_to_runtime": True, "active": True, "display_order": 2, "registry_key": "grandSport", "model_label": "Grand Sport"},
        ],
    )
    add_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active"],
        [
            {"model_key": "stingray", "source_role": "source_option_sheet", "sheet_name": "stingray_options", "active": True},
            {"model_key": "grand_sport", "source_role": "source_option_sheet", "sheet_name": "grandSport_options", "active": True},
        ],
    )
    option_headers = ["option_id", "rpo", "option_name", "section_id", "active", "selectable"]
    add_sheet(
        wb,
        "stingray_options",
        option_headers,
        [
            {"option_id": "opt_disp_001", "rpo": "DSP", "option_name": "Display Only", "section_id": "sec_disp_001", "active": True, "selectable": True},
            {"option_id": "opt_seb_001", "rpo": "SEB", "option_name": "Std Equip", "section_id": "sec_bucket_001", "active": True, "selectable": True},
            {"option_id": "opt_exist_001", "rpo": "EXI", "option_name": "Existing Row", "section_id": "sec_opt_001", "active": True, "selectable": True},
            {"option_id": "opt_req_001", "rpo": "REQ", "option_name": "Required Section", "section_id": "sec_req_001", "active": True, "selectable": True},
            {"option_id": "opt_repl_001", "rpo": "RPL", "option_name": "Replaceable Default", "section_id": "sec_repl_001", "active": True, "selectable": True},
            {"option_id": "opt_unma_001", "rpo": "UNM", "option_name": "Unmatched Section", "section_id": "sec_ghost_001", "active": True, "selectable": True},
            {"option_id": "opt_prec_001", "rpo": "PRC", "option_name": "Precedent Section", "section_id": "sec_opt_001", "active": True, "selectable": True},
            {"option_id": "opt_nopr_001", "rpo": "NPR", "option_name": "No Precedent", "section_id": "sec_nopr_001", "active": True, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "grandSport_options",
        option_headers,
        [
            {"option_id": "opt_exist_001", "rpo": "EXI", "option_name": "Existing Row GS", "section_id": "sec_opt_001", "active": True, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "section_master",
        ["section_id", "section_name", "selection_mode", "is_required", "display_order", "standard_behavior", "step_key"],
        [
            {"section_id": "sec_disp_001", "section_name": "Display", "selection_mode": "display_only", "is_required": False, "standard_behavior": "locked_included"},
            {"section_id": "sec_bucket_001", "section_name": "Bucketed", "selection_mode": "multi_select_opt", "is_required": False, "standard_behavior": "locked_included"},
            {"section_id": "sec_opt_001", "section_name": "Optional Covered", "selection_mode": "multi_select_opt", "is_required": False, "standard_behavior": "locked_included"},
            {"section_id": "sec_req_001", "section_name": "Required", "selection_mode": "single_select_req", "is_required": True, "standard_behavior": "locked_included"},
            {"section_id": "sec_repl_001", "section_name": "Replaceable", "selection_mode": "single_select_opt", "is_required": False, "standard_behavior": "replaceable_default"},
            {"section_id": "sec_nopr_001", "section_name": "No Precedent", "selection_mode": "multi_select_opt", "is_required": False, "standard_behavior": "locked_included"},
        ],
    )
    add_sheet(
        wb,
        "section_presentation",
        ["model_key", "section_id", "standard_equipment_bucket", "active"],
        [
            {"model_key": "stingray", "section_id": "sec_bucket_001", "standard_equipment_bucket": True, "active": True},
        ],
    )
    add_sheet(
        wb,
        "asset_map",
        ["model_key", "target_type", "target_id", "image_url", "image_alt", "image_fit", "image_position", "active", "notes"],
        [
            {
                "model_key": "stingray",
                "target_type": "option",
                "target_id": "opt_exist_001",
                "image_url": "https://example.test/exi.png",
                "image_alt": "Existing Row",
                "image_fit": "cover",
                "image_position": "center",
                "active": True,
                "notes": "existing",
            },
        ],
    )
    wb.save(path)
    wb.close()


def _coverage_inputs(workbook_path: Path):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sources = asset_map_sync.discover_promoted_option_sources(wb)
    desired = asset_map_sync.read_option_sheets(wb, sources)
    desired.update(asset_map_sync.read_model_targets(wb))
    desired.update(asset_map_sync.read_bodystyle_targets(sources))
    section_metadata = asset_map_sync.read_section_coverage_metadata(wb)
    ws = wb["asset_map"]
    header_index = {
        header: index
        for index, header in enumerate(
            [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        )
        if header
    }
    existing = asset_map_sync.existing_asset_rows(ws, header_index)
    return desired, section_metadata, existing


def test_coverage_classifier_rules_and_reasons(tmp_path: Path) -> None:
    workbook_path = tmp_path / "coverage.xlsx"
    make_coverage_workbook(workbook_path)
    desired, section_metadata, existing = _coverage_inputs(workbook_path)
    classify = asset_map_sync.build_coverage_classifier(desired, section_metadata, existing)

    cases = {
        ("stingray", "model", "stingray"): ("expected", "target_type:model"),
        ("stingray", "context_choice", "body_style__coupe"): ("expected", "target_type:context_choice"),
        ("stingray", "option", "opt_disp_001"): ("not_expected", "section-display-only:sec_disp_001"),
        ("stingray", "option", "opt_seb_001"): ("not_expected", "standard-equipment-bucket:sec_bucket_001"),
        # Universal-expected policy: coverage state, sibling coverage, section
        # requiredness, and media precedent no longer influence intent.
        ("stingray", "option", "opt_exist_001"): ("expected", "universal-expected"),
        ("grand_sport", "option", "opt_exist_001"): ("expected", "universal-expected"),
        ("stingray", "option", "opt_req_001"): ("expected", "universal-expected"),
        ("stingray", "option", "opt_repl_001"): ("expected", "universal-expected"),
        ("stingray", "option", "opt_unma_001"): ("expected", "unmatched-section"),
        ("stingray", "option", "opt_prec_001"): ("expected", "universal-expected"),
        ("stingray", "option", "opt_nopr_001"): ("expected", "universal-expected"),
    }
    for key, wanted in cases.items():
        assert classify(*key) == wanted, key
    for key in cases:
        intent, reason = classify(*key)
        assert intent in asset_map_sync.COVERAGE_INTENTS
        assert reason


def test_coverage_classifier_ignores_asset_coverage_state(tmp_path: Path) -> None:
    """Structural not_expected must never depend on media/asset coverage (no circularity)."""

    workbook_path = tmp_path / "coverage-state.xlsx"
    make_coverage_workbook(workbook_path)
    desired, section_metadata, existing = _coverage_inputs(workbook_path)

    classify_with = asset_map_sync.build_coverage_classifier(desired, section_metadata, existing)
    classify_without = asset_map_sync.build_coverage_classifier(desired, section_metadata, {})
    for key in sorted(desired):
        assert classify_with(*key) == classify_without(*key), key


def test_coverage_classifier_deterministic_under_input_order(tmp_path: Path) -> None:
    workbook_path = tmp_path / "coverage-order.xlsx"
    make_coverage_workbook(workbook_path)
    desired, section_metadata, existing = _coverage_inputs(workbook_path)

    keys = sorted(desired)
    shuffled_desired = {key: desired[key] for key in reversed(keys)}
    classify_a = asset_map_sync.build_coverage_classifier(desired, section_metadata, existing)
    classify_b = asset_map_sync.build_coverage_classifier(shuffled_desired, section_metadata, existing)
    for key in keys:
        assert classify_a(*key) == classify_b(*key), key


def test_missing_images_csv_is_actionable_queue_and_broad_report_keeps_all(tmp_path: Path) -> None:
    workbook_path = tmp_path / "coverage-report.xlsx"
    make_coverage_workbook(workbook_path)
    report_dir = tmp_path / "reports"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=[],
        apply=False,
        verify_existing=False,
        media_source="media-url-list",
    )

    report_rows = list(csv.DictReader(result.report_path.open(encoding="utf-8")))
    missing_rows = list(csv.DictReader(result.missing_path.open(encoding="utf-8")))
    for rows in (report_rows, missing_rows):
        assert rows
        assert "coverage_intent" in rows[0]
        assert "coverage_intent_reason" in rows[0]

    assert {row["coverage_intent"] for row in missing_rows} == {"expected"}
    broad_missing = [
        row for row in report_rows if row["action"] in asset_map_sync.MISSING_IMAGE_ACTIONS
    ]
    excluded = [row for row in broad_missing if row["coverage_intent"] == "not_expected"]
    assert excluded, "coverage fixture must produce not_expected missing rows"
    missing_ids = {(row["model_key"], row["target_type"], row["target_id"]) for row in missing_rows}
    for row in excluded:
        key = (row["model_key"], row["target_type"], row["target_id"])
        assert key not in missing_ids
        assert row["coverage_intent_reason"].startswith(("section-display-only", "standard-equipment-bucket"))
    assert len(missing_rows) == len(broad_missing) - len(excluded)
    sort_keys = [(row["model_key"], row["section_id"], row["target_id"]) for row in missing_rows]
    assert sort_keys == sorted(sort_keys), "queue must be sorted model -> section -> target"


def test_manifest_reports_coverage_breakdown_and_ruleset(tmp_path: Path) -> None:
    workbook_path = tmp_path / "coverage-manifest.xlsx"
    make_coverage_workbook(workbook_path)
    report_dir = tmp_path / "reports"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=[],
        apply=False,
        verify_existing=False,
        media_source="media-url-list",
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    coverage = manifest["coverage"]
    assert coverage["ruleset_version"] == asset_map_sync.COVERAGE_RULESET_VERSION
    assert coverage["ruleset"] == list(asset_map_sync.COVERAGE_RULESET)
    assert coverage["broad_missing_count"] == manifest["broad_missing_images_count"]
    assert manifest["missing_images_count"] == coverage["actionable_missing_count"]
    assert set(coverage["intent_counts"]) <= set(asset_map_sync.COVERAGE_INTENTS)
    assert coverage["intent_counts"].get("not_expected")
    sections = coverage["missing_by_model_section_intent"]
    assert "stingray" in sections
    assert any(
        intent_counts.get("not_expected")
        for section_counts in sections.values()
        for intent_counts in section_counts.values()
    )


def test_manifest_section_coverage_stats(tmp_path: Path) -> None:
    workbook_path = tmp_path / "coverage-stats.xlsx"
    make_coverage_workbook(workbook_path)
    report_dir = tmp_path / "reports"

    result = asset_map_sync.run_sync(
        workbook_path=workbook_path,
        report_dir=report_dir,
        media_urls=[],
        apply=False,
        verify_existing=False,
        media_source="media-url-list",
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    stats = manifest["coverage"]["section_coverage"]

    # Fully covered section: opt_exist_001 is stingray's only sec_opt_001 row
    # besides opt_prec_001 -> partial; grand_sport's sec_opt_001 is uncovered.
    stingray = stats["stingray"]
    sec_opt = stingray["sections"]["sec_opt_001"]
    assert sec_opt["total_targets"] == 2
    assert sec_opt["covered"] == 1
    assert sec_opt["missing"] == 1
    assert sec_opt["coverage_pct"] == 50.0

    grand_sport = stats["grand_sport"]
    gs_sec_opt = grand_sport["sections"]["sec_opt_001"]
    assert gs_sec_opt == {"total_targets": 1, "covered": 0, "missing": 1, "coverage_pct": 0.0}

    # Stats cover ALL desired option targets, including structurally
    # not_expected sections (display-only still has a row counted).
    assert "sec_disp_001" in stingray["sections"]

    # Model rollups are consistent with section sums.
    for model_stats in stats.values():
        section_totals = sum(s["total_targets"] for s in model_stats["sections"].values())
        section_covered = sum(s["covered"] for s in model_stats["sections"].values())
        assert model_stats["total_targets"] == section_totals
        assert model_stats["covered"] == section_covered
        assert model_stats["missing"] == section_totals - section_covered

    # Pure helper agrees with the manifest.
    desired, _, existing = _coverage_inputs(workbook_path)
    assert asset_map_sync.build_section_coverage_stats(desired, existing) == stats
