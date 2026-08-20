"""Pass 2 storage-split and legacy-recovery regressions."""

from __future__ import annotations

import hashlib
import json
import shutil
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "workbook-manager" / "backend"
SCRIPTS = ROOT / "scripts"
for path in (BACKEND, SCRIPTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from app import db as dbmod  # noqa: E402
from app import catalog  # noqa: E402
from app import importer  # noqa: E402
from workbook_manager_fixtures import (  # noqa: E402
    verified_manager_fixture,
    write_compact_missing_identifier_workbook,
    write_compact_missing_sheet_workbook,
    write_compact_unresolved_reference_workbook,
)


class TestSplitStoreMigration(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory(prefix="wbm-pass2-")
        self.root = Path(self.tempdir.name)
        self.state_path = self.root / "workbook_manager.sqlite3"
        self.projection_path = self.root / "workbook_projection.sqlite3"

    def tearDown(self):
        self.tempdir.cleanup()

    def _legacy_database(self) -> tuple[bytes, dict[str, int]]:
        conn = dbmod.connect(self.state_path)
        dbmod.init_schema(conn)
        conn.execute(
            "INSERT INTO models(model_key, registry_key, src_sheet, src_row) "
            "VALUES('legacy_model', 'legacyRegistry', 'model_master', 2)"
        )
        pending_id = conn.execute(
            "INSERT INTO pending_changes(ts, session_id, table_name, model_id, "
            "entity_key_json, op, old_json, new_json, status, validation_json, "
            "confirmed_dependencies) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-22T12:00:00+00:00",
                "legacy-session",
                "options",
                "legacy_model",
                '{"option_id":"opt_legacy"}',
                "update",
                '{"price":"1"}',
                '{"price":"2"}',
                "committed",
                '{"errors":[]}',
                1,
            ),
        ).lastrowid
        history_id = conn.execute(
            "INSERT INTO change_history(ts, actor, entity_type, entity_id, "
            "model_id, op, old_json, new_json, src_sheet, src_row, "
            "validation_result, status, sync_status, sync_detail, "
            "pending_change_id) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-22T12:01:00+00:00",
                "legacy-user",
                "options",
                "opt_legacy",
                "legacy_model",
                "update",
                '{"price":"1"}',
                '{"price":"2"}',
                "legacy_options",
                10,
                "passed",
                "committed",
                "pending",
                "",
                pending_id,
            ),
        ).lastrowid
        dbmod.set_meta(conn, "workbook_sha256", "abc123")
        dbmod.set_meta(conn, "workbook_mtime_ns", "456")
        dbmod.set_meta(conn, "last_import_run_id", "7")
        dbmod.set_meta(conn, "legacy_unknown_meta", "preserve-me")
        conn.commit()
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        conn.close()
        return self.state_path.read_bytes(), {
            "pending_changes": int(pending_id),
            "change_history": int(history_id),
        }

    def _assert_exact_recovery(self) -> None:
        state = dbmod.connect(self.state_path)
        try:
            rows = state.execute(
                "SELECT source_table, source_primary_key, row_json "
                "FROM legacy_recovery_records ORDER BY source_table"
            ).fetchall()
            self.assertEqual(len(rows), 2)
            self.assertEqual(
                {row["source_table"] for row in rows},
                {"pending_changes", "change_history"},
            )
            self.assertEqual(
                len({(row["source_table"], row["source_primary_key"]) for row in rows}),
                2,
            )
        finally:
            state.close()

    def test_first_start_splits_projection_and_preserves_exact_legacy_rows(self):
        legacy_bytes, ids = self._legacy_database()

        result = dbmod.bootstrap_storage(self.state_path, self.projection_path)

        self.assertEqual(result["status"], "migrated")
        self.assertTrue(self.projection_path.exists())
        archive = Path(result["archive_path"])
        self.assertTrue(archive.exists())
        self.assertEqual(archive.read_bytes(), legacy_bytes)
        self.assertIn(hashlib.sha256(legacy_bytes).hexdigest(), archive.name)

        projection = dbmod.connect(self.projection_path)
        state = dbmod.connect(self.state_path)
        try:
            self.assertEqual(
                projection.execute(
                    "SELECT registry_key FROM models WHERE model_key='legacy_model'"
                ).fetchone()["registry_key"],
                "legacyRegistry",
            )
            self.assertEqual(dbmod.get_meta(projection, "workbook_sha256"), "abc123")
            with self.assertRaises(sqlite3.OperationalError):
                projection.execute("SELECT * FROM pending_changes").fetchall()

            rows = state.execute(
                "SELECT source_table, source_primary_key, row_json "
                "FROM legacy_recovery_records ORDER BY source_table"
            ).fetchall()
            self.assertEqual(len(rows), 2)
            recovered = {row["source_table"]: json.loads(row["row_json"]) for row in rows}
            self.assertEqual(recovered["pending_changes"]["id"], ids["pending_changes"])
            self.assertEqual(recovered["pending_changes"]["confirmed_dependencies"], 1)
            self.assertEqual(recovered["change_history"]["id"], ids["change_history"])
            self.assertEqual(recovered["change_history"]["sync_status"], "pending")
            self.assertEqual(
                state.execute(
                    "SELECT value FROM durable_recovery_meta "
                    "WHERE key='legacy_unknown_meta'"
                ).fetchone()["value"],
                "preserve-me",
            )
            projection_marker = dbmod.storage_manifest(projection)
            state_marker = dbmod.storage_manifest(state)
            self.assertEqual(projection_marker["migration_id"], state_marker["migration_id"])
            self.assertEqual(projection_marker["source_sha256"], state_marker["source_sha256"])
            # Both stores must carry the version the code currently builds; a
            # literal here would go stale the moment a schema change lands.
            self.assertEqual(
                projection_marker["schema_version"], dbmod.SCHEMA_VERSION
            )
            self.assertEqual(state_marker["schema_version"], dbmod.SCHEMA_VERSION)
        finally:
            projection.close()
            state.close()

    def test_restart_is_idempotent_and_does_not_duplicate_recovery_rows(self):
        self._legacy_database()
        first = dbmod.bootstrap_storage(self.state_path, self.projection_path)
        state_before = self.state_path.read_bytes()
        projection_before = self.projection_path.read_bytes()

        second = dbmod.bootstrap_storage(self.state_path, self.projection_path)

        self.assertEqual(second["status"], "ready")
        self.assertEqual(second["migration_id"], first["migration_id"])
        self.assertEqual(self.state_path.read_bytes(), state_before)
        self.assertEqual(self.projection_path.read_bytes(), projection_before)
        state = dbmod.connect(self.state_path)
        try:
            self.assertEqual(
                state.execute("SELECT COUNT(*) c FROM legacy_recovery_records").fetchone()["c"],
                2,
            )
        finally:
            state.close()

    def test_new_install_initializes_two_matching_versioned_stores(self):
        result = dbmod.bootstrap_storage(self.state_path, self.projection_path)
        self.assertEqual(result["status"], "initialized")
        state = dbmod.connect(self.state_path)
        projection = dbmod.connect(self.projection_path)
        try:
            self.assertEqual(
                dbmod.storage_manifest(state)["migration_id"],
                dbmod.storage_manifest(projection)["migration_id"],
            )
            with self.assertRaises(sqlite3.OperationalError):
                projection.execute("SELECT * FROM pending_changes").fetchall()
            self.assertIsNotNone(
                state.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' "
                    "AND name='legacy_recovery_records'"
                ).fetchone()
            )
        finally:
            state.close()
            projection.close()

    def test_restart_recovers_partial_and_verified_archive_temporaries(self):
        legacy_bytes, _ = self._legacy_database()
        partial = self.root / f"{dbmod._ARCHIVE_TEMP_PREFIX}partial.tmp"
        partial.write_bytes(legacy_bytes[:100])
        verified = self.root / f"{dbmod._ARCHIVE_TEMP_PREFIX}verified.tmp"
        verified.write_bytes(legacy_bytes)

        result = dbmod.bootstrap_storage(self.state_path, self.projection_path)

        archive = Path(result["archive_path"])
        self.assertEqual(archive.read_bytes(), legacy_bytes)
        self.assertFalse(partial.exists())
        self.assertFalse(verified.exists())
        self._assert_exact_recovery()

    def test_multiple_verified_archive_temporaries_abort_as_ambiguous(self):
        legacy_bytes, _ = self._legacy_database()
        for suffix in ("one", "two"):
            path = self.root / f"{dbmod._ARCHIVE_TEMP_PREFIX}{suffix}.tmp"
            path.write_bytes(legacy_bytes)

        with self.assertRaisesRegex(RuntimeError, "ambiguous verified"):
            dbmod.bootstrap_storage(self.state_path, self.projection_path)

        self.assertEqual(self.state_path.read_bytes(), legacy_bytes)

    def test_restart_before_projection_replacement_keeps_legacy_authoritative(self):
        legacy_bytes, _ = self._legacy_database()
        with mock.patch.object(
            dbmod, "_replace_candidate", side_effect=RuntimeError("before replace")
        ):
            with self.assertRaisesRegex(RuntimeError, "before replace"):
                dbmod.bootstrap_storage(self.state_path, self.projection_path)
        self.assertEqual(self.state_path.read_bytes(), legacy_bytes)

        result = dbmod.bootstrap_storage(self.state_path, self.projection_path)
        self.assertEqual(result["status"], "migrated")
        self._assert_exact_recovery()

    def test_restart_after_projection_before_durable_replacement_is_idempotent(self):
        legacy_bytes, _ = self._legacy_database()
        real_replace = dbmod._replace_candidate
        calls = 0

        def fail_second(candidate, target):
            nonlocal calls
            calls += 1
            if calls == 2:
                raise RuntimeError("before durable replace")
            return real_replace(candidate, target)

        with mock.patch.object(dbmod, "_replace_candidate", side_effect=fail_second):
            with self.assertRaisesRegex(RuntimeError, "before durable replace"):
                dbmod.bootstrap_storage(self.state_path, self.projection_path)
        self.assertEqual(self.state_path.read_bytes(), legacy_bytes)

        result = dbmod.bootstrap_storage(self.state_path, self.projection_path)
        self.assertEqual(result["status"], "migrated")
        self._assert_exact_recovery()

    def test_restart_after_durable_replacement_uses_completed_marker(self):
        self._legacy_database()
        real_replace = dbmod._replace_candidate

        def fail_after_replace(candidate, target):
            real_replace(candidate, target)
            if Path(target) == self.state_path:
                raise RuntimeError("after durable replace")

        with mock.patch.object(dbmod, "_replace_candidate", side_effect=fail_after_replace):
            with self.assertRaisesRegex(RuntimeError, "after durable replace"):
                dbmod.bootstrap_storage(self.state_path, self.projection_path)

        result = dbmod.bootstrap_storage(self.state_path, self.projection_path)
        self.assertEqual(result["status"], "ready")
        self._assert_exact_recovery()


class TestAtomicProjectionPromotion(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory(prefix="wbm-pass4-promotion-")
        self.root = Path(self.tempdir.name)
        self.workbook = self.root / "source.xlsx"
        shutil.copy2(ROOT / "stingray_master.xlsx", self.workbook)
        self.projection = self.root / "workbook_projection.sqlite3"
        conn = dbmod.connect(self.projection)
        try:
            dbmod.init_projection_schema(conn)
            conn.execute(
                "INSERT INTO meta(key, value) VALUES('sentinel', 'prior-projection')"
            )
            conn.commit()
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        finally:
            conn.close()

    def tearDown(self):
        self.tempdir.cleanup()

    def _projection_hash(self) -> str:
        return hashlib.sha256(self.projection.read_bytes()).hexdigest()

    def _assert_complete_managed_row_dispositions(self) -> None:
        """Focused assertions over the already-promoted acceptance projection."""
        from openpyxl import load_workbook

        workbook = load_workbook(self.workbook, read_only=True, data_only=False)
        try:
            classifications = catalog.classify_workbook_sheets(workbook)
            expected_sheets = len(workbook.sheetnames)
            expected = 0
            for sheet_name, classification in classifications.items():
                if classification.spec is None:
                    continue
                for values in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
                    if any(value is not None and value != "" for value in values):
                        expected += 1
        finally:
            workbook.close()
        conn = dbmod.connect(self.projection)
        try:
            actual = conn.execute(
                "SELECT COUNT(*) c FROM managed_row_dispositions"
            ).fetchone()["c"]
            sheet_dispositions = conn.execute(
                "SELECT COUNT(*) c FROM sheet_dispositions"
            ).fetchone()["c"]
            duplicates = conn.execute(
                "SELECT COUNT(*) c FROM ("
                "SELECT sheet, src_row, family, COUNT(*) n "
                "FROM managed_row_dispositions GROUP BY sheet, src_row, family "
                "HAVING n <> 1)"
            ).fetchone()["c"]
            self.assertEqual(actual, expected)
            self.assertEqual(sheet_dispositions, expected_sheets)
            self.assertEqual(duplicates, 0)
        finally:
            conn.close()

    def test_incomplete_asset_rows_block_promotion(self):
        workbook = write_compact_missing_identifier_workbook(
            self.root / "missing-identifier.xlsx"
        )
        before = self._projection_hash()

        with mock.patch.object(dbmod, "_replace_projection") as replace:
            result = importer.promote_verified_projection(workbook, self.projection)

        self.assertFalse(result["promoted"])
        self.assertEqual(result["status"], "blocked")
        self.assertEqual(
            [issue["category"] for issue in result["issues"]],
            ["missing_identifier"],
        )
        replace.assert_not_called()
        self.assertEqual(self._projection_hash(), before)
        self.assertFalse(list(self.root.glob(".wbm-import-candidate-*")))

    def test_unresolved_reference_disposition_records_offending_value(self):
        workbook = write_compact_unresolved_reference_workbook(
            self.root / "unresolved.xlsx"
        )
        missing_target = "missing_target_for_checkpoint_5"

        result = importer.promote_verified_projection(workbook, self.projection)

        self.assertFalse(result["promoted"])
        disposition = next(
            row for row in result["import"]["managed_row_dispositions"]
            if row["sheet"] == "rule_mapping" and row["src_row"] == 2
        )
        self.assertEqual(disposition["disposition"], "excluded")
        self.assertEqual(disposition["field"], "target_id")
        self.assertEqual(disposition["value"], missing_target)
        self.assertTrue(disposition["blocking"])

    def test_post_replace_failure_restores_prior_projection_bytes(self):
        before = self.projection.read_bytes()
        candidate = self.root / "candidate.sqlite3"
        candidate.write_bytes(b"replacement bytes")
        real_replace = dbmod._replace_candidate

        def fail_after_replace(source, target):
            if source == candidate:
                source.replace(target)
                raise OSError("forced post-replace durability failure")
            return real_replace(source, target)

        with mock.patch.object(dbmod, "_replace_candidate", side_effect=fail_after_replace):
            with self.assertRaisesRegex(OSError, "post-replace"):
                dbmod._replace_projection(candidate, self.projection)

        self.assertEqual(self.projection.read_bytes(), before)
        self.assertFalse(list(self.root.glob(".workbook_projection.sqlite3.rollback-*")))

    def test_file_fsync_failure_restores_prior_projection_bytes(self):
        before = self.projection.read_bytes()
        candidate = self.root / "candidate-file-fsync.sqlite3"
        candidate.write_bytes(b"replacement bytes")
        real_fsync = dbmod._fsync_file
        failed = False

        def fail_target_fsync(path):
            nonlocal failed
            if path == self.projection and not failed:
                failed = True
                raise OSError("forced target file fsync failure")
            return real_fsync(path)

        with mock.patch.object(dbmod, "_fsync_file", side_effect=fail_target_fsync):
            with self.assertRaisesRegex(OSError, "file fsync"):
                dbmod._replace_projection(candidate, self.projection)
        self.assertEqual(self.projection.read_bytes(), before)

    def test_directory_fsync_failure_restores_prior_projection_bytes(self):
        before = self.projection.read_bytes()
        candidate = self.root / "candidate-dir-fsync.sqlite3"
        candidate.write_bytes(b"replacement bytes")
        real_fsync = dbmod._fsync_dir
        calls = 0

        def fail_replacement_dir_fsync(path):
            nonlocal calls
            calls += 1
            if calls == 3:
                raise OSError("forced directory fsync failure")
            return real_fsync(path)

        with mock.patch.object(dbmod, "_fsync_dir", side_effect=fail_replacement_dir_fsync):
            with self.assertRaisesRegex(OSError, "directory fsync"):
                dbmod._replace_projection(candidate, self.projection)
        self.assertEqual(self.projection.read_bytes(), before)

    def test_reopen_verification_failure_restores_prior_projection_bytes(self):
        before = self.projection.read_bytes()
        candidate = self.root / "candidate-reopen.sqlite3"
        candidate.write_bytes(b"replacement bytes")

        def fail_verification(_path):
            raise RuntimeError("forced manifest verification failure")

        with self.assertRaisesRegex(RuntimeError, "manifest verification"):
            dbmod._replace_projection(
                candidate,
                self.projection,
                verify=fail_verification,
            )
        self.assertEqual(self.projection.read_bytes(), before)

    def test_residual_sidecar_refuses_replacement_without_deleting_it(self):
        before = self.projection.read_bytes()
        candidate = self.root / "candidate-sidecar.sqlite3"
        candidate.write_bytes(b"replacement bytes")
        sidecar = Path(f"{self.projection}-wal")
        sidecar.write_bytes(b"residual WAL")

        class Checkpoint:
            def execute(self, _sql):
                return self

            def fetchone(self):
                return (1, 1, 0)

            def close(self):
                pass

        with mock.patch.object(dbmod, "connect", return_value=Checkpoint()):
            with self.assertRaisesRegex(
                dbmod.ProjectionBusyError, "could not be fully checkpointed"
            ):
                dbmod._replace_projection(candidate, self.projection)
        self.assertEqual(self.projection.read_bytes(), before)
        self.assertEqual(sidecar.read_bytes(), b"residual WAL")

    def test_malformed_candidate_leaves_prior_projection_byte_identical(self):
        workbook = write_compact_missing_sheet_workbook(
            self.root / "missing-sheet.xlsx"
        )
        before = self._projection_hash()

        result = importer.promote_verified_projection(workbook, self.projection)

        self.assertFalse(result["promoted"])
        self.assertEqual(result["status"], "blocked")
        self.assertEqual(self._projection_hash(), before)
        self.assertFalse(list(self.root.glob(".wbm-import-candidate-*")))

    def test_real_workbook_acceptance_promotes_verified_projection_once(self):
        """Own success, manifest, row reconciliation, and state-store isolation."""
        fixture = verified_manager_fixture()
        fixture.clone_workbook(self.workbook)
        result = fixture.promotion_report
        state = self.root / "workbook_manager.sqlite3"
        dbmod.bootstrap_storage(state, self.projection)
        conn = dbmod.connect(state, foreign_keys=True)
        try:
            conn.execute(
                "INSERT INTO pending_changes(ts, table_name, entity_key_json, op) "
                "VALUES('test', 'options', '{}', 'update')"
            )
            conn.commit()
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        finally:
            conn.close()
        state_before = state.read_bytes()
        fixture.clone_projection(self.projection)

        self.assertTrue(result["promoted"], result)
        self.assertEqual(result["status"], "promoted")
        self.assertEqual(result["blocking_findings"], 0)
        self.assertTrue(result["semantic_readback_verified"])
        self.assertTrue(result["package_valid"])
        self.assertTrue(result["schema_valid"])
        self.assertEqual(
            result["source_sha256"], hashlib.sha256(self.workbook.read_bytes()).hexdigest()
        )
        self.assertFalse(list(self.root.glob(".wbm-import-candidate-*")))

        conn = dbmod.connect(self.projection)
        try:
            self.assertEqual(
                dbmod.get_meta(conn, "workbook_sha256"), result["source_sha256"]
            )
            self.assertEqual(
                dbmod.storage_manifest(conn)["source_sha256"], result["source_sha256"]
            )
        finally:
            conn.close()
        promoted_hash = self._projection_hash()
        restart = dbmod.bootstrap_storage(state, self.projection)
        self.assertEqual(restart["status"], "ready")
        self.assertEqual(self._projection_hash(), promoted_hash)
        self.assertEqual(state.read_bytes(), state_before)
        self._assert_complete_managed_row_dispositions()

    def test_blocking_import_findings_never_call_atomic_replace(self):
        workbook = write_compact_missing_identifier_workbook(
            self.root / "blocking-findings.xlsx"
        )
        before = self._projection_hash()

        with mock.patch.object(dbmod, "_replace_projection") as replace:
            result = importer.promote_verified_projection(workbook, self.projection)

        self.assertFalse(result["promoted"])
        replace.assert_not_called()
        self.assertEqual(self._projection_hash(), before)

    def test_atomic_replace_failure_leaves_prior_projection_byte_identical(self):
        """Own the complete real-workbook atomic-replace fail-closed proof."""
        before = self._projection_hash()

        with mock.patch.object(
            dbmod, "_replace_projection", side_effect=OSError("forced replace failure")
        ):
            result = importer.promote_verified_projection(self.workbook, self.projection)

        self.assertFalse(result["promoted"])
        self.assertEqual(result["status"], "blocked")
        self.assertEqual(result["issues"][0]["category"], "candidate_build_exception")
        self.assertEqual(self._projection_hash(), before)
        self.assertFalse(list(self.root.glob(".wbm-import-candidate-*")))

    def test_source_identity_drift_blocks_before_replacement(self):
        """Own the complete real-workbook source-drift fail-closed proof."""
        before = self._projection_hash()
        real_identity = importer.workbook_identity
        calls = 0

        def drift_on_recheck(path):
            nonlocal calls
            calls += 1
            identity = real_identity(path)
            if calls > 1:
                return {**identity, "mtime_ns": identity["mtime_ns"] + 1}
            return identity

        with mock.patch.object(
            importer, "workbook_identity", side_effect=drift_on_recheck
        ):
            result = importer.promote_verified_projection(self.workbook, self.projection)

        self.assertFalse(result["promoted"])
        self.assertEqual(result["status"], "source_changed")
        self.assertEqual(self._projection_hash(), before)


class TestReconstructionPreservation(unittest.TestCase):
    def test_only_declared_registry_owned_cells_may_change(self):
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory(prefix="wbm-preservation-") as tempdir:
            workbook_path = Path(tempdir) / "source.xlsx"
            workbook = Workbook()
            managed = workbook.create_sheet("managed")
            workbook.remove(workbook["Sheet"])
            managed.append(["id", "owned", "opaque"])
            managed.append(["row_1", "before", "keep"])
            preserved = workbook.create_sheet("preserved")
            preserved["A1"] = "=1+1"
            workbook.save(workbook_path)
            workbook.close()
            before = importer._workbook_preservation_snapshot(workbook_path)

            workbook = load_workbook(workbook_path)
            workbook["managed"]["B2"] = "after"
            workbook.save(workbook_path)
            workbook.close()
            operations = [{
                "sheet": "managed",
                "_src_row": 2,
                "row": {"owned": "after"},
            }]
            self.assertEqual(
                importer._preservation_issues(
                    before,
                    importer._workbook_preservation_snapshot(workbook_path),
                    operations,
                    workbook_path,
                ),
                [],
            )

            workbook = load_workbook(workbook_path)
            workbook["preserved"]["A1"] = "=2+2"
            workbook.save(workbook_path)
            workbook.close()
            issues = importer._preservation_issues(
                before,
                importer._workbook_preservation_snapshot(workbook_path),
                operations,
                workbook_path,
            )
            self.assertIn("reconstruction_cell_drift", {
                issue["category"] for issue in issues
            })


if __name__ == "__main__":
    unittest.main()
