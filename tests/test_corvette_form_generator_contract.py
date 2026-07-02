#!/usr/bin/env python3
# pyright: reportMissingImports=false
"""Focused tests for shared contract-surface helpers."""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.contract import (  # noqa: E402
    ASSET_IMAGE_FIELDS,
    build_body_context_choices,
    build_model_context_choices,
    build_trim_context_choices,
    load_asset_map,
    merge_option_asset_fields,
)


class ContractHelperTests(unittest.TestCase):
    def test_build_model_context_choices_matches_existing_direct_call_shape(self) -> None:
        variants = [
            {
                "variant_id": "1lt_e07",
                "body_style": "coupe",
                "trim_level": "1LT",
                "display_name": "1LT Coupe",
                "base_price": 70000,
                "display_order": 2,
            },
            {
                "variant_id": "2lt_e67",
                "body_style": "convertible",
                "trim_level": "2LT",
                "display_name": "2LT Convertible",
                "base_price": 78000,
                "display_order": 1,
            },
        ]
        copy_rows = [
            {
                "model_key": "test_model",
                "context_type": "body_style",
                "value": "coupe",
                "body_style": "coupe",
                "info_tooltip": "Coupe tooltip",
            },
            {
                "model_key": "test_model",
                "context_type": "trim_level",
                "value": "2LT",
                "body_style": "convertible",
                "info_tooltip": "2LT convertible tooltip",
            },
        ]
        body_style_display_order = {"coupe": 1, "convertible": 2}
        bodystyle_assets = {
            "body_style__coupe": {
                "image_url": "coupe.jpg",
                "image_alt": "Coupe",
                "image_fit": "cover",
                "image_position": "center",
                "hover_image_url": "coupe-hover.jpg",
                "hover_image_alt": "Coupe hover",
                "hover_image_position": "center",
            }
        }

        expected = build_body_context_choices(
            variants,
            copy_rows,
            "test_model",
            body_style_display_order,
            bodystyle_assets,
        ) + build_trim_context_choices(variants, copy_rows, "test_model")

        self.assertEqual(
            build_model_context_choices(
                variants,
                copy_rows,
                "test_model",
                body_style_display_order,
                bodystyle_assets,
            ),
            expected,
        )

    def test_merge_option_asset_fields_copies_only_asset_fields_from_source_row(self) -> None:
        destination = {"option_id": "opt_q9i_001", "image_url": "stale-target.jpg", "label": "Kept"}
        source_rows = {
            "opt_q9i_001": {
                "image_url": "source.jpg",
                "image_alt": "Source image",
                "image_fit": "contain",
                "image_position": "top",
                "hover_image_url": "source-hover.jpg",
                "hover_image_alt": "Source hover",
                "hover_image_position": "bottom",
                "label": "Do not copy",
            }
        }

        merge_option_asset_fields(destination, source_rows, only_if_image_present=False)

        self.assertEqual(
            {field: destination[field] for field in ASSET_IMAGE_FIELDS},
            {field: source_rows["opt_q9i_001"].get(field, "") for field in ASSET_IMAGE_FIELDS},
        )
        self.assertEqual(destination["label"], "Kept")

    def test_merge_option_asset_fields_true_gates_on_source_row_image_url_not_destination(self) -> None:
        destination = {"option_id": "opt_q9i_001", "image_url": "target-only.jpg"}
        source_rows = {
            "opt_q9i_001": {
                "image_url": "",
                "image_alt": "Should not copy",
            }
        }

        merge_option_asset_fields(destination, source_rows, only_if_image_present=True)

        self.assertEqual(destination, {"option_id": "opt_q9i_001", "image_url": "target-only.jpg"})

        source_rows["opt_q9i_001"]["image_url"] = "source.jpg"
        merge_option_asset_fields(destination, source_rows, only_if_image_present=True)

        self.assertEqual(destination["image_url"], "source.jpg")
        self.assertEqual(destination["image_alt"], "Should not copy")

    def test_merge_option_asset_fields_is_noop_without_source_entry(self) -> None:
        destination = {"option_id": "opt_missing_001", "label": "No asset"}

        merge_option_asset_fields(destination, {}, only_if_image_present=False)
        merge_option_asset_fields(destination, {}, only_if_image_present=True)

        self.assertEqual(destination, {"option_id": "opt_missing_001", "label": "No asset"})


def _asset_map_workbook(rows: list[dict[str, object]]):
    from openpyxl import Workbook

    headers = [
        "model_key",
        "target_type",
        "target_id",
        "image_url",
        "image_alt",
        "image_fit",
        "image_position",
        "hover_image_url",
        "hover_image_alt",
        "hover_image_position",
        "active",
        "notes",
    ]
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    ws = wb.create_sheet("asset_map")
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    return wb


class LoadAssetMapWildcardTests(unittest.TestCase):
    def test_wildcard_only_row_applies_to_every_model(self) -> None:
        wb = _asset_map_workbook(
            [
                {"model_key": "*", "target_type": "option", "target_id": "opt_gba_001", "image_url": "shared.png", "active": True},
            ]
        )

        for model_key in ("stingray", "grand_sport", "z06"):
            assets = load_asset_map(wb, model_key)
            self.assertEqual(assets[("option", "opt_gba_001")]["image_url"], "shared.png", model_key)

    def test_exact_row_overlays_wildcard_row_regardless_of_row_order(self) -> None:
        rows = [
            {"model_key": "*", "target_type": "option", "target_id": "opt_j6d_001", "image_url": "shared.png", "active": True},
            {"model_key": "z06", "target_type": "option", "target_id": "opt_j6d_001", "image_url": "z06-exact.png", "active": True},
        ]
        for ordering in (rows, list(reversed(rows))):
            wb = _asset_map_workbook(ordering)
            self.assertEqual(
                load_asset_map(wb, "z06")[("option", "opt_j6d_001")]["image_url"],
                "z06-exact.png",
            )
            self.assertEqual(
                load_asset_map(wb, "stingray")[("option", "opt_j6d_001")]["image_url"],
                "shared.png",
            )

    def test_blank_model_key_row_stays_skipped(self) -> None:
        wb = _asset_map_workbook(
            [
                {"model_key": "", "target_type": "option", "target_id": "opt_gba_001", "image_url": "orphan.png", "active": True},
            ]
        )

        self.assertEqual(load_asset_map(wb, "stingray"), {})

    def test_wildcard_row_with_non_option_target_type_is_ignored(self) -> None:
        wb = _asset_map_workbook(
            [
                {"model_key": "*", "target_type": "context_choice", "target_id": "body_style__coupe", "image_url": "coupe.png", "active": True},
                {"model_key": "*", "target_type": "model", "target_id": "stingray", "image_url": "model.png", "active": True},
            ]
        )

        self.assertEqual(load_asset_map(wb, "stingray"), {})

    def test_inactive_wildcard_row_is_ignored(self) -> None:
        wb = _asset_map_workbook(
            [
                {"model_key": "*", "target_type": "option", "target_id": "opt_gba_001", "image_url": "shared.png", "active": False},
            ]
        )

        self.assertEqual(load_asset_map(wb, "stingray"), {})

    def test_exact_only_behavior_unchanged_with_mixed_case_target_id(self) -> None:
        wb = _asset_map_workbook(
            [
                {"model_key": "stingray", "target_type": "option", "target_id": "OPT_GBA_001", "image_url": "exact.png", "active": True},
            ]
        )

        assets = load_asset_map(wb, "stingray")
        self.assertEqual(assets[("option", "OPT_GBA_001")]["image_url"], "exact.png")
        self.assertEqual(load_asset_map(wb, "z06"), {})


if __name__ == "__main__":
    unittest.main()
