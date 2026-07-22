from __future__ import annotations

import importlib.util
import json
import os
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
ALLOWLIST_PATH = ROOT / "tests/fixtures/options-sheet-quality-allowlist.json"


def test_options_sheet_quality_module_exists() -> None:
    assert importlib.util.find_spec("corvette_form_generator.options_sheet_quality") is not None


def test_options_sheet_quality_allowlist_exists() -> None:
    assert ALLOWLIST_PATH.exists()


OPTION_HEADERS = [
    "option_id",
    "rpo",
    "price",
    "option_name",
    "description",
    "detail_raw",
    "section_id",
    "selectable",
    "display_order",
    "active",
]


def _quality_workbook(tmp_path: Path, rows: list[list[object]]) -> Path:
    path = tmp_path / "quality.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    sources = wb.create_sheet("model_workbook_sources")
    sources.append(["model_key", "source_role", "sheet_name", "active", "notes"])
    sources.append(["zr1", "source_option_sheet", "zr1_options", False, "inactive scaffold"])
    sections = wb.create_sheet("section_master")
    sections.append(["section_id", "selection_mode"])
    sections.append(["sec_std_001", "display_only"])
    sections.append(["sec_opt_001", "single_select_opt"])
    options = wb.create_sheet("zr1_options")
    options.append(OPTION_HEADERS)
    for row in rows:
        options.append(row)
    wb.save(path)
    wb.close()
    return path


def _valid_row(**overrides: object) -> list[object]:
    row = {
        "option_id": "opt_001",
        "rpo": "ABC",
        "price": None,
        "option_name": "Valid Equipment",
        "description": "Ancillary information",
        "detail_raw": "Preserved raw evidence",
        "section_id": "sec_std_001",
        "selectable": False,
        "display_order": 10,
        "active": True,
    }
    row.update(overrides)
    return [row[header] for header in OPTION_HEADERS]


@pytest.mark.parametrize(
    ("overrides", "check_id"),
    [
        (
            {"option_name": "Repeated customer copy", "description": "Repeated customer copy"},
            "option_name_equals_description",
        ),
        (
            {"description": "Duplicated raw evidence", "detail_raw": "Duplicated raw evidence"},
            "description_equals_detail_raw",
        ),
        ({"option_name": "Line one\nLine two"}, "option_name_multiline"),
        ({"option_name": "X" * 61}, "option_name_too_long"),
        ({"option_name": "LPO"}, "bare_lpo_option_name"),
        ({"option_id": "opt_std_0123456789abcdef"}, "hash_derived_option_id"),
        ({"display_order": None}, "active_option_missing_display_order"),
        ({"price": 995}, "standard_option_nonzero_price"),
        (
            {"section_id": "sec_opt_001", "price": None},
            "selectable_section_standard_missing_zero_price",
        ),
    ],
)
def test_quality_lint_reports_reference_proven_predicates_on_inactive_sheet(
    tmp_path: Path,
    overrides: dict[str, object],
    check_id: str,
) -> None:
    from corvette_form_generator.options_sheet_quality import lint_options_sheet_quality

    workbook = _quality_workbook(tmp_path, [_valid_row(**overrides)])

    issues = lint_options_sheet_quality(workbook)

    assert check_id in {issue.check_id for issue in issues}
    issue = next(issue for issue in issues if issue.check_id == check_id)
    assert issue.model == "zr1"
    assert issue.sheet == "zr1_options"
    assert issue.row == 2
    assert issue.option_id == (overrides.get("option_id") or "opt_001")


def test_quality_lint_reports_stub_count_above_reference_band(tmp_path: Path) -> None:
    from corvette_form_generator.options_sheet_quality import lint_options_sheet_quality

    rows = [
        _valid_row(option_id=f"opt_{number:03d}", rpo=f"R{number:02d}", option_name=f"Stub {number}")
        for number in range(1, 8)
    ]
    workbook = _quality_workbook(tmp_path, rows)

    issue = next(
        issue
        for issue in lint_options_sheet_quality(workbook)
        if issue.check_id == "stub_name_count_exceeds_reference_band"
    )

    assert issue.value == 7
    assert issue.row is None


def test_pure_quality_evaluator_grades_complete_projected_rows_without_workbook_io() -> None:
    from corvette_form_generator.options_sheet_quality import evaluate_options_sheet_quality

    row = dict(zip(OPTION_HEADERS, _valid_row(option_name="LPO")))
    issues = evaluate_options_sheet_quality(
        "zr1",
        "zr1_options",
        [row],
        {"sec_std_001": "display_only"},
    )

    assert {issue.check_id for issue in issues} == {"bare_lpo_option_name"}
    assert issues[0].option_id == "opt_001"


def test_display_only_included_row_may_use_explicit_zero_price() -> None:
    from corvette_form_generator.options_sheet_quality import evaluate_options_sheet_quality

    row = dict(zip(OPTION_HEADERS, _valid_row(price=0)))

    assert evaluate_options_sheet_quality(
        "zr1",
        "zr1_options",
        [row],
        {"sec_std_001": "display_only"},
    ) == []


def test_pure_quality_evaluator_reports_each_active_section_order_collision() -> None:
    from corvette_form_generator.options_sheet_quality import evaluate_options_sheet_quality

    rows = [
        dict(zip(OPTION_HEADERS, _valid_row(option_id="opt_001", option_name="First package"))),
        dict(zip(OPTION_HEADERS, _valid_row(option_id="opt_002", option_name="Second package"))),
    ]

    issues = evaluate_options_sheet_quality(
        "zr1",
        "zr1_options",
        rows,
        {"sec_std_001": "display_only"},
    )

    collisions = [issue for issue in issues if issue.check_id == "active_display_order_collision"]
    assert {issue.option_id for issue in collisions} == {"opt_001", "opt_002"}


def test_quality_allowlist_requires_exact_value_and_reason(tmp_path: Path) -> None:
    from corvette_form_generator.options_sheet_quality import lint_options_sheet_quality

    long_name = "A" * 61
    workbook = _quality_workbook(tmp_path, [_valid_row(option_name=long_name)])
    allowlist = tmp_path / "allowlist.json"
    allowlist.write_text(
        json.dumps(
            {
                "schemaVersion": "options-sheet-quality-allowlist-1",
                "entries": [
                    {
                        "model": "zr1",
                        "sheet": "zr1_options",
                        "optionId": "opt_001",
                        "checkId": "option_name_too_long",
                        "value": long_name,
                        "reason": "Reviewed reference exception.",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    assert not any(
        issue.check_id == "option_name_too_long"
        for issue in lint_options_sheet_quality(workbook, allowlist_path=allowlist)
    )

    wb = load_workbook(workbook)
    wb["zr1_options"]["D2"] = "B" * 61
    wb.save(workbook)
    wb.close()
    assert any(
        issue.check_id == "option_name_too_long"
        for issue in lint_options_sheet_quality(workbook, allowlist_path=allowlist)
    )


def test_quality_cli_returns_nonzero_and_machine_readable_findings(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    from corvette_form_generator.options_sheet_quality import main

    workbook = _quality_workbook(tmp_path, [_valid_row(option_name="LPO")])

    result = main(["--workbook", str(workbook), "--json"])
    payload = json.loads(capsys.readouterr().out)

    assert result == 1
    assert payload["status"] == "failed"
    assert payload["issueCount"] >= 1
    assert any(issue["check_id"] == "bare_lpo_option_name" for issue in payload["issues"])


def test_canonical_workbook_options_meet_customer_facing_quality_gate() -> None:
    from corvette_form_generator.options_sheet_quality import lint_options_sheet_quality

    workbook = Path(
        os.environ.get("OPTIONS_SHEET_QUALITY_WORKBOOK", str(ROOT / "stingray_master.xlsx"))
    )
    issues = lint_options_sheet_quality(workbook, allowlist_path=ALLOWLIST_PATH)
    counts = Counter((issue.model, issue.check_id) for issue in issues)

    assert not issues, "\n".join(
        [f"{len(issues)} option-sheet quality findings; run the documented CLI for row details."]
        + [
            f"{model} {check_id}: {count}"
            for (model, check_id), count in sorted(counts.items())
        ]
    )
