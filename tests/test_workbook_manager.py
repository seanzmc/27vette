"""Tests for workbook-manager: import fidelity, validation, staging, sync
batch construction, and comparison export.

Runs with pytest or plain unittest. API-layer tests skip automatically when
fastapi is not installed (install workbook-manager/backend/requirements.txt
into the repo venv to enable them).
"""

from __future__ import annotations

import json
import os
import shutil
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

REPO_ROOT = Path(__file__).resolve().parents[1]
BACKEND = REPO_ROOT / "workbook-manager" / "backend"
for p in (str(BACKEND), str(REPO_ROOT / "scripts")):
    if p not in sys.path:
        sys.path.insert(0, p)

from app import db as dbmod                     # noqa: E402
from app import naming, staging, sync as syncmod  # noqa: E402
from app.catalog import (  # noqa: E402
    SPEC_BY_TABLE,
    TABLE_SPECS,
    classify_workbook_sheets,
)
from app.staging import StagingError            # noqa: E402
from app.validation import find_dependents      # noqa: E402
from workbook_manager_fixtures import (  # noqa: E402
    clone_combined_projection,
    sha256_file,
    verified_manager_fixture,
)

WORKBOOK = REPO_ROOT / "stingray_master.xlsx"

try:
    import fastapi  # noqa: F401
    HAVE_FASTAPI = True
except ImportError:
    HAVE_FASTAPI = False


def fresh_db(tmpdir: Path) -> sqlite3.Connection:
    conn = dbmod.connect(tmpdir / "test.sqlite3")
    dbmod.init_schema(conn)
    return conn


def _sha256(path: Path) -> str:
    return sha256_file(path)


class ImportedWorkbookCase(unittest.TestCase):
    """Focused behavior fixture cloned from the shared verified projection."""

    tmpdir: Path
    conn: sqlite3.Connection
    report: dict

    @classmethod
    def setUpClass(cls):
        fixture = verified_manager_fixture()
        cls.tmpdir = Path(tempfile.mkdtemp(prefix="wbm-test-"))
        projection = cls.tmpdir / "test.sqlite3"
        _, report = clone_combined_projection(projection)
        cls.conn = dbmod.connect(projection)
        cls.report = report
        cls._fixture = fixture

    @classmethod
    def tearDownClass(cls):
        cls.conn.close()
        shutil.rmtree(cls.tmpdir, ignore_errors=True)
        cls._fixture.assert_unmutated()


class TestImportFidelity(ImportedWorkbookCase):
    def _workbook_row_count(self, sheet: str) -> int:
        from openpyxl import load_workbook
        wb = load_workbook(WORKBOOK, read_only=True)
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(it)]
        count = 0
        for row in it:
            if row is None:
                continue
            rec = {h: v for h, v in zip(headers, row) if h}
            if all(v is None or str(v).strip() == "" for v in rec.values()):
                continue
            count += 1
        wb.close()
        return count

    def test_no_silent_row_loss_for_options(self):
        for model, sheet in [("stingray", "stingray_options"),
                             ("grand_sport", "grandSport_options"),
                             ("z06", "z06_options")]:
            db_count = self.conn.execute(
                "SELECT COUNT(*) c FROM options WHERE model_id=?",
                (model,)).fetchone()["c"]
            dup_or_missing = self.conn.execute(
                "SELECT COUNT(*) c FROM import_issues WHERE run_id=? AND "
                "sheet=? AND category IN ('duplicate_id','missing_identifier')",
                (self.report["run"]["id"], sheet)).fetchone()["c"]
            self.assertEqual(db_count + dup_or_missing,
                             self._workbook_row_count(sheet),
                             f"row loss in {sheet}")

    def test_every_sheet_is_classified_without_a_second_preserved_cell_store(self):
        from openpyxl import load_workbook
        wb = load_workbook(WORKBOOK, read_only=True)
        names = set(wb.sheetnames)
        classifications = classify_workbook_sheets(wb)
        wb.close()
        self.assertEqual(
            self.conn.execute("SELECT COUNT(*) c FROM raw_sheet_rows").fetchone()["c"],
            0,
        )
        self.assertEqual(names, set(classifications))

    def test_model_scoped_option_uniqueness_enforced(self):
        dup = self.conn.execute(
            "SELECT model_id, option_id, COUNT(*) c FROM options "
            "GROUP BY model_id, option_id HAVING c > 1").fetchall()
        self.assertEqual([dict(d) for d in dup], [])

    def test_cross_model_option_ids_coexist(self):
        row = self.conn.execute(
            "SELECT COUNT(DISTINCT model_id) c FROM options WHERE "
            "option_id IN (SELECT option_id FROM options GROUP BY option_id "
            "HAVING COUNT(DISTINCT model_id) > 1)").fetchone()
        self.assertGreater(row["c"], 1,
                           "expected shared option_ids across models")

    def test_model_setup_copy_columns_are_managed_and_imported(self):
        expected = {
            "setup_card_subtitle", "setup_eyebrow", "setup_title", "setup_description",
            "setup_fact_1", "setup_fact_2", "setup_fact_3",
        }
        model_spec = SPEC_BY_TABLE["models"]
        self.assertTrue(expected.issubset({column.header for column in model_spec.columns}))
        rows = self.conn.execute(
            "SELECT * FROM models WHERE active IN ('True', '1', 'TRUE', 'true') "
            "ORDER BY model_key"
        ).fetchall()
        self.assertEqual(
            [row["model_key"] for row in rows],
            ["grand_sport", "grand_sport_x", "stingray", "z06", "zr1", "zr1x"],
        )
        for row in rows:
            self.assertTrue(
                all(row[column] for column in expected),
                f"active model {row['model_key']} must have complete setup copy",
            )

    def test_import_reports_are_queryable(self):
        self.assertIn(self.report["run"]["status"],
                      ("imported", "imported_with_issues"))
        for issue in self.report["issues"]:
            self.assertTrue(issue["message"])
            self.assertIn(issue["severity"], ("error", "warning"))

    def test_ovs_rows_reference_known_options(self):
        orphans = self.conn.execute(
            "SELECT COUNT(*) c FROM option_availability oa WHERE NOT EXISTS ("
            "SELECT 1 FROM options o WHERE o.model_id = oa.model_id AND "
            "o.option_id = oa.option_id)").fetchone()["c"]
        reported = self.conn.execute(
            "SELECT COUNT(*) c FROM import_issues WHERE run_id=? AND "
            "table_name='option_availability' AND category='unresolved_ref' "
            "AND field='option_id'",
            (self.report["run"]["id"],)).fetchone()["c"]
        self.assertEqual(orphans, reported,
                         "unresolved OVS references must all be reported")

    def test_shared_physical_source_rows_import_once_with_all_model_contexts(self):
        rows = self.conn.execute(
            "SELECT src_sheet, physical_key, model_context FROM interiors"
        ).fetchall()
        identities = {(row["src_sheet"], row["physical_key"]) for row in rows}
        self.assertEqual(len(rows), len(identities))
        shared = self.conn.execute(
            "SELECT model_context FROM interiors "
            "WHERE src_sheet='lt_interiors' LIMIT 1"
        ).fetchone()
        self.assertGreater(len(json.loads(shared["model_context"])), 1)


class TestNaming(unittest.TestCase):
    def test_humanize_examples(self):
        self.assertEqual(naming.humanize("stingray_exterior_options"),
                         "Stingray Exterior Options")
        self.assertEqual(naming.humanize("grandSport_options"),
                         "Grand Sport Options")
        self.assertEqual(naming.humanize("z06_ovs"), "Z06 OVS")
        self.assertEqual(naming.humanize("LZ_Interiors"), "LZ Interiors")

    def test_display_id_prefix_stripping_is_reversible(self):
        prefix, remainder = naming.strip_prefix("opt_z51_001", ("opt_",))
        self.assertEqual(prefix + remainder, "opt_z51_001")
        self.assertEqual(naming.display_id("opt_z51_001", ("opt_",)),
                         "Z51 001")

    def test_unconfirmed_prefix_is_not_stripped(self):
        self.assertEqual(naming.strip_prefix("sec_pain_001", ("opt_",)),
                         ("", "sec_pain_001"))


class TestStagingWorkflow(ImportedWorkbookCase):
    def setUp(self):
        self.conn.execute("DELETE FROM pending_changes")
        self.conn.execute("DELETE FROM change_history")
        self.conn.commit()

    def test_stage_validate_commit_add_option(self):
        record = {
            "option_id": "opt_test_901", "rpo": "TST", "price": "100",
            "option_name": "Test Option", "description": "", "detail_raw": "",
            "section_id": "sec_pain_001", "selectable": "True",
            "display_order": "999", "active": "True", "display_behavior": "",
        }
        change = staging.stage_change(
            self.conn, table="options", model_id="stingray", op="add",
            key={"option_id": "opt_test_901"}, record=record)
        self.assertEqual(change["status"], "staged")
        result = staging.commit_staged(self.conn, actor="test")
        self.assertTrue(result["ok"], result)
        row = self.conn.execute(
            "SELECT * FROM options WHERE model_id='stingray' AND "
            "option_id='opt_test_901'").fetchone()
        self.assertIsNotNone(row)
        hist = self.conn.execute(
            "SELECT * FROM change_history WHERE entity_id='opt_test_901'"
        ).fetchone()
        self.assertIsNotNone(hist, "committed change must appear in audit")
        self.assertEqual(hist["sync_status"], "pending")

    def test_invalid_reference_is_rejected_with_field_detail(self):
        record = {"option_id": "opt_test_902", "rpo": "TST",
                  "section_id": "sec_does_not_exist", "price": "0",
                  "option_name": "Bad", "selectable": "True",
                  "display_order": "1", "active": "True"}
        with self.assertRaises(StagingError) as ctx:
            staging.stage_change(
                self.conn, table="options", model_id="stingray", op="add",
                key={"option_id": "opt_test_902"}, record=record)
        fields = {e["field"] for e in ctx.exception.errors}
        self.assertIn("section_id", fields)

    def test_duplicate_key_rejected_in_model_scope(self):
        existing = self.conn.execute(
            "SELECT option_id FROM options WHERE model_id='stingray' "
            "LIMIT 1").fetchone()["option_id"]
        with self.assertRaises(StagingError):
            staging.stage_change(
                self.conn, table="options", model_id="stingray", op="add",
                key={"option_id": existing},
                record={"option_id": existing, "option_name": "Dup",
                        "rpo": "X", "price": "0", "selectable": "True",
                        "display_order": "1", "active": "True"})

    def test_delete_with_dependents_cannot_bypass_final_graph_validation(self):
        row = self.conn.execute(
            "SELECT o.option_id FROM options o JOIN option_availability oa "
            "ON oa.model_id=o.model_id AND oa.option_id=o.option_id "
            "WHERE o.model_id='stingray' LIMIT 1").fetchone()
        option_id = row["option_id"]
        with self.assertRaises(StagingError) as ctx:
            staging.stage_change(self.conn, table="options",
                                 model_id="stingray", op="delete",
                                 key={"option_id": option_id}, record=None)
        self.assertIn("dependent", ctx.exception.errors[0]["message"])
        with self.assertRaises(TypeError):
            staging.stage_change(
                self.conn, table="options", model_id="stingray", op="delete",
                key={"option_id": option_id}, record=None,
                confirm_dependencies=True)

    def test_undo_before_commit(self):
        change = staging.stage_change(
            self.conn, table="exclusive_groups", model_id="stingray",
            op="add", key={"group_id": "xg_test_1"},
            record={"group_id": "xg_test_1",
                    "selection_mode": "single_within_group",
                    "active": "True", "notes": ""})
        discarded = staging.discard_change(self.conn, change["id"])
        self.assertEqual(discarded["status"], "discarded")
        result = staging.commit_staged(self.conn)
        self.assertEqual(result["status"], "empty")

    def test_read_only_table_rejected(self):
        with self.assertRaises(StagingError) as ctx:
            staging.stage_change(
                self.conn, table="form_sections", model_id="", op="add",
                key={"section_id": "sec_test_1"},
                record={"section_id": "sec_test_1"})
        self.assertIn("read-only", ctx.exception.errors[0]["message"])

    def test_key_rename_on_update_rejected(self):
        """Keys are immutable on update (verifier finding 2026-07-15): the
        workbook write path cannot rename keys, so the DB must not either."""
        row = self.conn.execute(
            "SELECT * FROM options WHERE model_id='stingray' LIMIT 1"
        ).fetchone()
        record = {c.sql_name(): row[c.sql_name()]
                  for c in SPEC_BY_TABLE["options"].columns}
        record["option_id"] = "opt_renamed_999"
        with self.assertRaises(StagingError) as ctx:
            staging.stage_change(
                self.conn, table="options", model_id="stingray", op="update",
                key={"option_id": row["option_id"]}, record=record)
        self.assertIn("cannot change on update",
                      ctx.exception.errors[0]["message"])

    def test_duplicate_staged_adds_fail_batch_validation_not_commit(self):
        """Two staged adds with the same key must fail Validate All cleanly
        (verifier finding 2026-07-15), not explode at commit."""
        rec = {"group_id": "xg_dup_1",
               "selection_mode": "single_within_group",
               "active": "True", "notes": ""}
        for _ in range(2):
            staging.stage_change(
                self.conn, table="exclusive_groups", model_id="stingray",
                op="add", key={"group_id": "xg_dup_1"}, record=dict(rec))
        result = staging.commit_staged(self.conn)
        self.assertFalse(result["ok"])
        self.assertEqual(result["status"], "invalid")
        messages = [e["message"] for r in result["validation"]["results"]
                    for e in r["errors"]]
        self.assertTrue(any("duplicate key with staged change" in m
                            for m in messages), messages)
        # nothing committed, nothing in history
        self.assertEqual(self.conn.execute(
            "SELECT COUNT(*) c FROM change_history").fetchone()["c"], 0)
        for c in staging.list_changes(self.conn, "staged"):
            staging.discard_change(self.conn, c["id"])

    def test_unknown_model_content_rejected(self):
        with self.assertRaises(StagingError) as ctx:
            staging.stage_change(
                self.conn, table="options", model_id="not_a_model", op="add",
                key={"option_id": "opt_test_910"},
                record={"option_id": "opt_test_910"})
        self.assertIn("unknown model", ctx.exception.errors[0]["message"])

    def test_active_source_backed_unpublished_model_is_editable(self):
        source = self.conn.execute(
            "SELECT * FROM options WHERE model_id='zr1' LIMIT 1"
        ).fetchone()
        record = {
            column.sql_name(): source[column.sql_name()]
            for column in SPEC_BY_TABLE["options"].columns
        }
        record["option_id"] = "opt_test_911"
        change = staging.stage_change(
            self.conn,
            table="options",
            model_id="zr1",
            op="add",
            key={"option_id": record["option_id"]},
            record=record,
        )
        self.assertEqual(change["status"], "staged")
        staging.discard_change(self.conn, change["id"])

    def test_runtime_promotion_requires_canonical_model_generatability(self):
        model = "zr1"
        source = self.conn.execute(
            "SELECT * FROM model_registry_promotion WHERE model_key=?", (model,)
        ).fetchone()
        record = {
            column.sql_name(): source[column.sql_name()]
            for column in SPEC_BY_TABLE["model_registry_promotion"].columns
        }
        record["active"] = "True"
        record["promoted_to_runtime"] = "True"
        with mock.patch.object(
            staging, "discover_generation_model_configs", return_value={}
        ):
            with self.assertRaises(StagingError) as ctx:
                staging.stage_change(
                    self.conn,
                    table="model_registry_promotion",
                    model_id=model,
                    op="update",
                    key={"model_key": model},
                    record=record,
                )
        self.assertIn("not generatable", ctx.exception.errors[0]["message"])

    def test_inactive_model_content_rejected_but_topology_editable(self):
        model = "grand_sport_x"
        original_active = self.conn.execute(
            "SELECT active FROM models WHERE model_key=?", (model,)
        ).fetchone()["active"]
        self.conn.execute("UPDATE models SET active='False' WHERE model_key=?", (model,))
        self.conn.commit()
        try:
            source = self.conn.execute(
                "SELECT * FROM options WHERE model_id=? LIMIT 1", (model,)
            ).fetchone()
            record = {
                column.sql_name(): source[column.sql_name()]
                for column in SPEC_BY_TABLE["options"].columns
            }
            record["option_id"] = "opt_test_912"
            with self.assertRaises(StagingError) as ctx:
                staging.stage_change(
                    self.conn,
                    table="options",
                    model_id=model,
                    op="add",
                    key={"option_id": record["option_id"]},
                    record=record,
                )
            self.assertIn("inactive model", ctx.exception.errors[0]["message"])

            topology = self.conn.execute(
                "SELECT * FROM model_variants WHERE model_key=? LIMIT 1", (model,)
            ).fetchone()
            topology_record = {
                column.sql_name(): topology[column.sql_name()]
                for column in SPEC_BY_TABLE["model_variants"].columns
            }
            change = staging.stage_change(
                self.conn,
                table="model_variants",
                model_id=model,
                op="update",
                key={
                    "model_key": topology["model_key"],
                    "variant_id": topology["variant_id"],
                },
                record=topology_record,
            )
            staging.discard_change(self.conn, change["id"])
        finally:
            self.conn.execute(
                "UPDATE models SET active=? WHERE model_key=?", (original_active, model)
            )
            self.conn.commit()

    def test_asset_wildcard_is_the_only_writable_wildcard_model_scope(self):
        change = staging.stage_change(
            self.conn,
            table="assets",
            model_id="*",
            op="add",
            key={"model_key": "*", "target_type": "option", "target_id": "opt_test_asset"},
            record={
                "model_key": "*",
                "target_type": "option",
                "target_id": "opt_test_asset",
                "image_url": "https://example.invalid/test.png",
                "image_alt": "Test",
                "image_fit": "contain",
                "image_position": "center",
                "hover_image_url": "",
                "hover_image_alt": "",
                "hover_image_position": "",
                "active": "True",
                "notes": "test",
            },
        )
        staging.discard_change(self.conn, change["id"])
        with self.assertRaises(StagingError):
            staging.stage_change(
                self.conn,
                table="default_selection_rules",
                model_id="*",
                op="add",
                key={"model_key": "*", "rule_id": "default_test"},
                record={"model_key": "*", "rule_id": "default_test"},
            )

    def test_inactive_source_role_rejected_while_fixed_model_content_is_editable(self):
        model = "zr1"
        role = "source_option_sheet"
        original = self.conn.execute(
            "SELECT active FROM sheet_registry WHERE model_key=? AND source_role=?",
            (model, role),
        ).fetchone()["active"]
        self.conn.execute(
            "UPDATE sheet_registry SET active='False' WHERE model_key=? AND source_role=?",
            (model, role),
        )
        self.conn.commit()
        try:
            source = self.conn.execute(
                "SELECT * FROM options WHERE model_id=? LIMIT 1", (model,)
            ).fetchone()
            record = {
                column.sql_name(): source[column.sql_name()]
                for column in SPEC_BY_TABLE["options"].columns
            }
            record["option_id"] = "opt_test_913"
            with self.assertRaises(StagingError) as ctx:
                staging.stage_change(
                    self.conn,
                    table="options",
                    model_id=model,
                    op="add",
                    key={"option_id": record["option_id"]},
                    record=record,
                )
            self.assertIn("inactive source role", ctx.exception.errors[0]["message"])

            fixed = self.conn.execute(
                "SELECT * FROM form_steps WHERE model_key=? LIMIT 1", (model,)
            ).fetchone()
            fixed_record = {
                column.sql_name(): fixed[column.sql_name()]
                for column in SPEC_BY_TABLE["form_steps"].columns
            }
            change = staging.stage_change(
                self.conn,
                table="form_steps",
                model_id=model,
                op="update",
                key={"model_key": model, "step_key": fixed["step_key"]},
                record=fixed_record,
            )
            staging.discard_change(self.conn, change["id"])
        finally:
            self.conn.execute(
                "UPDATE sheet_registry SET active=? WHERE model_key=? AND source_role=?",
                (original, model, role),
            )
            self.conn.commit()


class TestSyncBatch(ImportedWorkbookCase):
    def setUp(self):
        self.conn.execute("DELETE FROM pending_changes")
        self.conn.execute("DELETE FROM change_history")
        self.conn.commit()

    def _commit_price_edit(self):
        row = self.conn.execute(
            "SELECT * FROM options WHERE model_id='stingray' AND rpo='Z51'"
        ).fetchone()
        record = {c.sql_name(): row[c.sql_name()]
                  for c in SPEC_BY_TABLE["options"].columns}
        record["price"] = str(int(record["price"] or 0) + 1)
        staging.stage_change(
            self.conn, table="options", model_id="stingray", op="update",
            key={"option_id": row["option_id"]}, record=record)
        result = staging.commit_staged(self.conn, actor="test")
        self.assertTrue(result["ok"], result)
        return row, record

    def _commit_model_setup_edit(self):
        row = self.conn.execute(
            "SELECT * FROM models WHERE model_key='stingray'"
        ).fetchone()
        record = {c.sql_name(): row[c.sql_name()]
                  for c in SPEC_BY_TABLE["models"].columns}
        record["setup_title"] = f"{record['setup_title']} (reviewed)"
        staging.stage_change(
            self.conn, table="models", model_id="", op="update",
            key={"model_key": row["model_key"]}, record=record)
        result = staging.commit_staged(self.conn, actor="test")
        self.assertTrue(result["ok"], result)
        return row, record

    def test_batch_targets_registered_sheet_with_header_names(self):
        row, record = self._commit_price_edit()
        batch = syncmod.build_batch(self.conn, WORKBOOK)
        self.assertEqual(len(batch["items"]), 1)
        op = batch["items"][0]
        self.assertEqual(op["sheet"], "stingray_options")
        self.assertEqual(op["action"], "update")
        self.assertEqual(op["key"], {"option_id": row["option_id"]})
        self.assertEqual(op["row"]["price"], record["price"])

    def test_model_setup_edit_builds_model_master_editor_op(self):
        row, record = self._commit_model_setup_edit()
        batch = syncmod.build_batch(self.conn, WORKBOOK)
        self.assertEqual(len(batch["items"]), 1)
        op = batch["items"][0]
        self.assertEqual(op["sheet"], "model_master")
        self.assertEqual(op["action"], "update")
        self.assertEqual(op["key"], {"model_key": row["model_key"]})
        self.assertEqual(op["row"]["setup_title"], record["setup_title"])

    def test_dry_run_batch_passes_editor_ops_validation(self):
        """Fast slice of the gate: batch preparation + validation only."""
        from corvette_form_generator import editor_ops
        self._commit_price_edit()
        batch = syncmod.build_batch(self.conn, WORKBOOK)
        extract = editor_ops.extract_workbook(WORKBOOK)
        errors, warnings, prepared = editor_ops._prepare_batch(
            extract, {"items": batch["items"]})
        self.assertEqual(errors, [])
        self.assertEqual(len(prepared), 1)

    @unittest.skipUnless(os.environ.get("WBM_SLOW_GATE") == "1",
                         "full dry-run gate is slow; set WBM_SLOW_GATE=1")
    def test_dry_run_through_editor_ops_gate(self):
        # Copy the workbook so even a bug cannot touch the real file.
        wb_copy = Path(self.tmpdir) / "scratch.xlsx"
        shutil.copy2(WORKBOOK, wb_copy)
        self._commit_price_edit()
        result = syncmod.sync_workbook(self.conn, wb_copy, write=False)
        self.assertEqual(result.get("status"), "validated",
                         f"dry-run failed: {result}")
        # dry run must not mark anything synced
        pending = self.conn.execute(
            "SELECT COUNT(*) c FROM change_history WHERE "
            "sync_status='pending'").fetchone()["c"]
        self.assertEqual(pending, 1)

    @unittest.skipUnless(os.environ.get("WBM_SLOW_GATE") == "1",
                         "full live-write gate is slow; set WBM_SLOW_GATE=1")
    def test_live_write_on_scratch_copy_creates_backup_and_marks_synced(self):
        from app import config

        wb_copy = Path(self.tmpdir) / "scratch3.xlsx"
        shutil.copy2(WORKBOOK, wb_copy)
        self._commit_price_edit()
        original_log_path = config.EDIT_LOG_PATH
        config.EDIT_LOG_PATH = Path(self.tmpdir) / "workbook-edit-log.jsonl"
        try:
            result = syncmod.sync_workbook(
                self.conn, wb_copy, write=True,
                expected_mtime_ns=str(wb_copy.stat().st_mtime_ns))
        finally:
            config.EDIT_LOG_PATH = original_log_path
        self.assertEqual(result.get("status"), "applied",
                         f"live write failed: {result}")
        self.assertTrue(result.get("backupPath"))
        pending = self.conn.execute(
            "SELECT COUNT(*) c FROM change_history WHERE "
            "sync_status='pending'").fetchone()["c"]
        self.assertEqual(pending, 0)

    def test_live_write_requires_matching_mtime(self):
        wb_copy = Path(self.tmpdir) / "scratch2.xlsx"
        shutil.copy2(WORKBOOK, wb_copy)
        self._commit_price_edit()
        result = syncmod.sync_workbook(self.conn, wb_copy, write=True,
                                       expected_mtime_ns="1")
        self.assertEqual(result["status"], "stale")


class TestComparisonExport(ImportedWorkbookCase):
    """Real acceptance exports; the unchanged success is executed exactly once."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from app import config  # type: ignore[import-not-found]

        fixture = verified_manager_fixture()
        cls.workbook = cls.tmpdir / "valid-source.xlsx"
        fixture.clone_workbook(cls.workbook)
        cls._previous_config = (
            config.VAR_DIR,
            config.EXPORT_DIR,
            config.DB_BACKUP_DIR,
        )
        config.VAR_DIR = cls.tmpdir / "var"
        config.EXPORT_DIR = config.VAR_DIR / "per-test-exports"
        config.DB_BACKUP_DIR = config.VAR_DIR / "db-backups"

    # The unchanged export is a real comparison export of the whole workbook,
    # measured at 67.91s. Only the acceptance test below reads it, but building
    # it in setUpClass charged it to all three tests, so a shard running just
    # the overlay proof paid 68s for an export it never opened.
    # unchanged_export_result() is process-cached, so this is still executed
    # exactly once, and _build_unchanged_export saves and restores config
    # itself, so calling it from a test rather than from setUpClass is safe.
    @property
    def unchanged_export(self) -> dict:
        return verified_manager_fixture().unchanged_export_result()

    @property
    def unchanged_export_dir(self) -> Path:
        return Path(self.unchanged_export["path"]).parent

    @classmethod
    def tearDownClass(cls):
        from app import config  # type: ignore[import-not-found]

        config.VAR_DIR, config.EXPORT_DIR, config.DB_BACKUP_DIR = cls._previous_config
        super().tearDownClass()

    def test_acceptance_export_is_disposable_and_preserves_unchanged_workbook(self):
        """Own the complete real-workbook unchanged comparison-export proof."""
        result = self.unchanged_export
        self.assertTrue(result["ok"], result)
        self.assertTrue(result["disposable"])
        self.assertIn("DISPOSABLE-comparison-", Path(result["path"]).name)
        self.assertTrue(
            Path(result["path"]).is_relative_to(self.unchanged_export_dir)
        )
        self.assertFalse(result["generated_contract_parity_verified"])
        self.assertTrue(result["byte_identical"])
        self.assertEqual(Path(result["path"]).read_bytes(), self.workbook.read_bytes())
        from openpyxl import load_workbook
        orig = load_workbook(self.workbook, read_only=True)
        regen = load_workbook(result["path"], read_only=True)
        try:
            self.assertEqual(set(orig.sheetnames), set(regen.sheetnames),
                             "regenerated workbook must keep every sheet")
            # unmanaged sheet content preserved verbatim (PriceRef)
            def rows(wb, sheet):
                return [tuple(str(c) if c is not None else "" for c in r)
                        for r in wb[sheet].iter_rows(values_only=True)]
            self.assertEqual(rows(orig, "PriceRef"), rows(regen, "PriceRef"))
            # managed sheet keeps its row count
            self.assertEqual(len(rows(orig, "stingray_options")),
                             len(rows(regen, "stingray_options")))
            self.assertEqual(
                rows(orig, "model_master"),
                rows(regen, "model_master"),
                "managed model_master setup copy must round-trip without drift",
            )
        finally:
            orig.close()
            regen.close()

    def test_export_overlays_registry_owned_projection_fields(self):
        """Own the complete real-workbook changed-overlay acceptance proof."""
        from app import config
        from openpyxl import load_workbook

        config.VAR_DIR = self.tmpdir / "var"
        config.EXPORT_DIR = config.VAR_DIR / "exports"
        config.DB_BACKUP_DIR = config.VAR_DIR / "db-backups"
        row = self.conn.execute(
            "SELECT id, src_sheet, src_row FROM sheet_registry "
            "ORDER BY id LIMIT 1"
        ).fetchone()
        self.assertIsNotNone(row)
        value = "pass4-overlay-proof"
        self.conn.execute(
            "UPDATE sheet_registry SET notes=? WHERE id=?",
            (value, row["id"]),
        )
        self.conn.commit()

        result = syncmod.export_comparison_workbook(self.conn, self.workbook)

        self.assertTrue(result["ok"], result)
        self.assertFalse(result["byte_identical"])
        self.assertEqual(result["rewritten"], {"registry_owned_fields": "overlaid"})
        exported = load_workbook(result["path"], read_only=True, data_only=False)
        try:
            sheet = exported[row["src_sheet"]]
            headers = [cell.value for cell in sheet[1]]
            column = headers.index("notes") + 1
            self.assertEqual(sheet.cell(row["src_row"], column).value, value)
        finally:
            exported.close()

    def test_export_refuses_source_drift_during_copy(self):
        """Own the real-workbook comparison-export source-drift refusal."""
        import os
        from unittest import mock
        from app import config

        config.VAR_DIR = self.tmpdir / "var"
        config.EXPORT_DIR = config.VAR_DIR / "exports"
        original = self.workbook.read_bytes()
        mtime_ns = self.workbook.stat().st_mtime_ns
        real_copy = shutil.copy2

        def copy_then_drift(source, destination):
            result = real_copy(source, destination)
            self.workbook.write_bytes(original + b"drift")
            os.utime(self.workbook, ns=(mtime_ns, mtime_ns))
            return result

        try:
            with mock.patch.object(syncmod.shutil, "copy2", side_effect=copy_then_drift):
                result = syncmod.export_comparison_workbook(self.conn, self.workbook)
            self.assertFalse(result["ok"])
            self.assertEqual(result["status"], "stale")
            self.assertFalse(list(config.EXPORT_DIR.glob("*.candidate.xlsx")))
        finally:
            self.workbook.write_bytes(original)
            os.utime(self.workbook, ns=(mtime_ns, mtime_ns))


class TestDependencyInspection(ImportedWorkbookCase):
    def test_option_dependents_span_tables(self):
        row = self.conn.execute(
            "SELECT o.option_id FROM options o JOIN exclusive_group_members "
            "egm ON egm.model_id=o.model_id AND egm.option_id=o.option_id "
            "WHERE o.model_id='stingray' LIMIT 1").fetchone()
        deps = find_dependents(self.conn, SPEC_BY_TABLE["options"],
                               "stingray", {"option_id": row["option_id"]})
        tables = {d["table"] for d in deps}
        self.assertIn("exclusive_group_members", tables)
        for d in deps:
            self.assertTrue(d["entity_key"])
            self.assertTrue(d["src_sheet"])


class TestPass1BrowserContainment(unittest.TestCase):
    def test_browser_uses_one_guarded_durable_apply_rebuild_control(self):
        app_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                      "App.jsx").read_text()
        sync_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                       "components" / "ChangesSync.jsx").read_text()
        operations_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                             "components" / "ModelOperations.jsx").read_text()
        structure_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                            "components" / "FormStructure.jsx").read_text()
        self.assertIn("Guarded workbook workflow", app_source)
        self.assertIn("Lock Draft for Validation", sync_source)
        self.assertIn("Approve Validated Changes", sync_source)
        self.assertIn("Write Approved Changes & Rebuild Form Data", sync_source)
        self.assertIn("api.applyRebuildDraft", sync_source)
        self.assertIn("APPLY AND REBUILD", sync_source)
        self.assertIn("api.saveDraftOperation", operations_source)
        self.assertIn("api.saveDraftOperation", structure_source)
        self.assertIn("await loadRows();\n    await onChanged();", operations_source)
        self.assertIn("await load(modelKey);\n    onChanged();", structure_source)
        self.assertNotIn("onChanged({ draft: Boolean(operation) })", operations_source)
        self.assertNotIn("onChanged({ draft: Boolean(operation) })", structure_source)
        self.assertNotIn("api.stage", operations_source)
        self.assertNotIn("api.stage", structure_source)
        self.assertNotIn("api.changes", sync_source)
        self.assertNotIn("api.commit(", sync_source)
        self.assertNotIn("api.sync", sync_source)
        self.assertNotIn("write: true", sync_source)
        self.assertNotIn("Write to stingray_master.xlsx", sync_source)

    def test_asset_workspace_is_lazy_bounded_and_uses_shared_draft_only(self):
        app_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                      "App.jsx").read_text()
        asset_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                        "components" / "AssetManager.jsx").read_text()
        self.assertIn('id: "assets", label: "Images"', app_source)
        self.assertIn("<AssetManager", app_source)
        self.assertIn("PAGE_SIZE = 24", asset_source)
        self.assertIn('loading="lazy"', asset_source)
        self.assertIn("data.controls?.image_fit", asset_source)
        self.assertIn("fitValues.map", asset_source)
        self.assertIn('/^[\\w\\s.%/-]+$/.test(position)', asset_source)
        self.assertIn("Show body-style hover media", asset_source)
        self.assertIn("not regenerated runtime proof", asset_source)
        self.assertIn("Add all safe matches to draft", asset_source)
        self.assertIn("Use explicitly selected candidate", asset_source)
        self.assertIn("Advanced: use manual URL", asset_source)
        self.assertIn("Assign to selected target", asset_source)
        self.assertIn("Ignore this exact media identity", asset_source)
        self.assertIn("Save presentation edits to draft", asset_source)
        self.assertIn("api.saveAssetResolution", asset_source)
        self.assertNotIn("api.apply", asset_source)
        self.assertNotIn("write: true", asset_source)

    def test_form_payload_uses_schema_model_context_for_model_key_families(self):
        record_form = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                       "components" / "RecordForm.jsx").read_text()
        structure = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                     "components" / "FormStructure.jsx").read_text()
        self.assertIn("schema.model_context?.required", record_form)
        self.assertIn("schema.model_context.value || modelKey", record_form)
        self.assertNotIn("model_id: schema.model_scoped ? modelKey", record_form)
        self.assertIn("saveFn={saveDraft}", structure)
        self.assertNotIn('model_id: ""', structure)

    def test_form_controls_follow_registry_controls_and_keep_setup_copy(self):
        record_form = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                       "components" / "RecordForm.jsx").read_text()
        structure = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                     "components" / "FormStructure.jsx").read_text()
        self.assertIn("CONTROL_RENDERERS[control.kind]", record_form)
        self.assertIn("Unsupported control kind", record_form)
        self.assertIn("api.referenceOptions", record_form)
        self.assertNotIn("field_kind", record_form)
        self.assertNotIn("finite_values", record_form)
        self.assertIn("Not specified / inherit", record_form)
        self.assertIn("Edit model metadata &amp; Vehicle Setup copy", structure)

    def test_checkpoint_one_shell_is_readiness_first_and_explorers_are_read_only(self):
        app_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                      "App.jsx").read_text()
        explorer_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                           "components" / "ConnectedExplorer.jsx").read_text()
        operations_source = (REPO_ROOT / "workbook-manager" / "frontend" / "src" /
                             "components" / "ModelOperations.jsx").read_text()

        for state in (
            "Starting Workbook Manager",
            "Loading and checking workbook data",
            "Ready to edit",
            "Workbook changed—reload latest data",
            "Draft requires attention",
            "Workbook recovery required",
            "Cannot reach Manager backend",
        ):
            self.assertIn(state, app_source)
        self.assertIn('id: "overview", label: "Form Overview"', app_source)
        self.assertIn('id: "options", label: "Options & Relationships"', app_source)
        self.assertIn('id: "groups", label: "Groups"', app_source)
        self.assertIn('id: "advanced", label: "Advanced & Recovery"', app_source)
        self.assertIn("System details", app_source)
        self.assertIn("Reload Latest Workbook Data", app_source)
        self.assertLess(
            app_source.index("await api.draftLifecycle(id)"),
            app_source.index("setDraftLifecycle(lifecycle)"),
        )
        # Exact-draft navigation must request /api/drafts/{id} directly and
        # never gate on the bounded /api/drafts list window (history records
        # can point outside it).
        self.assertNotIn("listed.drafts.some", app_source)
        self.assertIn("e.status === 404", app_source)
        self.assertIn("api.connectedOption", explorer_source)
        self.assertIn("api.connectedGroup", explorer_source)
        self.assertIn("api.explorerSearch", explorer_source)
        self.assertIn("api.explorerDiagnostics", explorer_source)
        self.assertIn("Where this option is used", explorer_source)
        self.assertIn("Show option relationships", explorer_source)
        self.assertIn("Where this group is used", explorer_source)
        self.assertNotIn("saveDraftOperation", explorer_source)
        self.assertNotIn("Record Comparison", operations_source)
        self.assertNotIn('type="checkbox"', operations_source)


@unittest.skipUnless(HAVE_FASTAPI, "fastapi not installed; install "
                     "workbook-manager/backend/requirements.txt")
class TestApi(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        import os
        cls.tmpdir = Path(tempfile.mkdtemp(prefix="wbm-api-"))
        fixture = verified_manager_fixture()
        cls.workbook = cls.tmpdir / "source.xlsx"
        fixture.clone_workbook(cls.workbook)
        cls.previous_workbook_env = os.environ.get("WBM_WORKBOOK")
        cls.previous_asset_media_env = os.environ.get("WBM_ASSET_MEDIA_URL_LIST")
        cls.previous_apply_output_env = os.environ.get("WBM_APPLY_OUTPUT_ROOT")
        cls.previous_projection_env = os.environ.get("WBM_PROJECTION_DB")
        cls.apply_output_root = cls.tmpdir / "apply-output"
        shutil.copytree(REPO_ROOT / "form-output", cls.apply_output_root / "form-output")
        (cls.apply_output_root / "form-app").mkdir(parents=True)
        for name in ("data.js", "index.html"):
            shutil.copy2(
                REPO_ROOT / "form-app" / name,
                cls.apply_output_root / "form-app" / name,
            )
        os.environ["WBM_DB"] = str(cls.tmpdir / "api.sqlite3")
        os.environ["WBM_VAR_DIR"] = str(cls.tmpdir / "var")
        os.environ["WBM_WORKBOOK"] = str(cls.workbook)
        os.environ["WBM_PROJECTION_DB"] = str(cls.tmpdir / "workbook_projection.sqlite3")
        os.environ["WBM_APPLY_OUTPUT_ROOT"] = str(cls.apply_output_root)
        os.environ["WBM_ASSET_MEDIA_URL_LIST"] = str(
            REPO_ROOT / "tests" / "fixtures" / "asset-map-sync-media-urls.txt"
        )
        # Force a FULL re-import of the app package so config re-reads the
        # env vars above. The bare "app" entry must be deleted too: leaving
        # the package object in sys.modules makes `from . import staging`
        # resolve to the stale module via the package attribute while
        # `from .staging import StagingError` re-imports a fresh copy —
        # two StagingError classes, and main.py's `except StagingError`
        # misses the raise (500 instead of 422).
        for mod in list(sys.modules):
            if mod == "app" or mod.startswith("app."):
                del sys.modules[mod]
        from fastapi.testclient import TestClient
        from app import main as mainmod
        cls.mainmod = mainmod
        # The client must be entered as a context manager: storage bootstrap
        # runs in the FastAPI lifespan, not lazily inside a request.
        cls.client = TestClient(mainmod.app)
        cls.client.__enter__()
        fixture.clone_projection(mainmod.config.DEFAULT_PROJECTION_DB)
        fixture.assert_unmutated()

    @classmethod
    def tearDownClass(cls):
        import os

        cls.client.__exit__(None, None, None)
        if cls.previous_workbook_env is None:
            os.environ.pop("WBM_WORKBOOK", None)
        else:
            os.environ["WBM_WORKBOOK"] = cls.previous_workbook_env
        if cls.previous_asset_media_env is None:
            os.environ.pop("WBM_ASSET_MEDIA_URL_LIST", None)
        else:
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = cls.previous_asset_media_env
        if cls.previous_apply_output_env is None:
            os.environ.pop("WBM_APPLY_OUTPUT_ROOT", None)
        else:
            os.environ["WBM_APPLY_OUTPUT_ROOT"] = cls.previous_apply_output_env
        if cls.previous_projection_env is None:
            os.environ.pop("WBM_PROJECTION_DB", None)
        else:
            os.environ["WBM_PROJECTION_DB"] = cls.previous_projection_env
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_checkpoint_2a_graph_and_atomic_plan_api(self):
        variants = self.client.get("/api/graph/option-create/z06")
        self.assertEqual(variants.status_code, 200, variants.text)
        self.assertEqual(
            [row["variant_id"] for row in variants.json()["active_variants"]],
            ["1lz_h07", "2lz_h07", "3lz_h07", "1lz_h67", "2lz_h67", "3lz_h67"],
        )

        plan = self.client.post(
            "/api/records/options/dependency-plan",
            params={"draft_id": "api-cp2a-plan"},
            json={"model_id": "stingray", "key": {"option_id": "opt_pcx_001"}},
        )
        self.assertEqual(plan.status_code, 200, plan.text)
        self.assertGreater(plan.json()["count"], 1)
        self.assertTrue(all(
            item["selected_action"] == "keep"
            for item in plan.json()["dependents"]
        ))

        response = self.client.post(
            "/api/drafts/api-cp2a-plan/operation-plan",
            json={
                "actor": "api-test",
                "session_id": "browser",
                "operations": [
                    {
                        "table": "options", "model_id": "stingray", "op": "update",
                        "key": {"option_id": "opt_pcx_001"},
                        "record": {"option_name": "Checkpoint 2A API test"},
                    }
                ],
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(len(response.json()["operations"]), 1)
        self.client.delete(
            f"/api/drafts/api-cp2a-plan/operations/{response.json()['operations'][0]['id']}"
        )

    def test_status_and_models(self):
        status = self.client.get("/api/status").json()
        self.assertGreater(status["tables"]["options"], 900)
        models = self.client.get("/api/models").json()["models"]
        keys = {m["model_key"] for m in models}
        self.assertLessEqual({"stingray", "grand_sport", "z06"}, keys)

    def test_asset_reconciliation_api_is_exact_shared_read_only_view(self):
        from corvette_form_generator import asset_map_sync

        fixture = REPO_ROOT / "tests" / "fixtures" / "asset-map-sync-media-urls.txt"
        snapshot = asset_map_sync.build_asset_manager_snapshot(
            self.workbook,
            asset_map_sync.read_media_url_list(fixture),
            media_source=f"media-url-list:{fixture.name}",
        )
        expected = asset_map_sync.filter_asset_manager_snapshot(
            snapshot,
            model_key="stingray",
            status="safe_proposal",
            offset=0,
            limit=24,
        )
        workbook_before = _sha256(self.workbook)
        drafts_before = self.client.get("/api/drafts", params={"limit": 200}).json()
        protected_paths = [
            self.mainmod.config.DEFAULT_PROJECTION_DB,
            self.mainmod.config.DEFAULT_DB,
            REPO_ROOT / "form-app" / "data.js",
            REPO_ROOT / "form-app" / "index.html",
            *(REPO_ROOT / "form-output" / "runtime").glob("*-runtime-contract.json"),
        ]
        protected_before = {path: _sha256(path) for path in protected_paths}
        with mock.patch.object(self.mainmod.drafts, "save_operation") as draft_write:
            response = self.client.get(
                "/api/assets/reconciliation",
                params={
                    "model": "stingray",
                    "status": "safe_proposal",
                    "limit": 24,
                    "refresh": "true",
                },
            )
        self.assertEqual(response.status_code, 200, response.text)
        actual = response.json()
        self.assertEqual(actual["fingerprints"], expected["fingerprints"])
        self.assertEqual(actual["media"], expected["media"])
        self.assertEqual(actual["coverage"], expected["coverage"])
        self.assertEqual(actual["status_counts"], expected["status_counts"])
        self.assertEqual(actual["queue"], expected["queue"])
        self.assertEqual(actual["controls"]["image_fit"], ["cover", "contain", "swatch"])
        draft_write.assert_not_called()
        self.assertEqual(_sha256(self.workbook), workbook_before)
        self.assertEqual(
            self.client.get("/api/drafts", params={"limit": 200}).json(),
            drafts_before,
        )
        self.assertEqual(
            {path: _sha256(path) for path in protected_paths},
            protected_before,
        )

    def test_safe_asset_resolution_uses_existing_draft_and_binds_evidence(self):
        draft_id = "asset-safe-api"
        protected = [
            self.workbook,
            REPO_ROOT / "form-app" / "data.js",
            *(REPO_ROOT / "form-output" / "runtime").glob("*-runtime-contract.json"),
        ]
        before = {path: _sha256(path) for path in protected}
        view = self.client.get(
            "/api/assets/reconciliation",
            params={"status": "safe_proposal", "draft_id": draft_id},
        ).json()
        self.assertEqual(view["status_counts"]["safe_proposal"], 1)
        item = view["queue"]["items"][0]
        response = self.client.post(
            f"/api/drafts/{draft_id}/asset-resolutions",
            json={
                "item_id": item["id"],
                "resolution_kind": "accept_safe",
                "fingerprints": view["fingerprints"],
                "values": {"image_fit": "contain", "image_position": "25% 40%"},
                "actor": "asset-test",
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        operation = response.json()
        self.assertEqual(operation["table_name"], "assets")
        self.assertEqual(operation["changed_fields"]["image_fit"]["after"], "contain")
        lifecycle = self.client.get(f"/api/drafts/{draft_id}").json()
        self.assertEqual(len(lifecycle["operations"]), 1)
        evidence = lifecycle["artifacts"]["asset_resolutions"]
        self.assertEqual(len(evidence), 1)
        self.assertEqual(evidence[0]["resolution_kind"], "accept_safe")
        self.assertEqual(evidence[0]["reconciliation_sha256"], view["fingerprints"]["reconciliation_sha256"])
        self.assertEqual(lifecycle["operations"][0]["asset_resolutions"][0]["id"], evidence[0]["id"])
        self.assertEqual({path: _sha256(path) for path in protected}, before)
        self.assertEqual(self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200)

    def test_operational_ignore_invalidates_when_inventory_fingerprint_changes(self):
        draft_id = "asset-ignore-api"
        fixture = self.tmpdir / "ignore-media.txt"
        fixture.write_text("https://example.test/readme.txt\n", encoding="utf-8")
        original_fixture = os.environ["WBM_ASSET_MEDIA_URL_LIST"]
        try:
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = str(fixture)
            self.mainmod.asset_workspace.clear_cache()
            view = self.client.get(
                "/api/assets/reconciliation",
                params={"status": "unparseable", "draft_id": draft_id, "refresh": True},
            ).json()
            item = view["queue"]["items"][0]
            response = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": item["id"],
                    "resolution_kind": "ignore",
                    "fingerprints": view["fingerprints"],
                },
            )
            self.assertEqual(response.status_code, 200, response.text)
            ignored = self.client.get(
                "/api/assets/reconciliation",
                params={"status": "ignored", "draft_id": draft_id},
            ).json()
            self.assertEqual(ignored["status_counts"]["ignored"], 1)

            fixture.write_text(
                "https://example.test/readme.txt\nhttps://example.test/another.txt\n",
                encoding="utf-8",
            )
            refreshed = self.client.get(
                "/api/assets/reconciliation",
                params={"status": "unparseable", "draft_id": draft_id, "refresh": True},
            ).json()
            self.assertEqual(refreshed["status_counts"]["ignored"], 0)
            self.assertGreaterEqual(refreshed["status_counts"]["unparseable"], 1)
            self.assertEqual(refreshed["draft_asset_resolutions"]["stale_count"], 1)
        finally:
            self.client.post(f"/api/drafts/{draft_id}/cancel")
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = original_fixture
            self.mainmod.asset_workspace.clear_cache()

    def test_bulk_asset_acceptance_adds_only_server_classified_safe_matches(self):
        draft_id = "asset-safe-bulk-api"
        view = self.client.get(
            "/api/assets/reconciliation",
            params={"draft_id": draft_id, "refresh": True},
        ).json()
        response = self.client.post(
            f"/api/drafts/{draft_id}/asset-resolutions/safe",
            json={"fingerprints": view["fingerprints"], "actor": "bulk-test"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["accepted"], view["status_counts"]["safe_proposal"])
        lifecycle = self.client.get(f"/api/drafts/{draft_id}").json()
        evidence = lifecycle["artifacts"]["asset_resolutions"]
        self.assertTrue(evidence)
        self.assertEqual({item["resolution_kind"] for item in evidence}, {"accept_safe"})
        self.assertFalse({
            item["evidence"]["source_status"] for item in evidence
        } & {"ambiguous", "stale_target", "wildcard_conflict", "unmatched", "unparseable"})
        self.assertEqual(self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200)

    def test_bulk_asset_acceptance_enforces_the_requested_model_scope(self):
        draft_id = "asset-safe-bulk-scoped-api"
        fixture = self.tmpdir / "scoped-bulk-media.txt"
        fixture.write_text("https://example.test/h-stx.png\n", encoding="utf-8")
        original_fixture = os.environ["WBM_ASSET_MEDIA_URL_LIST"]
        try:
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = str(fixture)
            self.mainmod.asset_workspace.clear_cache()
            view = self.client.get(
                "/api/assets/reconciliation",
                params={"status": "safe_proposal", "draft_id": draft_id, "refresh": True},
            ).json()
            self.assertEqual(view["status_counts"]["safe_proposal"], 1)
            item = view["queue"]["items"][0]
            other_models = [
                model["model_key"]
                for model in self.client.get("/api/models").json()["models"]
                if model["model_key"] != item["model_key"]
            ]
            self.assertTrue(other_models)
            refused = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions/safe",
                json={
                    "fingerprints": view["fingerprints"],
                    "model": other_models[0],
                    "actor": "bulk-scope-test",
                },
            )
            self.assertEqual(refused.status_code, 200, refused.text)
            self.assertEqual(refused.json()["accepted"], 0)
            self.assertEqual(
                self.client.get(f"/api/drafts/{draft_id}").status_code, 404
            )
            staged = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions/safe",
                json={
                    "fingerprints": view["fingerprints"],
                    "model": item["model_key"],
                    "actor": "bulk-scope-test",
                },
            )
            self.assertEqual(staged.status_code, 200, staged.text)
            self.assertEqual(staged.json()["accepted"], 1)
            lifecycle = self.client.get(f"/api/drafts/{draft_id}").json()
            evidence = lifecycle["artifacts"]["asset_resolutions"]
            self.assertEqual(len(evidence), 1)
            self.assertEqual(
                evidence[0]["evidence"]["workbook_target"]["model_key"],
                item["model_key"],
            )
        finally:
            self.client.post(f"/api/drafts/{draft_id}/cancel")
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = original_fixture
            self.mainmod.asset_workspace.clear_cache()

    def test_asset_fingerprint_drift_blocks_changeset_freeze(self):
        draft_id = "asset-stale-freeze-api"
        fixture = self.tmpdir / "stale-freeze-media.txt"
        fixture.write_text("https://example.test/h-stx.png\n", encoding="utf-8")
        original_fixture = os.environ["WBM_ASSET_MEDIA_URL_LIST"]
        try:
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = str(fixture)
            self.mainmod.asset_workspace.clear_cache()
            view = self.client.get(
                "/api/assets/reconciliation",
                params={"status": "safe_proposal", "draft_id": draft_id, "refresh": True},
            ).json()
            item = view["queue"]["items"][0]
            saved = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": item["id"],
                    "resolution_kind": "accept_safe",
                    "fingerprints": view["fingerprints"],
                },
            )
            self.assertEqual(saved.status_code, 200, saved.text)
            fixture.write_text(
                "https://example.test/h-stx.png\nhttps://example.test/readme.txt\n",
                encoding="utf-8",
            )
            committed = self.client.post(f"/api/drafts/{draft_id}/commit")
            self.assertEqual(committed.status_code, 422, committed.text)
            self.assertEqual(
                committed.json()["detail"]["status"], "asset_reconciliation_stale"
            )
            lifecycle = self.client.get(f"/api/drafts/{draft_id}").json()
            self.assertEqual(lifecycle["draft"]["status"], "draft")
            self.assertIsNone(lifecycle["artifacts"]["changeset"])
        finally:
            self.client.post(f"/api/drafts/{draft_id}/cancel")
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = original_fixture
            self.mainmod.asset_workspace.clear_cache()

    def test_ambiguous_asset_requires_explicit_equal_priority_candidate(self):
        draft_id = "asset-ambiguous-api"
        fixture = self.tmpdir / "ambiguous-media.txt"
        fixture.write_text(
            "\n".join([
                "https://example.test/a/h-stx.png",
                "https://example.test/b/h-stx.png",
            ]) + "\n",
            encoding="utf-8",
        )
        original_fixture = os.environ["WBM_ASSET_MEDIA_URL_LIST"]
        try:
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = str(fixture)
            self.mainmod.asset_workspace.clear_cache()
            view = self.client.get(
                "/api/assets/reconciliation",
                params={"model": "z06", "status": "ambiguous", "draft_id": draft_id, "refresh": True},
            ).json()
            item = view["queue"]["items"][0]
            alternatives = item["candidate"]["alternatives"]
            self.assertEqual(len(alternatives), 2)
            rejected = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": item["id"],
                    "resolution_kind": "select_candidate",
                    "fingerprints": view["fingerprints"],
                    "selected_url": "https://example.test/not-a-candidate.png",
                },
            )
            self.assertEqual(rejected.status_code, 422)
            accepted = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": item["id"],
                    "resolution_kind": "select_candidate",
                    "fingerprints": view["fingerprints"],
                    "selected_url": alternatives[1]["url"],
                    "values": {"image_fit": "contain", "image_position": "right center"},
                },
            )
            self.assertEqual(accepted.status_code, 200, accepted.text)
            lifecycle = self.client.get(f"/api/drafts/{draft_id}").json()
            evidence = lifecycle["artifacts"]["asset_resolutions"][0]
            self.assertEqual(evidence["resolution_kind"], "select_candidate")
            self.assertEqual(evidence["media_url"], alternatives[1]["url"])
        finally:
            self.client.post(f"/api/drafts/{draft_id}/cancel")
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = original_fixture
            self.mainmod.asset_workspace.clear_cache()

    def test_unmatched_media_can_only_assign_to_a_valid_snapshot_target(self):
        draft_id = "asset-assignment-api"
        fixture = self.tmpdir / "assignment-media.txt"
        media_url = "https://example.test/c-xyz.png"
        fixture.write_text(media_url + "\n", encoding="utf-8")
        original_fixture = os.environ["WBM_ASSET_MEDIA_URL_LIST"]
        try:
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = str(fixture)
            self.mainmod.asset_workspace.clear_cache()
            view = self.client.get(
                "/api/assets/reconciliation",
                params={"status": "unmatched", "draft_id": draft_id, "refresh": True},
            ).json()
            source = view["queue"]["items"][0]
            target = next(
                item for item in view["assignment_targets"]
                if item["model_key"] == "stingray" and item["target_type"] == "option"
            )
            rejected = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": source["id"],
                    "resolution_kind": "assign_media",
                    "fingerprints": view["fingerprints"],
                    "target_item_id": "not-a-target",
                },
            )
            self.assertEqual(rejected.status_code, 422)
            accepted = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": source["id"],
                    "resolution_kind": "assign_media",
                    "fingerprints": view["fingerprints"],
                    "target_item_id": target["item_id"],
                    "values": {"image_fit": "cover", "image_position": "center"},
                },
            )
            self.assertEqual(accepted.status_code, 200, accepted.text)
            operation = accepted.json()
            self.assertEqual(operation["table_name"], "assets")
            self.assertEqual(operation["final"]["image_url"], media_url)
            evidence = self.client.get(f"/api/drafts/{draft_id}").json()["artifacts"]["asset_resolutions"][0]
            self.assertEqual(evidence["resolution_kind"], "assign_media")
            self.assertEqual(evidence["evidence"]["target_item_id"], target["item_id"])
            other_target = next(
                item for item in view["assignment_targets"]
                if item["model_key"] == "z06"
                and (
                    item["target_type"], item["target_id"]
                ) != (target["target_type"], target["target_id"])
            )
            retargeted = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": source["id"],
                    "resolution_kind": "assign_media",
                    "fingerprints": view["fingerprints"],
                    "target_item_id": other_target["item_id"],
                },
            )
            self.assertEqual(retargeted.status_code, 422)
            self.assertEqual(
                retargeted.json()["detail"]["status"],
                "asset_resolution_retarget_rejected",
            )
        finally:
            self.client.post(f"/api/drafts/{draft_id}/cancel")
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = original_fixture
            self.mainmod.asset_workspace.clear_cache()

    def test_mixed_ordinary_and_asset_draft_reaches_one_approval_chain(self):
        draft_id = "asset-mixed-approval-api"
        fixture = self.tmpdir / "mixed-media.txt"
        fixture.write_text(
            "\n".join([
                "https://example.test/h-stx.png",
                "https://example.test/a/h-j6a.png",
                "https://example.test/b/h-j6a.png",
                "https://example.test/c-gba.png",
            ]) + "\n",
            encoding="utf-8",
        )
        original_fixture = os.environ["WBM_ASSET_MEDIA_URL_LIST"]
        protected = [
            self.workbook,
            REPO_ROOT / "form-app" / "data.js",
            REPO_ROOT / "form-app" / "index.html",
            *(REPO_ROOT / "form-output" / "runtime").glob("*-runtime-contract.json"),
        ]
        before = {path: _sha256(path) for path in protected}
        try:
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = str(fixture)
            self.mainmod.asset_workspace.clear_cache()
            ordinary = self.client.get(
                "/api/records/options", params={"model": "stingray", "limit": 1}
            ).json()["records"][0]
            saved_ordinary = self.client.post(
                f"/api/drafts/{draft_id}/operations",
                json={
                    "table": "options",
                    "model_id": "stingray",
                    "op": "update",
                    "key": {"option_id": ordinary["option_id"]},
                    "record": {"option_name": ordinary["option_name"] + " mixed proof"},
                },
            )
            self.assertEqual(saved_ordinary.status_code, 200, saved_ordinary.text)

            view = self.client.get(
                "/api/assets/reconciliation",
                params={"draft_id": draft_id, "refresh": True},
            ).json()
            safe = next(item for item in view["queue"]["items"] if item["status"] == "safe_proposal")
            saved_safe = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": safe["id"],
                    "resolution_kind": "accept_safe",
                    "fingerprints": view["fingerprints"],
                },
            )
            self.assertEqual(saved_safe.status_code, 200, saved_safe.text)

            ambiguous_view = self.client.get(
                "/api/assets/reconciliation",
                params={"model": "z06", "status": "ambiguous", "draft_id": draft_id},
            ).json()
            ambiguous = ambiguous_view["queue"]["items"][0]
            selected = ambiguous["candidate"]["alternatives"][0]["url"]
            saved_ambiguous = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": ambiguous["id"],
                    "resolution_kind": "select_candidate",
                    "fingerprints": ambiguous_view["fingerprints"],
                    "selected_url": selected,
                    "values": {"image_fit": "contain", "image_position": "right center"},
                },
            )
            self.assertEqual(saved_ambiguous.status_code, 200, saved_ambiguous.text)

            presentation_view = self.client.get(
                "/api/assets/reconciliation",
                params={"model": "stingray", "status": "missing", "draft_id": draft_id, "limit": 100},
            ).json()
            presentation = next(
                item for item in presentation_view["queue"]["items"]
                if item["coverage"]["kind"] == "exact"
                and item["current_values"]["image_url"]
            )
            current_fit = presentation["current_values"]["image_fit"]
            edited_fit = "contain" if current_fit != "contain" else "cover"
            saved_presentation = self.client.post(
                f"/api/drafts/{draft_id}/asset-resolutions",
                json={
                    "item_id": presentation["id"],
                    "resolution_kind": "manual_url",
                    "fingerprints": presentation_view["fingerprints"],
                    "selected_url": presentation["current_values"]["image_url"],
                    "values": {"image_fit": edited_fit, "image_position": "left center"},
                },
            )
            self.assertEqual(saved_presentation.status_code, 200, saved_presentation.text)

            lifecycle = self.client.get(f"/api/drafts/{draft_id}").json()
            self.assertEqual(len(lifecycle["artifacts"]["asset_resolutions"]), 3)
            self.assertTrue(any(op["table_name"] == "options" for op in lifecycle["operations"]))
            self.assertTrue(any(op["table_name"] == "assets" for op in lifecycle["operations"]))

            committed = self.client.post(f"/api/drafts/{draft_id}/commit")
            self.assertEqual(committed.status_code, 200, committed.text)
            changeset = committed.json()
            preview_artifact = {
                "ok": True,
                "schemaVersion": "workbook-change-preview-1",
                "changeSetId": changeset["changeSetId"],
                "semanticFingerprint": changeset["semanticFingerprint"],
                "workbook": changeset["workbook"],
                "status": "validated",
                "errors": [],
                "warnings": [],
                "warningPolicy": {"blockingIds": [], "confirmableIds": []},
                "previewFingerprint": "a" * 64,
            }
            approval_artifact = {
                "ok": True,
                "schemaVersion": "workbook-change-approval-1",
                "actor": "mixed-proof",
                "changeSetId": changeset["changeSetId"],
                "semanticFingerprint": changeset["semanticFingerprint"],
                "previewFingerprint": preview_artifact["previewFingerprint"],
                "workbook": changeset["workbook"],
                "acceptedWarningIds": [],
                "approvalFingerprint": "b" * 64,
            }
            with (
                mock.patch.object(
                    self.mainmod.drafts.workbook_service,
                    "preview_changeset",
                    return_value=preview_artifact,
                ),
                mock.patch.object(
                    self.mainmod.drafts.workbook_service,
                    "approve_changeset",
                    return_value=approval_artifact,
                ),
            ):
                preview = self.client.post(f"/api/drafts/{draft_id}/preview")
                self.assertEqual(preview.status_code, 200, preview.text)
                self.assertEqual(preview.json()["manager_state"], "preview_ready")
                approved = self.client.post(
                    f"/api/drafts/{draft_id}/approve",
                    json={"actor": "mixed-proof", "warning_ids": []},
                )
            self.assertEqual(approved.status_code, 200, approved.text)
            self.assertEqual(approved.json()["manager_state"], "approved")
            final = self.client.get(f"/api/drafts/{draft_id}").json()
            self.assertIsNotNone(final["artifacts"]["changeset"])
            self.assertEqual(len(final["artifacts"]["preview_attempts"]), 1)
            self.assertEqual(len(final["artifacts"]["approval_attempts"]), 1)
            self.assertEqual({path: _sha256(path) for path in protected}, before)
        finally:
            self.client.post(f"/api/drafts/{draft_id}/cancel")
            os.environ["WBM_ASSET_MEDIA_URL_LIST"] = original_fixture
            self.mainmod.asset_workspace.clear_cache()

    def test_status_reports_guarded_surfaces_separately(self):
        status = self.client.get("/api/status").json()
        self.assertEqual(status["mode"], "durable_apply_rebuild")
        self.assertEqual(status["projection"]["state"], "current")
        self.assertTrue(status["projection"]["active"])
        self.assertTrue(status["projection"]["reimport_allowed"])
        self.assertEqual(status["draft"]["state"], "clear")
        self.assertIn("state", status["workbook"])
        self.assertEqual(status["generated_artifacts"]["state"], "unverified")
        self.assertEqual(status["publication"]["state"], "unverified")

    def test_same_mtime_hash_drift_is_stale_and_allows_verified_reimport(self):
        import os

        original = self.workbook.read_bytes()
        mtime_ns = self.workbook.stat().st_mtime_ns
        try:
            mutated = bytearray(original)
            mutated[-1] ^= 1
            self.workbook.write_bytes(mutated)
            os.utime(self.workbook, ns=(mtime_ns, mtime_ns))
            status = self.client.get("/api/status").json()
            self.assertEqual(status["workbook"]["state"], "stale")
            self.assertEqual(status["projection"]["state"], "stale")
            self.assertTrue(status["projection"]["reimport_allowed"])
        finally:
            self.workbook.write_bytes(original)
            os.utime(self.workbook, ns=(mtime_ns, mtime_ns))

    def test_structure_and_collections(self):
        structure = self.client.get("/api/structure/stingray").json()
        self.assertTrue(structure["steps"])
        cols = self.client.get(
            "/api/models/stingray/collections").json()["collections"]
        tables = {c["table"] for c in cols}
        self.assertIn("options", tables)
        self.assertIn("pricing", tables)

    def test_structure_falls_back_to_master_section_step(self):
        structure = self.client.get("/api/structure/stingray").json()
        presentation = {
            section["section_id"]: section
            for section in structure["section_presentation"]
        }
        stripes = next(
            step for step in structure["steps"]
            if step["step_key"] == "aero_exhaust_stripes_accessories"
        )

        self.assertEqual(
            presentation["sec_1lte_001"]["step_key"],
            "standard_equipment",
        )
        self.assertIn(
            "sec_stri_001",
            {section["section_id"] for section in stripes["sections"]},
        )

    def test_catalog_routes_refuse_non_allowlisted_table_names(self):
        attempts = (
            self.client.get("/api/records/sqlite_master/schema"),
            self.client.get("/api/records/sqlite_master"),
            self.client.post(
                "/api/records/sqlite_master/dependencies",
                json={"model_id": "", "key": {}},
            ),
        )

        for response in attempts:
            self.assertEqual(response.status_code, 404, response.text)
            self.assertIn("unknown table", response.json()["detail"])

    def test_stage_endpoint_validation_error_shape(self):
        resp = self.client.post("/api/changes", json={
            "table": "options", "model_id": "stingray", "op": "add",
            "key": {"option_id": "opt_x"},
            "record": {"option_id": "opt_x", "price": "not-a-number"}})
        self.assertEqual(resp.status_code, 422)
        detail = resp.json()["detail"]
        self.assertTrue(any(e["field"] == "price"
                            for e in detail["errors"]))

    def test_records_search(self):
        resp = self.client.get(
            "/api/records/options", params={"model": "stingray",
                                            "search": "Z51"}).json()
        self.assertGreater(resp["total"], 0)

    def test_connected_option_detail_is_model_scoped_complete_and_read_only(self):
        protected = [
            self.workbook,
            self.mainmod.config.DEFAULT_PROJECTION_DB,
            self.mainmod.config.DEFAULT_DB,
            self.apply_output_root / "form-app" / "data.js",
        ]
        before = {path: _sha256(path) for path in protected}

        response = self.client.get(
            "/api/explorer/stingray/options/opt_5zu_001"
        )

        self.assertEqual(response.status_code, 200, response.text)
        detail = response.json()
        self.assertEqual(detail["model_key"], "stingray")
        self.assertEqual(detail["entity_type"], "option")
        self.assertEqual(detail["option"]["option_id"], "opt_5zu_001")
        self.assertEqual(detail["option"]["rpo"], "5ZU")
        self.assertEqual(detail["option"]["name"], "Body-Color High Wing Spoiler")
        self.assertEqual(detail["section"]["section_id"], "sec_spoi_001")
        self.assertEqual(len(detail["availability"]), 6)
        self.assertTrue(detail["exclusive_groups"])
        self.assertTrue(detail["rule_groups"])
        self.assertTrue(detail["rules"])
        self.assertTrue(detail["assets"])
        self.assertIn("technical", detail)
        self.assertEqual(detail["destination"], {
            "workspace": "options",
            "entity_type": "option",
            "entity_id": "opt_5zu_001",
        })
        self.assertEqual({path: _sha256(path) for path in protected}, before)

        missing = self.client.get(
            "/api/explorer/grand_sport/options/opt_5zu_001"
        )
        self.assertEqual(missing.status_code, 404)

    def test_connected_option_detail_overlays_coalesced_durable_draft_intent(self):
        base = self.client.get(
            "/api/explorer/stingray/options/opt_5zu_001"
        ).json()
        original_name = base["option"]["option_name"]
        draft_id = "connected-option-overlay"
        proposed_name = f"{original_name} draft overlay"
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "options",
                "model_id": "stingray",
                "op": "update",
                "key": {"option_id": "opt_5zu_001"},
                "record": {"option_name": proposed_name},
                "actor": "connected-overlay-test",
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)

        response = self.client.get(
            "/api/explorer/stingray/options/opt_5zu_001",
            params={"draft_id": draft_id},
        )

        self.assertEqual(response.status_code, 200, response.text)
        detail = response.json()
        self.assertEqual(detail["option"]["option_name"], original_name)
        self.assertEqual(detail["draft_overlay"]["draft_id"], draft_id)
        self.assertEqual(detail["draft_overlay"]["state"], "modified")
        self.assertEqual(
            detail["draft_overlay"]["base"]["option_name"], original_name
        )
        self.assertEqual(
            detail["draft_overlay"]["proposed"]["option_name"], proposed_name
        )
        self.assertEqual(
            detail["draft_overlay"]["effective"]["option_name"], proposed_name
        )
        self.assertEqual(detail["draft_overlay"]["conflicts"], [])
        self.assertEqual(
            self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200
        )

    def test_connected_detail_accepts_the_reserved_empty_browser_draft_id(self):
        response = self.client.get(
            "/api/explorer/stingray/options/opt_5zu_001",
            params={"draft_id": "reserved-before-first-save"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["draft_overlay"], {
            "draft_id": "reserved-before-first-save",
            "draft_revision": 0,
            "state": "unchanged",
            "base": None,
            "proposed": None,
            "effective": None,
            "conflicts": [],
        })

    def test_connected_overlay_fails_closed_when_draft_binding_is_stale(self):
        draft_id = "connected-option-stale-overlay"
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "options",
                "model_id": "stingray",
                "op": "update",
                "key": {"option_id": "opt_5zu_001"},
                "record": {"option_name": "Stale draft name"},
                "actor": "connected-overlay-test",
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)
        state = self.mainmod.open_state_connection()
        try:
            state.execute(
                "UPDATE workflow_drafts SET base_workbook_sha256=? WHERE id=?",
                ("stale-workbook-sha", draft_id),
            )
            state.commit()
        finally:
            state.close()

        detail = self.client.get(
            "/api/explorer/stingray/options/opt_5zu_001",
            params={"draft_id": draft_id},
        ).json()

        self.assertEqual(detail["draft_overlay"]["state"], "conflicted")
        self.assertEqual(detail["draft_overlay"]["effective"], None)
        self.assertEqual(
            detail["draft_overlay"]["conflicts"][0]["code"],
            "draft_binding_stale",
        )
        self.assertEqual(
            self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200
        )

    def test_connected_option_deep_link_resolves_a_draft_added_record_only(self):
        draft_id = "connected-option-added-overlay"
        option_id = "opt_connected_add_001"
        record = {
            "option_id": option_id,
            "rpo": "C3F",
            "price": "100",
            "option_name": "Checkpoint 3F proposed option",
            "description": "",
            "detail_raw": "",
            "section_id": "sec_pain_001",
            "selectable": "True",
            "display_order": "999",
            "active": "True",
            "display_behavior": "",
        }
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "options",
                "model_id": "stingray",
                "op": "add",
                "key": {"option_id": option_id},
                "record": record,
                "actor": "connected-overlay-test",
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)
        self.assertEqual(
            self.client.get(
                f"/api/explorer/stingray/options/{option_id}"
            ).status_code,
            404,
        )

        response = self.client.get(
            f"/api/explorer/stingray/options/{option_id}",
            params={"draft_id": draft_id},
        )

        self.assertEqual(response.status_code, 200, response.text)
        detail = response.json()
        self.assertEqual(detail["option"]["option_id"], option_id)
        self.assertEqual(detail["draft_overlay"]["state"], "added")
        self.assertEqual(detail["draft_overlay"]["base"], None)
        self.assertEqual(
            detail["draft_overlay"]["effective"]["option_name"],
            record["option_name"],
        )
        self.assertEqual(
            self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200
        )

    def test_connected_option_keeps_pending_deletion_inspectable(self):
        draft_id = "connected-option-delete-overlay"
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "options",
                "model_id": "stingray",
                "op": "update",
                "key": {"option_id": "opt_5zu_001"},
                "record": {"option_name": "Temporary update"},
                "actor": "connected-overlay-test",
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)
        state = self.mainmod.open_state_connection()
        try:
            state.execute(
                "UPDATE draft_operations SET action='delete', final_json=NULL "
                "WHERE draft_id=?",
                (draft_id,),
            )
            state.commit()
        finally:
            state.close()

        response = self.client.get(
            "/api/explorer/stingray/options/opt_5zu_001",
            params={"draft_id": draft_id},
        )

        self.assertEqual(response.status_code, 200, response.text)
        overlay = response.json()["draft_overlay"]
        self.assertEqual(overlay["state"], "pending_deletion")
        self.assertIsNotNone(overlay["base"])
        self.assertIsNone(overlay["proposed"])
        self.assertIsNone(overlay["effective"])
        self.assertEqual(
            self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200
        )

    def test_connected_group_detail_leads_with_description_and_named_members(self):
        group_id = "grand_sport_x_excl_1623e1da9d59"
        response = self.client.get(
            f"/api/explorer/grand_sport_x/groups/exclusive/{group_id}"
        )

        self.assertEqual(response.status_code, 200, response.text)
        detail = response.json()
        self.assertEqual(detail["model_key"], "grand_sport_x")
        self.assertEqual(detail["entity_type"], "group")
        self.assertEqual(detail["group_type"], "exclusive")
        self.assertEqual(detail["label"], "LS6 Engine Covers")
        self.assertEqual(detail["display_label"], "LS6 Engine Covers")
        self.assertEqual(detail["audience"], "customer")
        self.assertEqual(detail["label_status"], "authored")
        self.assertNotIn("1623e1da9d59", detail["label"])
        self.assertIn("Engine cover choices", detail["notes"])
        self.assertGreater(detail["member_count"], 1)
        self.assertTrue(all(member["rpo"] and member["option_name"]
                            for member in detail["members"]))
        self.assertEqual(detail["group"]["group_id"], group_id)
        self.assertEqual(detail["group"]["display_label"], "LS6 Engine Covers")
        self.assertTrue(all(member["physical_key"] and member["src_sheet"]
                            for member in detail["members"]))
        self.assertEqual(detail["editor"], {
            "group_table": "exclusive_groups",
            "group_id_field": "group_id",
            "member_table": "exclusive_group_members",
            "member_id_field": "option_id",
            "member_group_field": "group_id",
            "member_order_field": "display_order",
            "member_active_field": "active",
        })
        self.assertEqual(detail["technical"]["lineage"]["source_sheet"],
                         "grand_sport_x_exclusive_groups")
        self.assertEqual(detail["destination"]["entity_id"],
                         f"exclusive:{group_id}")

        rule_group_id = "grand_sport_x_group_dpb_excludes_any_096f31ba6bc8"
        rule_response = self.client.get(
            f"/api/explorer/grand_sport_x/groups/rule/{rule_group_id}"
        )
        self.assertEqual(rule_response.status_code, 200, rule_response.text)
        rule_detail = rule_response.json()
        self.assertEqual(
            rule_detail["label"],
            "DPB Carbon Flash/Blue Racing Stripes — Hash Mark and Z15 Exclusions",
        )
        self.assertEqual(rule_detail["audience"], "manager")
        self.assertEqual(rule_detail["label_status"], "authored")
        self.assertEqual(rule_detail["group"]["group_id"], rule_group_id)
        self.assertEqual(rule_detail["editor"], {
            "group_table": "rule_groups",
            "group_id_field": "group_id",
            "member_table": "rule_group_members",
            "member_id_field": "target_id",
            "member_group_field": "group_id",
            "member_order_field": "display_order",
            "member_active_field": "active",
        })

    def test_connected_group_detail_uses_the_same_durable_overlay_contract(self):
        group_id = "grand_sport_x_excl_1623e1da9d59"
        base = self.client.get(
            f"/api/explorer/grand_sport_x/groups/exclusive/{group_id}"
        ).json()
        draft_id = "connected-group-overlay"
        proposed_notes = f'{base["group"]["notes"]} draft overlay'
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "exclusive_groups",
                "model_id": "grand_sport_x",
                "op": "update",
                "key": {"group_id": group_id},
                "record": {"notes": proposed_notes},
                "actor": "connected-overlay-test",
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)

        detail = self.client.get(
            f"/api/explorer/grand_sport_x/groups/exclusive/{group_id}",
            params={"draft_id": draft_id},
        ).json()

        self.assertEqual(detail["draft_overlay"]["state"], "modified")
        self.assertEqual(
            detail["draft_overlay"]["effective"]["notes"], proposed_notes
        )
        self.assertEqual(detail["group"]["notes"], base["group"]["notes"])
        self.assertEqual(
            self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200
        )

    def test_cross_entity_search_is_ranked_typed_scoped_and_stable(self):
        option = self.client.get(
            "/api/explorer/stingray/search", params={"query": "5ZU"}
        ).json()
        self.assertEqual(option["model_key"], "stingray")
        self.assertEqual(option["results"][0]["entity_type"], "option")
        self.assertEqual(option["results"][0]["entity_id"], "opt_5zu_001")
        self.assertEqual(option["results"][0]["destination"], {
            "workspace": "options", "entity_type": "option",
            "entity_id": "opt_5zu_001",
        })
        other_model_ids = {
            row["entity_id"] for row in self.client.get(
                "/api/explorer/grand_sport/search", params={"query": "5ZU"}
            ).json()["results"] if row["entity_type"] == "option"
        }
        self.assertNotIn("opt_5zu_001", other_model_ids)

        for query, entity_type in (
            ("Engine cover choices", "group"),
            ("Engine Appearance", "section"),
            ("excludes", "rule"),
        ):
            result = self.client.get(
                "/api/explorer/grand_sport_x/search", params={"query": query}
            )
            self.assertEqual(result.status_code, 200, result.text)
            self.assertIn(entity_type, {
                row["entity_type"] for row in result.json()["results"]
            })

        section = self.client.get(
            "/api/explorer/grand_sport_x/sections/sec_engi_001"
        )
        self.assertEqual(section.status_code, 200, section.text)
        self.assertEqual(section.json()["destination"], {
            "workspace": "sections", "entity_type": "section",
            "entity_id": "sec_engi_001",
        })
        self.assertTrue(section.json()["options"])

        rule_search = self.client.get(
            "/api/explorer/grand_sport_x/search", params={"query": "excludes"}
        ).json()["results"]
        rule_result = next(row for row in rule_search if row["entity_type"] == "rule")
        rule = self.client.get(
            f"/api/explorer/grand_sport_x/rules/{rule_result['entity_id']}"
        )
        self.assertEqual(rule.status_code, 200, rule.text)
        self.assertEqual(rule.json()["destination"], rule_result["destination"])
        self.assertIn("source_option", rule.json())
        self.assertIn("target_option", rule.json())

    def test_named_diagnostics_are_bounded_defined_scoped_and_traceable(self):
        catalog = self.client.get(
            "/api/explorer/stingray/diagnostics"
        ).json()
        self.assertEqual([item["key"] for item in catalog["diagnostics"]], [
            "missing_required_images",
            "multiple_exclusive_groups",
            "where_used",
            "option_relationships",
            "variant_availability_differences",
        ])
        for item in catalog["diagnostics"]:
            self.assertTrue(item["label"])
            self.assertTrue(item["definition"])

        missing = self.client.get(
            "/api/explorer/stingray/diagnostics/missing_required_images",
            params={"limit": 5},
        ).json()
        self.assertEqual(missing["model_key"], "stingray")
        self.assertLessEqual(len(missing["results"]), 5)
        self.assertTrue(all(row["destination"]["workspace"] == "options"
                            for row in missing["results"]))

        multiple = self.client.get(
            "/api/explorer/stingray/diagnostics/multiple_exclusive_groups"
        ).json()
        self.assertEqual(multiple["model_key"], "stingray")

        used = self.client.get(
            "/api/explorer/stingray/diagnostics/where_used",
            params={"entity_id": "opt_5zu_001"},
        ).json()
        self.assertTrue(used["results"])
        self.assertTrue(all(row["technical"]["source_sheet"]
                            for row in used["results"]))

        group_used = self.client.get(
            "/api/explorer/grand_sport_x/diagnostics/where_used",
            params={"entity_id": "exclusive:grand_sport_x_excl_1623e1da9d59"},
        ).json()
        self.assertTrue(group_used["results"])
        self.assertTrue(all(row["technical"]["source_sheet"]
                            for row in group_used["results"]))

        relationships = self.client.get(
            "/api/explorer/stingray/diagnostics/option_relationships",
            params={"entity_id": "opt_5zu_001"},
        ).json()
        self.assertTrue(relationships["results"])

        differences = self.client.get(
            "/api/explorer/stingray/diagnostics/variant_availability_differences",
            params={"limit": 5},
        ).json()
        self.assertLessEqual(len(differences["results"]), 5)
        self.assertTrue(all(row["distinct_status_count"] > 1
                            for row in differences["results"]))

        missing_entity = self.client.get(
            "/api/explorer/stingray/diagnostics/where_used"
        )
        self.assertEqual(missing_entity.status_code, 422)

    def test_schema_exposes_final_shared_field_and_model_context_metadata(self):
        defaults = self.client.get(
            "/api/records/default_selection_rules/schema",
            params={"model": "stingray"},
        ).json()
        self.assertTrue(defaults["model_context"]["required"])
        by_name = {column["name"]: column for column in defaults["columns"]}
        self.assertEqual(by_name["condition_type"]["field_kind"], "finite")
        condition_ref = by_name["condition_id"]["reference"]
        self.assertEqual(condition_ref["kind"], "conditional")
        self.assertTrue(
            any(target["derived"] for target in condition_ref["targets"])
        )
        self.assertTrue(by_name["condition_id"]["optional"])
        self.assertEqual(by_name["notes"]["field_kind"], "free_text")

        mappings = self.client.get(
            "/api/records/rule_mappings/schema", params={"model": "stingray"}
        ).json()
        mapping_fields = {column["name"]: column for column in mappings["columns"]}
        self.assertEqual(mapping_fields["source_id"]["reference"]["kind"], "union")

        assets = self.client.get(
            "/api/records/assets/schema", params={"model": "stingray"}
        ).json()
        asset_fields = {column["name"]: column for column in assets["columns"]}
        self.assertEqual(asset_fields["image_fit"]["field_kind"], "finite")
        self.assertEqual(
            asset_fields["image_fit"]["finite_values"],
            ["cover", "contain", "swatch"],
        )

    def test_durable_draft_discovery_and_cancel_routes_preserve_history(self):
        row = self.client.get(
            "/api/records/options",
            params={"model": "stingray", "limit": 1},
        ).json()["records"][0]
        draft_id = "api-ui-draft-cancel"
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "options",
                "model_id": "stingray",
                "op": "update",
                "key": {"option_id": row["option_id"]},
                "record": {"option_name": f'{row["option_name"]} UI review'},
                "actor": "browser-test",
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)

        listed = self.client.get("/api/drafts", params={"limit": 200}).json()
        summary = next(item for item in listed["drafts"] if item["id"] == draft_id)
        self.assertEqual(summary["status"], "draft")
        self.assertEqual(summary["operation_count"], 1)

        cancelled = self.client.post(f"/api/drafts/{draft_id}/cancel")
        self.assertEqual(cancelled.status_code, 200, cancelled.text)
        self.assertEqual(cancelled.json()["status"], "cancelled")
        lifecycle = self.client.get(f"/api/drafts/{draft_id}").json()
        self.assertEqual(lifecycle["draft"]["status"], "cancelled")
        self.assertEqual(len(lifecycle["operations"]), 1)
        self.assertEqual(
            lifecycle["artifacts"]["cancellation"]["status"], "cancelled"
        )

    def test_vehicle_setup_copy_round_trips_through_browser_schema_and_draft(self):
        setup_fields = [
            "setup_card_subtitle", "setup_eyebrow", "setup_title",
            "setup_description", "setup_fact_1", "setup_fact_2", "setup_fact_3",
        ]
        schema = self.client.get(
            "/api/records/models/schema", params={"model": "stingray"}
        ).json()
        columns = {column["name"]: column for column in schema["columns"]}
        self.assertTrue(set(setup_fields) <= columns.keys())
        self.assertTrue(all(columns[name]["field_kind"] == "free_text"
                            for name in setup_fields))

        records = self.client.get(
            "/api/records/models",
            params={"model": "stingray", "search": "stingray"},
        ).json()["records"]
        model = next(row for row in records if row["model_key"] == "stingray")
        final_copy = {name: model[name] for name in setup_fields}
        final_copy["setup_title"] = f'{final_copy["setup_title"]} UI proof'
        draft_id = "api-setup-copy-ui"
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "models",
                "model_id": "stingray",
                "op": "update",
                "key": {"model_key": "stingray"},
                "record": final_copy,
                "actor": "browser-test",
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)
        operation = saved.json()
        self.assertEqual(
            {name: operation["final"][name] for name in setup_fields}, final_copy
        )

        committed = self.client.post(f"/api/drafts/{draft_id}/commit")
        self.assertEqual(committed.status_code, 200, committed.text)
        change = committed.json()["rowChanges"][0]
        self.assertEqual(set(change["fields"]), {"setup_title"})
        self.assertEqual(
            change["fields"]["setup_title"]["after"], final_copy["setup_title"]
        )
        self.assertEqual(
            self.client.post(f"/api/drafts/{draft_id}/cancel").status_code, 200
        )

    def test_real_model_row_round_trips_context_lineage_null_and_reference(self):
        conn = self.mainmod.open_projection_connection()
        try:
            row = conn.execute(
                "SELECT * FROM rule_mappings WHERE model_id='stingray' "
                "AND runtime_action IS NULL AND source_id IS NOT NULL "
                "AND original_detail_raw IS NOT NULL LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(row)
            source = dict(row)
        finally:
            conn.close()

        schema = self.client.get(
            "/api/records/rule_mappings/schema", params={"model": "stingray"}
        ).json()
        fields = {column["name"]: column for column in schema["columns"]}
        self.assertEqual(schema["model_context"]["value"], "stingray")
        self.assertEqual(fields["source_id"]["reference"]["kind"], "union")
        self.assertTrue(fields["runtime_action"]["optional"])

        projected = self.client.get(
            "/api/records/rule_mappings",
            params={"model": "stingray", "search": source["rule_id"]},
        ).json()["records"]
        record = next(item for item in projected if item["id"] == source["id"])
        self.assertEqual(record["model_context"], ["stingray"])
        self.assertEqual(record["src_sheet"], source["src_sheet"])
        self.assertEqual(record["src_row"], source["src_row"])
        self.assertEqual(record["physical_key"], source["physical_key"])
        self.assertIsNone(record["runtime_action"])
        self.assertEqual(record["source_id"], source["source_id"])

        draft_id = "api-context-round-trip"
        response = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "rule_mappings",
                "model_id": "stingray",
                "op": "update",
                "key": {"rule_id": source["rule_id"]},
                "record": {
                    "original_detail_raw": source["original_detail_raw"] + " context proof"
                },
                "session_id": "api-context-test",
                "actor": "test",
            },
        )
        try:
            self.assertEqual(response.status_code, 200, response.text)
            operation = response.json()
            self.assertEqual(operation["model_context"], ["stingray"])
            self.assertEqual(operation["source_sheet"], source["src_sheet"])
            self.assertEqual(operation["source_row"], source["src_row"])
            self.assertEqual(operation["physical_key"], source["physical_key"])
            self.assertIsNone(operation["original"]["runtime_action"])
            self.assertEqual(operation["original"]["source_id"], source["source_id"])
            self.assertIsNone(operation["final"]["runtime_action"])
            self.assertEqual(operation["final"]["source_id"], source["source_id"])

            lifecycle = self.client.get(f"/api/drafts/{draft_id}")
            self.assertEqual(lifecycle.status_code, 200, lifecycle.text)
            body = lifecycle.json()
            self.assertEqual(body["context"]["model_keys"], ["stingray"])
            self.assertEqual(body["operations"], [operation])
            self.assertEqual(
                body["context"]["physical_targets"][0]["physical_key"],
                source["physical_key"],
            )

            emitted = self.client.post(f"/api/drafts/{draft_id}/commit")
            self.assertEqual(emitted.status_code, 200, emitted.text)
            committed = self.client.get(f"/api/drafts/{draft_id}").json()
            self.assertEqual(
                committed["artifacts"]["changeset"]["artifact"], emitted.json()
            )
            self.assertEqual(
                committed["artifacts"]["changeset"]["change_set_id"],
                emitted.json()["changeSetId"],
            )
        finally:
            state = self.mainmod.open_state_connection()
            try:
                status = state.execute(
                    "SELECT status FROM workflow_drafts WHERE id=?", (draft_id,)
                ).fetchone()
                if status is not None and status["status"] != "cancelled":
                    self.mainmod.drafts.cancel_draft(state, draft_id=draft_id)
            finally:
                state.close()

    def test_live_sync_is_provisionally_read_only_even_when_fully_confirmed(self):
        current_mtime = str(WORKBOOK.stat().st_mtime_ns)
        resp = self.client.post("/api/sync", json={
            "write": True,
            "confirm": "SYNC",
            "expected_mtime_ns": current_mtime,
        })
        self.assertEqual(resp.status_code, 409)
        detail = resp.json()["detail"]
        self.assertEqual(detail["status"], "read_only_provisional")
        self.assertIn("Apply and Rebuild", detail["message"])

    def test_pass5_draft_api_roots_durable_update_without_projection_mutation(self):
        projected = self.client.get(
            "/api/records/options", params={"model": "stingray", "limit": 1}
        ).json()["records"][0]
        option_id = projected["option_id"]
        original_name = projected["option_name"]
        draft_id = "api-draft-test"
        response = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "options",
                "model_id": "stingray",
                "op": "update",
                "key": {"option_id": option_id},
                "record": {"option_name": f"{original_name} changed"},
                "session_id": "api-test",
                "actor": "test",
            },
        )
        try:
            self.assertEqual(response.status_code, 200, response.text)
            operation = response.json()
            self.assertEqual(operation["draft_id"], draft_id)
            self.assertEqual(operation["source_sheet"], "stingray_options")
            self.assertEqual(
                operation["changed_fields"],
                {
                    "option_name": {
                        "before": original_name,
                        "after": f"{original_name} changed",
                    }
                },
            )
            listed = self.client.get(f"/api/drafts/{draft_id}/operations")
            self.assertEqual(listed.status_code, 200, listed.text)
            self.assertEqual(len(listed.json()["operations"]), 1)
            status = self.client.get("/api/status").json()
            self.assertEqual(status["draft"]["active"], 1)
            self.assertEqual(status["draft"]["state"], "blocked")
            self.assertFalse(status["projection"]["reimport_allowed"])
            unchanged = self.client.get(
                "/api/records/options",
                params={"model": "stingray", "search": option_id, "limit": 1},
            ).json()["records"][0]
            self.assertEqual(unchanged["option_name"], original_name)
        finally:
            conn = self.mainmod.open_state_connection()
            try:
                conn.execute("DELETE FROM draft_operations WHERE draft_id=?", (draft_id,))
                conn.execute("DELETE FROM workflow_drafts WHERE id=?", (draft_id,))
                conn.commit()
            finally:
                conn.close()

    def test_reimport_endpoint_promotes_a_verified_replacement(self):
        fixture = verified_manager_fixture()
        conn = self.mainmod.open_projection_connection()
        before_manifest = self.mainmod._projection_manifest(conn)
        before_options = conn.execute("SELECT COUNT(*) c FROM options").fetchone()["c"]
        conn.close()

        def promote_fixture(_workbook, destination):
            candidate = self.tmpdir / "api-reimport-candidate.sqlite3"
            fixture.clone_projection(candidate)
            replacement = self.mainmod.dbmod.connect(candidate)
            try:
                replacement.execute(
                    "UPDATE storage_manifest SET migration_id='import-api-reimport'"
                )
                replacement.commit()
                replacement.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            finally:
                replacement.close()
            self.mainmod.dbmod._replace_projection(candidate, destination)
            return dict(fixture.promotion_report)

        with mock.patch.object(
            self.mainmod.importer,
            "promote_verified_projection",
            side_effect=promote_fixture,
        ) as promote:
            resp = self.client.post("/api/import")
        self.assertEqual(resp.status_code, 200, resp.text)
        self.assertTrue(resp.json()["promoted"])
        promote.assert_called_once_with(
            self.mainmod.config.DEFAULT_WORKBOOK,
            self.mainmod.config.DEFAULT_PROJECTION_DB,
        )
        conn = self.mainmod.open_projection_connection()
        self.addCleanup(conn.close)
        self.assertNotEqual(self.mainmod._projection_manifest(conn), before_manifest)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) c FROM options").fetchone()["c"], before_options
        )

    def test_current_projection_allows_verified_comparison_export(self):
        status = self.client.get("/api/status").json()
        self.assertEqual(status["projection"]["state"], "current")
        export = {
            "ok": True,
            "status": "exported",
            "byte_identical": True,
            "semantic_readback_verified": True,
            "path": str(self.tmpdir / "DISPOSABLE-api-export.xlsx"),
        }
        with mock.patch.object(
            self.mainmod.syncmod,
            "export_comparison_workbook",
            return_value=export,
        ) as export_workbook:
            resp = self.client.post("/api/export")
        self.assertEqual(resp.status_code, 200, resp.text)
        self.assertTrue(resp.json()["byte_identical"])
        self.assertTrue(resp.json()["semantic_readback_verified"])
        export_workbook.assert_called_once()

    def test_import_reports_all_unresolved_legacy_workflow_blockers(self):
        conn = self.mainmod.open_state_connection()
        self.addCleanup(conn.close)
        pending_id = conn.execute(
            "INSERT INTO pending_changes(ts, table_name, entity_key_json, op, "
            "status) VALUES('test', 'options', '{}', 'update', 'staged')"
        ).lastrowid
        pending_history_id = conn.execute(
            "INSERT INTO change_history(ts, entity_type, entity_id, op, status, "
            "sync_status) VALUES('test', 'options', 'one', 'update', "
            "'committed', 'pending')"
        ).lastrowid
        failed_history_id = conn.execute(
            "INSERT INTO change_history(ts, entity_type, entity_id, op, status, "
            "sync_status) VALUES('test', 'options', 'two', 'update', "
            "'committed', 'sync_failed')"
        ).lastrowid
        conn.commit()
        try:
            resp = self.client.post("/api/import")
            self.assertEqual(resp.status_code, 409)
            blockers = resp.json()["detail"]["import_blockers"]
            self.assertEqual(blockers["staged"], 1)
            self.assertEqual(blockers["committed_unsynchronized"], 1)
            self.assertEqual(blockers["failed"], 1)
            self.assertEqual(blockers["unresolved_total"], 3)
        finally:
            conn.execute("DELETE FROM pending_changes WHERE id=?", (pending_id,))
            conn.execute(
                "DELETE FROM change_history WHERE id IN (?, ?)",
                (pending_history_id, failed_history_id),
            )
            conn.commit()

    def test_zz_apply_rebuild_copied_workbook_mixed_draft_and_replay(self):
        draft_id = "api-apply-rebuild-mixed"
        canonical_paths = [
            WORKBOOK,
            REPO_ROOT / "form-app" / "data.js",
            REPO_ROOT / "form-app" / "index.html",
            *(REPO_ROOT / "form-output" / "runtime").glob("*-runtime-contract.json"),
        ]
        canonical_before = {path: _sha256(path) for path in canonical_paths}
        workbook_before = _sha256(self.workbook)
        ordinary = self.client.get(
            "/api/records/options", params={"model": "stingray", "limit": 1}
        ).json()["records"][0]
        saved = self.client.post(
            f"/api/drafts/{draft_id}/operations",
            json={
                "table": "options",
                "model_id": "stingray",
                "op": "update",
                "key": {"option_id": ordinary["option_id"]},
                "record": {"option_name": ordinary["option_name"] + " apply proof"},
            },
        )
        self.assertEqual(saved.status_code, 200, saved.text)

        asset_view = self.client.get(
            "/api/assets/reconciliation",
            params={"model": "stingray", "status": "covered", "draft_id": draft_id, "refresh": True},
        ).json()
        asset = asset_view["queue"]["items"][0]
        saved_asset = self.client.post(
            f"/api/drafts/{draft_id}/asset-resolutions",
            json={
                "item_id": asset["id"],
                "resolution_kind": "edit",
                "fingerprints": asset_view["fingerprints"],
                "values": {
                    "image_alt": asset["current_values"]["image_alt"] + " apply proof",
                },
            },
        )
        self.assertEqual(saved_asset.status_code, 200, saved_asset.text)

        committed = self.client.post(f"/api/drafts/{draft_id}/commit")
        self.assertEqual(committed.status_code, 200, committed.text)
        preview = self.client.post(f"/api/drafts/{draft_id}/preview")
        self.assertEqual(preview.status_code, 200, preview.text)
        confirmable = preview.json()["result"]["warningPolicy"]["confirmableIds"]
        approved = self.client.post(
            f"/api/drafts/{draft_id}/approve",
            json={"actor": "apply-proof", "warning_ids": confirmable},
        )
        self.assertEqual(approved.status_code, 200, approved.text)

        applied = self.client.post(
            f"/api/drafts/{draft_id}/apply-rebuild",
            json={"actor": "apply-proof", "confirm": "APPLY AND REBUILD"},
        )
        self.assertEqual(applied.status_code, 200, applied.text)
        attempt = applied.json()
        self.assertEqual(attempt["manager_state"], "applied")
        evidence = attempt["result"]["applyRebuild"]
        self.assertEqual(evidence["status"], "current")
        self.assertEqual(evidence["affected_models"], ["stingray"])
        self.assertEqual(evidence["workbook"]["state"], "applied")
        self.assertEqual(evidence["generated_contracts"]["state"], "current")
        self.assertEqual(evidence["publication"]["state"], "current")
        self.assertNotEqual(_sha256(self.workbook), workbook_before)

        replay = self.client.post(
            f"/api/drafts/{draft_id}/apply-rebuild",
            json={"actor": "apply-proof", "confirm": "APPLY AND REBUILD"},
        )
        self.assertEqual(replay.status_code, 200, replay.text)
        self.assertEqual(replay.json()["id"], attempt["id"])
        state = self.client.get("/api/status").json()
        self.assertEqual(state["projection"]["state"], "stale")
        self.assertEqual(state["generated_artifacts"]["state"], "current")
        self.assertEqual(state["publication"]["state"], "current")
        self.assertEqual({path: _sha256(path) for path in canonical_paths}, canonical_before)


if __name__ == "__main__":
    unittest.main(verbosity=2)
