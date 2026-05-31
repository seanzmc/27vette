from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_z_rule_audit import build_z_rule_audit


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


OPTION_HEADERS = ["option_id", "rpo", "option_name", "section_id", "selectable", "active", "display_behavior"]
OVS_HEADERS = ["option_id", "variant_id", "status"]
RULE_HEADERS = [
    "rule_id",
    "source_id",
    "rule_type",
    "target_id",
    "target_type",
    "original_detail_raw",
    "review_flag",
    "source_type",
    "target_selection_mode",
    "source_selection_mode",
    "target_section",
    "source_section",
    "generation_action",
    "body_style_scope",
    "runtime_action",
    "disabled_reason",
    "normalization_status",
    "normalization_reason",
    "replacement_group_id",
    "replacement_rule_id",
]
GROUP_HEADERS = ["group_id", "group_type", "source_id", "body_style_scope", "trim_level_scope", "variant_scope", "disabled_reason", "active", "notes"]
GROUP_MEMBER_HEADERS = ["group_id", "target_id", "display_order", "active"]
EXCLUSIVE_HEADERS = ["group_id", "selection_mode", "active", "notes"]
EXCLUSIVE_MEMBER_HEADERS = ["group_id", "option_id", "display_order", "active"]
DEFAULT_HEADERS = ["model_key", "rule_id", "target_option_id", "condition_type", "condition_id", "body_style_scope", "trim_level_scope", "variant_scope", "priority", "active", "notes"]
VARIANT_HEADERS = ["variant_id", "model_key", "body_style", "trim_level", "active"]


def audit_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    append_sheet(
        wb,
        "variant_master",
        VARIANT_HEADERS,
        [
            {"variant_id": "z06_coupe", "model_key": "z06", "body_style": "coupe", "trim_level": "1LZ", "active": True},
            {"variant_id": "z06_conv", "model_key": "z06", "body_style": "convertible", "trim_level": "1LZ", "active": True},
        ],
    )
    append_sheet(
        wb,
        "z06_options",
        OPTION_HEADERS,
        [
            {"option_id": "opt_a", "rpo": "AAA", "option_name": "Option A", "section_id": "sec_a", "selectable": True, "active": True},
            {"option_id": "opt_b", "rpo": "BBB", "option_name": "Option B", "section_id": "sec_b", "selectable": True, "active": True},
            {"option_id": "opt_c", "rpo": "CCC", "option_name": "Option C", "section_id": "sec_b", "selectable": True, "active": True, "display_behavior": "default_selected"},
            {"option_id": "opt_missing_status", "rpo": "DDD", "option_name": "Option D", "section_id": "sec_d", "selectable": True, "active": True},
            {"option_id": "opt_inactive", "rpo": "EEE", "option_name": "Inactive", "section_id": "sec_i", "selectable": True, "active": False},
        ],
    )
    append_sheet(
        wb,
        "z06_ovs",
        OVS_HEADERS,
        [
            {"option_id": "opt_a", "variant_id": "z06_coupe", "status": "available"},
            {"option_id": "opt_a", "variant_id": "z06_conv", "status": "available"},
            {"option_id": "opt_b", "variant_id": "z06_coupe", "status": "available"},
            {"option_id": "opt_b", "variant_id": "z06_conv", "status": "unavailable"},
            {"option_id": "opt_c", "variant_id": "z06_coupe", "status": "standard"},
            {"option_id": "opt_c", "variant_id": "z06_conv", "status": "standard"},
        ],
    )
    append_sheet(
        wb,
        "z06_rule_mapping",
        RULE_HEADERS,
        [
            {"rule_id": "r1", "source_id": "opt_a", "rule_type": "excludes", "target_id": "opt_b", "target_type": "option", "source_type": "option", "normalization_status": "active"},
            {"rule_id": "r2", "source_id": "opt_a", "rule_type": "excludes", "target_id": "opt_b", "target_type": "option", "source_type": "option", "normalization_status": "active"},
            {"rule_id": "r3", "source_id": "opt_a", "rule_type": "requires", "target_id": "missing", "target_type": "option", "source_type": "option", "normalization_status": "active"},
            {"rule_id": "r4", "source_id": "opt_inactive", "rule_type": "includes", "target_id": "opt_c", "target_type": "option", "source_type": "option", "normalization_status": "active"},
        ],
    )
    append_sheet(wb, "z06_rule_groups", GROUP_HEADERS, [{"group_id": "g1", "group_type": "requires_any", "source_id": "opt_a", "active": True}])
    append_sheet(wb, "z06_rule_group_members", GROUP_MEMBER_HEADERS, [{"group_id": "g1", "target_id": "opt_b", "active": True}, {"group_id": "g1", "target_id": "missing", "active": True}])
    append_sheet(wb, "z06_exclusive_groups", EXCLUSIVE_HEADERS, [{"group_id": "x1", "selection_mode": "single_within_group", "active": True}])
    append_sheet(wb, "z06_exclusive_members", EXCLUSIVE_MEMBER_HEADERS, [{"group_id": "x1", "option_id": "opt_a", "active": True}, {"group_id": "x1", "option_id": "opt_b", "active": True}])
    append_sheet(wb, "default_selection_rules", DEFAULT_HEADERS, [{"model_key": "z06", "rule_id": "d1", "target_option_id": "missing", "condition_type": "always", "active": True}])
    return wb


def test_z_rule_audit_reports_integrity_and_schema_level_rule_signals() -> None:
    audit = build_z_rule_audit(audit_workbook(), ["z06"])

    z06 = audit["models"]["z06"]
    assert z06["summary"]["ruleMappingRows"] == 4
    assert z06["summary"]["exclusiveGroups"] == 1
    assert z06["summary"]["ruleGroups"] == 1
    assert z06["ruleTypeCounts"] == {"excludes": 2, "includes": 1, "requires": 1}
    assert z06["focusedReviewCounts"]["duplicateSemanticRules"] == 1
    assert z06["focusedReviewCounts"]["directExcludesCoveredByExclusiveGroups"] == 2
    assert z06["focusedReviewCounts"]["missingOptionReferences"] == 1
    assert z06["focusedReviewCounts"]["inactiveOptionReferences"] == 1
    assert z06["focusedReviewCounts"]["missingRuleGroupMemberReferences"] == 1
    assert z06["focusedReviewCounts"]["missingDefaultRuleTargets"] == 1
    assert z06["focusedReviewCounts"]["optionsMissingVariantStatuses"] == 1
    assert z06["hotSpots"]["defaults"]["defaultSelectionRules"] == 1


def test_z_rule_audit_supports_empty_missing_optional_group_sheets() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    append_sheet(wb, "variant_master", VARIANT_HEADERS, [{"variant_id": "zr1_base", "model_key": "zr1", "active": True}])
    append_sheet(wb, "zr1_options", OPTION_HEADERS, [{"option_id": "opt_a", "rpo": "AAA", "section_id": "sec_a", "active": True}])
    append_sheet(wb, "zr1_ovs", OVS_HEADERS, [{"option_id": "opt_a", "variant_id": "zr1_base", "status": "available"}])
    append_sheet(wb, "zr1_rule_mapping", RULE_HEADERS, [])
    append_sheet(wb, "zr1_exclusive_groups", EXCLUSIVE_HEADERS, [])
    append_sheet(wb, "zr1_exclusive_members", EXCLUSIVE_MEMBER_HEADERS, [])
    append_sheet(wb, "default_selection_rules", DEFAULT_HEADERS, [])

    audit = build_z_rule_audit(wb, ["zr1"])

    assert audit["models"]["zr1"]["summary"]["ruleMappingRows"] == 0
    assert audit["models"]["zr1"]["summary"]["ruleGroups"] == 0
    assert audit["models"]["zr1"]["focusedReviewCounts"]["missingOptionReferences"] == 0
