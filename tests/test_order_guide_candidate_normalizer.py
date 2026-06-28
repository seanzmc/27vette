#!/usr/bin/env python3
"""Tests for the Pass 1 raw order-guide candidate normalizer."""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from collections import Counter
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.ingest.candidate_normalizer import (  # noqa: E402
    normalize_order_guide_candidates,
)
from corvette_form_generator.ingest.source_profiler import profile_order_guide  # noqa: E402


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def fixture_workbook(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "model_master",
        [
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "notes",
        ],
        [
            {
                "model_key": "stingray",
                "registry_key": "stingray",
                "model_label": "Stingray",
                "model_year": "2027",
                "dataset_name": "Fixture Stingray",
                "export_slug": "stingray",
                "expected_variant_count": 1,
                "default_model": True,
                "active": True,
            },
            {
                "model_key": "zr1",
                "registry_key": "zr1",
                "model_label": "ZR1",
                "model_year": "2027",
                "dataset_name": "Fixture ZR1",
                "export_slug": "zr1",
                "expected_variant_count": 1,
                "default_model": False,
                "active": False,
            },
        ],
    )
    append_sheet(
        wb,
        "variant_master",
        ["variant_id", "model_year", "trim_level", "body_style", "display_name", "base_price", "display_order", "active"],
        [
            {
                "variant_id": "1lt_c07",
                "model_year": "2027",
                "trim_level": "1lt",
                "body_style": "coupe",
                "display_name": "Corvette Stingray Coupe 1LT",
                "base_price": 71000,
                "display_order": 1,
                "active": True,
            },
            {
                "variant_id": "1lz_r07",
                "model_year": "2027",
                "trim_level": "1lz",
                "body_style": "coupe",
                "display_name": "Corvette ZR1 Coupe 1LZ",
                "base_price": 194700,
                "display_order": 2,
                "active": False,
            },
        ],
    )
    append_sheet(
        wb,
        "model_variants",
        ["model_key", "variant_id", "display_order", "active", "notes"],
        [
            {"model_key": "stingray", "variant_id": "1lt_c07", "display_order": 1, "active": True},
            {"model_key": "zr1", "variant_id": "1lz_r07", "display_order": 1, "active": False},
        ],
    )
    append_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        [
            {
                "model_key": "stingray",
                "source_role": "source_option_sheet",
                "sheet_name": "stingray_options",
                "active": True,
            }
        ],
    )
    append_sheet(
        wb,
        "stingray_options",
        [
            "option_id",
            "rpo",
            "price",
            "option_name",
            "description",
            "detail_raw",
            "section_id",
            "selectable",
            "display_order",
            "active",
            "display_behavior",
        ],
        [
            {
                "option_id": "opt_eri_001",
                "rpo": "ERI",
                "option_name": "Battery Protection Package",
                "section_id": "sec_add_001",
                "selectable": True,
                "display_order": 1,
                "active": True,
            },
            {
                "option_id": "opt_abc_001",
                "rpo": "ABC",
                "option_name": "Example Package",
                "section_id": "sec_pkg_001",
                "selectable": True,
                "display_order": 2,
                "active": True,
            },
        ],
    )
    wb.save(path)


def raw_export_fixture(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    ws = wb.create_sheet("Price Schedule")
    ws.append(["2027 CHEVROLET CORVETTE", "", "", ""])
    ws.append(["", "", "", ""])
    ws.append(["", "", "", ""])
    ws.append(["", "", "", ""])
    ws.append(["", "", "", ""])
    ws.append(["", "Model", "Model Description", "MSRP(c)"])
    ws.append(["", "1YC07", "Stingray Coupe", "$71,000"])

    ws = wb.create_sheet("Exterior 1")
    ws.append(["Stingray", "", "", ""])
    ws.append(["", "", "S = Standard Equipment  A = Available  -- = Not Available", ""])
    ws.append([
        "Orderable RPO Code",
        "Ref. Only RPO Code",
        "Description",
        "Coupe / 1YC07 / 1LT",
    ])
    ws.append(["Additional Options", "", "", ""])
    ws.append([
        "ERI",
        "",
        "Battery Protection Package / 1. Requires (ABC) Example Package. 2. Not available with (XYZ) Example Conflict.",
        "A1",
    ])
    ws.append(["", "EYT", "Carbon Flash Exterior Badge Package", "S"])

    ws = wb.create_sheet("Exterior 4")
    ws.append(["ZR1", "", "", ""])
    ws.append(["", "", "S = Standard Equipment  A = Available  -- = Not Available", ""])
    ws.append([
        "Orderable RPO Code",
        "Ref. Only RPO Code",
        "Description",
        "ZR1 Coupe / 1YR07 / 1LZ",
    ])
    ws.append(["TOM", "", "Carbon Fiber Aero Package", "A"])

    ws = wb.create_sheet("Color and Trim 1")
    ws.append(["Recommended", "", "", "", ""])
    ws.append(["A = Available  -- = Not Available", "", "", "", ""])
    ws.append(["", "", "", "Interior Colors", ""])
    ws.append(["Decor Level", "Seat Type", "Seat Code", "Seat Trim", "Jet Black"])
    ws.append(["1LT, 1LZ", "GT1 buckets", "AQ9", "Mulan leather", "HTA"])
    wb.save(path)


def build_evidence(tmp: Path) -> tuple[Path, Path]:
    workbook = tmp / "canonical.xlsx"
    raw_export = tmp / "raw.xlsx"
    evidence_dir = tmp / "evidence"
    fixture_workbook(workbook)
    raw_export_fixture(raw_export)
    profile_order_guide(
        raw_export=raw_export,
        workbook=workbook,
        output_dir=evidence_dir,
        run_id="unit-evidence",
        root=ROOT,
    )
    return workbook, evidence_dir


class OrderGuideCandidateNormalizerTests(unittest.TestCase):
    def test_normalizes_option_ovs_rule_and_out_of_scope_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir = build_evidence(tmp)
            output_dir = tmp / "candidates"

            result = normalize_order_guide_candidates(
                evidence_dir=evidence_dir,
                workbook=workbook,
                output_dir=output_dir,
                run_id="unit-candidates",
                root=ROOT,
            )

            self.assertEqual(result["status"], "passed")
            for name in [
                "candidate-options.json",
                "candidate-ovs.json",
                "candidate-rules.json",
                "candidate-price-rules.json",
                "candidate-summary.json",
                "unresolved-review.json",
                "unresolved-review.md",
            ]:
                self.assertTrue((output_dir / name).exists(), name)

            options = json.loads((output_dir / "candidate-options.json").read_text())
            eri = next(option for option in options if option["normalized_values"]["rpo"] == "ERI")
            self.assertEqual(eri["candidate_family"], "options")
            self.assertEqual(eri["normalized_values"]["candidate_option_ref"], "candopt-exterior-1-row-5-eri")
            self.assertEqual(eri["source_refs"][0]["source_sheet"], "Exterior 1")
            self.assertEqual(eri["source_refs"][0]["source_row_index"], 5)
            self.assertEqual(eri["normalized_values"]["canonical_option_match"]["option_id"], "opt_eri_001")
            self.assertEqual(eri["normalized_values"]["section_id_candidate"], "sec_add_001")
            self.assertEqual(eri["normalized_values"]["status_summary"]["stingray"]["1lt_c07"]["available"], 1)
            self.assertEqual(eri["raw_values"]["source_description_raw"].split(" / ")[0], "Battery Protection Package")

            ovs = json.loads((output_dir / "candidate-ovs.json").read_text())
            eri_ovs = next(row for row in ovs if row["normalized_values"]["candidate_option_ref"] == "candopt-exterior-1-row-5-eri")
            self.assertEqual(eri_ovs["normalized_values"]["variant_id"], "1lt_c07")
            self.assertEqual(eri_ovs["normalized_values"]["model_key"], "stingray")
            self.assertEqual(eri_ovs["normalized_values"]["raw_status"], "A1")
            self.assertEqual(eri_ovs["normalized_values"]["normalized_status_candidate"], "available")
            self.assertEqual(eri_ovs["normalized_values"]["status_marker"], "1")
            self.assertEqual(eri_ovs["normalized_values"]["source_cell"], "D5")

            rules = json.loads((output_dir / "candidate-rules.json").read_text())
            requires = next(row for row in rules if row["normalized_values"]["marker"] == "1")
            self.assertEqual(requires["resolution_status"], "needs_review")
            self.assertEqual(requires["normalized_values"]["relationship_hint"], "requires")
            self.assertEqual(requires["normalized_values"]["target_rpo_tokens"], ["ABC"])
            self.assertEqual(requires["normalized_values"]["target_match_status"], "exact")
            excludes = next(row for row in rules if row["normalized_values"]["marker"] == "2")
            self.assertEqual(excludes["normalized_values"]["relationship_hint"], "excludes")
            self.assertEqual(excludes["normalized_values"]["target_match_status"], "unresolved")

            price_rules = json.loads((output_dir / "candidate-price-rules.json").read_text())
            self.assertEqual(price_rules, [])
            unresolved = (output_dir / "unresolved-review.md").read_text()
            self.assertIn("price_schedule_rows_not_extracted", unresolved)
            self.assertIn("color_trim_rows_not_extracted", unresolved)
            self.assertIn("section_context_requires_review", unresolved)

            unresolved_json = json.loads((output_dir / "unresolved-review.json").read_text())
            self.assertEqual(unresolved_json["version"], 1)
            self.assertEqual(unresolved_json["run_id"], "unit-candidates")
            self.assertIn("price_schedule_rows_not_extracted", unresolved_json["unresolved_counts"])
            self.assertEqual(unresolved_json["unresolved_counts"], dict(Counter(item["reason"] for item in unresolved_json["items"])))
            by_reason = {item["reason"]: item for item in unresolved_json["items"]}
            price_item = by_reason["price_schedule_rows_not_extracted"]
            self.assertEqual(price_item["category"], "price_out_of_scope")
            self.assertEqual(price_item["severity"], "out_of_scope")
            self.assertTrue(price_item["source_refs"])
            self.assertIn("blocked_out_of_scope", price_item["suggested_decision_states"])
            target_item = by_reason["target_rpo_token_ambiguous_or_missing"]
            self.assertEqual(target_item["category"], "relationship_hint")
            self.assertTrue(target_item["source_refs"])
            self.assertIn("description_fragment", target_item["raw_values"])

            summary = json.loads((output_dir / "candidate-summary.json").read_text())
            self.assertEqual(summary["candidate_counts"]["price_rules"], 0)
            self.assertGreaterEqual(summary["candidate_counts"]["options"], 2)
            self.assertGreaterEqual(summary["unresolved_counts"]["price_schedule_rows_not_extracted"], 1)
            self.assertIn("unresolved-review.json", summary["artifact_files"])

    def test_rejects_failed_or_incomplete_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir = build_evidence(tmp)
            manifest_path = evidence_dir / "manifest.json"
            manifest = json.loads(manifest_path.read_text())
            manifest["status"] = "failed"
            manifest_path.write_text(json.dumps(manifest) + "\n")

            with self.assertRaisesRegex(ValueError, "status.*passed"):
                normalize_order_guide_candidates(
                    evidence_dir=evidence_dir,
                    workbook=workbook,
                    output_dir=tmp / "candidates",
                    run_id="unit-candidates",
                    root=ROOT,
                )

    def test_output_dir_guard_rejects_tracked_generated_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir = build_evidence(tmp)
            with self.assertRaisesRegex(ValueError, "form-output/ingest"):
                normalize_order_guide_candidates(
                    evidence_dir=evidence_dir,
                    workbook=workbook,
                    output_dir=ROOT / "form-output" / "runtime",
                    run_id="unit-candidates",
                    root=ROOT,
                )


if __name__ == "__main__":
    unittest.main()
