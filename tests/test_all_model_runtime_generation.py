#!/usr/bin/env python3
"""The six-model executable gate (spec Pass 2 requirement 10).

Every workbook-discovered active model is generated through the real operator
entrypoint into one isolated candidate root, and each written artifact is then
re-read from disk and put through the *same* strict validator that generation
uses. Nothing here trusts the generator's own report of its work.

The model set is checked from both sides: against ``model_master`` read directly
from the workbook, and against the named six. Neither alone is enough — the
workbook side catches a discovery regression, the named set catches a model
silently leaving the workbook and shrinking this file's coverage with it.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PYTHON = ROOT / ".venv" / "bin" / "python"
GENERATE_FORM = ROOT / "scripts" / "generate_form.py"
WORKBOOK = ROOT / "stingray_master.xlsx"
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_configs import discover_generation_model_configs  # noqa: E402
from corvette_form_generator.runtime_contract import assert_runtime_contract  # noqa: E402

# Membership, not only count. A count-only gate passes a workbook edit that
# renames or swaps a model key while leaving six rows active, which is exactly
# the silent-coverage-loss case requirement 10 exists to prevent.
EXPECTED_MODEL_KEYS = frozenset({"stingray", "grand_sport", "grand_sport_x", "z06", "zr1", "zr1x"})

# Every model reports the same keys. An absent key is a route change, not a
# per-model difference.
REQUIRED_SUMMARY_KEYS = frozenset(
    {
        "model_key",
        "model_label",
        "route_engine",
        "runtime_contract_json",
        "runtime_contract_artifacts",
        "inspection_artifacts",
        "preview_artifacts",
        "draft_artifacts",
        "counts",
        "validation_errors",
        "notes",
    }
)


def workbook_active_models() -> dict[str, int]:
    """Active model keys and their declared variant counts, read from the workbook.

    Deliberately does not call ``discover_generation_model_configs``, so the two
    can be compared. Note this reads ``active`` strictly (``true``/``1``/``yes``)
    while discovery treats a blank cell as active; the difference is why a blank
    ``active`` cell surfaces here rather than passing silently.
    """

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        rows = workbook["model_master"].iter_rows(values_only=True)
        headers = list(next(rows))
        key = headers.index("model_key")
        active = headers.index("active")
        variants = headers.index("expected_variant_count")
        return {
            str(row[key]).strip(): int(row[variants])
            for row in rows
            if row[key] and str(row[active]).strip().lower() in {"true", "1", "yes"}
        }
    finally:
        workbook.close()


def protected_hashes() -> dict[str, str]:
    paths = [WORKBOOK, ROOT / "form-app" / "data.js"]
    paths.extend(path for path in (ROOT / "form-output").rglob("*") if path.is_file())
    return {str(path.relative_to(ROOT)): hashlib.sha256(path.read_bytes()).hexdigest() for path in paths}


@pytest.fixture(scope="module")
def generated(tmp_path_factory) -> dict[str, object]:
    """Generate every active model once, into one isolated candidate root."""

    tmp_path = tmp_path_factory.mktemp("six-model-gate")
    workbook_dir = tmp_path / "workbook"
    workbook_dir.mkdir()
    # Keep the canonical filename: ``dataset.source_workbook`` records it, so a
    # renamed snapshot would show up as a spurious contract difference.
    snapshot = workbook_dir / WORKBOOK.name
    shutil.copy2(WORKBOOK, snapshot)
    candidate_root = tmp_path / "candidate"

    before = protected_hashes()
    configs = discover_generation_model_configs(snapshot, root=candidate_root)
    results: dict[str, dict] = {}
    for model_key in sorted(configs):
        completed = subprocess.run(
            [
                str(PYTHON),
                str(GENERATE_FORM),
                "--model",
                model_key,
                "--workbook",
                str(snapshot),
                "--output-root",
                str(candidate_root),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        )
        results[model_key] = {
            "returncode": completed.returncode,
            "stderr": completed.stderr,
            "summary": json.loads(completed.stdout) if completed.returncode == 0 else None,
        }

    return {
        "configs": configs,
        "results": results,
        "candidate_root": candidate_root,
        "hashes_before": before,
    }


def contract_for(generated: dict, model_key: str) -> dict:
    summary = generated["results"][model_key]["summary"]
    assert summary is not None, generated["results"][model_key]["stderr"]
    path = Path(summary["runtime_contract_json"])
    assert path.is_relative_to(generated["candidate_root"]), f"{model_key} wrote outside the candidate root: {path}"
    return json.loads(path.read_text(encoding="utf-8"))


def section_rpo_order(contract: dict, section_id: str) -> list[str]:
    """Return one RPO per active choice in its lowest authored display order."""

    order_by_rpo: dict[str, int] = {}
    for choice in contract["choices"]:
        if choice["section_id"] != section_id or choice["active"] != "True":
            continue
        rpo = choice["rpo"]
        order = int(choice["display_order"])
        order_by_rpo[rpo] = min(order, order_by_rpo.get(rpo, order))
    return [rpo for rpo, _ in sorted(order_by_rpo.items(), key=lambda item: (item[1], item[0]))]


def test_discovery_matches_the_workbooks_own_active_model_set(generated) -> None:
    """Breaks if `discover_generation_model_configs()` stops returning what the workbook activates.

    Scope, stated honestly: discovery *raises* rather than silently dropping a
    model whose sources are incomplete, so for a well-formed workbook both sides
    are the same set by construction and this comparison only catches a
    regression in discovery itself. A workbook that deactivates a model is
    caught by `test_named_models_are_active_and_green`, not here.
    """

    assert set(generated["configs"]) == set(workbook_active_models())


def test_named_models_are_active_and_green(generated) -> None:
    """Six named models, every one generated successfully.

    Asserts membership, not just the count: a workbook edit that renames a model
    key while leaving six rows active must fail here rather than shrink the
    parametrized coverage below in silence.
    """

    assert set(workbook_active_models()) == EXPECTED_MODEL_KEYS
    assert set(generated["configs"]) == EXPECTED_MODEL_KEYS
    failed = {key: value["stderr"] for key, value in generated["results"].items() if value["returncode"] != 0}
    assert failed == {}, f"{len(failed)} of {len(EXPECTED_MODEL_KEYS)} models failed generation"


@pytest.mark.parametrize("model_key", sorted(EXPECTED_MODEL_KEYS))
def test_written_artifact_passes_the_strict_validator(model_key, generated) -> None:
    """The gate and generation must call one validator; this re-runs it on the file on disk."""

    assert_runtime_contract(
        contract_for(generated, model_key),
        source=f"six-model gate {model_key}",
        config=generated["configs"][model_key],
    )


@pytest.mark.parametrize("model_key", sorted(EXPECTED_MODEL_KEYS))
def test_dataset_binds_to_the_candidate_snapshot(model_key, generated) -> None:
    """The spec's rejection matrix requires this; the strict validator does not yet enforce it.

    ``assert_runtime_contract`` checks dataset name/model/model_year but not
    ``source_workbook``, so a contract generated from one workbook and written
    into another candidate's root would pass it. Checked here until the
    validator owns it.
    """

    contract = contract_for(generated, model_key)

    assert contract["dataset"]["source_workbook"] == WORKBOOK.name


@pytest.mark.parametrize("model_key", sorted(EXPECTED_MODEL_KEYS))
def test_variant_count_matches_the_workbooks_declared_count(model_key, generated) -> None:
    """Breaks if a variant is added to or dropped from a model without updating model_master."""

    contract = contract_for(generated, model_key)
    variant_ids = [row["variant_id"] for row in contract["variants"]]

    assert len(variant_ids) == workbook_active_models()[model_key]
    assert len(set(variant_ids)) == len(variant_ids), f"{model_key} has duplicate variant ids"


@pytest.mark.parametrize("model_key", sorted(EXPECTED_MODEL_KEYS))
def test_reported_summary_agrees_with_the_written_artifact(model_key, generated) -> None:
    """Breaks if the CLI summary is ever derived from something other than the validated contract."""

    summary = generated["results"][model_key]["summary"]
    contract = contract_for(generated, model_key)

    assert REQUIRED_SUMMARY_KEYS <= set(summary)
    assert summary["model_key"] == model_key
    assert summary["validation_errors"] == 0
    assert summary["dataset_name"] == contract["dataset"]["name"]
    assert summary["status"] == contract["dataset"]["status"] == "runtime_active"
    for field, count in summary["counts"].items():
        assert count == len(contract[field]), f"{model_key}.{field}: summary {count} vs artifact {len(contract[field])}"


def test_fresh_unpublished_contracts_preserve_workbook_owned_roof_order(generated) -> None:
    """Migrate the only product assertion from the retained-artifact gate to fresh generation."""

    assert section_rpo_order(contract_for(generated, "grand_sport_x"), "sec_roof_001") == [
        "CF7",
        "C2Z",
        "CC3",
        "CM9",
        "CF8",
        "D84",
        "D86",
    ]
    zr1_contract = contract_for(generated, "zr1")
    assert section_rpo_order(zr1_contract, "sec_roof_001") == []
    assert section_rpo_order(zr1_contract, "sec_stan_001").count("C2Z") >= 1
    assert all(choice["rpo"] != "CFC" for choice in zr1_contract["choices"])
    zr1x_contract = contract_for(generated, "zr1x")
    assert section_rpo_order(zr1x_contract, "sec_roof_001") == []
    zr1x_standard_rpos = section_rpo_order(zr1x_contract, "sec_stan_001")
    assert zr1x_standard_rpos.count("C2Z") >= 1
    assert zr1x_standard_rpos.count("CFC") >= 1


def test_fresh_unpublished_contracts_preserve_generated_order_summary_metadata(generated) -> None:
    """Move retained-artifact metadata assertions into the current all-model generation owner."""

    for model_key in ("grand_sport_x", "zr1", "zr1x"):
        contract = contract_for(generated, model_key)
        expects_required_charges = model_key in {"zr1", "zr1x"}
        assert len(contract["steps"]) == 14
        assert len(contract["orderSummary"]["sections"]) == (12 if expects_required_charges else 11)
        assert len(contract["orderSummary"]["stepMap"]) == (14 if expects_required_charges else 13)
        assert contract["orderSummary"]["stepMap"]["base_interior"] == "seats_interior"
        assert ("standard_equipment" in contract["orderSummary"]["stepMap"]) is expects_required_charges


def test_strict_validation_rejects_what_a_status_and_error_scan_would_accept(generated) -> None:
    """Negative proof for the gate itself.

    Without this, the strict-validator assertions above could pass for a payload
    a much weaker check would also accept, and requirement 10 would be unproven.
    """

    contract = contract_for(generated, "stingray")
    config = generated["configs"]["stingray"]

    weakly_valid = dict(contract, steps=[])
    weakly_valid.pop("orderSummary", None)
    # The check the earlier CLI gate performed, and which this payload survives:
    assert weakly_valid["dataset"]["status"] == "runtime_active"
    assert [row for row in weakly_valid["validation"] if str(row.get("severity", "")).lower() == "error"] == []

    with pytest.raises(ValueError) as excinfo:
        assert_runtime_contract(weakly_valid, source="negative proof", config=config)

    message = str(excinfo.value)
    assert "steps must not be empty" in message
    assert "orderSummary must be an object" in message


def test_generation_left_every_protected_surface_untouched(generated) -> None:
    assert protected_hashes() == generated["hashes_before"]
