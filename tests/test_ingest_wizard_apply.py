#!/usr/bin/env python3
"""Pass D apply CLI/store tests.

Fixture workbooks only. The live workbook is never written here; compact
fixture workbooks disable schema validation because they are not schema-complete.
"""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from corvette_form_generator.ingest.wizard.session import (  # noqa: E402
    WizardError,
    WizardSessionStore,
    read_json,
    write_json,
)
from ingest_wizard_fixtures import build_master_workbook, build_raw_export  # noqa: E402


class ApplyFlowTest(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        build_raw_export(self.root / "raw.xlsx")
        self.master = build_master_workbook(self.root / "master.xlsx")
        self.store = WizardSessionStore(self.root, workbook_path=self.master)
        created = self.store.create_session("raw.xlsx")
        self.run_id = created["session"]["runId"]
        roles = {card["sheetName"]: card["recommendedRole"] for card in created["profile"]["sheets"]}
        self.store.confirm_roles(self.run_id, roles)
        self.store.run_parse(self.run_id)
        self.store.select_models(self.run_id, ["zr1", "zr1x"], {"zr1": "z06", "zr1x": "z06"})

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def complete_model(self, model: str) -> None:
        decisions = []
        reconciliation = self.store.reconciliation(self.run_id)["models"].get(model, {})
        if not reconciliation.get("agrees", True):
            decisions.append(
                {
                    "model": model,
                    "lane": "relationship",
                    "groupKey": "variant_reconciliation",
                    "action": "needs_product_decision",
                    "payload": {},
                    "resolution": "approved_for_plan",
                }
            )
        for candidate in self.store.review_queue(self.run_id, model, "section")["candidates"]:
            decisions.append(
                {
                    "model": model,
                    "lane": "section",
                    "candidateId": candidate["candidateId"],
                    "action": "assign_section",
                    "payload": {"sectionId": "sec_whee_001"},
                    "resolution": "approved_for_plan",
                }
            )
        for candidate in self.store.review_queue(self.run_id, model, "price")["candidates"]:
            decisions.append(
                {
                    "model": model,
                    "lane": "price",
                    "candidateId": candidate["candidateId"],
                    "action": "accept_exact_price" if candidate["priceMatch"] == "exact" else "confirm_no_price",
                    "payload": {},
                    "resolution": "approved_for_plan",
                }
            )
        for candidate in self.store.review_queue(self.run_id, model, "status_nuance")["candidates"]:
            decisions.append(
                {
                    "model": model,
                    "lane": "status_nuance",
                    "candidateId": candidate["candidateId"],
                    "action": "confirm_status",
                    "payload": {},
                    "resolution": "approved_for_plan",
                }
            )
        prefill = self.store.review_queue(self.run_id, model, "presentation")["prefill"]
        for sheet, proposals in prefill["sheets"].items():
            decisions.append(
                {
                    "model": model,
                    "lane": "presentation",
                    "groupKey": sheet,
                    "action": "approve_presentation_rows",
                    "payload": {"rows": [p["row"] for p in proposals], "templateModel": "z06"},
                    "resolution": "approved_for_plan",
                }
            )
        self.store.save_decisions(self.run_id, decisions)

    def approve_plan(self) -> Path:
        self.complete_model("zr1")
        self.complete_model("zr1x")
        self.store.mark_complete(self.run_id)
        result = self.store.build_apply_plan(self.run_id, schema_validation=False)
        self.assertTrue(result["dryRun"]["ok"], result["dryRun"])
        self.store.approve_plan(self.run_id, "sean")
        return self.store.run_dir(self.run_id)

    def test_default_dry_run_writes_report_and_leaves_workbook_unchanged(self) -> None:
        run_dir = self.approve_plan()
        before = self.master.read_bytes()

        result = self.store.apply_approved_plan(self.run_id, schema_validation=False)

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["status"], "validated")
        self.assertEqual(self.master.read_bytes(), before)
        report = read_json(run_dir / "apply-dry-run-report.json")
        plan = read_json(run_dir / "apply-plan.json")
        self.assertEqual(report["schemaVersion"], "pass-d-1")
        self.assertEqual(report["planSchemaVersion"], "pass-c-2")
        self.assertIn("startedAt", report)
        self.assertIn("completedAt", report)
        self.assertEqual(report["approvedBy"], "sean")
        self.assertIn("approvedAt", report)
        self.assertFalse(report["write"])
        self.assertEqual(report["status"], "validated")
        self.assertGreater(report["opCounts"]["stage1"], 0)
        self.assertGreater(report["opCounts"]["stage2"], 0)
        self.assertEqual(report["opCounts"]["combined"], len(plan["stage1"]["items"]) + len(plan["stage2"]["items"]))
        self.assertGreater(report["perSheetCounts"].get("model_workbook_sources", 0), 0)
        self.assertGreater(report["perSheetActionCounts"]["model_workbook_sources"]["add"], 0)
        self.assertEqual(report["boolHygieneResult"]["error_count"], 0)
        self.assertEqual(report["deploymentContinuity"]["zr1"]["status"], "not_deployment_ready")
        self.assertIn("registryLoadable", report["deploymentContinuity"]["zr1"])
        self.assertIn("registryError", report["deploymentContinuity"]["zr1"])
        self.assertIn(
            "price_rules_required_for_runtime",
            {item["kind"] for item in report["deploymentContinuity"]["zr1"]["deploymentBlockers"]},
        )
        self.assertEqual(
            report["deploymentContinuity"]["zr1"]["sourceCoverage"]["priceRuleAddOrUpdateCount"],
            0,
        )
        self.assertEqual(report["applyResult"]["status"], "validated")
        self.assertEqual(report["workbookBefore"], report["workbookAfter"])
        self.assertIsNone(report["backupPath"])
        self.assertEqual(report["verification"]["mismatches"], [])
        self.assertEqual(read_json(run_dir / "session.json")["state"], "plan_approved")

    def test_apply_refuses_stale_or_unapproved_inputs(self) -> None:
        with self.assertRaisesRegex(WizardError, "approved"):
            self.store.apply_approved_plan(self.run_id, schema_validation=False)
        run_dir = self.approve_plan()

        approval = read_json(run_dir / "plan-approval.json")
        approval["planSha"] = "0" * 64
        write_json(run_dir / "plan-approval.json", approval)
        with self.assertRaisesRegex(WizardError, "Plan approval hash"):
            self.store.apply_approved_plan(self.run_id, schema_validation=False)
        self.store.approve_plan(self.run_id, "sean")

        plan = read_json(run_dir / "apply-plan.json")
        plan["workbookFingerprint"]["sha256"] = "0" * 64
        write_json(run_dir / "apply-plan.json", plan)
        approval = read_json(run_dir / "plan-approval.json")
        approval["planSha"] = __import__("hashlib").sha256((run_dir / "apply-plan.json").read_bytes()).hexdigest()
        write_json(run_dir / "plan-approval.json", approval)
        with self.assertRaisesRegex(WizardError, "Workbook changed"):
            self.store.apply_approved_plan(self.run_id, schema_validation=False)

    def test_write_refuses_superseded_pass_c1_plan(self) -> None:
        run_dir = self.approve_plan()
        plan = read_json(run_dir / "apply-plan.json")
        plan["schemaVersion"] = "pass-c-1"
        write_json(run_dir / "apply-plan.json", plan)
        approval = read_json(run_dir / "plan-approval.json")
        approval["planSha"] = __import__("hashlib").sha256((run_dir / "apply-plan.json").read_bytes()).hexdigest()
        write_json(run_dir / "plan-approval.json", approval)

        with self.assertRaisesRegex(WizardError, "superseded"):
            self.store.apply_approved_plan(
                self.run_id,
                write=True,
                confirm_plan_warnings=True,
                schema_validation=False,
            )

    def test_dry_run_failure_still_writes_report(self) -> None:
        run_dir = self.approve_plan()
        before = self.master.read_bytes()

        with patch(
            "corvette_form_generator.editor_ops.apply_batch",
            return_value={
                "ok": False,
                "status": "bool_hygiene_failed",
                "errors": ["dry-run bool hygiene failed with 1 error(s)"],
                "warnings": [],
                "boolHygieneResult": {"error_count": 1},
            },
        ):
            result = self.store.apply_approved_plan(self.run_id, schema_validation=False)

        self.assertFalse(result["ok"])
        self.assertEqual(result["status"], "bool_hygiene_failed")
        self.assertEqual(self.master.read_bytes(), before)
        report = read_json(run_dir / "apply-dry-run-report.json")
        self.assertFalse(report["write"])
        self.assertEqual(report["status"], "bool_hygiene_failed")
        self.assertEqual(report["applyResult"]["status"], "bool_hygiene_failed")
        self.assertEqual(report["boolHygieneResult"]["error_count"], 1)
        self.assertEqual(report["workbookBefore"], report["workbookAfter"])
        self.assertEqual(report["deploymentContinuity"], {})
        self.assertEqual(read_json(run_dir / "session.json")["state"], "plan_approved")

    def test_write_requires_warning_confirmation(self) -> None:
        self.approve_plan()

        result = self.store.apply_approved_plan(self.run_id, write=True, schema_validation=False)

        self.assertFalse(result["ok"])
        self.assertEqual(result["status"], "needs_confirmation")
        self.assertTrue(result["warnings"])
        self.assertFalse((self.store.run_dir(self.run_id) / "apply-report.json").exists())
        self.assertEqual(self.master.parent.joinpath("workbook-edit-log.jsonl").exists(), False)

    def test_confirmed_write_applies_combined_plan_reports_and_locks_run(self) -> None:
        run_dir = self.approve_plan()

        result = self.store.apply_approved_plan(
            self.run_id,
            write=True,
            confirm_plan_warnings=True,
            schema_validation=False,
        )

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["status"], "applied")
        report = read_json(run_dir / "apply-report.json")
        self.assertTrue(report["write"])
        self.assertEqual(report["status"], "applied")
        self.assertTrue(Path(report["backupPath"]).is_file())
        self.assertTrue(Path(report["workbookEditLogPath"]).is_file())
        self.assertEqual(report["verification"]["mismatches"], [])
        self.assertEqual(read_json(run_dir / "session.json")["state"], "applied")

        wb = load_workbook(self.master, read_only=True, data_only=True)
        try:
            self.assertIn("zr1_rule_mapping", wb.sheetnames)
            self.assertIn("zr1x_rule_mapping", wb.sheetnames)
            self.assertFalse(
                any(row[8] for row in wb["model_master"].iter_rows(min_row=2, values_only=True) if row[0] in {"zr1", "zr1x"})
            )
            self.assertFalse(
                any(row[2] for row in wb["model_registry_promotion"].iter_rows(min_row=2, values_only=True) if row[0] in {"zr1", "zr1x"})
            )
        finally:
            wb.close()

        with self.assertRaisesRegex(WizardError, "already applied"):
            self.store.apply_approved_plan(self.run_id, write=True, confirm_plan_warnings=True, schema_validation=False)
        with self.assertRaisesRegex(WizardError, "already applied"):
            self.store.build_apply_plan(self.run_id, schema_validation=False)
        with self.assertRaisesRegex(WizardError, "already applied"):
            candidate = self.store.review_queue(self.run_id, "zr1", "section")["candidates"][0]
            self.store.save_decisions(
                self.run_id,
                [
                    {
                        "model": "zr1",
                        "lane": "section",
                        "candidateId": candidate["candidateId"],
                        "action": "assign_section",
                        "payload": {"sectionId": "sec_pain_001"},
                        "resolution": "approved_for_plan",
                    }
                ],
            )
        with self.assertRaisesRegex(WizardError, "already applied"):
            self.store.select_models(self.run_id, ["zr1"], {"zr1": "z06"})

    def test_cli_dry_run_outputs_json_and_report(self) -> None:
        run_dir = self.approve_plan()

        completed = subprocess.run(
            [
                sys.executable,
                str(ROOT / "scripts" / "ingest_wizard_apply.py"),
                "--root",
                str(self.root),
                "--workbook",
                str(self.master),
                "--run",
                self.run_id,
                "--no-schema-validation",
            ],
            check=False,
            text=True,
            capture_output=True,
        )

        self.assertEqual(completed.returncode, 0, completed.stderr + completed.stdout)
        payload = json.loads(completed.stdout)
        self.assertEqual(payload["status"], "validated")
        self.assertTrue((run_dir / "apply-dry-run-report.json").is_file())


if __name__ == "__main__":
    unittest.main()
