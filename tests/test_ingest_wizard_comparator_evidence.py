#!/usr/bin/env python3
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from corvette_form_generator.ingest.wizard.comparator_evidence import build_comparator_evidence  # noqa: E402
from ingest_wizard_fixtures import build_master_workbook  # noqa: E402


class ComparatorEvidenceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.workbook = build_master_workbook(Path(self.tmp.name) / "master.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(self.workbook)
        wb["z06_options"].append(["opt_bv4_001", "BV4", 395, "Personalized Plaque", "Fixture endpoint", "", "sec_whee_001", True, 15, True, ""])
        wb["default_selection_rules"].append(
            ["z06", "z06_default_pdb", "opt_pdb_001", "always", "", "", "", "", 1, True, "", "default_selected"]
        )
        wb.save(self.workbook)
        wb.close()

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_requires_generation_discoverable_active_comparator(self) -> None:
        with self.assertRaisesRegex(ValueError, "not generation-discoverable"):
            build_comparator_evidence(self.workbook, {"zr1": "zr1"})

    def test_indexes_runtime_effective_facts_by_rpo_not_option_id(self) -> None:
        artifact = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        self.assertEqual(artifact["schemaVersion"], "comparator-evidence-1")
        entry = artifact["targets"]["zr1"]
        self.assertEqual(entry["comparator"], "z06")
        types = {fact["factType"] for fact in entry["facts"]}
        self.assertTrue({"direct_rule", "rule_group", "exclusive_group", "price_rule", "default_selection"}.issubset(types))
        direct = next(fact for fact in entry["facts"] if fact["factType"] == "direct_rule")
        self.assertEqual(direct["signature"]["sourceRpo"], "PDB")
        self.assertEqual(direct["signature"]["targetRpo"], "BV4")
        self.assertNotIn("opt_pdb_001", str(direct["signature"]))
        default = next(fact for fact in entry["facts"] if fact["factType"] == "default_selection")
        self.assertIn("displayBehavior", default["signature"])

    def test_inactive_occurrence_is_preserved_but_not_portable(self) -> None:
        artifact = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        occurrences = artifact["targets"]["zr1"]["optionOccurrences"]
        xfr = next(item for item in occurrences if item["rpo"] == "XFR")
        self.assertFalse(xfr["runtimeActive"])
        self.assertEqual(xfr["disposition"], "inactive_context_only")

    def test_member_order_does_not_change_group_signature(self) -> None:
        artifact = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        group = next(fact for fact in artifact["targets"]["zr1"]["facts"] if fact["factType"] == "exclusive_group")
        self.assertEqual(group["signature"]["memberRpos"], ["BV4", "PDB"])

    def test_comparator_prices_are_context_only(self) -> None:
        artifact = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        price = next(fact for fact in artifact["targets"]["zr1"]["facts"] if fact["factType"] == "price_rule")
        self.assertEqual(price["context"]["priceValue"], "500")
        self.assertEqual(price["disposition"], "corroborating_context_only")

    def test_semantic_hash_is_stable(self) -> None:
        first = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        second = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        self.assertEqual(first["comparatorEvidenceSemanticSha"], second["comparatorEvidenceSemanticSha"])

    def test_semantic_hash_ignores_source_row_reordering(self) -> None:
        before = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        from openpyxl import load_workbook

        wb = load_workbook(self.workbook)
        ws = wb["z06_options"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        ws.delete_rows(2, len(rows))
        for row in reversed(rows):
            ws.append(list(row))
        wb.save(self.workbook)
        wb.close()
        after = build_comparator_evidence(self.workbook, {"zr1": "z06"})
        self.assertEqual(
            before["comparatorEvidenceSemanticSha"],
            after["comparatorEvidenceSemanticSha"],
        )


if __name__ == "__main__":
    unittest.main()
