#!/usr/bin/env python3
"""Contract, coverage, determinism, and refusal tests for the ingest
canonical-manifest ChangeSet emitter (workbook-changeset-1)."""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from corvette_form_generator.editor_ops import extract_workbook  # noqa: E402
from corvette_form_generator.ingest.wizard.changeset_emitter import (  # noqa: E402
    emit_manifest_changeset,
)
from corvette_form_generator.workbook_domain.changeset import canonical_json  # noqa: E402
from ingest_wizard_fixtures import build_master_workbook  # noqa: E402


@pytest.fixture()
def workbook(tmp_path):
    return build_master_workbook(tmp_path / "master.xlsx")


def _manifest_row(*, model, family, sheet, action, key, values, status="ready"):
    return {
        "model": model,
        "family": family,
        "sheet": sheet,
        "action": action,
        "key": key,
        "values": values,
        "status": status,
        "semanticSignature": {"fixture": [sheet, key]},
        "derivationVersion": "fixture",
    }


def _inputs(workbook_path, *, rows, targets, modes, comparator_targets=None):
    authority = {
        "fingerprint": "authority-sha",
        "bindings": {"compilerPolicyVersion": "fixture"},
    }
    return {
        "workbook_path": workbook_path,
        "run_id": "fixture-run",
        "manifest": {
            "schemaVersion": "canonical-row-manifest-v1",
            "manifestSemanticSha": "manifest-semantic-sha",
            "runAuthorityFingerprint": authority,
            "modelModes": modes,
            "rows": rows,
        },
        "compile_report": {
            "models": {
                target: {"mode": modes[target], "compileReady": True, "blockers": []}
                for target in targets
            },
            "deferrals": [],
            "runAuthorityFingerprint": authority,
            "queueSubjectFingerprint": "queue-sha",
            "comparatorEvidenceSemanticSha": "comparator-semantic-sha",
        },
        "selection": {
            "targets": targets,
            "comparators": {target: "z06" for target in targets},
            "sourceFingerprint": "source-sha",
            "candidatesFingerprint": "candidate-sha",
        },
        "compiler_bindings": {
            "canonicalManifestSha": "manifest-file-sha",
            "compileReportSha": "report-file-sha",
            "exceptionResolutionsSha": "resolution-file-sha",
            "exceptionQueueSha": "queue-file-sha",
            "comparatorEvidenceSha": "comparator-file-sha",
        },
        "authority_artifacts": {
            "exceptionQueue": {
                "queueSubjectFingerprint": "queue-sha",
                "comparatorEvidenceSemanticSha": "comparator-semantic-sha",
                "runAuthorityFingerprint": authority,
                "subjects": [],
            },
            "resolutions": {
                "queueSubjectFingerprint": "queue-sha",
                "validEntries": [],
            },
            "comparatorEvidence": {
                "comparatorEvidenceSemanticSha": "comparator-semantic-sha",
                "runAuthorityFingerprint": authority,
                "targets": {target: {} for target in (comparator_targets or targets)},
            },
        },
    }


def _zr1_rows(workbook_path):
    """noop (model_master) + update (options) + add (options) for zr1."""
    extract = extract_workbook(workbook_path)
    model_row = next(
        row
        for row in extract["sheets"]["model_master"]["rows"]
        if row["model_key"] == "zr1"
    )
    existing = dict(extract["sheets"]["zr1_options"]["rows"][0])
    updated = {**existing, "description": "Emitter updated description"}
    headers = extract["sheets"]["zr1_options"]["headers"]
    added = {header: None for header in headers}
    added.update(
        {
            "option_id": "opt_fixture_new_001",
            "rpo": "NEW",
            "price": 100,
            "option_name": "Fixture option",
            "description": "Fixture option",
            "section_id": "sec_whee_001",
            "selectable": True,
            "active": True,
        }
    )
    return [
        _manifest_row(
            model="zr1",
            family="model_master",
            sheet="model_master",
            action="noop",
            key={"model_key": "zr1"},
            values=model_row,
        ),
        _manifest_row(
            model="zr1",
            family="options",
            sheet="zr1_options",
            action="update",
            key={"option_id": existing["option_id"]},
            values=updated,
        ),
        _manifest_row(
            model="zr1",
            family="options",
            sheet="zr1_options",
            action="add",
            key={"option_id": added["option_id"]},
            values=added,
        ),
    ], existing


def test_emitter_covers_every_row_and_binds_manifest(workbook):
    rows, _existing = _zr1_rows(workbook)
    changeset = emit_manifest_changeset(
        **_inputs(workbook, rows=rows, targets=["zr1"], modes={"zr1": "reprocess"})
    )
    covered = {
        item["provenance"][0]["manifestRef"]
        for item in [*changeset["rowChanges"], *changeset["noops"]]
    }
    expected = {f"manifest-{index:05d}" for index in range(len(rows))}
    assert covered == expected
    assert changeset["bindings"]["canonicalManifestSha"] == "manifest-file-sha"
    assert changeset["source"] == {"kind": "ingest", "runId": "fixture-run"}
    assert changeset["changeSetId"] == changeset["semanticFingerprint"][:24]
    assert changeset["targets"] == ["zr1"]


def test_emitter_is_byte_deterministic_and_ignores_decisions(workbook, tmp_path):
    rows, _existing = _zr1_rows(workbook)
    inputs = _inputs(workbook, rows=rows, targets=["zr1"], modes={"zr1": "reprocess"})
    first = emit_manifest_changeset(**inputs)
    (tmp_path / "decisions.json").write_text('{"decisions":[{"id":"legacy"}]}')
    second = emit_manifest_changeset(**inputs)
    assert canonical_json(first) == canonical_json(second)


def test_update_fields_are_exact_deltas_with_workbook_before_values(workbook):
    rows, existing = _zr1_rows(workbook)
    changeset = emit_manifest_changeset(
        **_inputs(workbook, rows=rows, targets=["zr1"], modes={"zr1": "reprocess"})
    )
    update = next(
        change for change in changeset["rowChanges"] if change["action"] == "update"
    )
    assert set(update["fields"]) == {"description"}
    assert update["fields"]["description"] == {
        "before": existing["description"],
        "after": "Emitter updated description",
    }
    add = next(change for change in changeset["rowChanges"] if change["action"] == "add")
    assert add["fields"]
    assert all(pair["before"] is None for pair in add["fields"].values())
    assert all(pair["after"] is not None for pair in add["fields"].values())


def test_typed_normalization_noop_becomes_update(workbook):
    from openpyxl import load_workbook

    wb = load_workbook(workbook)
    ws = wb["zr1_options"]
    headers = {
        str(cell.value): index + 1
        for index, cell in enumerate(ws[1])
        if cell.value is not None
    }
    ws.cell(row=2, column=headers["selectable"], value="False")
    ws.cell(row=2, column=headers["active"], value="True")
    ws.cell(row=2, column=headers["price"], value="0")
    wb.save(workbook)
    wb.close()

    row = extract_workbook(workbook)["sheets"]["zr1_options"]["rows"][0]
    values = {**row, "selectable": False, "active": True, "price": 0}
    changeset = emit_manifest_changeset(
        **_inputs(
            workbook,
            rows=[
                _manifest_row(
                    model="zr1",
                    family="options",
                    sheet="zr1_options",
                    action="noop",
                    key={"option_id": row["option_id"]},
                    values=values,
                )
            ],
            targets=["zr1"],
            modes={"zr1": "reprocess"},
        )
    )
    assert changeset["noops"] == []
    assert len(changeset["rowChanges"]) == 1
    change = changeset["rowChanges"][0]
    assert change["action"] == "update"
    assert change["fields"]["selectable"] == {"before": "False", "after": False}
    assert change["fields"]["price"] == {"before": "0", "after": 0}


def test_greenfield_scaffold_is_one_named_non_manifest_row_change(workbook):
    extract = extract_workbook(workbook)
    master_headers = extract["sheets"]["model_master"]["headers"]
    gsx = {header: None for header in master_headers}
    gsx.update(
        {
            "model_key": "grand_sport_x",
            "registry_key": "grand_sport_x",
            "model_label": "Grand Sport X",
            "expected_variant_count": 6,
            "default_model": False,
            "active": False,
        }
    )
    changeset = emit_manifest_changeset(
        **_inputs(
            workbook,
            rows=[
                _manifest_row(
                    model="grand_sport_x",
                    family="model_master",
                    sheet="model_master",
                    action="add",
                    key={"model_key": "grand_sport_x"},
                    values=gsx,
                )
            ],
            targets=["grand_sport_x"],
            modes={"grand_sport_x": "greenfield"},
        )
    )
    scaffolds = [
        change
        for change in changeset["rowChanges"]
        if change["provenance"][0].get("kind") == "scaffold"
    ]
    assert len(scaffolds) == 1
    scaffold = scaffolds[0]
    assert scaffold["sheet"] == "model_registry_promotion"
    assert scaffold["family"] == "model_registry_promotion"
    assert scaffold["key"] == {"model_key": "grand_sport_x"}
    assert scaffold["fields"]["promoted_to_runtime"] == {"before": None, "after": False}
    assert scaffold["fields"]["active"] == {"before": None, "after": False}
    assert "manifestRef" not in scaffold["provenance"][0]


def _noop_model_row(workbook, model="zr1"):
    extract = extract_workbook(workbook)
    values = next(
        row
        for row in extract["sheets"]["model_master"]["rows"]
        if row["model_key"] == model
    )
    return _manifest_row(
        model=model,
        family="model_master",
        sheet="model_master",
        action="noop",
        key={"model_key": model},
        values=values,
    )


def test_refuses_unselected_model_row(workbook):
    row = _noop_model_row(workbook)
    row["model"] = "stingray"
    with pytest.raises(ValueError, match="unselected model"):
        emit_manifest_changeset(
            **_inputs(workbook, rows=[row], targets=["zr1"], modes={"zr1": "reprocess"})
        )


def test_refuses_non_ready_row(workbook):
    row = _noop_model_row(workbook)
    row["status"] = "blocked"
    with pytest.raises(ValueError, match="is not ready"):
        emit_manifest_changeset(
            **_inputs(workbook, rows=[row], targets=["zr1"], modes={"zr1": "reprocess"})
        )


def test_refuses_unknown_family(workbook):
    row = _noop_model_row(workbook)
    row["family"] = "not_a_family"
    with pytest.raises(ValueError, match="unknown family"):
        emit_manifest_changeset(
            **_inputs(workbook, rows=[row], targets=["zr1"], modes={"zr1": "reprocess"})
        )


def test_refuses_header_vector_mismatch(workbook):
    row = _noop_model_row(workbook)
    row["values"] = {**row["values"], "not_a_column": "x"}
    with pytest.raises(ValueError, match="header vector"):
        emit_manifest_changeset(
            **_inputs(workbook, rows=[row], targets=["zr1"], modes={"zr1": "reprocess"})
        )


def test_refuses_duplicate_physical_key(workbook):
    row = _noop_model_row(workbook)
    with pytest.raises(ValueError, match="more than once"):
        emit_manifest_changeset(
            **_inputs(
                workbook,
                rows=[row, dict(row)],
                targets=["zr1"],
                modes={"zr1": "reprocess"},
            )
        )


def test_refuses_unsupported_action(workbook):
    row = _noop_model_row(workbook)
    row["action"] = "mutate"
    with pytest.raises(ValueError, match="unsupported action"):
        emit_manifest_changeset(
            **_inputs(workbook, rows=[row], targets=["zr1"], modes={"zr1": "reprocess"})
        )


def test_refuses_stale_authority_binding(workbook):
    inputs = _inputs(
        workbook,
        rows=[_noop_model_row(workbook)],
        targets=["zr1"],
        modes={"zr1": "reprocess"},
    )
    inputs["compile_report"]["runAuthorityFingerprint"] = {
        "fingerprint": "different",
        "bindings": {},
    }
    with pytest.raises(ValueError, match="authority fingerprints do not agree"):
        emit_manifest_changeset(**inputs)


def test_refuses_stale_queue_binding(workbook):
    inputs = _inputs(
        workbook,
        rows=[_noop_model_row(workbook)],
        targets=["zr1"],
        modes={"zr1": "reprocess"},
    )
    inputs["authority_artifacts"]["resolutions"]["queueSubjectFingerprint"] = "stale"
    with pytest.raises(ValueError, match="queue and resolution fingerprints"):
        emit_manifest_changeset(**inputs)


def test_refuses_non_compile_ready_target(workbook):
    inputs = _inputs(
        workbook,
        rows=[_noop_model_row(workbook)],
        targets=["zr1"],
        modes={"zr1": "reprocess"},
    )
    inputs["compile_report"]["models"]["zr1"]["blockers"] = ["fixture-blocker"]
    with pytest.raises(ValueError, match="not compile-ready"):
        emit_manifest_changeset(**inputs)


def test_refuses_comparator_evidence_target_drift(workbook):
    with pytest.raises(ValueError, match="comparator evidence target"):
        emit_manifest_changeset(
            **_inputs(
                workbook,
                rows=[_noop_model_row(workbook)],
                targets=["zr1"],
                modes={"zr1": "reprocess"},
                comparator_targets=["zr1x"],
            )
        )


def test_refuses_greenfield_isolation_discarding_non_noop(workbook):
    rows = [
        _manifest_row(
            model="grand_sport_x",
            family="rule_mapping",
            sheet="grandSport_rule_mapping",
            action="update",
            key={"rule_id": "rule_outside_domain"},
            values={
                "rule_id": "rule_outside_domain",
                "source_id": "opt_outside_001",
                "rule_type": "requires",
                "target_id": "opt_outside_002",
                "original_detail_raw": "",
                "body_style_scope": "",
                "runtime_action": "",
                "disabled_reason": "",
            },
        )
    ]
    with pytest.raises(ValueError, match="cannot discard a non-noop"):
        emit_manifest_changeset(
            **_inputs(
                workbook,
                rows=rows,
                targets=["grand_sport_x"],
                modes={"grand_sport_x": "greenfield"},
            )
        )
