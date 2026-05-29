#!/usr/bin/env python3
"""Tests for dry-run future-model compatibility rebasing."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_compatibility import (  # noqa: E402
    EXCLUSIVE_GROUP_HEADERS,
    EXCLUSIVE_MEMBER_HEADERS,
    RULE_GROUP_HEADERS,
    RULE_GROUP_MEMBER_HEADERS,
    RULE_MAPPING_HEADERS,
    apply_future_compatibility_to_workbook,
    build_future_compatibility_preview,
)
from corvette_form_generator.future_model_ingest import OPTION_SOURCE_HEADERS  # noqa: E402


def append_sheet(wb: Workbook, name: str, headers: tuple[str, ...] | list[str], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    for row in rows or []:
        ws.append([row.get(header, "") for header in headers])


def sheet_rows(wb: Workbook, name: str) -> list[dict[str, object]]:
    ws = wb[name]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header: value for header, value in zip(headers, values) if header}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def option(option_id: str, rpo: str, *, active: object = True) -> dict[str, object]:
    return {
        "option_id": option_id,
        "rpo": rpo,
        "price": "",
        "option_name": f"{rpo} option",
        "description": "",
        "detail_raw": "",
        "section_id": "sec_test",
        "selectable": True,
        "display_order": 1,
        "active": active,
        "display_behavior": "card",
    }


def rule(rule_id: str, source_id: str, target_id: str, **overrides: object) -> dict[str, object]:
    row: dict[str, object] = {header: "" for header in RULE_MAPPING_HEADERS}
    row.update(
        {
            "rule_id": rule_id,
            "source_id": source_id,
            "rule_type": "excludes",
            "target_id": target_id,
            "target_type": "option",
            "review_flag": False,
            "source_type": "option",
            "normalization_status": "active",
        }
    )
    row.update(overrides)
    return row


def compatibility_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "grandSport_options",
        OPTION_SOURCE_HEADERS,
        [
            option("gs_src", "SRC"),
            option("gs_tgt", "TGT"),
            option("gs_missing", "MISS"),
            option("gs_dupe_a", "DUP"),
            option("gs_dupe_b", "DUP"),
            option("gs_ex1", "EX1"),
            option("gs_ex2", "EX2"),
            option("gs_ex3", "EX3"),
        ],
    )
    append_sheet(
        wb,
        "z06_options",
        OPTION_SOURCE_HEADERS,
        [
            option("z_src", "SRC"),
            option("z_tgt", "TGT"),
            option("z_dupe_a", "DUP"),
            option("z_dupe_b", "DUP"),
            option("z_ex1", "EX1"),
            option("z_ex2", "EX2"),
        ],
    )
    for model_key in ("zr1", "zr1x"):
        append_sheet(wb, f"{model_key}_options", OPTION_SOURCE_HEADERS, [])

    append_sheet(
        wb,
        "grandSport_rule_mapping",
        RULE_MAPPING_HEADERS,
        [
            rule("gs_rule_src_tgt", "gs_src", "gs_tgt"),
            rule("gs_rule_interior", "gs_src", "gs_tgt", source_type="interior"),
            rule("gs_rule_missing_target", "gs_src", "gs_missing"),
            rule("gs_rule_replaced", "gs_src", "gs_tgt", normalization_status="replaced"),
            rule("gs_rule_duplicate_source", "gs_dupe_a", "gs_tgt"),
        ],
    )
    append_sheet(
        wb,
        "grandSport_rule_groups",
        RULE_GROUP_HEADERS,
        [
            {
                "group_id": "gs_group_keep",
                "group_type": "requires_any",
                "source_id": "gs_src",
                "active": True,
                "notes": "kept with pruned members",
            },
            {
                "group_id": "gs_group_drop",
                "group_type": "requires_any",
                "source_id": "gs_missing",
                "active": True,
            },
        ],
    )
    append_sheet(
        wb,
        "grandSport_rule_group_members",
        RULE_GROUP_MEMBER_HEADERS,
        [
            {"group_id": "gs_group_keep", "target_id": "gs_tgt", "display_order": 10, "active": True},
            {"group_id": "gs_group_keep", "target_id": "gs_missing", "display_order": 20, "active": True},
            {"group_id": "gs_group_drop", "target_id": "gs_tgt", "display_order": 10, "active": True},
        ],
    )
    append_sheet(
        wb,
        "grandSport_exclusive_groups",
        EXCLUSIVE_GROUP_HEADERS,
        [
            {"group_id": "gs_excl_keep", "selection_mode": "single_within_group", "active": True},
            {"group_id": "gs_excl_drop", "selection_mode": "single_within_group", "active": True},
        ],
    )
    append_sheet(
        wb,
        "grandSport_exclusive_members",
        EXCLUSIVE_MEMBER_HEADERS,
        [
            {"group_id": "gs_excl_keep", "option_id": "gs_ex1", "display_order": 10, "active": True},
            {"group_id": "gs_excl_keep", "option_id": "gs_ex2", "display_order": 20, "active": True},
            {"group_id": "gs_excl_keep", "option_id": "gs_ex3", "display_order": 30, "active": True},
            {"group_id": "gs_excl_drop", "option_id": "gs_ex1", "display_order": 10, "active": True},
            {"group_id": "gs_excl_drop", "option_id": "gs_ex3", "display_order": 20, "active": True},
        ],
    )
    for model_key in ("z06", "zr1", "zr1x"):
        append_sheet(wb, f"{model_key}_rule_mapping", RULE_MAPPING_HEADERS, [])
        append_sheet(wb, f"{model_key}_rule_groups", RULE_GROUP_HEADERS, [])
        append_sheet(wb, f"{model_key}_rule_group_members", RULE_GROUP_MEMBER_HEADERS, [])
        append_sheet(wb, f"{model_key}_exclusive_groups", EXCLUSIVE_GROUP_HEADERS, [])
        append_sheet(wb, f"{model_key}_exclusive_members", EXCLUSIVE_MEMBER_HEADERS, [])
    return wb


class FutureModelCompatibilityRebaseTests(unittest.TestCase):
    def test_rebases_rules_by_unique_active_rpo_and_reports_skips(self) -> None:
        preview = build_future_compatibility_preview(compatibility_workbook(), ["z06"])
        z06 = preview["models"]["z06"]

        self.assertEqual(z06["proposed_row_counts"]["rule_mapping"], 1)
        rebased_rule = z06["proposed_rows"]["rule_mapping"][0]
        self.assertEqual(rebased_rule["rule_id"], "z06_rule_src_tgt")
        self.assertEqual(rebased_rule["source_id"], "z_src")
        self.assertEqual(rebased_rule["target_id"], "z_tgt")

        self.assertEqual(z06["skipped"]["rule_mapping"]["deferred_source_type_interior"], 1)
        self.assertEqual(z06["skipped"]["rule_mapping"]["inactive_or_replaced_source_rule"], 1)
        self.assertEqual(z06["skipped"]["rule_mapping"]["target_id:target_rpo_not_found"], 1)
        self.assertEqual(z06["skipped"]["rule_mapping"]["source_id:source_duplicate_active_rpo"], 1)
        self.assertEqual(z06["rebase"]["mapped_option_count"], 4)

    def test_prunes_rule_group_members_and_requires_two_exclusive_members(self) -> None:
        preview = build_future_compatibility_preview(compatibility_workbook(), ["z06"])
        z06 = preview["models"]["z06"]

        self.assertEqual(z06["proposed_row_counts"]["rule_groups"], 1)
        self.assertEqual(z06["proposed_rows"]["rule_groups"][0]["group_id"], "z06_group_keep")
        self.assertEqual(z06["proposed_row_counts"]["rule_group_members"], 1)
        self.assertEqual(z06["proposed_rows"]["rule_group_members"][0]["target_id"], "z_tgt")
        self.assertEqual(z06["skipped"]["rule_groups"]["source_id:target_rpo_not_found"], 1)
        self.assertEqual(z06["skipped"]["rule_group_members"]["target_id:target_rpo_not_found"], 1)

        self.assertEqual(z06["proposed_row_counts"]["exclusive_groups"], 1)
        self.assertEqual(z06["proposed_rows"]["exclusive_groups"][0]["group_id"], "z06_excl_keep")
        self.assertEqual([row["option_id"] for row in z06["proposed_rows"]["exclusive_members"]], ["z_ex1", "z_ex2"])
        self.assertEqual(z06["skipped"]["exclusive_groups"]["fewer_than_two_resolved_members"], 1)
        self.assertEqual(z06["skipped"]["exclusive_members"]["option_id:target_rpo_not_found"], 2)

    def test_preview_does_not_modify_workbook_file_or_target_sheets(self) -> None:
        wb = compatibility_workbook()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "preview.xlsx"
            wb.save(path)
            before_mtime = path.stat().st_mtime_ns
            before = load_workbook(path, read_only=True, data_only=True)
            before_rows = before["z06_rule_mapping"].max_row
            before.close()

            opened = load_workbook(path, read_only=False, data_only=True)
            preview = build_future_compatibility_preview(opened, ["z06"], include_rows=False)
            opened.close()

            after_mtime = path.stat().st_mtime_ns
            after = load_workbook(path, read_only=True, data_only=True)
            after_rows = after["z06_rule_mapping"].max_row
            after.close()

        self.assertEqual(preview["status"], "dry_run")
        self.assertNotIn("proposed_rows", preview["models"]["z06"])
        self.assertEqual(after_mtime, before_mtime)
        self.assertEqual(after_rows, before_rows)

    def test_write_mode_writes_selected_model_target_sheets_idempotently(self) -> None:
        wb = compatibility_workbook()
        preview = build_future_compatibility_preview(wb, ["z06"])

        first_report = apply_future_compatibility_to_workbook(wb, preview)
        second_report = apply_future_compatibility_to_workbook(wb, preview)

        self.assertEqual(first_report["status"], "applied")
        self.assertEqual(second_report["status"], "applied")
        self.assertEqual([row["rule_id"] for row in sheet_rows(wb, "z06_rule_mapping")], ["z06_rule_src_tgt"])
        self.assertEqual([row["group_id"] for row in sheet_rows(wb, "z06_rule_groups")], ["z06_group_keep"])
        self.assertEqual([row["target_id"] for row in sheet_rows(wb, "z06_rule_group_members")], ["z_tgt"])
        self.assertEqual([row["group_id"] for row in sheet_rows(wb, "z06_exclusive_groups")], ["z06_excl_keep"])
        self.assertEqual([row["option_id"] for row in sheet_rows(wb, "z06_exclusive_members")], ["z_ex1", "z_ex2"])
        self.assertEqual(sheet_rows(wb, "zr1_rule_mapping"), [])


if __name__ == "__main__":
    unittest.main()
