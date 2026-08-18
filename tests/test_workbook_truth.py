#!/usr/bin/env python3
"""Contract test for the independent workbook-truth snapshot.

Spec: docs/superpowers/specs/2026-08-17-fast-layered-validation-suite.md §6.2.

The snapshot is the expected side of every parity gate, so the thing worth
testing is not "does it produce rows" but the two properties that make it usable
as an oracle at all:

1. It is INDEPENDENT of the generator. `build_workbook_truth` may not reach
   generation, rule derivation, runtime cleanup, or business fallback logic,
   because a shared function cannot be both implementation and oracle. The
   import boundary is asserted against a freshly launched interpreter rather
   than against this process, which pytest has already polluted with generator
   imports from sibling modules.

2. Its two locally implemented cell helpers AGREE with the workbook contract the
   generator reads through. Independence is worth nothing if it silently means
   "a different answer" — so `clean` and `truthy` are pinned to
   `workbook.clean` / `workbook.workbook_truthy` over a value table that
   includes every representation class the workbook actually stores.

Everything else here checks that the snapshot says what it claims to say, and
each such check is followed by a forced mutation proving it can fail.
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = REPO_ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import build_workbook_truth as truth  # noqa: E402
from corvette_form_generator import workbook as generator_workbook  # noqa: E402
from corvette_form_generator.workbook_domain import registry as reg  # noqa: E402

WORKBOOK = REPO_ROOT / "stingray_master.xlsx"

# Modules that implement generation. The snapshot reaching any of them would
# make every parity gate circular, so the boundary is named explicitly rather
# than inferred from a prefix.
GENERATOR_MODULES = frozenset(
    {
        "corvette_form_generator.contract",
        "corvette_form_generator.inspection",
        "corvette_form_generator.model_configs",
        "corvette_form_generator.model_generation",
        "corvette_form_generator.output",
        "corvette_form_generator.production",
        "corvette_form_generator.registry_promotion",
        "corvette_form_generator.rule_derivation",
        "corvette_form_generator.rules",
        "corvette_form_generator.runtime_contract",
        "corvette_form_generator.schema_validation",
        "corvette_form_generator.source_assembly",
    }
)


@pytest.fixture(scope="module")
def snapshot() -> dict:
    return truth.build_workbook_truth(WORKBOOK)


# ── Property 1: independence from the generator ───────────────────────────────


def test_builder_imports_no_generation_module() -> None:
    """A fresh interpreter that builds a snapshot must not load generation."""

    program = (
        "import sys, json\n"
        f"sys.path.insert(0, {str(SCRIPTS_DIR)!r})\n"
        "import build_workbook_truth as t\n"
        f"t.build_workbook_truth({str(WORKBOOK)!r})\n"
        "print(json.dumps(sorted(sys.modules)))\n"
    )
    completed = subprocess.run(
        [sys.executable, "-c", program], cwd=REPO_ROOT, capture_output=True, text=True, check=True
    )
    loaded = set(json.loads(completed.stdout))
    assert not (loaded & GENERATOR_MODULES), sorted(loaded & GENERATOR_MODULES)


def test_local_cell_helpers_are_not_the_generator_functions() -> None:
    """Independence has to include representation, not only module names.

    `corvette_form_generator.workbook` is loaded in-process regardless — the
    shared `workbook_domain.registry` metadata imports it — so the module list
    above cannot see the failure this guards. Re-exporting `workbook.clean` or
    `workbook.workbook_truthy` here would leave every parity gate reading cells
    through the same code generation reads them through, and one representation
    bug would blind all of them at once while every test stayed green. The
    agreement table below is what keeps the two definitions equal; this is what
    keeps them two.
    """

    assert truth.clean is not generator_workbook.clean
    assert truth.truthy is not generator_workbook.workbook_truthy


def test_import_boundary_check_would_notice_a_generator_import() -> None:
    """The boundary test above is only worth its runtime if it can fail."""

    program = (
        "import sys, json\n"
        f"sys.path.insert(0, {str(SCRIPTS_DIR)!r})\n"
        "import build_workbook_truth as t\n"
        "import corvette_form_generator.runtime_contract\n"
        f"t.build_workbook_truth({str(WORKBOOK)!r})\n"
        "print(json.dumps(sorted(sys.modules)))\n"
    )
    completed = subprocess.run(
        [sys.executable, "-c", program], cwd=REPO_ROOT, capture_output=True, text=True, check=True
    )
    loaded = set(json.loads(completed.stdout))
    assert loaded & GENERATOR_MODULES == {"corvette_form_generator.runtime_contract"}


# ── Property 2: local helpers agree with the workbook contract ────────────────

CELL_VALUES = [
    None,
    True,
    False,
    "",
    "  padded  ",
    "True",
    "FALSE",
    "yes",
    "Y",
    "n",
    "1",
    "0",
    0,
    1,
    263,
    -5,
    10.0,
    10.5,
    0.0,
    "$1,695",
    "sec_perf_001",
    "Bright Red-Painted Calipers",
]


@pytest.mark.parametrize("value", CELL_VALUES)
def test_local_clean_matches_the_generator_representation(value) -> None:
    assert truth.clean(value) == generator_workbook.clean(value)


@pytest.mark.parametrize("value", CELL_VALUES)
def test_local_truthy_matches_the_generator_convention(value) -> None:
    assert truth.truthy(value) is generator_workbook.workbook_truthy(value)


# ── The snapshot says what it claims to say ───────────────────────────────────


def test_snapshot_carries_every_registered_sheet(snapshot: dict) -> None:
    """Registered sheet families resolve to sheets the workbook really has."""

    assert snapshot["registeredSheetsMissingFromWorkbook"] == []
    for sheet_name, family in snapshot["registeredSheetFamilies"].items():
        entry = snapshot["sheets"][sheet_name]
        assert entry["family"] == family
        assert entry["headers"], f"{sheet_name} has no headers"


def test_registered_sheets_are_discovered_not_hardcoded(snapshot: dict) -> None:
    """Per-model source sheets come from the workbook's own registration rows."""

    registered = snapshot["registeredSheetFamilies"]
    for row in snapshot["sheets"]["model_workbook_sources"]["rows"]:
        if not truth.truthy(row.get("active")):
            continue
        family = reg.SOURCE_ROLE_FAMILIES.get(row.get("source_role", ""))
        if not family:
            continue
        assert registered.get(row["sheet_name"]) == family


def test_dropping_a_registration_row_narrows_the_snapshot(tmp_path: Path) -> None:
    """Forced mutation behind the check above.

    Deactivating one registration must remove exactly that sheet from the
    snapshot. If the sheet list were hardcoded the snapshot would be unchanged,
    which is the defect this replaces.
    """

    copy = tmp_path / WORKBOOK.name
    copy.write_bytes(WORKBOOK.read_bytes())
    wb = load_workbook(copy)
    ws = wb["model_workbook_sources"]
    headers = [truth.clean(cell.value) for cell in ws[1]]
    model_column = headers.index("model_key") + 1
    role_column = headers.index("source_role") + 1
    active_column = headers.index("active") + 1
    sheet_column = headers.index("sheet_name") + 1
    dropped = ""
    for row in range(2, ws.max_row + 1):
        if (
            truth.clean(ws.cell(row, model_column).value) == "z06"
            and truth.clean(ws.cell(row, role_column).value) == "price_rules_sheet"
        ):
            dropped = truth.clean(ws.cell(row, sheet_column).value)
            ws.cell(row, active_column).value = False
            break
    wb.save(copy)
    wb.close()
    assert dropped, "fixture found no z06 price_rules_sheet registration to drop"

    mutated = truth.build_workbook_truth(copy)
    assert dropped not in mutated["registeredSheetFamilies"]
    assert "price_rules_sheet" not in mutated["models"]["z06"]["source_sheets"]


def test_rows_carry_registry_declared_identity(snapshot: dict) -> None:
    """Every keyed sheet exposes one identity per row, from its family key."""

    keyed = 0
    for sheet_name, entry in snapshot["sheets"].items():
        key_columns = entry["key_columns"]
        if not key_columns:
            continue
        keyed += 1
        assert len(entry["row_identities"]) == len(entry["rows"]), sheet_name
        assert key_columns == snapshot["familyKeyColumns"][entry["family"]]
    assert keyed > 20, "expected most registered sheets to declare a key"


def test_model_topology_matches_the_workbook_metadata_rows(snapshot: dict) -> None:
    """Variants come from model_variants joined to variant_master."""

    membership = {}
    for row in snapshot["sheets"]["model_variants"]["rows"]:
        if truth.truthy(row.get("active")):
            membership.setdefault(row["model_key"], set()).add(row["variant_id"])

    for model_key, model in snapshot["models"].items():
        assert {v["variant_id"] for v in model["variants"]} == membership.get(model_key, set())
        for variant in model["variants"]:
            assert variant["declared_in_variant_master"], f"{model_key} {variant['variant_id']}"
            assert variant["active_in_variant_master"], f"{model_key} {variant['variant_id']}"


def test_promotions_are_ordered_and_declare_one_default(snapshot: dict) -> None:
    promotions = snapshot["promotions"]
    keys = promotions["promoted_model_keys"]
    assert keys, "no model is promoted"
    assert len(set(keys)) == len(keys)
    orders = [int(row["display_order"] or 0) for row in promotions["rows"]]
    assert orders == sorted(orders)
    assert len(promotions["default_model_keys"]) == 1


def test_a_second_default_row_is_reported_not_resolved(tmp_path: Path) -> None:
    """Forced mutation: the snapshot reports both defaults rather than picking.

    Choosing a winner here would let the snapshot agree with whatever the
    generator did, including agreeing with it about a workbook defect.
    """

    copy = tmp_path / WORKBOOK.name
    copy.write_bytes(WORKBOOK.read_bytes())
    wb = load_workbook(copy)
    ws = wb["model_registry_promotion"]
    headers = [truth.clean(cell.value) for cell in ws[1]]
    model_column = headers.index("model_key") + 1
    default_column = headers.index("default_model") + 1
    for row in range(2, ws.max_row + 1):
        if truth.clean(ws.cell(row, model_column).value) == "z06":
            ws.cell(row, default_column).value = True
            break
    wb.save(copy)
    wb.close()

    mutated = truth.build_workbook_truth(copy)
    assert sorted(mutated["promotions"]["default_model_keys"]) == ["stingray", "z06"]


def test_asset_precedence_prefers_the_exact_model_row(snapshot: dict) -> None:
    """Wildcard rows are shared option media; an exact model row wins."""

    assert snapshot["assetConflicts"] == []
    asset_rows = snapshot["sheets"]["asset_map"]["rows"]
    wildcard_targets = {
        f"{row['target_type']}::{row['target_id']}"
        for row in asset_rows
        if row.get("model_key") == "*" and truth.truthy(row.get("active")) and row.get("image_url")
    }
    assert wildcard_targets, "fixture expects the workbook to use wildcard asset rows"

    for model_key, resolved in snapshot["assets"].items():
        exact = {
            f"{row['target_type']}::{row['target_id']}": row
            for row in asset_rows
            if row.get("model_key") == model_key
            and truth.truthy(row.get("active"))
            and row.get("image_url")
        }
        for target, row in exact.items():
            assert resolved[target]["image_url"] == row["image_url"], f"{model_key} {target}"
        for target in wildcard_targets - set(exact):
            assert target in resolved, f"{model_key} lost shared media for {target}"


def test_the_tracked_workbook_has_no_topology_conflicts(snapshot: dict) -> None:
    """Promotion and variant identity are assumed unique; check that they are."""

    assert snapshot["topologyConflicts"] == []


def test_duplicate_promotion_and_variant_rows_are_reported(tmp_path: Path) -> None:
    """Forced mutation: a second row for a key the topology indexes uniquely.

    A duplicate promotion is exactly as unadjudicable as a duplicate
    `asset_map` row, so it is reported rather than decided by row order.
    """

    copy = tmp_path / "topology.xlsx"
    wb = Workbook()
    sources = wb.active
    sources.title = "model_workbook_sources"
    sources.append(["model_key", "source_role", "sheet_name", "active"])

    master = wb.create_sheet("model_master")
    master.append(["model_key", "registry_key", "model_label", "active"])
    master.append(["stingray", "stingray", "Stingray", True])

    variants = wb.create_sheet("variant_master")
    variants.append(["variant_id", "body_style", "trim_level", "base_price", "active"])
    variants.append(["1lt_c07", "coupe", "1lt", "68300", True])
    variants.append(["1lt_c07", "convertible", "1lt", "75800", True])

    promotion = wb.create_sheet("model_registry_promotion")
    promotion.append(["model_key", "registry_key", "artifact_path", "promoted_to_runtime", "active"])
    promotion.append(["stingray", "stingray", "form-output/a.json", True, True])
    promotion.append(["stingray", "stingray", "form-output/b.json", True, True])
    wb.save(copy)
    wb.close()

    built = truth.build_workbook_truth(copy)
    assert built["topologyConflicts"] == [
        {"sheet": "variant_master", "variant_id": "1lt_c07"},
        {"sheet": "model_registry_promotion", "model_key": "stingray"},
    ]
    # First row wins the resolved view, which is safe to look at only because
    # the conflict list above is asserted empty for the tracked workbook.
    assert built["models"]["stingray"]["promotion"]["artifact_path"] == "form-output/a.json"


def test_conflicting_asset_rows_are_reported(tmp_path: Path) -> None:
    """Forced mutation: two active rows for one model and target conflict.

    Nothing in the workbook can adjudicate this, so it is surfaced instead of
    being resolved by row order.
    """

    copy = tmp_path / "conflict.xlsx"
    wb = Workbook()
    sources = wb.active
    sources.title = "model_workbook_sources"
    sources.append(["model_key", "source_role", "sheet_name", "active"])

    master = wb.create_sheet("model_master")
    master.append(["model_key", "registry_key", "model_label", "active"])
    master.append(["stingray", "stingray", "Stingray", True])

    assets = wb.create_sheet("asset_map")
    assets.append(["model_key", "target_type", "target_id", "image_url", "active"])
    assets.append(["stingray", "option", "opt_j6f_001", "https://example.test/a.webp", True])
    assets.append(["stingray", "option", "opt_j6f_001", "https://example.test/b.webp", True])
    wb.save(copy)
    wb.close()

    built = truth.build_workbook_truth(copy)
    assert built["assetConflicts"] == [{"model_key": "stingray", "target": "option::opt_j6f_001"}]
    # First row wins the resolved view; the conflict is what makes that safe to
    # look at, because a gate asserts the conflict list is empty.
    assert built["assets"]["stingray"]["option::opt_j6f_001"]["image_url"] == "https://example.test/a.webp"


def test_a_missing_registered_sheet_is_reported(tmp_path: Path) -> None:
    """A registration pointing at a sheet the workbook lacks is not silent."""

    copy = tmp_path / "missing.xlsx"
    wb = Workbook()
    sources = wb.active
    sources.title = "model_workbook_sources"
    sources.append(["model_key", "source_role", "sheet_name", "active"])
    sources.append(["stingray", "source_option_sheet", "stingray_options", True])
    wb.save(copy)
    wb.close()

    built = truth.build_workbook_truth(copy)
    assert "stingray_options" in built["registeredSheetsMissingFromWorkbook"]
    assert "stingray_options" not in built["sheets"]


def test_snapshot_binds_itself_to_the_workbook_it_read(snapshot: dict) -> None:
    """The snapshot records which workbook produced it, by content."""

    import hashlib

    assert snapshot["workbook"]["name"] == WORKBOOK.name
    assert snapshot["workbook"]["sha256"] == hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()


def test_cli_writes_a_snapshot_to_an_explicit_path(tmp_path: Path) -> None:
    """§6.2: node receives the snapshot through an explicit temporary path."""

    target = tmp_path / "nested" / "workbook-truth.json"
    assert truth.main(["--workbook", str(WORKBOOK), "--out", str(target)]) == 0
    written = json.loads(target.read_text(encoding="utf-8"))
    assert written["schemaVersion"] == truth.SCHEMA_VERSION
