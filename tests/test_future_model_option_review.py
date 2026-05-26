#!/usr/bin/env python3
"""Tests for simplified future_model_option_review generation."""

from __future__ import annotations

import importlib.util
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import FUTURE_MODEL_SOURCE_REVIEW_HEADERS  # noqa: E402

MODULE_PATH = ROOT / "scripts" / "create_future_model_option_review.py"
spec = importlib.util.spec_from_file_location("create_future_model_option_review", MODULE_PATH)
assert spec and spec.loader
option_review = importlib.util.module_from_spec(spec)
spec.loader.exec_module(option_review)


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def source_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {header: "" for header in FUTURE_MODEL_SOURCE_REVIEW_HEADERS}
    row.update(
        {
            "model_key": "z06",
            "raw_source_sheets": "z06_intextmec_raw",
            "raw_source_spans": "z06_intextmec_raw:10-10",
            "source_orderable_rpo": "ABC",
            "source_ref_rpo": "",
            "source_primary_rpo": "ABC",
            "source_option_description": "Visible carbon widget",
            "source_disclosure_raw": "Disclosure text",
            "candidate_option_id": "opt_abc_001",
            "approved_option_id": "opt_abc_001",
            "approved_option_name": "Visible Carbon Widget",
            "approved_section_id": "sec_test_001",
            "approved_selectable": "True",
            "approved_display_order": "10",
            "copy_from_model_key": "grand_sport",
            "copy_from_option_id": "opt_abc_001",
            "review_status": "approved",
            "active": "True",
            "raw_status_1lz_h07": "A",
            "raw_status_2lz_h07": "S",
            "status_1lz_h07": "available",
            "status_2lz_h07": "standard",
        }
    )
    row.update(overrides)
    return row


class FutureModelOptionReviewTests(unittest.TestCase):
    def test_simplified_review_rows_use_source_rpo_and_compact_status_summaries(self) -> None:
        wb = Workbook()
        del wb[wb.sheetnames[0]]
        append_sheet(
            wb,
            "future_model_source_review",
            list(FUTURE_MODEL_SOURCE_REVIEW_HEADERS),
            [
                source_row(),
                source_row(
                    raw_source_spans="z06_standard_raw:15-15",
                    raw_source_sheets="z06_standard_raw",
                    source_orderable_rpo="",
                    source_ref_rpo="REF",
                    source_primary_rpo="REF",
                    candidate_option_id="opt_ref_001",
                    approved_option_id="",
                    review_status="needs_section_review",
                    active="False",
                ),
            ],
        )

        rows = option_review.build_option_review_rows(wb)

        self.assertEqual(len(rows), 2)
        self.assertNotIn("source_group", option_review.OPTION_REVIEW_HEADERS)
        self.assertNotIn("approved_rpo", option_review.OPTION_REVIEW_HEADERS)
        first = rows[0]
        self.assertEqual(first["raw_source_sheet"], "z06_intextmec_raw")
        self.assertEqual(first["orderable_rpo"], "ABC")
        self.assertEqual(first["ref_only_rpo"], "")
        self.assertEqual(first["source_rpo"], "ABC")
        self.assertEqual(first["suggested_copy_from"], "grand_sport:opt_abc_001")
        self.assertEqual(first["final_option_id"], "opt_abc_001")
        self.assertIn("1lz_h07=A", first["raw_status_summary"])
        self.assertIn("2lz_h07=standard", first["normalized_status_summary"])
        second = rows[1]
        self.assertEqual(second["orderable_rpo"], "")
        self.assertEqual(second["ref_only_rpo"], "REF")
        self.assertEqual(second["source_rpo"], "REF")
        self.assertEqual(second["suggested_option_id"], "opt_ref_001")

    def test_existing_human_decisions_are_preserved_on_rerun(self) -> None:
        wb = Workbook()
        del wb[wb.sheetnames[0]]
        append_sheet(wb, "future_model_source_review", list(FUTURE_MODEL_SOURCE_REVIEW_HEADERS), [source_row(approved_option_id="")])
        append_sheet(
            wb,
            "future_model_option_review",
            option_review.OPTION_REVIEW_HEADERS,
            [
                {
                    "model_key": "z06",
                    "raw_source_sheet": "z06_intextmec_raw",
                    "raw_source_span": "z06_intextmec_raw:10-10",
                    "source_rpo": "ABC",
                    "final_option_id": "opt_manual_001",
                    "final_section_id": "sec_manual_001",
                    "review_status": "approved",
                    "active": "True",
                    "notes": "manual decision",
                }
            ],
        )

        rows = option_review.build_option_review_rows(wb)

        self.assertEqual(rows[0]["final_option_id"], "opt_manual_001")
        self.assertEqual(rows[0]["final_section_id"], "sec_manual_001")
        self.assertEqual(rows[0]["notes"], "manual decision")


if __name__ == "__main__":
    unittest.main()
