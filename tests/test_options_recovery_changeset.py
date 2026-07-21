from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from test_options_recovery_projection import _projection_fixture


def test_header_and_rows_streams_read_only_sheet_with_physical_row_numbers() -> None:
    from corvette_form_generator.ingest.options_recovery_changeset import _headers_and_rows

    class Cell:
        def __init__(self, value: object) -> None:
            self.value = value

    class ReadOnlySheet:
        max_row = 4

        def __getitem__(self, row: int) -> list[Cell]:
            assert row == 1
            return [Cell("option_id"), Cell("rpo")]

        def cell(self, row: int, column: int) -> None:
            raise AssertionError(f"random read attempted at {row}/{column}")

        def iter_rows(self, *, values_only: bool):
            assert values_only is True
            return iter([
                ("option_id", "rpo"),
                ("opt_001", "A1"),
                (None, None),
                ("opt_002", "A2"),
            ])

    class Workbook:
        sheetnames = ["options"]

        def __getitem__(self, sheet: str) -> ReadOnlySheet:
            assert sheet == "options"
            return ReadOnlySheet()

    assert _headers_and_rows(Workbook(), "options") == (
        ["option_id", "rpo"],
        {
            2: {"option_id": "opt_001", "rpo": "A1"},
            4: {"option_id": "opt_002", "rpo": "A2"},
        },
    )


def _checkpoint_fixture(tmp_path: Path) -> tuple[dict[str, Path], Path]:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
        record_checkpoint_1_decisions,
    )

    paths = _projection_fixture(tmp_path)
    workbook_path = paths["workbook_path"]
    workbook = load_workbook(workbook_path)
    workbook["model_workbook_sources"].append(
        ["grand_sport_x", "status_sheet", "grand_sport_x_ovs", False, "unpromoted"]
    )
    workbook["grand_sport_x_ovs"].append(["opt_n26_001", "1lt_g07", "available"])
    workbook.save(workbook_path)
    workbook.close()
    report_dir = tmp_path / "reports"
    generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    approval, packet = generate_checkpoint_1_packet(
        report_dir, workbook_path, reviewer="Sean", reviewed_at="2026-07-20T16:00:00-04:00"
    )
    decisions = []
    for group in packet["exceptionGroups"]:
        action = "accept"
        instruction = {"decisionGroupId": group["decisionGroupId"], "action": action}
        if group["identityKey"] == "N26" and group["lane"] == "full_review":
            instruction["action"] = "delete_option_and_owned_references"
        elif group["identityKey"] == "N26" and group["lane"] == "display_order":
            instruction["action"] = "not_applicable_due_to_delete"
        elif group["models"] == ["zr1", "zr1x"] and group["lane"] == "display_order":
            instruction.update(
                action="override_by_model",
                overrideByModel={
                    "zr1": {"display_order": 51},
                    "zr1x": {"display_order": 61},
                },
            )
        decisions.append(instruction)

    record_checkpoint_1_decisions(
        report_dir,
        workbook_path,
        decisions=decisions,
        bulk_overrides=[
            {
                "model": "grand_sport_x",
                "rpo": "",
                "lane": "price",
                "override": {"price": 77},
            }
        ],
        reviewer="Sean",
        reviewed_at="2026-07-20T17:00:00-04:00",
    )
    return paths, report_dir


def test_emits_deterministic_bounded_changeset_without_mutating_workbook(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_changeset import (
        emit_options_recovery_changeset,
    )

    paths, report_dir = _checkpoint_fixture(tmp_path)
    workbook_before = paths["workbook_path"].read_bytes()

    first = emit_options_recovery_changeset(paths["workbook_path"], report_dir)
    second = emit_options_recovery_changeset(paths["workbook_path"], report_dir)

    assert first == second
    assert paths["workbook_path"].read_bytes() == workbook_before
    assert first["targets"] == ["grand_sport_x", "zr1", "zr1x"]
    assert first["sheetCreates"] == []
    assert {change["sheet"] for change in first["rowChanges"]} <= {
        "grand_sport_x_options",
        "zr1_options",
        "zr1x_options",
        "grand_sport_x_ovs",
        "default_selection_rules",
    }
    assert all(change["provenance"] for change in first["rowChanges"])

    deleted = {
        (change["action"], change["sheet"], tuple(sorted(change["key"].items())))
        for change in first["rowChanges"]
        if change["action"] == "delete"
    }
    assert ("delete", "grand_sport_x_options", (("option_id", "opt_std_aaaaaaaaaaaaaaaa"),)) in deleted
    assert ("delete", "grand_sport_x_ovs", (("option_id", "opt_std_aaaaaaaaaaaaaaaa"), ("variant_id", "1lt_g07"))) in deleted

    added_option = next(
        change
        for change in first["rowChanges"]
        if change["action"] == "add"
        and change["sheet"] == "grand_sport_x_options"
        and change["key"] == {"option_id": "opt_001"}
    )
    assert added_option["fields"]["option_id"] == {"before": None, "after": "opt_001"}
    assert added_option["fields"]["detail_raw"]["after"] == "Air filtration system, with pollen filter"
    assert added_option["fields"]["price"]["after"] == 77
    assert not any(
        change["action"] == "update" and "option_id" in change["fields"]
        for change in first["rowChanges"]
    )
    assert any(
        change["action"] == "update"
        and change["sheet"] == "default_selection_rules"
        and change["fields"] == {
            "target_option_id": {
                "before": "opt_std_aaaaaaaaaaaaaaaa",
                "after": "opt_001",
            }
        }
        for change in first["rowChanges"]
    )
    assert any(
        change["action"] == "update"
        and change["sheet"] == "zr1_options"
        and change["fields"].get("display_order", {}).get("after") == 51
        for change in first["rowChanges"]
    )
    assert sum(
        change["action"] == "delete"
        and change["sheet"] == "grand_sport_x_options"
        and change["key"] == {"option_id": "opt_n26_001"}
        for change in first["rowChanges"]
    ) == 1
    assert sum(
        change["action"] == "delete"
        and change["sheet"] == "grand_sport_x_ovs"
        and change["key"] == {"option_id": "opt_n26_001", "variant_id": "1lt_g07"}
        for change in first["rowChanges"]
    ) == 1


def test_emits_reported_nonresidual_option_deletions_exactly_once(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_changeset import (
        emit_options_recovery_changeset,
    )

    paths, report_dir = _checkpoint_fixture(tmp_path)
    changeset = emit_options_recovery_changeset(paths["workbook_path"], report_dir)
    deleted_options = {
        (change["sheet"], change["key"]["option_id"])
        for change in changeset["rowChanges"]
        if change["action"] == "delete" and change["family"] == "options"
    }

    assert {
        ("zr1_options", "opt_feh_001"),
        ("zr1_options", "opt_tu7_001"),
        ("zr1x_options", "opt_tu7_001"),
    } <= deleted_options
    assert sum(
        change["action"] == "delete"
        and change["family"] == "options"
        and change["sheet"] == "zr1_options"
        and change["key"] == {"option_id": "opt_feh_001"}
        for change in changeset["rowChanges"]
    ) == 1


def test_refuses_incomplete_review_coverage(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_changeset import (
        emit_options_recovery_changeset,
    )
    from corvette_form_generator.ingest.options_recovery_projection import _canonical_sha

    paths, report_dir = _checkpoint_fixture(tmp_path)
    approval_path = report_dir / "checkpoint-1-bulk-approval.json"
    approval = json.loads(approval_path.read_text(encoding="utf-8"))
    approval["approvedReviewIds"] = approval["approvedReviewIds"][:-1]
    approval["approvedReviewCount"] -= 1
    approval["approvalFingerprint"] = _canonical_sha(
        {key: value for key, value in approval.items() if key != "approvalFingerprint"}
    )
    approval_path.write_text(json.dumps(approval, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    packet_path = report_dir / "checkpoint-1-exception-review.json"
    packet = json.loads(packet_path.read_text(encoding="utf-8"))
    packet["bulkApprovalFingerprint"] = approval["approvalFingerprint"]
    packet["packetFingerprint"] = _canonical_sha(
        {key: value for key, value in packet.items() if key != "packetFingerprint"}
    )
    packet_path.write_text(json.dumps(packet, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    decision_path = report_dir / "checkpoint-1-exception-decisions.json"
    decision = json.loads(decision_path.read_text(encoding="utf-8"))
    decision["bulkApprovalFingerprint"] = approval["approvalFingerprint"]
    decision["sourcePacketFingerprint"] = packet["packetFingerprint"]
    decision["decisionArtifactFingerprint"] = _canonical_sha(
        {key: value for key, value in decision.items() if key != "decisionArtifactFingerprint"}
    )
    decision_path.write_text(json.dumps(decision, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    pending_path = report_dir / "checkpoint-1-pending-review.json"
    pending = json.loads(pending_path.read_text(encoding="utf-8"))
    pending["bulkApprovalFingerprint"] = approval["approvalFingerprint"]
    pending["sourcePacketFingerprint"] = packet["packetFingerprint"]
    pending["decisionArtifactFingerprint"] = decision["decisionArtifactFingerprint"]
    pending["pendingPacketFingerprint"] = _canonical_sha(
        {key: value for key, value in pending.items() if key != "pendingPacketFingerprint"}
    )
    pending_path.write_text(json.dumps(pending, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    with pytest.raises(ValueError, match="coverage"):
        emit_options_recovery_changeset(paths["workbook_path"], report_dir)


def test_refuses_re_signed_artifact_with_tampered_nested_decision(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_changeset import (
        emit_options_recovery_changeset,
    )
    from corvette_form_generator.ingest.options_recovery_projection import _canonical_sha

    paths, report_dir = _checkpoint_fixture(tmp_path)
    decision_path = report_dir / "checkpoint-1-exception-decisions.json"
    decision = json.loads(decision_path.read_text(encoding="utf-8"))
    decision["decisions"][0]["decisionFingerprint"] = "0" * 64
    decision["decisionArtifactFingerprint"] = _canonical_sha(
        {key: value for key, value in decision.items() if key != "decisionArtifactFingerprint"}
    )
    decision_path.write_text(json.dumps(decision, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    pending_path = report_dir / "checkpoint-1-pending-review.json"
    pending = json.loads(pending_path.read_text(encoding="utf-8"))
    pending["decisionArtifactFingerprint"] = decision["decisionArtifactFingerprint"]
    pending["pendingPacketFingerprint"] = _canonical_sha(
        {key: value for key, value in pending.items() if key != "pendingPacketFingerprint"}
    )
    pending_path.write_text(json.dumps(pending, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    with pytest.raises(ValueError, match="decision fingerprint"):
        emit_options_recovery_changeset(paths["workbook_path"], report_dir)


def test_cli_refuses_immutable_output_overwrite(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_changeset import main

    paths, report_dir = _checkpoint_fixture(tmp_path)
    output = tmp_path / "repair.json"
    args = ["--workbook", str(paths["workbook_path"]), "--report-dir", str(report_dir), "--output", str(output)]

    assert main(args) == 0
    assert main(args) == 0
    output.write_text("{}\n", encoding="utf-8")
    with pytest.raises(ValueError, match="different content"):
        main(args)
