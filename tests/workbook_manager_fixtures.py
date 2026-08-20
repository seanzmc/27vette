"""Shared Workbook Manager test fixtures for Checkpoint 5.

One process-wide verified real-workbook projection/candidate is built only when
a Layer 3 owner asks for it. Negative cases stay on compact workbooks. Clones
are byte copies; the source hash is asserted after every consumer.
"""

from __future__ import annotations

import atexit
import hashlib
import shutil
import sqlite3
import sys
import tempfile
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "workbook-manager" / "backend"
SCRIPTS = ROOT / "scripts"
for path in (BACKEND, SCRIPTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from app import config as configmod  # noqa: E402
from app import db as dbmod  # noqa: E402
from app import importer  # noqa: E402
from app import sync as syncmod  # noqa: E402
from corvette_form_generator.schema_validation import REQUIRED_SHEETS  # noqa: E402
from corvette_form_generator.workbook_domain.registry import (  # noqa: E402
    GLOBAL_SHEET_FAMILIES,
    READONLY_SHEET_META,
    WRITABLE_COLUMNS,
)

CANONICAL_WORKBOOK = ROOT / "stingray_master.xlsx"
MISSING_REFERENCE_TARGET = "missing_target_for_checkpoint_5"

_INTERIOR_HEADERS = WRITABLE_COLUMNS["interiors"]
_COMPACT_SHEET_HEADERS: dict[str, tuple[str, ...]] = {
    family: WRITABLE_COLUMNS[family] for family in GLOBAL_SHEET_FAMILIES.values()
}
_COMPACT_SHEET_HEADERS["sections"] = READONLY_SHEET_META["sections"]["columns"]
_COMPACT_SHEET_HEADERS["interiors"] = _INTERIOR_HEADERS

_COMBINED_SCHEMA_TABLES = (
    "pending_changes",
    "change_history",
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clone_sqlite(source: Path, destination: Path) -> Path:
    """Copy one checkpointed SQLite file without opening the source."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return destination


def clone_workbook(source: Path, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return destination


def assert_fixture_unmutated(path: Path, expected_sha256: str) -> None:
    actual = sha256_file(path)
    if actual != expected_sha256:
        raise AssertionError(
            f"shared fixture mutated: {path} {actual} != {expected_sha256}"
        )


def _append(workbook: Workbook, name: str, headers: tuple[str, ...], rows: list[list[object]]) -> None:
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
    sheet = workbook.create_sheet(name)
    sheet.append(list(headers))
    for row in rows:
        padded = list(row) + [""] * (len(headers) - len(row))
        sheet.append(padded[: len(headers)])


def _empty_required_sheets(workbook: Workbook, *, omit: set[str] | None = None) -> None:
    skip = set(omit or ())
    for sheet in REQUIRED_SHEETS:
        if sheet in skip or sheet in workbook.sheetnames:
            continue
        if sheet in {"lt_interiors", "LZ_Interiors"}:
            _append(workbook, sheet, _INTERIOR_HEADERS, [])
        elif sheet == "section_master":
            _append(workbook, sheet, _COMPACT_SHEET_HEADERS["sections"], [])
        elif sheet == "PriceRef":
            _append(workbook, sheet, ("code", "label", "value"), [])
        else:
            family = GLOBAL_SHEET_FAMILIES[sheet]
            _append(workbook, sheet, _COMPACT_SHEET_HEADERS[family], [])


def write_compact_missing_identifier_workbook(path: Path) -> Path:
    """Compact workbook whose asset_map row is missing identifying keys."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    active = workbook.active
    if active is not None:
        workbook.remove(active)
    _empty_required_sheets(workbook)
    _append(
        workbook,
        "asset_map",
        WRITABLE_COLUMNS["asset_map"],
        [[
            "stingray",
            "",
            "",
            "https://example.invalid/unresolved-checkpoint-5.png",
            "",
            "cover",
            "center",
            "",
            "",
            "",
            False,
            "",
        ]],
    )
    workbook.save(path)
    workbook.close()
    return path


def write_compact_unresolved_reference_workbook(path: Path) -> Path:
    """Compact workbook with one rule_mapping row pointing at a missing option."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    active = workbook.active
    if active is not None:
        workbook.remove(active)
    _empty_required_sheets(workbook)
    _append(
        workbook,
        "model_master",
        WRITABLE_COLUMNS["model_master"],
        [[
            "stingray",
            "stingray",
            "Stingray",
            2027,
            "compact",
            "stingray",
            1,
            True,
            True,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]],
    )
    _append(
        workbook,
        "model_workbook_sources",
        WRITABLE_COLUMNS["model_workbook_sources"],
        [[
            "stingray",
            "source_option_sheet",
            "stingray_options",
            True,
            "",
        ], [
            "stingray",
            "rule_mapping_sheet",
            "rule_mapping",
            True,
            "",
        ]],
    )
    _append(
        workbook,
        "stingray_options",
        WRITABLE_COLUMNS["options"],
        [[
            "opt_src_001",
            "SRC",
            0,
            "Source option",
            "",
            "",
            "sec_test_001",
            True,
            1,
            True,
            "",
        ]],
    )
    _append(
        workbook,
        "rule_mapping",
        WRITABLE_COLUMNS["rule_mapping"],
        [[
            "rule_cp5_001",
            "opt_src_001",
            "requires",
            MISSING_REFERENCE_TARGET,
            "",
            "",
            "",
            "",
        ]],
    )
    workbook.save(path)
    workbook.close()
    return path


def write_compact_missing_sheet_workbook(path: Path) -> Path:
    """Compact workbook that omits a required sheet."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    active = workbook.active
    if active is not None:
        workbook.remove(active)
    _empty_required_sheets(workbook, omit={"model_master"})
    workbook.save(path)
    workbook.close()
    return path


def attach_combined_schema(connection: sqlite3.Connection) -> None:
    """Add legacy combined staging tables onto a cloned projection."""
    names = {
        row["name"]
        for row in connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )
    }
    if set(_COMBINED_SCHEMA_TABLES) <= names:
        return
    dbmod.init_schema(connection)


class VerifiedManagerFixture:
    def __init__(
        self,
        *,
        root: Path,
        workbook: Path,
        projection: Path,
        import_report: dict,
        promotion_report: dict,
        workbook_sha256: str,
        projection_sha256: str,
    ) -> None:
        self.root = root
        self.workbook = workbook
        self.projection = projection
        self.import_report = import_report
        self.promotion_report = promotion_report
        self.workbook_sha256 = workbook_sha256
        self.projection_sha256 = projection_sha256
        self._unchanged_export: Path | None = None
        self._export_sha256 = ""
        self._export_result: dict | None = None

    def clone_projection(self, destination: Path) -> Path:
        clone_sqlite(self.projection, destination)
        return destination

    def clone_workbook(self, destination: Path) -> Path:
        clone_workbook(self.workbook, destination)
        return destination

    def unchanged_export_result(self) -> dict:
        if self._export_result is None:
            self._build_unchanged_export()
        assert self._export_result is not None
        return dict(self._export_result)

    def clone_unchanged_export(self, destination: Path) -> Path:
        clone_workbook(Path(self.unchanged_export_result()["path"]), destination)
        return destination

    def imported_report(self) -> dict:
        return deepcopy(self.import_report)

    def assert_unmutated(self) -> None:
        assert_fixture_unmutated(self.workbook, self.workbook_sha256)
        assert_fixture_unmutated(self.projection, self.projection_sha256)
        if self._unchanged_export is not None:
            assert_fixture_unmutated(self._unchanged_export, self._export_sha256)

    def _build_unchanged_export(self) -> None:
        export_dir = self.root / "unchanged-export"
        export_dir.mkdir(exist_ok=True)
        connection = dbmod.connect(self.projection)
        previous = (configmod.VAR_DIR, configmod.EXPORT_DIR, configmod.DB_BACKUP_DIR)
        configmod.VAR_DIR = self.root / "var"
        configmod.EXPORT_DIR = export_dir
        configmod.DB_BACKUP_DIR = configmod.VAR_DIR / "db-backups"
        try:
            exported = syncmod.export_comparison_workbook(connection, self.workbook)
        finally:
            configmod.VAR_DIR, configmod.EXPORT_DIR, configmod.DB_BACKUP_DIR = previous
            connection.close()
        if not exported.get("ok"):
            raise AssertionError(
                f"verified fixture unchanged export failed: {exported}"
            )
        self._unchanged_export = Path(exported["path"])
        self._export_sha256 = sha256_file(self._unchanged_export)
        self._export_result = exported


_VERIFIED: VerifiedManagerFixture | None = None


def _checkpoint(path: Path) -> None:
    connection = dbmod.connect(path)
    try:
        connection.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    finally:
        connection.close()
    for sidecar in (Path(f"{path}-wal"), Path(f"{path}-shm")):
        sidecar.unlink(missing_ok=True)


def _build_verified_fixture() -> VerifiedManagerFixture:
    root = Path(tempfile.mkdtemp(prefix="wbm-verified-fixture-"))
    workbook = root / "source.xlsx"
    shutil.copy2(CANONICAL_WORKBOOK, workbook)
    projection = root / "workbook_projection.sqlite3"
    prior = dbmod.connect(projection)
    try:
        dbmod.init_projection_schema(prior)
        prior.execute(
            "INSERT INTO meta(key, value) VALUES('sentinel', 'prior-projection')"
        )
        prior.commit()
        prior.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    finally:
        prior.close()

    promotion = importer.promote_verified_projection(workbook, projection)
    if not promotion.get("promoted"):
        shutil.rmtree(root, ignore_errors=True)
        raise AssertionError(f"verified fixture promotion failed: {promotion}")
    _checkpoint(projection)

    imported = dbmod.connect(projection)
    try:
        report = importer.latest_report(imported)
        if report is None:
            raise AssertionError("verified fixture is missing its import report")
        errors = [
            issue for issue in report["issues"] if issue["severity"] == "error"
        ]
        if errors:
            raise AssertionError(f"verified fixture import has errors: {errors}")
    finally:
        imported.close()

    fixture = VerifiedManagerFixture(
        root=root,
        workbook=workbook,
        projection=projection,
        import_report=report,
        promotion_report=promotion,
        workbook_sha256=sha256_file(workbook),
        projection_sha256=sha256_file(projection),
    )
    atexit.register(_cleanup_verified_fixture)
    return fixture


def _cleanup_verified_fixture() -> None:
    global _VERIFIED
    if _VERIFIED is None:
        return
    try:
        _VERIFIED.assert_unmutated()
    finally:
        shutil.rmtree(_VERIFIED.root, ignore_errors=True)
        _VERIFIED = None


def verified_manager_fixture() -> VerifiedManagerFixture:
    """Return the process-wide verified real-workbook projection/candidate."""
    global _VERIFIED
    if _VERIFIED is None:
        _VERIFIED = _build_verified_fixture()
    _VERIFIED.assert_unmutated()
    return _VERIFIED


def clone_combined_projection(destination: Path) -> tuple[Path, dict]:
    """Clone the verified projection and attach combined staging tables."""
    fixture = verified_manager_fixture()
    clone_sqlite(fixture.projection, destination)
    connection = dbmod.connect(destination)
    try:
        attach_combined_schema(connection)
        connection.commit()
        connection.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    finally:
        connection.close()
    for sidecar in (Path(f"{destination}-wal"), Path(f"{destination}-shm")):
        sidecar.unlink(missing_ok=True)
    return destination, fixture.imported_report()
