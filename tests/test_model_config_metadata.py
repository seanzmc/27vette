#!/usr/bin/env python3
"""Focused tests for Phase 7 workbook-backed model configuration metadata."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "stingray_master.xlsx"
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_generation import generate_model_artifacts  # noqa: E402
from corvette_form_generator.model_configs import (  # noqa: E402
    REQUIRED_GENERATION_SOURCE_ROLES,
    base_model_config,
    discover_generation_model_configs,
)
from corvette_form_generator.runtime_metadata import (  # noqa: E402
    load_model_config_overrides,
    load_runtime_steps,
)

BASE_STINGRAY_CONFIG = base_model_config("stingray")


def workbook_with_model_metadata(
    *,
    model_rows: list[dict[str, object]] | None = None,
    source_rows: list[dict[str, object]] | None = None,
    variant_rows: list[dict[str, object]] | None = None,
) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "model_master",
        [
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "notes",
        ],
        model_rows or [],
    )
    append_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        source_rows or [],
    )
    append_sheet(
        wb,
        "model_variants",
        ["model_key", "variant_id", "display_order", "active", "notes"],
        variant_rows or [],
    )
    return wb


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def stingray_model_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "model_key": "stingray",
        "registry_key": "stingray",
        "model_label": "Workbook Stingray",
        "model_year": "2027",
        "dataset_name": "Workbook dataset",
        "expected_variant_count": 2,
        "active": True,
    }
    row.update(overrides)
    return row


def required_source_rows(model_key: str = "stingray", *, active: bool = True) -> list[dict[str, object]]:
    return [
        {
            "model_key": model_key,
            "source_role": role,
            "sheet_name": f"{model_key}_{role}",
            "active": active,
        }
        for role in REQUIRED_GENERATION_SOURCE_ROLES
    ]


def discover_temp_configs(wb: Workbook):
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "model-discovery.xlsx"
        wb.save(path)
        return discover_generation_model_configs(path)


class ModelConfigMetadataTests(unittest.TestCase):
    def test_generation_discovery_binds_configs_to_explicit_paths(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=required_source_rows("stingray"),
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
            ],
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            target_root = Path(tmpdir) / "candidate"
            workbook_path = Path(tmpdir) / "model-discovery.xlsx"
            output_dir = target_root / "generated"
            app_dir = target_root / "browser"
            wb.save(workbook_path)

            config = discover_generation_model_configs(
                workbook_path,
                root=target_root,
                output_dir=output_dir,
                app_dir=app_dir,
            )["stingray"]

        self.assertEqual(config.workbook_path, workbook_path)
        self.assertEqual(config.root, target_root)
        self.assertEqual(config.output_dir, output_dir)
        self.assertEqual(config.app_dir, app_dir)

    def test_header_only_metadata_falls_back_to_constants(self) -> None:
        wb = workbook_with_model_metadata()
        resolved = load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)
        self.assertEqual(resolved, BASE_STINGRAY_CONFIG)

    def test_workbook_metadata_overrides_config_fields(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=[
                {
                    "model_key": "stingray",
                    "source_role": "source_option_sheet",
                    "sheet_name": "workbook_options",
                    "active": True,
                },
                {
                    "model_key": "stingray",
                    "source_role": "status_sheet",
                    "sheet_name": "workbook_status",
                    "active": True,
                },
            ],
            variant_rows=[
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
            ],
        )
        resolved = load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)
        self.assertEqual(resolved.model_label, "Workbook Stingray")
        self.assertEqual(resolved.dataset_name, "Workbook dataset")
        self.assertEqual(resolved.source_option_sheet, "workbook_options")
        self.assertEqual(resolved.status_sheet, "workbook_status")
        self.assertEqual(resolved.variant_ids, ("a_variant", "z_variant"))
        self.assertEqual(resolved.expected_variant_count, 2)
        self.assertEqual(BASE_STINGRAY_CONFIG.source_option_sheet, "stingray_options")

    def test_workbook_metadata_overrides_interior_source_sheet(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=[
                {
                    "model_key": "stingray",
                    "source_role": "interior_source_sheet",
                    "sheet_name": "LZ_Interiors",
                    "active": True,
                }
            ],
            variant_rows=[
                {"model_key": "stingray", "variant_id": "1lz_h07", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "3lz_h07", "display_order": 2, "active": True},
            ],
        )

        resolved = load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)

        self.assertEqual(resolved.interior_source_sheet, "LZ_Interiors")
        self.assertEqual(BASE_STINGRAY_CONFIG.interior_source_sheet, "lt_interiors")

    def test_unknown_source_role_fails_fast(self) -> None:
        wb = workbook_with_model_metadata(
            source_rows=[
                {
                    "model_key": "stingray",
                    "source_role": "surprise_sheet",
                    "sheet_name": "surprise",
                    "active": True,
                }
            ]
        )
        with self.assertRaisesRegex(ValueError, "Unknown model_workbook_sources roles"):
            load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)

    def test_duplicate_source_role_fails_fast(self) -> None:
        wb = workbook_with_model_metadata(
            source_rows=[
                {
                    "model_key": "stingray",
                    "source_role": "source_option_sheet",
                    "sheet_name": "first",
                    "active": True,
                },
                {
                    "model_key": "stingray",
                    "source_role": "source_option_sheet",
                    "sheet_name": "second",
                    "active": True,
                },
            ]
        )
        with self.assertRaisesRegex(ValueError, "Duplicate active model_workbook_sources roles"):
            load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)

    def test_variant_count_mismatch_fails_fast(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row(expected_variant_count=2)],
            variant_rows=[{"model_key": "stingray", "variant_id": "only_one", "display_order": 1, "active": True}],
        )
        with self.assertRaisesRegex(ValueError, "expected 2 variants"):
            load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)

    def test_duplicate_variant_fails_fast(self) -> None:
        wb = workbook_with_model_metadata(
            variant_rows=[
                {"model_key": "stingray", "variant_id": "dup", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "dup", "display_order": 2, "active": True},
            ]
        )
        with self.assertRaisesRegex(ValueError, "Duplicate active model_variants rows"):
            load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)

    def test_registry_key_mismatch_fails_fast(self) -> None:
        wb = workbook_with_model_metadata(model_rows=[stingray_model_row(registry_key="wrong")])
        with self.assertRaisesRegex(ValueError, "registry_key"):
            load_model_config_overrides(wb, BASE_STINGRAY_CONFIG)

    def test_generation_discovery_finds_active_models_with_complete_exact_metadata(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=required_source_rows("stingray"),
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
            ],
        )

        configs = discover_temp_configs(wb)

        self.assertEqual(list(configs), ["stingray"])
        self.assertEqual(configs["stingray"].variant_ids, ("a_variant", "z_variant"))
        self.assertEqual(configs["stingray"].expected_variant_count, 2)

    def test_generation_discovery_excludes_inactive_scaffolds(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[
                stingray_model_row(),
                {
                    "model_key": "zr1",
                    "registry_key": "zr1",
                    "model_label": "ZR1",
                    "expected_variant_count": 2,
                    "active": False,
                },
            ],
            source_rows=[*required_source_rows("stingray"), *required_source_rows("zr1")],
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
                {"model_key": "zr1", "variant_id": "zr1_a", "display_order": 1, "active": True},
                {"model_key": "zr1", "variant_id": "zr1_b", "display_order": 2, "active": True},
            ],
        )

        configs = discover_temp_configs(wb)

        self.assertEqual(list(configs), ["stingray"])
        self.assertNotIn("zr1", configs)

    def test_generation_discovery_requires_complete_active_exact_source_roles(self) -> None:
        incomplete_rows = [
            row for row in required_source_rows("stingray") if row["source_role"] != "status_sheet"
        ]
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=incomplete_rows,
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
            ],
        )

        with self.assertRaisesRegex(ValueError, "missing required active model_workbook_sources roles.*status_sheet"):
            discover_temp_configs(wb)

    def test_generation_discovery_inactive_source_rows_do_not_satisfy_completeness(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=required_source_rows("stingray", active=False),
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
            ],
        )

        with self.assertRaisesRegex(ValueError, "missing required active model_workbook_sources roles"):
            discover_temp_configs(wb)

    def test_generation_discovery_global_source_rows_do_not_satisfy_model_completeness(self) -> None:
        global_rows = [
            {**row, "model_key": "shared"}
            for row in required_source_rows("stingray")
        ]
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=global_rows,
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
            ],
        )

        with self.assertRaisesRegex(ValueError, "missing required active model_workbook_sources roles"):
            discover_temp_configs(wb)

    def test_generation_discovery_requires_positive_expected_variant_count(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row(expected_variant_count=0)],
            source_rows=required_source_rows("stingray"),
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
            ],
        )

        with self.assertRaisesRegex(ValueError, "expected_variant_count"):
            discover_temp_configs(wb)

    def test_generation_discovery_requires_active_exact_variant_rows(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=required_source_rows("stingray"),
            variant_rows=[],
        )

        with self.assertRaisesRegex(ValueError, "requires active model_variants rows"):
            discover_temp_configs(wb)

    def test_generation_discovery_requires_variant_count_to_match_expected_count(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row(expected_variant_count=2)],
            source_rows=required_source_rows("stingray"),
            variant_rows=[{"model_key": "stingray", "variant_id": "only_one", "display_order": 1, "active": True}],
        )

        with self.assertRaisesRegex(ValueError, "expected 2 active model_variants rows; found 1"):
            discover_temp_configs(wb)

    def test_generation_discovery_does_not_require_runtime_promotion_metadata(self) -> None:
        wb = workbook_with_model_metadata(
            model_rows=[stingray_model_row()],
            source_rows=required_source_rows("stingray"),
            variant_rows=[
                {"model_key": "stingray", "variant_id": "a_variant", "display_order": 1, "active": True},
                {"model_key": "stingray", "variant_id": "z_variant", "display_order": 2, "active": True},
            ],
        )

        configs = discover_temp_configs(wb)

        self.assertIn("stingray", configs)


class WorkbookOwnedPresentationMetadataTests(unittest.TestCase):
    """Requirement 9: presentation metadata is workbook-owned for EVERY active model.

    Falling back to a Python constant when a workbook row is missing makes the
    runtime unpredictable from the source of truth. It must fail instead — and
    for unpromoted active models too, not only promoted ones.
    """

    def workbook_without(self, sheet_name: str, model_key: str) -> Path:
        """Copy the canonical workbook with one model's rows on one sheet removed."""

        target = Path(self.tmpdir.name) / f"without-{sheet_name}-{model_key}.xlsx"
        wb = load_workbook(WORKBOOK)
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        model_column = headers.index("model_key") + 1
        for row_index in range(ws.max_row, 1, -1):
            value = ws.cell(row_index, model_column).value
            if str(value or "").strip().lower() == model_key:
                ws.delete_rows(row_index)
        wb.save(target)
        wb.close()
        return target

    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)

    def test_missing_presentation_rows_fail_for_an_unpromoted_active_model(self) -> None:
        # zr1 is active and generatable but NOT promoted to the runtime registry.
        for sheet_name in ("runtime_steps", "context_section_master"):
            with self.subTest(sheet=sheet_name):
                snapshot = self.workbook_without(sheet_name, "zr1")
                configs = discover_generation_model_configs(snapshot, root=Path(self.tmpdir.name) / "out")
                self.assertIn("zr1", configs)

                with self.assertRaisesRegex(ValueError, sheet_name):
                    generate_model_artifacts(configs["zr1"])

    def workbook_without_step(self, model_key: str, step_key: str, source: Path | None = None) -> Path:
        target = Path(self.tmpdir.name) / f"without-step-{model_key}-{step_key}.xlsx"
        wb = load_workbook(source or WORKBOOK)
        ws = wb["runtime_steps"]
        headers = [cell.value for cell in ws[1]]
        model_column = headers.index("model_key") + 1
        step_column = headers.index("step_key") + 1
        for row_index in range(ws.max_row, 1, -1):
            same_model = str(ws.cell(row_index, model_column).value or "").strip().lower() == model_key
            same_step = str(ws.cell(row_index, step_column).value or "").strip() == step_key
            if same_model and same_step:
                ws.delete_rows(row_index)
        wb.save(target)
        wb.close()
        return target

    def assert_runtime_step_metadata_rejects(self, snapshot: Path, model_key: str) -> None:
        """Exercise the metadata owner directly; generation has a retained proof below."""

        wb = load_workbook(snapshot, read_only=True, data_only=True)
        try:
            with self.assertRaisesRegex(ValueError, "incomplete workbook-owned runtime_steps rows"):
                load_runtime_steps(wb, model_key)
        finally:
            wb.close()

    def test_deactivating_any_single_runtime_step_is_rejected(self) -> None:
        """Every authored step is load-bearing, for every active model.

        The earlier version of this check compared the workbook against a Python
        list of expected steps and only ran for promoted models. Its replacement
        must not be weaker: losing ANY one of ANY model's steps must fail,
        promoted or not.

        This sweep deactivates rows rather than deleting them, because it reuses
        one in-memory workbook instead of writing 84 snapshots. `active_rows()`
        treats the two identically, and the deleted-row path keeps its own proofs
        in the snapshot-based tests below. Generation is not run here either: it
        is the caller of `load_runtime_steps`, and the last test in this class
        holds that wiring against a real `generate_model_artifacts` run.
        """

        # Opened writable because the sweep flips cells, and never saved — the
        # canonical workbook is source data, not a fixture. Nothing below may
        # call wb.save().
        wb = load_workbook(WORKBOOK, data_only=True)
        try:
            ws = wb["runtime_steps"]
            headers = {cell.value: index for index, cell in enumerate(ws[1], start=1)}
            master = wb["model_master"]
            master_headers = {cell.value: index for index, cell in enumerate(master[1], start=1)}
            active_models = sorted(
                str(master.cell(row_index, master_headers["model_key"]).value or "").strip()
                for row_index in range(2, master.max_row + 1)
                if str(master.cell(row_index, master_headers["model_key"]).value or "").strip()
                and str(master.cell(row_index, master_headers["active"]).value or "").strip().lower()
                in {"true", "1", "yes"}
            )
            self.assertEqual(len(active_models), 6)

            for model_key in active_models:
                model_rows = [
                    row_index
                    for row_index in range(2, ws.max_row + 1)
                    if str(ws.cell(row_index, headers["model_key"]).value or "").strip().lower()
                    == model_key
                ]
                self.assertGreaterEqual(len(model_rows), 10)
                for row_index in model_rows:
                    step_key = str(ws.cell(row_index, headers["step_key"]).value or "")
                    with self.subTest(model=model_key, step=step_key):
                        active_cell = ws.cell(row_index, headers["active"])
                        original = active_cell.value
                        active_cell.value = False
                        try:
                            with self.assertRaisesRegex(
                                ValueError, "incomplete workbook-owned runtime_steps rows"
                            ):
                                load_runtime_steps(wb, model_key)
                        finally:
                            active_cell.value = original
        finally:
            wb.close()

    def test_dropping_one_step_from_two_models_still_fails(self) -> None:
        """The cross-model rule is a union, so surviving peers still demand the step.

        An intersection rule was defeated by dropping the same step from two
        models, or by deactivating one peer row. Both must fail.
        """

        snapshot = self.workbook_without_step("z06", "summary")
        snapshot = self.workbook_without_step("zr1", "summary", source=snapshot)

        for model_key in ("z06", "zr1"):
            with self.subTest(model=model_key):
                self.assert_runtime_step_metadata_rejects(snapshot, model_key)

    def test_workbook_scoped_sources_catch_a_step_no_peer_authors(self) -> None:
        """Pin the model-scoped sheets independently of the cross-model rule.

        Without this, a future change narrowing the check back to peer comparison
        alone would pass every other case in this class.
        """

        snapshot = WORKBOOK
        for model_key in sorted(discover_generation_model_configs(WORKBOOK, root=Path(self.tmpdir.name) / "peers")):
            snapshot = self.workbook_without_step(model_key, "paint", source=snapshot)

        # No peer authors `paint` any more, so only step_order_summary_map can catch it.
        self.assert_runtime_step_metadata_rejects(snapshot, "z06")

    def test_missing_presentation_rows_still_fail_for_a_promoted_model(self) -> None:
        snapshot = self.workbook_without("runtime_steps", "z06")
        configs = discover_generation_model_configs(snapshot, root=Path(self.tmpdir.name) / "out-z06")

        with self.assertRaisesRegex(ValueError, "runtime_steps"):
            generate_model_artifacts(configs["z06"])


if __name__ == "__main__":
    unittest.main()
