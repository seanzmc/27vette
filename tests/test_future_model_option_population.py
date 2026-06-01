#!/usr/bin/env python3
"""Tests for applying simplified future_model_option_review rows to future source sheets."""

from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from create_future_model_option_review import OPTION_REVIEW_HEADERS  # noqa: E402
from corvette_form_generator.future_model_ingest import OPTION_SOURCE_HEADERS, OVS_SOURCE_HEADERS  # noqa: E402

MODULE_PATH = ROOT / "scripts" / "apply_future_model_option_review.py"
spec = importlib.util.spec_from_file_location("apply_future_model_option_review", MODULE_PATH)
assert spec and spec.loader
option_apply = importlib.util.module_from_spec(spec)
spec.loader.exec_module(option_apply)


def append_sheet(wb: Workbook, name: str, headers: list[str] | tuple[str, ...], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    for row in rows or []:
        ws.append([row.get(header, "") for header in headers])


def option_review_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {header: "" for header in OPTION_REVIEW_HEADERS}
    row.update(
        {
            "model_key": "z06",
            "raw_source_sheet": "z06_intextmec_raw",
            "raw_source_span": "z06_intextmec_raw:10-10",
            "orderable_rpo": "ABC",
            "ref_only_rpo": "",
            "source_rpo": "ABC",
            "source_option_description": "Source Option Name",
            "source_disclosure_raw": "Source disclosure",
            "normalized_status_summary": "1lz_h07=available; 2lz_h07=standard; 3lz_h07=unavailable; 1lz_h67=available; 2lz_h67=standard; 3lz_h67=unavailable",
            "suggested_option_id": "opt_abc_001",
            "suggested_section_id": "wheels",
            "suggested_display_order": "12",
            "review_status": "approved",
            "active": "True",
            "final_option_id": "",
            "final_option_name": "Final Option Name",
            "final_description": "Final description",
            "final_detail_raw": "Final detail",
            "final_section_id": "",
            "final_selectable": "",
            "final_display_order": "14",
            "final_display_behavior": "option_card",
        }
    )
    row.update(overrides)
    return row


def population_workbook(existing_rows: list[dict[str, object]] | None = None) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "section_master",
        ["section_id", "section_name", "active"],
        [
            {"section_id": "wheels", "section_name": "Wheels", "active": True},
            {"section_id": "exterior", "section_name": "Exterior", "active": True},
        ],
    )
    append_sheet(wb, "future_model_option_review", OPTION_REVIEW_HEADERS, existing_rows or [])
    for model_key in ["z06", "zr1", "zr1x"]:
        append_sheet(wb, f"{model_key}_options", OPTION_SOURCE_HEADERS, [])
        append_sheet(wb, f"{model_key}_ovs", OVS_SOURCE_HEADERS, [])
    return wb


class FutureModelOptionPopulationTests(unittest.TestCase):
    def test_csv_section_filled_rows_materialize_options_and_ovs_with_spec_mapping(self) -> None:
        wb = population_workbook()
        csv_rows = [
            option_review_row(),
            option_review_row(raw_source_span="z06_intextmec_raw:20-20", source_rpo="REF", orderable_rpo="", ref_only_rpo="REF", suggested_option_id="opt_ref_001", suggested_section_id=""),
        ]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(plan["error_count"], 0)
        self.assertEqual(z06["eligible_option_count"], 1)
        self.assertEqual(z06["emitted_ovs_count"], 6)
        self.assertEqual(
            z06["option_rows"],
            [
                {
                    "option_id": "opt_abc_001",
                    "rpo": "ABC",
                    "price": "",
                    "option_name": "Final Option Name",
                    "description": "Final description",
                    "detail_raw": "Final detail",
                    "section_id": "wheels",
                    "selectable": "True",
                    "display_order": "14",
                    "active": "True",
                    "display_behavior": "option_card",
                }
            ],
        )
        self.assertEqual(
            [row["variant_id"] for row in z06["ovs_rows"]],
            ["1lz_h07", "2lz_h07", "3lz_h07", "1lz_h67", "2lz_h67", "3lz_h67"],
        )
        self.assertEqual([row["status"] for row in z06["ovs_rows"]], ["available", "standard", "unavailable", "available", "standard", "unavailable"])
        self.assertEqual(z06["blocked_counts"]["missing_resolved_section_id"], 1)

    def test_ref_only_rows_emit_as_not_selectable_when_section_filled(self) -> None:
        wb = population_workbook()
        csv_rows = [option_review_row(orderable_rpo="", ref_only_rpo="REF", source_rpo="REF", suggested_option_id="opt_ref_001", final_option_id="")]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        row = plan["models"]["z06"]["option_rows"][0]
        self.assertEqual(row["rpo"], "REF")
        self.assertEqual(row["selectable"], "False")

    def test_population_plan_preserves_existing_option_prices(self) -> None:
        wb = population_workbook()
        wb["z06_options"].append([
            "opt_abc_001" if header == "option_id" else
            "ABC" if header == "rpo" else
            "1495" if header == "price" else
            "Existing Option" if header == "option_name" else
            "wheels" if header == "section_id" else
            "True" if header in {"selectable", "active"} else
            ""
            for header in OPTION_SOURCE_HEADERS
        ])
        csv_rows = [option_review_row()]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        row = plan["models"]["z06"]["option_rows"][0]
        self.assertEqual(row["option_id"], "opt_abc_001")
        self.assertEqual(row["price"], "1495")

    def test_population_plan_preserves_existing_option_values_when_review_fields_are_blank(self) -> None:
        wb = population_workbook()
        wb["z06_options"].append([
            "opt_abc_001" if header == "option_id" else
            "ABC" if header == "rpo" else
            "1495" if header == "price" else
            "Existing Option" if header == "option_name" else
            "Existing description" if header == "description" else
            "Existing detail" if header == "detail_raw" else
            "wheels" if header == "section_id" else
            "55" if header == "display_order" else
            "option_card" if header == "display_behavior" else
            "True" if header in {"selectable", "active"} else
            ""
            for header in OPTION_SOURCE_HEADERS
        ])
        csv_rows = [
            option_review_row(
                source_disclosure_raw="",
                final_description="",
                final_detail_raw="",
                final_display_order="",
                suggested_display_order="",
                final_display_behavior="",
            )
        ]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        row = plan["models"]["z06"]["option_rows"][0]
        self.assertEqual(row["price"], "1495")
        self.assertEqual(row["description"], "Existing description")
        self.assertEqual(row["detail_raw"], "Existing detail")
        self.assertEqual(row["display_order"], "55")
        self.assertEqual(row["display_behavior"], "option_card")

    def test_overlay_csv_rows_preserves_existing_workbook_rows_missing_from_csv(self) -> None:
        existing = option_review_row(notes="keep me")
        csv_replacement = option_review_row(final_option_name="CSV Name")
        missing_existing = option_review_row(raw_source_span="z06_intextmec_raw:99-99", source_rpo="ZZZ", suggested_option_id="opt_zzz_001", notes="preserve absent row")

        merged = option_apply.merge_csv_with_existing_option_review_rows([existing, missing_existing], [csv_replacement])

        self.assertEqual(len(merged), 2)
        self.assertEqual(merged[0]["final_option_name"], "CSV Name")
        self.assertEqual(merged[1]["notes"], "preserve absent row")

    def test_duplicate_option_ids_unknown_sections_and_missing_statuses_are_errors(self) -> None:
        wb = population_workbook()
        csv_rows = [
            option_review_row(),
            option_review_row(raw_source_span="z06_intextmec_raw:11-11", final_option_name="Duplicate", normalized_status_summary="1lz_h07=available", suggested_section_id="missing_section"),
        ]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(plan["error_count"], 2)
        self.assertEqual(z06["eligible_option_count"], 0)
        self.assertTrue(any("duplicate option_id opt_abc_001" in error for error in z06["errors"]))
        self.assertTrue(any("resolved_section_id missing_section is not in section_master" in error for error in z06["errors"]))
        self.assertTrue(any("status_2lz_h07 is required" in error for error in z06["errors"]))

    def test_inactive_needs_section_review_rows_with_sections_are_blocked(self) -> None:
        wb = population_workbook()
        csv_rows = [
            option_review_row(
                review_status="needs_section_review",
                active="False",
                suggested_section_id="wheels",
                suggested_option_id="opt_blocked_001",
            )
        ]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(plan["error_count"], 0)
        self.assertEqual(z06["eligible_option_count"], 0)
        self.assertEqual(z06["emitted_ovs_count"], 0)
        self.assertEqual(z06["blocked_counts"]["needs_section_review"], 1)
        self.assertEqual(z06["blocked_counts"]["inactive"], 1)

    def test_final_section_and_selectable_override_suggested_defaults(self) -> None:
        wb = population_workbook()
        csv_rows = [
            option_review_row(
                suggested_section_id="wheels",
                final_section_id="exterior",
                orderable_rpo="ABC",
                final_selectable="False",
            )
        ]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(plan["error_count"], 0)
        self.assertEqual(z06["eligible_option_count"], 1)
        self.assertEqual(z06["option_rows"][0]["section_id"], "exterior")
        self.assertEqual(z06["option_rows"][0]["selectable"], "False")

    def test_blocked_deferred_rows_are_reported_separately_from_missing_sections(self) -> None:
        wb = population_workbook()
        csv_rows = [
            option_review_row(review_status="deferred", active="False", suggested_section_id="wheels"),
            option_review_row(
                raw_source_span="z06_intextmec_raw:21-21",
                source_rpo="DEF",
                suggested_option_id="opt_def_001",
                review_status="approved",
                active="True",
                suggested_section_id="",
            ),
        ]

        plan = option_apply.build_future_option_population_plan(wb, csv_rows, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(plan["error_count"], 0)
        self.assertEqual(z06["eligible_option_count"], 0)
        self.assertEqual(z06["blocked_counts"]["deferred"], 1)
        self.assertEqual(z06["blocked_counts"]["inactive"], 1)
        self.assertEqual(z06["blocked_counts"]["missing_resolved_section_id"], 1)


if __name__ == "__main__":
    unittest.main()
