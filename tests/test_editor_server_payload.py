#!/usr/bin/env python3
"""Tests for the workbook-editor server payload derivation (Phase 1)."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from workbook_editor_server import (  # noqa: E402
    build_payload,
    extract_workbook,
    sheet_payload,
)

REAL_WORKBOOK = ROOT / "stingray_master.xlsx"


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def build_fixture_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb, "model_master",
        ["model_key", "registry_key", "model_label", "model_year", "default_model", "active"],
        [
            {"model_key": "stingray", "registry_key": "stingray", "model_label": "Stingray",
             "model_year": "2027", "default_model": True, "active": True},
            {"model_key": "z06", "registry_key": "z06", "model_label": "Z06",
             "model_year": "2027", "default_model": False, "active": True},
            {"model_key": "zr1", "registry_key": "zr1", "model_label": "ZR1",
             "model_year": "2027", "default_model": False, "active": False},
        ],
    )
    append_sheet(
        wb, "model_registry_promotion",
        ["model_key", "promoted_to_runtime", "default_model", "display_order", "active"],
        [
            {"model_key": "stingray", "promoted_to_runtime": True, "default_model": True,
             "display_order": 1, "active": True},
            {"model_key": "z06", "promoted_to_runtime": True, "default_model": False,
             "display_order": 3, "active": True},
        ],
    )
    append_sheet(
        wb, "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        [
            {"model_key": "stingray", "source_role": "source_option_sheet",
             "sheet_name": "stingray_options", "active": True},
            {"model_key": "stingray", "source_role": "status_sheet",
             "sheet_name": "stingray_ovs", "active": True},
            {"model_key": "z06", "source_role": "source_option_sheet",
             "sheet_name": "z06_options", "active": True},
            {"model_key": "zr1", "source_role": "source_option_sheet",
             "sheet_name": "zr1_options", "active": False},
            {"model_key": "stingray", "source_role": "mystery_role",
             "sheet_name": "stingray_options", "active": True},
            {"model_key": "stingray", "source_role": "rule_groups_sheet",
             "sheet_name": "rule_groups", "active": True},
            {"model_key": "stingray", "source_role": "rule_group_members_sheet",
             "sheet_name": "rule_group_members", "active": True},
        ],
    )
    append_sheet(
        wb, "rule_groups",
        ["group_id", "group_type", "source_id", "active", "notes"],
        [{"group_id": "grp_alpha", "group_type": "requires_any",
          "source_id": "opt_z51_001", "active": True}],
    )
    append_sheet(
        wb, "rule_group_members",
        ["group_id", "target_id", "display_order", "active"],
        [{"group_id": "grp_alpha", "target_id": "opt_gkz_001", "display_order": 10, "active": True}],
    )
    append_sheet(
        wb, "runtime_steps",
        ["model_key", "step_key", "step_label", "runtime_order", "active"],
        [
            {"model_key": "stingray", "step_key": "paint", "step_label": "Exterior Paint",
             "runtime_order": 3, "active": True},
            {"model_key": "stingray", "step_key": "body_style", "step_label": "Body Style",
             "runtime_order": 1, "active": True},
            {"model_key": "stingray", "step_key": "old_step", "step_label": "Old",
             "runtime_order": 99, "active": False},
        ],
    )
    append_sheet(
        wb, "context_section_master",
        ["model_key", "context_type", "section_id", "section_name", "step_key", "active"],
        [
            {"model_key": "stingray", "context_type": "body_style",
             "section_id": "sec_context_body_style", "section_name": "Body Style",
             "step_key": "body_style", "active": True},
        ],
    )
    append_sheet(
        wb, "section_master",
        ["section_id", "section_name", "selection_mode", "is_required", "display_order",
         "standard_behavior", "step_key"],
        [
            {"section_id": "sec_pain_001", "section_name": "Paint",
             "selection_mode": "single", "is_required": True, "display_order": 10,
             "step_key": "paint"},
            {"section_id": "sec_whee_002", "section_name": "Wheels",
             "selection_mode": "single", "is_required": False, "display_order": 20,
             "step_key": "wheels"},
        ],
    )
    append_sheet(
        wb, "section_presentation",
        ["model_key", "section_id", "display_label", "step_key", "section_display_order", "active"],
        [
            {"model_key": "stingray", "section_id": "sec_pain_001",
             "display_label": "Exterior Paint", "step_key": "paint",
             "section_display_order": 10, "active": True},
        ],
    )
    append_sheet(
        wb, "variant_master",
        ["variant_id", "display_name", "active"],
        [
            {"variant_id": "1lt_c07", "display_name": "1LT Coupe", "active": True},
            {"variant_id": "2lt_c07", "display_name": "2LT Coupe", "active": True},
        ],
    )
    append_sheet(
        wb, "model_variants",
        ["model_key", "variant_id", "display_order", "active"],
        [
            {"model_key": "stingray", "variant_id": "1lt_c07", "display_order": 1, "active": True},
            {"model_key": "stingray", "variant_id": "2lt_c07", "display_order": 2, "active": True},
            {"model_key": "zr1", "variant_id": "zr1_c07", "display_order": 1, "active": False},
        ],
    )
    append_sheet(
        wb, "stingray_options",
        ["option_id", "rpo", "price", "option_name", "section_id", "selectable",
         "display_order", "active", "display_behavior"],
        [
            {"option_id": "opt_z51_001", "rpo": "Z51", "price": 5395,
             "option_name": "Z51 Performance Package", "section_id": "sec_pain_001",
             "selectable": True, "display_order": 30, "active": True},
            {"option_id": "opt_gkz_001", "rpo": "GKZ", "price": 0,
             "option_name": "Torch Red", "section_id": "sec_pain_001",
             "selectable": True, "display_order": 10, "active": True},
        ],
    )
    append_sheet(
        wb, "stingray_ovs",
        ["option_id", "variant_id", "status"],
        [{"option_id": "opt_z51_001", "variant_id": "1lt_c07", "status": "available"}],
    )
    append_sheet(
        wb, "z06_options",
        ["option_id", "rpo", "price", "option_name", "section_id"],
        [{"option_id": "opt_z07_001", "rpo": "Z07", "price": 9500,
          "option_name": "Z07 Performance Package", "section_id": "sec_pain_001"}],
    )
    append_sheet(
        wb, "form_steps",
        ["step_key", "label"],
        [{"step_key": "paint", "label": "Paint"}],
    )
    return wb


class PayloadTestBase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        wb = build_fixture_workbook()
        cls._tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(cls._tmp.name)
        cls.extract = extract_workbook(Path(cls._tmp.name))
        cls.payload = build_payload(cls.extract)

    @classmethod
    def tearDownClass(cls):
        Path(cls._tmp.name).unlink(missing_ok=True)


class ModelsTest(PayloadTestBase):
    def test_models_sorted_by_promotion_display_order(self):
        keys = [m["key"] for m in self.payload["models"]]
        self.assertEqual(keys, ["stingray", "z06", "zr1"])

    def test_model_flags(self):
        by_key = {m["key"]: m for m in self.payload["models"]}
        self.assertTrue(by_key["stingray"]["defaultModel"])
        self.assertTrue(by_key["stingray"]["promoted"])
        self.assertTrue(by_key["z06"]["promoted"])
        self.assertFalse(by_key["zr1"]["promoted"])
        self.assertFalse(by_key["zr1"]["active"])


class ModelSheetsTest(PayloadTestBase):
    def test_active_rows_with_known_roles_only(self):
        sheets = self.payload["modelSheets"]
        stingray = {e["sheet"]: e for e in sheets["stingray"]}
        self.assertEqual(
            set(stingray),
            {"stingray_options", "stingray_ovs", "rule_groups", "rule_group_members"},
        )
        self.assertEqual(stingray["stingray_options"]["family"], "options")
        self.assertNotIn("zr1", sheets)


class SheetClassificationTest(PayloadTestBase):
    def test_generated_and_unregistered_sheets_read_only(self):
        by_name = {s["name"]: s for s in self.payload["sheets"]}
        self.assertTrue(by_name["form_steps"]["readOnly"])
        self.assertIsNone(by_name["form_steps"]["family"])
        self.assertTrue(by_name["section_master"]["readOnly"])

    def test_registered_source_sheet_carries_meta(self):
        by_name = {s["name"]: s for s in self.payload["sheets"]}
        entry = by_name["stingray_options"]
        self.assertFalse(entry["readOnly"])
        self.assertEqual(entry["family"], "options")
        self.assertEqual(entry["keyCols"], ["option_id"])
        self.assertEqual(entry["types"]["display_order"], "int")
        self.assertIn("display_behavior", entry["enums"])
        self.assertEqual(entry["rowCount"], 2)


class StepsAndSectionsTest(PayloadTestBase):
    def test_steps_active_only(self):
        steps = [s for s in self.payload["steps"] if s["modelKey"] == "stingray"]
        self.assertEqual([s["stepKey"] for s in steps], ["body_style", "paint"])

    def test_section_surfaces_present(self):
        self.assertEqual(len(self.payload["sections"]), 2)
        self.assertEqual(len(self.payload["sectionPresentation"]), 1)
        self.assertEqual(len(self.payload["contextSections"]), 1)


class ReferenceDomainsTest(PayloadTestBase):
    def test_sections_domain(self):
        domain = self.payload["referenceDomains"]["sections"]
        self.assertIn({"id": "sec_pain_001", "name": "Paint"}, domain)

    def test_variants_by_model_excludes_inactive(self):
        variants = self.payload["referenceDomains"]["variantsByModel"]
        self.assertEqual([v["id"] for v in variants["stingray"]], ["1lt_c07", "2lt_c07"])
        self.assertNotIn("zr1", variants)

    def test_options_by_model(self):
        options = self.payload["referenceDomains"]["optionsByModel"]
        self.assertEqual(len(options["stingray"]), 2)
        self.assertEqual(options["z06"][0]["rpo"], "Z07")

    def test_group_and_interior_domains(self):
        dom = self.payload["referenceDomains"]
        self.assertEqual([g["id"] for g in dom["ruleGroupsByModel"]["stingray"]], ["grp_alpha"])
        self.assertEqual(dom["exclusiveGroupsByModel"], {})
        self.assertEqual(dom["interiorsByModel"], {})


class SheetPayloadTest(PayloadTestBase):
    def test_rows_preserve_json_types(self):
        payload = sheet_payload(self.extract, "stingray_options")
        row = payload["rows"][0]
        self.assertEqual(row["price"], 5395)
        self.assertIs(row["selectable"], True)

    def test_unknown_sheet_returns_none(self):
        self.assertIsNone(sheet_payload(self.extract, "nope"))


@unittest.skipUnless(REAL_WORKBOOK.exists(), "canonical workbook not present")
class RealWorkbookIntegrationTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.extract = extract_workbook(REAL_WORKBOOK)
        cls.payload = build_payload(cls.extract)

    def test_live_models_promoted(self):
        by_key = {m["key"]: m for m in self.payload["models"]}
        for key in ("stingray", "grand_sport", "z06"):
            self.assertTrue(by_key[key]["promoted"], key)
        self.assertTrue(by_key["stingray"]["defaultModel"])

    def test_model_sheet_registries(self):
        sheets = self.payload["modelSheets"]
        z06_names = {e["sheet"] for e in sheets["z06"]}
        self.assertIn("z06_options", z06_names)
        self.assertIn("LZ_Interiors", z06_names)

    def test_large_sheet_served(self):
        payload = sheet_payload(self.extract, "stingray_ovs")
        self.assertGreater(len(payload["rows"]), 1000)

    def test_form_sheets_read_only(self):
        for entry in self.payload["sheets"]:
            if entry["name"].startswith("form_"):
                self.assertTrue(entry["readOnly"], entry["name"])

    def test_group_domains_real(self):
        dom = self.payload["referenceDomains"]
        self.assertTrue(any(g["id"] for g in dom["exclusiveGroupsByModel"]["z06"]))
        self.assertTrue(any(i["id"] for i in dom["interiorsByModel"]["stingray"]))


if __name__ == "__main__":
    unittest.main()
