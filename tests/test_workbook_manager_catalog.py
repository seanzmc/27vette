"""Pass 2 shared-registry/catalog and requiredness regressions."""

from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "workbook-manager" / "backend"
SCRIPTS = ROOT / "scripts"
for path in (BACKEND, SCRIPTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from app import catalog  # noqa: E402
from corvette_form_generator import editor_ops  # noqa: E402
from corvette_form_generator.schema_validation import REQUIRED_SHEETS  # noqa: E402
from corvette_form_generator.workbook_domain.registry import (  # noqa: E402
    EDITOR_SHEET_META,
    GLOBAL_SHEET_FAMILIES,
    SOURCE_ROLE_FAMILIES,
)

WORKBOOK = ROOT / "stingray_master.xlsx"


class TestCatalogParity(unittest.TestCase):
    def test_structure_specs_follow_registered_fixed_sheet_specs(self):
        expected_families = {
            "model_registry_promotion",
            "model_workbook_sources",
            "variant_master",
            "model_variants",
            "order_summary_sections_meta",
            "step_order_summary_map_meta",
        }
        specs = catalog.structure_specs()
        self.assertTrue(expected_families.issubset({spec.family for spec in specs}))

        synthetic = catalog.TableSpec(
            table="synthetic_structure",
            family="synthetic_structure",
            sheet=("synthetic_structure",),
            editable=True,
        )
        followed = catalog.structure_specs((*catalog.WRITABLE_SPECS, synthetic))
        self.assertIn("synthetic_structure", {spec.table for spec in followed})

    def test_group_editor_metadata_follows_registry_derived_specs(self):
        group = replace(
            catalog.SPEC_BY_FAMILY["exclusive_groups"],
            table="groups_v2",
            key=("parent_key",),
        )
        members = replace(
            catalog.SPEC_BY_FAMILY["exclusive_members"],
            table="members_v2",
            key=("parent_key", "member_key"),
            columns=(
                catalog.ColumnSpec("parent_key"),
                catalog.ColumnSpec("member_key"),
                catalog.ColumnSpec("sequence", "int"),
                catalog.ColumnSpec("enabled", "bool"),
            ),
            refs=(catalog.RefSpec("parent_key", "groups_v2", "parent_key"),),
        )
        with mock.patch.dict(catalog.SPEC_BY_FAMILY, {
            "exclusive_groups": group,
            "exclusive_members": members,
        }):
            self.assertEqual(catalog.group_editor_metadata("exclusive"), {
                "group_table": "groups_v2",
                "group_id_field": "parent_key",
                "member_table": "members_v2",
                "member_id_field": "member_key",
                "member_group_field": "parent_key",
                "member_order_field": "sequence",
                "member_active_field": "enabled",
            })

    def test_every_writable_family_uses_shared_contract_metadata(self):
        self.assertEqual(set(catalog.WRITABLE_FAMILIES), set(EDITOR_SHEET_META))
        for family, shared in EDITOR_SHEET_META.items():
            spec = catalog.SPEC_BY_FAMILY[family]
            # spec.key holds SQL names (catalog._build_spec sanitizes the
            # registry headers); PriceRef is the first family whose key headers
            # differ from their SQL spelling.
            self.assertEqual(
                spec.key, tuple(catalog.sanitize_identifier(k) for k in shared["key"]), family
            )
            self.assertEqual(tuple(c.header for c in spec.columns), tuple(shared["columns"]), family)
            self.assertEqual(
                {c.header: c.ctype for c in spec.columns if c.ctype != "text"},
                dict(shared.get("types", {})),
                family,
            )
            self.assertEqual(
                {c.header: c.enum for c in spec.columns if c.enum},
                {key: tuple(values) for key, values in shared.get("enums", {}).items()},
                family,
            )
            self.assertEqual(spec.required_on_add, tuple(shared["required_on_add"]), family)
            self.assertEqual(
                spec.required_on_effective_active_row,
                tuple(shared["required_on_effective_active_row"]),
                family,
            )
            self.assertEqual(spec.optional_columns, tuple(shared["optional_columns"]), family)
            self.assertEqual(spec.ref_contract(), catalog.shared_ref_contract(shared), family)

    def test_sheet_catalog_classifies_every_live_workbook_sheet(self):
        wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
        try:
            classifications = catalog.classify_workbook_sheets(wb)
            self.assertEqual(set(classifications), set(wb.sheetnames))
            self.assertTrue(
                set(REQUIRED_SHEETS).issubset(classifications),
                set(REQUIRED_SHEETS) - set(classifications),
            )
            self.assertEqual(classifications["section_master"].disposition, "managed_read_only")
            for sheet in ("PriceRef", "context_choice_copy"):
                self.assertEqual(classifications[sheet].disposition, "managed_writable", sheet)
            for sheet in ("rule_phrase_map", "runtime_rule_exceptions"):
                self.assertEqual(classifications[sheet].disposition, "workbook_preserved_known", sheet)
            self.assertNotIn(
                "raw_rows",
                {classification.disposition for classification in classifications.values()},
            )
        finally:
            wb.close()

    def test_managed_columns_have_one_known_or_opaque_disposition(self):
        wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
        try:
            classifications = catalog.classify_workbook_sheets(wb)
            for sheet, classification in classifications.items():
                if not classification.family:
                    continue
                headers = [str(cell.value).strip() for cell in wb[sheet][1] if cell.value]
                reconciliation = catalog.reconcile_columns(classification.spec, headers)
                self.assertEqual(
                    set(headers),
                    set(reconciliation.known) | set(reconciliation.opaque),
                    sheet,
                )
                self.assertEqual(set(reconciliation.known) & set(reconciliation.opaque), set())
        finally:
            wb.close()

    def test_optional_blank_maps_to_none_and_required_blank_is_blocking(self):
        options = catalog.SPEC_BY_FAMILY["options"]
        price = options.column_by_header("price")
        selectable = options.column_by_header("selectable")
        self.assertIsNone(catalog.projection_value(price, None))
        self.assertIsNone(catalog.projection_value(price, ""))
        with self.assertRaises(catalog.RequiredValueError):
            catalog.projection_value(selectable, "")


class TestSharedRequiredness(unittest.TestCase):
    def test_prepare_batch_rejects_blank_required_field_on_effective_row(self):
        extract = editor_ops.extract_workbook(WORKBOOK)
        sheet = "stingray_options"
        row = next(row for row in extract["sheets"][sheet]["rows"] if row.get("option_id"))
        result = editor_ops._prepare_batch(
            extract,
            {
                "items": [
                    {
                        "action": "update",
                        "sheet": sheet,
                        "key": {"option_id": row["option_id"]},
                        "row": {"selectable": ""},
                    }
                ]
            },
        )
        errors = result[0]
        self.assertTrue(any("required field selectable is blank" in error for error in errors), errors)


if __name__ == "__main__":
    unittest.main()
