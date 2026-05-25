#!/usr/bin/env python3
"""Tests for applying future_model_source_review rows to normalized source sheets."""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SOURCE_REVIEW_HEADERS,
    OPTION_SOURCE_HEADERS,
    OVS_SOURCE_HEADERS,
    build_future_source_population_plan,
)


def append_sheet(wb: Workbook, name: str, headers: list[str] | tuple[str, ...], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    for row in rows or []:
        ws.append([row.get(header, "") for header in headers])


def review_row(**overrides: object) -> dict[str, object]:
    row = {header: "" for header in FUTURE_MODEL_SOURCE_REVIEW_HEADERS}
    row.update(
        {
            "model_key": "z06",
            "raw_source_sheets": "z06_standard_raw",
            "raw_source_spans": "z06_standard_raw:10-13",
            "source_primary_rpo": "AJ7",
            "candidate_option_id": "opt_aj7_001",
            "approved_option_id": "opt_aj7_001",
            "approved_rpo": "AJ7",
            "approved_price": "",
            "approved_option_name": "Airbags, frontal and side-impact",
            "approved_description": "",
            "approved_detail_raw": "Disclosure one\nDisclosure two",
            "approved_section_id": "standard_equipment",
            "approved_selectable": "False",
            "approved_display_behavior": "standard",
            "approved_display_order": "10",
            "review_status": "approved",
            "active": "True",
            "status_1lz_h07": "standard",
            "status_2lz_h07": "standard",
            "status_3lz_h07": "standard",
            "status_1lz_h67": "standard",
            "status_2lz_h67": "standard",
            "status_3lz_h67": "standard",
        }
    )
    row.update(overrides)
    return row


def population_workbook(rows: list[dict[str, object]]) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(wb, "section_master", ["section_id", "section_label", "active"], [
        {"section_id": "standard_equipment", "section_label": "Standard Equipment", "active": True},
        {"section_id": "wheels", "section_label": "Wheels", "active": True},
    ])
    append_sheet(wb, "future_model_source_review", FUTURE_MODEL_SOURCE_REVIEW_HEADERS, rows)
    for model_key in ["z06", "zr1", "zr1x"]:
        append_sheet(wb, f"{model_key}_options", OPTION_SOURCE_HEADERS, [])
        append_sheet(wb, f"{model_key}_ovs", OVS_SOURCE_HEADERS, [])
    return wb


class FutureModelSourcePopulationTests(unittest.TestCase):
    def test_approved_active_review_row_materializes_options_and_ovs_rows(self) -> None:
        wb = population_workbook([review_row()])

        plan = build_future_source_population_plan(wb, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(plan["error_count"], 0)
        self.assertEqual(z06["eligible_option_count"], 1)
        self.assertEqual(z06["emitted_ovs_count"], 6)
        self.assertEqual(z06["option_rows"], [
            {
                "option_id": "opt_aj7_001",
                "rpo": "AJ7",
                "price": "",
                "option_name": "Airbags, frontal and side-impact",
                "description": "",
                "detail_raw": "Disclosure one\nDisclosure two",
                "section_id": "standard_equipment",
                "selectable": "False",
                "display_order": "10",
                "active": "True",
                "display_behavior": "standard",
            }
        ])
        self.assertEqual(
            [row["variant_id"] for row in z06["ovs_rows"]],
            ["1lz_h07", "2lz_h07", "3lz_h07", "1lz_h67", "2lz_h67", "3lz_h67"],
        )
        self.assertTrue(all(row["status"] == "standard" for row in z06["ovs_rows"]))

    def test_needs_review_rows_are_blocked_without_validation_errors(self) -> None:
        wb = population_workbook([
            review_row(review_status="needs_review", active="False", review_flags="section_unresolved"),
        ])

        plan = build_future_source_population_plan(wb, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(plan["error_count"], 0)
        self.assertEqual(z06["eligible_option_count"], 0)
        self.assertEqual(z06["emitted_ovs_count"], 0)
        self.assertEqual(z06["blocked_counts"]["needs_review"], 1)
        self.assertEqual(z06["blocked_counts"]["inactive"], 1)
        self.assertEqual(z06["blocked_counts"]["section_unresolved"], 1)

    def test_approved_active_rows_with_blocking_flags_are_errors(self) -> None:
        wb = population_workbook([
            review_row(review_flags="section_unresolved"),
        ])

        plan = build_future_source_population_plan(wb, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(z06["eligible_option_count"], 0)
        self.assertEqual(plan["error_count"], 1)
        self.assertIn("section_unresolved", z06["errors"][0])

    def test_duplicate_approved_option_ids_are_rejected_per_model(self) -> None:
        wb = population_workbook([
            review_row(approved_display_order="10"),
            review_row(approved_display_order="20", approved_option_name="Duplicate Airbag Row"),
        ])

        plan = build_future_source_population_plan(wb, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(z06["eligible_option_count"], 0)
        self.assertEqual(plan["error_count"], 2)
        self.assertTrue(all("duplicate approved_option_id opt_aj7_001" in error for error in z06["errors"]))

    def test_missing_required_fields_and_statuses_are_rejected(self) -> None:
        wb = population_workbook([
            review_row(approved_rpo="", status_2lz_h07=""),
        ])

        plan = build_future_source_population_plan(wb, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(z06["eligible_option_count"], 0)
        self.assertEqual(plan["error_count"], 1)
        self.assertIn("approved_rpo is required", z06["errors"][0])
        self.assertIn("status_2lz_h07 is required", z06["errors"][0])

    def test_price_schedule_multiple_candidates_can_emit_only_with_blank_price(self) -> None:
        blank_price = review_row(
            approved_option_id="opt_blank_price",
            review_flags="price_schedule_multiple_candidates",
            approved_price="",
            approved_display_order="1",
        )
        explicit_price = review_row(
            approved_option_id="opt_explicit_price",
            review_flags="price_schedule_multiple_candidates",
            approved_price="1234",
            approved_display_order="2",
        )
        wb = population_workbook([blank_price, explicit_price])

        plan = build_future_source_population_plan(wb, ["z06"])

        z06 = plan["models"]["z06"]
        self.assertEqual(z06["eligible_option_count"], 1)
        self.assertEqual(z06["option_rows"][0]["option_id"], "opt_blank_price")
        self.assertEqual(plan["error_count"], 1)
        self.assertIn("price_schedule_multiple_candidates requires blank approved_price", z06["errors"][0])


if __name__ == "__main__":
    unittest.main()
