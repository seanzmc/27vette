from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from corvette_form_generator.ingest.wizard.decisions import candidate_fingerprint


def test_options_recovery_projection_module_exists() -> None:
    assert importlib.util.find_spec(
        "corvette_form_generator.ingest.options_recovery_projection"
    ) is not None


def test_identifying_copy_keeps_wheel_finish_in_name_and_size_in_description() -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        _identifying_copy_proposal,
    )

    proposal = _identifying_copy_proposal(
        {
            "option_name": (
                'Wheels, 20" x 10" (50.8 cm x 25.4 cm) front and '
                '21" x 13" (53.3 cm x 33 cm) rear 20-spoke, '
                "Edge Blue-painted forged aluminum"
            ),
            "description": "",
            "detail_raw": "",
            "section_id": "sec_whee_001",
        },
    )

    assert proposal["option_name"] == "20-Spoke Edge Blue-Painted Forged Aluminum Wheels"
    assert proposal["description"] == (
        '20" x 10" (50.8 cm x 25.4 cm) front and '
        '21" x 13" (53.3 cm x 33 cm) rear'
    )


def test_identifying_copy_uses_noun_final_caliper_name() -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        _identifying_copy_proposal,
    )

    proposal = _identifying_copy_proposal(
        {
            "option_name": "Calipers, Bronze-painted",
            "description": "",
            "detail_raw": "",
            "section_id": "sec_cali_001",
        },
    )

    assert proposal == {
        "option_name": "Bronze-Painted Calipers",
        "description": "",
        "sourceType": "identifying_copy_derivation",
        "flags": [],
    }


@pytest.mark.parametrize(
    ("raw", "expected_name", "expected_description"),
    [
        (
            "Transmission, 8-speed dual clutch, includes manual and auto modes",
            "8-Speed Dual-Clutch Transmission",
            "Includes manual and automatic modes",
        ),
        (
            "Competition Yellow custom leather stitch, includes seats, instrument panel, doors and console",
            "Competition Yellow Custom Leather Stitching",
            "Includes seats, instrument panel, doors, and console",
        ),
        (
            "NEW!  Royal Blue Full Length Dual Racing Stripes\n1. Not available with exterior color (GTR) Admiral Blue Metallic.",
            "Royal Blue Full-Length Dual Racing Stripes",
            "Not available with exterior color (GTR) Admiral Blue Metallic.",
        ),
        (
            "Brakes, 4-wheel antilock, 4-wheel disc, carbon ceramic, 10-piston front, 6-piston rear",
            "10-Piston Front / 6-Piston Rear Carbon Ceramic Brakes",
            "4-wheel antilock disc",
        ),
        (
            "Front axle, electrified propulsion (186 hp [138.7 kW], 145 lb-ft of front torque [196.6 N-m]). 1250 (932.1 kW) total combined hp with LT7 engine",
            "Electrified Front Axle",
            "186 hp [138.7 kW], 145 lb-ft of front torque [196.6 N-m]; 1250 hp [932.1 kW] total combined with LT7 engine",
        ),
    ],
)
def test_identifying_copy_splits_name_from_ancillary_information(
    raw: str,
    expected_name: str,
    expected_description: str,
) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        _identifying_copy_proposal,
    )

    proposal = _identifying_copy_proposal(
        {
            "option_name": raw,
            "description": "",
            "detail_raw": raw,
            "section_id": "sec_test_001",
        },
    )

    assert proposal["option_name"] == expected_name
    assert proposal["description"] == expected_description
    assert proposal["flags"] == []


OPTION_HEADERS = [
    "option_id",
    "rpo",
    "price",
    "option_name",
    "description",
    "detail_raw",
    "section_id",
    "selectable",
    "display_order",
    "active",
    "display_behavior",
]


def _sheet(workbook: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = workbook.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def _option(
    option_id: str,
    rpo: str,
    price: object,
    name: str,
    description: str,
    detail_raw: str,
    section_id: str,
    selectable: bool,
    display_order: object,
    active: bool = True,
) -> list[object]:
    return [
        option_id,
        rpo,
        price,
        name,
        description,
        detail_raw,
        section_id,
        selectable,
        display_order,
        active,
        "",
    ]


def _write_json(path: Path, payload: object) -> Path:
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return path


def _projection_fixture(tmp_path: Path):
    workbook_path = tmp_path / "current.xlsx"
    pre_path = tmp_path / "pre.xlsx"

    wb = Workbook()
    del wb[wb.sheetnames[0]]
    _sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        [
            ["grand_sport_x", "source_option_sheet", "grand_sport_x_options", False, "unpromoted"],
            ["zr1", "source_option_sheet", "zr1_options", False, "unpromoted"],
            ["zr1x", "source_option_sheet", "zr1x_options", False, "unpromoted"],
            ["grand_sport", "source_option_sheet", "grandSport_options", True, ""],
            ["z06", "source_option_sheet", "z06_options", True, ""],
        ],
    )
    _sheet(
        wb,
        "section_master",
        [
            "section_id",
            "section_name",
            "selection_mode",
            "is_required",
            "display_order",
            "standard_behavior",
            "step_key",
        ],
        [
            ["sec_stan_001", "Standard", "display_only", False, 10, "locked_included", "options"],
            ["sec_safe_001", "Safety", "display_only", False, 20, "locked_included", "options"],
            ["sec_exte_001", "Paint", "single_select_req", True, 30, "", "exterior"],
        ],
    )
    raw_aj7 = "Airbags, frontal and side-impact"
    raw_standard = "Air filtration system, with pollen filter"
    _sheet(
        wb,
        "grand_sport_x_options",
        OPTION_HEADERS,
        [
            _option("opt_aj7_001", "AJ7", 1295, raw_aj7, raw_aj7, raw_aj7, "sec_stan_001", True, ""),
            _option("opt_gba_001", "GBA", 0, "Black", "", "", "sec_exte_001", True, 20),
            _option(
                "opt_std_aaaaaaaaaaaaaaaa",
                "",
                0,
                "Air filtration system",
                raw_standard,
                raw_standard,
                "sec_stan_001",
                False,
                "",
            ),
            _option("opt_n26_001", "N26", "", "Fresh feature", "", "Fresh feature", "sec_stan_001", True, ""),
        ],
    )
    _sheet(
        wb,
        "grandSport_options",
        OPTION_HEADERS,
        [
            _option("opt_010", "", "", "Air Filtration System", "With pollen filter", "", "sec_stan_001", False, 160),
            _option("opt_aj7_001", "AJ7", "", "Airbags", "", "", "sec_safe_001", False, 40, True),
            _option("opt_n26_001", "N26", "", "Fresh feature", "", "", "sec_safe_001", True, 50, False),
        ],
    )
    _sheet(
        wb,
        "z06_options",
        OPTION_HEADERS,
        [
            _option("opt_tdm_001", "TDM", "", "Teen Driver", "Configurable driving settings", "", "sec_safe_001", False, 20),
            _option("opt_new_001", "NEW", "", "New Feature", "Comparator detail", "", "sec_safe_001", True, 50),
            _option("opt_085", "", "", "Air Filtration System", "Includes pollen filter", "", "sec_stan_001", False, 100),
            _option("opt_187", "", "", "Integral Front and Rear Antenna", "", "", "sec_stan_001", False, 114),
        ],
    )
    _sheet(
        wb,
        "zr1_options",
        OPTION_HEADERS,
        [
            _option("opt_tdm_001", "TDM", "", "Teen Driver raw paragraph", "Teen Driver raw paragraph", "Teen Driver raw paragraph", "sec_safe_001", False, 30),
            _option("opt_new_001", "NEW", "", "NEW! New Feature, with detail", "NEW! New Feature, with detail", "NEW! New Feature, with detail", "sec_safe_001", True, ""),
            _option("opt_unm_001", "UNM", "", "NEW! Unmatched Feature, target detail", "", "Requires package ABC.", "sec_safe_001", True, 60),
            _option("opt_feh_001", "FEH", "", "ZR1X-only suspension", "", "", "sec_safe_001", False, 70),
            _option("opt_707", "", "", "Antenna, integral front and rear", "", "", "sec_stan_001", False, 80),
            _option("opt_tu7_001", "TU7", "", "Seats, two-tone", "", "", "sec_stan_001", True, 90),
        ],
    )
    _sheet(
        wb,
        "zr1x_options",
        OPTION_HEADERS,
        [
            _option("opt_tdm_001", "TDM", "", "Teen Driver raw paragraph", "Teen Driver raw paragraph", "Teen Driver raw paragraph", "sec_safe_001", False, 30),
            _option("opt_feh_001", "FEH", "", "ZR1X-only suspension", "", "", "sec_safe_001", False, 40),
            _option("opt_708", "", "", "Antenna, integral front and rear", "", "", "sec_stan_001", False, 80),
            _option("opt_tu7_001", "TU7", "", "Seats, two-tone", "", "", "sec_stan_001", True, 90),
        ],
    )
    _sheet(
        wb,
        "grand_sport_x_ovs",
        ["option_id", "variant_id", "status"],
        [["opt_std_aaaaaaaaaaaaaaaa", "1lt_g07", "standard"]],
    )
    _sheet(
        wb,
        "default_selection_rules",
        ["model_key", "rule_id", "target_option_id", "active"],
        [["grand_sport_x", "default_std", "opt_std_aaaaaaaaaaaaaaaa", True]],
    )
    wb.save(workbook_path)

    pre = Workbook()
    del pre[pre.sheetnames[0]]
    _sheet(
        pre,
        "zr1_options",
        OPTION_HEADERS,
        [
            _option("opt_tdm_001", "TDM", "", "Unusable historical raw", "", "", "sec_safe_001", False, 20),
            _option("opt_unm_001", "UNM", "", "Unusable historical raw", "", "", "sec_safe_001", True, 10),
        ],
    )
    _sheet(
        pre,
        "zr1x_options",
        OPTION_HEADERS,
        [_option("opt_tdm_001", "TDM", "", "Unusable historical raw", "", "", "sec_safe_001", False, 20)],
    )
    pre.save(pre_path)

    candidate = {
        "candidateId": "Interior 3:4",
        "rpo": "AJ7",
        "refOnlyRpo": "",
        "description": raw_aj7,
        "statuses": [],
    }
    decision_id = "grand_sport_x:section:Interior 3:4"
    reviewed_candidates = _write_json(
        tmp_path / "option-candidates.json",
        {"schemaVersion": "pass-a-1", "candidates": [candidate], "skippedRows": []},
    )
    reconciliation_candidates = _write_json(
        tmp_path / "reconciliation-option-candidates.json",
        {
            "schemaVersion": "pass-a-1",
            "candidates": [
                {
                    "candidateId": "Mechanical 5:45",
                    "rpo": "",
                    "refOnlyRpo": "FEH",
                    "description": "ZR1X-only suspension",
                    "statuses": [
                        {"modelCode": "1YR07", "raw": "--", "status": "unavailable"},
                        {"modelCode": "1YR67", "raw": "--", "status": "unavailable"},
                        {"modelCode": "1YS07", "raw": "S", "status": "standard"},
                        {"modelCode": "1YS67", "raw": "S", "status": "standard"},
                    ],
                }
            ],
            "skippedRows": [],
        },
    )
    reviewed_decisions = _write_json(
        tmp_path / "decisions.json",
        {
            "schemaVersion": "pass-b-2",
            "candidatesFingerprint": "fixture",
            "decisions": [
                {
                    "model": "grand_sport_x",
                    "lane": "section",
                    "candidateId": candidate["candidateId"],
                    "candidateFingerprint": candidate_fingerprint(candidate),
                    "decisionId": decision_id,
                }
            ],
        },
    )
    reviewed_plan = _write_json(
        tmp_path / "apply-plan.json",
        {
            "schemaVersion": "pass-c-2",
            "stage2": {
                "items": [
                    {
                        "action": "add",
                        "sheet": "grandSportX_options",
                        "key": {"option_id": "opt_aj7_001"},
                        "row": {
                            "option_id": "opt_aj7_001",
                            "rpo": "AJ7",
                            "price": None,
                            "option_name": "Airbags",
                            "description": "Frontal and side-impact",
                            "detail_raw": raw_aj7,
                            "section_id": "sec_safe_001",
                            "selectable": False,
                            "display_order": 10,
                            "active": False,
                        },
                        "_decisions": [decision_id],
                    }
                ]
            },
        },
    )
    evidence_sha = "a" * 64
    subject_id = "subject:grand_sport_x:missing_section:fixture"
    exception_queue = _write_json(
        tmp_path / "exception-queue.json",
        {
            "schemaVersion": "exception-queue-1",
            "subjects": [
                {
                    "model": "grand_sport_x",
                    "subjectId": subject_id,
                    "subjectVersion": "v1",
                    "evidenceDependencies": [
                        {"evidenceId": f"target:grand_sport_x:candidate:{evidence_sha}"}
                    ],
                }
            ],
        },
    )
    exception_resolutions = _write_json(
        tmp_path / "exception-resolutions.json",
        {
            "schemaVersion": "exception-resolutions-1",
            "entries": [
                {
                    "action": "choose_section",
                    "disposition": "resolved",
                    "payload": {"sectionId": "sec_safe_001"},
                    "reviewer": "SeanM",
                    "subjectId": subject_id,
                    "subjectVersion": "v1",
                }
            ],
        },
    )
    canonical_manifest = _write_json(
        tmp_path / "canonical-row-manifest.json",
        {
            "schemaVersion": "canonical-rows-1",
            "rows": [
                {
                    "model": "grand_sport_x",
                    "family": "options",
                    "sheet": "grand_sport_x_options",
                    "evidenceDependencies": [
                        {"evidenceId": f"target:grand_sport_x:candidate:{evidence_sha}"}
                    ],
                    "values": {
                        "option_id": "opt_std_aaaaaaaaaaaaaaaa",
                        "rpo": "",
                    },
                }
            ],
        },
    )
    comparator_evidence = _write_json(
        tmp_path / "comparator-evidence.json",
        {
            "schemaVersion": "comparator-evidence-1",
            "targets": {
                "grand_sport_x": {"comparator": "grand_sport"},
                "zr1": {"comparator": "z06"},
                "zr1x": {"comparator": "z06"},
            },
        },
    )
    return {
        "workbook_path": workbook_path,
        "pre_integration_workbook_path": pre_path,
        "reviewed_plan_path": reviewed_plan,
        "reviewed_candidates_path": reviewed_candidates,
        "reconciliation_candidates_path": reconciliation_candidates,
        "reviewed_decisions_path": reviewed_decisions,
        "exception_queue_path": exception_queue,
        "exception_resolutions_path": exception_resolutions,
        "canonical_manifest_path": canonical_manifest,
        "comparator_evidence_path": comparator_evidence,
    }


def test_target_all_unavailable_row_is_a_deletion_not_a_copy_review(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    reports = generate_recovery_reports(ProjectionInputs(**paths), tmp_path / "reports")

    zr1 = reports["zr1"]
    deletion = zr1["targetApplicabilityDeletions"][0]
    assert deletion["rpo"] == "FEH"
    assert deletion["optionId"] == "opt_feh_001"
    assert deletion["sourceCandidateIds"] == ["Mechanical 5:45"]
    assert deletion["targetStatuses"] == ["--", "--"]
    assert not any(row["identity"]["rpo"] == "FEH" for row in zr1["residualRows"])

    zr1x = reports["zr1x"]
    assert zr1x["targetApplicabilityDeletions"] == []
    assert any(row["identity"]["rpo"] == "FEH" for row in zr1x["residualRows"])


def test_recorded_gsx_deletions_also_remove_same_rpo_from_zr_targets(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    reports = generate_recovery_reports(ProjectionInputs(**paths), tmp_path / "reports")

    for model in ("zr1", "zr1x"):
        report = reports[model]
        deletion = next(
            item
            for item in report["recordedInstructionDeletions"]
            if item["rpo"] == "TU7"
        )
        assert deletion["optionId"] == "opt_tu7_001"
        assert deletion["reason"] == "checkpoint_1_delete_rpo_across_targets"
        assert not any(row["identity"]["rpo"] == "TU7" for row in report["residualRows"])


def test_generate_reports_is_read_only_and_builds_residual_proposals(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    workbook_before = paths["workbook_path"].read_bytes()
    output_dir = tmp_path / "reports"

    reports = generate_recovery_reports(ProjectionInputs(**paths), output_dir)

    assert paths["workbook_path"].read_bytes() == workbook_before
    assert set(reports) == {"grand_sport_x", "zr1", "zr1x"}
    assert sorted(path.name for path in output_dir.iterdir()) == [
        "grand_sport_x-recovery-projection.json",
        "grand_sport_x-recovery-projection.md",
        "zr1-recovery-projection.json",
        "zr1-recovery-projection.md",
        "zr1x-recovery-projection.json",
        "zr1x-recovery-projection.md",
    ]

    gsx = reports["grand_sport_x"]
    by_rpo = {row["identity"]["rpo"]: row for row in gsx["residualRows"] if row["identity"]["rpo"]}
    assert by_rpo["AJ7"]["after"] == {
        "option_id": "opt_aj7_001",
        "option_name": "Airbags",
        "description": "Frontal and side-impact",
        "detail_raw": "Airbags, frontal and side-impact",
        "section_id": "sec_safe_001",
        "price": 1295,
        "active": True,
        "selectable": True,
        "display_order": 10,
    }
    assert "GBA" not in by_rpo
    assert by_rpo["N26"]["partition"] == "fresh_rpo_full_review"
    gsx_no_rpo = next(row for row in gsx["residualRows"] if not row["identity"]["rpo"])
    assert gsx_no_rpo["after"]["option_name"] == "Air Filtration System"
    assert gsx_no_rpo["after"]["description"] == "Includes pollen filter"
    assert gsx_no_rpo["fieldProvenance"]["option_name"]["reference"]["model"] == "z06"
    assert not any(item["lane"] == "no_rpo_mapping" for item in gsx_no_rpo["reviewItems"])

    id_repair = gsx["idRepairs"][0]
    assert id_repair["oldOptionId"] == "opt_std_aaaaaaaaaaaaaaaa"
    assert id_repair["proposedOptionId"] == "opt_001"
    assert {(ref["sheet"], ref["column"]) for ref in id_repair["cascade"]} == {
        ("grand_sport_x_ovs", "option_id"),
        ("default_selection_rules", "target_option_id"),
    }
    assert gsx["sectionReconciliation"][0]["landedSectionId"] == "sec_stan_001"
    assert gsx["sectionReconciliation"][0]["decidedSectionId"] == "sec_safe_001"
    assert gsx["sectionReconciliationCheck"] == {
        "decisionCount": 1,
        "matchCount": 0,
        "mismatchCount": 1,
        "missingLandedRowCount": 0,
    }
    assert "preIntegrationUsage" not in gsx["sources"]
    zr1_source = reports["zr1"]["sources"]
    assert zr1_source["preIntegrationUsage"] == "row_existence_only"
    assert "preIntegrationWorkbookSha256" in zr1_source
    assert "preIntegrationWorkbookPath" not in zr1_source
    gsx_markdown = (output_dir / "grand_sport_x-recovery-projection.md").read_text(encoding="utf-8")
    assert "| Decisions checked | 1 |" in gsx_markdown
    assert "| Landed matches | 0 |" in gsx_markdown
    assert gsx["summary"]["pendingReviewCount"] > 0

    zr1 = reports["zr1"]
    zr1_by_rpo = {row["identity"]["rpo"]: row for row in zr1["residualRows"]}
    assert zr1_by_rpo["TDM"]["after"]["option_name"] == "Teen Driver"
    assert zr1_by_rpo["TDM"]["after"]["description"] == "Configurable driving settings"
    assert zr1_by_rpo["TDM"]["after"]["detail_raw"] == "Teen Driver raw paragraph"
    assert zr1_by_rpo["TDM"]["after"]["display_order"] == 30
    assert zr1_by_rpo["TDM"]["partition"] == "forward_copy_repair"
    assert zr1_by_rpo["NEW"]["after"]["option_name"] == "New Feature"
    assert zr1_by_rpo["NEW"]["after"]["description"] == "Comparator detail"
    assert zr1_by_rpo["NEW"]["after"]["display_order"] == 50
    assert zr1_by_rpo["UNM"]["after"]["option_name"] == "NEW! Unmatched Feature, target detail"
    assert zr1_by_rpo["UNM"]["after"]["description"] == ""
    assert zr1_by_rpo["UNM"]["after"]["detail_raw"] == "Requires package ABC."
    assert zr1_by_rpo["UNM"]["after"]["display_order"] == 60
    assert "identifying_copy_review" in [item["lane"] for item in zr1_by_rpo["UNM"]["reviewItems"]]
    identifying_review = next(
        item
        for item in zr1_by_rpo["UNM"]["reviewItems"]
        if item["lane"] == "identifying_copy_review"
    )
    assert identifying_review["provenance"]["flags"] == ["needs_curated_copy"]
    zr1_blank = next(row for row in zr1["residualRows"] if not row["identity"]["rpo"])
    assert zr1_blank["after"]["option_name"] == "Integral Front and Rear Antenna"
    assert zr1_blank["after"]["description"] == ""
    assert zr1_blank["fieldProvenance"]["option_name"]["reference"]["model"] == "z06"
    assert not any(item["lane"] == "copy_split" for item in zr1_blank["reviewItems"])
    assert zr1["bulkDecisions"]["comparatorCopy"]["reviewIds"]
    assert not zr1["bulkDecisions"]["copySplit"]["reviewIds"]
    zr1_markdown = (output_dir / "zr1-recovery-projection.md").read_text(encoding="utf-8")
    assert "## Bulk decision sets" in zr1_markdown
    assert "| Comparator copy: bulk-safe |" in zr1_markdown
    assert "| Comparator copy: material-disagreement exclusions | 0 |" in zr1_markdown


def test_reviewed_plan_reuse_fails_closed_on_candidate_fingerprint_mismatch(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    decisions = json.loads(paths["reviewed_decisions_path"].read_text(encoding="utf-8"))
    decisions["decisions"][0]["candidateFingerprint"] = "stale"
    _write_json(paths["reviewed_decisions_path"], decisions)

    report = generate_recovery_reports(ProjectionInputs(**paths), tmp_path / "reports")["grand_sport_x"]

    aj7 = next(row for row in report["residualRows"] if row["identity"]["rpo"] == "AJ7")
    assert aj7["partition"] == "reviewed_plan_fingerprint_mismatch"
    assert aj7["after"]["option_name"] == "Airbags, frontal and side-impact"
    assert aj7["after"]["section_id"] == "sec_stan_001"
    assert aj7["reviewItems"][0]["lane"] == "full_review"
    assert "candidate_fingerprint_mismatch" in aj7["reviewItems"][0]["provenance"]["reasons"][0]


def test_material_comparator_copy_disagreement_is_not_bulk_eligible(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    wb = load_workbook(paths["workbook_path"])
    ws = wb["z06_options"]
    ws.cell(2, OPTION_HEADERS.index("option_name") + 1).value = "Carbon Fiber Wheels"
    ws.cell(2, OPTION_HEADERS.index("description") + 1).value = "Visible carbon-fiber wheel design"
    wb.save(paths["workbook_path"])
    wb.close()

    report = generate_recovery_reports(ProjectionInputs(**paths), tmp_path / "reports")["zr1"]

    tdm = next(row for row in report["residualRows"] if row["identity"]["rpo"] == "TDM")
    copy_review = next(item for item in tdm["reviewItems"] if item["lane"] == "comparator_copy_material_disagreement")
    assert copy_review["provenance"]["comparison"]["materialDisagreement"] is True
    assert copy_review["provenance"]["bulkEligible"] is False
    assert copy_review["reviewId"] not in report["bulkDecisions"]["comparatorCopy"]["reviewIds"]
    assert copy_review["reviewId"] in report["bulkDecisions"]["comparatorCopy"]["excludedMaterialDisagreementReviewIds"]


def test_comparator_copy_uses_complete_target_text_for_disagreement_check(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    wb = load_workbook(paths["workbook_path"])
    ws = wb["zr1_options"]
    ws.cell(2, OPTION_HEADERS.index("detail_raw") + 1).value = "Restrictions and availability only."
    wb.save(paths["workbook_path"])
    wb.close()

    report = generate_recovery_reports(ProjectionInputs(**paths), tmp_path / "reports")["zr1"]

    tdm = next(row for row in report["residualRows"] if row["identity"]["rpo"] == "TDM")
    copy_review = next(item for item in tdm["reviewItems"] if item["lane"] == "comparator_copy")
    assert copy_review["provenance"]["comparison"]["materialDisagreement"] is False
    assert copy_review["reviewId"] in report["bulkDecisions"]["comparatorCopy"]["reviewIds"]


def test_checkpoint_1_bulk_approval_partitions_every_review_item_and_writes_packet(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    report_dir = tmp_path / "reports"
    reports = generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    workbook_before = paths["workbook_path"].read_bytes()

    approval, packet = generate_checkpoint_1_packet(
        report_dir,
        paths["workbook_path"],
        reviewer="Sean",
        reviewed_at="2026-07-20T16:00:00-04:00",
    )

    all_review_ids = {
        item["reviewId"]
        for report in reports.values()
        for row in report["residualRows"]
        for item in row["reviewItems"]
    }
    approved_ids = set(approval["approvedReviewIds"])
    pending_ids = {
        review_id
        for group in packet["exceptionGroups"]
        for review_id in group["reviewIds"]
    }
    assert paths["workbook_path"].read_bytes() == workbook_before
    assert approved_ids.isdisjoint(pending_ids)
    assert approved_ids | pending_ids == all_review_ids
    assert approval["schemaVersion"] == "options-recovery-checkpoint-1-approval-1"
    assert approval["status"] == "bulk_approved_exceptions_pending"
    assert approval["reviewer"] == "Sean"
    assert approval["reviewedAt"] == "2026-07-20T16:00:00-04:00"
    assert approval["approvedReviewCount"] == len(approved_ids)
    assert packet["pendingReviewCount"] == len(pending_ids)
    assert packet["bulkApprovalFingerprint"] == approval["approvalFingerprint"]
    full_review_group = next(
        group for group in packet["exceptionGroups"] if group["lane"] == "full_review"
    )
    assert set(full_review_group["proposedState"]) == {
        "option_name",
        "description",
        "section_id",
        "active",
        "selectable",
        "display_order",
    }
    assert not any(group["lane"] == "no_rpo_mapping" for group in packet["exceptionGroups"])
    assert sorted(path.name for path in report_dir.glob("checkpoint-1-*")) == [
        "checkpoint-1-bulk-approval.json",
        "checkpoint-1-exception-review.json",
        "checkpoint-1-exception-review.md",
    ]
    markdown = (report_dir / "checkpoint-1-exception-review.md").read_text(encoding="utf-8")
    assert "# Checkpoint 1 Exception-Only Review" in markdown
    assert "Bulk-approved review records" in markdown


def test_checkpoint_1_keeps_unresolved_identifying_copy_in_one_shared_zr_choice(
    tmp_path: Path,
) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    for sheet_name in ("zr1_options", "zr1x_options"):
        wb = load_workbook(paths["workbook_path"])
        ws = wb[sheet_name]
        target_row = 4 if sheet_name == "zr1_options" else 2
        ws.cell(target_row, OPTION_HEADERS.index("rpo") + 1).value = "WHE"
        ws.cell(target_row, OPTION_HEADERS.index("option_name") + 1).value = "Wheels, forged aluminum"
        ws.cell(target_row, OPTION_HEADERS.index("description") + 1).value = ""
        ws.cell(target_row, OPTION_HEADERS.index("detail_raw") + 1).value = "Restrictions only."
        wb.save(paths["workbook_path"])
        wb.close()

    report_dir = tmp_path / "reports"
    reports = generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    approval, packet = generate_checkpoint_1_packet(
        report_dir,
        paths["workbook_path"],
        reviewer="Sean",
        reviewed_at="2026-07-20T16:00:00-04:00",
    )

    flagged_ids = {
        item["reviewId"]
        for report in reports.values()
        for row in report["residualRows"]
        for item in row["reviewItems"]
        if row["identity"]["rpo"] == "WHE"
        and item["lane"] == "identifying_copy_review"
        and item["provenance"]["flags"]
    }
    flagged_groups = [
        group
        for group in packet["exceptionGroups"]
        if flagged_ids & set(group["reviewIds"])
    ]
    assert flagged_ids
    assert flagged_ids.isdisjoint(approval["approvedReviewIds"])
    assert len(flagged_groups) == 1
    assert flagged_groups[0]["models"] == ["zr1", "zr1x"]
    assert flagged_groups[0]["flags"] == ["needs_curated_copy"]


def test_checkpoint_1_bulk_approval_rejects_tampered_report(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
    )

    paths = _projection_fixture(tmp_path)
    report_dir = tmp_path / "reports"
    generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    report_path = report_dir / "zr1-recovery-projection.json"
    report = json.loads(report_path.read_text(encoding="utf-8"))
    report["summary"]["pendingReviewCount"] += 1
    _write_json(report_path, report)

    with pytest.raises(ValueError, match="report fingerprint"):
        generate_checkpoint_1_packet(
            report_dir,
            paths["workbook_path"],
            reviewer="Sean",
            reviewed_at="2026-07-20T16:00:00-04:00",
        )
    assert not (report_dir / "checkpoint-1-bulk-approval.json").exists()


def test_checkpoint_1_bulk_approval_cli_uses_existing_reports_without_regenerating(
    tmp_path: Path,
) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_recovery_reports,
        main,
    )

    paths = _projection_fixture(tmp_path)
    report_dir = tmp_path / "reports"
    generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    report_hashes_before = {
        path.name: path.read_bytes()
        for path in report_dir.glob("*-recovery-projection.json")
    }

    result = main(
        [
            "--root",
            str(tmp_path),
            "--workbook",
            str(paths["workbook_path"]),
            "--output-dir",
            str(report_dir),
            "--approve-checkpoint-1",
            "--reviewer",
            "Sean",
            "--reviewed-at",
            "2026-07-20T16:00:00-04:00",
        ]
    )

    assert result == 0
    assert {
        path.name: path.read_bytes()
        for path in report_dir.glob("*-recovery-projection.json")
    } == report_hashes_before
    assert (report_dir / "checkpoint-1-bulk-approval.json").exists()


def test_checkpoint_1_decisions_record_deletes_cascades_and_emit_full_pending_review(
    tmp_path: Path,
) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
        record_checkpoint_1_decisions,
    )

    paths = _projection_fixture(tmp_path)
    wb = load_workbook(paths["workbook_path"])
    wb["grand_sport_x_ovs"].append(["opt_n26_001", "1lt_g07", "available"])
    _sheet(
        wb,
        "grand_sport_x_price_rules",
        ["rule_id", "condition_option_id", "target_option_id", "price"],
        [["price_n26", "opt_n26_001", "opt_n26_001", 0]],
    )
    long_text = "Air filtration system, " + ("complete unabridged source detail " * 8).strip()
    ws = wb["grand_sport_x_options"]
    ws.cell(4, OPTION_HEADERS.index("description") + 1).value = long_text
    ws.cell(4, OPTION_HEADERS.index("detail_raw") + 1).value = long_text
    wb.save(paths["workbook_path"])
    wb.close()

    report_dir = tmp_path / "reports"
    generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    _, packet = generate_checkpoint_1_packet(
        report_dir,
        paths["workbook_path"],
        reviewer="Sean",
        reviewed_at="2026-07-20T16:00:00-04:00",
    )
    delete_group = next(
        group
        for group in packet["exceptionGroups"]
        if group["lane"] == "full_review" and group["identityKey"] == "N26"
    )
    decisions = []
    for group in packet["exceptionGroups"]:
        if group["lane"] == "active":
            decisions.append({"decisionGroupId": group["decisionGroupId"], "action": "accept"})
        elif group["decisionGroupId"] == delete_group["decisionGroupId"]:
            decisions.append(
                {
                    "decisionGroupId": group["decisionGroupId"],
                    "action": "delete_option_and_owned_references",
                }
            )
        elif group["lane"] == "display_order":
            action = (
                "not_applicable_due_to_delete"
                if group["identityKey"] == "N26"
                else "accept"
            )
            decisions.append({"decisionGroupId": group["decisionGroupId"], "action": action})

    decision_artifact, pending = record_checkpoint_1_decisions(
        report_dir,
        paths["workbook_path"],
        decisions=decisions,
        reviewer="Sean",
        reviewed_at="2026-07-20T17:00:00-04:00",
    )

    recorded_delete = next(
        decision
        for decision in decision_artifact["decisions"]
        if decision["decisionGroupId"] == delete_group["decisionGroupId"]
    )
    assert recorded_delete["action"] == "delete_option_and_owned_references"
    assert recorded_delete["deleteScope"]["optionRows"] == [
        {
            "sheet": "grand_sport_x_options",
            "rowNumber": 5,
            "optionId": "opt_n26_001",
            "rpo": "N26",
        }
    ]
    assert {
        (row["sheet"], row["rowNumber"])
        for row in recorded_delete["deleteScope"]["ownedReferenceRows"]
    } == {
        ("grand_sport_x_ovs", 3),
        ("grand_sport_x_price_rules", 2),
    }
    assert decision_artifact["resolvedGroupCount"] == len(decisions)
    assert pending["pendingGroupCount"] == packet["exceptionGroupCount"] - len(decisions)
    assert pending["sourcePacketFingerprint"] == packet["packetFingerprint"]
    pending_markdown = (report_dir / "checkpoint-1-pending-review.md").read_text(encoding="utf-8")
    assert long_text in pending_markdown
    assert "..." not in pending_markdown
    assert not any(
        group["identityKey"] == "N26" and group["lane"] in {"full_review", "display_order"}
        for group in pending["exceptionGroups"]
    )


def test_checkpoint_1_decisions_support_shared_zr_deletes_and_model_scoped_overrides(
    tmp_path: Path,
) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
        record_checkpoint_1_decisions,
    )

    paths = _projection_fixture(tmp_path)
    for sheet_name, target_row in (("zr1_options", 4), ("zr1x_options", 2)):
        wb = load_workbook(paths["workbook_path"])
        ws = wb[sheet_name]
        ws.cell(target_row, OPTION_HEADERS.index("rpo") + 1).value = "WHE"
        ws.cell(target_row, OPTION_HEADERS.index("option_id") + 1).value = "opt_whe_001"
        ws.cell(target_row, OPTION_HEADERS.index("option_name") + 1).value = "Wheels, forged aluminum"
        ws.cell(target_row, OPTION_HEADERS.index("description") + 1).value = ""
        ws.cell(target_row, OPTION_HEADERS.index("detail_raw") + 1).value = "Restrictions only."
        wb.save(paths["workbook_path"])
        wb.close()

    report_dir = tmp_path / "reports"
    generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    _, packet = generate_checkpoint_1_packet(
        report_dir,
        paths["workbook_path"],
        reviewer="Sean",
        reviewed_at="2026-07-20T16:00:00-04:00",
    )
    group = next(
        group
        for group in packet["exceptionGroups"]
        if group["identityKey"] == "WHE" and group["lane"] == "identifying_copy_review"
    )

    delete_artifact, _ = record_checkpoint_1_decisions(
        report_dir,
        paths["workbook_path"],
        decisions=[
            {
                "decisionGroupId": group["decisionGroupId"],
                "action": "delete_option_and_owned_references",
            }
        ],
        reviewer="Sean",
        reviewed_at="2026-07-20T17:00:00-04:00",
    )
    assert delete_artifact["decisions"][0]["deleteScope"]["optionRows"] == [
        {
            "sheet": "zr1_options",
            "rowNumber": 4,
            "optionId": "opt_whe_001",
            "rpo": "WHE",
        },
        {
            "sheet": "zr1x_options",
            "rowNumber": 2,
            "optionId": "opt_whe_001",
            "rpo": "WHE",
        },
    ]

    overrides = {
        "zr1": {"option_name": "ZR1 Wheels", "description": "ZR1 description"},
        "zr1x": {"option_name": "ZR1X Wheels", "description": "ZR1X description"},
    }
    override_artifact, _ = record_checkpoint_1_decisions(
        report_dir,
        paths["workbook_path"],
        decisions=[
            {
                "decisionGroupId": group["decisionGroupId"],
                "action": "override_by_model",
                "overrideByModel": overrides,
            }
        ],
        reviewer="Sean",
        reviewed_at="2026-07-20T17:00:00-04:00",
    )
    assert override_artifact["decisions"][0]["overrideByModel"] == overrides


def test_checkpoint_1_decisions_can_override_an_explicit_bulk_approved_price(
    tmp_path: Path,
) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
        record_checkpoint_1_decisions,
    )

    paths = _projection_fixture(tmp_path)
    wb = load_workbook(paths["workbook_path"])
    ws = wb["zr1_options"]
    ws.cell(2, OPTION_HEADERS.index("price") + 1).value = 1495
    wb.save(paths["workbook_path"])
    wb.close()

    report_dir = tmp_path / "reports"
    generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    generate_checkpoint_1_packet(
        report_dir,
        paths["workbook_path"],
        reviewer="Sean",
        reviewed_at="2026-07-20T16:00:00-04:00",
    )

    artifact, _ = record_checkpoint_1_decisions(
        report_dir,
        paths["workbook_path"],
        decisions=[],
        bulk_overrides=[
            {
                "model": "zr1",
                "rpo": "TDM",
                "lane": "price",
                "override": {"price": None},
            }
        ],
        reviewer="Sean",
        reviewed_at="2026-07-20T17:00:00-04:00",
    )

    assert artifact["bulkApprovalOverrideCount"] == 1
    assert artifact["bulkApprovalOverrides"][0]["model"] == "zr1"
    assert artifact["bulkApprovalOverrides"][0]["identity"]["rpo"] == "TDM"
    assert artifact["bulkApprovalOverrides"][0]["override"] == {"price": None}


def test_checkpoint_1_decisions_reject_tampered_exception_packet(tmp_path: Path) -> None:
    from corvette_form_generator.ingest.options_recovery_projection import (
        ProjectionInputs,
        generate_checkpoint_1_packet,
        generate_recovery_reports,
        record_checkpoint_1_decisions,
    )

    paths = _projection_fixture(tmp_path)
    report_dir = tmp_path / "reports"
    generate_recovery_reports(ProjectionInputs(**paths), report_dir)
    _, packet = generate_checkpoint_1_packet(
        report_dir,
        paths["workbook_path"],
        reviewer="Sean",
        reviewed_at="2026-07-20T16:00:00-04:00",
    )
    packet_path = report_dir / "checkpoint-1-exception-review.json"
    tampered = json.loads(packet_path.read_text(encoding="utf-8"))
    tampered["pendingReviewCount"] += 1
    _write_json(packet_path, tampered)

    with pytest.raises(ValueError, match="exception packet fingerprint"):
        record_checkpoint_1_decisions(
            report_dir,
            paths["workbook_path"],
            decisions=[
                {
                    "decisionGroupId": packet["exceptionGroups"][0]["decisionGroupId"],
                    "action": "accept",
                }
            ],
            reviewer="Sean",
            reviewed_at="2026-07-20T17:00:00-04:00",
        )
    assert not (report_dir / "checkpoint-1-exception-decisions.json").exists()
