#!/usr/bin/env python3
"""Checkpoint 2A/2B RED contracts: workbook-owned group display_label.

Covers the shared registry contract (registry.py), the workbook schema
validator's label gate, and the generator's group loaders. The schema tests
reuse the registry-shaped fixture from test_schema_validation_metadata.py so a
drift in that fixture cannot mask drift here.
"""

from __future__ import annotations

import csv
import json
import runpy
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from corvette_form_generator.model_config import ModelConfig  # noqa: E402
from corvette_form_generator.rules import (  # noqa: E402
    load_exclusive_groups,
    load_rule_groups,
)
from corvette_form_generator.workbook_domain import registry  # noqa: E402

sys.path.insert(0, "tests")

GROUP_FAMILIES = ("exclusive_groups", "rule_groups")
ROOT = Path(__file__).resolve().parents[1]
REVIEW_DIR = ROOT / "workbook-manager" / "review"
# Spec §7.2 placeholders that an approved label must never equal.
PLACEHOLDER_LABELS = tuple(sorted(registry.GROUP_DISPLAY_LABEL_PLACEHOLDERS))


def _registry_columns(family: str) -> list[str]:
    return list(registry.WRITABLE_COLUMNS[family])


def _z06_model_config() -> ModelConfig:
    """A z06 config resolved against a real canonical-workbook openpyxl load."""

    from openpyxl import load_workbook

    from corvette_form_generator.model_configs import base_model_config
    from corvette_form_generator.runtime_metadata import load_model_config_overrides

    wb = load_workbook("stingray_master.xlsx", read_only=True, data_only=True)
    try:
        return load_model_config_overrides(wb, base_model_config("z06"))
    finally:
        wb.close()


def _minimal_group_workbook():
    """The shared registry-shaped fixture plus one z06 group row per family."""

    from test_schema_validation_metadata import registry_shaped_workbook

    wb = registry_shaped_workbook()
    for sheet, values in (
        ("z06_rule_groups", {"group_id": "grp_r1", "display_label": "", "group_type": "requires_any",
                             "source_id": "opt_z06", "active": True}),
        ("z06_exclusive_groups", {"group_id": "grp_e1", "display_label": "",
                                  "selection_mode": "single_within_group", "active": True}),
    ):
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        ws.append([values.get(header) for header in headers])
    return wb


def _validate(wb):
    from test_schema_validation_metadata import validate_temp_workbook

    return validate_temp_workbook(wb)


class GroupDisplayLabelRegistryTests(unittest.TestCase):
    """The shared registry owns the display_label column for both group families."""

    def test_registry_declares_display_label_after_group_id(self) -> None:
        for family in GROUP_FAMILIES:
            columns = registry.WRITABLE_COLUMNS[family]
            self.assertIn("display_label", columns, family)
            self.assertEqual(
                columns.index("display_label"),
                columns.index("group_id") + 1,
                f"{family}: display_label must sit immediately after group_id",
            )

    def test_display_label_is_optional_not_required_on_add(self) -> None:
        for family in GROUP_FAMILIES:
            meta = registry.EDITOR_SHEET_META[family]
            self.assertIn("display_label", registry.OPTIONAL_COLUMNS[family], family)
            self.assertIn("display_label", meta["optional_columns"], family)
            self.assertNotIn("display_label", meta["required_on_add"], family)

    def test_active_exclusive_group_requires_a_label_on_the_write_path(self) -> None:
        """§7.1: the editor cannot clear the label of an active group.

        `display_label` stays optional overall -- blank means "label pending"
        on an inactive row -- so the requirement is carried by
        `required_on_effective_active_row`, which `editor_ops` unions in when
        the effective row is active.
        """
        self.assertIn(
            "display_label",
            registry.EDITOR_SHEET_META["exclusive_groups"][
                "required_on_effective_active_row"
            ],
        )
        self.assertNotIn(
            "display_label",
            registry.EDITOR_SHEET_META["rule_groups"][
                "required_on_effective_active_row"
            ],
            "rule-group labels stay Manager-facing per §7.1",
        )


class GroupDisplayLabelSchemaValidationTests(unittest.TestCase):
    """§7.2 label rules enforced by validate_group_display_labels()."""

    def test_registry_shaped_group_sheets_accept_display_label(self) -> None:
        wb = _minimal_group_workbook()
        for sheet, label in (("z06_rule_groups", "Required wheels"), ("z06_exclusive_groups", "Wheel selection")):
            headers = [c.value for c in wb[sheet][1]]
            last = wb[sheet].max_row
            wb[sheet].cell(row=last, column=headers.index("display_label") + 1).value = label
        issues = _validate(wb)
        self.assertFalse(
            [i for i in issues if i.severity == "error"],
            [(i.check_id, i.message) for i in issues if i.severity == "error"],
        )

    def test_blank_label_on_active_exclusive_group_is_rejected(self) -> None:
        """The migration is complete, so blank is no longer a free pass.

        Pre-migration sheets are distinguished by the column being absent,
        not by the value being blank.
        """
        wb = _minimal_group_workbook()
        headers = [c.value for c in wb["z06_rule_groups"][1]]
        wb["z06_rule_groups"].cell(
            row=wb["z06_rule_groups"].max_row,
            column=headers.index("display_label") + 1,
        ).value = "Required wheels"
        issues = _validate(wb)
        missing = [i for i in issues if i.check_id == "group_display_label_missing"]
        self.assertEqual(1, len(missing), [(i.check_id, i.message) for i in issues])
        self.assertEqual("z06_exclusive_groups", missing[0].sheet)
        self.assertEqual("error", missing[0].severity)

    def test_blank_label_on_inactive_exclusive_group_is_allowed(self) -> None:
        wb = _minimal_group_workbook()
        sheet = "z06_exclusive_groups"
        headers = [c.value for c in wb[sheet][1]]
        wb[sheet].cell(
            row=wb[sheet].max_row, column=headers.index("active") + 1
        ).value = False
        rg_headers = [c.value for c in wb["z06_rule_groups"][1]]
        wb["z06_rule_groups"].cell(
            row=wb["z06_rule_groups"].max_row,
            column=rg_headers.index("display_label") + 1,
        ).value = "Required wheels"
        issues = _validate(wb)
        self.assertEqual(
            [], [i for i in issues if i.check_id == "group_display_label_missing"]
        )

    def test_blank_label_on_active_rule_group_is_allowed(self) -> None:
        """Rule-group labels are Manager-facing per §7.1, not customer copy."""
        wb = _minimal_group_workbook()
        sheet = "z06_exclusive_groups"
        headers = [c.value for c in wb[sheet][1]]
        wb[sheet].cell(
            row=wb[sheet].max_row, column=headers.index("display_label") + 1
        ).value = "Wheel selection"
        issues = _validate(wb)
        self.assertEqual(
            [], [i for i in issues if i.check_id == "group_display_label_missing"]
        )

    def test_display_label_must_not_equal_placeholder(self) -> None:
        wb = _minimal_group_workbook()
        sheet = "z06_exclusive_groups"
        headers = [c.value for c in wb[sheet][1]]
        col = headers.index("display_label") + 1
        for placeholder in PLACEHOLDER_LABELS:
            wb[sheet].cell(row=wb[sheet].max_row, column=col).value = placeholder
            issues = _validate(wb)
            self.assertTrue(
                any(i.check_id == "group_display_label_invalid" for i in issues),
                f"placeholder {placeholder!r} not rejected",
            )
            wb[sheet].cell(row=wb[sheet].max_row, column=col).value = None

    def test_display_label_must_not_equal_group_id(self) -> None:
        wb = _minimal_group_workbook()
        sheet = "z06_exclusive_groups"
        headers = [c.value for c in wb[sheet][1]]
        col = headers.index("display_label") + 1
        wb[sheet].cell(row=wb[sheet].max_row, column=col).value = "grp_e1"
        issues = _validate(wb)
        self.assertTrue(any(i.check_id == "group_display_label_invalid" for i in issues), issues)

    def test_display_label_must_not_copy_terminal_hash_token(self) -> None:
        wb = _minimal_group_workbook()
        sheet = "z06_exclusive_groups"
        headers = [c.value for c in wb[sheet][1]]
        row = wb[sheet].max_row
        wb[sheet].cell(row=row, column=headers.index("group_id") + 1).value = (
            "z06_excl_engine_covers_1623e1da9d59"
        )
        wb[sheet].cell(row=row, column=headers.index("display_label") + 1).value = (
            "Engine Covers 1623e1da9d59"
        )

        issues = _validate(wb)

        self.assertTrue(any(i.check_id == "group_display_label_invalid" for i in issues), issues)

    def test_display_label_rejects_untrimmed_multiline_and_bad_length(self) -> None:
        wb = _minimal_group_workbook()
        sheet = "z06_exclusive_groups"
        headers = [c.value for c in wb[sheet][1]]
        col = headers.index("display_label") + 1
        for bad in (" padded", "two\nlines", "ab", "x" * 101):
            wb[sheet].cell(row=wb[sheet].max_row, column=col).value = bad
            issues = _validate(wb)
            self.assertTrue(
                any(i.check_id == "group_display_label_invalid" for i in issues),
                f"value {bad!r} not rejected",
            )
            wb[sheet].cell(row=wb[sheet].max_row, column=col).value = None

    def test_valid_display_label_passes_all_label_checks(self) -> None:
        wb = _minimal_group_workbook()
        sheet = "z06_exclusive_groups"
        headers = [c.value for c in wb[sheet][1]]
        col = headers.index("display_label") + 1
        wb[sheet].cell(row=wb[sheet].max_row, column=col).value = "Wheel selection"
        issues = _validate(wb)
        self.assertFalse(any(i.check_id == "group_display_label_invalid" for i in issues), issues)


class GroupDisplayLabelGeneratorTests(unittest.TestCase):
    """Generator group loaders carry display_label into their group dicts."""

    def setUp(self) -> None:
        self.config = _z06_model_config()

    def test_load_rule_groups_carries_display_label(self) -> None:
        groups = load_rule_groups(_minimal_group_workbook(), self.config)
        matching = [g for g in groups if g.get("group_id") == "grp_r1"]
        self.assertEqual(len(matching), 1, groups)
        self.assertEqual(matching[0].get("display_label"), "")

    def test_load_exclusive_groups_carries_display_label(self) -> None:
        groups = load_exclusive_groups(_minimal_group_workbook(), self.config)
        matching = [g for g in groups if g.get("group_id") == "grp_e1"]
        self.assertEqual(len(matching), 1, groups)
        self.assertEqual(matching[0].get("display_label"), "")

    def test_missing_label_loads_as_blank_not_invented(self) -> None:
        wb = _minimal_group_workbook()
        headers = [c.value for c in wb["z06_rule_groups"][1]]
        col = headers.index("display_label") + 1
        wb["z06_rule_groups"].cell(row=wb["z06_rule_groups"].max_row, column=col).value = "Required wheels"
        groups = load_rule_groups(wb, self.config)
        matching = [g for g in groups if g.get("group_id") == "grp_r1"]
        self.assertEqual(len(matching), 1, groups)
        self.assertEqual(matching[0].get("display_label"), "Required wheels")


class GroupDisplayLabelReviewToolingTests(unittest.TestCase):
    """§7.3 review tooling: accurate evidence, immutable evidence fields."""

    @staticmethod
    def _generator():
        import importlib.util

        path = ROOT / "workbook-manager" / "review" / "generate_group_display_label_review.py"
        spec = importlib.util.spec_from_file_location("_gen_review", path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    @staticmethod
    def _syncer():
        import importlib.util

        path = ROOT / "workbook-manager" / "review" / "sync_group_display_label_review.py"
        spec = importlib.util.spec_from_file_location("_sync_review", path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def test_hash_suffixed_group_reports_the_manager_section_fallback(self) -> None:
        """The Manager renders `Exclusive group . <section>`, not the placeholder."""
        gen = self._generator()
        label = gen._exclusive_fallback_label(
            "z06_excl_engine_covers_1623e1da9d59",
            ["opt_a", "opt_b"],
            {"opt_a": "sec_engi_001", "opt_b": "sec_engi_001"},
            {"sec_engi_001": "Engine Appearance"},
        )
        self.assertEqual("Exclusive group \u00b7 Engine Appearance", label)
        self.assertNotEqual(gen.PLACEHOLDER_FALLBACK, label)

    def test_hash_suffixed_group_spanning_sections_keeps_the_placeholder(self) -> None:
        gen = self._generator()
        label = gen._exclusive_fallback_label(
            "z06_excl_mixed_1623e1da9d59",
            ["opt_a", "opt_b"],
            {"opt_a": "sec_engi_001", "opt_b": "sec_whee_001"},
            {"sec_engi_001": "Engine Appearance", "sec_whee_001": "Wheels"},
        )
        self.assertEqual(gen.PLACEHOLDER_FALLBACK, label)

    def test_unhashed_group_still_uses_the_humanized_id(self) -> None:
        gen = self._generator()
        self.assertEqual(
            "Excl Engine Covers",
            gen._exclusive_fallback_label("excl_engine_covers", [], {}, {}),
        )

    def test_sync_rejects_edits_to_generated_evidence_fields(self) -> None:
        sync = self._syncer()
        record = {
            "model_key": "z06", "group_type": "exclusive", "group_id": "g1",
            "notes": "generated note", "active": True, "member_count": 3,
        }
        row = dict(
            model_key="z06", group_type="exclusive", group_id="g1",
            notes="hand edited", active="TRUE", member_count="3",
        )
        with self.assertRaises(ValueError) as caught:
            sync.assert_evidence_unchanged(row, record)
        self.assertIn("notes", str(caught.exception))

    def test_sync_tolerates_spreadsheet_boolean_case(self) -> None:
        """A spreadsheet rewrites `true` as `TRUE`; that is not an edit."""
        sync = self._syncer()
        record = {"model_key": "z06", "group_type": "exclusive", "group_id": "g1",
                  "active": True}
        row = {"model_key": "z06", "group_type": "exclusive", "group_id": "g1",
               "active": "TRUE"}
        sync.assert_evidence_unchanged(row, record)

    def test_sync_allows_decision_field_edits(self) -> None:
        sync = self._syncer()
        record = {"model_key": "z06", "group_type": "exclusive", "group_id": "g1",
                  "notes": "n", "proposed_display_label": "", "review_status": "pending",
                  "reviewer_note": "", "audience": "customer", "customer_visible": True}
        row = {"model_key": "z06", "group_type": "exclusive", "group_id": "g1",
               "notes": "n", "proposed_display_label": "Engine Appearance",
               "review_status": "approved", "reviewer_note": "ok",
               "audience": "customer", "customer_visible": "FALSE"}
        sync.assert_evidence_unchanged(row, record)


class GroupDisplayLabelReviewArtifactTests(unittest.TestCase):
    def test_csv_and_json_are_exact_approved_decision_companions(self) -> None:
        csv_path = REVIEW_DIR / "group-display-label-review.csv"
        self.assertNotIn(b"\r\n", csv_path.read_bytes())
        with csv_path.open(
            newline="", encoding="utf-8-sig"
        ) as fh:
            csv_rows = list(csv.DictReader(fh))
        payload = json.loads(
            (REVIEW_DIR / "group-display-label-review.json").read_text(encoding="utf-8")
        )
        json_rows = payload["records"]

        def identity(row: dict) -> tuple[str, str, str]:
            return row["model_key"], row["group_type"], row["group_id"]

        def stable_key(row: dict) -> tuple[str, str, str, int, str]:
            sheet, row_number = row["source_sheet_row"].rsplit("!", 1)
            return row["model_key"], row["group_type"], sheet, int(row_number), row["group_id"]

        self.assertEqual(payload["record_count"], len(csv_rows))
        self.assertEqual([identity(row) for row in csv_rows], [identity(row) for row in json_rows])
        self.assertEqual(csv_rows, sorted(csv_rows, key=stable_key))
        self.assertEqual({row["review_status"] for row in csv_rows}, {"approved"})
        self.assertTrue(
            all(
                registry.GROUP_DISPLAY_LABEL_MIN_LENGTH
                <= len(row["proposed_display_label"])
                <= registry.GROUP_DISPLAY_LABEL_MAX_LENGTH
                for row in csv_rows
            )
        )
        for csv_row, json_row in zip(csv_rows, json_rows, strict=True):
            for field in (
                "proposed_display_label",
                "review_status",
                "reviewer_note",
                "audience",
            ):
                self.assertEqual(csv_row[field], json_row[field], (identity(csv_row), field))

    def test_approved_review_labels_match_canonical_workbook(self) -> None:
        from openpyxl import load_workbook

        with (REVIEW_DIR / "group-display-label-review.csv").open(
            newline="", encoding="utf-8-sig"
        ) as fh:
            rows = list(csv.DictReader(fh))
        wb = load_workbook(ROOT / "stingray_master.xlsx", read_only=True, data_only=True)
        try:
            for row in rows:
                sheet_name, row_number = row["source_sheet_row"].rsplit("!", 1)
                ws = wb[sheet_name]
                headers = [str(cell.value or "") for cell in ws[1]]
                self.assertEqual(
                    headers.index("display_label"),
                    headers.index("group_id") + 1,
                    sheet_name,
                )
                actual = ws.cell(
                    row=int(row_number),
                    column=headers.index("display_label") + 1,
                ).value
                self.assertEqual(
                    actual,
                    row["proposed_display_label"],
                    (row["model_key"], row["group_type"], row["group_id"]),
                )
        finally:
            wb.close()

    def test_review_generator_refuses_to_overwrite_approved_decisions(self) -> None:
        script_path = REVIEW_DIR / "generate_group_display_label_review.py"
        module_globals = runpy.run_path(str(script_path), run_name="group_label_review_generator")

        with tempfile.TemporaryDirectory() as tmp:
            review_dir = Path(tmp)
            paths = []
            for name in ("group-display-label-review.csv", "group-display-label-review.json"):
                source = REVIEW_DIR / name
                target = review_dir / name
                shutil.copy2(source, target)
                paths.append(target)
            before = {path.name: path.read_bytes() for path in paths}
            module_globals["REVIEW_DIR"] = review_dir

            with self.assertRaisesRegex(RuntimeError, "reviewed decisions"):
                module_globals["main"]()

            self.assertEqual({path.name: path.read_bytes() for path in paths}, before)


class Checkpoint5OperatorTerminologyTests(unittest.TestCase):
    """§12 operator actions use outcomes, not lifecycle implementation terms."""

    def test_review_and_apply_actions_use_the_approved_operator_labels(self) -> None:
        source = (
            ROOT
            / "workbook-manager"
            / "frontend"
            / "src"
            / "components"
            / "ChangesSync.jsx"
        ).read_text(encoding="utf-8").replace("&amp;", "&")
        for label in (
            "Lock Draft for Validation",
            "Validate Draft Against Workbook",
            "Approve Validated Changes",
            "Write Approved Changes & Rebuild Form Data",
            "Cancel Draft and Keep Audit Record",
            "Reload Latest Workbook Data",
            "Export Workbook Review Copy",
            "Back Up Drafts & History",
            "Refresh Screen Status",
        ):
            self.assertIn(label, source)
        for stale in (
            "> Freeze ChangeSet",
            '"Run Workbook Preview"',
            "> Approve Exact Preview",
            '"Apply and Rebuild" : "Retry Apply and Rebuild"',
            "> Re-Import Workbook",
            "> Export Disposable Comparison",
            "> Backup Manager State",
        ):
            self.assertNotIn(stale, source)
        self.assertNotRegex(source, r">\s*Cancel Draft\s*<")
        self.assertNotRegex(source, r">\s*Refresh\s*<")

    def test_manager_uses_review_and_apply_and_names_the_image_inventory(self) -> None:
        source_root = ROOT / "workbook-manager" / "frontend" / "src"
        components = (
            "AssetManager.jsx",
            "ConnectedExplorer.jsx",
            "ModelOperations.jsx",
            "OptionEditor.jsx",
        )
        combined = "\n".join(
            (source_root / "components" / name).read_text(encoding="utf-8")
            for name in components
        )
        self.assertNotIn("Draft Review", combined)
        self.assertNotIn("until Apply and Rebuild", combined)
        self.assertIn("Review & Apply", combined.replace("&amp;", "&"))
        asset_source = (source_root / "components" / "AssetManager.jsx").read_text(
            encoding="utf-8"
        )
        self.assertIn("Refresh WordPress Image Inventory", asset_source)
        self.assertNotIn("> Refresh inventory", asset_source)


if __name__ == "__main__":
    unittest.main()
