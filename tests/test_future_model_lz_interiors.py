#!/usr/bin/env python3
"""Tests for dry-run future Z-family LZ interior readiness previews."""

from __future__ import annotations

import json
import subprocess
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_lz_interiors import (  # noqa: E402
    apply_lz_interiors_to_workbook,
    build_lz_interiors_preview,
)

LZ_HEADERS = (
    "interior_id",
    "Interior Name",
    "Material",
    "Price",
    "Detail from Disclosure",
    "Color Overrides",
    "Trim",
    "Seat",
    "Interior Code",
    "Suede",
    "Stitch",
    "Two Tone",
    "section_id",
    "active_for_stingray",
    "requires_r6x",
    "included_option_id",
)

PRICE_REF_HEADERS = ("OptionType", "Trim", "Code", "Price")
SECTION_HEADERS = ("section_id", "section_name", "selection_mode", "is_required", "display_order", "standard_behavior", "help_text", "step_key")
SCOPE_HEADERS = ("model_key", "interior_id", "trim_level", "active", "requires_option_id", "notes")
COMPONENT_HEADERS = (
    "model_key",
    "interior_id",
    "rpo",
    "component_type",
    "label",
    "price_ref_type",
    "price_ref_code",
    "price_trim_scope",
    "display_order",
    "active",
    "notes",
)


def append_sheet(wb: Workbook, name: str, headers: tuple[str, ...], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def lz_row(interior_id: str, trim: str, seat: str = "AQ9", **overrides: object) -> dict[str, object]:
    row: dict[str, object] = {header: "" for header in LZ_HEADERS}
    row.update(
        {
            "interior_id": interior_id,
            "Interior Name": interior_id,
            "Material": "Mulan leather",
            "Price": 0,
            "Trim": trim,
            "Seat": seat,
            "Interior Code": "HTA",
            "active_for_stingray": False,
            "requires_r6x": trim.endswith("_R6X"),
        }
    )
    row.update(overrides)
    return row


def preview_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "section_master",
        SECTION_HEADERS,
        [
            {
                "section_id": "sec_intc_001",
                "section_name": "1LT Interior",
                "selection_mode": "single_select_req",
                "is_required": False,
                "display_order": 15,
                "standard_behavior": "locked_included",
                "step_key": "base_interior",
            }
        ],
    )
    append_sheet(
        wb,
        "LZ_Interiors",
        LZ_HEADERS,
        [
            lz_row("1LZ_AQ9_HTA", "1LZ"),
            lz_row("2LZ_AE4_HTA", "2LZ", seat="AE4"),
            lz_row("3LZ_AE4_HTA_N26_36S", "3LZ", seat="AE4", Suede="N26", Stitch="36S"),
            lz_row("3LZ_R6X_AQ9_HTA_R6X", "3LZ_R6X"),
        ],
    )
    append_sheet(
        wb,
        "PriceRef",
        PRICE_REF_HEADERS,
        [
            {"OptionType": "Seat", "Trim": "2LZ", "Code": "AE4", "Price": 2095},
            {"OptionType": "Seat", "Trim": "3LZ", "Code": "AE4", "Price": 500},
            {"OptionType": "Suede", "Code": "N26", "Price": 695},
            {"OptionType": "Stitching", "Code": "36S", "Price": 495},
            {"OptionType": "Seat", "Trim": "3LZ_R6X", "Code": "AQ9", "Price": 9995},
        ],
    )
    append_sheet(
        wb,
        "model_interior_scope",
        SCOPE_HEADERS,
        [{"model_key": "grand_sport", "interior_id": "existing_gs", "trim_level": "1LT", "active": True}],
    )
    append_sheet(
        wb,
        "interior_components",
        COMPONENT_HEADERS,
        [
            {
                "model_key": "stingray",
                "interior_id": "existing_stingray",
                "rpo": "AE4",
                "component_type": "seat",
                "active": True,
            }
        ],
    )
    return wb


def sheet_rows(wb: Workbook, name: str) -> list[dict[str, object]]:
    ws = wb[name]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header: value for header, value in zip(headers, values) if header}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


class FutureModelLzInteriorsTests(unittest.TestCase):
    def test_preview_maps_sections_and_model_trim_scopes_without_mutating_workbook(self) -> None:
        wb = preview_workbook()

        preview = build_lz_interiors_preview(wb, ["all"], include_details=True)

        self.assertEqual(preview["status"], "dry_run")
        self.assertFalse(preview["would_write_workbook"])
        self.assertFalse(preview["would_mutate_generated_runtime_data"])
        self.assertEqual(preview["source_row_count"], 4)
        self.assertEqual(preview["current_section_counts"], {"": 4})
        self.assertEqual(preview["proposed_section_counts"], {"sec_lzint_001": 1, "sec_lzint_002": 1, "sec_lzint_003": 2})
        self.assertEqual(preview["models"]["z06"]["interior_scope_row_count"], 4)
        self.assertEqual(preview["models"]["zr1"]["interior_scope_row_count"], 3)
        self.assertNotIn("2LZ", preview["models"]["zr1"]["trim_counts"])
        self.assertEqual(preview["models"]["zr1x"]["interior_scope_row_count"], 3)

    def test_component_projection_uses_legacy_lz_component_semantics(self) -> None:
        wb = preview_workbook()

        preview = build_lz_interiors_preview(wb, ["z06"], include_details=True)
        z06 = preview["models"]["z06"]

        self.assertEqual(z06["component_type_counts"], {"r6x": 1, "seat": 2, "stitching": 1, "suede": 1})
        self.assertEqual(z06["component_row_count"], 5)
        self.assertEqual(
            {(row["interior_id"], row["rpo"], row["component_type"]) for row in z06["component_rows"]},
            {
                ("2LZ_AE4_HTA", "AE4", "seat"),
                ("3LZ_AE4_HTA_N26_36S", "AE4", "seat"),
                ("3LZ_AE4_HTA_N26_36S", "N26", "suede"),
                ("3LZ_AE4_HTA_N26_36S", "36S", "stitching"),
                ("3LZ_R6X_AQ9_HTA_R6X", "R6X", "r6x"),
            },
        )

    def test_write_mode_applies_lz_sections_and_z_rows_idempotently(self) -> None:
        wb = preview_workbook()
        preview = build_lz_interiors_preview(wb, ["z06"], include_details=True)

        first_report = apply_lz_interiors_to_workbook(wb, preview)
        second_report = apply_lz_interiors_to_workbook(wb, preview)

        self.assertEqual(first_report["status"], "applied")
        self.assertEqual(second_report["status"], "applied")
        sections = {row["section_id"]: row for row in sheet_rows(wb, "section_master")}
        self.assertEqual(sections["sec_lzint_001"]["section_name"], "1LZ Interior")
        self.assertEqual(sections["sec_lzint_002"]["section_name"], "2LZ Interior")
        self.assertEqual(sections["sec_lzint_003"]["section_name"], "3LZ Interior")
        self.assertEqual(sections["sec_lzint_001"]["step_key"], "base_interior")

        self.assertEqual(
            [row["section_id"] for row in sheet_rows(wb, "LZ_Interiors")],
            ["sec_lzint_001", "sec_lzint_002", "sec_lzint_003", "sec_lzint_003"],
        )
        z06_scope = [row for row in sheet_rows(wb, "model_interior_scope") if row["model_key"] == "z06"]
        z06_components = [row for row in sheet_rows(wb, "interior_components") if row["model_key"] == "z06"]
        self.assertEqual(len(z06_scope), 4)
        self.assertEqual(len(z06_components), 5)
        self.assertEqual(len([row for row in sheet_rows(wb, "model_interior_scope") if row["model_key"] == "grand_sport"]), 1)

    def test_live_workbook_preview_matches_approved_spec_b_counts(self) -> None:
        workbook_path = ROOT / "stingray_master.xlsx"
        if not workbook_path.exists():
            self.skipTest("stingray_master.xlsx is not available")

        wb = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            preview = build_lz_interiors_preview(wb, ["all"])
        finally:
            wb.close()

        self.assertEqual(preview["source_row_count"], 132)
        self.assertEqual(preview["source_trim_counts"], {"1LZ": 4, "2LZ": 40, "3LZ": 73, "3LZ_R6X": 15})
        self.assertIn(
            preview["current_section_counts"],
            [
                {"": 132},
                {"sec_lzint_001": 4, "sec_lzint_002": 40, "sec_lzint_003": 88},
            ],
        )
        self.assertEqual(preview["models"]["z06"]["interior_scope_row_count"], 132)
        self.assertEqual(preview["models"]["zr1"]["interior_scope_row_count"], 92)
        self.assertEqual(preview["models"]["zr1x"]["interior_scope_row_count"], 92)
        self.assertEqual(preview["models"]["z06"]["component_row_count"], 198)
        self.assertEqual(preview["models"]["zr1"]["component_row_count"], 128)
        self.assertEqual(preview["models"]["zr1x"]["component_row_count"], 128)

    def test_cli_defaults_to_dry_run_json_stdout(self) -> None:
        command = [str(ROOT / ".venv" / "bin" / "python"), str(ROOT / "scripts" / "apply_future_model_lz_interiors.py"), "--model-key", "zr1"]

        result = subprocess.run(command, cwd=ROOT, check=True, capture_output=True, text=True)
        report = json.loads(result.stdout)

        self.assertEqual(report["status"], "dry_run")
        self.assertEqual(report["selected_model_keys"], ["zr1"])
        self.assertFalse(report["would_write_workbook"])
        self.assertNotIn("interior_scope_rows", report["models"]["zr1"])


if __name__ == "__main__":
    unittest.main()
