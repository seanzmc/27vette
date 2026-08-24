#!/usr/bin/env python3
"""Tests for the workbook-editor op engine (Phase 2)."""

from __future__ import annotations

import copy
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import corvette_form_generator.editor_ops as editor_ops  # noqa: E402
from corvette_form_generator.editor_ops import (  # noqa: E402
    apply_batch,
    coalesce_ops,
    coerce_value,
    extract_workbook,
    flatten_items,
    gate_reminders,
    validate_batch,
)
from corvette_form_generator.schema_validation import SchemaIssue  # noqa: E402

REAL_WORKBOOK = ROOT / "stingray_master.xlsx"


def op(action, sheet, key, row=None, **extra):
    out = {"action": action, "sheet": sheet, "key": key}
    if row is not None:
        out["row"] = row
    out.update(extra)
    return out


class FlattenTest(unittest.TestCase):
    def test_composite_members_carry_label(self):
        items = [{"kind": "composite", "label": "Add XYZ", "ops": [
            op("add", "s", {"option_id": "a"}, {"option_id": "a"})]}]
        flat = flatten_items(items)
        self.assertEqual(flat[0]["_composite"], "Add XYZ")
        self.assertEqual(flat[0]["_raw_effects"], [
            {"index": 0, "action": "add", "fields": ["option_id"]},
        ])

    def test_plain_ops_pass_through(self):
        flat = flatten_items([op("delete", "s", {"option_id": "a"})])
        self.assertEqual(flat[0]["action"], "delete")


class CoalesceTest(unittest.TestCase):
    K = {"option_id": "a"}

    def test_update_update_merges_but_marks_replaced_effect(self):
        out = coalesce_ops([op("update", "s", self.K, {"price": 1, "rpo": "X"}),
                            op("update", "s", self.K, {"price": 2})])
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0]["row"], {"price": 2, "rpo": "X"})
        self.assertTrue(any("contradictory raw effects" in error for error in out[0]["_coverage_errors"]))

    def test_add_update_merges_but_marks_replaced_effect(self):
        out = coalesce_ops([op("add", "s", self.K, {"option_id": "a", "price": 1}),
                            op("update", "s", self.K, {"price": 2})])
        self.assertEqual(out[0]["action"], "add")
        self.assertEqual(out[0]["row"]["price"], 2)
        self.assertTrue(any("contradictory raw effects" in error for error in out[0]["_coverage_errors"]))

    def test_add_delete_is_retained_as_a_dropped_effect_error(self):
        out = coalesce_ops([op("add", "s", self.K, {"option_id": "a"}),
                            op("delete", "s", self.K)])
        self.assertEqual(out[0]["_raw_indices"], [0, 1])
        self.assertTrue(any("dropped raw effects" in error for error in out[0]["_coverage_errors"]))

    def test_update_delete_becomes_delete(self):
        out = coalesce_ops([op("update", "s", self.K, {"price": 1}),
                            op("delete", "s", self.K)])
        self.assertEqual([o["action"] for o in out], ["delete"])
        self.assertTrue(any("dropped raw effect" in error for error in out[0]["_coverage_errors"]))

    def test_delete_is_a_barrier(self):
        out = coalesce_ops([op("delete", "s", self.K),
                            op("add", "s", self.K, {"option_id": "a"})])
        self.assertEqual([o["action"] for o in out], ["delete", "add"])

    def test_delete_followed_by_update_is_a_dropped_effect_error(self):
        out = coalesce_ops([op("delete", "s", self.K),
                            op("update", "s", self.K, {"price": 1})])
        self.assertTrue(any("delete followed by update" in error for error in out[0]["_coverage_errors"]))

    def test_repeated_delete_is_a_dropped_effect_error(self):
        out = coalesce_ops([op("delete", "s", self.K),
                            op("delete", "s", self.K)])
        self.assertTrue(any("delete followed by delete" in error for error in out[0]["_coverage_errors"]))

    def test_different_keys_untouched(self):
        out = coalesce_ops([op("update", "s", {"option_id": "a"}, {"price": 1}),
                            op("update", "s", {"option_id": "b"}, {"price": 2})])
        self.assertEqual(len(out), 2)

    def test_disjoint_updates_preserve_each_raw_effect(self):
        out = coalesce_ops([
            op("update", "s", self.K, {"price": 1}),
            op("update", "s", self.K, {"description": "x"}),
        ])

        self.assertEqual(out[0]["_raw_effects"], [
            {"index": 0, "action": "update", "fields": ["price"]},
            {"index": 1, "action": "update", "fields": ["description"]},
        ])


class CoerceTest(unittest.TestCase):
    def test_int(self):
        self.assertEqual(coerce_value("options", "price", 500), 500)
        self.assertEqual(coerce_value("options", "price", "500"), 500)
        self.assertIsNone(coerce_value("options", "price", ""))
        with self.assertRaises(ValueError):
            coerce_value("options", "price", "abc")
        with self.assertRaises(ValueError):
            coerce_value("options", "price", True)

    def test_bool(self):
        self.assertIs(coerce_value("options", "selectable", True), True)
        self.assertIs(coerce_value("options", "selectable", "False"), False)
        with self.assertRaises(ValueError):
            coerce_value("options", "selectable", "yes")

    def test_enum(self):
        self.assertEqual(coerce_value("ovs", "status", "standard"), "standard")
        with self.assertRaises(ValueError):
            coerce_value("ovs", "status", "maybe")
        # blank allowed only when "" is in the domain
        self.assertIsNone(coerce_value("options", "display_behavior", ""))
        with self.assertRaises(ValueError):
            coerce_value("ovs", "status", "")

    def test_variant_override_selectable_uses_real_excel_boolean(self):
        self.assertIs(coerce_value("variant_overrides", "selectable", "True"), True)
        self.assertIs(coerce_value("variant_overrides", "selectable", False), False)

    def test_free_text_stripped(self):
        self.assertEqual(coerce_value("options", "option_name", "  X  "), "X")
        self.assertIsNone(coerce_value("options", "description", ""))


class GateRemindersTest(unittest.TestCase):
    def test_grand_sport_reminders_use_the_current_candidate_lane(self):
        reminders = gate_reminders({"grand_sport"})

        self.assertIn(".venv/bin/python scripts/generate_form.py --model grand_sport", reminders)
        self.assertIn(".venv/bin/python scripts/validate_workbook_package.py stingray_master.xlsx", reminders)
        self.assertIn(".venv/bin/python scripts/validate_workbook_schema.py stingray_master.xlsx", reminders)
        self.assertIn(
            ".venv/bin/python scripts/verify_workbook_candidate.py --workbook stingray_master.xlsx "
            "--changed-model grand_sport --report /tmp/27vette-workbook-candidate.json",
            reminders,
        )
        candidate_index = next(index for index, command in enumerate(reminders) if "verify_workbook_candidate.py" in command)
        generation_index = reminders.index(".venv/bin/python scripts/generate_form.py --model grand_sport")
        publication_index = reminders.index(".venv/bin/python scripts/generate_registry.py")
        self.assertLess(candidate_index, generation_index)
        self.assertLess(generation_index, publication_index)
        self.assertFalse(any("preview" in command or "draft" in command or "rule-audit" in command for command in reminders), reminders)

    def test_multi_model_reminders_dedupe_shared_validation_and_scope_candidate_report(self):
        reminders = gate_reminders({"stingray", "grand_sport"})

        self.assertEqual(
            reminders.count(".venv/bin/python scripts/validate_workbook_schema.py stingray_master.xlsx"),
            1,
        )
        candidate = [command for command in reminders if "verify_workbook_candidate.py" in command]
        self.assertEqual(len(candidate), 1)
        self.assertIn("--changed-model grand_sport", candidate[0])
        self.assertIn("--changed-model stingray", candidate[0])


# ─────────────────────────────────────────────────────────────
# Fixture workbook for validation/apply tests
# ─────────────────────────────────────────────────────────────

def append_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    return ws


OPTION_HEADERS = ["option_id", "rpo", "price", "option_name", "description", "detail_raw",
                  "section_id", "selectable", "display_order", "active", "display_behavior"]


def option_row(oid, rpo, section, order):
    return {"option_id": oid, "rpo": rpo, "price": 0, "option_name": rpo,
            "section_id": section, "selectable": True, "display_order": order, "active": True}


def build_ops_fixture() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(wb, "model_master",
                 ["model_key", "model_label", "default_model", "active"],
                 [{"model_key": "stingray", "model_label": "Stingray", "default_model": True, "active": True},
                  {"model_key": "zr1", "model_label": "ZR1", "default_model": False, "active": False}])
    append_sheet(wb, "model_registry_promotion",
                 ["model_key", "promoted_to_runtime", "display_order", "active"],
                 [{"model_key": "stingray", "promoted_to_runtime": True, "display_order": 1, "active": True}])
    src = [
        ("stingray", "source_option_sheet", "stingray_options"),
        ("stingray", "status_sheet", "stingray_ovs"),
        ("stingray", "rule_mapping_sheet", "rule_mapping"),
        ("stingray", "rule_groups_sheet", "rule_groups"),
        ("stingray", "rule_group_members_sheet", "rule_group_members"),
        ("stingray", "exclusive_groups_sheet", "exclusive_groups"),
        ("stingray", "exclusive_group_members_sheet", "exclusive_group_members"),
        ("stingray", "price_rules_sheet", "price_rules"),
        ("stingray", "variant_option_overrides_sheet", "variant_overrides"),
        ("stingray", "color_overrides_sheet", "color_overrides"),
        ("stingray", "interior_source_sheet", "lt_interiors"),
        ("zr1", "source_option_sheet", "zr1_options"),
        ("zr1", "status_sheet", "zr1_ovs"),
    ]
    append_sheet(wb, "model_workbook_sources",
                 ["model_key", "source_role", "sheet_name", "active"],
                 [{"model_key": m, "source_role": r, "sheet_name": s, "active": True} for m, r, s in src])
    append_sheet(wb, "variant_master", ["variant_id", "display_name", "active"],
                 [{"variant_id": "1lt", "display_name": "1LT", "active": True},
                  {"variant_id": "2lt", "display_name": "2LT", "active": True},
                  {"variant_id": "zr1_c", "display_name": "ZR1", "active": False}])
    append_sheet(wb, "model_variants", ["model_key", "variant_id", "display_order", "active"],
                 [{"model_key": "stingray", "variant_id": "1lt", "display_order": 1, "active": True},
                  {"model_key": "stingray", "variant_id": "2lt", "display_order": 2, "active": True},
                  {"model_key": "zr1", "variant_id": "zr1_c", "display_order": 1, "active": True}])
    append_sheet(wb, "section_master", ["section_id", "section_name", "step_key"],
                 [{"section_id": "sec_a", "section_name": "A", "step_key": "paint"},
                  {"section_id": "sec_b", "section_name": "B", "step_key": "wheels"}])
    ws = append_sheet(wb, "stingray_options", OPTION_HEADERS,
                      [option_row("opt_one_001", "ONE", "sec_a", 10),
                       option_row("opt_two_001", "TWO", "sec_a", 20),
                       option_row("opt_thr_001", "THR", "sec_b", 10),
                       option_row("opt_rem_001", "REM", "sec_b", 20)])
    # Deliberately stale table ref: data goes to row 5, ref claims row 3.
    ws.add_table(Table(displayName="tbl_fixture_options", ref="A1:K3"))
    append_sheet(wb, "stingray_ovs", ["option_id", "variant_id", "status"],
                 [{"option_id": "opt_one_001", "variant_id": "1lt", "status": "available"},
                  {"option_id": "opt_one_001", "variant_id": "2lt", "status": "standard"}])
    append_sheet(
        wb,
        "rule_mapping",
        ["rule_id", "source_id", "rule_type", "target_id", "original_detail_raw",
         "body_style_scope", "runtime_action", "disabled_reason"],
        [
            {"rule_id": "rule_int", "source_id": "int_one_001", "rule_type": "requires",
             "target_id": "opt_thr_001"},
            {"rule_id": "rule_remap", "source_id": "opt_one_001", "rule_type": "requires",
             "target_id": "opt_rem_001"},
        ],
    )
    append_sheet(wb, "rule_groups", ["group_id", "group_type", "source_id", "active", "notes"],
                 [{"group_id": "grp_one", "group_type": "requires_any", "source_id": "opt_one_001", "active": True}])
    append_sheet(wb, "rule_group_members", ["group_id", "target_id", "display_order", "active"],
                 [{"group_id": "grp_one", "target_id": "opt_two_001", "display_order": 10, "active": True}])
    append_sheet(wb, "exclusive_groups", ["group_id", "selection_mode", "active", "notes"],
                 [{"group_id": "excl_one", "selection_mode": "single_within_group", "active": True}])
    append_sheet(wb, "exclusive_group_members", ["group_id", "option_id", "display_order", "active"],
                 [{"group_id": "excl_one", "option_id": "opt_one_001", "display_order": 10, "active": True},
                  {"group_id": "excl_one", "option_id": "opt_two_001", "display_order": 20, "active": True}])
    append_sheet(wb, "price_rules",
                 ["price_rule_id", "condition_option_id", "price_rule_type", "target_option_id",
                  "price_value", "body_style_scope", "trim_level_scope", "notes"],
                 [{"price_rule_id": "price_int", "condition_option_id": "int_one_001",
                   "price_rule_type": "override", "target_option_id": "opt_thr_001",
                   "price_value": 100}])
    append_sheet(
        wb,
        "variant_overrides",
        ["option_id", "variant_id", "selectable", "display_behavior", "section_id", "active", "note"],
        [],
    )
    append_sheet(
        wb,
        "lt_interiors",
        ["interior_id", "Interior Name", "Material", "Price", "Detail from Disclosure",
         "Color Overrides", "Trim", "Seat", "Interior Code", "Suede", "Stitch", "Two Tone",
         "section_id", "active_for_stingray", "requires_r6x", "included_option_id"],
        [{"interior_id": "int_one_001", "Interior Name": "Jet Black", "Price": 0,
          "section_id": "sec_b", "active_for_stingray": True,
          "requires_r6x": False, "included_option_id": "opt_two_001"}],
    )
    append_sheet(
        wb,
        "color_overrides",
        ["interior_id", "option_id", "rule_type", "adds_rpo"],
        [{"interior_id": "int_one_001", "option_id": "opt_thr_001",
          "rule_type": "requires", "adds_rpo": "opt_two_001"}],
    )
    append_sheet(
        wb,
        "model_interior_scope",
        ["model_key", "interior_id", "trim_level", "active", "requires_option_id", "notes"],
        [{"model_key": "stingray", "interior_id": "int_one_001", "trim_level": "1lt",
          "active": True, "requires_option_id": "opt_thr_001"}],
    )
    append_sheet(
        wb,
        "default_selection_rules",
        ["model_key", "rule_id", "target_option_id", "condition_type", "condition_id",
         "body_style_scope", "trim_level_scope", "variant_scope", "priority", "active", "notes",
         "display_behavior"],
        [
            {"model_key": "stingray", "rule_id": "default_target", "target_option_id": "opt_thr_001",
             "condition_type": "always", "priority": 1, "active": True},
            {"model_key": "stingray", "rule_id": "default_condition", "target_option_id": "opt_one_001",
             "condition_type": "when_selected_unless_selected_section", "condition_id": "opt_two_001",
             "priority": 2, "active": True},
            {"model_key": "stingray", "rule_id": "default_rpo_reference", "target_option_id": "opt_two_001",
             "condition_type": "unless_selected_rpo", "condition_id": "THR",
             "priority": 3, "active": True},
        ],
    )
    append_sheet(
        wb,
        "asset_map",
        ["model_key", "target_type", "target_id", "image_url", "image_fit", "active", "notes"],
        [{"model_key": "stingray", "target_type": "option", "target_id": "opt_thr_001",
          "image_url": "https://example.test/thr.png", "image_fit": "cover", "active": True}],
    )
    append_sheet(
        wb,
        "interior_components",
        ["model_key", "interior_id", "rpo", "component_type", "label", "price_ref_type",
         "price_ref_code", "price_trim_scope", "display_order", "active", "notes"],
        [{"model_key": "stingray", "interior_id": "int_one_001", "rpo": "AQ9",
          "component_type": "seat", "label": "GT1 seat", "display_order": 1, "active": True}],
    )
    append_sheet(wb, "zr1_options", OPTION_HEADERS, [option_row("opt_zzz_001", "ZZZ", "sec_a", 10)])
    append_sheet(wb, "zr1_ovs", ["option_id", "variant_id", "status"],
                 [{"option_id": "opt_zzz_001", "variant_id": "zr1_c", "status": "available"}])
    append_sheet(wb, "readonly_generated_fixture", ["step_key"], [{"step_key": "paint"}])
    return wb


class OpsFixtureBase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._dir = tempfile.TemporaryDirectory()
        cls.wb_path = Path(cls._dir.name) / "fixture.xlsx"
        build_ops_fixture().save(cls.wb_path)
        cls.extract = extract_workbook(cls.wb_path)

    @classmethod
    def tearDownClass(cls):
        cls._dir.cleanup()


def batch(*items, mtime=None):
    return {"version": 1, "workbook": "fixture.xlsx", "workbookMtimeNs": mtime, "items": list(items)}


def add_option_composite(oid="opt_new_001", statuses=("available", "available"), section="sec_a", order=30):
    row = option_row(oid, oid[4:7].upper(), section, order)
    row.update({"description": None, "detail_raw": None, "display_behavior": None})
    ops = [op("add", "stingray_options", {"option_id": oid}, row)]
    for vid, status in zip(("1lt", "2lt"), statuses):
        if status is not None:
            ops.append(op("add", "stingray_ovs", {"option_id": oid, "variant_id": vid},
                          {"option_id": oid, "variant_id": vid, "status": status}))
    return {"kind": "composite", "label": f"Add {oid}", "ops": ops}


class ValidateBatchTest(OpsFixtureBase):
    def errors_of(self, *items):
        return validate_batch(self.extract, batch(*items))["errors"]

    def warnings_of(self, *items):
        return validate_batch(self.extract, batch(*items))["warnings"]

    def test_clean_add_option_composite(self):
        result = validate_batch(self.extract, batch(add_option_composite()))
        self.assertEqual(result["errors"], [])

    def test_readonly_and_unknown_sheet(self):
        self.assertTrue(self.errors_of(op("update", "readonly_generated_fixture", {"step_key": "paint"}, {"step_key": "x"})))
        self.assertTrue(self.errors_of(op("update", "section_master", {"section_id": "sec_a"}, {"section_name": "X"})))
        self.assertTrue(self.errors_of(op("update", "nope", {"option_id": "a"}, {})))

    def test_rogue_physical_column_outside_the_registry_is_rejected(self):
        """A column that exists in Excel but not in the shared registry is not writable."""

        drifted = copy.deepcopy(self.extract)
        drifted["sheets"]["stingray_options"]["headers"].append("rogue_column")

        errors = validate_batch(
            drifted,
            batch(op("update", "stingray_options", {"option_id": "opt_one_001"}, {"rogue_column": "x"})),
        )["errors"]

        self.assertTrue(any("rogue_column" in error for error in errors), errors)

    def test_unknown_column_and_immutable_key(self):
        self.assertTrue(self.errors_of(op("update", "stingray_options", {"option_id": "opt_one_001"}, {"bogus": 1})))
        self.assertTrue(self.errors_of(op("update", "stingray_options", {"option_id": "opt_one_001"}, {"option_id": "x"})))

    def test_key_existence(self):
        self.assertTrue(self.errors_of(op("update", "stingray_options", {"option_id": "missing"}, {"price": 1})))
        self.assertTrue(self.errors_of(op("add", "stingray_options", {"option_id": "opt_one_001"},
                                          option_row("opt_one_001", "ONE", "sec_a", 99))))

    def test_type_enum_ref_violations(self):
        self.assertTrue(self.errors_of(op("update", "stingray_options", {"option_id": "opt_one_001"}, {"price": "abc"})))
        self.assertTrue(self.errors_of(op("update", "stingray_ovs",
                                          {"option_id": "opt_one_001", "variant_id": "1lt"}, {"status": "maybe"})))
        self.assertTrue(self.errors_of(op("update", "stingray_options", {"option_id": "opt_one_001"},
                                          {"section_id": "sec_nope"})))

    def test_batch_aware_refs(self):
        result = validate_batch(self.extract, batch(add_option_composite()))
        self.assertEqual(result["errors"], [])
        self.assertTrue(self.errors_of(op("add", "stingray_ovs",
                                          {"option_id": "opt_ghost_001", "variant_id": "1lt"},
                                          {"option_id": "opt_ghost_001", "variant_id": "1lt", "status": "available"})))

    def test_delete_does_not_validate_its_own_orphaned_outgoing_reference(self):
        orphaned = copy.deepcopy(self.extract)
        orphaned["sheets"]["stingray_options"]["rows"] = [
            row
            for row in orphaned["sheets"]["stingray_options"]["rows"]
            if row.get("option_id") != "opt_one_001"
        ]

        result = validate_batch(
            orphaned,
            batch(
                op(
                    "delete",
                    "stingray_ovs",
                    {"option_id": "opt_one_001", "variant_id": "1lt"},
                )
            ),
        )

        self.assertEqual(result["errors"], [], result)

    def test_direct_and_price_endpoints_accept_option_or_interior_entities(self):
        result = validate_batch(
            self.extract,
            batch(
                op(
                    "add",
                    "rule_mapping",
                    {"rule_id": "rule_union"},
                    {"rule_id": "rule_union", "source_id": "opt_one_001", "rule_type": "requires",
                     "target_id": "int_one_001"},
                ),
                op(
                    "add",
                    "price_rules",
                    {"price_rule_id": "price_union"},
                    {"price_rule_id": "price_union", "condition_option_id": "int_one_001",
                     "price_rule_type": "override", "target_option_id": "int_one_001", "price_value": 50},
                ),
            ),
        )

        self.assertEqual(result["errors"], [], result)

    def test_model_scoped_and_conditional_global_option_refs_are_validated(self):
        valid_items = [
            op(
                "update",
                "model_interior_scope",
                {"model_key": "stingray", "interior_id": "int_one_001", "trim_level": "1lt"},
                {"requires_option_id": "opt_two_001"},
            ),
            op(
                "add",
                "default_selection_rules",
                {"model_key": "stingray", "rule_id": "default_rpo"},
                {"model_key": "stingray", "rule_id": "default_rpo", "target_option_id": "opt_one_001",
                 "condition_type": "unless_selected_rpo", "condition_id": "TWO", "priority": 3, "active": True},
            ),
            op(
                "add",
                "default_selection_rules",
                {"model_key": "stingray", "rule_id": "default_section"},
                {"model_key": "stingray", "rule_id": "default_section", "target_option_id": "opt_one_001",
                 "condition_type": "unless_selected_section", "condition_id": "sec_b", "priority": 4,
                 "active": True},
            ),
            op(
                "update",
                "color_overrides",
                {"interior_id": "int_one_001", "option_id": "opt_thr_001"},
                {"adds_rpo": "opt_one_001"},
            ),
            op(
                "add",
                "asset_map",
                {"model_key": "stingray", "target_type": "context_choice", "target_id": "body_style__coupe"},
                {"model_key": "stingray", "target_type": "context_choice", "target_id": "body_style__coupe",
                 "image_url": "https://example.test/coupe.png", "image_fit": "cover", "active": True},
            ),
        ]
        self.assertEqual(self.errors_of(*valid_items), [])

        invalid_items = [
            op(
                "update",
                "model_interior_scope",
                {"model_key": "stingray", "interior_id": "int_one_001", "trim_level": "1lt"},
                {"requires_option_id": "opt_missing"},
            ),
            op(
                "add",
                "default_selection_rules",
                {"model_key": "stingray", "rule_id": "bad_target"},
                {"model_key": "stingray", "rule_id": "bad_target", "target_option_id": "opt_missing",
                 "condition_type": "always", "priority": 5, "active": True},
            ),
            op(
                "add",
                "default_selection_rules",
                {"model_key": "stingray", "rule_id": "bad_always"},
                {"model_key": "stingray", "rule_id": "bad_always", "target_option_id": "opt_one_001",
                 "condition_type": "always", "condition_id": "opt_two_001", "priority": 6, "active": True},
            ),
            op(
                "update",
                "color_overrides",
                {"interior_id": "int_one_001", "option_id": "opt_thr_001"},
                {"adds_rpo": "opt_missing"},
            ),
            op(
                "add",
                "asset_map",
                {"model_key": "stingray", "target_type": "option", "target_id": "opt_missing"},
                {"model_key": "stingray", "target_type": "option", "target_id": "opt_missing",
                 "image_url": "https://example.test/missing.png", "image_fit": "cover", "active": True},
            ),
        ]
        for item in invalid_items:
            with self.subTest(item=item):
                self.assertTrue(self.errors_of(item))

    def test_ovs_coverage_enforced(self):
        partial = add_option_composite(statuses=("available", None))  # missing 2lt
        errors = self.errors_of(partial)
        self.assertTrue(any("OVS coverage" in e and "2lt" in e for e in errors))

    def test_group_integrity(self):
        lone_group = op("add", "rule_groups", {"group_id": "grp_new"},
                        {"group_id": "grp_new", "group_type": "requires_any",
                         "source_id": "opt_one_001", "active": True, "notes": None})
        self.assertTrue(any("member" in e for e in self.errors_of(lone_group)))
        lone_excl = op("add", "exclusive_groups", {"group_id": "excl_new"},
                       {"group_id": "excl_new", "selection_mode": "single_within_group",
                        "active": True, "notes": None})
        self.assertTrue(any("member" in e for e in self.errors_of(lone_excl)),
                        self.errors_of(lone_excl))

    def test_active_exclusive_group_requires_display_label_once_migrated(self):
        """With the column present, blank is rejected on an active group."""
        import copy

        migrated = copy.deepcopy(self.extract)
        sheet = migrated["sheets"]["exclusive_groups"]
        sheet["headers"] = list(sheet["headers"]) + ["display_label"]
        for row in sheet.get("rows", []):
            row.setdefault("display_label", "Existing label")

        unlabeled = op("add", "exclusive_groups", {"group_id": "excl_blank"},
                       {"group_id": "excl_blank",
                        "selection_mode": "single_within_group",
                        "active": True, "notes": None})
        errors = validate_batch(migrated, batch(unlabeled))["errors"]
        self.assertTrue(
            [e for e in errors if "display_label" in e], errors
        )

        labeled = op("add", "exclusive_groups", {"group_id": "excl_ok"},
                     {"group_id": "excl_ok", "display_label": "Wheel selection",
                      "selection_mode": "single_within_group",
                      "active": True, "notes": None})
        labeled_errors = validate_batch(migrated, batch(labeled))["errors"]
        self.assertFalse(
            [e for e in labeled_errors if "display_label" in e], labeled_errors
        )

        inactive = op("add", "exclusive_groups", {"group_id": "excl_off"},
                      {"group_id": "excl_off",
                       "selection_mode": "single_within_group",
                       "active": False, "notes": None})
        inactive_errors = validate_batch(migrated, batch(inactive))["errors"]
        self.assertFalse(
            [e for e in inactive_errors if "display_label" in e], inactive_errors
        )

        # Deleting a row must never require filling in its label first.
        removal = op("delete", "exclusive_groups", {"group_id": "excl_one"}, {})
        delete_errors = validate_batch(migrated, batch(removal))["errors"]
        self.assertFalse(
            [e for e in delete_errors if "display_label" in e], delete_errors
        )

    def test_premigration_sheet_without_display_label_stays_writable(self):
        """§7.1's label requirement keys on the column, not the value.

        This fixture's `exclusive_groups` sheet predates the Checkpoint 2
        migration and has no `display_label` column, so an active group must
        still be addable without one.
        """
        self.assertNotIn(
            "display_label",
            set(self.extract["sheets"]["exclusive_groups"]["headers"]),
        )
        labeled = op("add", "exclusive_groups", {"group_id": "excl_new"},
                     {"group_id": "excl_new", "selection_mode": "single_within_group",
                      "active": True, "notes": None})
        self.assertFalse(
            [e for e in self.errors_of(labeled) if "display_label" in e],
            self.errors_of(labeled),
        )

    def test_display_order_collision_warns(self):
        warnings = self.warnings_of(op("update", "stingray_options", {"option_id": "opt_two_001"},
                                       {"display_order": 10}))
        self.assertTrue(any(w["id"].startswith("dorder:") for w in warnings))

    def test_referenced_delete_warns(self):
        warnings = self.warnings_of(op("delete", "stingray_options", {"option_id": "opt_one_001"}))
        self.assertTrue(any(w["id"].startswith("refdel:") for w in warnings))
        items = [op("delete", "stingray_ovs", {"option_id": "opt_one_001", "variant_id": "1lt"}),
                 op("delete", "stingray_ovs", {"option_id": "opt_one_001", "variant_id": "2lt"}),
                 op("delete", "rule_groups", {"group_id": "grp_one"}),
                 op("delete", "rule_group_members", {"group_id": "grp_one", "target_id": "opt_two_001"}),
                 op("delete", "exclusive_group_members", {"group_id": "excl_one", "option_id": "opt_one_001"}),
                 op("delete", "rule_mapping", {"rule_id": "rule_remap"}),
                 op("delete", "default_selection_rules", {"model_key": "stingray", "rule_id": "default_condition"}),
                 op("delete", "stingray_options", {"option_id": "opt_one_001"})]
        warnings = self.warnings_of(*items)
        self.assertFalse(any(w["id"] == "refdel:stingray_options:opt_one_001" for w in warnings))

    def test_complete_option_and_interior_incoming_reference_surfaces_warn(self):
        option_warning = next(
            warning
            for warning in self.warnings_of(
                op("delete", "stingray_options", {"option_id": "opt_thr_001"})
            )
            if warning["id"] == "refdel:stingray_options:opt_thr_001"
        )
        for reference in (
            "rule_mapping.target_id",
            "price_rules.target_option_id",
            "color_overrides.option_id",
            "model_interior_scope.requires_option_id",
            "default_selection_rules.target_option_id",
            "default_selection_rules.condition_id",
            "asset_map.target_id",
        ):
            self.assertIn(reference, option_warning["message"])

        interior_warning = next(
            warning
            for warning in self.warnings_of(
                op("delete", "lt_interiors", {"interior_id": "int_one_001"})
            )
            if warning["id"] == "refdel:lt_interiors:int_one_001"
        )
        for reference in (
            "rule_mapping.source_id",
            "price_rules.condition_option_id",
            "color_overrides.interior_id",
            "model_interior_scope.interior_id",
            "interior_components.interior_id",
        ):
            self.assertIn(reference, interior_warning["message"])

    def test_rpo_reference_survives_one_duplicate_rpo_option_delete(self):
        duplicate_rpo = copy.deepcopy(self.extract)
        duplicate_rpo["sheets"]["stingray_options"]["rows"].append(
            option_row("opt_thr_002", "THR", "sec_b", 30)
        )

        warning = next(
            warning
            for warning in validate_batch(
                duplicate_rpo,
                batch(op("delete", "stingray_options", {"option_id": "opt_thr_001"})),
            )["warnings"]
            if warning["id"] == "refdel:stingray_options:opt_thr_001"
        )

        self.assertNotIn("default_selection_rules.condition_id", warning["message"])

    def test_final_reference_state_is_built_once_for_multiple_deletes(self):
        with patch.object(
            editor_ops,
            "_final_rows_by_sheet",
            wraps=editor_ops._final_rows_by_sheet,
        ) as final_state_builder:
            result = validate_batch(
                self.extract,
                batch(
                    op("delete", "stingray_options", {"option_id": "opt_thr_001"}),
                    op("delete", "stingray_options", {"option_id": "opt_rem_001"}),
                ),
            )

        self.assertEqual(result["errors"], [], result)
        self.assertEqual(final_state_builder.call_count, 1)

    def test_same_batch_reference_remap_closes_delete_warning(self):
        warnings = self.warnings_of(
            op("update", "rule_mapping", {"rule_id": "rule_remap"}, {"target_id": "opt_two_001"}),
            op("delete", "stingray_options", {"option_id": "opt_rem_001"}),
        )

        self.assertFalse(any(w["id"] == "refdel:stingray_options:opt_rem_001" for w in warnings), warnings)

    def test_scaffold_model_warns(self):
        warnings = self.warnings_of(op("update", "zr1_options", {"option_id": "opt_zzz_001"}, {"price": 1}))
        self.assertTrue(any(w["id"] == "scaffold:zr1_options" for w in warnings))


class ApplyBatchTest(unittest.TestCase):
    def setUp(self):
        self._dir = tempfile.TemporaryDirectory()
        self.path = Path(self._dir.name) / "fixture.xlsx"
        build_ops_fixture().save(self.path)
        self.log = Path(self._dir.name) / "log.jsonl"

    def tearDown(self):
        self._dir.cleanup()

    def run_batch(self, *items, write=True, confirmed=(), allow_stale=False, mtime=None):
        b = batch(*items, mtime=self.path.stat().st_mtime_ns if mtime is None else mtime)
        schema_enabled = bool(write)
        with patch("corvette_form_generator.editor_ops.validate_workbook_schema", return_value=[]):
            return apply_batch(self.path, b, write=write, confirmed_warnings=confirmed,
                               log_path=self.log, run_schema_validation=schema_enabled,
                               allow_stale=allow_stale)

    def test_add_option_round_trip_and_table_heal(self):
        result = self.run_batch(add_option_composite())
        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path)
        ws = wb["stingray_options"]
        rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
        self.assertIn("opt_new_001", rows)
        r = rows["opt_new_001"]
        self.assertIsInstance(ws.cell(row=r, column=3).value, int)   # price typed
        self.assertIsInstance(ws.cell(row=r, column=8).value, bool)  # selectable typed
        self.assertEqual(ws.tables["tbl_fixture_options"].ref, "A1:K6")  # healed 3 -> 6
        self.assertEqual(wb["stingray_ovs"].max_row, 5)  # 2 ovs rows added
        wb.close()
        self.assertTrue(self.log.exists())
        entry = json.loads(self.log.read_text().splitlines()[-1])
        self.assertEqual(entry["opCount"], 3)
        self.assertIn("stingray_options", entry["sheets"])
        backups = Path(self._dir.name) / "backups"
        self.assertTrue(any(backups.iterdir()))

    def test_pass_c3_batch_forces_real_excel_booleans_over_text_convention(self):
        wb = load_workbook(self.path)
        ws = wb["stingray_options"]
        headers = {
            str(cell.value): index + 1
            for index, cell in enumerate(ws[1])
            if cell.value is not None
        }
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=headers["selectable"], value="True")
            ws.cell(row=row, column=headers["active"], value="True")
        wb.save(self.path)
        wb.close()

        payload = batch(
            {
                "action": "update",
                "sheet": "stingray_options",
                "key": {"option_id": "opt_one_001"},
                "row": {"selectable": False, "active": True},
            },
            mtime=self.path.stat().st_mtime_ns,
        )
        payload["forceTypedBools"] = True
        with patch(
            "corvette_form_generator.editor_ops.validate_workbook_schema",
            return_value=[],
        ):
            result = apply_batch(
                self.path,
                payload,
                write=True,
                log_path=self.log,
                run_schema_validation=True,
            )

        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb["stingray_options"]
            headers = {
                str(cell.value): index
                for index, cell in enumerate(ws[1])
                if cell.value is not None
            }
            row = next(
                values
                for values in ws.iter_rows(min_row=2, values_only=True)
                if values[headers["option_id"]] == "opt_one_001"
            )
            self.assertIs(row[headers["selectable"]], False)
            self.assertIs(row[headers["active"]], True)
        finally:
            wb.close()

    def test_update_and_delete(self):
        # mtime passed as a string, as the browser must send it (JS precision)
        result = self.run_batch(
            op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": 777, "description": "x"}),
            op("delete", "stingray_ovs", {"option_id": "opt_one_001", "variant_id": "2lt"}),
            mtime=str(self.path.stat().st_mtime_ns),
        )
        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path)
        ws = wb["stingray_options"]
        r = next(r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == "opt_thr_001")
        self.assertEqual(ws.cell(row=r, column=3).value, 777)
        self.assertEqual(wb["stingray_ovs"].max_row, 2)
        wb.close()

    def test_blanking_a_cell(self):
        result = self.run_batch(op("update", "stingray_options", {"option_id": "opt_thr_001"},
                                   {"option_name": ""}))
        self.assertTrue(result["ok"], result)
        wb = load_workbook(self.path)
        ws = wb["stingray_options"]
        r = next(r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == "opt_thr_001")
        self.assertIsNone(ws.cell(row=r, column=4).value)
        wb.close()

    def test_validate_only_makes_no_changes(self):
        before = self.path.stat().st_mtime_ns
        result = self.run_batch(add_option_composite(), write=False)
        self.assertTrue(result["ok"])
        self.assertEqual(result["status"], "validated")
        self.assertEqual(self.path.stat().st_mtime_ns, before)
        self.assertFalse(self.log.exists())

    def test_coalesced_updates_report_complete_raw_coverage_and_exact_readback(self):
        before = self.path.read_bytes()
        result = self.run_batch(
            op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": "777"}),
            op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"description": "  exact  "}),
            write=False,
        )

        self.assertTrue(result["ok"], result)
        self.assertEqual(
            result["operationCoverage"],
            {"rawCount": 2, "rawCovered": 2, "preparedCount": 1},
        )
        self.assertTrue(result["verification"]["ok"], result)
        self.assertEqual(result["verification"]["preparedChecked"], 1)
        self.assertEqual(result["verification"]["preparedCount"], 1)
        self.assertEqual(self.path.read_bytes(), before)

    def test_exact_readback_indexes_each_sheet_once_for_multiple_operations(self):
        wb = load_workbook(self.path, read_only=True)
        try:
            data_rows = wb["stingray_options"].max_row - 1
        finally:
            wb.close()
        original = editor_ops._canonical_readback_key
        with patch.object(
            editor_ops,
            "_canonical_readback_key",
            wraps=original,
        ) as readback_key:
            result = self.run_batch(
                op(
                    "update",
                    "stingray_options",
                    {"option_id": "opt_one_001"},
                    {"price": 11},
                ),
                op(
                    "update",
                    "stingray_options",
                    {"option_id": "opt_thr_001"},
                    {"price": 12},
                ),
                write=False,
            )

        self.assertTrue(result["ok"], result)
        self.assertEqual(readback_key.call_count, data_rows)

    def test_composite_members_each_count_as_raw_operations(self):
        composite = add_option_composite()
        composite["ops"][0]["row"].update(
            {"price": "42", "selectable": "False", "display_order": "30", "active": "True"}
        )
        result = self.run_batch(composite, write=False)

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["operationCoverage"]["rawCount"], 3)
        self.assertEqual(result["operationCoverage"]["rawCovered"], 3)
        self.assertEqual(result["operationCoverage"]["preparedCount"], 3)
        self.assertEqual(result["verification"]["preparedChecked"], 3)

    def test_contradictory_coalesced_effects_are_invalid(self):
        result = self.run_batch(
            op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": 1}),
            op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": 2}),
            write=False,
        )

        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "invalid")
        self.assertTrue(any("contradictory raw effects" in error for error in result["errors"]), result)

    def test_dropped_update_effect_is_invalid(self):
        result = self.run_batch(
            op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": 1}),
            op("delete", "stingray_options", {"option_id": "opt_thr_001"}),
            write=False,
        )

        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "invalid")
        self.assertTrue(any("dropped raw effect" in error for error in result["errors"]), result)

    def test_tampered_scratch_fields_fail_exact_readback_before_live_mutation(self):
        before = self.path.read_bytes()

        def save_and_tamper(workbook, scratch_path):
            workbook.save(scratch_path)
            tampered = load_workbook(scratch_path)
            ws = tampered["stingray_options"]
            headers = {cell.value: cell.column for cell in ws[1]}
            row = next(
                row_number
                for row_number in range(2, ws.max_row + 1)
                if ws.cell(row=row_number, column=headers["option_id"]).value == "opt_thr_001"
            )
            ws.cell(row=row, column=headers["price"], value=778)
            ws.cell(row=row, column=headers["selectable"], value="False")
            tampered.save(scratch_path)
            tampered.close()

        item = op(
            "update",
            "stingray_options",
            {"option_id": "opt_thr_001"},
            {"price": "777", "selectable": "False"},
        )
        with patch.object(editor_ops, "_save_scratch_workbook", side_effect=save_and_tamper, create=True):
            result = self.run_batch(item, write=False)

        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "readback_failed")
        self.assertTrue(any("price" in error for error in result["errors"]), result)
        self.assertTrue(any("selectable" in error for error in result["errors"]), result)
        self.assertLess(result["verification"]["preparedChecked"], result["operationCoverage"]["preparedCount"])
        self.assertEqual(self.path.read_bytes(), before)

    def test_tampered_scratch_delete_fails_absence_readback(self):
        before = self.path.read_bytes()

        def save_and_restore_deleted_row(workbook, scratch_path):
            workbook.save(scratch_path)
            tampered = load_workbook(scratch_path)
            tampered["stingray_ovs"].append(["opt_one_001", "2lt", "standard"])
            tampered.save(scratch_path)
            tampered.close()

        item = op(
            "delete",
            "stingray_ovs",
            {"option_id": "opt_one_001", "variant_id": "2lt"},
        )
        with patch.object(
            editor_ops,
            "_save_scratch_workbook",
            side_effect=save_and_restore_deleted_row,
            create=True,
        ):
            result = self.run_batch(item, write=False)

        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "readback_failed")
        self.assertTrue(any("delete" in error and "still exists" in error for error in result["errors"]), result)
        self.assertEqual(self.path.read_bytes(), before)

    def test_tampered_live_save_exposes_apply_verification_failed(self):
        real_safe_save = editor_ops.save_workbook_safely

        def save_and_tamper_live(
            workbook,
            workbook_path,
            *,
            loaded_mtime_ns,
            approved_bool_type_migrations=None,
        ):
            backup_path = real_safe_save(
                workbook,
                workbook_path,
                loaded_mtime_ns=loaded_mtime_ns,
                approved_bool_type_migrations=approved_bool_type_migrations,
            )
            tampered = load_workbook(workbook_path)
            ws = tampered["stingray_options"]
            headers = {cell.value: cell.column for cell in ws[1]}
            row = next(
                row_number
                for row_number in range(2, ws.max_row + 1)
                if ws.cell(row=row_number, column=headers["option_id"]).value == "opt_thr_001"
            )
            ws.cell(row=row, column=headers["price"], value=999)
            tampered.save(workbook_path)
            tampered.close()
            return backup_path

        item = op(
            "update",
            "stingray_options",
            {"option_id": "opt_thr_001"},
            {"price": 777},
        )
        before = self.path.read_bytes()
        with patch(
            "corvette_form_generator.editor_ops.save_workbook_safely",
            side_effect=save_and_tamper_live,
        ):
            result = self.run_batch(item, write=True)

        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "apply_verification_failed_rolled_back")
        self.assertEqual(result["workbookState"], "restored")
        self.assertTrue(Path(result["backupPath"]).exists())
        self.assertTrue(any("price" in error for error in result["errors"]), result)
        self.assertEqual(result["failure"]["kind"], "returned_failure")
        self.assertEqual(result["failure"]["phase"], "live_readback")
        self.assertTrue(result["restoration"]["verified"])
        self.assertEqual(
            result["restoration"]["workbookSha256"],
            result["restoration"]["backupSha256"],
        )
        self.assertFalse(self.log.exists())
        self.assertEqual(self.path.read_bytes(), before)

    def test_thrown_post_save_failures_restore_and_preserve_phase_detail(self):
        item = op(
            "update",
            "stingray_options",
            {"option_id": "opt_thr_001"},
            {"price": 777},
        )
        original_verify = editor_ops.verify_prepared_workbook
        original_package = editor_ops.assert_valid_workbook_package

        def run_fault(phase):
            before = self.path.read_bytes()
            verify_calls = {"count": 0}
            package_calls = {"count": 0}
            schema_calls = {"count": 0}

            def verify(path, prepared):
                verify_calls["count"] += 1
                if phase == "live_readback" and verify_calls["count"] == 2:
                    raise RuntimeError("forced live readback exception")
                return original_verify(path, prepared)

            def package(path):
                package_calls["count"] += 1
                if phase == "live_package" and package_calls["count"] == 2:
                    raise RuntimeError("forced live package exception")
                return original_package(path)

            def schema(*args, **kwargs):
                schema_calls["count"] += 1
                if phase == "live_schema" and schema_calls["count"] == 2:
                    raise RuntimeError("forced live schema exception")
                return []

            def append_log(*args, **kwargs):
                if phase == "write_log":
                    raise OSError("forced write log exception")

            batch_payload = batch(item, mtime=self.path.stat().st_mtime_ns)
            with (
                patch.object(editor_ops, "verify_prepared_workbook", side_effect=verify),
                patch.object(editor_ops, "assert_valid_workbook_package", side_effect=package),
                patch.object(editor_ops, "validate_workbook_schema", side_effect=schema),
                patch.object(editor_ops, "_append_edit_log", side_effect=append_log, create=True),
            ):
                result = apply_batch(
                    self.path,
                    batch_payload,
                    write=True,
                    log_path=self.log,
                    run_schema_validation=True,
                )

            self.assertFalse(result["ok"], result)
            self.assertEqual(result["status"], "apply_verification_failed_rolled_back")
            self.assertEqual(result["workbookState"], "restored")
            self.assertEqual(result["failure"]["phase"], phase)
            self.assertEqual(result["failure"]["kind"], "exception")
            self.assertIn("forced", result["failure"]["detail"])
            self.assertTrue(result["restoration"]["verified"])
            self.assertIsNone(result["restoration"]["error"])
            self.assertEqual(
                result["restoration"]["workbookSha256"],
                result["restoration"]["backupSha256"],
            )
            self.assertEqual(self.path.read_bytes(), before)
            self.assertFalse(self.log.exists())

        for phase in ("live_readback", "live_package", "live_schema", "write_log"):
            with self.subTest(phase=phase):
                run_fault(phase)

    def test_returned_live_schema_failure_restores_and_preserves_schema_result(self):
        item = op(
            "update",
            "stingray_options",
            {"option_id": "opt_thr_001"},
            {"price": 777},
        )
        before = self.path.read_bytes()
        schema_calls = {"count": 0}

        def schema(*args, **kwargs):
            schema_calls["count"] += 1
            if schema_calls["count"] == 2:
                return [SchemaIssue("error", "forced_live_schema", message="forced returned failure")]
            return []

        with patch.object(editor_ops, "validate_workbook_schema", side_effect=schema):
            result = apply_batch(
                self.path,
                batch(item, mtime=self.path.stat().st_mtime_ns),
                write=True,
                log_path=self.log,
                run_schema_validation=True,
            )

        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "apply_verification_failed_rolled_back")
        self.assertEqual(result["workbookState"], "restored")
        self.assertEqual(result["failure"]["phase"], "live_schema")
        self.assertEqual(result["failure"]["kind"], "returned_failure")
        self.assertEqual(result["schemaResult"]["error_count"], 1)
        self.assertTrue(result["restoration"]["verified"])
        self.assertEqual(
            result["restoration"]["workbookSha256"],
            result["restoration"]["backupSha256"],
        )
        self.assertEqual(self.path.read_bytes(), before)
        self.assertFalse(self.log.exists())

    def test_unproved_post_save_restoration_preserves_both_failure_records(self):
        item = op(
            "update",
            "stingray_options",
            {"option_id": "opt_thr_001"},
            {"price": 777},
        )
        original_verify = editor_ops.verify_prepared_workbook
        original_restore = editor_ops.restore_workbook_backup

        for restoration_mode in ("hash_mismatch", "exception"):
            with self.subTest(restoration_mode=restoration_mode):
                before = self.path.read_bytes()
                verify_calls = {"count": 0}

                def fail_returned_live_readback(path, prepared):
                    verify_calls["count"] += 1
                    if verify_calls["count"] == 2:
                        return {
                            "ok": False,
                            "preparedChecked": 0,
                            "preparedCount": 1,
                            "errors": ["forced original readback failure"],
                        }
                    return original_verify(path, prepared)

                def fail_restore(path, backup_path):
                    if restoration_mode == "exception":
                        raise RuntimeError("forced restoration exception")
                    original_restore(path, backup_path)
                    path.write_bytes(path.read_bytes() + b"forced-hash-mismatch")

                with (
                    patch.object(
                        editor_ops,
                        "verify_prepared_workbook",
                        side_effect=fail_returned_live_readback,
                    ),
                    patch.object(editor_ops, "restore_workbook_backup", side_effect=fail_restore),
                ):
                    result = self.run_batch(item, write=True)

                self.assertFalse(result["ok"], result)
                self.assertEqual(result["status"], "workbook_restore_failed")
                self.assertEqual(result["workbookState"], "unknown")
                self.assertEqual(result["failure"]["phase"], "live_readback")
                self.assertEqual(result["failure"]["kind"], "returned_failure")
                self.assertIn("forced original", result["failure"]["detail"])
                self.assertFalse(result["restoration"]["verified"])
                self.assertTrue(result["restoration"]["backupSha256"])
                self.assertTrue(result["restoration"]["error"])
                self.assertIn("forced original", result["errors"][0])
                if restoration_mode == "hash_mismatch":
                    self.assertTrue(result["restoration"]["workbookSha256"])
                    self.assertIn("does not match", result["restoration"]["error"])
                else:
                    self.assertIsNone(result["restoration"]["workbookSha256"])
                    self.assertIn("forced restoration", result["restoration"]["error"])

                self.path.write_bytes(before)

    def test_backup_hash_failure_does_not_claim_restore_was_attempted(self):
        item = op(
            "update",
            "stingray_options",
            {"option_id": "opt_thr_001"},
            {"price": 777},
        )
        original_verify = editor_ops.verify_prepared_workbook
        original_read_bytes = Path.read_bytes
        verify_calls = {"count": 0}

        def fail_returned_live_readback(path, prepared):
            verify_calls["count"] += 1
            if verify_calls["count"] == 2:
                return {
                    "ok": False,
                    "preparedChecked": 0,
                    "preparedCount": 1,
                    "errors": ["forced original readback failure"],
                }
            return original_verify(path, prepared)

        def fail_backup_read(path):
            if path.parent.name == "backups":
                raise OSError("forced backup read failure")
            return original_read_bytes(path)

        with (
            patch.object(
                editor_ops,
                "verify_prepared_workbook",
                side_effect=fail_returned_live_readback,
            ),
            patch.object(Path, "read_bytes", autospec=True, side_effect=fail_backup_read),
            patch.object(editor_ops, "restore_workbook_backup") as restore,
        ):
            result = self.run_batch(item, write=True)

        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "workbook_restore_failed")
        self.assertEqual(result["workbookState"], "unknown")
        self.assertFalse(result["restoration"]["attempted"])
        self.assertFalse(result["restoration"]["verified"])
        self.assertIsNone(result["restoration"]["backupSha256"])
        self.assertIn("forced backup read failure", result["restoration"]["error"])
        restore.assert_not_called()

    def test_warning_requires_confirmation(self):
        item = op("update", "zr1_options", {"option_id": "opt_zzz_001"}, {"price": 1})
        result = self.run_batch(item)
        self.assertFalse(result["ok"])
        self.assertEqual(result["status"], "needs_confirmation")
        wid = result["warnings"][0]["id"]
        result = self.run_batch(item, confirmed=(wid,))
        self.assertTrue(result["ok"], result)

    def test_unconfirmable_and_stale_warning_ids_cannot_reach_write(self):
        cases = [
            (
                op("update", "stingray_options", {"option_id": "opt_two_001"}, {"display_order": 10}),
                "dorder:",
            ),
            (op("delete", "stingray_options", {"option_id": "opt_thr_001"}), "refdel:"),
        ]
        for item, prefix in cases:
            with self.subTest(prefix=prefix):
                before = self.path.read_bytes()
                preview = self.run_batch(item, write=False)
                warning_id = next(w["id"] for w in preview["warnings"] if w["id"].startswith(prefix))
                result = self.run_batch(item, confirmed=(warning_id,))
                self.assertFalse(result["ok"], result)
                self.assertEqual(result["status"], "warning_blocked")
                self.assertEqual(self.path.read_bytes(), before)

        clean_item = op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": 1})
        result = self.run_batch(clean_item, confirmed=("scaffold:stale_sheet",))
        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "warning_confirmation_mismatch")

    def test_unknown_warning_kind_blocks_write(self):
        item = op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": 1})
        prepared = [{
            "action": "update",
            "sheet": "stingray_options",
            "_raw_indices": [0],
            "_raw_effects": [{"index": 0, "action": "update", "fields": ["price"]}],
            "_coerced_row": {"price": 1},
        }]
        with patch(
            "corvette_form_generator.editor_ops._prepare_batch",
            return_value=([], [{"id": "mystery:stingray_options", "message": "unknown"}], prepared),
        ):
            result = self.run_batch(item, confirmed=("mystery:stingray_options",))
        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "warning_blocked")

    def test_schema_disabled_write_is_refused_before_mutation(self):
        item = op("update", "stingray_options", {"option_id": "opt_thr_001"}, {"price": 1})
        before = self.path.read_bytes()
        result = apply_batch(
            self.path,
            batch(item, mtime=self.path.stat().st_mtime_ns),
            write=True,
            run_schema_validation=False,
            log_path=self.log,
        )
        self.assertFalse(result["ok"], result)
        self.assertEqual(result["status"], "schema_validation_required")
        self.assertEqual(self.path.read_bytes(), before)
        self.assertFalse(self.log.exists())

    def test_warning_classification_and_fingerprint_are_finite_and_stable(self):
        warnings = [
            {"id": "scaffold:zr1_options", "message": "scaffold"},
            {"id": "refdel:stingray_options:opt_one_001", "message": "reference"},
        ]
        self.assertEqual(editor_ops.CONFIRMABLE_WARNING_KINDS, {"scaffold"})
        classified = editor_ops.classify_warnings(warnings)
        self.assertEqual(classified["confirmableIds"], ["scaffold:zr1_options"])
        self.assertEqual(classified["blockingIds"], ["refdel:stingray_options:opt_one_001"])
        self.assertEqual(
            editor_ops.warning_fingerprint(reversed(warnings)),
            editor_ops.warning_fingerprint(warnings),
        )

    def test_invalid_batch_refused(self):
        result = self.run_batch(op("update", "stingray_options", {"option_id": "opt_one_001"},
                                   {"section_id": "sec_nope"}))
        self.assertEqual(result["status"], "invalid")

    def test_lock_and_stale_refusal(self):
        lock = self.path.with_name(f"~${self.path.name}")
        lock.write_text("")
        self.assertEqual(self.run_batch(add_option_composite())["status"], "locked")
        lock.unlink()
        self.assertEqual(self.run_batch(add_option_composite(), mtime=1)["status"], "stale")
        result = self.run_batch(add_option_composite(), mtime=1, allow_stale=True)
        self.assertTrue(result["ok"], result)


@unittest.skipUnless(REAL_WORKBOOK.exists(), "canonical workbook not present")
class RealWorkbookApplyTest(unittest.TestCase):
    def setUp(self):
        self._dir = tempfile.TemporaryDirectory()
        self.path = Path(self._dir.name) / "copy.xlsx"
        shutil.copy2(REAL_WORKBOOK, self.path)
        self.log = Path(self._dir.name) / "log.jsonl"

    def tearDown(self):
        self._dir.cleanup()

    def test_gs_update_heals_stale_table_refs(self):
        # Inject the ingest-era stale ref on the temp copy: the canonical
        # workbook's own ref was healed by a real editor apply on 2026-06-12,
        # so the stale precondition can no longer be assumed from the file.
        wb = load_workbook(self.path)
        ws = wb["grandSport_options"]
        ws.tables["tbl_grandSport_options"].ref = "A1:K267"
        last_data_row = max(
            (r for r in range(1, ws.max_row + 1)
             if any(c.value is not None for c in ws[r])), default=2)
        wb.save(self.path)
        wb.close()
        b = {"version": 1, "workbookMtimeNs": self.path.stat().st_mtime_ns, "items": [
            {"action": "update", "sheet": "grandSport_options", "key": {"option_id": "opt_fey_001"},
             "row": {"detail_raw": "editor apply test"}}]}
        result = apply_batch(self.path, b, write=True, log_path=self.log, source="test")
        self.assertTrue(result["ok"], result)
        self.assertEqual((result["schemaResult"] or {}).get("error_count"), 0)
        wb = load_workbook(self.path)
        self.assertEqual(wb["grandSport_options"].tables["tbl_grandSport_options"].ref,
                         f"A1:K{max(last_data_row, 2)}")
        wb.close()

    def test_bad_ref_refused_on_real_workbook(self):
        b = {"version": 1, "workbookMtimeNs": self.path.stat().st_mtime_ns, "items": [
            {"action": "update", "sheet": "stingray_options", "key": {"option_id": "opt_z51_001"},
             "row": {"section_id": "sec_nope"}}]}
        result = apply_batch(self.path, b, write=True, log_path=self.log, source="test")
        self.assertEqual(result["status"], "invalid")


if __name__ == "__main__":
    unittest.main()
