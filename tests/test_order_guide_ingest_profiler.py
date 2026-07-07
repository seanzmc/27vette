#!/usr/bin/env python3
"""Tests for the Pass 0 raw order-guide source profiler."""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.ingest.source_profiler import (  # noqa: E402
    parse_status,
    profile_order_guide,
    validate_output_dir,
)


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def fixture_workbook(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "model_master",
        [
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "notes",
        ],
        [
            {
                "model_key": "stingray",
                "registry_key": "stingray",
                "model_label": "Stingray",
                "model_year": "2027",
                "dataset_name": "Fixture Stingray",
                "export_slug": "stingray",
                "expected_variant_count": 1,
                "default_model": True,
                "active": True,
            },
            {
                "model_key": "zr1",
                "registry_key": "zr1",
                "model_label": "ZR1",
                "model_year": "2027",
                "dataset_name": "Fixture ZR1",
                "export_slug": "zr1",
                "expected_variant_count": 1,
                "default_model": False,
                "active": False,
            },
            {
                "model_key": "zr1x",
                "registry_key": "zr1x",
                "model_label": "ZR1X",
                "model_year": "2027",
                "dataset_name": "Fixture ZR1X",
                "export_slug": "zr1x",
                "expected_variant_count": 1,
                "default_model": False,
                "active": False,
            },
        ],
    )
    append_sheet(
        wb,
        "variant_master",
        ["variant_id", "model_year", "trim_level", "body_style", "display_name", "base_price", "display_order", "active"],
        [
            {
                "variant_id": "1lt_c07",
                "model_year": "2027",
                "trim_level": "1lt",
                "body_style": "coupe",
                "display_name": "Corvette Stingray Coupe 1LT",
                "base_price": 71000,
                "display_order": 1,
                "active": True,
            },
            {
                "variant_id": "1lz_r07",
                "model_year": "2027",
                "trim_level": "1lz",
                "body_style": "coupe",
                "display_name": "Corvette ZR1 Coupe 1LZ",
                "base_price": 194700,
                "display_order": 2,
                "active": False,
            },
            {
                "variant_id": "1lz_s07",
                "model_year": "2027",
                "trim_level": "1lz",
                "body_style": "coupe",
                "display_name": "Corvette ZR1X Coupe 1LZ",
                "base_price": 224900,
                "display_order": 3,
                "active": False,
            },
        ],
    )
    append_sheet(
        wb,
        "model_variants",
        ["model_key", "variant_id", "display_order", "active", "notes"],
        [
            {"model_key": "stingray", "variant_id": "1lt_c07", "display_order": 1, "active": True},
            {"model_key": "zr1", "variant_id": "1lz_r07", "display_order": 1, "active": False},
            {"model_key": "zr1x", "variant_id": "1lz_s07", "display_order": 1, "active": False},
        ],
    )
    wb.save(path)


def raw_export_fixture(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    ws = wb.create_sheet("Exterior 1")
    ws.append(["Stingray", "", "", ""])
    ws.append(["", "", "S = Standard Equipment  A = Available  -- = Not Available", ""])
    ws.append([
        "Orderable RPO Code",
        "Ref. Only RPO Code",
        "Description",
        "Coupe / 1YC07 / 1LT",
    ])
    ws.append(["Additional Options", "", "", ""])
    ws.append([
        "ERI",
        "",
        "Battery Protection Package / 1. Requires (ABC) Example Package. 2. Not available with (XYZ) Example Conflict.",
        "A1",
    ])
    ws.append(["", "EYT", "Carbon Flash Exterior Badge Package", "S"])

    ws = wb.create_sheet("Exterior 4")
    ws.append(["ZR1 and ZR1X", "", "", "", ""])
    ws.append(["", "", "S = Standard Equipment  A = Available  -- = Not Available", "", ""])
    ws.append([
        "Orderable RPO Code",
        "Ref. Only RPO Code",
        "Description",
        "ZR1 Coupe / 1YR07 / 1LZ",
        "ZR1X Coupe / 1YS07 / 1LZ",
    ])
    ws.append(["TOM", "", "Carbon Fiber Aero Package", "A", "--"])

    ws = wb.create_sheet("Color and Trim 1")
    ws.append(["Recommended", "", "", "", ""])
    ws.append(["A = Available  -- = Not Available", "", "", "", ""])
    ws.append(["", "", "", "Interior Colors", ""])
    ws.append(["Decor Level", "Seat Type", "Seat Code", "Seat Trim", "Jet Black"])
    ws.append(["1LT, 1LZ", "GT1 buckets", "AQ9", "Mulan leather", "HTA"])
    wb.save(path)


class OrderGuideIngestProfilerTests(unittest.TestCase):
    def test_status_parser_maps_availability_and_unavailable_variants(self) -> None:
        for raw in ("A/D", "A/D1", "a/d", "A / D", "A-D2"):
            parsed = parse_status(raw)
            self.assertEqual(parsed["parsed_base_status"], "available", raw)
            self.assertIn("dealer_installed_or_adi", parsed["status_flags"], raw)

        for raw in ("--", "--1", "—", "–", "- -"):
            parsed = parse_status(raw)
            self.assertEqual(parsed["parsed_base_status"], "unavailable", raw)

    def test_profiles_source_layout_variants_rows_and_disclosures(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook = tmp / "canonical.xlsx"
            raw_export = tmp / "raw.xlsx"
            output_dir = tmp / "profile-output"
            fixture_workbook(workbook)
            raw_export_fixture(raw_export)

            result = profile_order_guide(
                raw_export=raw_export,
                workbook=workbook,
                output_dir=output_dir,
                run_id="unit-test",
                root=ROOT,
            )

            self.assertEqual(result["status"], "passed")
            self.assertEqual(result["source_sheet_count"], 3)
            self.assertEqual(result["parsed_matrix_sheet_count"], 2)
            self.assertGreaterEqual(result["raw_row_count"], 3)
            self.assertTrue((output_dir / "source-layout.json").exists())
            self.assertTrue((output_dir / "variant-matrix.json").exists())
            self.assertTrue((output_dir / "raw-rows.json").exists())
            self.assertTrue((output_dir / "disclosure-links.json").exists())
            self.assertTrue((output_dir / "checkpoint-report.md").exists())

            layout = json.loads((output_dir / "source-layout.json").read_text())
            self.assertEqual(layout[0]["sheet_type"], "matrix")
            self.assertEqual(layout[0]["header_row"], 3)
            self.assertEqual(layout[2]["sheet_type"], "color_trim")

            variants = json.loads((output_dir / "variant-matrix.json").read_text())
            variant_ids = {row["parsed_variant_id"]: row for row in variants}
            self.assertEqual(variant_ids["1lt_c07"]["parsed_target_model"], "stingray")
            self.assertEqual(variant_ids["1lz_r07"]["parsed_target_model"], "zr1")
            self.assertEqual(variant_ids["1lz_s07"]["parsed_target_model"], "zr1x")

            raw_rows = json.loads((output_dir / "raw-rows.json").read_text())
            battery = next(row for row in raw_rows if row["primary_rpo_candidate"] == "ERI")
            self.assertEqual(battery["status_cells"][0]["raw_status"], "A1")
            self.assertEqual(battery["status_cells"][0]["parsed_base_status"], "available")
            self.assertEqual(battery["status_cells"][0]["status_marker"], "1")
            self.assertEqual(battery["description_disclosure_markers"], ["1", "2"])

            links = json.loads((output_dir / "disclosure-links.json").read_text())
            marker_one = next(link for link in links if link["source_row_index"] == 5 and link["marker"] == "1")
            self.assertEqual(marker_one["candidate_relationship_hint"], "requires")
            marker_two = next(link for link in links if link["source_row_index"] == 5 and link["marker"] == "2")
            self.assertEqual(marker_two["candidate_relationship_hint"], "excludes")

    def test_output_dir_guard_rejects_tracked_generated_output_paths(self) -> None:
        with self.assertRaisesRegex(ValueError, "form-output/ingest"):
            validate_output_dir(ROOT / "form-output" / "runtime", root=ROOT)
        with self.assertRaisesRegex(ValueError, "form-output/ingest"):
            validate_output_dir(ROOT / "form-output" / "stingray-form-data.json", root=ROOT)

    def test_output_dir_guard_allows_tmp_and_ingest_run_paths(self) -> None:
        validate_output_dir(Path(tempfile.gettempdir()) / "27vette-ingest-test", root=ROOT)
        validate_output_dir(ROOT / "form-output" / "ingest" / "unit-test", root=ROOT)


if __name__ == "__main__":
    unittest.main()
