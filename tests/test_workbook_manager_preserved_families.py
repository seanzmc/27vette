"""Checkpoint 2D-B — direct management of PriceRef and context_choice_copy.

Owns PRES-01–05 for the two approved families. Every test here either reads
the canonical workbook read-only or works on an isolated copy; nothing writes
to tracked or canonical data. Registry literals are recomputed from the
registry so a registry edit fails the gate for the right reason.
"""
from __future__ import annotations

import json
import shutil
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(ROOT / "workbook-manager" / "backend"))

from corvette_form_generator import editor_ops  # noqa: E402
from corvette_form_generator.schema_validation import REQUIRED_SHEETS  # noqa: E402
from corvette_form_generator.workbook_domain import registry  # noqa: E402
from app import apply_rebuild, catalog, db as dbmod, drafts, importer, main, staging  # noqa: E402

WORKBOOK = ROOT / "stingray_master.xlsx"
FAMILIES = ("price_ref", "context_choice_copy")
SHEETS = {"price_ref": "PriceRef", "context_choice_copy": "context_choice_copy"}
STILL_PRESERVED = ("rule_phrase_map", "runtime_rule_exceptions")


def _fresh_projection(workbook: Path) -> tuple[sqlite3.Connection, dict, Path]:
    tmp = Path(tempfile.mkdtemp(prefix="wbm-2db-"))
    conn = dbmod.connect(tmp / "projection.sqlite3")
    conn.row_factory = sqlite3.Row
    dbmod.init_projection_schema(conn)
    report = importer.import_workbook(conn, workbook)
    return conn, report, tmp


def _state(tmp: Path) -> sqlite3.Connection:
    state = dbmod.connect(tmp / "state.sqlite3")
    state.row_factory = sqlite3.Row
    dbmod.init_durable_schema(state)
    return state


def _workbook_rows(path: Path, sheet: str) -> list[tuple]:
    wb = load_workbook(path, read_only=True)
    try:
        return [tuple(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()


class TestRegistryContract(unittest.TestCase):
    """PRES-01 registry half: the exact approved contract, and nothing invented."""

    def test_families_are_registered_once_with_the_approved_keys(self):
        self.assertEqual(registry.EDITOR_SHEET_META["price_ref"]["key"], ("OptionType", "Trim", "Code"))
        self.assertEqual(
            registry.EDITOR_SHEET_META["context_choice_copy"]["key"],
            ("model_key", "context_type", "value", "body_style"),
        )
        self.assertEqual(registry.GLOBAL_SHEET_FAMILIES["PriceRef"], "price_ref")
        self.assertEqual(registry.GLOBAL_SHEET_FAMILIES["context_choice_copy"], "context_choice_copy")
        for family in STILL_PRESERVED:
            self.assertNotIn(family, registry.EDITOR_SHEET_META)
            self.assertNotIn(family, registry.GLOBAL_SHEET_FAMILIES)

    def test_price_ref_has_no_invented_option_type_enum(self):
        # pricing.py normalizes OptionType spellings and interior_components
        # authors both two_tone and twotone, so no closed enum is proven.
        self.assertEqual(registry.EDITOR_SHEET_META["price_ref"]["enums"], {})
        self.assertEqual(registry.FIELD_CONTROLS["price_ref"]["Price"]["kind"], "money")
        self.assertEqual(registry.FIELD_CONTROLS["price_ref"]["OptionType"]["kind"], "short_text")

    def test_optional_key_is_the_only_blank_permitting_key_column(self):
        """One rule: a key column may be blank iff the registry marks it optional."""
        blank_keys = {
            (family, column)
            for family, meta in registry.EDITOR_SHEET_META.items()
            for column in meta["key"]
            if column in meta["optional_columns"]
        }
        self.assertEqual(blank_keys, {("price_ref", "Trim")})
        self.assertEqual(registry.FIELD_CONTROLS["price_ref"]["Trim"]["blank"], "optional_key")
        self.assertTrue(registry.FIELD_CONTROLS["price_ref"]["Trim"]["immutable_on_edit"])
        for family, meta in registry.EDITOR_SHEET_META.items():
            for column in meta["key"]:
                if (family, column) == ("price_ref", "Trim"):
                    continue
                self.assertEqual(
                    registry.FIELD_CONTROLS[family][column]["blank"], "never_blank_key", (family, column)
                )

    def test_context_choice_copy_domains_come_from_the_consumer(self):
        meta = registry.EDITOR_SHEET_META["context_choice_copy"]
        # contract.py:172,207 call with exactly these two context types.
        self.assertEqual(meta["enums"]["context_type"], ("body_style", "trim_level"))
        self.assertEqual(meta["enums"]["body_style"], ("*", "coupe", "convertible"))
        self.assertEqual(meta["conditional_ref"], {"discriminator": "context_type", "column": "value"})
        self.assertEqual(
            meta["conditional_refs"],
            {"trim_level": "variant_trim_levels", "body_style": "variant_body_styles"},
        )
        self.assertEqual(meta["types"], {"active": "bool"})
        self.assertEqual(
            registry.EDITOR_SHEET_META["context_choice_copy"]["required_on_add"],
            ("model_key", "context_type", "value", "body_style", "active"),
        )
        # contract.py:113 drops active rows with a blank tooltip, so the
        # tooltip is required exactly once the row is effectively active.
        self.assertIn(
            "info_tooltip",
            registry.EDITOR_SHEET_META["context_choice_copy"]["required_on_effective_active_row"],
        )
        self.assertEqual(registry.FIELD_CONTROLS["context_choice_copy"]["info_tooltip"]["kind"], "long_text")

    def test_wildcard_model_families_are_registry_owned(self):
        self.assertEqual(registry.WILDCARD_MODEL_FAMILIES, frozenset({"asset_map", "context_choice_copy"}))


class TestCatalogAndClassification(unittest.TestCase):
    """PRES-01 projection half and the KNOWN_PRESERVED_SHEETS move."""

    def test_sheets_moved_out_of_preserved_and_into_managed(self):
        self.assertEqual(set(catalog.KNOWN_PRESERVED_SHEETS), set(STILL_PRESERVED))
        wb = load_workbook(WORKBOOK, read_only=True)
        try:
            classes = catalog.classify_workbook_sheets(wb)
        finally:
            wb.close()
        for family, sheet in SHEETS.items():
            self.assertEqual(classes[sheet].disposition, "managed_writable", sheet)
            self.assertEqual(classes[sheet].family, family)
        for sheet in STILL_PRESERVED:
            self.assertEqual(classes[sheet].disposition, "workbook_preserved_known")
        self.assertIn("PriceRef", REQUIRED_SHEETS)

    def test_routing_and_surface_classification(self):
        self.assertEqual(catalog.SPEC_BY_FAMILY["price_ref"].sheet, ("PriceRef",))
        self.assertFalse(catalog.SPEC_BY_FAMILY["price_ref"].has_model_key_column)
        self.assertIn("price_ref", catalog.SHARED_TABLES)
        self.assertTrue(catalog.SPEC_BY_FAMILY["context_choice_copy"].has_model_key_column)
        self.assertIn("context_choice_copy", catalog.MODEL_COLLECTIONS)
        for family in FAMILIES:
            self.assertNotIn(catalog.SPEC_BY_FAMILY[family].table, catalog.STRUCTURE_TABLES)

    def test_derived_reference_domains_present_as_variant_columns(self):
        for name, column in (("variant_trim_levels", "trim_level"), ("variant_body_styles", "body_style")):
            presentation = catalog.REFERENCE_OPTION_PRESENTATION[name]
            self.assertEqual(presentation["table"], "variants")
            self.assertEqual(presentation["value"], column)


class TestProjectionAndOwnership(unittest.TestCase):
    """PRES-02 import half plus the D1/D2 ownership rule."""

    @classmethod
    def setUpClass(cls):
        cls.conn, cls.report, cls.tmp = _fresh_projection(WORKBOOK)

    @classmethod
    def tearDownClass(cls):
        cls.conn.close()
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def test_every_row_projects_with_lineage_and_no_errors(self):
        errors = [i for i in self.report["issues"] if i["severity"] == "error"
                  and i["sheet"] in SHEETS.values()]
        self.assertEqual(errors, [])
        self.assertEqual(self.conn.execute("SELECT COUNT(*) c FROM price_ref").fetchone()["c"], 21)
        self.assertEqual(self.conn.execute("SELECT COUNT(*) c FROM context_choice_copy").fetchone()["c"], 6)
        blank_trim = self.conn.execute(
            "SELECT COUNT(*) c FROM price_ref WHERE trim IS NULL"
        ).fetchone()["c"]
        self.assertEqual(blank_trim, 7)
        row = self.conn.execute(
            "SELECT * FROM price_ref WHERE optiontype='Stitching' AND code='36S'"
        ).fetchone()
        self.assertEqual(row["src_sheet"], "PriceRef")
        self.assertEqual(row["src_row"], 16)
        self.assertEqual(json.loads(row["physical_key"]), ["Stitching", "", "36S"])
        self.assertEqual(row["price"], "495")

    def test_global_and_wildcard_rows_are_owned_by_every_active_model(self):
        active = sorted(
            r["model_key"] for r in self.conn.execute("SELECT model_key FROM models WHERE active='True'")
        )
        self.assertEqual(len(active), 6)
        # Fixed families without a model_key column (PriceRef, variant_master)
        # are owned by every active model; role-registered shared sheets such
        # as interiors keep their exact source-registration ownership.
        for table in ("price_ref", "variants"):
            contexts = {
                r["model_context"] for r in self.conn.execute(f'SELECT model_context FROM "{table}"')
            }
            self.assertEqual(contexts, {json.dumps(active)}, table)
        interior_contexts = {
            r["model_context"] for r in self.conn.execute('SELECT model_context FROM "interiors"')
        }
        self.assertEqual(
            interior_contexts,
            {'["grand_sport", "grand_sport_x", "stingray"]', '["z06", "zr1", "zr1x"]'},
        )
        wildcard = {
            r["model_context"] for r in self.conn.execute(
                "SELECT model_context FROM context_choice_copy WHERE model_key='*'"
            )
        }
        self.assertEqual(wildcard, {json.dumps(active)})
        concrete = {
            r["model_context"] for r in self.conn.execute(
                "SELECT model_context FROM context_choice_copy WHERE model_key='z06'"
            )
        }
        self.assertEqual(concrete, {'["z06"]'})
        asset_wildcard = {
            r["model_context"] for r in self.conn.execute(
                "SELECT model_context FROM assets WHERE model_key='*'"
            )
        }
        self.assertEqual(asset_wildcard, {json.dumps(active)})

    def test_affected_models_are_never_empty_for_a_real_operation(self):
        promoted = ["stingray", "grand_sport", "grand_sport_x", "z06", "zr1", "zr1x"]
        row = self.conn.execute("SELECT model_context FROM price_ref LIMIT 1").fetchone()
        derived = apply_rebuild.derive_affected_models(
            [{"model_id": "", "model_context": json.loads(row["model_context"])}],
            promoted_models=promoted,
        )
        self.assertEqual(derived, sorted(promoted))
        row = self.conn.execute(
            "SELECT model_context FROM context_choice_copy WHERE model_key='*' LIMIT 1"
        ).fetchone()
        derived = apply_rebuild.derive_affected_models(
            [{"model_id": "*", "model_context": json.loads(row["model_context"])}],
            promoted_models=promoted,
        )
        self.assertEqual(derived, sorted(promoted))

    def test_context_choice_reference_options_keep_workbook_trim_casing(self):
        response = main._reference_options(
            self.conn,
            catalog.SPEC_BY_TABLE["context_choice_copy"],
            "value",
            "stingray",
            "",
            "trim_level",
            100,
            0,
        )
        self.assertEqual(
            [option["value"] for option in response["options"]],
            ["1LT", "2LT", "3LT"],
        )

    def test_wildcard_write_allowed_only_for_registered_families(self):
        allowed = staging._editable_guard(
            self.conn, catalog.SPEC_BY_TABLE["context_choice_copy"], "*",
            op="add", key={"model_key": "*"}, record={"model_key": "*"},
        )
        self.assertEqual(allowed, [])
        refused = staging._editable_guard(
            self.conn, catalog.SPEC_BY_TABLE["default_selection_rules"], "*",
            op="add", key={"model_key": "*"}, record={"model_key": "*"},
        )
        self.assertTrue(refused)
        self.assertIn("wildcard", refused[0]["message"])


class TestDraftOperations(unittest.TestCase):
    """PRES-03 draft half (M2–M13): behavior and persisted draft results."""

    def setUp(self):
        self.conn, _, self.tmp = _fresh_projection(WORKBOOK)
        self.state = _state(self.tmp)
        self.identity = {
            "projection_state": "current",
            "base_workbook_sha256": dbmod.get_meta(self.conn, "workbook_sha256"),
            "base_workbook_mtime_ns": dbmod.get_meta(self.conn, "workbook_mtime_ns"),
        }

    def tearDown(self):
        self.conn.close()
        self.state.close()
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _save(self, draft_id, **kwargs):
        kwargs.setdefault("record", None)
        return drafts.save_operation(self.conn, self.state, draft_id=draft_id, **self.identity, **kwargs)

    def test_price_ref_update_and_blank_trim_add_persist_exact_intent(self):
        op = self._save("d1", table="price_ref", model_id="", op="update",
                        key={"optiontype": "Seat", "trim": "1LT", "code": "AE4"},
                        record={"price": "1195"})
        self.assertEqual(op["changed_fields"], {"price": {"before": "1095", "after": "1195"}})
        self.assertEqual(sorted(op["model_context"]), sorted(
            r["model_key"] for r in self.conn.execute("SELECT model_key FROM models WHERE active='True'")
        ))
        # M2: unreferenced component row with a blank Trim key
        op = self._save("d1", table="price_ref", model_id="", op="add",
                        key={"optiontype": "Stitching", "trim": "", "code": "39S"},
                        record={"optiontype": "Stitching", "trim": "", "code": "39S", "price": "495"})
        self.assertEqual(op["action"], "add")
        self.assertEqual(op["final"]["trim"], None)
        self.assertEqual(json.loads(op["physical_key"]), ["Stitching", "", "39S"])

    def test_price_ref_refuses_blank_required_key_and_duplicate(self):
        with self.assertRaises(drafts.DraftError) as ctx:
            self._save("d2", table="price_ref", model_id="", op="add",
                       key={"optiontype": "Stitching", "trim": "", "code": ""},
                       record={"optiontype": "Stitching", "trim": "", "code": "", "price": "1"})
        self.assertEqual(ctx.exception.code, "invalid_entity_key")
        with self.assertRaises(drafts.DraftError) as ctx:
            self._save("d2", table="price_ref", model_id="", op="add",
                       key={"optiontype": "Seat", "trim": "1LT", "code": "AE4"},
                       record={"optiontype": "Seat", "trim": "1LT", "code": "AE4", "price": "1"})
        self.assertEqual(ctx.exception.code, "duplicate_record")
        # Blank-Trim duplicate is also a duplicate.
        with self.assertRaises(drafts.DraftError) as ctx:
            self._save("d2", table="price_ref", model_id="", op="add",
                       key={"optiontype": "Suede", "trim": "", "code": "N26"},
                       record={"optiontype": "Suede", "trim": "", "code": "N26", "price": "1"})
        self.assertEqual(ctx.exception.code, "duplicate_record")

    def test_context_choice_copy_wildcard_row_is_writable_and_widely_owned(self):
        key = {"model_key": "*", "context_type": "trim_level", "value": "1LT", "body_style": "*"}
        op = self._save("d3", table="context_choice_copy", model_id="*", op="update",
                        key=key, record={"info_tooltip": "1LT test copy"})
        self.assertEqual(op["changed_fields"]["info_tooltip"]["after"], "1LT test copy")
        self.assertEqual(len(op["model_context"]), 6)
        emitted = drafts.emit_changeset(self.state, draft_id="d3")
        self.assertEqual(len(emitted["targets"]), 6)
        self.assertNotIn("*", emitted["targets"])

    def test_context_choice_copy_concrete_add_and_full_reversion(self):
        key = {"model_key": "zr1", "context_type": "trim_level", "value": "1LZ", "body_style": "*"}
        record = {**key, "info_tooltip": "ZR1 1LZ copy", "active": "True", "notes": ""}
        op = self._save("d4", table="context_choice_copy", model_id="zr1", op="add", key=key, record=record)
        self.assertEqual(op["model_context"], ["zr1"])
        # M12: delete of a draft-added row coalesces to nothing
        result = self._save("d4", table="context_choice_copy", model_id="zr1", op="delete", key=key)
        self.assertIsNone(result)
        self.assertEqual(drafts.list_operations(self.state, "d4"), [])

    def test_key_change_and_unknown_field_fail_closed(self):
        key = {"model_key": "z06", "context_type": "trim_level", "value": "1LZ", "body_style": "*"}
        with self.assertRaises(drafts.DraftError) as ctx:
            self._save("d5", table="context_choice_copy", model_id="z06", op="update",
                       key=key, record={"value": "2LZ"})
        self.assertEqual(ctx.exception.code, "key_change_rejected")
        with self.assertRaises(drafts.DraftError) as ctx:
            self._save("d5", table="price_ref", model_id="", op="update",
                       key={"optiontype": "Seat", "trim": "1LT", "code": "AE4"},
                       record={"notes": "x"})
        self.assertEqual(ctx.exception.code, "unknown_fields")


class TestWorkbookWriteAndParity(unittest.TestCase):
    """PRES-02 export half, PRES-03 workbook half, PRES-04 (M1, M3, M11, M15)."""

    def _apply(self, workbook: Path, items: list[dict]) -> dict:
        identity = importer.workbook_identity(workbook)
        return editor_ops.apply_batch(
            workbook,
            {"workbookMtimeNs": str(identity["mtime_ns"]), "workbookSha256": identity["sha256"],
             "items": items},
            write=True,
            source="test-2d-b",
            log_path=workbook.parent / "edits.jsonl",
        )

    def test_unchanged_export_is_byte_identical_and_projects_equal(self):
        with tempfile.TemporaryDirectory(prefix="wbm-2db-m1-") as raw:
            copy = Path(raw) / "copy.xlsx"
            shutil.copy2(WORKBOOK, copy)
            conn, _, tmp = _fresh_projection(copy)
            try:
                ok_package, ok_schema, semantic_equal, issues = importer._validate_reconstruction(copy, conn)
                self.assertTrue(ok_package and ok_schema and semantic_equal, issues)
            finally:
                conn.close()
                shutil.rmtree(tmp, ignore_errors=True)
            self.assertEqual(copy.read_bytes(), WORKBOOK.read_bytes())

    def test_isolated_edits_touch_exactly_the_intended_cells(self):
        with tempfile.TemporaryDirectory(prefix="wbm-2db-m3-") as raw:
            copy = Path(raw) / "copy.xlsx"
            shutil.copy2(WORKBOOK, copy)
            before = {s: _workbook_rows(copy, s) for s in ("PriceRef", "context_choice_copy", "rule_phrase_map")}
            result = self._apply(copy, [
                {"action": "update", "sheet": "PriceRef",
                 "key": {"OptionType": "Seat", "Trim": "1LT", "Code": "AE4"}, "row": {"Price": 1195}},
                {"action": "add", "sheet": "PriceRef",
                 "key": {"OptionType": "Stitching", "Trim": "", "Code": "39S"},
                 "row": {"OptionType": "Stitching", "Trim": None, "Code": "39S", "Price": 495}},
                {"action": "update", "sheet": "context_choice_copy",
                 "key": {"model_key": "*", "context_type": "trim_level", "value": "1LT", "body_style": "*"},
                 "row": {"info_tooltip": "1LT test copy"}},
                {"action": "delete", "sheet": "context_choice_copy",
                 "key": {"model_key": "z06", "context_type": "trim_level", "value": "3LZ", "body_style": "*"}},
            ])
            self.assertTrue(result["ok"], result)
            after = {s: _workbook_rows(copy, s) for s in before}
            self.assertEqual(after["rule_phrase_map"], before["rule_phrase_map"])
            price_diff = [(i, a, b) for i, (a, b) in enumerate(zip(before["PriceRef"], after["PriceRef"])) if a != b]
            self.assertEqual(price_diff, [(1, ("Seat", "1LT", "AE4", 1095), ("Seat", "1LT", "AE4", 1195))])
            self.assertEqual(after["PriceRef"][-1], ("Stitching", None, "39S", 495))
            self.assertEqual(len(after["PriceRef"]), len(before["PriceRef"]) + 1)
            wb = load_workbook(copy)
            try:
                self.assertEqual(str(wb["PriceRef"].tables["Table_8"].ref), "A1:D23")
                self.assertEqual(wb["PriceRef"]["D2"].data_type, "n")
                self.assertEqual(wb["context_choice_copy"]["F2"].value, "True")
            finally:
                wb.close()
            copy_rows = after["context_choice_copy"]
            self.assertEqual(len(copy_rows), len(before["context_choice_copy"]) - 1)
            self.assertEqual(copy_rows[1][4], "1LT test copy")
            self.assertFalse(any(r[2] == "3LZ" for r in copy_rows[1:]))

    def test_blank_trim_seat_and_duplicate_normalized_key_fail_closed_in_the_writer(self):
        with tempfile.TemporaryDirectory(prefix="wbm-2db-m5-") as raw:
            copy = Path(raw) / "copy.xlsx"
            shutil.copy2(WORKBOOK, copy)
            result = self._apply(copy, [
                {"action": "add", "sheet": "PriceRef",
                 "key": {"OptionType": "Seat", "Trim": "", "Code": "AQ9"},
                 "row": {"OptionType": "Seat", "Trim": None, "Code": "AQ9", "Price": 100}},
            ])
            self.assertFalse(result["ok"])
            self.assertTrue(any("Seat rows require Trim" in e for e in result["errors"]), result["errors"])
            result = self._apply(copy, [
                {"action": "add", "sheet": "PriceRef",
                 "key": {"OptionType": "Two Tone", "Trim": "", "Code": "TU7"},
                 "row": {"OptionType": "Two Tone", "Trim": None, "Code": "TU7", "Price": 1}},
            ])
            self.assertFalse(result["ok"])
            self.assertTrue(any("normalized PriceRef key" in e for e in result["errors"]), result["errors"])
            result = self._apply(copy, [
                {"action": "add", "sheet": "context_choice_copy",
                 "key": {"model_key": "*", "context_type": "colour", "value": "1LT", "body_style": "*"},
                 "row": {"model_key": "*", "context_type": "colour", "value": "1LT", "body_style": "*",
                         "info_tooltip": "x", "active": True}},
            ])
            self.assertFalse(result["ok"])
            result = self._apply(copy, [
                {"action": "add", "sheet": "context_choice_copy",
                 "key": {"model_key": "*", "context_type": "trim_level", "value": "4LT", "body_style": "*"},
                 "row": {"model_key": "*", "context_type": "trim_level", "value": "4LT", "body_style": "*",
                         "info_tooltip": "x", "active": True}},
            ])
            self.assertFalse(result["ok"])
            self.assertTrue(any("value=" in e and "4LT" in e for e in result["errors"]), result["errors"])
            self.assertEqual(copy.read_bytes(), WORKBOOK.read_bytes())

    def test_context_choice_reference_is_scoped_to_active_model_variants(self):
        with tempfile.TemporaryDirectory(prefix="wbm-2db-model-ref-") as raw:
            copy = Path(raw) / "copy.xlsx"
            shutil.copy2(WORKBOOK, copy)
            result = self._apply(copy, [
                {"action": "add", "sheet": "context_choice_copy",
                 "key": {"model_key": "stingray", "context_type": "trim_level", "value": "1LZ", "body_style": "*"},
                 "row": {"model_key": "stingray", "context_type": "trim_level", "value": "1LZ", "body_style": "*",
                         "info_tooltip": "unreachable copy", "active": True}},
            ])
            self.assertFalse(result["ok"])
            self.assertTrue(any("value=" in error and "1LZ" in error for error in result["errors"]), result["errors"])
            self.assertEqual(copy.read_bytes(), WORKBOOK.read_bytes())


if __name__ == "__main__":
    unittest.main()
