#!/usr/bin/env python3
"""Fault-injection and lifecycle tests for workbook_domain.service.

Fixture workbooks only. The compact build_ops_fixture workbook is not
schema-clean, so schema scans are mocked clean exactly like
tests/test_editor_ops_apply.py does; the live workbook is never used.
"""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

import apply_workbook_changeset as cli  # noqa: E402
import corvette_form_generator.editor_ops as editor_ops  # noqa: E402
from corvette_form_generator.workbook_domain.changeset import (  # noqa: E402
    changeset_fingerprint,
)
from corvette_form_generator.workbook_domain.service import (  # noqa: E402
    apply_changeset,
    approve_changeset,
    preview_changeset,
)
from test_editor_ops_apply import build_ops_fixture  # noqa: E402


@pytest.fixture(autouse=True)
def _schema_scan_clean(monkeypatch):
    """The compact fixture is not schema-clean; mock the scan like ApplyBatchTest."""
    monkeypatch.setattr(
        editor_ops, "validate_workbook_schema", lambda *args, **kwargs: []
    )


def make_workbook(tmp_path):
    path = tmp_path / "fixture.xlsx"
    build_ops_fixture().save(path)
    return path


def make_valid_changeset(path):
    payload = {
        "schemaVersion": "workbook-changeset-1",
        "source": {"kind": "editor", "runId": "service-test"},
        "targets": ["stingray"],
        "workbook": {
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            "mtimeNs": str(path.stat().st_mtime_ns),
        },
        "sheetCreates": [],
        "rowChanges": [{
            "action": "update",
            "sheet": "stingray_options",
            "family": "options",
            "key": {"option_id": "opt_one_001"},
            # build_ops_fixture sets price 0 on every option row.
            "fields": {"price": {"before": 0, "after": 101}},
            "provenance": [{"kind": "editor", "id": "service-test:price"}],
        }],
        "noops": [],
        "warningAcknowledgementsRequested": [],
        "bindings": {},
    }
    payload["semanticFingerprint"] = changeset_fingerprint(payload)
    payload["changeSetId"] = payload["semanticFingerprint"][:24]
    return payload


def _resign(payload):
    payload["semanticFingerprint"] = changeset_fingerprint(payload)
    payload["changeSetId"] = payload["semanticFingerprint"][:24]
    return payload


def _zr1_scaffold_changeset(path):
    """Update on the non-promoted zr1 sheet emits confirmable scaffold warning."""
    payload = make_valid_changeset(path)
    payload["rowChanges"] = [{
        "action": "update",
        "sheet": "zr1_options",
        "family": "options",
        "key": {"option_id": "opt_zzz_001"},
        "fields": {"price": {"before": 0, "after": 1}},
        "provenance": [{"kind": "editor", "id": "service-test:zzz-price"}],
    }]
    return _resign(payload)


def _display_order_changeset(path):
    """display_order update emits a blocking ``dorder:`` warning."""
    payload = make_valid_changeset(path)
    payload["rowChanges"] = [{
        "action": "update",
        "sheet": "stingray_options",
        "family": "options",
        "key": {"option_id": "opt_two_001"},
        "fields": {"display_order": {"before": 20, "after": 10}},
        "provenance": [{"kind": "editor", "id": "service-test:dorder"}],
    }]
    return _resign(payload)


# ── Plan Step 1 fault-injection tests ─────────────────────────────────


def test_live_apply_rechecks_original_reviewed_fingerprint(tmp_path, monkeypatch):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    assert preview["ok"], preview
    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])
    assert approval["ok"], approval

    original_prepare = editor_ops._prepare_batch

    def mutate_after_prepare(extract, batch):
        result = original_prepare(extract, batch)
        workbook.touch()
        return result

    monkeypatch.setattr(editor_ops, "_prepare_batch", mutate_after_prepare)

    receipt = apply_changeset(
        workbook, changeset, preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )
    assert receipt["status"] == "stale_before_save"
    assert receipt["workbookState"] == "untouched"


def test_failed_live_readback_restores_and_verifies_backup(tmp_path, monkeypatch):
    workbook = make_workbook(tmp_path)
    before = workbook.read_bytes()
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    assert preview["ok"], preview
    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])
    assert approval["ok"], approval

    # Fail only the LIVE readback: the first (scratch) call delegates to the
    # real verifier so the write path proceeds to the live save.
    real_verify = editor_ops.verify_prepared_workbook
    calls = {"count": 0}

    def fail_live_verify(path, prepared):
        calls["count"] += 1
        if calls["count"] == 1:
            return real_verify(path, prepared)
        return {
            "ok": False,
            "preparedChecked": 0,
            "preparedCount": 1,
            "errors": ["forced"],
        }

    monkeypatch.setattr(editor_ops, "verify_prepared_workbook", fail_live_verify)

    receipt = apply_changeset(
        workbook, changeset, preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )
    assert receipt["status"] == "apply_verification_failed_rolled_back"
    assert receipt["workbookState"] == "restored"
    assert workbook.read_bytes() == before


# ── Service lifecycle tests ───────────────────────────────────────────


def test_preview_approve_apply_happy_path(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    assert preview["ok"], preview
    assert preview["status"] == "validated"
    assert preview["schemaVersion"] == "workbook-change-preview-1"
    assert preview["semanticFingerprint"] == changeset["semanticFingerprint"]
    assert preview["previewFingerprint"]

    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])
    assert approval["ok"], approval
    assert approval["schemaVersion"] == "workbook-change-approval-1"
    assert approval["previewFingerprint"] == preview["previewFingerprint"]
    assert approval["approvalFingerprint"]
    assert approval["acceptedWarningIds"] == []

    receipt = apply_changeset(
        workbook, changeset, preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )
    assert receipt["ok"], receipt
    assert receipt["status"] == "applied"
    assert receipt["workbookState"] == "saved"
    assert receipt["schemaVersion"] == "workbook-change-receipt-1"
    assert Path(receipt["backupPath"]).exists()
    assert (tmp_path / "edit-log.jsonl").exists()

    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["stingray_options"]
    headers = {cell.value: index + 1 for index, cell in enumerate(ws[1])}
    price = next(
        ws.cell(row=row, column=headers["price"]).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=headers["option_id"]).value == "opt_one_001"
    )
    wb.close()
    assert price == 101


def test_apply_refuses_mismatched_preview_binding(tmp_path):
    workbook = make_workbook(tmp_path)
    before = workbook.read_bytes()
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])

    tampered_preview = dict(preview)
    tampered_preview["previewFingerprint"] = "0" * 64
    receipt = apply_changeset(
        workbook, changeset, tampered_preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )
    assert receipt["ok"] is False
    assert receipt["status"] == "binding_mismatch"
    assert receipt["workbookState"] == "untouched"
    assert workbook.read_bytes() == before


def test_approve_refuses_preview_changed_after_fingerprinting(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    preview["createdAt"] = "tampered-after-preview"

    approval = approve_changeset(
        changeset, preview, actor="Sean", warning_ids=[],
    )

    assert approval["ok"] is False
    assert approval["status"] == "binding_mismatch"


def test_apply_refuses_preview_changed_after_approval(tmp_path):
    workbook = make_workbook(tmp_path)
    before = workbook.read_bytes()
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])
    preview["createdAt"] = "tampered-after-approval"

    receipt = apply_changeset(
        workbook, changeset, preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )

    assert receipt["ok"] is False
    assert receipt["status"] == "binding_mismatch"
    assert receipt["workbookState"] == "untouched"
    assert workbook.read_bytes() == before


def test_apply_refuses_approval_changed_after_fingerprinting(tmp_path):
    workbook = make_workbook(tmp_path)
    before = workbook.read_bytes()
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])
    approval["actor"] = "tampered-after-approval"

    receipt = apply_changeset(
        workbook, changeset, preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )

    assert receipt["ok"] is False
    assert receipt["status"] == "approval_invalid"
    assert receipt["workbookState"] == "untouched"
    assert workbook.read_bytes() == before


def test_approve_requires_exact_warning_acceptance(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = _zr1_scaffold_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    assert preview["ok"], preview
    assert preview["warningPolicy"]["confirmableIds"] == ["scaffold:zr1_options"]

    refusal = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])
    assert refusal["ok"] is False

    extra = approve_changeset(
        changeset, preview, actor="Sean",
        warning_ids=["scaffold:zr1_options", "scaffold:bogus"],
    )
    assert extra["ok"] is False

    approval = approve_changeset(
        changeset, preview, actor="Sean",
        warning_ids=["scaffold:zr1_options"],
    )
    assert approval["ok"], approval
    assert approval["acceptedWarningIds"] == ["scaffold:zr1_options"]


def test_approve_refuses_blocking_warnings(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = _display_order_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    assert preview["ok"], preview
    blocking = preview["warningPolicy"]["blockingIds"]
    assert blocking and all(str(wid).startswith("dorder:") for wid in blocking)

    refusal = approve_changeset(
        changeset, preview, actor="Sean", warning_ids=list(blocking),
    )
    assert refusal["ok"] is False


def test_preview_refuses_stale_workbook(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    # Re-saving the fixture changes both mtime and bytes after signing.
    build_ops_fixture().save(workbook)
    preview = preview_changeset(workbook, changeset)
    assert preview["ok"] is False
    assert preview["status"] == "stale"


def test_preview_refuses_invalid_changeset(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    changeset["semanticFingerprint"] = "0" * 64
    preview = preview_changeset(workbook, changeset)
    assert preview["ok"] is False
    assert preview["status"] == "invalid_changeset"


def test_failed_restore_reports_workbook_restore_failed(tmp_path, monkeypatch):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])

    real_verify = editor_ops.verify_prepared_workbook
    calls = {"count": 0}

    def fail_live_verify(path, prepared):
        calls["count"] += 1
        if calls["count"] == 1:
            return real_verify(path, prepared)
        return {
            "ok": False,
            "preparedChecked": 0,
            "preparedCount": 1,
            "errors": ["forced"],
        }

    monkeypatch.setattr(editor_ops, "verify_prepared_workbook", fail_live_verify)

    def broken_restore(path, backup_path):
        raise RuntimeError("backup unreadable")

    monkeypatch.setattr(editor_ops, "restore_workbook_backup", broken_restore)

    receipt = apply_changeset(
        workbook, changeset, preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )
    assert receipt["ok"] is False
    assert receipt["status"] == "workbook_restore_failed"
    assert receipt["workbookState"] == "unknown"
    assert str(workbook) in receipt["errors"][0]
    assert receipt["backupPath"] in receipt["errors"][0]

    # No restore occurred: the live save remains applied (price 101).
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["stingray_options"]
    headers = {cell.value: index + 1 for index, cell in enumerate(ws[1])}
    price = next(
        ws.cell(row=row, column=headers["price"]).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=headers["option_id"]).value == "opt_one_001"
    )
    wb.close()
    assert price == 101


def test_apply_refuses_workbook_drifted_after_approval(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    preview = preview_changeset(workbook, changeset)
    approval = approve_changeset(changeset, preview, actor="Sean", warning_ids=[])

    # Workbook drifts between approval and apply.
    build_ops_fixture().save(workbook)
    drifted = workbook.read_bytes()

    receipt = apply_changeset(
        workbook, changeset, preview, approval,
        log_path=tmp_path / "edit-log.jsonl",
    )
    assert receipt["ok"] is False
    assert receipt["status"] == "stale"
    assert receipt["workbookState"] == "untouched"
    assert workbook.read_bytes() == drifted


# ── Plan Step 6 operator CLI tests ────────────────────────────────────


def _write_changeset_json(tmp_path, changeset, name="change-set.json"):
    path = tmp_path / name
    with path.open("w", encoding="utf-8") as handle:
        json.dump(changeset, handle)
    return path


def _fixture_price(workbook, option_id):
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["stingray_options"]
    headers = {cell.value: index + 1 for index, cell in enumerate(ws[1])}
    price = next(
        ws.cell(row=row, column=headers["price"]).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=headers["option_id"]).value == option_id
    )
    wb.close()
    return price


def test_cli_preview_approve_write_happy_path(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset_path = _write_changeset_json(tmp_path, make_valid_changeset(workbook))

    preview_path = tmp_path / "preview.json"
    rc = cli.main([
        str(changeset_path), "--workbook", str(workbook),
        "--preview-out", str(preview_path),
    ])
    assert rc == 0
    preview = json.loads(preview_path.read_text(encoding="utf-8"))
    assert preview["ok"] is True
    assert preview["status"] == "validated"

    approval_path = tmp_path / "approval.json"
    rc = cli.main([
        str(changeset_path), "--workbook", str(workbook),
        "--approve", "Sean", "--preview", str(preview_path),
        "--approval-out", str(approval_path),
    ])
    assert rc == 0
    approval = json.loads(approval_path.read_text(encoding="utf-8"))
    assert approval["ok"] is True

    receipt_path = tmp_path / "receipt.json"
    rc = cli.main([
        str(changeset_path), "--workbook", str(workbook),
        "--write", "--preview", str(preview_path),
        "--approval", str(approval_path),
        "--receipt-out", str(receipt_path),
        "--log-path", str(tmp_path / "edit-log.jsonl"),
    ])
    assert rc == 0
    receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
    assert receipt["ok"] is True
    assert receipt["status"] == "applied"
    assert receipt["workbookState"] == "saved"
    assert _fixture_price(workbook, "opt_one_001") == 101


def test_cli_write_refuses_without_approval_file(tmp_path, capsys):
    workbook = make_workbook(tmp_path)
    changeset_path = _write_changeset_json(tmp_path, make_valid_changeset(workbook))
    preview_path = tmp_path / "preview.json"
    assert cli.main([
        str(changeset_path), "--workbook", str(workbook),
        "--preview-out", str(preview_path),
    ]) == 0

    before = workbook.read_bytes()
    rc = cli.main([
        str(changeset_path), "--workbook", str(workbook),
        "--write", "--preview", str(preview_path),
    ])
    assert rc == 2
    assert "--approval" in capsys.readouterr().err
    assert workbook.read_bytes() == before


def test_cli_approve_requires_preview_artifact(tmp_path, capsys):
    workbook = make_workbook(tmp_path)
    changeset_path = _write_changeset_json(tmp_path, make_valid_changeset(workbook))
    before = workbook.read_bytes()
    rc = cli.main([
        str(changeset_path), "--workbook", str(workbook), "--approve", "Sean",
    ])
    assert rc == 2
    assert "--preview" in capsys.readouterr().err
    assert workbook.read_bytes() == before


def test_cli_preview_default_invalid_changeset_exits_1(tmp_path):
    workbook = make_workbook(tmp_path)
    changeset = make_valid_changeset(workbook)
    changeset["semanticFingerprint"] = "0" * 64  # tampered after signing
    changeset_path = _write_changeset_json(tmp_path, changeset)
    rc = cli.main([
        str(changeset_path), "--workbook", str(workbook),
        "--preview-out", str(tmp_path / "preview.json"),
    ])
    assert rc == 1
