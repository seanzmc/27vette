#!/usr/bin/env python3
"""Tests for workbook-owned runtime registry promotion metadata."""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook_domain.registry import (  # noqa: E402
    DEFAULT_REGISTRY_PROMOTION_ARTIFACT_TYPE,
    REGISTRY_PROMOTION_ARTIFACT_TYPES,
)
from corvette_form_generator.registry_promotion import (  # noqa: E402
    MODEL_REGISTRY_PROMOTION_HEADERS,
    build_registry_from_artifacts,
    live_contract_data,
    load_registry_promotions,
)


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows or []:
        ws.append([row.get(header, None) for header in headers])


SETUP_COPY = {
    "setup_card_subtitle": "Workbook-authored card subtitle",
    "setup_eyebrow": "WORKBOOK-AUTHORED EYEBROW",
    "setup_title": "Workbook-authored title",
    "setup_description": "Workbook-authored description.",
    "setup_fact_1": "Fact one",
    "setup_fact_2": "Fact two",
    "setup_fact_3": "Fact three",
}


def runtime_contract_data(
    model_label: str,
    source_sheet: str,
    *,
    choices: list[dict[str, object]] | None = None,
    rules: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    data: dict[str, object] = {
        "dataset": {
            "name": f"2027 Corvette {model_label} operational form",
            "model": model_label,
            "model_year": "2027",
            "source_sheet": source_sheet,
            "status": "runtime_active",
        },
        "orderSummary": {},
    }
    for field in (
        "variants",
        "steps",
        "sections",
        "contextChoices",
        "choices",
        "standardEquipment",
        "ruleGroups",
        "exclusiveGroups",
        "rules",
        "priceRules",
        "interiors",
        "colorOverrides",
        "defaultSelectionRules",
        "validation",
    ):
        data[field] = []
    data["variants"] = [{"variant_id": "test-variant"}]
    data["steps"] = [{"step_key": "test-step"}]
    data["sections"] = [{"section_id": "test-section"}]
    data["contextChoices"] = [{"context_choice_id": "test-context"}]
    data["choices"] = choices or [{"choice_id": "test-choice"}]
    data["rules"] = rules or []
    return data


def workbook_with_promotions(promotion_rows: list[dict[str, object]] | None = None) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "model_master",
        [
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "setup_card_subtitle",
            "setup_eyebrow",
            "setup_title",
            "setup_description",
            "setup_fact_1",
            "setup_fact_2",
            "setup_fact_3",
            "notes",
        ],
        [
            {
                "model_key": "stingray",
                "registry_key": "stingray",
                "model_label": "Stingray",
                "export_slug": "stingray",
                "active": True,
                **SETUP_COPY,
            },
            {
                "model_key": "grand_sport",
                "registry_key": "grandSport",
                "model_label": "Grand Sport",
                "export_slug": "grand-sport",
                "active": True,
                **SETUP_COPY,
            },
            {
                "model_key": "z06",
                "registry_key": "z06",
                "model_label": "Z06",
                "export_slug": "z06",
                "active": False,
                **SETUP_COPY,
            },
        ],
    )
    append_sheet(wb, "model_registry_promotion", MODEL_REGISTRY_PROMOTION_HEADERS, promotion_rows or [])
    return wb


def promoted_stingray_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "model_key": "stingray",
        "registry_key": "stingray",
        "promoted_to_runtime": True,
        "default_model": True,
        "artifact_path": "form-output/runtime/stingray-runtime-contract.json",
        "artifact_type": "runtime_contract",
        "legacy_alias": "STINGRAY_FORM_DATA",
        "active": True,
        "display_order": 1,
        "notes": "Stingray runtime contract.",
    }
    row.update(overrides)
    return row


def promoted_grand_sport_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "model_key": "grand_sport",
        "registry_key": "grandSport",
        "promoted_to_runtime": True,
        "default_model": False,
        "artifact_path": "form-output/runtime/grand-sport-runtime-contract.json",
        "artifact_type": "runtime_contract",
        "legacy_alias": "",
        "active": True,
        "display_order": 2,
        "notes": "Grand Sport runtime contract promoted to runtime registry.",
    }
    row.update(overrides)
    return row


class RegistryPromotionMetadataTests(unittest.TestCase):
    def test_header_only_promotion_sheet_refuses_to_build_a_registry(self) -> None:
        """Breaks if an empty promotion sheet is ever allowed to publish something.

        `build_registry_from_promotions` used to answer None here and let the
        caller decide. It is gone (Pass 3 requirement 8); the surviving builder
        refuses outright rather than guessing.
        """

        wb = workbook_with_promotions([])

        self.assertEqual(load_registry_promotions(wb), [])
        with self.assertRaisesRegex(RuntimeError, "no promoted rows"):
            build_registry_from_artifacts(wb, model_assets={}, root=ROOT)

    def test_promoted_rows_build_ordered_registry_and_aliases_from_workbook_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            artifact_path = root / "form-output" / "runtime" / "grand-sport-runtime-contract.json"
            artifact_path.parent.mkdir(parents=True)
            artifact_path.write_text(
                json.dumps(
                    runtime_contract_data(
                        "Grand Sport",
                        "grandSport_options",
                        choices=[
                            {
                                "choice_id": "gs-choice",
                            }
                        ],
                        rules=[
                            {
                                "rule_id": "rule-1",
                                "source_id": "opt_sht_001",
                                "source_type": "option",
                                "source_note": "runtime rule note",
                            }
                        ],
                    )
                ),
                encoding="utf-8",
            )
            stingray_path = root / "form-output" / "runtime" / "stingray-runtime-contract.json"
            stingray_path.write_text(
                json.dumps(runtime_contract_data("Stingray", "stingray_options")), encoding="utf-8"
            )
            wb = workbook_with_promotions(
                [
                    promoted_grand_sport_row(display_order=2),
                    {**promoted_stingray_row(display_order=1)},
                    {
                        "model_key": "z06",
                        "registry_key": "z06",
                        "promoted_to_runtime": False,
                        "default_model": False,
                        "active": False,
                        "display_order": 3,
                    },
                ]
            )

            registry = build_registry_from_artifacts(
                wb,
                model_assets={"stingray": {"image_url": "stingray.png"}, "grandSport": {"image_url": "gs.png"}},
                root=root,
            )

        assert registry is not None
        self.assertEqual(registry["defaultModelKey"], "stingray")
        self.assertEqual(list(registry["models"].keys()), ["stingray", "grandSport"])
        self.assertEqual(registry["models"]["stingray"]["label"], "Stingray")
        self.assertEqual(registry["models"]["stingray"]["exportSlug"], "stingray")
        self.assertEqual(registry["models"]["stingray"]["image_url"], "stingray.png")
        self.assertEqual(registry["models"]["grandSport"]["label"], "Grand Sport")
        self.assertEqual(registry["models"]["grandSport"]["exportSlug"], "grand-sport")
        self.assertEqual(
            registry["models"]["grandSport"]["vehicleSetup"],
            {
                "cardSubtitle": "Workbook-authored card subtitle",
                "eyebrow": "WORKBOOK-AUTHORED EYEBROW",
                "title": "Workbook-authored title",
                "description": "Workbook-authored description.",
                "facts": ["Fact one", "Fact two", "Fact three"],
            },
        )
        self.assertEqual(registry["models"]["grandSport"]["data"]["dataset"]["source_sheet"], "grandSport_options")
        choice = registry["models"]["grandSport"]["data"]["choices"][0]
        self.assertEqual(choice["choice_id"], "gs-choice")
        self.assertEqual(registry["legacyAliases"], {"STINGRAY_FORM_DATA": "stingray"})

    def test_live_contract_data_trims_choice_rows_without_stripping_section_or_group_modes(self) -> None:
        cleaned = live_contract_data(
            {
                "dataset": {"status": "draft_not_runtime_active", "name": "Grand Sport form data draft"},
                "sections": [
                    {
                        "section_id": "sec_wheels",
                        "choice_mode": "single",
                        "selection_mode": "single_select_req",
                        "selection_mode_label": "Required single selection",
                    }
                ],
                "choices": [
                    {
                        "choice_id": "choice-1",
                        "source_detail_raw": "raw note",
                        "choice_mode": "single",
                        "selection_mode": "single_select_req",
                        "selection_mode_label": "Required single selection",
                    }
                ],
                "standardEquipment": [{"equipment_id": "std-1", "source_detail_raw": "raw note"}],
                "exclusiveGroups": [{"group_id": "excl-1", "selection_mode": "required_single_within_group"}],
            }
        )

        self.assertEqual(cleaned["dataset"]["status"], "runtime_active")
        self.assertEqual(cleaned["sections"][0]["choice_mode"], "single")
        self.assertEqual(cleaned["sections"][0]["selection_mode"], "single_select_req")
        self.assertEqual(cleaned["sections"][0]["selection_mode_label"], "Required single selection")
        self.assertEqual(cleaned["exclusiveGroups"][0]["selection_mode"], "required_single_within_group")
        self.assertEqual(cleaned["choices"], [{"choice_id": "choice-1"}])
        self.assertEqual(cleaned["standardEquipment"], [{"equipment_id": "std-1"}])

    def test_every_promoted_row_is_loaded_from_its_named_runtime_contract(self) -> None:
        """Breaks if a promoted row can ever publish anything but the file it names.

        Was `..._loads_current_generation_and_runtime_contract_artifacts`: one of
        the two rows used to resolve to whatever sat at a generator output path
        instead of naming its artifact. Both rows now name one.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            stingray_path = root / "form-output" / "runtime" / "stingray-runtime-contract.json"
            gs_path = root / "form-output" / "runtime" / "grand-sport-runtime-contract.json"
            gs_path.parent.mkdir(parents=True)
            stingray_path.write_text(
                json.dumps(runtime_contract_data("Stingray", "stingray_options")),
                encoding="utf-8",
            )
            gs_path.write_text(
                json.dumps(runtime_contract_data("Grand Sport", "grandSport_options")),
                encoding="utf-8",
            )
            wb = workbook_with_promotions(
                [
                    promoted_grand_sport_row(display_order=2),
                    promoted_stingray_row(display_order=1),
                ]
            )

            registry = build_registry_from_artifacts(
                wb,
                model_assets={"stingray": {"image_url": "stingray.png"}, "grandSport": {"image_url": "gs.png"}},
                root=root,
            )

        self.assertEqual(registry["defaultModelKey"], "stingray")
        self.assertEqual(list(registry["models"].keys()), ["stingray", "grandSport"])
        self.assertEqual(registry["models"]["stingray"]["data"]["dataset"]["source_sheet"], "stingray_options")
        self.assertEqual(registry["models"]["grandSport"]["data"]["dataset"]["status"], "runtime_active")
        self.assertEqual(registry["models"]["stingray"]["vehicleSetup"]["facts"], ["Fact one", "Fact two", "Fact three"])
        self.assertEqual(registry["legacyAliases"], {"STINGRAY_FORM_DATA": "stingray"})

    def test_promoted_model_requires_complete_vehicle_setup_copy(self) -> None:
        wb = workbook_with_promotions([promoted_stingray_row()])
        headers = [cell.value for cell in wb["model_master"][1]]
        title_col = headers.index("setup_title") + 1
        wb["model_master"].cell(row=2, column=title_col).value = None

        with self.assertRaisesRegex(ValueError, "setup_title"):
            load_registry_promotions(wb)

    def test_promoted_artifact_with_draft_fields_fails_fast(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            artifact_path = root / "form-output" / "runtime" / "grand-sport-runtime-contract.json"
            artifact_path.parent.mkdir(parents=True)
            artifact_path.write_text(
                json.dumps(
                    {
                        "draftMetadata": {"inspection": True},
                        "dataset": {"status": "draft_not_runtime_active"},
                        "choices": [{"choice_id": "gs-choice", "source_option_name": "draft only"}],
                    }
                ),
                encoding="utf-8",
            )
            stingray_path = root / "form-output" / "runtime" / "stingray-runtime-contract.json"
            stingray_path.write_text(
                json.dumps(runtime_contract_data("Stingray", "stingray_options")), encoding="utf-8"
            )
            wb = workbook_with_promotions(
                [promoted_grand_sport_row(display_order=2), promoted_stingray_row(display_order=1)]
            )
            with self.assertRaisesRegex(ValueError, "not publishable"):
                build_registry_from_artifacts(wb, model_assets={}, root=root)

    def test_promotions_require_exactly_one_default_model(self) -> None:
        wb = workbook_with_promotions([promoted_stingray_row(default_model=False), promoted_grand_sport_row()])

        with self.assertRaisesRegex(ValueError, "exactly one promoted default model"):
            load_registry_promotions(wb)

    def test_duplicate_promoted_registry_keys_fail_fast(self) -> None:
        wb = workbook_with_promotions(
            [
                promoted_stingray_row(display_order=1),
                promoted_stingray_row(display_order=2, default_model=False),
            ]
        )

        with self.assertRaisesRegex(ValueError, "Duplicate promoted registry_key"):
            load_registry_promotions(wb)

    def test_every_promoted_row_requires_an_artifact_path(self) -> None:
        """Breaks if a promoted row can ever publish without naming its artifact.

        Scope, stated precisely: this does NOT catch the old
        `!= "current_generation"` exemption coming back, because the vocabulary
        check above raises first and makes that branch unreachable here. The
        schema layer accumulates rather than raising, so the exemption is live
        there — `test_a_retired_type_with_a_blank_path_reports_both_defects` in
        `tests/test_schema_validation_metadata.py` is what covers it.
        """

        wb = workbook_with_promotions([promoted_stingray_row(artifact_path=""), promoted_grand_sport_row()])

        with self.assertRaisesRegex(ValueError, "artifact_path"):
            load_registry_promotions(wb)

    def test_a_retired_artifact_type_is_rejected_by_name(self) -> None:
        """Breaks if `current_generation` or `draft_artifact` is ever accepted again.

        Both used to publish something other than a strictly validated runtime
        contract. The check is on the shared vocabulary, so narrowing it in one
        module and not another would also fail here.
        """

        for retired in ("current_generation", "draft_artifact"):
            with self.subTest(artifact_type=retired):
                wb = workbook_with_promotions([promoted_stingray_row(artifact_type=retired)])
                with self.assertRaisesRegex(ValueError, "Unsupported model_registry_promotion artifact_type"):
                    load_registry_promotions(wb)

    def test_runtime_contract_is_the_only_promotable_artifact_type(self) -> None:
        """Breaks if the vocabulary widens anywhere; this is the single authority."""

        self.assertEqual(REGISTRY_PROMOTION_ARTIFACT_TYPES, ("runtime_contract",))
        self.assertEqual(DEFAULT_REGISTRY_PROMOTION_ARTIFACT_TYPE, "runtime_contract")

    def test_a_blank_artifact_type_defaults_to_runtime_contract(self) -> None:
        """Breaks if a blank cell ever silently means a retired type again.

        It used to mean `draft_artifact` — a review artifact promoted to
        production by omission.
        """

        wb = workbook_with_promotions([promoted_stingray_row(artifact_type=None)])

        promotions = load_registry_promotions(wb)

        self.assertEqual([p.artifact_type for p in promotions], ["runtime_contract"])

    def test_promoted_registry_keys_must_match_model_master(self) -> None:
        wb = workbook_with_promotions([promoted_stingray_row(registry_key="wrong")])

        with self.assertRaisesRegex(ValueError, "registry_key"):
            load_registry_promotions(wb)


if __name__ == "__main__":
    unittest.main()
