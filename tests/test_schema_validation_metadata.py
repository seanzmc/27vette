#!/usr/bin/env python3
"""Tests for workbook metadata-derived schema validation."""

from __future__ import annotations

import sys
import tempfile
import unittest
import json
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_configs import REQUIRED_GENERATION_SOURCE_ROLES  # noqa: E402
from corvette_form_generator.schema_validation import live_contract_provenance_leaks, validate_workbook_schema  # noqa: E402


OPTION_HEADERS = ["option_id", "rpo", "selectable", "active", "price", "section_id", "display_order"]
OVS_HEADERS = ["option_id", "variant_id", "status"]
RULE_MAPPING_HEADERS = [
    "rule_id",
    "source_id",
    "target_id",
    "rule_type",
]
PRICE_RULE_HEADERS = ["rule_id", "target_id", "price_rule_type", "price_value"]
ACTIVE_GROUP_HEADERS = ["group_id", "active"]
ACTIVE_MEMBER_HEADERS = ["group_id", "option_id", "active"]
VARIANT_OVERRIDE_HEADERS = ["option_id", "variant_id", "active", "selectable"]
INTERIOR_HEADERS = ["interior_id", "Trim", "Price", "active_for_stingray", "requires_r6x", "Seat"]


def source_rows(model_key: str, role_to_sheet: dict[str, str], *, active: bool = True) -> list[dict[str, object]]:
    return [
        {"model_key": model_key, "source_role": role, "sheet_name": role_to_sheet[role], "active": active}
        for role in REQUIRED_GENERATION_SOURCE_ROLES
    ]


def z06_source_map(*, option_sheet: str = "z06_options", status_sheet: str = "z06_ovs") -> dict[str, str]:
    return {
        "source_option_sheet": option_sheet,
        "status_sheet": status_sheet,
        "rule_mapping_sheet": "z06_rule_mapping",
        "price_rules_sheet": "z06_price_rules",
        "rule_groups_sheet": "z06_rule_groups",
        "rule_group_members_sheet": "z06_rule_group_members",
        "exclusive_groups_sheet": "z06_exclusive_groups",
        "exclusive_group_members_sheet": "z06_exclusive_members",
        "color_overrides_sheet": "color_overrides",
        "interior_source_sheet": "LZ_Interiors",
    }


def stingray_source_map() -> dict[str, str]:
    return {
        "source_option_sheet": "stingray_options",
        "status_sheet": "stingray_ovs",
        "rule_mapping_sheet": "rule_mapping",
        "price_rules_sheet": "price_rules",
        "rule_groups_sheet": "rule_groups",
        "rule_group_members_sheet": "rule_group_members",
        "exclusive_groups_sheet": "exclusive_groups",
        "exclusive_group_members_sheet": "exclusive_group_members",
        "color_overrides_sheet": "color_overrides",
        "interior_source_sheet": "lt_interiors",
    }


def z06_source_sheets(
    *,
    option_headers: list[str] | None = None,
    option_rows: list[dict[str, object]] | None = None,
) -> dict[str, tuple[list[str], list[dict[str, object]]]]:
    return {
        "z06_options": (
            option_headers or OPTION_HEADERS,
            option_rows or [{"option_id": "opt_z06_known", "rpo": "Z06", "selectable": True, "active": True, "price": 0}],
        ),
        "z06_ovs": (OVS_HEADERS, [{"option_id": "opt_z06_known", "variant_id": "1lz_h07", "status": "available"}]),
        "z06_rule_mapping": (RULE_MAPPING_HEADERS, []),
        "z06_price_rules": (PRICE_RULE_HEADERS, []),
        "z06_rule_groups": (ACTIVE_GROUP_HEADERS, []),
        "z06_rule_group_members": (ACTIVE_MEMBER_HEADERS, []),
        "z06_exclusive_groups": (ACTIVE_GROUP_HEADERS, []),
        "z06_exclusive_members": (ACTIVE_MEMBER_HEADERS, []),
    }


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows or []:
        ws.append([row.get(header, None) for header in headers])


def minimal_schema_workbook(
    *,
    extra_model_rows: list[dict[str, object]] | None = None,
    extra_source_rows: list[dict[str, object]] | None = None,
    extra_sheets: dict[str, tuple[list[str], list[dict[str, object]]]] | None = None,
) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    append_sheet(wb, "variant_master", ["variant_id", "active"])
    append_sheet(wb, "section_master", ["section_id", "active"])
    append_sheet(
        wb,
        "model_master",
        [
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "notes",
        ],
        extra_model_rows or [],
    )
    append_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        extra_source_rows or [],
    )
    append_sheet(wb, "model_variants", ["model_key", "variant_id", "display_order", "active", "notes"])

    for name in ("stingray_options", "grandSport_options"):
        append_sheet(wb, name, OPTION_HEADERS, [{"option_id": "opt_known", "rpo": "ABC", "selectable": True, "active": True, "price": 0}])
    for name in ("stingray_ovs", "grandSport_ovs"):
        append_sheet(wb, name, OVS_HEADERS, [{"option_id": "opt_known", "variant_id": "v1", "status": "available"}])
    for name in ("rule_mapping", "grandSport_rule_mapping"):
        append_sheet(wb, name, RULE_MAPPING_HEADERS)
    for name in ("price_rules", "grandSport_price_rules"):
        append_sheet(wb, name, PRICE_RULE_HEADERS)
    for name in ("rule_groups", "grandSport_rule_groups", "exclusive_groups", "grandSport_exclusive_groups"):
        append_sheet(wb, name, ACTIVE_GROUP_HEADERS)
    for name in ("rule_group_members", "grandSport_rule_group_members", "exclusive_group_members", "grandSport_exclusive_members"):
        append_sheet(wb, name, ACTIVE_MEMBER_HEADERS)
    append_sheet(wb, "color_overrides", ["model_key", "interior_id", "option_id", "active"])
    append_sheet(wb, "lt_interiors", INTERIOR_HEADERS)
    append_sheet(wb, "LZ_Interiors", INTERIOR_HEADERS)
    append_sheet(wb, "model_interior_scope", ["model_key", "interior_id", "active"])
    append_sheet(wb, "interior_components", ["model_key", "interior_id", "rpo", "active"])
    append_sheet(wb, "PriceRef", ["RPO", "Price"])

    for name, (headers, rows) in (extra_sheets or {}).items():
        if name in wb.sheetnames:
            del wb[name]
        append_sheet(wb, name, headers, rows)
    return wb


def validate_temp_workbook(wb: Workbook):
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "metadata-schema.xlsx"
        wb.save(path)
        return validate_workbook_schema(path, check_live_contract=False)


class SchemaValidationMetadataTests(unittest.TestCase):
    def test_model_master_asset_map_shaped_headers_are_rejected_directly(self) -> None:
        wb = minimal_schema_workbook(
            extra_sheets={
                "model_master": (
                    [
                        "model_key",
                        "target_type",
                        "target_id",
                        "image_url",
                        "image_alt",
                        "image_fit",
                        "image_position",
                        "active",
                        "notes",
                        "notes",
                    ],
                    [
                        {
                            "model_key": "stingray",
                            "target_type": "model",
                            "target_id": "stingray",
                            "active": True,
                        }
                    ],
                )
            }
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "model_master_header_drift" for issue in issues), issues)

    def test_model_master_rejects_duplicate_active_model_keys(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "stingray", "registry_key": "stingray", "active": True},
                {"model_key": "stingray", "registry_key": "stingray", "active": True},
            ]
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "duplicate_active_model_master_row" for issue in issues), issues)

    def test_asset_map_rejects_duplicate_active_rows_for_same_identity(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "stingray", "registry_key": "stingray", "active": True}],
            extra_sheets={
                "asset_map": (
                    ["model_key", "target_type", "target_id", "image_url", "active"],
                    [
                        {"model_key": "stingray", "target_type": "option", "target_id": "opt_gba_001", "image_url": "https://example.test/a.png", "active": True},
                        {"model_key": "stingray", "target_type": "option", "target_id": "opt_gba_001", "image_url": "https://example.test/b.png", "active": True},
                    ],
                )
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "duplicate_active_asset_map_row" for issue in issues), issues)

    def test_asset_map_ignores_duplicate_blank_or_inactive_rows(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "stingray", "registry_key": "stingray", "active": True}],
            extra_sheets={
                "asset_map": (
                    ["model_key", "target_type", "target_id", "image_url", "active"],
                    [
                        {"model_key": "stingray", "target_type": "option", "target_id": "blank_duplicate", "image_url": "https://example.test/a.png", "active": ""},
                        {"model_key": "stingray", "target_type": "option", "target_id": "blank_duplicate", "image_url": "https://example.test/b.png", "active": ""},
                        {"model_key": "stingray", "target_type": "option", "target_id": "inactive_duplicate", "image_url": "https://example.test/c.png", "active": False},
                        {"model_key": "stingray", "target_type": "option", "target_id": "inactive_duplicate", "image_url": "https://example.test/d.png", "active": False},
                    ],
                )
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertFalse(any(issue.check_id == "duplicate_active_asset_map_row" for issue in issues), issues)

    def test_asset_map_allows_same_target_id_under_different_target_types(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "stingray", "registry_key": "stingray", "active": True}],
            extra_sheets={
                "asset_map": (
                    ["model_key", "target_type", "target_id", "image_url", "active"],
                    [
                        {"model_key": "stingray", "target_type": "option", "target_id": "shared_id", "image_url": "https://example.test/a.png", "active": True},
                        {"model_key": "stingray", "target_type": "model", "target_id": "shared_id", "image_url": "https://example.test/b.png", "active": True},
                    ],
                )
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertFalse(any(issue.check_id == "duplicate_active_asset_map_row" for issue in issues), issues)

    def test_default_selection_display_behavior_rejects_invalid_values(self) -> None:
        wb = minimal_schema_workbook(
            extra_sheets={
                "default_selection_rules": (
                    [
                        "model_key",
                        "rule_id",
                        "target_option_id",
                        "condition_type",
                        "condition_id",
                        "body_style_scope",
                        "trim_level_scope",
                        "variant_scope",
                        "priority",
                        "active",
                        "display_behavior",
                        "notes",
                    ],
                    [
                        {
                            "model_key": "stingray",
                            "rule_id": "default_bc7",
                            "target_option_id": "opt_bc7_001",
                            "condition_type": "always",
                            "active": True,
                            "display_behavior": "default-selected",
                        },
                        {
                            "model_key": "grand_sport",
                            "rule_id": "gs_default_bc7_coupe",
                            "target_option_id": "opt_bc7_001",
                            "condition_type": "always",
                            "active": True,
                            "display_behavior": "default_selected",
                        },
                        {
                            "model_key": "z06",
                            "rule_id": "z06_default_r8e_tax",
                            "target_option_id": "opt_r8e_002",
                            "condition_type": "always",
                            "active": True,
                            "display_behavior": "",
                        },
                    ],
                )
            }
        )

        issues = validate_temp_workbook(wb)

        invalid_issues = [issue for issue in issues if issue.check_id == "invalid_default_selection_display_behavior"]
        self.assertEqual(len(invalid_issues), 1, invalid_issues)
        self.assertEqual(invalid_issues[0].sheet, "default_selection_rules")
        self.assertEqual(invalid_issues[0].column, "display_behavior")
        self.assertEqual(invalid_issues[0].value, "default-selected")

    def test_model_workbook_sources_sheet_is_required(self) -> None:
        wb = minimal_schema_workbook(extra_model_rows=[{"model_key": "stingray", "registry_key": "stingray", "active": True}])
        del wb["model_workbook_sources"]

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "missing_required_sheet" and issue.sheet == "model_workbook_sources" for issue in issues),
            issues,
        )

    def test_active_model_missing_required_source_option_role_is_error(self) -> None:
        role_map = z06_source_map()
        rows = [row for row in source_rows("z06", role_map) if row["source_role"] != "source_option_sheet"]
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": True}],
            extra_source_rows=rows,
            extra_sheets=z06_source_sheets(),
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(
                issue.check_id == "missing_model_source_role"
                and issue.sheet == "model_workbook_sources"
                and issue.value == {"model_key": "z06", "source_role": "source_option_sheet"}
                for issue in issues
            ),
            issues,
        )

    def test_active_model_missing_required_status_role_is_error(self) -> None:
        role_map = z06_source_map()
        rows = [row for row in source_rows("z06", role_map) if row["source_role"] != "status_sheet"]
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": True}],
            extra_source_rows=rows,
            extra_sheets=z06_source_sheets(),
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(
                issue.check_id == "missing_model_source_role"
                and issue.sheet == "model_workbook_sources"
                and issue.value == {"model_key": "z06", "source_role": "status_sheet"}
                for issue in issues
            ),
            issues,
        )

    def test_shared_source_roles_do_not_satisfy_active_model_exact_match_requirements(self) -> None:
        role_map = z06_source_map()
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": True}],
            extra_source_rows=source_rows("shared", role_map),
            extra_sheets=z06_source_sheets(),
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(
                issue.check_id == "missing_model_source_role"
                and issue.value == {"model_key": "z06", "source_role": "source_option_sheet"}
                for issue in issues
            ),
            issues,
        )

    def test_active_model_complete_source_roles_validate_without_legacy_seeds(self) -> None:
        role_map = z06_source_map()
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": True}],
            extra_source_rows=source_rows("z06", role_map),
            extra_sheets=z06_source_sheets(),
        )

        issues = validate_temp_workbook(wb)

        self.assertFalse(any(issue.check_id == "missing_model_source_role" for issue in issues), issues)
        self.assertFalse(any(issue.check_id == "missing_model_source_sheet" for issue in issues), issues)

    def test_active_model_variant_must_reference_active_variant_master_row(self) -> None:
        role_map = z06_source_map()
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "z06", "registry_key": "z06", "expected_variant_count": 1, "active": True}
            ],
            extra_source_rows=source_rows("z06", role_map),
            extra_sheets={
                **z06_source_sheets(),
                "variant_master": (["variant_id", "active"], [{"variant_id": "1lz_h07", "active": False}]),
                "model_variants": (
                    ["model_key", "variant_id", "display_order", "active", "notes"],
                    [{"model_key": "z06", "variant_id": "1lz_h07", "display_order": 1, "active": True}],
                ),
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "model_variant_inactive_variant_master" for issue in issues),
            issues,
        )

    def test_active_model_variant_must_reference_existing_variant_master_row(self) -> None:
        role_map = z06_source_map()
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "z06", "registry_key": "z06", "expected_variant_count": 1, "active": True}
            ],
            extra_source_rows=source_rows("z06", role_map),
            extra_sheets={
                **z06_source_sheets(),
                "variant_master": (["variant_id", "active"], []),
                "model_variants": (
                    ["model_key", "variant_id", "display_order", "active", "notes"],
                    [{"model_key": "z06", "variant_id": "1lz_h07", "display_order": 1, "active": True}],
                ),
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "model_variant_unknown_variant_master" for issue in issues), issues)

    def test_active_model_variants_require_unique_display_order(self) -> None:
        role_map = z06_source_map()
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "z06", "registry_key": "z06", "expected_variant_count": 2, "active": True}
            ],
            extra_source_rows=source_rows("z06", role_map),
            extra_sheets={
                **z06_source_sheets(),
                "variant_master": (
                    ["variant_id", "active"],
                    [{"variant_id": "1lz_h07", "active": True}, {"variant_id": "2lz_h07", "active": True}],
                ),
                "model_variants": (
                    ["model_key", "variant_id", "display_order", "active", "notes"],
                    [
                        {"model_key": "z06", "variant_id": "1lz_h07", "display_order": 1, "active": True},
                        {"model_key": "z06", "variant_id": "2lz_h07", "display_order": 1, "active": True},
                    ],
                ),
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "duplicate_model_variant_display_order" for issue in issues), issues)

    def test_active_model_variant_count_matches_model_master_expected_count(self) -> None:
        role_map = z06_source_map()
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "z06", "registry_key": "z06", "expected_variant_count": 2, "active": True}
            ],
            extra_source_rows=source_rows("z06", role_map),
            extra_sheets={
                **z06_source_sheets(),
                "variant_master": (["variant_id", "active"], [{"variant_id": "1lz_h07", "active": True}]),
                "model_variants": (
                    ["model_key", "variant_id", "display_order", "active", "notes"],
                    [{"model_key": "z06", "variant_id": "1lz_h07", "display_order": 1, "active": True}],
                ),
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "model_variant_count_mismatch" for issue in issues), issues)

    def test_live_contract_provenance_leaks_flags_future_model_review_lineage_only(self) -> None:
        data = {
            "dataset": {"source_sheet": "grandSport_options"},
            "label": "Grand Sport",
            "choices": [
                {
                    "choice_id": "choice-1",
                    "source_detail_raw": "customer-facing source detail",
                    "choice_mode": "single",
                    "selection_mode": "single_select_req",
                    "selection_mode_label": "Required single selection",
                    "suggested_copy_from": "grand_sport:opt_abc_001",
                }
            ],
            "standardEquipment": [{"equipment_id": "std-1", "source_detail_raw": "standard source detail"}],
            "sections": [
                {
                    "section_id": "sec_wheels",
                    "choice_mode": "single",
                    "selection_mode": "single_select_req",
                    "selection_mode_label": "Required single selection",
                }
            ],
            "exclusiveGroups": [{"group_id": "excl-1", "selection_mode": "required_single_within_group"}],
            "rules": [
                {
                    "source_id": "opt_abc_001",
                    "source_type": "option",
                    "copy_from_model_key": "grand_sport",
                }
            ],
        }

        leaks = list(live_contract_provenance_leaks(data))

        self.assertEqual(
            {path for path, _, _ in leaks},
            {
                "$.choices[0].source_detail_raw",
                "$.choices[0].choice_mode",
                "$.choices[0].selection_mode",
                "$.choices[0].selection_mode_label",
                "$.choices[0].suggested_copy_from",
                "$.standardEquipment[0].source_detail_raw",
                "$.rules[0].copy_from_model_key",
            },
        )
        self.assertFalse(any(path == "$.dataset.source_sheet" for path, _, _ in leaks))
        self.assertFalse(any(path == "$.label" for path, _, _ in leaks))
        self.assertFalse(any(path.startswith("$.sections[0].") for path, _, _ in leaks))
        self.assertFalse(any(path == "$.exclusiveGroups[0].selection_mode" for path, _, _ in leaks))

    def test_metadata_discovered_ovs_sheet_validates_option_ids(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": True}],
            extra_source_rows=[
                {"model_key": "z06", "source_role": "source_option_sheet", "sheet_name": "z06_options", "active": True},
                {"model_key": "z06", "source_role": "status_sheet", "sheet_name": "z06_ovs", "active": True},
            ],
            extra_sheets={
                "z06_options": (OPTION_HEADERS, [{"option_id": "opt_z06_known", "rpo": "Z06", "selectable": True, "active": True, "price": 0}]),
                "z06_ovs": (OVS_HEADERS, [{"option_id": "opt_missing", "variant_id": "1lz_h07", "status": "available"}]),
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "ovs_unknown_option_id" and issue.sheet == "z06_ovs" for issue in issues),
            [issue for issue in issues if issue.check_id == "ovs_unknown_option_id"],
        )

    def test_metadata_discovered_source_sheet_must_exist(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "zr1", "registry_key": "zr1", "active": True}],
            extra_source_rows=[
                {"model_key": "zr1", "source_role": "source_option_sheet", "sheet_name": "zr1_options", "active": True},
            ],
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "missing_model_source_sheet" and issue.sheet == "zr1_options" for issue in issues),
            issues,
        )

    def test_metadata_discovered_option_headers_match_by_role(self) -> None:
        role_map = z06_source_map()
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "stingray", "registry_key": "stingray", "active": True},
                {"model_key": "z06", "registry_key": "z06", "active": True},
            ],
            extra_source_rows=[*source_rows("stingray", stingray_source_map()), *source_rows("z06", role_map)],
            extra_sheets=z06_source_sheets(
                option_headers=["option_id", "rpo", "selectable", "active", "price", "unexpected_extra"],
                option_rows=[{"option_id": "opt_z06_known", "rpo": "Z06", "selectable": True, "active": True, "price": 0}],
            ),
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "source_role_header_drift" and issue.sheet == "source_option_sheet" for issue in issues),
            issues,
        )

    def test_option_display_order_duplicates_are_rejected_in_standard_sections(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": True}],
            extra_source_rows=[
                {"model_key": "z06", "source_role": "source_option_sheet", "sheet_name": "z06_options", "active": True},
            ],
            extra_sheets={
                "z06_options": (
                    OPTION_HEADERS,
                    [
                        {
                            "option_id": "opt_u80_001",
                            "rpo": "U80",
                            "selectable": False,
                            "active": True,
                            "price": 0,
                            "section_id": "sec_stan_001",
                            "display_order": 20,
                        },
                        {
                            "option_id": "opt_wub_001",
                            "rpo": "WUB",
                            "selectable": False,
                            "active": True,
                            "price": 0,
                            "section_id": "sec_stan_001",
                            "display_order": 20,
                        },
                    ],
                ),
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "duplicate_option_display_order" and issue.sheet == "z06_options" for issue in issues),
            issues,
        )

    def test_future_scaffold_option_display_order_duplicates_are_rejected(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "zr1", "registry_key": "zr1", "active": False}],
            extra_source_rows=[
                {
                    "model_key": "zr1",
                    "source_role": "source_option_sheet",
                    "sheet_name": "zr1_options",
                    "active": False,
                },
            ],
            extra_sheets={
                "zr1_options": (
                    OPTION_HEADERS,
                    [
                        {
                            "option_id": "opt_u80_001",
                            "rpo": "U80",
                            "selectable": False,
                            "active": True,
                            "price": 0,
                            "section_id": "sec_stan_001",
                            "display_order": "20",
                        },
                        {
                            "option_id": "opt_wub_001",
                            "rpo": "WUB",
                            "selectable": False,
                            "active": True,
                            "price": 0,
                            "section_id": "sec_stan_001",
                            "display_order": 20,
                        },
                    ],
                ),
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(
                issue.check_id == "duplicate_future_scaffold_option_display_order"
                and issue.sheet == "zr1_options"
                for issue in issues
            ),
            issues,
        )

    def test_metadata_discovered_interior_source_sheet_role_is_known(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": True}],
            extra_source_rows=[
                {"model_key": "z06", "source_role": "interior_source_sheet", "sheet_name": "LZ_Interiors", "active": True},
            ],
        )

        issues = validate_temp_workbook(wb)

        self.assertFalse(any(issue.check_id == "unknown_model_source_role" for issue in issues), issues)

    def test_metadata_discovered_interior_source_sheet_must_exist(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "zr1", "registry_key": "zr1", "active": True}],
            extra_source_rows=[
                {
                    "model_key": "zr1",
                    "source_role": "interior_source_sheet",
                    "sheet_name": "missing_zr1_interiors",
                    "active": True,
                },
            ],
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "missing_model_source_sheet" and issue.sheet == "missing_zr1_interiors" for issue in issues),
            issues,
        )

    def test_metadata_discovered_lz_interiors_validates_by_interior_role(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "zr1x", "registry_key": "zr1x", "active": True}],
            extra_source_rows=[
                {"model_key": "zr1x", "source_role": "interior_source_sheet", "sheet_name": "LZ_Interiors", "active": True},
            ],
            extra_sheets={
                "LZ_Interiors": (
                    INTERIOR_HEADERS,
                    [
                        {
                            "interior_id": "1LZ_AQ9_HTE",
                            "Trim": "1LZ",
                            "Price": "not numeric",
                            "active_for_stingray": True,
                            "requires_r6x": False,
                            "Seat": "AQ9",
                        }
                    ],
                )
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(
            any(issue.check_id == "price_type_drift" and issue.sheet == "LZ_Interiors" and issue.column == "Price" for issue in issues),
            issues,
        )

    def test_inactive_metadata_source_rows_are_ignored(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "future", "registry_key": "future", "active": True}],
            extra_source_rows=[
                {"model_key": "future", "source_role": "source_option_sheet", "sheet_name": "missing_future_options", "active": False},
            ],
        )

        issues = validate_temp_workbook(wb)

        self.assertFalse(any(issue.sheet == "missing_future_options" for issue in issues), issues)

    def test_inactive_future_model_rows_do_not_require_missing_source_sheets(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "stingray", "registry_key": "stingray", "active": True},
                {"model_key": "z06", "registry_key": "z06", "active": False},
            ],
            extra_source_rows=[
                {"model_key": "z06", "source_role": "source_option_sheet", "sheet_name": "z06_options", "active": True},
                {"model_key": "z06", "source_role": "status_sheet", "sheet_name": "z06_ovs", "active": True},
                {"model_key": "z06", "source_role": "interior_source_sheet", "sheet_name": "LZ_Interiors", "active": True},
            ],
        )

        issues = validate_temp_workbook(wb)

        self.assertFalse(any(issue.sheet in {"z06_options", "z06_ovs"} for issue in issues), issues)
        self.assertFalse(any(issue.check_id == "source_role_header_drift" for issue in issues), issues)

    def test_model_registry_promotion_requires_exactly_one_promoted_default(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[
                {"model_key": "stingray", "registry_key": "stingray", "active": True},
                {"model_key": "grand_sport", "registry_key": "grandSport", "active": True},
            ],
            extra_sheets={
                "model_registry_promotion": (
                    [
                        "model_key",
                        "registry_key",
                        "promoted_to_runtime",
                        "default_model",
                        "artifact_path",
                        "artifact_type",
                        "legacy_alias",
                        "active",
                        "display_order",
                        "notes",
                    ],
                    [
                        {
                            "model_key": "stingray",
                            "registry_key": "stingray",
                            "promoted_to_runtime": True,
                            "default_model": False,
                            "artifact_type": "current_generation",
                            "active": True,
                            "display_order": 1,
                        },
                        {
                            "model_key": "grand_sport",
                            "registry_key": "grandSport",
                            "promoted_to_runtime": True,
                            "default_model": False,
                            "artifact_path": "form-output/runtime/grand-sport-runtime-contract.json",
                            "artifact_type": "runtime_contract",
                            "active": True,
                            "display_order": 2,
                        },
                    ],
                )
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "registry_promotion_default_count" for issue in issues), issues)

    def test_model_registry_promotion_rejects_future_promoted_registry_key_drift(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "z06", "registry_key": "z06", "active": False}],
            extra_sheets={
                "model_registry_promotion": (
                    [
                        "model_key",
                        "registry_key",
                        "promoted_to_runtime",
                        "default_model",
                        "artifact_path",
                        "artifact_type",
                        "legacy_alias",
                        "active",
                        "display_order",
                        "notes",
                    ],
                    [
                        {
                            "model_key": "z06",
                            "registry_key": "wrong",
                            "promoted_to_runtime": True,
                            "default_model": True,
                            "artifact_path": "form-output/runtime/z06-runtime-contract.json",
                            "artifact_type": "runtime_contract",
                            "active": True,
                            "display_order": 3,
                        }
                    ],
                )
            },
        )

        issues = validate_temp_workbook(wb)

        self.assertTrue(any(issue.check_id == "registry_promotion_registry_key_mismatch" for issue in issues), issues)

    def test_live_app_registry_must_match_promoted_artifacts(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "stingray", "registry_key": "stingray", "active": True}],
            extra_sheets={
                "model_registry_promotion": (
                    [
                        "model_key",
                        "registry_key",
                        "promoted_to_runtime",
                        "default_model",
                        "artifact_path",
                        "artifact_type",
                        "legacy_alias",
                        "active",
                        "display_order",
                        "notes",
                    ],
                    [
                        {
                            "model_key": "stingray",
                            "registry_key": "stingray",
                            "promoted_to_runtime": True,
                            "default_model": True,
                            "artifact_type": "current_generation",
                            "legacy_alias": "STINGRAY_FORM_DATA",
                            "active": True,
                            "display_order": 1,
                        }
                    ],
                )
            },
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "stingray_master.xlsx"
            app_dir = root / "form-app"
            output_dir = root / "form-output"
            app_dir.mkdir()
            output_dir.mkdir()
            fresh_data = {"dataset": {"source_sheet": "stingray_options"}, "choices": [{"choice_id": "fresh"}]}
            stale_registry = {
                "defaultModelKey": "stingray",
                "models": {
                    "stingray": {
                        "key": "stingray",
                        "label": "Stingray",
                        "modelName": "Corvette Stingray",
                        "exportSlug": "stingray",
                        "data": {"dataset": {"source_sheet": "stingray_options"}, "choices": [{"choice_id": "stale"}]},
                    }
                },
            }
            wb.save(workbook_path)
            (output_dir / "stingray-form-data.json").write_text(json.dumps(fresh_data), encoding="utf-8")
            (app_dir / "data.js").write_text(
                f"window.CORVETTE_FORM_DATA = {json.dumps(stale_registry)};\n"
                "window.STINGRAY_FORM_DATA = window.CORVETTE_FORM_DATA.models.stingray.data;\n",
                encoding="utf-8",
            )

            issues = validate_workbook_schema(workbook_path, check_live_contract=True)

        self.assertTrue(any(issue.check_id == "app_registry_stale" for issue in issues), issues)

    def test_live_app_registry_freshness_ignores_generated_timestamps(self) -> None:
        wb = minimal_schema_workbook(
            extra_model_rows=[{"model_key": "stingray", "registry_key": "stingray", "active": True}],
            extra_sheets={
                "model_registry_promotion": (
                    [
                        "model_key",
                        "registry_key",
                        "promoted_to_runtime",
                        "default_model",
                        "artifact_path",
                        "artifact_type",
                        "legacy_alias",
                        "active",
                        "display_order",
                        "notes",
                    ],
                    [
                        {
                            "model_key": "stingray",
                            "registry_key": "stingray",
                            "promoted_to_runtime": True,
                            "default_model": True,
                            "artifact_type": "current_generation",
                            "legacy_alias": "STINGRAY_FORM_DATA",
                            "active": True,
                            "display_order": 1,
                        }
                    ],
                )
            },
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "stingray_master.xlsx"
            app_dir = root / "form-app"
            output_dir = root / "form-output"
            app_dir.mkdir()
            output_dir.mkdir()
            fresh_data = {
                "dataset": {"source_sheet": "stingray_options", "generated_at": "2026-06-15T20:31:02+00:00"},
                "choices": [{"choice_id": "same"}],
            }
            timestamp_only_registry = {
                "defaultModelKey": "stingray",
                "models": {
                    "stingray": {
                        "key": "stingray",
                        "label": "Stingray",
                        "modelName": "Corvette Stingray",
                        "exportSlug": "stingray",
                        "data": {
                            "dataset": {"source_sheet": "stingray_options", "generated_at": "2026-06-15T19:00:28+00:00"},
                            "choices": [{"choice_id": "same"}],
                        },
                    }
                },
            }
            wb.save(workbook_path)
            (output_dir / "stingray-form-data.json").write_text(json.dumps(fresh_data), encoding="utf-8")
            (app_dir / "data.js").write_text(
                f"window.CORVETTE_FORM_DATA = {json.dumps(timestamp_only_registry)};\n"
                "window.STINGRAY_FORM_DATA = window.CORVETTE_FORM_DATA.models.stingray.data;\n",
                encoding="utf-8",
            )

            issues = validate_workbook_schema(workbook_path, check_live_contract=True)

        self.assertFalse(any(issue.check_id == "app_registry_stale" for issue in issues), issues)


if __name__ == "__main__":
    unittest.main()
