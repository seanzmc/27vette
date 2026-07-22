#!/usr/bin/env python3
"""Compact fixture workbooks for wizard tests: a raw GM order-guide export
(Pass A) and a canonical-workbook stand-in (Pass B read-only pickers)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

STINGRAY_VARIANTS = ["Coupe\n1YC07\n1LT", "Coupe\n1YC07\n2LT", "Convertible\n1YC67\n1LT"]
ZR1_VARIANTS = ["ZR1 Coupe\n1YR07\n1LZ", "ZR1X Coupe\n1YS07\n1LZ"]
BASE_HEADERS = ["Orderable RPO Code", "Ref. Only RPO Code", "Description"]


def matrix_sheet(wb: Workbook, name: str, family_title: str, variants: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append([family_title])
    ws.append(["", "", "S = Standard Equipment  A = Available"])
    ws.append(BASE_HEADERS + variants)
    for row in rows:
        ws.append(row)


def build_raw_export(path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Price Schedule")
    ws.append(["2027 CHEVROLET CORVETTE"])
    ws.append([])
    ws.append(["Base Model Prices"])
    ws.append(["", "Model", "Model Description", "List", "Factory", "MSRP(c)", "", "", "", "DFC"])
    ws.append(["", "1YC07", "Corvette Stingray Coupe 1LT", 71000, 0, 71000, "", "", "", 2495])
    ws.append(["", "1YR07", "Corvette ZR1 Coupe 1LZ", 194700, 0, 194700, "", "", "", 2495])
    ws.append([])
    ws.append(["Additional Options"])
    ws.append(["", "Option Code", "Description", "List", "Factory", "MSRP(c)"])
    ws.append(["", "Additional Options:"])
    ws.append(["", "BV4", "Personalized Plaque", 395, 0, 395])
    ws.append(["", "PDB", "Carbon Wheel Package", "with ROY wheels", 16000, 0, 16000])
    ws.append(["", "PDB", "Carbon Wheel Package", "with ROZ wheels", 17000, 0, 17000])
    ws.append(["", "YYY", "Orphan priced option", "", 500, 0, 500])

    matrix_sheet(
        wb,
        "Exterior 1",
        "Stingray",
        STINGRAY_VARIANTS,
        [
            ["Equipment Groups"],
            ["", "UQH", "Audio system feature, Bose premium", "--", "■", "--"],
            ["BV4", "", "Personalized Plaque. Not available with (PDB).", "A1", "A", "A"],
            ["E60", "", "Front Lift", "A/D1", "A", "--"],
            ["ZZZ", "", "Mystery option with odd status", "?", "A", "A"],
            ["", "", "Narrative-only detail row without any RPO", "", "", ""],
        ],
    )
    matrix_sheet(
        wb,
        "Mechanical 4",
        "ZR1 and ZR1X",
        ZR1_VARIANTS,
        [
            # Before any section-label row: sectionLabel stays empty, so the
            # review source-group filter must fall back to the sheet name.
            ["CC3", "", "Roof panel, transparent", "A", "A"],
            ["Equipment Groups"],
            ["PDB", "", "Carbon Wheel Package", "A", "A"],
            ["C2Z", "", "ZR1 only cosmetic pack", "A", "--"],
            # Same comma-rule name as CC3 ("Roof panel") — must be flagged as
            # a duplicate proposal in the copy-split queue.
            ["CC2", "", "Roof panel, painted body color", "A", "A"],
            # Ref-only with an available status on ZR1 but not ZR1X: excluded
            # from ZR1's standard-equipment queue, included in ZR1X's.
            ["", "XFR", "Tires, high performance, ref only", "A", "--"],
            # Ref-only standard on both: legitimate standard-equipment row.
            ["", "AJ7", "Airbags, ref only", "S", "S"],
        ],
    )
    matrix_sheet(
        wb,
        "Standard Equipment 1",
        "Stingray",
        STINGRAY_VARIANTS,
        [
            ["", "AJ7", "Airbags, frontal and side-impact", "S1", "S1", "S1"],
            ["", "CJ2", "Air conditioning, dual-zone", "S", "S", "S"],
            ["", "UQH", "Audio, standard", "S", "S", "S"],
            ["EYT", "", "Rare orderable row on SE sheet", "A", "S", "S"],
        ],
    )

    ws = wb.create_sheet("Color and Trim 1")
    ws.append(["Recommended"])
    ws.append(["Some", "unrelated", "layout"])

    wb.save(path)
    return path


def _table(wb: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def build_master_workbook(path: Path) -> Path:
    """Canonical-workbook stand-in for Pass B: model metadata, section picker,
    and z06 presentation rows used as the template model. Read-only in tests."""

    wb = Workbook()
    wb.remove(wb.active)
    _table(
        wb,
        "model_master",
        ["model_key", "registry_key", "model_label", "model_year", "dataset_name", "export_slug", "expected_variant_count", "default_model", "active", "notes"],
        [
            ["z06", "z06", "Z06", 2027, "z06 dataset", "z06", 1, False, True, ""],
            ["zr1", "zr1", "ZR1", 2027, "zr1 scaffold", "zr1", 2, False, False, ""],
        ],
    )
    _table(
        wb,
        "model_registry_promotion",
        ["model_key", "registry_key", "promoted_to_runtime", "default_model", "artifact_path", "artifact_type", "legacy_alias", "active", "display_order", "notes"],
        [
            ["z06", "z06", True, False, "form-output/runtime/z06-runtime-contract.json", "runtime_contract", "", True, 1, ""],
            ["zr1", "zr1", False, False, "", "", "", False, 2, ""],
        ],
    )
    _table(
        wb,
        "variant_master",
        ["variant_id", "model_year", "trim_level", "body_style", "display_name", "base_price", "display_order", "active"],
        [
            ["1lz_r07", 2027, "1lz", "coupe", "Corvette ZR1 Coupe 1LZ", 197195, 25, False],
            ["3lz_r67", 2027, "3lz", "convertible", "Corvette ZR1 Convertible 3LZ", 218195, 28, False],
            ["1lz_s07", 2027, "1lz", "coupe", "Corvette ZR1X Coupe 1LZ", 227395, 29, False],
            ["1lz_h07", 2027, "1lz", "coupe", "Corvette Z06 Coupe 1LZ", 112100, 20, True],
        ],
    )
    _table(
        wb,
        "model_variants",
        ["model_key", "variant_id", "display_order", "active", "notes"],
        [
            ["z06", "1lz_h07", 1, True, ""],
            ["zr1", "1lz_r07", 1, False, ""],
            ["zr1", "3lz_r67", 2, False, ""],
            ["zr1x", "1lz_s07", 1, False, ""],
        ],
    )
    _table(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        [
            ["z06", "source_option_sheet", "z06_options", True, ""],
            ["z06", "status_sheet", "z06_ovs", True, ""],
            ["z06", "rule_mapping_sheet", "z06_rule_mapping", True, ""],
            ["z06", "price_rules_sheet", "z06_price_rules", True, ""],
            ["z06", "rule_groups_sheet", "z06_rule_groups", True, ""],
            ["z06", "rule_group_members_sheet", "z06_rule_group_members", True, ""],
            ["z06", "exclusive_groups_sheet", "z06_exclusive_groups", True, ""],
            ["z06", "exclusive_group_members_sheet", "z06_exclusive_members", True, ""],
            ["z06", "color_overrides_sheet", "color_overrides", True, ""],
            ["z06", "interior_source_sheet", "LZ_Interiors", True, ""],
            ["z06", "variant_option_overrides_sheet", "z06_variant_overrides", True, ""],
            ["zr1", "source_option_sheet", "zr1_options", False, "inactive scaffold, must be ignored"],
        ],
    )
    option_headers = ["option_id", "rpo", "price", "option_name", "description", "detail_raw", "section_id", "selectable", "display_order", "active", "display_behavior"]
    _table(
        wb,
        "z06_options",
        option_headers,
        [
            ["opt_gba_001", "GBA", 0, "Black", "Touch-Up Paint Number WA-8555", "", "sec_pain_001", True, 5, True, ""],
            ["opt_pdb_001", "PDB", 16000, "Carbon Fiber Wheel Package", "Visible carbon", "", "sec_whee_001", "", 10, True, ""],
            ["opt_xfr_001", "XFR", "", "High Performance Tires", "Reference-only tires", "", "sec_whee_001", False, 20, False, ""],
        ],
    )
    _table(
        wb,
        "z06_ovs",
        ["option_id", "variant_id", "status"],
        [
            ["opt_gba_001", "1lz_h07", "available"],
            ["opt_pdb_001", "1lz_h07", "available"],
        ],
    )
    _table(wb, "z06_rule_mapping", ["rule_id", "source_id", "rule_type", "target_id", "original_detail_raw", "body_style_scope", "runtime_action", "disabled_reason"], [["z06_rule_pdb_requires_bv4", "opt_pdb_001", "requires", "opt_bv4_001", "", "", "block", ""]])
    _table(wb, "z06_price_rules", ["price_rule_id", "condition_option_id", "price_rule_type", "target_option_id", "price_value", "body_style_scope", "trim_level_scope", "variant_scope", "notes"], [["z06_pr_pdb_bv4", "opt_pdb_001", "override", "opt_bv4_001", 500, "coupe", "1lz", "", ""]])
    _table(wb, "z06_rule_groups", ["group_id", "group_type", "source_id", "body_style_scope", "trim_level_scope", "variant_scope", "disabled_reason", "active", "notes"], [["z06_group_pdb", "requires_any", "opt_pdb_001", "", "", "", "", True, ""]])
    _table(wb, "z06_rule_group_members", ["group_id", "target_id", "display_order", "active"], [["z06_group_pdb", "opt_bv4_001", 1, True]])
    _table(wb, "z06_exclusive_groups", ["group_id", "selection_mode", "active", "notes"], [["z06_excl_wheels", "single", True, ""]])
    _table(wb, "z06_exclusive_members", ["group_id", "option_id", "display_order", "active"], [["z06_excl_wheels", "opt_pdb_001", 1, True], ["z06_excl_wheels", "opt_bv4_001", 2, True]])
    _table(wb, "z06_variant_overrides", ["option_id", "variant_id", "selectable", "display_behavior", "section_id", "active", "note"], [])
    _table(
        wb,
        "color_overrides",
        ["interior_id", "option_id", "rule_type", "adds_rpo"],
        [["3LZ_AQ9_HTA", "opt_gba_001", "requires", ""]],
    )
    # The inactive source's sheet must actually exist with a conflicting row,
    # so the active-flag filter is the only thing excluding it.
    _table(
        wb,
        "zr1_options",
        option_headers,
        [
            # Deliberately uses the id the plan builder would mint first for
            # PDB: locks in the add-after-delete seeding fix (new ids must
            # avoid keys that exist pre-batch even when deleted in-batch).
            ["opt_pdb_001", "PDB", 99999, "WRONG Scaffold Package", "must never surface", "", "sec_pain_001", "", 10, True, ""],
        ],
    )
    _table(
        wb,
        "section_master",
        ["section_id", "section_name", "selection_mode", "is_required", "display_order", "standard_behavior", "step_key"],
        [
            ["sec_pain_001", "Exterior Color", "single", True, 1, "", "paint"],
            ["sec_whee_001", "Wheels", "single_select_req", True, 2, "", "wheels"],
            # Standard-behavior section: rows assigned here feed the
            # standard-equipment queue (unless they carry a price — spec B9).
            ["sec_std_001", "Standard Equipment", "informational", False, 3, "included", "standard"],
        ],
    )
    _table(
        wb,
        "rule_phrase_map",
        ["phrase", "rule_type", "direction", "stop_phrases", "review_flag_default", "active", "notes"],
        [
            ["not available with", "excludes", "source_to_mentioned", "", True, True, "fixture"],
            ["requires", "requires", "source_to_mentioned", " or included with| included with", True, True, "fixture"],
            ["includes", "includes", "source_to_mentioned", " requires", True, True, "fixture"],
            ["included with", "includes", "mentioned_to_source", "", True, True, "fixture"],
            ["replaces", "replaces", "source_to_mentioned", "", True, True, "fixture"],
        ],
    )
    interior_headers = [
        "interior_id", "Interior Name", "Material", "Price", "Detail from Disclosure", "Color Overrides",
        "Trim", "Seat", "Interior Code", "Suede", "Stitch", "Two Tone", "section_id",
        "active_for_stingray", "requires_r6x", "included_option_id",
    ]
    _table(
        wb,
        "LZ_Interiors",
        interior_headers,
        [
            ["1LZ_AQ9_HTA", "Jet Black", "Mulan leather", 0, "", "", "1LZ", "AQ9", "HTA", "N2Z", "", "", "sec_lzint_001", False, False, ""],
            ["3LZ_AQ9_HTA", "Jet Black", "Napa leather", 0, "", "", "3LZ", "AQ9", "HTA", "", "", "", "sec_lzint_001", False, False, ""],
        ],
    )
    scope_headers = [
        "model_key", "interior_id", "trim_level", "active", "requires_option_id", "notes",
        "interior_seat_label", "interior_color_family", "interior_material_family", "interior_variant_label",
        "interior_group_display_order", "interior_material_display_order", "interior_choice_display_order",
        "interior_hierarchy_levels", "interior_parent_group_label", "interior_leaf_label", "interior_reference_order",
        "grouping_source",
    ]
    _table(
        wb,
        "model_interior_scope",
        scope_headers,
        [
            ["z06", "1LZ_AQ9_HTA", "1LZ", "True", "", "Workbook-owned interior trim scope metadata.", "AQ9 GT1 Bucket Seats", "HTA Jet Black", "Mulan leather", "HTA Jet Black", 1, 1, 1, '["1LZ", "AQ9 GT1 Bucket Seats", "HTA Jet Black"]', "AQ9 GT1 Bucket Seats", "HTA Jet Black", 1, "fixture"],
            ["z06", "3LZ_AQ9_HTA", "3LZ", "True", "", "Workbook-owned interior trim scope metadata.", "AQ9 GT1 Bucket Seats", "HTA Jet Black", "Napa leather", "HTA Jet Black", 2, 2, 2, '["3LZ", "AQ9 GT1 Bucket Seats", "HTA Jet Black"]', "AQ9 GT1 Bucket Seats", "HTA Jet Black", 2, "fixture"],
            ["zr1", "1LZ_AQ9_HTA", "1LZ", "True", "", "Workbook-owned interior trim scope metadata.", "AQ9 GT1 Bucket Seats", "HTA Jet Black", "Mulan leather", "HTA Jet Black", 1, 1, 1, '["1LZ", "AQ9 GT1 Bucket Seats", "HTA Jet Black"]', "AQ9 GT1 Bucket Seats", "HTA Jet Black", 1, "fixture"],
        ],
    )
    _table(
        wb,
        "interior_components",
        [
            "model_key", "interior_id", "rpo", "component_type", "label",
            "price_ref_type", "price_ref_code", "price_trim_scope",
            "display_order", "active", "notes",
        ],
        [
            ["z06", "1LZ_AQ9_HTA", "HTA", "base_color", "Jet Black", "", "", "1LZ", 10, True, "fixture"],
            ["z06", "3LZ_AQ9_HTA", "HTA", "base_color", "Jet Black", "", "", "3LZ", 10, True, "fixture"],
        ],
    )
    _table(
        wb,
        "default_selection_rules",
        ["model_key", "rule_id", "target_option_id", "condition_type", "condition_id", "body_style_scope", "trim_level_scope", "variant_scope", "priority", "active", "notes", "display_behavior"],
        [],
    )
    # Headers mirror the live workbook (probe 2026-07-06).
    presentation_rows = {
        "runtime_steps": (
            ["model_key", "step_key", "step_label", "runtime_order", "source", "active", "notes"],
            [["z06", "paint", "Paint", 1, "", True, ""], ["z06", "wheels", "Wheels", 2, "", True, ""]],
        ),
        "section_presentation": (
            ["model_key", "section_id", "display_label", "step_key", "display_behavior", "section_display_order", "standard_equipment_bucket", "standard_equipment_group_type", "auto_added_bucket", "active", "notes"],
            [["z06", "sec_pain_001", "Exterior Color", "paint", "", 1, "", "", "", True, ""]],
        ),
        "context_section_master": (
            ["model_key", "context_type", "section_id", "section_name", "selection_mode", "choice_mode", "is_required", "standard_behavior", "section_display_order", "step_key", "step_label", "active", "notes"],
            [["z06", "body", "ctx_body", "Body", "single", "", True, "", 1, "body", "Body", True, ""]],
        ),
        "order_summary_sections": (
            ["model_key", "section_key", "section_label", "display_order", "active", "notes"],
            [["z06", "sum_paint", "Exterior", 1, True, ""]],
        ),
        "step_order_summary_map": (
            ["model_key", "step_key", "section_key", "active", "notes"],
            [["z06", "paint", "sum_paint", True, ""]],
        ),
    }
    for sheet, (headers, rows) in presentation_rows.items():
        _table(wb, sheet, headers, rows)
    wb.save(path)
    return path
