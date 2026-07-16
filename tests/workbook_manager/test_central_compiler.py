from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.central_compiler import compile_central_tables
from app.compile_types import DecisionRequired
from app.workbook_profile import profile_workbook


@pytest.fixture
def compiled_central(real_workbook):
    profile = profile_workbook(real_workbook)
    return compile_central_tables(profile, real_workbook)


def test_variants_link_models_body_styles_and_trims(compiled_central):
    tables = {table.name: table for table in compiled_central}
    assert {row.values["body_style"] for row in tables["variants"].rows} == {
        "coupe",
        "convertible",
    }
    assert {row.values["trim_level"] for row in tables["variants"].rows} <= {
        row.values["trim_level"] for row in tables["trim_levels"].rows
    }
    assert {row.values["model_key"] for row in tables["model_variants"].rows} >= {
        "stingray",
        "grand_sport",
        "z06",
    }


def test_runtime_structure_has_one_model_aware_route(compiled_central):
    tables = {table.name: table for table in compiled_central}
    step_keys = {
        (row.values["model_key"], row.values["step_key"])
        for row in tables["runtime_steps"].rows
    }
    assert ("stingray", "body_style") in step_keys
    assert ("grand_sport", "body_style") in step_keys
    assert ("z06", "body_style") in step_keys


def test_compiles_every_central_table_with_declared_primary_keys(compiled_central):
    assert [(table.name, table.primary_key) for table in compiled_central] == [
        ("models", ("model_key",)),
        ("model_registry_promotion", ("model_key",)),
        ("body_styles", ("body_style",)),
        ("trim_levels", ("trim_level",)),
        ("variants", ("variant_id",)),
        ("model_variants", ("model_key", "variant_id")),
        ("sections", ("section_id",)),
        ("section_presentation", ("model_key", "section_id")),
        ("runtime_route_keys", ("model_key", "route_key")),
        ("runtime_steps", ("model_key", "step_key")),
        (
            "runtime_context_sections",
            ("model_key", "context_type", "section_id"),
        ),
        ("runtime_context_choices", ("model_key", "context_choice_id")),
        ("runtime_summary_sections", ("model_key", "section_key")),
        (
            "runtime_step_summary_map",
            ("model_key", "step_key"),
        ),
        ("model_assets", ("model_key",)),
        ("price_ref", ("option_type", "trim_level", "code")),
        ("rule_phrase_map", ("phrase",)),
    ]
    for table in compiled_central:
        assert table.rows
        assert len(
            {
                tuple(row.values[column] for column in table.primary_key)
                for row in table.rows
            }
        ) == len(table.rows)


def test_runtime_route_keys_distinguish_visible_and_hidden_routes(
    compiled_central,
):
    tables = {table.name: table for table in compiled_central}
    routes = {
        (row.values["model_key"], row.values["route_key"]): row.values[
            "route_kind"
        ]
        for row in tables["runtime_route_keys"].rows
    }
    assert routes[("stingray", "body_style")] == "visible_step"
    assert routes[("grand_sport", "body_style")] == "visible_step"
    assert routes[("z06", "body_style")] == "visible_step"
    assert routes[("z06", "standard_equipment")] == "hidden_summary_bucket"
    assert ("z06", "standard_equipment") not in {
        (row.values["model_key"], row.values["step_key"])
        for row in tables["runtime_steps"].rows
    }


def test_context_choices_compile_complete_generator_inventory_with_lineage(
    compiled_central,
):
    tables = {table.name: table for table in compiled_central}
    choices = tables["runtime_context_choices"].rows
    assert len(choices) == 24
    assert {
        model: sum(row.values["model_key"] == model for row in choices)
        for model in ("stingray", "grand_sport", "z06")
    } == {"stingray": 8, "grand_sport": 8, "z06": 8}

    body = next(
        row
        for row in choices
        if row.values["model_key"] == "stingray"
        and row.values["context_choice_id"] == "body_style__coupe"
    )
    assert dict(body.values) == {
        "model_key": "stingray",
        "context_choice_id": "body_style__coupe",
        "context_type": "body_style",
        "value": "coupe",
        "label": "Coupe",
        "description": "3 trims available",
        "info_tooltip": "",
        "section_id": "sec_context_body_style",
        "step_key": "body_style",
        "body_style": "coupe",
        "trim_level": None,
        "variant_id": None,
        "base_price": None,
        "display_order": 1,
        "active": True,
        "notes": "",
    }
    assert body.source_sheet == "model_variants"
    assert body.lineage_role == "normalized"
    assert body.mapping_parameters["context_choice_id"]["transform"] == (
        "derive_body_context_choice_id"
    )

    trim = next(
        row
        for row in choices
        if row.values["model_key"] == "grand_sport"
        and row.values["context_choice_id"] == "trim_level__convertible__1lt"
    )
    assert trim.values["value"] == "1LT"
    assert trim.values["label"] == "1LT"
    assert trim.values["trim_level"] == "1lt"
    assert trim.values["variant_id"] == "1lt_e67"
    assert trim.values["base_price"] == 95495
    assert trim.values["display_order"] == 10
    assert trim.values["section_id"] == "sec_context_trim_level"
    assert trim.values["step_key"] == "trim_level"
    assert trim.source_sheet == "model_variants"
    assert trim.mapping_parameters["info_tooltip"]["source_sheet"] == (
        "context_choice_copy"
    )
    assert trim.mapping_parameters["info_tooltip"]["source_row"] == 2

    assert {
        row.values["context_choice_id"]
        for row in choices
        if row.values["model_key"] == "z06"
    } == {
        "body_style__coupe",
        "body_style__convertible",
        "trim_level__coupe__1lz",
        "trim_level__coupe__2lz",
        "trim_level__coupe__3lz",
        "trim_level__convertible__1lz",
        "trim_level__convertible__2lz",
        "trim_level__convertible__3lz",
    }


def test_context_choices_reference_model_body_trim_variant_section_and_route(
    compiled_central,
):
    tables = {table.name: table for table in compiled_central}
    bodies = {row.values["body_style"] for row in tables["body_styles"].rows}
    trims = {row.values["trim_level"] for row in tables["trim_levels"].rows}
    variants = {
        (row.values["model_key"], row.values["variant_id"])
        for row in tables["model_variants"].rows
    }
    sections = {
        (row.values["model_key"], row.values["context_type"], row.values["section_id"])
        for row in tables["runtime_context_sections"].rows
    }
    routes = {
        (row.values["model_key"], row.values["route_key"])
        for row in tables["runtime_route_keys"].rows
    }
    for row in tables["runtime_context_choices"].rows:
        assert row.values["body_style"] in bodies
        assert row.values["trim_level"] is None or row.values["trim_level"] in trims
        assert row.values["variant_id"] is None or (
            row.values["model_key"], row.values["variant_id"]
        ) in variants
        assert (
            row.values["model_key"],
            row.values["context_type"],
            row.values["section_id"],
        ) in sections
        assert (row.values["model_key"], row.values["step_key"]) in routes


def test_normalization_preserves_original_values_in_lineage_evidence(
    compiled_central,
):
    tables = {table.name: table for table in compiled_central}

    unrestricted_price = next(
        row
        for row in tables["price_ref"].rows
        if row.values["option_type"] == "r6x" and row.values["code"] == "R6X"
    )
    assert unrestricted_price.values["trim_level"] is None


def test_context_sections_stay_distinct_and_share_runtime_step_keys(
    compiled_central,
):
    tables = {table.name: table for table in compiled_central}
    option_section_ids = {
        row.values["section_id"] for row in tables["sections"].rows
    }
    runtime_step_keys = {
        (row.values["model_key"], row.values["step_key"])
        for row in tables["runtime_steps"].rows
    }
    context_rows = tables["runtime_context_sections"].rows
    assert not {
        row.values["section_id"] for row in context_rows
    } & option_section_ids
    assert all(
        (row.values["model_key"], row.values["step_key"]) in runtime_step_keys
        for row in context_rows
    )


def test_only_active_live_rows_are_compiled(compiled_central):
    tables = {table.name: table for table in compiled_central}
    assert {row.values["model_key"] for row in tables["models"].rows} == {
        "stingray",
        "grand_sport",
        "z06",
    }
    assert len(tables["variants"].rows) == 18
    assert len(tables["model_variants"].rows) == 18
    assert len(tables["model_assets"].rows) == 3


def _copy_with_broken_model_variant(
    source: Path, destination: Path
) -> Path:
    workbook = load_workbook(source)
    sheet = workbook["model_variants"]
    sheet.cell(row=2, column=2, value="variant_missing_from_master")
    workbook.save(destination)
    workbook.close()
    return destination


def test_broken_central_reference_hard_stops_as_decision_required(
    tmp_path, real_workbook
):
    path = _copy_with_broken_model_variant(
        real_workbook, tmp_path / "broken-central-reference.xlsx"
    )
    profile = profile_workbook(path)
    with pytest.raises(DecisionRequired) as error:
        compile_central_tables(profile, path)
    assert error.value.code == "model_variant_reference_missing"
    assert error.value.source_sheet == "model_variants"
    assert error.value.source_row == 2


def test_summary_only_route_does_not_require_an_option_section_key(
    tmp_path, real_workbook
):
    path = tmp_path / "summary-only-route.xlsx"
    workbook = load_workbook(real_workbook)
    workbook["step_order_summary_map"].append(
        (
            "grand_sport",
            "hidden_summary_test",
            "pricing_summary",
            True,
            "Test-only hidden summary route.",
        )
    )
    workbook.save(path)
    workbook.close()

    tables = {
        table.name: table
        for table in compile_central_tables(profile_workbook(path), path)
    }
    route = next(
        row
        for row in tables["runtime_route_keys"].rows
        if row.values["model_key"] == "grand_sport"
        and row.values["route_key"] == "hidden_summary_test"
    )
    assert route.values["route_kind"] == "hidden_summary_bucket"


def test_context_section_can_reference_the_shared_route_domain(
    tmp_path, real_workbook
):
    path = tmp_path / "context-hidden-route.xlsx"
    workbook = load_workbook(real_workbook)
    sheet = workbook["context_section_master"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    for row_number in range(2, sheet.max_row + 1):
        if (
            sheet.cell(row_number, headers["model_key"]).value == "z06"
            and sheet.cell(row_number, headers["context_type"]).value
            == "body_style"
        ):
            sheet.cell(row_number, headers["step_key"]).value = (
                "standard_equipment"
            )
            break
    else:
        raise AssertionError("missing z06 body_style context row")
    workbook.save(path)
    workbook.close()

    tables = {
        table.name: table
        for table in compile_central_tables(profile_workbook(path), path)
    }
    row = next(
        row
        for row in tables["runtime_context_sections"].rows
        if row.values["model_key"] == "z06"
        and row.values["context_type"] == "body_style"
    )
    assert row.values["step_key"] == "standard_equipment"


def test_derived_route_lineage_preserves_pre_normalized_source_value(
    tmp_path, real_workbook
):
    path = tmp_path / "normalized-route.xlsx"
    workbook = load_workbook(real_workbook)
    sheet = workbook["runtime_steps"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    for row_number in range(2, sheet.max_row + 1):
        if (
            sheet.cell(row_number, headers["model_key"]).value == "stingray"
            and sheet.cell(row_number, headers["step_key"]).value
            == "body_style"
        ):
            sheet.cell(row_number, headers["step_key"]).value = "BODY_STYLE"
            break
    else:
        raise AssertionError("missing stingray body_style runtime step")
    workbook.save(path)
    workbook.close()

    tables = {
        table.name: table
        for table in compile_central_tables(profile_workbook(path), path)
    }
    route = next(
        row
        for row in tables["runtime_route_keys"].rows
        if row.values["model_key"] == "stingray"
        and row.values["route_key"] == "body_style"
    )
    assert route.mapping_parameters["route_key"] == {
        "original": "BODY_STYLE",
        "transform": "lowercase_then_derived_from_runtime_steps.step_key",
    }


def test_active_model_variant_with_unknown_model_fails_closed(
    tmp_path, real_workbook
):
    path = tmp_path / "unknown-model-variant.xlsx"
    workbook = load_workbook(real_workbook)
    workbook["model_variants"].cell(row=2, column=1, value="stingraye")
    workbook.save(path)
    workbook.close()

    with pytest.raises(DecisionRequired) as error:
        compile_central_tables(profile_workbook(path), path)
    assert error.value.code == "model_variant_model_reference_missing"
    assert error.value.source_sheet == "model_variants"
    assert error.value.source_row == 2
    assert error.value.source_column == "model_key"
    assert error.value.value == "stingraye"


def test_live_model_variant_count_must_match_model_contract(
    tmp_path, real_workbook
):
    path = tmp_path / "missing-model-variant.xlsx"
    workbook = load_workbook(real_workbook)
    workbook["model_variants"].cell(row=2, column=4, value=False)
    workbook.save(path)
    workbook.close()

    with pytest.raises(DecisionRequired) as error:
        compile_central_tables(profile_workbook(path), path)
    assert error.value.code == "model_variant_count_mismatch"
    assert error.value.source_sheet == "model_variants"
    assert error.value.value == {
        "model_key": "stingray",
        "expected": 6,
        "actual": 5,
    }


def test_step_summary_route_has_exactly_one_destination(
    tmp_path, real_workbook
):
    path = tmp_path / "duplicate-step-summary-route.xlsx"
    workbook = load_workbook(real_workbook)
    workbook["step_order_summary_map"].append(
        (
            "stingray",
            "body_style",
            "pricing_summary",
            True,
            "Test-only conflicting destination.",
        )
    )
    duplicate_row = workbook["step_order_summary_map"].max_row
    workbook.save(path)
    workbook.close()

    with pytest.raises(DecisionRequired) as error:
        compile_central_tables(profile_workbook(path), path)
    assert error.value.code == "step_summary_route_duplicate"
    assert error.value.source_sheet == "step_order_summary_map"
    assert error.value.source_row == duplicate_row
    assert error.value.value == ("stingray", "body_style")


def test_route_lineage_preserves_exact_raw_case_and_whitespace(
    tmp_path, real_workbook
):
    path = tmp_path / "raw-route-lineage.xlsx"
    workbook = load_workbook(real_workbook)
    sheet = workbook["runtime_steps"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    for row_number in range(2, sheet.max_row + 1):
        if (
            sheet.cell(row_number, headers["model_key"]).value == "stingray"
            and sheet.cell(row_number, headers["step_key"]).value
            == "body_style"
        ):
            sheet.cell(row_number, headers["model_key"]).value = "STINGRAY"
            sheet.cell(row_number, headers["step_key"]).value = " BODY_STYLE "
            break
    else:
        raise AssertionError("missing stingray body_style runtime step")
    workbook.save(path)
    workbook.close()

    tables = {
        table.name: table
        for table in compile_central_tables(profile_workbook(path), path)
    }
    route = next(
        row
        for row in tables["runtime_route_keys"].rows
        if row.values["model_key"] == "stingray"
        and row.values["route_key"] == "body_style"
    )
    assert route.mapping_parameters["model_key"] == {
        "original": "STINGRAY",
        "transform": "lowercase_then_derived_from_runtime_steps.model_key",
    }
    assert route.mapping_parameters["route_key"] == {
        "original": " BODY_STYLE ",
        "transform": (
            "trim_then_lowercase_then_derived_from_runtime_steps.step_key"
        ),
    }
    assert route.mapping_parameters["step_key"] == {
        "original": " BODY_STYLE ",
        "transform": "trim_then_lowercase",
    }
