"""Private destructive Stage 1 loader used only by reserved legacy tests."""

from __future__ import annotations

import hashlib
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app import db as dbmod
from app.specs import RAW_PRESERVED_SHEETS, TABLE_SPECS, TableSpec
from app.validation import check_references

def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sheet_rows(ws) -> tuple[list[str], list[tuple[int, dict]]]:
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        return [], []
    headers = [str(h) if h is not None else "" for h in header]
    rows: list[tuple[int, dict]] = []
    for idx, values in enumerate(it, start=2):
        if values is None:
            continue
        record = {h: _cell_text(v) for h, v in zip(headers, values) if h}
        if all(v == "" for v in record.values()):
            continue
        rows.append((idx, record))
    return headers, rows


def _registry_rows(wb) -> list[dict]:
    _, rows = _sheet_rows(wb["model_workbook_sources"])
    return [r for _, r in rows]


def _truthy(value: str) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y", "x"}


def _clear_imported_test_data(conn: sqlite3.Connection) -> None:
    for spec in TABLE_SPECS:
        conn.execute(f"DELETE FROM {spec.table}")
    conn.execute("DELETE FROM raw_sheet_rows")
    conn.commit()


class _LegacyImporterForTests:
    def __init__(self, conn: sqlite3.Connection, workbook_path: Path):
        self.conn = conn
        self.workbook_path = Path(workbook_path)
        self.issues: list[dict] = []
        self.row_counts: dict[str, int] = {}

    # ── issue reporting ───────────────────────────────────────────────
    def issue(self, severity, category, *, sheet="", src_row=None,
              table="", model="", key="", field="", message=""):
        self.issues.append({
            "severity": severity, "category": category, "sheet": sheet,
            "src_row": src_row, "table_name": table, "model_id": model,
            "entity_key": key, "field": field, "message": message,
        })

    # ── main entry ────────────────────────────────────────────────────
    def run(self) -> dict:
        digest = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        mtime_ns = str(self.workbook_path.stat().st_mtime_ns)
        wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet_names = set(wb.sheetnames)

        _clear_imported_test_data(self.conn)

        registry = _registry_rows(wb)
        handled_sheets: set[str] = set()

        for spec in TABLE_SPECS:
            if spec.role:
                self._import_model_scoped(wb, spec, registry, handled_sheets,
                                          sheet_names)
            else:
                self._import_fixed(wb, spec, handled_sheets, sheet_names)

        # Preserve every remaining sheet verbatim (declared raw + anything new).
        for name in wb.sheetnames:
            if name in handled_sheets:
                continue
            if name not in RAW_PRESERVED_SHEETS:
                self.issue("warning", "unmanaged_sheet", sheet=name,
                           message=f"sheet {name!r} has no normalized table; "
                                   "imported verbatim into raw_sheet_rows")
            self._import_raw(wb, name)

        wb.close()

        ref_issues = check_references(self.conn)
        for iss in ref_issues:
            self.issues.append(iss)

        issue_counts = {"error": 0, "warning": 0}
        for iss in self.issues:
            issue_counts[iss["severity"]] = issue_counts.get(iss["severity"], 0) + 1

        cur = self.conn.execute(
            "INSERT INTO import_runs(ts, workbook_path, workbook_mtime_ns, "
            "workbook_sha256, status, row_counts_json, issue_counts_json) "
            "VALUES(?,?,?,?,?,?,?)",
            (_now(), str(self.workbook_path), mtime_ns, digest,
             "imported_with_issues" if issue_counts["error"] else "imported",
             json.dumps(self.row_counts), json.dumps(issue_counts)),
        )
        run_id = cur.lastrowid
        self.conn.executemany(
            "INSERT INTO import_issues(run_id, severity, category, sheet, "
            "src_row, table_name, model_id, entity_key, field, message) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)",
            [(run_id, i["severity"], i["category"], i["sheet"], i["src_row"],
              i["table_name"], i["model_id"], i["entity_key"], i["field"],
              i["message"]) for i in self.issues],
        )
        dbmod.set_meta(self.conn, "workbook_mtime_ns", mtime_ns)
        dbmod.set_meta(self.conn, "workbook_sha256", digest)
        dbmod.set_meta(self.conn, "last_import_run_id", str(run_id))
        self.conn.commit()
        return self.report(run_id)

    def report(self, run_id: int) -> dict:
        run = self.conn.execute(
            "SELECT * FROM import_runs WHERE id=?", (run_id,)).fetchone()
        issues = self.conn.execute(
            "SELECT * FROM import_issues WHERE run_id=? ORDER BY id",
            (run_id,)).fetchall()
        return {
            "run": dict(run),
            "issues": [dict(i) for i in issues],
        }

    # ── sheet import paths ────────────────────────────────────────────
    def _import_fixed(self, wb, spec: TableSpec, handled: set, names: set):
        total = 0
        for sheet in spec.sheet:
            if sheet not in names:
                self.issue("error", "missing_sheet", sheet=sheet,
                           table=spec.table,
                           message=f"expected sheet {sheet!r} not found")
                continue
            handled.add(sheet)
            total += self._ingest_sheet(wb, sheet, spec, model_id=None)
        self.row_counts[spec.table] = total

    def _import_model_scoped(self, wb, spec: TableSpec, registry, handled,
                             names):
        total = 0
        seen_sheets: set[tuple[str, str]] = set()
        for row in registry:
            if row.get("source_role") != spec.role:
                continue
            model = row.get("model_key", "")
            sheet = row.get("sheet_name", "")
            if not model or not sheet:
                continue
            if (model, sheet) in seen_sheets:
                continue
            seen_sheets.add((model, sheet))
            if sheet not in names:
                self.issue("error", "missing_sheet", sheet=sheet, model=model,
                           table=spec.table,
                           message=f"registered sheet {sheet!r} for model "
                                   f"{model!r} not found in workbook")
                continue
            # Shared sheets (color_overrides, interiors) are fixed-table specs;
            # a model-scoped spec re-registering a shared sheet is skipped when
            # another table already owns that sheet.
            handled.add(sheet)
            total += self._ingest_sheet(wb, sheet, spec, model_id=model)
        self.row_counts[spec.table] = total

    def _import_raw(self, wb, sheet: str):
        headers, rows = _sheet_rows(wb[sheet])
        self.conn.executemany(
            "INSERT INTO raw_sheet_rows(sheet, src_row, data_json) VALUES(?,?,?)",
            [(sheet, idx, json.dumps(rec, ensure_ascii=False))
             for idx, rec in rows],
        )
        self.row_counts[f"raw:{sheet}"] = len(rows)

    def _ingest_sheet(self, wb, sheet: str, spec: TableSpec,
                      model_id: str | None) -> int:
        headers, rows = _sheet_rows(wb[sheet])
        expected = [c.header for c in spec.columns]
        missing = [h for h in expected if h not in headers]
        extra = [h for h in headers if h and h not in expected]
        if missing:
            self.issue("error", "missing_columns", sheet=sheet,
                       table=spec.table, model=model_id or "",
                       message=f"missing columns {missing}; these fields "
                               "import as blank")
        if extra:
            self.issue("warning", "unmapped_columns", sheet=sheet,
                       table=spec.table, model=model_id or "",
                       message=f"unmapped columns {extra} preserved via raw "
                               "export copy only")

        sql_cols = ["src_sheet", "src_row"]
        if spec.model_scoped:
            sql_cols.append("model_id")
        sql_cols += [c.sql_name() for c in spec.columns]
        placeholders = ",".join("?" for _ in sql_cols)
        collist = ",".join(f'"{c}"' for c in sql_cols)
        insert = f"INSERT INTO {spec.table} ({collist}) VALUES ({placeholders})"

        seen_keys: set[tuple] = set()
        count = 0
        for idx, rec in rows:
            keyvals = tuple(rec.get(k, "") for k in spec.key)
            key_id = (model_id or "",) + keyvals if spec.model_scoped else keyvals
            key_label = "/".join(v for v in keyvals if v) or f"row{idx}"
            if any(v == "" for v in keyvals):
                self.issue("error", "missing_identifier", sheet=sheet,
                           src_row=idx, table=spec.table,
                           model=model_id or "", key=key_label,
                           field=",".join(spec.key),
                           message=f"blank key column(s) in {sheet} row {idx}")
            if key_id in seen_keys:
                self.issue("error", "duplicate_id", sheet=sheet, src_row=idx,
                           table=spec.table, model=model_id or "",
                           key=key_label,
                           message=f"duplicate key {key_label!r} in {sheet} "
                                   f"row {idx}; first occurrence kept")
                continue
            seen_keys.add(key_id)
            values = [sheet, idx]
            if spec.model_scoped:
                values.append(model_id)
            values += [rec.get(c.header, "") for c in spec.columns]
            try:
                self.conn.execute(insert, values)
                count += 1
            except sqlite3.IntegrityError as exc:
                # Fixed-sheet tables with global keys (e.g. interiors across
                # two sheets) can still collide; report instead of failing.
                self.issue("error", "duplicate_id", sheet=sheet, src_row=idx,
                           table=spec.table, model=model_id or "",
                           key=key_label, message=str(exc))
        return count



def import_workbook_for_legacy_tests(
    conn: sqlite3.Connection, workbook_path: Path
) -> dict:
    return _LegacyImporterForTests(conn, workbook_path).run()
