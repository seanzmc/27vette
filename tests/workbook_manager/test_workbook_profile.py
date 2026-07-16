import shutil

import pytest
from openpyxl import load_workbook

from app.catalog import LIVE_MODELS
from app.workbook_profile import profile_workbook


def _copy_with_extra_sheet(real_workbook, tmp_path, sheet_name):
    copied = tmp_path / f"{sheet_name}.xlsx"
    shutil.copyfile(real_workbook, copied)
    workbook = load_workbook(copied)
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(("Exact Header", "Another Header"))
    sheet.append(("value", 1))
    workbook.save(copied)
    workbook.close()
    return copied


def _copy_with_source_binding(
    real_workbook,
    tmp_path,
    *,
    model_key,
    existing_role,
    source_role,
    sheet_name,
):
    copied = tmp_path / f"{model_key}-{source_role}.xlsx"
    shutil.copyfile(real_workbook, copied)
    workbook = load_workbook(copied)
    sources = workbook["model_workbook_sources"]
    headers = {cell.value: cell.column for cell in sources[1]}
    for row_number in range(2, sources.max_row + 1):
        if (
            sources.cell(row_number, headers["model_key"]).value == model_key
            and sources.cell(row_number, headers["source_role"]).value
            == existing_role
        ):
            sources.cell(row_number, headers["source_role"]).value = source_role
            sources.cell(row_number, headers["sheet_name"]).value = sheet_name
            break
    else:
        raise AssertionError((model_key, existing_role))
    workbook.save(copied)
    workbook.close()
    return copied


def test_every_workbook_sheet_has_one_disposition(real_workbook):
    profile = profile_workbook(real_workbook)

    assert len(profile.sheets) == 65
    assert len({sheet.source_sheet for sheet in profile.sheets}) == 65
    assert not [sheet for sheet in profile.sheets if not sheet.disposition]
    assert not [sheet for sheet in profile.sheets if sheet.disposition == "decision_required"]


def test_live_model_sources_have_identical_roles(real_workbook):
    profile = profile_workbook(real_workbook)
    role_sets = {
        model: frozenset(profile.active_sources[model])
        for model in LIVE_MODELS
    }

    assert profile.active_models == LIVE_MODELS
    assert len(set(role_sets.values())) == 1
    assert all(
        source_sheet in {sheet.source_sheet for sheet in profile.sheets}
        for sources in profile.active_sources.values()
        for source_sheet in sources.values()
    )


def test_current_workbook_has_no_retired_generated_sources(real_workbook):
    profile = profile_workbook(real_workbook)
    by_name = {sheet.source_sheet: sheet for sheet in profile.sheets}

    assert not [name for name in by_name if name.startswith("form_")]
    assert not [
        sheet
        for sheet in profile.sheets
        if sheet.disposition == "generated_artifact_validation"
    ]
    assert by_name["zr1_options"].disposition == "inactive_future_source"


def test_retired_generated_sheet_is_validation_only(real_workbook, tmp_path):
    copied = _copy_with_extra_sheet(real_workbook, tmp_path, "form_choices")

    profile = profile_workbook(copied)
    sheet = next(s for s in profile.sheets if s.source_sheet == "form_choices")

    assert sheet.disposition == "generated_artifact_validation"
    assert sheet.destination_tables == ()
    assert any(
        finding.status == "contract_mismatch"
        and finding.code == "retired_generated_sheet_present"
        and finding.source_sheet == "form_choices"
        for finding in profile.findings
    )


def test_unknown_sheet_fails_closed_as_decision_required(real_workbook, tmp_path):
    copied = _copy_with_extra_sheet(real_workbook, tmp_path, "unexpected_source")

    profile = profile_workbook(copied)
    sheet = next(s for s in profile.sheets if s.source_sheet == "unexpected_source")

    assert sheet.disposition == "decision_required"
    assert sheet.destination_tables == ()
    assert any(
        finding.status == "decision_required"
        and finding.code == "source_sheet_unclassified"
        and finding.source_sheet == "unexpected_source"
        for finding in profile.findings
    )


def test_unknown_inactive_source_role_requires_a_decision(real_workbook, tmp_path):
    copied = tmp_path / "unknown-future-role.xlsx"
    shutil.copyfile(real_workbook, copied)
    workbook = load_workbook(copied)
    sources = workbook["model_workbook_sources"]
    headers = {cell.value: cell.column for cell in sources[1]}
    for row_number in range(2, sources.max_row + 1):
        if sources.cell(row_number, headers["sheet_name"]).value == "zr1_options":
            sources.cell(row_number, headers["source_role"]).value = (
                "unapproved_future_role"
            )
            break
    workbook.save(copied)
    workbook.close()

    profile = profile_workbook(copied)
    sheet = next(s for s in profile.sheets if s.source_sheet == "zr1_options")

    assert sheet.disposition == "decision_required"
    assert any(
        finding.status == "decision_required"
        and finding.code == "source_role_unclassified"
        and finding.value == "unapproved_future_role"
        for finding in profile.findings
    )


@pytest.mark.parametrize(
    "control_sheet",
    ("model_master", "model_registry_promotion", "model_workbook_sources"),
)
def test_duplicate_control_header_is_rejected(
    real_workbook,
    tmp_path,
    control_sheet,
):
    copied = tmp_path / f"duplicate-{control_sheet}.xlsx"
    shutil.copyfile(real_workbook, copied)
    workbook = load_workbook(copied)
    sheet = workbook[control_sheet]
    sheet.cell(1, sheet.max_column + 1).value = "active"
    workbook.save(copied)
    workbook.close()

    with pytest.raises(
        ValueError,
        match=rf"Sheet {control_sheet!r} has duplicate header 'active'",
    ):
        profile_workbook(copied)


def test_known_registry_role_cannot_claim_a_central_sheet(real_workbook, tmp_path):
    copied = _copy_with_source_binding(
        real_workbook,
        tmp_path,
        model_key="stingray",
        existing_role="source_option_sheet",
        source_role="source_option_sheet",
        sheet_name="model_master",
    )

    profile = profile_workbook(copied)
    sheet = next(s for s in profile.sheets if s.source_sheet == "model_master")

    assert sheet.disposition == "decision_required"
    assert sheet.destination_tables == ()
    assert any(
        finding.code == "source_ownership_conflict"
        and finding.source_sheet == "model_master"
        and finding.model_key == "stingray"
        and finding.value["source_role"] == "source_option_sheet"
        for finding in profile.findings
    )


def test_unknown_registry_role_cannot_claim_a_central_sheet(real_workbook, tmp_path):
    copied = _copy_with_source_binding(
        real_workbook,
        tmp_path,
        model_key="zr1",
        existing_role="source_option_sheet",
        source_role="unapproved_future_role",
        sheet_name="model_master",
    )

    profile = profile_workbook(copied)
    sheet = next(s for s in profile.sheets if s.source_sheet == "model_master")

    assert sheet.disposition == "decision_required"
    assert sheet.destination_tables == ()
    assert any(
        finding.code == "source_ownership_conflict"
        and finding.source_sheet == "model_master"
        and finding.model_key == "zr1"
        and finding.value["source_role"] == "unapproved_future_role"
        for finding in profile.findings
    )


def test_profile_preserves_exact_sheet_names_headers_and_data_row_counts(
    real_workbook,
):
    profile = profile_workbook(real_workbook)
    by_name = {sheet.source_sheet: sheet for sheet in profile.sheets}

    assert by_name["PriceRef"].headers == (
        "OptionType",
        "Trim",
        "Code",
        "Price",
    )
    assert by_name["PriceRef"].row_count == 21
    assert by_name["LZ_Interiors"].headers[1] == "Interior Name"


def test_profile_records_current_workbook_hash(real_workbook):
    profile = profile_workbook(real_workbook)

    assert profile.workbook_path == real_workbook
    assert len(profile.workbook_sha256) == 64
    assert set(profile.workbook_sha256) <= set("0123456789abcdef")
