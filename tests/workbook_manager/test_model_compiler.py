from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.catalog import LIVE_MODELS, physical_table
from app.central_compiler import compile_central_tables
from app.compile_types import DecisionRequired
from app.model_compiler import compile_direct_model_tables
from app.workbook_profile import profile_workbook


DIRECT_ROLES = (
    "options",
    "option_availability",
    "rule_mapping",
    "price_rules",
    "rule_groups",
    "rule_group_members",
    "exclusive_groups",
    "exclusive_group_members",
    "variant_overrides",
)


@pytest.fixture
def compiled_central(real_workbook):
    profile = profile_workbook(real_workbook)
    return compile_central_tables(profile, real_workbook)


@pytest.fixture
def compiled_models(real_workbook, compiled_central):
    profile = profile_workbook(real_workbook)
    return compile_direct_model_tables(profile, real_workbook, compiled_central)


def test_model_collections_have_identical_direct_roles(compiled_models):
    assert tuple((table.model_key, table.role) for table in compiled_models) == tuple(
        (model, role) for model in LIVE_MODELS for role in DIRECT_ROLES
    )
    assert all(
        table.name == physical_table(table.model_key, table.role)
        for table in compiled_models
    )


def test_options_and_availability_reconcile_real_workbook(compiled_models):
    by_name = {table.name: table for table in compiled_models}
    assert {
        model: len(by_name[f"{model}_options"].rows) for model in LIVE_MODELS
    } == {"stingray": 242, "grand_sport": 241, "z06": 244}
    assert {
        model: len(by_name[f"{model}_option_availability"].rows)
        for model in LIVE_MODELS
    } == {"stingray": 1452, "grand_sport": 1446, "z06": 1464}
    assert all(
        row.values["model_key"] == table.model_key
        for table in compiled_models
        for row in table.rows
    )


def test_direct_rows_preserve_registered_source_and_workbook_row(compiled_models):
    table = next(table for table in compiled_models if table.name == "grand_sport_options")
    first = table.rows[0]
    assert first.source_sheet == "grandSport_options"
    assert first.source_row == 2
    assert first.values["option_id"] == "opt_uvb_001"
    assert first.lineage_role == "normalized"
    assert first.mapping_parameters["price"] == {
        "original": None,
        "canonical": 0,
        "transform": "blank_to_zero",
        "reverse_transform": "restore_original_number_from_lineage",
    }


def test_live_trimmed_text_preserves_exact_raw_lineage(compiled_models):
    table = next(table for table in compiled_models if table.name == "stingray_options")
    trailing = next(row for row in table.rows if row.source_row == 19)
    leading = next(row for row in table.rows if row.source_row == 135)

    assert trailing.values["description"] == (
        "Illuminate the engine bay. Included with Engine Appearance package "
        "and accessory engine covers."
    )
    assert trailing.mapping_parameters["description"] == {
        "original": (
            "Illuminate the engine bay. Included with Engine Appearance package "
            "and accessory engine covers. "
        ),
        "canonical": trailing.values["description"],
        "transform": "trim_text",
        "reverse_transform": "restore_original_text_from_lineage",
    }
    assert leading.values["detail_raw"].startswith("includes (B4Z)")
    assert leading.mapping_parameters["detail_raw"] == {
        "original": (
            " includes (B4Z) Performance Traction Management 1. Requires "
            "(Z51) Z51 Performance Package."
        ),
        "canonical": leading.values["detail_raw"],
        "transform": "trim_text",
        "reverse_transform": "restore_original_text_from_lineage",
    }


def test_polymorphic_sources_are_typed_by_exact_membership(compiled_models):
    by_name = {table.name: table for table in compiled_models}
    for model in LIVE_MODELS:
        rules = by_name[f"{model}_rule_mapping"].rows
        prices = by_name[f"{model}_price_rules"].rows
        assert all(
            (row.values["source_option_id"] is None)
            != (row.values["source_interior_id"] is None)
            for row in rules
        )
        assert all(
            (row.values["condition_option_id"] is None)
            != (row.values["condition_interior_id"] is None)
            for row in prices
        )


def test_aliases_and_scope_normalization_retain_reversible_evidence(compiled_models):
    by_name = {table.name: table for table in compiled_models}
    rule = next(
        row
        for row in by_name["stingray_rule_mapping"].rows
        if row.values["source_interior_id"] is not None
    )
    assert rule.mapping_parameters["source_interior_id"] == {
        "source_column": "source_id",
        "original": rule.values["source_interior_id"],
        "canonical": rule.values["source_interior_id"],
        "transform": "typed_entity_reference",
        "reverse_transform": "coalesce_typed_entity_reference",
    }
    assert rule.mapping_parameters["target_option_id"] == {
        "source_column": "target_id",
        "original": rule.values["target_option_id"],
        "canonical": rule.values["target_option_id"],
        "transform": "option_reference",
        "reverse_transform": "restore_original_text_from_lineage",
    }

    unrestricted = next(
        row
        for row in by_name["z06_price_rules"].rows
        if row.mapping_parameters.get("body_style_scope", {}).get("original") == "*"
    )
    assert unrestricted.values["body_style_scope"] is None
    assert unrestricted.mapping_parameters["body_style_scope"] == {
        "original": "*",
        "canonical": None,
        "transform": "asterisk_to_null",
        "reverse_transform": "restore_original_scope_from_lineage",
    }
    restricted = next(
        row
        for row in by_name["z06_price_rules"].rows
        if row.values["trim_level_scope"] == "1lz"
    )
    assert restricted.mapping_parameters["trim_level_scope"] == {
        "original": "1LZ",
        "canonical": "1lz",
        "transform": "lowercase",
        "reverse_transform": "restore_original_scope_from_lineage",
    }


def test_schema_mappings_state_actual_column_contracts(compiled_models):
    mappings = {
        (table.role, mapping.source_column, mapping.destination_column): (
            mapping.transform,
            mapping.reverse_transform,
        )
        for table in compiled_models
        if table.model_key == "stingray"
        for mapping in table.schema_mappings
    }
    assert mappings[("options", "description", "description")] == (
        "blank_to_empty_else_trim_text",
        "restore_original_text_from_lineage",
    )
    assert mappings[("options", "price", "price")] == (
        "blank_to_zero_else_normalize_integer",
        "restore_original_number_from_lineage",
    )
    assert mappings[("options", "selectable", "selectable")] == (
        "normalize_workbook_boolean",
        "restore_original_boolean_from_lineage",
    )
    assert mappings[("price_rules", "body_style_scope", "body_style_scope")] == (
        "blank_or_asterisk_to_null_else_trim_and_lowercase",
        "restore_original_scope_from_lineage",
    )
    assert mappings[("rule_mapping", "source_id", "source_option_id")] == (
        "require_nonblank_trim_text_then_typed_entity_reference",
        "coalesce_typed_entity_reference_then_restore_original_text_from_lineage",
    )
    assert mappings[("rule_group_members", "target_id", "target_option_id")] == (
        "require_nonblank_trim_text_then_option_reference",
        "restore_original_text_from_lineage",
    )


def test_mapping_evidence_is_deeply_immutable(compiled_models):
    row = next(
        row
        for table in compiled_models
        if table.name == "grand_sport_options"
        for row in table.rows
        if "price" in row.mapping_parameters
    )
    with pytest.raises(TypeError):
        row.mapping_parameters["price"]["transform"] = "mutated"


def _copy_with_changed_cell(
    source: Path,
    destination: Path,
    sheet_name: str,
    column: str,
    value: object,
) -> tuple[Path, int]:
    workbook = load_workbook(source)
    sheet = workbook[sheet_name]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers[column], value)
    workbook.save(destination)
    workbook.close()
    return destination, 2


@pytest.mark.parametrize(
    ("sheet", "column", "value", "code"),
    (
        ("stingray_ovs", "option_id", "missing_option", "option_reference_missing"),
        ("rule_mapping", "source_id", "missing_source", "entity_reference_ambiguous_or_missing"),
        ("price_rules", "target_option_id", "missing_target", "option_reference_missing"),
        ("rule_groups", "body_style_scope", "sedan", "scope_reference_missing"),
    ),
)
def test_unresolved_direct_relationships_fail_closed_with_source_evidence(
    tmp_path, real_workbook, sheet, column, value, code
):
    path, source_row = _copy_with_changed_cell(
        real_workbook, tmp_path / f"broken-{sheet}.xlsx", sheet, column, value
    )
    profile = profile_workbook(path)
    central = compile_central_tables(profile, path)
    with pytest.raises(DecisionRequired) as error:
        compile_direct_model_tables(profile, path, central)
    assert error.value.code == code
    assert error.value.source_sheet == sheet
    assert error.value.source_row == source_row
    assert error.value.source_column == column
    assert error.value.value == value
