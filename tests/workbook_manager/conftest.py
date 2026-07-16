from pathlib import Path
import sqlite3
import shutil
import sys

import pytest
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = REPO_ROOT / "workbook-manager" / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app import db  # noqa: E402


@pytest.fixture
def repo_root() -> Path:
    return REPO_ROOT


@pytest.fixture
def real_workbook(repo_root: Path) -> Path:
    return repo_root / "stingray_master.xlsx"


@pytest.fixture
def broken_fk_workbook(tmp_path, real_workbook) -> Path:
    destination = tmp_path / "broken-fk.xlsx"
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    options = workbook["stingray_options"]
    ovs = workbook["stingray_ovs"]
    option_headers = {cell.value: cell.column for cell in options[1]}
    ovs_headers = {cell.value: cell.column for cell in ovs[1]}
    referenced = {
        ovs.cell(row, ovs_headers["option_id"]).value
        for row in range(2, ovs.max_row + 1)
    }
    for row in range(2, options.max_row + 1):
        if options.cell(row, option_headers["option_id"]).value in referenced:
            options.delete_rows(row)
            break
    else:
        raise AssertionError("no referenced Stingray option row found")
    workbook.save(destination)
    workbook.close()
    return destination


@pytest.fixture
def legacy_db_path(tmp_path) -> Path:
    path = tmp_path / "legacy.sqlite3"
    conn = sqlite3.connect(path)
    conn.execute(
        "CREATE TABLE pending_changes ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "ts TEXT NOT NULL,"
        "session_id TEXT NOT NULL DEFAULT '',"
        "table_name TEXT NOT NULL,"
        "model_id TEXT NOT NULL DEFAULT '',"
        "entity_key_json TEXT NOT NULL,"
        "op TEXT NOT NULL,"
        "old_json TEXT,"
        "new_json TEXT,"
        "status TEXT NOT NULL DEFAULT 'staged',"
        "validation_json TEXT NOT NULL DEFAULT '{}',"
        "confirmed_dependencies INTEGER NOT NULL DEFAULT 0"
        ")"
    )
    conn.commit()
    conn.close()
    return path


@pytest.fixture
def unowned_shared_row_workbook(tmp_path, real_workbook) -> Path:
    destination = tmp_path / "unowned-shared-row.xlsx"
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    interiors = workbook["lt_interiors"]
    interior_headers = {cell.value: cell.column for cell in interiors[1]}
    interior_row = interiors.max_row + 1
    interiors.cell(interior_row, interior_headers["interior_id"], "int_unowned_test")
    interiors.cell(interior_row, interior_headers["Interior Name"], "Unowned Test")
    interiors.cell(interior_row, interior_headers["Material"], "Test material")
    interiors.cell(interior_row, interior_headers["Price"], 0)
    interiors.cell(interior_row, interior_headers["Trim"], "1LT")
    interiors.cell(interior_row, interior_headers["Seat"], "AQ9")
    interiors.cell(interior_row, interior_headers["Interior Code"], "ZZZ")
    interiors.cell(interior_row, interior_headers["section_id"], "sec_intc_001")
    interiors.cell(interior_row, interior_headers["active_for_stingray"], True)
    interiors.cell(interior_row, interior_headers["requires_r6x"], False)

    overrides = workbook["color_overrides"]
    override_headers = {cell.value: cell.column for cell in overrides[1]}
    override_row = overrides.max_row + 1
    overrides.cell(override_row, override_headers["interior_id"], "int_unowned_test")
    overrides.cell(override_row, override_headers["option_id"], "opt_g26_001")
    overrides.cell(override_row, override_headers["rule_type"], "requires")
    overrides.cell(override_row, override_headers["adds_rpo"], "opt_d30_001")
    workbook.save(destination)
    workbook.close()
    return destination


@pytest.fixture
def wildcard_asset_overlay_workbook(tmp_path, real_workbook) -> Path:
    destination = tmp_path / "wildcard-asset-overlay.xlsx"
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    assets = workbook["asset_map"]
    headers = [cell.value for cell in assets[1]]
    row = {header: None for header in headers}
    row.update(
        {
            "model_key": "stingray",
            "target_type": "option",
            "target_id": "opt_gba_001",
            "image_url": "https://example.test/stingray-gba-overlay.png",
            "image_alt": "Stingray black overlay",
            "image_fit": "contain",
            "image_position": "top",
            "active": True,
            "notes": "Task 5 exact-model overlay fixture.",
        }
    )
    assets.append(tuple(row[header] for header in headers))
    workbook.save(destination)
    workbook.close()
    return destination


@pytest.fixture
def unsupported_wildcard_context_asset_workbook(tmp_path, real_workbook) -> Path:
    destination = tmp_path / "wildcard-context-asset.xlsx"
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    assets = workbook["asset_map"]
    headers = [cell.value for cell in assets[1]]
    row = {header: None for header in headers}
    row.update(
        {
            "model_key": "*",
            "target_type": "context_choice",
            "target_id": "body_style__coupe",
            "image_url": "https://example.test/unsupported-context.png",
            "active": True,
        }
    )
    assets.append(tuple(row[header] for header in headers))
    workbook.save(destination)
    workbook.close()
    return destination


def _copy_with_color_override_value(
    tmp_path, real_workbook, *, name: str, column: str, value: str
) -> Path:
    destination = tmp_path / f"{name}.xlsx"
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    sheet = workbook["color_overrides"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers[column]).value = value
    workbook.save(destination)
    workbook.close()
    return destination


@pytest.fixture
def invalid_color_added_option_workbook(tmp_path, real_workbook) -> Path:
    return _copy_with_color_override_value(
        tmp_path,
        real_workbook,
        name="invalid-color-added-option",
        column="adds_rpo",
        value="opt_missing_added",
    )


@pytest.fixture
def invalid_color_interior_workbook(tmp_path, real_workbook) -> Path:
    return _copy_with_color_override_value(
        tmp_path,
        real_workbook,
        name="invalid-color-interior",
        column="interior_id",
        value="int_missing_color",
    )


@pytest.fixture
def invalid_color_option_workbook(tmp_path, real_workbook) -> Path:
    return _copy_with_color_override_value(
        tmp_path,
        real_workbook,
        name="invalid-color-option",
        column="option_id",
        value="opt_missing_color",
    )


@pytest.fixture
def conflicting_color_owners_workbook(tmp_path, real_workbook) -> Path:
    destination = tmp_path / "conflicting-color-owners.xlsx"
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    sheet = workbook["color_overrides"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["interior_id"]).value = "3LT_AE4_EL9"
    sheet.cell(2, headers["option_id"]).value = "opt_085"
    workbook.save(destination)
    workbook.close()
    return destination


def _append_asset_rows(destination: Path, real_workbook: Path, rows) -> Path:
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    sheet = workbook["asset_map"]
    headers = [cell.value for cell in sheet[1]]
    for values in rows:
        row = {header: None for header in headers}
        row.update(values)
        sheet.append(tuple(row[header] for header in headers))
    workbook.save(destination)
    workbook.close()
    return destination


@pytest.fixture
def duplicate_wildcard_asset_workbook(tmp_path, real_workbook) -> Path:
    return _append_asset_rows(
        tmp_path / "duplicate-wildcard-asset.xlsx",
        real_workbook,
        (
            {
                "model_key": "*",
                "target_type": "option",
                "target_id": "opt_gba_001",
                "image_url": "https://example.test/duplicate-wildcard.png",
                "active": True,
            },
        ),
    )


@pytest.fixture
def duplicate_exact_asset_workbook(tmp_path, real_workbook) -> Path:
    return _append_asset_rows(
        tmp_path / "duplicate-exact-asset.xlsx",
        real_workbook,
        (
            {
                "model_key": "stingray",
                "target_type": "option",
                "target_id": "opt_gba_001",
                "image_url": "https://example.test/first-exact.png",
                "active": True,
            },
            {
                "model_key": "stingray",
                "target_type": "option",
                "target_id": "opt_gba_001",
                "image_url": "https://example.test/duplicate-exact.png",
                "active": True,
            },
        ),
    )


@pytest.fixture
def unknown_shared_model_workbook(tmp_path, real_workbook) -> Path:
    destination = tmp_path / "unknown-shared-model.xlsx"
    shutil.copyfile(real_workbook, destination)
    workbook = load_workbook(destination)
    sheet = workbook["model_interior_scope"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["model_key"]).value = "unknown_future"
    workbook.save(destination)
    workbook.close()
    return destination


@pytest.fixture
def connection(tmp_path):
    conn = db.connect(tmp_path / "test.sqlite3")
    try:
        yield conn
    finally:
        conn.close()
