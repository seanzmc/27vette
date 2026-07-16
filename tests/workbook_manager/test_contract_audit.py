import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app import contract_audit, db, export_adapter, importer
from app.contract_audit import ContractAudit, ContractDifference
from corvette_form_generator import production
from corvette_form_generator.model_configs import discover_generation_model_configs
from corvette_form_generator.model_generation import generate_model_artifacts


TIMESTAMP_KEYS = {"generated_at", "sourceGeneratedAt", "generatedAt"}


def load_rows(path: Path, sheet_name: str) -> tuple[dict[str, object], ...]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook[sheet_name]
    values = worksheet.iter_rows(values_only=True)
    headers = tuple(str(value) if value is not None else "" for value in next(values))
    rows = tuple(
        {header: value for header, value in zip(headers, row) if header}
        for row in values
        if any(value is not None and value != "" for value in row)
    )
    workbook.close()
    return rows


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_scope_and_semantic_aliases_round_trip(imported_db, real_workbook, tmp_path):
    output = tmp_path / "roundtrip.xlsx"
    export_adapter.export_comparison_workbook(imported_db, real_workbook, output)

    assert load_rows(output, "color_overrides") == load_rows(
        real_workbook, "color_overrides"
    )
    for sheet in ("rule_mapping", "grandSport_rule_mapping", "z06_rule_mapping"):
        assert load_rows(output, sheet) == load_rows(real_workbook, sheet)


def test_all_promoted_contracts_match_except_timestamps(
    imported_db, real_workbook, tmp_path, repo_root
):
    repo_artifacts = tuple(
        path
        for root in (repo_root / "form-output", repo_root / "form-app")
        for path in root.rglob("*")
        if path.is_file()
    )
    before = {path: _digest(path) for path in repo_artifacts}
    audit = contract_audit.audit_runtime_contracts(
        imported_db, real_workbook, tmp_path
    )

    assert audit.models == ("stingray", "grand_sport", "z06")
    assert audit.differences == ()
    assert set(audit.generated_paths) == set(audit.models)
    for model_key, generated_path in audit.generated_paths.items():
        assert generated_path.is_file(), model_key
        assert generated_path.is_relative_to(tmp_path)
    assert before == {path: _digest(path) for path in repo_artifacts}


def test_changed_sql_value_reports_exact_json_path(
    imported_db, real_workbook, tmp_path
):
    imported_db.execute(
        "UPDATE stingray_options SET option_name=? WHERE option_id=?",
        ("Task 7 deliberate contract difference", "opt_gba_001"),
    )
    imported_db.commit()

    audit = contract_audit.audit_runtime_contracts(
        imported_db, real_workbook, tmp_path
    )

    stingray = tuple(
        difference
        for difference in audit.differences
        if difference.model_key == "stingray"
    )
    assert stingray
    assert any(
        difference.json_path.endswith(".label")
        and difference.candidate_value == "Task 7 deliberate contract difference"
        for difference in stingray
    )
    assert all(
        not any(key in difference.json_path for key in TIMESTAMP_KEYS)
        for difference in audit.differences
    )


def test_recursive_diff_reports_missing_and_extra_paths():
    differences = contract_audit.diff_contracts(
        "stingray",
        {"choices": [{"id": "a", "label": "A"}]},
        {"choices": [{"id": "a", "label": "B"}], "extra": True},
    )

    assert tuple(difference.json_path for difference in differences) == (
        "$.choices[0].label",
        "$.extra",
    )
    assert differences[0].baseline_value == "A"
    assert differences[0].candidate_value == "B"
    assert differences[1].baseline_value is contract_audit.MISSING
    assert differences[1].candidate_value is True


def test_unreconstructable_sql_field_fails_closed(
    imported_db, real_workbook, tmp_path
):
    imported_db.execute(
        "UPDATE model_table_registry SET source_filter=? "
        "WHERE model_key='stingray' AND table_role='options'",
        ("invented_filter",),
    )
    imported_db.commit()

    with pytest.raises(export_adapter.ReverseMappingError):
        export_adapter.export_comparison_workbook(
            imported_db, real_workbook, tmp_path / "must-fail.xlsx"
        )


def test_stingray_generation_honors_comparison_config_paths(
    tmp_path, real_workbook, monkeypatch
):
    comparison = tmp_path / real_workbook.name
    comparison.write_bytes(real_workbook.read_bytes())
    workbook = load_workbook(comparison)
    options = workbook["stingray_options"]
    headers = {cell.value: cell.column for cell in options[1]}
    changed_label = "Task 7 isolated comparison label"
    for row in range(2, options.max_row + 1):
        if options.cell(row, headers["option_id"]).value == "opt_gba_001":
            options.cell(row, headers["option_name"], changed_label)
            break
    else:
        raise AssertionError("fixture option opt_gba_001 not found")
    workbook.save(comparison)
    workbook.close()

    # Keep the RED safe: old module-global writers are redirected away from
    # the repository while the assertion proves they ignored the config.
    monkeypatch.setattr(production, "OUTPUT_DIR", tmp_path / "forbidden-global-output")
    monkeypatch.setattr(production, "APP_DIR", tmp_path / "forbidden-global-app")
    config_root = tmp_path / "configured-root"
    config = discover_generation_model_configs(comparison)["stingray"].with_overrides(
        root=config_root,
        workbook_path=comparison,
        output_dir=config_root / "output",
        app_dir=config_root / "app",
    )

    result = generate_model_artifacts(config)
    runtime_path = Path(result["runtime_contract_json"])
    compatibility_path = Path(result["compatibility_artifacts"]["json"])
    payload = json.loads(runtime_path.read_text(encoding="utf-8"))

    assert compatibility_path.is_relative_to(config.output_dir)
    assert any(choice.get("label") == changed_label for choice in payload["choices"])


def test_failed_contract_audit_does_not_promote_or_write_repo_artifacts(
    tmp_path, real_workbook, repo_root, monkeypatch
):
    destination = tmp_path / "must-not-promote.sqlite3"
    artifact_paths = tuple(
        path
        for root in (repo_root / "form-output", repo_root / "form-app")
        for path in root.rglob("*")
        if path.is_file()
    )
    before = {path: _digest(path) for path in artifact_paths}

    def reject_contracts(conn, source_workbook, temp_dir):
        return ContractAudit(
            models=("stingray", "grand_sport", "z06"),
            differences=(
                ContractDifference(
                    model_key="stingray",
                    json_path="$.choices[0].label",
                    baseline_value="baseline",
                    candidate_value="candidate",
                ),
            ),
            generated_paths={},
        )

    monkeypatch.setattr(contract_audit, "audit_runtime_contracts", reject_contracts)

    report = importer.import_workbook(destination, real_workbook)

    assert report.status == "contract_mismatch"
    assert report.promoted_path is None
    assert not destination.exists()
    assert before == {path: _digest(path) for path in artifact_paths}
    assert not list(repo_root.glob(".*.candidate-*.sqlite3*"))
