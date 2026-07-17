import hashlib
import inspect
import json
import os
import shutil
import sqlite3
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app import db, importer, migration
from app.catalog import LIVE_MODELS, MODEL_TABLE_ROLES
from app.compile_types import freeze_mapping
from app.contract_audit import ContractAudit


def _digest(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _task6_import(destination, workbook):
    return importer.import_workbook(destination, workbook)


def test_public_import_is_path_only_and_has_no_audit_bypass(
    tmp_path, real_workbook
):
    signature = inspect.signature(importer.import_workbook)
    assert "audit_contracts" not in signature.parameters
    destination = tmp_path / "public.sqlite3"

    assert not hasattr(importer, "_import_workbook_for_task6_tests")
    assert not hasattr(importer, "_TASK6_TEST_CAPABILITY")
    conn = sqlite3.connect(":memory:")
    try:
        with pytest.raises(TypeError):
            importer.import_workbook(conn, real_workbook)
    finally:
        conn.close()
    assert not hasattr(importer, "Importer")
    with pytest.raises(PermissionError):
        migration._promote_audited_candidate(
            tmp_path / "candidate.sqlite3",
            destination,
            migration.capture_destination_snapshot(destination),
            audit=object(),
        )


@pytest.mark.parametrize(
    "public_promote",
    (migration.promote_candidate, importer.promote_candidate),
)
def test_public_low_level_promotion_requires_audited_importer(
    tmp_path, real_workbook, public_promote
):
    candidate = tmp_path / f"{public_promote.__module__}.candidate.sqlite3"
    destination = tmp_path / f"{public_promote.__module__}.destination.sqlite3"
    importer.load_candidate(importer.compile_workbook(real_workbook), candidate)
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "destination_marker", "must-survive")
    conn.commit()
    conn.close()
    snapshot = migration.capture_destination_snapshot(destination)
    before = _digest(destination)

    with pytest.raises(PermissionError, match="contract_audit_required"):
        public_promote(candidate, destination, snapshot)

    assert _digest(destination) == before
    assert candidate.exists()


def test_constructed_empty_audit_cannot_replace_destination(
    tmp_path, real_workbook
):
    candidate = tmp_path / "fake-audit-candidate.sqlite3"
    destination = tmp_path / "fake-audit-destination.sqlite3"
    importer.load_candidate(importer.compile_workbook(real_workbook), candidate)
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "destination_marker", "must-survive")
    conn.commit()
    conn.close()
    before = _digest(destination)
    snapshot = migration.capture_destination_snapshot(destination)
    fake = ContractAudit(
        models=LIVE_MODELS, differences=(), generated_paths={}
    )

    with pytest.raises(PermissionError, match="[Cc]ompleted audit"):
        migration._promote_audited_candidate(
            candidate, destination, snapshot, audit=fake
        )

    assert _digest(destination) == before
    assert candidate.exists()


def test_post_audit_wal_mutation_cannot_replace_destination(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "wal-audit-destination.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "destination_marker", "must-survive")
    conn.commit()
    conn.close()
    before = _digest(destination)
    writers = []
    real_audit = importer.contract_audit.audit_runtime_contracts

    def audit_then_write_wal(conn, source_workbook, temp_dir):
        audit = real_audit(conn, source_workbook, temp_dir)
        candidate = Path(
            next(
                row[2]
                for row in conn.execute("PRAGMA database_list")
                if row[1] == "main"
            )
        )
        writer = sqlite3.connect(candidate)
        writer.execute("PRAGMA journal_mode=WAL")
        writer.execute(
            "UPDATE stingray_options SET option_name=? WHERE option_id=?",
            ("unaudited WAL mutation", "opt_gba_001"),
        )
        writer.commit()
        writers.append(writer)
        assert Path(str(candidate) + "-wal").exists()
        return audit

    monkeypatch.setattr(
        importer.contract_audit, "audit_runtime_contracts", audit_then_write_wal
    )
    try:
        report = importer.import_workbook(destination, real_workbook)
    finally:
        for writer in writers:
            writer.close()

    assert report.status == "contract_mismatch"
    assert report.finding_codes == ("candidate_changed_after_contract_audit",)
    assert _digest(destination) == before
    assert not list(tmp_path.glob(".*.candidate-*.sqlite3*"))
    assert not list(tmp_path.glob("*.backup-*"))


def test_candidate_wal_after_backup_is_rechecked_before_replace(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "post-backup-candidate.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "destination_marker", "must-survive")
    conn.commit()
    conn.close()
    before = _digest(destination)
    writers = []
    real_backup = migration._backup_destination

    def backup_then_mutate_candidate(path, snapshot):
        backup = real_backup(path, snapshot)
        candidate = next(tmp_path.glob(".*.candidate-*.sqlite3"))
        writer = sqlite3.connect(candidate)
        writer.execute("PRAGMA journal_mode=WAL")
        writer.execute(
            "UPDATE stingray_options SET option_name=? WHERE option_id=?",
            ("changed after backup", "opt_gba_001"),
        )
        writer.commit()
        writers.append(writer)
        return backup

    monkeypatch.setattr(migration, "_backup_destination", backup_then_mutate_candidate)
    try:
        report = importer.import_workbook(destination, real_workbook)
    finally:
        for writer in writers:
            writer.close()

    assert report.status == "contract_mismatch"
    assert report.finding_codes == ("candidate_changed_after_contract_audit",)
    assert _digest(destination) == before
    assert not list(tmp_path.glob(".*.candidate-*.sqlite3*"))


def test_api_import_uses_path_only_audited_import_without_opening_live_connection(
    tmp_path, real_workbook, monkeypatch
):
    from app import main

    destination = tmp_path / "api.sqlite3"
    monkeypatch.setattr(main.config, "DEFAULT_DB", destination)
    monkeypatch.setattr(main.config, "DEFAULT_WORKBOOK", real_workbook)
    monkeypatch.setattr(
        main,
        "get_conn",
        lambda: (_ for _ in ()).throw(AssertionError("must stay unopened")),
    )

    calls = []

    def audited_import(db_path, workbook_path):
        calls.append((db_path, workbook_path))
        return importer.ImportReport(
            status="validated", live_models=("stingray",), findings=(),
            decision_required=(), contract_differences=(),
            candidate_path=None, promoted_path=db_path,
        )

    monkeypatch.setattr(importer, "import_workbook", audited_import)

    response = main.run_import()
    assert calls == [(destination, real_workbook)]
    assert response["status"] == "validated"
    assert response["run"]["status"] == "imported"


def test_canonical_orchestrator_cannot_skip_contract_audit(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "canonical-audit-boundary.sqlite3"

    def reject_contracts(*_args, **_kwargs):
        from app.contract_audit import ContractAudit, ContractDifference

        return ContractAudit(
            models=("stingray", "grand_sport", "z06"),
            differences=(
                ContractDifference(
                    "stingray", "$.choices[0].label", "old", "new"
                ),
            ),
            generated_paths={},
        )

    monkeypatch.setattr(
        importer.contract_audit, "audit_runtime_contracts", reject_contracts
    )
    monkeypatch.setattr(
        migration,
        "_promote_audited_candidate",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("contract mismatch must not promote")
        ),
    )

    report = importer._canonical_import_workbook(destination, real_workbook)

    assert report.status == "contract_mismatch"
    assert report.finding_codes == ("runtime_contract_difference",)
    assert report.promoted_path is None
    assert not destination.exists()


def test_workbook_change_during_contract_audit_blocks_promotion(
    tmp_path, real_workbook, monkeypatch
):
    workbook_path = tmp_path / "drifting-source" / real_workbook.name
    workbook_path.parent.mkdir()
    workbook_path.write_bytes(real_workbook.read_bytes())
    shutil.copytree(
        real_workbook.parent / "form-output" / "runtime",
        workbook_path.parent / "form-output" / "runtime",
    )
    destination = tmp_path / "must-not-promote-drift.sqlite3"
    real_audit = importer.contract_audit.audit_runtime_contracts

    def audit_then_change(conn, source_workbook, temp_dir):
        audit = real_audit(conn, source_workbook, temp_dir)
        workbook = load_workbook(workbook_path)
        sheet = workbook["zr1_options"]
        sheet.cell(2, sheet.max_column, "changed during contract audit")
        workbook.save(workbook_path)
        workbook.close()
        return audit

    monkeypatch.setattr(
        importer.contract_audit, "audit_runtime_contracts", audit_then_change
    )

    report = importer.import_workbook(destination, workbook_path)

    assert report.status == "decision_required"
    assert report.finding_codes == ("workbook_changed_during_contract_audit",)
    assert report.promoted_path is None
    assert not destination.exists()


def test_failed_candidate_does_not_replace_verified_database(
    tmp_path, broken_fk_workbook
):
    destination = tmp_path / "workbook.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "verification_marker", "keep")
    conn.commit()
    conn.close()
    before = _digest(destination)

    report = _task6_import(destination, broken_fk_workbook)

    assert report.status == "decision_required"
    assert _digest(destination) == before
    assert not list(tmp_path.glob("*.candidate*"))


def test_successful_candidate_has_complete_lineage(tmp_path, real_workbook):
    destination = tmp_path / "workbook.sqlite3"

    report = _task6_import(destination, real_workbook)

    assert report.status == "validated"
    assert report.live_models == LIVE_MODELS
    assert report.promoted_path == destination
    conn = db.connect(destination)
    try:
        assert conn.execute("PRAGMA foreign_key_check").fetchall() == []
        assert conn.execute("SELECT COUNT(*) FROM import_lineage").fetchone()[0] == 8922
        assert conn.execute("SELECT COUNT(*) FROM source_table_catalog").fetchone()[0] == 65
        assert conn.execute("SELECT COUNT(*) FROM import_issues").fetchone()[0] == 444
        assert conn.execute("SELECT COUNT(*) FROM model_table_registry").fetchone()[0] == 51
        assert conn.execute("SELECT COUNT(*) FROM schema_mapping").fetchone()[0] == 646
        assert conn.execute(
            "SELECT COUNT(DISTINCT sql_table) FROM schema_mapping"
        ).fetchone()[0] == 68
        run = conn.execute("SELECT * FROM import_runs").fetchone()
        assert run["workbook_path"] == str(real_workbook)
        assert run["workbook_sha256"] == _digest(real_workbook)
        assert run["status"] == "validated"
        assert conn.execute(
            "SELECT value FROM meta WHERE key='trusted_workbook_sha256'"
        ).fetchone()[0] == _digest(real_workbook)
        assert conn.execute(
            "SELECT COUNT(*) FROM source_row_disposition"
        ).fetchone()[0] == 11180
        assert conn.execute(
            "SELECT COUNT(*) FROM source_row_disposition "
            "WHERE disposition='emitted'"
        ).fetchone()[0] == 8424
        assert conn.execute(
            "SELECT COUNT(*) FROM source_row_disposition WHERE "
            "disposition IN ('emission_required', 'decision_required', "
            "'contract_mismatch') OR reason=''"
        ).fetchone()[0] == 0
        assert {
            row["disposition"]: row["count"]
            for row in conn.execute(
                "SELECT disposition, COUNT(*) AS count "
                "FROM source_row_disposition GROUP BY disposition"
            )
        } == {
            "emitted": 8424,
            "inactive_future_source": 2225,
            "mapped": 444,
            "metadata": 59,
            "inactive": 28,
        }
        assert conn.execute(
            "SELECT COUNT(*) FROM source_row_disposition "
            "WHERE evidence_json='{}'"
        ).fetchone()[0] == 0
        index_sql = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='index' AND "
            "name='schema_mapping_null_safe_unique'"
        ).fetchone()[0]
        assert "COALESCE(model_key" in index_sql
    finally:
        conn.close()


def test_legacy_pending_changes_block_replacement(legacy_db_path, real_workbook):
    conn = sqlite3.connect(legacy_db_path)
    conn.execute(
        "INSERT INTO pending_changes("
        "ts, session_id, table_name, model_id, entity_key_json, op, old_json, "
        "new_json, status, validation_json, confirmed_dependencies"
        ") VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            "2026-07-16T12:00:00+00:00",
            "migration-test",
            "options",
            "stingray",
            '{"option_id":"opt_pending"}',
            "add",
            None,
            '{"option_id":"opt_pending"}',
            "staged",
            "{}",
            0,
        ),
    )
    conn.commit()
    conn.close()
    before = _digest(legacy_db_path)

    report = _task6_import(legacy_db_path, real_workbook)

    assert report.status == "decision_required"
    assert report.finding_codes == ("legacy_pending_changes",)
    assert _digest(legacy_db_path) == before


def test_legacy_unsynced_history_blocks_replacement(tmp_path, real_workbook):
    destination = tmp_path / "legacy-history.sqlite3"
    conn = sqlite3.connect(destination)
    conn.execute(
        "CREATE TABLE change_history ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, ts TEXT NOT NULL, "
        "actor TEXT NOT NULL DEFAULT '', entity_type TEXT NOT NULL, "
        "entity_id TEXT NOT NULL, model_id TEXT NOT NULL DEFAULT '', "
        "op TEXT NOT NULL, old_json TEXT, new_json TEXT, "
        "src_sheet TEXT NOT NULL DEFAULT '', src_row INTEGER, "
        "validation_result TEXT NOT NULL DEFAULT '', status TEXT NOT NULL, "
        "sync_status TEXT NOT NULL DEFAULT 'pending', "
        "sync_detail TEXT NOT NULL DEFAULT '', pending_change_id INTEGER)"
    )
    conn.execute(
        "INSERT INTO change_history(ts, entity_type, entity_id, op, status) "
        "VALUES('2026-07-16T12:00:00+00:00', 'options', 'opt_pending', "
        "'add', 'committed')"
    )
    conn.commit()
    conn.close()
    before = _digest(destination)

    report = _task6_import(destination, real_workbook)

    assert report.status == "decision_required"
    assert report.finding_codes == ("legacy_unsynced_change_history",)
    assert _digest(destination) == before


def test_reimport_preserves_unambiguous_canonical_edit_history(
    imported_db_path, real_workbook
):
    conn = db.connect(imported_db_path)
    try:
        option = dict(conn.execute(
            "SELECT * FROM stingray_options ORDER BY option_id LIMIT 1"
        ).fetchone())
        option_json = json.dumps(option, sort_keys=True)
        pending_id = conn.execute(
            "INSERT INTO pending_changes("
            "ts, session_id, model_key, table_role, sql_table, "
            "entity_key_json, op, old_json, new_json, status, "
            "validation_json, confirmed_dependencies) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-16T12:00:00+00:00", "preserve-test", "stingray",
                "options", "stingray_options",
                json.dumps({"option_id": option["option_id"]}, sort_keys=True),
                "update", option_json, option_json, "committed", "{}", 0,
            ),
        ).lastrowid
        conn.execute(
            "INSERT INTO pending_changes("
            "ts, session_id, model_key, table_role, sql_table, "
            "entity_key_json, op, old_json, new_json, status, "
            "validation_json, confirmed_dependencies) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-16T12:00:01+00:00", "preserve-test", "stingray",
                "options", "stingray_options", '{"option_id":"discarded"}',
                "add", None,
                json.dumps(
                    {**option, "option_id": "discarded"}, sort_keys=True
                ),
                "discarded", "{}", 0,
            ),
        )
        committed_id = conn.execute(
            "INSERT INTO change_history("
            "ts, actor, model_key, table_role, sql_table, entity_id, op, "
            "old_json, new_json, src_sheet, src_row, validation_result, "
            "status, sync_status, pending_change_id) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-16T12:00:02+00:00", "preserve-test", "stingray",
                "options", "stingray_options", option["option_id"], "update",
                option_json, option_json, "stingray_options", 2, "passed",
                "committed", "pending", pending_id,
            ),
        ).lastrowid
        conn.execute(
            "INSERT INTO change_history("
            "ts, actor, model_key, table_role, sql_table, entity_id, op, "
            "old_json, new_json, src_sheet, src_row, validation_result, "
            "status, sync_status, sync_detail, related_history_id) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-16T12:00:03+00:00", "workbook-sync", "stingray",
                "options", "stingray_options", option["option_id"], "update",
                option_json, option_json, "stingray_options", 2, "passed",
                "sync_succeeded", "synced", '{"status":"applied"}',
                committed_id,
            ),
        )
        conn.execute(
            "UPDATE sqlite_sequence SET seq=seq+7 "
            "WHERE name IN ('pending_changes','change_history')"
        )
        conn.commit()
        expected_pending = tuple(
            tuple(row) for row in conn.execute(
                "SELECT * FROM pending_changes ORDER BY id"
            )
        )
        expected_history = tuple(
            tuple(row) for row in conn.execute(
                "SELECT * FROM change_history ORDER BY id"
            )
        )
        expected_sequences = tuple(
            tuple(row) for row in conn.execute(
                "SELECT name, seq FROM sqlite_sequence "
                "WHERE name IN ('pending_changes','change_history') ORDER BY name"
            )
        )
    finally:
        conn.close()

    report = importer.import_workbook(imported_db_path, real_workbook)

    assert report.status == "validated"
    conn = db.connect(imported_db_path)
    try:
        assert tuple(
            tuple(row) for row in conn.execute(
                "SELECT * FROM pending_changes ORDER BY id"
            )
        ) == expected_pending
        assert tuple(
            tuple(row) for row in conn.execute(
                "SELECT * FROM change_history ORDER BY id"
            )
        ) == expected_history
        assert tuple(
            tuple(row) for row in conn.execute(
                "SELECT name, seq FROM sqlite_sequence "
                "WHERE name IN ('pending_changes','change_history') ORDER BY name"
            )
        ) == expected_sequences
        with pytest.raises(sqlite3.IntegrityError, match="append-only"):
            conn.execute("UPDATE change_history SET actor='tampered'")
    finally:
        conn.close()


def test_reimport_rejects_malformed_synced_history(
    imported_db_path, real_workbook
):
    conn = db.connect(imported_db_path)
    try:
        conn.execute(
            "INSERT INTO change_history("
            "ts, actor, model_key, table_role, sql_table, entity_id, op, "
            "validation_result, status, sync_status) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-16T12:00:00+00:00", "legacy", "stingray",
                "options", "stingray_options", "opt_malformed", "update",
                "passed", "committed", "synced",
            ),
        )
        conn.commit()
    finally:
        conn.close()
    before = _digest(imported_db_path)

    report = importer.import_workbook(imported_db_path, real_workbook)

    assert report.status == "decision_required"
    assert report.finding_codes == ("canonical_edit_state_incompatible",)
    assert _digest(imported_db_path) == before


def test_carry_forward_rejects_invalid_edit_state_domains(imported_db):
    cases = (
        ("teleport", "{}", 0),
        ("add", "{malformed", 0),
        ("add", "{}", 2),
    )
    for op, new_json, confirmed in cases:
        imported_db.execute(
            "INSERT INTO pending_changes("
            "ts, model_key, table_role, sql_table, entity_key_json, op, "
            "new_json, status, validation_json, confirmed_dependencies) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)",
            (
                "2026-07-16T12:00:00+00:00", "stingray", "options",
                "stingray_options", '{"option_id":"invalid"}', op,
                new_json, "discarded", "{}", confirmed,
            ),
        )
        with pytest.raises(migration.EditStateCarryForwardError):
            migration._edit_state_rows(imported_db)
        imported_db.execute("DELETE FROM pending_changes")
        imported_db.commit()


def test_compiler_returns_complete_canonical_contract(real_workbook):
    compiled = importer.compile_workbook(real_workbook)

    assert len(compiled.tables) == 68
    assert sum(len(table.rows) for table in compiled.tables) == 8922
    assert len(compiled.source_catalog) == 65
    assert len(compiled.lineage) == 8922
    assert len(compiled.schema_mappings) == 646
    assert sum(len(source.rows) for source in compiled.source_catalog) == 11180
    assert len([finding for finding in compiled.findings if finding.status == "mapped"]) == 444
    assert {
        (table.model_key, table.role)
        for table in compiled.tables
        if table.model_key
    } == {
        (model, role) for model in LIVE_MODELS for role in MODEL_TABLE_ROLES
    }


def test_reconciliation_rejects_a_missing_exact_source_disposition(
    tmp_path, real_workbook
):
    candidate = tmp_path / "candidate.sqlite3"
    importer.load_candidate(importer.compile_workbook(real_workbook), candidate)
    conn = db.connect(candidate)
    try:
        conn.execute(
            "DELETE FROM source_row_disposition WHERE source_sheet=? "
            "AND source_row=?",
            ("model_workbook_sources", 2),
        )
        conn.commit()
        codes = {
            error["code"] for error in migration.reconcile_source_counts(conn)
        }
    finally:
        conn.close()
    assert "source_row_reconciliation_mismatch" in codes


def test_mapping_coverage_is_independent_of_mapping_row_count(
    tmp_path, real_workbook
):
    candidate = tmp_path / "candidate.sqlite3"
    importer.load_candidate(importer.compile_workbook(real_workbook), candidate)
    conn = db.connect(candidate)
    try:
        conn.execute(
            "UPDATE schema_mapping SET sql_column='missing_column' "
            "WHERE id=(SELECT MIN(id) FROM schema_mapping)"
        )
        conn.commit()
        codes = {
            error["code"] for error in migration.validate_mapping_coverage(conn)
        }
    finally:
        conn.close()
    assert "schema_mapping_destination_column_missing" in codes
    assert "schema_mapping_destination_coverage_missing" in codes


def test_structurally_failed_candidate_is_removed_and_destination_is_stable(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "verified.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "verification_marker", "keep")
    conn.commit()
    conn.close()
    before = _digest(destination)

    compiled = importer.compile_workbook(real_workbook)
    tables = list(compiled.tables)
    table_index = next(
        index
        for index, table in enumerate(tables)
        if table.name == "stingray_option_availability"
    )
    table = tables[table_index]
    rows = list(table.rows)
    values = dict(rows[0].values)
    values["option_id"] = "opt_missing_candidate_fk"
    rows[0] = replace(rows[0], values=freeze_mapping(values))
    tables[table_index] = replace(table, rows=tuple(rows))
    broken = replace(compiled, tables=tuple(tables))
    monkeypatch.setattr(importer, "compile_workbook", lambda _: broken)

    report = _task6_import(destination, real_workbook)

    assert report.status == "contract_mismatch"
    assert report.finding_codes == ("candidate_build_or_promotion_failed",)
    assert _digest(destination) == before
    assert not list(tmp_path.glob(".*.candidate-*"))
    assert not list(tmp_path.glob("*.backup-*"))


def test_existing_destination_is_backed_up_before_atomic_replacement(
    tmp_path, real_workbook
):
    destination = tmp_path / "verified.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "verification_marker", "backup-proof")
    conn.commit()
    conn.close()

    report = _task6_import(destination, real_workbook)

    assert report.status == "validated"
    backups = list(tmp_path.glob("verified.sqlite3.backup-*"))
    assert len(backups) == 1
    backup = db.connect(backups[0])
    try:
        assert db.get_meta(backup, "verification_marker") == "backup-proof"
        assert backup.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
    finally:
        backup.close()


def test_destination_sidecars_block_before_candidate_build(
    tmp_path, real_workbook
):
    destination = tmp_path / "verified.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "verification_marker", "keep")
    conn.commit()
    conn.close()
    before = _digest(destination)
    wal = tmp_path / "verified.sqlite3-wal"
    wal.write_bytes(b"active-or-stale-wal-must-not-be-guessed")

    report = _task6_import(destination, real_workbook)

    assert report.status == "decision_required"
    assert report.finding_codes == ("legacy_destination_sidecars",)
    assert _digest(destination) == before
    assert wal.read_bytes() == b"active-or-stale-wal-must-not-be-guessed"
    assert not list(tmp_path.glob(".*.candidate-*"))


def test_post_replace_failure_restores_byte_identical_destination(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "verified.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "verification_marker", "keep")
    conn.commit()
    conn.close()
    before = _digest(destination)
    real_replace = os.replace

    def replace_then_fail(source, target):
        real_replace(source, target)
        if ".candidate-" in str(source):
            raise OSError("injected failure after candidate replacement")

    monkeypatch.setattr("app.migration.os.replace", replace_then_fail)

    report = _task6_import(destination, real_workbook)

    assert report.status == "contract_mismatch"
    assert _digest(destination) == before
    assert not list(tmp_path.glob(".*.candidate-*"))
    assert not list(tmp_path.glob(".*.rollback-*"))


def test_destination_change_during_compile_blocks_before_backup(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "verified.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    conn.commit()
    conn.close()
    real_compile = importer.compile_workbook

    def compile_then_change(path):
        compiled = real_compile(path)
        changed = db.connect(destination)
        db.set_meta(changed, "concurrent_writer", "present")
        changed.commit()
        changed.close()
        return compiled

    monkeypatch.setattr(importer, "compile_workbook", compile_then_change)
    report = _task6_import(destination, real_workbook)

    assert report.status == "decision_required"
    assert report.finding_codes == ("destination_changed_during_import",)
    conn = db.connect(destination)
    try:
        assert db.get_meta(conn, "concurrent_writer") == "present"
    finally:
        conn.close()
    assert not list(tmp_path.glob("*.backup-*"))


def test_destination_change_after_backup_blocks_replace(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "verified.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    conn.commit()
    conn.close()
    real_backup = migration._backup_destination

    def backup_then_change(path, snapshot):
        backup = real_backup(path, snapshot)
        changed = db.connect(path)
        db.set_meta(changed, "after_backup_writer", "present")
        changed.commit()
        changed.close()
        return backup

    monkeypatch.setattr(migration, "_backup_destination", backup_then_change)
    report = _task6_import(destination, real_workbook)

    assert report.status == "decision_required"
    assert report.finding_codes == ("destination_changed_during_import",)
    conn = db.connect(destination)
    try:
        assert db.get_meta(conn, "after_backup_writer") == "present"
    finally:
        conn.close()


def test_failed_backup_verification_removes_backup_and_preserves_destination(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "verified.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    conn.commit()
    conn.close()
    before = _digest(destination)

    def reject_backup(*_args):
        raise migration.BackupVerificationError("injected verification failure")

    monkeypatch.setattr(migration, "_verify_backup", reject_backup)
    report = _task6_import(destination, real_workbook)

    assert report.finding_codes == ("backup_verification_failed",)
    assert _digest(destination) == before
    assert not list(tmp_path.glob("*.backup-*"))


def test_busy_candidate_checkpoint_removes_disposable_candidate_set(
    tmp_path, real_workbook, monkeypatch
):
    destination = tmp_path / "destination.sqlite3"
    conn = db.connect(destination)
    db.create_canonical_schema(conn)
    db.set_meta(conn, "destination_marker", "must-survive")
    conn.commit()
    conn.close()
    before = _digest(destination)
    candidates = []

    def reject_busy_checkpoint(candidate):
        candidates.append(Path(candidate))
        Path(str(candidate) + "-wal").write_bytes(b"busy-wal")
        Path(str(candidate) + "-shm").write_bytes(b"busy-shm")
        raise migration.CandidateCheckpointError(
            "candidate_checkpoint_incomplete: (1, 3, 2)"
        )

    monkeypatch.setattr(
        migration, "_checkpoint_and_verify_candidate", reject_busy_checkpoint
    )

    report = _task6_import(destination, real_workbook)

    assert report.finding_codes == ("candidate_checkpoint_incomplete",)
    assert _digest(destination) == before
    assert len(candidates) == 1
    candidate = candidates[0]
    assert not candidate.exists()
    assert not Path(str(candidate) + "-wal").exists()
    assert not Path(str(candidate) + "-shm").exists()


def test_corrupt_destination_audit_is_structured(tmp_path, real_workbook):
    destination = tmp_path / "corrupt.sqlite3"
    destination.write_bytes(b"not-a-sqlite-database")

    report = _task6_import(destination, real_workbook)

    assert report.status == "decision_required"
    assert report.finding_codes == ("legacy_destination_audit_failed",)
