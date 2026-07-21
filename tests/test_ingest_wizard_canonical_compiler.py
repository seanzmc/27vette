#!/usr/bin/env python3
from __future__ import annotations

import copy
import hashlib
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from corvette_form_generator.editor_ops import EDITOR_SHEET_META  # noqa: E402
from corvette_form_generator.ingest.wizard.canonical_rows import canonical_bytes, semantic_hash, validate_artifact_graph  # noqa: E402
from corvette_form_generator.ingest.wizard import compiler as compiler_module  # noqa: E402
from corvette_form_generator.ingest.wizard.comparator_evidence import build_comparator_evidence  # noqa: E402
from corvette_form_generator.ingest.wizard.compiler import (  # noqa: E402
    _candidate_feature_index,
    _profile_reconciled_price,
    _reconcile_represented_comparator_facts,
    _status_feature_index,
    build_family_registry,
    compile_canonical_rows,
)
from corvette_form_generator.ingest.wizard.decisions import model_scoped_statuses  # noqa: E402
from corvette_form_generator.ingest.wizard.identity import option_occurrence_signature  # noqa: E402
from corvette_form_generator.ingest.wizard.joiner import join_prices  # noqa: E402
from corvette_form_generator.ingest.wizard.parser import parse_confirmed_sheets  # noqa: E402
from corvette_form_generator.ingest.wizard.profiler import profile_workbook  # noqa: E402
from corvette_form_generator.options_sheet_quality import (  # noqa: E402
    DEFAULT_ALLOWLIST_PATH,
    DEFAULT_ALLOWLIST_RELATIVE_PATH,
)
from ingest_wizard_fixtures import build_master_workbook, build_raw_export  # noqa: E402

ROLES = {
    "Exterior 1": "exclude",
    "Mechanical 4": "options",
    "Price Schedule": "price",
    "Standard Equipment 1": "exclude",
    "Color and Trim 1": "exclude",
}


class CanonicalCompilerTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        base = Path(self.tmp.name)
        self.raw = build_raw_export(base / "raw.xlsx")
        self.master = build_master_workbook(base / "master.xlsx")
        from openpyxl import load_workbook
        raw_wb = load_workbook(self.raw)
        raw_wb["Mechanical 4"].append(["BV4", "", "Personalized Plaque", "A", "A"])
        raw_wb.save(self.raw)
        raw_wb.close()
        master_wb = load_workbook(self.master)
        master_wb["z06_options"].append(["opt_bv4_001", "BV4", 395, "Personalized Plaque", "Fixture endpoint", "", "sec_whee_001", True, 15, True, ""])
        master_wb["zr1_options"].append(["opt_bv4_001", "BV4", 395, "Personalized Plaque", "Existing target identity", "", "sec_whee_001", False, 20, False, ""])
        master_wb.save(self.master)
        master_wb.close()
        parsed = parse_confirmed_sheets(self.raw, ROLES)
        report = join_prices(parsed["candidates"], parsed["priceRows"])
        self.option_payload = {
            "schemaVersion": "pass-a-1",
            "candidates": parsed["candidates"],
            "skippedRows": parsed["skippedRows"],
        }
        self.price_payload = {"schemaVersion": "pass-a-1", "priceRows": parsed["priceRows"], "baseModelPriceRows": parsed["baseModelPriceRows"], "skippedPriceRows": parsed["skippedPriceRows"]}
        self.join_report = report
        self.roles_payload = {"schemaVersion": "pass-a-1", "roles": ROLES}
        self.sheet_profile = profile_workbook(self.raw)
        self.selection = {"targets": ["zr1"], "comparators": {"zr1": "z06"}}
        authority_bindings = {
            "sourceSha256": "a" * 64,
            "workbookSha256": "b" * 64,
            "compilerPolicyVersion": "options-recurrence-prevention-4.4-v1",
            "optionsSheetQualityAllowlist": {
                "path": DEFAULT_ALLOWLIST_RELATIVE_PATH.as_posix(),
                "sha256": hashlib.sha256(DEFAULT_ALLOWLIST_PATH.read_bytes()).hexdigest(),
            },
        }
        self.authority = {
            "fingerprint": hashlib.sha256(canonical_bytes(authority_bindings)).hexdigest(),
            "bindings": authority_bindings,
        }
        self.comparator = build_comparator_evidence(self.master, self.selection["comparators"], run_authority_fingerprint=self.authority)

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def compile(self, **overrides):
        args = {
            "workbook_path": self.master,
            "option_payload": self.option_payload,
            "price_payload": self.price_payload,
            "join_report": self.join_report,
            "roles_payload": self.roles_payload,
            "sheet_profile": self.sheet_profile,
            "selection": self.selection,
            "comparator_artifact": self.comparator,
            "run_authority_fingerprint": self.authority,
            "resolution_entries": [],
        }
        args.update(overrides)
        return compile_canonical_rows(**args)

    def compile_with_ready_comparator_relationship(self, **overrides):
        preliminary = self.compile()
        subject = next(
            item
            for item in preliminary["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "comparator_only_relationship_proposal"
        )
        proposal = subject["proposedRows"][0]
        required_rpos = {proposal["sourceRpo"], proposal["targetRpo"]}
        option_payload = copy.deepcopy(self.option_payload)
        price_payload = copy.deepcopy(self.price_payload)
        for candidate in option_payload["candidates"]:
            if candidate.get("rpo") in required_rpos:
                candidate["sectionLabel"] = "Wheels"
                candidate["priceMatch"] = None
                candidate["listPrice"] = None
                candidate["priceRows"] = []
        price_payload["priceRows"] = [
            row
            for row in price_payload["priceRows"]
            if str(row.get("rpo") or "").upper() not in required_rpos
        ]
        join_report = join_prices(option_payload["candidates"], price_payload["priceRows"])
        result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            **overrides,
        )
        current_subject = next(
            item
            for item in result["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "comparator_only_relationship_proposal"
            and item["proposedRows"] == subject["proposedRows"]
        )
        return option_payload, price_payload, join_report, result, current_subject

    def test_family_registry_uses_roles_and_exact_live_headers(self) -> None:
        registry = build_family_registry(self.master, ["zr1"])
        self.assertEqual(registry["zr1"]["source_option_sheet"]["sheetName"], "zr1_options")
        self.assertEqual(registry["zr1"]["source_option_sheet"]["headers"][0], "option_id")
        self.assertIn("rule_group_members_sheet", registry["zr1"])

    def test_family_registry_isolates_greenfield_model_rule_sheets(self) -> None:
        registry = build_family_registry(self.master, ["future_x"])["future_x"]

        self.assertEqual(registry["price_rules_sheet"]["sheetName"], "future_x_price_rules")
        self.assertEqual(registry["rule_mapping_sheet"]["sheetName"], "future_x_rule_mapping")
        self.assertEqual(registry["rule_groups_sheet"]["sheetName"], "future_x_rule_groups")
        self.assertEqual(
            registry["rule_group_members_sheet"]["sheetName"],
            "future_x_rule_members",
        )
        self.assertEqual(
            registry["exclusive_groups_sheet"]["sheetName"],
            "future_x_exclusive_groups",
        )
        self.assertEqual(
            registry["exclusive_group_members_sheet"]["sheetName"],
            "future_x_exclusive_members",
        )
        self.assertEqual(
            registry["color_overrides_sheet"]["sheetName"],
            "future_x_color_overrides",
        )
        self.assertEqual(registry["interior_source_sheet"]["sheetName"], "lt_interiors")
        self.assertTrue(
            all(len(entry["sheetName"]) <= 31 for entry in registry.values())
        )

    def test_unchanged_detail_raw_preserves_valid_curated_target_copy_without_review(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        sheet = workbook["zr1_options"]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == "opt_bv4_001":
                row[3].value = "Curated Personalized Plaque"
                row[4].value = "Customer-provided personalization text"
                row[5].value = "Personalized Plaque"
                break
        workbook.save(self.master)
        workbook.close()

        result = self.compile()
        option = next(
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["family"] == "options" and row["values"].get("option_id") == "opt_bv4_001"
        )

        self.assertEqual(option["values"]["option_name"], "Curated Personalized Plaque")
        self.assertEqual(option["values"]["description"], "Customer-provided personalization text")
        self.assertEqual(option["values"]["detail_raw"], "Personalized Plaque")
        self.assertFalse(
            any(
                subject["reasonCode"] in {"copy_review_required", "comparator_copy_conflict"}
                and "opt_bv4_001" in str(subject.get("proposedRows") or [])
                for subject in result["exception-queue.json"]["subjects"]
            )
        )

    def test_interior_profile_metadata_does_not_replace_target_option_evidence(self) -> None:
        option_payload = copy.deepcopy(self.option_payload)
        component = copy.deepcopy(
            next(candidate for candidate in option_payload["candidates"] if candidate.get("rpo") == "PDB")
        )
        component.update(
            {
                "candidateId": "Mechanical 4:999",
                "rpo": "N2Z",
                "rpoNormalized": "N2Z",
                "rpoDisplay": "N2Z",
                "description": "Suede-wrapped interior trim",
                "detailRaw": "Suede-wrapped interior trim",
                "listPrice": 9999,
                "priceMatch": "exact",
                "priceRows": [
                    {
                        "rpo": "N2Z",
                        "qualifier": "ZR1 only",
                        "listPrice": 9999,
                        "priceColumnEvidence": [
                            {"headerText": "List", "value": 9999}
                        ],
                    }
                ],
            }
        )
        component["sourceEvidence"] = {
            **component["sourceEvidence"],
            "rowIndex": 999,
        }
        option_payload["candidates"].append(component)
        component_feature_id = _candidate_feature_index(option_payload["candidates"])[component["candidateId"]]
        status_feature_ids = {
            feature_id
            for feature_ids in _status_feature_index(option_payload["candidates"])[component["candidateId"]].values()
            for feature_id in feature_ids
        }
        price_payload = copy.deepcopy(self.price_payload)
        price_payload["priceRows"].extend(copy.deepcopy(component["priceRows"]))
        join_report = join_prices(option_payload["candidates"], price_payload["priceRows"])
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["z06_options"].append(
            [
                "opt_n2z_001", "N2Z", 895, "Suede-wrapped interior trim", "", "",
                "sec_whee_001", True, 18, True, "",
            ]
        )
        workbook.save(self.master)
        workbook.close()

        result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )
        manifest = result["canonical-row-manifest.json"]
        queue = result["exception-queue.json"]
        report = result["compile-report.json"]

        option = next(
            (
                row
                for row in manifest["rows"]
                if row["model"] == "zr1"
                and row["family"] == "options"
                and row["values"].get("rpo") == "N2Z"
            ),
            None,
        )
        self.assertIsNotNone(
            option,
            {
                "options": [
                    row["values"].get("rpo")
                    for row in manifest["rows"]
                    if row["model"] == "zr1" and row["family"] == "options"
                ],
                "subjects": queue["subjects"],
            },
        )
        assert option is not None
        self.assertEqual(option["values"]["price"], 9999)
        self.assertEqual(option["status"], "ready")
        self.assertFalse(
            any("N2Z" in str(subject.get("proposedRows") or subject.get("question")) for subject in queue["subjects"])
        )
        component_features = [
            feature
            for feature in report["sourceFeatureCoverage"]
            if component_feature_id in (feature.get("evidenceIds") or [])
        ]
        self.assertTrue(component_features)
        self.assertEqual(
            {feature["disposition"] for feature in component_features},
            {"compiled"},
        )
        status_dispositions = {
            feature["disposition"]
            for feature in report["sourceFeatureCoverage"]
            if feature["featureId"] in status_feature_ids
        }
        self.assertIn("compiled", status_dispositions)
        price_feature = next(
            feature
            for feature in report["sourceFeatureCoverage"]
            if feature["featureId"]
            == f"price:zr1:N2Z:{semantic_hash(component['priceRows'][0])}"
        )
        self.assertEqual(price_feature["disposition"], "compiled")

    def test_existing_single_exclusive_group_represents_comparator_single_within_group(self) -> None:
        rows = [
            {
                "model": "zr1", "family": "options", "status": "ready",
                "values": {"option_id": "opt_a", "rpo": "AAA"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "options", "status": "ready",
                "values": {"option_id": "opt_b", "rpo": "BBB"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "exclusive_groups", "status": "ready",
                "values": {"group_id": "group_ab", "selection_mode": "single"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "exclusive_members", "status": "ready",
                "values": {"group_id": "group_ab", "option_id": "opt_a"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "exclusive_members", "status": "ready",
                "values": {"group_id": "group_ab", "option_id": "opt_b"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
        ]
        evidence_id = "comparator:exclusive_group:fixture"
        subject = {
            "model": "zr1",
            "reasonCode": "comparator_only_exclusive_group_proposal",
            "evidenceReferences": [evidence_id],
        }
        filtered, matched, _ = _reconcile_represented_comparator_facts(
            rows,
            [subject],
            {
                "targets": {
                    "zr1": {
                        "facts": [
                            {
                                "factType": "exclusive_group",
                                "disposition": "corroborating_context_only",
                                "evidenceId": evidence_id,
                                "signature": {
                                    "selectionMode": "single_within_group",
                                    "memberRpos": ["AAA", "BBB"],
                                },
                                "context": {},
                            }
                        ]
                    }
                }
            },
            ["zr1"],
        )
        self.assertEqual(filtered, [])
        self.assertEqual(matched, {("zr1", evidence_id)})

    def test_semantic_gate_blocks_exclusive_subset_and_removes_historical_resolution_rows(self) -> None:
        result = self.compile()
        manifest_rows = copy.deepcopy(result["canonical-row-manifest.json"]["rows"])
        subjects = copy.deepcopy(result["exception-queue.json"]["subjects"])
        subject = next(
            item
            for item in subjects
            if item["reasonCode"] == "comparator_only_exclusive_group_proposal"
        )
        option_ids = {
            str(row["values"].get("rpo") or "").upper(): str(row["values"].get("option_id") or "")
            for row in manifest_rows
            if row["model"] == "zr1" and row["family"] == "options"
        }
        retained_group = "target_excl_wheels"
        manifest_rows.extend(
            [
                {
                    "model": "zr1", "family": "exclusive_groups", "status": "ready",
                    "sheet": "zr1_exclusive_groups", "action": "noop",
                    "key": {"group_id": retained_group},
                    "values": {"group_id": retained_group, "selection_mode": "single", "active": True, "notes": ""},
                    "semanticSignature": {}, "evidenceDependencies": [],
                },
                *[
                    {
                        "model": "zr1", "family": "exclusive_members", "status": "ready",
                        "sheet": "zr1_exclusive_members", "action": "noop",
                        "key": {"group_id": retained_group, "option_id": option_ids[rpo]},
                        "values": {"group_id": retained_group, "option_id": option_ids[rpo], "display_order": order, "active": True},
                        "semanticSignature": {}, "evidenceDependencies": [],
                    }
                    for order, rpo in enumerate(("PDB", "BV4", "GBA"), 1)
                ],
                {
                    "model": "zr1", "family": "exclusive_groups", "status": "ready",
                    "sheet": "zr1_exclusive_groups", "action": "add",
                    "key": {"group_id": "historical_subgroup"},
                    "values": {"group_id": "historical_subgroup", "selection_mode": "single", "active": True, "notes": ""},
                    "semanticSignature": {},
                    "evidenceDependencies": [{"evidenceId": f"resolution:{subject['subjectId']}", "evidenceVersion": "fixture"}],
                },
            ]
        )

        filtered_rows, gated_subjects, conflict_subject_ids, _ = (
            compiler_module._apply_comparator_semantic_gate(
                manifest_rows,
                subjects,
                self.comparator,
                self.selection["targets"],
            )
        )

        conflict = next(
            item
            for item in gated_subjects
            if item.get("originalReasonCode") == "comparator_only_exclusive_group_proposal"
        )
        self.assertEqual(conflict["reasonCode"], "semantic_group_overlap")
        self.assertEqual(conflict["semanticConflict"]["overlapKind"], "proposed_subset")
        self.assertEqual(conflict["semanticConflict"]["affectedSheets"], ["zr1_exclusive_groups", "zr1_exclusive_members"])
        self.assertEqual(conflict["allowedActions"], ["mark_not_applicable"])
        self.assertIn(subject["subjectId"], conflict_subject_ids)
        self.assertFalse(
            any(
                dependency["evidenceId"] == f"resolution:{subject['subjectId']}"
                for row in filtered_rows
                for dependency in row.get("evidenceDependencies") or []
            )
        )

    def test_semantic_gate_blocks_same_direction_different_relationship_type(self) -> None:
        _, _, _, result, subject = self.compile_with_ready_comparator_relationship()
        manifest_rows = copy.deepcopy(result["canonical-row-manifest.json"]["rows"])
        option_ids = {
            str(row["values"].get("rpo") or "").upper(): str(row["values"].get("option_id") or "")
            for row in manifest_rows
            if row["model"] == "zr1" and row["family"] == "options"
        }
        proposal = subject["proposedRows"][0]
        manifest_rows.append(
            {
                "model": "zr1", "family": "rule_mapping", "status": "ready",
                "sheet": "zr1_rule_mapping", "action": "noop",
                "key": {"rule_id": "target_conflicting_rule"},
                "values": {
                    "rule_id": "target_conflicting_rule",
                    "source_id": option_ids[proposal["sourceRpo"]],
                    "rule_type": "includes",
                    "target_id": option_ids[proposal["targetRpo"]],
                    "original_detail_raw": "", "body_style_scope": "",
                    "runtime_action": "auto_add", "disabled_reason": "",
                },
                "semanticSignature": {}, "evidenceDependencies": [],
            }
        )

        _, gated_subjects, conflict_subject_ids, _ = compiler_module._apply_comparator_semantic_gate(
            manifest_rows,
            copy.deepcopy(result["exception-queue.json"]["subjects"]),
            self.comparator,
            self.selection["targets"],
        )

        conflict = next(
            item
            for item in gated_subjects
            if item.get("originalReasonCode") == "comparator_only_relationship_proposal"
        )
        self.assertEqual(conflict["reasonCode"], "semantic_relationship_conflict")
        self.assertEqual(conflict["semanticConflict"]["overlapKind"], "same_direction_different_type")
        self.assertEqual(conflict["allowedActions"], ["mark_not_applicable"])
        self.assertIn(subject["subjectId"], conflict_subject_ids)

    def test_semantic_gate_blocks_rule_group_member_mismatch(self) -> None:
        result = self.compile()
        manifest_rows = copy.deepcopy(result["canonical-row-manifest.json"]["rows"])
        subjects = copy.deepcopy(result["exception-queue.json"]["subjects"])
        subject = next(
            item
            for item in subjects
            if item["reasonCode"] == "comparator_only_rule_group_proposal"
        )
        option_ids = {
            str(row["values"].get("rpo") or "").upper(): str(row["values"].get("option_id") or "")
            for row in manifest_rows
            if row["model"] == "zr1" and row["family"] == "options"
        }
        group_id = "target_requires_any"
        manifest_rows.extend(
            [
                {
                    "model": "zr1", "family": "rule_groups", "status": "ready",
                    "sheet": "zr1_rule_groups", "action": "noop",
                    "key": {"group_id": group_id},
                    "values": {
                        "group_id": group_id, "group_type": "requires_any",
                        "source_id": option_ids["PDB"], "body_style_scope": "",
                        "trim_level_scope": "", "variant_scope": "",
                        "disabled_reason": "", "active": True, "notes": "",
                    },
                    "semanticSignature": {}, "evidenceDependencies": [],
                },
                *[
                    {
                        "model": "zr1", "family": "rule_group_members", "status": "ready",
                        "sheet": "zr1_rule_group_members", "action": "noop",
                        "key": {"group_id": group_id, "target_id": option_ids[rpo]},
                        "values": {"group_id": group_id, "target_id": option_ids[rpo], "display_order": order, "active": True},
                        "semanticSignature": {}, "evidenceDependencies": [],
                    }
                    for order, rpo in enumerate(("BV4", "GBA"), 1)
                ],
            ]
        )

        _, gated_subjects, conflict_subject_ids, _ = compiler_module._apply_comparator_semantic_gate(
            manifest_rows,
            subjects,
            self.comparator,
            self.selection["targets"],
        )

        conflict = next(
            item
            for item in gated_subjects
            if item.get("originalReasonCode") == "comparator_only_rule_group_proposal"
        )
        self.assertEqual(conflict["reasonCode"], "semantic_group_overlap")
        self.assertEqual(conflict["semanticConflict"]["overlapKind"], "member_set_mismatch")
        self.assertEqual(conflict["allowedActions"], ["mark_not_applicable"])
        self.assertIn(subject["subjectId"], conflict_subject_ids)

    def test_requirement_for_target_standard_on_every_variant_is_not_a_rule_question(self) -> None:
        rows = [
            {
                "model": "zr1", "family": "options", "status": "ready",
                "values": {"option_id": "opt_src", "rpo": "SRC"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "options", "status": "ready",
                "values": {"option_id": "opt_std", "rpo": "STD"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "model_variants", "status": "ready",
                "values": {"model_key": "zr1", "variant_id": "v1"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "ovs", "status": "ready",
                "values": {
                    "option_id": "opt_std", "variant_id": "v1", "status": "standard",
                },
                "semanticSignature": {}, "evidenceDependencies": [],
            },
        ]
        evidence_id = "comparator:direct_rule:standard-target"
        subject = {
            "model": "zr1",
            "reasonCode": "comparator_only_relationship_proposal",
            "evidenceReferences": [evidence_id],
        }
        filtered, matched, _ = _reconcile_represented_comparator_facts(
            rows,
            [subject],
            {
                "targets": {
                    "zr1": {
                        "facts": [
                            {
                                "factType": "direct_rule",
                                "disposition": "corroborating_context_only",
                                "evidenceId": evidence_id,
                                "signature": {
                                    "sourceRpo": "SRC", "ruleType": "requires",
                                    "targetRpo": "STD", "bodyStyleScope": "*",
                                    "trimLevelScope": "*", "variantScope": "*",
                                },
                                "context": {},
                            }
                        ]
                    }
                }
            },
            ["zr1"],
        )
        self.assertEqual(filtered, [])
        self.assertEqual(matched, {("zr1", evidence_id)})

    def test_broader_existing_price_scope_represents_narrower_comparator_fact(self) -> None:
        rows = [
            {
                "model": "zr1", "family": "options", "status": "ready",
                "values": {"option_id": "opt_cnd", "rpo": "CND"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "options", "status": "ready",
                "values": {"option_id": "opt_tgt", "rpo": "TGT"},
                "semanticSignature": {}, "evidenceDependencies": [],
            },
            {
                "model": "zr1", "family": "price_rules", "status": "ready",
                "values": {
                    "condition_option_id": "opt_cnd", "price_rule_type": "override",
                    "target_option_id": "opt_tgt", "price_value": 0,
                    "body_style_scope": "", "trim_level_scope": "",
                    "variant_scope": "",
                },
                "semanticSignature": {}, "evidenceDependencies": [],
            },
        ]
        evidence_id = "comparator:price_rule:covered-scope"
        subject = {
            "model": "zr1",
            "reasonCode": "comparator_only_price_rule_proposal",
            "evidenceReferences": [evidence_id],
        }
        filtered, matched, _ = _reconcile_represented_comparator_facts(
            rows,
            [subject],
            {
                "targets": {
                    "zr1": {
                        "facts": [
                            {
                                "factType": "price_rule",
                                "disposition": "corroborating_context_only",
                                "evidenceId": evidence_id,
                                "signature": {
                                    "conditionRpo": "CND", "priceRuleType": "override",
                                    "targetRpo": "TGT", "bodyStyleScope": "coupe",
                                    "trimLevelScope": "*", "variantScope": "*",
                                },
                                "context": {"priceValue": 0},
                            }
                        ]
                    }
                }
            },
            ["zr1"],
        )
        self.assertEqual(filtered, [])
        self.assertEqual(matched, {("zr1", evidence_id)})

    def test_selected_comparator_compiles_shared_color_interior_and_presentation_profiles(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook.save(self.master)
        workbook.close()

        result = self.compile()
        rows = result["canonical-row-manifest.json"]["rows"]
        subjects = result["exception-queue.json"]["subjects"]

        paint = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "GBA"
        )
        self.assertEqual(paint["values"]["option_id"], "opt_gba_001")
        self.assertEqual(
            {
                row["values"]["variant_id"]
                for row in rows
                if row["model"] == "zr1"
                and row["family"] == "ovs"
                and row["values"].get("option_id") == "opt_gba_001"
            },
            {"1lz_r07", "3lz_r67"},
        )
        self.assertEqual(
            {
                (row["values"]["interior_id"], row["values"]["trim_level"])
                for row in rows
                if row["model"] == "zr1" and row["family"] == "model_interior_scope"
            },
            {("1LZ_AQ9_HTA", "1LZ"), ("3LZ_AQ9_HTA", "3LZ")},
        )
        self.assertEqual(
            {
                row["values"]["interior_id"]
                for row in rows
                if row["model"] == "zr1" and row["family"] == "interior_components"
            },
            {"1LZ_AQ9_HTA", "3LZ_AQ9_HTA"},
        )
        color_override = next(
            row
            for row in rows
            if row["family"] == "color_overrides"
            and row["values"].get("interior_id") == "3LZ_AQ9_HTA"
        )
        self.assertEqual(color_override["values"]["option_id"], "opt_gba_001")
        self.assertEqual(color_override["values"]["rule_type"], "requires")
        for family in (
            "runtime_steps_meta",
            "section_presentation_meta",
            "context_section_master_meta",
            "order_summary_sections_meta",
            "step_order_summary_map_meta",
        ):
            self.assertTrue(
                any(row["model"] == "zr1" and row["family"] == family for row in rows),
                family,
            )
        self.assertFalse(
            any(subject["reasonCode"] == "unsupported_color_trim_source" for subject in subjects)
        )
        compiled_global_families = {
            "model_master",
            "model_workbook_sources",
            "model_interior_scope",
            "interior_components",
            "runtime_steps_meta",
            "section_presentation_meta",
            "context_section_master_meta",
            "order_summary_sections_meta",
            "step_order_summary_map_meta",
        }
        self.assertFalse(
            any(
                subject["reasonCode"] == "unsupported_global_family"
                and subject["model"] == "zr1"
                and subject["family"] in compiled_global_families
                for subject in subjects
            )
        )

    def test_target_status_comparator_placement_and_existing_default_are_not_false_blockers(self) -> None:
        from openpyxl import load_workbook

        raw = load_workbook(self.raw)
        raw["Mechanical 4"].append(
            ["SEC", "", "Comparator-placed target option", "A", "A"]
        )
        raw["Mechanical 4"].append(
            ["CON", "", "Standard condition", "S", "S"]
        )
        raw["Mechanical 4"].append(
            ["DFT", "", "Standard dependent default", "S", "S"]
        )
        raw.save(self.raw)
        raw.close()

        workbook = load_workbook(self.master)
        workbook["z06_options"].append(
            [
                "opt_sec_001", "SEC", 0, "Comparator-placed option",
                "Fixture canonical placement", "", "sec_whee_001",
                True, 30, True, "",
            ]
        )
        workbook["z06_options"].append(
            [
                "opt_con_001", "CON", 0, "Standard condition", "", "",
                "sec_whee_001", False, 31, True, "",
            ]
        )
        workbook["z06_options"].append(
            [
                "opt_dft_001", "DFT", 6000, "Standard dependent default", "", "",
                "sec_whee_001", True, 32, True, "",
            ]
        )
        workbook["default_selection_rules"].append(
            [
                "z06", "z06_default_bv4", "opt_bv4_001", "always", "",
                "*", "*", "*", 5, True, "Fixture comparator default", "",
            ]
        )
        workbook["default_selection_rules"].append(
            [
                "zr1", "zr1_default_bv4", "opt_bv4_001", "always", "",
                "*", "*", "*", 105, True, "Existing target default", "default_selected",
            ]
        )
        workbook["default_selection_rules"].append(
            [
                "z06", "z06_default_dft_with_con", "opt_dft_001",
                "when_selected_unless_selected_section", "opt_con_001",
                "*", "*", "*", 6, True,
                "Fixture source-supported default", "",
            ]
        )
        model_variants = workbook["model_variants"]
        for row_index in range(model_variants.max_row, 1, -1):
            if (
                model_variants.cell(row_index, 1).value == "zr1"
                and model_variants.cell(row_index, 2).value == "3lz_r67"
            ):
                model_variants.delete_rows(row_index)
        workbook.save(self.master)
        workbook.close()

        parsed = parse_confirmed_sheets(self.raw, ROLES)
        option_payload = {
            "schemaVersion": "pass-a-1",
            "candidates": parsed["candidates"],
            "skippedRows": parsed["skippedRows"],
        }
        price_payload = {
            "schemaVersion": "pass-a-1",
            "priceRows": parsed["priceRows"],
            "baseModelPriceRows": parsed["baseModelPriceRows"],
            "skippedPriceRows": parsed["skippedPriceRows"],
        }
        comparator = build_comparator_evidence(
            self.master,
            self.selection["comparators"],
            run_authority_fingerprint=self.authority,
        )
        result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_prices(parsed["candidates"], parsed["priceRows"]),
            comparator_artifact=comparator,
        )
        rows = result["canonical-row-manifest.json"]["rows"]
        subjects = result["exception-queue.json"]["subjects"]

        sec = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "SEC"
        )
        self.assertEqual(sec["values"]["section_id"], "sec_whee_001")
        self.assertFalse(
            any(
                subject["reasonCode"] == "missing_section"
                and "SEC" in str(subject)
                for subject in subjects
            )
        )

        bv4 = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "BV4"
        )
        self.assertFalse(
            bv4["values"]["active"],
            "target source status must not silently reactivate an existing inactive option",
        )
        self.assertTrue(
            any(
                subject["reasonCode"] == "option_behavior_conflict"
                and subject["model"] == "zr1"
                and "opt_bv4_001" in str(subject.get("proposedRows") or [])
                for subject in subjects
            )
        )
        behavior_subject = next(
            subject
            for subject in subjects
            if subject["reasonCode"] == "option_behavior_conflict"
            and subject["model"] == "zr1"
            and "opt_bv4_001" in str(subject.get("proposedRows") or [])
        )
        self.assertEqual(
            behavior_subject["proposedRows"][0]["exactTargetDefaultEvidence"][0]["model_key"],
            "zr1",
        )
        retained_default = next(
            row
            for row in rows
            if row["family"] == "default_selection_rules"
            and row["values"].get("rule_id") == "zr1_default_bv4"
        )
        self.assertFalse(
            any(
                dependency["evidenceId"].startswith("comparator:default_selection:")
                for dependency in retained_default["evidenceDependencies"]
            ),
            "comparator-only default evidence is not target behavior authority",
        )
        source_supported_default = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "default_selection_rules"
            and row["values"].get("target_option_id") == "opt_dft_001"
        )
        dft = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "DFT"
        )
        self.assertFalse(
            dft["values"]["selectable"],
            "a same-compile comparator proposal cannot replace exact target default evidence",
        )
        self.assertEqual(
            dft["values"]["price"],
            0,
            "target-standard status owns the included price over an optional comparator price",
        )
        self.assertEqual(
            source_supported_default["values"]["condition_id"], "opt_con_001"
        )
        self.assertFalse(
            any(
                subject["reasonCode"] == "option_behavior_conflict"
                and "opt_dft_001" in str(subject.get("proposedRows") or [])
                for subject in subjects
            ),
            "an optional single-select section does not require target default evidence",
        )

    def test_ambiguous_target_price_uses_raw_base_and_canonical_conditional_rules(self) -> None:
        from openpyxl import load_workbook

        raw = load_workbook(self.raw)
        raw["Mechanical 4"].append(
            ["CND", "", "Conditional price trigger", "A", "A"]
        )
        raw["Mechanical 4"].append(
            ["ALT", "", "Conditionally priced option", "A", "A"]
        )
        raw["Price Schedule"].cell(row=9, column=4, value="Qualifier")
        raw["Price Schedule"].cell(row=9, column=5, value="List")
        raw["Price Schedule"].cell(row=9, column=6, value="Factory")
        raw["Price Schedule"].cell(row=9, column=7, value="MSRP(c)")
        raw["Price Schedule"].append(
            ["", "ALT", "Conditionally priced option", "with CND", 75, 0, 75]
        )
        raw["Price Schedule"].append(
            ["", "ALT", "Conditionally priced option", "without CND", 100, 0, 100]
        )
        raw.save(self.raw)
        raw.close()

        workbook = load_workbook(self.master)
        workbook["z06_options"].append(
            [
                "opt_cnd_001", "CND", 0, "Conditional price trigger", "", "",
                "sec_whee_001", True, 31, True, "",
            ]
        )
        workbook["z06_options"].append(
            [
                "opt_alt_001", "ALT", 100, "Conditionally priced option", "", "",
                "sec_whee_001", True, 32, True, "",
            ]
        )
        workbook["z06_price_rules"].append(
            [
                "z06_pr_cnd_alt", "opt_cnd_001", "override", "opt_alt_001",
                75, "", "", "", "Fixture conditional price",
            ]
        )
        workbook.save(self.master)
        workbook.close()

        parsed = parse_confirmed_sheets(self.raw, ROLES)
        result = self.compile(
            option_payload={
                "schemaVersion": "pass-a-1",
                "candidates": parsed["candidates"],
                "skippedRows": parsed["skippedRows"],
            },
            price_payload={
                "schemaVersion": "pass-a-1",
                "priceRows": parsed["priceRows"],
                "baseModelPriceRows": parsed["baseModelPriceRows"],
                "skippedPriceRows": parsed["skippedPriceRows"],
            },
            join_report=join_prices(parsed["candidates"], parsed["priceRows"]),
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )
        option = next(
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "ALT"
        )
        self.assertEqual(option["values"]["price"], 100)
        self.assertFalse(
            any(
                subject["reasonCode"] == "unresolved_price_scope"
                and "ALT" in str(subject)
                for subject in result["exception-queue.json"]["subjects"]
            )
        )

    def test_trim_qualified_target_prices_emit_comparator_corroborated_self_overrides(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["z06_options"].append(
            [
                "opt_ae4_002", "AE4", 1095, "Competition Sport Bucket Seats", "", "",
                "sec_whee_001", True, 40, True, "",
            ]
        )
        workbook["z06_price_rules"].append(
            [
                "z06_pr_3lz_ae4", "opt_ae4_002", "override", "opt_ae4_002",
                595, "*", "3LZ", "*", "3LZ AE4 price",
            ]
        )
        workbook.save(self.master)
        workbook.close()

        option_payload = copy.deepcopy(self.option_payload)
        candidate = copy.deepcopy(
            next(item for item in option_payload["candidates"] if item.get("rpo") == "PDB")
        )
        candidate.update(
            {
                "candidateId": "Mechanical 4:998",
                "rpo": "AE4",
                "rpoNormalized": "AE4",
                "rpoDisplay": "AE4",
                "description": "Competition Sport Bucket Seats",
                "priceMatch": "ambiguous",
                "listPrice": None,
                "priceRows": [
                    {
                        "rpo": "AE4",
                        "qualifier": "1LT/LZ Only",
                        "priceColumnEvidence": [{"headerText": "List", "value": 1095}],
                    },
                    {
                        "rpo": "AE4",
                        "qualifier": "2LT/LZ Only",
                        "priceColumnEvidence": [{"headerText": "List", "value": 2095}],
                    },
                    {
                        "rpo": "AE4",
                        "qualifier": "3LT/LZ Only",
                        "priceColumnEvidence": [{"headerText": "List", "value": 595}],
                    },
                ],
            }
        )
        candidate["sourceEvidence"] = {
            **candidate["sourceEvidence"],
            "rowIndex": 998,
        }
        option_payload["candidates"].append(candidate)
        price_payload = copy.deepcopy(self.price_payload)
        three_lz_price = copy.deepcopy(
            next(
                row
                for row in price_payload["baseModelPriceRows"]
                if str(row.get("modelCode") or "").startswith("1YR")
            )
        )
        three_lz_price.update(
            {
                "modelCode": "1YR67",
                "description": "Corvette ZR1 Convertible 3LZ",
                "priceColumnEvidence": [{"headerText": "List", "value": 215700}],
                "sourceEvidence": {
                    **three_lz_price["sourceEvidence"],
                    "rowIndex": 999,
                },
            }
        )
        price_payload["baseModelPriceRows"].append(three_lz_price)
        price_payload["priceRows"].extend(copy.deepcopy(candidate["priceRows"]))
        join_report = join_prices(
            option_payload["candidates"],
            price_payload["priceRows"],
        )
        comparator = build_comparator_evidence(
            self.master,
            self.selection["comparators"],
            run_authority_fingerprint=self.authority,
        )
        result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            comparator_artifact=comparator,
        )

        ae4_options = [
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "AE4"
        ]
        self.assertEqual(len(ae4_options), 1, ae4_options)
        option = ae4_options[0]
        self.assertEqual(option["values"]["price"], 1095, option)
        override = next(
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "price_rules"
            and row["values"].get("target_option_id") == option["values"]["option_id"]
            and row["values"].get("trim_level_scope") == "3LZ"
        )
        self.assertEqual(override["values"]["price_value"], 595)
        self.assertFalse(
            any(
                subject["reasonCode"] == "unresolved_price_scope"
                and option_occurrence_signature(
                    {**candidate, "statuses": model_scoped_statuses(candidate, "zr1")}
                )
                in subject.get("evidenceReferences", [])
                for subject in result["exception-queue.json"]["subjects"]
            )
        )
        price_dispositions = {
            row["qualifier"]: next(
                feature["disposition"]
                for feature in result["compile-report.json"]["sourceFeatureCoverage"]
                if feature["featureId"].endswith(f":AE4:{semantic_hash(row)}")
            )
            for row in candidate["priceRows"]
        }
        self.assertEqual(
            price_dispositions,
            {
                "1LT/LZ Only": "compiled",
                "2LT/LZ Only": "resolved_not_applicable",
                "3LT/LZ Only": "compiled",
            },
        )

        workbook = load_workbook(self.master)
        workbook["z06_options"].append(
            [
                "opt_other_001", "OTHER", 0, "Unrelated condition", "", "",
                "sec_whee_001", True, 42, True, "",
            ]
        )
        for row in workbook["z06_price_rules"].iter_rows(min_row=2):
            if row[0].value == "z06_pr_3lz_ae4":
                workbook["z06_price_rules"].cell(
                    row=row[0].row,
                    column=2,
                    value="opt_other_001",
                )
        workbook.save(self.master)
        workbook.close()
        unrelated_result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )
        unrelated_option = next(
            row
            for row in unrelated_result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "AE4"
        )
        self.assertEqual(unrelated_option["status"], "blocked")
        self.assertTrue(
            any(
                subject["reasonCode"] == "unresolved_price_scope"
                and subject["model"] == "zr1"
                and option_occurrence_signature(
                    {**candidate, "statuses": model_scoped_statuses(candidate, "zr1")}
                )
                in subject.get("evidenceReferences", [])
                for subject in unrelated_result["exception-queue.json"]["subjects"]
            )
        )

        workbook = load_workbook(self.master)
        for row in workbook["z06_price_rules"].iter_rows(min_row=2):
            if row[0].value == "z06_pr_3lz_ae4":
                workbook["z06_price_rules"].cell(
                    row=row[0].row,
                    column=2,
                    value="opt_ae4_001",
                )
                workbook["z06_price_rules"].cell(
                    row=row[0].row,
                    column=3,
                    value="discount",
                )
        workbook.save(self.master)
        workbook.close()
        wrong_type_result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )
        wrong_type_option = next(
            row
            for row in wrong_type_result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "AE4"
        )
        self.assertEqual(wrong_type_option["status"], "blocked")
        self.assertTrue(
            any(
                subject["reasonCode"] == "unresolved_price_scope"
                and subject["model"] == "zr1"
                and option_occurrence_signature(
                    {**candidate, "statuses": model_scoped_statuses(candidate, "zr1")}
                )
                in subject.get("evidenceReferences", [])
                for subject in wrong_type_result["exception-queue.json"]["subjects"]
            )
        )

    def test_profile_price_reconciliation_requires_complete_rule_identity(self) -> None:
        candidate = {
            "rpo": "AAA",
            "priceRows": [
                {
                    "rpo": "AAA",
                    "qualifier": "3LT/LZ Only",
                    "listPrice": 100,
                    "priceColumnEvidence": [{"headerText": "List", "value": 100}],
                },
                {
                    "rpo": "AAA",
                    "qualifier": "Coupes with AAA, 1LT/LZ Only",
                    "listPrice": 80,
                    "priceColumnEvidence": [{"headerText": "List", "value": 80}],
                },
            ],
        }
        rule = {
            "conditionRpo": "AAA",
            "targetRpo": "AAA",
            "priceRuleType": "override",
            "priceValue": 80,
            "bodyStyleScope": "coupe",
            "trimLevelScope": "1LZ",
            "variantScope": "*",
            "evidenceId": "comparator:override",
            "source": {"price_rule_type": "override"},
        }
        precedent = {
            "basePrice": 100,
            "conditionalPriceRules": [rule],
        }
        variants = [
            {"trim_level": "1lz", "body_style": "coupe"},
            {"trim_level": "3lz", "body_style": "convertible"},
        ]
        self.assertIsNotNone(
            _profile_reconciled_price(candidate, precedent, "zr1", variants)
        )
        for field, value in (
            ("bodyStyleScope", "convertible"),
            ("trimLevelScope", "3LZ"),
            ("targetRpo", "WRONG"),
            ("priceRuleType", "discount"),
        ):
            mismatched_rule = {**rule, field: value}
            self.assertIsNone(
                _profile_reconciled_price(
                    candidate,
                    {**precedent, "conditionalPriceRules": [mismatched_rule]},
                    "zr1",
                    variants,
                ),
                (field, value),
            )
        combined_candidate = {
            "rpo": "AAA",
            "priceRows": [
                {
                    "rpo": "AAA",
                    "qualifier": "Coupes without B6P",
                    "listPrice": 695,
                    "priceColumnEvidence": [{"headerText": "List", "value": 695}],
                },
                {
                    "rpo": "AAA",
                    "qualifier": "Convertible requires ZZ3; Coupes with B6P",
                    "listPrice": 595,
                    "priceColumnEvidence": [{"headerText": "List", "value": 595}],
                },
            ],
        }
        combined_rules = [
            {
                **rule,
                "conditionRpo": "ZZ3",
                "priceValue": 595,
                "bodyStyleScope": "convertible",
                "trimLevelScope": "*",
                "evidenceId": "comparator:convertible",
            },
            {
                **rule,
                "conditionRpo": "B6P",
                "priceValue": 595,
                "bodyStyleScope": "coupe",
                "trimLevelScope": "*",
                "evidenceId": "comparator:coupe",
            },
        ]
        self.assertIsNotNone(
            _profile_reconciled_price(
                combined_candidate,
                {"basePrice": 695, "conditionalPriceRules": combined_rules},
                "grand_sport_x",
                variants,
            )
        )

    def test_model_qualified_price_rows_select_the_exact_target_price(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["z06_options"].append(
            [
                "opt_ztk_001", "ZTK", 5995, "Track Performance Package", "", "",
                "sec_whee_001", True, 41, True, "",
            ]
        )
        workbook.save(self.master)
        workbook.close()

        option_payload = copy.deepcopy(self.option_payload)
        candidate = copy.deepcopy(
            next(item for item in option_payload["candidates"] if item.get("rpo") == "PDB")
        )
        candidate.update(
            {
                "candidateId": "Mechanical 4:997",
                "rpo": "ZTK",
                "rpoNormalized": "ZTK",
                "rpoDisplay": "ZTK",
                "description": "Track Performance Package",
                "detailRaw": "Track Performance Package",
                "listPrice": None,
                "priceRows": [],
            }
        )
        candidate["sourceEvidence"] = {
            **candidate["sourceEvidence"],
            "rowIndex": 997,
        }
        option_payload["candidates"].append(candidate)
        price_payload = copy.deepcopy(self.price_payload)
        qualified_rows = [
            {
                "rpo": "ZTK",
                "qualifier": "ZR1X only; requires TOM",
                "listPrice": 1500,
                "priceColumnEvidence": [{"headerText": "List", "value": 1500}],
            },
            {
                "rpo": "ZTK",
                "qualifier": "ZR1 only; requires TOM",
                "listPrice": 5995,
                "priceColumnEvidence": [{"headerText": "List", "value": 5995}],
            },
        ]
        price_payload["priceRows"].extend(copy.deepcopy(qualified_rows))
        join_report = join_prices(
            option_payload["candidates"],
            price_payload["priceRows"],
        )
        result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )

        option = next(
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "ZTK"
        )
        self.assertEqual(option["values"]["price"], 5995)
        self.assertFalse(
            any(
                subject["reasonCode"] == "unresolved_price_scope"
                and subject["model"] == "zr1"
                and option_occurrence_signature(
                    {**candidate, "statuses": model_scoped_statuses(candidate, "zr1")}
                )
                in subject.get("evidenceReferences", [])
                for subject in result["exception-queue.json"]["subjects"]
            )
        )
        dispositions = {
            row["qualifier"]: next(
                feature["disposition"]
                for feature in result["compile-report.json"]["sourceFeatureCoverage"]
                if feature["featureId"].endswith(f":ZTK:{semantic_hash(row)}")
            )
            for row in candidate["priceRows"]
        }
        self.assertEqual(
            dispositions,
            {
                "ZR1X only; requires TOM": "resolved_not_applicable",
                "ZR1 only; requires TOM": "compiled",
            },
        )
        price_features = [
            feature
            for feature in result["compile-report.json"]["sourceFeatureCoverage"]
            if feature["featureId"].endswith(
                tuple(
                    f":ZTK:{semantic_hash(row)}"
                    for row in candidate["priceRows"]
                )
            )
        ]
        self.assertEqual(
            {(feature["model"], feature["disposition"]) for feature in price_features},
            {("zr1", "compiled"), ("zr1x", "resolved_not_applicable")},
        )

        other_target_only = copy.deepcopy(option_payload)
        other_target_candidate = next(
            item
            for item in other_target_only["candidates"]
            if item.get("candidateId") == candidate["candidateId"]
        )
        other_target_candidate["priceMatch"] = "exact"
        other_target_candidate["priceRows"] = [copy.deepcopy(qualified_rows[0])]
        other_target_price_payload = copy.deepcopy(self.price_payload)
        other_target_price_payload["priceRows"].append(copy.deepcopy(qualified_rows[0]))
        other_target_result = self.compile(
            option_payload=other_target_only,
            price_payload=other_target_price_payload,
            join_report=join_prices(
                other_target_only["candidates"],
                other_target_price_payload["priceRows"],
            ),
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )
        other_target_option = next(
            row
            for row in other_target_result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "ZTK"
        )
        self.assertEqual(other_target_option["status"], "blocked")
        self.assertNotEqual(other_target_option["values"]["price"], 1500)
        self.assertTrue(
            any(
                subject["reasonCode"] == "unresolved_price_scope"
                and subject["model"] == "zr1"
                and option_occurrence_signature(
                    {
                        **other_target_candidate,
                        "statuses": model_scoped_statuses(other_target_candidate, "zr1"),
                    }
                )
                in subject.get("evidenceReferences", [])
                for subject in other_target_result["exception-queue.json"]["subjects"]
            )
        )

        same_price_payload = copy.deepcopy(option_payload)
        same_price_candidate = next(
            item
            for item in same_price_payload["candidates"]
            if item.get("candidateId") == candidate["candidateId"]
        )
        same_price_rows = [
            {
                "rpo": "ZTK",
                "qualifier": "ZR1 only; requires TOM",
                "listPrice": 777,
                "priceColumnEvidence": [{"headerText": "List", "value": 777}],
            },
            {
                "rpo": "ZTK",
                "qualifier": "ZR1 only; requires S47",
                "listPrice": 777,
                "priceColumnEvidence": [{"headerText": "List", "value": 777}],
            },
        ]
        same_price_candidate["priceMatch"] = "ambiguous"
        same_price_candidate["priceRows"] = copy.deepcopy(same_price_rows)
        same_price_source = copy.deepcopy(self.price_payload)
        same_price_source["priceRows"].extend(copy.deepcopy(same_price_rows))
        same_price_result = self.compile(
            option_payload=same_price_payload,
            price_payload=same_price_source,
            join_report=join_prices(
                same_price_payload["candidates"],
                same_price_source["priceRows"],
            ),
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )
        same_price_option = next(
            row
            for row in same_price_result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "ZTK"
        )
        self.assertEqual(same_price_option["status"], "ready")
        self.assertEqual(same_price_option["values"]["price"], 777)
        self.assertFalse(
            any(
                subject["reasonCode"] == "unresolved_price_scope"
                and subject["model"] == "zr1"
                and option_occurrence_signature(
                    {**same_price_candidate, "statuses": model_scoped_statuses(same_price_candidate, "zr1")}
                )
                in subject.get("evidenceReferences", [])
                for subject in same_price_result["exception-queue.json"]["subjects"]
            )
        )

    def test_ref_only_target_option_materializes_target_qualified_price_evidence(self) -> None:
        from openpyxl import load_workbook

        raw = load_workbook(self.raw)
        raw["Mechanical 4"].append(
            ["", "R8E", "Gas guzzler tax, mandatory federal tax", "A", "A"]
        )
        raw["Price Schedule"].append(
            ["", "R8E", "Gas Guzzler Tax", "ZR1 only", 3000, 0, 3000]
        )
        raw.save(self.raw)
        raw.close()

        workbook = load_workbook(self.master)
        for sheet in ("z06_options", "zr1_options"):
            workbook[sheet].append(
                [
                    "opt_r8e_002",
                    "R8E",
                    3000,
                    "Gas Guzzler Tax",
                    "Mandatory federal tax",
                    "",
                    "sec_whee_001",
                    False,
                    44,
                    True,
                    "",
                ]
            )
        workbook.save(self.master)
        workbook.close()

        parsed = parse_confirmed_sheets(self.raw, ROLES)
        report = join_prices(parsed["candidates"], parsed["priceRows"])
        candidate = next(
            item for item in parsed["candidates"] if item.get("refOnlyRpo") == "R8E"
        )
        self.assertFalse(candidate.get("priceRows"))
        price_row = next(
            row
            for row in parsed["priceRows"]
            if row.get("rpo") == "R8E" and row.get("qualifier") == "ZR1 only"
        )
        result = self.compile(
            option_payload={
                "schemaVersion": "pass-a-1",
                "candidates": parsed["candidates"],
                "skippedRows": parsed["skippedRows"],
            },
            price_payload={
                "schemaVersion": "pass-a-1",
                "priceRows": parsed["priceRows"],
                "baseModelPriceRows": parsed["baseModelPriceRows"],
                "skippedPriceRows": parsed["skippedPriceRows"],
            },
            join_report=report,
            sheet_profile=profile_workbook(self.raw),
            comparator_artifact=build_comparator_evidence(
                self.master,
                self.selection["comparators"],
                run_authority_fingerprint=self.authority,
            ),
        )
        option = next(
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "R8E"
        )
        price_evidence_id = f"price:R8E:{semantic_hash(price_row)}"
        self.assertEqual(option["values"]["price"], 3000)
        self.assertIn(
            price_evidence_id,
            {dependency["evidenceId"] for dependency in option["evidenceDependencies"]},
        )
        price_feature = next(
            feature
            for feature in result["compile-report.json"]["sourceFeatureCoverage"]
            if feature["featureId"] == f"price:zr1:R8E:{semantic_hash(price_row)}"
        )
        self.assertEqual(price_feature["model"], "zr1")
        self.assertEqual(price_feature["disposition"], "compiled")

    def test_profile_effect_relationship_compiles_through_source_ledger(self) -> None:
        from openpyxl import load_workbook

        raw = load_workbook(self.raw)
        raw["Mechanical 4"].append(
            ["Z25", "", "Includes (HTA) Jet Black interior.", "A", "A"]
        )
        raw["Price Schedule"].append(
            ["", "Z25", "Launch Edition", 1995, 0, 1995]
        )
        raw.save(self.raw)
        raw.close()

        workbook = load_workbook(self.master)
        workbook["z06_options"].append(
            [
                "opt_z25_001", "Z25", 0, "Launch Edition",
                "Includes Jet Black interior.", "", "sec_whee_001",
                False, 30, True, "auto_only",
            ]
        )
        workbook["z06_ovs"].append(["opt_z25_001", "1lz_h07", "available"])
        scope_sheet = workbook["model_interior_scope"]
        for row in scope_sheet.iter_rows(min_row=2):
            if row[0].value == "z06":
                scope_sheet[f"E{row[0].row}"] = "opt_z25_001"
        interior_sheet = workbook["LZ_Interiors"]
        for row_index in range(2, interior_sheet.max_row + 1):
            interior_sheet.cell(row=row_index, column=4, value=1995)
        model_variants = workbook["model_variants"]
        for row_index in range(model_variants.max_row, 1, -1):
            if (
                model_variants.cell(row_index, 1).value == "zr1"
                and model_variants.cell(row_index, 2).value == "3lz_r67"
            ):
                model_variants.delete_rows(row_index)
        workbook.save(self.master)
        workbook.close()

        parsed = parse_confirmed_sheets(self.raw, ROLES)
        report = join_prices(parsed["candidates"], parsed["priceRows"])
        option_payload = {
            "schemaVersion": "pass-a-1",
            "candidates": parsed["candidates"],
            "skippedRows": parsed["skippedRows"],
        }
        price_payload = {
            "schemaVersion": "pass-a-1",
            "priceRows": parsed["priceRows"],
            "baseModelPriceRows": parsed["baseModelPriceRows"],
            "skippedPriceRows": parsed["skippedPriceRows"],
        }
        comparator = build_comparator_evidence(
            self.master,
            self.selection["comparators"],
            run_authority_fingerprint=self.authority,
        )
        result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=report,
            comparator_artifact=comparator,
        )
        copy_subject = next(
            item
            for item in result["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "comparator_copy_conflict"
            and item["proposedRows"][0].get("optionId") == "opt_z25_001"
        )
        proposal = copy_subject["proposedRows"][0]
        self.assertEqual(
            {
                "targetStatuses",
                "behaviorEvidence",
                "placementEvidence",
                "priceEvidence",
            }
            - set(proposal),
            set(),
        )
        result = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=report,
            comparator_artifact=comparator,
            resolution_entries=[
                {
                    "subjectId": copy_subject["subjectId"],
                    "subjectVersion": copy_subject["subjectVersion"],
                    "action": "provide_option_copy",
                    "payload": {
                        "optionName": proposal["proposedOptionName"],
                        "description": proposal["proposedDescription"],
                    },
                    "disposition": "resolved",
                }
            ],
        )
        self.assertTrue(
            any(
                item["family"] == "rule_mapping"
                and item["disposition"] == "compiled"
                for item in result["compile-report.json"]["sourceFeatureCoverage"]
            )
        )
        rows = result["canonical-row-manifest.json"]["rows"]
        compiled_relationship_features = [
            item
            for item in result["compile-report.json"]["sourceFeatureCoverage"]
            if item["family"] == "rule_mapping"
            and item["disposition"] == "compiled"
            and item["featureId"].startswith("relationship:")
        ]
        for feature in compiled_relationship_features:
            declared = set(feature["evidenceIds"])
            self.assertTrue(
                any(
                    declared
                    <= {
                        dependency["evidenceId"]
                        for dependency in row["evidenceDependencies"]
                    }
                    for row in rows
                ),
                feature,
            )
        z25_option = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "Z25"
            and row["status"] == "ready"
        )
        self.assertEqual(z25_option["values"]["price"], 0)
        self.assertTrue(z25_option["values"]["selectable"])
        self.assertEqual(z25_option["values"]["display_behavior"], "auto_only")
        self.assertTrue(
            any(
                dependency["evidenceId"].startswith("price:Z25:")
                for dependency in z25_option["evidenceDependencies"]
            )
        )
        self.assertTrue(
            any(
                dependency["evidenceId"].endswith(":price-allocation")
                for dependency in z25_option["evidenceDependencies"]
            )
        )
        self.assertEqual(
            {
                row["values"]["Price"]
                for row in rows
                if row["model"] == "zr1"
                and row["family"] == "interiors"
                and row["values"].get("interior_id")
                in {"1LZ_AQ9_HTA", "3LZ_AQ9_HTA"}
            },
            {1995},
        )
        z25_ovs = [
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "ovs"
            and row["values"].get("option_id") == "opt_z25_001"
        ]
        self.assertTrue(z25_ovs)
        self.assertTrue(
            all(
                any(
                    dependency["evidenceId"].startswith("status:")
                    for dependency in row["evidenceDependencies"]
                )
                for row in z25_ovs
            )
        )
        self.assertEqual(
            {
                row["values"]["variant_id"]
                for row in rows
                if row["model"] == "zr1"
                and row["family"] == "ovs"
                and row["values"].get("option_id") == "opt_z25_001"
                and row["values"].get("status") == "available"
            },
            {"1lz_r07"},
        )
        z25_signatures = {
            option_occurrence_signature(candidate)
            for candidate in parsed["candidates"]
            if candidate.get("rpo") == "Z25"
        }
        self.assertFalse(
            any(
                subject["reasonCode"] == "missing_section"
                and z25_signatures.intersection(subject["evidenceReferences"])
                for subject in result["exception-queue.json"]["subjects"]
            )
        )

    def test_color_trim_source_requires_source_authority_and_tracks_its_fingerprint(self) -> None:
        result = self.compile()
        coverage = next(
            item
            for item in result["compile-report.json"]["sourceFeatureCoverage"]
            if item["featureId"] == "source-sheet:Color and Trim 1"
        )
        self.assertEqual(coverage["disposition"], "resolved_not_a_workbook_fact")
        self.assertTrue(
            any(
                evidence_id.startswith("source-sheet-content:Color and Trim 1:")
                for evidence_id in coverage["evidenceIds"]
            )
        )

        missing_bindings = copy.deepcopy(self.authority["bindings"])
        del missing_bindings["sourceSha256"]
        missing_authority = {
            "fingerprint": hashlib.sha256(canonical_bytes(missing_bindings)).hexdigest(),
            "bindings": missing_bindings,
        }
        missing_comparator = build_comparator_evidence(
            self.master,
            self.selection["comparators"],
            run_authority_fingerprint=missing_authority,
        )
        missing_result = self.compile(
            run_authority_fingerprint=missing_authority,
            comparator_artifact=missing_comparator,
        )
        self.assertTrue(
            any(
                subject["reasonCode"] == "unsupported_color_trim_source"
                for subject in missing_result["exception-queue.json"]["subjects"]
            )
        )

        changed_bindings = copy.deepcopy(self.authority["bindings"])
        changed_bindings["sourceSha256"] = "c" * 64
        changed_authority = {
            "fingerprint": hashlib.sha256(canonical_bytes(changed_bindings)).hexdigest(),
            "bindings": changed_bindings,
        }
        changed_comparator = build_comparator_evidence(
            self.master,
            self.selection["comparators"],
            run_authority_fingerprint=changed_authority,
        )
        changed_result = self.compile(
            run_authority_fingerprint=changed_authority,
            comparator_artifact=changed_comparator,
        )
        changed_coverage = next(
            item
            for item in changed_result["compile-report.json"]["sourceFeatureCoverage"]
            if item["featureId"] == "source-sheet:Color and Trim 1"
        )
        self.assertNotEqual(coverage["evidenceIds"], changed_coverage["evidenceIds"])

    def test_color_trim_policy_does_not_close_a_nonexistent_excluded_sheet(self) -> None:
        cases = (
            ("Color and Trim Ghost", "exclude"),
            ("Color and Trim 1", "options"),
        )
        for sheet_name, role in cases:
            with self.subTest(sheet_name=sheet_name, role=role):
                roles_payload = copy.deepcopy(self.roles_payload)
                roles_payload["roles"][sheet_name] = role

                result = self.compile(roles_payload=roles_payload)
                coverage = next(
                    item
                    for item in result["compile-report.json"]["sourceFeatureCoverage"]
                    if item["featureId"] == f"source-sheet:{sheet_name}"
                )

                self.assertEqual(coverage["disposition"], "exception_open")
                self.assertFalse(
                    any(
                        evidence_id.startswith(f"source-sheet-content:{sheet_name}:")
                        for evidence_id in coverage["evidenceIds"]
                    )
                )
                self.assertTrue(
                    any(
                        subject["reasonCode"] == "unsupported_color_trim_source"
                        and f"sheet-role:{sheet_name}" in subject["evidenceReferences"]
                        for subject in result["exception-queue.json"]["subjects"]
                    )
                )

    def test_existing_option_id_is_reused_and_no_family_is_cleared(self) -> None:
        result = self.compile()
        rows = result["canonical-row-manifest.json"]["rows"]
        pdb = next(row for row in rows if row["family"] == "options" and row["values"].get("rpo") == "PDB")
        self.assertEqual(pdb["values"]["option_id"], "opt_pdb_001")
        self.assertNotIn("clear", {row["action"] for row in rows})
        self.assertTrue(all(row["action"] != "delete" for row in rows))

    def test_compilation_does_not_mutate_shared_candidate_input(self) -> None:
        before = copy.deepcopy(self.option_payload)
        self.compile()
        self.assertEqual(self.option_payload, before)

    def test_every_option_has_explicit_boolean_values(self) -> None:
        result = self.compile()
        options = [row for row in result["canonical-row-manifest.json"]["rows"] if row["family"] == "options" and row["status"] == "ready"]
        self.assertTrue(options)
        for row in options:
            self.assertIsInstance(row["values"]["active"], bool)
            self.assertIsInstance(row["values"]["selectable"], bool)
        bv4 = next(row for row in options if row["values"]["rpo"] == "BV4")
        self.assertEqual(bv4["values"]["price"], 395)

        variant = next(
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["family"] == "variant_master" and row["key"] == {"variant_id": "1lz_r07"}
        )
        self.assertEqual(variant["values"]["base_price"], 197195)
        self.assertIs(variant["values"]["active"], False)
        self.assertTrue(
            any(
                row["family"] == "model_variants"
                and row["key"] == {"model_key": "zr1", "variant_id": "1lz_r07"}
                for row in result["canonical-row-manifest.json"]["rows"]
            )
        )

    def test_status_bearing_no_rpo_standard_candidate_compiles_to_option_and_ovs(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["zr1_options"].append(
            [
                "opt_std_existing",
                "",
                0,
                "Air filtration system with pollen filter",
                "",
                "Air filtration system with pollen filter",
                "sec_whee_001",
                False,
                40,
                False,
                "",
            ]
        )
        workbook["zr1_options"].append(
            [
                "opt_std_other",
                "",
                0,
                "Unrelated no-RPO standard equipment",
                "",
                "Unrelated no-RPO standard equipment",
                "sec_stan_001",
                False,
                41,
                False,
                "",
            ]
        )
        workbook.save(self.master)
        workbook.close()
        option_payload = copy.deepcopy(self.option_payload)
        option_payload["candidates"].append(
            {
                "candidateId": "Interior 5:4",
                "sheetName": "Interior 5",
                "rowIndex": 4,
                "modelFamily": "ZR1",
                "modelFamilies": ["ZR1"],
                "sectionLabel": "",
                "rowKind": "standard_no_rpo",
                "rpo": "",
                "refOnlyRpo": "",
                "description": "Air filtration system with pollen filter",
                "statuses": [
                    {
                        "columnLetter": "D",
                        "variantLabel": "ZR1 Coupe",
                        "modelCode": "1YR07",
                        "trim": "1LZ",
                        "bodyStyle": "coupe",
                        "raw": "S",
                        "status": "standard",
                        "disclosureMarker": "",
                        "flags": [],
                    }
                ],
                "sourceEvidence": {
                    "sheetName": "Interior 5",
                    "rowIndex": 4,
                    "cells": {"C4": "Air filtration system with pollen filter", "D4": "S"},
                },
                "priceMatch": None,
                "listPrice": None,
                "priceRows": [],
            }
        )

        result = self.compile(option_payload=option_payload)
        rows = result["canonical-row-manifest.json"]["rows"]
        option = next(
            row
            for row in rows
            if row["family"] == "options"
            and row["values"].get("option_name") == "Air filtration system with pollen filter"
        )

        self.assertEqual(option["values"]["option_id"], "opt_std_existing")
        self.assertEqual(option["values"]["section_id"], "sec_whee_001")
        self.assertIs(option["values"]["active"], False)
        self.assertIs(option["values"]["selectable"], False)
        ovs = next(
            row
            for row in rows
            if row["family"] == "ovs"
            and row["values"].get("option_id") == option["values"]["option_id"]
        )
        self.assertEqual(ovs["values"]["status"], "standard")
        self.assertFalse(
            any(
                subject["reasonCode"] == "unsupported_source_feature"
                and "Interior 5:4" in subject.get("evidenceReferences", [])
                for subject in result["exception-queue.json"]["subjects"]
            )
        )

    def test_all_unavailable_duplicate_rpo_does_not_compete_for_target_identity(self) -> None:
        option_payload = copy.deepcopy(self.option_payload)
        target_candidate = next(
            candidate
            for candidate in option_payload["candidates"]
            if candidate.get("rpo") == "BV4"
        )
        sibling = copy.deepcopy(target_candidate)
        sibling["candidateId"] = "Mechanical 4:999"
        sibling["description"] = "ZR1X Personalized Plaque"
        sibling["sourceEvidence"] = {
            **sibling["sourceEvidence"],
            "rowIndex": 999,
        }
        sibling["statuses"] = [
            {
                **status,
                "raw": "--",
                "status": "unavailable",
                "disclosureMarker": "",
                "flags": [],
            }
            for status in sibling["statuses"]
        ]
        option_payload["candidates"].append(sibling)
        join_report = join_prices(
            option_payload["candidates"],
            self.price_payload["priceRows"],
        )

        result = self.compile(
            option_payload=option_payload,
            join_report=join_report,
        )

        self.assertFalse(
            any(
                subject["reasonCode"] == "ambiguous_existing_identity"
                and any(row.get("rpo") == "BV4" for row in subject["proposedRows"])
                for subject in result["exception-queue.json"]["subjects"]
            )
        )
        options = [
            row
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("rpo") == "BV4"
        ]
        self.assertEqual(len(options), 1)
        self.assertEqual(options[0]["values"]["option_id"], "opt_bv4_001")

    def test_all_unavailable_target_occurrence_is_not_an_exception(self) -> None:
        option_payload = copy.deepcopy(self.option_payload)
        unavailable = copy.deepcopy(
            next(
                candidate
                for candidate in option_payload["candidates"]
                if candidate.get("rpo") == "BV4"
            )
        )
        unavailable.update(
            {
                "candidateId": "Mechanical 4:998",
                "rpo": "ZZZ",
                "refOnlyRpo": "",
                "description": "Unavailable target-only option",
                "priceMatch": None,
                "listPrice": None,
                "priceRows": [],
            }
        )
        unavailable["sourceEvidence"] = {
            **unavailable["sourceEvidence"],
            "rowIndex": 998,
        }
        unavailable["statuses"] = [
            {
                **status,
                "raw": "--",
                "status": "unavailable",
                "disclosureMarker": "",
                "flags": [],
            }
            for status in unavailable["statuses"]
        ]
        option_payload["candidates"].append(unavailable)
        join_report = join_prices(
            option_payload["candidates"],
            self.price_payload["priceRows"],
        )
        scoped_signature = option_occurrence_signature(
            {
                **unavailable,
                "statuses": model_scoped_statuses(unavailable, "zr1"),
            }
        )

        result = self.compile(
            option_payload=option_payload,
            join_report=join_report,
        )

        self.assertFalse(
            any(
                scoped_signature in subject.get("evidenceReferences", [])
                for subject in result["exception-queue.json"]["subjects"]
            )
        )
        self.assertFalse(
            any(
                row["model"] == "zr1"
                and row["family"] == "options"
                and row["values"].get("rpo") == "ZZZ"
                for row in result["canonical-row-manifest.json"]["rows"]
            )
        )
        source_feature = next(
            item
            for item in result["compile-report.json"]["sourceFeatureCoverage"]
            if item["family"] == "options"
            and item["featureId"]
            == f"candidate:zr1:{_candidate_feature_index(option_payload['candidates'])[unavailable['candidateId']]}"
        )
        self.assertEqual(source_feature["disposition"], "resolved_not_applicable")

    def test_all_unavailable_target_occurrence_deletes_existing_option_and_ovs(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["zr1_options"].append(
            [
                "opt_zzz_001",
                "ZZZ",
                0,
                "Unavailable target-only option",
                "",
                "",
                "sec_whee_001",
                False,
                30,
                True,
                "",
            ]
        )
        ovs = workbook.create_sheet("zr1_ovs")
        ovs.append(["option_id", "variant_id", "status"])
        ovs.append(["opt_zzz_001", "1lz_r07", "unavailable"])
        ovs.append(["opt_zzz_001", "3lz_r67", "unavailable"])
        rules = workbook.create_sheet("zr1_rule_mapping")
        rules.append(
            [
                "rule_id",
                "source_id",
                "rule_type",
                "target_id",
                "original_detail_raw",
                "body_style_scope",
                "runtime_action",
                "disabled_reason",
            ]
        )
        rules.append(
            [
                "zr1_rule_existing_requires_zzz",
                "opt_bv4_001",
                "requires",
                "opt_zzz_001",
                "",
                "",
                "",
                "",
            ]
        )
        workbook.save(self.master)
        workbook.close()

        option_payload = copy.deepcopy(self.option_payload)
        unavailable = copy.deepcopy(
            next(
                candidate
                for candidate in option_payload["candidates"]
                if candidate.get("rpo") == "BV4"
            )
        )
        unavailable.update(
            {
                "candidateId": "Mechanical 4:998",
                "rpo": "ZZZ",
                "refOnlyRpo": "",
                "description": "Unavailable target-only option",
                "priceMatch": None,
                "listPrice": None,
                "priceRows": [],
            }
        )
        unavailable["sourceEvidence"] = {
            **unavailable["sourceEvidence"],
            "rowIndex": 998,
        }
        unavailable["statuses"] = [
            {
                **status,
                "raw": "--",
                "status": "unavailable",
                "disclosureMarker": "",
                "flags": [],
            }
            for status in unavailable["statuses"]
        ]
        option_payload["candidates"].append(unavailable)

        result = self.compile(
            option_payload=option_payload,
            join_report=join_prices(
                option_payload["candidates"],
                self.price_payload["priceRows"],
            ),
        )
        rows = result["canonical-row-manifest.json"]["rows"]

        option = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["values"].get("option_id") == "opt_zzz_001"
        )
        self.assertEqual(option["action"], "delete")
        self.assertEqual(option["disposition"], "resolved_not_applicable")
        target_ovs = [
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "ovs"
            and row["values"].get("option_id") == "opt_zzz_001"
        ]
        self.assertEqual(len(target_ovs), 2)
        self.assertEqual({row["action"] for row in target_ovs}, {"delete"})
        self.assertEqual(
            {row["disposition"] for row in target_ovs},
            {"resolved_not_applicable"},
        )
        relationship = next(
            row
            for row in rows
            if row["model"] == "zr1"
            and row["family"] == "rule_mapping"
            and row["values"].get("rule_id") == "zr1_rule_existing_requires_zzz"
        )
        self.assertEqual(relationship["action"], "delete")
        self.assertEqual(relationship["disposition"], "resolved_not_applicable")

    def test_all_unavailable_deletion_blocks_when_target_identity_is_ambiguous(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        for option_id in ("opt_zzz_001", "opt_zzz_002"):
            workbook["zr1_options"].append(
                [option_id, "ZZZ", 0, "Ambiguous unavailable option", "", "", "sec_whee_001", False, 30, False, ""]
            )
        workbook.save(self.master)
        workbook.close()

        option_payload = copy.deepcopy(self.option_payload)
        unavailable = copy.deepcopy(
            next(candidate for candidate in option_payload["candidates"] if candidate.get("rpo") == "BV4")
        )
        unavailable.update(
            {
                "candidateId": "Mechanical 4:998",
                "rpo": "ZZZ",
                "refOnlyRpo": "",
                "description": "Ambiguous unavailable option",
                "detailRaw": "Ambiguous unavailable option",
                "priceMatch": None,
                "listPrice": None,
                "priceRows": [],
                "statuses": [
                    {**status, "raw": "--", "status": "unavailable", "disclosureMarker": "", "flags": []}
                    for status in unavailable["statuses"]
                ],
            }
        )
        option_payload["candidates"].append(unavailable)

        result = self.compile(
            option_payload=option_payload,
            join_report=join_prices(option_payload["candidates"], self.price_payload["priceRows"]),
        )

        self.assertTrue(
            any(
                subject["reasonCode"] == "ambiguous_deletion_identity"
                and subject["allowedActions"] == []
                for subject in result["exception-queue.json"]["subjects"]
            )
        )
        self.assertFalse(
            any(
                row["family"] == "options"
                and row["values"].get("rpo") == "ZZZ"
                and row["action"] == "delete"
                for row in result["canonical-row-manifest.json"]["rows"]
            )
        )

    def test_retain_existing_resolution_consumes_one_ambiguous_option_identity(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        for option_id in ("opt_dup_001", "opt_dup_002"):
            workbook["zr1_options"].append(
                [option_id, "DUP", 0, "Duplicate identity", "", "", "sec_whee_001", False, 30, False, ""]
            )
        workbook.save(self.master)
        workbook.close()

        option_payload = copy.deepcopy(self.option_payload)
        option_payload["candidates"].append(
            {
                "candidateId": "Equipment Groups 4:99",
                "sheetName": "Equipment Groups 4",
                "rowIndex": 99,
                "modelFamily": "ZR1",
                "modelFamilies": ["ZR1"],
                "sectionLabel": "Wheels",
                "rowKind": "orderable",
                "rpo": "DUP",
                "refOnlyRpo": "",
                "description": "Duplicate identity",
                "statuses": [
                    {
                        "columnLetter": "D",
                        "variantLabel": "ZR1 Coupe",
                        "modelCode": "1YR07",
                        "trim": "1LZ",
                        "bodyStyle": "coupe",
                        "raw": "A",
                        "status": "available",
                        "disclosureMarker": "",
                        "flags": [],
                    }
                ],
                "sourceEvidence": {
                    "sheetName": "Equipment Groups 4",
                    "rowIndex": 99,
                    "cells": {"A99": "DUP", "C99": "Duplicate identity", "D99": "A"},
                },
                "priceMatch": None,
                "listPrice": None,
                "priceRows": [],
            }
        )

        first = self.compile(option_payload=option_payload)
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "ambiguous_existing_identity"
            and item["model"] == "zr1"
        )
        self.assertEqual(subject["allowedActions"], ["retain_existing"])
        self.assertEqual(
            {row["existingId"] for row in subject["proposedRows"]},
            {"opt_dup_001", "opt_dup_002"},
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "retain_existing",
            "payload": {"existingId": "opt_dup_002"},
            "disposition": "retained_existing",
        }

        second = self.compile(option_payload=option_payload, resolution_entries=[resolution])
        option = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "options"
            and row["values"].get("rpo") == "DUP"
            and any(
                dep["evidenceId"] == f"resolution:{subject['subjectId']}"
                for dep in row["evidenceDependencies"]
            )
        )
        self.assertEqual(option["values"]["option_id"], "opt_dup_002")
        self.assertTrue(
            any(
                dep["evidenceId"] == f"resolution:{subject['subjectId']}"
                for dep in option["evidenceDependencies"]
            )
        )
        self.assertNotIn(
            subject["subjectId"],
            {
                blocker["subjectId"]
                for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )

    def test_choose_relationship_resolution_materializes_direct_rule(self) -> None:
        option_payload, price_payload, join_report, first, subject = (
            self.compile_with_ready_comparator_relationship()
        )
        option_rows = [
            row
            for row in first["canonical-row-manifest.json"]["rows"]
            if row["family"] == "options" and row["status"] == "ready"
        ]
        pdb_id = next(row["values"]["option_id"] for row in option_rows if row["values"].get("rpo") == "PDB")
        bv4_id = next(row["values"]["option_id"] for row in option_rows if row["values"].get("rpo") == "BV4")
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "choose_relationship",
            "payload": {
                "sourceOptionId": pdb_id,
                "ruleType": "requires",
                "targetOptionId": bv4_id,
            },
            "disposition": "resolved",
        }

        second = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            resolution_entries=[resolution],
        )
        rule = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "rule_mapping"
            and row["values"].get("source_id") == pdb_id
            and row["values"].get("target_id") == bv4_id
        )
        self.assertEqual(rule["values"]["rule_type"], "requires")
        self.assertTrue(
            any(
                dep["evidenceId"] == f"resolution:{subject['subjectId']}"
                for dep in rule["evidenceDependencies"]
            )
        )
        self.assertNotIn(
            subject["subjectId"],
            {
                blocker["subjectId"]
                for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )

    def test_mark_not_applicable_consumes_comparator_proposal_without_a_row(self) -> None:
        option_payload, price_payload, join_report, _, subject = (
            self.compile_with_ready_comparator_relationship()
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "mark_not_applicable",
            "payload": {"reason": "Target order guide does not establish this comparator rule."},
            "disposition": "resolved_not_applicable",
        }

        second = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            resolution_entries=[resolution],
        )
        self.assertNotIn(
            subject["subjectId"],
            {
                blocker["subjectId"]
                for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )
        self.assertFalse(
            any(
                dep["evidenceId"] == f"resolution:{subject['subjectId']}"
                for row in second["canonical-row-manifest.json"]["rows"]
                for dep in row["evidenceDependencies"]
            )
        )
        self.assertTrue(
            any(
                item["disposition"] == "resolved_not_applicable"
                and item["evidenceId"] in subject["evidenceReferences"]
                for item in second["compile-report.json"]["comparatorDispositions"]
            )
        )
        for evidence_id in subject["evidenceReferences"]:
            ledger_entries = [
                item
                for item in second["compile-report.json"]["sourceFeatureCoverage"]
                if evidence_id in item["evidenceIds"]
            ]
            self.assertEqual(len(ledger_entries), 1)
            self.assertEqual(ledger_entries[0]["disposition"], "resolved_not_applicable")

    def test_mark_not_applicable_updates_comparator_evidence_disposition(self) -> None:
        evidence_id = "comparator:direct_rule:variant-scoped"
        subject = compiler_module._typed_exception(
            "zr1",
            "rule_mapping",
            "comparator_only_relationship_proposal",
            ["PBC", "requires", "ZZ3"],
            [],
            evidence_references=[evidence_id],
            proposed_rows=[
                {
                    "sourceRpo": "PBC",
                    "ruleType": "requires",
                    "targetRpo": "ZZ3",
                    "bodyStyleScope": "*",
                    "trimLevelScope": "*",
                    "variantScope": "*",
                }
            ],
            allowed_actions=["choose_relationship", "mark_not_applicable"],
            question="Confirm or reject this comparator relationship.",
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "mark_not_applicable",
            "payload": {"reason": "The target status matrix makes the rule unnecessary."},
            "disposition": "resolved_not_applicable",
        }
        relationship_result = {
            "rows": [],
            "exceptions": [subject],
            "dispositions": [
                {
                    "featureId": "derived:relationship:proposal",
                    "evidenceIds": [evidence_id],
                    "disposition": "proposed_exception",
                }
            ],
        }

        _, _, consumed, overrides = compiler_module._relationship_rows(
            {"sheets": {"zr1_rule_mapping": {"headers": [], "rows": []}}},
            "zr1",
            {"rule_mapping_sheet": {"sheetName": "zr1_rule_mapping", "headers": []}},
            {},
            set(),
            set(),
            relationship_result,
            [resolution],
        )

        self.assertIn(subject["subjectId"], consumed)
        self.assertEqual(overrides[evidence_id], "resolved_not_applicable")
        self.assertEqual(
            overrides["derived:relationship:proposal"],
            "resolved_not_applicable",
        )

    def test_comparator_coverage_disposition_is_scoped_by_target(self) -> None:
        evidence_id = "comparator:direct_rule:shared"
        relationship_dispositions = [
            {
                "featureId": evidence_id,
                "model": "zr1",
                "evidenceIds": [evidence_id],
                "disposition": "resolved_not_applicable",
            },
            {
                "featureId": evidence_id,
                "model": "zr1x",
                "evidenceIds": [evidence_id],
                "disposition": "proposed_exception",
            },
        ]
        comparator_artifact = {
            "targets": {
                model: {
                    "facts": [
                        {
                            "evidenceId": evidence_id,
                            "factType": "direct_rule",
                        }
                    ]
                }
                for model in ("zr1", "zr1x")
            }
        }

        ledger = compiler_module._source_feature_ledger(
            ["zr1", "zr1x"],
            {"candidates": [], "skippedRows": []},
            {"priceRows": [], "baseModelPriceRows": [], "skippedPriceRows": []},
            {"roles": {}},
            {"sheets": []},
            {},
            relationship_dispositions,
            comparator_artifact,
            set(),
            set(),
            set(),
            set(),
            set(),
            set(),
            {},
            {},
        )
        dispositions = {
            item["model"]: item["disposition"]
            for item in ledger
            if evidence_id in item["evidenceIds"]
            and item["featureId"].startswith("comparator:")
        }

        self.assertEqual(dispositions["zr1"], "resolved_not_applicable")
        self.assertEqual(dispositions["zr1x"], "exception_open")

    def test_choose_relationship_replaces_uses_canonical_excludes_runtime_action(self) -> None:
        option_payload, price_payload, join_report, first, subject = (
            self.compile_with_ready_comparator_relationship()
        )
        options_by_rpo = {
            str(row["values"].get("rpo") or "").upper(): str(row["values"].get("option_id") or "")
            for row in first["canonical-row-manifest.json"]["rows"]
            if row["family"] == "options" and row["status"] == "ready"
        }
        proposal = subject["proposedRows"][0]
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "choose_relationship",
            "payload": {
                "sourceOptionId": options_by_rpo[proposal["sourceRpo"]],
                "ruleType": "replaces",
                "targetOptionId": options_by_rpo[proposal["targetRpo"]],
            },
            "disposition": "resolved",
        }
        second = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            resolution_entries=[resolution],
        )
        rule = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "rule_mapping"
            and any(
                dependency["evidenceId"] == f"resolution:{subject['subjectId']}"
                for dependency in row["evidenceDependencies"]
            )
        )
        self.assertEqual(rule["values"]["rule_type"], "excludes")
        self.assertEqual(rule["values"]["runtime_action"], "replace")

    def test_comparator_group_exclusive_and_price_confirmations_materialize_rows(self) -> None:
        initial = self.compile()
        proposal_subjects = [
            subject
            for subject in initial["exception-queue.json"]["subjects"]
            if subject["reasonCode"] in {
                "comparator_only_rule_group_proposal",
                "comparator_only_exclusive_group_proposal",
                "comparator_only_price_rule_proposal",
            }
        ]
        required_rpos = {
            str(value).upper()
            for subject in proposal_subjects
            for proposal in subject["proposedRows"]
            for value in (
                proposal.get("sourceRpo"),
                proposal.get("conditionRpo"),
                proposal.get("targetRpo"),
                *(proposal.get("memberRpos") or []),
            )
            if str(value or "")
        }
        option_payload = copy.deepcopy(self.option_payload)
        price_payload = copy.deepcopy(self.price_payload)
        price_payload["priceRows"] = [
            row
            for row in price_payload["priceRows"]
            if str(row.get("rpo") or "").upper() not in required_rpos
        ]
        for candidate in option_payload["candidates"]:
            if str(candidate.get("rpo") or "").upper() in required_rpos:
                candidate["priceMatch"] = {"status": "none"}
                candidate.pop("priceRows", None)
        join_report = join_prices(option_payload["candidates"], price_payload["priceRows"])
        first = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
        )
        price_proposal = next(
            subject["proposedRows"][0]
            for subject in first["exception-queue.json"]["subjects"]
            if subject["reasonCode"] == "comparator_only_price_rule_proposal"
        )
        option_ids_by_rpo = {
            str(row["values"].get("rpo") or "").upper(): str(row["values"].get("option_id") or "")
            for row in first["canonical-row-manifest.json"]["rows"]
            if row["family"] == "options" and row["status"] == "ready"
        }
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        sheet = workbook["z06_price_rules"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        mismatched = {
            "price_rule_id": "price_variant_mismatch",
            "condition_option_id": option_ids_by_rpo[str(price_proposal["conditionRpo"]).upper()],
            "price_rule_type": price_proposal["priceRuleType"],
            "target_option_id": option_ids_by_rpo[str(price_proposal["targetRpo"]).upper()],
            "price_value": 1,
            "body_style_scope": "coupe",
            "trim_level_scope": "1lz",
            "variant_scope": "different_variant",
        }
        sheet.append([mismatched.get(header, "") for header in headers])
        workbook.save(self.master)
        workbook.close()
        subjects = {
            subject["reasonCode"]: subject
            for subject in first["exception-queue.json"]["subjects"]
            if subject["reasonCode"] in {
                "comparator_only_rule_group_proposal",
                "comparator_only_exclusive_group_proposal",
                "comparator_only_price_rule_proposal",
            }
        }
        self.assertEqual(len(subjects), 3)
        resolutions = []
        for reason, subject in subjects.items():
            payload: dict[str, object] = {"decision": "confirm_proposal"}
            if reason == "comparator_only_exclusive_group_proposal":
                payload["selectionMode"] = "single_within_group"
            if reason == "comparator_only_price_rule_proposal":
                payload.update(
                    {
                        "priceValue": 995,
                        "bodyStyleScope": "coupe",
                        "trimLevelScope": "1lz",
                        "variantScope": "*",
                    }
                )
            resolutions.append(
                {
                    "subjectId": subject["subjectId"],
                    "subjectVersion": subject["subjectVersion"],
                    "action": "provide_typed_value",
                    "payload": payload,
                    "disposition": "resolved",
                }
            )

        second = self.compile(
            option_payload=option_payload,
            price_payload=price_payload,
            join_report=join_report,
            resolution_entries=resolutions,
        )
        resolved_subject_ids = {subject["subjectId"] for subject in subjects.values()}
        self.assertTrue(
            resolved_subject_ids
            <= {
                subject["subjectId"]
                for subject in second["exception-queue.json"]["subjects"]
            }
        )
        self.assertTrue(
            resolved_subject_ids
            <= {
                entry["subjectId"]
                for entry in second["exception-resolutions.json"]["validEntries"]
            }
        )
        self.assertTrue(
            resolved_subject_ids.isdisjoint(
                entry["subjectId"]
                for entry in second["exception-resolutions.json"]["supersededEntries"]
            )
        )
        rows = second["canonical-row-manifest.json"]["rows"]
        self.assertTrue(
            all(row["family"] in EDITOR_SHEET_META for row in rows),
            sorted({row["family"] for row in rows} - set(EDITOR_SHEET_META)),
        )
        for family in (
            "rule_groups",
            "rule_group_members",
            "exclusive_groups",
            "exclusive_members",
            "price_rules",
        ):
            with self.subTest(family=family):
                self.assertTrue(
                    any(
                        row["family"] == family
                        and row["status"] == "ready"
                        and any(
                            dep["evidenceId"].startswith("resolution:")
                            for dep in row["evidenceDependencies"]
                        )
                        for row in rows
                    )
                )
        price = next(
            row
            for row in rows
            if row["family"] == "price_rules"
            and any(dep["evidenceId"].startswith("resolution:") for dep in row["evidenceDependencies"])
        )
        self.assertEqual(price["values"]["price_value"], 995)
        self.assertNotEqual(price["values"]["price_rule_id"], "price_variant_mismatch")
        self.assertEqual(price["values"]["variant_scope"], "")
        blocker_ids = {
            blocker["subjectId"]
            for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
        }
        self.assertTrue(
            {subject["subjectId"] for subject in subjects.values()}.isdisjoint(blocker_ids)
        )
        for subject in subjects.values():
            for evidence_id in subject["evidenceReferences"]:
                ledger_entries = [
                    item
                    for item in second["compile-report.json"]["sourceFeatureCoverage"]
                    if evidence_id in item["evidenceIds"]
                ]
                self.assertEqual(len(ledger_entries), 1)
                self.assertEqual(ledger_entries[0]["disposition"], "compiled")
                report_entry = next(
                    item
                    for item in second["compile-report.json"]["comparatorDispositions"]
                    if item["target"] == "zr1" and item["evidenceId"] == evidence_id
                )
                self.assertEqual(report_entry["disposition"], "corroborated_target_match")

    def test_comparator_confirmations_remain_pending_when_any_option_endpoint_is_blocked(self) -> None:
        first = self.compile()
        subjects = [
            subject
            for subject in first["exception-queue.json"]["subjects"]
            if subject["reasonCode"] in {
                "comparator_only_rule_group_proposal",
                "comparator_only_exclusive_group_proposal",
                "comparator_only_price_rule_proposal",
            }
        ]
        self.assertEqual(len(subjects), 3)
        resolutions = []
        for subject in subjects:
            reason = subject["reasonCode"]
            payload: dict[str, object] = {"decision": "confirm_proposal"}
            if reason == "comparator_only_exclusive_group_proposal":
                payload["selectionMode"] = "single_within_group"
            if reason == "comparator_only_price_rule_proposal":
                payload.update(
                    {
                        "priceValue": 995,
                        "bodyStyleScope": "*",
                        "trimLevelScope": "*",
                        "variantScope": "*",
                    }
                )
            resolutions.append(
                {
                    "subjectId": subject["subjectId"],
                    "subjectVersion": subject["subjectVersion"],
                    "action": "provide_typed_value",
                    "payload": payload,
                    "disposition": "resolved",
                }
            )
        second = self.compile(resolution_entries=resolutions)
        blocker_ids = {
            blocker["subjectId"]
            for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
        }
        self.assertTrue({subject["subjectId"] for subject in subjects} <= blocker_ids)
        self.assertFalse(
            any(
                dependency["evidenceId"].startswith("resolution:")
                for row in second["canonical-row-manifest.json"]["rows"]
                for dependency in row["evidenceDependencies"]
            )
        )

    def test_comparator_default_confirmation_requires_and_uses_target_priority(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["default_selection_rules"].append(
            [
                "z06",
                "z06_default_bv4",
                "opt_bv4_001",
                "always",
                "",
                "",
                "",
                "",
                1,
                True,
                "Comparator context only",
                "default_selected",
            ]
        )
        workbook.save(self.master)
        workbook.close()
        comparator = build_comparator_evidence(
            self.master,
            self.selection["comparators"],
            run_authority_fingerprint=self.authority,
        )

        first = self.compile(comparator_artifact=comparator)
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "comparator_only_default_selection_proposal"
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "provide_typed_value",
            "payload": {
                "decision": "confirm_proposal",
                "priority": 40,
                "displayBehavior": "default_selected",
            },
            "disposition": "resolved",
        }

        second = self.compile(
            comparator_artifact=comparator,
            resolution_entries=[resolution],
        )
        default = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "default_selection_rules"
            and row["values"].get("model_key") == "zr1"
            and any(
                dep["evidenceId"] == f"resolution:{subject['subjectId']}"
                for dep in row["evidenceDependencies"]
            )
        )
        self.assertEqual(default["values"]["priority"], 40)
        self.assertEqual(default["values"]["display_behavior"], "default_selected")
        self.assertNotIn(
            subject["subjectId"],
            {
                blocker["subjectId"]
                for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )

    def test_comparator_default_confirmation_stays_pending_for_blocked_target_option(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["default_selection_rules"].append(
            [
                "z06",
                "z06_default_pdb",
                "opt_pdb_001",
                "always",
                "",
                "",
                "",
                "",
                1,
                True,
                "Comparator context only",
                "default_selected",
            ]
        )
        workbook.save(self.master)
        workbook.close()
        comparator = build_comparator_evidence(
            self.master,
            self.selection["comparators"],
            run_authority_fingerprint=self.authority,
        )
        first = self.compile(comparator_artifact=comparator)
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "comparator_only_default_selection_proposal"
            and item["proposedRows"][0]["targetRpo"] == "PDB"
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "provide_typed_value",
            "payload": {
                "decision": "confirm_proposal",
                "priority": 40,
                "displayBehavior": "default_selected",
            },
            "disposition": "resolved",
        }
        second = self.compile(
            comparator_artifact=comparator,
            resolution_entries=[resolution],
        )
        self.assertIn(
            subject["subjectId"],
            {
                blocker["subjectId"]
                for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )
        self.assertFalse(
            any(
                dependency["evidenceId"] == f"resolution:{subject['subjectId']}"
                for row in second["canonical-row-manifest.json"]["rows"]
                for dependency in row["evidenceDependencies"]
            )
        )

    def test_every_source_feature_and_family_has_one_disposition(self) -> None:
        result = self.compile()
        report = result["compile-report.json"]
        feature_ids = [item["featureId"] for item in report["sourceFeatureCoverage"]]
        self.assertEqual(len(feature_ids), len(set(feature_ids)))
        self.assertTrue(
            {item["disposition"] for item in report["sourceFeatureCoverage"]}
            <= {
                "compiled",
                "retained_existing",
                "exception_open",
                "resolved_not_a_workbook_fact",
                "resolved_not_applicable",
                "allowed_deferral",
                "unsupported_blocker",
            }
        )
        self.assertTrue(
            any(
                "Color and Trim 1" in item["featureId"]
                and item["disposition"] == "resolved_not_a_workbook_fact"
                and any(
                    evidence_id.startswith("source-sheet-content:Color and Trim 1:")
                    for evidence_id in item["evidenceIds"]
                )
                for item in report["sourceFeatureCoverage"]
            )
        )
        color_subjects = [
            subject
            for subject in result["exception-queue.json"]["subjects"]
            if "Color and Trim 1" in subject.get("evidenceReferences", [])
        ]
        self.assertEqual(color_subjects, [])
        family_ids = [item["featureId"] for item in report["familyCoverage"]]
        self.assertEqual(len(family_ids), len(set(family_ids)))
        self.assertTrue(any(item["family"] == "interiors" for item in report["familyCoverage"]))
        self.assertTrue(
            any(
                item["family"] in {"price_rules", "variant_overrides"}
                and item["disposition"] == "explicit_empty"
                for item in report["familyCoverage"]
            )
        )
        for item in report["familyCoverage"]:
            if item["disposition"] != "retained_existing":
                continue
            self.assertTrue(
                any(
                    row["family"] == item["family"]
                    and row["model"] in {item["model"], "*"}
                    and row.get("disposition") == "retained_existing"
                    for row in result["canonical-row-manifest.json"]["rows"]
                ),
                item,
            )
        self.assertEqual(result["canonical-row-manifest.json"]["modelModes"]["zr1"], "reprocess")
        self.assertEqual(report["models"]["zr1"]["mode"], "reprocess")
        self.assertEqual(report["incomingReferenceImpact"]["status"], "no_deletions_proposed")
        self.assertTrue(report["incomingReferenceImpact"]["graphBuilt"])
        self.assertGreater(report["incomingReferenceImpact"]["referenceEdges"], 0)
        self.assertTrue(report["manifestCounts"])
        compiled_status_ids = {
            item["featureId"]
            for item in report["sourceFeatureCoverage"]
            if item["family"] == "ovs" and item["disposition"] == "compiled"
        }
        emitted_status_ids = {
            dependency["evidenceId"]
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["family"] == "ovs" and row["status"] == "ready"
            for dependency in row["evidenceDependencies"]
            if str(dependency["evidenceId"]).startswith("status:")
        }
        self.assertEqual(compiled_status_ids, emitted_status_ids)
        self.assertTrue(
            all(
                row["values"].get("runtime_action") in {"", "replace"}
                for row in result["canonical-row-manifest.json"]["rows"]
                if row["family"] == "rule_mapping" and row["status"] == "ready"
            )
        )
        rule_signatures = {
            (row["model"], semantic_hash(row["semanticSignature"]))
            for row in result["canonical-row-manifest.json"]["rows"]
            if row["family"] == "rule_mapping" and row["status"] == "ready"
        }
        facts = {
            (target, fact["evidenceId"]): fact
            for target, entry in result["comparator-evidence.json"]["targets"].items()
            for fact in entry["facts"]
        }
        for item in report["comparatorDispositions"]:
            if item["disposition"] != "corroborated_target_match":
                continue
            fact = facts[(item["target"], item["evidenceId"])]
            signature = fact["signature"]
            normalized = {
                "sourceRpo": signature.get("sourceRpo"),
                "ruleType": signature.get("ruleType"),
                "targetRpo": signature.get("targetRpo"),
                "bodyStyleScope": signature.get("bodyStyleScope") or "*",
                "trimLevelScope": signature.get("trimLevelScope") or "*",
                "variantScope": signature.get("variantScope") or "*",
            }
            fingerprint = semantic_hash(normalized)
            self.assertTrue(
                (item["target"], fingerprint) in rule_signatures
                or ("*", fingerprint) in rule_signatures,
                item,
            )
        for field in (
            "targetEvidenceFingerprint",
            "comparatorEvidenceFingerprint",
            "phraseEvidenceFingerprint",
            "workbookEvidenceFingerprint",
        ):
            self.assertEqual(result["canonical-row-manifest.json"][field], result["exception-queue.json"][field])
        self.assertTrue(result["canonical-row-manifest.json"]["workbookEvidenceFingerprint"])
        self.assertTrue(all(result["canonical-row-manifest.json"]["targetEvidenceFingerprint"].values()))
        self.assertNotIn(
            semantic_hash([]),
            result["canonical-row-manifest.json"]["targetEvidenceFingerprint"].values(),
        )
        self.assertTrue(
            all(
                evidence_id == evidence_id.strip()
                for evidence_id in result["canonical-row-manifest.json"]["workbookEvidenceFingerprint"]
            )
        )
        for row in result["canonical-row-manifest.json"]["rows"]:
            for column, kind in EDITOR_SHEET_META.get(row["family"], {}).get("types", {}).items():
                if kind == "int" and row["values"].get(column) not in (None, ""):
                    self.assertIsInstance(row["values"][column], int)
                    self.assertNotIsInstance(row["values"][column], bool)


    def test_missing_section_and_conditional_price_are_typed_blockers(self) -> None:
        result = self.compile()
        reasons = {item["reasonCode"] for item in result["exception-queue.json"]["subjects"]}
        self.assertIn("missing_section", reasons)
        self.assertIn("unresolved_price_scope", reasons)
        self.assertEqual(result["compile-report.json"]["models"]["zr1"]["compileReady"], False)
        self.assertFalse(result["compile-report.json"]["models"]["zr1"]["planReady"])

    def test_artifact_graph_rejects_invalid_and_conflicting_current_resolutions(self) -> None:
        result = self.compile()
        malformed_queue = copy.deepcopy(result["exception-queue.json"])
        malformed_queue["subjects"][0]["reasonCode"] = "unknown_reason"
        malformed_queue["subjects"][0]["allowedActions"] = ["choose_section"]
        with self.assertRaisesRegex(ValueError, "does not support actions"):
            validate_artifact_graph(
                result["canonical-row-manifest.json"],
                result["compile-report.json"],
                result["comparator-evidence.json"],
                malformed_queue,
                result["exception-resolutions.json"],
            )
        cyclic_queue = copy.deepcopy(result["exception-queue.json"])
        cyclic_queue["resolutionSemanticSha"] = result["exception-resolutions.json"]["resolutionSemanticSha"]
        with self.assertRaisesRegex(ValueError, "may not depend on resolution state"):
            validate_artifact_graph(
                result["canonical-row-manifest.json"],
                result["compile-report.json"],
                result["comparator-evidence.json"],
                cyclic_queue,
                result["exception-resolutions.json"],
            )
        subject = next(
            item
            for item in result["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "missing_section"
        )
        invalid_entry = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "choose_section",
            "payload": {"sectionId": 7},
            "disposition": "resolved",
        }
        invalid = copy.deepcopy(result["exception-resolutions.json"])
        invalid["entries"] = [invalid_entry]
        invalid["validEntries"] = [invalid_entry]
        with self.assertRaisesRegex(ValueError, "non-empty string sectionId"):
            validate_artifact_graph(
                result["canonical-row-manifest.json"],
                result["compile-report.json"],
                result["comparator-evidence.json"],
                result["exception-queue.json"],
                invalid,
            )
        first = {**invalid_entry, "payload": {"sectionId": "sec_a"}}
        second = {**invalid_entry, "payload": {"sectionId": "sec_b"}}
        conflicting = copy.deepcopy(result["exception-resolutions.json"])
        conflicting["entries"] = [first, second]
        conflicting["validEntries"] = [first, second]
        with self.assertRaisesRegex(ValueError, "Conflicting current resolutions"):
            validate_artifact_graph(
                result["canonical-row-manifest.json"],
                result["compile-report.json"],
                result["comparator-evidence.json"],
                result["exception-queue.json"],
                conflicting,
            )

    def test_choose_section_resolution_is_consumed_into_manifest(self) -> None:
        first = self.compile()
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "missing_section"
        )
        candidate_id = subject["evidenceReferences"][0]
        candidate = next(
            item
            for item in self.option_payload["candidates"]
            if option_occurrence_signature(
                {**item, "statuses": model_scoped_statuses(item, "zr1")}
            )
            == candidate_id
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "choose_section",
            "payload": {"sectionId": "sec_whee_001"},
            "disposition": "resolved",
        }
        second = self.compile(resolution_entries=[resolution])
        option = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "options"
            and row["values"].get("rpo") == candidate.get("rpo")
            and row["values"].get("section_id") == "sec_whee_001"
        )
        self.assertEqual(option["status"], "blocked")
        self.assertTrue(
            any(
                item["reasonCode"] == "copy_review_required"
                and candidate_id in item.get("evidenceReferences", [])
                for item in second["exception-queue.json"]["subjects"]
            )
        )
        blocker_ids = {
            item["subjectId"]
            for item in second["compile-report.json"]["models"]["zr1"]["blockers"]
        }
        self.assertNotIn(subject["subjectId"], blocker_ids)

    def test_missing_section_can_be_resolved_by_omitting_the_source_option(self) -> None:
        option_payload = copy.deepcopy(self.option_payload)
        omitted_candidate = next(
            item for item in option_payload["candidates"] if item.get("rpo") == "CC3"
        )
        omitted_candidate["description"] = (
            "Roof panel, transparent\n1. Requires (CC2) painted roof panel."
        )
        join_report = join_prices(
            option_payload["candidates"], self.price_payload["priceRows"]
        )
        first = self.compile(
            option_payload=option_payload,
            join_report=join_report,
        )
        candidate_id = option_occurrence_signature(
            {
                **omitted_candidate,
                "statuses": model_scoped_statuses(omitted_candidate, "zr1"),
            }
        )
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "missing_section"
            and item["evidenceReferences"] == [candidate_id]
        )
        candidate = omitted_candidate
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "mark_not_applicable",
            "payload": {"reason": "Reviewer omitted this target option."},
            "disposition": "resolved_not_applicable",
        }

        second = self.compile(
            option_payload=option_payload,
            join_report=join_report,
            resolution_entries=[resolution],
        )

        self.assertFalse(
            any(
                row["family"] == "options"
                and row["values"].get("rpo") == candidate.get("rpo")
                for row in second["canonical-row-manifest.json"]["rows"]
            )
        )
        self.assertNotIn(
            subject["subjectId"],
            {
                blocker["subjectId"]
                for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )
        self.assertIn(
            subject["subjectId"],
            {
                entry["subjectId"]
                for entry in second["exception-resolutions.json"]["validEntries"]
            },
        )
        omitted_status_coverage = [
            item
            for item in second["compile-report.json"]["sourceFeatureCoverage"]
            if item["family"] == "ovs"
            and str(item["featureId"]).startswith(
                f"status:{option_occurrence_signature(candidate)}:"
            )
        ]
        self.assertTrue(omitted_status_coverage)
        self.assertEqual(
            {item["disposition"] for item in omitted_status_coverage},
            {"resolved_not_applicable"},
        )
        omitted_relationship_coverage = [
            item
            for item in second["compile-report.json"]["sourceFeatureCoverage"]
            if item["family"] == "rule_mapping"
            and f"candidate:{option_occurrence_signature(candidate)}"
            in item["evidenceIds"]
        ]
        self.assertTrue(omitted_relationship_coverage)
        self.assertEqual(
            {item["disposition"] for item in omitted_relationship_coverage},
            {"resolved_not_applicable"},
        )

    def test_missing_section_can_keep_an_inactive_nonselectable_unpriced_option(self) -> None:
        first = self.compile()
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "missing_section"
        )
        candidate_id = subject["evidenceReferences"][0]
        candidate = next(
            item
            for item in self.option_payload["candidates"]
            if option_occurrence_signature(
                {**item, "statuses": model_scoped_statuses(item, "zr1")}
            )
            == candidate_id
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "keep_inactive_option",
            "payload": {"sectionId": "sec_whee_001"},
            "disposition": "resolved",
        }

        second = self.compile(resolution_entries=[resolution])

        option = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "options"
            and row["values"].get("rpo") == candidate.get("rpo")
        )
        self.assertEqual(option["status"], "blocked")
        self.assertEqual(option["values"]["section_id"], "sec_whee_001")
        self.assertEqual(option["values"]["price"], "")
        self.assertFalse(option["values"]["active"])
        self.assertFalse(option["values"]["selectable"])
        self.assertTrue(
            any(
                item["reasonCode"] == "copy_review_required"
                and candidate_id in item.get("evidenceReferences", [])
                for item in second["exception-queue.json"]["subjects"]
            )
        )
        self.assertNotIn(
            subject["subjectId"],
            {
                blocker["subjectId"]
                for blocker in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )

    def test_typed_price_scope_resolution_is_consumed_into_price_rule(self) -> None:
        first = self.compile()
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "unresolved_price_scope"
        )
        resolution = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "provide_typed_value",
            "payload": {"trimLevelScope": "1LT", "priceValue": 1234},
            "disposition": "resolved",
        }
        second = self.compile(resolution_entries=[resolution])
        price_rule = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "price_rules"
            and row["values"].get("trim_level_scope") == "1LT"
            and row["values"].get("price_value") == 1234
        )
        self.assertEqual(price_rule["status"], "ready")
        self.assertEqual(
            price_rule["values"]["condition_option_id"],
            price_rule["values"]["target_option_id"],
        )
        blocker_ids = {
            item["subjectId"]
            for item in second["compile-report.json"]["models"]["zr1"]["blockers"]
        }
        self.assertNotIn(subject["subjectId"], blocker_ids)

    def test_current_resolution_projects_when_stale_history_has_same_subject_id(self) -> None:
        first = self.compile()
        subject = next(
            item
            for item in first["exception-queue.json"]["subjects"]
            if item["reasonCode"] == "unresolved_price_scope"
        )
        stale = {
            "subjectId": subject["subjectId"],
            "subjectVersion": "superseded-subject-version",
            "action": "provide_typed_value",
            "payload": {"priceValue": 695},
            "disposition": "resolved",
        }
        current = {
            "subjectId": subject["subjectId"],
            "subjectVersion": subject["subjectVersion"],
            "action": "provide_typed_value",
            "payload": {
                "bodyStyleScope": "*",
                "trimLevelScope": "*",
                "variantScope": "*",
                "priceValue": 695,
            },
            "disposition": "resolved",
        }

        second = self.compile(resolution_entries=[stale, current])

        price_rule = next(
            row
            for row in second["canonical-row-manifest.json"]["rows"]
            if row["family"] == "price_rules"
            and row["values"].get("body_style_scope") == "*"
            and row["values"].get("trim_level_scope") == "*"
            and row["values"].get("variant_scope") == "*"
            and row["values"].get("price_value") == 695
        )
        self.assertEqual(price_rule["status"], "ready")
        self.assertNotIn(
            subject["subjectId"],
            {
                item["subjectId"]
                for item in second["compile-report.json"]["models"]["zr1"]["blockers"]
            },
        )

    def test_artifacts_are_deterministic_and_resolution_independent_queue(self) -> None:
        first = self.compile()
        second = self.compile()
        for name in ("comparator-evidence.json", "canonical-row-manifest.json", "exception-queue.json", "exception-resolutions.json", "compile-report.json"):
            self.assertEqual(canonical_bytes(first[name]), canonical_bytes(second[name]), name)
        self.assertNotIn("resolutionSemanticSha", canonical_bytes(first["exception-queue.json"]).decode())
        projected_options = [
            row["values"]
            for row in first["canonical-row-manifest.json"]["rows"]
            if row["model"] == "zr1"
            and row["family"] == "options"
            and row["action"] != "delete"
        ]
        for section_id in {str(row.get("section_id") or "") for row in projected_options}:
            orders = [
                int(row["display_order"])
                for row in projected_options
                if str(row.get("section_id") or "") == section_id
                and str(row.get("display_order") or "").isdigit()
            ]
            self.assertEqual(len(orders), len(set(orders)), section_id)

    def test_compiler_refuses_join_report_drift(self) -> None:
        drifted = copy.deepcopy(self.join_report)
        drifted["exactMatches"] += 1
        with self.assertRaisesRegex(ValueError, "Join report exactMatches"):
            self.compile(join_report=drifted)

    def test_compiler_refuses_stale_options_quality_allowlist_authority(self) -> None:
        stale_authority = copy.deepcopy(self.authority)
        stale_authority["bindings"]["optionsSheetQualityAllowlist"]["sha256"] = "0" * 64
        stale_authority["fingerprint"] = hashlib.sha256(
            canonical_bytes(stale_authority["bindings"])
        ).hexdigest()

        with self.assertRaisesRegex(ValueError, "quality allowlist path and bytes"):
            self.compile(run_authority_fingerprint=stale_authority)

    def test_complete_projected_options_quality_blocks_retained_invalid_row(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.master)
        workbook["zr1_options"].append(
            [
                "opt_bad_001",
                "BAD",
                0,
                "LPO",
                "Retained invalid fixture",
                "",
                "sec_whee_001",
                True,
                "",
                True,
                "",
            ]
        )
        workbook.save(self.master)
        workbook.close()

        result = self.compile()

        quality_subjects = [
            subject
            for subject in result["exception-queue.json"]["subjects"]
            if subject["reasonCode"] == "projected_options_quality"
            and subject["proposedRows"][0].get("optionId") == "opt_bad_001"
        ]
        self.assertEqual(
            {subject["proposedRows"][0]["checkId"] for subject in quality_subjects},
            {"active_option_missing_display_order", "bare_lpo_option_name"},
        )
        self.assertTrue(all(subject["allowedActions"] == [] for subject in quality_subjects))
        self.assertFalse(result["compile-report.json"]["models"]["zr1"]["compileReady"])

    def test_source_row_reorder_preserves_option_semantics(self) -> None:
        reordered = copy.deepcopy(self.option_payload)
        reordered["candidates"] = list(reversed(reordered["candidates"]))
        first = self.compile()["canonical-row-manifest.json"]
        second = self.compile(option_payload=reordered)["canonical-row-manifest.json"]
        first_options = {(row["values"].get("rpo"), semantic_hash(row["semanticSignature"])) for row in first["rows"] if row["family"] == "options"}
        second_options = {(row["values"].get("rpo"), semantic_hash(row["semanticSignature"])) for row in second["rows"] if row["family"] == "options"}
        self.assertEqual(first_options, second_options)

    def test_price_row_reorder_preserves_all_semantic_hashes(self) -> None:
        first = self.compile()
        reordered = copy.deepcopy(self.price_payload)
        reordered["priceRows"] = list(reversed(reordered["priceRows"]))
        reordered["baseModelPriceRows"] = list(reversed(reordered["baseModelPriceRows"]))
        second = self.compile(price_payload=reordered)
        self.assertEqual(
            first["exception-queue.json"]["queueSubjectFingerprint"],
            second["exception-queue.json"]["queueSubjectFingerprint"],
        )
        self.assertEqual(
            first["canonical-row-manifest.json"]["manifestSemanticSha"],
            second["canonical-row-manifest.json"]["manifestSemanticSha"],
        )
        self.assertEqual(
            first["compile-report.json"]["compileReportSemanticSha"],
            second["compile-report.json"]["compileReportSemanticSha"],
        )

    def test_comparator_fact_reorder_preserves_compile_report_semantics(self) -> None:
        first = self.compile()
        reordered = copy.deepcopy(self.comparator)
        for entry in reordered["targets"].values():
            entry["facts"] = list(reversed(entry["facts"]))
        second = self.compile(comparator_artifact=reordered)
        self.assertEqual(
            first["compile-report.json"]["compileReportSemanticSha"],
            second["compile-report.json"]["compileReportSemanticSha"],
        )

    def test_source_coordinate_drift_preserves_subject_and_derivation_semantics(self) -> None:
        shifted = copy.deepcopy(self.option_payload)
        for index, candidate in enumerate(shifted["candidates"], start=100):
            sheet = str(candidate["candidateId"]).split(":", 1)[0]
            candidate["candidateId"] = f"{sheet}:{index}"
            candidate["rowIndex"] = index
        first = self.compile()
        second = self.compile(option_payload=shifted)
        self.assertEqual(
            first["canonical-row-manifest.json"]["manifestSemanticSha"],
            second["canonical-row-manifest.json"]["manifestSemanticSha"],
        )
        self.assertEqual(
            first["exception-queue.json"]["queueSubjectFingerprint"],
            second["exception-queue.json"]["queueSubjectFingerprint"],
        )
        self.assertEqual(
            first["compile-report.json"]["compileReportSemanticSha"],
            second["compile-report.json"]["compileReportSemanticSha"],
        )

    def test_comparator_price_never_populates_target_option(self) -> None:
        result = self.compile()
        bv4 = next(row for row in result["canonical-row-manifest.json"]["rows"] if row["family"] == "options" and row["values"].get("rpo") == "BV4")
        self.assertEqual(bv4["values"]["price"], 395)
        self.assertNotEqual(bv4["values"]["price"], 500)


if __name__ == "__main__":
    unittest.main()
