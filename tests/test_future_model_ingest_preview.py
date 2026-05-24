#!/usr/bin/env python3
"""Tests for non-mutating future-model archive ingestion previews."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SPECS,
    build_future_model_preview,
    build_preview_for_model,
    build_section_candidates,
    normalize_status,
    resolve_section,
)

OPTION_HEADERS = [
    "option_id",
    "rpo",
    "price",
    "option_name",
    "description",
    "detail_raw",
    "section_id",
    "selectable",
    "display_order",
    "active",
    "display_behavior",
]
OVS_HEADERS = ["option_id", "variant_id", "status"]
ARCHIVE_HEADERS = [
    "RPO",
    "Price",
    "Option Name",
    "Description",
    "Detail",
    "Category",
    "",
    "1LZ Coupe",
    "2LZ Coupe",
    "3LZ Coupe",
    "1LZ Convertible",
    "2LZ Convertible",
    "3LZ Convertible",
]


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows or []:
        ws.append([row.get(header, None) for header in headers])


def minimal_preview_workbook(*, archive_rows: list[dict[str, object]] | None = None) -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(wb, "stingray_options", OPTION_HEADERS, [
        {
            "option_id": "opt_exact",
            "rpo": "ABC",
            "option_name": "Exact Match Name",
            "section_id": "sec_exact",
            "active": True,
            "display_behavior": "card",
        },
        {
            "option_id": "opt_conflict_one",
            "rpo": "MULTI",
            "option_name": "First Multi Name",
            "section_id": "sec_alpha",
            "active": True,
        },
    ])
    append_sheet(wb, "grandSport_options", OPTION_HEADERS, [
        {
            "option_id": "opt_conflict_two",
            "rpo": "MULTI",
            "option_name": "Second Multi Name",
            "section_id": "sec_beta",
            "active": True,
        },
        {
            "option_id": "opt_inactive_ignored",
            "rpo": "INACTIVE",
            "option_name": "Inactive Should Not Resolve",
            "section_id": "sec_inactive",
            "active": False,
        },
    ])
    append_sheet(wb, "section_master", ["section_id", "active"], [
        {"section_id": "sec_exact", "active": True},
        {"section_id": "sec_alpha", "active": True},
        {"section_id": "sec_beta", "active": True},
    ])
    append_sheet(wb, "archive_Z06_Ingest", ARCHIVE_HEADERS, archive_rows or [])
    append_sheet(wb, "archive_ZR1_Ingest", ARCHIVE_HEADERS[:7] + ["1LZ Coupe", "3LZ Coupe", "1LZ Convertible", "3LZ Convertible"], [])
    append_sheet(wb, "archive_ZR1X_Ingest", ARCHIVE_HEADERS[:7] + ["1LZ Coupe", "3LZ Coupe", "1LZ Convertible", "3LZ Convertible"], [])
    append_sheet(wb, "z06_options", OPTION_HEADERS, [])
    append_sheet(wb, "z06_ovs", OVS_HEADERS, [])
    append_sheet(wb, "zr1_options", OPTION_HEADERS, [])
    append_sheet(wb, "zr1_ovs", OVS_HEADERS, [])
    append_sheet(wb, "zr1x_options", OPTION_HEADERS, [])
    append_sheet(wb, "zr1x_ovs", OVS_HEADERS, [])
    return wb


class FutureModelIngestPreviewTests(unittest.TestCase):
    def test_status_normalization_preserves_unknown_as_review_only(self) -> None:
        self.assertEqual(normalize_status("Standard"), ("standard", None))
        self.assertEqual(normalize_status("Available"), ("available", None))
        self.assertEqual(normalize_status("Not Available"), ("unavailable", None))
        self.assertEqual(normalize_status(""), (None, None))
        self.assertEqual(normalize_status("Dealer Installed"), (None, "unknown_status"))

    def test_future_model_variant_mappings_are_explicit(self) -> None:
        self.assertEqual(FUTURE_MODEL_SPECS["z06"].variant_columns["2LZ Convertible"], "2lz_h67")
        self.assertEqual(tuple(FUTURE_MODEL_SPECS["z06"].variant_columns.values()), (
            "1lz_h07",
            "2lz_h07",
            "3lz_h07",
            "1lz_h67",
            "2lz_h67",
            "3lz_h67",
        ))
        self.assertEqual(tuple(FUTURE_MODEL_SPECS["zr1"].variant_columns.values()), (
            "1lz_r07",
            "3lz_r07",
            "1lz_r67",
            "3lz_r67",
        ))
        self.assertEqual(tuple(FUTURE_MODEL_SPECS["zr1x"].variant_columns.values()), (
            "1lz_s07",
            "3lz_s07",
            "1lz_s67",
            "3lz_s67",
        ))

    def test_option_and_ovs_projection_keeps_source_shapes_and_price_semantics(self) -> None:
        wb = minimal_preview_workbook(
            archive_rows=[
                {
                    "RPO": "001",
                    "Price": "",
                    "Option Name": "Numeric Looking RPO",
                    "Description": "desc",
                    "Detail": "detail",
                    "Category": "Mechanical",
                    "1LZ Coupe": "Available",
                    "2LZ Coupe": "Standard",
                    "3LZ Coupe": "Not Available",
                },
                {
                    "RPO": "ZERO",
                    "Price": 0,
                    "Option Name": "Zero Price",
                    "Category": "Exterior",
                    "1LZ Coupe": "Standard",
                },
            ]
        )

        preview = build_preview_for_model(wb, FUTURE_MODEL_SPECS["z06"])

        first_option = preview["proposed_options"][0]
        self.assertEqual(first_option["rpo"], "001")
        self.assertIsNone(first_option["price"])
        self.assertEqual(first_option["option_name"], "Numeric Looking RPO")
        self.assertEqual(first_option["description"], "desc")
        self.assertEqual(first_option["detail_raw"], "detail")
        self.assertEqual(first_option["display_order"], 1)
        self.assertTrue(first_option["selectable"])
        self.assertTrue(first_option["active"])
        self.assertTrue(first_option["option_id"].startswith("opt_001_"))

        second_option = preview["proposed_options"][1]
        self.assertEqual(second_option["price"], 0)

        ovs = preview["proposed_ovs"][:3]
        self.assertEqual([row["variant_id"] for row in ovs], ["1lz_h07", "2lz_h07", "3lz_h07"])
        self.assertEqual([row["status"] for row in ovs], ["available", "standard", "unavailable"])
        self.assertEqual(ovs[0]["option_id"], first_option["option_id"])
        self.assertEqual(ovs[0]["source_variant_column"], "1LZ Coupe")

    def test_section_resolution_exact_conflict_and_unresolved(self) -> None:
        wb = minimal_preview_workbook()
        candidates = build_section_candidates(wb)

        exact = resolve_section({"rpo": "ABC", "option_name": "Exact Match Name"}, candidates)
        self.assertEqual(exact["section_resolution"], "resolved")
        self.assertEqual(exact["section_id"], "sec_exact")
        self.assertEqual(exact["display_behavior"], "card")

        conflict = resolve_section({"rpo": "MULTI", "option_name": "Different Name"}, candidates)
        self.assertEqual(conflict["section_resolution"], "conflict")
        self.assertEqual(conflict["section_id"], "")
        self.assertEqual(set(conflict["section_candidates"]), {"sec_alpha", "sec_beta"})

        unresolved = resolve_section({"rpo": "NOPE", "option_name": "No Match"}, candidates)
        self.assertEqual(unresolved["section_resolution"], "unresolved")
        self.assertEqual(unresolved["section_candidates"], [])

    def test_review_classifications_include_duplicates_missing_rpo_unknown_status_and_blank_variants(self) -> None:
        wb = minimal_preview_workbook(
            archive_rows=[
                {"RPO": "DUP", "Price": "abc", "Option Name": "First", "Category": "Exterior", "1LZ Coupe": "Available"},
                {"RPO": "DUP", "Option Name": "Second", "Category": "Exterior", "1LZ Coupe": "Dealer Installed"},
                {"RPO": "", "Option Name": "No RPO", "Category": "Interior"},
            ]
        )

        preview = build_preview_for_model(wb, FUTURE_MODEL_SPECS["z06"])
        summary = preview["summary"]

        self.assertGreaterEqual(summary["review_counts"]["duplicate_rpo"], 2)
        self.assertEqual(summary["review_counts"]["missing_rpo"], 1)
        self.assertEqual(summary["review_counts"]["unknown_status"], 1)
        self.assertGreaterEqual(summary["review_counts"]["blank_variant_status"], 1)
        self.assertEqual(summary["review_counts"]["price_type_issue"], 1)
        review_reasons = "\n".join(row["review_reason"] for row in preview["review_rows"])
        self.assertIn("duplicate_rpo", review_reasons)
        self.assertIn("unknown_status", review_reasons)

    def test_build_future_preview_does_not_modify_workbook_file_or_header_only_future_sheets(self) -> None:
        wb = minimal_preview_workbook(
            archive_rows=[{"RPO": "ABC", "Option Name": "Exact Match Name", "Category": "Exterior", "1LZ Coupe": "Available"}]
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "preview.xlsx"
            wb.save(path)
            before_mtime = path.stat().st_mtime_ns
            before = load_workbook(path, read_only=True, data_only=True)
            before_z06_rows = before["z06_options"].max_row
            before.close()

            opened = load_workbook(path, read_only=False, data_only=True)
            preview = build_future_model_preview(opened)
            opened.close()

            after_mtime = path.stat().st_mtime_ns
            after = load_workbook(path, read_only=True, data_only=True)
            after_z06_rows = after["z06_options"].max_row
            after.close()

        self.assertIn("z06", preview["models"])
        self.assertEqual(after_mtime, before_mtime)
        self.assertEqual(after_z06_rows, before_z06_rows)


if __name__ == "__main__":
    unittest.main()
