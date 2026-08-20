"""Pass 4 primary-runtime-only generated-contract acceptance regressions."""

from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from unittest import mock
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "workbook-manager" / "backend"
SCRIPTS = ROOT / "scripts"
for path in (BACKEND, SCRIPTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from app.contract_parity import (  # noqa: E402
    generate_contract_snapshot,
    promoted_runtime_models,
    validate_primary_runtime_parity,
)
from workbook_manager_fixtures import verified_manager_fixture  # noqa: E402

WORKBOOK = ROOT / "stingray_master.xlsx"


def _hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _protected_hashes() -> dict[str, str]:
    paths = [WORKBOOK, ROOT / "form-app" / "data.js"]
    paths.extend(
        sorted(path for path in (ROOT / "form-output").rglob("*") if path.is_file())
    )
    return {str(path.relative_to(ROOT)): _hash(path) for path in paths}


class TestWorkbookManagerGeneratedParity(unittest.TestCase):
    def test_generation_discovery_uses_canonical_default_context(self):
        class DiscoveryObserved(RuntimeError):
            pass

        def observe_discovery(workbook_path, **kwargs):
            self.assertEqual(Path(workbook_path), WORKBOOK)
            self.assertEqual(kwargs, {})
            raise DiscoveryObserved

        with tempfile.TemporaryDirectory(prefix="wbm-pass4-discovery-") as tempdir:
            with mock.patch.dict(
                generate_contract_snapshot.__globals__,
                {"discover_generation_model_configs": observe_discovery},
            ), self.assertRaises(DiscoveryObserved):
                generate_contract_snapshot(
                    WORKBOOK,
                    Path(tempdir),
                    "stingray",
                )

    def test_runtime_contract_drift_is_a_blocking_finding(self):
        with tempfile.TemporaryDirectory(prefix="wbm-pass4-drift-") as tempdir:
            root = Path(tempdir)
            source = root / "source.xlsx"
            reconstructed = root / "reconstructed.xlsx"
            source.write_bytes(b"source")
            reconstructed.write_bytes(b"reconstructed")

            def write_contract(workbook_path, output_root, _model_key, **_kwargs):
                output = Path(output_root) / "contract.json"
                output.parent.mkdir(parents=True, exist_ok=True)
                payload = {
                    "generated_at": "ignored",
                    "value": Path(workbook_path).parent.name,
                }
                output.write_text(json.dumps(payload), encoding="utf-8")
                return output

            with mock.patch.dict(
                validate_primary_runtime_parity.__globals__,
                {
                    "promoted_runtime_models": mock.Mock(
                        return_value=("stingray",)
                    ),
                    "generate_contract_snapshot": write_contract,
                },
            ):
                issues = validate_primary_runtime_parity(source, reconstructed, ROOT)

        self.assertEqual([issue["category"] for issue in issues], ["generated_contract_drift"])

    def test_strict_preflight_rejects_historical_artifact_type(self):
        with tempfile.TemporaryDirectory(prefix="wbm-pass4-preflight-") as tempdir:
            workbook_path = Path(tempdir) / "source.xlsx"
            shutil.copy2(WORKBOOK, workbook_path)
            workbook = load_workbook(workbook_path)
            sheet = workbook["model_registry_promotion"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            model_key_column = headers["model_key"]
            artifact_type_column = headers["artifact_type"]
            assert isinstance(model_key_column, int)
            assert isinstance(artifact_type_column, int)
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, model_key_column).value == "stingray":
                    sheet.cell(row, artifact_type_column).value = "current_generation"
                    break
            workbook.save(workbook_path)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "accepts only 'runtime_contract'"):
                promoted_runtime_models(workbook_path, ROOT)

    def test_source_and_identity_reconstruction_runtime_contracts_match(self):
        before = _protected_hashes()
        with tempfile.TemporaryDirectory(prefix="wbm-pass4-parity-") as tempdir:
            temp = Path(tempdir)
            source = temp / "source" / "candidate.xlsx"
            reconstruction = temp / "reconstruction" / "candidate.xlsx"
            source.parent.mkdir()
            reconstruction.parent.mkdir()
            fixture = verified_manager_fixture()
            fixture.clone_workbook(source)
            fixture.clone_unchanged_export(reconstruction)
            models = promoted_runtime_models(source, ROOT)
            self.assertEqual(
                models,
                (
                    "stingray",
                    "grand_sport",
                    "grand_sport_x",
                    "z06",
                    "zr1",
                    "zr1x",
                ),
            )

            for model_key in models:
                source_root = temp / "source-output" / model_key
                reconstructed_root = temp / "reconstructed-output" / model_key
                source_contract = generate_contract_snapshot(
                    source, source_root, model_key
                )
                reconstructed_contract = generate_contract_snapshot(
                    reconstruction,
                    reconstructed_root,
                    model_key,
                )
                subprocess.run(
                    [
                        "node",
                        str(ROOT / "scripts" / "compare-generated-contracts.mjs"),
                        str(source_contract),
                        str(reconstructed_contract),
                    ],
                    cwd=ROOT,
                    check=True,
                    capture_output=True,
                    text=True,
                )
                self.assertEqual(
                    [path.relative_to(source_root).as_posix() for path in source_root.rglob("*") if path.is_file()],
                    [f"form-output/runtime/{source_contract.name}"],
                )
                self.assertEqual(
                    [path.relative_to(reconstructed_root).as_posix() for path in reconstructed_root.rglob("*") if path.is_file()],
                    [f"form-output/runtime/{reconstructed_contract.name}"],
                )
        self.assertEqual(_protected_hashes(), before)


if __name__ == "__main__":
    unittest.main()
