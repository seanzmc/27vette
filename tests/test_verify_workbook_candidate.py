#!/usr/bin/env python3
"""Proofs required of the composed candidate lane (spec Pass 3 requirement 12).

Each test names the change it would catch. The lane is expensive — it generates
six models — so one canonical full run and one controlled-drift full run are
module-scoped fixtures. Stage/report mechanics use compact or early-failure
inputs instead of rebuilding an equivalent candidate.
"""

from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "stingray_master.xlsx"
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from verify_workbook_candidate import (  # noqa: E402
    HARNESS_DATA_JS_ENV,
    REPORT_SCHEMA_VERSION,
    STAGES,
    WORKBOOK_TRUTH_ENV,
    protected_surface_hashes,
    run_browser_harness,
    semantic_drift,
    verify_candidate,
)

REQUIRED_MODEL_FIELDS = {
    "model_key",
    "generated",
    "validation_findings",
    "contract_sha256",
    "declared_changed",
    "semantic_drift_vs_retained",
}

ALL_MODEL_KEYS = {"stingray", "grand_sport", "grand_sport_x", "z06", "zr1", "zr1x"}


# The lane generates six models per invocation, so every full run in this file
# is module-scoped and shared. Adding an unshared full run costs about a minute.
@pytest.fixture(scope="module")
def canonical_run(tmp_path_factory) -> dict:
    """The canonical workbook, all models declared, including browser proof."""

    report_path = tmp_path_factory.mktemp("canonical") / "readiness.json"
    report = verify_candidate(
        WORKBOOK, changed_models=["*"], report_path=report_path, run_harness=True
    )
    report["_report_path"] = str(report_path)
    return report


DRIFT_MODEL_KEY = "zr1"
RETAINED_DRIFT_CONTRACT = ROOT / "form-output" / "runtime" / "zr1-runtime-contract.json"


def _live_drift_probe() -> tuple[str, list[str]]:
    """An option id proven to reach the retained contract, and the collections it moves.

    Checkpoint 1 of the fast layered validation suite (spec §9) replaced the
    hardcoded EFR probe these canaries used. `zr1_options.opt_efr_001` is
    `active=True, selectable=False` and appears in NEITHER the choices nor the
    standardEquipment of the retained zr1 contract, so renaming it could not
    move anything: both forcing tests measured an empty drift set while
    asserting a non-empty one, and the `semantic_drift` stage the release gate
    depends on had no live positive proof. (Why an active workbook option row
    emits nothing is a separate workbook/generator question, recorded as
    `finding.dead_semantic_drift_canary`; §12 forbids answering it in test code.)

    The probe is chosen from the retained artifact itself, so it cannot go inert
    again silently — if no option reaches both collections, this raises instead
    of quietly asserting nothing.
    """

    contract = json.loads(RETAINED_DRIFT_CONTRACT.read_text())
    in_choices = {str(row.get("option_id") or "") for row in contract["choices"]}
    in_standard = {str(row.get("option_id") or "") for row in contract["standardEquipment"]}
    reaches_both = sorted(option_id for option_id in in_choices & in_standard if option_id)
    if not reaches_both:
        raise AssertionError(
            f"no {DRIFT_MODEL_KEY} option reaches both retained collections; "
            "the drift canary has no probe target"
        )
    return reaches_both[0], ["choices", "standardEquipment"]


DRIFT_PROBE_OPTION_ID, DRIFT_COLLECTIONS = _live_drift_probe()


def workbook_with_a_drifting_model(directory: Path) -> Path:
    """A candidate workbook whose zr1 output differs from the retained artifact.

    Drift used to be supplied for free by three stale retained contracts. Those
    were regenerated on 2026-07-27, so the drift is manufactured here instead —
    which is the better shape anyway: the test now controls its own input rather
    than depending on a defect elsewhere in the tree staying broken.
    """

    candidate = directory / WORKBOOK.name
    shutil.copy2(WORKBOOK, candidate)
    workbook = load_workbook(candidate)
    sheet = workbook[f"{DRIFT_MODEL_KEY}_options"]
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1) if cell.value}
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row, headers["option_id"]).value or "").strip() == DRIFT_PROBE_OPTION_ID:
            cell = sheet.cell(row, headers["option_name"])
            cell.value = f"{cell.value} (drift probe)"
            break
    else:  # pragma: no cover - the row exists; this guards a silent no-op
        raise AssertionError(
            f"{DRIFT_MODEL_KEY}_options has no {DRIFT_PROBE_OPTION_ID} row to perturb"
        )
    workbook.save(candidate)
    workbook.close()
    return candidate


@pytest.fixture(scope="module")
def drifting_undeclared(tmp_path_factory) -> dict:
    """The drifting workbook with nothing declared: zr1 must be unexpected_drift."""

    directory = tmp_path_factory.mktemp("drift-undeclared")
    return verify_candidate(workbook_with_a_drifting_model(directory), run_harness=False)


def test_every_stage_runs_in_order_against_a_candidate_copy(canonical_run) -> None:
    """Breaks if a stage is reordered, dropped, or silently skipped."""

    ran = canonical_run["stagesRun"]
    expected = list(STAGES)

    assert ran == expected, f"stage order changed: {ran}"
    assert canonical_run["stages"][0]["detail"]["candidate"] != str(WORKBOOK)
    assert Path(canonical_run["stages"][0]["detail"]["candidate"]).name == WORKBOOK.name


def test_a_workbook_defect_fails_at_the_earliest_applicable_stage(tmp_path) -> None:
    """Breaks if a defect stages 2-4 can see is allowed to reach generation."""

    broken = tmp_path / WORKBOOK.name
    shutil.copy2(WORKBOOK, broken)
    workbook = load_workbook(broken)
    del workbook["section_master"]
    workbook.save(broken)
    workbook.close()

    report = verify_candidate(broken, report_path=tmp_path / "readiness.json", run_harness=False)

    assert report["ok"] is False
    assert report["failedStage"] == "workbook_schema"
    # The point of "earliest applicable": generation must not have been attempted.
    assert "generate_models" not in report["stagesRun"]
    assert "generate_models" in report["stagesNotRun"]
    assert report["models"] == {}


def test_undeclared_semantic_drift_is_reported_and_fails(drifting_undeclared) -> None:
    """Breaks if a model whose output moved without being declared stops failing the run.

    This is the check that catches a bad global-family edit: one workbook cell
    changed, nobody declared it, the run must refuse.
    """

    assert drifting_undeclared["partition"]["unexpected_drift"] == ["zr1"]
    assert drifting_undeclared["ok"] is False
    assert drifting_undeclared["failedStage"] == "semantic_drift"
    assert drifting_undeclared["models"]["zr1"]["declared_changed"] is False
    # The probe option is standard on zr1, so renaming it moves both collections
    # that carry its label. Asserting the exact set, not just "non-empty", keeps
    # this from passing on unrelated drift.
    assert drifting_undeclared["models"]["zr1"]["semantic_drift_vs_retained"] == DRIFT_COLLECTIONS
    # Only the edited model moves; the other five must stay clean, or the drift
    # signal is noise rather than a pointer.
    assert set(drifting_undeclared["partition"]["unchanged"]) == ALL_MODEL_KEYS - {"zr1"}


def test_declaring_a_changed_model_does_not_reduce_the_generated_set(canonical_run) -> None:
    """Breaks if the touched-model set is ever used as a generation filter (§3.7.1.5)."""

    assert set(canonical_run["models"]) == ALL_MODEL_KEYS
    assert set(canonical_run["stages"][5]["detail"]["generated"]) == ALL_MODEL_KEYS
    assert set(canonical_run["declaredChangedModels"]) == ALL_MODEL_KEYS
    assert all(row["generated"] for row in canonical_run["models"].values())


def test_the_canonical_workbook_has_no_undeclared_drift(canonical_run) -> None:
    """The tree is clean: every retained contract matches what the workbook generates.

    Breaks the moment a retained artifact goes stale again — which is exactly the
    class of defect that went unnoticed until this lane existed.
    """

    assert canonical_run["partition"]["unexpected_drift"] == []
    assert set(canonical_run["partition"]["changed"]) == ALL_MODEL_KEYS
    assert canonical_run["ok"] is True


def test_all_models_marker_declares_every_model() -> None:
    """§3.7.1.1: a global-family row marks the touched set as all models."""

    from verify_workbook_candidate import declared_changed_set

    assert declared_changed_set(["*"], ALL_MODEL_KEYS) == ALL_MODEL_KEYS


def test_an_unknown_changed_model_fails_rather_than_being_ignored(tmp_path) -> None:
    """Breaks if a typo'd --changed-model silently declares nothing."""

    report = verify_candidate(
        WORKBOOK,
        changed_models=["corvette_zora"],
        report_path=tmp_path / "readiness.json",
        run_harness=False,
    )

    assert report["ok"] is False
    assert report["failedStage"] == "discover_models"
    assert "generate_models" not in report["stagesRun"]


def test_report_is_machine_readable_with_a_stable_schema_and_field_set(canonical_run) -> None:
    """Breaks if the database workflow's interface changes shape (§3.7.1.4)."""

    report_path = Path(canonical_run["_report_path"])
    expected = {key: value for key, value in canonical_run.items() if key != "_report_path"}

    assert report_path.exists()
    written = json.loads(report_path.read_text(encoding="utf-8"))
    assert written == expected
    assert written["schemaVersion"] == REPORT_SCHEMA_VERSION
    for model_key, row in written["models"].items():
        assert REQUIRED_MODEL_FIELDS <= set(row), f"{model_key} missing {REQUIRED_MODEL_FIELDS - set(row)}"


def test_protected_surfaces_are_byte_identical_after_a_passing_and_a_failing_run(
    tmp_path, canonical_run
) -> None:
    """Breaks if any code path — including failure paths — writes a tracked file."""

    before = protected_surface_hashes(ROOT)

    passing = canonical_run  # a full passing run, already executed
    broken = tmp_path / WORKBOOK.name
    shutil.copy2(WORKBOOK, broken)
    workbook = load_workbook(broken)
    del workbook["section_master"]
    workbook.save(broken)
    workbook.close()
    failing = verify_candidate(broken, run_harness=False)

    assert passing["boundaryViolations"] == []
    assert failing["boundaryViolations"] == []
    assert protected_surface_hashes(ROOT) == before


def test_drift_detection_ignores_order_but_not_content() -> None:
    """Breaks if drift is ever keyed on array position again.

    A section reorder re-sorts thousands of downstream rows; keyed on position it
    reports drift for all of them and the signal is worthless.
    """

    base = {
        "dataset": {"name": "x", "generated_at": "2026-01-01"},
        "sections": [{"section_id": "a", "display_order": 1}, {"section_id": "b", "display_order": 2}],
        "choices": [{"choice_id": "c1", "label": "L"}],
    }
    reordered = {
        "dataset": {"name": "x", "generated_at": "2026-06-06"},
        "sections": [{"section_id": "b", "display_order": 2}, {"section_id": "a", "display_order": 1}],
        "choices": [{"choice_id": "c1", "label": "L"}],
    }
    changed = json.loads(json.dumps(reordered))
    changed["choices"][0]["label"] = "different"

    assert semantic_drift(reordered, base) == []
    assert semantic_drift(changed, base) == ["choices"]


def test_the_browser_stage_reads_the_candidate_registry_not_the_published_one(tmp_path) -> None:
    """The failure this exists to catch: a harness that silently falls back.

    If the harness ignored the override and read the tracked `form-app/data.js`,
    the `browser_harness` stage would pass against a registry the candidate never
    produced — a green
    readiness verdict proving nothing. Pointed at a deliberately broken registry
    it must fail, and the published file must be untouched.
    """

    before = protected_surface_hashes(ROOT)
    broken_data_js = tmp_path / "data.js"
    broken_data_js.write_text("window.CORVETTE_FORM_DATA = {};\n", encoding="utf-8")

    result = run_browser_harness(broken_data_js, ROOT / "tests" / "multi-model-runtime-switching.test.mjs")

    assert result.ok is False, "harness passed against an empty registry; it is not reading the override"
    assert result.detail["data_js"] == str(broken_data_js)
    assert protected_surface_hashes(ROOT) == before


def test_the_harness_override_env_var_is_the_one_the_harness_reads() -> None:
    """Breaks if the override name drifts on either side of the contract."""

    switching = (ROOT / "tests" / "multi-model-runtime-switching.test.mjs").read_text(encoding="utf-8")
    matrix = (ROOT / "tests" / "lib" / "runtime-harness.mjs").read_text(encoding="utf-8")

    assert f"process.env.{HARNESS_DATA_JS_ENV}" in switching
    assert f"process.env.{HARNESS_DATA_JS_ENV}" in matrix
    matrix_truth = (ROOT / "tests" / "lib" / "workbook-truth.mjs").read_text(encoding="utf-8")
    assert f"process.env[{WORKBOOK_TRUTH_ENV!r}]" in matrix_truth or f"process.env.{WORKBOOK_TRUTH_ENV}" in matrix_truth or WORKBOOK_TRUTH_ENV in matrix_truth


def test_the_browser_stage_receives_the_already_built_snapshot(tmp_path) -> None:
    """The matrix must not rebuild the snapshot inside Layer 1.

    CI has no `.venv/bin/python`. If browser_harness omits CORVETTE_WORKBOOK_TRUTH,
    the matrix falls through to that hardcoded interpreter and the lane dies with
    ENOENT after ten expensive stages. The snapshot already exists from stage 9.
    """

    data_js = tmp_path / "data.js"
    data_js.write_text("window.CORVETTE_FORM_DATA = {};\n", encoding="utf-8")
    truth = tmp_path / "workbook-truth.json"
    truth.write_text("{}", encoding="utf-8")

    result = run_browser_harness(
        data_js,
        ROOT / "tests" / "multi-model-runtime-switching.test.mjs",
        truth_path=truth,
    )

    assert result.detail[WORKBOOK_TRUTH_ENV] == str(truth)
    assert result.detail["data_js"] == str(data_js)


def test_protected_surface_hashes_ignore_macos_finder_metadata(tmp_path) -> None:
    form_output = tmp_path / "form-output"
    form_output.mkdir()
    (form_output / ".DS_Store").write_bytes(b"finder metadata")

    assert "form-output/.DS_Store" not in protected_surface_hashes(tmp_path)


def test_the_lane_detects_and_reports_a_protected_path_write(monkeypatch) -> None:
    """Proves the in-tool boundary check can actually fire.

    Every other assertion here checks `boundaryViolations == []`, which a lane
    that never computes it would satisfy trivially. This one makes a stage write
    a tracked file and requires the lane to notice, then restores the file.
    """

    import verify_workbook_candidate as lane

    victim = ROOT / "form-output" / "runtime" / "z06-runtime-contract.json"
    original = victim.read_bytes()
    real_stage = lane.run_stage_options_quality

    def writing_stage(candidate):
        victim.write_bytes(original + b"\n")
        result = real_stage(candidate)
        result.ok = False
        result.findings.append({"message": "forced early stop after boundary write"})
        return result

    monkeypatch.setattr(lane, "run_stage_options_quality", writing_stage)
    try:
        report = lane.verify_candidate(WORKBOOK, changed_models=["*"], run_harness=False)
    finally:
        victim.write_bytes(original)

    assert report["boundaryViolations"] == ["form-output/runtime/z06-runtime-contract.json"]
    assert report["ok"] is False
    assert victim.read_bytes() == original


def test_the_lane_runs_the_browser_stage_against_a_temporary_registry(canonical_run) -> None:
    """Breaks if `browser_harness` is removed or stops using the candidate's data.js.

    Every other test in this file passes `run_harness=False`, so without this one
    deleting that stage outright would be invisible to the whole suite.
    """

    report = canonical_run
    harness_stage = next(stage for stage in report["stages"] if stage["stage"] == "browser_harness")

    assert report["stagesRun"] == list(STAGES)
    assert report["stagesSkippedByCaller"] == []
    assert harness_stage["ok"] is True
    assert harness_stage["detail"]["returncode"] == 0
    # The registry it tested exists only inside the lane's temporary root, which
    # is deleted on exit — so it cannot have been the published form-app/data.js.
    data_js = Path(harness_stage["detail"]["data_js"])
    assert not data_js.is_relative_to(ROOT), data_js
    assert not data_js.exists()
    assert report["ok"] is True
