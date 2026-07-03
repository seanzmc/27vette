#!/usr/bin/env python3
"""Tests for the Pass 3 order-guide expert interpreter."""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
TESTS_DIR = ROOT / "tests"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))
if str(TESTS_DIR) not in sys.path:
    sys.path.insert(0, str(TESTS_DIR))

from corvette_form_generator.ingest.candidate_normalizer import normalize_order_guide_candidates  # noqa: E402
from corvette_form_generator.ingest.expert_interpreter import interpret_order_guide_candidates  # noqa: E402
from corvette_form_generator.ingest.source_profiler import profile_order_guide  # noqa: E402
from test_order_guide_candidate_normalizer import build_evidence as build_focused_evidence  # noqa: E402


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def fixture_workbook(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "model_master",
        [
            "model_key",
            "registry_key",
            "model_label",
            "model_year",
            "dataset_name",
            "export_slug",
            "expected_variant_count",
            "default_model",
            "active",
            "notes",
        ],
        [{
            "model_key": "stingray",
            "registry_key": "stingray",
            "model_label": "Stingray",
            "model_year": "2027",
            "dataset_name": "Fixture Stingray",
            "export_slug": "stingray",
            "expected_variant_count": 1,
            "default_model": True,
            "active": True,
        }],
    )
    append_sheet(
        wb,
        "variant_master",
        ["variant_id", "model_year", "trim_level", "body_style", "display_name", "base_price", "display_order", "active"],
        [{
            "variant_id": "1lt_c07",
            "model_year": "2027",
            "trim_level": "1lt",
            "body_style": "coupe",
            "display_name": "Corvette Stingray Coupe 1LT",
            "base_price": 71000,
            "display_order": 1,
            "active": True,
        }],
    )
    append_sheet(
        wb,
        "model_variants",
        ["model_key", "variant_id", "display_order", "active", "notes"],
        [{"model_key": "stingray", "variant_id": "1lt_c07", "display_order": 1, "active": True}],
    )
    append_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active", "notes"],
        [
            {"model_key": "stingray", "source_role": "source_option_sheet", "sheet_name": "stingray_options", "active": True},
            {"model_key": "stingray", "source_role": "status_sheet", "sheet_name": "stingray_ovs", "active": True},
        ],
    )
    append_sheet(
        wb,
        "stingray_options",
        [
            "option_id",
            "rpo",
            "price",
            "option_name",
            "description",
            "detail_raw",
            "section_id",
            "selectable",
            "display_order",
            "active",
            "display_behavior",
        ],
        [
            {"option_id": "opt_safe_001", "rpo": "SAF", "option_name": "Friendly Safe Copy", "section_id": "sec_ext_001", "active": True},
            {"option_id": "opt_dup_001", "rpo": "DUP", "option_name": "Friendly Duplicate Copy", "section_id": "sec_ext_001", "active": True},
            {"option_id": "opt_adi_001", "rpo": "ADI", "option_name": "Friendly Dealer Copy", "section_id": "sec_ext_001", "active": True},
            {"option_id": "opt_inc_001", "rpo": "INC", "option_name": "Friendly Included Copy", "section_id": "sec_eqp_001", "active": True},
            {"option_id": "opt_fot_001", "rpo": "FOT", "option_name": "Friendly Footnote Copy", "section_id": "sec_ext_001", "active": True},
        ],
    )
    append_sheet(
        wb,
        "stingray_ovs",
        ["option_id", "variant_id", "status"],
        [
            {"option_id": "opt_safe_001", "variant_id": "1lt_c07", "status": "available"},
            {"option_id": "opt_dup_001", "variant_id": "1lt_c07", "status": "available"},
            {"option_id": "opt_adi_001", "variant_id": "1lt_c07", "status": "available"},
            {"option_id": "opt_inc_001", "variant_id": "1lt_c07", "status": "standard"},
            {"option_id": "opt_fot_001", "variant_id": "1lt_c07", "status": "available"},
        ],
    )
    wb.save(path)


def raw_export_fixture(path: Path) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    ws = wb.create_sheet("Price Schedule")
    ws.append(["2027 CHEVROLET CORVETTE", "", "", ""])
    ws.append(["", "Model", "Model Description", "MSRP(c)"])
    ws.append(["", "1YC07", "Stingray Coupe", "$71,000"])

    ws = wb.create_sheet("Equipment Groups 1")
    ws.append(["Stingray", "", "", ""])
    ws.append(["", "", "S = Standard Equipment  A = Available  -- = Not Available", ""])
    ws.append(["Orderable RPO Code", "Ref. Only RPO Code", "Description", "Coupe / 1YC07 / 1LT"])
    ws.append(["SAF", "", "Raw source copy intentionally differs from friendly workbook copy", "A"])
    ws.append(["DUP", "", "Duplicate source row", "A"])
    ws.append(["ADI", "", "Dealer-installed source row", "A/D"])
    ws.append(["INC", "", "Included equipment source row", "■"])
    ws.append(["FOT", "", "Footnote source row / 1. Requires (SAF) Safe source option.", "A1"])
    ws.append(["NEW", "", "New source option", "A"])

    ws = wb.create_sheet("Exterior 1")
    ws.append(["Stingray", "", "", ""])
    ws.append(["", "", "S = Standard Equipment  A = Available  -- = Not Available", ""])
    ws.append(["Orderable RPO Code", "Ref. Only RPO Code", "Description", "Coupe / 1YC07 / 1LT"])
    ws.append(["DUP", "", "Duplicate source row", "A"])
    wb.save(path)


def build_candidates(tmp: Path) -> tuple[Path, Path, Path]:
    workbook = tmp / "canonical.xlsx"
    raw_export = tmp / "raw.xlsx"
    evidence_dir = tmp / "evidence"
    candidates_dir = tmp / "candidates"
    fixture_workbook(workbook)
    raw_export_fixture(raw_export)
    profile_order_guide(
        raw_export=raw_export,
        workbook=workbook,
        output_dir=evidence_dir,
        run_id="unit-evidence",
        root=ROOT,
    )
    normalize_order_guide_candidates(
        evidence_dir=evidence_dir,
        workbook=workbook,
        output_dir=candidates_dir,
        run_id="unit-candidates",
        root=ROOT,
    )
    return workbook, evidence_dir, candidates_dir


class OrderGuideIngestInterpreterTests(unittest.TestCase):
    def test_interprets_model_rpo_units_and_strict_auto_confirmed_gates(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir, candidates_dir = build_candidates(tmp)
            output_dir = tmp / "interpretation"

            result = interpret_order_guide_candidates(
                evidence_dir=evidence_dir,
                candidates_dir=candidates_dir,
                workbook=workbook,
                output_dir=output_dir,
                run_id="unit-interpretation",
                root=ROOT,
            )

            self.assertEqual(result["status"], "passed")
            for name in [
                "interpretation-summary.json",
                "interpreted-options.json",
                "review-queue.json",
                "duplicate-rpo-report.json",
                "duplicate-rpo-report.md",
                "source-sheet-coverage.json",
                "source-sheet-coverage.md",
                "blocked-interpretation.json",
            ]:
                self.assertTrue((output_dir / name).exists(), name)

            summary = json.loads((output_dir / "interpretation-summary.json").read_text())
            self.assertIn(summary["reduction_status"], {"material_reduction", "insufficient_reduction"})
            self.assertEqual(summary["raw_candidate_counts"]["options"], 7)
            self.assertEqual(summary["interpreted_option_count"], 6)
            self.assertEqual(summary["hidden_auto_confirmed_count"], 2)
            self.assertGreater(summary["visible_review_queue_count"], 0)
            self.assertEqual(summary["duplicate_rpo_count"], 1)
            self.assertIn("stingray", summary["source_sheet_coverage"])

            interpreted = {
                item["rpo"]: item
                for item in json.loads((output_dir / "interpreted-options.json").read_text())
            }
            self.assertEqual(interpreted["SAF"]["interpretation_confidence"], "auto_confirmed")
            self.assertEqual(interpreted["SAF"]["copy_comparison_status"], "not_compared_by_design")
            self.assertEqual(interpreted["DUP"]["duplicate_classification"], "redundant_duplicates")
            self.assertEqual(interpreted["DUP"]["interpretation_confidence"], "auto_confirmed")
            self.assertEqual(interpreted["ADI"]["interpretation_confidence"], "review_needed")
            self.assertIn("dealer_installed_or_adi", interpreted["ADI"]["review_reason_codes"])
            self.assertEqual(interpreted["INC"]["interpretation_confidence"], "review_needed")
            self.assertIn("equipment_group_inclusion", interpreted["INC"]["review_reason_codes"])
            self.assertEqual(interpreted["FOT"]["interpretation_confidence"], "review_needed")
            self.assertIn("footnote_or_disclosure_marker", interpreted["FOT"]["review_reason_codes"])
            self.assertEqual(interpreted["NEW"]["interpretation_confidence"], "mechanical_safe")
            self.assertEqual(interpreted["NEW"]["workbook_identity_match"]["match_status"], "missing_in_workbook")

            review_queue = json.loads((output_dir / "review-queue.json").read_text())
            self.assertNotIn("SAF", {item["rpo"] for item in review_queue})
            self.assertIn("ADI", {item["rpo"] for item in review_queue})
            self.assertIn("NEW", {item["rpo"] for item in review_queue})

    def test_rejects_protected_generated_output_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir, candidates_dir = build_candidates(tmp)
            with self.assertRaisesRegex(ValueError, "form-output"):
                interpret_order_guide_candidates(
                    evidence_dir=evidence_dir,
                    candidates_dir=candidates_dir,
                    workbook=workbook,
                    output_dir=ROOT / "form-output" / "runtime" / "bad-interpretation",
                    run_id="bad-output",
                    root=ROOT,
                )

    def test_focused_model_interpretation_writes_workbook_build_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir = build_focused_evidence(tmp)
            candidates_dir = tmp / "focused-candidates"
            normalize_order_guide_candidates(
                evidence_dir=evidence_dir,
                workbook=workbook,
                output_dir=candidates_dir,
                run_id="focused-candidates",
                root=ROOT,
                selected_models=["zr1"],
            )
            output_dir = tmp / "focused-interpretation"

            result = interpret_order_guide_candidates(
                evidence_dir=evidence_dir,
                candidates_dir=candidates_dir,
                workbook=workbook,
                output_dir=output_dir,
                run_id="focused-interpretation",
                root=ROOT,
                selected_models=["zr1"],
                primary_models=["zr1"],
            )

            self.assertEqual(result["status"], "passed")
            for name in ["model-selection.json", "workbook-build-summary.json", "workbook-build-review-units.json"]:
                self.assertTrue((output_dir / name).exists(), name)
                self.assertIn(name, result["artifact_files"])
            summary = json.loads((output_dir / "workbook-build-summary.json").read_text())
            self.assertEqual(summary["review_mode"], "focused_workbook_build")
            self.assertEqual(summary["selection_metadata"]["selected_models"], ["zr1"])
            self.assertTrue(summary["cross_check_status"]["ok"])
            self.assertGreaterEqual(summary["lane_counts"]["option_rows"], 1)
            self.assertGreaterEqual(summary["lane_counts"]["ovs_rows"], 1)

            units = json.loads((output_dir / "workbook-build-review-units.json").read_text())
            option_unit = next(row for row in units if row["lane"] == "option_rows" and row["rpo"] == "TOM")
            self.assertEqual(option_unit["target_sheet"], "zr1_options")
            self.assertEqual(option_unit["workbook_presence"], "existing_inactive_scaffold")
            self.assertEqual(option_unit["proposed_workbook_action"], "verify_existing_option_row")
            ovs_unit = next(row for row in units if row["lane"] == "ovs_rows" and row["rpo"] == "TOM")
            self.assertEqual(ovs_unit["target_sheet"], "zr1_ovs")
            self.assertEqual(ovs_unit["proposed_workbook_action"], "verify_status_matrix")

    def test_focused_interpretation_marks_comparator_units_and_never_targets_primary_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir = build_focused_evidence(tmp)
            candidates_dir = tmp / "focused-candidates"
            normalize_order_guide_candidates(
                evidence_dir=evidence_dir,
                workbook=workbook,
                output_dir=candidates_dir,
                run_id="focused-comparator-candidates",
                root=ROOT,
                selected_models=["zr1", "stingray"],
                primary_models=["zr1"],
                comparator_models=["stingray"],
            )
            output_dir = tmp / "focused-interpretation"
            interpret_order_guide_candidates(
                evidence_dir=evidence_dir,
                candidates_dir=candidates_dir,
                workbook=workbook,
                output_dir=output_dir,
                run_id="focused-comparator-interpretation",
                root=ROOT,
                selected_models=["zr1", "stingray"],
                primary_models=["zr1"],
                comparator_models=["stingray"],
            )

            units = json.loads((output_dir / "workbook-build-review-units.json").read_text())
            comparator_units = [unit for unit in units if unit["model_key"] == "stingray"]
            primary_units = [unit for unit in units if unit["model_key"] == "zr1"]
            self.assertTrue(comparator_units)
            self.assertTrue(primary_units)
            for unit in comparator_units:
                self.assertEqual(unit["model_role"], "comparator")
                self.assertTrue(unit["comparator_context"]["comparator_only"])
                self.assertFalse(unit["target_sheet"].startswith("zr1"), unit["target_sheet"])
            for unit in primary_units:
                self.assertEqual(unit["model_role"], "primary")

            summary = json.loads((output_dir / "workbook-build-summary.json").read_text())
            self.assertGreaterEqual(summary["model_role_counts"]["comparator"], 1)
            self.assertGreaterEqual(summary["model_role_counts"]["primary"], 1)
            self.assertIn("evidence", summary["artifact_fingerprints"])
            self.assertIn("candidates", summary["artifact_fingerprints"])
            self.assertIn("interpretation", summary["artifact_fingerprints"])
            self.assertIn("model-selection.json", summary["artifact_fingerprints"]["candidates"])

    def test_focused_interpretation_fails_closed_on_evidence_fingerprint_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir = build_focused_evidence(tmp)
            candidates_dir = tmp / "focused-candidates"
            normalize_order_guide_candidates(
                evidence_dir=evidence_dir,
                workbook=workbook,
                output_dir=candidates_dir,
                run_id="focused-candidates",
                root=ROOT,
                selected_models=["zr1"],
            )
            matrix_path = evidence_dir / "variant-matrix.json"
            matrix_path.write_text(matrix_path.read_text() + "\n")

            with self.assertRaisesRegex(ValueError, "evidence fingerprint mismatch"):
                interpret_order_guide_candidates(
                    evidence_dir=evidence_dir,
                    candidates_dir=candidates_dir,
                    workbook=workbook,
                    output_dir=tmp / "focused-interpretation",
                    run_id="focused-interpretation",
                    root=ROOT,
                )

    def test_focused_interpretation_requires_candidate_selection_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir, candidates_dir = build_candidates(tmp)

            with self.assertRaisesRegex(ValueError, "model-selection.json"):
                interpret_order_guide_candidates(
                    evidence_dir=evidence_dir,
                    candidates_dir=candidates_dir,
                    workbook=workbook,
                    output_dir=tmp / "focused-interpretation",
                    run_id="focused-interpretation",
                    root=ROOT,
                    selected_models=["stingray"],
                )

    def test_cli_emits_summary_json(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            workbook, evidence_dir, candidates_dir = build_candidates(tmp)
            output_dir = tmp / "cli-interpretation"
            result = subprocess.run(
                [
                    sys.executable,
                    str(ROOT / "scripts" / "order_guide_ingest_interpreter.py"),
                    "--evidence-dir",
                    str(evidence_dir),
                    "--candidates-dir",
                    str(candidates_dir),
                    "--workbook",
                    str(workbook),
                    "--run-id",
                    "cli-interpretation",
                    "--output-dir",
                    str(output_dir),
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            payload = json.loads(result.stdout)
            self.assertEqual(payload["status"], "passed")
            self.assertEqual(payload["output_dir"], str(output_dir.resolve()))
            self.assertTrue((output_dir / "interpretation-summary.json").exists())


if __name__ == "__main__":
    unittest.main()
