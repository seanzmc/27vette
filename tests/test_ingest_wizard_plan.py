#!/usr/bin/env python3
"""Historical Pass C projection-library characterization tests.

The current ingest API no longer exposes plan mutation or approval. These
fixture-only tests retain coverage for the read-only ``build_plan`` library
used to interpret historical evidence; no current session enters plan state.
"""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from corvette_form_generator.ingest.wizard.plan_builder import build_plan  # noqa: E402
from corvette_form_generator.ingest.wizard.session import (  # noqa: E402
    WizardSessionStore,
    read_json,
)
from ingest_wizard_fixtures import build_master_workbook, build_raw_export  # noqa: E402


class PlanFlowTest(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self._schema_patch = patch(
            "corvette_form_generator.editor_ops.validate_workbook_schema",
            return_value=[],
        )
        self._schema_patch.start()
        self.root = Path(self._tmp.name)
        build_raw_export(self.root / "raw.xlsx")
        self.master = build_master_workbook(self.root / "master.xlsx")
        self.store = WizardSessionStore(self.root, workbook_path=self.master)
        created = self.store.create_session("raw.xlsx")
        self.run_id = created["session"]["runId"]
        roles = {card["sheetName"]: card["recommendedRole"] for card in created["profile"]["sheets"]}
        self.store.confirm_roles(self.run_id, roles)
        self.store.run_parse(self.run_id)
        self.store.select_models(self.run_id, ["zr1", "zr1x"], {"zr1": "z06", "zr1x": "z06"})

    def tearDown(self) -> None:
        self._schema_patch.stop()
        self._tmp.cleanup()

    def complete_model(
        self,
        model: str,
        *,
        extra: list[dict] | None = None,
        skip_section_ids: set[str] | None = None,
    ) -> None:
        queue = self.store.review_queue(self.run_id, model, "section")
        decisions = []
        skip_section_ids = skip_section_ids or set()
        reconciliation = self.store.reconciliation(self.run_id)["models"].get(model, {})
        if not reconciliation.get("agrees", True):
            decisions.append(
                {
                    "model": model,
                    "lane": "relationship",
                    "groupKey": "variant_reconciliation",
                    "action": "needs_product_decision",
                    "payload": {},
                    "resolution": "approved_for_plan",
                }
            )
        for candidate in queue["candidates"]:
            if candidate["candidateId"] in skip_section_ids:
                continue
            decisions.append(
                {
                    "model": model,
                    "lane": "section",
                    "candidateId": candidate["candidateId"],
                    "action": "assign_section",
                    "payload": {"sectionId": "sec_whee_001"},
                    "resolution": "approved_for_plan",
                }
            )
        for candidate in self.store.review_queue(self.run_id, model, "price")["candidates"]:
            action = "accept_exact_price" if candidate["priceMatch"] == "exact" else "confirm_no_price"
            decisions.append(
                {
                    "model": model,
                    "lane": "price",
                    "candidateId": candidate["candidateId"],
                    "action": action,
                    "payload": {},
                    "resolution": "approved_for_plan",
                }
            )
        for candidate in self.store.review_queue(self.run_id, model, "status_nuance")["candidates"]:
            decisions.append(
                {
                    "model": model,
                    "lane": "status_nuance",
                    "candidateId": candidate["candidateId"],
                    "action": "confirm_status",
                    "payload": {},
                    "resolution": "approved_for_plan",
                }
            )
        prefill = self.store.review_queue(self.run_id, model, "presentation")["prefill"]
        for sheet, proposals in prefill["sheets"].items():
            decisions.append(
                {
                    "model": model,
                    "lane": "presentation",
                    "groupKey": sheet,
                    "action": "approve_presentation_rows",
                    "payload": {"rows": [p["row"] for p in proposals], "templateModel": "z06"},
                    "resolution": "approved_for_plan",
                }
            )
        decisions.extend(extra or [])
        self.store.save_decisions(self.run_id, decisions)

    def complete_all(self, *, zr1_extra: list[dict] | None = None) -> None:
        self.complete_model("zr1", extra=zr1_extra)
        self.complete_model("zr1x")

    def plan_inputs(self) -> dict:
        run_dir = self.store.run_dir(self.run_id)
        selection = read_json(run_dir / "model-selection.json")
        candidates = read_json(run_dir / "option-candidates.json")["candidates"]
        decisions = {r["decisionId"]: r for r in read_json(run_dir / "decisions.json")["decisions"]}
        return {
            "workbook_path": self.master,
            "selection": selection,
            "candidates": candidates,
            "decisions": decisions,
            "candidates_fingerprint": selection["candidatesFingerprint"],
        }


    def test_plan_deterministic_and_covered(self) -> None:
        self.complete_all()
        inputs = self.plan_inputs()
        plan_a = build_plan(**inputs)
        plan_b = build_plan(**inputs)
        self.assertEqual(plan_a["schemaVersion"], "pass-c-2")
        self.assertEqual(json.dumps(plan_a, sort_keys=True), json.dumps(plan_b, sort_keys=True))
        self.assertTrue(plan_a["valid"], plan_a["report"]["blockingGaps"])
        self.assertEqual(plan_a["coverage"]["uncoveredApprovedDecisions"], [])
        # Clean reprocess recorded: the zr1_options scaffold row is cleared.
        self.assertEqual(plan_a["report"]["clearedRows"].get("zr1_options"), 1)
        # Presentation rows land as adds on all five sheets.
        for sheet in ("runtime_steps", "section_presentation", "context_section_master", "order_summary_sections", "step_order_summary_map"):
            self.assertIn("add", plan_a["report"]["perSheetCounts"].get(sheet, {}), sheet)
            self.assertIn("add", plan_a["report"]["perSheetActionCounts"].get(sheet, {}), sheet)
        # No timestamps anywhere in the plan payload (determinism guard).
        import re

        payload = json.dumps(plan_a)
        self.assertIsNone(re.search(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}", payload))

    def test_grand_sport_x_registry_key_uses_runtime_metadata_key(self) -> None:
        inputs = self.plan_inputs()
        inputs["selection"] = {**inputs["selection"], "targets": ["grand_sport_x"]}
        plan = build_plan(**inputs)
        rows = [
            item["row"]
            for item in plan["stage1"]["items"]
            if item.get("sheet") in {"model_master", "model_registry_promotion"}
        ]

        self.assertTrue(rows)
        self.assertTrue(all(row["registry_key"] == "grand_sport_x" for row in rows))

    def test_relationship_and_exclusive_ops(self) -> None:
        extra = [
            {
                "model": "zr1",
                "lane": "relationship",
                "groupKey": "pdb-excludes-c2z",
                "action": "create_relationship_candidate",
                "payload": {"kind": "not_available_with", "sourceRpo": "PDB", "targetRpos": ["C2Z"]},
                "resolution": "approved_for_plan",
            },
            {
                "model": "zr1",
                "lane": "exclusive_group",
                "groupKey": "zr1-wheel-pack",
                "action": "create_exclusive_group",
                "payload": {"members": ["PDB", "C2Z"]},
                "resolution": "approved_for_plan",
            },
        ]
        self.complete_all(zr1_extra=extra)
        plan = build_plan(**self.plan_inputs())
        self.assertTrue(plan["valid"])
        items = plan["stage2"]["items"]
        rule = next(i for i in items if i["sheet"] == "zr1_rule_mapping" and i["action"] == "add")
        self.assertEqual(rule["row"]["rule_type"], "excludes")
        group = next(i for i in items if i["sheet"] == "zr1_exclusive_groups" and i["action"] == "add")
        self.assertEqual(group["row"]["group_id"], "excl_zr1wheelpack")
        members = [i for i in items if i["sheet"] == "zr1_exclusive_members" and i["action"] == "add"]
        self.assertEqual(len(members), 2)

    def test_relationship_and_exclusive_ops_dry_run_against_live_like_headers(self) -> None:
        from openpyxl import load_workbook

        extra = [
            {
                "model": "zr1",
                "lane": "relationship",
                "groupKey": "pdb-excludes-c2z",
                "action": "create_relationship_candidate",
                "payload": {"kind": "not_available_with", "sourceRpo": "PDB", "targetRpos": ["C2Z"]},
                "resolution": "approved_for_plan",
            },
            {
                "model": "zr1",
                "lane": "exclusive_group",
                "groupKey": "zr1-wheel-pack",
                "action": "create_exclusive_group",
                "payload": {"members": ["PDB", "C2Z"]},
                "resolution": "approved_for_plan",
            },
        ]
        self.complete_all(zr1_extra=extra)

        wb = load_workbook(self.master, read_only=True, data_only=True)
        try:
            rule_headers = [cell.value for cell in wb["z06_rule_mapping"][1]]
            group_headers = [cell.value for cell in wb["z06_exclusive_groups"][1]]
        finally:
            wb.close()
        self.assertNotIn("active", rule_headers)
        self.assertNotIn("group_name", group_headers)
        self.assertIn("notes", group_headers)

        plan = build_plan(**self.plan_inputs())
        self.assertTrue(plan["valid"], plan["report"]["blockingGaps"])
        rule = next(i for i in plan["stage2"]["items"] if i["sheet"] == "zr1_rule_mapping" and i["action"] == "add")
        group = next(i for i in plan["stage2"]["items"] if i["sheet"] == "zr1_exclusive_groups" and i["action"] == "add")
        self.assertNotIn("active", rule["row"])
        self.assertNotIn("group_name", group["row"])
        self.assertEqual(group["row"]["notes"], "Review group: zr1-wheel-pack")

    def test_default_selection_rule_emits_from_reviewed_exclusive_payload(self) -> None:
        extra = [
            {
                "model": "zr1",
                "lane": "exclusive_group",
                "groupKey": "zr1-wheel-pack",
                "action": "create_exclusive_group",
                "payload": {"members": ["PDB", "C2Z"], "defaultRpo": "PDB"},
                "resolution": "approved_for_plan",
            }
        ]
        self.complete_all(zr1_extra=extra)
        plan = build_plan(**self.plan_inputs())
        self.assertTrue(plan["valid"], plan["report"]["blockingGaps"])
        row = next(
            item["row"]
            for item in plan["stage2"]["items"]
            if item["sheet"] == "default_selection_rules" and item["action"] == "add"
        )
        self.assertEqual(row["model_key"], "zr1")
        self.assertEqual(row["target_option_id"], "opt_pdb_002")
        self.assertEqual(row["condition_type"], "always")
        self.assertEqual(row["display_behavior"], "default_selected")

    def test_required_default_selection_without_default_blocks_plan(self) -> None:
        extra = [
            {
                "model": "zr1",
                "lane": "exclusive_group",
                "groupKey": "zr1-wheel-pack",
                "action": "create_exclusive_group",
                "payload": {"members": ["PDB", "C2Z"], "requiresDefaultSelection": True},
                "resolution": "approved_for_plan",
            }
        ]
        self.complete_all(zr1_extra=extra)
        plan = build_plan(**self.plan_inputs())
        self.assertFalse(plan["valid"])
        self.assertIn("default_selection_rules_missing", {gap["kind"] for gap in plan["report"]["blockingGaps"]})

    def test_relationship_resolves_against_retained_existing_option_rows(self) -> None:
        from openpyxl import load_workbook

        wb = load_workbook(self.master)
        wb["zr1_options"].append(["opt_tyz_001", "TYZ", "", "Existing target", "", "", "sec_whee_001", "", 30, True, ""])
        wb.save(self.master)
        wb.close()
        extra = [
            {
                "model": "zr1",
                "lane": "relationship",
                "groupKey": "pdb-excludes-tyz",
                "action": "create_relationship_candidate",
                "payload": {"kind": "not_available_with", "sourceRpo": "PDB", "targetRpos": ["TYZ"]},
                "resolution": "approved_for_plan",
            }
        ]
        self.complete_all(zr1_extra=extra)
        plan = build_plan(**self.plan_inputs())
        self.assertTrue(plan["valid"], plan["report"]["blockingGaps"])
        deletes = [
            item
            for item in plan["stage2"]["items"]
            if item["sheet"] == "zr1_options" and item["action"] == "delete" and item["key"].get("option_id") == "opt_tyz_001"
        ]
        self.assertEqual(deletes, [])
        rule = next(item for item in plan["stage2"]["items"] if item["sheet"] == "zr1_rule_mapping" and item["action"] == "add")
        self.assertEqual(rule["row"]["target_id"], "opt_tyz_001")

    def test_model_interior_scope_uses_existing_row_strategy(self) -> None:
        self.complete_all()
        plan = build_plan(**self.plan_inputs())
        scope_ops = [item for item in plan["stage2"]["items"] if item["sheet"] == "model_interior_scope"]
        self.assertFalse(
            any(item["key"] == {"model_key": "zr1", "interior_id": "1LZ_AQ9_HTA", "trim_level": "1LZ"} for item in scope_ops)
        )
        self.assertTrue(
            any(
                item["action"] == "add"
                and item["key"] == {"model_key": "zr1", "interior_id": "3LZ_AQ9_HTA", "trim_level": "3LZ"}
                for item in scope_ops
            )
        )
        self.assertTrue(
            any(
                item["action"] == "add"
                and item["key"] == {"model_key": "zr1x", "interior_id": "1LZ_AQ9_HTA", "trim_level": "1LZ"}
                for item in scope_ops
            )
        )

    def test_legacy_standard_equipment_inclusion_becomes_standard_option(self) -> None:
        se_queue = self.store.review_queue(self.run_id, "zr1", "standard_equipment")
        target = se_queue["candidates"][0]
        extra = [
            {
                "model": "zr1",
                "lane": "standard_equipment",
                "candidateId": target["candidateId"],
                "action": "include_standard_equipment",
                "payload": {},
                "resolution": "approved_for_plan",
            }
        ]
        self.complete_model("zr1", extra=extra, skip_section_ids={target["candidateId"]})
        self.complete_model("zr1x")
        plan = build_plan(**self.plan_inputs())
        self.assertTrue(plan["valid"], plan["report"]["blockingGaps"] or plan["coverage"]["uncoveredApprovedDecisions"])
        adds = [i for i in plan["stage2"]["items"] if i["sheet"] == "zr1_options" and i["action"] == "add"]
        se_row = next(i["row"] for i in adds if i["row"]["rpo"] == target["refOnlyRpo"])
        self.assertIs(se_row["selectable"], False)
        self.assertIsNone(se_row["section_id"])

    def test_ref_only_section_assignment_becomes_sectioned_option(self) -> None:
        self.complete_all()
        target = next(
            candidate
            for candidate in self.store.review_queue(self.run_id, "zr1", "section")["candidates"]
            if candidate["refOnlyRpo"] == "XFR"
        )
        self.store.save_decisions(
            self.run_id,
            [
                {
                    "model": "zr1",
                    "lane": "section",
                    "candidateId": target["candidateId"],
                    "action": "assign_section",
                    "payload": {"sectionId": "sec_pain_001", "selectable": False},
                    "resolution": "approved_for_plan",
                }
            ],
        )
        plan = build_plan(**self.plan_inputs())
        self.assertTrue(plan["valid"], plan["report"]["blockingGaps"] or plan["coverage"]["uncoveredApprovedDecisions"])
        row = next(
            i["row"]
            for i in plan["stage2"]["items"]
            if i["sheet"] == "zr1_options" and i["action"] == "add" and i["row"]["rpo"] == "XFR"
        )
        self.assertIsNone(row["price"])
        self.assertEqual(row["section_id"], "sec_pain_001")
        self.assertIs(row["selectable"], False)
        ovs_rows = [
            i["row"]
            for i in plan["stage2"]["items"]
            if i["sheet"] == "zr1_ovs" and i["action"] == "add" and i["row"]["option_id"] == row["option_id"]
        ]
        self.assertEqual(
            {ovs["variant_id"]: ovs["status"] for ovs in ovs_rows},
            {"1lz_r07": "available", "3lz_r67": "unavailable"},
        )

    def test_skipped_row_stays_out_of_plan_and_plan_stays_valid(self) -> None:
        # Full review first, then the reviewer flips one row to Skip: the row
        # must produce no ops and not strand its already-approved price
        # decision as "uncovered".
        self.complete_all()
        skipped = self.store.review_queue(self.run_id, "zr1", "section")["candidates"][0]
        self.store.save_decisions(
            self.run_id,
            [
                {
                    "model": "zr1",
                    "lane": "section",
                    "candidateId": skipped["candidateId"],
                    "action": "exclude_row",
                    "payload": {},
                    "resolution": "not_needed",
                }
            ],
        )
        plan = build_plan(**self.plan_inputs())
        self.assertTrue(plan["valid"], plan["report"]["blockingGaps"] or plan["coverage"]["uncoveredApprovedDecisions"])
        self.assertEqual(plan["coverage"]["uncoveredApprovedDecisions"], [])
        zr1_adds = [
            i["row"]["rpo"]
            for i in plan["stage2"]["items"]
            if i["sheet"] == "zr1_options" and i["action"] == "add"
        ]
        self.assertNotIn(skipped["rpo"], zr1_adds)

    def test_selectable_and_active_flags_flow_into_option_rows(self) -> None:
        self.complete_all()
        target = self.store.review_queue(self.run_id, "zr1", "section")["candidates"][0]
        self.store.save_decisions(
            self.run_id,
            [
                {
                    "model": "zr1",
                    "lane": "section",
                    "candidateId": target["candidateId"],
                    "action": "assign_section",
                    "payload": {"sectionId": "sec_whee_001", "selectable": False, "active": False},
                    "resolution": "approved_for_plan",
                }
            ],
        )
        plan = build_plan(**self.plan_inputs())
        row = next(
            i["row"]
            for i in plan["stage2"]["items"]
            if i["sheet"] == "zr1_options" and i["action"] == "add" and i["row"]["rpo"] == target["rpo"]
        )
        self.assertIs(row["selectable"], False)
        self.assertIs(row["active"], False)

    def test_unresolvable_variants_block_the_plan(self) -> None:
        self.complete_all()
        # Same decisions against a workbook with no variant rows at all: the
        # plan must fail closed instead of writing options without OVS rows.
        import shutil

        from openpyxl import load_workbook

        crippled = self.root / "master-no-variants.xlsx"
        shutil.copy2(self.master, crippled)
        wb = load_workbook(crippled)
        for sheet in ("model_variants", "variant_master"):
            ws = wb[sheet]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
        wb.save(crippled)
        plan = build_plan(**{**self.plan_inputs(), "workbook_path": crippled})
        self.assertFalse(plan["valid"])
        self.assertIn("no_variants_mapped", {g["kind"] for g in plan["report"]["blockingGaps"]})
