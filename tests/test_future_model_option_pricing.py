#!/usr/bin/env python3
"""Tests for filling first-pass future-model option prices from price_sched_raw."""

from __future__ import annotations

import importlib.util
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import OPTION_SOURCE_HEADERS  # noqa: E402

MODULE_PATH = ROOT / "scripts" / "apply_future_model_option_prices.py"
spec = importlib.util.spec_from_file_location("apply_future_model_option_prices", MODULE_PATH)
assert spec and spec.loader
option_prices = importlib.util.module_from_spec(spec)
sys.modules[spec.name] = option_prices
spec.loader.exec_module(option_prices)


def append_sheet(wb: Workbook, name: str, headers: list[str] | tuple[str, ...], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    for row in rows or []:
        ws.append([row.get(header, "") for header in headers])


def option_row(option_id: str, rpo: str, section_id: str, price: object = "") -> dict[str, object]:
    row: dict[str, object] = {header: "" for header in OPTION_SOURCE_HEADERS}
    row.update(
        {
            "option_id": option_id,
            "rpo": rpo,
            "price": price,
            "option_name": f"{rpo} option",
            "section_id": section_id,
            "selectable": "True",
            "display_order": "10",
            "active": "True",
            "display_behavior": "option_card",
        }
    )
    return row


def pricing_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(
        wb,
        "section_master",
        ["section_id", "section_name", "selection_mode", "active"],
        [
            {"section_id": "sec_stan_001", "section_name": "Standard", "selection_mode": "display_only", "active": "True"},
            {"section_id": "sec_engi_001", "section_name": "Engine", "selection_mode": "single", "active": "True"},
            {"section_id": "sec_lpoe_001", "section_name": "LPO Exterior", "selection_mode": "multi", "active": "True"},
            {"section_id": "sec_perf_z52_001", "section_name": "Performance Package", "selection_mode": "single", "active": "True"},
            {"section_id": "sec_perf_aero_001", "section_name": "Aero", "selection_mode": "single", "active": "True"},
            {"section_id": "sec_seat_002", "section_name": "Seats", "selection_mode": "single", "active": "True"},
        ],
    )
    append_sheet(
        wb,
        "z06_options",
        OPTION_SOURCE_HEADERS,
        [
            option_row("opt_b6p_001", "B6P", "sec_engi_001"),
            option_row("opt_std_001", "STD", "sec_stan_001", 999),
            option_row("opt_pdb_001", "PDB", "sec_perf_z52_001"),
            option_row("opt_nope_001", "NOPE", "sec_lpoe_001", 123),
            option_row("opt_5v5_001", "5V5", "sec_perf_aero_001"),
        ],
    )
    append_sheet(
        wb,
        "zr1_options",
        OPTION_SOURCE_HEADERS,
        [
            option_row("opt_ztk_001", "ZTK", "sec_perf_z52_001"),
            option_row("opt_ae4_001", "AE4", "sec_seat_002", 777),
        ],
    )
    append_sheet(
        wb,
        "zr1x_options",
        OPTION_SOURCE_HEADERS,
        [
            option_row("opt_ztk_001", "ZTK", "sec_perf_z52_001"),
            option_row("opt_pcq_001", "PCQ", "sec_lpoe_001"),
            option_row("opt_vwe_001", "VWE", "sec_lpoe_001"),
        ],
    )
    ws = wb.create_sheet("price_sched_raw")
    ws.append(["Additional Options"])
    ws.append(["", "Option Code", "Description", "Application", "List Price"])
    price_rows = [
        ("B6P", "Coupe Engine Appearance Package", "", 1895),
        ("PDB", "Carbon Fiber Wheel and Brake Package", "Z06 with ROY Carbon Fiber Wheel", 16000),
        ("PDB", "Carbon Fiber Wheel and Brake Package", "Z06 with ROZ Carbon Fiber Wheel", 17000),
        ("5V5", "Exposed Carbon Fiber spoiler", "Z06 & Grand Sport X", 5995),
        ("5V5", "Exposed Carbon Fiber spoiler", "Stingray", 8250),
        ("ZTK", "Track Performance Package", "ZR1X only; requires TOM", 1500),
        ("ZTK", "ZTK Performance Package", "ZR1 only; requires TOM; Includes J59 Ceramic Brakes", 5995),
        ("AE4", "Competition Sport Bucket Seats", "3LT/LZ Only", 595),
        ("AE4", "Competition Sport Bucket Seats", "1LT/LZ Only", 1095),
        ("PCQ", "Grille Screen Protection Package", "Stingray", 1375),
        ("PCQ", "Grille Screen Protection Package", "", 1675),
        ("VWE", "Front Grille Protection Screens", "Stingray", 695),
        ("VWE", "Front Grille Protection Screens", "", 950),
    ]
    for code, description, application, price in price_rows:
        ws.append(["", code, description, application, price])
    return wb


class FutureModelOptionPricingTests(unittest.TestCase):
    def test_plan_sets_unique_and_safe_model_specific_prices(self) -> None:
        wb = pricing_workbook()

        plan = option_prices.build_future_option_price_plan(wb, ["all"])

        z06 = plan["models"]["z06"]
        zr1 = plan["models"]["zr1"]
        zr1x = plan["models"]["zr1x"]
        self.assertEqual(z06["desired_prices"]["opt_b6p_001"], 1895)
        self.assertEqual(z06["desired_prices"]["opt_5v5_001"], 5995)
        self.assertEqual(zr1["desired_prices"]["opt_ztk_001"], 5995)
        self.assertEqual(zr1x["desired_prices"]["opt_ztk_001"], 1500)
        self.assertEqual(zr1x["desired_prices"]["opt_pcq_001"], 1675)
        self.assertEqual(zr1x["desired_prices"]["opt_vwe_001"], 950)

    def test_plan_leaves_standard_no_match_and_ambiguous_prices_blank(self) -> None:
        wb = pricing_workbook()

        plan = option_prices.build_future_option_price_plan(wb, ["all"])

        z06 = plan["models"]["z06"]
        zr1 = plan["models"]["zr1"]
        self.assertEqual(z06["desired_prices"]["opt_std_001"], "")
        self.assertEqual(z06["desired_prices"]["opt_nope_001"], "")
        self.assertEqual(z06["desired_prices"]["opt_pdb_001"], "")
        self.assertEqual(zr1["desired_prices"]["opt_ae4_001"], "")
        self.assertEqual(z06["resolution_counts"]["display_only_blank"], 1)
        self.assertEqual(z06["resolution_counts"]["no_price_match"], 1)
        self.assertEqual(z06["resolution_counts"]["ambiguous_price_deferred"], 1)
        self.assertEqual(zr1["resolution_counts"]["ambiguous_price_deferred"], 1)

    def test_apply_plan_updates_only_price_cells(self) -> None:
        wb = pricing_workbook()
        plan = option_prices.build_future_option_price_plan(wb, ["all"])

        option_prices.apply_price_plan_to_workbook(wb, plan)

        z06_rows = option_prices.rows_from_sheet(wb, "z06_options")
        zr1x_rows = option_prices.rows_from_sheet(wb, "zr1x_options")
        by_id = {row["option_id"]: row for row in z06_rows}
        self.assertEqual(by_id["opt_b6p_001"]["price"], "1895")
        self.assertEqual(by_id["opt_std_001"]["price"], "")
        self.assertEqual(by_id["opt_pdb_001"]["price"], "")
        self.assertEqual(by_id["opt_nope_001"]["price"], "")
        self.assertEqual(by_id["opt_b6p_001"]["option_name"], "B6P option")
        self.assertEqual({row["option_id"]: row["price"] for row in zr1x_rows}["opt_ztk_001"], "1500")


if __name__ == "__main__":
    unittest.main()
