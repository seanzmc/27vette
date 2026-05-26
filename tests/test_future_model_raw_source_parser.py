#!/usr/bin/env python3
"""Tests for raw order-guide source parsing used by future-model ingestion."""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import (  # noqa: E402
    build_price_schedule_rows,
    build_raw_source_blocks,
    normalize_raw_status,
)


def create_raw_sheet(wb: Workbook, name: str, variant_headers: list[tuple[str, str, str]]):
    ws = wb.create_sheet(name)
    ws.cell(7, 1, "Orderable RPO Code")
    ws.cell(7, 2, "Ref. Only RPO Code")
    ws.cell(7, 3, "Description")
    for offset, (model_label, model_code, trim) in enumerate(variant_headers, start=4):
        ws.cell(7, offset, model_label)
        ws.cell(8, offset, model_code)
        ws.cell(9, offset, trim)
    return ws


class FutureModelRawSourceParserTests(unittest.TestCase):
    def test_normalize_raw_status_preserves_note_refs_and_special_tokens(self) -> None:
        self.assertEqual(normalize_raw_status("S1"), ("standard", "1", None))
        self.assertEqual(normalize_raw_status("S2"), ("standard", "2", None))
        self.assertEqual(normalize_raw_status("A1"), ("available", "1", None))
        self.assertEqual(normalize_raw_status("A2"), ("available", "2", None))
        self.assertEqual(normalize_raw_status("--"), ("unavailable", "", None))
        self.assertEqual(normalize_raw_status("D"), ("available", "", "dealer_installed_status"))
        self.assertEqual(normalize_raw_status("A/D"), ("available", "", "dealer_installed_status"))
        self.assertEqual(normalize_raw_status("A/D1"), ("available", "1", "dealer_installed_status"))
        self.assertEqual(normalize_raw_status("A/D2"), ("available", "2", "dealer_installed_status"))
        self.assertEqual(normalize_raw_status("■"), ("standard", "", "included_in_equipment_group"))
        self.assertEqual(normalize_raw_status("■1"), ("standard", "1", "included_in_equipment_group"))
        self.assertEqual(normalize_raw_status("□"), ("standard", "", "included_in_equipment_group_upgradeable"))
        self.assertEqual(normalize_raw_status("mystery"), (None, "", "unknown_status"))

    def test_merged_z06_option_block_attaches_disclosures_and_maps_status_suffix(self) -> None:
        wb = Workbook()
        del wb[wb.sheetnames[0]]
        ws = create_raw_sheet(
            wb,
            "z06_standard_raw",
            [
                ("Z06 Coupe", "1YH07", "1LZ"),
                ("Z06 Coupe", "1YH07", "2LZ"),
                ("Z06 Coupe", "1YH07", "3LZ"),
                ("Z06 Convertible", "1YH67", "1LZ"),
                ("Z06 Convertible", "1YH67", "2LZ"),
                ("Z06 Convertible", "1YH67", "3LZ"),
            ],
        )
        ws.cell(10, 2, "AJ7")
        ws.cell(10, 3, "Airbags, frontal and side-impact")
        ws.cell(12, 3, "First disclosure should map to 1")
        ws.cell(13, 3, "Second disclosure should map to 2")
        for col in [1, 2, 4, 5, 6, 7, 8, 9]:
            ws.merge_cells(start_row=10, start_column=col, end_row=13, end_column=col)
        for col in range(4, 10):
            ws.cell(10, col, "S2")

        blocks = build_raw_source_blocks(wb)

        self.assertEqual(len(blocks), 1)
        block = blocks[0]
        self.assertEqual(block["model_key"], "z06")
        self.assertEqual(block["raw_source_sheet"], "z06_standard_raw")
        self.assertEqual(block["raw_start_row"], 10)
        self.assertEqual(block["raw_end_row"], 13)
        self.assertEqual(block["source_ref_rpo"], "AJ7")
        self.assertEqual(block["source_primary_rpo"], "AJ7")
        self.assertEqual(block["source_option_description"], "Airbags, frontal and side-impact")
        self.assertEqual(block["source_disclosure_rows"], [12, 13])
        self.assertEqual(block["source_disclosure_map"]["1"], "First disclosure should map to 1")
        self.assertEqual(block["source_disclosure_map"]["2"], "Second disclosure should map to 2")
        self.assertEqual(block["raw_statuses"]["1lz_h07"], "S2")
        self.assertEqual(block["statuses"]["1lz_h07"], "standard")
        self.assertEqual(block["status_notes"]["1lz_h07"], "2")
        self.assertEqual(block["status_note_texts"]["1lz_h07"], "Second disclosure should map to 2")

    def test_consumed_merged_child_rows_do_not_become_context_or_option_starts(self) -> None:
        wb = Workbook()
        del wb[wb.sheetnames[0]]
        ws = create_raw_sheet(
            wb,
            "z06_intextmec_raw",
            [
                ("Z06 Coupe", "1YH07", "1LZ"),
                ("Z06 Coupe", "1YH07", "2LZ"),
                ("Z06 Coupe", "1YH07", "3LZ"),
                ("Z06 Convertible", "1YH67", "1LZ"),
                ("Z06 Convertible", "1YH67", "2LZ"),
                ("Z06 Convertible", "1YH67", "3LZ"),
            ],
        )
        ws.cell(10, 1, "AAA")
        ws.cell(10, 3, "First option")
        ws.cell(12, 3, "Misleading disclosure text that is not a category")
        for col in [1, 2, 4, 5, 6, 7, 8, 9]:
            ws.merge_cells(start_row=10, start_column=col, end_row=13, end_column=col)
        for col in range(4, 10):
            ws.cell(10, col, "A1")
        ws.cell(14, 1, "BBB")
        ws.cell(14, 3, "Second option")
        for col in range(4, 10):
            ws.cell(14, col, "S")

        blocks = build_raw_source_blocks(wb)

        self.assertEqual([block["raw_start_row"] for block in blocks], [10, 14])
        self.assertEqual(blocks[0]["source_disclosure_rows"], [12])
        self.assertEqual(blocks[0]["source_disclosure_raw"], "Misleading disclosure text that is not a category")
        self.assertEqual(blocks[0]["raw_category_context"], "")
        self.assertEqual(blocks[1]["raw_category_context"], "")
        self.assertNotIn("Misleading disclosure", blocks[1].get("raw_category_context", ""))

    def test_combined_zr1_zr1x_raw_sheet_splits_statuses_by_model(self) -> None:
        wb = Workbook()
        del wb[wb.sheetnames[0]]
        ws = create_raw_sheet(
            wb,
            "zr1_zr1x_standard_raw",
            [
                ("ZR1 Coupe", "1YR07", "1LZ"),
                ("ZR1 Coupe", "1YR07", "3LZ"),
                ("ZR1 Convertible", "1YR67", "1LZ"),
                ("ZR1 Convertible", "1YR67", "3LZ"),
                ("ZR1X Coupe", "1YS07", "1LZ"),
                ("ZR1X Coupe", "1YS07", "3LZ"),
                ("ZR1X Convertible", "1YS67", "1LZ"),
                ("ZR1X Convertible", "1YS67", "3LZ"),
            ],
        )
        ws.cell(10, 2, "UQS")
        ws.cell(10, 3, "Bose Premium audio")
        for col, value in enumerate(["S", "--", "S", "--", "A", "S", "A", "S"], start=4):
            ws.cell(10, col, value)

        blocks = build_raw_source_blocks(wb)

        by_model = {block["model_key"]: block for block in blocks}
        self.assertEqual(set(by_model), {"zr1", "zr1x"})
        self.assertEqual(by_model["zr1"]["statuses"], {
            "1lz_r07": "standard",
            "3lz_r07": "unavailable",
            "1lz_r67": "standard",
            "3lz_r67": "unavailable",
        })
        self.assertEqual(by_model["zr1x"]["statuses"], {
            "1lz_s07": "available",
            "3lz_s07": "standard",
            "1lz_s67": "available",
            "3lz_s67": "standard",
        })

    def test_price_schedule_uses_section_list_price_adds_dfc_and_marks_gas_guzzler_placeholder(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "price_sched_raw"
        ws.cell(5, 1, "Base Model Prices")
        ws.cell(6, 2, "Model")
        ws.cell(6, 3, "Model Description")
        ws.cell(6, 4, "List")
        ws.cell(6, 10, "DFC")
        ws.cell(7, 4, "Price")
        ws.cell(9, 2, "1YH07")
        ws.cell(9, 3, "Corvette Z06 Coupe 1LZ")
        ws.cell(9, 4, 118900)
        ws.cell(9, 10, 2495)
        ws.cell(42, 1, "Additional Options")
        ws.cell(43, 2, "Option Code")
        ws.cell(43, 3, "Description")
        ws.cell(43, 5, "List")
        ws.cell(44, 5, "Price")
        ws.cell(47, 2, "J57")
        ws.cell(47, 3, "Carbon Ceramic Brakes")
        ws.cell(47, 4, "Included in Z07")
        ws.cell(47, 5, 9000)
        ws.cell(196, 1, "Gas Guzzler Tax")
        ws.cell(197, 2, "GGT")
        ws.cell(197, 3, "Gas Guzzler Tax")
        ws.cell(197, 5, 0)

        prices = build_price_schedule_rows(wb)

        self.assertEqual(prices["base_model_prices"][0]["model_code"], "1YH07")
        self.assertEqual(prices["base_model_prices"][0]["list_price"], 118900)
        self.assertEqual(prices["base_model_prices"][0]["dfc"], 2495)
        self.assertEqual(prices["base_model_prices"][0]["total_price"], 121395)
        self.assertEqual(prices["option_price_rows"][0]["price_rpo"], "J57")
        self.assertEqual(prices["option_price_rows"][0]["price_list"], 9000)
        self.assertTrue(prices["gas_guzzler_rows"][0]["pending_certification_placeholder"])


if __name__ == "__main__":
    unittest.main()
