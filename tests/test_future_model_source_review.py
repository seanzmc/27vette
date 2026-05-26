#!/usr/bin/env python3
"""Tests for future model source review row generation."""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SOURCE_REVIEW_HEADERS,
    build_future_model_preview,
    build_source_review_rows,
)

OPTION_HEADERS = [
    "option_id",
    "rpo",
    "price",
    "option_name",
    "description",
    "detail_raw",
    "section_id",
    "selectable",
    "display_order",
    "active",
    "display_behavior",
]


def append_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, object]] | None = None) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows or []:
        ws.append([row.get(header, None) for header in headers])


def create_review_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(wb, "stingray_options", OPTION_HEADERS, [])
    append_sheet(wb, "grandSport_options", OPTION_HEADERS, [
        {
            "option_id": "opt_aj7_existing",
            "rpo": "AJ7",
            "option_name": "Airbags, frontal and side-impact",
            "section_id": "standard_equipment",
            "active": True,
        }
    ])
    ws = wb.create_sheet("z06_standard_raw")
    ws.cell(7, 1, "Orderable RPO Code")
    ws.cell(7, 2, "Ref. Only RPO Code")
    ws.cell(7, 3, "Description")
    for col, (model, code, trim) in enumerate([
        ("Z06 Coupe", "1YH07", "1LZ"),
        ("Z06 Coupe", "1YH07", "2LZ"),
        ("Z06 Coupe", "1YH07", "3LZ"),
        ("Z06 Convertible", "1YH67", "1LZ"),
        ("Z06 Convertible", "1YH67", "2LZ"),
        ("Z06 Convertible", "1YH67", "3LZ"),
    ], start=4):
        ws.cell(7, col, model)
        ws.cell(8, col, code)
        ws.cell(9, col, trim)
    ws.cell(10, 2, "AJ7")
    ws.cell(10, 3, "Airbags, frontal and side-impact")
    ws.cell(12, 3, "Disclosure one")
    ws.cell(13, 3, "Disclosure two")
    for col in [1, 2, 4, 5, 6, 7, 8, 9]:
        ws.merge_cells(start_row=10, start_column=col, end_row=13, end_column=col)
        ws.cell(10, col, "S1" if col >= 4 else ws.cell(10, col).value)
    ws.cell(15, 3, "Missing RPO row")
    for col in range(4, 10):
        ws.cell(15, col, "A")

    eq = wb.create_sheet("z06_eqgrps_raw")
    eq.cell(7, 1, "Orderable RPO Code")
    eq.cell(7, 2, "Ref. Only RPO Code")
    eq.cell(7, 3, "Description")
    for col, (model, code, trim) in enumerate([
        ("Z06 Coupe", "1YH07", "1LZ"),
        ("Z06 Coupe", "1YH07", "2LZ"),
        ("Z06 Coupe", "1YH07", "3LZ"),
        ("Z06 Convertible", "1YH67", "1LZ"),
        ("Z06 Convertible", "1YH67", "2LZ"),
        ("Z06 Convertible", "1YH67", "3LZ"),
    ], start=4):
        eq.cell(7, col, model)
        eq.cell(8, col, code)
        eq.cell(9, col, trim)
    eq.cell(10, 2, "EQG")
    eq.cell(10, 3, "Equipment group duplicate source should be ignored")
    for col in range(4, 10):
        eq.cell(10, col, "S")
    return wb


class FutureModelSourceReviewTests(unittest.TestCase):
    def test_review_rows_preserve_raw_provenance_status_notes_and_approval_state(self) -> None:
        wb = create_review_workbook()

        rows = build_source_review_rows(wb)

        self.assertEqual(len(rows), 2)
        self.assertFalse(any(row["raw_source_sheets"] == "z06_eqgrps_raw" for row in rows))
        self.assertIn("source_disclosure_map", FUTURE_MODEL_SOURCE_REVIEW_HEADERS)
        self.assertIn("status_note_1lz_h07", FUTURE_MODEL_SOURCE_REVIEW_HEADERS)
        first = rows[0]
        self.assertEqual(first["model_key"], "z06")
        self.assertEqual(first["source_primary_rpo"], "AJ7")
        self.assertEqual(first["raw_source_spans"], "z06_standard_raw:10-13")
        self.assertIn("1=Disclosure one", first["source_disclosure_map"])
        self.assertEqual(first["raw_status_1lz_h07"], "S1")
        self.assertEqual(first["status_1lz_h07"], "standard")
        self.assertEqual(first["status_note_1lz_h07"], "1")
        self.assertEqual(first["approved_detail_raw"], "")
        self.assertEqual(first["raw_category_context"], "")
        self.assertEqual(first["candidate_section_id"], "")
        self.assertEqual(first["candidate_section_resolution"], "")
        self.assertEqual(first["approved_section_id"], "")
        self.assertEqual(first["review_status"], "needs_section_review")
        self.assertFalse(first["active"])

        missing = next(row for row in rows if row["source_option_description"] == "Missing RPO row")
        self.assertIn("missing_rpo", missing["review_flags"])
        self.assertEqual(missing["review_status"], "needs_section_review")
        self.assertFalse(missing["active"])

    def test_future_model_preview_uses_raw_sources_when_present(self) -> None:
        wb = create_review_workbook()

        preview = build_future_model_preview(wb, generated_at="2026-01-01T00:00:00+00:00")

        self.assertEqual(preview["source_mode"], "raw_order_guide")
        self.assertIn("Raw order-guide sheets were read", preview["notes"][0])
        z06 = preview["models"]["z06"]
        self.assertEqual(z06["summary"]["raw_source_block_count"], 2)
        self.assertEqual(z06["summary"]["review_row_count"], 2)
        self.assertEqual(z06["summary"]["section_resolution_counts"], {"not_assigned": 2})
        self.assertEqual(z06["review_rows"][0]["source_primary_rpo"], "AJ7")


if __name__ == "__main__":
    unittest.main()
