#!/usr/bin/env python3
"""Tests for workbook bool-like cell storage hygiene."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import save_workbook_safely  # noqa: E402
from corvette_form_generator.workbook_bool_hygiene import (  # noqa: E402
    compare_bool_like_workbooks,
    snapshot_bool_like_cells,
)


def write_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "source_rows"
    headers = ["option_id", "active", "notes"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    wb.save(path)


class WorkbookBoolHygieneTests(unittest.TestCase):
    def test_snapshot_captures_bool_like_cells_with_type_and_coordinate(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "book.xlsx"
            write_workbook(path, [{"option_id": "a", "active": "TRUE"}, {"option_id": "b", "active": False}])

            cells = snapshot_bool_like_cells(path)

        payloads = {(cell.coordinate, cell.value, cell.python_type, cell.storage_family) for cell in cells}
        self.assertIn(("B2", "TRUE", "str", "text"), payloads)
        self.assertIn(("B3", False, "bool", "excel_boolean"), payloads)

    def test_same_logical_value_text_to_excel_bool_is_rejected_by_stable_key(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            before = Path(tmpdir) / "before.xlsx"
            after = Path(tmpdir) / "after.xlsx"
            write_workbook(before, [{"option_id": "a", "active": "TRUE"}, {"option_id": "b", "active": "TRUE"}])
            # Insert a row so coordinate-only comparison would compare the wrong rows.
            write_workbook(
                after,
                [
                    {"option_id": "new", "active": "TRUE"},
                    {"option_id": "a", "active": True},
                    {"option_id": "b", "active": "TRUE"},
                ],
            )

            issues = compare_bool_like_workbooks(before, after)

        self.assertTrue(any(issue.check_id == "bool_type_family_changed" for issue in issues), issues)
        changed = next(issue for issue in issues if issue.check_id == "bool_type_family_changed")
        self.assertEqual(changed.before["row_key_values"], ["a"])
        self.assertEqual(changed.after["row_key_values"], ["a"])

    def test_row_insertion_with_matching_convention_does_not_create_coordinate_false_positive(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            before = Path(tmpdir) / "before.xlsx"
            after = Path(tmpdir) / "after.xlsx"
            write_workbook(before, [{"option_id": "a", "active": "TRUE"}, {"option_id": "b", "active": "TRUE"}])
            write_workbook(
                after,
                [
                    {"option_id": "new", "active": "TRUE"},
                    {"option_id": "a", "active": "TRUE"},
                    {"option_id": "b", "active": "TRUE"},
                ],
            )

            issues = compare_bool_like_workbooks(before, after)

        self.assertEqual(issues, [])

    def test_added_row_must_match_unambiguous_sheet_column_convention(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            before = Path(tmpdir) / "before.xlsx"
            after = Path(tmpdir) / "after.xlsx"
            write_workbook(before, [{"option_id": "a", "active": "TRUE"}])
            write_workbook(after, [{"option_id": "a", "active": "TRUE"}, {"option_id": "b", "active": True}])

            issues = compare_bool_like_workbooks(before, after)

        self.assertTrue(any(issue.check_id == "added_bool_type_convention_mismatch" for issue in issues), issues)

    def test_exact_approved_sheet_column_migration_is_allowed(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            before = Path(tmpdir) / "before.xlsx"
            after = Path(tmpdir) / "after.xlsx"
            write_workbook(before, [{"option_id": "a", "active": "TRUE"}])
            write_workbook(after, [{"option_id": "a", "active": True}])

            issues = compare_bool_like_workbooks(before, after, approved_bool_type_migrations=[("source_rows", "active")])

        self.assertEqual(issues, [])

    def test_save_workbook_safely_refuses_unapproved_bool_storage_flip(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "book.xlsx"
            write_workbook(path, [{"option_id": "a", "active": "TRUE"}])
            original = path.read_bytes()
            wb = load_workbook(path)
            try:
                wb["source_rows"]["B2"] = True
                with self.assertRaisesRegex(RuntimeError, "bool-like cell storage changed"):
                    save_workbook_safely(wb, path, loaded_mtime_ns=path.stat().st_mtime_ns)
            finally:
                wb.close()

            self.assertEqual(path.read_bytes(), original)
            self.assertFalse((Path(tmpdir) / "backups").exists())

    def test_save_workbook_safely_allows_approved_bool_storage_flip(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "book.xlsx"
            write_workbook(path, [{"option_id": "a", "active": "TRUE"}])
            wb = load_workbook(path)
            try:
                wb["source_rows"]["B2"] = True
                backup = save_workbook_safely(
                    wb,
                    path,
                    loaded_mtime_ns=path.stat().st_mtime_ns,
                    approved_bool_type_migrations=[("source_rows", "active")],
                )
            finally:
                wb.close()

            self.assertTrue(backup.exists())
            saved = load_workbook(path, read_only=True, data_only=False)
            try:
                self.assertIs(saved["source_rows"]["B2"].value, True)
            finally:
                saved.close()


if __name__ == "__main__":
    unittest.main()
