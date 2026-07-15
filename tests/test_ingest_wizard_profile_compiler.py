#!/usr/bin/env python3
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
for entry in (ROOT / "scripts", ROOT / "tests"):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from corvette_form_generator.editor_ops import extract_workbook  # noqa: E402
from corvette_form_generator.ingest.wizard.compiler import build_family_registry  # noqa: E402
from corvette_form_generator.ingest.wizard.profile_compiler import build_target_profile  # noqa: E402
from ingest_wizard_fixtures import build_master_workbook  # noqa: E402


class ProfileCompilerTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.workbook = build_master_workbook(Path(self.tmp.name) / "master.xlsx")

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def _build_zr1_profile(self):
        extract = extract_workbook(self.workbook)
        registry = build_family_registry(self.workbook, ["zr1"])["zr1"]
        return build_target_profile(
            extract,
            registry,
            target="zr1",
            comparator="z06",
            variants=[
                {
                    "variant_id": "1lz_r07",
                    "model_year": 2027,
                    "trim_level": "1lz",
                    "body_style": "coupe",
                },
                {
                    "variant_id": "3lz_r67",
                    "model_year": 2027,
                    "trim_level": "3lz",
                    "body_style": "convertible",
                },
            ],
        )

    def test_grand_sport_x_lt_profile_includes_grand_sport_exclusive_interior(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.workbook)
        sources = workbook["model_workbook_sources"]
        z06_sources = [tuple(cell.value for cell in row) for row in sources.iter_rows(min_row=2)]
        for row in z06_sources:
            if row[0] == "z06":
                sources.append(["grand_sport", *row[1:]])
        model_variants = workbook["model_variants"]
        model_variants.append(["grand_sport", "1lz_h07", 1, True, "fixture"])

        lz = workbook["LZ_Interiors"]
        lt = workbook.create_sheet("lt_interiors")
        lt.append([cell.value for cell in lz[1]])
        lt.append(
            [
                "3LT_AE4_EL9",
                "Santorini Blue Dipped with Torch Red accents",
                "Napa leather seating surfaces",
                1995,
                "Included and only available with (Z25) Grand Sport Launch Edition.",
                "G26, G4Z, GBK, GPH",
                "3LT",
                "AE4",
                "EL9",
                "",
                "",
                "",
                "sec_intc_003",
                False,
                False,
                "",
            ]
        )

        scope = workbook["model_interior_scope"]
        scope.append(
            [
                "grand_sport",
                "3LT_AE4_EL9",
                "3LT",
                True,
                "opt_z25_001",
                "Grand Sport launch interior.",
                "AE4 Competition Sport Bucket Seats",
                "EL9 Santorini Blue Dipped",
                "Napa leather",
                "EL9 Santorini Blue Dipped",
                1,
                1,
                1,
                '["3LT", "AE4 Competition Sport Bucket Seats", "EL9 Santorini Blue Dipped"]',
                "AE4 Competition Sport Bucket Seats",
                "EL9 Santorini Blue Dipped",
                1,
                "fixture",
            ]
        )
        workbook["z06_options"].append(
            [
                "opt_z25_001",
                "Z25",
                0,
                "Launch Edition",
                "Grand Sport Launch Edition",
                "",
                "sec_whee_001",
                True,
                30,
                True,
                "",
            ]
        )
        workbook["z06_ovs"].append(["opt_z25_001", "1lz_h07", "available"])
        for sheet_name in (
            "runtime_steps",
            "section_presentation",
            "context_section_master",
            "order_summary_sections",
            "step_order_summary_map",
        ):
            sheet = workbook[sheet_name]
            rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2)]
            for row in rows:
                if row[0] == "z06":
                    sheet.append(["grand_sport", *row[1:]])
        workbook.save(self.workbook)
        workbook.close()

        extract = extract_workbook(self.workbook)
        registry = build_family_registry(self.workbook, ["grand_sport_x"])["grand_sport_x"]
        profile = build_target_profile(
            extract,
            registry,
            target="grand_sport_x",
            comparator="grand_sport",
            variants=[
                {
                    "variant_id": "3lt_gsx_r07",
                    "model_year": 2027,
                    "trim_level": "3lt",
                    "body_style": "coupe",
                }
            ],
        )

        self.assertEqual(profile["trimFamily"], "LT")
        self.assertEqual(profile["interiorSheet"], "lt_interiors")
        self.assertIn("3LT_AE4_EL9", profile["interiorIds"])
        el9_scope = next(
            row
            for row in profile["rows"]
            if row["family"] == "model_interior_scope"
            and row["values"].get("interior_id") == "3LT_AE4_EL9"
        )
        self.assertEqual(el9_scope["values"]["model_key"], "grand_sport_x")
        self.assertEqual(el9_scope["values"]["trim_level"], "3LT")

        self.assertEqual(profile["requiredOptionRpoIds"], {"Z25": "opt_z25_001"})
        self.assertEqual(
            profile["requiredOptions"]["Z25"]["optionId"],
            "opt_z25_001",
        )
        self.assertEqual(
            profile["requiredOptions"]["Z25"]["sectionId"],
            "sec_whee_001",
        )
        self.assertEqual(
            profile["requiredOptions"]["Z25"]["statusByVariant"],
            {"3lt_gsx_r07": "available"},
        )
        self.assertFalse(
            any(
                row["family"] in {"options", "ovs"}
                and (
                    row["values"].get("rpo") == "Z25"
                    or row["values"].get("option_id") == "opt_z25_001"
                )
                for row in profile["rows"]
            ),
            "The profile supplies placement/identity only; target source owns option and OVS facts.",
        )

    def test_existing_target_metadata_identity_is_preserved(self) -> None:
        profile = self._build_zr1_profile()
        model = next(row for row in profile["rows"] if row["family"] == "model_master")
        self.assertEqual(model["values"]["model_label"], "ZR1")
        self.assertEqual(model["values"]["registry_key"], "zr1")

    def test_target_transformed_profile_dependencies_have_unique_evidence_ids(self) -> None:
        extract = extract_workbook(self.workbook)
        registry = build_family_registry(self.workbook, ["zr1", "zr1x"])
        profiles = [
            build_target_profile(
                extract,
                registry[target],
                target=target,
                comparator="z06",
                variants=[
                    {
                        "variant_id": f"1lz_{suffix}07",
                        "model_year": 2027,
                        "trim_level": "1lz",
                        "body_style": "coupe",
                    },
                    {
                        "variant_id": f"3lz_{suffix}67",
                        "model_year": 2027,
                        "trim_level": "3lz",
                        "body_style": "convertible",
                    },
                ],
            )
            for target, suffix in (("zr1", "r"), ("zr1x", "s"))
        ]
        fingerprints_by_id: dict[str, set[str]] = {}
        for profile in profiles:
            for row in profile["rows"]:
                for dependency in row["evidenceDependencies"]:
                    fingerprints_by_id.setdefault(dependency["evidenceId"], set()).add(
                        dependency["semanticFingerprint"]
                    )
        conflicts = {
            evidence_id: sorted(fingerprints)
            for evidence_id, fingerprints in fingerprints_by_id.items()
            if len(fingerprints) != 1
        }
        self.assertEqual(conflicts, {})

    def test_invalid_presentation_section_reference_fails_closed(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.workbook)
        workbook["section_presentation"]["B2"] = "sec_does_not_exist"
        workbook.save(self.workbook)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "presentation section"):
            self._build_zr1_profile()

    def test_shared_paint_ignores_other_paint_step_sections(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.workbook)
        workbook["section_master"].append(
            ["sec_other_paint", "Other Paint Content", "single", False, 99, "", "paint"]
        )
        workbook["z06_options"].append(
            [
                "opt_zzz_001",
                "ZZZ",
                0,
                "Unrelated paint-step content",
                "Must not copy with exterior paints",
                "",
                "sec_other_paint",
                True,
                99,
                True,
                "",
            ]
        )
        workbook["z06_ovs"].append(["opt_zzz_001", "1lz_h07", "available"])
        workbook.save(self.workbook)
        workbook.close()

        profile = self._build_zr1_profile()
        self.assertNotIn("ZZZ", profile["optionRpoIds"])

    def test_shared_paint_refuses_incomplete_comparator_availability(self) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(self.workbook)
        status_sheet = workbook["z06_ovs"]
        self.assertEqual(status_sheet["A2"].value, "opt_gba_001")
        status_sheet["C2"] = "unavailable"
        workbook.save(self.workbook)
        workbook.close()

        extract = extract_workbook(self.workbook)
        registry = build_family_registry(self.workbook, ["zr1"])["zr1"]
        with self.assertRaisesRegex(ValueError, "paint availability"):
            build_target_profile(
                extract,
                registry,
                target="zr1",
                comparator="z06",
                variants=[
                    {
                        "variant_id": "1lz_r07",
                        "model_year": 2027,
                        "trim_level": "1lz",
                        "body_style": "coupe",
                    },
                    {
                        "variant_id": "3lz_r67",
                        "model_year": 2027,
                        "trim_level": "3lz",
                        "body_style": "convertible",
                    },
                ],
            )


if __name__ == "__main__":
    unittest.main()
