#!/usr/bin/env python3
"""Focused tests for the guarded asset display override command."""

from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import set_asset_display  # noqa: E402


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def make_workbook(path: Path) -> Path:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    add_sheet(
        wb,
        "model_registry_promotion",
        ["model_key", "promoted_to_runtime", "active", "display_order"],
        [
            {"model_key": "stingray", "promoted_to_runtime": True, "active": True, "display_order": 1},
            {"model_key": "grand_sport_x", "promoted_to_runtime": True, "active": True, "display_order": 2},
        ],
    )
    add_sheet(
        wb,
        "model_workbook_sources",
        ["model_key", "source_role", "sheet_name", "active"],
        [
            {"model_key": "stingray", "source_role": "source_option_sheet", "sheet_name": "stingray_options", "active": True},
            {"model_key": "grand_sport_x", "source_role": "source_option_sheet", "sheet_name": "grandSportX_options", "active": True},
        ],
    )
    option_headers = ["option_id", "rpo", "option_name", "section_id", "active", "selectable"]
    add_sheet(
        wb,
        "stingray_options",
        option_headers,
        [
            {"option_id": "opt_aq9_001", "rpo": "AQ9", "option_name": "GT1 Seats", "section_id": "sec_seat_002", "active": True, "selectable": True},
            {"option_id": "opt_eri_001", "rpo": "ERI", "option_name": "Battery Protection Package", "section_id": "sec_perf_001", "active": True, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "grandSportX_options",
        option_headers,
        [
            {"option_id": "opt_aq9_001", "rpo": "AQ9", "option_name": "GT1 Seats", "section_id": "sec_seat_002", "active": True, "selectable": True},
            {"option_id": "opt_ae4_001", "rpo": "AE4", "option_name": "Competition Seats", "section_id": "sec_seat_002", "active": True, "selectable": True},
        ],
    )
    add_sheet(
        wb,
        "asset_map",
        ["model_key", "target_type", "target_id", "image_url", "image_alt", "image_fit", "image_position", "hover_image_url", "hover_image_alt", "hover_image_position", "active", "notes"],
        [
            {"model_key": "*", "target_type": "option", "target_id": "opt_aq9_001", "image_url": "https://example.test/aq9.png", "image_fit": "cover", "image_position": "center", "active": True},
            {"model_key": "*", "target_type": "option", "target_id": "opt_eri_001", "image_url": "https://example.test/eri.png", "image_fit": "cover", "image_position": "center", "active": True},
            {"model_key": "grand_sport_x", "target_type": "option", "target_id": "opt_ae4_001", "image_url": "https://example.test/ae4.png", "image_fit": "contain", "image_position": "center", "active": True},
        ],
    )
    wb.save(path)
    wb.close()
    return path


def test_resolve_updates_deduplicates_wildcards_and_preserves_exact_rows(tmp_path: Path) -> None:
    workbook = make_workbook(tmp_path / "master.xlsx")
    result = set_asset_display.resolve_updates(
        workbook,
        rpos=["AQ9", "AE4"],
        models=[],
        image_fit="contain",
        image_position=None,
    )

    assert len(result["resolved_rows"]) == 2
    assert result["resolved_rows"][0]["key"] == {
        "model_key": "*", "target_type": "option", "target_id": "opt_aq9_001"
    }
    assert result["resolved_rows"][0]["updates"] == {"image_fit": "contain"}
    assert result["resolved_rows"][0]["covers"] == [
        {"model_key": "grand_sport_x", "rpo": "AQ9"},
        {"model_key": "stingray", "rpo": "AQ9"},
    ]
    assert result["resolved_rows"][1]["key"]["model_key"] == "grand_sport_x"
    assert result["resolved_rows"][1]["updates"] == {}
    assert len(result["operations"]) == 1


def test_main_dry_run_routes_update_through_guarded_editor_ops(tmp_path: Path, monkeypatch) -> None:
    workbook = make_workbook(tmp_path / "master.xlsx")
    monkeypatch.setattr(
        sys,
        "argv",
        ["set_asset_display.py", "--workbook", str(workbook), "--rpo", "ERI", "--fit", "contain"],
    )
    with patch.object(set_asset_display, "apply_batch", return_value={"ok": True, "status": "dry_run"}) as guarded:
        assert set_asset_display.main() == 0

    batch = guarded.call_args.args[1]
    assert guarded.call_args.kwargs == {"write": False, "source": "asset-display-cli"}
    assert batch["items"] == [
        {
            "action": "update",
            "sheet": "asset_map",
            "key": {"model_key": "*", "target_type": "option", "target_id": "opt_eri_001"},
            "row": {"image_fit": "contain"},
        }
    ]


def test_resolve_updates_rejects_unknown_rpo(tmp_path: Path) -> None:
    workbook = make_workbook(tmp_path / "master.xlsx")
    try:
        set_asset_display.resolve_updates(
            workbook,
            rpos=["ZZZ"],
            models=[],
            image_fit="contain",
            image_position=None,
        )
    except ValueError as exc:
        assert "RPOs were not found" in str(exc)
    else:
        raise AssertionError("unknown RPO should fail closed")
