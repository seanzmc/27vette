#!/usr/bin/env python3
"""Add canonical exterior paint choices to Z-family option and OVS sheets.

Dry-run is the default. Use --write to safely save stingray_master.xlsx.
This intentionally does not call apply_future_model_option_review.py because that
writer currently blanks option prices.
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from create_future_model_option_review import OPTION_REVIEW_HEADERS  # noqa: E402
from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SOURCE_REVIEW_HEADERS,
    FUTURE_MODEL_SPECS,
    OPTION_SOURCE_HEADERS,
    OVS_SOURCE_HEADERS,
    RAW_VARIANT_IDS,
)
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely  # noqa: E402

PAINT_SECTION_ID = "sec_pain_001"
SOURCE_OPTION_SHEETS = ("stingray_options", "grandSport_options")
TARGET_MODELS = ("z06", "zr1", "zr1x")
TARGET_OPTION_SHEETS = {model: FUTURE_MODEL_SPECS[model].target_option_sheet for model in TARGET_MODELS}
TARGET_OVS_SHEETS = {model: FUTURE_MODEL_SPECS[model].target_ovs_sheet for model in TARGET_MODELS}
REVIEW_SOURCE_SHEET = "future_model_source_review"
REVIEW_OPTION_SHEET = "future_model_option_review"
PROVENANCE_SOURCE = "stingray_options; grandSport_options; user_confirmed_same_across_models"
NOTES = "User confirmed exterior paint name/RPO/description/price/compatibility are identical across Stingray, Grand Sport, Z06, ZR1, and ZR1X."
EXPECTED_PAINT_RPOS = ("G8G", "GBA", "GKA", "GBK", "GTR", "GEC", "GPH", "G4Z", "G26", "GKZ")
CANONICAL_DISPLAY_ORDERS = {rpo: str((idx + 1) * 10) for idx, rpo in enumerate(EXPECTED_PAINT_RPOS)}


def headers_for(ws) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def header_index(ws, expected: list[str] | tuple[str, ...] | None = None) -> dict[str, int]:
    headers = headers_for(ws)
    if expected is not None and headers != list(expected):
        raise RuntimeError(f"{ws.title} header drift: expected {list(expected)}, found {headers}")
    return {header: col for col, header in enumerate(headers, start=1) if header}


def row_record(ws, index: dict[str, int], row_number: int) -> dict[str, Any]:
    return {header: ws.cell(row_number, col).value for header, col in index.items()}


def normalize_price(value: Any) -> str:
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        return "0"
    numeric = float(text)
    if numeric.is_integer():
        return str(int(numeric))
    return text


def load_canonical_paint_rows(wb) -> list[dict[str, Any]]:
    source_by_sheet: dict[str, dict[str, dict[str, Any]]] = {}
    for sheet_name in SOURCE_OPTION_SHEETS:
        ws = wb[sheet_name]
        index = header_index(ws, OPTION_SOURCE_HEADERS)
        by_rpo: dict[str, dict[str, Any]] = {}
        for row_number in range(2, ws.max_row + 1):
            section_id = clean(ws.cell(row_number, index["section_id"]).value)
            if section_id != PAINT_SECTION_ID:
                continue
            record = row_record(ws, index, row_number)
            rpo = clean(record.get("rpo"))
            if not rpo:
                raise RuntimeError(f"{sheet_name} row {row_number} is a paint row without an RPO")
            if rpo in by_rpo:
                raise RuntimeError(f"{sheet_name} has duplicate paint RPO {rpo}")
            by_rpo[rpo] = record
        source_by_sheet[sheet_name] = by_rpo

    for sheet_name, rows in source_by_sheet.items():
        rpos = tuple(rows)
        if tuple(rows.keys()) != EXPECTED_PAINT_RPOS:
            raise RuntimeError(f"{sheet_name} paint RPO drift: expected {EXPECTED_PAINT_RPOS}, found {rpos}")

    canonical: list[dict[str, Any]] = []
    stingray = source_by_sheet["stingray_options"]
    grand_sport = source_by_sheet["grandSport_options"]
    comparable_fields = ("option_id", "rpo", "price", "option_name", "description", "detail_raw", "section_id", "selectable", "active", "display_behavior")
    for rpo in EXPECTED_PAINT_RPOS:
        s_row = stingray[rpo]
        gs_row = grand_sport[rpo]
        for field in comparable_fields:
            left = normalize_price(s_row.get(field)) if field == "price" else clean(s_row.get(field))
            right = normalize_price(gs_row.get(field)) if field == "price" else clean(gs_row.get(field))
            if left != right:
                raise RuntimeError(f"Paint parity mismatch for {rpo} field {field}: Stingray={left!r}, Grand Sport={right!r}")
        canonical.append(
            {
                "option_id": clean(gs_row.get("option_id")),
                "rpo": rpo,
                "price": normalize_price(gs_row.get("price")),
                "option_name": clean(gs_row.get("option_name")),
                "description": clean(gs_row.get("description")),
                "detail_raw": clean(gs_row.get("detail_raw")),
                "section_id": PAINT_SECTION_ID,
                "selectable": "True",
                "display_order": CANONICAL_DISPLAY_ORDERS[rpo],
                "active": "True",
                "display_behavior": clean(gs_row.get("display_behavior")),
            }
        )
    return canonical


def value_equal(current: Any, desired: Any, field: str) -> bool:
    if field in {"price", "display_order"}:
        return normalize_price(current) == normalize_price(desired)
    return clean(current) == clean(desired)


def set_cell(ws, row_number: int, index: dict[str, int], field: str, value: Any, changes: list[dict[str, Any]], *, write: bool, reason: str) -> None:
    current = ws.cell(row_number, index[field]).value
    if value_equal(current, value, field):
        return
    changes.append({"sheet": ws.title, "row": row_number, "field": field, "current": clean(current), "desired": clean(value), "reason": reason})
    if write:
        ws.cell(row_number, index[field]).value = value


def append_row(ws, index: dict[str, int], headers: list[str], values: dict[str, Any], changes: list[dict[str, Any]], *, write: bool, reason: str) -> int:
    row_number = ws.max_row + 1
    changes.append({"sheet": ws.title, "row": row_number, "field": "<row>", "current": "", "desired": clean(values.get(headers[0])), "reason": reason})
    if write:
        for header in headers:
            ws.cell(row_number, index[header]).value = values.get(header, "")
    return row_number


def apply_option_rows(wb, canonical_rows: list[dict[str, Any]], *, write: bool) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for model, sheet_name in TARGET_OPTION_SHEETS.items():
        ws = wb[sheet_name]
        headers = headers_for(ws)
        index = header_index(ws, OPTION_SOURCE_HEADERS)
        locations: dict[str, int] = {}
        for row_number in range(2, ws.max_row + 1):
            option_id = clean(ws.cell(row_number, index["option_id"]).value)
            if not option_id:
                continue
            if option_id in locations:
                raise RuntimeError(f"{sheet_name} has duplicate option_id {option_id}")
            locations[option_id] = row_number
        for canonical in canonical_rows:
            option_id = canonical["option_id"]
            row_number = locations.get(option_id)
            if row_number is None:
                row_number = append_row(ws, index, headers, canonical, changes, write=write, reason=f"add {model} exterior paint option")
                locations[option_id] = row_number
                continue
            for field in OPTION_SOURCE_HEADERS:
                set_cell(ws, row_number, index, field, canonical[field], changes, write=write, reason=f"canonicalize {model} exterior paint option")
    return changes


def apply_ovs_rows(wb, canonical_rows: list[dict[str, Any]], *, write: bool) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for model, sheet_name in TARGET_OVS_SHEETS.items():
        ws = wb[sheet_name]
        headers = headers_for(ws)
        index = header_index(ws, OVS_SOURCE_HEADERS)
        locations: dict[tuple[str, str], int] = {}
        for row_number in range(2, ws.max_row + 1):
            key = (clean(ws.cell(row_number, index["option_id"]).value), clean(ws.cell(row_number, index["variant_id"]).value))
            if not any(key):
                continue
            if key in locations:
                raise RuntimeError(f"{sheet_name} has duplicate OVS key {key}")
            locations[key] = row_number
        for canonical in canonical_rows:
            option_id = canonical["option_id"]
            for variant_id in FUTURE_MODEL_SPECS[model].variant_columns.values():
                desired = {"option_id": option_id, "variant_id": variant_id, "status": "available"}
                key = (option_id, variant_id)
                row_number = locations.get(key)
                if row_number is None:
                    row_number = append_row(ws, index, headers, desired, changes, write=write, reason=f"add {model} paint OVS availability")
                    locations[key] = row_number
                    continue
                for field in OVS_SOURCE_HEADERS:
                    set_cell(ws, row_number, index, field, desired[field], changes, write=write, reason=f"canonicalize {model} paint OVS availability")
    return changes


def source_review_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (clean(row.get("model_key")), clean(row.get("source_group")), clean(row.get("approved_option_id") or row.get("candidate_option_id")), clean(row.get("source_primary_rpo") or row.get("source_orderable_rpo")))


def option_review_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (clean(row.get("model_key")), clean(row.get("raw_source_sheet")), clean(row.get("raw_source_span")), clean(row.get("source_rpo")))


def status_summary(model: str) -> str:
    return "; ".join(f"{variant_id}=available" for variant_id in FUTURE_MODEL_SPECS[model].variant_columns.values())


def source_review_row(model: str, canonical: dict[str, Any]) -> dict[str, Any]:
    row = {header: "" for header in FUTURE_MODEL_SOURCE_REVIEW_HEADERS}
    option_id = clean(canonical["option_id"])
    rpo = clean(canonical["rpo"])
    row.update(
        {
            "model_key": model,
            "source_group": "exterior_paint",
            "raw_source_sheets": PROVENANCE_SOURCE,
            "raw_source_spans": f"{model}:{option_id}",
            "raw_category_context": "Paint",
            "source_orderable_rpo": rpo,
            "source_primary_rpo": rpo,
            "source_option_description": clean(canonical["option_name"]),
            "source_disclosure_raw": clean(canonical["description"]),
            "source_detail_raw": clean(canonical["detail_raw"]),
            "candidate_option_id": option_id,
            "candidate_section_id": PAINT_SECTION_ID,
            "candidate_section_resolution": "user_confirmed_canonical",
            "candidate_section_candidates": PAINT_SECTION_ID,
            "candidate_price": clean(canonical["price"]),
            "price_candidate_summary": f"canonical_price={clean(canonical['price'])}",
            "approved_option_id": option_id,
            "approved_rpo": rpo,
            "approved_price": clean(canonical["price"]),
            "approved_option_name": clean(canonical["option_name"]),
            "approved_description": clean(canonical["description"]),
            "approved_detail_raw": clean(canonical["detail_raw"]),
            "approved_section_id": PAINT_SECTION_ID,
            "approved_selectable": "True",
            "approved_display_behavior": clean(canonical["display_behavior"]),
            "approved_display_order": clean(canonical["display_order"]),
            "copy_from_model_key": "grand_sport",
            "copy_from_option_id": option_id,
            "review_status": "approved",
            "review_reason": "user_confirmed_same_across_models",
            "active": "True",
            "notes": NOTES,
        }
    )
    for variant_id in RAW_VARIANT_IDS:
        if variant_id in FUTURE_MODEL_SPECS[model].variant_columns.values():
            row[f"raw_status_{variant_id}"] = "available"
            row[f"status_{variant_id}"] = "available"
    return row


def option_review_row(model: str, canonical: dict[str, Any]) -> dict[str, Any]:
    option_id = clean(canonical["option_id"])
    rpo = clean(canonical["rpo"])
    row = {header: "" for header in OPTION_REVIEW_HEADERS}
    row.update(
        {
            "model_key": model,
            "raw_source_sheet": PROVENANCE_SOURCE,
            "raw_source_span": f"{model}:{option_id}",
            "orderable_rpo": rpo,
            "source_rpo": rpo,
            "source_option_description": clean(canonical["option_name"]),
            "source_disclosure_raw": clean(canonical["description"]),
            "raw_status_summary": status_summary(model),
            "normalized_status_summary": status_summary(model),
            "suggested_option_id": option_id,
            "suggested_section_id": PAINT_SECTION_ID,
            "suggested_display_order": clean(canonical["display_order"]),
            "suggested_copy_from": f"grand_sport:{option_id}",
            "final_option_id": option_id,
            "final_option_name": clean(canonical["option_name"]),
            "final_description": clean(canonical["description"]),
            "final_detail_raw": clean(canonical["detail_raw"]),
            "final_section_id": PAINT_SECTION_ID,
            "final_selectable": "True",
            "final_display_order": clean(canonical["display_order"]),
            "final_display_behavior": clean(canonical["display_behavior"]),
            "review_status": "approved",
            "active": "True",
            "notes": NOTES,
        }
    )
    return row


def apply_review_rows(wb, canonical_rows: list[dict[str, Any]], *, write: bool) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    specs = [
        (REVIEW_SOURCE_SHEET, list(FUTURE_MODEL_SOURCE_REVIEW_HEADERS), source_review_key, source_review_row, "source-review"),
        (REVIEW_OPTION_SHEET, list(OPTION_REVIEW_HEADERS), option_review_key, option_review_row, "option-review"),
    ]
    for sheet_name, expected_headers, key_fn, builder, label in specs:
        ws = wb[sheet_name]
        headers = headers_for(ws)
        if headers != expected_headers:
            raise RuntimeError(f"{sheet_name} header drift: expected {expected_headers}, found {headers}")
        index = header_index(ws)
        locations: dict[tuple[str, str, str, str], int] = {}
        for row_number in range(2, ws.max_row + 1):
            record = row_record(ws, index, row_number)
            key = key_fn(record)
            if not any(key):
                continue
            if key in locations:
                # Only fail on duplicate keys for rows this helper owns. Existing raw-review
                # data can duplicate broader blank/provisional keys.
                if clean(record.get("source_group")) == "exterior_paint" or clean(record.get("raw_source_sheet")) == PROVENANCE_SOURCE:
                    raise RuntimeError(f"{sheet_name} has duplicate owned review key {key}")
                continue
            locations[key] = row_number
        for model in TARGET_MODELS:
            for canonical in canonical_rows:
                desired = builder(model, canonical)
                key = key_fn(desired)
                row_number = locations.get(key)
                if row_number is None:
                    row_number = append_row(ws, index, headers, desired, changes, write=write, reason=f"add Z exterior paint {label} row")
                    locations[key] = row_number
                    continue
                for field in headers:
                    set_cell(ws, row_number, index, field, desired.get(field, ""), changes, write=write, reason=f"canonicalize Z exterior paint {label} row")
    return changes


def verify_saved(path: Path, canonical_rows: list[dict[str, Any]]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        paint_ids = {row["option_id"] for row in canonical_rows}
        expected_by_model = {model: len(list(FUTURE_MODEL_SPECS[model].variant_columns.values())) * len(canonical_rows) for model in TARGET_MODELS}
        verification: dict[str, Any] = {"option_paint_rows": {}, "ovs_paint_rows": {}, "review_rows": {}}
        for model, sheet_name in TARGET_OPTION_SHEETS.items():
            ws = wb[sheet_name]
            index = header_index(ws, OPTION_SOURCE_HEADERS)
            rows: dict[str, dict[str, str]] = {}
            for row_number in range(2, ws.max_row + 1):
                record = {field: clean(ws.cell(row_number, index[field]).value) for field in OPTION_SOURCE_HEADERS}
                if record["option_id"] in paint_ids:
                    rows[record["option_id"]] = record
            if len(rows) != len(canonical_rows):
                raise RuntimeError(f"{sheet_name}: expected {len(canonical_rows)} paint rows, found {len(rows)}")
            for canonical in canonical_rows:
                actual = rows.get(canonical["option_id"])
                if not actual:
                    raise RuntimeError(f"{sheet_name}: missing {canonical['option_id']}")
                for field in OPTION_SOURCE_HEADERS:
                    if not value_equal(actual[field], canonical[field], field):
                        raise RuntimeError(f"{sheet_name}: {canonical['option_id']} field {field} expected {canonical[field]!r}, found {actual[field]!r}")
            verification["option_paint_rows"][sheet_name] = len(rows)
        for model, sheet_name in TARGET_OVS_SHEETS.items():
            ws = wb[sheet_name]
            index = header_index(ws, OVS_SOURCE_HEADERS)
            count = 0
            bad: list[str] = []
            for row_number in range(2, ws.max_row + 1):
                option_id = clean(ws.cell(row_number, index["option_id"]).value)
                if option_id not in paint_ids:
                    continue
                count += 1
                status = clean(ws.cell(row_number, index["status"]).value)
                if status != "available":
                    bad.append(f"row {row_number} {option_id} status={status}")
            if count != expected_by_model[model]:
                raise RuntimeError(f"{sheet_name}: expected {expected_by_model[model]} paint OVS rows, found {count}")
            if bad:
                raise RuntimeError(f"{sheet_name}: bad paint OVS statuses: {bad[:5]}")
            verification["ovs_paint_rows"][sheet_name] = count
        for sheet_name, expected_headers, key_fn, _builder, _label in [
            (REVIEW_SOURCE_SHEET, list(FUTURE_MODEL_SOURCE_REVIEW_HEADERS), source_review_key, source_review_row, ""),
            (REVIEW_OPTION_SHEET, list(OPTION_REVIEW_HEADERS), option_review_key, option_review_row, ""),
        ]:
            ws = wb[sheet_name]
            headers = headers_for(ws)
            if headers != expected_headers:
                raise RuntimeError(f"{sheet_name}: header drift after save")
            index = header_index(ws)
            owned = 0
            for row_number in range(2, ws.max_row + 1):
                record = {header: clean(ws.cell(row_number, index[header]).value) for header in headers}
                if sheet_name == REVIEW_SOURCE_SHEET and record.get("source_group") == "exterior_paint" and record.get("model_key") in TARGET_MODELS:
                    owned += 1
                if sheet_name == REVIEW_OPTION_SHEET and record.get("raw_source_sheet") == PROVENANCE_SOURCE and record.get("model_key") in TARGET_MODELS:
                    owned += 1
            if owned < len(TARGET_MODELS) * len(canonical_rows):
                raise RuntimeError(f"{sheet_name}: expected at least {len(TARGET_MODELS) * len(canonical_rows)} owned paint review rows, found {owned}")
            verification["review_rows"][sheet_name] = owned
        return verification
    finally:
        wb.close()


def summarize_changes(changes: list[dict[str, Any]]) -> dict[str, Any]:
    by_sheet = Counter(change["sheet"] for change in changes)
    by_reason = Counter(change["reason"] for change in changes)
    return {
        "total_changes": len(changes),
        "by_sheet": dict(sorted(by_sheet.items())),
        "by_reason": dict(sorted(by_reason.items())),
        "sample": changes[:25],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="Persist changes to stingray_master.xlsx. Default is dry-run only.")
    parser.add_argument("--dry-run", action="store_true", help="Explicit dry-run flag for readability; this is the default.")
    args = parser.parse_args(argv)
    write = bool(args.write)

    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to {'write' if write else 'inspect'} workbook while Excel lock file exists: {lock_path}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    backup_path = None
    try:
        canonical_rows = load_canonical_paint_rows(wb)
        option_changes = apply_option_rows(wb, canonical_rows, write=write)
        ovs_changes = apply_ovs_rows(wb, canonical_rows, write=write)
        review_changes = apply_review_rows(wb, canonical_rows, write=write)
        all_changes = option_changes + ovs_changes + review_changes
        if write and all_changes:
            backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved(WORKBOOK_PATH, canonical_rows) if write else None
    print(
        json.dumps(
            {
                "mode": "write" if write else "dry-run",
                "workbook": str(WORKBOOK_PATH),
                "backup": str(backup_path) if backup_path else None,
                "canonical_paint_rows": [{"option_id": row["option_id"], "rpo": row["rpo"], "price": row["price"], "display_order": row["display_order"]} for row in canonical_rows],
                "changes": summarize_changes(all_changes),
                "verification": verification,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
