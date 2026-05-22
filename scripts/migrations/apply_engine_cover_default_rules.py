#!/usr/bin/env python3
"""Restore workbook-owned coupe engine-cover default-selection rules.

The runtime already serves workbook-authored default_selection_rules generically for
Stingray suspension (FE1) and exhaust (NGA).  Coupe LS6 engine covers should use
that same source-data contract: BC7 is the default/restored coupe engine cover,
and the existing engine-cover exclusive group prevents BC7 from being re-added
while a paid cover peer is selected.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[1]
ROOT = SCRIPTS_DIR.parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import (  # noqa: E402
    clean,
    excel_lock_path,
    rows_from_sheet,
    save_workbook_safely,
    write_sheet,
)

WORKBOOK_PATH = ROOT / "stingray_master.xlsx"
SHEET_NAME = "default_selection_rules"
HEADERS = [
    "model_key",
    "rule_id",
    "target_option_id",
    "condition_type",
    "condition_id",
    "body_style_scope",
    "trim_level_scope",
    "variant_scope",
    "priority",
    "active",
    "notes",
]

ENGINE_COVER_DEFAULT_RULES: list[dict[str, Any]] = [
    {
        "model_key": "stingray",
        "rule_id": "default_bc7",
        "target_option_id": "opt_bc7_001",
        "condition_type": "always",
        "condition_id": "",
        "body_style_scope": "coupe",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "priority": 40,
        "active": "TRUE",
        "notes": (
            "Workbook-owned coupe BC7 default/restoration rule. Runtime serves this through "
            "generic defaultSelectionRules, while grp_ls6_engine_covers prevents BC7 from "
            "re-adding when a paid engine-cover peer is selected."
        ),
    },
    {
        "model_key": "grand_sport",
        "rule_id": "gs_default_bc7_coupe",
        "target_option_id": "opt_bc7_001",
        "condition_type": "always",
        "condition_id": "",
        "body_style_scope": "coupe",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "priority": 40,
        "active": "TRUE",
        "notes": (
            "Workbook-owned Grand Sport coupe BC7 default/restoration rule. Runtime serves "
            "this through generic defaultSelectionRules, while gs_excl_ls6_engine_covers "
            "prevents BC7 from re-adding when a paid engine-cover peer is selected."
        ),
    },
]


def row_key(row: dict[str, Any]) -> tuple[str, str]:
    return (clean(row.get("model_key")).lower(), clean(row.get("rule_id")).lower())


def upsert_rules(rows: list[dict[str, Any]]) -> dict[str, int]:
    existing = {row_key(row): row for row in rows}
    counts = {"inserted": 0, "updated": 0, "skipped": 0}
    for rule in ENGINE_COVER_DEFAULT_RULES:
        key = row_key(rule)
        existing_row = existing.get(key)
        if existing_row is None:
            rows.append(dict(rule))
            existing[key] = rows[-1]
            counts["inserted"] += 1
            continue

        changed = False
        for field in HEADERS:
            next_value = rule.get(field, "")
            if clean(existing_row.get(field)) != clean(next_value):
                existing_row[field] = next_value
                changed = True
        counts["updated" if changed else "skipped"] += 1
    return counts


def verify_saved_rows(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        rows = rows_from_sheet(wb, SHEET_NAME)
        by_key = {row_key(row): row for row in rows}
        missing: list[str] = []
        wrong: list[str] = []
        for rule in ENGINE_COVER_DEFAULT_RULES:
            key = row_key(rule)
            row = by_key.get(key)
            if row is None:
                missing.append(rule["rule_id"])
                continue
            for field in ("target_option_id", "condition_type", "body_style_scope", "trim_level_scope", "variant_scope", "active"):
                if clean(row.get(field)) != clean(rule[field]):
                    wrong.append(f"{rule['rule_id']}.{field}: expected {rule[field]!r}, found {row.get(field)!r}")
        if missing or wrong:
            details = "; ".join([*(f"missing {item}" for item in missing), *wrong])
            raise RuntimeError(f"Saved workbook verification failed: {details}")
    finally:
        wb.close()


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to run; Excel lock file is present: {lock_path}. Close Excel first.")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    try:
        rows = rows_from_sheet(wb, SHEET_NAME) if SHEET_NAME in wb.sheetnames else []
        counts = upsert_rules(rows)
        write_sheet(wb, SHEET_NAME, HEADERS, rows)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verify_saved_rows(WORKBOOK_PATH)
    print(f"Workbook updated: {WORKBOOK_PATH}")
    print(f"Backup created: {backup_path}")
    print(f"{SHEET_NAME}: inserted={counts['inserted']} updated={counts['updated']} skipped={counts['skipped']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
