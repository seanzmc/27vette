#!/usr/bin/env python3
"""Add workbook-authored Z06 Z07/PDB strict peer exclusion groups.

This is a targeted, idempotent source-workbook migration for the approved
z07-rule-fix spec. It only upserts rows in z06_rule_groups and
z06_rule_group_members, then verifies the saved workbook on disk.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[1]
ROOT = SCRIPTS_DIR.parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import (  # noqa: E402
    clean,
    excel_lock_path,
    rows_from_sheet,
    save_workbook_safely,
    write_sheet,
)

WORKBOOK_PATH = ROOT / "stingray_master.xlsx"
GROUP_SHEET = "z06_rule_groups"
MEMBER_SHEET = "z06_rule_group_members"

GROUP_ROWS: list[dict[str, Any]] = [
    {
        "group_id": "z06_group_z07_excludes_non_z07_aero",
        "group_type": "excludes_any",
        "source_id": "opt_z07_001",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "disabled_reason": "Not available while Z07 Performance Package is selected. Choose T0F or T0G.",
        "active": "True",
        "notes": "Z07 permits only T0F/T0G aero choices; T0E/5ZV should remain disabled even after switching from default T0F to T0G.",
    },
    {
        "group_id": "z06_group_z07_excludes_j56_brakes",
        "group_type": "excludes_any",
        "source_id": "opt_z07_001",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "disabled_reason": "J56 performance brakes are not available while Z07 is selected; Z07 includes J57 carbon ceramic brakes.",
        "active": "True",
        "notes": "Z07 always includes J57; J56 must not be selectable or restorable while Z07 remains selected.",
    },
    {
        "group_id": "z06_group_pdb_excludes_j56_brakes",
        "group_type": "excludes_any",
        "source_id": "opt_pdb_001",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "disabled_reason": "J56 performance brakes are not available while PDB is selected; PDB includes J57 carbon ceramic brakes.",
        "active": "True",
        "notes": "PDB includes J57 directly and must not allow switching back to J56 while selected.",
    },
]

MEMBER_ROWS: list[dict[str, Any]] = [
    {"group_id": "z06_group_z07_excludes_non_z07_aero", "target_id": "opt_t0e_001", "active": "True"},
    {"group_id": "z06_group_z07_excludes_non_z07_aero", "target_id": "opt_5zv_001", "active": "True"},
    {"group_id": "z06_group_z07_excludes_j56_brakes", "target_id": "opt_j56_001", "active": "True"},
    {"group_id": "z06_group_pdb_excludes_j56_brakes", "target_id": "opt_j56_001", "active": "True"},
]


def headers_for(wb, sheet_name: str) -> list[str]:
    ws = wb[sheet_name]
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1) if clean(ws.cell(1, col).value)]


def require_headers(headers: list[str], required: set[str], sheet_name: str) -> None:
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"{sheet_name} is missing required headers: {', '.join(missing)}")


def row_diff(existing: dict[str, Any] | None, desired: dict[str, Any], headers: list[str]) -> dict[str, tuple[str, str]]:
    diff: dict[str, tuple[str, str]] = {}
    for field, desired_value in desired.items():
        if field not in headers:
            continue
        old_value = "" if existing is None else clean(existing.get(field))
        new_value = clean(desired_value)
        if old_value != new_value:
            diff[field] = (old_value, new_value)
    return diff


def upsert_group_rows(rows: list[dict[str, Any]], headers: list[str]) -> dict[str, int]:
    require_headers(headers, {"group_id", "group_type", "source_id", "active"}, GROUP_SHEET)
    existing = {clean(row.get("group_id")): row for row in rows}
    counts = {"inserted": 0, "updated": 0, "skipped": 0}
    for desired in GROUP_ROWS:
        key = clean(desired["group_id"])
        row = existing.get(key)
        if row is None:
            rows.append({header: desired.get(header, "") for header in headers})
            existing[key] = rows[-1]
            counts["inserted"] += 1
            continue
        changed = False
        for field, value in desired.items():
            if field in headers and clean(row.get(field)) != clean(value):
                row[field] = value
                changed = True
        counts["updated" if changed else "skipped"] += 1
    return counts


def member_key(row: dict[str, Any]) -> tuple[str, str]:
    return clean(row.get("group_id")), clean(row.get("target_id"))


def upsert_member_rows(rows: list[dict[str, Any]], headers: list[str]) -> dict[str, int]:
    require_headers(headers, {"group_id", "target_id"}, MEMBER_SHEET)
    existing = {member_key(row): row for row in rows}
    counts = {"inserted": 0, "updated": 0, "skipped": 0}
    for desired in MEMBER_ROWS:
        key = member_key(desired)
        row = existing.get(key)
        if row is None:
            rows.append({header: desired.get(header, "") for header in headers})
            existing[key] = rows[-1]
            counts["inserted"] += 1
            continue
        changed = False
        for field, value in desired.items():
            if field in headers and clean(row.get(field)) != clean(value):
                row[field] = value
                changed = True
        counts["updated" if changed else "skipped"] += 1
    return counts


def verify_saved_rows(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        group_headers = headers_for(wb, GROUP_SHEET)
        member_headers = headers_for(wb, MEMBER_SHEET)
        group_rows = rows_from_sheet(wb, GROUP_SHEET)
        member_rows = rows_from_sheet(wb, MEMBER_SHEET)
        groups = {clean(row.get("group_id")): row for row in group_rows}
        members = {member_key(row) for row in member_rows}

        errors: list[str] = []
        for desired in GROUP_ROWS:
            row = groups.get(clean(desired["group_id"]))
            if row is None:
                errors.append(f"missing group {desired['group_id']}")
                continue
            for field in ["group_type", "source_id", "body_style_scope", "trim_level_scope", "variant_scope", "disabled_reason", "active"]:
                if field in group_headers and clean(row.get(field)) != clean(desired[field]):
                    errors.append(f"{desired['group_id']}.{field}: expected {desired[field]!r}, found {row.get(field)!r}")
        for desired in MEMBER_ROWS:
            if member_key(desired) not in members:
                errors.append(f"missing member {desired['group_id']} -> {desired['target_id']}")
            elif "active" in member_headers:
                row = next(row for row in member_rows if member_key(row) == member_key(desired))
                if clean(row.get("active")) != clean(desired["active"]):
                    errors.append(f"{desired['group_id']}->{desired['target_id']}.active: expected True, found {row.get('active')!r}")
        if errors:
            raise RuntimeError("Saved workbook verification failed: " + "; ".join(errors))
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="Persist workbook changes. Defaults to dry-run.")
    args = parser.parse_args()

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to run; Excel lock file is present: {lock_path}. Close Excel first.")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    try:
        group_headers = headers_for(wb, GROUP_SHEET)
        member_headers = headers_for(wb, MEMBER_SHEET)
        group_rows = rows_from_sheet(wb, GROUP_SHEET)
        member_rows = rows_from_sheet(wb, MEMBER_SHEET)
        group_counts = upsert_group_rows(group_rows, group_headers)
        member_counts = upsert_member_rows(member_rows, member_headers)

        print(f"{GROUP_SHEET}: inserted={group_counts['inserted']} updated={group_counts['updated']} skipped={group_counts['skipped']}")
        print(f"{MEMBER_SHEET}: inserted={member_counts['inserted']} updated={member_counts['updated']} skipped={member_counts['skipped']}")
        if not args.write:
            print("Dry run only; pass --write to save workbook changes.")
            return 0

        write_sheet(wb, GROUP_SHEET, group_headers, group_rows)
        write_sheet(wb, MEMBER_SHEET, member_headers, member_rows)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verify_saved_rows(WORKBOOK_PATH)
    print(f"Workbook updated: {WORKBOOK_PATH}")
    print(f"Backup created: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
