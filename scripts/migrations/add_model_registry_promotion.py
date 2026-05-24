#!/usr/bin/env python3
"""Add workbook-owned runtime registry promotion metadata for Phase 6."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.registry_promotion import MODEL_REGISTRY_PROMOTION_HEADERS  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, rows_from_sheet, save_workbook_safely  # noqa: E402

WORKBOOK_PATH = ROOT / "stingray_master.xlsx"

PROMOTION_ROWS = [
    {
        "model_key": "stingray",
        "registry_key": "stingray",
        "promoted_to_runtime": True,
        "default_model": True,
        "artifact_path": "",
        "artifact_type": "current_generation",
        "legacy_alias": "STINGRAY_FORM_DATA",
        "active": True,
        "display_order": 1,
        "notes": "Phase 6 current generated Stingray registry entry.",
    },
    {
        "model_key": "grand_sport",
        "registry_key": "grandSport",
        "promoted_to_runtime": True,
        "default_model": False,
        "artifact_path": "form-output/inspection/grand-sport-form-data-draft.json",
        "artifact_type": "draft_artifact",
        "legacy_alias": "",
        "active": True,
        "display_order": 2,
        "notes": "Phase 6 Grand Sport draft artifact promoted to branch runtime registry.",
    },
    {
        "model_key": "z06",
        "registry_key": "z06",
        "promoted_to_runtime": False,
        "default_model": False,
        "artifact_path": "",
        "artifact_type": "",
        "legacy_alias": "",
        "active": False,
        "display_order": 3,
        "notes": "Future model remains unpromoted until source data review is complete.",
    },
    {
        "model_key": "zr1",
        "registry_key": "zr1",
        "promoted_to_runtime": False,
        "default_model": False,
        "artifact_path": "",
        "artifact_type": "",
        "legacy_alias": "",
        "active": False,
        "display_order": 4,
        "notes": "Future model remains unpromoted until source data review is complete.",
    },
    {
        "model_key": "zr1x",
        "registry_key": "zr1x",
        "promoted_to_runtime": False,
        "default_model": False,
        "artifact_path": "",
        "artifact_type": "",
        "legacy_alias": "",
        "active": False,
        "display_order": 5,
        "notes": "Future model remains unpromoted until source data review is complete.",
    },
]


def ensure_sheet(wb: Any, sheet_name: str, headers: list[str]) -> Any:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        return ws
    ws = wb[sheet_name]
    existing_headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    for header in headers:
        if header not in existing_headers:
            ws.cell(1, len(existing_headers) + 1, header)
            existing_headers.append(header)
    return ws


def header_indexes(ws: Any) -> dict[str, int]:
    return {clean(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1) if clean(ws.cell(1, col).value)}


def upsert_rows(wb: Any) -> int:
    ws = ensure_sheet(wb, "model_registry_promotion", list(MODEL_REGISTRY_PROMOTION_HEADERS))
    indexes = header_indexes(ws)
    existing: dict[str, int] = {}
    for row_number in range(2, ws.max_row + 1):
        model_key = clean(ws.cell(row_number, indexes["model_key"]).value)
        if model_key:
            existing[model_key] = row_number

    changed = 0
    for desired in PROMOTION_ROWS:
        model_key = clean(desired["model_key"])
        row_number = existing.get(model_key)
        if row_number is None:
            row_number = ws.max_row + 1
            existing[model_key] = row_number
        for header in MODEL_REGISTRY_PROMOTION_HEADERS:
            value = desired.get(header, "")
            column = indexes[header]
            if clean(ws.cell(row_number, column).value) != clean(value):
                ws.cell(row_number, column, value)
                changed += 1
    return changed


def verify_saved_rows(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        headers = [clean(wb["model_registry_promotion"].cell(1, col).value) for col in range(1, wb["model_registry_promotion"].max_column + 1)]
        if headers != list(MODEL_REGISTRY_PROMOTION_HEADERS):
            raise RuntimeError(f"model_registry_promotion header drift: {headers}")
        rows = rows_from_sheet(wb, "model_registry_promotion")
        promoted = [row for row in rows if row.get("promoted_to_runtime") == "True" and row.get("active") == "True"]
        future_unpromoted = {
            row.get("model_key"): row
            for row in rows
            if row.get("model_key") in {"z06", "zr1", "zr1x"}
        }
        if [row.get("model_key") for row in promoted] != ["stingray", "grand_sport"]:
            raise RuntimeError(f"Expected only Stingray and Grand Sport promoted; found {promoted}")
        if sum(1 for row in promoted if row.get("default_model") == "True") != 1:
            raise RuntimeError("Expected exactly one promoted default model")
        if any(row.get("promoted_to_runtime") != "False" or row.get("active") != "False" for row in future_unpromoted.values()):
            raise RuntimeError(f"Future model rows must remain inactive/unpromoted: {future_unpromoted}")
        return {
            "headers": headers,
            "promoted_models": [row.get("model_key") for row in promoted],
            "default_model": next(row.get("model_key") for row in promoted if row.get("default_model") == "True"),
            "future_models_unpromoted": sorted(str(model_key) for model_key in future_unpromoted),
        }
    finally:
        wb.close()


def main() -> None:
    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {lock_path}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    try:
        changed = upsert_rows(wb)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_rows(WORKBOOK_PATH)
    print(
        json.dumps(
            {
                "workbook": str(WORKBOOK_PATH),
                "backup": str(backup_path),
                "cell_updates": changed,
                "verification": verification,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
