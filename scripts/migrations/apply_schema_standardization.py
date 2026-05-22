#!/usr/bin/env python3
"""Apply approved workbook schema standardization migrations.

This script only edits workbook source sheets and saves through save_workbook_safely().
It is intentionally idempotent so later schema-standardization passes can re-run it.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SCRIPTS_DIR = Path(__file__).resolve().parents[1]
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import save_workbook_safely  # noqa: E402

CANONICAL_INTERIOR_HEADERS = [
    "interior_id",
    "Interior Name",
    "Material",
    "Price",
    "Detail from Disclosure",
    "Color Overrides",
    "Trim",
    "Seat",
    "Interior Code",
    "Suede",
    "Stitch",
    "Two Tone",
    "section_id",
    "active_for_stingray",
    "requires_r6x",
    "included_option_id",
]

BOOLEAN_COLUMNS: dict[str, tuple[str, ...]] = {
    "stingray_options": ("selectable", "active"),
    "grandSport_options": ("selectable", "active"),
    "rule_mapping": ("review_flag",),
    "grandSport_rule_mapping": ("review_flag",),
    "price_rules": ("review_flag",),
    "grandSport_price_rules": ("review_flag",),
    "rule_groups": ("active",),
    "grandSport_rule_groups": ("active",),
    "rule_group_members": ("active",),
    "grandSport_rule_group_members": ("active",),
    "exclusive_groups": ("active",),
    "grandSport_exclusive_groups": ("active",),
    "exclusive_group_members": ("active",),
    "grandSport_exclusive_members": ("active",),
    "grandSport_variant_overrides": ("active", "selectable"),
    "lt_interiors": ("active_for_stingray", "requires_r6x"),
    "LZ_Interiors": ("active_for_stingray", "requires_r6x"),
    "model_interior_scope": ("active",),
    "interior_components": ("active",),
}

PRICE_COLUMNS: dict[str, tuple[str, ...]] = {
    "stingray_options": ("price",),
    "grandSport_options": ("price",),
    "price_rules": ("price_value",),
    "grandSport_price_rules": ("price_value",),
    "PriceRef": ("Price",),
    "lt_interiors": ("Price",),
    "LZ_Interiors": ("Price",),
}

RPO_COLUMNS: dict[str, tuple[str, ...]] = {
    "stingray_options": ("rpo",),
    "grandSport_options": ("rpo",),
    "interior_components": ("rpo",),
}

LIFECYCLE_COLUMNS = ["normalization_status", "normalization_reason", "replacement_group_id", "replacement_rule_id"]

ACTION_REASONS = {
    "omit_grouped_requirement": "Retained source row is represented by a workbook-authored grouped requirement.",
    "omit_grouped_exclusion": "Retained source row is represented by a workbook-authored grouped exclusion.",
    "omit_replaced_by_d3v_include": "Retained source row is replaced by the direct D3V include rule.",
    "omit_soft_defaulted_caliper": "Retained source row is represented by a workbook-authored soft default instead of a hard include.",
    "omit_redundant_scoped_duplicate": "Retained source row is a scoped duplicate of an active workbook rule.",
    "preserve_runtime_exclude": "Source row is intentionally preserved as a runtime exclusion despite grouped-rule cleanup.",
    "omit_replaced_by_brake_exclusive_group": "Retained source row is represented by the workbook-authored performance brakes exclusive group.",
}

ACTION_STATUS = {
    "omit_grouped_requirement": "omitted",
    "omit_grouped_exclusion": "omitted",
    "omit_replaced_by_d3v_include": "replaced",
    "omit_soft_defaulted_caliper": "replaced",
    "omit_redundant_scoped_duplicate": "replaced",
    "preserve_runtime_exclude": "preserved",
    "omit_replaced_by_brake_exclusive_group": "replaced",
}


def headers(ws) -> list[str]:
    return [str(ws.cell(1, col).value).strip() if ws.cell(1, col).value is not None else "" for col in range(1, ws.max_column + 1)]


def header_map(ws) -> dict[str, int]:
    return {header: index for index, header in enumerate(headers(ws), start=1) if header}


def iter_records(ws):
    mapping = header_map(ws)
    for row_number in range(2, ws.max_row + 1):
        if not any(ws.cell(row_number, col).value is not None for col in range(1, ws.max_column + 1)):
            continue
        yield row_number, {header: ws.cell(row_number, col).value for header, col in mapping.items()}


def as_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "1", "y"}:
        return True
    if text in {"false", "no", "0", "n"}:
        return False
    raise ValueError(f"Cannot convert {value!r} to boolean")


def as_number_or_blank(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError(f"Boolean {value!r} is not a valid price")
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return None
    number = float(text)
    return int(number) if number.is_integer() else number


def normalize_boolean_columns(wb) -> int:
    changed = 0
    for sheet, columns in BOOLEAN_COLUMNS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        mapping = header_map(ws)
        for column in columns:
            if column not in mapping:
                continue
            col_idx = mapping[column]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is None or cell.value == "":
                    continue
                converted = as_bool(cell.value)
                if cell.value is not converted:
                    cell.value = converted
                    changed += 1
    return changed


def normalize_rpo_columns(wb) -> int:
    changed = 0
    for sheet, columns in RPO_COLUMNS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        mapping = header_map(ws)
        for column in columns:
            if column not in mapping:
                continue
            col_idx = mapping[column]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is None or cell.value == "":
                    continue
                text = str(int(cell.value)) if isinstance(cell.value, float) and cell.value.is_integer() else str(cell.value).strip()
                text = text.upper()
                if cell.value != text:
                    cell.value = text
                    changed += 1
                cell.number_format = "@"
    return changed


def normalize_price_columns(wb) -> int:
    changed = 0
    for sheet, columns in PRICE_COLUMNS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        mapping = header_map(ws)
        for column in columns:
            if column not in mapping:
                continue
            col_idx = mapping[column]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is None or cell.value == "":
                    continue
                converted = as_number_or_blank(cell.value)
                if cell.value != converted or type(cell.value) is not type(converted):
                    cell.value = converted
                    changed += 1
    return changed


def normalize_lz_interiors(wb) -> int:
    if "LZ_Interiors" not in wb.sheetnames:
        return 0
    ws = wb["LZ_Interiors"]
    old_mapping = header_map(ws)
    source_rows = [record for _, record in iter_records(ws)]
    normalized_rows: list[dict[str, Any]] = []
    for row in source_rows:
        interior_id = row.get("interior_id") or row.get("ID") or ""
        trim = row.get("Trim") or ""
        normalized_rows.append(
            {
                "interior_id": interior_id,
                "Interior Name": row.get("Interior Name"),
                "Material": row.get("Material"),
                "Price": as_number_or_blank(row.get("Price") if "Price" in row else row.get("Cost")),
                "Detail from Disclosure": row.get("Detail from Disclosure"),
                "Color Overrides": row.get("Color Overrides"),
                "Trim": trim,
                "Seat": row.get("Seat"),
                "Interior Code": row.get("Interior Code"),
                "Suede": row.get("Suede"),
                "Stitch": row.get("Stitch"),
                "Two Tone": row.get("Two Tone"),
                "section_id": row.get("section_id") or None,
                "active_for_stingray": False if row.get("active_for_stingray") in (None, "") else as_bool(row.get("active_for_stingray")),
                "requires_r6x": "_R6X" in str(trim) or str(interior_id).endswith("_R6X"),
                "included_option_id": row.get("included_option_id") or None,
            }
        )

    for col_idx, header in enumerate(CANONICAL_INTERIOR_HEADERS, start=1):
        ws.cell(1, col_idx).value = header
    for col_idx in range(len(CANONICAL_INTERIOR_HEADERS) + 1, ws.max_column + 1):
        ws.cell(1, col_idx).value = None

    for row_idx, row in enumerate(normalized_rows, start=2):
        for col_idx, header in enumerate(CANONICAL_INTERIOR_HEADERS, start=1):
            ws.cell(row_idx, col_idx).value = row.get(header)
        for col_idx in range(len(CANONICAL_INTERIOR_HEADERS) + 1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None
    for row_idx in range(len(normalized_rows) + 2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None

    for col_idx in range(1, len(CANONICAL_INTERIOR_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(ws.column_dimensions[get_column_letter(col_idx)].width or 12, 12)

    refresh_table_refs(ws)
    return 1 if list(old_mapping) != CANONICAL_INTERIOR_HEADERS else 0


def active(value: Any) -> bool:
    converted = as_bool(value)
    return True if converted is None else converted


def rule_group_lookup(wb, group_sheet: str, member_sheet: str) -> dict[tuple[str, str], str]:
    if group_sheet not in wb.sheetnames or member_sheet not in wb.sheetnames:
        return {}
    groups = {row["group_id"]: row for _, row in iter_records(wb[group_sheet]) if row.get("group_id") and active(row.get("active"))}
    lookup: dict[tuple[str, str], str] = {}
    for _, member in iter_records(wb[member_sheet]):
        if not active(member.get("active")):
            continue
        group_id = member.get("group_id")
        target_id = member.get("target_id")
        group = groups.get(group_id)
        if group and target_id:
            lookup[(str(group.get("source_id") or ""), str(target_id))] = str(group_id)
    return lookup


def rule_id_lookup(wb, sheet: str) -> dict[tuple[str, str, str], str]:
    lookup: dict[tuple[str, str, str], str] = {}
    if sheet not in wb.sheetnames:
        return lookup
    for _, row in iter_records(wb[sheet]):
        action = str(row.get("generation_action") or "").strip()
        if action.startswith("omit"):
            continue
        key = (str(row.get("source_id") or ""), str(row.get("rule_type") or "").lower(), str(row.get("target_id") or ""))
        if all(key):
            lookup[key] = str(row.get("rule_id") or "")
    return lookup


def replacement_fields(action: str, row: dict[str, Any], *, group_lookup: dict[tuple[str, str], str], active_rules: dict[tuple[str, str, str], str]) -> tuple[str | None, str | None]:
    source_id = str(row.get("source_id") or "")
    rule_type = str(row.get("rule_type") or "").lower()
    target_id = str(row.get("target_id") or "")
    if action == "omit_grouped_requirement":
        return group_lookup.get((source_id, target_id)) or ("grp_5v7_spoiler_requirement" if source_id == "opt_5v7_001" else "grp_5zu_paint_requirement"), None
    if action == "omit_grouped_exclusion":
        return "gs_group_z15_excludes_non_center_stripes", None
    if action == "omit_replaced_by_brake_exclusive_group":
        return "gs_excl_performance_brakes", None
    if action == "omit_replaced_by_d3v_include":
        return None, active_rules.get((source_id, "includes", "opt_d3v_001")) or "gs_rule_direct_d3v_include"
    if action == "omit_redundant_scoped_duplicate":
        return None, active_rules.get((source_id, rule_type, target_id))
    if action == "omit_soft_defaulted_caliper":
        return None, "gs_default_j6d_with_j57"
    return None, None


def add_lifecycle_columns(wb) -> int:
    changed = 0
    sheet_config = {
        "rule_mapping": ("rule_groups", "rule_group_members"),
        "grandSport_rule_mapping": ("grandSport_rule_groups", "grandSport_rule_group_members"),
    }
    for sheet, (group_sheet, member_sheet) in sheet_config.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        mapping = header_map(ws)
        for column in LIFECYCLE_COLUMNS:
            if column not in mapping:
                target_col = next((col for col in range(1, ws.max_column + 2) if ws.cell(1, col).value in (None, "")), ws.max_column + 1)
                ws.cell(1, target_col).value = column
                changed += 1
                mapping = header_map(ws)
        group_lookup = rule_group_lookup(wb, group_sheet, member_sheet)
        active_rules = rule_id_lookup(wb, sheet)
        for row_number, row in iter_records(ws):
            action = str(row.get("generation_action") or "").strip()
            status = ACTION_STATUS.get(action, "active")
            reason = "" if status == "active" else str(row.get("disabled_reason") or "").strip() or ACTION_REASONS[action]
            group_id, rule_id = replacement_fields(action, row, group_lookup=group_lookup, active_rules=active_rules)
            values = {
                "normalization_status": status,
                "normalization_reason": reason or None,
                "replacement_group_id": group_id,
                "replacement_rule_id": rule_id,
            }
            for column, value in values.items():
                cell = ws.cell(row_number, mapping[column])
                if cell.value != value:
                    cell.value = value
                    changed += 1
        refresh_table_refs(ws)
    return changed


def populated_ref(ws) -> str:
    max_col = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value not in (None, ""):
            max_col = col
    max_row = 1
    for row in range(2, ws.max_row + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, max_col + 1)):
            max_row = row
    return f"A1:{get_column_letter(max_col)}{max_row}"


def refresh_table_refs(ws) -> None:
    ref = populated_ref(ws)
    if ws.tables:
        for table in ws.tables.values():
            table.ref = ref
            table.tableColumns = []
            table._initialise_columns()
            for index, column in enumerate(table.tableColumns, start=1):
                column.id = index
                column.name = str(ws.cell(1, index).value or f"Column{index}")
    ws.auto_filter.ref = ref


def refresh_all_table_refs(wb) -> None:
    for ws in wb.worksheets:
        refresh_table_refs(ws)


def apply_schema_standardization(workbook: Path) -> dict[str, Any]:
    loaded_mtime_ns = workbook.stat().st_mtime_ns
    wb = load_workbook(workbook)
    try:
        changes = {
            "lz_interiors_header_normalizations": normalize_lz_interiors(wb),
            "boolean_cells_normalized": normalize_boolean_columns(wb),
            "rpo_cells_normalized": normalize_rpo_columns(wb),
            "price_cells_normalized": normalize_price_columns(wb),
            "lifecycle_cells_populated": add_lifecycle_columns(wb),
        }
        refresh_all_table_refs(wb)
        backup_path = save_workbook_safely(wb, workbook, loaded_mtime_ns=loaded_mtime_ns)
        return {"workbook": str(workbook), "backup_path": str(backup_path), "changes": changes}
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", nargs="?", default="stingray_master.xlsx")
    args = parser.parse_args()
    workbook = Path(args.workbook)
    result = apply_schema_standardization(workbook)
    for key, value in result.items():
        print(f"{key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
