#!/usr/bin/env python3
"""Backfill workbook-owned model configuration metadata for Phase 7."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import clean, excel_lock_path, rows_from_sheet, save_workbook_safely  # noqa: E402

WORKBOOK_PATH = ROOT / "stingray_master.xlsx"

MODEL_MASTER_HEADERS = [
    "model_key",
    "registry_key",
    "model_label",
    "model_year",
    "dataset_name",
    "export_slug",
    "expected_variant_count",
    "default_model",
    "active",
    "notes",
]
MODEL_WORKBOOK_SOURCES_HEADERS = ["model_key", "source_role", "sheet_name", "active", "notes"]
MODEL_VARIANTS_HEADERS = ["model_key", "variant_id", "display_order", "active", "notes"]

MODEL_MASTER_ROWS = [
    {
        "model_key": "stingray",
        "registry_key": "stingray",
        "model_label": "Stingray",
        "model_year": "2027",
        "dataset_name": "2027 Corvette Stingray operational form",
        "export_slug": "stingray",
        "expected_variant_count": 6,
        "default_model": True,
        "active": True,
        "notes": "Phase 7 parity row matching existing STINGRAY_MODEL.",
    },
    {
        "model_key": "grand_sport",
        "registry_key": "grandSport",
        "model_label": "Grand Sport",
        "model_year": "2027",
        "dataset_name": "2027 Corvette Grand Sport operational form",
        "export_slug": "grand-sport",
        "expected_variant_count": 6,
        "default_model": False,
        "active": True,
        "notes": "Phase 7 parity row matching existing GRAND_SPORT_MODEL.",
    },
]

SOURCE_ROWS = [
    ("stingray", "source_option_sheet", "stingray_options", "Matches current STINGRAY_MODEL.source_option_sheet."),
    ("stingray", "status_sheet", "stingray_ovs", "Matches current STINGRAY_MODEL.status_sheet."),
    ("stingray", "rule_mapping_sheet", "rule_mapping", "Matches current default rule mapping sheet."),
    ("stingray", "price_rules_sheet", "price_rules", "Matches current default price rules sheet."),
    ("stingray", "rule_groups_sheet", "rule_groups", "Matches current default rule groups sheet."),
    ("stingray", "rule_group_members_sheet", "rule_group_members", "Matches current default rule group members sheet."),
    ("stingray", "exclusive_groups_sheet", "exclusive_groups", "Matches current default exclusive groups sheet."),
    ("stingray", "exclusive_group_members_sheet", "exclusive_group_members", "Matches current default exclusive group members sheet."),
    ("stingray", "color_overrides_sheet", "color_overrides", "Matches current default color overrides sheet."),
    ("grand_sport", "source_option_sheet", "grandSport_options", "Matches current GRAND_SPORT_MODEL.source_option_sheet."),
    ("grand_sport", "status_sheet", "grandSport_ovs", "Matches current GRAND_SPORT_MODEL.status_sheet."),
    ("grand_sport", "rule_mapping_sheet", "grandSport_rule_mapping", "Matches current GRAND_SPORT_MODEL.rule_mapping_sheet."),
    ("grand_sport", "price_rules_sheet", "grandSport_price_rules", "Matches current GRAND_SPORT_MODEL.price_rules_sheet."),
    ("grand_sport", "rule_groups_sheet", "grandSport_rule_groups", "Matches current GRAND_SPORT_MODEL.rule_groups_sheet."),
    ("grand_sport", "rule_group_members_sheet", "grandSport_rule_group_members", "Matches current GRAND_SPORT_MODEL.rule_group_members_sheet."),
    ("grand_sport", "exclusive_groups_sheet", "grandSport_exclusive_groups", "Matches current GRAND_SPORT_MODEL.exclusive_groups_sheet."),
    ("grand_sport", "exclusive_group_members_sheet", "grandSport_exclusive_members", "Matches current GRAND_SPORT_MODEL.exclusive_group_members_sheet."),
    ("grand_sport", "color_overrides_sheet", "color_overrides", "Matches current shared color overrides sheet."),
    ("grand_sport", "variant_option_overrides_sheet", "grandSport_variant_overrides", "Matches current GRAND_SPORT_MODEL.variant_option_overrides_sheet."),
]
MODEL_WORKBOOK_SOURCES_ROWS = [
    {"model_key": model_key, "source_role": role, "sheet_name": sheet_name, "active": True, "notes": notes}
    for model_key, role, sheet_name, notes in SOURCE_ROWS
]

VARIANT_ROWS = [
    ("stingray", "1lt_c07", 1),
    ("stingray", "2lt_c07", 2),
    ("stingray", "3lt_c07", 3),
    ("stingray", "1lt_c67", 4),
    ("stingray", "2lt_c67", 5),
    ("stingray", "3lt_c67", 6),
    ("grand_sport", "1lt_e07", 1),
    ("grand_sport", "2lt_e07", 2),
    ("grand_sport", "3lt_e07", 3),
    ("grand_sport", "1lt_e67", 4),
    ("grand_sport", "2lt_e67", 5),
    ("grand_sport", "3lt_e67", 6),
]
MODEL_VARIANTS_ROWS = [
    {
        "model_key": model_key,
        "variant_id": variant_id,
        "display_order": display_order,
        "active": True,
        "notes": f"Matches current {'STINGRAY_MODEL' if model_key == 'stingray' else 'GRAND_SPORT_MODEL'}.variant_ids.",
    }
    for model_key, variant_id, display_order in VARIANT_ROWS
]


def ensure_sheet(wb: Any, sheet_name: str, headers: list[str]) -> Any:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        return ws
    ws = wb[sheet_name]
    existing_headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    for header in headers:
        if header not in existing_headers:
            ws.cell(1, len(existing_headers) + 1, header)
            existing_headers.append(header)
    return ws


def header_indexes(ws: Any) -> dict[str, int]:
    return {clean(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1) if clean(ws.cell(1, col).value)}


def row_key(ws: Any, row_number: int, key_fields: tuple[str, ...], headers: dict[str, int]) -> tuple[str, ...]:
    return tuple(clean(ws.cell(row_number, headers[field]).value) for field in key_fields)


def upsert_rows(wb: Any, sheet_name: str, headers: list[str], key_fields: tuple[str, ...], desired_rows: list[dict[str, Any]]) -> int:
    ws = ensure_sheet(wb, sheet_name, headers)
    indexes = header_indexes(ws)
    existing: dict[tuple[str, ...], int] = {}
    for row_number in range(2, ws.max_row + 1):
        key = row_key(ws, row_number, key_fields, indexes)
        if any(key):
            existing[key] = row_number

    changed = 0
    for desired in desired_rows:
        key = tuple(clean(desired[field]) for field in key_fields)
        row_number = existing.get(key)
        if row_number is None:
            row_number = ws.max_row + 1
            existing[key] = row_number
        for header, value in desired.items():
            column = indexes[header]
            if clean(ws.cell(row_number, column).value) != clean(value):
                ws.cell(row_number, column, value)
                changed += 1
    return changed


def verify_saved_rows(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        model_rows = [row for row in rows_from_sheet(wb, "model_master") if row.get("active") == "True"]
        source_rows = [row for row in rows_from_sheet(wb, "model_workbook_sources") if row.get("active") == "True"]
        variant_rows = [row for row in rows_from_sheet(wb, "model_variants") if row.get("active") == "True"]
        required_roles = {
            "stingray": {row["source_role"] for row in MODEL_WORKBOOK_SOURCES_ROWS if row["model_key"] == "stingray"},
            "grand_sport": {row["source_role"] for row in MODEL_WORKBOOK_SOURCES_ROWS if row["model_key"] == "grand_sport"},
        }
        source_roles = {
            model_key: {row["source_role"] for row in source_rows if row.get("model_key") == model_key}
            for model_key in required_roles
        }
        variant_ids = {
            model_key: [row["variant_id"] for row in variant_rows if row.get("model_key") == model_key]
            for model_key in required_roles
        }
        for model_key in ("stingray", "grand_sport"):
            if sum(1 for row in model_rows if row.get("model_key") == model_key) != 1:
                raise RuntimeError(f"Expected exactly one active model_master row for {model_key}")
            missing_roles = sorted(required_roles[model_key] - source_roles[model_key])
            if missing_roles:
                raise RuntimeError(f"Missing active model_workbook_sources roles for {model_key}: {missing_roles}")
            if len(variant_ids[model_key]) != 6:
                raise RuntimeError(f"Expected 6 active model_variants rows for {model_key}; found {len(variant_ids[model_key])}")
        return {
            "active_model_master_rows": {model_key: sum(1 for row in model_rows if row.get("model_key") == model_key) for model_key in required_roles},
            "active_source_role_counts": {model_key: len(source_roles[model_key]) for model_key in required_roles},
            "active_variant_ids": variant_ids,
        }
    finally:
        wb.close()


def main() -> None:
    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {lock_path}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    try:
        changed = 0
        changed += upsert_rows(wb, "model_master", MODEL_MASTER_HEADERS, ("model_key",), MODEL_MASTER_ROWS)
        changed += upsert_rows(
            wb,
            "model_workbook_sources",
            MODEL_WORKBOOK_SOURCES_HEADERS,
            ("model_key", "source_role"),
            MODEL_WORKBOOK_SOURCES_ROWS,
        )
        changed += upsert_rows(wb, "model_variants", MODEL_VARIANTS_HEADERS, ("model_key", "variant_id"), MODEL_VARIANTS_ROWS)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_rows(WORKBOOK_PATH)
    print(json.dumps({"workbook": str(WORKBOOK_PATH), "backup": str(backup_path), "cell_updates": changed, "verification": verification}, indent=2))


if __name__ == "__main__":
    main()
