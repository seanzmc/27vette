#!/usr/bin/env python3
"""Backfill workbook-owned Grand Sport audit parser metadata.

Idempotent: appends only missing metadata rows and preserves existing workbook
rows.  The audit/source-builder reads these rows before falling back to legacy
Python constants.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[1]
ROOT = SCRIPTS_DIR.parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import (  # noqa: E402
    excel_lock_path,
    rows_from_sheet,
    save_workbook_safely,
    write_sheet,
)

WORKBOOK_PATH = ROOT / "stingray_master.xlsx"

SHEET_HEADERS: dict[str, list[str]] = {
    "rule_phrase_map": [
        "phrase",
        "rule_type",
        "direction",
        "stop_phrases",
        "review_flag_default",
        "active",
        "notes",
    ],
    "option_audit_groups": ["group_id", "group_label", "active", "notes"],
    "option_audit_group_members": ["group_id", "rpo", "option_id", "active", "notes"],
    "rule_review_groups": ["model_key", "group_id", "rpo", "review_reason", "active", "notes"],
}

BACKFILL_ROWS: dict[str, list[dict[str, Any]]] = {
    "rule_phrase_map": [
        {
            "phrase": "not available with",
            "rule_type": "excludes",
            "direction": "source_to_mentioned",
            "stop_phrases": "",
            "review_flag_default": "TRUE",
            "active": "TRUE",
            "notes": "Mirrors current not-available parser branch.",
        },
        {
            "phrase": "requires",
            "rule_type": "requires",
            "direction": "source_to_mentioned",
            "stop_phrases": " or included with| included with",
            "review_flag_default": "TRUE",
            "active": "TRUE",
            "notes": "Mirrors current requires parser branch and stop phrase behavior.",
        },
        {
            "phrase": "includes",
            "rule_type": "includes",
            "direction": "source_to_mentioned",
            "stop_phrases": " requires",
            "review_flag_default": "TRUE",
            "active": "TRUE",
            "notes": "Applies only when fragment does not also contain included with, preserving current behavior.",
        },
        {
            "phrase": "included and only available with",
            "rule_type": "includes",
            "direction": "mentioned_to_source",
            "stop_phrases": "",
            "review_flag_default": "TRUE",
            "active": "TRUE",
            "notes": "Mentioned option includes the current source option.",
        },
        {
            "phrase": "included with",
            "rule_type": "includes",
            "direction": "mentioned_to_source",
            "stop_phrases": "",
            "review_flag_default": "TRUE",
            "active": "TRUE",
            "notes": "Mentioned option includes the current source option.",
        },
        {
            "phrase": "only available with",
            "rule_type": "requires",
            "direction": "source_to_mentioned",
            "stop_phrases": "",
            "review_flag_default": "TRUE",
            "active": "TRUE",
            "notes": "Applies only when fragment does not contain included, preserving current behavior.",
        },
    ],
    "option_audit_groups": [
        {
            "group_id": "engine_cover",
            "group_label": "Engine cover options",
            "active": "TRUE",
            "notes": "Grand Sport rule audit focus group formerly hardcoded as ENGINE_COVER_RPOS.",
        },
    ],
    "option_audit_group_members": [
        {
            "group_id": "engine_cover",
            "rpo": rpo,
            "option_id": "",
            "active": "TRUE",
            "notes": "Former ENGINE_COVER_RPOS member.",
        }
        for rpo in ("BC7", "BCP", "BCS", "BC4", "B6P", "ZZ3", "D3V", "SL9")
    ],
    "rule_review_groups": [
        {
            "model_key": "grand_sport",
            "group_id": "special_package_review",
            "rpo": rpo,
            "review_reason": "Verify package/rule mentions for Grand Sport audit.",
            "active": "TRUE",
            "notes": "Former GRAND_SPORT_MODEL.special_rule_review_rpos member.",
        }
        for rpo in ("EL9", "Z25", "FEY", "Z15")
    ],
}

KEY_FIELDS: dict[str, tuple[str, ...]] = {
    "rule_phrase_map": ("phrase",),
    "option_audit_groups": ("group_id",),
    "option_audit_group_members": ("group_id", "rpo", "option_id"),
    "rule_review_groups": ("model_key", "group_id", "rpo"),
}


def row_key(row: dict[str, Any], fields: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(str(row.get(field, "")).strip().lower() for field in fields)


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to run; Excel lock file is present: {lock_path}. Close Excel first.")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    inserted: dict[str, int] = {}
    skipped: dict[str, int] = {}

    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name in wb.sheetnames:
            rows = rows_from_sheet(wb, sheet_name)
        else:
            rows = []
        fields = KEY_FIELDS[sheet_name]
        existing = {row_key(row, fields) for row in rows}
        for backfill_row in BACKFILL_ROWS[sheet_name]:
            key = row_key(backfill_row, fields)
            if key in existing:
                skipped[sheet_name] = skipped.get(sheet_name, 0) + 1
                continue
            rows.append(backfill_row)
            existing.add(key)
            inserted[sheet_name] = inserted.get(sheet_name, 0) + 1
        write_sheet(wb, sheet_name, headers, rows)

    backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    wb.close()

    print(f"Workbook updated: {WORKBOOK_PATH}")
    print(f"Backup created: {backup_path}")
    for sheet_name in SHEET_HEADERS:
        print(f"{sheet_name}: inserted={inserted.get(sheet_name, 0)} skipped={skipped.get(sheet_name, 0)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
