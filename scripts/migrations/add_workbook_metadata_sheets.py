#!/usr/bin/env python3
"""Add workbook-owned runtime/business metadata sheets to stingray_master.xlsx.

This migration is intentionally data-empty: it creates or normalizes headers for
metadata source sheets while preserving any existing rows on those sheets.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[1]
ROOT = SCRIPTS_DIR.parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import (  # noqa: E402
    excel_lock_path,
    rows_from_sheet,
    save_workbook_safely,
    write_sheet,
)

WORKBOOK_PATH = ROOT / "stingray_master.xlsx"

SHEET_HEADERS: dict[str, list[str]] = {
    "runtime_steps": [
        "model_key",
        "step_key",
        "step_label",
        "runtime_order",
        "source",
        "active",
        "notes",
    ],
    "context_section_master": [
        "model_key",
        "context_type",
        "section_id",
        "section_name",
        "selection_mode",
        "choice_mode",
        "is_required",
        "standard_behavior",
        "section_display_order",
        "step_key",
        "step_label",
        "active",
        "notes",
    ],
    "section_presentation": [
        "model_key",
        "section_id",
        "display_label",
        "step_key",
        "presentation_bucket",
        "display_behavior",
        "section_display_order",
        "standard_equipment_bucket",
        "standard_equipment_group_type",
        "active",
        "notes",
    ],
    "model_master": [
        "model_key",
        "registry_key",
        "model_label",
        "model_year",
        "dataset_name",
        "export_slug",
        "expected_variant_count",
        "default_model",
        "active",
        "notes",
    ],
    "model_workbook_sources": [
        "model_key",
        "source_role",
        "sheet_name",
        "active",
        "notes",
    ],
    "model_variants": [
        "model_key",
        "variant_id",
        "display_order",
        "active",
        "notes",
    ],
    "variant_option_overrides": [
        "model_key",
        "option_id",
        "variant_id",
        "status",
        "selectable",
        "active",
        "display_behavior",
        "notes",
    ],
    "default_selection_rules": [
        "model_key",
        "rule_id",
        "target_option_id",
        "condition_type",
        "condition_id",
        "body_style_scope",
        "trim_level_scope",
        "variant_scope",
        "priority",
        "active",
        "notes",
    ],
    "runtime_rule_exceptions": [
        "model_key",
        "exception_id",
        "source_option_id",
        "target_option_id",
        "exception_type",
        "body_style_scope",
        "trim_level_scope",
        "variant_scope",
        "disabled_reason",
        "active",
        "notes",
    ],
    "interior_components": [
        "model_key",
        "interior_id",
        "rpo",
        "component_type",
        "label",
        "price_ref_type",
        "price_ref_code",
        "price_trim_scope",
        "display_order",
        "active",
        "notes",
    ],
    "component_price_rules": [
        "model_key",
        "price_rule_id",
        "condition_option_id",
        "target_component_rpo",
        "price_rule_type",
        "price_value",
        "body_style_scope",
        "trim_level_scope",
        "variant_scope",
        "active",
        "notes",
    ],
    "standard_equipment_groups": [
        "model_key",
        "section_id",
        "group_type",
        "default_open",
        "canonical_rank",
        "duplicate_group_key",
        "active",
        "notes",
    ],
    "order_summary_sections": [
        "model_key",
        "section_key",
        "section_label",
        "display_order",
        "active",
        "notes",
    ],
    "step_order_summary_map": [
        "model_key",
        "step_key",
        "section_key",
        "active",
        "notes",
    ],
    "rule_phrase_map": [
        "phrase",
        "rule_type",
        "direction",
        "stop_phrases",
        "review_flag_default",
        "active",
        "notes",
    ],
    "option_audit_groups": [
        "group_id",
        "group_label",
        "active",
        "notes",
    ],
    "option_audit_group_members": [
        "group_id",
        "rpo",
        "option_id",
        "active",
        "notes",
    ],
    "rule_review_groups": [
        "model_key",
        "group_id",
        "rpo",
        "review_reason",
        "active",
        "notes",
    ],
    "model_interior_scope": [
        "model_key",
        "interior_id",
        "trim_level",
        "active",
        "requires_option_id",
        "notes",
    ],
}


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(
            f"Refusing to run; Excel lock file is present: {lock_path}. Close Excel first."
        )

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    created: list[str] = []
    normalized: list[str] = []

    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name in wb.sheetnames:
            rows = rows_from_sheet(wb, sheet_name)
            normalized.append(f"{sheet_name} ({len(rows)} existing rows preserved)")
        else:
            rows = []
            created.append(sheet_name)
        write_sheet(wb, sheet_name, headers, rows)

    backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    wb.close()

    print(f"Workbook updated: {WORKBOOK_PATH}")
    print(f"Backup created: {backup_path}")
    if created:
        print("Sheets created:")
        for sheet_name in created:
            print(f"- {sheet_name}")
    else:
        print("Sheets created: none")
    if normalized:
        print("Sheets normalized:")
        for item in normalized:
            print(f"- {item}")
    else:
        print("Sheets normalized: none")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
