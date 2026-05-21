#!/usr/bin/env python3
"""Backfill Grand Sport NGA default-selected metadata.

NGA is standard on all Grand Sport variants and should seed the runtime build
like other workbook-authored default_selected standard options. This migration
adds variant-scoped display_behavior overrides without changing exhaust rules or
runtime JavaScript.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[1]
ROOT = SCRIPTS_DIR.parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook import (  # noqa: E402
    excel_lock_path,
    rows_from_sheet,
    save_workbook_safely,
    write_sheet,
)

WORKBOOK_PATH = ROOT / "stingray_master.xlsx"
SHEET_NAME = "grandSport_variant_overrides"
HEADERS = ["option_id", "variant_id", "selectable", "display_behavior", "section_id", "active", "note"]
VARIANT_IDS = ("1lt_e07", "2lt_e07", "3lt_e07", "1lt_e67", "2lt_e67", "3lt_e67")
BACKFILL_ROWS: list[dict[str, Any]] = [
    {
        "option_id": "opt_nga_001",
        "variant_id": variant_id,
        "selectable": "",
        "display_behavior": "default_selected",
        "section_id": "",
        "active": "True",
        "note": "Default black exhaust tips; NGA is standard on Grand Sport.",
    }
    for variant_id in VARIANT_IDS
]


def row_key(row: dict[str, Any]) -> tuple[str, str]:
    return (str(row.get("option_id", "")).strip().lower(), str(row.get("variant_id", "")).strip().lower())


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to run; Excel lock file is present: {lock_path}. Close Excel first.")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    rows = rows_from_sheet(wb, SHEET_NAME) if SHEET_NAME in wb.sheetnames else []
    existing = {row_key(row): row for row in rows}
    inserted = 0
    updated = 0
    skipped = 0

    for backfill_row in BACKFILL_ROWS:
        key = row_key(backfill_row)
        existing_row = existing.get(key)
        if existing_row is None:
            rows.append(backfill_row)
            existing[key] = backfill_row
            inserted += 1
            continue
        changed = False
        for field in ("display_behavior", "active", "note"):
            if existing_row.get(field, "") != backfill_row[field]:
                existing_row[field] = backfill_row[field]
                changed = True
        if changed:
            updated += 1
        else:
            skipped += 1

    write_sheet(wb, SHEET_NAME, HEADERS, rows)
    backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    wb.close()

    print(f"Workbook updated: {WORKBOOK_PATH}")
    print(f"Backup created: {backup_path}")
    print(f"{SHEET_NAME}: inserted={inserted} updated={updated} skipped={skipped}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
