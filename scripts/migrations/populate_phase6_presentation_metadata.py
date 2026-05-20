#!/usr/bin/env python3
"""Populate Phase 6 workbook-owned step, context, and presentation metadata."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path

from openpyxl import load_workbook

from corvette_form_generator.model_configs import (
    CONTEXT_SECTIONS,
    GRAND_SPORT_SECTION_LABEL_OVERRIDES,
    STANDARD_SECTIONS,
    STEP_LABELS,
    STEP_ORDER,
)
from corvette_form_generator.workbook import clean, rows_from_sheet, save_workbook_safely, write_sheet

WORKBOOK_PATH = Path("stingray_master.xlsx")
MODEL_KEYS = ("stingray", "grand_sport")
TRIM_EQUIPMENT_SECTIONS = frozenset({"sec_1lte_001", "sec_2lte_001", "sec_3lte_001"})
STINGRAY_SECTION_DISPLAY_ORDER_OVERRIDES = {
    "sec_stri_001": 30,
    "sec_gsha_001": 50,
    "sec_gsce_001": 51,
}

SHEET_HEADERS: dict[str, list[str]] = {
    "runtime_steps": ["model_key", "step_key", "step_label", "runtime_order", "source", "active", "notes"],
    "context_section_master": [
        "model_key",
        "context_type",
        "section_id",
        "section_name",
        "selection_mode",
        "choice_mode",
        "is_required",
        "standard_behavior",
        "section_display_order",
        "step_key",
        "step_label",
        "active",
        "notes",
    ],
    "section_presentation": [
        "model_key",
        "section_id",
        "display_label",
        "step_key",
        "presentation_bucket",
        "display_behavior",
        "section_display_order",
        "standard_equipment_bucket",
        "standard_equipment_group_type",
        "active",
        "notes",
    ],
    "standard_equipment_groups": [
        "model_key",
        "section_id",
        "group_type",
        "default_open",
        "canonical_rank",
        "duplicate_group_key",
        "active",
        "notes",
    ],
}


def replace_keyed_rows(
    rows: Sequence[Mapping[str, object]],
    new_rows: Sequence[Mapping[str, object]],
    *,
    key_fields: tuple[str, ...],
) -> list[dict[str, object]]:
    new_keys = {tuple(clean(row.get(field)) for field in key_fields) for row in new_rows}
    kept = [
        dict(row)
        for row in rows
        if tuple(clean(row.get(field)) for field in key_fields) not in new_keys
    ]
    return kept + [dict(row) for row in new_rows]


def runtime_step_rows() -> list[dict[str, object]]:
    return [
        {
            "model_key": model_key,
            "step_key": step_key,
            "step_label": STEP_LABELS[step_key],
            "runtime_order": index,
            "source": "workbook_phase6",
            "active": "True",
            "notes": "Backfilled from model_configs.STEP_ORDER/STEP_LABELS for Phase 6 parity.",
        }
        for model_key in MODEL_KEYS
        for index, step_key in enumerate(STEP_ORDER, start=1)
    ]


def context_type_for_section(section: Mapping[str, object]) -> str:
    step_key = clean(section.get("step_key"))
    if step_key in {"body_style", "trim_level"}:
        return step_key
    section_id = clean(section.get("section_id"))
    if section_id.endswith("body_style"):
        return "body_style"
    if section_id.endswith("trim_level"):
        return "trim_level"
    return step_key


def context_section_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for model_key in MODEL_KEYS:
        for section in CONTEXT_SECTIONS:
            rows.append(
                {
                    "model_key": model_key,
                    "context_type": context_type_for_section(section),
                    "section_id": clean(section.get("section_id")),
                    "section_name": clean(section.get("section_name")),
                    "selection_mode": clean(section.get("selection_mode")),
                    "choice_mode": clean(section.get("choice_mode")),
                    "is_required": clean(section.get("is_required")),
                    "standard_behavior": clean(section.get("standard_behavior")),
                    "section_display_order": section.get("section_display_order", ""),
                    "step_key": clean(section.get("step_key")),
                    "step_label": clean(section.get("step_label")),
                    "active": "True",
                    "notes": "Backfilled from model_configs.CONTEXT_SECTIONS for Phase 6 parity.",
                }
            )
    return rows


def section_presentation_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for model_key in MODEL_KEYS:
        for section_id in sorted(STANDARD_SECTIONS):
            group_type = "trim_equipment" if section_id in TRIM_EQUIPMENT_SECTIONS else ""
            rows.append(
                {
                    "model_key": model_key,
                    "section_id": section_id,
                    "display_label": "",
                    "step_key": "",
                    "presentation_bucket": "",
                    "display_behavior": "",
                    "section_display_order": "",
                    "standard_equipment_bucket": "True",
                    "standard_equipment_group_type": group_type,
                    "active": "True",
                    "notes": "Backfilled from model_configs.STANDARD_SECTIONS for Phase 6 parity.",
                }
            )

    for section_id, display_order in STINGRAY_SECTION_DISPLAY_ORDER_OVERRIDES.items():
        rows.append(
            {
                "model_key": "stingray",
                "section_id": section_id,
                "display_label": "",
                "step_key": "",
                "presentation_bucket": "",
                "display_behavior": "",
                "section_display_order": display_order,
                "standard_equipment_bucket": "",
                "standard_equipment_group_type": "",
                "active": "True",
                "notes": "Backfilled from STINGRAY_SECTION_DISPLAY_ORDER_OVERRIDES for Phase 6 parity.",
            }
        )

    for section_id, display_label in GRAND_SPORT_SECTION_LABEL_OVERRIDES.items():
        rows.append(
            {
                "model_key": "grand_sport",
                "section_id": section_id,
                "display_label": display_label,
                "step_key": "",
                "presentation_bucket": "",
                "display_behavior": "",
                "section_display_order": "",
                "standard_equipment_bucket": "",
                "standard_equipment_group_type": "",
                "active": "True",
                "notes": "Backfilled from GRAND_SPORT_SECTION_LABEL_OVERRIDES for Phase 6 parity.",
            }
        )
    return rows


def standard_equipment_group_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for model_key in MODEL_KEYS:
        for index, section_id in enumerate(sorted(STANDARD_SECTIONS), start=1):
            group_type = "trim_equipment" if section_id in TRIM_EQUIPMENT_SECTIONS else ""
            rows.append(
                {
                    "model_key": model_key,
                    "section_id": section_id,
                    "group_type": group_type,
                    "default_open": "True" if group_type == "trim_equipment" else "",
                    "canonical_rank": index,
                    "duplicate_group_key": "",
                    "active": "True",
                    "notes": "Backfilled from model_configs.STANDARD_SECTIONS for Phase 6 parity.",
                }
            )
    return rows


def main() -> None:
    if WORKBOOK_PATH.with_name(f"~${WORKBOOK_PATH.name}").exists():
        raise SystemExit("Excel lock file exists; close stingray_master.xlsx first.")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)

    updates: dict[str, tuple[list[dict[str, object]], tuple[str, ...]]] = {
        "runtime_steps": (runtime_step_rows(), ("model_key", "step_key")),
        "context_section_master": (context_section_rows(), ("model_key", "section_id")),
        "section_presentation": (section_presentation_rows(), ("model_key", "section_id")),
        "standard_equipment_groups": (standard_equipment_group_rows(), ("model_key", "section_id")),
    }

    for sheet_name, (new_rows, key_fields) in updates.items():
        existing = rows_from_sheet(wb, sheet_name) if sheet_name in wb.sheetnames else []
        rows = replace_keyed_rows(existing, new_rows, key_fields=key_fields)
        write_sheet(wb, sheet_name, SHEET_HEADERS[sheet_name], rows)
        print(f"{sheet_name}: upserted={len(new_rows)} total={len(rows)}")

    backup = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    print(f"Phase 6 presentation metadata populated; backup={backup}")


if __name__ == "__main__":
    main()
