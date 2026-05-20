#!/usr/bin/env python3
"""Populate Phase 3 workbook-owned runtime metadata for Stingray."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path

from openpyxl import load_workbook

from corvette_form_generator.workbook import rows_from_sheet, save_workbook_safely, write_sheet

WORKBOOK_PATH = Path("stingray_master.xlsx")

SHEET_HEADERS: dict[str, list[str]] = {
    "default_selection_rules": [
        "model_key",
        "rule_id",
        "target_option_id",
        "condition_type",
        "condition_id",
        "body_style_scope",
        "trim_level_scope",
        "variant_scope",
        "priority",
        "active",
        "notes",
    ],
    "runtime_rule_exceptions": [
        "model_key",
        "exception_id",
        "source_option_id",
        "target_option_id",
        "exception_type",
        "body_style_scope",
        "trim_level_scope",
        "variant_scope",
        "disabled_reason",
        "active",
        "notes",
    ],
    "order_summary_sections": ["model_key", "section_key", "section_label", "display_order", "active", "notes"],
    "step_order_summary_map": ["model_key", "step_key", "section_key", "active", "notes"],
}

DEFAULT_SELECTION_RULES = [
    {
        "model_key": "stingray",
        "rule_id": "default_fe1",
        "target_option_id": "opt_fe1_001",
        "condition_type": "unless_selected_rpo",
        "condition_id": "Z51",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "priority": 10,
        "active": "TRUE",
        "notes": "Default standard suspension unless Z51 is selected.",
    },
    {
        "model_key": "stingray",
        "rule_id": "default_nga",
        "target_option_id": "opt_nga_001",
        "condition_type": "unless_selected_rpo",
        "condition_id": "NWI",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "priority": 20,
        "active": "TRUE",
        "notes": "Default black exhaust tips unless NWI center exhaust is selected.",
    },
    {
        "model_key": "stingray",
        "rule_id": "default_719",
        "target_option_id": "opt_719_001",
        "condition_type": "unless_selected_section",
        "condition_id": "sec_seat_001",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "priority": 30,
        "active": "TRUE",
        "notes": "Default black seat belt when no seat-belt option is selected or auto-added.",
    },
    {
        "model_key": "stingray",
        "rule_id": "default_bc7",
        "target_option_id": "opt_bc7_001",
        "condition_type": "always",
        "condition_id": "",
        "body_style_scope": "coupe",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "priority": 40,
        "active": "TRUE",
        "notes": "Default coupe black LS6 engine cover.",
    },
]

RUNTIME_RULE_EXCEPTIONS = [
    {
        "model_key": "stingray",
        "exception_id": "ex_z51_fe1",
        "source_option_id": "opt_z51_001",
        "target_option_id": "opt_fe1_001",
        "exception_type": "remove_target_when_source_selected",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "disabled_reason": "Replaced by FE3 Z51 performance suspension.",
        "active": "TRUE",
        "notes": "Z51 replaces FE1 standard suspension.",
    },
    {
        "model_key": "stingray",
        "exception_id": "ex_z51_fe2",
        "source_option_id": "opt_z51_001",
        "target_option_id": "opt_fe2_001",
        "exception_type": "remove_target_when_source_selected",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "disabled_reason": "Not available with Z51 Performance Package.",
        "active": "TRUE",
        "notes": "Z51 suppresses FE2.",
    },
    {
        "model_key": "stingray",
        "exception_id": "ex_nwi_nga",
        "source_option_id": "opt_nwi_001",
        "target_option_id": "opt_nga_001",
        "exception_type": "remove_target_when_source_selected",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "disabled_reason": "Replaced by NWI center exhaust.",
        "active": "TRUE",
        "notes": "NWI replaces NGA exhaust tips.",
    },
    {
        "model_key": "stingray",
        "exception_id": "ex_gba_zyc",
        "source_option_id": "opt_gba_001",
        "target_option_id": "opt_zyc_001",
        "exception_type": "remove_target_when_source_selected",
        "body_style_scope": "*",
        "trim_level_scope": "*",
        "variant_scope": "*",
        "disabled_reason": "Black exterior paint is not available with body-color accents.",
        "active": "TRUE",
        "notes": "Black paint removes body-color accents.",
    },
]

ORDER_SUMMARY_SECTIONS = [
    ("vehicle", "Vehicle"),
    ("exterior_paint", "Exterior Paint"),
    ("exterior_appearance", "Exterior Appearance"),
    ("wheels_brakes", "Wheels & Brakes"),
    ("performance_mechanical", "Performance & Mechanical"),
    ("stripes", "Stripes"),
    ("seats_interior", "Seats & Interior"),
    ("accessories", "Accessories"),
    ("delivery", "Delivery"),
    ("auto_added_required", "Auto-Added / Required"),
    ("pricing_summary", "Pricing Summary"),
]

STEP_ORDER_SUMMARY_MAP = [
    ("body_style", "vehicle"),
    ("trim_level", "vehicle"),
    ("paint", "exterior_paint"),
    ("exterior_appearance", "exterior_appearance"),
    ("wheels", "wheels_brakes"),
    ("packages_performance", "performance_mechanical"),
    ("aero_exhaust_stripes_accessories", "stripes"),
    ("seat", "seats_interior"),
    ("base_interior", "seats_interior"),
    ("seat_belt", "seats_interior"),
    ("interior_trim", "seats_interior"),
    ("accessories", "accessories"),
    ("delivery", "delivery"),
]


def replace_model_rows(rows: Sequence[Mapping[str, object]], new_rows: Sequence[Mapping[str, object]], *, key_field: str) -> list[dict[str, object]]:
    new_keys = {(row.get("model_key"), row.get(key_field)) for row in new_rows}
    kept = [dict(row) for row in rows if (row.get("model_key"), row.get(key_field)) not in new_keys]
    return kept + [dict(row) for row in new_rows]


def main() -> None:
    if WORKBOOK_PATH.with_name(f"~${WORKBOOK_PATH.name}").exists():
        raise SystemExit("Excel lock file exists; close stingray_master.xlsx first.")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)

    updates = {
        "default_selection_rules": (DEFAULT_SELECTION_RULES, "rule_id"),
        "runtime_rule_exceptions": (RUNTIME_RULE_EXCEPTIONS, "exception_id"),
        "order_summary_sections": (
            [
                {
                    "model_key": "stingray",
                    "section_key": section_key,
                    "section_label": section_label,
                    "display_order": index,
                    "active": "TRUE",
                    "notes": "Runtime order summary grouping.",
                }
                for index, (section_key, section_label) in enumerate(ORDER_SUMMARY_SECTIONS, start=1)
            ],
            "section_key",
        ),
        "step_order_summary_map": (
            [
                {
                    "model_key": "stingray",
                    "step_key": step_key,
                    "section_key": section_key,
                    "active": "TRUE",
                    "notes": "Runtime step-to-summary-section grouping.",
                }
                for step_key, section_key in STEP_ORDER_SUMMARY_MAP
            ],
            "step_key",
        ),
    }

    for sheet_name, (new_rows, key_field) in updates.items():
        existing = rows_from_sheet(wb, sheet_name) if sheet_name in wb.sheetnames else []
        rows = replace_model_rows(existing, new_rows, key_field=key_field)
        write_sheet(wb, sheet_name, SHEET_HEADERS[sheet_name], rows)

    backup = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    print(f"Phase 3 runtime metadata populated; backup={backup}")


if __name__ == "__main__":
    main()
