#!/usr/bin/env python3
"""Apply approved Z06 Pass 3 performance/package interaction corrections.

Dry-run is the default. Use --write to safely save stingray_master.xlsx.
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely  # noqa: E402

OPTION_SHEET = "z06_options"
RULE_MAPPING_SHEET = "z06_rule_mapping"
EXCLUSIVE_GROUPS_SHEET = "z06_exclusive_groups"
EXCLUSIVE_MEMBERS_SHEET = "z06_exclusive_members"

PACKAGE_GROUP_ID = "z06_excl_carbon_wheel_packages"
AERO_GROUP_ID = "z06_excl_aero_packages"
EXHAUST_GROUP_ID = "z06_excl_exhaust_tips"

PACKAGE_RPOS = ("PDB", "PDD", "PDF")
AERO_RPOS = ("T0E", "T0F", "T0G", "5ZV")
GROUND_RPOS = ("CFL", "CFZ", "CFV")
EXHAUST_RPOS = ("NGA", "NWI")

# These rows modeled radio-like peers as hard conflicts. Pass 3 restores
# workbook-owned exclusive groups so peers stay clickable and replace each other.
OMIT_EXCLUDE_PAIRS = {
    ("PDB", "PDD"),
    ("PDB", "PDF"),
    ("PDD", "PDB"),
    ("PDD", "PDF"),
    ("PDF", "PDB"),
    ("PDF", "PDD"),
    ("5ZV", "T0F"),
    ("PDD", "T0G"),
    ("PDF", "T0F"),
}


@dataclass(frozen=True)
class Change:
    sheet: str
    row_number: int | str
    key: str
    field: str
    current: Any
    desired: Any
    reason: str


def headers(ws) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def index(ws) -> dict[str, int]:
    return {header: offset + 1 for offset, header in enumerate(headers(ws)) if header}


def rows(ws):
    idx = index(ws)
    for row_number in range(2, ws.max_row + 1):
        row = {header: ws.cell(row_number, col).value for header, col in idx.items()}
        if any(value not in (None, "") for value in row.values()):
            yield row_number, row


def values_equal(current: Any, desired: Any) -> bool:
    return clean(current) == clean(desired)


def set_cell(ws, row_number: int, field: str, desired: Any, changes: list[Change], *, key: str, reason: str) -> None:
    idx = index(ws)
    if field not in idx:
        raise RuntimeError(f"{ws.title} missing required column {field!r}")
    current = ws.cell(row_number, idx[field]).value
    if not values_equal(current, desired):
        changes.append(Change(ws.title, row_number, key, field, current, desired, reason))
        ws.cell(row_number, idx[field]).value = desired


def option_maps(wb) -> tuple[dict[str, dict[str, Any]], dict[str, str], dict[str, str]]:
    by_rpo: dict[str, dict[str, Any]] = {}
    id_by_rpo: dict[str, str] = {}
    rpo_by_id: dict[str, str] = {}
    for row_number, row in rows(wb[OPTION_SHEET]):
        rpo = clean(row.get("rpo"))
        option_id = clean(row.get("option_id"))
        if rpo and option_id:
            by_rpo[rpo] = {**row, "_row_number": row_number}
            id_by_rpo[rpo] = option_id
            rpo_by_id[option_id] = rpo
    return by_rpo, id_by_rpo, rpo_by_id


def require_rpos(id_by_rpo: dict[str, str], rpos: tuple[str, ...] | list[str]) -> None:
    missing = [rpo for rpo in rpos if rpo not in id_by_rpo]
    if missing:
        raise RuntimeError(f"Missing required Z06 RPO(s): {', '.join(missing)}")


def ensure_group(wb, changes: list[Change], group_id: str, selection_mode: str, notes: str) -> None:
    ws = wb[EXCLUSIVE_GROUPS_SHEET]
    row_number = None
    for candidate_number, row in rows(ws):
        if clean(row.get("group_id")) == group_id:
            row_number = candidate_number
            break
    values = {
        "group_id": group_id,
        "selection_mode": selection_mode,
        "active": "True",
        "notes": notes,
    }
    if row_number is None:
        row_number = ws.max_row + 1
        idx = index(ws)
        for field, value in values.items():
            ws.cell(row_number, idx[field]).value = value
        changes.append(Change(ws.title, f"new:{row_number}", group_id, "row", None, values, "approved Z06 Pass 3 exclusive group"))
    else:
        for field, value in values.items():
            if field == "group_id":
                continue
            set_cell(ws, row_number, field, value, changes, key=group_id, reason="approved Z06 Pass 3 exclusive group")


def ensure_group_member(wb, changes: list[Change], group_id: str, option_id: str, display_order: int) -> None:
    ws = wb[EXCLUSIVE_MEMBERS_SHEET]
    row_number = None
    for candidate_number, row in rows(ws):
        if clean(row.get("group_id")) == group_id and clean(row.get("option_id")) == option_id:
            row_number = candidate_number
            break
    key = f"{group_id}:{option_id}"
    values = {
        "group_id": group_id,
        "option_id": option_id,
        "display_order": display_order,
        "active": "True",
    }
    if row_number is None:
        row_number = ws.max_row + 1
        idx = index(ws)
        for field, value in values.items():
            ws.cell(row_number, idx[field]).value = value
        changes.append(Change(ws.title, f"new:{row_number}", key, "row", None, values, "approved Z06 Pass 3 exclusive group member"))
    else:
        for field, value in values.items():
            if field in {"group_id", "option_id"}:
                continue
            set_cell(ws, row_number, field, value, changes, key=key, reason="approved Z06 Pass 3 exclusive group member")


def apply_exclusive_groups(wb, changes: list[Change]) -> None:
    _, id_by_rpo, _ = option_maps(wb)
    require_rpos(id_by_rpo, [*PACKAGE_RPOS, *AERO_RPOS, *GROUND_RPOS, *EXHAUST_RPOS])
    ensure_group(
        wb,
        changes,
        PACKAGE_GROUP_ID,
        "single_within_group",
        "PDB, PDD, and PDF are radio-like package peers; selecting one replaces the others without greying them out.",
    )
    for display_order, rpo in enumerate(PACKAGE_RPOS, start=1):
        ensure_group_member(wb, changes, PACKAGE_GROUP_ID, id_by_rpo[rpo], display_order * 10)

    ensure_group(
        wb,
        changes,
        AERO_GROUP_ID,
        "single_within_group",
        "Z06 aero choices are radio-like peers; T0E is the default and T0F/T0G/5ZV replace it without hard disabling peers.",
    )
    for display_order, rpo in enumerate(AERO_RPOS, start=1):
        ensure_group_member(wb, changes, AERO_GROUP_ID, id_by_rpo[rpo], display_order * 10)

    ensure_group(
        wb,
        changes,
        EXHAUST_GROUP_ID,
        "single_within_group",
        "Z06 exhaust tips are mutually exclusive; NGA is the default/restored tip and NWI replaces it when selected.",
    )
    for display_order, rpo in enumerate(EXHAUST_RPOS, start=1):
        ensure_group_member(wb, changes, EXHAUST_GROUP_ID, id_by_rpo[rpo], display_order * 10)


def apply_default_option_metadata(wb, changes: list[Change]) -> None:
    by_rpo, _, _ = option_maps(wb)
    nga = by_rpo["NGA"]
    row_number = int(nga["_row_number"])
    set_cell(wb[OPTION_SHEET], row_number, "selectable", "True", changes, key="NGA", reason="NGA should seed/restore as the default Z06 exhaust-tip peer")
    set_cell(wb[OPTION_SHEET], row_number, "display_behavior", "default_selected", changes, key="NGA", reason="NGA should seed/restore as the default Z06 exhaust-tip peer")
    set_cell(wb[OPTION_SHEET], row_number, "price", 0, changes, key="NGA", reason="NGA default exhaust tips should remain zero-price")


def omit_legacy_excludes(wb, changes: list[Change]) -> None:
    _, id_by_rpo, rpo_by_id = option_maps(wb)
    ws = wb[RULE_MAPPING_SHEET]
    for row_number, row in rows(ws):
        source_rpo = rpo_by_id.get(clean(row.get("source_id")), "")
        target_rpo = rpo_by_id.get(clean(row.get("target_id")), "")
        if clean(row.get("rule_type")).lower() != "excludes" or (source_rpo, target_rpo) not in OMIT_EXCLUDE_PAIRS:
            continue
        key = clean(row.get("rule_id")) or f"{source_rpo}->{target_rpo}"
        set_cell(ws, row_number, "generation_action", "omit_grouped_exclusion", changes, key=key, reason="Pass 3 replaces hard conflict with exclusive-group peer switching")
        set_cell(ws, row_number, "normalization_status", "omitted", changes, key=key, reason="Pass 3 replaces hard conflict with exclusive-group peer switching")
        set_cell(ws, row_number, "normalization_reason", "Pass 3: peer is represented by an active exclusive group, not a hard disable rule.", changes, key=key, reason="Pass 3 replaces hard conflict with exclusive-group peer switching")
        set_cell(ws, row_number, "replacement_group_id", AERO_GROUP_ID if source_rpo in {"5ZV", "PDD", "PDF"} and target_rpo in {"T0F", "T0G"} else PACKAGE_GROUP_ID, changes, key=key, reason="Pass 3 replaces hard conflict with exclusive-group peer switching")


def build_plan(wb) -> list[Change]:
    required = [OPTION_SHEET, RULE_MAPPING_SHEET, EXCLUSIVE_GROUPS_SHEET, EXCLUSIVE_MEMBERS_SHEET]
    missing = [sheet for sheet in required if sheet not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Missing required sheets: {', '.join(missing)}")
    changes: list[Change] = []
    apply_exclusive_groups(wb, changes)
    apply_default_option_metadata(wb, changes)
    omit_legacy_excludes(wb, changes)
    return changes


def verify_saved(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        by_rpo, id_by_rpo, rpo_by_id = option_maps(wb)
        checks: dict[str, Any] = {}
        # Exclusive groups.
        groups = {clean(row.get("group_id")): row for _, row in rows(wb[EXCLUSIVE_GROUPS_SHEET])}
        members_by_group: dict[str, set[str]] = {}
        for _, row in rows(wb[EXCLUSIVE_MEMBERS_SHEET]):
            if clean(row.get("active")) == "True":
                members_by_group.setdefault(clean(row.get("group_id")), set()).add(clean(row.get("option_id")))
        expected = {
            PACKAGE_GROUP_ID: {id_by_rpo[rpo] for rpo in PACKAGE_RPOS},
            AERO_GROUP_ID: {id_by_rpo[rpo] for rpo in AERO_RPOS},
            EXHAUST_GROUP_ID: {id_by_rpo[rpo] for rpo in EXHAUST_RPOS},
        }
        for group_id, option_ids in expected.items():
            if clean(groups.get(group_id, {}).get("active")) != "True":
                raise RuntimeError(f"{group_id} should be active")
            if clean(groups.get(group_id, {}).get("selection_mode")) != "single_within_group":
                raise RuntimeError(f"{group_id} should be single_within_group")
            if members_by_group.get(group_id, set()) != option_ids:
                raise RuntimeError(f"{group_id} members mismatch: {members_by_group.get(group_id, set())}")
            checks[group_id] = sorted(option_ids)
        nga = by_rpo["NGA"]
        if clean(nga.get("selectable")) != "True" or clean(nga.get("display_behavior")) != "default_selected":
            raise RuntimeError("NGA should be selectable default_selected")
        checks["nga_default"] = {"selectable": clean(nga.get("selectable")), "display_behavior": clean(nga.get("display_behavior"))}
        omitted = 0
        for _, row in rows(wb[RULE_MAPPING_SHEET]):
            source_rpo = rpo_by_id.get(clean(row.get("source_id")), "")
            target_rpo = rpo_by_id.get(clean(row.get("target_id")), "")
            if (source_rpo, target_rpo) in OMIT_EXCLUDE_PAIRS:
                if clean(row.get("generation_action")) != "omit_grouped_exclusion" or clean(row.get("normalization_status")) != "omitted":
                    raise RuntimeError(f"legacy exclude {source_rpo}->{target_rpo} should be omitted")
                omitted += 1
        checks["omitted_legacy_excludes"] = omitted
        if omitted != len(OMIT_EXCLUDE_PAIRS):
            raise RuntimeError(f"Expected {len(OMIT_EXCLUDE_PAIRS)} omitted legacy excludes, found {omitted}")
        return checks
    finally:
        wb.close()


def summarize(changes: list[Change]) -> dict[str, Any]:
    by_sheet: dict[str, int] = {}
    by_field: dict[str, int] = {}
    for change in changes:
        by_sheet[change.sheet] = by_sheet.get(change.sheet, 0) + 1
        by_field[change.field] = by_field.get(change.field, 0) + 1
    return {"total_changes": len(changes), "changes_by_sheet": dict(sorted(by_sheet.items())), "changes_by_field": dict(sorted(by_field.items()))}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="Write workbook changes. Default is dry-run.")
    parser.add_argument("--include-changes", action="store_true", help="Include per-cell changes in JSON report.")
    args = parser.parse_args(argv)

    if args.write and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")
    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    backup_path = None
    try:
        changes = build_plan(wb)
        if not args.write:
            report: dict[str, Any] = {"status": "dry_run", **summarize(changes), "workbook": str(WORKBOOK_PATH)}
            if args.include_changes:
                report["changes"] = [change.__dict__ for change in changes]
            print(json.dumps(report, indent=2, default=str))
            return 0
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved(WORKBOOK_PATH)
    report = {"status": "written", **summarize(changes), "workbook": str(WORKBOOK_PATH), "backup": str(backup_path), "verification": verification}
    if args.include_changes:
        report["changes"] = [change.__dict__ for change in changes]
    print(json.dumps(report, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
