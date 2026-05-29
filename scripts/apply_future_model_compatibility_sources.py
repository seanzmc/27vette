#!/usr/bin/env python3
"""Preview or write Grand Sport compatibility sources rebased to Z models.

Dry-run mode reads stingray_master.xlsx and prints JSON or Markdown describing
proposed Z06/ZR1/ZR1X rule/group/exclusive rows. Explicit --write persists only
the successfully mapped rows into future-model workbook source sheets; it does
not mutate generated form_* sheets or runtime app data.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_compatibility import (  # noqa: E402
    FUTURE_COMPATIBILITY_SPECS,
    apply_future_compatibility_to_workbook,
    build_future_compatibility_preview,
    render_compatibility_preview_markdown,
    verify_saved_compatibility_workbook,
)
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import excel_lock_path, save_workbook_safely  # noqa: E402


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model-key", choices=("all", *FUTURE_COMPATIBILITY_SPECS), default="all")
    parser.add_argument("--format", choices=("json", "markdown"), default="json")
    parser.add_argument("--include-rows", action="store_true", help="Include proposed row payloads in JSON output.")
    parser.add_argument("--dry-run", action="store_true", help="Report projected rows without saving the workbook. This is the default unless --write is passed.")
    parser.add_argument("--write", action="store_true", help="Persist proposed rows into the selected future-model source sheets.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=WORKBOOK_PATH,
        help="Workbook to inspect or write. Default: stingray_master.xlsx in the repository root.",
    )
    args = parser.parse_args(argv)

    workbook_path = args.workbook
    selected = list(FUTURE_COMPATIBILITY_SPECS) if args.model_key == "all" else [args.model_key]

    if not args.write:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            preview = build_future_compatibility_preview(wb, selected, include_rows=args.include_rows)
        finally:
            wb.close()

        if args.format == "markdown":
            print(render_compatibility_preview_markdown(preview))
        else:
            print(json.dumps(preview, indent=2, ensure_ascii=False))
        return 0

    if excel_lock_path(workbook_path).exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {excel_lock_path(workbook_path)}")

    loaded_mtime_ns = workbook_path.stat().st_mtime_ns
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    backup_path = None
    try:
        preview = build_future_compatibility_preview(wb, selected, include_rows=True)
        apply_report = apply_future_compatibility_to_workbook(wb, preview)
        backup_path = save_workbook_safely(wb, workbook_path, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_compatibility_workbook(workbook_path, preview)
    output = {
        **preview,
        "status": "written",
        "workbook": str(workbook_path),
        "backup": str(backup_path),
        "apply_report": apply_report,
        "verification": verification,
        "notes": [
            "Workbook source sheets were written for selected future models.",
            "Generated form_* sheets and runtime app data were not written.",
            *preview.get("notes", []),
        ],
    }
    if not args.include_rows:
        for model in output["models"].values():
            model.pop("proposed_rows", None)
    if args.format == "markdown":
        print(render_compatibility_preview_markdown(output))
    else:
        print(json.dumps(output, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
