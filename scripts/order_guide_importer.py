"""Shared helpers for Chevrolet order guide import staging scripts."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STATUS_SYMBOLS = ROOT / "data" / "import_maps" / "chevrolet_common" / "status_symbols.csv"
DEFAULT_PHRASE_PATTERNS = ROOT / "data" / "import_maps" / "chevrolet_common" / "phrase_patterns.csv"
DEFAULT_SHEET_ROLES = ROOT / "data" / "import_maps" / "corvette_2027" / "sheet_roles.csv"
DEFAULT_MODEL_CODES = ROOT / "data" / "import_maps" / "corvette_2027" / "model_codes.csv"

PRIMARY_MATRIX_ROLES = {"primary_variant_matrix"}
STATUS_CONTEXTS = {"variant_availability", "color_trim_compatibility", "availability_matrix"}
VARIANT_HEADER_RE = re.compile(r"(?P<body_style>[A-Za-z ]+?)\s+(?P<body_code>\d?[A-Z]{2}\d{2})\s+(?P<trim>[0-9A-Z]{2,3})")
RPO_RE = re.compile(r"\b[A-Z0-9]{3}\b")
TRAILING_FOOTNOTE_RE = re.compile(r"^(?P<base>.*?)(?P<footnotes>\d+)$")
STATUS_RE = re.compile(r"^(?P<symbol>A|S|D|--|■|□|\*)(?P<footnotes>\d*)$")

MODEL_HEADER_ALIASES = {
    "stingray": ("stingray",),
    "grand_sport": ("grand sport", "grand-sport"),
    "z06": ("z06",),
    "zr1": ("zr1",),
    "zr1x": ("zr1x",),
}

SECTION_ALIAS_DEFAULTS = {
    "color_trim_interior_header": ("decor level", "seat type", "seat code", "seat trim"),
    "color_trim_compatibility_header": ("exterior solid paint", "color code", "touch-up paint number"),
    "color_trim_interior_colors": ("interior colors",),
}

STAGING_OUTPUTS = {
    "sheets": [
        "sheet_name",
        "sheet_role",
        "section_family",
        "model_group_index",
        "guide_family",
        "model_key",
        "scope_type",
        "creates_canonical_candidates",
        "notes",
    ],
    "sheet_sections": [
        "source_sheet",
        "section_role",
        "section_index",
        "start_row",
        "end_row",
        "header_row",
        "data_start_row",
        "data_end_row",
        "guide_family",
        "model_key",
        "scope_type",
        "confidence",
        "notes",
    ],
    "variants": [
        "source_sheet",
        "guide_family",
        "model_key",
        "variant_label",
        "body_code",
        "body_style",
        "trim_level",
        "inferred_variant_id",
        "source_cell_range",
        "confidence",
        "notes",
    ],
    "variant_matrix_rows": [
        "source_sheet",
        "source_row",
        "guide_family",
        "model_key",
        "model_key_confidence",
        "section_family",
        "orderable_rpo",
        "ref_rpo",
        "description",
        "variant_id",
        "body_code",
        "body_style",
        "trim_level",
        "raw_status",
        "status_symbol",
        "footnote_refs",
        "canonical_status",
        "source_detail_raw",
    ],
    "color_trim_rows": [
        "source_sheet",
        "source_row",
        "guide_family",
        "model_key",
        "scope_type",
        "exterior_color_rpo",
        "exterior_color_name",
        "interior_code",
        "interior_label",
        "seat_code",
        "trim_level",
        "raw_status",
        "status_symbol",
        "footnote_refs",
        "canonical_status",
        "source_detail_raw",
    ],
    "color_trim_interior_rows": [
        "source_sheet",
        "source_row",
        "source_cell_range",
        "guide_family",
        "model_key",
        "scope_type",
        "trim_level",
        "seat_type",
        "seat_code",
        "seat_trim_raw",
        "seat_trim",
        "interior_color_name",
        "interior_rpo_raw",
        "interior_rpo",
        "footnote_refs",
        "footnote_scope",
        "confidence",
        "source_detail_raw",
    ],
    "color_trim_compatibility_rows": [
        "source_sheet",
        "source_row",
        "source_cell_range",
        "guide_family",
        "model_key",
        "scope_type",
        "exterior_color_name_raw",
        "exterior_color_name",
        "exterior_color_rpo_raw",
        "exterior_color_rpo",
        "touch_up_paint_code",
        "interior_rpo_raw",
        "interior_rpo",
        "raw_status",
        "status_symbol",
        "footnote_refs",
        "canonical_status",
        "confidence",
        "source_detail_raw",
    ],
    "color_trim_disclosures": [
        "source_sheet",
        "source_row",
        "source_cell_range",
        "guide_family",
        "model_key",
        "raw_text",
        "extracted_rpos",
        "phrase_type",
        "applies_to_section_role",
        "confidence",
        "review_status",
    ],
    "equipment_group_rows": [
        "source_sheet",
        "source_row",
        "guide_family",
        "model_key",
        "equipment_group_rpo",
        "orderable_rpo",
        "ref_rpo",
        "description",
        "row_kind",
        "matched_primary_row_key",
        "match_status",
        "source_detail_raw",
    ],
    "price_rows": [
        "source_sheet",
        "source_row",
        "guide_family",
        "model_key",
        "price_block_label",
        "raw_values",
        "notes",
    ],
    "status_symbols": [
        "source_sheet",
        "status_context",
        "status_symbol",
        "raw_status_examples",
        "count",
        "canonical_status",
    ],
    "rule_phrase_candidates": [
        "source_sheet",
        "source_row",
        "model_key",
        "source_field",
        "raw_text",
        "phrase_type",
        "extracted_rpos",
        "confidence",
        "review_status",
        "notes",
    ],
    "unresolved_rows": [
        "source_sheet",
        "source_row",
        "raw_values",
        "reason",
        "review_status",
    ],
    "ignored_rows": [
        "source_sheet",
        "source_row",
        "raw_values",
        "reason",
    ],
}


@dataclass(frozen=True)
class Section:
    source_sheet: str
    section_role: str
    section_index: int
    start_row: int
    end_row: int
    header_row: int
    data_start_row: int
    data_end_row: int
    guide_family: str
    model_key: str
    scope_type: str
    confidence: str
    notes: str

    def as_row(self) -> dict[str, Any]:
        return {
            "source_sheet": self.source_sheet,
            "section_role": self.section_role,
            "section_index": self.section_index,
            "start_row": self.start_row,
            "end_row": self.end_row,
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
            "data_end_row": self.data_end_row,
            "guide_family": self.guide_family,
            "model_key": self.model_key,
            "scope_type": self.scope_type,
            "confidence": self.confidence,
            "notes": self.notes,
        }


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized_text(value: str) -> str:
    return re.sub(r"\s+", " ", clean(value).replace("\n", " ")).strip().lower()


def raw_row_values(values: list[Any]) -> str:
    return json.dumps([clean(value) for value in values], ensure_ascii=False, separators=(",", ":"))


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def load_status_map(path: Path = DEFAULT_STATUS_SYMBOLS) -> dict[str, dict[str, str]]:
    return {row["raw_symbol"]: row for row in load_csv_rows(path)}


def load_phrase_patterns(path: Path = DEFAULT_PHRASE_PATTERNS) -> list[dict[str, Any]]:
    return [{**row, "compiled": re.compile(row["pattern"], re.IGNORECASE)} for row in load_csv_rows(path)]


def load_sheet_role_rows(path: Path = DEFAULT_SHEET_ROLES) -> list[dict[str, Any]]:
    return [{**row, "compiled": re.compile(row["sheet_name_pattern"])} for row in load_csv_rows(path)]


def load_model_codes(path: Path = DEFAULT_MODEL_CODES) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    return {row["body_code"]: row for row in load_csv_rows(path)}


def classify_sheet(sheet_name: str, role_rows: list[dict[str, Any]]) -> dict[str, str]:
    for row in role_rows:
        match = row["compiled"].match(sheet_name)
        if not match:
            continue
        model_group_index = clean(row.get("model_group_index", ""))
        if not model_group_index and row.get("model_group_index_source") == "sheet_number":
            model_group_index = clean(match.groupdict().get("index", ""))
        return {
            "sheet_name": sheet_name,
            "sheet_role": row["sheet_role"],
            "section_family": row["section_family"],
            "scope_type": row["scope_type"],
            "creates_canonical_candidates": row["creates_canonical_candidates"].lower(),
            "guide_family": row.get("guide_family", row.get("model_key", "")),
            "model_key": row.get("model_key", "") if row.get("model_key", "") != row.get("guide_family", "") else "",
            "model_group_index": model_group_index,
            "notes": row["notes"],
        }
    return {
        "sheet_name": sheet_name,
        "sheet_role": "ignored_or_unknown",
        "section_family": "unknown",
        "scope_type": "unknown",
        "creates_canonical_candidates": "false",
        "guide_family": "",
        "model_key": "",
        "model_group_index": "",
        "notes": "No matching sheet role map row.",
    }


def load_workbook_readonly(source: Path):
    return load_workbook(source, read_only=True, data_only=True)


def worksheet_rows(ws) -> list[tuple[int, list[str]]]:
    return [(row_index, [clean(value) for value in values]) for row_index, values in enumerate(ws.iter_rows(values_only=True), start=1)]


def is_blank(values: list[str]) -> bool:
    return not any(clean(value) for value in values)


def value_at(values: list[str], index: int) -> str:
    if index >= len(values):
        return ""
    return clean(values[index])


def column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def cell_ref(row_index: int, column_index: int) -> str:
    return f"{column_letter(column_index)}{row_index}"


def row_contains_aliases(values: list[str], aliases: tuple[str, ...]) -> bool:
    normalized_values = {normalized_text(value) for value in values if value}
    joined = " | ".join(sorted(normalized_values))
    return all(alias in joined for alias in aliases)


def find_matrix_header(rows: list[tuple[int, list[str]]]) -> tuple[int, list[str]] | None:
    for row_index, values in rows:
        normalized = [normalized_text(value) for value in values]
        if "orderable rpo code" in normalized and "description" in normalized:
            return row_index, values
    return None


def detect_sections(sheet_name: str, classification: dict[str, str], rows: list[tuple[int, list[str]]]) -> list[Section]:
    role = classification["sheet_role"]
    guide_family = classification["guide_family"]
    if role in PRIMARY_MATRIX_ROLES:
        header = find_matrix_header(rows)
        if not header:
            return [unknown_section(sheet_name, classification, rows, "No variant matrix header found.")]
        header_row = header[0]
        return [
            Section(sheet_name, "header_or_legend", 1, 1, header_row, 0, 1, header_row, guide_family, "", "metadata", "high", "Rows before and including matrix header."),
            Section(
                sheet_name,
                "variant_matrix",
                2,
                header_row + 1,
                rows[-1][0] if rows else header_row,
                header_row,
                header_row + 1,
                rows[-1][0] if rows else header_row,
                guide_family,
                "",
                "variant_scoped",
                "high",
                "Primary variant availability matrix.",
            ),
        ]
    if role == "derived_equipment_summary":
        header = find_matrix_header(rows)
        if not header:
            return [unknown_section(sheet_name, classification, rows, "No Equipment Groups matrix header found.")]
        header_row = header[0]
        return [
            Section(sheet_name, "header_or_legend", 1, 1, header_row, 0, 1, header_row, guide_family, "", "metadata", "high", "Rows before and including derived matrix header."),
            Section(
                sheet_name,
                "derived_equipment_summary",
                2,
                header_row + 1,
                rows[-1][0] if rows else header_row,
                header_row,
                header_row + 1,
                rows[-1][0] if rows else header_row,
                guide_family,
                "",
                "derived_cross_check",
                "high",
                "Equipment Groups derived/cross-check rows only.",
            ),
        ]
    if role == "model_global_matrix":
        return detect_color_trim_sections(sheet_name, classification, rows)
    if role == "price_source":
        return detect_price_sections(sheet_name, classification, rows)
    return [unknown_section(sheet_name, classification, rows, "No section detector for sheet role.")]


def unknown_section(sheet_name: str, classification: dict[str, str], rows: list[tuple[int, list[str]]], notes: str) -> Section:
    first_row = rows[0][0] if rows else 0
    last_row = rows[-1][0] if rows else 0
    return Section(sheet_name, "ignored_or_unknown", 1, first_row, last_row, 0, first_row, last_row, classification["guide_family"], "", classification["scope_type"], "needs_review", notes)


def detect_color_trim_sections(sheet_name: str, classification: dict[str, str], rows: list[tuple[int, list[str]]]) -> list[Section]:
    interior_headers = [row_index for row_index, values in rows if row_contains_aliases(values, SECTION_ALIAS_DEFAULTS["color_trim_interior_header"])]
    compatibility_headers = [row_index for row_index, values in rows if row_contains_aliases(values, SECTION_ALIAS_DEFAULTS["color_trim_compatibility_header"])]
    disclosure_rows = [row_index for row_index, values in rows if normalized_text(" ".join(values)).startswith(("•", "note:", "requires option code")) or "• note:" in normalized_text(" ".join(values))]
    sections: list[Section] = []
    guide_family = classification["guide_family"]
    section_index = 1
    if interior_headers:
        header_row = interior_headers[0]
        next_boundary = min([row for row in compatibility_headers + disclosure_rows if row > header_row], default=rows[-1][0] + 1)
        sections.append(
            Section(
                sheet_name,
                "color_trim_interior_matrix",
                section_index,
                max(1, header_row - 1),
                next_boundary - 1,
                header_row,
                header_row + 1,
                next_boundary - 1,
                guide_family,
                "",
                "model_global",
                "high",
                "Detected from Decor Level/Seat Type/Seat Code/Seat Trim header aliases.",
            )
        )
        section_index += 1
    for header_row in compatibility_headers:
        next_boundary = min([row for row in disclosure_rows if row > header_row], default=rows[-1][0] + 1)
        sections.append(
            Section(
                sheet_name,
                "color_trim_compatibility_matrix",
                section_index,
                max(1, header_row - 1),
                next_boundary - 1,
                header_row,
                header_row + 1,
                next_boundary - 1,
                guide_family,
                "",
                "model_global",
                "high",
                "Detected from Exterior Solid Paint/Color Code/Touch-Up Paint Number header aliases.",
            )
        )
        section_index += 1
    for row_index in disclosure_rows:
        sections.append(
            Section(
                sheet_name,
                "color_trim_disclosure",
                section_index,
                row_index,
                row_index,
                0,
                row_index,
                row_index,
                guide_family,
                "",
                "model_global",
                "medium",
                "Detected disclosure/rule note row.",
            )
        )
        section_index += 1
    if not sections:
        sections.append(unknown_section(sheet_name, classification, rows, "Color and Trim section anchors not found."))
    return sections


def detect_price_sections(sheet_name: str, classification: dict[str, str], rows: list[tuple[int, list[str]]]) -> list[Section]:
    sections = []
    guide_family = classification["guide_family"]
    section_index = 1
    block_start = rows[0][0] if rows else 0
    for row_index, values in rows:
        if "price" in normalized_text(" ".join(values)) or "base model prices" in normalized_text(" ".join(values)):
            sections.append(
                Section(sheet_name, "price_block", section_index, block_start, rows[-1][0], row_index, row_index + 1, rows[-1][0], guide_family, "", "model_global", "medium", "High-level price block; staging evidence only.")
            )
            section_index += 1
            break
    if not sections:
        sections.append(Section(sheet_name, "price_block", 1, block_start, rows[-1][0] if rows else 0, 0, block_start, rows[-1][0] if rows else 0, guide_family, "", "model_global", "low", "Price sheet without clear block header."))
    return sections


def row_map(rows: list[tuple[int, list[str]]]) -> dict[int, list[str]]:
    return {row_index: values for row_index, values in rows}


def parse_variant_header(value: str, model_codes: dict[str, dict[str, str]] | None = None) -> dict[str, str]:
    text = re.sub(r"\s+", " ", clean(value).replace("\n", " ")).strip()
    match = VARIANT_HEADER_RE.search(text)
    if not match:
        return {
            "variant_label": text,
            "body_code": "",
            "body_style": "",
            "trim_level": "",
            "model_key": "",
            "model_key_confidence": "needs_review",
            "inferred_variant_id": "",
            "confidence": "needs_review",
            "notes": "Could not parse variant header.",
        }
    body_style = normalized_body_style(match.group("body_style"))
    body_code = match.group("body_code")
    trim = match.group("trim")
    model_result = resolve_model_key(body_code, text, model_codes or {})
    model_key = model_result["model_key"]
    variant_prefix = model_key if model_key else "unresolved"
    body_suffix = body_code[-3:].lower()
    inferred_variant_id = f"{variant_prefix}_{trim.lower()}_{body_suffix}" if model_key else ""
    confidence = "high" if model_result["confidence"] == "high" else "needs_review"
    return {
        "variant_label": text,
        "body_code": body_code,
        "body_style": body_style,
        "trim_level": trim,
        "model_key": model_key,
        "model_key_confidence": model_result["confidence"],
        "inferred_variant_id": inferred_variant_id,
        "confidence": confidence,
        "notes": model_result["notes"],
    }


def normalized_body_style(value: str) -> str:
    text = normalized_text(value)
    if "convertible" in text:
        return "convertible"
    if "coupe" in text:
        return "coupe"
    return text.replace(" ", "_")


def resolve_model_key(body_code: str, variant_label: str, model_codes: dict[str, dict[str, str]]) -> dict[str, str]:
    row = model_codes.get(body_code)
    if not row:
        return {"model_key": "", "confidence": "needs_review", "notes": f"Body code {body_code} is not mapped; model_key left blank."}
    mapped_key = row["model_key"]
    label = normalized_text(variant_label)
    mentioned_keys = [key for key, aliases in MODEL_HEADER_ALIASES.items() if any(alias in label for alias in aliases)]
    if mentioned_keys and mapped_key not in mentioned_keys:
        return {
            "model_key": "",
            "confidence": "needs_review",
            "notes": f"Body code {body_code} maps to {mapped_key}, but header text suggests {', '.join(mentioned_keys)}.",
        }
    return {
        "model_key": mapped_key,
        "confidence": row.get("confidence", "high") or "high",
        "notes": f"Model key resolved from workbook body code {body_code}.",
    }


def variant_columns(header_values: list[str], model_codes: dict[str, dict[str, str]]) -> list[dict[str, str]]:
    columns = []
    for index, value in enumerate(header_values, start=1):
        parsed = parse_variant_header(value, model_codes)
        if parsed["body_code"] or "\n" in value:
            columns.append({**parsed, "column_index": str(index), "source_cell_range": column_letter(index)})
    return columns


def parse_status(raw_status: str, status_map: dict[str, dict[str, str]], context: str) -> dict[str, str]:
    raw = clean(raw_status)
    if not raw:
        return {"raw_status": "", "status_symbol": "", "footnote_refs": "", "canonical_status": "", "confidence": ""}
    if context not in STATUS_CONTEXTS:
        return {"raw_status": raw, "status_symbol": "", "footnote_refs": "", "canonical_status": "", "confidence": "rejected_context"}
    match = STATUS_RE.match(raw)
    if not match:
        return {"raw_status": raw, "status_symbol": "", "footnote_refs": "", "canonical_status": "", "confidence": "rejected_pattern"}
    symbol = match.group("symbol")
    if symbol not in status_map:
        return {"raw_status": raw, "status_symbol": symbol, "footnote_refs": match.group("footnotes"), "canonical_status": "unknown", "confidence": "needs_review"}
    return {
        "raw_status": raw,
        "status_symbol": symbol,
        "footnote_refs": match.group("footnotes"),
        "canonical_status": status_map[symbol]["canonical_status"],
        "confidence": "high",
    }


def split_footnote(raw_value: str, field_context: str) -> dict[str, str]:
    raw = clean(raw_value)
    if not raw:
        return {"raw_value": "", "parsed_value": "", "footnote_refs": "", "footnote_scope": "", "confidence": ""}
    if field_context in {"seat_trim", "interior_color_header", "exterior_color_name"}:
        match = TRAILING_FOOTNOTE_RE.match(raw)
        if match and match.group("base").strip():
            return {
                "raw_value": raw,
                "parsed_value": match.group("base").strip(),
                "footnote_refs": match.group("footnotes"),
                "footnote_scope": field_context,
                "confidence": "high",
            }
    if field_context in {"interior_rpo", "exterior_color_rpo"}:
        if re.fullmatch(r"[A-Z0-9]{3}", raw):
            return {"raw_value": raw, "parsed_value": raw, "footnote_refs": "", "footnote_scope": "", "confidence": "high"}
        if re.fullmatch(r"[A-Z0-9]{3}\d+", raw):
            return {
                "raw_value": raw,
                "parsed_value": raw[:3],
                "footnote_refs": raw[3:],
                "footnote_scope": field_context,
                "confidence": "medium",
            }
        return {"raw_value": raw, "parsed_value": raw, "footnote_refs": "", "footnote_scope": "", "confidence": "needs_review"}
    return {"raw_value": raw, "parsed_value": raw, "footnote_refs": "", "footnote_scope": "", "confidence": "high"}


def rule_phrase_candidates(
    *,
    source_sheet: str,
    source_row: int,
    model_key: str,
    source_field: str,
    raw_text: str,
    patterns: list[dict[str, Any]],
) -> list[dict[str, str]]:
    text = clean(raw_text)
    if not text:
        return []
    candidates = []
    extracted_rpos = "|".join(sorted(set(RPO_RE.findall(text))))
    for pattern in patterns:
        if pattern["compiled"].search(text):
            candidates.append(
                {
                    "source_sheet": source_sheet,
                    "source_row": str(source_row),
                    "model_key": model_key,
                    "source_field": source_field,
                    "raw_text": text,
                    "phrase_type": pattern["phrase_type"],
                    "extracted_rpos": extracted_rpos,
                    "confidence": pattern["confidence"],
                    "review_status": "needs_review",
                    "notes": "Staging candidate only; no canonical rule generated.",
                }
            )
    return candidates


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    normalized_rows = [{header: clean(row.get(header, "")) for header in headers} for row in rows]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, lineterminator="\n")
        writer.writeheader()
        writer.writerows(normalized_rows)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")


def profile_workbook(source: Path) -> dict[str, Any]:
    wb = load_workbook_readonly(source)
    role_rows = load_sheet_role_rows()
    status_map = load_status_map()
    patterns = load_phrase_patterns()
    model_codes = load_model_codes()
    sheets = []
    status_counter: Counter[tuple[str, str, str]] = Counter()
    status_rejections: Counter[str] = Counter()
    section_counter: Counter[str] = Counter()
    model_confidence_counter: Counter[str] = Counter()
    rule_candidate_count = 0
    variant_count = 0
    price_row_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        classification = classify_sheet(sheet_name, role_rows)
        rows = worksheet_rows(ws)
        sections = detect_sections(sheet_name, classification, rows)
        for section in sections:
            section_counter[section.section_role] += 1
        header = find_matrix_header(rows)
        variants = []
        if header:
            variants = variant_columns(header[1], model_codes)
            variant_count += len(variants)
            for variant in variants:
                model_confidence_counter[variant["model_key_confidence"]] += 1
        for row_index, values in rows:
            if classification["sheet_role"] == "price_source" and not is_blank(values):
                price_row_count += 1
            rule_candidate_count += len(
                rule_phrase_candidates(
                    source_sheet=sheet_name,
                    source_row=row_index,
                    model_key=classification["model_key"],
                    source_field="row",
                    raw_text=" ".join(value for value in values if value),
                    patterns=patterns,
                )
            )
            if header and row_index > header[0] and classification["sheet_role"] in PRIMARY_MATRIX_ROLES | {"derived_equipment_summary"}:
                for variant in variants:
                    parsed = parse_status(value_at(values, int(variant["column_index"]) - 1), status_map, "variant_availability")
                    record_status_counter(status_counter, status_rejections, sheet_name, "variant_availability", parsed)
        sheets.append(
            {
                **classification,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "likely_variant_headers": [variant["variant_label"] for variant in variants],
                "likely_rpo_columns": ["Orderable RPO Code", "Ref. Only RPO Code"] if header else [],
            }
        )

    return {
        "source_path": str(source),
        "sheet_count": len(wb.sheetnames),
        "sheets": sheets,
        "detected": {
            "model_keys": sorted({variant_key for sheet in sheets for variant_key in [sheet.get("model_key", "")] if variant_key}),
            "section_families": sorted({sheet["section_family"] for sheet in sheets if sheet["section_family"] != "unknown"}),
            "sheet_section_role_counts": dict(sorted(section_counter.items())),
            "variant_header_count": variant_count,
            "model_key_confidence_counts": dict(sorted(model_confidence_counter.items())),
            "status_symbols": [
                {"source_sheet": sheet, "status_context": context, "status_symbol": symbol, "canonical_status": canonical, "count": count}
                for (sheet, context, symbol, canonical), count in sorted(status_counter.items())
            ],
            "status_parse_rejections": dict(sorted(status_rejections.items())),
            "price_schedule_rows": price_row_count,
            "rule_phrase_candidate_count": rule_candidate_count,
        },
    }


def record_status_counter(
    status_counter: Counter[tuple[str, str, str, str]],
    status_rejections: Counter[str],
    sheet_name: str,
    context: str,
    parsed: dict[str, str],
) -> None:
    if parsed.get("canonical_status"):
        status_counter[(sheet_name, context, parsed["status_symbol"], parsed["canonical_status"])] += 1
    elif parsed.get("raw_status") and parsed.get("confidence", "").startswith("rejected"):
        status_rejections[parsed["raw_status"]] += 1


def extract_staging(source: Path) -> dict[str, list[dict[str, Any]] | dict[str, Any]]:
    wb = load_workbook_readonly(source)
    role_rows = load_sheet_role_rows()
    status_map = load_status_map()
    patterns = load_phrase_patterns()
    model_codes = load_model_codes()
    rows_by_name: dict[str, list[dict[str, Any]]] = {name: [] for name in STAGING_OUTPUTS}
    status_examples: dict[tuple[str, str, str, str], Counter[str]] = {}
    status_rejections: Counter[str] = Counter()
    footnote_scope_counts: Counter[str] = Counter()
    model_confidence_counts: Counter[str] = Counter()
    unresolved_reasons: Counter[str] = Counter()
    primary_keys: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        classification = classify_sheet(sheet_name, role_rows)
        rows = worksheet_rows(ws)
        rows_by_name["sheets"].append(classification)
        sections = detect_sections(sheet_name, classification, rows)
        for section in sections:
            rows_by_name["sheet_sections"].append(section.as_row())
        header = find_matrix_header(rows)
        header_row_index = header[0] if header else 0
        headers = header[1] if header else []
        variants = variant_columns(headers, model_codes) if header else []
        for variant in variants:
            model_confidence_counts[variant["model_key_confidence"]] += 1
            rows_by_name["variants"].append(
                {
                    "source_sheet": sheet_name,
                    "guide_family": classification["guide_family"],
                    "model_key": variant["model_key"],
                    **{key: variant[key] for key in ["variant_label", "body_code", "body_style", "trim_level", "inferred_variant_id", "source_cell_range", "confidence"]},
                    "notes": variant["notes"],
                }
            )
            if variant["model_key_confidence"] == "needs_review":
                unresolved_reasons["model_key_needs_review"] += 1

        row_lookup = row_map(rows)
        handled_rows: set[int] = set()
        for section in sections:
            if section.section_role == "variant_matrix":
                extract_variant_matrix_section(rows_by_name, status_examples, status_rejections, primary_keys, classification, section, row_lookup, variants, status_map, patterns)
                handled_rows.update(range(section.start_row, section.end_row + 1))
            elif section.section_role == "derived_equipment_summary":
                extract_equipment_group_section(rows_by_name, primary_keys, classification, section, row_lookup, patterns)
                handled_rows.update(range(section.start_row, section.end_row + 1))
            elif section.section_role == "color_trim_interior_matrix":
                extract_color_trim_interior_section(rows_by_name, status_rejections, footnote_scope_counts, classification, section, row_lookup)
                handled_rows.update(range(section.start_row, section.end_row + 1))
            elif section.section_role == "color_trim_compatibility_matrix":
                extract_color_trim_compatibility_section(rows_by_name, status_examples, status_rejections, footnote_scope_counts, classification, section, row_lookup, status_map)
                handled_rows.update(range(section.start_row, section.end_row + 1))
            elif section.section_role == "color_trim_disclosure":
                extract_color_trim_disclosure(rows_by_name, classification, section, row_lookup, patterns)
                handled_rows.update(range(section.start_row, section.end_row + 1))
            elif section.section_role == "price_block":
                extract_price_section(rows_by_name, classification, section, row_lookup)
                handled_rows.update(range(section.start_row, section.end_row + 1))

        for row_index, values in rows:
            if row_index in handled_rows:
                continue
            if is_blank(values):
                rows_by_name["ignored_rows"].append({"source_sheet": sheet_name, "source_row": row_index, "raw_values": raw_row_values(values), "reason": "blank_row"})
            elif row_index <= header_row_index and header:
                rows_by_name["ignored_rows"].append({"source_sheet": sheet_name, "source_row": row_index, "raw_values": raw_row_values(values), "reason": "header_or_legend_row"})
            elif classification["sheet_role"] == "ignored_or_unknown":
                reason = "ignored_or_unknown_sheet_with_content"
                unresolved_reasons[reason] += 1
                rows_by_name["rule_phrase_candidates"].extend(
                    rule_phrase_candidates(
                        source_sheet=sheet_name,
                        source_row=row_index,
                        model_key="",
                        source_field="row",
                        raw_text=" ".join(value for value in values if value),
                        patterns=patterns,
                    )
                )
                rows_by_name["unresolved_rows"].append({"source_sheet": sheet_name, "source_row": row_index, "raw_values": raw_row_values(values), "reason": reason, "review_status": "needs_review"})

    for (sheet_name, context, symbol, canonical_status), examples in sorted(status_examples.items()):
        rows_by_name["status_symbols"].append(
            {
                "source_sheet": sheet_name,
                "status_context": context,
                "status_symbol": symbol,
                "raw_status_examples": "|".join(sorted(examples)),
                "count": sum(examples.values()),
                "canonical_status": canonical_status,
            }
        )

    report = {
        "source_path": str(source),
        "row_counts": {name: len(rows) for name, rows in rows_by_name.items()},
        "section_role_counts": dict(sorted(Counter(row["section_role"] for row in rows_by_name["sheet_sections"]).items())),
        "model_key_confidence_counts": dict(sorted(model_confidence_counts.items())),
        "unresolved_model_key_count": model_confidence_counts.get("needs_review", 0),
        "color_trim_interior_rows": len(rows_by_name["color_trim_interior_rows"]),
        "color_trim_compatibility_rows": len(rows_by_name["color_trim_compatibility_rows"]),
        "color_trim_disclosure_rows": len(rows_by_name["color_trim_disclosures"]),
        "status_parse_context_counts": dict(sorted(Counter(row["status_context"] for row in rows_by_name["status_symbols"]).items())),
        "status_parse_rejections": dict(sorted(status_rejections.items())),
        "footnote_scope_counts": dict(sorted(footnote_scope_counts.items())),
        "unresolved_rows_by_reason": dict(sorted(unresolved_reasons.items())),
        "notes": [
            "Staging output is structured raw evidence, not canonical truth.",
            "No canonical proposal rows are generated by this scaffold.",
        ],
    }
    return {**rows_by_name, "report": report}


def add_status_example(
    status_examples: dict[tuple[str, str, str, str], Counter[str]],
    status_rejections: Counter[str],
    sheet_name: str,
    context: str,
    parsed: dict[str, str],
) -> None:
    if parsed.get("canonical_status"):
        key = (sheet_name, context, parsed["status_symbol"], parsed["canonical_status"])
        status_examples.setdefault(key, Counter())[parsed["raw_status"]] += 1
    elif parsed.get("raw_status") and parsed.get("confidence", "").startswith("rejected"):
        status_rejections[parsed["raw_status"]] += 1


def extract_variant_matrix_section(
    rows_by_name: dict[str, list[dict[str, Any]]],
    status_examples: dict[tuple[str, str, str, str], Counter[str]],
    status_rejections: Counter[str],
    primary_keys: set[str],
    classification: dict[str, str],
    section: Section,
    row_lookup: dict[int, list[str]],
    variants: list[dict[str, str]],
    status_map: dict[str, dict[str, str]],
    patterns: list[dict[str, Any]],
) -> None:
    for row_index in range(section.data_start_row, section.data_end_row + 1):
        values = row_lookup.get(row_index, [])
        if is_blank(values):
            rows_by_name["ignored_rows"].append({"source_sheet": section.source_sheet, "source_row": row_index, "raw_values": raw_row_values(values), "reason": "blank_row"})
            continue
        orderable_rpo = value_at(values, 0)
        ref_rpo = value_at(values, 1)
        description = value_at(values, 2)
        if not (orderable_rpo or ref_rpo or description):
            rows_by_name["ignored_rows"].append({"source_sheet": section.source_sheet, "source_row": row_index, "raw_values": raw_row_values(values), "reason": "matrix_row_without_rpo_or_description"})
            continue
        rows_by_name["rule_phrase_candidates"].extend(
            rule_phrase_candidates(
                source_sheet=section.source_sheet,
                source_row=row_index,
                model_key="",
                source_field="description",
                raw_text=description,
                patterns=patterns,
            )
        )
        for variant in variants:
            column_index = int(variant["column_index"])
            raw_status = value_at(values, column_index - 1)
            parsed = parse_status(raw_status, status_map, "variant_availability")
            add_status_example(status_examples, status_rejections, section.source_sheet, "variant_availability", parsed)
            row_key = f"{variant['model_key']}|{classification['section_family']}|{orderable_rpo}|{ref_rpo}|{description}"
            primary_keys.add(row_key)
            rows_by_name["variant_matrix_rows"].append(
                {
                    "source_sheet": section.source_sheet,
                    "source_row": row_index,
                    "guide_family": classification["guide_family"],
                    "model_key": variant["model_key"],
                    "model_key_confidence": variant["model_key_confidence"],
                    "section_family": classification["section_family"],
                    "orderable_rpo": orderable_rpo,
                    "ref_rpo": ref_rpo,
                    "description": description,
                    "variant_id": variant["inferred_variant_id"],
                    "body_code": variant["body_code"],
                    "body_style": variant["body_style"],
                    "trim_level": variant["trim_level"],
                    "raw_status": parsed["raw_status"],
                    "status_symbol": parsed["status_symbol"],
                    "footnote_refs": parsed["footnote_refs"],
                    "canonical_status": parsed["canonical_status"],
                    "source_detail_raw": description,
                }
            )


def extract_equipment_group_section(
    rows_by_name: dict[str, list[dict[str, Any]]],
    primary_keys: set[str],
    classification: dict[str, str],
    section: Section,
    row_lookup: dict[int, list[str]],
    patterns: list[dict[str, Any]],
) -> None:
    for row_index in range(section.data_start_row, section.data_end_row + 1):
        values = row_lookup.get(row_index, [])
        if is_blank(values):
            rows_by_name["ignored_rows"].append({"source_sheet": section.source_sheet, "source_row": row_index, "raw_values": raw_row_values(values), "reason": "blank_row"})
            continue
        orderable_rpo = value_at(values, 0)
        ref_rpo = value_at(values, 1)
        description = value_at(values, 2)
        if not (orderable_rpo or ref_rpo or description) or orderable_rpo == "Equipment Groups":
            reason = "equipment_group_label_row" if orderable_rpo == "Equipment Groups" else "equipment_group_row_without_rpo_or_description"
            rows_by_name["ignored_rows"].append({"source_sheet": section.source_sheet, "source_row": row_index, "raw_values": raw_row_values(values), "reason": reason})
            continue
        rows_by_name["rule_phrase_candidates"].extend(
            rule_phrase_candidates(
                source_sheet=section.source_sheet,
                source_row=row_index,
                model_key="",
                source_field="description",
                raw_text=description,
                patterns=patterns,
            )
        )
        row_key_suffix = f"|{classification['section_family']}|{orderable_rpo}|{ref_rpo}|{description}"
        matched_keys = [key for key in primary_keys if key.endswith(row_key_suffix)]
        rows_by_name["equipment_group_rows"].append(
            {
                "source_sheet": section.source_sheet,
                "source_row": row_index,
                "guide_family": classification["guide_family"],
                "model_key": "",
                "equipment_group_rpo": orderable_rpo,
                "orderable_rpo": orderable_rpo,
                "ref_rpo": ref_rpo,
                "description": description,
                "row_kind": "derived_cross_check",
                "matched_primary_row_key": "|".join(sorted(matched_keys)),
                "match_status": "matched_primary" if matched_keys else "unmatched_primary_review",
                "source_detail_raw": description,
            }
        )


def extract_color_trim_interior_section(
    rows_by_name: dict[str, list[dict[str, Any]]],
    status_rejections: Counter[str],
    footnote_scope_counts: Counter[str],
    classification: dict[str, str],
    section: Section,
    row_lookup: dict[int, list[str]],
) -> None:
    header_values = row_lookup.get(section.header_row, [])
    interior_headers: list[dict[str, str]] = []
    for column_index, header in enumerate(header_values[4:], start=5):
        if not header:
            continue
        split = split_footnote(header, "interior_color_header")
        if split["footnote_scope"]:
            footnote_scope_counts[split["footnote_scope"]] += 1
        interior_headers.append(
            {
                "column_index": str(column_index),
                "interior_color_name": split["parsed_value"],
                "header_footnote_refs": split["footnote_refs"],
            }
        )
    for row_index in range(section.data_start_row, section.data_end_row + 1):
        values = row_lookup.get(row_index, [])
        if is_blank(values) or normalized_text(value_at(values, 0)) in {"decor level", "exterior solid paint"}:
            continue
        seat_trim_split = split_footnote(value_at(values, 3), "seat_trim")
        if seat_trim_split["footnote_scope"]:
            footnote_scope_counts[seat_trim_split["footnote_scope"]] += 1
        for header in interior_headers:
            column_index = int(header["column_index"])
            raw_rpo = value_at(values, column_index - 1)
            if not raw_rpo or raw_rpo == "--":
                continue
            status_rejections[raw_rpo] += 1
            rpo_split = split_footnote(raw_rpo, "interior_rpo")
            footnote_refs = "|".join(filter(None, [header["header_footnote_refs"], seat_trim_split["footnote_refs"], rpo_split["footnote_refs"]]))
            footnote_scope = "|".join(filter(None, ["interior_color_header" if header["header_footnote_refs"] else "", seat_trim_split["footnote_scope"], rpo_split["footnote_scope"]]))
            if rpo_split["footnote_scope"]:
                footnote_scope_counts[rpo_split["footnote_scope"]] += 1
            rows_by_name["color_trim_interior_rows"].append(
                {
                    "source_sheet": section.source_sheet,
                    "source_row": row_index,
                    "source_cell_range": cell_ref(row_index, column_index),
                    "guide_family": classification["guide_family"],
                    "model_key": "",
                    "scope_type": "model_global",
                    "trim_level": value_at(values, 0),
                    "seat_type": value_at(values, 1),
                    "seat_code": value_at(values, 2),
                    "seat_trim_raw": seat_trim_split["raw_value"],
                    "seat_trim": seat_trim_split["parsed_value"],
                    "interior_color_name": header["interior_color_name"],
                    "interior_rpo_raw": raw_rpo,
                    "interior_rpo": rpo_split["parsed_value"],
                    "footnote_refs": footnote_refs,
                    "footnote_scope": footnote_scope,
                    "confidence": "medium" if rpo_split["confidence"] == "medium" else "high",
                    "source_detail_raw": raw_row_values(values),
                }
            )
            rows_by_name["color_trim_rows"].append(
                {
                    "source_sheet": section.source_sheet,
                    "source_row": row_index,
                    "guide_family": classification["guide_family"],
                    "model_key": "",
                    "scope_type": "model_global",
                    "exterior_color_rpo": "",
                    "exterior_color_name": "",
                    "interior_code": raw_rpo,
                    "interior_label": seat_trim_split["parsed_value"],
                    "seat_code": value_at(values, 2),
                    "trim_level": value_at(values, 0),
                    "raw_status": "",
                    "status_symbol": "",
                    "footnote_refs": footnote_refs,
                    "canonical_status": "",
                    "source_detail_raw": raw_row_values(values),
                }
            )


def extract_color_trim_compatibility_section(
    rows_by_name: dict[str, list[dict[str, Any]]],
    status_examples: dict[tuple[str, str, str, str], Counter[str]],
    status_rejections: Counter[str],
    footnote_scope_counts: Counter[str],
    classification: dict[str, str],
    section: Section,
    row_lookup: dict[int, list[str]],
    status_map: dict[str, dict[str, str]],
) -> None:
    header_values = row_lookup.get(section.header_row, [])
    interior_headers = []
    for column_index, header in enumerate(header_values[4:], start=5):
        if not header:
            continue
        split = split_footnote(header.replace(" / en-us", ""), "interior_color_header")
        if split["footnote_scope"]:
            footnote_scope_counts[split["footnote_scope"]] += 1
        interior_headers.append({"column_index": str(column_index), "raw": header, "parsed": split["parsed_value"], "footnote_refs": split["footnote_refs"]})
    for row_index in range(section.data_start_row, section.data_end_row + 1):
        values = row_lookup.get(row_index, [])
        if is_blank(values):
            continue
        exterior_name_split = split_footnote(value_at(values, 0), "exterior_color_name")
        if exterior_name_split["footnote_scope"]:
            footnote_scope_counts[exterior_name_split["footnote_scope"]] += 1
        exterior_rpo_split = split_footnote(value_at(values, 2), "exterior_color_rpo")
        for header in interior_headers:
            column_index = int(header["column_index"])
            raw_status = value_at(values, column_index - 1)
            if not raw_status:
                continue
            parsed = parse_status(raw_status, status_map, "color_trim_compatibility")
            add_status_example(status_examples, status_rejections, section.source_sheet, "color_trim_compatibility", parsed)
            if not parsed["raw_status"]:
                continue
            rows_by_name["color_trim_compatibility_rows"].append(
                {
                    "source_sheet": section.source_sheet,
                    "source_row": row_index,
                    "source_cell_range": cell_ref(row_index, column_index),
                    "guide_family": classification["guide_family"],
                    "model_key": "",
                    "scope_type": "model_global",
                    "exterior_color_name_raw": exterior_name_split["raw_value"],
                    "exterior_color_name": exterior_name_split["parsed_value"],
                    "exterior_color_rpo_raw": exterior_rpo_split["raw_value"],
                    "exterior_color_rpo": exterior_rpo_split["parsed_value"],
                    "touch_up_paint_code": value_at(values, 3),
                    "interior_rpo_raw": header["raw"],
                    "interior_rpo": header["parsed"],
                    "raw_status": parsed["raw_status"],
                    "status_symbol": parsed["status_symbol"],
                    "footnote_refs": parsed["footnote_refs"],
                    "canonical_status": parsed["canonical_status"],
                    "confidence": parsed["confidence"] if parsed["canonical_status"] else "needs_review",
                    "source_detail_raw": raw_row_values(values),
                }
            )


def extract_color_trim_disclosure(
    rows_by_name: dict[str, list[dict[str, Any]]],
    classification: dict[str, str],
    section: Section,
    row_lookup: dict[int, list[str]],
    patterns: list[dict[str, Any]],
) -> None:
    values = row_lookup.get(section.data_start_row, [])
    raw_text = " ".join(value for value in values if value)
    candidates = rule_phrase_candidates(
        source_sheet=section.source_sheet,
        source_row=section.data_start_row,
        model_key="",
        source_field="color_trim_disclosure",
        raw_text=raw_text,
        patterns=patterns,
    )
    if candidates:
        for candidate in candidates:
            rows_by_name["color_trim_disclosures"].append(
                {
                    "source_sheet": section.source_sheet,
                    "source_row": section.data_start_row,
                    "source_cell_range": f"A{section.data_start_row}",
                    "guide_family": classification["guide_family"],
                    "model_key": "",
                    "raw_text": raw_text,
                    "extracted_rpos": candidate["extracted_rpos"],
                    "phrase_type": candidate["phrase_type"],
                    "applies_to_section_role": "color_trim",
                    "confidence": candidate["confidence"],
                    "review_status": "needs_review",
                }
            )
    elif raw_text:
        rows_by_name["color_trim_disclosures"].append(
            {
                "source_sheet": section.source_sheet,
                "source_row": section.data_start_row,
                "source_cell_range": f"A{section.data_start_row}",
                "guide_family": classification["guide_family"],
                "model_key": "",
                "raw_text": raw_text,
                "extracted_rpos": "|".join(sorted(set(RPO_RE.findall(raw_text)))),
                "phrase_type": "",
                "applies_to_section_role": "color_trim",
                "confidence": "medium",
                "review_status": "needs_review",
            }
        )


def extract_price_section(
    rows_by_name: dict[str, list[dict[str, Any]]],
    classification: dict[str, str],
    section: Section,
    row_lookup: dict[int, list[str]],
) -> None:
    for row_index in range(section.data_start_row, section.data_end_row + 1):
        values = row_lookup.get(row_index, [])
        if is_blank(values):
            rows_by_name["ignored_rows"].append({"source_sheet": section.source_sheet, "source_row": row_index, "raw_values": raw_row_values(values), "reason": "blank_row"})
            continue
        rows_by_name["price_rows"].append(
            {
                "source_sheet": section.source_sheet,
                "source_row": row_index,
                "guide_family": classification["guide_family"],
                "model_key": "",
                "price_block_label": infer_price_block(values),
                "raw_values": raw_row_values(values),
                "notes": "Price Schedule extraction in this pass is staging evidence only. Do not classify final canonical price rules, price books, package pricing, or included-zero pricing yet.",
            }
        )


def infer_price_block(values: list[str]) -> str:
    joined = normalized_text(" ".join(value for value in values if value))
    if "base model prices" in joined:
        return "base_model_prices"
    if "model" in joined and "price" in joined:
        return "price_header"
    return "raw_price_schedule_row"


def write_staging_outputs(out_dir: Path, staging: dict[str, list[dict[str, Any]] | dict[str, Any]]) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    for name, headers in STAGING_OUTPUTS.items():
        rows = staging[name]
        assert isinstance(rows, list)
        write_csv(out_dir / f"staging_{name}.csv", headers, rows)
    report = staging["report"]
    assert isinstance(report, dict)
    write_json(out_dir / "import_report.json", report)
