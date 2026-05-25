#!/usr/bin/env python3
"""Apply approved future_model_source_review rows to normalized future source sheets."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SPECS,
    OPTION_SOURCE_HEADERS,
    OVS_SOURCE_HEADERS,
    build_future_source_population_plan,
)
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely, write_sheet  # noqa: E402

FUTURE_MODEL_KEYS = tuple(FUTURE_MODEL_SPECS)


def _strip_rows_from_plan(plan: dict[str, Any]) -> dict[str, Any]:
    """Return a JSON report without embedding every row to be written."""

    report = {
        "selected_model_keys": plan["selected_model_keys"],
        "error_count": plan["error_count"],
        "models": {},
    }
    for model_key, model_plan in plan["models"].items():
        report["models"][model_key] = {
            key: value
            for key, value in model_plan.items()
            if key not in {"option_rows", "ovs_rows"}
        }
    return report


def _sheet_headers(ws) -> list[str]:
    return [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]


def _assert_future_metadata_inactive(wb) -> None:
    future_models = set(FUTURE_MODEL_KEYS)
    checks = {
        "model_master": ("model_key", ["active"]),
        "model_workbook_sources": ("model_key", ["active"]),
        "model_registry_promotion": ("model_key", ["promoted_to_runtime", "active"]),
    }
    for sheet_name, (model_field, inactive_fields) in checks.items():
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Missing metadata sheet {sheet_name}")
        ws = wb[sheet_name]
        headers = _sheet_headers(ws)
        index = {header: offset + 1 for offset, header in enumerate(headers)}
        if model_field not in index:
            raise RuntimeError(f"{sheet_name} missing {model_field}")
        for field in inactive_fields:
            if field not in index:
                raise RuntimeError(f"{sheet_name} missing {field}")
        for row_number in range(2, ws.max_row + 1):
            model_key = clean(ws.cell(row_number, index[model_field]).value)
            if model_key not in future_models:
                continue
            for field in inactive_fields:
                value = clean(ws.cell(row_number, index[field]).value).casefold()
                if value in {"true", "1", "yes", "y", "active"}:
                    raise RuntimeError(f"{sheet_name} row {row_number} unexpectedly has active {field}={value}")


def verify_saved_workbook(path: Path, plan: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        verification: dict[str, Any] = {"models": {}}
        for model_key, model_plan in plan["models"].items():
            option_sheet = model_plan["target_option_sheet"]
            ovs_sheet = model_plan["target_ovs_sheet"]
            if option_sheet not in wb.sheetnames:
                raise RuntimeError(f"Missing target option sheet {option_sheet}")
            if ovs_sheet not in wb.sheetnames:
                raise RuntimeError(f"Missing target OVS sheet {ovs_sheet}")
            option_ws = wb[option_sheet]
            ovs_ws = wb[ovs_sheet]
            option_headers = _sheet_headers(option_ws)
            ovs_headers = _sheet_headers(ovs_ws)
            if option_headers != list(OPTION_SOURCE_HEADERS):
                raise RuntimeError(f"{option_sheet} header drift: {option_headers}")
            if ovs_headers != list(OVS_SOURCE_HEADERS):
                raise RuntimeError(f"{ovs_sheet} header drift: {ovs_headers}")
            option_rows = max(option_ws.max_row - 1, 0)
            ovs_rows = max(ovs_ws.max_row - 1, 0)
            expected_options = model_plan["eligible_option_count"]
            expected_ovs = model_plan["emitted_ovs_count"]
            if option_rows != expected_options:
                raise RuntimeError(f"{option_sheet} expected {expected_options} rows, found {option_rows}")
            if ovs_rows != expected_ovs:
                raise RuntimeError(f"{ovs_sheet} expected {expected_ovs} rows, found {ovs_rows}")
            verification["models"][model_key] = {
                "option_sheet": option_sheet,
                "option_rows": option_rows,
                "ovs_sheet": ovs_sheet,
                "ovs_rows": ovs_rows,
            }
        _assert_future_metadata_inactive(wb)
        verification["future_metadata_inactive"] = True
        return verification
    finally:
        wb.close()


def apply_plan_to_workbook(wb, plan: dict[str, Any]) -> None:
    for model_plan in plan["models"].values():
        write_sheet(wb, model_plan["target_option_sheet"], list(OPTION_SOURCE_HEADERS), model_plan["option_rows"])
        write_sheet(wb, model_plan["target_ovs_sheet"], list(OVS_SOURCE_HEADERS), model_plan["ovs_rows"])


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model-key", choices=("all", *FUTURE_MODEL_KEYS), default="all")
    parser.add_argument("--dry-run", action="store_true", help="Report rows that would be written without saving the workbook.")
    args = parser.parse_args(argv)

    if not args.dry_run and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    backup_path = None
    try:
        selected = list(FUTURE_MODEL_KEYS) if args.model_key == "all" else [args.model_key]
        plan = build_future_source_population_plan(wb, selected)
        report = _strip_rows_from_plan(plan)
        if plan["error_count"]:
            print(json.dumps({"status": "blocked", **report}, indent=2))
            return 1
        if args.dry_run:
            print(json.dumps({"status": "dry_run", **report}, indent=2))
            return 0
        apply_plan_to_workbook(wb, plan)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_workbook(WORKBOOK_PATH, plan)
    print(json.dumps({
        "status": "written",
        "workbook": str(WORKBOOK_PATH),
        "backup": str(backup_path),
        **_strip_rows_from_plan(plan),
        "verification": verification,
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
