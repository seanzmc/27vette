"""Future-model compatibility source preview and workbook write helpers.

Grand Sport compatibility rows are treated as the workbook-owned template. This
module rebases those rows to Z06/ZR1/ZR1X by matching unique active RPOs between
Grand Sport options and each future model's option sheet. Dry-run preview remains
non-mutating. Explicit write-mode callers can persist only the successfully
mapped rows into the future-model workbook source sheets.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from corvette_form_generator.future_model_ingest import FUTURE_MODEL_SPECS
from corvette_form_generator.workbook import clean, rows_from_sheet

RULE_MAPPING_HEADERS = (
    "rule_id",
    "source_id",
    "rule_type",
    "target_id",
    "target_type",
    "original_detail_raw",
    "review_flag",
    "source_type",
    "target_selection_mode",
    "source_selection_mode",
    "target_section",
    "source_section",
    "generation_action",
    "body_style_scope",
    "runtime_action",
    "disabled_reason",
    "normalization_status",
    "normalization_reason",
    "replacement_group_id",
    "replacement_rule_id",
)
RULE_GROUP_HEADERS = (
    "group_id",
    "group_type",
    "source_id",
    "body_style_scope",
    "trim_level_scope",
    "variant_scope",
    "disabled_reason",
    "active",
    "notes",
)
RULE_GROUP_MEMBER_HEADERS = ("group_id", "target_id", "display_order", "active")
EXCLUSIVE_GROUP_HEADERS = ("group_id", "selection_mode", "active", "notes")
EXCLUSIVE_MEMBER_HEADERS = ("group_id", "option_id", "display_order", "active")
COMPATIBILITY_HEADERS = {
    "rule_mapping": RULE_MAPPING_HEADERS,
    "rule_groups": RULE_GROUP_HEADERS,
    "rule_group_members": RULE_GROUP_MEMBER_HEADERS,
    "exclusive_groups": EXCLUSIVE_GROUP_HEADERS,
    "exclusive_members": EXCLUSIVE_MEMBER_HEADERS,
}

SOURCE_SHEETS = {
    "options": "grandSport_options",
    "rule_mapping": "grandSport_rule_mapping",
    "rule_groups": "grandSport_rule_groups",
    "rule_group_members": "grandSport_rule_group_members",
    "exclusive_groups": "grandSport_exclusive_groups",
    "exclusive_members": "grandSport_exclusive_members",
}

TARGET_COMPATIBILITY_SHEETS = {
    "rule_mapping": "{model_key}_rule_mapping",
    "rule_groups": "{model_key}_rule_groups",
    "rule_group_members": "{model_key}_rule_group_members",
    "exclusive_groups": "{model_key}_exclusive_groups",
    "exclusive_members": "{model_key}_exclusive_members",
}

_TRUE_VALUES = {"true", "1", "yes", "y", "active"}
_FALSE_VALUES = {"false", "0", "no", "n", "inactive"}
_BOOL_FIELDS = {"active", "review_flag", "selectable"}
OPTION_ID_FIELDS = {"source_id", "target_id", "option_id"}


@dataclass(frozen=True)
class FutureCompatibilitySpec:
    model_key: str
    target_option_sheet: str
    target_sheets: dict[str, str]


FUTURE_COMPATIBILITY_SPECS: dict[str, FutureCompatibilitySpec] = {
    model_key: FutureCompatibilitySpec(
        model_key=model_key,
        target_option_sheet=spec.target_option_sheet,
        target_sheets={key: pattern.format(model_key=model_key) for key, pattern in TARGET_COMPATIBILITY_SHEETS.items()},
    )
    for model_key, spec in FUTURE_MODEL_SPECS.items()
}


def active_bool(value: Any, *, default: bool = True) -> bool:
    text = clean(value).casefold()
    if not text:
        return default
    if text in _TRUE_VALUES:
        return True
    if text in _FALSE_VALUES:
        return False
    return default


def _safe_rows_from_sheet(wb, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    return rows_from_sheet(wb, sheet_name)


def _active_option_rows(wb, sheet_name: str) -> list[dict[str, str]]:
    return [row for row in _safe_rows_from_sheet(wb, sheet_name) if active_bool(row.get("active"))]


def _option_indexes(rows: list[dict[str, str]]) -> dict[str, Any]:
    by_id: dict[str, dict[str, str]] = {}
    rpo_counts: Counter[str] = Counter()
    by_rpo: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        option_id = clean(row.get("option_id"))
        rpo = clean(row.get("rpo"))
        if option_id:
            by_id[option_id] = row
        if rpo:
            rpo_counts[rpo] += 1
            by_rpo[rpo].append(row)
    unique_by_rpo = {rpo: matches[0] for rpo, matches in by_rpo.items() if rpo_counts[rpo] == 1}
    duplicate_rpos = {rpo for rpo, count in rpo_counts.items() if count > 1}
    return {"by_id": by_id, "unique_by_rpo": unique_by_rpo, "duplicate_rpos": duplicate_rpos}


def build_option_rebase_map(wb, target_option_sheet: str) -> dict[str, Any]:
    """Map Grand Sport option_id -> future option_id by unique active RPO."""

    source_index = _option_indexes(_active_option_rows(wb, SOURCE_SHEETS["options"]))
    target_index = _option_indexes(_active_option_rows(wb, target_option_sheet))
    mapping: dict[str, str] = {}
    unresolved: dict[str, str] = {}
    source_rpos: dict[str, str] = {}

    for source_option_id, source_row in source_index["by_id"].items():
        rpo = clean(source_row.get("rpo"))
        source_rpos[source_option_id] = rpo
        if not rpo:
            unresolved[source_option_id] = "source_missing_rpo"
            continue
        if rpo in source_index["duplicate_rpos"]:
            unresolved[source_option_id] = "source_duplicate_active_rpo"
            continue
        if rpo in target_index["duplicate_rpos"]:
            unresolved[source_option_id] = "target_duplicate_active_rpo"
            continue
        target_row = target_index["unique_by_rpo"].get(rpo)
        if not target_row:
            unresolved[source_option_id] = "target_rpo_not_found"
            continue
        mapping[source_option_id] = clean(target_row.get("option_id"))

    return {
        "mapping": mapping,
        "unresolved": unresolved,
        "source_rpos": source_rpos,
        "source_active_option_count": len(source_index["by_id"]),
        "target_active_option_count": len(target_index["by_id"]),
        "mapped_option_count": len(mapping),
        "source_duplicate_rpos": sorted(source_index["duplicate_rpos"]),
        "target_duplicate_rpos": sorted(target_index["duplicate_rpos"]),
    }


def _rebase_group_id(group_id: str, model_key: str) -> str:
    text = clean(group_id)
    if text.startswith("gs_"):
        return f"{model_key}_{text[3:]}"
    if text.startswith("grandSport_"):
        return f"{model_key}_{text[len('grandSport_'):]}".replace(" ", "")
    return f"{model_key}_{text}"


def _rebase_rule_id(rule_id: str, model_key: str) -> str:
    text = clean(rule_id)
    if text.startswith("gs_"):
        return f"{model_key}_{text[3:]}"
    return f"{model_key}_{text}"


def _rebase_option_id(option_id: str, rebase_map: dict[str, Any]) -> tuple[str, str | None]:
    text = clean(option_id)
    if not text:
        return "", "missing_option_id"
    mapped = rebase_map["mapping"].get(text)
    if mapped:
        return mapped, None
    return "", rebase_map["unresolved"].get(text, "source_option_id_not_found")


def _skip(counter: Counter[str], reason: str) -> None:
    counter[reason] += 1


def _row_status_is_active(row: dict[str, Any]) -> bool:
    status = clean(row.get("normalization_status")).casefold()
    if status and status != "active":
        return False
    if clean(row.get("disabled_reason")):
        return False
    return True


def _copy_headers(row: dict[str, Any], headers: Iterable[str]) -> dict[str, str]:
    return {header: clean(row.get(header)) for header in headers}


def _build_rule_mapping_rows(wb, model_key: str, rebase_map: dict[str, Any]) -> dict[str, Any]:
    proposed: list[dict[str, str]] = []
    skipped: Counter[str] = Counter()
    rows = _safe_rows_from_sheet(wb, SOURCE_SHEETS["rule_mapping"])
    for row in rows:
        if clean(row.get("source_type")).casefold() == "interior":
            _skip(skipped, "deferred_source_type_interior")
            continue
        if not _row_status_is_active(row):
            _skip(skipped, "inactive_or_replaced_source_rule")
            continue
        source_id, source_reason = _rebase_option_id(clean(row.get("source_id")), rebase_map)
        if source_reason:
            _skip(skipped, f"source_id:{source_reason}")
            continue
        target_id, target_reason = _rebase_option_id(clean(row.get("target_id")), rebase_map)
        if target_reason:
            _skip(skipped, f"target_id:{target_reason}")
            continue
        rebased = _copy_headers(row, RULE_MAPPING_HEADERS)
        rebased["rule_id"] = _rebase_rule_id(clean(row.get("rule_id")), model_key)
        rebased["source_id"] = source_id
        rebased["target_id"] = target_id
        if clean(row.get("replacement_group_id")):
            rebased["replacement_group_id"] = _rebase_group_id(clean(row.get("replacement_group_id")), model_key)
        if clean(row.get("replacement_rule_id")):
            rebased["replacement_rule_id"] = _rebase_rule_id(clean(row.get("replacement_rule_id")), model_key)
        proposed.append(rebased)
    return {"rows": proposed, "skipped": dict(sorted(skipped.items())), "source_row_count": len(rows)}


def _build_rule_group_rows(wb, model_key: str, rebase_map: dict[str, Any]) -> dict[str, Any]:
    groups = _safe_rows_from_sheet(wb, SOURCE_SHEETS["rule_groups"])
    members = _safe_rows_from_sheet(wb, SOURCE_SHEETS["rule_group_members"])
    proposed_groups: list[dict[str, str]] = []
    proposed_members: list[dict[str, str]] = []
    skipped_groups: Counter[str] = Counter()
    skipped_members: Counter[str] = Counter()
    members_by_group: dict[str, list[dict[str, str]]] = defaultdict(list)
    for member in members:
        members_by_group[clean(member.get("group_id"))].append(member)

    for group in groups:
        group_id = clean(group.get("group_id"))
        if not active_bool(group.get("active")):
            _skip(skipped_groups, "inactive_group")
            continue
        source_id, source_reason = _rebase_option_id(clean(group.get("source_id")), rebase_map)
        if source_reason:
            _skip(skipped_groups, f"source_id:{source_reason}")
            continue
        new_group_id = _rebase_group_id(group_id, model_key)
        resolved_members: list[dict[str, str]] = []
        for member in members_by_group.get(group_id, []):
            if not active_bool(member.get("active")):
                _skip(skipped_members, "inactive_member")
                continue
            target_id, target_reason = _rebase_option_id(clean(member.get("target_id")), rebase_map)
            if target_reason:
                _skip(skipped_members, f"target_id:{target_reason}")
                continue
            new_member = _copy_headers(member, RULE_GROUP_MEMBER_HEADERS)
            new_member["group_id"] = new_group_id
            new_member["target_id"] = target_id
            resolved_members.append(new_member)
        if not resolved_members:
            _skip(skipped_groups, "no_resolved_members")
            continue
        new_group = _copy_headers(group, RULE_GROUP_HEADERS)
        new_group["group_id"] = new_group_id
        new_group["source_id"] = source_id
        proposed_groups.append(new_group)
        proposed_members.extend(resolved_members)

    return {
        "groups": proposed_groups,
        "members": proposed_members,
        "skipped_groups": dict(sorted(skipped_groups.items())),
        "skipped_members": dict(sorted(skipped_members.items())),
        "source_group_count": len(groups),
        "source_member_count": len(members),
    }


def _build_exclusive_rows(wb, model_key: str, rebase_map: dict[str, Any]) -> dict[str, Any]:
    groups = _safe_rows_from_sheet(wb, SOURCE_SHEETS["exclusive_groups"])
    members = _safe_rows_from_sheet(wb, SOURCE_SHEETS["exclusive_members"])
    proposed_groups: list[dict[str, str]] = []
    proposed_members: list[dict[str, str]] = []
    skipped_groups: Counter[str] = Counter()
    skipped_members: Counter[str] = Counter()
    members_by_group: dict[str, list[dict[str, str]]] = defaultdict(list)
    for member in members:
        members_by_group[clean(member.get("group_id"))].append(member)

    for group in groups:
        group_id = clean(group.get("group_id"))
        if not active_bool(group.get("active")):
            _skip(skipped_groups, "inactive_group")
            continue
        new_group_id = _rebase_group_id(group_id, model_key)
        resolved_members: list[dict[str, str]] = []
        for member in members_by_group.get(group_id, []):
            if not active_bool(member.get("active")):
                _skip(skipped_members, "inactive_member")
                continue
            option_id, option_reason = _rebase_option_id(clean(member.get("option_id")), rebase_map)
            if option_reason:
                _skip(skipped_members, f"option_id:{option_reason}")
                continue
            new_member = _copy_headers(member, EXCLUSIVE_MEMBER_HEADERS)
            new_member["group_id"] = new_group_id
            new_member["option_id"] = option_id
            resolved_members.append(new_member)
        if len(resolved_members) < 2:
            _skip(skipped_groups, "fewer_than_two_resolved_members")
            continue
        new_group = _copy_headers(group, EXCLUSIVE_GROUP_HEADERS)
        new_group["group_id"] = new_group_id
        proposed_groups.append(new_group)
        proposed_members.extend(resolved_members)

    return {
        "groups": proposed_groups,
        "members": proposed_members,
        "skipped_groups": dict(sorted(skipped_groups.items())),
        "skipped_members": dict(sorted(skipped_members.items())),
        "source_group_count": len(groups),
        "source_member_count": len(members),
    }


def _current_row_count(wb, sheet_name: str) -> int:
    return len(_safe_rows_from_sheet(wb, sheet_name))


def build_compatibility_preview_for_model(wb, spec: FutureCompatibilitySpec) -> dict[str, Any]:
    rebase_map = build_option_rebase_map(wb, spec.target_option_sheet)
    rules = _build_rule_mapping_rows(wb, spec.model_key, rebase_map)
    rule_groups = _build_rule_group_rows(wb, spec.model_key, rebase_map)
    exclusives = _build_exclusive_rows(wb, spec.model_key, rebase_map)
    target_sheets = spec.target_sheets

    proposed_counts = {
        "rule_mapping": len(rules["rows"]),
        "rule_groups": len(rule_groups["groups"]),
        "rule_group_members": len(rule_groups["members"]),
        "exclusive_groups": len(exclusives["groups"]),
        "exclusive_members": len(exclusives["members"]),
    }
    current_counts = {key: _current_row_count(wb, sheet_name) for key, sheet_name in target_sheets.items()}
    return {
        "model_key": spec.model_key,
        "target_option_sheet": spec.target_option_sheet,
        "target_sheets": target_sheets,
        "current_target_row_counts": current_counts,
        "proposed_row_counts": proposed_counts,
        "rebase": {key: value for key, value in rebase_map.items() if key != "mapping"},
        "skipped": {
            "rule_mapping": rules["skipped"],
            "rule_groups": rule_groups["skipped_groups"],
            "rule_group_members": rule_groups["skipped_members"],
            "exclusive_groups": exclusives["skipped_groups"],
            "exclusive_members": exclusives["skipped_members"],
        },
        "proposed_rows": {
            "rule_mapping": rules["rows"],
            "rule_groups": rule_groups["groups"],
            "rule_group_members": rule_groups["members"],
            "exclusive_groups": exclusives["groups"],
            "exclusive_members": exclusives["members"],
        },
    }


def _selected_model_keys(model_keys: Iterable[str] | None = None) -> list[str]:
    requested = [clean(model_key) for model_key in (model_keys or []) if clean(model_key)]
    if not requested or requested == ["all"] or "all" in requested:
        return list(FUTURE_COMPATIBILITY_SPECS)
    unknown = [model_key for model_key in requested if model_key not in FUTURE_COMPATIBILITY_SPECS]
    if unknown:
        raise ValueError(f"Unknown future model key(s): {', '.join(unknown)}")
    return requested


def build_future_compatibility_preview(
    wb,
    model_keys: Iterable[str] | None = None,
    *,
    generated_at: str | None = None,
    include_rows: bool = True,
) -> dict[str, Any]:
    """Build a non-mutating compatibility-source rebase preview."""

    selected = _selected_model_keys(model_keys)
    models = {
        model_key: build_compatibility_preview_for_model(wb, FUTURE_COMPATIBILITY_SPECS[model_key])
        for model_key in selected
    }
    if not include_rows:
        for model in models.values():
            model.pop("proposed_rows", None)
    return {
        "status": "dry_run",
        "generated_at": generated_at or datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_sheets": SOURCE_SHEETS,
        "selected_model_keys": selected,
        "models": models,
        "notes": [
            "Dry-run only: stingray_master.xlsx was read but not saved.",
            "Grand Sport compatibility source rows are rebased by unique active RPO matches, not raw option_id equality.",
            "source_type=interior rule_mapping rows are deferred and not proposed.",
            "Exclusive groups are proposed only when at least two members survive rebasing.",
        ],
    }


def _sheet_headers(ws: Any) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def _header_index(wb: Any, sheet_name: str, required_headers: Iterable[str]) -> dict[str, int]:
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Missing required target sheet {sheet_name}")
    ws = wb[sheet_name]
    headers = _sheet_headers(ws)
    index = {header: offset + 1 for offset, header in enumerate(headers) if header}
    missing = [header for header in required_headers if header not in index]
    if missing:
        raise RuntimeError(f"{sheet_name} missing required header(s): {', '.join(missing)}")
    return index


def _excel_value(header: str, value: Any) -> Any:
    text = clean(value)
    if header in _BOOL_FIELDS:
        if text.casefold() in _TRUE_VALUES:
            return True
        if text.casefold() in _FALSE_VALUES:
            return False
    return value


def _replace_target_sheet_rows(wb: Any, sheet_name: str, headers: tuple[str, ...], rows: list[dict[str, Any]]) -> dict[str, int]:
    _header_index(wb, sheet_name, headers)
    ws = wb[sheet_name]
    deleted = max(ws.max_row - 1, 0)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append([_excel_value(header, row.get(header, "")) for header in headers])
    return {"deleted": deleted, "inserted": len(rows)}


def apply_future_compatibility_to_workbook(wb: Any, preview: dict[str, Any]) -> dict[str, Any]:
    """Write proposed compatibility rows from a preview into selected target sheets."""

    selected = _selected_model_keys(preview.get("selected_model_keys"))
    report: dict[str, Any] = {"status": "applied", "selected_model_keys": selected, "models": {}, "would_mutate_generated_runtime_data": False}
    for model_key in selected:
        model = preview["models"][model_key]
        proposed_rows = model.get("proposed_rows")
        if proposed_rows is None:
            raise RuntimeError(f"Preview for {model_key} is missing proposed_rows. Rebuild preview with include_rows=True for write mode.")
        sheet_reports: dict[str, dict[str, int]] = {}
        for area, sheet_name in model["target_sheets"].items():
            headers = COMPATIBILITY_HEADERS[area]
            sheet_reports[area] = _replace_target_sheet_rows(wb, sheet_name, headers, proposed_rows[area])
        report["models"][model_key] = {
            "target_sheets": model["target_sheets"],
            "written_row_counts": {area: data["inserted"] for area, data in sheet_reports.items()},
            "sheet_reports": sheet_reports,
            "skipped": model.get("skipped", {}),
        }
    return report


def assert_future_metadata_inactive(wb: Any) -> None:
    future_models = set(FUTURE_COMPATIBILITY_SPECS)
    checks = {
        "model_master": ("model_key", ["active"]),
        "model_workbook_sources": ("model_key", ["active"]),
        "model_registry_promotion": ("model_key", ["promoted_to_runtime", "active"]),
    }
    for sheet_name, (model_field, inactive_fields) in checks.items():
        if sheet_name not in wb.sheetnames:
            continue
        index = _header_index(wb, sheet_name, (model_field, *inactive_fields))
        ws = wb[sheet_name]
        for row_number in range(2, ws.max_row + 1):
            model_key = clean(ws.cell(row_number, index[model_field]).value)
            if model_key not in future_models:
                continue
            for field in inactive_fields:
                value = clean(ws.cell(row_number, index[field]).value).casefold()
                if value in _TRUE_VALUES:
                    raise RuntimeError(f"{sheet_name} row {row_number} unexpectedly has active {field}={value}")


def verify_saved_compatibility_workbook(path: Path, preview: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        selected = _selected_model_keys(preview.get("selected_model_keys"))
        verification: dict[str, Any] = {"models": {}}
        for model_key in selected:
            model = preview["models"][model_key]
            model_verification: dict[str, Any] = {"target_sheets": {}, "written_row_counts": {}}
            for area, sheet_name in model["target_sheets"].items():
                expected = model["proposed_row_counts"][area]
                actual = len(rows_from_sheet(wb, sheet_name))
                if actual != expected:
                    raise RuntimeError(f"{sheet_name} expected {expected} persisted rows, found {actual}")
                model_verification["target_sheets"][area] = sheet_name
                model_verification["written_row_counts"][area] = actual
            verification["models"][model_key] = model_verification
        assert_future_metadata_inactive(wb)
        verification["future_metadata_inactive"] = True
        return verification
    finally:
        wb.close()


def render_compatibility_preview_markdown(preview: dict[str, Any]) -> str:
    lines = ["# Future Model Compatibility Preview", ""]
    lines.append(f"Generated: {preview['generated_at']}")
    lines.append("")
    if preview.get("status") == "written":
        lines.append("Workbook source sheets were written. Generated/runtime app data was not written.")
    else:
        lines.append("Dry-run only: workbook and runtime app data were not written.")
    lines.append("")
    for model_key, model in preview["models"].items():
        lines.append(f"## {model_key}")
        lines.append(f"- Target options: `{model['target_option_sheet']}`")
        lines.append(f"- Active option RPO matches: {model['rebase']['mapped_option_count']} / {model['rebase']['source_active_option_count']} Grand Sport active options")
        lines.append("- Proposed row counts:")
        for key, count in model["proposed_row_counts"].items():
            current = model["current_target_row_counts"].get(key, 0)
            sheet = model["target_sheets"].get(key, key)
            lines.append(f"  - `{sheet}`: {count} proposed (currently {current})")
        lines.append("- Skipped/unresolved reasons:")
        any_skips = False
        for area, reasons in model["skipped"].items():
            for reason, count in reasons.items():
                any_skips = True
                lines.append(f"  - {area} {reason}: {count}")
        if not any_skips:
            lines.append("  - none")
        lines.append("")
    lines.append("Notes:")
    for note in preview.get("notes", []):
        lines.append(f"- {note}")
    lines.append("")
    return "\n".join(lines)
