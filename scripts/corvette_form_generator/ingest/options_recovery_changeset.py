#!/usr/bin/env python3
"""Emit the reviewed options-recovery projection as ``workbook-changeset-1``.

This is deliberately a read-only projection: it binds the reviewed reports
to the current workbook and returns a strict ChangeSet, but never previews,
applies, or saves the workbook.
"""

from __future__ import annotations

import argparse
import json
import re
from copy import deepcopy
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from corvette_form_generator.ingest.options_recovery_projection import (
    CHECKPOINT_1_APPROVAL_SCHEMA_VERSION,
    CHECKPOINT_1_DECISIONS_SCHEMA_VERSION,
    CHECKPOINT_1_EXCEPTION_SCHEMA_VERSION,
    CHECKPOINT_1_PENDING_SCHEMA_VERSION,
    OPTION_VIEW_FIELDS,
    SCHEMA_VERSION as REPORT_SCHEMA_VERSION,
    TARGET_MODELS,
    _canonical_sha,
    _sha256,
    _validate_checkpoint_report,
    _validate_fingerprinted_artifact,
)
from corvette_form_generator.workbook import clean, workbook_truthy
from corvette_form_generator.workbook_domain.changeset import (
    changeset_fingerprint,
    parse_changeset,
)
from corvette_form_generator.workbook_domain.registry import (
    EDITOR_SHEET_META,
    GLOBAL_SHEET_FAMILIES,
    SOURCE_ROLE_FAMILIES,
)


_TOKEN_CACHE: dict[str, re.Pattern[str]] = {}


def _json(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise ValueError(f"Missing required options recovery artifact: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Options recovery artifact must be a JSON object: {path}")
    return payload


def _unsigned_fingerprint(payload: Mapping[str, Any], field: str, label: str) -> str:
    return _validate_fingerprinted_artifact(payload, field, label)


def _headers_and_rows(workbook: Any, sheet: str) -> tuple[list[str], dict[int, dict[str, Any]]]:
    if sheet not in workbook.sheetnames:
        raise ValueError(f"Recovery artifact references missing worksheet: {sheet}")
    ws = workbook[sheet]
    values = ws.iter_rows(values_only=True)
    try:
        headers = [clean(value) for value in next(values)]
    except StopIteration:
        return [], {}
    if not any(headers):
        raise ValueError(f"Recovery artifact references headerless worksheet: {sheet}")
    rows: dict[int, dict[str, Any]] = {}
    for row_number, raw_row in enumerate(values, start=2):
        row = {
            header: raw_row[index] if index < len(raw_row) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value is not None for value in row.values()):
            rows[row_number] = row
    return headers, rows


def _family_map(workbook: Any) -> tuple[dict[str, str], dict[str, str]]:
    """Discover all model families, including inactive/unpromoted sources."""

    if "model_workbook_sources" not in workbook.sheetnames:
        raise ValueError("Workbook is missing model_workbook_sources")
    _, source_rows = _headers_and_rows(workbook, "model_workbook_sources")
    source_sheets: dict[str, str] = {}
    families = dict(GLOBAL_SHEET_FAMILIES)
    for row in source_rows.values():
        model = clean(row.get("model_key")).lower()
        role = clean(row.get("source_role"))
        sheet = clean(row.get("sheet_name"))
        if not model or not role or not sheet:
            continue
        family = SOURCE_ROLE_FAMILIES.get(role)
        if family:
            families[sheet] = family
            if role == "source_option_sheet":
                source_sheets[model] = sheet
    missing = [model for model in TARGET_MODELS if not source_sheets.get(model)]
    if missing:
        raise ValueError(f"Missing target source_option_sheet mappings: {', '.join(missing)}")
    return source_sheets, families


def _semantic_equal(actual: Any, projected: Any, family: str, column: str) -> bool:
    if actual is None or clean(actual) == "":
        return projected is None or clean(projected) == ""
    if projected is None or clean(projected) == "":
        return False
    kind = (EDITOR_SHEET_META.get(family, {}).get("types") or {}).get(column)
    if kind == "bool":
        return workbook_truthy(actual) == workbook_truthy(projected)
    if kind == "int":
        try:
            return float(str(actual).strip()) == float(str(projected).strip())
        except (TypeError, ValueError):
            return False
    return actual == projected


def _row_key(row: Mapping[str, Any], family: str) -> dict[str, Any]:
    keys = EDITOR_SHEET_META[family]["key"]
    key = {column: row.get(column) for column in keys}
    if any(value is None or clean(value) == "" for value in key.values()):
        raise ValueError(f"Recovery cascade row has a blank {family} key: {key}")
    return key


def _full_delete_fields(row: Mapping[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        column: {"before": value, "after": None}
        for column, value in row.items()
        if value is not None and clean(value) != ""
    }


def _add_fields(row: Mapping[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        column: {"before": None, "after": value}
        for column, value in row.items()
        if value is not None and clean(value) != ""
    }


def _token_pattern(old: str) -> re.Pattern[str]:
    return _TOKEN_CACHE.setdefault(
        old,
        re.compile(rf"(?<![A-Za-z0-9_]){re.escape(old)}(?![A-Za-z0-9_])"),
    )


def _replacement(value: Any, old: str, new: str, match_type: str) -> Any:
    text = clean(value)
    if match_type == "exact":
        if text != old:
            raise ValueError(f"Expected exact id reference {old!r}, found {value!r}")
        return new
    replaced, count = _token_pattern(old).subn(new, text)
    if not count:
        raise ValueError(f"Expected embedded id reference {old!r}, found {value!r}")
    return replaced


def _artifact_bindings(
    workbook_path: Path,
    report_dir: Path,
    workbook: Any,
) -> tuple[dict[str, dict[str, Any]], dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any], dict[str, str], dict[str, str]]:
    workbook_sha = _sha256(workbook_path)
    source_sheets, families = _family_map(workbook)
    reports: dict[str, dict[str, Any]] = {}
    for model in TARGET_MODELS:
        path = report_dir / f"{model}-recovery-projection.json"
        report = _json(path)
        _validate_checkpoint_report(report, model, path)
        if report.get("schemaVersion") != REPORT_SCHEMA_VERSION:
            raise ValueError(f"Unsupported recovery report schema: {path}")
        if clean((report.get("sources") or {}).get("workbookSha256")) != workbook_sha:
            raise ValueError(f"Current workbook SHA does not match recovery report: {path}")
        if clean((report.get("sources") or {}).get("targetSheet")) != source_sheets[model]:
            raise ValueError(f"Recovery report target sheet binding mismatch for {model}")
        reports[model] = report

    approval = _json(report_dir / "checkpoint-1-bulk-approval.json")
    packet = _json(report_dir / "checkpoint-1-exception-review.json")
    decisions = _json(report_dir / "checkpoint-1-exception-decisions.json")
    pending = _json(report_dir / "checkpoint-1-pending-review.json")
    if approval.get("schemaVersion") != CHECKPOINT_1_APPROVAL_SCHEMA_VERSION:
        raise ValueError("Unsupported Checkpoint 1 bulk approval schema")
    if packet.get("schemaVersion") != CHECKPOINT_1_EXCEPTION_SCHEMA_VERSION:
        raise ValueError("Unsupported Checkpoint 1 exception packet schema")
    if decisions.get("schemaVersion") != CHECKPOINT_1_DECISIONS_SCHEMA_VERSION:
        raise ValueError("Unsupported Checkpoint 1 decisions schema")
    if pending.get("schemaVersion") != CHECKPOINT_1_PENDING_SCHEMA_VERSION:
        raise ValueError("Unsupported Checkpoint 1 pending packet schema")
    approval_fp = _unsigned_fingerprint(approval, "approvalFingerprint", "Checkpoint 1 bulk approval")
    packet_fp = _unsigned_fingerprint(packet, "packetFingerprint", "Checkpoint 1 exception packet")
    decisions_fp = _unsigned_fingerprint(decisions, "decisionArtifactFingerprint", "Checkpoint 1 decisions")
    pending_fp = _unsigned_fingerprint(pending, "pendingPacketFingerprint", "Checkpoint 1 pending packet")
    for decision in decisions.get("decisions") or []:
        _unsigned_fingerprint(decision, "decisionFingerprint", "Checkpoint 1 decision")
    for override in decisions.get("bulkApprovalOverrides") or []:
        _unsigned_fingerprint(override, "decisionFingerprint", "Checkpoint 1 bulk override")
    for name, artifact in (("bulk approval", approval), ("exception packet", packet), ("decisions", decisions), ("pending packet", pending)):
        if clean(artifact.get("workbookSha256")) != workbook_sha:
            raise ValueError(f"Current workbook SHA does not match Checkpoint 1 {name}")
    if packet.get("bulkApprovalFingerprint") != approval_fp:
        raise ValueError("Checkpoint 1 exception packet is not bound to bulk approval")
    if decisions.get("bulkApprovalFingerprint") != approval_fp or decisions.get("sourcePacketFingerprint") != packet_fp:
        raise ValueError("Checkpoint 1 decisions are not bound to approval and exception packet")
    if (
        pending.get("bulkApprovalFingerprint") != approval_fp
        or pending.get("sourcePacketFingerprint") != packet_fp
        or pending.get("decisionArtifactFingerprint") != decisions_fp
    ):
        raise ValueError("Checkpoint 1 pending packet is not bound to approved decisions")
    if decisions.get("status") != "complete" or decisions.get("pendingReviewCount") != 0:
        raise ValueError("Checkpoint 1 decisions are not complete")
    if pending.get("status") != "complete" or pending.get("pendingReviewCount") != 0:
        raise ValueError("Checkpoint 1 pending review is not complete")
    if pending_fp != pending.get("pendingPacketFingerprint"):
        raise ValueError("Checkpoint 1 pending packet fingerprint mismatch")

    bindings = {str(model): dict(value or {}) for model, value in (approval.get("sourceReports") or {}).items()}
    if set(bindings) != set(TARGET_MODELS) or set(packet.get("sourceReports") or {}) != set(TARGET_MODELS):
        raise ValueError("Checkpoint 1 source report bindings are incomplete")
    for model, report in reports.items():
        binding = bindings[model]
        if binding.get("reportFingerprint") != report.get("reportFingerprint") or binding.get("workbookSha256") != workbook_sha:
            raise ValueError(f"Checkpoint 1 source report binding mismatch for {model}")
        packet_binding = (packet.get("sourceReports") or {}).get(model) or {}
        if packet_binding.get("reportFingerprint") != report.get("reportFingerprint"):
            raise ValueError(f"Checkpoint 1 exception packet report binding mismatch for {model}")
    return reports, approval, packet, decisions, pending, source_sheets, families


def emit_options_recovery_changeset(workbook_path: Path, report_dir: Path) -> dict[str, Any]:
    """Return one deterministic, immutable ChangeSet from completed Checkpoint 1."""

    workbook_path = Path(workbook_path)
    report_dir = Path(report_dir)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        reports, approval, packet, decisions, pending, source_sheets, families = _artifact_bindings(
            workbook_path, report_dir, workbook
        )
        sheet_rows: dict[str, tuple[list[str], dict[int, dict[str, Any]]]] = {}
        for sheet in set(source_sheets.values()) | set(families):
            if sheet in workbook.sheetnames:
                sheet_rows[sheet] = _headers_and_rows(workbook, sheet)

        review_records: dict[str, tuple[str, Mapping[str, Any]]] = {}
        desired: dict[tuple[str, int], dict[str, Any]] = {}
        option_rows: dict[tuple[str, int], dict[str, Any]] = {}
        for model, report in reports.items():
            sheet = source_sheets[model]
            _, physical = sheet_rows[sheet]
            for residual in report.get("residualRows") or []:
                identity = residual.get("identity") or {}
                row_number = int(identity.get("rowNumber") or 0)
                actual = physical.get(row_number)
                if not actual or clean(actual.get("option_id")) != clean(identity.get("optionId")):
                    raise ValueError(f"Recovery residual option row no longer matches: {model}/{row_number}")
                if clean(actual.get("rpo")).upper() != clean(identity.get("rpo")).upper():
                    raise ValueError(f"Recovery residual option RPO no longer matches: {model}/{row_number}")
                key = (sheet, row_number)
                if key in desired:
                    raise ValueError(f"Duplicate recovery residual option row: {sheet}/{row_number}")
                desired[key] = deepcopy(dict(residual.get("after") or {}))
                option_rows[key] = actual
                for item in residual.get("reviewItems") or []:
                    review_id = clean(item.get("reviewId"))
                    if not review_id or review_id in review_records:
                        raise ValueError(f"Duplicate or blank recovery review ID: {review_id}")
                    review_records[review_id] = (model, residual)

        all_review_ids = set(review_records)
        approved_ids = [clean(value) for value in approval.get("approvedReviewIds") or []]
        resolved_ids = [clean(value) for value in decisions.get("resolvedReviewIds") or []]
        if not all(approved_ids) or not all(resolved_ids) or len(set(approved_ids)) != len(approved_ids) or len(set(resolved_ids)) != len(resolved_ids):
            raise ValueError("Checkpoint 1 review coverage has blank or duplicate IDs")
        if set(approved_ids) & set(resolved_ids) or set(approved_ids) | set(resolved_ids) != all_review_ids:
            raise ValueError("Checkpoint 1 review coverage is incomplete or overlapping")
        packet_ids = {
            clean(review_id)
            for group in packet.get("exceptionGroups") or []
            for review_id in group.get("reviewIds") or []
        }
        if packet_ids != set(resolved_ids) or set(approval.get("exceptionReviewIds") or []) != packet_ids:
            raise ValueError("Checkpoint 1 exception review coverage binding mismatch")

        for override in decisions.get("bulkApprovalOverrides") or []:
            review_id = clean(override.get("reviewId"))
            if review_id not in set(approved_ids):
                raise ValueError("Bulk override is not bound to an approved review ID")
            model, residual = review_records[review_id]
            identity = residual.get("identity") or {}
            key = (source_sheets[model], int(identity.get("rowNumber") or 0))
            if clean(override.get("model")).lower() != model or dict(override.get("identity") or {}) != dict(identity):
                raise ValueError("Bulk override identity binding mismatch")
            fields = override.get("override")
            if not isinstance(fields, Mapping) or set(fields) - set(OPTION_VIEW_FIELDS):
                raise ValueError("Bulk override has unsupported option fields")
            desired[key].update(deepcopy(dict(fields)))

        deleted: dict[tuple[str, int], list[dict[str, Any]]] = {}
        delete_refs: dict[tuple[str, int], list[dict[str, Any]]] = {}

        def add_delete(option: Mapping[str, Any], refs: Any, report_id: str) -> None:
            sheet = clean(option.get("sheet") or option.get("targetSheet"))
            row_number = int(option.get("rowNumber") or 0)
            actual = sheet_rows.get(sheet, ([], {}))[1].get(row_number)
            if not actual or clean(actual.get("option_id")) != clean(option.get("optionId")):
                raise ValueError(f"Recovery delete option row no longer matches: {sheet}/{row_number}")
            if "rpo" in option and clean(actual.get("rpo")).upper() != clean(option.get("rpo")).upper():
                raise ValueError(f"Recovery delete option RPO no longer matches: {sheet}/{row_number}")
            deleted.setdefault((sheet, row_number), []).append({"kind": "report", "id": report_id})
            for ref in refs or []:
                ref_sheet = clean(ref.get("sheet"))
                ref_row = int(ref.get("rowNumber") or 0)
                actual_ref = sheet_rows.get(ref_sheet, ([], {}))[1].get(ref_row)
                if not actual_ref:
                    raise ValueError(f"Recovery delete reference row no longer exists: {ref_sheet}/{ref_row}")
                matches = ref.get("matches") or [ref]
                for match in matches:
                    column = clean(match.get("column"))
                    if not column or clean(actual_ref.get(column)) != clean(match.get("value")):
                        raise ValueError(f"Recovery delete reference value mismatch: {ref_sheet}/{ref_row}/{column}")
                delete_refs.setdefault((ref_sheet, ref_row), []).append({"kind": "report", "id": report_id})

        for report in reports.values():
            report_id = str(report["reportFingerprint"])
            for field in ("targetApplicabilityDeletions", "recordedInstructionDeletions"):
                for entry in report.get(field) or []:
                    add_delete(entry, entry.get("ownedReferenceRows"), report_id)
        for decision in decisions.get("decisions") or []:
            if decision.get("action") == "delete_option_and_owned_references":
                scope = decision.get("deleteScope") or {}
                for option in scope.get("optionRows") or []:
                    add_delete(option, scope.get("ownedReferenceRows"), str(decision.get("decisionGroupId")))

        for decision in decisions.get("decisions") or []:
            action = decision.get("action")
            decision_ids = [clean(value) for value in decision.get("reviewIds") or []]
            if not decision_ids or not set(decision_ids) <= set(resolved_ids):
                raise ValueError("Checkpoint 1 decision has unbound review IDs")
            if action not in {"accept", "override", "override_by_model", "delete_option_and_owned_references", "not_applicable_due_to_delete"}:
                raise ValueError(f"Unsupported Checkpoint 1 decision action: {action}")
            for review_id in decision_ids:
                model, residual = review_records[review_id]
                identity = residual.get("identity") or {}
                key = (source_sheets[model], int(identity.get("rowNumber") or 0))
                if action == "override":
                    fields = decision.get("override")
                elif action == "override_by_model":
                    fields = (decision.get("overrideByModel") or {}).get(model)
                else:
                    fields = None
                if fields is not None:
                    if not isinstance(fields, Mapping) or set(fields) - set(OPTION_VIEW_FIELDS):
                        raise ValueError("Checkpoint 1 decision has unsupported option fields")
                    desired[key].update(deepcopy(dict(fields)))

        remaps: dict[tuple[str, int], str] = {}
        cascade_edits: dict[tuple[str, int], dict[str, Any]] = {}
        for model, report in reports.items():
            for repair in report.get("idRepairs") or []:
                sheet = clean(repair.get("targetSheet"))
                row_number = int(repair.get("targetRowNumber") or 0)
                old = clean(repair.get("oldOptionId"))
                new = clean(repair.get("proposedOptionId"))
                actual = sheet_rows.get(sheet, ([], {}))[1].get(row_number)
                if not old or not new or not actual or clean(actual.get("option_id")) != old:
                    raise ValueError(f"Recovery id repair no longer matches: {sheet}/{row_number}")
                key = (sheet, row_number)
                if key in remaps and remaps[key] != new:
                    raise ValueError(f"Conflicting recovery id repair: {sheet}/{row_number}")
                remaps[key] = new
                if key not in desired:
                    raise ValueError(f"Recovery id repair is missing residual option state: {sheet}/{row_number}")
                desired[key]["option_id"] = new
                for ref in repair.get("cascade") or []:
                    ref_sheet = clean(ref.get("sheet"))
                    ref_row = int(ref.get("rowNumber") or 0)
                    column = clean(ref.get("column"))
                    actual_ref = sheet_rows.get(ref_sheet, ([], {}))[1].get(ref_row)
                    if not actual_ref or ref_sheet not in families or column not in actual_ref:
                        raise ValueError(f"Recovery id cascade row is not a bound workbook row: {ref_sheet}/{ref_row}")
                    expected = clean(ref.get("value"))
                    if clean(actual_ref.get(column)) != expected:
                        raise ValueError(f"Recovery id cascade value mismatch: {ref_sheet}/{ref_row}/{column}")
                    replacement = _replacement(actual_ref[column], old, new, clean(ref.get("matchType")))
                    edits = cascade_edits.setdefault((ref_sheet, ref_row), {})
                    if column in edits and edits[column] != replacement:
                        raise ValueError(f"Conflicting recovery id cascade edit: {ref_sheet}/{ref_row}/{column}")
                    edits[column] = replacement

        base_provenance = [
            {"kind": "options_recovery_report", "id": reports[model]["reportFingerprint"]}
            for model in TARGET_MODELS
        ] + [
            {"kind": "checkpoint_1_bulk_approval", "id": approval["approvalFingerprint"]},
            {"kind": "checkpoint_1_decisions", "id": decisions["decisionArtifactFingerprint"]},
        ]
        row_changes: list[dict[str, Any]] = []

        def change(action: str, sheet: str, family: str, key: dict[str, Any], fields: dict[str, Any], provenance: list[dict[str, Any]]) -> None:
            if not fields:
                return
            row_changes.append({"action": action, "sheet": sheet, "family": family, "key": key, "fields": fields, "provenance": provenance})

        for (sheet, row_number), actual in sorted(option_rows.items()):
            family = families.get(sheet)
            if family != "options":
                raise ValueError(f"Recovery option sheet has invalid family: {sheet}")
            if (sheet, row_number) in deleted:
                change("delete", sheet, family, _row_key(actual, family), _full_delete_fields(actual), deleted[(sheet, row_number)] + base_provenance)
                continue
            after = dict(actual)
            after.update(desired[(sheet, row_number)])
            if (sheet, row_number) in remaps:
                old_key = _row_key(actual, family)
                new_key = {"option_id": remaps[(sheet, row_number)]}
                change("delete", sheet, family, old_key, _full_delete_fields(actual), base_provenance)
                change("add", sheet, family, new_key, _add_fields(after), base_provenance)
                continue
            if clean(after.get("option_id")) != clean(actual.get("option_id")):
                raise ValueError("Recovery emitter refuses immutable option_id update without an idRepair")
            fields = {
                column: {"before": actual.get(column), "after": value}
                for column, value in after.items()
                if column in actual and not _semantic_equal(actual.get(column), value, family, column)
            }
            change("update", sheet, family, _row_key(actual, family), fields, base_provenance)

        for (sheet, row_number), provenance in sorted(deleted.items()):
            if (sheet, row_number) in option_rows:
                continue
            family = families.get(sheet)
            actual = sheet_rows.get(sheet, ([], {}))[1].get(row_number)
            if family != "options" or not actual:
                raise ValueError(f"Recovery delete option row is not a bound option row: {sheet}/{row_number}")
            change("delete", sheet, family, _row_key(actual, family), _full_delete_fields(actual), provenance + base_provenance)

        for (sheet, row_number), provenance in sorted(delete_refs.items()):
            if (sheet, row_number) in deleted:
                continue
            family = families.get(sheet)
            if not family:
                raise ValueError(f"Recovery delete reference sheet is not registered: {sheet}")
            actual = sheet_rows[sheet][1][row_number]
            change("delete", sheet, family, _row_key(actual, family), _full_delete_fields(actual), provenance + base_provenance)

        for (sheet, row_number), edits in sorted(cascade_edits.items()):
            if (sheet, row_number) in delete_refs:
                continue
            family = families[sheet]
            actual = sheet_rows[sheet][1][row_number]
            after = dict(actual)
            after.update(edits)
            key_columns = set(EDITOR_SHEET_META[family]["key"])
            if any(column in key_columns for column in edits):
                old_key = _row_key(actual, family)
                new_key = _row_key(after, family)
                change("delete", sheet, family, old_key, _full_delete_fields(actual), base_provenance)
                change("add", sheet, family, new_key, _add_fields(after), base_provenance)
            else:
                fields = {
                    column: {"before": actual.get(column), "after": after.get(column)}
                    for column in edits
                    if not _semantic_equal(actual.get(column), after.get(column), family, column)
                }
                change("update", sheet, family, _row_key(actual, family), fields, base_provenance)

        order = {"delete": 0, "update": 1, "add": 2}
        row_changes.sort(key=lambda item: (item["sheet"], order[item["action"]], json.dumps(item["key"], sort_keys=True)))
        payload: dict[str, Any] = {
            "schemaVersion": "workbook-changeset-1",
            "source": "options-recovery-checkpoint-1",
            "targets": list(TARGET_MODELS),
            "workbook": {"sha256": _sha256(workbook_path), "mtimeNs": str(workbook_path.stat().st_mtime_ns)},
            "sheetCreates": [],
            "rowChanges": row_changes,
            "noops": [],
            "warningAcknowledgementsRequested": [],
            "bindings": {
                "reports": {model: reports[model]["reportFingerprint"] for model in TARGET_MODELS},
                "bulkApprovalFingerprint": approval["approvalFingerprint"],
                "exceptionPacketFingerprint": packet["packetFingerprint"],
                "decisionArtifactFingerprint": decisions["decisionArtifactFingerprint"],
                "pendingPacketFingerprint": pending["pendingPacketFingerprint"],
            },
        }
        payload["semanticFingerprint"] = changeset_fingerprint(payload)
        payload["changeSetId"] = payload["semanticFingerprint"][:24]
        return parse_changeset(payload)
    finally:
        workbook.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--report-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args(argv)
    payload = emit_options_recovery_changeset(args.workbook, args.report_dir)
    encoded = json.dumps(payload, indent=2, sort_keys=True, ensure_ascii=False) + "\n"
    if args.output.exists():
        if args.output.read_text(encoding="utf-8") != encoded:
            raise ValueError(f"Refusing to overwrite ChangeSet with different content: {args.output}")
        return 0
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(encoded, encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
