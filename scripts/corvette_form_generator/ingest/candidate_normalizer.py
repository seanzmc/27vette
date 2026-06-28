"""Read-only candidate normalizer for Pass 0 order-guide evidence artifacts."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.ingest.source_profiler import rows_from_sheet, validate_output_dir
from corvette_form_generator.workbook import clean, workbook_truthy

REQUIRED_ARTIFACTS = {
    "source-layout.json": {"source_sheet", "sheet_type"},
    "variant-matrix.json": {"source_sheet", "raw_variant_header", "resolution_status"},
    "raw-rows.json": {
        "source_sheet",
        "source_row_index",
        "source_row_span",
        "primary_rpo_candidate",
        "status_cells",
    },
    "disclosure-links.json": {"source_sheet", "source_row_index", "marker"},
    "manifest.json": {"status", "artifact_files"},
}
ALLOWED_STATUS = {"standard", "available", "unavailable", "unresolved"}
TARGET_RPO_RE = re.compile(r"\(([A-Z0-9]{2,4})\)")


def normalize_order_guide_candidates(
    *,
    evidence_dir: Path,
    workbook: Path,
    output_dir: Path,
    run_id: str,
    root: Path | None = None,
) -> dict[str, Any]:
    """Normalize Pass 0 evidence into transient review candidate artifacts."""

    root = Path(root or Path.cwd()).resolve()
    evidence_dir = Path(evidence_dir).resolve()
    workbook = Path(workbook).resolve()
    output_dir = Path(output_dir).resolve()
    validate_output_dir(output_dir, root=root)

    evidence = load_evidence(evidence_dir)
    option_index = load_workbook_option_index(workbook)

    unresolved: list[dict[str, Any]] = []
    candidate_options, option_by_source_row = build_option_candidates(evidence["raw-rows"], option_index, unresolved)
    candidate_ovs = build_ovs_candidates(evidence["raw-rows"], option_by_source_row, unresolved)
    candidate_rules = build_rule_candidates(evidence["disclosure-links"], option_by_source_row, option_index, unresolved)
    candidate_price_rules: list[dict[str, Any]] = []
    add_layout_unresolved(evidence["source-layout"], unresolved)

    unresolved_counts = Counter(item["reason"] for item in unresolved)
    candidate_counts = {
        "options": len(candidate_options),
        "ovs": len(candidate_ovs),
        "rules": len(candidate_rules),
        "price_rules": len(candidate_price_rules),
    }
    artifact_files = [
        "candidate-options.json",
        "candidate-ovs.json",
        "candidate-rules.json",
        "candidate-price-rules.json",
        "candidate-summary.json",
        "unresolved-review.json",
        "unresolved-review.md",
    ]
    summary = {
        "input_evidence_dir": str(evidence_dir),
        "workbook": str(workbook),
        "run_id": run_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "status": "passed",
        "candidate_counts": candidate_counts,
        "unresolved_counts": dict(sorted(unresolved_counts.items())),
        "status_vocabulary": status_vocabulary(evidence["source-layout"]),
        "model_variant_coverage": model_variant_coverage(evidence["variant-matrix"]),
        "invariant_failures": [],
        "artifact_files": artifact_files,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "candidate-options.json", candidate_options)
    write_json(output_dir / "candidate-ovs.json", candidate_ovs)
    write_json(output_dir / "candidate-rules.json", candidate_rules)
    write_json(output_dir / "candidate-price-rules.json", candidate_price_rules)
    write_json(output_dir / "candidate-summary.json", summary)
    write_json(output_dir / "unresolved-review.json", unresolved_review_payload(
        unresolved,
        run_id=run_id,
        generated_at=summary["generated_at"],
        evidence_dir=evidence_dir,
        workbook=workbook,
    ))
    (output_dir / "unresolved-review.md").write_text(render_unresolved_review(unresolved) + "\n")
    return summary


def load_evidence(evidence_dir: Path) -> dict[str, Any]:
    evidence: dict[str, Any] = {}
    for filename, required_keys in REQUIRED_ARTIFACTS.items():
        path = evidence_dir / filename
        if not path.exists():
            raise ValueError(f"Missing required Pass 0 artifact: {path}")
        payload = json.loads(path.read_text())
        key = filename.replace(".json", "").replace("-", "_")
        artifact_name = filename.removesuffix(".json")
        validate_artifact_shape(filename, payload, required_keys)
        evidence[artifact_name] = payload
        evidence[key] = payload
    manifest = evidence["manifest"]
    if manifest.get("status") != "passed":
        raise ValueError("Pass 1 requires manifest.json status to be passed before candidate normalization.")
    return evidence


def validate_artifact_shape(filename: str, payload: Any, required_keys: set[str]) -> None:
    if filename == "manifest.json":
        if not isinstance(payload, dict):
            raise ValueError(f"{filename} must be an object")
        missing = sorted(required_keys - set(payload))
        if missing:
            raise ValueError(f"{filename} missing required keys: {missing}")
        return
    if not isinstance(payload, list):
        raise ValueError(f"{filename} must be a list")
    for index, row in enumerate(payload, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"{filename} row {index} must be an object")
        missing = sorted(required_keys - set(row))
        if missing:
            raise ValueError(f"{filename} row {index} missing required keys: {missing}")


def load_workbook_option_index(workbook: Path) -> dict[str, Any]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        model_sheet_rows = [
            row
            for row in rows_from_sheet(wb, "model_workbook_sources")
            if clean(row.get("source_role")) == "source_option_sheet" and workbook_truthy(row.get("active"))
        ]
        by_model_rpo: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
        by_rpo: dict[str, list[dict[str, str]]] = defaultdict(list)
        for source in model_sheet_rows:
            model_key = clean(source.get("model_key"))
            sheet_name = clean(source.get("sheet_name"))
            if not model_key or not sheet_name or sheet_name not in wb.sheetnames:
                continue
            for row in rows_from_sheet(wb, sheet_name):
                rpo = clean(row.get("rpo")).upper()
                if not rpo:
                    continue
                payload = {key: clean(value) for key, value in row.items()}
                payload["model_key"] = model_key
                payload["sheet_name"] = sheet_name
                by_model_rpo[model_key][rpo].append(payload)
                by_rpo[rpo].append(payload)
        return {"by_model_rpo": by_model_rpo, "by_rpo": by_rpo}
    finally:
        wb.close()


def build_option_candidates(
    raw_rows: list[dict[str, Any]],
    option_index: dict[str, Any],
    unresolved: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[tuple[str, int], dict[str, Any]]]:
    candidates: list[dict[str, Any]] = []
    option_by_source_row: dict[tuple[str, int], dict[str, Any]] = {}
    for row in raw_rows:
        source_sheet = row["source_sheet"]
        source_row = row["source_row_index"]
        source_ref = row_source_ref(row)
        if "section_context_row" in row.get("row_flags", []):
            unresolved.append(
                unresolved_item(
                    reason="section_context_requires_review",
                    source_refs=[source_ref],
                    raw_values=raw_values_for_row(row),
                    message="Source row is a section/context row, not an option candidate.",
                )
            )
            continue
        rpo = clean(row.get("primary_rpo_candidate")).upper()
        if not rpo:
            unresolved.append(
                unresolved_item(
                    reason="missing_or_invalid_primary_rpo",
                    source_refs=[source_ref],
                    raw_values=raw_values_for_row(row),
                    message="No valid primary RPO candidate was available for this row.",
                )
            )
            continue
        candidate_ref = candidate_option_ref(source_sheet, source_row, rpo)
        model_keys = sorted(
            {
                clean(cell.get("model_key_candidate"))
                for cell in row.get("status_cells", [])
                if clean(cell.get("model_key_candidate"))
            }
        )
        workbook_match = exact_option_match(rpo, model_keys, option_index)
        status_summary = summarize_status(row.get("status_cells", []))
        normalized = {
            "candidate_option_ref": candidate_ref,
            "model_key_candidates": model_keys,
            "rpo": rpo,
            "orderable_rpo_raw": row.get("orderable_rpo", {}).get("raw_value", ""),
            "ref_only_rpo_raw": row.get("ref_only_rpo", {}).get("raw_value", ""),
            "source_description_raw": row.get("description", {}).get("raw_value", ""),
            "source_option_name_candidate": source_name_candidate(row.get("description", {}).get("raw_value", "")),
            "section_context_raw": row.get("section_context", ""),
            "section_id_candidate": workbook_match.get("section_id", "") if workbook_match else "",
            "status_summary": status_summary,
            "canonical_option_match": workbook_match or None,
        }
        candidate = candidate_envelope(
            candidate_id=candidate_ref,
            family="options",
            resolution_status="candidate" if workbook_match else "needs_review",
            confidence="mechanical" if workbook_match else "source_hint",
            source_refs=[source_ref],
            raw_values={
                "orderable_rpo_raw": normalized["orderable_rpo_raw"],
                "ref_only_rpo_raw": normalized["ref_only_rpo_raw"],
                "source_description_raw": normalized["source_description_raw"],
                "section_context_raw": normalized["section_context_raw"],
            },
            normalized_values=normalized,
            workbook_match=workbook_match or None,
            review_notes=[] if workbook_match else ["No exact current workbook option match; review before apply."],
        )
        candidates.append(candidate)
        option_by_source_row[(source_sheet, source_row)] = candidate
    return candidates, option_by_source_row


def build_ovs_candidates(
    raw_rows: list[dict[str, Any]],
    option_by_source_row: dict[tuple[str, int], dict[str, Any]],
    unresolved: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for row in raw_rows:
        option_candidate = option_by_source_row.get((row["source_sheet"], row["source_row_index"]))
        if not option_candidate:
            continue
        option_ref = option_candidate["normalized_values"]["candidate_option_ref"]
        for cell in row.get("status_cells", []):
            status = clean(cell.get("parsed_base_status")) or "unresolved"
            source_ref = cell_source_ref(row, cell)
            if status not in ALLOWED_STATUS:
                raise ValueError(f"Unsupported normalized status candidate {status!r} at {cell.get('coordinate')}")
            variant_id = clean(cell.get("variant_id_candidate"))
            model_key = clean(cell.get("model_key_candidate"))
            if not variant_id or not model_key:
                unresolved.append(
                    unresolved_item(
                        reason="unmatched_or_ambiguous_variant_evidence",
                        source_refs=[source_ref],
                        candidate_refs=[option_ref],
                        raw_values={
                            "raw_status": clean(cell.get("raw_status")),
                            "raw_variant_header": clean(cell.get("raw_variant_header")),
                            "status_marker": clean(cell.get("status_marker")),
                        },
                        normalized_values={"normalized_status_candidate": status},
                        message="Status cell lacks matched variant/model evidence.",
                    )
                )
                continue
            normalized = {
                "candidate_option_ref": option_ref,
                "variant_id": variant_id,
                "model_key": model_key,
                "raw_status": clean(cell.get("raw_status")),
                "normalized_status_candidate": status,
                "status_marker": clean(cell.get("status_marker")),
                "status_flags": cell.get("status_flags", []),
                "source_cell": clean(cell.get("coordinate")),
            }
            candidates.append(
                candidate_envelope(
                    candidate_id=f"candovs-{slug(row['source_sheet'])}-cell-{clean(cell.get('coordinate')).lower()}",
                    family="ovs",
                    resolution_status="candidate" if status != "unresolved" else "needs_review",
                    confidence="mechanical" if status != "unresolved" else "unresolved",
                    source_refs=[source_ref],
                    raw_values={"raw_status": normalized["raw_status"]},
                    normalized_values=normalized,
                    workbook_match=option_candidate.get("workbook_match"),
                    review_notes=[] if not normalized["status_marker"] else ["Status marker preserved; review linked disclosure before apply."],
                )
            )
    return candidates


def build_rule_candidates(
    disclosure_links: list[dict[str, Any]],
    option_by_source_row: dict[tuple[str, int], dict[str, Any]],
    option_index: dict[str, Any],
    unresolved: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for link in disclosure_links:
        relationship_hint = clean(link.get("candidate_relationship_hint"))
        phrase_hints = link.get("phrase_hints", [])
        if not relationship_hint and not phrase_hints:
            continue
        source_sheet = link["source_sheet"]
        source_row = link["source_row_index"]
        option_candidate = option_by_source_row.get((source_sheet, source_row))
        source_ref = {
            "source_sheet": source_sheet,
            "source_row_index": source_row,
            "marker": clean(link.get("marker")),
            "status_cells": link.get("status_cells", []),
        }
        if not option_candidate:
            unresolved.append(
                unresolved_item(
                    reason="disclosure_relationship_requires_review",
                    source_refs=[source_ref],
                    raw_values={
                        "marker": clean(link.get("marker")),
                        "description_fragment": clean(link.get("description_fragment")),
                        "phrase_hints": phrase_hints,
                    },
                    message="Disclosure relationship hint could not be linked to an option candidate.",
                )
            )
            continue
        fragment = clean(link.get("description_fragment"))
        target_tokens = sorted(set(TARGET_RPO_RE.findall(fragment)))
        target_match_status = target_match_status_for_tokens(target_tokens, option_index)
        normalized = {
            "candidate_rule_ref": f"candrule-{slug(source_sheet)}-row-{source_row}-m{clean(link.get('marker'))}",
            "source_candidate_option_ref": option_candidate["normalized_values"]["candidate_option_ref"],
            "marker": clean(link.get("marker")),
            "description_fragment": fragment,
            "phrase_hints": phrase_hints,
            "relationship_hint": relationship_hint,
            "target_rpo_tokens": target_tokens,
            "target_match_status": target_match_status,
            "recommended_review_action": recommended_rule_review_action(relationship_hint, target_match_status),
        }
        if target_match_status != "exact":
            unresolved.append(
                unresolved_item(
                    reason="target_rpo_token_ambiguous_or_missing",
                    source_refs=[source_ref],
                    candidate_refs=[normalized["candidate_rule_ref"]],
                    raw_values={
                        "marker": normalized["marker"],
                        "description_fragment": fragment,
                        "target_rpo_tokens": target_tokens,
                    },
                    normalized_values={
                        "relationship_hint": relationship_hint,
                        "target_match_status": target_match_status,
                    },
                    message="Disclosure relationship hint requires target RPO review before apply.",
                )
            )
        candidates.append(
            candidate_envelope(
                candidate_id=normalized["candidate_rule_ref"],
                family="rules",
                resolution_status="needs_review",
                confidence="source_hint",
                source_refs=[source_ref],
                raw_values={"description_fragment": fragment},
                normalized_values=normalized,
                workbook_match=None,
                review_notes=["Review-only relationship hint; do not apply as workbook rule without human approval."],
            )
        )
    return candidates


def add_layout_unresolved(source_layout: list[dict[str, Any]], unresolved: list[dict[str, Any]]) -> None:
    for layout in source_layout:
        sheet_type = clean(layout.get("sheet_type"))
        if sheet_type == "price_schedule":
            unresolved.append(
                unresolved_item(
                    reason="price_schedule_rows_not_extracted",
                    source_refs=[{"source_sheet": layout.get("source_sheet"), "header_row": layout.get("header_row")}],
                    raw_values={"sheet_type": sheet_type, "source_sheet": layout.get("source_sheet")},
                    message="Price Schedule is layout evidence only in Pass 0; no price candidates emitted in Pass 1.",
                )
            )
        elif sheet_type == "color_trim":
            unresolved.append(
                unresolved_item(
                    reason="color_trim_rows_not_extracted",
                    source_refs=[{"source_sheet": layout.get("source_sheet"), "header_row": layout.get("header_row")}],
                    raw_values={"sheet_type": sheet_type, "source_sheet": layout.get("source_sheet")},
                    message="Color and Trim is layout evidence only in Pass 0; no interior/color candidates emitted in Pass 1.",
                )
            )
        elif sheet_type not in {"matrix", "price_schedule", "color_trim"}:
            unresolved.append(
                unresolved_item(
                    reason="non_matrix_sheet_evidence_only",
                    source_refs=[{"source_sheet": layout.get("source_sheet"), "header_row": layout.get("header_row")}],
                    raw_values={"sheet_type": sheet_type, "source_sheet": layout.get("source_sheet")},
                    message="Non-matrix source sheet preserved as evidence only.",
                )
            )


def exact_option_match(rpo: str, model_keys: list[str], option_index: dict[str, Any]) -> dict[str, str] | None:
    matches: list[dict[str, str]] = []
    by_model_rpo = option_index["by_model_rpo"]
    for model_key in model_keys:
        matches.extend(by_model_rpo.get(model_key, {}).get(rpo, []))
    if len(matches) != 1:
        return None
    match = matches[0]
    return {
        "model_key": match.get("model_key", ""),
        "sheet_name": match.get("sheet_name", ""),
        "option_id": match.get("option_id", ""),
        "rpo": match.get("rpo", ""),
        "section_id": match.get("section_id", ""),
    }


def target_match_status_for_tokens(tokens: list[str], option_index: dict[str, Any]) -> str:
    if not tokens:
        return "unresolved"
    statuses = []
    for token in tokens:
        matches = option_index["by_rpo"].get(token, [])
        if len(matches) == 1:
            statuses.append("exact")
        elif len(matches) > 1:
            statuses.append("ambiguous")
        else:
            statuses.append("unresolved")
    if all(status == "exact" for status in statuses):
        return "exact"
    if any(status == "ambiguous" for status in statuses):
        return "ambiguous"
    return "unresolved"


def summarize_status(cells: list[dict[str, Any]]) -> dict[str, dict[str, dict[str, int]]]:
    summary: dict[str, dict[str, dict[str, int]]] = {}
    for cell in cells:
        model_key = clean(cell.get("model_key_candidate"))
        variant_id = clean(cell.get("variant_id_candidate"))
        status = clean(cell.get("parsed_base_status")) or "unresolved"
        if not model_key or not variant_id:
            continue
        if status not in ALLOWED_STATUS:
            raise ValueError(f"Unsupported normalized status candidate {status!r} at {cell.get('coordinate')}")
        summary.setdefault(model_key, {}).setdefault(variant_id, {}).setdefault(status, 0)
        summary[model_key][variant_id][status] += 1
    return summary


def source_name_candidate(description: str) -> str:
    return clean(description).split(" / ", 1)[0].split("\n", 1)[0]


def row_source_ref(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "source_sheet": row.get("source_sheet"),
        "source_row_index": row.get("source_row_index"),
        "source_row_span": row.get("source_row_span"),
        "orderable_rpo_cell": row.get("orderable_rpo", {}).get("coordinate"),
        "ref_only_rpo_cell": row.get("ref_only_rpo", {}).get("coordinate"),
        "description_cell": row.get("description", {}).get("coordinate"),
    }


def cell_source_ref(row: dict[str, Any], cell: dict[str, Any]) -> dict[str, Any]:
    return {
        "source_sheet": row.get("source_sheet"),
        "source_row_index": row.get("source_row_index"),
        "source_cell": cell.get("coordinate"),
        "raw_status": cell.get("raw_status"),
        "raw_variant_header": cell.get("raw_variant_header"),
    }


def candidate_envelope(
    *,
    candidate_id: str,
    family: str,
    resolution_status: str,
    confidence: str,
    source_refs: list[dict[str, Any]],
    raw_values: dict[str, Any],
    normalized_values: dict[str, Any],
    workbook_match: dict[str, Any] | None,
    review_notes: list[str],
) -> dict[str, Any]:
    if not source_refs:
        raise ValueError(f"Candidate {candidate_id} has no source references")
    return {
        "candidate_id": candidate_id,
        "candidate_family": family,
        "resolution_status": resolution_status,
        "confidence": confidence,
        "source_refs": source_refs,
        "raw_values": raw_values,
        "normalized_values": normalized_values,
        "workbook_match": workbook_match,
        "review_notes": review_notes,
    }


CATEGORY_BY_REASON = {
    "section_context_requires_review": "section_context",
    "missing_or_invalid_primary_rpo": "rpo_identity",
    "unmatched_or_ambiguous_variant_evidence": "source_shape",
    "disclosure_relationship_requires_review": "relationship_hint",
    "target_rpo_token_ambiguous_or_missing": "relationship_hint",
    "price_schedule_rows_not_extracted": "price_out_of_scope",
    "color_trim_rows_not_extracted": "color_trim_out_of_scope",
    "non_matrix_sheet_evidence_only": "source_shape",
}
OUT_OF_SCOPE_REASONS = {
    "price_schedule_rows_not_extracted",
    "color_trim_rows_not_extracted",
    "non_matrix_sheet_evidence_only",
}


def unresolved_item(
    *,
    reason: str,
    source_refs: list[dict[str, Any]],
    message: str,
    raw_values: dict[str, Any] | None = None,
    normalized_values: dict[str, Any] | None = None,
    candidate_refs: list[str] | None = None,
) -> dict[str, Any]:
    severity = "out_of_scope" if reason in OUT_OF_SCOPE_REASONS else "review"
    return {
        "reason": reason,
        "category": CATEGORY_BY_REASON.get(reason, "source_shape"),
        "severity": severity,
        "candidate_refs": candidate_refs or [],
        "source_refs": source_refs,
        "raw_values": raw_values or {},
        "normalized_values": normalized_values or {},
        "blocked_decision": message,
        "suggested_decision_states": suggested_decision_states(severity),
        "message": message,
    }


def suggested_decision_states(severity: str) -> list[str]:
    if severity == "out_of_scope":
        return ["blocked_out_of_scope", "needs_source_review", "skip"]
    return ["needs_source_review", "edit_before_apply", "skip"]


def raw_values_for_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "orderable_rpo_raw": row.get("orderable_rpo", {}).get("raw_value", ""),
        "ref_only_rpo_raw": row.get("ref_only_rpo", {}).get("raw_value", ""),
        "source_description_raw": row.get("description", {}).get("raw_value", ""),
        "section_context_raw": row.get("section_context", ""),
    }


def unresolved_review_payload(
    unresolved: list[dict[str, Any]],
    *,
    run_id: str,
    generated_at: str,
    evidence_dir: Path,
    workbook: Path,
) -> dict[str, Any]:
    counts = Counter(item["reason"] for item in unresolved)
    items = []
    for index, item in enumerate(unresolved, start=1):
        reason = item["reason"]
        items.append({
            "unresolved_id": f"unres-{index:05d}-{slug(reason)}",
            "reason": reason,
            "category": item.get("category") or CATEGORY_BY_REASON.get(reason, "source_shape"),
            "severity": item.get("severity") or ("out_of_scope" if reason in OUT_OF_SCOPE_REASONS else "review"),
            "candidate_refs": item.get("candidate_refs", []),
            "source_refs": item.get("source_refs", []),
            "raw_values": item.get("raw_values", {}),
            "normalized_values": item.get("normalized_values", {}),
            "blocked_decision": item.get("blocked_decision") or item.get("message", ""),
            "suggested_decision_states": item.get("suggested_decision_states") or suggested_decision_states(item.get("severity", "review")),
        })
    return {
        "version": 1,
        "run_id": run_id,
        "generated_at": generated_at,
        "input_evidence_dir": str(evidence_dir),
        "workbook": str(workbook),
        "unresolved_counts": dict(sorted(counts.items())),
        "items": items,
    }


def recommended_rule_review_action(relationship_hint: str, target_match_status: str) -> str:
    if target_match_status != "exact":
        return "review_target_rpo"
    if relationship_hint in {"requires", "excludes", "includes"}:
        return "review_direct_rule"
    return "review_group_or_exclusive_rule"


def candidate_option_ref(source_sheet: str, source_row: int, rpo: str) -> str:
    return f"candopt-{slug(source_sheet)}-row-{source_row}-{slug(rpo)}"


def slug(value: Any) -> str:
    text = clean(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "blank"


def status_vocabulary(source_layout: list[dict[str, Any]]) -> dict[str, int]:
    counts: Counter[str] = Counter()
    for sheet in source_layout:
        counts.update(sheet.get("status_vocabulary", {}))
    return dict(sorted(counts.items()))


def model_variant_coverage(variant_matrix: list[dict[str, Any]]) -> dict[str, Any]:
    coverage: dict[str, Any] = defaultdict(lambda: {"variant_ids": set(), "source_columns": 0})
    unresolved = 0
    for variant in variant_matrix:
        model_key = clean(variant.get("parsed_target_model")) or "unmatched"
        variant_id = clean(variant.get("parsed_variant_id")) or "unmatched"
        if variant.get("resolution_status") != "matched":
            unresolved += 1
        coverage[model_key]["variant_ids"].add(variant_id)
        coverage[model_key]["source_columns"] += 1
    return {
        model_key: {"variant_ids": sorted(values["variant_ids"]), "source_columns": values["source_columns"]}
        for model_key, values in sorted(coverage.items())
    } | {"unresolved_variant_columns": unresolved}


def render_unresolved_review(unresolved: list[dict[str, Any]]) -> str:
    lines = ["# Pass 1 unresolved review", ""]
    if not unresolved:
        lines.append("No unresolved items.")
        return "\n".join(lines)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in unresolved:
        grouped[item["reason"]].append(item)
    for reason in sorted(grouped):
        lines.extend([f"## {reason}", ""])
        for item in grouped[reason]:
            refs = "; ".join(format_source_ref(ref) for ref in item["source_refs"])
            lines.append(f"- {item['message']} Source: {refs}")
        lines.append("")
    return "\n".join(lines).rstrip()


def format_source_ref(ref: dict[str, Any]) -> str:
    parts = []
    for key in ("source_sheet", "source_row_index", "source_cell", "header_row", "marker"):
        value = ref.get(key)
        if value not in (None, ""):
            parts.append(f"{key}={value}")
    return ", ".join(parts) or str(ref)


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
