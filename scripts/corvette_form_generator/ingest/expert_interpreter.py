"""Read-only expert interpretation and review reduction for order-guide ingest."""

from __future__ import annotations

import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from itertools import combinations
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.ingest.model_selection import (
    assert_selection_matches,
    build_model_selection,
    read_model_selection,
    selection_fingerprint,
)
from corvette_form_generator.ingest.source_profiler import rows_from_sheet, validate_output_dir
from corvette_form_generator.workbook import clean, workbook_truthy

EVIDENCE_FILES = {
    "source-layout.json": "source_layout",
    "variant-matrix.json": "variant_matrix",
    "raw-rows.json": "raw_rows",
    "disclosure-links.json": "disclosure_links",
    "manifest.json": "manifest",
}
CANDIDATE_FILES = {
    "candidate-options.json": "candidate_options",
    "candidate-ovs.json": "candidate_ovs",
    "candidate-rules.json": "candidate_rules",
    "candidate-price-rules.json": "candidate_price_rules",
    "candidate-summary.json": "candidate_summary",
    "unresolved-review.json": "unresolved_review",
}
OUTPUT_FILES = [
    "interpretation-summary.json",
    "interpreted-options.json",
    "review-queue.json",
    "duplicate-rpo-report.json",
    "duplicate-rpo-report.md",
    "source-sheet-coverage.json",
    "source-sheet-coverage.md",
    "blocked-interpretation.json",
]
WORKBOOK_BUILD_FILES = ["model-selection.json", "workbook-build-summary.json", "workbook-build-review-units.json"]
AUTO_ALLOWED_DUPLICATE_CLASSES = {"single_source", "redundant_duplicates"}
VISIBLE_CONFIDENCES = {"mechanical_safe", "review_needed", "blocked"}


def interpret_order_guide_candidates(
    *,
    evidence_dir: Path,
    candidates_dir: Path,
    workbook: Path,
    output_dir: Path,
    run_id: str,
    root: Path | None = None,
    selected_models: list[str] | str | None = None,
    primary_models: list[str] | str | None = None,
    comparator_models: list[str] | str | None = None,
) -> dict[str, Any]:
    """Aggregate Pass 1 raw candidates into read-only model/RPO review units."""

    root = Path(root or Path.cwd()).resolve()
    evidence_dir = Path(evidence_dir).resolve()
    candidates_dir = Path(candidates_dir).resolve()
    workbook = Path(workbook).resolve()
    output_dir = Path(output_dir).resolve()
    validate_output_dir(output_dir, root=root)

    evidence = load_artifacts(evidence_dir, EVIDENCE_FILES)
    candidates = load_artifacts(candidates_dir, CANDIDATE_FILES)
    validate_inputs(evidence, candidates)
    selection_metadata = load_or_validate_selection(
        evidence_dir=evidence_dir,
        candidates_dir=candidates_dir,
        evidence=evidence,
        run_id=run_id,
        selected_models=selected_models,
        primary_models=primary_models,
        comparator_models=comparator_models,
    )
    workbook_index = load_workbook_identity_and_status(workbook)

    interpreted_options = build_interpreted_options(candidates, workbook_index)
    if selection_metadata:
        selected = set(selection_metadata["selected_models"])
        interpreted_options = [item for item in interpreted_options if item["model_key"] in selected]
    source_sheet_coverage = build_source_sheet_coverage(interpreted_options)
    duplicate_report = build_duplicate_report(interpreted_options)
    review_queue = [item for item in interpreted_options if item["interpretation_confidence"] in VISIBLE_CONFIDENCES]
    blocked_interpretation = build_blocked_interpretation(interpreted_options, candidates["unresolved_review"])
    workbook_build_units = build_workbook_build_units(interpreted_options, blocked_interpretation, selection_metadata)
    summary = build_summary(
        run_id=run_id,
        evidence_dir=evidence_dir,
        candidates_dir=candidates_dir,
        workbook=workbook,
        interpreted_options=interpreted_options,
        review_queue=review_queue,
        blocked_interpretation=blocked_interpretation,
        duplicate_report=duplicate_report,
        source_sheet_coverage=source_sheet_coverage,
        candidate_summary=candidates["candidate_summary"],
        unresolved_review=candidates["unresolved_review"],
        selection_metadata=selection_metadata,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    if selection_metadata:
        shutil.copyfile(candidates_dir / "model-selection.json", output_dir / "model-selection.json")
    write_json(output_dir / "interpreted-options.json", interpreted_options)
    write_json(output_dir / "review-queue.json", review_queue)
    write_json(output_dir / "duplicate-rpo-report.json", duplicate_report)
    (output_dir / "duplicate-rpo-report.md").write_text(render_duplicate_report(duplicate_report) + "\n")
    write_json(output_dir / "source-sheet-coverage.json", source_sheet_coverage)
    (output_dir / "source-sheet-coverage.md").write_text(render_source_sheet_coverage(source_sheet_coverage) + "\n")
    write_json(output_dir / "blocked-interpretation.json", blocked_interpretation)
    if selection_metadata:
        write_json(output_dir / "workbook-build-review-units.json", workbook_build_units)
        workbook_build_summary = build_workbook_build_summary(
            run_id=run_id,
            selection_metadata=selection_metadata,
            workbook_build_units=workbook_build_units,
            output_dir=output_dir,
        )
        write_json(output_dir / "workbook-build-summary.json", workbook_build_summary)
    summary["artifact_files"] = OUTPUT_FILES + (WORKBOOK_BUILD_FILES if selection_metadata else [])
    write_json(output_dir / "interpretation-summary.json", summary)

    return {
        "status": "passed",
        "output_dir": str(output_dir),
        "raw_candidate_counts": summary["raw_candidate_counts"],
        "raw_candidate_total": summary["raw_candidate_total"],
        "interpreted_option_count": summary["interpreted_option_count"],
        "hidden_auto_confirmed_count": summary["hidden_auto_confirmed_count"],
        "visible_review_queue_count": summary["visible_review_queue_count"],
        "blocked_count": summary["blocked_count"],
        "duplicate_rpo_count": summary["duplicate_rpo_count"],
        "conflicting_duplicate_count": summary["conflicting_duplicate_count"],
        "reduction_status": summary["reduction_status"],
        "reduction_reason_codes": summary["reduction_reason_codes"],
        "artifact_files": summary["artifact_files"],
    }


def load_artifacts(directory: Path, file_map: dict[str, str]) -> dict[str, Any]:
    artifacts: dict[str, Any] = {}
    for filename, key in file_map.items():
        path = directory / filename
        if not path.exists():
            raise ValueError(f"Missing required ingest artifact: {path}")
        artifacts[key] = json.loads(path.read_text())
    return artifacts


def load_or_validate_selection(
    *,
    evidence_dir: Path,
    candidates_dir: Path,
    evidence: dict[str, Any],
    run_id: str,
    selected_models: list[str] | str | None,
    primary_models: list[str] | str | None,
    comparator_models: list[str] | str | None,
) -> dict[str, Any] | None:
    selection_path = candidates_dir / "model-selection.json"
    if selected_models or primary_models or comparator_models:
        selection = read_model_selection(selection_path)
        expected = build_model_selection(
            evidence_dir=evidence_dir,
            variant_matrix=evidence["variant_matrix"],
            run_id=run_id,
            selected_models=selected_models or selection["selected_models"],
            primary_models=primary_models or selection["primary_models"],
            comparator_models=comparator_models or selection["comparator_models"],
        )
        assert_selection_matches(selection, expected, left="candidate model-selection.json", right="interpreter selected-model args")
        return selection
    if selection_path.exists():
        return read_model_selection(selection_path)
    return None


def validate_inputs(evidence: dict[str, Any], candidates: dict[str, Any]) -> None:
    manifest = evidence["manifest"]
    if not isinstance(manifest, dict) or manifest.get("status") != "passed":
        raise ValueError("Pass 3 requires Pass 0 manifest.json status to be passed.")
    summary = candidates["candidate_summary"]
    if not isinstance(summary, dict) or summary.get("status") != "passed":
        raise ValueError("Pass 3 requires Pass 1 candidate-summary.json status to be passed.")
    unresolved = candidates["unresolved_review"]
    if not isinstance(unresolved, dict) or not isinstance(unresolved.get("items"), list):
        raise ValueError("Pass 3 requires unresolved-review.json with an items list.")


def load_workbook_identity_and_status(workbook: Path) -> dict[str, Any]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        model_active = {
            clean(row.get("model_key")): workbook_truthy(row.get("active"))
            for row in rows_from_sheet(wb, "model_master")
            if clean(row.get("model_key"))
        }
        source_rows = list(rows_from_sheet(wb, "model_workbook_sources"))
        option_sources = [row for row in source_rows if clean(row.get("source_role")) == "source_option_sheet"]
        status_sources = [row for row in source_rows if clean(row.get("source_role")) == "status_sheet"]
        by_model_rpo: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
        by_model_option_id: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)
        inactive_by_model_rpo: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
        for source in option_sources:
            source_active = workbook_truthy(source.get("active"))
            model_key = clean(source.get("model_key"))
            sheet_name = clean(source.get("sheet_name"))
            if not model_key or sheet_name not in wb.sheetnames:
                continue
            for row in rows_from_sheet(wb, sheet_name):
                rpo = clean(row.get("rpo")).upper()
                option_id = clean(row.get("option_id"))
                if not rpo or not option_id:
                    continue
                payload = {key: clean(value) for key, value in row.items()}
                payload["model_key"] = model_key
                payload["sheet_name"] = sheet_name
                by_model_option_id[model_key][option_id] = payload
                if source_active and (clean(row.get("active")) == "" or workbook_truthy(row.get("active"))):
                    by_model_rpo[model_key][rpo].append(payload)
                else:
                    inactive_by_model_rpo[model_key][rpo].append(payload)
        status_by_model_option: dict[str, dict[str, dict[str, str]]] = defaultdict(lambda: defaultdict(dict))
        for source in status_sources:
            model_key = clean(source.get("model_key"))
            sheet_name = clean(source.get("sheet_name"))
            if not model_key or sheet_name not in wb.sheetnames:
                continue
            for row in rows_from_sheet(wb, sheet_name):
                option_id = clean(row.get("option_id"))
                variant_id = clean(row.get("variant_id"))
                status = clean(row.get("status"))
                if option_id and variant_id:
                    status_by_model_option[model_key][option_id][variant_id] = status
        return {
            "model_active": model_active,
            "by_model_rpo": by_model_rpo,
            "inactive_by_model_rpo": inactive_by_model_rpo,
            "status_by_model_option": status_by_model_option,
        }
    finally:
        wb.close()


def build_interpreted_options(candidates: dict[str, Any], workbook_index: dict[str, Any]) -> list[dict[str, Any]]:
    option_by_ref = {
        clean(row.get("normalized_values", {}).get("candidate_option_ref")): row
        for row in candidates["candidate_options"]
    }
    ovs_by_option_ref: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in candidates["candidate_ovs"]:
        option_ref = clean(row.get("normalized_values", {}).get("candidate_option_ref"))
        if option_ref:
            ovs_by_option_ref[option_ref].append(row)
    rules_by_option_ref: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in candidates["candidate_rules"]:
        option_ref = clean(row.get("normalized_values", {}).get("source_candidate_option_ref"))
        if option_ref:
            rules_by_option_ref[option_ref].append(row)
    unresolved_by_candidate_ref = unresolved_by_ref(candidates["unresolved_review"])

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for option_ref, option in option_by_ref.items():
        normalized = option.get("normalized_values", {})
        rpo = clean(normalized.get("rpo")).upper()
        if not rpo:
            continue
        model_keys = normalized.get("model_key_candidates") or []
        for model_key in model_keys:
            model_key = clean(model_key)
            if not model_key:
                continue
            group = grouped.setdefault(
                (model_key, rpo),
                {"model_key": model_key, "rpo": rpo, "occurrences": [], "ovs": [], "rules": [], "unresolved": []},
            )
            group["occurrences"].append(source_occurrence(option_ref, option))
            group["ovs"].extend(row for row in ovs_by_option_ref.get(option_ref, []) if clean(row.get("normalized_values", {}).get("model_key")) == model_key)
            group["rules"].extend(rules_by_option_ref.get(option_ref, []))
            group["unresolved"].extend(unresolved_by_candidate_ref.get(option_ref, []))
            for rule in rules_by_option_ref.get(option_ref, []):
                group["unresolved"].extend(unresolved_by_candidate_ref.get(clean(rule.get("candidate_id")), []))

    interpreted = [interpret_group(group, workbook_index) for group in grouped.values()]
    return sorted(interpreted, key=lambda item: (item["model_key"], item["rpo"]))


def unresolved_by_ref(unresolved_review: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    by_ref: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in unresolved_review.get("items", []):
        for candidate_ref in item.get("candidate_refs", []):
            by_ref[clean(candidate_ref)].append(item)
    return by_ref


def source_occurrence(option_ref: str, option: dict[str, Any]) -> dict[str, Any]:
    source_ref = (option.get("source_refs") or [{}])[0]
    raw_values = option.get("raw_values", {})
    normalized = option.get("normalized_values", {})
    return {
        "candidate_id": option.get("candidate_id"),
        "candidate_option_ref": option_ref,
        "source_sheet": clean(source_ref.get("source_sheet")),
        "source_row_index": source_ref.get("source_row_index"),
        "source_row_span": source_ref.get("source_row_span"),
        "source_refs": option.get("source_refs", []),
        "raw_values": raw_values,
        "source_description_raw": clean(raw_values.get("source_description_raw") or normalized.get("source_description_raw")),
        "section_context_raw": clean(raw_values.get("section_context_raw") or normalized.get("section_context_raw")),
    }


def interpret_group(group: dict[str, Any], workbook_index: dict[str, Any]) -> dict[str, Any]:
    model_key = group["model_key"]
    rpo = group["rpo"]
    source_occurrences = sorted(group["occurrences"], key=lambda item: (item["source_sheet"], item.get("source_row_index") or 0))
    availability_matrix, status_conflicts = build_availability_matrix(group["ovs"])
    duplicate_classification = classify_duplicates(source_occurrences, availability_matrix, status_conflicts)
    workbook_identity_match = workbook_identity(model_key, rpo, workbook_index)
    workbook_status_match = workbook_status(model_key, workbook_identity_match, availability_matrix, workbook_index)
    categories, reason_codes = interpretation_categories(group, availability_matrix, duplicate_classification, workbook_identity_match, workbook_status_match, status_conflicts)
    confidence = confidence_for(workbook_identity_match, workbook_status_match, duplicate_classification, reason_codes)
    primary = primary_source_occurrence(source_occurrences, duplicate_classification)
    return {
        "interpretation_id": f"interpopt-{slug(model_key)}-{slug(rpo)}",
        "model_key": model_key,
        "rpo": rpo,
        "source_occurrences": source_occurrences,
        "primary_source_occurrence": primary,
        "duplicate_classification": duplicate_classification,
        "source_sheet_roles": source_sheet_roles(source_occurrences, primary, duplicate_classification),
        "availability_matrix": availability_matrix,
        "status_pattern_summary": status_pattern_summary(availability_matrix),
        "disclosure_evidence": disclosure_evidence(group["rules"], group["unresolved"]),
        "workbook_identity_match": workbook_identity_match,
        "workbook_status_match": workbook_status_match,
        "copy_comparison_status": "not_compared_by_design",
        "interpretation_confidence": confidence,
        "interpretation_categories": categories,
        "review_reason_codes": sorted(reason_codes),
        "expert_summary": expert_summary(confidence, duplicate_classification, workbook_identity_match, workbook_status_match, reason_codes),
    }


def build_availability_matrix(ovs_rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    by_variant: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in ovs_rows:
        normalized = row.get("normalized_values", {})
        variant_id = clean(normalized.get("variant_id"))
        if not variant_id:
            continue
        by_variant[variant_id].append({
            "candidate_id": row.get("candidate_id"),
            "normalized_status_candidate": clean(normalized.get("normalized_status_candidate")),
            "raw_status": clean(normalized.get("raw_status")),
            "status_marker": clean(normalized.get("status_marker")),
            "status_flags": normalized.get("status_flags", []),
            "source_cell": clean(normalized.get("source_cell")),
            "source_refs": row.get("source_refs", []),
        })
    matrix: dict[str, Any] = {}
    conflicts = []
    for variant_id, rows in sorted(by_variant.items()):
        statuses = sorted({row["normalized_status_candidate"] for row in rows})
        if len(statuses) > 1:
            conflicts.append(variant_id)
        raw_statuses = sorted({row["raw_status"] for row in rows})
        matrix[variant_id] = {
            "normalized_status": statuses[0] if len(statuses) == 1 else "conflict",
            "raw_statuses": raw_statuses,
            "status_markers": sorted({row["status_marker"] for row in rows if row["status_marker"]}),
            "status_flags": sorted({flag for row in rows for flag in row.get("status_flags", [])}),
            "source_cells": sorted({row["source_cell"] for row in rows if row["source_cell"]}),
            "source_refs": [ref for row in rows for ref in row.get("source_refs", [])],
            "candidate_ids": [row["candidate_id"] for row in rows],
        }
    return matrix, conflicts


def classify_duplicates(source_occurrences: list[dict[str, Any]], availability_matrix: dict[str, Any], status_conflicts: list[str]) -> str:
    if len(source_occurrences) <= 1:
        return "single_source"
    if status_conflicts or any(values.get("normalized_status") == "conflict" for values in availability_matrix.values()):
        return "conflicting_duplicates"
    descriptions = {clean(item.get("source_description_raw")) for item in source_occurrences}
    if len(descriptions) <= 1:
        return "redundant_duplicates"
    return "complementary_duplicates"


def workbook_identity(model_key: str, rpo: str, workbook_index: dict[str, Any]) -> dict[str, Any]:
    active_matches = workbook_index["by_model_rpo"].get(model_key, {}).get(rpo, [])
    if len(active_matches) == 1:
        match = active_matches[0]
        return {
            "match_status": "unique_rpo_match",
            "model_key": model_key,
            "option_id": match.get("option_id", ""),
            "rpo": match.get("rpo", ""),
            "section_id": match.get("section_id", ""),
            "sheet_name": match.get("sheet_name", ""),
        }
    if len(active_matches) > 1:
        return {
            "match_status": "duplicate_workbook_rpo",
            "matches": [match_payload(match) for match in active_matches],
        }
    inactive_matches = workbook_index["inactive_by_model_rpo"].get(model_key, {}).get(rpo, [])
    if inactive_matches:
        return {"match_status": "inactive_or_scaffold_match", "matches": [match_payload(match) for match in inactive_matches]}
    return {
        "match_status": "missing_in_workbook",
        "matches": [],
        "model_active": workbook_index["model_active"].get(model_key, False),
    }


def match_payload(match: dict[str, str]) -> dict[str, str]:
    return {
        "model_key": match.get("model_key", ""),
        "option_id": match.get("option_id", ""),
        "rpo": match.get("rpo", ""),
        "section_id": match.get("section_id", ""),
        "sheet_name": match.get("sheet_name", ""),
    }


def workbook_status(
    model_key: str,
    identity_match: dict[str, Any],
    availability_matrix: dict[str, Any],
    workbook_index: dict[str, Any],
) -> dict[str, Any]:
    if identity_match.get("match_status") != "unique_rpo_match":
        return {"match_status": "not_compared", "reason": identity_match.get("match_status")}
    option_id = clean(identity_match.get("option_id"))
    workbook_statuses = workbook_index["status_by_model_option"].get(model_key, {}).get(option_id, {})
    if not availability_matrix:
        return {"match_status": "no_source_status", "workbook_statuses": workbook_statuses}
    mismatches = []
    for variant_id, source_status in availability_matrix.items():
        expected = clean(workbook_statuses.get(variant_id))
        observed = clean(source_status.get("normalized_status"))
        if not expected or expected != observed:
            mismatches.append({"variant_id": variant_id, "source_status": observed, "workbook_status": expected})
    return {
        "match_status": "exact" if not mismatches else "mismatch",
        "mismatches": mismatches,
        "workbook_statuses": workbook_statuses,
    }


def interpretation_categories(
    group: dict[str, Any],
    availability_matrix: dict[str, Any],
    duplicate_classification: str,
    workbook_identity_match: dict[str, Any],
    workbook_status_match: dict[str, Any],
    status_conflicts: list[str],
) -> tuple[list[str], set[str]]:
    categories: set[str] = set()
    reasons: set[str] = set()
    raw_statuses = {raw for data in availability_matrix.values() for raw in data.get("raw_statuses", [])}
    markers = {marker for data in availability_matrix.values() for marker in data.get("status_markers", [])}
    if not markers and not group["rules"] and not status_conflicts:
        categories.add("plain_availability")
    if any("D" in raw or "A/D" in raw for raw in raw_statuses):
        categories.add("dealer_installed_or_adi")
        reasons.add("dealer_installed_or_adi")
    if any("■" in raw for raw in raw_statuses):
        categories.add("equipment_group_inclusion")
        reasons.add("equipment_group_inclusion")
    if any("□" in raw for raw in raw_statuses):
        categories.add("upgradeable_equipment_group_review")
        reasons.add("upgradeable_equipment_group_review")
    if markers:
        reasons.add("footnote_or_disclosure_marker")
    for rule in group["rules"]:
        hint = clean(rule.get("normalized_values", {}).get("relationship_hint"))
        if hint == "requires":
            categories.add("requires_relationship_hint")
        elif hint == "includes":
            categories.add("includes_relationship_hint")
        elif hint == "excludes":
            categories.add("excludes_relationship_hint")
        elif hint == "included_only_available_with":
            categories.add("included_only_available_with_hint")
        if hint:
            reasons.add("relationship_hint")
    if any(item.get("category") == "relationship_hint" for item in group["unresolved"]):
        reasons.add("unresolved_target_rpo")
    if duplicate_classification == "conflicting_duplicates":
        reasons.add("duplicate_conflict")
    elif duplicate_classification in {"complementary_duplicates", "blocked_duplicate_review"}:
        reasons.add("duplicate_review")
    if workbook_identity_match.get("match_status") == "missing_in_workbook":
        reasons.add("missing_workbook_match")
    elif workbook_identity_match.get("match_status") == "duplicate_workbook_rpo":
        reasons.add("duplicate_workbook_rpo")
    elif workbook_identity_match.get("match_status") in {"inactive_or_scaffold_match", "out_of_scope_model"}:
        reasons.add(workbook_identity_match.get("match_status", "workbook_identity_review"))
    if workbook_status_match.get("match_status") == "mismatch":
        reasons.add("status_matrix_changed")
    elif workbook_status_match.get("match_status") in {"no_source_status", "not_compared"} and workbook_identity_match.get("match_status") == "unique_rpo_match":
        reasons.add("ovs_comparison_not_trusted")
    return sorted(categories), reasons


def confidence_for(
    workbook_identity_match: dict[str, Any],
    workbook_status_match: dict[str, Any],
    duplicate_classification: str,
    reason_codes: set[str],
) -> str:
    if workbook_identity_match.get("match_status") == "duplicate_workbook_rpo" or duplicate_classification == "conflicting_duplicates":
        return "blocked"
    auto_blockers = {
        "dealer_installed_or_adi",
        "equipment_group_inclusion",
        "upgradeable_equipment_group_review",
        "footnote_or_disclosure_marker",
        "relationship_hint",
        "unresolved_target_rpo",
        "missing_workbook_match",
        "duplicate_workbook_rpo",
        "duplicate_conflict",
        "duplicate_review",
        "inactive_or_scaffold_match",
        "out_of_scope_model",
        "status_matrix_changed",
        "ovs_comparison_not_trusted",
    }
    if (
        workbook_identity_match.get("match_status") == "unique_rpo_match"
        and workbook_status_match.get("match_status") == "exact"
        and duplicate_classification in AUTO_ALLOWED_DUPLICATE_CLASSES
        and not (reason_codes & auto_blockers)
    ):
        return "auto_confirmed"
    if workbook_identity_match.get("match_status") == "missing_in_workbook" and not (
        reason_codes - {"missing_workbook_match"}
    ) and duplicate_classification in {"single_source", "redundant_duplicates", "complementary_duplicates"}:
        return "mechanical_safe"
    return "review_needed"


def primary_source_occurrence(source_occurrences: list[dict[str, Any]], duplicate_classification: str) -> dict[str, Any] | None:
    if duplicate_classification not in {"single_source", "redundant_duplicates", "complementary_duplicates"}:
        return None
    if not source_occurrences:
        return None
    return sorted(source_occurrences, key=lambda item: (source_priority(item.get("source_sheet", "")), item.get("source_row_index") or 0))[0]


def source_priority(sheet: str) -> int:
    lower = clean(sheet).lower()
    if lower.startswith("equipment groups"):
        return 0
    if lower.startswith("exterior"):
        return 1
    if lower.startswith("interior"):
        return 2
    if lower.startswith("mechanical"):
        return 3
    if lower.startswith("standard equipment"):
        return 4
    return 5


def source_sheet_roles(
    source_occurrences: list[dict[str, Any]],
    primary: dict[str, Any] | None,
    duplicate_classification: str,
) -> list[dict[str, str]]:
    primary_key = occurrence_key(primary) if primary else ""
    roles = []
    for occurrence in source_occurrences:
        sheet = clean(occurrence.get("source_sheet"))
        if occurrence_key(occurrence) == primary_key:
            role = "candidate_primary"
        elif duplicate_classification == "redundant_duplicates":
            role = "redundant_duplicate"
        elif duplicate_classification == "conflicting_duplicates":
            role = "conflict"
        elif sheet.lower().startswith("standard equipment"):
            role = "standard_equipment_context"
        elif sheet.lower().startswith("equipment groups"):
            role = "equipment_group_context"
        else:
            role = "section_specific_context"
        roles.append({"source_sheet": sheet, "source_row_index": occurrence.get("source_row_index"), "role": role})
    return roles


def occurrence_key(occurrence: dict[str, Any] | None) -> str:
    if not occurrence:
        return ""
    return f"{occurrence.get('source_sheet')}:{occurrence.get('source_row_index')}"


def status_pattern_summary(availability_matrix: dict[str, Any]) -> dict[str, int]:
    counts = Counter(data.get("normalized_status", "unresolved") for data in availability_matrix.values())
    return dict(sorted(counts.items()))


def disclosure_evidence(rules: list[dict[str, Any]], unresolved: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "rule_candidates": [
            {
                "candidate_id": rule.get("candidate_id"),
                "relationship_hint": clean(rule.get("normalized_values", {}).get("relationship_hint")),
                "target_match_status": clean(rule.get("normalized_values", {}).get("target_match_status")),
                "description_fragment": clean(rule.get("normalized_values", {}).get("description_fragment")),
                "source_refs": rule.get("source_refs", []),
            }
            for rule in rules
        ],
        "unresolved_items": [
            {
                "unresolved_id": item.get("unresolved_id"),
                "reason": item.get("reason"),
                "category": item.get("category"),
                "source_refs": item.get("source_refs", []),
            }
            for item in unresolved
        ],
    }


def expert_summary(
    confidence: str,
    duplicate_classification: str,
    workbook_identity_match: dict[str, Any],
    workbook_status_match: dict[str, Any],
    reason_codes: set[str],
) -> str:
    parts = [f"confidence={confidence}", f"duplicates={duplicate_classification}"]
    parts.append(f"workbook_identity={workbook_identity_match.get('match_status')}")
    parts.append(f"workbook_status={workbook_status_match.get('match_status')}")
    if reason_codes:
        parts.append("reasons=" + ",".join(sorted(reason_codes)))
    return "; ".join(parts)


def build_source_sheet_coverage(interpreted_options: list[dict[str, Any]]) -> dict[str, Any]:
    per_model: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    row_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for item in interpreted_options:
        model_key = item["model_key"]
        rpo = item["rpo"]
        for occurrence in item["source_occurrences"]:
            sheet = clean(occurrence.get("source_sheet"))
            per_model[model_key][sheet].add(rpo)
            row_counts[model_key][sheet] += 1
    result: dict[str, Any] = {}
    for model_key, sheet_map in sorted(per_model.items()):
        all_rpos = set().union(*sheet_map.values()) if sheet_map else set()
        sheet_rows = [
            {
                "source_sheet": sheet,
                "unique_rpo_count": len(rpos),
                "row_count": row_counts[model_key][sheet],
                "duplicate_ref_count": max(row_counts[model_key][sheet] - len(rpos), 0),
            }
            for sheet, rpos in sorted(sheet_map.items())
        ]
        best_sets = best_sheet_sets(sheet_map, all_rpos)
        result[model_key] = {
            "total_unique_rpos": len(all_rpos),
            "sheets": sorted(sheet_rows, key=lambda row: (-row["unique_rpo_count"], row["source_sheet"])),
            "best_sheet_sets": best_sets,
        }
    return result


def best_sheet_sets(sheet_map: dict[str, set[str]], all_rpos: set[str]) -> list[dict[str, Any]]:
    if not sheet_map:
        return []
    candidates = []
    sheets = sorted(sheet_map)
    for size in range(1, min(3, len(sheets)) + 1):
        for sheet_set in combinations(sheets, size):
            covered = set().union(*(sheet_map[sheet] for sheet in sheet_set))
            missing = sorted(all_rpos - covered)
            duplicate_refs = sum(len(sheet_map[sheet]) for sheet in sheet_set) - len(covered)
            candidates.append({
                "source_sheets": list(sheet_set),
                "unique_rpo_count": len(covered),
                "coverage_ratio": round(len(covered) / len(all_rpos), 4) if all_rpos else 0,
                "duplicate_ref_count": duplicate_refs,
                "missing_rpos": missing,
            })
    return sorted(candidates, key=lambda row: (-row["unique_rpo_count"], len(row["source_sheets"]), row["duplicate_ref_count"], row["source_sheets"]))[:10]


def build_duplicate_report(interpreted_options: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in interpreted_options:
        if len(item["source_occurrences"]) <= 1:
            continue
        rows.append({
            "model_key": item["model_key"],
            "rpo": item["rpo"],
            "duplicate_classification": item["duplicate_classification"],
            "source_occurrence_count": len(item["source_occurrences"]),
            "source_sheets": sorted({occurrence["source_sheet"] for occurrence in item["source_occurrences"]}),
            "source_occurrences": item["source_occurrences"],
            "interpretation_confidence": item["interpretation_confidence"],
            "review_reason_codes": item["review_reason_codes"],
        })
    return sorted(rows, key=lambda row: (row["model_key"], row["rpo"]))


def build_blocked_interpretation(interpreted_options: list[dict[str, Any]], unresolved_review: dict[str, Any]) -> dict[str, Any]:
    out_of_scope = [
        item for item in unresolved_review.get("items", [])
        if item.get("severity") == "out_of_scope"
    ]
    blocked_options = [item for item in interpreted_options if item["interpretation_confidence"] == "blocked"]
    return {"blocked_options": blocked_options, "out_of_scope_unresolved_items": out_of_scope}


def build_workbook_build_units(
    interpreted_options: list[dict[str, Any]],
    blocked_interpretation: dict[str, Any],
    selection_metadata: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    if not selection_metadata:
        return []
    primary = set(selection_metadata.get("primary_models") or [])
    comparator = set(selection_metadata.get("comparator_models") or [])
    units: list[dict[str, Any]] = []
    for item in interpreted_options:
        model_key = item["model_key"]
        model_role = "comparator" if model_key in comparator else "primary" if model_key in primary else "selected"
        presence = workbook_presence_for(item["workbook_identity_match"])
        base = {
            "model_key": model_key,
            "model_role": model_role,
            "rpo": item.get("rpo"),
            "source_sheets": sorted({occ.get("source_sheet") for occ in item.get("source_occurrences", []) if occ.get("source_sheet")}),
            "source_refs": [ref for occ in item.get("source_occurrences", []) for ref in occ.get("source_refs", [])],
            "raw_source_snapshot": (item.get("primary_source_occurrence") or {}).get("raw_values", {}),
            "status_matrix_summary": item.get("status_pattern_summary", {}),
            "relationship_hint_summary": item.get("disclosure_evidence", {}),
            "workbook_presence": presence,
            "required_fields_missing": required_fields_missing(item),
            "comparator_context": {"comparator_only": model_role == "comparator"},
        }
        option_sheet = f"{model_key}_options"
        ovs_sheet = f"{model_key}_ovs"
        units.append({
            **base,
            "review_unit_id": f"wb-option-{slug(model_key)}-{slug(item.get('rpo'))}",
            "lane": "option_rows",
            "target_sheet": option_sheet,
            "target_workbook_surface": option_sheet,
            "proposed_workbook_action": "create_option_row" if presence == "missing" else "verify_existing_option_row",
        })
        if item.get("availability_matrix"):
            units.append({
                **base,
                "review_unit_id": f"wb-ovs-{slug(model_key)}-{slug(item.get('rpo'))}",
                "lane": "ovs_rows",
                "target_sheet": ovs_sheet,
                "target_workbook_surface": ovs_sheet,
                "proposed_workbook_action": "create_ovs_rows" if presence == "missing" else "verify_status_matrix",
            })
        if item.get("disclosure_evidence", {}).get("rule_candidates") or item.get("disclosure_evidence", {}).get("unresolved_items"):
            units.append({
                **base,
                "review_unit_id": f"wb-rel-{slug(model_key)}-{slug(item.get('rpo'))}",
                "lane": "relationships",
                "target_sheet": f"{model_key}_rule_mapping",
                "target_workbook_surface": f"{model_key}_rule_mapping",
                "proposed_workbook_action": "create_relationship_candidate",
            })
        if item.get("duplicate_classification") != "single_source":
            units.append({
                **base,
                "review_unit_id": f"wb-dup-{slug(model_key)}-{slug(item.get('rpo'))}",
                "lane": "duplicates_and_source_coverage",
                "target_sheet": "source_sheet_coverage",
                "target_workbook_surface": "source_sheet_coverage",
                "proposed_workbook_action": "classify_duplicate_source",
            })
    for index, item in enumerate(blocked_interpretation.get("out_of_scope_unresolved_items", []), start=1):
        category = item.get("category") or "source_shape"
        action = "defer_price_extractor" if category == "price_out_of_scope" else "defer_color_trim_extractor" if category == "color_trim_out_of_scope" else "blocked_unsupported_source_structure"
        units.append({
            "review_unit_id": f"wb-blocked-{index:05d}-{slug(category)}",
            "lane": "blocked_extractor_gaps",
            "model_key": "",
            "model_role": "all_selected",
            "rpo": "",
            "target_sheet": "extractor_gap",
            "target_workbook_surface": category,
            "proposed_workbook_action": action,
            "workbook_presence": "not_applicable",
            "required_fields_missing": [],
            "source_sheets": sorted({ref.get("source_sheet") for ref in item.get("source_refs", []) if ref.get("source_sheet")}),
            "source_refs": item.get("source_refs", []),
            "raw_source_snapshot": item.get("raw_values", {}),
            "status_matrix_summary": {},
            "relationship_hint_summary": {},
            "comparator_context": {},
        })
    lane_order = {
        "option_rows": 0,
        "ovs_rows": 1,
        "relationships": 2,
        "pricing": 3,
        "duplicates_and_source_coverage": 4,
        "blocked_extractor_gaps": 5,
    }
    return sorted(units, key=lambda row: (lane_order.get(row["lane"], 99), row.get("model_key") or "", row.get("rpo") or "", row["review_unit_id"]))


def workbook_presence_for(identity_match: dict[str, Any]) -> str:
    status = identity_match.get("match_status")
    if status == "unique_rpo_match":
        return "existing_active"
    if status == "inactive_or_scaffold_match":
        return "existing_inactive_scaffold"
    if status == "duplicate_workbook_rpo":
        return "duplicate_existing"
    if status == "missing_in_workbook":
        return "missing"
    return "not_applicable"


def required_fields_missing(item: dict[str, Any]) -> list[str]:
    missing = []
    primary = item.get("primary_source_occurrence") or {}
    raw = primary.get("raw_values", {})
    if not item.get("rpo"):
        missing.append("rpo")
    if not clean(raw.get("source_description_raw")):
        missing.append("source_description_raw")
    if not item.get("availability_matrix"):
        missing.append("availability_matrix")
    return missing


def build_workbook_build_summary(
    *,
    run_id: str,
    selection_metadata: dict[str, Any],
    workbook_build_units: list[dict[str, Any]],
    output_dir: Path,
) -> dict[str, Any]:
    lane_counts = Counter(unit["lane"] for unit in workbook_build_units)
    model_counts = Counter(unit.get("model_key") or "all_selected" for unit in workbook_build_units)
    role_counts = Counter(unit.get("model_role") or "selected" for unit in workbook_build_units)
    artifact_fingerprints = {}
    units_path = output_dir / "workbook-build-review-units.json"
    if units_path.exists():
        artifact_fingerprints["workbook-build-review-units.json"] = sha256_file(units_path)
    selection_path = output_dir / "model-selection.json"
    if selection_path.exists():
        artifact_fingerprints["model-selection.json"] = sha256_file(selection_path)
    return {
        "version": 1,
        "review_mode": "focused_workbook_build",
        "run_id": run_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "selection_metadata": selection_metadata,
        "selection_fingerprint": selection_fingerprint(selection_metadata),
        "artifact_fingerprints": artifact_fingerprints,
        "lane_counts": {lane: lane_counts.get(lane, 0) for lane in ["option_rows", "ovs_rows", "relationships", "pricing", "duplicates_and_source_coverage", "blocked_extractor_gaps"]},
        "model_counts": dict(sorted(model_counts.items())),
        "model_role_counts": dict(sorted(role_counts.items())),
        "cross_check_status": {"ok": True, "errors": []},
    }


def sha256_file(path: Path) -> str:
    import hashlib

    return hashlib.sha256(path.read_bytes()).hexdigest()
 
 
def build_summary(
    *,
    run_id: str,
    evidence_dir: Path,
    candidates_dir: Path,
    workbook: Path,
    interpreted_options: list[dict[str, Any]],
    review_queue: list[dict[str, Any]],
    blocked_interpretation: dict[str, Any],
    duplicate_report: list[dict[str, Any]],
    source_sheet_coverage: dict[str, Any],
    candidate_summary: dict[str, Any],
    unresolved_review: dict[str, Any],
    selection_metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    candidate_counts = dict(candidate_summary.get("candidate_counts", {}))
    unresolved_count = len(unresolved_review.get("items", []))
    raw_candidate_counts = {
        "options": int(candidate_counts.get("options", 0)),
        "ovs": int(candidate_counts.get("ovs", 0)),
        "rules": int(candidate_counts.get("rules", 0)),
        "price_rules": int(candidate_counts.get("price_rules", 0)),
        "unresolved": unresolved_count,
    }
    raw_total = sum(raw_candidate_counts.values())
    confidence_counts = Counter(item["interpretation_confidence"] for item in interpreted_options)
    blocked_count = confidence_counts.get("blocked", 0) + len(blocked_interpretation["out_of_scope_unresolved_items"])
    visible_count = confidence_counts.get("mechanical_safe", 0) + confidence_counts.get("review_needed", 0)
    duplicate_rpo_count = len(duplicate_report)
    conflicting_duplicate_count = sum(1 for row in duplicate_report if row["duplicate_classification"] == "conflicting_duplicates")
    target_visible = raw_total * 0.3
    material = (visible_count + blocked_count) <= target_visible if raw_total else True
    reason_codes = [] if material else reduction_reason_codes(interpreted_options, duplicate_report)
    summary = {
        "version": 1,
        "run_id": run_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_evidence_dir": str(evidence_dir),
        "input_candidates_dir": str(candidates_dir),
        "workbook": str(workbook),
        "status": "passed",
        "raw_candidate_counts": raw_candidate_counts,
        "raw_candidate_total": raw_total,
        "interpreted_option_count": len(interpreted_options),
        "hidden_auto_confirmed_count": confidence_counts.get("auto_confirmed", 0),
        "visible_review_queue_count": visible_count,
        "mechanical_safe_count": confidence_counts.get("mechanical_safe", 0),
        "review_needed_count": confidence_counts.get("review_needed", 0),
        "blocked_count": blocked_count,
        "duplicate_rpo_count": duplicate_rpo_count,
        "conflicting_duplicate_count": conflicting_duplicate_count,
        "source_sheet_coverage": compact_coverage(source_sheet_coverage),
        "reduction_status": "material_reduction" if material else "insufficient_reduction",
        "reduction_reason_codes": reason_codes,
        "reduction_components": {
            "aggregation_reduction": raw_total - len(interpreted_options),
            "auto_confirmed_hidden": confidence_counts.get("auto_confirmed", 0),
            "visible_plus_blocked": visible_count + blocked_count,
            "target_visible_plus_blocked_max": int(target_visible),
        },
        "artifact_files": OUTPUT_FILES,
    }
    if selection_metadata:
        summary["selection_metadata"] = selection_metadata
        summary["selection_fingerprint"] = selection_fingerprint(selection_metadata)
    return summary


def reduction_reason_codes(interpreted_options: list[dict[str, Any]], duplicate_report: list[dict[str, Any]]) -> list[str]:
    reasons = []
    if any(row["duplicate_classification"] in {"conflicting_duplicates", "blocked_duplicate_review"} for row in duplicate_report):
        reasons.append("too_many_conflicts")
    if any("footnote_or_disclosure_marker" in item["review_reason_codes"] for item in interpreted_options):
        reasons.append("too_many_footnotes")
    if any("ovs_comparison_not_trusted" in item["review_reason_codes"] for item in interpreted_options):
        reasons.append("ovs_comparison_not_trusted")
    if any(item["workbook_identity_match"].get("match_status") in {"duplicate_workbook_rpo", "inactive_or_scaffold_match"} for item in interpreted_options):
        reasons.append("workbook_match_too_ambiguous")
    if not reasons:
        reasons.append("duplicate_classification_too_weak")
    return sorted(set(reasons))


def compact_coverage(source_sheet_coverage: dict[str, Any]) -> dict[str, Any]:
    return {
        model_key: {
            "total_unique_rpos": values["total_unique_rpos"],
            "top_sheets": values["sheets"][:5],
            "best_sheet_set": values["best_sheet_sets"][0] if values["best_sheet_sets"] else None,
        }
        for model_key, values in source_sheet_coverage.items()
    }


def render_duplicate_report(rows: list[dict[str, Any]]) -> str:
    lines = ["# Pass 3 duplicate RPO report", ""]
    if not rows:
        lines.append("No duplicate source RPO rows found.")
        return "\n".join(lines)
    for row in rows:
        lines.extend([
            f"## {row['model_key']} {row['rpo']}",
            "",
            f"- Classification: {row['duplicate_classification']}",
            f"- Confidence: {row['interpretation_confidence']}",
            f"- Source sheets: {', '.join(row['source_sheets'])}",
            f"- Review reasons: {', '.join(row['review_reason_codes']) or 'none'}",
            "",
        ])
        for occurrence in row["source_occurrences"]:
            lines.append(f"  - {occurrence['source_sheet']} row {occurrence['source_row_index']}: {occurrence['source_description_raw']}")
        lines.append("")
    return "\n".join(lines).rstrip()


def render_source_sheet_coverage(coverage: dict[str, Any]) -> str:
    lines = ["# Pass 3 source-sheet coverage", ""]
    for model_key, values in coverage.items():
        lines.extend([f"## {model_key}", "", f"- Total unique RPOs: {values['total_unique_rpos']}", "", "### Top sheets", ""])
        for sheet in values["sheets"][:10]:
            lines.append(
                f"- {sheet['source_sheet']}: unique_rpos={sheet['unique_rpo_count']}, rows={sheet['row_count']}, duplicate_refs={sheet['duplicate_ref_count']}"
            )
        lines.extend(["", "### Best sheet sets", ""])
        for sheet_set in values["best_sheet_sets"][:5]:
            lines.append(
                f"- {', '.join(sheet_set['source_sheets'])}: unique_rpos={sheet_set['unique_rpo_count']}, coverage={sheet_set['coverage_ratio']}, duplicate_refs={sheet_set['duplicate_ref_count']}, missing={len(sheet_set['missing_rpos'])}"
            )
        lines.append("")
    return "\n".join(lines).rstrip()


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")


def slug(value: Any) -> str:
    text = clean(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "blank"
