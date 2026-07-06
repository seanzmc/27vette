#!/usr/bin/env python3
"""Model scoping and decision capture for the ingest wizard (Pass B).

The script owns structure: model-code scoping, deterministic fingerprints,
lane bookkeeping, template prefill from workbook-authored rows, and the
completeness math. The reviewer owns every business decision. Nothing here
writes the canonical workbook; it is opened read-only for pickers, variant
reconciliation, and presentation-template prefill.
"""

from __future__ import annotations

import hashlib
import json
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.workbook import rows_from_sheet, workbook_truthy

SCHEMA_VERSION_B = "pass-b-1"
# Additive revision (batchId on decision records); loaders accept both.
SCHEMA_VERSION_B2 = "pass-b-2"

# Model-code prefix -> canonical model_key. Derived from the 2027 export
# Price Schedule / variant headers (1YG confirmed = Grand Sport X) and the
# workbook model rows; a prefix missing here surfaces as "unknown", never a
# guessed model.
MODEL_CODE_PREFIXES: dict[str, str] = {
    "1YC": "stingray",
    "1YE": "grand_sport",
    "1YG": "grand_sport_x",
    "1YH": "z06",
    "1YR": "zr1",
    "1YS": "zr1x",
}

MODEL_LABELS: dict[str, str] = {
    "stingray": "Stingray",
    "grand_sport": "Grand Sport",
    "grand_sport_x": "Grand Sport X",
    "z06": "Z06",
    "zr1": "ZR1",
    "zr1x": "ZR1X",
}

# Resolved product decision 2026-07-05: comparators are structure-check only.
TARGET_MODELS = ("grand_sport_x", "zr1", "zr1x")
COMPARATOR_DEFAULTS: dict[str, str] = {
    "grand_sport_x": "grand_sport",
    "zr1": "z06",
    "zr1x": "z06",
}

PRESENTATION_SHEETS = (
    "runtime_steps",
    "section_presentation",
    "context_section_master",
    "order_summary_sections",
    "step_order_summary_map",
)

LANES: tuple[dict[str, Any], ...] = (
    {"lane": "section", "label": "Section assignment", "perCandidate": True},
    {"lane": "price", "label": "Price resolution", "perCandidate": True},
    {"lane": "exclusive_group", "label": "Exclusive groups", "perCandidate": False},
    {"lane": "relationship", "label": "Rule groups / relationships", "perCandidate": False},
    {"lane": "copy_split", "label": "Copy split", "perCandidate": True},
    {"lane": "status_nuance", "label": "Status nuances", "perCandidate": True},
    {"lane": "duplicate", "label": "Duplicates / cross-sheet", "perCandidate": False},
    {"lane": "standard_equipment", "label": "Standard equipment", "perCandidate": True},
    {"lane": "interior_media_deferral", "label": "Interiors / colors / media", "perCandidate": False},
    {"lane": "presentation", "label": "Presentation metadata", "perCandidate": False},
)
LANE_KEYS = {lane["lane"] for lane in LANES}

RESOLUTIONS = ("approved_for_plan", "hold_for_question", "not_needed")

# Lanes whose resolution is mandatory per in-scope orderable candidate before
# decisions_complete (spec B3). status_nuance is only mandatory for candidates
# that carry flagged/unresolved statuses.
MANDATORY_CANDIDATE_LANES = ("section", "price")


def canonical_json(payload: Any) -> str:
    return json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"))


def candidate_fingerprint(candidate: dict[str, Any]) -> str:
    evidence = candidate.get("sourceEvidence") or {}
    return hashlib.sha256(canonical_json(evidence).encode("utf-8")).hexdigest()


def artifact_fingerprint(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def status_needs_review(status: dict[str, Any]) -> bool:
    """True when an availability symbol needs a human read: unknown symbols,
    □ upgradeable-group rows, unresolved parses, and D / A/D dealer-install
    nuance. Plain footnote digits (A1/S2) are the copy-split lane's job."""

    flags = status.get("flags") or []
    if "unknown_status_symbol" in flags or "upgradeable_equipment_group_review" in flags:
        return True
    if status.get("status") == "unresolved":
        return True
    base = str(status.get("raw") or "").rstrip("0123456789")
    return base in ("D", "A/D")


def candidate_needs_status_review(candidate: dict[str, Any]) -> bool:
    return any(status_needs_review(status) for status in candidate.get("statuses", []))


def candidate_model_keys(candidate: dict[str, Any]) -> list[str]:
    """Model keys a candidate can belong to, from its status model codes."""

    keys: list[str] = []
    for status in candidate.get("statuses", []):
        prefix = str(status.get("modelCode") or "")[:3]
        key = MODEL_CODE_PREFIXES.get(prefix)
        if key and key not in keys:
            keys.append(key)
    return keys


def model_scoped_statuses(candidate: dict[str, Any], model_key: str) -> list[dict[str, Any]]:
    """Only the status cells belonging to `model_key`'s own variant columns.

    Mixed sheets carry every model's columns; per-model reads (like the
    standard-equipment gate) must not judge a row by another model's cells.
    """

    return [
        status
        for status in candidate.get("statuses", [])
        if MODEL_CODE_PREFIXES.get(str(status.get("modelCode") or "")[:3]) == model_key
    ]


def candidate_is_availability_row(candidate: dict[str, Any], model_key: str) -> bool:
    """True when any of the model's own statuses parse as orderable-available
    (A / A-with-footnote / A/D). Such rows are options, never standard
    equipment (field note 16, 2026-07-06)."""

    return any(
        status.get("status") == "available" for status in model_scoped_statuses(candidate, model_key)
    )


def scope_candidates(candidates: list[dict[str, Any]], model_key: str) -> list[dict[str, Any]]:
    return [c for c in candidates if model_key in candidate_model_keys(c)]


def detect_model_options(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Detected model families with candidate counts, ordered targets first."""

    counts: dict[str, int] = {}
    for candidate in candidates:
        for key in candidate_model_keys(candidate):
            counts[key] = counts.get(key, 0) + 1
    options = []
    for key in sorted(counts, key=lambda k: (k not in TARGET_MODELS, k)):
        options.append(
            {
                "modelKey": key,
                "label": MODEL_LABELS.get(key, key),
                "candidateCount": counts[key],
                "isDefaultTarget": key in TARGET_MODELS,
                "comparatorDefault": COMPARATOR_DEFAULTS.get(key, ""),
            }
        )
    return options


def validate_selection(
    candidates: list[dict[str, Any]],
    targets: list[str],
    comparators: dict[str, str],
) -> None:
    detected = {option["modelKey"] for option in detect_model_options(candidates)}
    if not targets:
        raise ValueError("Select at least one target model.")
    for target in targets:
        if target not in MODEL_LABELS:
            raise ValueError(f"Unknown target model: {target}")
        if target not in detected:
            raise ValueError(f"No candidates detected for target model: {target}")
    for target, comparator in comparators.items():
        if target not in targets:
            raise ValueError(f"Comparator set for non-target model: {target}")
        if comparator and comparator not in MODEL_LABELS:
            raise ValueError(f"Unknown comparator model: {comparator}")
        if comparator in targets:
            raise ValueError(f"Comparator may not also be a target: {comparator}")


# --------------------------------------------------------------- workbook
def workbook_option_reference(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Read-only index of live model option rows keyed by RPO.

    The reviewer compares candidates against data that already exists in the
    workbook (live `*_options` rows via active `model_workbook_sources`
    roles), never against another ingested sheet family.
    """

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        section_names = {
            row.get("section_id", ""): row.get("section_name", "")
            for row in rows_from_sheet(wb, "section_master")
        }
        source_rows = [
            row
            for row in rows_from_sheet(wb, "model_workbook_sources")
            if row.get("source_role") == "source_option_sheet" and workbook_truthy(row.get("active", ""))
        ]
        by_rpo: dict[str, list[dict[str, Any]]] = {}
        for source in source_rows:
            sheet_name = source.get("sheet_name", "")
            model_key = source.get("model_key", "")
            if not sheet_name or sheet_name not in wb.sheetnames:
                continue
            for row in rows_from_sheet(wb, sheet_name):
                rpo = row.get("rpo", "").upper()
                if not rpo:
                    continue
                by_rpo.setdefault(rpo, []).append(
                    {
                        "modelKey": model_key,
                        "optionName": row.get("option_name", ""),
                        "sectionId": row.get("section_id", ""),
                        "sectionName": section_names.get(row.get("section_id", ""), ""),
                        "price": row.get("price", ""),
                        "description": row.get("description", ""),
                    }
                )
        return by_rpo
    finally:
        wb.close()


def workbook_sections(workbook_path: Path) -> list[dict[str, Any]]:
    """Read-only section picker feed: section_master joined to step keys."""

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows = rows_from_sheet(wb, "section_master")
    finally:
        wb.close()
    sections = []
    for row in rows:
        if not row.get("section_id"):
            continue
        sections.append(
            {
                "sectionId": row.get("section_id", ""),
                "sectionName": row.get("section_name", ""),
                "stepKey": row.get("step_key", ""),
                "selectionMode": row.get("selection_mode", ""),
                "standardBehavior": row.get("standard_behavior", ""),
            }
        )
    return sections


def duplicate_rpo_groups(scoped: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """In-file RPO collisions: the same orderable RPO on >1 source sheet
    within the selected model scope. Workbook comparison is the reference
    line's job, not this lane's."""

    by_rpo: dict[str, list[dict[str, Any]]] = {}
    for candidate in scoped:
        if candidate["rowKind"] == "orderable" and candidate["rpo"]:
            by_rpo.setdefault(candidate["rpo"], []).append(candidate)
    groups = []
    for rpo in sorted(by_rpo):
        occurrences = by_rpo[rpo]
        sheets = {c["sheetName"] for c in occurrences}
        if len(occurrences) > 1 and len(sheets) > 1:
            groups.append(
                {
                    "rpo": rpo,
                    "candidateIds": [c["candidateId"] for c in occurrences],
                    "sheets": sorted(sheets),
                }
            )
    return groups


def variant_reconciliation(
    workbook_path: Path,
    candidates: list[dict[str, Any]],
    targets: list[str],
) -> dict[str, Any]:
    """Export variant headers vs workbook variant scaffolds, per target."""

    export_variants: dict[str, dict[str, dict[str, str]]] = {key: {} for key in targets}
    for candidate in candidates:
        for status in candidate.get("statuses", []):
            code = str(status.get("modelCode") or "")
            key = MODEL_CODE_PREFIXES.get(code[:3])
            if key in export_variants:
                export_variants[key][f"{code}:{status.get('trim', '')}"] = {
                    "modelCode": code,
                    "trim": str(status.get("trim") or ""),
                    "bodyStyle": str(status.get("bodyStyle") or ""),
                    "label": str(status.get("variantLabel") or ""),
                }
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        variant_master = rows_from_sheet(wb, "variant_master")
        model_variants = rows_from_sheet(wb, "model_variants")
    finally:
        wb.close()
    variants_by_model: dict[str, list[dict[str, str]]] = {}
    master_by_id = {row.get("variant_id", ""): row for row in variant_master}
    for row in model_variants:
        variants_by_model.setdefault(row.get("model_key", ""), []).append(row)

    models: dict[str, Any] = {}
    for target in targets:
        workbook_rows = []
        for row in variants_by_model.get(target, []):
            master = master_by_id.get(row.get("variant_id", ""), {})
            workbook_rows.append(
                {
                    "variantId": row.get("variant_id", ""),
                    "trim": master.get("trim_level", ""),
                    "bodyStyle": master.get("body_style", ""),
                    "active": workbook_truthy(row.get("active", "")),
                }
            )
        export_rows = sorted(export_variants[target].values(), key=lambda v: (v["modelCode"], v["trim"]))
        matches, export_only, workbook_only = _reconcile(export_rows, workbook_rows)
        models[target] = {
            "exportVariants": export_rows,
            "workbookVariants": workbook_rows,
            "matched": matches,
            "exportOnly": export_only,
            "workbookOnly": workbook_only,
            "agrees": not export_only and not workbook_only,
            # New models have no model_variants scaffold yet; the UI frames
            # that as "confirm the export's variants", not as an error.
            "workbookHasScaffold": bool(workbook_rows),
        }
    return {"schemaVersion": SCHEMA_VERSION_B, "models": models}


def _reconcile(
    export_rows: list[dict[str, str]],
    workbook_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, Any]]]:
    def export_key(row: dict[str, str]) -> str:
        body = "07" if "coupe" in row["bodyStyle"].lower() else "67" if row["bodyStyle"] else row["modelCode"][3:]
        return f"{row['trim'].lower()}:{body}"

    def workbook_key(row: dict[str, Any]) -> str:
        body = "07" if row["bodyStyle"].lower() == "coupe" else "67"
        return f"{row['trim'].lower()}:{body}"

    workbook_by_key = {workbook_key(row): row for row in workbook_rows}
    matched, export_only = [], []
    seen = set()
    for row in export_rows:
        key = export_key(row)
        if key in workbook_by_key:
            matched.append(row)
            seen.add(key)
        else:
            export_only.append(row)
    workbook_only = [row for key, row in workbook_by_key.items() if key not in seen]
    return matched, export_only, workbook_only


def presentation_prefill(workbook_path: Path, template_model: str, target_model: str) -> dict[str, Any]:
    """Template-derived proposal rows for the five presentation sheets.

    Copies workbook-authored rows for `template_model`, swaps the model key to
    the target, and records provenance. Proposals only — the reviewer approves,
    edits, or deletes every row in the presentation lane.
    """

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    proposals: dict[str, Any] = {}
    try:
        for sheet in PRESENTATION_SHEETS:
            rows = rows_from_sheet(wb, sheet)
            sheet_rows = []
            for row in rows:
                if row.get("model_key", "") != template_model:
                    continue
                proposal = dict(row)
                proposal["model_key"] = target_model
                sheet_rows.append(
                    {
                        "row": proposal,
                        "provenance": {"templateModel": template_model, "sheet": sheet},
                    }
                )
            proposals[sheet] = sheet_rows
    finally:
        wb.close()
    return {
        "schemaVersion": SCHEMA_VERSION_B,
        "templateModel": template_model,
        "targetModel": target_model,
        "sheets": proposals,
    }


# ------------------------------------------------------------------- copy
def copy_decisions(
    decisions: dict[str, dict[str, Any]],
    candidates: list[dict[str, Any]],
    from_model: str,
    to_model: str,
    *,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Reviewer-triggered copy of one model's decisions onto another.

    Same-candidate copy when the candidate is in scope of both models (shared
    mixed sheets); otherwise RPO-identity matching with an exactly-one-match
    requirement — ambiguous or missing matches are reported, never guessed.
    Existing target decisions are skipped unless `overwrite`. Copied records
    carry `copiedFrom` provenance and presentation rows get their `model_key`
    swapped. Returns new records plus a skip report; the caller persists.
    """

    if from_model == to_model:
        raise ValueError("Copy source and target must differ.")
    to_scope = {c["candidateId"]: c for c in scope_candidates(candidates, to_model)}
    by_identity: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for candidate in to_scope.values():
        key = (candidate["rpo"], candidate["refOnlyRpo"], candidate["rowKind"])
        by_identity.setdefault(key, []).append(candidate)
    from_candidates = {c["candidateId"]: c for c in scope_candidates(candidates, from_model)}

    copied: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    produced_ids: set[str] = set()
    for record in decisions.values():
        if record["model"] != from_model:
            continue
        lane = record["lane"]
        new_record = dict(record)
        new_record["model"] = to_model
        new_record["copiedFrom"] = from_model
        new_record["copiedAt"] = time.strftime("%Y-%m-%dT%H:%M:%S")
        candidate_id = record.get("candidateId")
        if candidate_id:
            target_candidate = None
            if candidate_id in to_scope:
                target_candidate = to_scope[candidate_id]
            else:
                source_candidate = from_candidates.get(candidate_id)
                if source_candidate is None:
                    skipped.append({"decisionId": record["decisionId"], "reason": "source_candidate_missing"})
                    continue
                key = (
                    source_candidate["rpo"],
                    source_candidate["refOnlyRpo"],
                    source_candidate["rowKind"],
                )
                matches = by_identity.get(key, [])
                if len(matches) == 1:
                    target_candidate = matches[0]
                else:
                    skipped.append(
                        {
                            "decisionId": record["decisionId"],
                            "reason": "no_unique_rpo_match",
                            "matches": len(matches),
                        }
                    )
                    continue
            new_record["candidateId"] = target_candidate["candidateId"]
            new_record["candidateFingerprint"] = candidate_fingerprint(target_candidate)
            new_record["decisionId"] = f"{to_model}:{lane}:{target_candidate['candidateId']}"
        else:
            new_record["decisionId"] = f"{to_model}:{lane}:{record['groupKey']}"
            if lane == "presentation" and isinstance(new_record.get("payload"), dict):
                payload = dict(new_record["payload"])
                payload["rows"] = [
                    {**row, "model_key": to_model} if isinstance(row, dict) and "model_key" in row else row
                    for row in payload.get("rows", [])
                ]
                new_record["payload"] = payload
        if not overwrite and new_record["decisionId"] in decisions:
            skipped.append({"decisionId": record["decisionId"], "reason": "target_already_decided"})
            continue
        # Two source candidates can share an RPO identity and resolve to the
        # same target candidate within one copy call — first wins, rest report.
        if new_record["decisionId"] in produced_ids:
            skipped.append({"decisionId": record["decisionId"], "reason": "duplicate_target_in_batch"})
            continue
        produced_ids.add(new_record["decisionId"])
        copied.append(new_record)

    by_lane: dict[str, int] = {}
    for record in copied:
        by_lane[record["lane"]] = by_lane.get(record["lane"], 0) + 1
    return {"copied": copied, "skipped": skipped, "copiedByLane": by_lane}


# --------------------------------------------------------------- decisions
def validate_decision(decision: dict[str, Any], candidates_by_id: dict[str, dict[str, Any]]) -> dict[str, Any]:
    model = str(decision.get("model") or "")
    lane = str(decision.get("lane") or "")
    resolution = str(decision.get("resolution") or "")
    if model not in MODEL_LABELS:
        raise ValueError(f"Decision has unknown model: {model!r}")
    if lane not in LANE_KEYS:
        raise ValueError(f"Decision has unknown lane: {lane!r}")
    if resolution not in RESOLUTIONS:
        raise ValueError(f"Decision has unknown resolution: {resolution!r}")
    candidate_id = str(decision.get("candidateId") or "")
    lane_config = next(l for l in LANES if l["lane"] == lane)
    record: dict[str, Any] = {
        "model": model,
        "lane": lane,
        "resolution": resolution,
        "action": str(decision.get("action") or ""),
        "payload": decision.get("payload") if isinstance(decision.get("payload"), (dict, list)) else {},
        "reviewerNote": str(decision.get("reviewerNote") or ""),
        "decidedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
    }
    if lane_config["perCandidate"]:
        if not candidate_id:
            raise ValueError(f"Lane {lane} decisions must name a candidateId.")
        candidate = candidates_by_id.get(candidate_id)
        if candidate is None:
            raise ValueError(f"Decision references unknown candidate: {candidate_id}")
        if model not in candidate_model_keys(candidate):
            raise ValueError(f"Candidate {candidate_id} is not in scope for model {model}.")
        record["candidateId"] = candidate_id
        record["candidateFingerprint"] = candidate_fingerprint(candidate)
        record["decisionId"] = f"{model}:{lane}:{candidate_id}"
    else:
        group_key = str(decision.get("groupKey") or "")
        if not group_key:
            raise ValueError(f"Lane {lane} decisions must name a groupKey.")
        record["groupKey"] = group_key
        record["decisionId"] = f"{model}:{lane}:{group_key}"
    return record


def load_decision_state(
    snapshot: dict[str, Any] | None,
    candidates_by_id: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Current decisions keyed by decisionId, with fingerprint invalidation."""

    decisions: dict[str, dict[str, Any]] = {}
    invalidated: list[dict[str, Any]] = []
    for record in (snapshot or {}).get("decisions", []):
        candidate_id = record.get("candidateId")
        if candidate_id:
            candidate = candidates_by_id.get(candidate_id)
            if candidate is None or candidate_fingerprint(candidate) != record.get("candidateFingerprint"):
                invalidated.append(record)
                continue
        decisions[record["decisionId"]] = record
    return {"decisions": decisions, "invalidated": invalidated}


# Group key reviewers must decide (any group lane, normally `relationship`
# with action `needs_product_decision`) when a model's export variants
# disagree with its workbook scaffold. Spec B1: disagreements are mandatory
# blocking decisions, not just surfaced report rows.
VARIANT_RECONCILIATION_KEY = "variant_reconciliation"


def completeness(
    candidates: list[dict[str, Any]],
    targets: list[str],
    decisions: dict[str, dict[str, Any]],
    reconciliation: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Spec B3: per-model, per-lane progress plus the blocking report."""

    reconciliation_models = (reconciliation or {}).get("models", {})
    models: dict[str, Any] = {}
    for target in targets:
        scoped = [c for c in scope_candidates(candidates, target) if c["rowKind"] == "orderable"]
        # A skipped section decision (resolution not_needed) means the row is
        # not carried into the plan — no price or status read is owed for it.
        skipped_ids = {
            d.get("candidateId")
            for d in decisions.values()
            if d["model"] == target and d["lane"] == "section" and d["resolution"] == "not_needed"
        }
        blockers: list[dict[str, Any]] = []
        holds_report: list[dict[str, Any]] = []
        lanes: dict[str, Any] = {}
        for lane_config in LANES:
            lane = lane_config["lane"]
            lane_decisions = [d for d in decisions.values() if d["model"] == target and d["lane"] == lane]
            holds = [d for d in lane_decisions if d["resolution"] == "hold_for_question"]
            entry: dict[str, Any] = {"decisions": len(lane_decisions), "holds": len(holds)}
            if lane in MANDATORY_CANDIDATE_LANES:
                required = (
                    scoped if lane == "section" else [c for c in scoped if c["candidateId"] not in skipped_ids]
                )
                decided_ids = {d.get("candidateId") for d in lane_decisions}
                missing = [c["candidateId"] for c in required if c["candidateId"] not in decided_ids]
                entry["required"] = len(required)
                entry["missing"] = len(missing)
                for candidate_id in missing:
                    blockers.append({"lane": lane, "candidateId": candidate_id, "reason": "no_decision"})
            if lane == "status_nuance":
                flagged = [
                    c
                    for c in scoped
                    if candidate_needs_status_review(c) and c["candidateId"] not in skipped_ids
                ]
                decided_ids = {d.get("candidateId") for d in lane_decisions}
                missing = [c["candidateId"] for c in flagged if c["candidateId"] not in decided_ids]
                entry["required"] = len(flagged)
                entry["missing"] = len(missing)
                for candidate_id in missing:
                    blockers.append({"lane": lane, "candidateId": candidate_id, "reason": "flagged_status_undecided"})
            if lane == "presentation":
                approved_sheets = {
                    d["groupKey"]
                    for d in lane_decisions
                    if d["resolution"] == "approved_for_plan" and d.get("groupKey") in PRESENTATION_SHEETS
                }
                missing_sheets = [s for s in PRESENTATION_SHEETS if s not in approved_sheets]
                entry["required"] = len(PRESENTATION_SHEETS)
                entry["missing"] = len(missing_sheets)
                for sheet in missing_sheets:
                    blockers.append({"lane": lane, "groupKey": sheet, "reason": "presentation_sheet_unapproved"})
            # Holds satisfy the per-candidate requirement (explicit hold/defer)
            # but every hold stays enumerated for the blocking report readers.
            for hold in holds:
                holds_report.append(
                    {"lane": lane, "candidateId": hold.get("candidateId"), "groupKey": hold.get("groupKey"), "note": hold.get("reviewerNote", "")}
                )
            lanes[lane] = entry
        reconciliation_entry = reconciliation_models.get(target)
        if reconciliation_entry and not reconciliation_entry.get("agrees", True):
            decided = any(
                record["model"] == target and record.get("groupKey") == VARIANT_RECONCILIATION_KEY
                for record in decisions.values()
            )
            if not decided:
                blockers.append(
                    {
                        "lane": "relationship",
                        "groupKey": VARIANT_RECONCILIATION_KEY,
                        "reason": "variant_reconciliation_disagreement_undecided",
                    }
                )
        models[target] = {
            "inScopeOrderable": len(scoped),
            "lanes": lanes,
            "blockers": blockers,
            "holds": holds_report,
            "complete": not blockers,
        }
    return {
        "schemaVersion": SCHEMA_VERSION_B,
        "models": models,
        "allComplete": all(entry["complete"] for entry in models.values()) if models else False,
    }
