#!/usr/bin/env python3
"""Sheet profiling for the interactive ingest wizard (Pass A, step 2).

Turns a raw GM order-guide export into friendly "sheet cards": detected sheet
type, model family, variant columns, row/header stats, confidence, and a
recommended role. Detection is structure-derived only — no product guessing
(core principle in docs/ingest/pass-a/interactive-ingest-wizard-pass-a-spec.md).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from corvette_form_generator.ingest.source_profiler import (
    ORDER_GUIDE_BASE_HEADERS,
    RPO_RE,
    STATUS_RE,
)
from corvette_form_generator.workbook import clean

SCHEMA_VERSION = "pass-a-1"
SHEET_TYPE_OPTIONS = "options_matrix"
SHEET_TYPE_PRICE = "price_sheet"
SHEET_TYPE_UNSUPPORTED = "unsupported"
SUBTYPE_ORDERABLE = "orderable_options"
SUBTYPE_STANDARD = "standard_equipment"
ROLE_OPTIONS = "options"
ROLE_PRICE = "price"
ROLE_EXCLUDE = "exclude"
PRICE_SECTION_BASE = "Base Model Prices"
PRICE_SECTION_OPTIONS = "Additional Options"
MODEL_CODE_RE = re.compile(r"^1Y[A-Z]\d{2}$")
# Observed 2027 export: Standard Equipment sheets have S-share >= 0.80, every
# other options matrix <= 0.47 (spec "Diagnosis"). 0.60 splits with margin.
STANDARD_EQUIPMENT_S_SHARE = 0.60
CANONICAL_OPTION_SHEET_RE = re.compile(r"^(Interior|Exterior|Mechanical)\s+\d+$", re.IGNORECASE)


def canonical_option_sheet_eligible(sheet_name: str) -> bool:
    """Return whether a raw matrix is an approved canonical option source."""

    return bool(CANONICAL_OPTION_SHEET_RE.fullmatch(str(sheet_name or "").strip()))


def profile_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    try:
        sheets = [profile_sheet(ws.title, sheet_values(ws)) for ws in wb.worksheets]
    finally:
        wb.close()
    return {"schemaVersion": SCHEMA_VERSION, "sourceFile": Path(path).name, "sheets": sheets}


def sheet_values(ws) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def profile_sheet(sheet_name: str, values: list[list[Any]]) -> dict[str, Any]:
    header_row = find_matrix_header_row(values)
    if header_row is not None:
        return matrix_card(sheet_name, values, header_row)
    if find_price_sections(values):
        return price_card(sheet_name, values)
    return unsupported_card(sheet_name)


def find_matrix_header_row(values: list[list[Any]]) -> int | None:
    """Return the 1-based header row carrying the order-guide base headers."""
    for index, row in enumerate(values[:10], start=1):
        cleaned = {clean(value) for value in row}
        if set(ORDER_GUIDE_BASE_HEADERS).issubset(cleaned):
            return index
    return None


def find_price_sections(values: list[list[Any]]) -> dict[str, int]:
    """Map price-schedule section labels to their 1-based row indexes."""
    sections: dict[str, int] = {}
    for index, row in enumerate(values, start=1):
        first = clean(row[0]) if row else ""
        if first in (PRICE_SECTION_BASE, PRICE_SECTION_OPTIONS) and first not in sections:
            sections[first] = index
    return sections if PRICE_SECTION_OPTIONS in sections else {}


def cell_text(row: list[Any], index: int) -> str:
    return clean(row[index]) if index < len(row) else ""


def matrix_card(sheet_name: str, values: list[list[Any]], header_row: int) -> dict[str, Any]:
    headers = values[header_row - 1]
    description_index = next(
        index for index, value in enumerate(headers) if clean(value) == "Description"
    )
    variant_columns = [
        parse_variant_header(index, clean(value))
        for index, value in enumerate(headers)
        if index > description_index and clean(value)
    ]
    stats, vocabulary = matrix_row_stats(values, header_row, variant_columns)
    family = detect_model_family(values[0][0] if values and values[0] else "")
    subtype = (
        SUBTYPE_STANDARD
        if stats["statusCells"] and stats["standardShare"] >= STANDARD_EQUIPMENT_S_SHARE
        else SUBTYPE_ORDERABLE
    )
    confidence, reasons = matrix_confidence(variant_columns, vocabulary)
    canonical_option_source = canonical_option_sheet_eligible(sheet_name)
    if not canonical_option_source:
        role = ROLE_EXCLUDE
        reason = "Canonical processing uses only Interior, Exterior, and Mechanical sheets."
    elif subtype == SUBTYPE_STANDARD:
        role = ROLE_EXCLUDE
        reason = "Mostly standard-equipment rows (S statuses); include manually if needed."
    elif not stats["orderableRpoRows"]:
        role = ROLE_EXCLUDE
        reason = "No orderable RPO rows detected."
    else:
        role = ROLE_OPTIONS
        reason = f"{stats['orderableRpoRows']} orderable option rows detected."
    return {
        "sheetName": sheet_name,
        "sheetType": SHEET_TYPE_OPTIONS,
        "contentSubtype": subtype,
        "modelFamily": family["modelFamily"],
        "modelFamilies": family["modelFamilies"],
        "headerRow": header_row,
        "variantColumns": variant_columns,
        "rowStats": stats,
        "statusVocabulary": vocabulary,
        "confidence": confidence,
        "confidenceReasons": reasons,
        "recommendedRole": role,
        "recommendedReason": reason,
        "canonicalOptionSource": canonical_option_source,
    }


def parse_variant_header(column_index0: int, raw_header: str) -> dict[str, Any]:
    parts = [part.strip() for part in str(raw_header).split("\n") if part.strip()]
    label = parts[0] if parts else str(raw_header)
    model_code = parts[1] if len(parts) > 1 else ""
    trim = parts[2] if len(parts) > 2 else ""
    lowered = label.lower()
    body_style = "convertible" if "convertible" in lowered else "coupe" if "coupe" in lowered else ""
    return {
        "columnIndex": column_index0 + 1,
        "columnLetter": get_column_letter(column_index0 + 1),
        "rawHeader": str(raw_header),
        "label": label,
        "modelCode": clean(model_code),
        "trim": clean(trim),
        "bodyStyle": body_style,
    }


def matrix_row_stats(
    values: list[list[Any]],
    header_row: int,
    variant_columns: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, int]]:
    orderable = ref_only = section = data_rows = status_cells = standard_cells = 0
    vocabulary: dict[str, int] = {}
    for row in values[header_row:]:
        first = cell_text(row, 0)
        second = cell_text(row, 1)
        third = cell_text(row, 2)
        row_statuses = [
            cell_text(row, column["columnIndex"] - 1)
            for column in variant_columns
            if cell_text(row, column["columnIndex"] - 1)
        ]
        if not first and not second and not third and not row_statuses:
            continue
        data_rows += 1
        if RPO_RE.fullmatch(first.upper()):
            orderable += 1
        elif not first and RPO_RE.fullmatch(second.upper()):
            ref_only += 1
        elif first and not second and not third:
            section += 1
        for raw in row_statuses:
            status_cells += 1
            vocabulary[raw] = vocabulary.get(raw, 0) + 1
            match = STATUS_RE.fullmatch(raw)
            if match and match.group(1) == "S":
                standard_cells += 1
    stats = {
        "orderableRpoRows": orderable,
        "refOnlyRpoRows": ref_only,
        "sectionRows": section,
        "dataRows": data_rows,
        "statusCells": status_cells,
        "standardShare": round(standard_cells / status_cells, 4) if status_cells else 0.0,
    }
    return stats, dict(sorted(vocabulary.items()))


def matrix_confidence(
    variant_columns: list[dict[str, Any]],
    vocabulary: dict[str, int],
) -> tuple[str, list[str]]:
    reasons: list[str] = []
    if not variant_columns:
        reasons.append("No variant columns detected after the Description column.")
    incomplete = [
        column["columnLetter"]
        for column in variant_columns
        if not column["modelCode"] or not column["trim"]
    ]
    if incomplete:
        reasons.append(
            "Variant headers missing model code or trim in columns: " + ", ".join(incomplete) + "."
        )
    unknown = sorted(symbol for symbol in vocabulary if not STATUS_RE.fullmatch(symbol))
    if unknown:
        reasons.append("Unrecognized status symbols: " + ", ".join(unknown[:8]) + ".")
    if not reasons:
        return "high", []
    return ("medium" if variant_columns else "low"), reasons


def detect_model_family(title_cell: Any) -> dict[str, Any]:
    title = clean(title_cell)
    if not title:
        return {"modelFamily": "unknown", "modelFamilies": []}
    parts = [part.strip() for part in title.split(" and ") if part.strip()]
    if len(parts) > 1:
        return {"modelFamily": "mixed", "modelFamilies": parts}
    return {"modelFamily": title, "modelFamilies": [title]}


def price_card(sheet_name: str, values: list[list[Any]]) -> dict[str, Any]:
    sections = find_price_sections(values)
    reasons: list[str] = []
    confidence = "high"
    if PRICE_SECTION_BASE not in sections:
        confidence = "medium"
        reasons.append("Base Model Prices section not found.")
    return {
        "sheetName": sheet_name,
        "sheetType": SHEET_TYPE_PRICE,
        "contentSubtype": "",
        "modelFamily": "all",
        "modelFamilies": [],
        "headerRow": None,
        "variantColumns": [],
        "rowStats": price_row_stats(values, sections),
        "statusVocabulary": {},
        "confidence": confidence,
        "confidenceReasons": reasons,
        "recommendedRole": ROLE_PRICE,
        "recommendedReason": "Detected price-schedule sections.",
    }


def price_row_stats(values: list[list[Any]], sections: dict[str, int]) -> dict[str, Any]:
    option_rows = base_rows = data_rows = 0
    options_start = sections.get(PRICE_SECTION_OPTIONS, 0)
    base_start = sections.get(PRICE_SECTION_BASE, 0)
    for index, row in enumerate(values, start=1):
        code = cell_text(row, 1)
        if not any(clean(value) for value in row):
            continue
        data_rows += 1
        if options_start and index > options_start and RPO_RE.fullmatch(code.upper()):
            option_rows += 1
        elif base_start and index > base_start and (not options_start or index < options_start):
            if MODEL_CODE_RE.fullmatch(code.upper()):
                base_rows += 1
    return {"optionPriceRows": option_rows, "baseModelRows": base_rows, "dataRows": data_rows}


def unsupported_card(sheet_name: str) -> dict[str, Any]:
    return {
        "sheetName": sheet_name,
        "sheetType": SHEET_TYPE_UNSUPPORTED,
        "contentSubtype": "",
        "modelFamily": "unknown",
        "modelFamilies": [],
        "headerRow": None,
        "variantColumns": [],
        "rowStats": {},
        "statusVocabulary": {},
        "confidence": "low",
        "confidenceReasons": ["No recognized options-matrix headers or price-schedule sections."],
        "recommendedRole": ROLE_EXCLUDE,
        "recommendedReason": "Unsupported layout in Pass A; handle manually.",
    }
