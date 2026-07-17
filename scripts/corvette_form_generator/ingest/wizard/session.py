#!/usr/bin/env python3
"""Run-state persistence and fail-closed state machine for the ingest wizard.

States: profiled -> roles_confirmed -> parsed (Pass A), then
models_selected -> decisions_in_progress -> decisions_complete (Pass B), then
plan_built -> dry_run_approved -> dry_run_validated_* (Pass C/D evidence).
Every transition persists JSON artifacts under
form-output/ingest-wizard/<run-id>/ so a run can be reopened and later passes
can consume the output. The canonical workbook is opened read-only for
pickers, variant reconciliation, presentation prefill, and Pass C dry-runs;
it and the raw source file are never written by this store.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import threading
import time
import uuid
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from corvette_form_generator.ingest.wizard.decisions import (
    LANES,
    SCHEMA_VERSION_B,
    SCHEMA_VERSION_B2,
    VARIANT_RECONCILIATION_KEY,
    artifact_fingerprint,
    candidate_has_price,
    candidate_is_availability_row,
    candidate_needs_section_decision,
    candidate_needs_status_review,
    completeness,
    copy_decisions,
    detect_model_options,
    duplicate_rpo_groups,
    load_decision_state,
    model_scoped_statuses,
    presentation_prefill,
    scope_candidates,
    validate_decision,
    validate_selection,
    variant_reconciliation,
    workbook_option_reference,
    workbook_sections,
)
from corvette_form_generator.ingest.wizard.copy_split import FLAG_DUPLICATE_NAME, propose_copy_split
from corvette_form_generator.ingest.wizard.canonical_rows import (
    COMPILER_POLICY_VERSION,
    canonical_bytes,
    canonical_text,
    semantic_hash,
    validate_artifact_graph,
)
from corvette_form_generator.ingest.wizard.comparator_evidence import build_comparator_evidence
from corvette_form_generator.ingest.wizard.compiler import (
    build_family_registry,
    compile_canonical_rows as run_canonical_compiler,
)
from corvette_form_generator.ingest.wizard.exceptions import (
    ACTION_DISPOSITIONS,
    ALLOWED_DEFERRAL_KINDS,
    append_audit_event_once,
    build_audit_event,
    validate_resolution,
)
from corvette_form_generator.ingest.wizard.hints import scan_candidates
from corvette_form_generator.ingest.wizard.identity import option_occurrence_signature
from corvette_form_generator.ingest.wizard.plan_builder import (
    artifact_sha,
    plan_markdown,
    plan_summary,
)
from corvette_form_generator.ingest.wizard.joiner import join_prices
from corvette_form_generator.ingest.wizard.parser import parse_confirmed_sheets
from corvette_form_generator.ingest.wizard.profiler import (
    ROLE_EXCLUDE,
    ROLE_OPTIONS,
    ROLE_PRICE,
    SCHEMA_VERSION,
    SHEET_TYPE_OPTIONS,
    SHEET_TYPE_PRICE,
    canonical_option_sheet_eligible,
    profile_workbook,
)

STATE_PROFILED = "profiled"
STATE_ROLES_CONFIRMED = "roles_confirmed"
STATE_PARSED = "parsed"
STATE_MODELS_SELECTED = "models_selected"
STATE_COMPILED_READY = "compiled_ready"
STATE_COMPILED = STATE_COMPILED_READY
STATE_COMPILED_WITH_EXCEPTIONS = "compiled_with_exceptions"
STATE_DECISIONS_IN_PROGRESS = "decisions_in_progress"
STATE_DECISIONS_COMPLETE = "decisions_complete"
STATE_PLAN_BUILT = "plan_built"
# Historical runs used this state and may still produce dry-run evidence. It
# is never live-write authority.
STATE_PLAN_APPROVED = "plan_approved"
STATE_DRY_RUN_APPROVED = "dry_run_approved"
STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED = "dry_run_validated_write_blocked"
STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE = "dry_run_validated_write_eligible"
STATE_WRITE_APPROVED = "write_approved"
STATE_APPLIED = "applied"
STATE_APPLY_VERIFICATION_FAILED = "apply_verification_failed"
PLAN_APPROVAL_SCHEMA = "plan-approval-2"
PLAN_APPROVAL_SCOPE = "dry_run_evidence"
WRITE_APPROVAL_SCHEMA = "write-approval-1"
WRITE_APPROVAL_SCOPE = "deployment_ready_write"
SCHEMA_VERSION_D = "pass-d-2"
WRITABLE_PLAN_SCHEMA = "pass-c-3"
ALLOWED_WRITE_DEFERRAL_KINDS = {"asset_map_media_missing"}
COMPILE_READINESS_FIELDS = (
    "compileReady",
    "planReady",
    "writeReady",
    "deploymentReady",
)
NOT_APPLICABLE_DEPLOYMENT_FAMILIES = {
    "asset_map",
    "color_overrides",
    "interiors",
    "interior_components",
}
COMPILER_ARTIFACT_BINDINGS = (
    ("canonical-row-manifest.json", "canonicalManifestSha"),
    ("compile-report.json", "compileReportSha"),
    ("exception-resolutions.json", "exceptionResolutionsSha"),
)
DECISION_STATES = (
    STATE_MODELS_SELECTED,
    STATE_DECISIONS_IN_PROGRESS,
    STATE_DECISIONS_COMPLETE,
    STATE_PLAN_BUILT,
    STATE_PLAN_APPROVED,
    STATE_DRY_RUN_APPROVED,
    STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED,
    STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE,
    STATE_WRITE_APPROVED,
    STATE_APPLIED,
    STATE_APPLY_VERIFICATION_FAILED,
)
VALID_ROLES = {ROLE_OPTIONS, ROLE_PRICE, ROLE_EXCLUDE}
RUN_ID_RE = re.compile(r"^[0-9]{8}-[0-9]{6}-[0-9a-f]{6}$")
PARSE_ARTIFACTS = ("option-candidates.json", "price-rows.json", "join-report.json")
COMPILER_ARTIFACTS = (
    "comparator-evidence.json",
    "canonical-row-manifest.json",
    "exception-queue.json",
    "exception-resolutions.json",
    "compile-report.json",
)
COMPILER_CACHE_ARTIFACTS = tuple(
    name for name in COMPILER_ARTIFACTS if name != "exception-resolutions.json"
)
COMPILER_MUTATION_FILES = (*COMPILER_ARTIFACTS, "exception-log.jsonl", "session.json")
COMPILER_STATES = (STATE_COMPILED, STATE_COMPILED_WITH_EXCEPTIONS)
COMPILER_DOWNSTREAM_ARTIFACTS = (
    "apply-plan.json",
    "apply-plan-dryrun.json",
    "apply-plan.md",
    "plan-approval.json",
    "apply-dry-run-report.json",
    "dry-run-approval.json",
    "write-approval.json",
    "apply-report.json",
    "scratch-apply-log.jsonl",
    "apply-workbook-edit-log.jsonl",
    "approval-log.jsonl",
)


class WizardError(ValueError):
    """User-visible wizard failure; maps to an HTTP 4xx response."""

    def __init__(self, message: str, status: int = 400) -> None:
        super().__init__(message)
        self.status = status


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def replace_json_artifact_set(run_dir: Path, payloads: dict[str, Any]) -> None:
    """Validate then replace only changed JSON artifacts with rollback."""

    token = uuid.uuid4().hex
    temporary: dict[str, Path] = {}
    originals: dict[str, bytes | None] = {}
    replaced: list[str] = []
    try:
        for filename, payload in payloads.items():
            encoded = json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n"
            json.loads(encoded)
            target = run_dir / filename
            encoded_bytes = encoded.encode("utf-8")
            if target.is_file() and target.read_bytes() == encoded_bytes:
                continue
            temp = run_dir / f".{filename}.{token}.tmp"
            with temp.open("wb") as handle:
                handle.write(encoded_bytes)
                handle.flush()
                os.fsync(handle.fileno())
            temporary[filename] = temp
            originals[filename] = target.read_bytes() if target.is_file() else None
        for filename in sorted(temporary):
            os.replace(temporary[filename], run_dir / filename)
            replaced.append(filename)
        directory_fd = os.open(run_dir, os.O_RDONLY)
        try:
            os.fsync(directory_fd)
        finally:
            os.close(directory_fd)
    except Exception:
        for filename in replaced:
            target = run_dir / filename
            original = originals[filename]
            if original is None:
                target.unlink(missing_ok=True)
            else:
                target.write_bytes(original)
        raise
    finally:
        for temp in temporary.values():
            temp.unlink(missing_ok=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def file_fingerprint(path: Path) -> dict[str, Any]:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    stat = path.stat()
    return {"sha256": digest, "sizeBytes": stat.st_size, "mtimeNs": stat.st_mtime_ns}


class WizardSessionStore:
    def __init__(self, root: Path, workbook_path: Path | None = None) -> None:
        self.root = Path(root)
        self.base = self.root / "form-output" / "ingest-wizard"
        self.uploads = self.base / "uploads"
        self.workbook_path = Path(workbook_path) if workbook_path else self.root / "stingray_master.xlsx"
        self._reference_cache: tuple[int, dict] | None = None
        self._run_locks: defaultdict[str, threading.RLock] = defaultdict(threading.RLock)

    def _option_reference(self) -> dict:
        path = self._require_workbook()
        mtime = path.stat().st_mtime_ns
        if self._reference_cache is None or self._reference_cache[0] != mtime:
            self._reference_cache = (mtime, workbook_option_reference(path))
        return self._reference_cache[1]

    def _require_workbook(self) -> Path:
        if not self.workbook_path.is_file():
            raise WizardError(
                f"Canonical workbook not found (read-only pickers need it): {self.workbook_path.name}",
                status=409,
            )
        return self.workbook_path

    # ------------------------------------------------------------- files
    def list_source_files(self) -> list[dict[str, Any]]:
        files: list[dict[str, Any]] = []
        for directory, origin in ((self.root, "project"), (self.uploads, "upload")):
            if not directory.is_dir():
                continue
            for path in sorted(directory.glob("*.xlsx")):
                if path.name == "stingray_master.xlsx" or path.name.startswith("~$"):
                    continue
                files.append(
                    {"name": path.name, "origin": origin, "sizeBytes": path.stat().st_size}
                )
        return files

    def save_upload(self, filename: str, data: bytes) -> dict[str, Any]:
        name = Path(str(filename)).name
        if (
            not name
            or name != filename
            or not name.lower().endswith(".xlsx")
            or name.startswith("~$")
        ):
            raise WizardError("Upload filename must be a plain .xlsx basename.")
        self.uploads.mkdir(parents=True, exist_ok=True)
        target = self.uploads / name
        target.write_bytes(data)
        return {"name": name, "origin": "upload", "sizeBytes": target.stat().st_size}

    def resolve_source(self, name: str) -> Path:
        if Path(str(name)).name != name:
            raise WizardError("Source file must be a plain basename.")
        for directory in (self.uploads, self.root):
            path = directory / name
            if path.is_file():
                return path
        raise WizardError(f"Source file not found: {name}")

    # ---------------------------------------------------------- sessions
    def create_session(self, file_name: str) -> dict[str, Any]:
        source = self.resolve_source(file_name)
        run_id = time.strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:6]
        run_dir = self.base / run_id
        run_dir.mkdir(parents=True)
        profile = profile_workbook(source)
        session = {
            "schemaVersion": SCHEMA_VERSION,
            "runId": run_id,
            "state": STATE_PROFILED,
            "sourceFile": file_name,
            "sourcePath": str(source),
            "fingerprint": file_fingerprint(source),
        }
        write_json(run_dir / "session.json", session)
        write_json(run_dir / "sheet-profile.json", profile)
        return {"session": session, "profile": profile}

    def run_dir(self, run_id: str) -> Path:
        if not RUN_ID_RE.fullmatch(str(run_id)):
            raise WizardError(f"Invalid run id: {run_id}")
        run_dir = self.base / run_id
        if not (run_dir / "session.json").is_file():
            raise WizardError(f"Unknown run: {run_id}", status=404)
        return run_dir

    @staticmethod
    def _snapshot_run_files(run_dir: Path, names: tuple[str, ...]) -> dict[str, bytes | None]:
        return {
            name: (run_dir / name).read_bytes() if (run_dir / name).is_file() else None
            for name in names
        }

    @staticmethod
    def _restore_run_files(run_dir: Path, snapshot: dict[str, bytes | None]) -> None:
        """Atomically restore one coherent pre-mutation file snapshot."""

        for name, content in snapshot.items():
            path = run_dir / name
            if content is None:
                path.unlink(missing_ok=True)
                continue
            temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.restore")
            temp.write_bytes(content)
            os.replace(temp, path)

    def load_session(self, run_id: str, *, verify_source: bool = True) -> dict[str, Any]:
        session = read_json(self.run_dir(run_id) / "session.json")
        if verify_source:
            source = Path(session["sourcePath"])
            if not source.is_file():
                raise WizardError("Source file for this run is missing; start a new run.")
            if file_fingerprint(source)["sha256"] != session["fingerprint"]["sha256"]:
                raise WizardError("Source file changed since it was profiled; start a new run.")
        return session

    def list_sessions(self) -> list[dict[str, Any]]:
        if not self.base.is_dir():
            return []
        sessions = []
        for run_dir in sorted(self.base.iterdir(), reverse=True):
            session_file = run_dir / "session.json"
            if session_file.is_file():
                sessions.append(read_json(session_file))
        return sessions

    def session_detail(self, run_id: str) -> dict[str, Any]:
        run_dir = self.run_dir(run_id)
        roles_file = run_dir / "sheet-roles.json"
        report_file = run_dir / "join-report.json"
        selection_file = run_dir / "model-selection.json"
        return {
            "session": self.load_session(run_id, verify_source=False),
            "profile": read_json(run_dir / "sheet-profile.json"),
            "roles": read_json(roles_file)["roles"] if roles_file.is_file() else None,
            "joinReport": read_json(report_file) if report_file.is_file() else None,
            "modelSelection": read_json(selection_file) if selection_file.is_file() else None,
        }

    def _refuse_if_applied(self, session: dict[str, Any]) -> None:
        if session.get("state") == STATE_APPLIED:
            raise WizardError("Run already applied; start a new run for further changes.", status=409)
        if session.get("state") == STATE_APPLY_VERIFICATION_FAILED:
            raise WizardError(
                "The last apply failed verification; restore or reconcile it before continuing.",
                status=409,
            )

    def _invalidate_compiler_artifacts(self, run_dir: Path, *, changed: str) -> None:
        """Evict only transitive dependents; preserve resolutions and audit history."""

        dependents = {
            "source": COMPILER_CACHE_ARTIFACTS,
            "selection": COMPILER_CACHE_ARTIFACTS,
            "workbook": COMPILER_CACHE_ARTIFACTS,
        }
        if changed not in dependents:
            raise ValueError(f"Unknown compiler dependency node: {changed}")
        for filename in dependents[changed]:
            (run_dir / filename).unlink(missing_ok=True)

    def _compiler_artifact_bindings(self, run_dir: Path) -> dict[str, str]:
        """Return only compiler hashes that actually exist for this run.

        Milestone 0 diagnostic runs predate the compiler, so absent hashes are
        deliberately omitted rather than represented by sentinel values.
        """

        bindings: dict[str, str] = {}
        for filename, field in COMPILER_ARTIFACT_BINDINGS:
            path = run_dir / filename
            if path.is_file():
                bindings[field] = hashlib.sha256(path.read_bytes()).hexdigest()
        return bindings

    def _require_compiler_bindings(
        self,
        run_dir: Path,
        plan: dict[str, Any],
        *approvals: tuple[str, dict[str, Any]],
    ) -> dict[str, str]:
        current: dict[str, str] = {}
        for filename, field in COMPILER_ARTIFACT_BINDINGS:
            path = run_dir / filename
            if not path.is_file():
                raise WizardError(
                    f"{WRITABLE_PLAN_SCHEMA} requires current {filename}.",
                    status=409,
                )
            sha = hashlib.sha256(path.read_bytes()).hexdigest()
            current[field] = sha
            if plan.get(field) != sha:
                raise WizardError(f"Plan compiler binding {field} is absent or stale.", status=409)
            for label, approval in approvals:
                if approval.get(field) != sha:
                    raise WizardError(
                        f"{label} compiler binding {field} is absent or stale.",
                        status=409,
                    )
        return current

    def _compile_report_allowed_deferrals(
        self,
        run_dir: Path,
        targets: list[str],
    ) -> list[dict[str, Any]]:
        """Return the closed-policy deferrals declared by the bound compile report."""

        report_file = run_dir / "compile-report.json"
        if not report_file.is_file():
            raise WizardError("Write eligibility requires the current compile report.", status=409)
        report = read_json(report_file)
        declared = report.get("deferrals") or []
        if not isinstance(declared, list):
            raise WizardError("Compile report deferrals must be a list.", status=409)
        allowed: list[dict[str, Any]] = []
        seen_ids: set[str] = set()
        target_set = set(targets)
        for item in declared:
            if not isinstance(item, dict) or item.get("disposition") != "allowed_deferral":
                continue
            kind = str(item.get("kind") or "")
            deferral_id = str(item.get("deferralId") or "")
            model = str(item.get("model") or "")
            if kind not in ALLOWED_WRITE_DEFERRAL_KINDS:
                raise WizardError(
                    f"Compile report contains non-allowlisted deferral kind {kind!r}.",
                    status=409,
                )
            if model not in target_set:
                raise WizardError(
                    "Compile report allowed deferrals require an exact selected target model.",
                    status=409,
                )
            if not deferral_id or deferral_id in seen_ids:
                raise WizardError(
                    "Compile report allowed deferrals require unique stable deferralId values.",
                    status=409,
                )
            seen_ids.add(deferral_id)
            allowed.append(dict(item))
        return allowed

    def _compile_report_not_applicable_families(
        self,
        run_dir: Path,
        targets: list[str],
    ) -> set[tuple[str, str]]:
        """Return model/family pairs whose zero-row result is explicitly proven complete."""

        report_file = run_dir / "compile-report.json"
        if not report_file.is_file():
            return set()
        coverage = read_json(report_file).get("sourceFeatureCoverage") or []
        if not isinstance(coverage, list):
            raise WizardError("Compile report sourceFeatureCoverage must be a list.", status=409)
        result: set[tuple[str, str]] = set()
        target_set = set(targets)
        for item in coverage:
            if not isinstance(item, dict) or item.get("disposition") != "resolved_not_applicable":
                continue
            model = str(item.get("model") or "")
            family = str(item.get("family") or "")
            evidence_ids = item.get("evidenceIds") or []
            valid_evidence = (
                isinstance(evidence_ids, list)
                and bool(evidence_ids)
                and all(isinstance(value, str) and value.strip() for value in evidence_ids)
                and len(set(evidence_ids)) == len(evidence_ids)
            )
            if model not in target_set or family not in NOT_APPLICABLE_DEPLOYMENT_FAMILIES:
                raise WizardError(
                    "Compile report resolved_not_applicable coverage must bind a selected target "
                    "and known deployment family.",
                    status=409,
                )
            if not valid_evidence:
                raise WizardError(
                    "Compile report resolved_not_applicable coverage requires unique nonblank "
                    "string evidenceIds.",
                    status=409,
                )
            result.add((model, family))
        return result

    def _require_compile_report_readiness(
        self,
        run_dir: Path,
        targets: list[str],
    ) -> dict[str, dict[str, Any]]:
        """Require exact per-target compiler readiness and an empty blocker set."""

        report_file = run_dir / "compile-report.json"
        if not report_file.is_file():
            raise WizardError("Write eligibility requires the current compile report.", status=409)
        report = read_json(report_file)
        if report.get("schemaVersion") != "compile-report-1":
            raise WizardError("Write eligibility requires compile-report-1.", status=409)
        models = report.get("models")
        if not isinstance(models, dict) or set(models) != set(targets):
            raise WizardError(
                "Compile report readiness must cover the exact selected target set.",
                status=409,
            )
        for model in targets:
            entry = models.get(model)
            if not isinstance(entry, dict):
                raise WizardError(f"Compile report readiness is missing for {model}.", status=409)
            missing = [field for field in COMPILE_READINESS_FIELDS if entry.get(field) is not True]
            if missing:
                raise WizardError(
                    f"Compile report target {model} is not ready: {', '.join(missing)}.",
                    status=409,
                )
            blockers = entry.get("blockers")
            if not isinstance(blockers, list) or blockers != []:
                raise WizardError(
                    f"Compile report target {model} still contains blockers.",
                    status=409,
                )
        return {str(model): dict(models[model]) for model in targets}

    def _validate_report_deferrals(
        self,
        run_dir: Path,
        deferrals: list[dict[str, Any]],
        targets: list[str],
    ) -> list[dict[str, Any]]:
        """Require every report deferral to be an exact compile-report declaration."""

        declared = {
            str(item["deferralId"]): item
            for item in self._compile_report_allowed_deferrals(run_dir, targets)
        }
        validated: list[dict[str, Any]] = []
        seen_ids: set[str] = set()
        for item in deferrals:
            if not isinstance(item, dict):
                raise WizardError("Dry-run deferrals must be structured records.", status=409)
            kind = str(item.get("kind") or "")
            deferral_id = str(item.get("deferralId") or "")
            if kind not in ALLOWED_WRITE_DEFERRAL_KINDS:
                raise WizardError(
                    f"Dry-run report contains unapproved deferral kind {kind!r}.",
                    status=409,
                )
            if not deferral_id or deferral_id in seen_ids:
                raise WizardError(
                    "Dry-run deferrals require unique stable deferralId values.",
                    status=409,
                )
            expected = declared.get(deferral_id)
            if expected != item:
                raise WizardError(
                    f"Dry-run deferral {deferral_id!r} does not match the bound compile report.",
                    status=409,
                )
            seen_ids.add(deferral_id)
            validated.append(dict(item))
        return validated

    def _approval_bindings(
        self,
        run_id: str,
        session: dict[str, Any],
        plan: dict[str, Any],
        run_dir: Path,
    ) -> dict[str, Any]:
        bindings = {
            "runId": run_id,
            "targets": list(plan.get("targets") or []),
            "planSchemaVersion": plan.get("schemaVersion"),
            "planSha": hashlib.sha256((run_dir / "apply-plan.json").read_bytes()).hexdigest(),
            "workbookFingerprint": dict(plan.get("workbookFingerprint") or {}),
            "sourceFingerprint": dict(session.get("fingerprint") or {}),
            "candidatesFingerprint": plan.get("candidatesFingerprint"),
            "decisionsFingerprint": plan.get("decisionsFingerprint"),
            **self._compiler_artifact_bindings(run_dir),
        }
        selection_file = run_dir / "model-selection.json"
        if selection_file.is_file():
            bindings["modelSelectionSha"] = hashlib.sha256(selection_file.read_bytes()).hexdigest()
        return bindings

    def _append_approval_audit(self, run_dir: Path, record: dict[str, Any]) -> None:
        with (run_dir / "approval-log.jsonl").open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")

    # ------------------------------------------------------------- roles
    def confirm_roles(self, run_id: str, roles: dict[str, str]) -> dict[str, Any]:
        with self._run_locks[run_id]:
            return self._confirm_roles_locked(run_id, roles)

    def _confirm_roles_locked(self, run_id: str, roles: dict[str, str]) -> dict[str, Any]:
        session = self.load_session(run_id)
        run_dir = self.run_dir(run_id)
        profile = read_json(run_dir / "sheet-profile.json")
        cards = {card["sheetName"]: card for card in profile["sheets"]}
        confirmed: dict[str, str] = {}
        for sheet, role in (roles or {}).items():
            if sheet not in cards:
                raise WizardError(f"Unknown sheet: {sheet}")
            if role not in VALID_ROLES:
                raise WizardError(f"Invalid role for {sheet}: {role}")
            confirmed[sheet] = role
        for sheet in cards:
            confirmed.setdefault(sheet, ROLE_EXCLUDE)
        for sheet, role in confirmed.items():
            sheet_type = cards[sheet]["sheetType"]
            if role == ROLE_OPTIONS and sheet_type != SHEET_TYPE_OPTIONS:
                raise WizardError(
                    f"{sheet} was not detected as an options matrix; it cannot take the options role."
                )
            if role == ROLE_OPTIONS and not canonical_option_sheet_eligible(sheet):
                raise WizardError(
                    f"{sheet} is not a canonical option source; choose an Interior, Exterior, or Mechanical sheet."
                )
            if role == ROLE_PRICE and sheet_type != SHEET_TYPE_PRICE:
                raise WizardError(
                    f"{sheet} was not detected as a price sheet; it cannot take the price role."
                )
        options = [sheet for sheet, role in confirmed.items() if role == ROLE_OPTIONS]
        price = [sheet for sheet, role in confirmed.items() if role == ROLE_PRICE]
        if not options:
            raise WizardError("Confirm at least one options sheet.")
        if len(price) != 1:
            raise WizardError("Confirm exactly one price sheet.")
        self._invalidate_compiler_artifacts(run_dir, changed="source")
        write_json(run_dir / "sheet-roles.json", {"schemaVersion": SCHEMA_VERSION, "roles": confirmed})
        for stale in PARSE_ARTIFACTS:
            (run_dir / stale).unlink(missing_ok=True)
        session["state"] = STATE_ROLES_CONFIRMED
        write_json(run_dir / "session.json", session)
        return session

    # ------------------------------------------------------------- parse
    def run_parse(self, run_id: str) -> dict[str, Any]:
        with self._run_locks[run_id]:
            return self._run_parse_locked(run_id)

    def _run_parse_locked(self, run_id: str) -> dict[str, Any]:
        session = self.load_session(run_id)
        if session["state"] not in (STATE_ROLES_CONFIRMED, STATE_PARSED) + DECISION_STATES + COMPILER_STATES:
            raise WizardError("Confirm sheet roles before parsing.")
        run_dir = self.run_dir(run_id)
        roles = read_json(run_dir / "sheet-roles.json")["roles"]
        parsed = parse_confirmed_sheets(Path(session["sourcePath"]), roles)
        report = join_prices(parsed["candidates"], parsed["priceRows"])
        self._invalidate_compiler_artifacts(run_dir, changed="source")
        write_json(
            run_dir / "option-candidates.json",
            {
                "schemaVersion": SCHEMA_VERSION,
                "candidates": parsed["candidates"],
                "skippedRows": parsed["skippedRows"],
            },
        )
        write_json(
            run_dir / "price-rows.json",
            {
                "schemaVersion": SCHEMA_VERSION,
                "priceRows": parsed["priceRows"],
                "baseModelPriceRows": parsed["baseModelPriceRows"],
                "skippedPriceRows": parsed["skippedPriceRows"],
            },
        )
        write_json(run_dir / "join-report.json", report)
        session["state"] = STATE_PARSED
        write_json(run_dir / "session.json", session)
        return {"session": session, "joinReport": report}

    # -------------------------------------------------------- candidates
    def candidates(
        self,
        run_id: str,
        *,
        sheet: str = "",
        price_match: str = "",
        family: str = "",
        query: str = "",
    ) -> dict[str, Any]:
        session = self.load_session(run_id, verify_source=False)
        if session["state"] not in (STATE_PARSED,) + DECISION_STATES + COMPILER_STATES:
            raise WizardError("Run the parse before requesting candidates.")
        run_dir = self.run_dir(run_id)
        payload = read_json(run_dir / "option-candidates.json")
        report = read_json(run_dir / "join-report.json")
        rows = payload["candidates"]
        if sheet:
            rows = [row for row in rows if row["sheetName"] == sheet]
        if family:
            rows = [row for row in rows if row["modelFamily"] == family]
        if price_match:
            rows = [row for row in rows if (row["priceMatch"] or "") == price_match]
        if query:
            needle = query.lower()
            rows = [
                row
                for row in rows
                if needle in row["rpo"].lower()
                or needle in row["refOnlyRpo"].lower()
                or needle in row["description"].lower()
            ]
        return {
            "session": session,
            "total": len(payload["candidates"]),
            "matched": len(rows),
            "candidates": rows,
            "skippedRows": payload["skippedRows"],
            "unmatchedPriceRows": report["unmatchedPriceRows"],
        }

    # ------------------------------------------------- pass b: model scoping
    def _parsed_candidates(
        self, run_id: str, *, allow_compiled: bool = False
    ) -> tuple[dict[str, Any], Path, list[dict[str, Any]]]:
        session = self.load_session(run_id, verify_source=False)
        allowed = (STATE_PARSED,) + DECISION_STATES + (COMPILER_STATES if allow_compiled else ())
        if session["state"] not in allowed:
            raise WizardError("Run the parse before Pass B model selection.")
        candidates_file = self.run_dir(run_id) / "option-candidates.json"
        return session, candidates_file, read_json(candidates_file)["candidates"]

    def _load_selection(self, run_id: str, candidates_file: Path) -> dict[str, Any]:
        selection_file = self.run_dir(run_id) / "model-selection.json"
        if not selection_file.is_file():
            raise WizardError("Select target models before reviewing decisions.")
        selection = read_json(selection_file)
        if selection.get("candidatesFingerprint") != artifact_fingerprint(candidates_file):
            raise WizardError(
                "Candidates were re-parsed after model selection; re-select target models. "
                "Existing decisions with matching evidence fingerprints will be kept.",
                status=409,
            )
        return selection

    def model_options(self, run_id: str) -> dict[str, Any]:
        session, _, candidates = self._parsed_candidates(run_id, allow_compiled=True)
        run_dir = self.run_dir(run_id)
        selection_file = run_dir / "model-selection.json"
        return {
            "session": session,
            "models": detect_model_options(candidates),
            "selection": read_json(selection_file) if selection_file.is_file() else None,
        }

    def select_models(self, run_id: str, targets: list[str], comparators: dict[str, str]) -> dict[str, Any]:
        with self._run_locks[run_id]:
            return self._select_models_locked(run_id, targets, comparators)

    def _select_models_locked(
        self,
        run_id: str,
        targets: list[str],
        comparators: dict[str, str],
    ) -> dict[str, Any]:
        session, candidates_file, candidates = self._parsed_candidates(run_id, allow_compiled=True)
        self._refuse_if_applied(session)
        try:
            validate_selection(candidates, targets, comparators)
        except ValueError as exc:
            raise WizardError(str(exc)) from exc
        run_dir = self.run_dir(run_id)
        self._invalidate_compiler_artifacts(run_dir, changed="selection")
        selection = {
            "schemaVersion": SCHEMA_VERSION_B,
            "targets": list(targets),
            "comparators": dict(comparators),
            "sourceFingerprint": session["fingerprint"],
            "candidatesFingerprint": artifact_fingerprint(candidates_file),
            "selectedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
        }
        write_json(run_dir / "model-selection.json", selection)
        reconciliation = variant_reconciliation(self._require_workbook(), candidates, list(targets))
        write_json(run_dir / "variant-reconciliation.json", reconciliation)
        # Keep only decisions that still resolve: in-target model and matching
        # candidate evidence fingerprint (re-parse invalidation, spec B2).
        state = self._decision_state(run_id, candidates, selection, prune_to_targets=True)
        self._write_decisions(run_id, state["decisions"], selection)
        session["state"] = STATE_MODELS_SELECTED if not state["decisions"] else STATE_DECISIONS_IN_PROGRESS
        write_json(run_dir / "session.json", session)
        return {
            "session": session,
            "selection": selection,
            "reconciliation": self._attach_reconciliation_decisions(dict(reconciliation), state["decisions"]),
        }

    def compile_canonical_rows(self, run_id: str) -> dict[str, Any]:
        """Compile one run while serializing all compiler artifact mutations."""

        with self._run_locks[run_id]:
            return self._compile_canonical_rows_locked(run_id)

    def _compile_canonical_rows_locked(self, run_id: str) -> dict[str, Any]:
        """Compile one selected run without mutating the canonical workbook."""

        session = self.load_session(run_id)
        if session["state"] not in (STATE_MODELS_SELECTED,) + COMPILER_STATES:
            raise WizardError("Select target and comparator models before compiling.", status=409)
        run_dir = self.run_dir(run_id)
        downstream_name = re.compile(
            r"apply|plan|approval|dry[-_]?run|write|backup|edit[-_]?log|promotion",
            re.IGNORECASE,
        )
        downstream = sorted(
            {
                *(name for name in COMPILER_DOWNSTREAM_ARTIFACTS if (run_dir / name).exists()),
                *(path.name for path in run_dir.glob("apply-plan*")),
                *(path.name for path in run_dir.iterdir() if downstream_name.search(path.name)),
            }
        )
        if session.get("dryRunReport") or session.get("applyReport"):
            downstream.append("session downstream report binding")
        decisions_file = run_dir / "decisions.json"
        if decisions_file.is_file() and (read_json(decisions_file).get("decisions") or []):
            downstream.append("nonempty decisions.json")
        if downstream:
            raise WizardError(
                "Compiler refuses a run with downstream plan/apply evidence; start a fresh run: "
                + ", ".join(downstream),
                status=409,
            )
        candidates_file = run_dir / "option-candidates.json"
        selection = self._load_selection(run_id, candidates_file)
        workbook = self._require_workbook()
        input_paths = {
            "source": Path(session["sourcePath"]),
            "workbook": workbook,
            "sheetRoles": run_dir / "sheet-roles.json",
            "sheetProfile": run_dir / "sheet-profile.json",
            "optionCandidates": candidates_file,
            "priceRows": run_dir / "price-rows.json",
            "joinReport": run_dir / "join-report.json",
            "modelSelection": run_dir / "model-selection.json",
        }
        missing = [label for label, path in input_paths.items() if not path.is_file()]
        if missing:
            raise WizardError(f"Compiler inputs are missing: {', '.join(missing)}.", status=409)
        before = {label: file_fingerprint(path) for label, path in input_paths.items()}
        from corvette_form_generator.editor_ops import extract_workbook, rows_of
        from corvette_form_generator.ingest.wizard.relationship_compiler import load_compiler_phrase_map

        workbook_extract = extract_workbook(workbook)
        selected_models = {
            *[str(model) for model in selection.get("targets") or []],
            *[str(model) for model in (selection.get("comparators") or {}).values()],
        }
        source_role_rows = sorted(
            (
                {str(key): value for key, value in row.items()}
                for row in rows_of(workbook_extract, "model_workbook_sources")
                if str(row.get("model_key") or "") in selected_models
            ),
            key=lambda row: (
                str(row.get("model_key") or ""),
                str(row.get("source_role") or ""),
                str(row.get("sheet_name") or ""),
            ),
        )
        phrase_rows = load_compiler_phrase_map(workbook)
        authority_bindings = {
            "compilerPolicyVersion": COMPILER_POLICY_VERSION,
            "files": before,
            "modelWorkbookSources": source_role_rows,
            "rulePhraseMap": phrase_rows,
        }
        run_authority = {
            "fingerprint": hashlib.sha256(canonical_bytes(authority_bindings)).hexdigest(),
            "bindings": authority_bindings,
        }
        resolution_entries: list[dict[str, Any]] = []
        resolution_file = run_dir / "exception-resolutions.json"
        if resolution_file.is_file():
            previous_resolutions = read_json(resolution_file)
            seen: set[str] = set()
            for field in ("entries", "validEntries", "staleEntries", "supersededEntries"):
                for entry in previous_resolutions.get(field) or []:
                    identity = semantic_hash(entry)
                    if identity not in seen:
                        seen.add(identity)
                        resolution_entries.append(dict(entry))
        try:
            comparator = build_comparator_evidence(
                workbook,
                selection["comparators"],
                run_authority_fingerprint=run_authority,
            )
            artifacts = run_canonical_compiler(
                workbook_path=workbook,
                option_payload=read_json(input_paths["optionCandidates"]),
                price_payload=read_json(input_paths["priceRows"]),
                join_report=read_json(input_paths["joinReport"]),
                roles_payload=read_json(input_paths["sheetRoles"]),
                sheet_profile=read_json(input_paths["sheetProfile"]),
                selection=selection,
                comparator_artifact=comparator,
                run_authority_fingerprint=run_authority,
                resolution_entries=resolution_entries,
            )
        except (ValueError, KeyError) as exc:
            raise WizardError(f"Compiler refused the run: {exc}", status=409) from exc
        after = {label: file_fingerprint(path) for label, path in input_paths.items()}
        if before != after:
            raise WizardError("Compiler input changed during compilation; no artifacts were replaced.", status=409)
        payloads = {name: artifacts[name] for name in COMPILER_ARTIFACTS}
        replace_json_artifact_set(run_dir, payloads)
        log_path = run_dir / "exception-log.jsonl"
        log_path.touch(exist_ok=True)
        queue = artifacts["exception-queue.json"]
        resolution_artifact = artifacts["exception-resolutions.json"]
        current_subjects = {item["subjectId"]: item for item in queue.get("subjects") or []}
        for valid in resolution_artifact.get("validEntries") or []:
            disposition = str(valid.get("disposition") or "resolved")
            next_state = disposition if disposition != "resolved" else "resolved_pending_projection"
            event = build_audit_event(
                queue_subject_fingerprint=queue["queueSubjectFingerprint"],
                subject_id_value=str(valid.get("subjectId") or ""),
                subject_version_value=str(valid.get("subjectVersion") or ""),
                event_type="resolution_recorded",
                prior_state="open",
                next_state=next_state,
                cause_fingerprint=semantic_hash(valid),
                resolution_entry_semantic_sha=semantic_hash(valid),
                reviewer=str(valid.get("reviewer") or ""),
            )
            append_audit_event_once(log_path, event)
        for stale in resolution_artifact.get("staleEntries") or []:
            current = current_subjects.get(str(stale.get("subjectId") or ""))
            cause = str(current.get("subjectVersion") if current else "subject_removed")
            event = build_audit_event(
                queue_subject_fingerprint=queue["queueSubjectFingerprint"],
                subject_id_value=str(stale.get("subjectId") or ""),
                subject_version_value=str(stale.get("subjectVersion") or ""),
                event_type="resolution_became_stale",
                prior_state="resolved",
                next_state="stale",
                cause_fingerprint=cause,
                resolution_entry_semantic_sha=semantic_hash(stale),
                reviewer=str(stale.get("reviewer") or ""),
            )
            append_audit_event_once(log_path, event)
        for superseded in resolution_artifact.get("supersededEntries") or []:
            event = build_audit_event(
                queue_subject_fingerprint=queue["queueSubjectFingerprint"],
                subject_id_value=str(superseded.get("subjectId") or ""),
                subject_version_value=str(superseded.get("subjectVersion") or ""),
                event_type="resolution_superseded",
                prior_state="resolved",
                next_state="superseded",
                cause_fingerprint="subject_removed",
                resolution_entry_semantic_sha=semantic_hash(superseded),
                reviewer=str(superseded.get("reviewer") or ""),
            )
            append_audit_event_once(log_path, event)
        report = artifacts["compile-report.json"]
        has_blockers = any(entry.get("blockers") for entry in report["models"].values())
        session["state"] = STATE_COMPILED_WITH_EXCEPTIONS if has_blockers else STATE_COMPILED
        session["compiler"] = {
            "runAuthorityFingerprint": run_authority["fingerprint"],
            "manifestSemanticSha": artifacts["canonical-row-manifest.json"]["manifestSemanticSha"],
            "queueSubjectFingerprint": queue["queueSubjectFingerprint"],
            "resolutionSemanticSha": artifacts["exception-resolutions.json"]["resolutionSemanticSha"],
        }
        replace_json_artifact_set(run_dir, {"session.json": session})
        return {
            "session": session,
            "compileReport": report,
            "manifest": artifacts["canonical-row-manifest.json"],
            "exceptionQueue": queue,
            "resolutions": artifacts["exception-resolutions.json"],
        }

    def compiler_detail(self, run_id: str) -> dict[str, Any]:
        """Load one coherent compiler snapshot while mutations are excluded."""

        with self._run_locks[run_id]:
            return self._compiler_detail_locked(run_id)

    def _compiler_detail_locked(self, run_id: str) -> dict[str, Any]:
        """Load and cross-validate one coherent compiled artifact set."""

        session = self.load_session(run_id, verify_source=False)
        if session.get("state") not in COMPILER_STATES:
            raise WizardError("Compile canonical rows first.", status=404)
        run_dir = self.run_dir(run_id)
        missing = [name for name in COMPILER_ARTIFACTS if not (run_dir / name).is_file()]
        if missing:
            raise WizardError(f"Compiled artifact set is incomplete: {', '.join(missing)}.", status=409)
        comparator = read_json(run_dir / "comparator-evidence.json")
        manifest = read_json(run_dir / "canonical-row-manifest.json")
        queue = read_json(run_dir / "exception-queue.json")
        resolutions = read_json(run_dir / "exception-resolutions.json")
        report = read_json(run_dir / "compile-report.json")
        try:
            validate_artifact_graph(manifest, report, comparator, queue, resolutions)
        except ValueError as exc:
            raise WizardError(f"Compiler artifact graph is invalid: {exc}", status=409) from exc
        compiler_state = session.get("compiler") or {}
        expected = {
            "runAuthorityFingerprint": str(
                (manifest.get("runAuthorityFingerprint") or {}).get("fingerprint") or ""
            ),
            "manifestSemanticSha": manifest.get("manifestSemanticSha"),
            "queueSubjectFingerprint": queue.get("queueSubjectFingerprint"),
            "resolutionSemanticSha": resolutions.get("resolutionSemanticSha"),
        }
        if any(compiler_state.get(key) != value for key, value in expected.items()):
            raise WizardError("Compiled artifact bindings do not match session state.", status=409)
        if report.get("manifestSemanticSha") != manifest.get("manifestSemanticSha"):
            raise WizardError("Compile report does not bind the current manifest.", status=409)
        if manifest.get("queueSubjectFingerprint") != queue.get("queueSubjectFingerprint"):
            raise WizardError("Manifest does not bind the current exception queue.", status=409)
        if manifest.get("resolutionSemanticSha") != resolutions.get("resolutionSemanticSha"):
            raise WizardError("Manifest does not bind the current resolutions.", status=409)
        if manifest.get("comparatorEvidenceSemanticSha") != comparator.get("comparatorEvidenceSemanticSha"):
            raise WizardError("Manifest does not bind the current comparator evidence.", status=409)
        return {
            "session": session,
            "comparatorEvidence": comparator,
            "manifest": manifest,
            "exceptionQueue": queue,
            "resolutions": resolutions,
            "compileReport": report,
        }

    def compiler_summary(self, run_id: str) -> dict[str, Any]:
        """Return one coherent compact summary while mutations are excluded."""

        with self._run_locks[run_id]:
            return self._compiler_summary_locked(run_id)

    def _compiler_summary_locked(self, run_id: str) -> dict[str, Any]:
        """Return the compact browser summary for one validated compiler run."""

        detail = self.compiler_detail(run_id)
        report = detail["compileReport"]
        queue = detail["exceptionQueue"]
        resolutions = detail["resolutions"]
        valid_subjects = {
            str(entry.get("subjectId") or "")
            for entry in resolutions.get("validEntries") or []
        }
        subjects = list(queue.get("subjects") or [])
        blocker_subjects = {
            str(blocker.get("subjectId") or "")
            for model in (report.get("models") or {}).values()
            for blocker in model.get("blockers") or []
        }

        manifest_actions: Counter[str] = Counter()
        manifest_statuses: Counter[str] = Counter()
        ready_option_rpos: dict[str, set[str]] = defaultdict(set)
        for row in detail["manifest"].get("rows") or []:
            if row.get("family") != "options" or row.get("status") != "ready":
                continue
            rpo = str((row.get("values") or {}).get("rpo") or "").upper()
            if rpo:
                ready_option_rpos[str(row.get("model") or "")].add(rpo)
        for key, count in (report.get("manifestCounts") or {}).items():
            parts = str(key).split("|", 3)
            if len(parts) != 4:
                continue
            _model, _family, action, status = parts
            manifest_actions[action] += int(count)
            manifest_statuses[status] += int(count)

        models = {}
        for model, entry in sorted((report.get("models") or {}).items()):
            models[model] = {
                "mode": entry.get("mode"),
                "compileReady": bool(entry.get("compileReady")),
                "planReady": bool(entry.get("planReady")),
                "writeReady": bool(entry.get("writeReady")),
                "deploymentReady": bool(entry.get("deploymentReady")),
                "blockerCount": len(entry.get("blockers") or []),
                "deferralCount": len(entry.get("deferrals") or []),
                "boundaryReasons": list(entry.get("boundaryReasons") or []),
            }

        exception_states = Counter()
        actionable_count = 0
        for subject in subjects:
            subject_id_value = str(subject.get("subjectId") or "")
            if subject_id_value in valid_subjects:
                item_state = (
                    "resolved_pending_projection"
                    if subject_id_value in blocker_subjects
                    else "resolved"
                )
            else:
                item_state = "open"
            exception_states[item_state] += 1
            actionable_count += bool(
                self._projectable_exception_actions(
                    subject,
                    ready_option_rpos.get(str(subject.get("model") or ""), set()),
                )
            )
        freshness = self._compiler_freshness(run_id, report)
        return {
            "session": detail["session"],
            "compiler": dict(detail["session"].get("compiler") or {}),
            "models": models,
            "counts": {
                "manifest": {
                    "total": sum(int(value) for value in (report.get("manifestCounts") or {}).values()),
                    "byAction": dict(sorted(manifest_actions.items())),
                    "byStatus": dict(sorted(manifest_statuses.items())),
                    "byFamily": dict(sorted((report.get("familyCounts") or {}).items())),
                },
                "exceptions": {
                    "total": len(subjects),
                    "byState": dict(sorted(exception_states.items())),
                    "actionable": actionable_count,
                    "byReason": dict(sorted(Counter(str(subject.get("reasonCode") or "") for subject in subjects).items())),
                    "byFamily": dict(sorted(Counter(str(subject.get("family") or "") for subject in subjects).items())),
                },
                "sourceFeatures": dict(
                    sorted(
                        Counter(
                            str(item.get("disposition") or "")
                            for item in report.get("sourceFeatureCoverage") or []
                        ).items()
                    )
                ),
                "familyCoverage": dict(
                    sorted(
                        Counter(
                            str(item.get("disposition") or "")
                            for item in report.get("familyCoverage") or []
                        ).items()
                    )
                ),
            },
            "freshness": freshness,
        }

    @staticmethod
    def _projectable_exception_actions(
        subject: dict[str, Any],
        available_option_rpos: set[str] | None = None,
    ) -> list[str]:
        """Expose only actions whose typed outcome is complete after recompilation."""

        reason = str(subject.get("reasonCode") or "")
        required_proposal_rpos = {
            str(value).upper()
            for row in subject.get("proposedRows") or []
            for value in (
                row.get("sourceRpo"),
                row.get("targetRpo"),
                row.get("conditionRpo"),
                *(row.get("memberRpos") or []),
            )
            if str(value or "")
        }
        comparator_reasons = {
            "comparator_only_relationship_proposal",
            "comparator_only_rule_group_proposal",
            "comparator_only_exclusive_group_proposal",
            "comparator_only_price_rule_proposal",
            "comparator_only_default_selection_proposal",
        }
        proposal_catalog_complete = (
            reason not in comparator_reasons
            or (
                bool(required_proposal_rpos)
                and available_option_rpos is not None
                and required_proposal_rpos <= available_option_rpos
            )
        )
        projectable: list[str] = []
        unsupported_row_actions: list[str] = []
        for raw_action in subject.get("allowedActions") or []:
            action = str(raw_action)
            if action == "choose_section":
                projectable.append(action)
            elif action == "choose_relationship" and reason in {
                "unsupported_relationship_type",
                "unsupported_relationship_direction",
                "comparator_only_relationship_proposal",
            } and proposal_catalog_complete:
                projectable.append(action)
            elif action == "retain_existing" and reason == "ambiguous_existing_identity":
                projectable.append(action)
            elif action == "provide_typed_value" and reason in {
                "unresolved_price_scope",
                "comparator_only_rule_group_proposal",
                "comparator_only_exclusive_group_proposal",
                "comparator_only_price_rule_proposal",
                "comparator_only_default_selection_proposal",
            } and proposal_catalog_complete:
                projectable.append(action)
            elif action == "mark_not_applicable" and reason in {
                "unsupported_relationship_type",
                "unsupported_relationship_direction",
                "comparator_only_relationship_proposal",
                "comparator_only_rule_group_proposal",
                "comparator_only_exclusive_group_proposal",
                "comparator_only_price_rule_proposal",
                "comparator_only_default_selection_proposal",
            } and proposal_catalog_complete:
                projectable.append(action)
            else:
                unsupported_row_actions.append(action)
        if unsupported_row_actions:
            return []
        return projectable

    @staticmethod
    def _unique_option_rpos(options: list[dict[str, Any]]) -> set[str]:
        """Return RPOs backed by exactly one ready target option identity."""

        counts = Counter(
            str(option.get("rpo") or "").upper()
            for option in options
            if str(option.get("rpo") or "")
        )
        return {rpo for rpo, count in counts.items() if count == 1}

    @staticmethod
    def _presentation_option(
        rpo: str,
        model_options: list[dict[str, Any]],
        source_rows: list[dict[str, Any]],
    ) -> dict[str, str]:
        """Return non-authoritative display copy for one structured option reference."""

        normalized = str(rpo or "").strip().upper()
        matches = [
            option
            for option in model_options
            if str(option.get("rpo") or "").strip().upper() == normalized
        ]
        source_match = next(
            (
                row
                for row in source_rows
                if str(row.get("rpo") or row.get("refOnlyRpo") or "").strip().upper()
                == normalized
                and str(row.get("description") or "").strip()
            ),
            None,
        )
        label = str(
            (matches[0].get("description") if matches else "")
            or ((source_match or {}).get("description") or "")
            or (matches[0].get("name") if matches else "")
            or "Description unavailable"
        ).strip()
        return {
            "optionId": str(matches[0].get("optionId") or "") if len(matches) == 1 else "",
            "rpo": normalized,
            "label": label,
        }

    @staticmethod
    def _presentation_option_text(option: dict[str, str]) -> str:
        rpo = str(option.get("rpo") or "").strip()
        label = str(option.get("label") or "Description unavailable").strip()
        return f"{rpo} — {label}" if rpo else label

    def _exception_presentation(
        self,
        item: dict[str, Any],
        model_options: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Project one conservative, display-only decision description."""

        subject = item["subject"]
        evidence = item.get("evidence") or {}
        source_rows = list(evidence.get("sourceEvidence") or [])
        proposed_rows = list(subject.get("proposedRows") or [])
        proposed = proposed_rows[0] if proposed_rows else {}
        decision_type = str(item.get("decisionType") or "source_or_tooling")
        model = str(subject.get("model") or "target").replace("_", " ").upper()

        rpos: list[str] = []
        for row in proposed_rows:
            for value in (
                row.get("sourceRpo"),
                row.get("targetRpo"),
                row.get("conditionRpo"),
                *(row.get("memberRpos") or []),
            ):
                normalized = str(value or "").strip().upper()
                if normalized and normalized not in rpos:
                    rpos.append(normalized)
        if not rpos:
            for row in source_rows:
                normalized = str(row.get("rpo") or row.get("refOnlyRpo") or "").strip().upper()
                if normalized and normalized not in rpos:
                    rpos.append(normalized)

        options = [
            self._presentation_option(rpo, model_options, source_rows) for rpo in rpos
        ]
        by_rpo = {option["rpo"]: option for option in options}

        def option_text(rpo: Any) -> str:
            normalized = str(rpo or "").strip().upper()
            option = by_rpo.get(normalized)
            if option is None:
                option = self._presentation_option(normalized, model_options, source_rows)
            return self._presentation_option_text(option)

        def value(record: dict[str, Any], camel: str, snake: str) -> Any:
            return record.get(camel) if record.get(camel) not in (None, "") else record.get(snake)

        def group_phrase(group_type: Any) -> str:
            return {
                "requires_any": "requires at least one of",
                "requires_all": "requires all of",
                "includes_all": "includes",
                "includes_any": "includes one of",
                "excludes_all": "makes unavailable",
                "excludes_any": "makes any of these unavailable",
            }.get(str(group_type or ""), "applies this rule to")

        def fact_summary(row: dict[str, Any]) -> str:
            values = row.get("values") or row.get("signature") or row
            source_rpo = value(values, "sourceRpo", "source_rpo")
            target_rpo = value(values, "targetRpo", "target_rpo")
            condition_rpo = value(values, "conditionRpo", "condition_rpo")
            members = value(values, "memberRpos", "member_rpos") or []
            group_type = value(values, "groupType", "group_type")
            selection_mode = value(values, "selectionMode", "selection_mode")
            rule_type = value(values, "ruleType", "rule_type")
            if source_rpo and members:
                member_text = ", ".join(option_text(rpo) for rpo in members)
                return f"Selecting {option_text(source_rpo)} {group_phrase(group_type)}: {member_text}."
            if members:
                member_text = ", ".join(option_text(rpo) for rpo in members)
                exact = str(selection_mode or "").startswith("required")
                return f"Choose {'exactly' if exact else 'at most'} one of: {member_text}."
            if source_rpo and target_rpo:
                behavior = {
                    "requires": "requires",
                    "includes": "includes",
                    "excludes": "makes unavailable",
                    "replaces": "replaces",
                }.get(str(rule_type or ""), "has a relationship with")
                return f"Selecting {option_text(source_rpo)} {behavior} {option_text(target_rpo)}."
            if target_rpo or condition_rpo:
                target_text = option_text(target_rpo or condition_rpo)
                return f"Applies a target rule to {target_text}."
            return "The target workbook contains a rule with a different structured effect."

        decision_names = {
            "section": "section placement",
            "identity": "option identity",
            "relationship": "option relationship",
            "rule_group": "group rule",
            "exclusive_group": "exclusive choices",
            "default": "default selection",
            "price": "price rule",
            "source_or_tooling": "source requirement",
        }
        title = f"{model} {decision_names.get(decision_type, 'decision')}"
        comparison = None

        if item.get("reviewState") == "semantic_conflict":
            conflict = subject.get("semanticConflict") or {}
            overlap_kind = str(conflict.get("overlapKind") or "different_behavior")
            overlap = {
                "member_set_mismatch": "different member set",
                "subset": "proposed members are a subset",
                "superset": "proposed members are a superset",
                "partial_overlap": "partial overlap",
                "different_relationship": "different relationship",
                "reverse_relationship": "reverse relationship",
            }.get(overlap_kind, overlap_kind.replace("_", " "))
            summary = "The proposal and target workbook describe incompatible behavior."
            existing = list(evidence.get("existingWorkbookRows") or [])
            derived = list(evidence.get("alreadyDerivedRows") or [])
            existing_members = list(conflict.get("existingMemberRpos") or [])
            proposed_members = list(conflict.get("proposedMemberRpos") or [])
            if existing_members or proposed_members:
                group_row = next(
                    (
                        row.get("values") or {}
                        for row in existing
                        if (row.get("values") or {}).get("group_type")
                    ),
                    {},
                )
                source_id = str(group_row.get("source_id") or "")
                source_rpo = next(
                    (
                        str(option.get("rpo") or "")
                        for option in model_options
                        if str(option.get("optionId") or "") == source_id
                    ),
                    str(proposed.get("sourceRpo") or ""),
                )
                existing_fact = {
                    "sourceRpo": source_rpo,
                    "memberRpos": existing_members,
                    "groupType": group_row.get("group_type") or proposed.get("groupType"),
                }
                proposed_fact = {
                    **proposed,
                    "memberRpos": proposed_members or proposed.get("memberRpos") or [],
                }
                existing_text = fact_summary(existing_fact)
                proposed_text = fact_summary(proposed_fact)
            else:
                existing_text = " ".join(fact_summary(row) for row in (existing or derived)) if (existing or derived) else "No exact target row is available."
                proposed_text = " ".join(fact_summary(row) for row in proposed_rows) if proposed_rows else "No exact proposal row is available."
            comparison = {
                "existing": existing_text,
                "proposed": proposed_text,
                "difference": overlap,
            }
        elif decision_type == "section":
            subject_option = options[0] if options else self._presentation_option("", [], source_rows)
            if not options and source_rows:
                subject_option = {
                    "optionId": "",
                    "rpo": "",
                    "label": str(source_rows[0].get("description") or "Description unavailable"),
                }
            summary = f"Place {self._presentation_option_text(subject_option)} in one target section."
        elif decision_type == "identity":
            summary = f"Match {option_text(rpos[0] if rpos else '')} to one existing target option."
        elif decision_type == "relationship":
            behavior = {
                "requires": "requires",
                "includes": "includes",
                "excludes": "makes unavailable",
                "replaces": "replaces",
            }.get(str(proposed.get("ruleType") or ""), "has a relationship with")
            summary = (
                f"When {option_text(proposed.get('sourceRpo'))} is selected, "
                f"it {behavior} {option_text(proposed.get('targetRpo'))}."
            )
        elif decision_type == "rule_group":
            members = ", ".join(option_text(rpo) for rpo in proposed.get("memberRpos") or [])
            summary = (
                f"Selecting {option_text(proposed.get('sourceRpo'))} "
                f"{group_phrase(proposed.get('groupType'))}: {members or 'Description unavailable'}."
            )
        elif decision_type == "exclusive_group":
            members = ", ".join(option_text(rpo) for rpo in proposed.get("memberRpos") or [])
            exact = str(proposed.get("selectionMode") or "").startswith("required")
            summary = f"Choose {'exactly' if exact else 'at most'} one of: {members or 'Description unavailable'}."
        elif decision_type == "default":
            summary = f"{option_text(proposed.get('targetRpo') or proposed.get('rpo'))} is selected by default."
        elif decision_type == "price":
            target = proposed.get("targetRpo") or (rpos[0] if rpos else "")
            condition = proposed.get("conditionRpo")
            comparator = list(evidence.get("comparator") or [])
            price = proposed.get("priceValue")
            if price in (None, "") and comparator:
                price = (comparator[0].get("context") or {}).get("priceValue")
            price_text = f"${price}" if price not in (None, "") else "the reviewed whole-dollar price"
            condition_text = f"When {option_text(condition)} applies, " if condition else ""
            summary = f"{condition_text}{option_text(target)} costs {price_text} for the selected target scope."
        else:
            summary = str(subject.get("question") or "Provide the missing target evidence before this can compile.")

        source_label = ""
        for row in source_rows:
            source_rpo = str(row.get("rpo") or row.get("refOnlyRpo") or "").strip()
            source_description = str(row.get("description") or "").strip()
            if source_rpo:
                source_label = self._presentation_option_text(
                    self._presentation_option(source_rpo, model_options, source_rows)
                )
            elif source_description:
                source_label = source_description
            if source_label:
                break
        if source_label:
            why_asked = f"The target source identifies {source_label}."
        elif evidence.get("workbookReferences"):
            why_asked = "The target workbook contains evidence that still needs an exact decision."
        else:
            why_asked = "The current target evidence does not produce one complete workbook-safe answer."
        if evidence.get("comparator"):
            why_asked += " Comparator suggestion is supporting evidence only."

        return {
            "title": title,
            "summary": summary,
            "whyAsked": why_asked,
            "options": options,
            "comparison": comparison,
        }

    def _compiler_freshness(self, run_id: str, report: dict[str, Any]) -> dict[str, Any]:
        """Compare current compiler inputs to the authority bound into the report."""

        run_dir = self.run_dir(run_id)
        session = self.load_session(run_id, verify_source=False)
        paths = {
            "source": Path(session["sourcePath"]),
            "workbook": self.workbook_path,
            "sheetRoles": run_dir / "sheet-roles.json",
            "optionCandidates": run_dir / "option-candidates.json",
            "priceRows": run_dir / "price-rows.json",
            "joinReport": run_dir / "join-report.json",
            "modelSelection": run_dir / "model-selection.json",
        }
        expected = (
            report.get("runAuthorityFingerprint", {})
            .get("bindings", {})
            .get("files", {})
        )
        reasons = []
        for label, path in paths.items():
            if not path.is_file():
                reasons.append(f"{label} is missing")
                continue
            if expected.get(label) != file_fingerprint(path):
                reasons.append(f"{label} changed after compile")
        return {
            "stale": bool(reasons),
            "recompileRequired": bool(reasons),
            "reasons": reasons,
        }

    def exception_queue_view(
        self,
        run_id: str,
        *,
        model: str = "",
        decision_type: str = "",
        affected_sheet: str = "",
        review_state: str = "",
        family: str = "",
        reason: str = "",
        severity: str = "",
        state: str = "",
        actionable: str = "",
        query: str = "",
        offset: int = 0,
        limit: int = 50,
    ) -> dict[str, Any]:
        """Return one coherent exception page while mutations are excluded."""

        with self._run_locks[run_id]:
            return self._exception_queue_view_locked(
                run_id,
                model=model,
                decision_type=decision_type,
                affected_sheet=affected_sheet,
                review_state=review_state,
                family=family,
                reason=reason,
                severity=severity,
                state=state,
                actionable=actionable,
                query=query,
                offset=offset,
                limit=limit,
            )

    def _exception_queue_view_locked(
        self,
        run_id: str,
        *,
        model: str = "",
        decision_type: str = "",
        affected_sheet: str = "",
        review_state: str = "",
        family: str = "",
        reason: str = "",
        severity: str = "",
        state: str = "",
        actionable: str = "",
        query: str = "",
        offset: int = 0,
        limit: int = 50,
    ) -> dict[str, Any]:
        """Return one deterministic browser page of validated exception subjects."""

        if state not in {"", "open", "resolved", "resolved_pending_projection", "stale_reopened"}:
            raise WizardError(f"Unknown exception state filter: {state}")
        if actionable not in {"", "yes", "no"}:
            raise WizardError(f"Unknown actionable filter: {actionable}")
        if review_state not in {
            "",
            "needs_decision",
            "prerequisite_blocked",
            "semantic_conflict",
            "source_or_tooling_blocked",
            "resolved",
            "pending_projection",
            "stale_reopened",
        }:
            raise WizardError(f"Unknown review state filter: {review_state}")
        try:
            offset = max(0, int(offset))
            limit = min(100, max(1, int(limit)))
        except (TypeError, ValueError) as exc:
            raise WizardError("Exception offset and limit must be integers.") from exc

        detail = self.compiler_detail(run_id)
        freshness = self._compiler_freshness(run_id, detail["compileReport"])
        if freshness["stale"]:
            raise WizardError(
                "Compiler inputs changed after this run was compiled. Recompile before reviewing exceptions.",
                status=409,
            )
        subjects = list(detail["exceptionQueue"].get("subjects") or [])
        run_dir = self.run_dir(run_id)
        candidates_payload = read_json(run_dir / "option-candidates.json")
        selection = read_json(run_dir / "model-selection.json")
        manifest_rows = list(detail["manifest"].get("rows") or [])
        raw_candidates: dict[tuple[str, str], list[dict[str, Any]]] = {}
        target_variant_pairs: dict[str, set[tuple[str, str]]] = {}
        for target in selection.get("targets") or []:
            for candidate in scope_candidates(candidates_payload.get("candidates") or [], str(target)):
                scoped = {
                    **candidate,
                    "targetModel": str(target),
                    "statuses": model_scoped_statuses(candidate, str(target)),
                }
                for status_entry in scoped["statuses"]:
                    body_style = str(status_entry.get("bodyStyle") or "").strip()
                    trim_level = str(status_entry.get("trim") or "").strip()
                    if body_style or trim_level:
                        target_variant_pairs.setdefault(str(target), set()).add(
                            (body_style, trim_level)
                        )
                signatures = {
                    option_occurrence_signature(candidate),
                    option_occurrence_signature(scoped),
                }
                for signature in signatures:
                    raw_candidates.setdefault((str(target), signature), []).append(dict(candidate))
        target_price_scopes: dict[str, list[dict[str, str]]] = {}
        for target in selection.get("targets") or []:
            pairs = target_variant_pairs.get(str(target), set())
            scope_triples = {("*", "*", "*")}
            scope_triples.update((body_style or "*", trim_level or "*", "*") for body_style, trim_level in pairs)
            scope_triples.update((body_style, "*", "*") for body_style, _ in pairs if body_style)
            scope_triples.update(("*", trim_level, "*") for _, trim_level in pairs if trim_level)
            scope_triples.update(
                (
                    str(values.get("body_style") or "*"),
                    str(values.get("trim_level") or "*"),
                    str(values.get("variant_id") or "*"),
                )
                for row in manifest_rows
                if row.get("model") == target
                and row.get("family") == "variant_master"
                and row.get("status") == "ready"
                for values in [row.get("values") or {}]
                if str(values.get("variant_id") or "")
            )
            target_price_scopes[str(target)] = [
                {
                    "bodyStyleScope": body_style,
                    "trimLevelScope": trim_level,
                    "variantScope": variant_scope,
                    "label": (
                        "All target variants (*)"
                        if (body_style, trim_level, variant_scope) == ("*", "*", "*")
                        else " · ".join(value for value in (body_style, trim_level, variant_scope) if value != "*")
                    ),
                }
                for body_style, trim_level, variant_scope in sorted(scope_triples)
            ]
        sections = workbook_sections(self._require_workbook())
        comparator_facts = {
            str(fact.get("evidenceId") or ""): dict(fact)
            for entry in (detail["comparatorEvidence"].get("targets") or {}).values()
            for fact in entry.get("facts") or []
        }
        target_options: dict[str, list[dict[str, Any]]] = {}
        for target in selection.get("targets") or []:
            option_rows = [
                dict(row.get("values") or {})
                for row in manifest_rows
                if row.get("model") == target and row.get("family") == "options"
                and row.get("status") == "ready"
            ]
            by_id: dict[str, dict[str, Any]] = {}
            for row in option_rows:
                option_id = str(row.get("option_id") or "")
                if option_id:
                    by_id[option_id] = {
                        "optionId": option_id,
                        "rpo": str(row.get("rpo") or ""),
                        "name": str(
                            row.get("name")
                            or row.get("option_name")
                            or row.get("description")
                            or ""
                        ),
                        "description": str(
                            row.get("description")
                            or row.get("detail_raw")
                            or row.get("name")
                            or row.get("option_name")
                            or ""
                        ),
                        "sectionId": str(row.get("section_id") or ""),
                    }
            target_options[str(target)] = [by_id[key] for key in sorted(by_id)]
        target_unique_option_rpos = {
            target: self._unique_option_rpos(options)
            for target, options in target_options.items()
        }
        registry = build_family_registry(self._require_workbook(), selection.get("targets") or [])
        family_sheets: dict[tuple[str, str], set[str]] = defaultdict(set)
        for target, entries in registry.items():
            for entry in entries.values():
                family_sheets[(str(target), str(entry.get("family") or ""))].add(
                    str(entry.get("sheetName") or "")
                )
        family_sheets.update(
            {
                (str(target), "default_selection_rules"): {"default_selection_rules"}
                for target in selection.get("targets") or []
            }
        )
        valid_by_subject = {
            str(entry.get("subjectId") or ""): dict(entry)
            for entry in detail["resolutions"].get("validEntries") or []
        }
        blocker_subjects = {
            str(blocker.get("subjectId") or "")
            for model_entry in (detail["compileReport"].get("models") or {}).values()
            for blocker in model_entry.get("blockers") or []
        }
        stale_by_subject: dict[str, list[dict[str, Any]]] = {}
        superseded_by_subject: dict[str, list[dict[str, Any]]] = {}
        for field, target in (
            ("staleEntries", stale_by_subject),
            ("supersededEntries", superseded_by_subject),
        ):
            for entry in detail["resolutions"].get(field) or []:
                target.setdefault(str(entry.get("subjectId") or ""), []).append(dict(entry))

        items = []
        decision_type_catalog: set[str] = set()
        affected_sheet_catalog: set[str] = set()
        review_state_catalog: set[str] = set()
        for subject in sorted(subjects, key=lambda item: str(item.get("subjectId") or "")):
            subject_id_value = str(subject.get("subjectId") or "")
            if subject_id_value in valid_by_subject:
                item_state = (
                    "resolved_pending_projection"
                    if subject_id_value in blocker_subjects
                    else "resolved"
                )
            elif stale_by_subject.get(subject_id_value):
                item_state = "stale_reopened"
            else:
                item_state = "open"
            model_options = target_options.get(str(subject.get("model") or ""), [])
            available_actions = self._projectable_exception_actions(
                subject,
                target_unique_option_rpos.get(str(subject.get("model") or ""), set()),
            )
            item = {
                "subject": dict(subject),
                "state": item_state,
                "availableActions": available_actions,
                "resolution": valid_by_subject.get(subject_id_value),
                "history": {
                    "stale": stale_by_subject.get(subject_id_value, []),
                    "superseded": superseded_by_subject.get(subject_id_value, []),
                },
            }
            evidence_ids = {
                str(dependency.get("evidenceId") or "")
                for dependency in subject.get("evidenceDependencies") or []
            }
            evidence_references = {
                str(reference) for reference in subject.get("evidenceReferences") or []
            }
            candidate_signatures = {
                evidence_id.removeprefix("candidate:")
                for evidence_id in evidence_ids
                if evidence_id.startswith("candidate:")
            }
            candidate_signatures.update(
                evidence_id.split(":candidate:", 1)[1]
                for evidence_id in evidence_ids
                if ":candidate:" in evidence_id
            )
            raw_rows: list[dict[str, Any]] = []
            seen_candidates: set[str] = set()
            for signature in sorted(candidate_signatures):
                for candidate in raw_candidates.get((str(subject.get("model") or ""), signature), []):
                    candidate_id = str(candidate.get("candidateId") or "")
                    if candidate_id not in seen_candidates:
                        seen_candidates.add(candidate_id)
                        raw_rows.append(candidate)
            target_rows = [
                row
                for row in manifest_rows
                if row.get("model") in {subject.get("model"), "*"}
                and evidence_ids
                & {
                    str(dependency.get("evidenceId") or "")
                    for dependency in row.get("evidenceDependencies") or []
                }
            ]
            conflict_rows = list((subject.get("semanticConflict") or {}).get("conflictingRows") or [])
            conflict_keys = {
                (str(row.get("sheet") or ""), canonical_text(row.get("key") or {}))
                for row in conflict_rows
            }
            existing_rows = [
                row
                for row in conflict_rows
                if (
                    row.get("action") == "noop"
                    or row.get("disposition") == "retained_existing"
                )
            ]
            already_derived_rows = [
                row
                for row in conflict_rows
                if row not in existing_rows
            ]
            shared_context = [
                row
                for row in target_rows
                if (str(row.get("sheet") or ""), canonical_text(row.get("key") or {})) not in conflict_keys
            ]
            item["evidence"] = {
                "sourceEvidence": raw_rows,
                "existingWorkbookRows": existing_rows,
                "alreadyDerivedRows": already_derived_rows,
                "sharedContext": shared_context,
                "comparator": [
                    comparator_facts[evidence_id]
                    for evidence_id in sorted(evidence_ids | evidence_references)
                    if evidence_id in comparator_facts
                ],
                "workbookReferences": sorted(
                    evidence_id
                    for evidence_id in evidence_ids
                    if evidence_id.startswith("workbook:")
                ),
            }
            reason_code = str(subject.get("reasonCode") or "")
            if reason_code in {"missing_section"}:
                decision_name = "section"
                effect_families = {"options"}
            elif reason_code == "ambiguous_existing_identity":
                decision_name = "identity"
                effect_families = {"options"}
            elif "exclusive_group" in reason_code or (
                reason_code == "semantic_group_overlap"
                and str(subject.get("originalReasonCode") or "").endswith("exclusive_group_proposal")
            ):
                decision_name = "exclusive_group"
                effect_families = {"exclusive_groups", "exclusive_members"}
            elif "rule_group" in reason_code or reason_code == "semantic_group_overlap":
                decision_name = "rule_group"
                effect_families = {"rule_groups", "rule_group_members"}
            elif "relationship" in reason_code or reason_code == "semantic_relationship_conflict":
                decision_name = "relationship"
                effect_families = {"rule_mapping"}
            elif "price" in reason_code:
                decision_name = "price"
                effect_families = {"price_rules" if "price_rule" in reason_code else "options"}
            elif "default" in reason_code:
                decision_name = "default"
                effect_families = {"default_selection_rules"}
            else:
                decision_name = "source_or_tooling"
                effect_families = {str(subject.get("family") or "")}
            affected_sheets = {
                str(value)
                for value in (subject.get("semanticConflict") or {}).get("affectedSheets") or []
                if str(value)
            }
            for effect_family in effect_families:
                affected_sheets.update(
                    family_sheets.get((str(subject.get("model") or ""), effect_family), set())
                )
            required_rpos = {
                str(value).upper()
                for row in subject.get("proposedRows") or []
                for value in (
                    row.get("sourceRpo"),
                    row.get("targetRpo"),
                    row.get("conditionRpo"),
                    *(row.get("memberRpos") or []),
                )
                if str(value or "")
            }
            missing_or_ambiguous = sorted(
                required_rpos
                - target_unique_option_rpos.get(str(subject.get("model") or ""), set())
            )
            item["decisionType"] = decision_name
            item["affectedSheets"] = sorted(affected_sheets)
            item["prerequisites"] = {
                "missingOrAmbiguousRpos": missing_or_ambiguous,
                "message": (
                    "Resolve a unique ready target option for: " + ", ".join(missing_or_ambiguous)
                    if missing_or_ambiguous
                    else ""
                ),
            }
            allowed_actions = set(subject.get("allowedActions") or [])
            item["choices"] = {
                "sections": list(sections) if "choose_section" in allowed_actions else [],
                "relationshipRuleTypes": (
                    ["requires", "includes", "excludes", "replaces"]
                    if "choose_relationship" in allowed_actions
                    else []
                ),
                "targetOptions": model_options if "choose_relationship" in allowed_actions else [],
                "existingOptions": (
                    [
                        option
                        for option in model_options
                        if option["optionId"]
                        in {
                            str(row.get("existingId") or "")
                            for row in subject.get("proposedRows") or []
                        }
                    ]
                    if "retain_existing" in allowed_actions
                    else []
                ),
                "exclusiveSelectionModes": (
                    ["single_within_group", "required_single_within_group"]
                    if subject.get("reasonCode") == "comparator_only_exclusive_group_proposal"
                    and "provide_typed_value" in available_actions
                    else []
                ),
                "defaultDisplayBehaviors": (
                    ["default_selected", ""]
                    if subject.get("reasonCode") == "comparator_only_default_selection_proposal"
                    and "provide_typed_value" in available_actions
                    else []
                ),
                "deferralKinds": (
                    sorted(ALLOWED_DEFERRAL_KINDS)
                    if "record_allowed_deferral" in allowed_actions
                    else []
                ),
                "priceScopes": (
                    target_price_scopes.get(str(subject.get("model") or ""), [])
                    if "provide_typed_value" in available_actions
                    else []
                ),
            }
            if item["state"] == "resolved":
                item["reviewState"] = "resolved"
            elif item["state"] == "resolved_pending_projection":
                item["reviewState"] = "pending_projection"
            elif item["state"] == "stale_reopened":
                item["reviewState"] = "stale_reopened"
            elif str(subject.get("reasonCode") or "").startswith("semantic_"):
                item["reviewState"] = "semantic_conflict"
            elif not available_actions and missing_or_ambiguous:
                item["reviewState"] = "prerequisite_blocked"
            elif not available_actions:
                item["reviewState"] = "source_or_tooling_blocked"
            else:
                item["reviewState"] = "needs_decision"
            item["presentation"] = self._exception_presentation(item, model_options)
            decision_type_catalog.add(item["decisionType"])
            affected_sheet_catalog.update(item["affectedSheets"])
            review_state_catalog.add(item["reviewState"])
            if model and subject.get("model") != model:
                continue
            if decision_type and item["decisionType"] != decision_type:
                continue
            if affected_sheet and affected_sheet not in item["affectedSheets"]:
                continue
            if review_state and item["reviewState"] != review_state:
                continue
            if family and subject.get("family") != family:
                continue
            if reason and subject.get("reasonCode") != reason:
                continue
            if severity and subject.get("severity") != severity:
                continue
            if state and item_state != state:
                continue
            is_actionable = bool(available_actions)
            if actionable == "yes" and not is_actionable:
                continue
            if actionable == "no" and is_actionable:
                continue
            needle = str(query or "").strip().lower()
            if needle:
                search_item = {
                    **item,
                    "evidence": {
                        key: value
                        for key, value in (item.get("evidence") or {}).items()
                        if key != "sharedContext"
                    },
                }
                haystack = canonical_text(search_item).lower()
                if needle not in haystack:
                    continue
            items.append(item)

        filters = {
            "models": sorted({str(subject.get("model") or "") for subject in subjects}),
            "decisionTypes": sorted(decision_type_catalog),
            "affectedSheets": sorted(affected_sheet_catalog),
            "reviewStates": sorted(review_state_catalog),
            "families": sorted({str(subject.get("family") or "") for subject in subjects}),
            "reasons": sorted({str(subject.get("reasonCode") or "") for subject in subjects}),
            "severities": sorted({str(subject.get("severity") or "") for subject in subjects}),
        }
        return {
            "runId": run_id,
            "queueSubjectFingerprint": detail["exceptionQueue"].get("queueSubjectFingerprint"),
            "freshness": freshness,
            "total": len(items),
            "offset": offset,
            "limit": limit,
            "items": items[offset : offset + limit],
            "filters": filters,
        }

    def _validated_exception_resolution_locked(
        self,
        run_id: str,
        *,
        subject_id: str,
        subject_version: str,
        action: str,
        payload: dict[str, Any],
        reviewer: str,
    ) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any]]:
        """Validate one current typed choice for both preview and persistence."""

        detail = self.compiler_detail(run_id)
        subjects = {
            str(subject.get("subjectId") or ""): subject
            for subject in detail["exceptionQueue"].get("subjects") or []
        }
        subject = subjects.get(str(subject_id))
        if subject is None:
            raise WizardError("Exception subject is not current; reload the queue.", status=409)
        if str(subject.get("subjectVersion") or "") != str(subject_version):
            raise WizardError("Exception subject version is stale; reload the queue.", status=409)
        freshness = self._compiler_freshness(run_id, detail["compileReport"])
        if freshness["stale"]:
            raise WizardError(
                "Compiler inputs changed; recompile before resolving exceptions: "
                + ", ".join(freshness["reasons"]),
                status=409,
            )
        current_view = self._exception_queue_view_locked(
            run_id,
            query=str(subject_id),
            limit=100,
        )
        subject_view = next(
            (
                item
                for item in current_view["items"]
                if item["subject"].get("subjectId") == subject_id
            ),
            None,
        )
        if subject_view is None:
            raise WizardError("Exception subject is not current; reload the queue.", status=409)
        if action not in subject_view["availableActions"]:
            raise WizardError(
                "This typed action is not yet projectable into a complete canonical outcome.",
                status=409,
            )
        reviewer = str(reviewer or "").strip()
        if not reviewer:
            raise WizardError("Exception resolution needs a reviewer name.")
        if any(
            entry.get("subjectId") == subject_id
            and entry.get("subjectVersion") == subject_version
            for entry in detail["resolutions"].get("validEntries") or []
        ):
            raise WizardError("Exception is already resolved; reopen it before replacing the answer.", status=409)
        resolution = {
            "subjectId": str(subject_id),
            "subjectVersion": str(subject_version),
            "action": str(action),
            "payload": dict(payload or {}),
            "disposition": ACTION_DISPOSITIONS.get(str(action), ""),
            "reviewer": reviewer,
            "resolvedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "evidenceReferences": list(subject.get("evidenceReferences") or []),
        }
        try:
            resolution = validate_resolution(resolution, subject)
        except ValueError as exc:
            raise WizardError(f"Invalid exception resolution: {exc}") from exc
        if action == "choose_section":
            allowed_sections = {
                str(section.get("sectionId") or "")
                for section in workbook_sections(self._require_workbook())
            }
            if resolution["payload"]["sectionId"] not in allowed_sections:
                raise WizardError("Section is not a current canonical workbook choice.")
        if action == "provide_typed_value" and subject.get("reasonCode") in {
            "missing_price_scope",
            "unresolved_price_scope",
            "comparator_only_price_rule_proposal",
        }:
            allowed_scopes = {
                (
                    str(scope.get("bodyStyleScope") or ""),
                    str(scope.get("trimLevelScope") or ""),
                    str(scope.get("variantScope") or ""),
                )
                for scope in subject_view["choices"].get("priceScopes") or []
            }
            requested_scope = (
                str(resolution["payload"].get("bodyStyleScope") or ""),
                str(resolution["payload"].get("trimLevelScope") or ""),
                str(resolution["payload"].get("variantScope") or ""),
            )
            if requested_scope not in allowed_scopes:
                raise WizardError("Price scope is not a current target variant or canonical wildcard choice.")
        if action in {"choose_relationship", "retain_existing"}:
            choices = subject_view["choices"]
            if action == "choose_relationship":
                allowed_ids = {
                    str(option.get("optionId") or "")
                    for option in choices.get("targetOptions") or []
                }
                if (
                    resolution["payload"]["sourceOptionId"] not in allowed_ids
                    or resolution["payload"]["targetOptionId"] not in allowed_ids
                    or resolution["payload"]["ruleType"]
                    not in set(choices.get("relationshipRuleTypes") or [])
                ):
                    raise WizardError("Relationship resolution is not made from current target choices.")
            if action == "retain_existing":
                allowed_ids = {
                    str(option.get("optionId") or "")
                    for option in choices.get("existingOptions") or []
                }
                if resolution["payload"]["existingId"] not in allowed_ids:
                    raise WizardError("Existing row is not a current target identity choice.")
        return detail, subject, subject_view, resolution

    def preview_exception(
        self,
        run_id: str,
        *,
        subject_id: str,
        subject_version: str,
        action: str,
        payload: dict[str, Any],
    ) -> dict[str, Any]:
        """Compile a typed choice in memory and return its exact physical delta."""

        with self._run_locks[run_id]:
            detail, subject, subject_view, resolution = self._validated_exception_resolution_locked(
                run_id,
                subject_id=subject_id,
                subject_version=subject_version,
                action=action,
                payload=payload,
                reviewer="preview",
            )
            run_dir = self.run_dir(run_id)
            selection = read_json(run_dir / "model-selection.json")
            existing_entries: list[dict[str, Any]] = []
            seen: set[str] = set()
            for field in ("entries", "validEntries", "staleEntries", "supersededEntries"):
                for entry in detail["resolutions"].get(field) or []:
                    identity = semantic_hash(entry)
                    if identity not in seen:
                        seen.add(identity)
                        existing_entries.append(dict(entry))
            try:
                staged = run_canonical_compiler(
                    workbook_path=self._require_workbook(),
                    option_payload=read_json(run_dir / "option-candidates.json"),
                    price_payload=read_json(run_dir / "price-rows.json"),
                    join_report=read_json(run_dir / "join-report.json"),
                    roles_payload=read_json(run_dir / "sheet-roles.json"),
                    sheet_profile=read_json(run_dir / "sheet-profile.json"),
                    selection=selection,
                    comparator_artifact=detail["comparatorEvidence"],
                    run_authority_fingerprint=detail["manifest"]["runAuthorityFingerprint"],
                    resolution_entries=[*existing_entries, resolution],
                )
            except (ValueError, KeyError) as exc:
                raise WizardError(f"Compiler refused the preview: {exc}", status=409) from exc

            def row_map(rows: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
                return {
                    (str(row.get("sheet") or ""), canonical_text(row.get("key") or {})): row
                    for row in rows
                }

            before_rows = row_map(list(detail["manifest"].get("rows") or []))
            after_rows = row_map(list(staged["canonical-row-manifest.json"].get("rows") or []))
            row_effects = []
            for key in sorted(set(before_rows) | set(after_rows)):
                before_row = before_rows.get(key)
                after_row = after_rows.get(key)
                if before_row == after_row:
                    continue
                effect = "added" if before_row is None else "removed" if after_row is None else "changed"
                row_effects.append({"effect": effect, "before": before_row, "after": after_row})

            def blocker_ids(report: dict[str, Any]) -> set[str]:
                return {
                    str(blocker.get("subjectId") or "")
                    for model in (report.get("models") or {}).values()
                    for blocker in model.get("blockers") or []
                }

            before_blockers = blocker_ids(detail["compileReport"])
            after_blockers = blocker_ids(staged["compile-report.json"])
            removed_blockers = sorted(before_blockers - after_blockers)
            added_blockers = sorted(after_blockers - before_blockers)
            model = str(subject.get("model") or "")
            return {
                "runId": str(run_id),
                "subjectId": str(subject["subjectId"]),
                "subjectVersion": str(subject["subjectVersion"]),
                "action": action,
                "projectable": True,
                "affectedSheets": list(subject_view.get("affectedSheets") or []),
                "conflicts": list(subject.get("semanticConflicts") or []),
                "prerequisites": dict(subject_view.get("prerequisites") or {}),
                "decisionEffect": {
                    "writesRows": bool(row_effects),
                    "rows": row_effects,
                    "suppressedProposalRows": (
                        list(subject.get("proposedRows") or [])
                        if action == "mark_not_applicable"
                        else []
                    ),
                    "blockers": {
                        "removed": removed_blockers,
                        "added": added_blockers,
                        "changed": [],
                    },
                    "readiness": {
                        "before": dict(
                            (detail["compileReport"].get("models") or {}).get(model) or {}
                        ),
                        "after": dict(
                            (staged["compile-report.json"].get("models") or {}).get(model) or {}
                        ),
                    },
                    "dispositions": {"changed": []},
                    "removedBlockerSubjectIds": removed_blockers,
                    "addedBlockerSubjectIds": added_blockers,
                    "writesWorkbook": False,
                },
            }

    def resolve_exception(
        self,
        run_id: str,
        *,
        subject_id: str,
        subject_version: str,
        action: str,
        payload: dict[str, Any],
        reviewer: str,
    ) -> dict[str, Any]:
        """Serialize one exact resolution with the aggregate compiler rewrite."""

        with self._run_locks[run_id]:
            return self._resolve_exception_locked(
                run_id,
                subject_id=subject_id,
                subject_version=subject_version,
                action=action,
                payload=payload,
                reviewer=reviewer,
            )

    def _resolve_exception_locked(
        self,
        run_id: str,
        *,
        subject_id: str,
        subject_version: str,
        action: str,
        payload: dict[str, Any],
        reviewer: str,
    ) -> dict[str, Any]:
        """Record one exact typed resolution and recompile the coherent run."""

        detail = self.compiler_detail(run_id)
        subjects = {
            str(subject.get("subjectId") or ""): subject
            for subject in detail["exceptionQueue"].get("subjects") or []
        }
        subject = subjects.get(str(subject_id))
        if subject is None:
            raise WizardError("Exception subject is not current; reload the queue.", status=409)
        if str(subject.get("subjectVersion") or "") != str(subject_version):
            raise WizardError("Exception subject version is stale; reload the queue.", status=409)
        freshness = self._compiler_freshness(run_id, detail["compileReport"])
        if freshness["stale"]:
            raise WizardError(
                "Compiler inputs changed; recompile before resolving exceptions: "
                + ", ".join(freshness["reasons"]),
                status=409,
            )
        current_view = self._exception_queue_view_locked(
            run_id,
            query=str(subject_id),
            limit=100,
        )
        subject_view = next(
            (
                item
                for item in current_view["items"]
                if item["subject"].get("subjectId") == subject_id
            ),
            None,
        )
        if subject_view is None:
            raise WizardError("Exception subject is not current; reload the queue.", status=409)
        if action not in subject_view["availableActions"]:
            raise WizardError(
                "This typed action is not yet projectable into a complete canonical outcome.",
                status=409,
            )
        reviewer = str(reviewer or "").strip()
        if not reviewer:
            raise WizardError("Exception resolution needs a reviewer name.")
        if any(
            entry.get("subjectId") == subject_id
            and entry.get("subjectVersion") == subject_version
            for entry in detail["resolutions"].get("validEntries") or []
        ):
            raise WizardError("Exception is already resolved; reopen it before replacing the answer.", status=409)
        resolution = {
            "subjectId": str(subject_id),
            "subjectVersion": str(subject_version),
            "action": str(action),
            "payload": dict(payload or {}),
            "disposition": ACTION_DISPOSITIONS.get(str(action), ""),
            "reviewer": reviewer,
            "resolvedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "evidenceReferences": list(subject.get("evidenceReferences") or []),
        }
        try:
            resolution = validate_resolution(resolution, subject)
        except ValueError as exc:
            raise WizardError(f"Invalid exception resolution: {exc}") from exc
        if action == "choose_section":
            allowed_sections = {
                str(section.get("sectionId") or "")
                for section in workbook_sections(self._require_workbook())
            }
            if resolution["payload"]["sectionId"] not in allowed_sections:
                raise WizardError("Section is not a current canonical workbook choice.")
        if action == "provide_typed_value" and subject.get("reasonCode") in {
            "missing_price_scope",
            "unresolved_price_scope",
            "comparator_only_price_rule_proposal",
        }:
            allowed_scopes = {
                (
                    str(scope.get("bodyStyleScope") or ""),
                    str(scope.get("trimLevelScope") or ""),
                    str(scope.get("variantScope") or ""),
                )
                for scope in subject_view["choices"].get("priceScopes") or []
            }
            requested_scope = (
                str(resolution["payload"].get("bodyStyleScope") or ""),
                str(resolution["payload"].get("trimLevelScope") or ""),
                str(resolution["payload"].get("variantScope") or ""),
            )
            if requested_scope not in allowed_scopes:
                raise WizardError("Price scope is not a current target variant or canonical wildcard choice.")
        if action in {"choose_relationship", "retain_existing"}:
            choices = subject_view["choices"]
            if action == "choose_relationship":
                allowed_ids = {
                    str(option.get("optionId") or "")
                    for option in choices.get("targetOptions") or []
                }
                if (
                    resolution["payload"]["sourceOptionId"] not in allowed_ids
                    or resolution["payload"]["targetOptionId"] not in allowed_ids
                    or resolution["payload"]["ruleType"]
                    not in set(choices.get("relationshipRuleTypes") or [])
                ):
                    raise WizardError("Relationship resolution is not made from current target choices.")
            if action == "retain_existing":
                allowed_ids = {
                    str(option.get("optionId") or "")
                    for option in choices.get("existingOptions") or []
                }
                if resolution["payload"]["existingId"] not in allowed_ids:
                    raise WizardError("Existing row is not a current target identity choice.")

        run_dir = self.run_dir(run_id)
        resolution_path = run_dir / "exception-resolutions.json"
        snapshot = self._snapshot_run_files(run_dir, COMPILER_MUTATION_FILES)
        current = read_json(resolution_path)
        current["entries"] = [
            *(current.get("entries") or []),
            resolution,
        ]
        try:
            replace_json_artifact_set(run_dir, {"exception-resolutions.json": current})
            self.compile_canonical_rows(run_id)
            refreshed_detail = self.compiler_detail(run_id)
            accepted_resolution = next(
                (
                    entry
                    for entry in refreshed_detail["resolutions"].get("validEntries") or []
                    if entry.get("subjectId") == subject_id
                    and entry.get("subjectVersion") == subject_version
                ),
                None,
            )
            if accepted_resolution is None:
                raise WizardError("Compiler did not accept the saved exception resolution.", status=409)
            blocker_subjects = {
                str(blocker.get("subjectId") or "")
                for model_entry in (refreshed_detail["compileReport"].get("models") or {}).values()
                for blocker in model_entry.get("blockers") or []
            }
            pending_projection = subject_id in blocker_subjects
            subject_view = {
                **subject_view,
                "state": "resolved_pending_projection" if pending_projection else "resolved",
                "reviewState": "pending_projection" if pending_projection else "resolved",
                "resolution": dict(accepted_resolution),
            }
            result = {"summary": self.compiler_summary(run_id), "subject": subject_view}
        except Exception:
            self._restore_run_files(run_dir, snapshot)
            raise
        return result

    def reopen_exception(
        self,
        run_id: str,
        *,
        subject_id: str,
        subject_version: str,
        reviewer: str,
    ) -> dict[str, Any]:
        """Serialize one reopen with the aggregate compiler rewrite."""

        with self._run_locks[run_id]:
            return self._reopen_exception_locked(
                run_id,
                subject_id=subject_id,
                subject_version=subject_version,
                reviewer=reviewer,
            )

    def _reopen_exception_locked(
        self,
        run_id: str,
        *,
        subject_id: str,
        subject_version: str,
        reviewer: str,
    ) -> dict[str, Any]:
        """Remove one exact current resolution, recompile, and audit the reopen."""

        detail = self.compiler_detail(run_id)
        subjects = {
            str(subject.get("subjectId") or ""): subject
            for subject in detail["exceptionQueue"].get("subjects") or []
        }
        subject = subjects.get(str(subject_id))
        if subject is None or str(subject.get("subjectVersion") or "") != str(subject_version):
            raise WizardError("Exception subject is stale or no longer current; reload the queue.", status=409)
        freshness = self._compiler_freshness(run_id, detail["compileReport"])
        if freshness["stale"]:
            raise WizardError(
                "Compiler inputs changed; recompile before reopening exceptions: "
                + ", ".join(freshness["reasons"]),
                status=409,
            )
        reviewer = str(reviewer or "").strip()
        if not reviewer:
            raise WizardError("Reopening an exception needs a reviewer name.")
        matches = [
            dict(entry)
            for entry in detail["resolutions"].get("validEntries") or []
            if entry.get("subjectId") == subject_id
            and entry.get("subjectVersion") == subject_version
        ]
        if len(matches) != 1:
            raise WizardError("Exception is not currently resolved.", status=409)
        previous_resolution = matches[0]

        run_dir = self.run_dir(run_id)
        resolution_path = run_dir / "exception-resolutions.json"
        snapshot = self._snapshot_run_files(run_dir, COMPILER_MUTATION_FILES)
        current = read_json(resolution_path)
        for field in ("entries", "validEntries", "staleEntries", "supersededEntries"):
            current[field] = [
                entry
                for entry in current.get(field) or []
                if not (
                    entry.get("subjectId") == subject_id
                    and entry.get("subjectVersion") == subject_version
                )
            ]
        try:
            replace_json_artifact_set(run_dir, {"exception-resolutions.json": current})
            self.compile_canonical_rows(run_id)
            refreshed_detail = self.compiler_detail(run_id)
            event = build_audit_event(
                queue_subject_fingerprint=refreshed_detail["exceptionQueue"]["queueSubjectFingerprint"],
                subject_id_value=str(subject_id),
                subject_version_value=str(subject_version),
                event_type="resolution_reopened",
                prior_state="resolved",
                next_state="open",
                cause_fingerprint=semantic_hash(previous_resolution),
                resolution_entry_semantic_sha=semantic_hash(previous_resolution),
                reviewer=reviewer,
            )
            append_audit_event_once(run_dir / "exception-log.jsonl", event)
            refreshed = self.exception_queue_view(run_id, query=str(subject_id), limit=100)
            subject_view = next(
                (
                    item
                    for item in refreshed["items"]
                    if item["subject"].get("subjectId") == subject_id
                ),
                None,
            )
            if subject_view is None:
                raise WizardError("Reopened subject disappeared from the current queue.", status=409)
            result = {"summary": self.compiler_summary(run_id), "subject": subject_view}
        except Exception:
            self._restore_run_files(run_dir, snapshot)
            raise
        return result

    def _attach_reconciliation_decisions(
        self, reconciliation: dict[str, Any], decisions: dict[str, dict[str, Any]]
    ) -> dict[str, Any]:
        """Show, per model, the recorded variant_reconciliation decision (any
        lane — completeness accepts any group decision with that key)."""

        for model, entry in (reconciliation.get("models") or {}).items():
            record = next(
                (
                    r
                    for r in decisions.values()
                    if r["model"] == model and r.get("groupKey") == VARIANT_RECONCILIATION_KEY
                ),
                None,
            )
            entry["decision"] = (
                {
                    "decisionId": record["decisionId"],
                    "action": record["action"],
                    "resolution": record["resolution"],
                    "reviewerNote": record.get("reviewerNote", ""),
                }
                if record
                else None
            )
        return reconciliation

    def reconciliation(self, run_id: str) -> dict[str, Any]:
        _, candidates_file, candidates = self._parsed_candidates(run_id)
        selection = self._load_selection(run_id, candidates_file)
        state = self._decision_state(run_id, candidates, selection)
        reconciliation = read_json(self.run_dir(run_id) / "variant-reconciliation.json")
        return self._attach_reconciliation_decisions(reconciliation, state["decisions"])

    # ---------------------------------------------------- pass b: decisions
    def _decision_state(
        self,
        run_id: str,
        candidates: list[dict[str, Any]],
        selection: dict[str, Any],
        *,
        prune_to_targets: bool = False,
    ) -> dict[str, Any]:
        decisions_file = self.run_dir(run_id) / "decisions.json"
        snapshot = read_json(decisions_file) if decisions_file.is_file() else None
        candidates_by_id = {c["candidateId"]: c for c in candidates}
        state = load_decision_state(snapshot, candidates_by_id)
        if prune_to_targets:
            targets = set(selection["targets"])
            state["decisions"] = {
                key: record for key, record in state["decisions"].items() if record["model"] in targets
            }
        return state

    def _write_decisions(
        self, run_id: str, decisions: dict[str, dict[str, Any]], selection: dict[str, Any]
    ) -> None:
        write_json(
            self.run_dir(run_id) / "decisions.json",
            {
                "schemaVersion": SCHEMA_VERSION_B2,
                "candidatesFingerprint": selection["candidatesFingerprint"],
                "decisions": sorted(decisions.values(), key=lambda record: record["decisionId"]),
            },
        )

    def review_queue(
        self,
        run_id: str,
        model: str,
        lane: str,
        *,
        query: str = "",
        template: str = "",
        source_section: str = "",
        price_match: str = "",
        decision_state: str = "",
        price_presence: str = "",
        workbook_ref: str = "",
        section_state: str = "",
    ) -> dict[str, Any]:
        session, candidates_file, candidates = self._parsed_candidates(run_id)
        selection = self._load_selection(run_id, candidates_file)
        if model not in selection["targets"]:
            raise WizardError(f"Model {model} is not a selected target.")
        if lane not in {entry["lane"] for entry in LANES}:
            raise WizardError(f"Unknown lane: {lane}")
        state = self._decision_state(run_id, candidates, selection)
        skipped_section_ids = {
            record.get("candidateId")
            for record in state["decisions"].values()
            if record["model"] == model
            and record["lane"] == "section"
            and record["resolution"] == "not_needed"
            and record.get("candidateId")
        }
        scoped = scope_candidates(candidates, model)
        if lane == "standard_equipment":
            # Rows without an orderable RPO, plus rows the reviewer assigned to
            # a standard-behavior section — never plain selectable options.
            # Ref-only rows that are available-to-order on THIS model's own
            # variant columns (A statuses) are options, not standard equipment.
            # Priced rows (joined price rows or a list price) are options or
            # price problems, never standard equipment (spec B9).
            standard_sections = {
                s["sectionId"] for s in workbook_sections(self._require_workbook()) if s["standardBehavior"]
            }
            # Same approved-assignment predicate as the sectionState filter —
            # a held/skipped section decision is not an assignment yet.
            standard_assigned = {
                record.get("candidateId")
                for record in state["decisions"].values()
                if record["model"] == model
                and record["lane"] == "section"
                and record["resolution"] == "approved_for_plan"
                and record.get("action") == "assign_section"
                and (record.get("payload") or {}).get("sectionId") in standard_sections
            }
            scoped = [
                c
                for c in scoped
                if not candidate_has_price(c)
                and (
                    (c["rowKind"] == "ref_only" and not candidate_is_availability_row(c, model))
                    or c["candidateId"] in standard_assigned
                )
            ]
        elif lane == "section":
            scoped = [c for c in scoped if candidate_needs_section_decision(c)]
        elif lane == "exclusive_group":
            selectable_section_ids = {
                record.get("candidateId")
                for record in state["decisions"].values()
                if record["model"] == model
                and record["lane"] == "section"
                and record["resolution"] == "approved_for_plan"
                and record.get("action") == "assign_section"
                and (record.get("payload") or {}).get("sectionId")
                and (record.get("payload") or {}).get("selectable", True) is not False
                and (record.get("payload") or {}).get("active", True) is not False
            }
            scoped = [
                c
                for c in scoped
                if (c["rpo"] or c["refOnlyRpo"])
                and c["candidateId"] in selectable_section_ids
            ]
        else:
            scoped = [c for c in scoped if c["rowKind"] == "orderable"]
        if lane != "section" and skipped_section_ids:
            scoped = [c for c in scoped if c["candidateId"] not in skipped_section_ids]
        if lane == "status_nuance":
            # Only rows whose availability symbols actually need a human read.
            scoped = [c for c in scoped if candidate_needs_status_review(c)]

        # Source group = the export's own section label where one exists, the
        # sheet name otherwise (Interior/Exterior/Mechanical sheets carry no
        # section-label rows — field note 4).
        def source_group(candidate: dict[str, Any]) -> str:
            return candidate["sectionLabel"] or candidate["sheetName"]

        source_sections = sorted({source_group(c) for c in scoped})
        if source_section:
            scoped = [c for c in scoped if source_group(c) == source_section]
        if price_match:
            scoped = [c for c in scoped if (c.get("priceMatch") or "") == price_match]
        if price_presence:
            if price_presence not in {"priced", "unpriced"}:
                raise WizardError(f"Invalid price presence filter: {price_presence}")
            want_priced = price_presence == "priced"
            scoped = [c for c in scoped if candidate_has_price(c) == want_priced]
        if workbook_ref:
            if workbook_ref not in {"in_workbook", "new"}:
                raise WizardError(f"Invalid workbook reference filter: {workbook_ref}")
            reference_index = self._option_reference()
            want_known = workbook_ref == "in_workbook"
            scoped = [
                c
                for c in scoped
                if bool(reference_index.get(c["rpo"] or c["refOnlyRpo"])) == want_known
            ]
        if section_state:
            if section_state not in {"assigned", "unassigned"}:
                raise WizardError(f"Invalid section state filter: {section_state}")
            assigned_ids = {
                record.get("candidateId")
                for record in state["decisions"].values()
                if record["model"] == model
                and record["lane"] == "section"
                and record["resolution"] == "approved_for_plan"
                and record.get("action") == "assign_section"
                and record.get("candidateId")
            }
            if section_state == "assigned":
                scoped = [c for c in scoped if c["candidateId"] in assigned_ids]
            else:
                scoped = [c for c in scoped if c["candidateId"] not in assigned_ids]
        if decision_state:
            if decision_state not in {"undecided", "decided"}:
                raise WizardError(f"Invalid decision state filter: {decision_state}")
            lane_decided_ids = {
                record.get("candidateId")
                for record in state["decisions"].values()
                if record["model"] == model and record["lane"] == lane and record.get("candidateId")
            }
            if decision_state == "decided":
                scoped = [c for c in scoped if c["candidateId"] in lane_decided_ids]
            else:
                scoped = [
                    c
                    for c in scoped
                    if c["candidateId"] not in lane_decided_ids
                ]
        if query:
            needle = query.lower()
            scoped = [
                c
                for c in scoped
                if needle in c["rpo"].lower()
                or needle in c["refOnlyRpo"].lower()
                or needle in c["description"].lower()
            ]
        payload: dict[str, Any] = {
            "session": session,
            "model": model,
            "lane": lane,
            "sourceSections": source_sections,
            "candidates": scoped,
            "decisions": [
                record
                for record in state["decisions"].values()
                if record["model"] == model and record["lane"] == lane
            ],
            "invalidated": [
                record
                for record in state["invalidated"]
                if record.get("model") == model and record.get("lane") == lane
            ],
        }
        lane_config = next(entry for entry in LANES if entry["lane"] == lane)
        if lane_config["perCandidate"]:
            reference = self._option_reference()
            preferred = selection["comparators"].get(model, "")
            payload["workbookReference"] = {
                rpo: sorted(rows, key=lambda row: row["modelKey"] != preferred)
                for rpo in {c["rpo"] or c["refOnlyRpo"] for c in scoped if c["rpo"] or c["refOnlyRpo"]}
                if (rows := reference.get(rpo))
            }
        if lane in ("section", "exclusive_group"):
            payload["sections"] = workbook_sections(self._require_workbook())
        if lane == "exclusive_group":
            # Pool picker context: each option's decided section, so the UI can
            # show selection modes and group accurately.
            payload["sectionDecisions"] = {
                record["candidateId"]: (record.get("payload") or {}).get("sectionId", "")
                for record in state["decisions"].values()
                if record["model"] == model and record["lane"] == "section" and record.get("candidateId")
            }
        if lane == "relationship":
            payload["hints"] = scan_candidates(scoped)
        if lane == "copy_split":
            for candidate in scoped:
                candidate["proposedSplit"] = propose_copy_split(candidate)
            # Name collisions ("Seats" — one per seat option) need a human
            # rebuild. Count distinct RPOs per name over the FULL model lane
            # scope (not the filtered view, or filters would hide collisions);
            # GM lists the same RPO on two sheets (category sheet + Additional
            # Options) — same-RPO pairs are expected, not collisions.
            rpos_by_name: dict[str, set[str]] = {}
            for candidate in scope_candidates(candidates, model):
                if candidate["rowKind"] != "orderable":
                    continue
                if candidate["candidateId"] in skipped_section_ids:
                    continue
                key = propose_copy_split(candidate)["name"].strip().lower()
                if key:
                    rpos_by_name.setdefault(key, set()).add(candidate["rpo"] or candidate["refOnlyRpo"])
            for candidate in scoped:
                split = candidate["proposedSplit"]
                if len(rpos_by_name.get(split["name"].strip().lower(), set())) > 1:
                    split["flags"].append(FLAG_DUPLICATE_NAME)
        if lane == "duplicate":
            payload["duplicateGroups"] = duplicate_rpo_groups(scoped)
        if lane == "interior_media_deferral":
            payload["suggestedDeferrals"] = [
                {
                    "groupKey": f"{model}-interior-scope",
                    "kind": "interior",
                    "label": "Interior scope rows",
                    "why": "Color & Trim sheets aren't parsed; model_interior_scope rows must be authored in the apply pass.",
                },
                {
                    "groupKey": f"{model}-color-overrides",
                    "kind": "color",
                    "label": "Color overrides check",
                    "why": "Confirm shared color_overrides rows cover this model or record what's missing.",
                },
                {
                    "groupKey": f"{model}-asset-images",
                    "kind": "asset",
                    "label": "Asset images (asset_map)",
                    "why": "Not a go-live blocker (resolved decision 5) — record so it stays on the list after promotion.",
                },
            ]
        if lane == "presentation":
            template_model = template or selection["comparators"].get(model, "")
            if not template_model:
                raise WizardError("Presentation prefill needs a template model.")
            payload["prefill"] = presentation_prefill(self._require_workbook(), template_model, model)
        return payload

    def save_decisions(self, run_id: str, decisions: list[dict[str, Any]]) -> dict[str, Any]:
        session, candidates_file, candidates = self._parsed_candidates(run_id)
        self._refuse_if_applied(session)
        selection = self._load_selection(run_id, candidates_file)
        if not isinstance(decisions, list) or not decisions:
            raise WizardError("Request must carry a non-empty decisions list.")
        candidates_by_id = {c["candidateId"]: c for c in candidates}
        state = self._decision_state(run_id, candidates, selection)
        batch_id = uuid.uuid4().hex[:12]
        accepted: list[dict[str, Any]] = []
        for decision in decisions:
            try:
                record = validate_decision(decision, candidates_by_id)
            except ValueError as exc:
                raise WizardError(str(exc)) from exc
            if record["model"] not in selection["targets"]:
                raise WizardError(f"Decision targets non-selected model: {record['model']}")
            record["batchId"] = batch_id
            state["decisions"][record["decisionId"]] = record
            accepted.append(record)
        run_dir = self.run_dir(run_id)
        with (run_dir / "decisions-log.jsonl").open("a", encoding="utf-8") as log:
            for record in accepted:
                log.write(json.dumps(record, ensure_ascii=False) + "\n")
        self._write_decisions(run_id, state["decisions"], selection)
        if session["state"] in (
            STATE_MODELS_SELECTED,
            STATE_DECISIONS_COMPLETE,
            STATE_PLAN_BUILT,
            STATE_PLAN_APPROVED,
            STATE_DRY_RUN_APPROVED,
            STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED,
            STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE,
            STATE_WRITE_APPROVED,
        ):
            session["state"] = STATE_DECISIONS_IN_PROGRESS
            write_json(run_dir / "session.json", session)
        return {
            "session": session,
            "batchId": batch_id,
            "accepted": [record["decisionId"] for record in accepted],
        }

    def delete_decisions(
        self,
        run_id: str,
        *,
        decision_ids: list[str] | None = None,
        batch_id: str = "",
    ) -> dict[str, Any]:
        session, candidates_file, candidates = self._parsed_candidates(run_id)
        self._refuse_if_applied(session)
        selection = self._load_selection(run_id, candidates_file)
        if not decision_ids and not batch_id:
            raise WizardError("Delete needs decisionIds or a batchId.")
        state = self._decision_state(run_id, candidates, selection)
        targets = set(decision_ids or [])
        deleted: list[str] = []
        for key, record in list(state["decisions"].items()):
            if key in targets or (batch_id and record.get("batchId") == batch_id):
                del state["decisions"][key]
                deleted.append(key)
        if deleted:
            run_dir = self.run_dir(run_id)
            with (run_dir / "decisions-log.jsonl").open("a", encoding="utf-8") as log:
                log.write(
                    json.dumps(
                        {
                            "deleted": deleted,
                            "batchId": batch_id or None,
                            "at": time.strftime("%Y-%m-%dT%H:%M:%S"),
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )
            self._write_decisions(run_id, state["decisions"], selection)
            if session["state"] in (
                STATE_DECISIONS_COMPLETE,
                STATE_PLAN_BUILT,
                STATE_PLAN_APPROVED,
                STATE_DRY_RUN_APPROVED,
                STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED,
                STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE,
                STATE_WRITE_APPROVED,
            ):
                session["state"] = STATE_DECISIONS_IN_PROGRESS
                write_json(run_dir / "session.json", session)
        return {"session": session, "deleted": deleted}

    def copy_model_decisions(
        self, run_id: str, from_model: str, to_model: str, *, overwrite: bool = False
    ) -> dict[str, Any]:
        session, candidates_file, candidates = self._parsed_candidates(run_id)
        self._refuse_if_applied(session)
        selection = self._load_selection(run_id, candidates_file)
        if from_model not in selection["targets"] or to_model not in selection["targets"]:
            raise WizardError("Copy source and target must both be selected target models.")
        state = self._decision_state(run_id, candidates, selection)
        try:
            report = copy_decisions(
                state["decisions"], candidates, from_model, to_model, overwrite=overwrite
            )
        except ValueError as exc:
            raise WizardError(str(exc)) from exc
        run_dir = self.run_dir(run_id)
        batch_id = uuid.uuid4().hex[:12]
        if report["copied"]:
            for record in report["copied"]:
                record["batchId"] = batch_id
                state["decisions"][record["decisionId"]] = record
            with (run_dir / "decisions-log.jsonl").open("a", encoding="utf-8") as log:
                for record in report["copied"]:
                    log.write(json.dumps(record, ensure_ascii=False) + "\n")
            self._write_decisions(run_id, state["decisions"], selection)
            if session["state"] in (STATE_MODELS_SELECTED, STATE_DECISIONS_COMPLETE):
                session["state"] = STATE_DECISIONS_IN_PROGRESS
                write_json(run_dir / "session.json", session)
        return {
            "session": session,
            "batchId": batch_id if report["copied"] else None,
            "copied": len(report["copied"]),
            "copiedByLane": report["copiedByLane"],
            "skipped": report["skipped"],
        }

    def progress(self, run_id: str) -> dict[str, Any]:
        session, candidates_file, candidates = self._parsed_candidates(run_id)
        selection = self._load_selection(run_id, candidates_file)
        state = self._decision_state(run_id, candidates, selection)
        reconciliation_file = self.run_dir(run_id) / "variant-reconciliation.json"
        reconciliation = read_json(reconciliation_file) if reconciliation_file.is_file() else None
        report = completeness(candidates, selection["targets"], state["decisions"], reconciliation)
        report["session"] = session
        report["invalidatedDecisions"] = len(state["invalidated"])
        return report

    # ------------------------------------------------------ pass c: plan
    def build_apply_plan(self, run_id: str, *, schema_validation: bool = True) -> dict[str, Any]:
        """Build the two-stage op plan and dry-run it. Stage 1 validates
        against the live extract; stage 2 against a scratch copy with stage 1
        applied. The live workbook is never written here.

        schema_validation=False exists for fixture-scale tests whose compact
        workbooks are not schema-complete; real runs keep it on."""

        import shutil
        import tempfile

        from corvette_form_generator.editor_ops import apply_batch
        from corvette_form_generator.ingest.wizard.plan_builder import build_plan

        session = self.load_session(run_id, verify_source=False)
        self._refuse_if_applied(session)
        if session["state"] not in (
            STATE_DECISIONS_COMPLETE,
            STATE_PLAN_BUILT,
            STATE_PLAN_APPROVED,
            STATE_DRY_RUN_APPROVED,
            STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED,
            STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE,
        ):
            raise WizardError("Mark decisions complete before building the apply plan.", status=409)
        run_dir = self.run_dir(run_id)
        _, candidates_file, candidates = self._parsed_candidates(run_id)
        selection = self._load_selection(run_id, candidates_file)
        state = self._decision_state(run_id, candidates, selection)
        workbook = self._require_workbook()
        plan = build_plan(
            workbook_path=workbook,
            selection=selection,
            candidates=candidates,
            decisions=state["decisions"],
            candidates_fingerprint=selection["candidatesFingerprint"],
        )

        def summarize(result: dict[str, Any]) -> dict[str, Any]:
            return {
                "ok": result.get("ok", False),
                "status": result.get("status", ""),
                "errors": result.get("errors", []),
                "warnings": result.get("warnings", []),
                "opCount": result.get("opCount"),
                "schemaErrors": (result.get("schemaResult") or {}).get("error_count"),
            }

        dry_run: dict[str, Any] = {}
        stage1_batch = {
            "workbookMtimeNs": str(workbook.stat().st_mtime_ns),
            "items": plan["stage1"]["items"],
        }
        dry_run["stage1"] = summarize(
            apply_batch(workbook, stage1_batch, write=False, run_schema_validation=False)
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            scratch = Path(tmp_dir) / workbook.name
            shutil.copy2(workbook, scratch)
            stage1_scratch = dict(stage1_batch, workbookMtimeNs=str(scratch.stat().st_mtime_ns))
            applied = apply_batch(
                # A scratch save still goes through the service-level write
                # boundary. Compact tests patch the schema scan to an empty
                # issue list; production never disables it for a write call.
                scratch, stage1_scratch, write=True, run_schema_validation=True,
                confirmed_warnings=[w["id"] for w in dry_run["stage1"]["warnings"]],
                log_path=run_dir / "scratch-apply-log.jsonl",
            )
            dry_run["stage1Scratch"] = summarize(applied)
            if applied.get("ok"):
                stage2_batch = {
                    "workbookMtimeNs": str(scratch.stat().st_mtime_ns),
                    "items": plan["stage2"]["items"],
                }
                dry_run["stage2"] = summarize(
                    apply_batch(scratch, stage2_batch, write=False, run_schema_validation=schema_validation)
                )
            else:
                dry_run["stage2"] = {"ok": False, "status": "skipped", "errors": ["stage 1 scratch apply failed"], "warnings": []}
        dry_run["ok"] = bool(
            plan["valid"] and dry_run["stage1"]["ok"] and dry_run["stage1Scratch"]["ok"] and dry_run["stage2"]["ok"]
        )

        write_json(run_dir / "apply-plan.json", plan)
        write_json(run_dir / "apply-plan-dryrun.json", dry_run)
        (run_dir / "apply-plan.md").write_text(plan_markdown(plan, dry_run), encoding="utf-8")
        session["state"] = STATE_PLAN_BUILT if dry_run["ok"] else STATE_DECISIONS_COMPLETE
        write_json(run_dir / "session.json", session)
        return {"session": session, "plan": plan_summary(plan), "dryRun": dry_run}

    def plan_detail(self, run_id: str) -> dict[str, Any]:
        run_dir = self.run_dir(run_id)
        plan_file = run_dir / "apply-plan.json"
        if not plan_file.is_file():
            raise WizardError("Build the apply plan first.", status=404)
        plan = read_json(plan_file)
        dry_file = run_dir / "apply-plan-dryrun.json"
        approval_file = run_dir / "plan-approval.json"
        return {
            "session": self.load_session(run_id, verify_source=False),
            "plan": plan_summary(plan),
            "dryRun": read_json(dry_file) if dry_file.is_file() else None,
            "approval": read_json(approval_file) if approval_file.is_file() else None,
        }

    def approve_plan(self, run_id: str, approver: str) -> dict[str, Any]:
        session = self.load_session(run_id, verify_source=False)
        if session["state"] not in (
            STATE_PLAN_BUILT,
            STATE_PLAN_APPROVED,
            STATE_DRY_RUN_APPROVED,
            STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED,
            STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE,
        ):
            raise WizardError("Plan must be built (and dry-run clean) before approval.", status=409)
        if not str(approver or "").strip():
            raise WizardError("Approval needs a reviewer name.")
        run_dir = self.run_dir(run_id)
        plan = read_json(run_dir / "apply-plan.json")
        # Fail closed if anything shifted since the plan was built.
        _, candidates_file, candidates = self._parsed_candidates(run_id)
        selection = self._load_selection(run_id, candidates_file)
        state = self._decision_state(run_id, candidates, selection)
        if plan["decisionsFingerprint"] != artifact_sha(state["decisions"]):
            raise WizardError("Decisions changed after the plan was built; rebuild the plan.", status=409)
        workbook = self._require_workbook()
        if plan["workbookFingerprint"]["mtimeNs"] != str(workbook.stat().st_mtime_ns) or plan[
            "workbookFingerprint"
        ]["sha256"] != hashlib.sha256(workbook.read_bytes()).hexdigest():
            raise WizardError("Workbook changed after the plan was built; rebuild the plan.", status=409)
        if plan.get("schemaVersion") == WRITABLE_PLAN_SCHEMA:
            self._require_compiler_bindings(run_dir, plan)
            semantic_blockers = self._option_semantic_blockers(plan)
            if semantic_blockers:
                raise WizardError(
                    f"{WRITABLE_PLAN_SCHEMA} plan approval requires explicit typed "
                    "selectable and active values on every emitted option row.",
                    status=409,
                )
        approval = {
            "schemaVersion": PLAN_APPROVAL_SCHEMA,
            "scope": PLAN_APPROVAL_SCOPE,
            "approvedBy": str(approver).strip(),
            "approvedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
            **self._approval_bindings(run_id, session, plan, run_dir),
        }
        write_json(run_dir / "plan-approval.json", approval)
        self._append_approval_audit(run_dir, approval)
        (run_dir / "write-approval.json").unlink(missing_ok=True)
        session["state"] = STATE_DRY_RUN_APPROVED
        write_json(run_dir / "session.json", session)
        return {"session": session, "approval": approval}

    # ------------------------------------------------------ pass d: apply
    def _combined_plan_batch(self, plan: dict[str, Any], workbook: Path) -> dict[str, Any]:
        return {
            "workbookMtimeNs": str(workbook.stat().st_mtime_ns),
            "items": [*plan["stage1"]["items"], *plan["stage2"]["items"]],
        }

    def _per_sheet_action_counts(self, items: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
        counts: dict[str, dict[str, int]] = {}
        for item in items:
            sheet = item.get("sheet")
            action = item.get("action")
            if not sheet or not action:
                continue
            sheet_counts = counts.setdefault(str(sheet), {})
            sheet_counts[str(action)] = sheet_counts.get(str(action), 0) + 1
        return {sheet: dict(sorted(actions.items())) for sheet, actions in sorted(counts.items())}

    def _deployment_continuity_from_plan(self, plan: dict[str, Any]) -> dict[str, Any]:
        """Action-aware source-coverage diagnostic for the dry-run report.

        The temp generation probe added by Pass D.1 can enrich these entries;
        this source-op layer is intentionally conservative and never treats
        create/delete-only sheet activity as runtime coverage.
        """

        report = plan.get("report") or {}
        plan_continuity = report.get("runtimeContinuity") or {}
        output: dict[str, Any] = {}
        for model in plan.get("targets") or []:
            source_ops = (plan_continuity.get(model) or {}).get("sourceOps") or {}
            price_adds = sum((source_ops.get("priceRules") or {}).get(action, 0) for action in ("add", "update"))
            rule_group_adds = sum((source_ops.get("ruleGroups") or {}).get(action, 0) for action in ("add", "update"))
            color_adds = sum((source_ops.get("colorOverrides") or {}).get(action, 0) for action in ("add", "update"))
            component_adds = sum((source_ops.get("interiorComponents") or {}).get(action, 0) for action in ("add", "update"))
            asset_adds = sum((source_ops.get("assetMap") or {}).get(action, 0) for action in ("add", "update"))
            blockers: list[dict[str, str]] = []
            deferrals: list[dict[str, str]] = []
            if model in {"zr1", "zr1x"} and price_adds == 0:
                blockers.append({"kind": "price_rules_required_for_runtime", "detail": "no price-rule add/update ops"})
            if model in {"zr1", "zr1x"} and rule_group_adds == 0:
                blockers.append({"kind": "rule_groups_required_for_runtime", "detail": "no rule-group add/update ops"})
            output[model] = {
                "status": "not_deployment_ready" if blockers else "source_ops_diagnostic",
                "registryLoadable": None,
                "registryLoadableNote": "temp generation probe not run in source-op diagnostic layer",
                "sourceOps": source_ops,
                "sourceCoverage": {
                    "priceRuleAddOrUpdateCount": price_adds,
                    "ruleGroupAddOrUpdateCount": rule_group_adds,
                    "colorOverrideAddOrUpdateCount": color_adds,
                    "interiorComponentAddOrUpdateCount": component_adds,
                    "assetMapAddOrUpdateCount": asset_adds,
                },
                "deploymentBlockers": blockers,
                "deploymentDeferrals": deferrals,
            }
        return output

    def _activate_probe_models(self, workbook: Path, models: list[str]) -> None:
        from openpyxl import load_workbook

        wb = load_workbook(workbook)
        try:
            targets = set(models)
            variant_ids: set[str] = set()
            if "model_variants" in wb.sheetnames:
                ws = wb["model_variants"]
                headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value is not None}
                model_col = headers.get("model_key")
                variant_col = headers.get("variant_id")
                if model_col and variant_col:
                    for row in range(2, ws.max_row + 1):
                        model_key = str(ws.cell(row=row, column=model_col).value or "").strip().lower()
                        if model_key in targets:
                            variant_id = str(ws.cell(row=row, column=variant_col).value or "").strip()
                            if variant_id:
                                variant_ids.add(variant_id)
            for sheet_name, fields in {
                "model_master": {"active": True},
                "model_variants": {"active": True},
                "model_workbook_sources": {"active": True},
                "model_registry_promotion": {"active": True, "promoted_to_runtime": True},
            }.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value is not None}
                model_col = headers.get("model_key")
                if not model_col:
                    continue
                for row in range(2, ws.max_row + 1):
                    model_key = str(ws.cell(row=row, column=model_col).value or "").strip().lower()
                    if model_key not in targets:
                        if sheet_name == "model_master" and "active" in headers:
                            ws.cell(row=row, column=headers["active"], value=False)
                        continue
                    for column, value in fields.items():
                        if column in headers:
                            ws.cell(row=row, column=headers[column], value=value)
                    if sheet_name == "model_registry_promotion":
                        if "artifact_type" in headers:
                            ws.cell(row=row, column=headers["artifact_type"], value="runtime_contract")
                        if "artifact_path" in headers:
                            ws.cell(
                                row=row,
                                column=headers["artifact_path"],
                                value=f"form-output/runtime/{model_key.replace('_', '-')}-runtime-contract.json",
                            )
            if variant_ids and "variant_master" in wb.sheetnames:
                ws = wb["variant_master"]
                headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value is not None}
                variant_col = headers.get("variant_id")
                active_col = headers.get("active")
                if variant_col and active_col:
                    for row in range(2, ws.max_row + 1):
                        variant_id = str(ws.cell(row=row, column=variant_col).value or "").strip()
                        if variant_id in variant_ids:
                            ws.cell(row=row, column=active_col, value=True)
            wb.save(workbook)
        finally:
            wb.close()

    def _source_count(self, payload: dict[str, Any], key: str) -> int:
        value = payload.get(key)
        return len(value) if isinstance(value, list) else 0

    def _deployment_continuity_probe(
        self,
        workbook: Path,
        batch: dict[str, Any],
        plan: dict[str, Any],
        *,
        schema_validation: bool,
    ) -> dict[str, Any]:
        """Apply the plan to a temp workbook and run temp-only generation probes."""

        import shutil
        import tempfile

        from corvette_form_generator.contract import ASSET_IMAGE_FIELDS
        from corvette_form_generator.editor_ops import apply_batch
        from corvette_form_generator.model_configs import discover_generation_model_configs
        from corvette_form_generator.source_assembly import assemble_model_source

        targets = [str(model) for model in plan.get("targets") or []]
        continuity = self._deployment_continuity_from_plan(plan)
        if not targets:
            return continuity
        with tempfile.TemporaryDirectory(prefix="ingest-d1-deployment-") as tmp_dir:
            tmp_root = Path(tmp_dir)
            tmp_workbook = tmp_root / workbook.name
            shutil.copy2(workbook, tmp_workbook)
            temp_batch = dict(batch, workbookMtimeNs=str(tmp_workbook.stat().st_mtime_ns))
            preview = apply_batch(tmp_workbook, temp_batch, write=False, run_schema_validation=schema_validation)
            confirmed = [warning["id"] for warning in preview.get("warnings", [])]
            applied = apply_batch(
                tmp_workbook,
                temp_batch,
                write=True,
                confirmed_warnings=confirmed,
                source="ingest_wizard_deployment_probe",
                log_path=tmp_root / "probe-edit-log.jsonl",
                # This is a saved scratch workbook, so the writer boundary
                # always keeps schema validation enabled. Compact tests patch
                # the issue scan rather than disabling the write gate.
                run_schema_validation=True,
            )
            if not applied.get("ok"):
                error = "; ".join(applied.get("errors", [])) or str(applied.get("status"))
                for model in targets:
                    entry = continuity.setdefault(model, {})
                    blockers = list(entry.get("deploymentBlockers") or [])
                    blockers.append({"kind": "deployment_probe_apply_failed", "detail": error})
                    entry.update(
                        {
                            "status": "probe_apply_failed",
                            "registryLoadable": False,
                            "registryError": error,
                            "deploymentBlockers": blockers,
                        }
                    )
                return continuity

            self._activate_probe_models(tmp_workbook, targets)
            try:
                configs = discover_generation_model_configs(tmp_workbook)
            except Exception as exc:
                configs = {}
                discovery_error = str(exc)
            else:
                discovery_error = ""

            for model in targets:
                entry = continuity.setdefault(model, {})
                blockers = list(entry.get("deploymentBlockers") or [])
                deferrals = list(entry.get("deploymentDeferrals") or [])
                config = configs.get(model)
                if config is None:
                    error = discovery_error or f"model {model!r} was not discoverable after temp activation"
                    blockers.append({"kind": "registry_load_failed", "detail": error})
                    entry.update(
                        {
                            "status": "not_deployment_ready",
                            "registryLoadable": False,
                            "registryError": error,
                            "deploymentBlockers": blockers,
                            "deploymentDeferrals": deferrals,
                        }
                    )
                    continue

                config = config.with_overrides(
                    workbook_path=tmp_workbook,
                    output_dir=tmp_root / "form-output",
                    app_dir=tmp_root / "form-app",
                )
                try:
                    assembly = assemble_model_source(config)
                except Exception as exc:
                    blockers.append({"kind": "registry_load_failed", "detail": str(exc)})
                    entry.update(
                        {
                            "status": "not_deployment_ready",
                            "registryLoadable": False,
                            "registryError": str(exc),
                            "deploymentBlockers": blockers,
                            "deploymentDeferrals": deferrals,
                        }
                    )
                    continue

                source = assembly.source_data
                raw_validation = source.get("validation")
                validation: list[dict[str, Any]] = raw_validation if isinstance(raw_validation, list) else []
                pricing_deferred = any(row.get("check_id") == "pricing_deferred" for row in validation if isinstance(row, dict))
                raw_choices = source.get("choices")
                choices: list[dict[str, Any]] = raw_choices if isinstance(raw_choices, list) else []
                raw_interiors = source.get("interiors")
                interiors: list[dict[str, Any]] = raw_interiors if isinstance(raw_interiors, list) else []
                media_count = sum(
                    1
                    for choice in choices
                    if isinstance(choice, dict) and any(choice.get(field) for field in ASSET_IMAGE_FIELDS)
                )
                component_count = sum(
                    len(interior.get("interior_components") or [])
                    for interior in interiors
                    if isinstance(interior, dict) and isinstance(interior.get("interior_components"), list)
                )
                counts = {
                    "choices": len(choices),
                    "directRules": self._source_count(source, "rules"),
                    "ruleGroups": self._source_count(source, "ruleGroups"),
                    "exclusiveGroups": self._source_count(source, "exclusiveGroups"),
                    "priceRules": self._source_count(source, "priceRules"),
                    "pricingDeferred": pricing_deferred,
                    "colorOverrides": self._source_count(source, "colorOverrides"),
                    "interiors": len(interiors),
                    "interiorComponentLineItems": component_count,
                    "optionMediaCoveredChoices": media_count,
                    "optionMediaTotalChoices": len(choices),
                    "validationWarnings": sum(1 for row in validation if isinstance(row, dict) and row.get("severity") == "warning"),
                    "validationErrors": sum(1 for row in validation if isinstance(row, dict) and row.get("severity") == "error"),
                }
                if counts["validationErrors"]:
                    blockers.append(
                        {
                            "kind": "generated_validation_errors",
                            "detail": (
                                f"generated contract has {counts['validationErrors']} validation error(s)"
                            ),
                        }
                    )
                if model in {"zr1", "zr1x"} and counts["priceRules"] == 0 and pricing_deferred:
                    blockers.append({"kind": "price_rules_required_for_runtime", "detail": "generated contract has pricing_deferred and zero priceRules"})
                if model in {"zr1", "zr1x"} and counts["ruleGroups"] == 0:
                    blockers.append({"kind": "rule_groups_required_for_runtime", "detail": "generated contract has zero ruleGroups"})
                if counts["colorOverrides"] == 0:
                    blockers.append(
                        {
                            "kind": "color_overrides_missing_or_unproven",
                            "detail": "generated contract has zero colorOverrides without not-applicable proof",
                        }
                    )
                if counts["interiors"] == 0:
                    blockers.append(
                        {
                            "kind": "interiors_missing_or_unproven",
                            "detail": "generated contract has zero interiors without not-applicable proof",
                        }
                    )
                if counts["interiorComponentLineItems"] == 0:
                    blockers.append(
                        {
                            "kind": "interior_components_missing_or_unproven",
                            "detail": (
                                "generated contract has zero interior component line items "
                                "without not-applicable proof"
                            ),
                        }
                    )
                if counts["optionMediaCoveredChoices"] == 0:
                    blockers.append(
                        {
                            "kind": "asset_map_media_missing",
                            "detail": "generated choices have zero option/card asset fields",
                        }
                    )
                entry.update(
                    {
                        "status": "not_deployment_ready" if blockers else "deployment_probe_passed",
                        "registryLoadable": True,
                        "registryError": "",
                        "counts": counts,
                        "deploymentBlockers": blockers,
                        "deploymentDeferrals": deferrals,
                    }
                )
            return continuity

    def _validate_plan_approval(
        self,
        run_id: str,
        session: dict[str, Any],
        plan: dict[str, Any],
        approval: dict[str, Any],
        run_dir: Path,
        *,
        allow_legacy: bool,
    ) -> bool:
        """Validate current scoped authority; return True for a legacy record."""

        plan_sha = hashlib.sha256((run_dir / "apply-plan.json").read_bytes()).hexdigest()
        if approval.get("schemaVersion") != PLAN_APPROVAL_SCHEMA:
            if (
                allow_legacy
                and approval.get("schemaVersion") in {"pass-c-1", "pass-c-2"}
                and not approval.get("scope")
                and approval.get("planSha") == plan_sha
            ):
                return True
            raise WizardError(
                f"Apply requires {PLAN_APPROVAL_SCHEMA} {PLAN_APPROVAL_SCOPE} approval.",
                status=409,
            )
        if approval.get("scope") != PLAN_APPROVAL_SCOPE:
            raise WizardError("Plan approval scope must be dry_run_evidence.", status=409)
        expected = self._approval_bindings(run_id, session, plan, run_dir)
        for field, value in expected.items():
            if approval.get(field) != value:
                raise WizardError(
                    f"Plan approval binding {field} changed; rebuild and re-approve.",
                    status=409,
                )
        return False

    def _validate_current_plan_inputs(
        self,
        run_id: str,
        session: dict[str, Any],
        plan: dict[str, Any],
    ) -> tuple[Path, dict[str, Any]]:
        _, candidates_file, candidates = self._parsed_candidates(run_id)
        selection = self._load_selection(run_id, candidates_file)
        state = self._decision_state(run_id, candidates, selection)
        if plan.get("decisionsFingerprint") != artifact_sha(state["decisions"]):
            raise WizardError("Decisions changed after the plan was approved; rebuild the plan.", status=409)
        source_fingerprint = plan.get("sourceFingerprint") or selection.get("sourceFingerprint") or {}
        if source_fingerprint.get("sha256") != session.get("fingerprint", {}).get("sha256"):
            raise WizardError("Source fingerprint changed after the plan was approved; start a new run.", status=409)
        workbook = self._require_workbook()
        before = file_fingerprint(workbook)
        expected = plan.get("workbookFingerprint") or {}
        if expected.get("mtimeNs") != str(before["mtimeNs"]) or expected.get("sha256") != before["sha256"]:
            raise WizardError("Workbook changed after the plan was approved; rebuild the plan.", status=409)
        return workbook, before

    def _option_semantic_blockers(self, plan: dict[str, Any]) -> list[dict[str, Any]]:
        from corvette_form_generator.editor_ops import flatten_items
        from corvette_form_generator.ingest.wizard.plan_builder import MODEL_PLAN_CONFIG

        blockers: list[dict[str, Any]] = []
        items = flatten_items(
            [
                *(plan.get("stage1", {}).get("items") or []),
                *(plan.get("stage2", {}).get("items") or []),
            ]
        )
        for model in plan.get("targets") or []:
            config = MODEL_PLAN_CONFIG.get(model) or {}
            options_sheet = f"{config.get('sheetPrefix', '')}options"
            missing: list[dict[str, Any]] = []
            for item in items:
                if item.get("sheet") != options_sheet or item.get("action") not in {"add", "update"}:
                    continue
                values = item.get("row") or {}
                absent = [field for field in ("selectable", "active") if not isinstance(values.get(field), bool)]
                if absent:
                    missing.append(
                        {
                            "optionId": str((item.get("key") or {}).get("option_id") or values.get("option_id") or ""),
                            "fields": absent,
                        }
                    )
            if missing:
                blockers.append(
                    {
                        "kind": "blank_option_semantics",
                        "model": model,
                        "detail": f"{len(missing)} emitted option row(s) lack typed selectable/active values",
                        "examples": missing[:10],
                    }
                )
        return blockers

    def _identity_churn_blockers(
        self,
        workbook: Path,
        plan: dict[str, Any],
    ) -> list[dict[str, Any]]:
        from corvette_form_generator.editor_ops import extract_workbook, flatten_items
        from corvette_form_generator.ingest.wizard.plan_builder import MODEL_PLAN_CONFIG

        extract = extract_workbook(workbook)
        items = flatten_items(
            [
                *(plan.get("stage1", {}).get("items") or []),
                *(plan.get("stage2", {}).get("items") or []),
            ]
        )
        blockers: list[dict[str, Any]] = []
        for model in plan.get("targets") or []:
            config = MODEL_PLAN_CONFIG.get(model) or {}
            sheet = f"{config.get('sheetPrefix', '')}options"
            existing = {
                str(row.get("option_id") or "").strip(): str(row.get("rpo") or "").strip().upper()
                for row in extract.get("sheets", {}).get(sheet, {}).get("rows", [])
            }
            deleted = {
                option_id: existing.get(option_id, "")
                for item in items
                if item.get("sheet") == sheet and item.get("action") == "delete"
                if (option_id := str((item.get("key") or {}).get("option_id") or "").strip())
            }
            additions = [
                item.get("row") or {}
                for item in items
                if item.get("sheet") == sheet and item.get("action") == "add"
            ]
            churn: list[dict[str, str]] = []
            for old_id, rpo in deleted.items():
                if not rpo:
                    continue
                for row in additions:
                    new_id = str(row.get("option_id") or "").strip()
                    if str(row.get("rpo") or "").strip().upper() == rpo and new_id and new_id != old_id:
                        churn.append({"rpo": rpo, "oldOptionId": old_id, "newOptionId": new_id})
            if churn:
                blockers.append(
                    {
                        "kind": "identity_churn",
                        "model": model,
                        "detail": f"{len(churn)} matched option identity replacement(s) require reconciliation",
                        "examples": churn[:10],
                    }
                )
        return blockers

    def _write_eligibility(
        self,
        *,
        workbook: Path,
        plan: dict[str, Any],
        approval: dict[str, Any],
        legacy_approval: bool,
        schema_validation: bool,
        result: dict[str, Any],
        deployment: dict[str, Any],
    ) -> dict[str, Any]:
        from corvette_form_generator.editor_ops import classify_warnings

        global_blockers: list[dict[str, Any]] = []
        target_blockers: dict[str, list[dict[str, Any]]] = {
            str(model): [] for model in plan.get("targets") or []
        }
        target_deferrals: dict[str, list[dict[str, Any]]] = {
            str(model): [] for model in plan.get("targets") or []
        }
        compile_deferrals: list[dict[str, Any]] = []
        not_applicable_families: set[tuple[str, str]] = set()
        if not schema_validation or result.get("schemaResult") is None:
            global_blockers.append(
                {
                    "kind": "schema_validation_not_run",
                    "detail": "schema-disabled diagnostics cannot establish live-write eligibility",
                }
            )
        if plan.get("schemaVersion") != WRITABLE_PLAN_SCHEMA:
            global_blockers.append(
                {
                    "kind": "plan_schema_not_writable",
                    "detail": (
                        f"plan schema {plan.get('schemaVersion')!r} is permanently diagnostic; "
                        f"{WRITABLE_PLAN_SCHEMA} is required"
                    ),
                }
            )
        if legacy_approval:
            global_blockers.append(
                {
                    "kind": "approval_schema_not_writable",
                    "detail": f"historical plan approval is diagnostic only; {PLAN_APPROVAL_SCHEMA} is required",
                }
            )
        compiler_bindings = (
            self._compiler_artifact_bindings(self.run_dir(str(approval["runId"])))
            if approval.get("runId")
            else {}
        )
        if plan.get("schemaVersion") == WRITABLE_PLAN_SCHEMA:
            for field in ("canonicalManifestSha", "compileReportSha", "exceptionResolutionsSha"):
                if not compiler_bindings.get(field) or plan.get(field) != compiler_bindings.get(field):
                    global_blockers.append(
                        {"kind": "compiler_binding_missing", "detail": f"{field} is absent or stale"}
                    )
            try:
                targets = [str(model) for model in plan.get("targets") or []]
                self._require_compile_report_readiness(
                    self.run_dir(str(approval["runId"])),
                    targets,
                )
                compile_deferrals = self._compile_report_allowed_deferrals(
                    self.run_dir(str(approval["runId"])),
                    targets,
                )
                not_applicable_families = self._compile_report_not_applicable_families(
                    self.run_dir(str(approval["runId"])),
                    targets,
                )
            except WizardError as exc:
                global_blockers.append(
                    {"kind": "compile_deferral_policy_invalid", "detail": str(exc)}
                )

        def media_deferral_for(model: str) -> dict[str, Any] | None:
            matches = [
                item
                for item in compile_deferrals
                if item.get("kind") == "asset_map_media_missing"
                and str(item.get("model") or "") == model
            ]
            return dict(matches[0]) if len(matches) == 1 else None

        semantic_blockers = self._option_semantic_blockers(plan) + self._identity_churn_blockers(workbook, plan)
        for blocker in semantic_blockers:
            model = str(blocker.get("model") or "")
            if model in target_blockers:
                target_blockers[model].append(blocker)
            else:
                global_blockers.append(blocker)

        coverage = result.get("operationCoverage") or {}
        if coverage.get("rawCovered") != coverage.get("rawCount"):
            global_blockers.append(
                {"kind": "operation_coverage_failed", "detail": "raw operation coverage is incomplete"}
            )
        verification = result.get("verification") or {}
        if not verification.get("ok") or verification.get("preparedChecked") != coverage.get("preparedCount"):
            global_blockers.append(
                {"kind": "readback_failed", "detail": "prepared operation readback is incomplete"}
            )

        warning_policy = result.get("warningPolicy") or classify_warnings(result.get("warnings") or [])
        for warning_id in warning_policy.get("blockingIds") or []:
            warning_kind = str(warning_id).partition(":")[0]
            kind = {
                "refdel": "referenced_delete",
                "dorder": "display_order_collision",
            }.get(warning_kind, "unknown_warning")
            global_blockers.append({"kind": kind, "detail": str(warning_id), "warningId": str(warning_id)})

        effective_deployment_blockers: dict[str, list[dict[str, Any]]] = {
            model: [] for model in target_blockers
        }
        for model in target_blockers:
            entry = deployment.get(model) or {}
            for blocker in entry.get("deploymentBlockers") or []:
                not_applicable_family = {
                    "color_overrides_missing_or_unproven": "color_overrides",
                    "interiors_missing_or_unproven": "interiors",
                    "interior_components_missing_or_unproven": "interior_components",
                    "asset_map_media_missing": "asset_map",
                }.get(str(blocker.get("kind") or ""))
                if not_applicable_family and (model, not_applicable_family) in not_applicable_families:
                    continue
                if blocker.get("kind") == "asset_map_media_missing":
                    media_deferral = media_deferral_for(model)
                    if media_deferral is not None:
                        if media_deferral not in target_deferrals[model]:
                            target_deferrals[model].append(media_deferral)
                        continue
                effective_deployment_blockers[model].append(
                    {
                        "kind": str(blocker.get("kind") or "deployment_blocked"),
                        "model": model,
                        "detail": str(blocker.get("detail") or "deployment continuity failed"),
                    }
                )
            for deferral in entry.get("deploymentDeferrals") or []:
                effective_deployment_blockers[model].append(
                    {
                        "kind": "unknown_deferral",
                        "model": model,
                        "detail": str(deferral.get("kind") or ""),
                    }
                )
            target_blockers[model].extend(effective_deployment_blockers[model])

        accepted_warning_ids: list[str] = []
        for warning_id in warning_policy.get("confirmableIds") or []:
            matched_model = next(
                (
                    model
                    for model in target_blockers
                    if str(warning_id).partition(":")[2].startswith(
                        {"grand_sport_x": "grandSportX_"}.get(model, f"{model}_")
                    )
                ),
                "",
            )
            entry = deployment.get(matched_model) or {}
            if (
                matched_model
                and not effective_deployment_blockers.get(matched_model)
                and entry.get("registryLoadable") is True
            ):
                accepted_warning_ids.append(str(warning_id))
            else:
                global_blockers.append(
                    {
                        "kind": "scaffold_warning_not_confirmable",
                        "detail": f"{warning_id} lacks a clean target scratch-generation probe",
                        "warningId": str(warning_id),
                    }
                )

        targets: dict[str, Any] = {}
        all_blockers = list(global_blockers)
        all_deferrals: list[dict[str, Any]] = []
        for model in target_blockers:
            blockers = target_blockers[model]
            deferrals = target_deferrals[model]
            all_blockers.extend(blockers)
            all_deferrals.extend(deferrals)
            targets[model] = {
                "eligible": not global_blockers and not blockers,
                "blockers": blockers,
                "deferrals": deferrals,
            }
        return {
            "eligible": bool(targets) and not all_blockers and all(
                entry["eligible"] for entry in targets.values()
            ),
            "blockers": all_blockers,
            "deferrals": all_deferrals,
            "targets": targets,
            "acceptedWarningIds": sorted(accepted_warning_ids),
            "warningFingerprint": warning_policy.get("fingerprint"),
        }

    def _eligible_report_contract(
        self,
        run_id: str,
        plan: dict[str, Any],
        report: dict[str, Any],
        run_dir: Path,
    ) -> None:
        from corvette_form_generator.editor_ops import classify_warnings, flatten_items

        if report.get("schemaVersion") != SCHEMA_VERSION_D:
            raise WizardError(f"Write approval requires a {SCHEMA_VERSION_D} dry-run report.", status=409)
        if not report.get("ok") or report.get("status") != "validated_write_eligible":
            raise WizardError("Write approval requires a validated_write_eligible dry-run report.", status=409)
        eligibility = report.get("writeEligibility") or {}
        if not eligibility.get("eligible"):
            raise WizardError("Dry-run report is not write eligible.", status=409)
        if eligibility.get("blockers"):
            raise WizardError("Write-eligible report still contains blockers.", status=409)
        targets = list(plan.get("targets") or [])
        report_targets = eligibility.get("targets") or {}
        if set(report_targets) != set(targets) or any(
            not (report_targets.get(model) or {}).get("eligible")
            or bool((report_targets.get(model) or {}).get("blockers"))
            for model in targets
        ):
            raise WizardError("Every target in the atomic plan must be write eligible.", status=409)
        plan_sha = hashlib.sha256((run_dir / "apply-plan.json").read_bytes()).hexdigest()
        if report.get("runId") != run_id or report.get("planSha") != plan_sha:
            raise WizardError("Dry-run report is stale for the current plan.", status=409)
        if report.get("planSchemaVersion") != WRITABLE_PLAN_SCHEMA:
            raise WizardError(f"Write approval requires {WRITABLE_PLAN_SCHEMA}.", status=409)
        if report.get("write"):
            raise WizardError("Write approval must bind a dry-run report.", status=409)
        if (
            report.get("schemaValidationEnabled") is not True
            or (report.get("schemaResult") or {}).get("error_count") != 0
        ):
            raise WizardError("Write approval requires a schema-validated eligible report.", status=409)
        reported_policy = report.get("warningPolicy") or (report.get("applyResult") or {}).get("warningPolicy")
        current_policy = classify_warnings(report.get("warnings") or [])
        if not reported_policy or reported_policy.get("fingerprint") != current_policy.get("fingerprint"):
            raise WizardError("Dry-run warning policy fingerprint is inconsistent.", status=409)
        if current_policy.get("blockingIds") or current_policy.get("unknownIds"):
            raise WizardError("Dry-run report contains blocking or unknown warnings.", status=409)
        if sorted(eligibility.get("acceptedWarningIds") or []) != sorted(
            current_policy.get("confirmableIds") or []
        ) or eligibility.get("warningFingerprint") != current_policy.get("fingerprint"):
            raise WizardError("Dry-run warning agreement is stale.", status=409)
        coverage = report.get("operationCoverage") or {}
        verification = report.get("verification") or {}
        apply_result = report.get("applyResult") or {}
        if coverage != (apply_result.get("operationCoverage") or {}):
            raise WizardError("Dry-run report does not match editor operation coverage.", status=409)
        if verification != (apply_result.get("verification") or {}):
            raise WizardError("Dry-run report does not match editor prepared readback.", status=409)
        combined_raw_count = len(
            flatten_items(
                [
                    *(plan.get("stage1", {}).get("items") or []),
                    *(plan.get("stage2", {}).get("items") or []),
                ]
            )
        )
        if coverage.get("rawCount") != combined_raw_count:
            raise WizardError("Dry-run raw operation count does not match the bound plan.", status=409)
        if coverage.get("rawCovered") != coverage.get("rawCount"):
            raise WizardError("Dry-run raw operation coverage is incomplete.", status=409)
        if not verification.get("ok") or verification.get("preparedChecked") != coverage.get("preparedCount"):
            raise WizardError("Dry-run prepared readback is incomplete.", status=409)
        deferrals = eligibility.get("deferrals") or []
        nested_deferrals = [
            item
            for model in targets
            for item in (report_targets.get(model) or {}).get("deferrals") or []
        ]
        def normalized(values: list[dict[str, Any]]) -> list[str]:
            return sorted(
                json.dumps(item, sort_keys=True, separators=(",", ":"))
                for item in values
            )
        if normalized(deferrals) != normalized(nested_deferrals):
            raise WizardError("Dry-run target deferrals do not match the atomic plan total.", status=409)
        self._require_compile_report_readiness(run_dir, targets)
        self._compile_report_not_applicable_families(run_dir, targets)
        self._validate_report_deferrals(run_dir, deferrals, targets)

    def approve_write(self, run_id: str, approver: str) -> dict[str, Any]:
        """Create write authority solely from current stored proof artifacts."""

        session = self.load_session(run_id)
        self._refuse_if_applied(session)
        if not str(approver or "").strip():
            raise WizardError("Write approval needs a reviewer name.")
        run_dir = self.run_dir(run_id)
        plan_file = run_dir / "apply-plan.json"
        plan_approval_file = run_dir / "plan-approval.json"
        report_file = run_dir / "apply-dry-run-report.json"
        if not plan_file.is_file() or not plan_approval_file.is_file() or not report_file.is_file():
            raise WizardError("Write approval requires plan, scoped approval, and dry-run report.", status=409)
        plan = read_json(plan_file)
        if plan.get("schemaVersion") != WRITABLE_PLAN_SCHEMA:
            raise WizardError(
                f"Write approval requires {WRITABLE_PLAN_SCHEMA}; older plans are diagnostic only.",
                status=409,
            )
        plan_approval = read_json(plan_approval_file)
        self._validate_plan_approval(
            run_id, session, plan, plan_approval, run_dir, allow_legacy=False
        )
        workbook, before = self._validate_current_plan_inputs(run_id, session, plan)
        del workbook
        self._require_compiler_bindings(
            run_dir,
            plan,
            (PLAN_APPROVAL_SCHEMA, plan_approval),
        )
        report = read_json(report_file)
        if report.get("approval") != plan_approval:
            raise WizardError("Dry-run report does not bind the current plan approval.", status=409)
        self._eligible_report_contract(run_id, plan, report, run_dir)
        if report.get("workbookBefore") != before or report.get("workbookAfter") != before:
            raise WizardError("Dry-run report workbook fingerprint is stale.", status=409)
        eligibility = report["writeEligibility"]
        approval = {
            "schemaVersion": WRITE_APPROVAL_SCHEMA,
            "scope": WRITE_APPROVAL_SCOPE,
            "approvedBy": str(approver).strip(),
            "approvedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
            **self._approval_bindings(run_id, session, plan, run_dir),
            "planApprovalSha": hashlib.sha256(plan_approval_file.read_bytes()).hexdigest(),
            "eligibleDryRunReportSha": hashlib.sha256(report_file.read_bytes()).hexdigest(),
            "acceptedWarningIds": list(eligibility.get("acceptedWarningIds") or []),
            "warningFingerprint": eligibility.get("warningFingerprint"),
            "allowedDeferrals": list(eligibility.get("deferrals") or []),
        }
        write_json(run_dir / "write-approval.json", approval)
        self._append_approval_audit(run_dir, approval)
        session["state"] = STATE_WRITE_APPROVED
        write_json(run_dir / "session.json", session)
        return {"session": session, "approval": approval}

    def _prewrite_authority(
        self,
        run_id: str,
        session: dict[str, Any],
        plan: dict[str, Any],
        plan_approval: dict[str, Any],
        run_dir: Path,
        *,
        schema_validation: bool,
    ) -> tuple[Path, dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any]]:
        """Repeat every refusal check before the first possible live mutation."""

        from corvette_form_generator.editor_ops import apply_batch, classify_warnings
        from corvette_form_generator.workbook import excel_lock_path

        if not schema_validation:
            raise WizardError("Live write requires schema validation.", status=409)
        if plan.get("schemaVersion") != WRITABLE_PLAN_SCHEMA:
            raise WizardError(
                f"Live write requires {WRITABLE_PLAN_SCHEMA}; older plans are permanently dry-run-only.",
                status=409,
            )
        self._validate_plan_approval(
            run_id, session, plan, plan_approval, run_dir, allow_legacy=False
        )
        if session.get("state") != STATE_WRITE_APPROVED:
            raise WizardError("Live write requires deployment_ready_write approval.", status=409)
        write_approval_file = run_dir / "write-approval.json"
        report_file = run_dir / "apply-dry-run-report.json"
        if not write_approval_file.is_file():
            raise WizardError("Live write requires write-approval.json.", status=409)
        if not report_file.is_file():
            raise WizardError("Live write requires the eligible dry-run report.", status=409)
        write_approval = read_json(write_approval_file)
        if write_approval.get("schemaVersion") != WRITE_APPROVAL_SCHEMA:
            raise WizardError(f"Live write requires {WRITE_APPROVAL_SCHEMA}.", status=409)
        if write_approval.get("scope") != WRITE_APPROVAL_SCOPE:
            raise WizardError("Write approval scope must be deployment_ready_write.", status=409)
        self._require_compiler_bindings(
            run_dir,
            plan,
            (PLAN_APPROVAL_SCHEMA, plan_approval),
            (WRITE_APPROVAL_SCHEMA, write_approval),
        )
        expected_bindings = self._approval_bindings(run_id, session, plan, run_dir)
        for field, value in expected_bindings.items():
            if write_approval.get(field) != value:
                raise WizardError(f"Write approval binding {field} changed.", status=409)
        if write_approval.get("planApprovalSha") != hashlib.sha256(
            (run_dir / "plan-approval.json").read_bytes()
        ).hexdigest():
            raise WizardError("Plan approval changed after write approval.", status=409)
        report_sha = hashlib.sha256(report_file.read_bytes()).hexdigest()
        if write_approval.get("eligibleDryRunReportSha") != report_sha:
            raise WizardError("Eligible dry-run report SHA changed after write approval.", status=409)
        report = read_json(report_file)
        if report.get("approval") != plan_approval:
            raise WizardError("Eligible report no longer matches plan approval.", status=409)
        self._eligible_report_contract(run_id, plan, report, run_dir)

        workbook, before = self._validate_current_plan_inputs(run_id, session, plan)
        if excel_lock_path(workbook).exists():
            raise WizardError("Excel lock file is present; close Excel before live write.", status=409)
        if report.get("workbookBefore") != before or report.get("workbookAfter") != before:
            raise WizardError("Eligible report workbook fingerprint changed.", status=409)
        eligibility = report["writeEligibility"]
        approved_deferrals = write_approval.get("allowedDeferrals") or []
        if approved_deferrals != (eligibility.get("deferrals") or []):
            raise WizardError("Allowed deferral set drifted after write approval.", status=409)
        self._validate_report_deferrals(
            run_dir,
            approved_deferrals,
            [str(model) for model in plan.get("targets") or []],
        )

        batch = self._combined_plan_batch(plan, workbook)
        preview = apply_batch(
            workbook,
            batch,
            write=False,
            run_schema_validation=True,
        )
        if not preview.get("ok"):
            raise WizardError(
                "Pre-write temporary-workbook proof no longer passes: "
                + "; ".join(preview.get("errors") or [str(preview.get("status"))]),
                status=409,
            )
        current_policy = preview.get("warningPolicy") or classify_warnings(preview.get("warnings") or [])
        if current_policy.get("blockingIds"):
            raise WizardError("Pre-write proof emitted blocking or unknown warnings.", status=409)
        if write_approval.get("warningFingerprint") != current_policy.get("fingerprint"):
            raise WizardError("Warning set drifted after write approval.", status=409)
        if sorted(write_approval.get("acceptedWarningIds") or []) != sorted(
            current_policy.get("confirmableIds") or []
        ):
            raise WizardError("Accepted warning IDs drifted after write approval.", status=409)
        coverage = preview.get("operationCoverage") or {}
        verification = preview.get("verification") or {}
        if coverage != (report.get("operationCoverage") or {}):
            raise WizardError("Pre-write operation coverage drifted from the eligible report.", status=409)
        if verification != (report.get("verification") or {}):
            raise WizardError("Pre-write prepared readback drifted from the eligible report.", status=409)
        if coverage.get("rawCovered") != coverage.get("rawCount"):
            raise WizardError("Pre-write raw operation coverage is incomplete.", status=409)
        if not verification.get("ok") or verification.get("preparedChecked") != coverage.get("preparedCount"):
            raise WizardError("Pre-write prepared readback is incomplete.", status=409)
        return workbook, before, write_approval, preview, report

    def apply_approved_plan(
        self,
        run_id: str,
        *,
        write: bool = False,
        confirm_plan_warnings: bool = False,
        schema_validation: bool = True,
    ) -> dict[str, Any]:
        """Produce bound dry-run evidence or execute separately approved write authority."""

        from corvette_form_generator.editor_ops import apply_batch, flatten_items

        started_at = datetime.now().isoformat(timespec="seconds")
        session = self.load_session(run_id)
        self._refuse_if_applied(session)
        if session["state"] not in {
            STATE_PLAN_APPROVED,
            STATE_DRY_RUN_APPROVED,
            STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED,
            STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE,
            STATE_WRITE_APPROVED,
        }:
            raise WizardError("Apply requires an approved plan.", status=409)

        run_dir = self.run_dir(run_id)
        plan_file = run_dir / "apply-plan.json"
        approval_file = run_dir / "plan-approval.json"
        if not plan_file.is_file() or not approval_file.is_file():
            raise WizardError("Apply requires apply-plan.json and plan-approval.json.", status=409)

        plan = read_json(plan_file)
        approval = read_json(approval_file)
        plan_sha = hashlib.sha256(plan_file.read_bytes()).hexdigest()
        if not plan.get("valid"):
            raise WizardError("Apply plan is not valid; rebuild before applying.", status=409)
        if write and not schema_validation:
            raise WizardError("Live write requires schema validation.", status=409)
        if write and plan.get("schemaVersion") != WRITABLE_PLAN_SCHEMA:
            raise WizardError(
                f"Live write requires {WRITABLE_PLAN_SCHEMA}; older plans are permanently dry-run-only.",
                status=409,
            )
        if write and confirm_plan_warnings:
            raise WizardError("Blanket plan-warning confirmation is retired.", status=409)

        legacy_approval = self._validate_plan_approval(
            run_id,
            session,
            plan,
            approval,
            run_dir,
            allow_legacy=not write,
        )
        if write:
            workbook, before, write_approval, _preview, approved_report = self._prewrite_authority(
                run_id,
                session,
                plan,
                approval,
                run_dir,
                schema_validation=schema_validation,
            )
        else:
            workbook, before = self._validate_current_plan_inputs(run_id, session, plan)
            write_approval = None
            approved_report = None

        batch = self._combined_plan_batch(plan, workbook)
        per_sheet_action_counts = self._per_sheet_action_counts(batch["items"])
        per_sheet_counts: dict[str, int] = {}
        for item in batch["items"]:
            sheet = item.get("sheet")
            if sheet:
                per_sheet_counts[str(sheet)] = per_sheet_counts.get(str(sheet), 0) + 1
        confirmed_warnings = list((write_approval or {}).get("acceptedWarningIds") or [])
        log_path = run_dir / "apply-workbook-edit-log.jsonl"
        result = apply_batch(
            workbook,
            batch,
            write=write,
            source="ingest_wizard_apply",
            confirmed_warnings=confirmed_warnings,
            log_path=log_path,
            run_schema_validation=schema_validation,
        )
        # A pre-mutation refusal must not create or replace report evidence.
        if write and not result.get("ok") and not result.get("backupPath"):
            return result
        after = file_fingerprint(workbook)
        completed_at = datetime.now().isoformat(timespec="seconds")
        verification = result.get("verification") or {
            "ok": False,
            "preparedChecked": 0,
            "preparedCount": 0,
            "errors": ["editor result did not include prepared readback"],
        }
        if result.get("ok") and write:
            deployment_continuity = dict((approved_report or {}).get("deploymentContinuity") or {})
        elif result.get("ok"):
            deployment_continuity = self._deployment_continuity_probe(
                workbook,
                batch,
                plan,
                schema_validation=schema_validation,
            )
        else:
            deployment_continuity = {}
        if result.get("ok"):
            if write:
                write_eligibility = dict((approved_report or {}).get("writeEligibility") or {})
            else:
                write_eligibility = self._write_eligibility(
                    workbook=workbook,
                    plan=plan,
                    approval=approval,
                    legacy_approval=legacy_approval,
                    schema_validation=schema_validation,
                    result=result,
                    deployment=deployment_continuity,
                )
        else:
            execution_blocker = {
                "kind": "mechanical_validation_failed",
                "detail": "; ".join(result.get("errors") or [str(result.get("status"))]),
            }
            write_eligibility = {
                "eligible": False,
                "blockers": [execution_blocker],
                "deferrals": [],
                "targets": {
                    str(model): {"eligible": False, "blockers": [execution_blocker], "deferrals": []}
                    for model in plan.get("targets") or []
                },
                "acceptedWarningIds": [],
                "warningFingerprint": (result.get("warningPolicy") or {}).get("fingerprint"),
            }
        diagnostic_status = (
            "validated_write_eligible" if write_eligibility["eligible"] else "validated_write_blocked"
        )
        status = result.get("status") if not result.get("ok") or write else diagnostic_status
        blocked_reason = None
        if not write_eligibility["eligible"]:
            blockers = write_eligibility.get("blockers") or []
            blocked_reason = str((blockers[0] if blockers else {}).get("detail") or "write eligibility is blocked")
        report = {
            "schemaVersion": SCHEMA_VERSION_D,
            "planSchemaVersion": plan.get("schemaVersion"),
            "planSupersededForWrite": plan.get("schemaVersion") != WRITABLE_PLAN_SCHEMA,
            "liveWriteBlockedReason": blocked_reason,
            "writeEligibility": write_eligibility,
            "runId": run_id,
            "startedAt": started_at,
            "completedAt": completed_at,
            "appliedAt": completed_at,
            "write": write,
            "schemaValidationEnabled": bool(schema_validation),
            "status": status,
            "ok": result["ok"],
            "planSha": plan_sha,
            "approvedBy": approval.get("approvedBy"),
            "approvedAt": approval.get("approvedAt"),
            "approval": approval,
            "opCounts": {
                "stage1": len(flatten_items(plan["stage1"]["items"])),
                "stage2": len(flatten_items(plan["stage2"]["items"])),
                "combinedRaw": (result.get("operationCoverage") or {}).get("rawCount", 0),
                "prepared": (result.get("operationCoverage") or {}).get(
                    "preparedCount", 0
                ),
            },
            "perSheetCounts": dict(sorted(per_sheet_counts.items())),
            "perSheetActionCounts": per_sheet_action_counts,
            "runtimeContinuity": (plan.get("report") or {}).get("runtimeContinuity", {}),
            "deploymentContinuity": deployment_continuity,
            "workbookBefore": before,
            "workbookAfter": after,
            "warnings": result.get("warnings", []),
            "confirmedWarnings": confirmed_warnings,
            "warningsConfirmed": confirmed_warnings,
            "applyResult": result,
            "warningPolicy": result.get("warningPolicy"),
            "operationCoverage": result.get("operationCoverage"),
            "schemaResult": result.get("schemaResult"),
            "boolHygieneResult": result.get("boolHygieneResult"),
            "gateReminders": result.get("gateReminders", []),
            "sheets": result.get("sheets", []),
            "backupPath": result.get("backupPath"),
            "workbookEditLogPath": result.get("logPath"),
            "verification": verification,
        }
        report_path = run_dir / ("apply-report.json" if write else "apply-dry-run-report.json")
        write_json(report_path, report)
        if not result.get("ok"):
            result["reportPath"] = str(report_path)
            result["verification"] = verification
            result["writeEligibility"] = write_eligibility
            result["liveWriteBlockedReason"] = blocked_reason
            if write and result.get("status") == "apply_verification_failed":
                session["state"] = STATE_APPLY_VERIFICATION_FAILED
                session["applyReport"] = "apply-report.json"
                write_json(run_dir / "session.json", session)
            return result
        if write:
            session["state"] = STATE_APPLIED
            session["appliedAt"] = report["appliedAt"]
            session["applyReport"] = "apply-report.json"
            write_json(run_dir / "session.json", session)
        else:
            session["state"] = (
                STATE_DRY_RUN_VALIDATED_WRITE_ELIGIBLE
                if write_eligibility["eligible"]
                else STATE_DRY_RUN_VALIDATED_WRITE_BLOCKED
            )
            session["dryRunReport"] = "apply-dry-run-report.json"
            write_json(run_dir / "session.json", session)
        result["reportPath"] = str(report_path)
        result["verification"] = verification
        result["status"] = status
        result["writeEligibility"] = write_eligibility
        result["liveWriteBlockedReason"] = blocked_reason
        return result

    def mark_complete(self, run_id: str) -> dict[str, Any]:
        session = self.load_session(run_id, verify_source=False)
        self._refuse_if_applied(session)
        report = self.progress(run_id)
        if not report["allComplete"]:
            blockers = {
                model: entry["blockers"] for model, entry in report["models"].items() if entry["blockers"]
            }
            raise WizardError(
                "Decisions are not complete; blocking items remain: " + json.dumps(blockers)[:2000],
                status=409,
            )
        run_dir = self.run_dir(run_id)
        session = self.load_session(run_id, verify_source=False)
        session["state"] = STATE_DECISIONS_COMPLETE
        write_json(run_dir / "session.json", session)
        report["session"] = session
        return report
