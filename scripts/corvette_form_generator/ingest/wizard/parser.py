#!/usr/bin/env python3
"""Deterministic option/price extraction for the ingest wizard (Pass A, step 4).

Reads only user-confirmed sheets. Preserves raw cells as evidence, parses
per-variant OVS statuses, and extracts price-schedule rows. Rows that fail
deterministic parsing are reported, never silently dropped.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from corvette_form_generator.ingest.source_profiler import RPO_RE, parse_status
from corvette_form_generator.ingest.wizard.profiler import (
    MODEL_CODE_RE,
    PRICE_SECTION_BASE,
    PRICE_SECTION_OPTIONS,
    ROLE_OPTIONS,
    ROLE_PRICE,
    SCHEMA_VERSION,
    SHEET_TYPE_OPTIONS,
    cell_text,
    find_price_sections,
    profile_sheet,
    sheet_values,
)
from corvette_form_generator.workbook import clean


def parse_confirmed_sheets(path: Path, roles: dict[str, str]) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    candidates: list[dict[str, Any]] = []
    skipped_rows: dict[str, list[dict[str, Any]]] = {}
    price_rows: list[dict[str, Any]] = []
    base_rows: list[dict[str, Any]] = []
    skipped_price: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            role = roles.get(ws.title, "")
            if role == ROLE_OPTIONS:
                values = sheet_values(ws)
                card = profile_sheet(ws.title, values)
                sheet_candidates, sheet_skipped = extract_option_candidates(ws.title, values, card)
                candidates.extend(sheet_candidates)
                if sheet_skipped:
                    skipped_rows[ws.title] = sheet_skipped
            elif role == ROLE_PRICE:
                values = sheet_values(ws)
                sheet_prices, sheet_base, sheet_skipped_price = extract_price_rows(ws.title, values)
                price_rows.extend(sheet_prices)
                base_rows.extend(sheet_base)
                skipped_price.extend(sheet_skipped_price)
    finally:
        wb.close()
    return {
        "schemaVersion": SCHEMA_VERSION,
        "candidates": candidates,
        "priceRows": price_rows,
        "baseModelPriceRows": base_rows,
        "skippedRows": skipped_rows,
        "skippedPriceRows": skipped_price,
    }


def extract_option_candidates(
    sheet_name: str,
    values: list[list[Any]],
    card: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if card["sheetType"] != SHEET_TYPE_OPTIONS or not card["headerRow"]:
        return [], [{"rowIndex": None, "reason": "sheet_not_options_matrix", "description": ""}]
    variant_columns = card["variantColumns"]
    candidates: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    section_label = ""
    for row_number in range(card["headerRow"] + 1, len(values) + 1):
        row = values[row_number - 1]
        first = cell_text(row, 0)
        second = cell_text(row, 1)
        third = cell_text(row, 2)
        statuses = extract_statuses(row, variant_columns)
        if not first and not second and not third and not statuses:
            continue
        if first and not second and not third and not RPO_RE.fullmatch(first.upper()):
            section_label = first
            continue
        rpo = first.upper() if RPO_RE.fullmatch(first.upper()) else ""
        ref_only = second.upper() if RPO_RE.fullmatch(second.upper()) else ""
        if not rpo and not ref_only:
            skipped.append(
                {"rowIndex": row_number, "reason": "no_rpo_on_content_row", "description": third}
            )
            continue
        candidates.append(
            {
                "candidateId": f"{sheet_name}:{row_number}",
                "sheetName": sheet_name,
                "rowIndex": row_number,
                "modelFamily": card["modelFamily"],
                "modelFamilies": card["modelFamilies"],
                "sectionLabel": section_label,
                "rowKind": "orderable" if rpo else "ref_only",
                "rpo": rpo,
                "refOnlyRpo": ref_only,
                "description": third,
                "statuses": statuses,
                "sourceEvidence": row_evidence(sheet_name, row_number, row),
            }
        )
    return candidates, skipped


def extract_statuses(row: list[Any], variant_columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    statuses: list[dict[str, Any]] = []
    for column in variant_columns:
        raw = cell_text(row, column["columnIndex"] - 1)
        if not raw:
            continue
        parsed = parse_status(raw)
        statuses.append(
            {
                "columnLetter": column["columnLetter"],
                "variantLabel": column["label"],
                "modelCode": column["modelCode"],
                "trim": column["trim"],
                "bodyStyle": column["bodyStyle"],
                "raw": raw,
                "status": parsed["parsed_base_status"],
                "disclosureMarker": parsed["status_marker"],
                "flags": parsed["status_flags"],
            }
        )
    return statuses


def row_evidence(sheet_name: str, row_number: int, row: list[Any]) -> dict[str, Any]:
    cells = {
        f"{get_column_letter(index + 1)}{row_number}": str(value)
        for index, value in enumerate(row)
        if value is not None and clean(value) != ""
    }
    return {"sheetName": sheet_name, "rowIndex": row_number, "cells": cells}


def numeric_value(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def numeric_columns(row: list[Any], row_number: int, start_index: int) -> dict[str, float]:
    columns: dict[str, float] = {}
    for index in range(start_index, len(row)):
        number = numeric_value(row[index])
        if number is not None:
            columns[f"{get_column_letter(index + 1)}{row_number}"] = number
    return columns


def extract_price_rows(
    sheet_name: str,
    values: list[list[Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    sections = find_price_sections(values)
    options_start = sections.get(PRICE_SECTION_OPTIONS, 0)
    base_start = sections.get(PRICE_SECTION_BASE, 0)
    option_rows: list[dict[str, Any]] = []
    base_rows: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    if base_start:
        base_end = options_start - 1 if options_start > base_start else len(values)
        for row_number in range(base_start + 1, base_end + 1):
            row = values[row_number - 1]
            code = cell_text(row, 1).upper()
            if not MODEL_CODE_RE.fullmatch(code):
                continue
            price_columns = numeric_columns(row, row_number, 3)
            if not price_columns:
                skipped.append(
                    {"sheetName": sheet_name, "rowIndex": row_number, "reason": "no_numeric_price"}
                )
                continue
            base_rows.append(
                {
                    "modelCode": code,
                    "description": cell_text(row, 2),
                    "listPrice": next(iter(price_columns.values())),
                    "sourceEvidence": row_evidence(sheet_name, row_number, row),
                }
            )
    if options_start:
        for row_number in range(options_start + 1, len(values) + 1):
            row = values[row_number - 1]
            if cell_text(row, 0) in (PRICE_SECTION_BASE, PRICE_SECTION_OPTIONS):
                break
            code = cell_text(row, 1).upper()
            if not RPO_RE.fullmatch(code):
                continue
            qualifier_cell = row[3] if len(row) > 3 else None
            qualifier = (
                clean(qualifier_cell)
                if numeric_value(qualifier_cell) is None and clean(qualifier_cell)
                else ""
            )
            price_columns = numeric_columns(row, row_number, 3)
            if not price_columns:
                skipped.append(
                    {"sheetName": sheet_name, "rowIndex": row_number, "reason": "no_numeric_price"}
                )
                continue
            option_rows.append(
                {
                    "rpo": code,
                    "description": cell_text(row, 2),
                    "qualifier": qualifier,
                    "listPrice": next(iter(price_columns.values())),
                    "priceColumns": price_columns,
                    "sourceEvidence": row_evidence(sheet_name, row_number, row),
                }
            )
    return option_rows, base_rows, skipped
