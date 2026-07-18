#!/usr/bin/env python3
"""Pass C plan builder: reviewed decisions → a two-stage workbook op plan.

Deterministic translation only — same decisions in, byte-identical plan out
(no timestamps inside the plan). Stage 1 scaffolds model metadata and sheets;
stage 2 writes data rows. Ops are editor_ops batch items joined by stable IDs
(RPO / variant_id / section_id / option_id), never row numbers. The live
workbook is opened read-only here; applying is Pass D's job.
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.editor_ops import (
    EDITOR_SHEET_META,
    GLOBAL_SHEET_FAMILIES,
    SOURCE_ROLE_FAMILIES,
    extract_workbook,
    model_sheet_registry,
)
from corvette_form_generator.ingest.wizard.copy_split import propose_copy_split
from corvette_form_generator.ingest.wizard.decisions import (
    MODEL_LABELS,
    PRESENTATION_SHEETS,
    canonical_json,
    candidate_needs_section_decision,
    scope_candidates,
)
from corvette_form_generator.model_configs import base_model_config
from corvette_form_generator.workbook import rows_from_sheet, workbook_truthy

SCHEMA_VERSION_C = "pass-c-2"

# Per-model identity, per the approved end-to-end spec (resolved decisions).
MODEL_PLAN_CONFIG: dict[str, dict[str, Any]] = {
    "grand_sport_x": {
        "sheetPrefix": "grandSportX_",
        "templatePrefix": "grandSport_",
        "registryKey": "grand_sport_x",
        "exportSlug": "grand-sport-x",
        "label": "Grand Sport X",
        "interiorSheet": "lt_interiors",
        "isNewModel": True,
    },
    "zr1": {
        "sheetPrefix": "zr1_",
        "templatePrefix": "z06_",
        "registryKey": "zr1",
        "exportSlug": "zr1",
        "label": "ZR1",
        "interiorSheet": "LZ_Interiors",
        "isNewModel": False,
    },
    "zr1x": {
        "sheetPrefix": "zr1x_",
        "templatePrefix": "z06_",
        "registryKey": "zr1x",
        "exportSlug": "zr1x",
        "label": "ZR1X",
        "interiorSheet": "LZ_Interiors",
        "isNewModel": False,
    },
}

MODEL_SHEET_ROLES: tuple[tuple[str, str], ...] = (
    ("source_option_sheet", "options"),
    ("status_sheet", "ovs"),
    ("rule_mapping_sheet", "rule_mapping"),
    ("price_rules_sheet", "price_rules"),
    ("rule_groups_sheet", "rule_groups"),
    ("rule_group_members_sheet", "rule_group_members"),
    ("exclusive_groups_sheet", "exclusive_groups"),
    ("exclusive_group_members_sheet", "exclusive_members"),
    ("variant_option_overrides_sheet", "variant_overrides"),
)

ROLE_SHEET_SUFFIX: dict[str, str] = {
    "source_option_sheet": "options",
    "status_sheet": "ovs",
    "rule_mapping_sheet": "rule_mapping",
    "price_rules_sheet": "price_rules",
    "rule_groups_sheet": "rule_groups",
    "rule_group_members_sheet": "rule_group_members",
    "exclusive_groups_sheet": "exclusive_groups",
    "exclusive_group_members_sheet": "exclusive_members",
    "variant_option_overrides_sheet": "variant_overrides",
}

# ZR1/ZR1X use grandSport-style "exclusive_members" suffix already; template
# sheet names for headers come from the template prefix + same suffix.
RELATIONSHIP_KIND_TO_RULE_TYPE = {
    "requires": "requires",
    "only_available_with": "requires",
    "requires_additional_equipment": "requires",
    "not_available_with": "excludes",
    "includes": "includes",
    "included_with": "includes",
}

BLOCKING_GAP_KINDS = {
    "presentation_missing",
    "missing_mandatory_decision",
    "no_variants_mapped",
    "option_display_order_missing",
    "option_display_order_collision",
    "price_rule_unresolved_required",
    "rule_group_unresolved_required",
    "relationship_option_identity_unresolved",
    "relationship_target_missing",
    "relationship_unmappable",
    "exclusive_members_missing",
    "exterior_paint_rows_missing_required",
    "model_interior_scope_missing_required",
    "model_interior_scope_conflict",
    "default_selection_rules_missing",
    "default_selection_rule_unresolved_required",
    "missing_template",
}


class PlanGap:
    """Named open item the plan cannot express — carried, never silent."""

    def __init__(self, model: str, kind: str, detail: str, decision_id: str = "") -> None:
        self.entry = {"model": model, "kind": kind, "detail": detail, "decisionId": decision_id}


def artifact_sha(payload: Any) -> str:
    return hashlib.sha256(canonical_json(payload).encode("utf-8")).hexdigest()


def slug_rpo(rpo: str) -> str:
    return re.sub(r"[^a-z0-9]", "", rpo.lower())


def _workbook_context(workbook_path: Path, targets: list[str]) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        global_sheets = {
            "asset_map",
            "color_overrides",
            "default_selection_rules",
            "interior_components",
            "model_interior_scope",
        }
        for target in targets:
            global_sheets.add(MODEL_PLAN_CONFIG[target]["interiorSheet"])
        context: dict[str, Any] = {
            "sheetnames": list(wb.sheetnames),
            "model_master": rows_from_sheet(wb, "model_master"),
            "model_variants": rows_from_sheet(wb, "model_variants"),
            "variant_master": rows_from_sheet(wb, "variant_master"),
            "model_workbook_sources": rows_from_sheet(wb, "model_workbook_sources"),
            "model_registry_promotion": rows_from_sheet(wb, "model_registry_promotion"),
            "existing_rows": {},
            "global_rows": {},
        }
        for sheet in sorted(global_sheets):
            if sheet in wb.sheetnames:
                context["global_rows"][sheet] = rows_from_sheet(wb, sheet)
        for target in targets:
            config = MODEL_PLAN_CONFIG[target]
            for _, suffix in MODEL_SHEET_ROLES:
                sheet = config["sheetPrefix"] + suffix
                if sheet in wb.sheetnames:
                    context["existing_rows"][sheet] = rows_from_sheet(wb, sheet)
        return context
    finally:
        wb.close()


def _variant_map(context: dict[str, Any], model: str) -> dict[tuple[str, str], str]:
    """(trim_lower, body_digits) -> variant_id from variant_master rows whose
    display name belongs to this model (model_variants when present)."""

    ids = {row["variant_id"] for row in context["model_variants"] if row.get("model_key") == model}
    label = MODEL_PLAN_CONFIG[model]["label"].lower()
    mapping: dict[tuple[str, str], str] = {}
    for row in context["variant_master"]:
        vid = row.get("variant_id", "")
        if not vid:
            continue
        if ids:
            if vid not in ids:
                continue
        elif label not in row.get("display_name", "").lower():
            continue
        body = "07" if row.get("body_style") == "coupe" else "67"
        mapping[(row.get("trim_level", "").lower(), body)] = vid
    return mapping


def _status_for_variant(candidate: dict[str, Any], model_code_suffix: str, trim: str) -> str | None:
    for status in candidate.get("statuses", []):
        code = str(status.get("modelCode") or "")
        if code.endswith(model_code_suffix) and str(status.get("trim") or "").lower() == trim:
            value = status.get("status")
            return value if value in ("standard", "available", "unavailable") else None
    return None


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _norm(value: Any) -> str:
    if isinstance(value, bool):
        return "True" if value else "False"
    return _clean(value)


def _intish(value: Any, default: int = 0) -> int:
    if isinstance(value, bool):
        return default
    if isinstance(value, int):
        return value
    text = _clean(value)
    return int(text) if text.lstrip("-").isdigit() else default


def _rpo_token(value: Any) -> str:
    return _clean(value).upper()


def _listish(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    return [value]


def _planned_relationship_rpos(records: dict[str, dict[str, Any]]) -> set[str]:
    rpos: set[str] = set()
    for record in records.values():
        if record.get("resolution") != "approved_for_plan":
            continue
        payload = record.get("payload") or {}
        if record.get("lane") == "relationship" and record.get("action") == "create_relationship_candidate":
            source = _rpo_token(payload.get("sourceRpo"))
            if source:
                rpos.add(source)
            rpos.update(_rpo_token(rpo) for rpo in payload.get("targetRpos") or [] if _rpo_token(rpo))
        if record.get("lane") == "exclusive_group" and record.get("action") == "create_exclusive_group":
            rpos.update(_rpo_token(rpo) for rpo in payload.get("members") or [] if _rpo_token(rpo))
    return rpos


def plan_summary(plan: dict[str, Any]) -> dict[str, Any]:
    """UI-facing summary: everything except the raw op lists."""

    summary = {
        "schemaVersion": plan["schemaVersion"],
        "targets": plan["targets"],
        "valid": plan["valid"],
        "stage1Count": len(plan["stage1"]["items"]),
        "stage2Count": len(plan["stage2"]["items"]),
        "report": plan["report"],
        "uncoveredApprovedDecisions": plan["coverage"]["uncoveredApprovedDecisions"],
        "workbookFingerprint": plan["workbookFingerprint"],
        "decisionsFingerprint": plan["decisionsFingerprint"],
    }
    if plan.get("schemaVersion") == "pass-c-3":
        summary["planReadiness"] = plan.get("planReadiness")
        summary["canonicalManifestSemanticSha"] = plan.get(
            "canonicalManifestSemanticSha"
        )
    return summary


def plan_markdown(plan: dict[str, Any], dry_run: dict[str, Any]) -> str:
    report = plan["report"]
    lines = [
        "# Ingest apply plan",
        "",
        f"Targets: {', '.join(plan['targets'])}",
        f"Stage 1 (scaffolding): {len(plan['stage1']['items'])} ops · Stage 2 (data): {len(plan['stage2']['items'])} ops",
        f"Plan valid: {plan['valid']} · Dry run ok: {dry_run.get('ok')}",
        "",
        "## Per-sheet ops",
    ]
    for sheet in sorted(report["perSheetCounts"]):
        counts = report["perSheetCounts"][sheet]
        summary = ", ".join(f"{action} {count}" for action, count in sorted(counts.items()))
        lines.append(f"- {sheet}: {summary}")
    if report["clearedRows"]:
        lines += ["", "## Scaffold rows cleared (clean reprocess)"]
        lines += [f"- {sheet}: {count} rows" for sheet, count in sorted(report["clearedRows"].items())]
    if report["unreviewedSplits"]:
        lines += ["", "## Script splits carried unreviewed"]
        lines += [
            f"- {model}: {len(rpos)} options ({', '.join(rpos[:12])}{'…' if len(rpos) > 12 else ''})"
            for model, rpos in sorted(report["unreviewedSplits"].items())
        ]
    if report["holds"]:
        lines += ["", "## Holds (answers still owed)"]
        lines += [f"- {h['model']} · {h['decisionId']} — {h['note']}" for h in report["holds"]]
    if report["deferrals"]:
        lines += ["", "## Deferred work items"]
        lines += [f"- {d['model']} · {d['groupKey']}" for d in report["deferrals"]]
    if report["gaps"]:
        lines += ["", "## Gaps"]
        lines += [f"- [{g['kind']}] {g['model']}: {g['detail']}" for g in report["gaps"]]
    return "\n".join(lines) + "\n"


MANIFEST_STAGE1_FAMILIES = frozenset(
    {
        "model_master",
        "model_variants",
        "variant_master",
        "model_workbook_sources",
        "model_registry_promotion",
    }
)
MANIFEST_ACTION_ORDER = {"add": 0, "update": 1, "delete": 2, "noop": 3}


def _manifest_key_text(key: dict[str, Any]) -> str:
    return canonical_json({str(column): key[column] for column in sorted(key)})


def _manifest_normalized_row(family: str, row: dict[str, Any]) -> dict[str, Any]:
    types = EDITOR_SHEET_META[family].get("types") or {}
    normalized: dict[str, Any] = {}
    for column, raw_value in row.items():
        value = None if raw_value == "" else raw_value
        if value is not None and types.get(column) == "int":
            text = str(value).strip()
            value = int(text) if text.lstrip("-").isdigit() else value
        elif value is not None and types.get(column) == "bool":
            value = workbook_truthy(value)
        normalized[column] = value
    return normalized


def build_manifest_plan(
    *,
    workbook_path: Path,
    manifest: dict[str, Any],
    compile_report: dict[str, Any],
    selection: dict[str, Any],
    compiler_bindings: dict[str, str],
    authority_artifacts: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Project an exact-current canonical manifest into a pass-c-3 plan.

    This function validates and translates. It does not rematch identity,
    derive business behavior, or clean target sheets beyond manifest actions.
    """

    workbook_path = Path(workbook_path)
    targets = [str(target) for target in selection.get("targets") or []]
    if not targets or len(targets) != len(set(targets)):
        raise ValueError("Canonical plan needs a nonempty unique target selection.")
    modes = {str(key): str(value) for key, value in (manifest.get("modelModes") or {}).items()}
    report_models = {
        str(key): dict(value or {})
        for key, value in (compile_report.get("models") or {}).items()
    }
    if set(targets) != set(modes) or set(targets) != set(report_models):
        raise ValueError("Canonical plan target set does not match compiler artifacts.")
    required_authority_artifacts = {
        "exceptionQueue",
        "resolutions",
        "comparatorEvidence",
    }
    if set(authority_artifacts) != required_authority_artifacts:
        raise ValueError(
            "Canonical plan requires the complete exception queue, resolutions, "
            "and comparator evidence artifact set."
        )
    queue = authority_artifacts["exceptionQueue"]
    resolutions = authority_artifacts["resolutions"]
    comparator_evidence = authority_artifacts["comparatorEvidence"]
    if set(comparator_evidence.get("targets") or {}) != set(targets):
        raise ValueError(
            "Canonical plan comparator evidence target set does not match selection."
        )
    authority = manifest.get("runAuthorityFingerprint")
    if not authority or any(
        artifact.get("runAuthorityFingerprint") != authority
        for artifact in (compile_report, queue, comparator_evidence)
    ):
        raise ValueError("Canonical plan authority fingerprints do not agree.")
    queue_fingerprint = str(queue.get("queueSubjectFingerprint") or "")
    if not queue_fingerprint or any(
        str(value or "") != queue_fingerprint
        for value in (
            compile_report.get("queueSubjectFingerprint"),
            resolutions.get("queueSubjectFingerprint"),
        )
    ):
        raise ValueError("Canonical plan queue and resolution fingerprints do not agree.")
    comparator_semantic_sha = str(
        comparator_evidence.get("comparatorEvidenceSemanticSha") or ""
    )
    if not comparator_semantic_sha or any(
        str(value or "") != comparator_semantic_sha
        for value in (
            compile_report.get("comparatorEvidenceSemanticSha"),
            queue.get("comparatorEvidenceSemanticSha"),
        )
    ):
        raise ValueError("Canonical plan comparator evidence bindings do not agree.")
    unknown_subject_models = {
        str(subject.get("model") or "")
        for subject in queue.get("subjects") or []
        if str(subject.get("model") or "") not in set(targets)
    }
    if unknown_subject_models:
        raise ValueError(
            "Canonical plan exception queue contains subjects outside the selected "
            f"targets: {sorted(unknown_subject_models)}"
        )
    current_subjects = {
        (str(subject.get("subjectId") or ""), str(subject.get("subjectVersion") or ""))
        for subject in queue.get("subjects") or []
    }
    orphan_resolutions = [
        entry
        for entry in resolutions.get("validEntries") or []
        if (
            str(entry.get("subjectId") or ""),
            str(entry.get("subjectVersion") or ""),
        )
        not in current_subjects
    ]
    if orphan_resolutions:
        raise ValueError(
            "Canonical plan resolutions do not map to the exact current exception queue."
        )
    for target in targets:
        model_report = report_models[target]
        if not model_report.get("compileReady") or model_report.get("blockers"):
            raise ValueError(f"Canonical target {target} is not compile-ready.")
        if str(model_report.get("mode") or "") != modes[target]:
            raise ValueError(f"Canonical target {target} mode does not match the manifest.")
    if compile_report.get("deferrals"):
        raise ValueError("Canonical plan cannot project unresolved compiler deferrals.")

    extract = extract_workbook(workbook_path)
    sheets = extract.get("sheets") or {}
    registry, registered_family = model_sheet_registry(extract)
    registered_family = {**GLOBAL_SHEET_FAMILIES, **registered_family}
    comparators = {
        str(key): str(value)
        for key, value in (selection.get("comparators") or {}).items()
    }
    if set(comparators) != set(targets):
        raise ValueError("Canonical plan comparator targets do not match the selection.")

    rows = [dict(row) for row in manifest.get("rows") or []]
    if not rows:
        raise ValueError("Canonical manifest has no rows to project.")
    greenfield_targets = {
        model for model, mode in modes.items() if mode == "greenfield"
    }
    isolated_role_by_family = {
        family: role for role, family in MODEL_SHEET_ROLES
    }
    greenfield_sheets = {
        (model, family): str(getattr(base_model_config(model), role))
        for model in greenfield_targets
        for family, role in isolated_role_by_family.items()
    }
    target_option_ids = {
        model: {
            str(row.get("key", {}).get("option_id") or "")
            for row in rows
            if row.get("model") == model and row.get("family") == "options"
        }
        for model in greenfield_targets
    }
    target_interior_ids = {
        model: {
            str(row.get("key", {}).get("interior_id") or "")
            for row in rows
            if row.get("model") == model
            and row.get("family") == "model_interior_scope"
        }
        for model in greenfield_targets
    }
    rule_group_ids_with_applicable_members = {
        model: {
            str(row.get("values", {}).get("group_id") or "")
            for row in rows
            if row.get("model") == model
            and row.get("family") == "rule_group_members"
            and str(row.get("values", {}).get("target_id") or "")
            in target_option_ids[model] | target_interior_ids[model]
        }
        for model in greenfield_targets
    }
    applicable_rule_group_ids = {
        model: {
            str(row.get("values", {}).get("group_id") or "")
            for row in rows
            if row.get("model") == model
            and row.get("family") == "rule_groups"
            and str(row.get("values", {}).get("group_id") or "")
            in rule_group_ids_with_applicable_members[model]
            and str(row.get("values", {}).get("source_id") or "")
            in target_option_ids[model] | target_interior_ids[model]
        }
        for model in greenfield_targets
    }
    applicable_exclusive_member_counts = {
        model: {
            group_id: sum(
                1
                for row in rows
                if row.get("model") == model
                and row.get("family") == "exclusive_members"
                and str(row.get("values", {}).get("group_id") or "") == group_id
                and str(row.get("values", {}).get("option_id") or "")
                in target_option_ids[model]
            )
            for group_id in {
                str(row.get("values", {}).get("group_id") or "")
                for row in rows
                if row.get("model") == model
                and row.get("family") == "exclusive_members"
            }
        }
        for model in greenfield_targets
    }
    applicable_exclusive_group_ids = {
        model: {
            group_id
            for group_id, count in applicable_exclusive_member_counts[model].items()
            if count >= 2
        }
        for model in greenfield_targets
    }

    def greenfield_row_is_applicable(row: dict[str, Any]) -> bool:
        model = str(row.get("model") or "")
        family = str(row.get("family") or "")
        values = row.get("values") or {}
        option_ids = target_option_ids.get(model, set())
        reference_ids = option_ids | target_interior_ids.get(model, set())
        if family == "price_rules":
            return {
                str(values.get("condition_option_id") or ""),
                str(values.get("target_option_id") or ""),
            } <= reference_ids
        if family == "rule_mapping":
            return {
                str(values.get("source_id") or ""),
                str(values.get("target_id") or ""),
            } <= reference_ids
        if family == "rule_group_members":
            return (
                str(values.get("target_id") or "") in reference_ids
                and str(values.get("group_id") or "")
                in applicable_rule_group_ids.get(model, set())
            )
        if family == "exclusive_members":
            return (
                str(values.get("option_id") or "") in option_ids
                and str(values.get("group_id") or "")
                in applicable_exclusive_group_ids.get(model, set())
            )
        if family == "rule_groups":
            return str(values.get("group_id") or "") in applicable_rule_group_ids.get(
                model, set()
            )
        if family == "exclusive_groups":
            return str(values.get("group_id") or "") in applicable_exclusive_group_ids.get(
                model, set()
            )
        return True

    projection_migrations: list[dict[str, Any]] = []
    projection_scope_exclusions: list[dict[str, Any]] = []
    migrated_rows: list[dict[str, Any]] = []
    for index, original in enumerate(rows):
        row = dict(original)
        model = str(row.get("model") or "")
        family = str(row.get("family") or "")
        if model in greenfield_targets and family == "model_workbook_sources":
            values = dict(row.get("values") or {})
            role = str(values.get("source_role") or "")
            source_family = SOURCE_ROLE_FAMILIES.get(role)
            target_sheet = greenfield_sheets.get((model, source_family or ""))
            if target_sheet and str(values.get("sheet_name") or "") != target_sheet:
                previous_sheet = str(values.get("sheet_name") or "")
                values["sheet_name"] = target_sheet
                row["values"] = values
                projection_migrations.append(
                    {
                        "manifestIndex": index,
                        "model": model,
                        "family": family,
                        "sourceRole": role,
                        "fromSheet": previous_sheet,
                        "toSheet": target_sheet,
                        "fromAction": str(row.get("action") or ""),
                        "toAction": str(row.get("action") or ""),
                        "reason": "greenfield_target_sheet_isolation",
                    }
                )
        elif model in greenfield_targets and family in isolated_role_by_family:
            target_sheet = greenfield_sheets[(model, family)]
            previous_sheet = str(row.get("sheet") or "")
            previous_action = str(row.get("action") or "")
            if previous_sheet != target_sheet:
                if not greenfield_row_is_applicable(row):
                    if previous_action != "noop":
                        raise ValueError(
                            "Greenfield target-sheet isolation cannot discard a non-noop "
                            f"{model}/{family} row outside the target reference domain."
                        )
                    row["projectionScopeDisposition"] = (
                        "retained_existing_noop_outside_target_domain"
                    )
                    projection_scope_exclusions.append(
                        {
                            "manifestIndex": index,
                            "model": model,
                            "family": family,
                            "sheet": previous_sheet,
                            "action": previous_action,
                            "key": dict(row.get("key") or {}),
                            "reason": "outside_greenfield_target_reference_domain",
                        }
                    )
                    migrated_rows.append(row)
                    continue
                row["sheet"] = target_sheet
                if target_sheet not in sheets and previous_action in {
                    "noop",
                    "update",
                }:
                    row["action"] = "add"
                projection_migrations.append(
                    {
                        "manifestIndex": index,
                        "model": model,
                        "family": family,
                        "fromSheet": previous_sheet,
                        "toSheet": target_sheet,
                        "fromAction": previous_action,
                        "toAction": str(row.get("action") or ""),
                        "reason": "greenfield_target_sheet_isolation",
                    }
                )
        migrated_rows.append(row)
    rows = migrated_rows
    source_sheet_by_model_family: dict[tuple[str, str], str] = {}
    source_sheet = sheets.get("model_workbook_sources") or {}
    for source_row in source_sheet.get("rows") or []:
        source_family = SOURCE_ROLE_FAMILIES.get(
            str(source_row.get("source_role") or "")
        )
        source_model = str(source_row.get("model_key") or "")
        source_name = str(source_row.get("sheet_name") or "")
        if source_model and source_family and source_name:
            source_sheet_by_model_family[(source_model, source_family)] = source_name
    for manifest_row in rows:
        if manifest_row.get("family") != "model_workbook_sources":
            continue
        source_values = manifest_row.get("values") or {}
        source_model = str(source_values.get("model_key") or "")
        source_family = SOURCE_ROLE_FAMILIES.get(
            str(source_values.get("source_role") or "")
        )
        source_name = str(source_values.get("sheet_name") or "")
        if not (source_model and source_family):
            continue
        if manifest_row.get("action") == "delete":
            source_sheet_by_model_family.pop((source_model, source_family), None)
        elif source_name:
            source_sheet_by_model_family[(source_model, source_family)] = source_name
    seen_keys: set[tuple[str, str]] = set()
    sheet_families: dict[str, str] = {}
    checked_rows: list[dict[str, Any]] = []

    for index, row in enumerate(rows):
        model = str(row.get("model") or "")
        family = str(row.get("family") or "")
        sheet = str(row.get("sheet") or "")
        action = str(row.get("action") or "")
        status = str(row.get("status") or "")
        key = dict(row.get("key") or {})
        values = dict(row.get("values") or {})
        if model not in targets and model != "*":
            raise ValueError(f"Manifest row {index} has unselected model {model!r}.")
        if status != "ready":
            raise ValueError(f"Manifest row {index} is not ready.")
        if family not in EDITOR_SHEET_META:
            raise ValueError(f"Manifest row {index} has unknown family {family!r}.")
        if action not in MANIFEST_ACTION_ORDER:
            raise ValueError(f"Manifest row {index} has unsupported action {action!r}.")
        if not sheet:
            raise ValueError(f"Manifest row {index} has no physical sheet.")
        prior_family = sheet_families.setdefault(sheet, family)
        if prior_family != family:
            raise ValueError(f"Manifest sheet {sheet} maps to multiple families.")
        expected_source_sheet = source_sheet_by_model_family.get((model, family))
        if (
            expected_source_sheet
            and sheet != expected_source_sheet
            and row.get("projectionScopeDisposition")
            != "retained_existing_noop_outside_target_domain"
        ):
            raise ValueError(
                f"Manifest row {index} routes {model!r}/{family!r} to {sheet!r}; "
                f"model_workbook_sources requires {expected_source_sheet!r}."
            )
        key_columns = list(EDITOR_SHEET_META[family]["key"])
        if sorted(key) != sorted(key_columns) or any(
            not str(value or "").strip() for value in key.values()
        ):
            raise ValueError(
                f"Manifest row {index} key must be exactly nonblank {key_columns}."
            )
        physical_key = (sheet, _manifest_key_text(key))
        if physical_key in seen_keys:
            raise ValueError(f"Manifest maps physical key more than once: {physical_key!r}.")
        seen_keys.add(physical_key)
        for column in key_columns:
            if values.get(column) != key[column]:
                raise ValueError(
                    f"Manifest row {index} values do not preserve key column {column}."
                )

        sheet_data = sheets.get(sheet)
        if sheet_data is None:
            if action != "add":
                raise ValueError(
                    f"Manifest row {index} cannot {action} absent sheet {sheet}."
                )
            headers = list(values)
        else:
            headers = list(sheet_data.get("headers") or [])
            if registered_family.get(sheet) not in {None, family}:
                raise ValueError(f"Manifest family disagrees with workbook registry for {sheet}.")
        if set(values) != set(headers):
            raise ValueError(
                f"Manifest row {index} values do not match the exact {sheet} header vector."
            )

        existing = None
        if sheet_data is not None:
            matching = [
                candidate
                for candidate in sheet_data.get("rows") or []
                if all(
                    str(candidate.get(column) or "").strip()
                    == str(key[column] or "").strip()
                    for column in key_columns
                )
            ]
            if len(matching) > 1:
                raise ValueError(f"Workbook key is ambiguous before projection: {physical_key!r}.")
            existing = matching[0] if matching else None
        if action == "add" and existing is not None:
            raise ValueError(f"Manifest add already exists in workbook: {physical_key!r}.")
        if action in {"update", "delete", "noop"} and existing is None:
            raise ValueError(f"Manifest {action} is missing in workbook: {physical_key!r}.")
        if action == "noop" and _manifest_normalized_row(
            family, existing or {}
        ) != _manifest_normalized_row(family, values):
            raise ValueError(f"Manifest noop does not equal workbook state: {physical_key!r}.")
        typed_normalization_columns = [
            column
            for column, kind in EDITOR_SHEET_META[family].get("types", {}).items()
            if existing is not None
            and (
                (
                    kind == "bool"
                    and isinstance(values.get(column), bool)
                    and not isinstance(existing.get(column), bool)
                )
                or (
                    kind == "int"
                    and isinstance(values.get(column), int)
                    and not isinstance(values.get(column), bool)
                    and (
                        not isinstance(existing.get(column), int)
                        or isinstance(existing.get(column), bool)
                    )
                )
            )
        ]
        checked_rows.append(
            {
                **row,
                "model": model,
                "family": family,
                "sheet": sheet,
                "action": action,
                "key": key,
                "values": values,
                "projectionPhysicalAction": (
                    "update"
                    if action == "noop" and typed_normalization_columns
                    else action
                ),
                "typedNormalizationColumns": typed_normalization_columns,
            }
        )

    missing_sheets = sorted(sheet for sheet in sheet_families if sheet not in sheets)
    creates: list[dict[str, Any]] = []
    for sheet in missing_sheets:
        family = sheet_families[sheet]
        sheet_rows = [row for row in checked_rows if row["sheet"] == sheet]
        expected_headers = set(sheet_rows[0]["values"])
        target = str(sheet_rows[0]["model"])
        comparator = comparators[target]
        candidates = sorted(
            entry["sheet"]
            for entry in registry.get(comparator, [])
            if entry.get("family") == family
            and set((sheets.get(entry["sheet"]) or {}).get("headers") or [])
            == expected_headers
        )
        if len(candidates) != 1:
            raise ValueError(
                f"Missing sheet {sheet} needs one exact comparator header template; found {candidates}."
            )
        creates.append(
            {
                "action": "create_sheet",
                "sheet": sheet,
                "family": family,
                "headersFrom": candidates[0],
                "_manifestRefs": [],
                "_scaffoldRule": "canonical_manifest_missing_sheet",
            }
        )

    planned_sheets = set(sheets) | {item["sheet"] for item in creates}
    for migration in projection_migrations:
        if migration.get("family") != "model_workbook_sources":
            continue
        target_sheet = str(migration.get("toSheet") or "")
        if not target_sheet or target_sheet in planned_sheets:
            continue
        source_sheet_name = str(migration.get("fromSheet") or "")
        source_role = str(migration.get("sourceRole") or "")
        source_family = SOURCE_ROLE_FAMILIES.get(source_role)
        model = str(migration.get("model") or "")
        key_columns = list(EDITOR_SHEET_META.get(source_family or "", {}).get("key") or [])
        if source_sheet_name not in sheets and source_family:
            comparator = comparators.get(model, "")
            candidates = sorted(
                entry["sheet"]
                for entry in registry.get(comparator, [])
                if entry.get("family") == source_family
                and set(key_columns).issubset(
                    set((sheets.get(entry["sheet"]) or {}).get("headers") or [])
                )
            )
            if len(candidates) == 1:
                source_sheet_name = candidates[0]
        template = sheets.get(source_sheet_name) or {}
        if not source_family or not template.get("headers"):
            raise ValueError(
                f"Greenfield source registration {source_role!r} cannot create "
                f"{target_sheet!r} from {source_sheet_name!r}."
            )
        missing_keys = [
            key
            for key in EDITOR_SHEET_META[source_family]["key"]
            if key not in template["headers"]
        ]
        if missing_keys:
            raise ValueError(
                f"Greenfield source template {source_sheet_name!r} lacks key "
                f"column(s) {missing_keys}."
            )
        creates.append(
            {
                "action": "create_sheet",
                "sheet": target_sheet,
                "family": source_family,
                "headersFrom": source_sheet_name,
                "_manifestRefs": [
                    f"manifest-{int(migration['manifestIndex']):05d}"
                ],
                "_scaffoldRule": "canonical_manifest_source_sheet",
            }
        )
        planned_sheets.add(target_sheet)

    for manifest_index, row in enumerate(checked_rows):
        if (
            row.get("family") != "model_workbook_sources"
            or row.get("action") == "delete"
        ):
            continue
        values = row.get("values") or {}
        target_sheet = str(values.get("sheet_name") or "")
        source_role = str(values.get("source_role") or "")
        source_family = SOURCE_ROLE_FAMILIES.get(source_role)
        model = str(values.get("model_key") or row.get("model") or "")
        if (
            not target_sheet
            or not source_family
            or target_sheet in planned_sheets
        ):
            continue
        comparator = comparators.get(model, "")
        key_columns = set(EDITOR_SHEET_META[source_family].get("key") or [])
        candidates = sorted(
            entry["sheet"]
            for entry in registry.get(comparator, [])
            if entry.get("role") == source_role
            and entry.get("family") == source_family
            and key_columns.issubset(
                set((sheets.get(entry["sheet"]) or {}).get("headers") or [])
            )
        )
        if len(candidates) != 1:
            raise ValueError(
                f"Canonical source registration {source_role!r} cannot create "
                f"{target_sheet!r}; expected one exact comparator template, "
                f"found {candidates}."
            )
        creates.append(
            {
                "action": "create_sheet",
                "sheet": target_sheet,
                "family": source_family,
                "headersFrom": candidates[0],
                "_manifestRefs": [f"manifest-{manifest_index:05d}"],
                "_scaffoldRule": "canonical_manifest_registered_source_sheet",
            }
        )
        planned_sheets.add(target_sheet)

    promotion_sheet = sheets.get("model_registry_promotion") or {}
    promotion_headers = list(promotion_sheet.get("headers") or [])
    existing_promotions = {
        str(row.get("model_key") or "")
        for row in promotion_sheet.get("rows") or []
    }
    next_promotion_order = 1 + max(
        [
            int(row.get("display_order") or 0)
            for row in promotion_sheet.get("rows") or []
        ]
        or [0]
    )
    for model in sorted(greenfield_targets):
        if model in existing_promotions:
            continue
        config = MODEL_PLAN_CONFIG[model]
        promotion_values: dict[str, Any] = {
            header: None for header in promotion_headers
        }
        promotion_values.update(
            {
                "model_key": model,
                "registry_key": config["registryKey"],
                "promoted_to_runtime": False,
                "default_model": False,
                "artifact_path": (
                    f"form-output/runtime/{config['exportSlug']}-runtime-contract.json"
                ),
                "artifact_type": "runtime_contract",
                "active": False,
                "display_order": next_promotion_order,
                "notes": "Inactive greenfield deployment-proof scaffold.",
            }
        )
        creates.append(
            {
                "action": "add",
                "sheet": "model_registry_promotion",
                "family": "model_registry_promotion",
                "key": {"model_key": model},
                "row": promotion_values,
                "_manifestRefs": [],
                "_scaffoldRule": "pass_c3_greenfield_registry_promotion",
            }
        )
        next_promotion_order += 1

    ordered_rows = sorted(
        enumerate(checked_rows),
        key=lambda pair: (
            0 if pair[1]["family"] in MANIFEST_STAGE1_FAMILIES else 1,
            pair[1]["sheet"],
            MANIFEST_ACTION_ORDER[pair[1]["action"]],
            _manifest_key_text(pair[1]["key"]),
            pair[0],
        ),
    )
    stage1 = sorted(creates, key=lambda item: item["sheet"])
    stage2: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []
    noop_receipts: list[dict[str, Any]] = []
    plan_ref = 0
    for original_index, row in ordered_rows:
        manifest_ref = f"manifest-{original_index:05d}"
        receipt = {
            "manifestRef": manifest_ref,
            "model": row["model"],
            "family": row["family"],
            "sheet": row["sheet"],
            "action": row["action"],
            "key": row["key"],
            "semanticSignature": row.get("semanticSignature"),
            "derivationVersion": row.get("derivationVersion"),
            "projectionScopeDisposition": row.get("projectionScopeDisposition"),
            "projectionPhysicalAction": row.get("projectionPhysicalAction"),
            # Preserve the compiler-authored row verbatim for bidirectional
            # temporary-runtime parity. This is evidence only; operations remain
            # the sole workbook mutation surface.
            "canonicalValues": dict(row.get("values") or {}),
            "evidenceDependencies": list(row.get("evidenceDependencies") or []),
        }
        physical_action = str(row.get("projectionPhysicalAction") or row["action"])
        if row["action"] == "noop" and physical_action == "noop":
            noop_ref = f"noop-{len(noop_receipts):05d}"
            noop_receipts.append({**receipt, "noopRef": noop_ref})
            coverage_rows.append({**receipt, "noopRef": noop_ref})
            continue
        operation: dict[str, Any] = {
            "action": physical_action,
            "sheet": row["sheet"],
            "key": row["key"],
            "_manifestRef": manifest_ref,
            "_planRef": f"op-{plan_ref:05d}",
        }
        if physical_action == "add":
            operation["row"] = row["values"]
        elif physical_action == "update":
            key_columns = set(EDITOR_SHEET_META[row["family"]]["key"])
            typed_normalization_columns = set(
                row.get("typedNormalizationColumns") or []
            )
            operation["row"] = {
                column: value
                for column, value in row["values"].items()
                if column not in key_columns
                and (
                    not typed_normalization_columns
                    or column in typed_normalization_columns
                )
            }
        destination = (
            stage1 if row["family"] in MANIFEST_STAGE1_FAMILIES else stage2
        )
        destination.append(operation)
        coverage_rows.append({**receipt, "planRef": operation["_planRef"]})
        plan_ref += 1

    for scaffold in stage1:
        if scaffold.get("action") == "create_sheet":
            derived_refs = [
                entry["manifestRef"]
                for entry in coverage_rows
                if entry["sheet"] == scaffold["sheet"]
            ]
            scaffold["_manifestRefs"] = sorted(
                set(scaffold.get("_manifestRefs") or []) | set(derived_refs)
            )
        if not scaffold.get("_planRef"):
            scaffold["_planRef"] = f"op-{plan_ref:05d}"
            plan_ref += 1

    per_sheet: dict[str, dict[str, int]] = {}
    for item in [*stage1, *stage2]:
        counts = per_sheet.setdefault(item["sheet"], {})
        counts[item["action"]] = counts.get(item["action"], 0) + 1
    for receipt in noop_receipts:
        counts = per_sheet.setdefault(receipt["sheet"], {})
        counts["noop"] = counts.get("noop", 0) + 1

    continuity_names = {
        "price_rules": "priceRules",
        "rule_groups": "ruleGroups",
        "color_overrides": "colorOverrides",
        "interior_components": "interiorComponents",
        "asset_map": "assetMap",
    }
    runtime_continuity: dict[str, dict[str, Any]] = {}
    for model in targets:
        source_ops: dict[str, dict[str, int]] = {}
        for row in checked_rows:
            if row["model"] != model or row["action"] not in {
                "add",
                "update",
                "noop",
            }:
                continue
            label = continuity_names.get(row["family"])
            if not label:
                continue
            actions = source_ops.setdefault(label, {})
            actions[row["action"]] = actions.get(row["action"], 0) + 1
        runtime_continuity[model] = {"sourceOps": source_ops}

    workbook_fingerprint = {
        "sha256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        "mtimeNs": str(workbook_path.stat().st_mtime_ns),
    }
    bound_hashes = dict(sorted(compiler_bindings.items()))
    source_feature_rows = list(compile_report.get("sourceFeatureCoverage") or [])
    source_feature_summary: dict[str, Any] = {
        "semanticSha": hashlib.sha256(
            json.dumps(
                source_feature_rows,
                sort_keys=True,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest(),
        "byModel": {},
    }
    allowed_source_dispositions = {
        "compiled",
        "resolved_not_applicable",
        "resolved_not_a_workbook_fact",
    }
    for model in targets:
        scoped_rows = [
            row
            for row in source_feature_rows
            if str(row.get("model") or "") in {model, "*"}
        ]
        disposition_counts: dict[str, int] = {}
        for row in scoped_rows:
            disposition = str(row.get("disposition") or "")
            disposition_counts[disposition] = disposition_counts.get(disposition, 0) + 1
        source_feature_summary["byModel"][model] = {
            "featureCount": len(scoped_rows),
            "dispositionCounts": dict(sorted(disposition_counts.items())),
            "blockingFeatures": [
                {
                    "featureId": str(row.get("featureId") or ""),
                    "family": str(row.get("family") or ""),
                    "disposition": str(row.get("disposition") or ""),
                }
                for row in scoped_rows
                if str(row.get("disposition") or "")
                not in allowed_source_dispositions
            ][:50],
        }
    return {
        "schemaVersion": "pass-c-3",
        "targets": targets,
        "targetModes": modes,
        "sourceFingerprint": selection.get("sourceFingerprint"),
        "candidatesFingerprint": selection.get("candidatesFingerprint"),
        "decisionsFingerprint": compiler_bindings.get("exceptionResolutionsSha"),
        "canonicalManifestSemanticSha": manifest.get("manifestSemanticSha"),
        "runAuthorityFingerprint": manifest.get("runAuthorityFingerprint"),
        "compilerBindings": bound_hashes,
        "sourceFeatureCoverage": source_feature_summary,
        **bound_hashes,
        "workbookFingerprint": workbook_fingerprint,
        "projectionMigrations": {
            "policy": "greenfield_target_sheet_isolation",
            "rowCount": len(projection_migrations),
            "rows": projection_migrations,
            "scopeExclusionCount": len(projection_scope_exclusions),
            "scopeExclusions": projection_scope_exclusions,
        },
        "stage1": {"items": stage1},
        "stage2": {"items": stage2},
        "coverage": {
            "manifestRows": coverage_rows,
            "noops": noop_receipts,
            "uncoveredManifestRows": [],
            "uncoveredApprovedDecisions": [],
        },
        "planReadiness": {
            "planReady": True,
            "manifestRowCount": len(coverage_rows),
            "mutationCount": len(stage1) + len(stage2),
            "noopCount": len(noop_receipts),
            "uncoveredCount": 0,
        },
        "report": {
            "perSheetCounts": per_sheet,
            "perSheetActionCounts": per_sheet,
            "runtimeContinuity": runtime_continuity,
            "clearedRows": {},
            "holds": [],
            "deferrals": [],
            "unreviewedSplits": {},
            "gaps": [],
            "blockingGaps": [],
        },
        "valid": True,
    }


def build_plan(
    *,
    workbook_path: Path,
    selection: dict[str, Any],
    candidates: list[dict[str, Any]],
    decisions: dict[str, dict[str, Any]],
    candidates_fingerprint: str,
) -> dict[str, Any]:
    targets = [t for t in selection["targets"] if t in MODEL_PLAN_CONFIG]
    unknown_targets = [t for t in selection["targets"] if t not in MODEL_PLAN_CONFIG]
    context = _workbook_context(workbook_path, targets)
    sheetnames = set(context["sheetnames"])

    stage1: list[dict[str, Any]] = []
    stage2: list[dict[str, Any]] = []
    coverage: dict[str, list[str]] = {record["decisionId"]: [] for record in decisions.values()}
    gaps: list[dict[str, Any]] = [
        {"model": t, "kind": "unknown_target", "detail": f"no plan config for {t}", "decisionId": ""}
        for t in unknown_targets
    ]
    holds: list[dict[str, Any]] = []
    deferrals: list[dict[str, Any]] = []
    unreviewed_splits: dict[str, list[str]] = {}
    cleared_rows: dict[str, int] = {}
    # Rows whose section decision is skip/hold produce no ops on purpose;
    # their other approved lane decisions must not count as uncovered.
    inert_candidates: set[tuple[str, str]] = set()
    op_index = 0

    def op(stage: list, item: dict[str, Any], decision_ids: list[str], rule: str = "") -> None:
        nonlocal op_index
        item = dict(item)
        item["_planRef"] = f"op-{op_index:05d}"
        item["_decisions"] = sorted(decision_ids)
        if rule:
            item["_scaffoldRule"] = rule
        op_index += 1
        stage.append(item)
        for decision_id in decision_ids:
            coverage.setdefault(decision_id, []).append(item["_planRef"])

    existing_model_keys = {row.get("model_key") for row in context["model_master"]}
    existing_promotion = {row.get("model_key") for row in context["model_registry_promotion"]}
    next_display_order = 1 + max(
        [int(row.get("display_order") or 0) for row in context["model_registry_promotion"]] or [0]
    )
    source_index = {
        (row.get("model_key"), row.get("source_role")): row for row in context["model_workbook_sources"]
    }

    for model in targets:
        config = MODEL_PLAN_CONFIG[model]
        model_decisions = {k: d for k, d in decisions.items() if d["model"] == model}
        scoped = [c for c in scope_candidates(candidates, model)]
        orderable = [c for c in scoped if c["rowKind"] == "orderable"]
        section_candidates = [c for c in scoped if candidate_needs_section_decision(c)]
        by_candidate_lane: dict[tuple[str, str], dict[str, Any]] = {}
        for record in model_decisions.values():
            if record.get("candidateId"):
                by_candidate_lane[(record["lane"], record["candidateId"])] = record
            if record["resolution"] == "hold_for_question":
                holds.append(
                    {"model": model, "decisionId": record["decisionId"], "note": record.get("reviewerNote", "")}
                )
            if record["lane"] == "interior_media_deferral":
                deferrals.append(
                    {"model": model, "groupKey": record.get("groupKey", ""), "payload": record.get("payload", {})}
                )
                kind = (record.get("payload") or {}).get("kind")
                group_key = record.get("groupKey", "")
                if kind == "color" or "color-overrides" in group_key:
                    gaps.append(
                        {"model": model, "kind": "color_overrides_deferred", "detail": group_key, "decisionId": record["decisionId"]}
                    )
                elif kind == "asset" or "asset-images" in group_key:
                    gaps.append(
                        {"model": model, "kind": "asset_map_deferred", "detail": group_key, "decisionId": record["decisionId"]}
                    )
                elif kind == "component" or "components" in group_key:
                    gaps.append(
                        {"model": model, "kind": "interior_components_deferred", "detail": group_key, "decisionId": record["decisionId"]}
                    )

        variant_map = _variant_map(context, model)
        options_sheet = config["sheetPrefix"] + "options"
        ovs_sheet = config["sheetPrefix"] + "ovs"
        existing_options = context["existing_rows"].get(options_sheet, [])
        existing_option_by_rpo = {
            _rpo_token(row.get("rpo")): row
            for row in existing_options
            if _rpo_token(row.get("rpo")) and row.get("option_id")
        }
        approved_planned_rpos = {
            _rpo_token(candidate["rpo"] or candidate["refOnlyRpo"])
            for candidate in section_candidates
            if (
                by_candidate_lane.get(("section", candidate["candidateId"]))
                and by_candidate_lane[("section", candidate["candidateId"])]["resolution"] == "approved_for_plan"
            )
        }
        relationship_rpos = _planned_relationship_rpos(model_decisions)
        retained_existing_option_ids = {
            _clean(existing_option_by_rpo[rpo].get("option_id"))
            for rpo in relationship_rpos - approved_planned_rpos
            if rpo in existing_option_by_rpo
        }
        if orderable and not variant_map:
            # Options without any OVS rows are unusable — fail at plan time.
            gaps.append(
                {
                    "model": model,
                    "kind": "no_variants_mapped",
                    "detail": "no variant_master/model_variants rows resolve for this model",
                    "decisionId": "",
                }
            )
        model_code_suffixes = sorted({(s.get("modelCode") or "")[-2:] for c in scoped for s in c["statuses"]})

        # ---------------------------------------------------------- stage 1
        for role, suffix in MODEL_SHEET_ROLES:
            sheet = config["sheetPrefix"] + suffix
            template = config["templatePrefix"] + ("exclusive_members" if suffix == "exclusive_members" else suffix)
            family = ROLE_SHEET_SUFFIX[role]
            family = {
                "options": "options",
                "ovs": "ovs",
                "rule_mapping": "rule_mapping",
                "price_rules": "price_rules",
                "rule_groups": "rule_groups",
                "rule_group_members": "rule_group_members",
                "exclusive_groups": "exclusive_groups",
                "exclusive_members": "exclusive_members",
                "variant_overrides": "variant_overrides",
            }[family]
            if sheet not in sheetnames:
                if template not in sheetnames:
                    gaps.append(
                        {"model": model, "kind": "missing_template", "detail": f"{template} for {sheet}", "decisionId": ""}
                    )
                    continue
                op(
                    stage1,
                    {"action": "create_sheet", "sheet": sheet, "family": family, "headersFrom": template},
                    [],
                    rule=f"{model}:create_missing_model_sheet",
                )
            source_row = source_index.get((model, role))
            if source_row is None:
                op(
                    stage1,
                    {
                        "action": "add",
                        "sheet": "model_workbook_sources",
                        "key": {"model_key": model, "source_role": role},
                        "row": {"model_key": model, "source_role": role, "sheet_name": sheet, "active": True},
                    },
                    [],
                    rule=f"{model}:register_source_role",
                )
            elif not workbook_truthy(source_row.get("active")):
                op(
                    stage1,
                    {
                        "action": "update",
                        "sheet": "model_workbook_sources",
                        "key": {"model_key": model, "source_role": role},
                        "row": {"active": True},
                    },
                    [],
                    rule=f"{model}:activate_source_role",
                )
        for role, sheet in (("color_overrides_sheet", "color_overrides"), ("interior_source_sheet", config["interiorSheet"])):
            source_row = source_index.get((model, role))
            if source_row is None:
                op(
                    stage1,
                    {
                        "action": "add",
                        "sheet": "model_workbook_sources",
                        "key": {"model_key": model, "source_role": role},
                        "row": {"model_key": model, "source_role": role, "sheet_name": sheet, "active": True},
                    },
                    [],
                    rule=f"{model}:register_source_role",
                )
            elif not workbook_truthy(source_row.get("active")):
                op(
                    stage1,
                    {
                        "action": "update",
                        "sheet": "model_workbook_sources",
                        "key": {"model_key": model, "source_role": role},
                        "row": {"active": True},
                    },
                    [],
                    rule=f"{model}:activate_source_role",
                )

        if model not in existing_model_keys:
            op(
                stage1,
                {
                    "action": "add",
                    "sheet": "model_master",
                    "key": {"model_key": model},
                    "row": {
                        "model_key": model,
                        "registry_key": config["registryKey"],
                        "model_label": config["label"],
                        "model_year": "2027",
                        "dataset_name": f"2027 Corvette {config['label']} dataset",
                        "export_slug": config["exportSlug"],
                        "expected_variant_count": len(variant_map),
                        "default_model": False,
                        "active": False,
                        "notes": "Created by ingest apply plan; activation happens at promotion.",
                    },
                },
                [],
                rule=f"{model}:model_master_row",
            )
        existing_model_variant_ids = {
            row.get("variant_id") for row in context["model_variants"] if row.get("model_key") == model
        }
        for order, ((trim, body), vid) in enumerate(sorted(variant_map.items(), key=lambda kv: kv[1]), start=1):
            if vid not in existing_model_variant_ids:
                op(
                    stage1,
                    {
                        "action": "add",
                        "sheet": "model_variants",
                        "key": {"model_key": model, "variant_id": vid},
                        "row": {"model_key": model, "variant_id": vid, "display_order": order, "active": False},
                    },
                    [],
                    rule=f"{model}:model_variant_row",
                )
        if model not in existing_promotion:
            op(
                stage1,
                {
                    "action": "add",
                    "sheet": "model_registry_promotion",
                    "key": {"model_key": model},
                    "row": {
                        "model_key": model,
                        "registry_key": config["registryKey"],
                        "promoted_to_runtime": False,
                        "default_model": False,
                        "active": False,
                        "display_order": next_display_order,
                        "notes": "Unpromoted until Pass F.",
                    },
                },
                [],
                rule=f"{model}:promotion_row",
            )
            next_display_order += 1

        # ---------------------------------------------------------- stage 2
        # Clean reprocess: existing scaffold data rows are replaced wholesale.
        for _, suffix in MODEL_SHEET_ROLES:
            sheet = config["sheetPrefix"] + suffix
            rows = context["existing_rows"].get(sheet) or []
            if not rows:
                continue
            keycols = {
                "options": ("option_id",),
                "ovs": ("option_id", "variant_id"),
                "rule_mapping": ("rule_id",),
                "price_rules": ("price_rule_id",),
                "rule_groups": ("group_id",),
                "rule_group_members": ("group_id", "target_id"),
                "exclusive_groups": ("group_id",),
                "exclusive_members": ("group_id", "option_id"),
                "variant_overrides": ("option_id", "variant_id"),
            }[suffix]
            cleared = 0
            for row in rows:
                key = {k: row.get(k, "") for k in keycols}
                if not all(str(v).strip() for v in key.values()):
                    continue
                if suffix == "options" and _clean(row.get("option_id")) in retained_existing_option_ids:
                    continue
                if suffix == "ovs" and _clean(row.get("option_id")) in retained_existing_option_ids:
                    continue
                op(
                    stage2,
                    {"action": "delete", "sheet": sheet, "key": key},
                    [],
                    rule=f"{model}:clear_scaffold_rows",
                )
                cleared += 1
            if cleared:
                cleared_rows[sheet] = cleared

        # Duplicate handling: same_option groups collapse to the first
        # occurrence; distinct_by_context keeps every occurrence.
        duplicate_decision_by_rpo = {
            record.get("groupKey"): record
            for record in model_decisions.values()
            if record["lane"] == "duplicate" and record.get("groupKey")
        }
        option_id_by_rpo: dict[str, str] = {}
        for rpo, row in existing_option_by_rpo.items():
            option_id = _clean(row.get("option_id"))
            if option_id and option_id in retained_existing_option_ids:
                option_id_by_rpo[rpo] = option_id
        existing_default_rule_ids = {
            _clean(row.get("rule_id"))
            for row in context["global_rows"].get("default_selection_rules", [])
            if row.get("model_key") == model and row.get("rule_id")
        }
        used_default_rule_ids = set(existing_default_rule_ids)
        # Seed with ids already on the scaffold sheets: editor_ops rejects an
        # add whose key still exists at validation time, even with a delete of
        # that key in the same batch — new rows must take fresh ids.
        used_option_ids: set[str] = {
            row.get("option_id", "")
            for row in context["existing_rows"].get(options_sheet, [])
            if row.get("option_id")
        }
        planned_rpos: set[str] = set()
        primary_ref_by_rpo: dict[str, str] = {}

        def option_id_for(rpo: str) -> str:
            base = f"opt_{slug_rpo(rpo)}"
            n = 1
            while f"{base}_{n:03d}" in used_option_ids:
                n += 1
            oid = f"{base}_{n:03d}"
            used_option_ids.add(oid)
            return oid

        def resolve_option_id_by_rpo(rpo: Any, record: dict[str, Any], role: str) -> str | None:
            token = _rpo_token(rpo)
            if not token:
                gaps.append(
                    {
                        "model": model,
                        "kind": "relationship_option_identity_unresolved",
                        "detail": f"{record['groupKey']}: blank {role} RPO",
                        "decisionId": record["decisionId"],
                    }
                )
                return None
            option_id = option_id_by_rpo.get(token)
            if option_id:
                return option_id
            detail = f"{record['groupKey']}: {role} RPO {token} is not in planned options"
            if token in existing_option_by_rpo:
                detail += " and its existing row is not retained in this plan"
            gaps.append(
                {
                    "model": model,
                    "kind": "relationship_option_identity_unresolved",
                    "detail": detail,
                    "decisionId": record["decisionId"],
                }
            )
            return None

        def default_rule_id_for(rpo: str) -> str:
            base = f"default_{slug_rpo(rpo)}"
            candidate = base
            n = 1
            while candidate in used_default_rule_ids:
                n += 1
                candidate = f"{base}_{n:03d}"
            used_default_rule_ids.add(candidate)
            return candidate

        used_display_orders: set[tuple[str, int]] = set()
        next_display_order_by_section: dict[str, int] = {}
        for existing in existing_options:
            section_id = _clean(existing.get("section_id"))
            order = _intish(existing.get("display_order"), 0)
            if section_id and order:
                used_display_orders.add((section_id, order))
                next_display_order_by_section[section_id] = max(
                    next_display_order_by_section.get(section_id, 0), order
                )

        def allocate_display_order(section_id: str, requested: Any = None) -> int:
            if requested not in (None, ""):
                return _intish(requested, 0)
            current = next_display_order_by_section.get(section_id, 0)
            return ((current // 10) + 1) * 10

        def reserve_display_order(section_id: str, order: int, decision_id: str) -> None:
            if not section_id or not order:
                gaps.append(
                    {
                        "model": model,
                        "kind": "option_display_order_missing",
                        "detail": f"section={section_id!r} order={order!r}",
                        "decisionId": decision_id,
                    }
                )
                return
            if (section_id, order) in used_display_orders:
                gaps.append(
                    {
                        "model": model,
                        "kind": "option_display_order_collision",
                        "detail": f"{options_sheet}: {section_id} display_order {order}",
                        "decisionId": decision_id,
                    }
                )
                return
            used_display_orders.add((section_id, order))
            next_display_order_by_section[section_id] = max(next_display_order_by_section.get(section_id, 0), order)

        for candidate in section_candidates:
            rpo = candidate["rpo"] or candidate["refOnlyRpo"]
            duplicate = duplicate_decision_by_rpo.get(rpo)
            if rpo in planned_rpos and (
                duplicate is None or (duplicate.get("payload") or {}).get("classification") == "same_option"
            ):
                # Collapsed into the first occurrence — this occurrence's
                # decisions are covered by the primary option op.
                primary_ref = primary_ref_by_rpo.get(rpo)
                if primary_ref:
                    for lane in ("section", "price", "copy_split", "status_nuance"):
                        record = by_candidate_lane.get((lane, candidate["candidateId"]))
                        if record:
                            coverage.setdefault(record["decisionId"], []).append(primary_ref)
                continue
            section_decision = by_candidate_lane.get(("section", candidate["candidateId"]))
            if section_decision is None:
                standard_record = by_candidate_lane.get(("standard_equipment", candidate["candidateId"]))
                if (
                    candidate["rowKind"] == "ref_only"
                    and standard_record
                    and standard_record["resolution"] == "approved_for_plan"
                    and standard_record["action"] == "include_standard_equipment"
                ):
                    # Legacy Pass B/C decisions included ref-only rows through
                    # Standard equipment before Section assignment owned them.
                    # Keep those plans valid; new reviews should prefer a
                    # section decision so selectable/section placement is known.
                    continue
                gaps.append(
                    {
                        "model": model,
                        "kind": "missing_mandatory_decision",
                        "detail": f"{rpo or candidate['candidateId']} lacks a section decision",
                        "decisionId": "",
                    }
                )
                continue
            if section_decision["resolution"] != "approved_for_plan":
                inert_candidates.add((model, candidate["candidateId"]))
                continue  # held/skipped rows stay out of the plan (no price owed); holds are reported
            price_decision = by_candidate_lane.get(("price", candidate["candidateId"]))
            if candidate["rowKind"] == "orderable" and price_decision is None:
                gaps.append(
                    {
                        "model": model,
                        "kind": "missing_mandatory_decision",
                        "detail": f"{rpo or candidate['candidateId']} lacks a price decision",
                        "decisionId": "",
                    }
                )
                continue
            split_decision = by_candidate_lane.get(("copy_split", candidate["candidateId"]))
            if split_decision and split_decision["resolution"] == "approved_for_plan":
                split = split_decision.get("payload") or {}
                split_ids = [split_decision["decisionId"]]
            else:
                split = propose_copy_split(candidate)
                split_ids = []
                unreviewed_splits.setdefault(model, []).append(rpo)
            price_value = None
            price_action = price_decision.get("action") if price_decision else ""
            price_payload = price_decision.get("payload") if price_decision else {}
            price_payload = price_payload or {}
            if price_decision and price_decision["resolution"] == "approved_for_plan":
                if price_action == "accept_exact_price":
                    price_value = candidate.get("listPrice")
                    if price_value is None and len(candidate.get("priceRows") or []) == 1:
                        price_value = candidate["priceRows"][0].get("listPrice")
                elif price_action == "choose_price_row":
                    index = price_payload.get("priceRowIndex")
                    rows = candidate.get("priceRows") or []
                    price_value = rows[int(index)].get("listPrice") if index is not None and int(index) < len(rows) else None
                elif price_action == "set_reviewed_price":
                    price_value = price_payload.get("reviewedPrice")
                elif price_action in ("confirm_no_price", "defer_price_extractor"):
                    price_value = None
                    if price_action == "defer_price_extractor":
                        gaps.append(
                            {
                                "model": model,
                                "kind": "price_deferred",
                                "detail": rpo,
                                "decisionId": price_decision["decisionId"],
                            }
                        )
            oid = option_id_for(rpo)
            option_id_by_rpo.setdefault(rpo, oid)
            planned_rpos.add(rpo)
            decision_ids = [section_decision["decisionId"], *split_ids]
            if price_decision:
                decision_ids.append(price_decision["decisionId"])
            if duplicate:
                decision_ids.append(duplicate["decisionId"])
            section_payload = section_decision.get("payload") or {}
            section_id = _clean(section_payload.get("sectionId"))
            display_order = allocate_display_order(section_id, section_payload.get("displayOrder"))
            reserve_display_order(section_id, display_order, section_decision["decisionId"])
            row = {
                "option_id": oid,
                "rpo": rpo,
                "price": int(price_value) if price_value is not None else None,
                "option_name": split.get("name") or candidate["description"].split("\n")[0],
                "description": split.get("description") or None,
                "detail_raw": split.get("detailRaw") or candidate["description"],
                "section_id": section_id,
                "display_order": display_order,
                # Reviewer-set flags (field note 7): display-only rows carry
                # selectable=False; inactive rows are written but hidden.
                "active": section_payload.get("active", True),
            }
            if section_payload.get("selectable") is False:
                row["selectable"] = False
            op(stage2, {"action": "add", "sheet": options_sheet, "key": {"option_id": oid}, "row": row}, decision_ids)
            primary_ref_by_rpo[rpo] = stage2[-1]["_planRef"]
            status_ids = []
            status_decision = by_candidate_lane.get(("status_nuance", candidate["candidateId"]))
            if status_decision:
                status_ids.append(status_decision["decisionId"])
                if status_decision.get("action") == "mark_unresolved_blocked":
                    gaps.append(
                        {
                            "model": model,
                            "kind": "status_blocked",
                            "detail": rpo,
                            "decisionId": status_decision["decisionId"],
                        }
                    )
            for (trim, body), vid in sorted(variant_map.items(), key=lambda kv: kv[1]):
                status_value = _status_for_variant(candidate, body, trim) or "unavailable"
                op(
                    stage2,
                    {
                        "action": "add",
                        "sheet": ovs_sheet,
                        "key": {"option_id": oid, "variant_id": vid},
                        "row": {"option_id": oid, "variant_id": vid, "status": status_value},
                    },
                    decision_ids + status_ids,
                )

        # Standard-equipment inclusions: ref-only rows the reviewer included
        # become standard, non-orderable option rows (runtime derives the
        # standard-equipment list from status == standard).
        for candidate in scoped:
            if candidate["rowKind"] != "ref_only":
                continue
            record = by_candidate_lane.get(("standard_equipment", candidate["candidateId"]))
            if not record or record["resolution"] != "approved_for_plan" or record["action"] != "include_standard_equipment":
                continue
            section_record = by_candidate_lane.get(("section", candidate["candidateId"]))
            if section_record and section_record["resolution"] != "approved_for_plan":
                continue
            rpo = candidate["refOnlyRpo"]
            if rpo in planned_rpos:
                continue
            split = propose_copy_split(candidate)
            oid = option_id_for(rpo)
            planned_rpos.add(rpo)
            op(
                stage2,
                {
                    "action": "add",
                    "sheet": options_sheet,
                    "key": {"option_id": oid},
                    "row": {
                        "option_id": oid,
                        "rpo": rpo,
                        "price": None,
                        "option_name": split.get("name") or candidate["description"].split("\n")[0],
                        "description": split.get("description") or None,
                        "detail_raw": split.get("detailRaw"),
                        "section_id": None,
                        "selectable": False,
                        "active": True,
                    },
                },
                [record["decisionId"]],
            )
            for (trim, body), vid in sorted(variant_map.items(), key=lambda kv: kv[1]):
                status_value = _status_for_variant(candidate, body, trim) or "standard"
                op(
                    stage2,
                    {
                        "action": "add",
                        "sheet": ovs_sheet,
                        "key": {"option_id": oid, "variant_id": vid},
                        "row": {"option_id": oid, "variant_id": vid, "status": status_value},
                    },
                    [record["decisionId"]],
                )

        # Global model interior scope rows are keyed rows, not clean-reprocess
        # model sheets. Compare existing rows and emit only add/update/no-op.
        interior_sheet = config["interiorSheet"]
        interior_rows = context["global_rows"].get(interior_sheet, [])
        scope_rows = context["global_rows"].get("model_interior_scope", [])
        if not interior_rows:
            gaps.append(
                {
                    "model": model,
                    "kind": "model_interior_scope_missing_required",
                    "detail": f"{interior_sheet} has no rows available for scope planning",
                    "decisionId": "",
                }
            )
        else:
            trims = {trim.upper() for trim, _body in variant_map}
            existing_scope_by_key = {
                (_clean(row.get("model_key")), _clean(row.get("interior_id")), _clean(row.get("trim_level"))): row
                for row in scope_rows
                if row.get("model_key") and row.get("interior_id") and row.get("trim_level")
            }
            template_scope_by_key = {
                (_clean(row.get("interior_id")), _clean(row.get("trim_level"))): row
                for row in scope_rows
                if row.get("interior_id") and row.get("trim_level")
            }
            for interior in interior_rows:
                interior_id = _clean(interior.get("interior_id"))
                trim_level = _clean(interior.get("Trim") or interior.get("trim_level"))
                if not interior_id or not trim_level or (trims and trim_level.upper() not in trims):
                    continue
                template = template_scope_by_key.get((interior_id, trim_level))
                if not template:
                    gaps.append(
                        {
                            "model": model,
                            "kind": "model_interior_scope_missing_required",
                            "detail": f"{interior_sheet}: no model_interior_scope template for {interior_id} / {trim_level}",
                            "decisionId": "",
                        }
                    )
                    continue
                desired = dict(template)
                desired.update(
                    {
                        "model_key": model,
                        "interior_id": interior_id,
                        "trim_level": trim_level,
                        "active": True,
                        "requires_option_id": interior.get("included_option_id") or template.get("requires_option_id"),
                        "notes": template.get("notes") or "Workbook-owned interior trim scope metadata.",
                    }
                )
                key = {"model_key": model, "interior_id": interior_id, "trim_level": trim_level}
                existing = existing_scope_by_key.get((model, interior_id, trim_level))
                if existing is None:
                    op(stage2, {"action": "add", "sheet": "model_interior_scope", "key": key, "row": desired}, [])
                    continue
                changed = {
                    column: value
                    for column, value in desired.items()
                    if column not in key and _norm(existing.get(column)) != _norm(value)
                }
                if changed:
                    op(stage2, {"action": "update", "sheet": "model_interior_scope", "key": key, "row": changed}, [])

        # Relationships -> rule_mapping rows.
        rule_sheet = config["sheetPrefix"] + "rule_mapping"
        rule_counter = 1
        for record in sorted(model_decisions.values(), key=lambda r: r["decisionId"]):
            if record["lane"] != "relationship" or record["resolution"] != "approved_for_plan":
                continue
            if record.get("action") != "create_relationship_candidate":
                continue  # needs_product_decision / reconciliation notes are report items
            payload = record.get("payload") or {}
            rule_type = RELATIONSHIP_KIND_TO_RULE_TYPE.get(payload.get("kind", ""))
            source_oid = resolve_option_id_by_rpo(payload.get("sourceRpo"), record, "source")
            if rule_type is None or source_oid is None:
                gaps.append(
                    {
                        "model": model,
                        "kind": "relationship_unmappable",
                        "detail": f"{record['groupKey']}: kind={payload.get('kind')} source={payload.get('sourceRpo')}",
                        "decisionId": record["decisionId"],
                    }
                )
                continue
            for target_rpo in payload.get("targetRpos") or []:
                target_oid = resolve_option_id_by_rpo(target_rpo, record, "target")
                if target_oid is None:
                    continue
                rid = f"rule_{slug_rpo(payload.get('sourceRpo', ''))}_{rule_counter:03d}"
                rule_counter += 1
                op(
                    stage2,
                    {
                        "action": "add",
                        "sheet": rule_sheet,
                        "key": {"rule_id": rid},
                        "row": {
                            "rule_id": rid,
                            "rule_type": rule_type,
                            "source_id": source_oid,
                            "target_id": target_oid,
                        },
                    },
                    [record["decisionId"]],
                )

        # Exclusive groups.
        excl_groups_sheet = config["sheetPrefix"] + "exclusive_groups"
        excl_members_sheet = config["sheetPrefix"] + "exclusive_members"
        for record in sorted(model_decisions.values(), key=lambda r: r["decisionId"]):
            if record["lane"] != "exclusive_group" or record["resolution"] != "approved_for_plan":
                continue
            payload = record.get("payload") or {}
            members = [resolve_option_id_by_rpo(m, record, "exclusive member") for m in payload.get("members") or []]
            missing = [m for m, oid in zip(payload.get("members") or [], members) if oid is None]
            members = [m for m in members if m]
            if missing or len(members) < 2:
                gaps.append(
                    {
                        "model": model,
                        "kind": "exclusive_members_missing",
                        "detail": f"{record['groupKey']}: unplanned members {missing}",
                        "decisionId": record["decisionId"],
                    }
                )
                continue
            existing_group_ids = {
                row.get("group_id", "")
                for row in context["existing_rows"].get(excl_groups_sheet, [])
            }
            gid = f"excl_{slug_rpo(record['groupKey'])}"
            while gid in existing_group_ids:
                gid += "x"
            op(
                stage2,
                {
                    "action": "add",
                    "sheet": excl_groups_sheet,
                    "key": {"group_id": gid},
                    "row": {"group_id": gid, "selection_mode": "single_within_group", "active": True, "notes": f"Review group: {record['groupKey']}"},
                },
                [record["decisionId"]],
            )
            for order, member in enumerate(members, start=1):
                op(
                    stage2,
                    {
                        "action": "add",
                        "sheet": excl_members_sheet,
                        "key": {"group_id": gid, "option_id": member},
                        "row": {"group_id": gid, "option_id": member, "display_order": order, "active": True},
                    },
                    [record["decisionId"]],
                )
            default_rpos = _listish(payload.get("defaultRpos") or payload.get("defaultSelectedRpos"))
            if payload.get("defaultRpo"):
                default_rpos = [*default_rpos, payload.get("defaultRpo")]
            if payload.get("requiresDefaultSelection") and not default_rpos:
                gaps.append(
                    {
                        "model": model,
                        "kind": "default_selection_rules_missing",
                        "detail": f"{record['groupKey']}: default selection required but no default RPO supplied",
                        "decisionId": record["decisionId"],
                    }
                )
                continue
            if default_rpos and "default_selection_rules" not in sheetnames:
                gaps.append(
                    {
                        "model": model,
                        "kind": "default_selection_rules_missing",
                        "detail": "default_selection_rules sheet is absent",
                        "decisionId": record["decisionId"],
                    }
                )
                continue
            for default_rpo in default_rpos:
                target_oid = resolve_option_id_by_rpo(default_rpo, record, "default selection target")
                if target_oid is None:
                    gaps.append(
                        {
                            "model": model,
                            "kind": "default_selection_rule_unresolved_required",
                            "detail": f"{record['groupKey']}: default RPO {default_rpo} did not resolve to an option_id",
                            "decisionId": record["decisionId"],
                        }
                    )
                    continue
                rid = default_rule_id_for(_rpo_token(default_rpo))
                op(
                    stage2,
                    {
                        "action": "add",
                        "sheet": "default_selection_rules",
                        "key": {"model_key": model, "rule_id": rid},
                        "row": {
                            "model_key": model,
                            "rule_id": rid,
                            "target_option_id": target_oid,
                            "condition_type": "always",
                            "condition_id": None,
                            "body_style_scope": "*",
                            "trim_level_scope": "*",
                            "variant_scope": "*",
                            "priority": 10,
                            "active": True,
                            "notes": f"Default selection from ingest review group {record['groupKey']}.",
                            "display_behavior": "default_selected",
                        },
                    },
                    [record["decisionId"]],
                )

        # Presentation rows (approved lane-10 payloads).
        for sheet in PRESENTATION_SHEETS:
            record = model_decisions.get(f"{model}:presentation:{sheet}")
            if not record or record["resolution"] != "approved_for_plan":
                gaps.append(
                    {
                        "model": model,
                        "kind": "presentation_missing",
                        "detail": sheet,
                        "decisionId": record["decisionId"] if record else "",
                    }
                )
                continue
            for row in (record.get("payload") or {}).get("rows") or []:
                keycols = {
                    "runtime_steps": ("model_key", "step_key"),
                    "section_presentation": ("model_key", "section_id"),
                    "context_section_master": ("model_key", "context_type", "section_id"),
                    "order_summary_sections": ("model_key", "section_key"),
                    "step_order_summary_map": ("model_key", "step_key", "section_key"),
                }[sheet]
                key = {k: row.get(k, "") for k in keycols}
                op(
                    stage2,
                    {"action": "add", "sheet": sheet, "key": key, "row": dict(row)},
                    [record["decisionId"]],
                )

    uncovered = sorted(
        decision_id
        for decision_id, refs in coverage.items()
        if not refs
        and decisions[decision_id]["resolution"] == "approved_for_plan"
        and decisions[decision_id]["lane"] not in ("interior_media_deferral", "status_nuance", "duplicate", "relationship")
        and (decisions[decision_id]["model"], decisions[decision_id].get("candidateId", "")) not in inert_candidates
    )

    per_sheet: dict[str, dict[str, int]] = {}
    for item in stage1 + stage2:
        sheet_counts = per_sheet.setdefault(item["sheet"], {})
        sheet_counts[item["action"]] = sheet_counts.get(item["action"], 0) + 1

    runtime_continuity: dict[str, dict[str, Any]] = {}
    for model in targets:
        config = MODEL_PLAN_CONFIG[model]
        model_counts: dict[str, Any] = {"sourceOps": {}}
        for surface, suffix in (
            ("priceRules", "price_rules"),
            ("ruleGroups", "rule_groups"),
            ("ruleGroupMembers", "rule_group_members"),
            ("exclusiveGroups", "exclusive_groups"),
            ("exclusiveMembers", "exclusive_members"),
            ("options", "options"),
        ):
            sheet = config["sheetPrefix"] + suffix
            model_counts["sourceOps"][surface] = dict(per_sheet.get(sheet, {}))
        model_counts["sourceOps"]["colorOverrides"] = dict(per_sheet.get("color_overrides", {}))
        model_counts["sourceOps"]["interiorComponents"] = dict(per_sheet.get("interior_components", {}))
        model_counts["sourceOps"]["assetMap"] = dict(per_sheet.get("asset_map", {}))
        runtime_continuity[model] = model_counts

    blocking = [gap for gap in gaps if gap["kind"] in BLOCKING_GAP_KINDS]

    return {
        "schemaVersion": SCHEMA_VERSION_C,
        "targets": targets,
        "candidatesFingerprint": candidates_fingerprint,
        "decisionsFingerprint": artifact_sha(decisions),
        "workbookFingerprint": {
            "sha256": hashlib.sha256(Path(workbook_path).read_bytes()).hexdigest(),
            "mtimeNs": str(Path(workbook_path).stat().st_mtime_ns),
        },
        "stage1": {"items": stage1},
        "stage2": {"items": stage2},
        "coverage": {"decisionToOps": coverage, "uncoveredApprovedDecisions": uncovered},
        "report": {
            "perSheetCounts": per_sheet,
            "perSheetActionCounts": per_sheet,
            "runtimeContinuity": runtime_continuity,
            "clearedRows": cleared_rows,
            "holds": holds,
            "deferrals": deferrals,
            "unreviewedSplits": unreviewed_splits,
            "gaps": gaps,
            "blockingGaps": blocking,
        },
        "valid": not blocking and not uncovered,
    }
