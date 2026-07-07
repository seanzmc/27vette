"""Read-only source profiler for raw GM Corvette order-guide exports.

This module implements the Pass 0 ingest contract: source-layout evidence,
variant matrix reconciliation, raw row preservation, disclosure links, and a
checkpoint report. It intentionally does not emit approved candidate workbook
rows or mutate the canonical workbook/generated outputs.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from corvette_form_generator.workbook import clean, workbook_truthy

ORDER_GUIDE_BASE_HEADERS = ("Orderable RPO Code", "Ref. Only RPO Code", "Description")
COLOR_TRIM_HEADERS = ("Decor Level", "Seat Type", "Seat Code", "Seat Trim")
PRICE_SCHEDULE_HEADERS = ("Model", "Model Description")
VALID_BASE_STATUSES = {
    "S": "standard",
    "A": "available",
    "D": "available",
    "A/D": "available",
    "--": "unavailable",
    "■": "standard",
}
UNRESOLVED_BASE_STATUSES = {"□"}
STATUS_RE = re.compile(r"^(A/D|--|[SAD■□])(\d*)$", re.IGNORECASE)
STATUS_MARKER_SUFFIX_RE = re.compile(r"^(.*?)(\d*)$")
RPO_RE = re.compile(r"^[A-Z0-9]{2,4}$")
DISCLOSURE_MARKER_RE = re.compile(r"(?:^|[\s/])(\d+)\.\s*")
RELATIONSHIP_HINTS = (
    ("requires", "requires"),
    ("not available with", "excludes"),
    ("included and only available with", "included_only_available_with"),
    ("included and only", "included_only_available_with"),
    ("included with", "includes"),
    ("includes", "includes"),
    ("only available with", "requires"),
)


@dataclass(frozen=True)
class VariantReference:
    variant_id: str
    model_key: str
    model_active: bool
    variant_active: bool
    membership_active: bool
    display_name: str
    body_style: str
    trim_level: str


def profile_order_guide(
    *,
    raw_export: Path,
    workbook: Path,
    output_dir: Path,
    run_id: str,
    root: Path | None = None,
) -> dict[str, Any]:
    """Profile a raw export and write Pass 0 run-scoped artifacts."""

    root = Path(root or Path.cwd()).resolve()
    raw_export = Path(raw_export).resolve()
    workbook = Path(workbook).resolve()
    output_dir = Path(output_dir).resolve()
    validate_output_dir(output_dir, root=root)

    variant_refs = load_variant_references(workbook)
    raw_wb = load_workbook(raw_export, read_only=True, data_only=True)
    try:
        source_layout: list[dict[str, Any]] = []
        variant_matrix: list[dict[str, Any]] = []
        raw_rows: list[dict[str, Any]] = []
        disclosure_links: list[dict[str, Any]] = []
        invariant_failures: list[dict[str, Any]] = []

        for ws in raw_wb.worksheets:
            layout = profile_sheet_layout(ws)
            source_layout.append(layout)
            if layout["sheet_type"] != "matrix":
                continue

            sheet_variants = [
                parse_variant_column(ws.title, column, raw_header, variant_refs)
                for column, raw_header in layout["variant_columns"]
            ]
            variant_matrix.extend(sheet_variants)
            unresolved_variants = [
                row for row in sheet_variants if row["resolution_status"] in {"unmatched", "ambiguous"}
            ]
            if unresolved_variants:
                invariant_failures.append(
                    {
                        "source_sheet": ws.title,
                        "check_id": "variant_header_resolution",
                        "message": "One or more variant headers did not reconcile to workbook variant metadata.",
                        "variants": unresolved_variants,
                    }
                )

            sheet_rows = extract_matrix_rows(ws, layout, sheet_variants)
            raw_rows.extend(sheet_rows)
            disclosure_links.extend(extract_disclosure_links(sheet_rows))

        checkpoint = build_checkpoint_report(
            raw_export=raw_export,
            workbook=workbook,
            run_id=run_id,
            source_layout=source_layout,
            variant_matrix=variant_matrix,
            raw_rows=raw_rows,
            disclosure_links=disclosure_links,
            invariant_failures=invariant_failures,
        )
        status = "failed" if invariant_failures else "passed"
        artifacts = {
            "source-layout.json": source_layout,
            "variant-matrix.json": variant_matrix,
            "raw-rows.json": raw_rows,
            "disclosure-links.json": disclosure_links,
            "manifest.json": {
                "tool": "scripts/order_guide_ingest_profiler.py",
                "run_id": run_id,
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "raw_export": str(raw_export),
                "workbook": str(workbook),
                "status": status,
                "artifact_files": [
                    "source-layout.json",
                    "variant-matrix.json",
                    "raw-rows.json",
                    "disclosure-links.json",
                    "checkpoint-report.md",
                ],
                "invariant_failures": invariant_failures,
            },
        }

        output_dir.mkdir(parents=True, exist_ok=True)
        for filename, payload in artifacts.items():
            (output_dir / filename).write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
        (output_dir / "checkpoint-report.md").write_text(checkpoint + "\n")

        return {
            "status": status,
            "output_dir": str(output_dir),
            "source_sheet_count": len(source_layout),
            "parsed_matrix_sheet_count": sum(1 for sheet in source_layout if sheet["sheet_type"] == "matrix"),
            "raw_row_count": len(raw_rows),
            "variant_column_count": len(variant_matrix),
            "disclosure_link_count": len(disclosure_links),
            "invariant_failures": invariant_failures,
        }
    finally:
        raw_wb.close()


def validate_output_dir(output_dir: Path, *, root: Path) -> None:
    """Refuse output paths that could mutate tracked generated outputs."""

    output_dir = output_dir.resolve()
    root = root.resolve()
    form_output = root / "form-output"
    ingest_root = form_output / "ingest"
    if _is_relative_to(output_dir, form_output) and not _is_relative_to(output_dir, ingest_root):
        raise ValueError(
            "Pass 0 output under form-output is only allowed below form-output/ingest/<run-id>; "
            f"got {output_dir}"
        )
    forbidden_paths = {
        root / "stingray_master.xlsx",
        root / "form-app" / "data.js",
    }
    if output_dir in {path.resolve() for path in forbidden_paths}:
        raise ValueError(f"Refusing to use protected output path: {output_dir}")


def _is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def load_variant_references(workbook: Path) -> dict[str, list[VariantReference]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        model_active = {
            clean(row.get("model_key")): workbook_truthy(row.get("active"))
            for row in rows_from_sheet(wb, "model_master")
        }
        variant_facts = {
            clean(row.get("variant_id")): row
            for row in rows_from_sheet(wb, "variant_master")
            if clean(row.get("variant_id"))
        }
        refs: dict[str, list[VariantReference]] = {}
        for row in rows_from_sheet(wb, "model_variants"):
            variant_id = clean(row.get("variant_id"))
            model_key = clean(row.get("model_key"))
            if not variant_id or not model_key:
                continue
            fact = variant_facts.get(variant_id, {})
            refs.setdefault(variant_id, []).append(
                VariantReference(
                    variant_id=variant_id,
                    model_key=model_key,
                    model_active=model_active.get(model_key, False),
                    variant_active=workbook_truthy(fact.get("active")),
                    membership_active=workbook_truthy(row.get("active")),
                    display_name=clean(fact.get("display_name")),
                    body_style=clean(fact.get("body_style")),
                    trim_level=clean(fact.get("trim_level")),
                )
            )
        return refs
    finally:
        wb.close()


def rows_from_sheet(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows_iter, ())]
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        record = {header: value for header, value in zip(headers, values) if header}
        if any(value is not None and value != "" for value in record.values()):
            rows.append(record)
    return rows


def profile_sheet_layout(ws) -> dict[str, Any]:
    header_row, headers = detect_header_row(ws)
    if header_row is None:
        return {
            "source_sheet": ws.title,
            "sheet_type": "unknown",
            "model_label": clean(ws.cell(1, 1).value),
            "header_row": None,
            "base_columns": [],
            "variant_columns": [],
            "section_rows": [],
            "data_row_count": 0,
            "status_vocabulary": {},
            "skipped_reason": "No supported header shape detected in the first 10 rows.",
            "invariant_warnings": [],
        }

    sheet_type = classify_headers(headers)
    if sheet_type != "matrix":
        return {
            "source_sheet": ws.title,
            "sheet_type": sheet_type,
            "model_label": clean(ws.cell(1, 1).value),
            "header_row": header_row,
            "base_columns": nonblank_columns(headers),
            "variant_columns": [],
            "section_rows": [],
            "data_row_count": max(ws.max_row - header_row, 0),
            "status_vocabulary": {},
            "skipped_reason": "Non-option-matrix source shape; preserved as layout evidence only.",
            "invariant_warnings": [],
        }

    variant_columns = [(column, clean(header)) for column, header in enumerate(headers, start=1) if column >= 4 and clean(header)]
    section_rows = detect_section_rows(ws, header_row)
    status_vocabulary = status_counts(ws, header_row, [column for column, _ in variant_columns])
    warnings = []
    missing = [header for header in ORDER_GUIDE_BASE_HEADERS if header not in {clean(value) for value in headers}]
    if missing:
        warnings.append({"check_id": "missing_base_header", "headers": missing})
    if not variant_columns:
        warnings.append({"check_id": "missing_variant_columns", "message": "No variant matrix columns detected."})

    return {
        "source_sheet": ws.title,
        "sheet_type": "matrix",
        "model_label": clean(ws.cell(1, 1).value),
        "header_row": header_row,
        "base_columns": [
            {"index": index, "letter": get_column_letter(index), "header": clean(header)}
            for index, header in enumerate(headers, start=1)
            if index <= 3 and clean(header)
        ],
        "variant_columns": variant_columns,
        "section_rows": section_rows,
        "data_row_count": max(ws.max_row - header_row, 0),
        "status_vocabulary": status_vocabulary,
        "skipped_reason": "",
        "invariant_warnings": warnings,
    }


def detect_header_row(ws) -> tuple[int | None, list[Any]]:
    for row_index in range(1, min(ws.max_row, 10) + 1):
        values = [ws.cell(row_index, column).value for column in range(1, ws.max_column + 1)]
        cleaned = {clean(value) for value in values}
        if set(ORDER_GUIDE_BASE_HEADERS).issubset(cleaned):
            return row_index, values
        if set(COLOR_TRIM_HEADERS).issubset(cleaned):
            return row_index, values
        if set(PRICE_SCHEDULE_HEADERS).issubset(cleaned):
            return row_index, values
    return None, []


def classify_headers(headers: list[Any]) -> str:
    cleaned = {clean(value) for value in headers}
    if set(ORDER_GUIDE_BASE_HEADERS).issubset(cleaned):
        return "matrix"
    if set(COLOR_TRIM_HEADERS).issubset(cleaned):
        return "color_trim"
    if set(PRICE_SCHEDULE_HEADERS).issubset(cleaned):
        return "price_schedule"
    return "unknown"


def nonblank_columns(headers: list[Any]) -> list[dict[str, Any]]:
    return [
        {"index": index, "letter": get_column_letter(index), "header": clean(header)}
        for index, header in enumerate(headers, start=1)
        if clean(header)
    ]


def detect_section_rows(ws, header_row: int) -> list[dict[str, Any]]:
    section_rows: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        orderable = clean(ws.cell(row_index, 1).value)
        ref_only = clean(ws.cell(row_index, 2).value)
        description = clean(ws.cell(row_index, 3).value)
        if orderable and not ref_only and not description and not RPO_RE.fullmatch(orderable):
            section_rows.append(
                {
                    "row_index": row_index,
                    "coordinate": f"A{row_index}",
                    "label": orderable,
                }
            )
    for index, section in enumerate(section_rows):
        next_row = section_rows[index + 1]["row_index"] if index + 1 < len(section_rows) else ws.max_row + 1
        section["row_span"] = [section["row_index"], next_row - 1]
    return section_rows


def status_counts(ws, header_row: int, variant_columns: list[int]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row_index in range(header_row + 1, ws.max_row + 1):
        for column in variant_columns:
            value = clean(ws.cell(row_index, column).value)
            if value:
                counts[value] = counts.get(value, 0) + 1
    return dict(sorted(counts.items()))


def parse_variant_column(
    source_sheet: str,
    column: int,
    raw_header: str,
    variant_refs: dict[str, list[VariantReference]],
) -> dict[str, Any]:
    parts = [part.strip() for part in raw_header.replace("\n", " / ").split("/")]
    parts = [part for part in parts if part]
    label = parts[0] if parts else raw_header
    model_code = parts[1] if len(parts) >= 2 else ""
    trim = parts[2] if len(parts) >= 3 else ""
    body_style = "convertible" if "convertible" in label.lower() else "coupe" if "coupe" in label.lower() else ""
    variant_id = candidate_variant_id(model_code, trim)
    matches = variant_refs.get(variant_id, []) if variant_id else []
    active_matches = [ref for ref in matches if ref.membership_active]
    resolution_status = "matched" if len(matches) == 1 else "ambiguous" if len(matches) > 1 else "unmatched"
    selected = matches[0] if matches else None
    if len(active_matches) == 1:
        selected = active_matches[0]
        resolution_status = "matched"
    elif len(active_matches) > 1:
        selected = active_matches[0]
        resolution_status = "ambiguous"

    return {
        "source_sheet": source_sheet,
        "source_column_index": column,
        "source_column_letter": get_column_letter(column),
        "raw_variant_header": raw_header,
        "parsed_body_style": body_style,
        "parsed_model_code": model_code,
        "parsed_trim": trim.lower(),
        "parsed_target_model": selected.model_key if selected else "",
        "parsed_variant_id": variant_id,
        "matched_variant_master": variant_ref_payload(selected) if selected else None,
        "matched_model_variants": [variant_ref_payload(ref) for ref in matches],
        "resolution_status": resolution_status,
        "evidence_notes": variant_evidence_notes(raw_header, variant_id, matches),
    }


def candidate_variant_id(model_code: str, trim: str) -> str:
    model_code = clean(model_code).lower()
    trim = clean(trim).lower()
    if not model_code or not trim:
        return ""
    suffix = model_code[2:] if len(model_code) >= 3 and model_code.startswith("1y") else model_code
    return f"{trim}_{suffix}"


def variant_ref_payload(ref: VariantReference | None) -> dict[str, Any] | None:
    if ref is None:
        return None
    return {
        "variant_id": ref.variant_id,
        "model_key": ref.model_key,
        "model_active": ref.model_active,
        "variant_active": ref.variant_active,
        "membership_active": ref.membership_active,
        "display_name": ref.display_name,
        "body_style": ref.body_style,
        "trim_level": ref.trim_level,
    }


def variant_evidence_notes(raw_header: str, variant_id: str, matches: list[VariantReference]) -> list[str]:
    notes = []
    if not variant_id:
        notes.append("Could not derive a candidate variant_id from the raw header.")
    if not matches:
        notes.append("No workbook variant metadata matched the derived candidate variant_id.")
    if len(matches) > 1:
        notes.append("Multiple workbook model memberships matched the derived candidate variant_id.")
    if "zr1" in raw_header.lower() and "zr1x" in raw_header.lower():
        notes.append("Header appears to name a combined model label; rely on model code for split.")
    return notes


def extract_matrix_rows(ws, layout: dict[str, Any], sheet_variants: list[dict[str, Any]]) -> list[dict[str, Any]]:
    header_row = layout["header_row"]
    variant_by_column = {variant["source_column_index"]: variant for variant in sheet_variants}
    section_for_row = section_lookup(layout["section_rows"], ws.max_row)
    rows: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        base_values = [clean(ws.cell(row_index, col).value) for col in (1, 2, 3)]
        status_values = [clean(ws.cell(row_index, col).value) for col in variant_by_column]
        if not any(base_values + status_values):
            continue
        orderable, ref_only, description = base_values
        status_cells = []
        for column, variant in variant_by_column.items():
            raw_status = clean(ws.cell(row_index, column).value)
            if not raw_status:
                continue
            status_cells.append(
                {
                    "coordinate": f"{get_column_letter(column)}{row_index}",
                    "source_column_index": column,
                    "raw_variant_header": variant["raw_variant_header"],
                    "variant_id_candidate": variant["parsed_variant_id"],
                    "model_key_candidate": variant["parsed_target_model"],
                    "raw_status": raw_status,
                    **parse_status(raw_status),
                }
            )
        row_flags = []
        primary_rpo = primary_rpo_candidate(orderable, ref_only)
        if (orderable or ref_only) and not primary_rpo:
            row_flags.append("unresolved_rpo_format")
        if is_section_row(orderable, ref_only, description):
            row_flags.append("section_context_row")
        rows.append(
            {
                "source_sheet": ws.title,
                "source_row_index": row_index,
                "source_row_span": section_for_row.get(row_index, {}).get("row_span", [row_index, row_index]),
                "section_context": section_for_row.get(row_index, {}).get("label", ""),
                "orderable_rpo": {"coordinate": f"A{row_index}", "raw_value": orderable},
                "ref_only_rpo": {"coordinate": f"B{row_index}", "raw_value": ref_only},
                "description": {"coordinate": f"C{row_index}", "raw_value": description},
                "primary_rpo_candidate": primary_rpo,
                "status_cells": status_cells,
                "description_disclosure_markers": description_markers(description),
                "row_flags": row_flags,
                "candidate_target_family_hints": target_family_hints(layout, orderable, ref_only, description),
            }
        )
    return rows


def section_lookup(section_rows: list[dict[str, Any]], max_row: int) -> dict[int, dict[str, Any]]:
    lookup: dict[int, dict[str, Any]] = {}
    if not section_rows:
        return lookup
    for section in section_rows:
        start, end = section.get("row_span", [section["row_index"], max_row])
        for row_index in range(start, end + 1):
            lookup[row_index] = section
    return lookup


def parse_status(raw_status: str) -> dict[str, Any]:
    symbol, marker = normalized_status_symbol(raw_status)
    if not symbol:
        return {
            "parsed_base_status": "unresolved",
            "status_marker": "",
            "status_flags": ["unknown_status_symbol"],
        }
    flags = []
    if marker:
        flags.append("status_disclosure_marker")
    if symbol in {"D", "A/D"}:
        flags.append("dealer_installed_or_adi")
    if symbol in UNRESOLVED_BASE_STATUSES:
        flags.append("upgradeable_equipment_group_review")
    return {
        "parsed_base_status": VALID_BASE_STATUSES.get(symbol, "unresolved"),
        "status_marker": marker,
        "status_flags": flags,
    }


def normalized_status_symbol(raw_status: str) -> tuple[str, str]:
    text = clean(raw_status)
    match = STATUS_MARKER_SUFFIX_RE.fullmatch(text)
    body = match.group(1) if match else text
    marker = match.group(2) if match else ""
    compact = re.sub(r"\s+", "", body).upper()
    compact = compact.replace("—", "-").replace("–", "-").replace("−", "-")
    if compact in {"A/D", "A-D", "AD"}:
        return "A/D", marker
    if compact in {"--", "-"}:
        return "--", marker
    match = STATUS_RE.fullmatch(compact + marker)
    if match:
        symbol, parsed_marker = match.groups()
        return symbol.upper(), parsed_marker
    return "", ""


def primary_rpo_candidate(orderable: str, ref_only: str) -> str:
    for value in (orderable, ref_only):
        candidate = clean(value).upper()
        if RPO_RE.fullmatch(candidate):
            return candidate
    return ""


def is_section_row(orderable: str, ref_only: str, description: str) -> bool:
    return bool(orderable and not ref_only and not description and not RPO_RE.fullmatch(orderable))


def description_markers(description: str) -> list[str]:
    return sorted(set(DISCLOSURE_MARKER_RE.findall(description)), key=int)


def target_family_hints(layout: dict[str, Any], orderable: str, ref_only: str, description: str) -> list[str]:
    if is_section_row(orderable, ref_only, description):
        return ["section_context"]
    hints = ["options", "ovs"]
    if any(phrase in description.lower() for phrase, _ in RELATIONSHIP_HINTS):
        hints.append("rule_relationship_review")
    if layout["source_sheet"].lower().startswith("interior"):
        hints.append("interior_review")
    return hints


def extract_disclosure_links(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    links: list[dict[str, Any]] = []
    for row in raw_rows:
        by_marker: dict[str, list[dict[str, Any]]] = {}
        for cell in row["status_cells"]:
            marker = cell.get("status_marker")
            if marker:
                by_marker.setdefault(marker, []).append(cell)
        markers = sorted(set(by_marker) | set(row["description_disclosure_markers"]), key=int)
        if not markers:
            continue
        fragments = disclosure_fragments(row["description"]["raw_value"])
        for marker in markers:
            fragment = fragments.get(marker, "")
            hint = relationship_hint(fragment)
            links.append(
                {
                    "source_sheet": row["source_sheet"],
                    "source_row_index": row["source_row_index"],
                    "marker": marker,
                    "status_cells": [
                        {
                            "coordinate": cell["coordinate"],
                            "raw_status": cell["raw_status"],
                            "variant_id_candidate": cell["variant_id_candidate"],
                            "model_key_candidate": cell["model_key_candidate"],
                        }
                        for cell in by_marker.get(marker, [])
                    ],
                    "description_fragment": fragment,
                    "phrase_hints": hint["phrase_hints"],
                    "candidate_relationship_hint": hint["candidate_relationship_hint"],
                    "review_state": "needs_review" if hint["phrase_hints"] or marker not in fragments else "evidence_only",
                }
            )
    return links


def disclosure_fragments(description: str) -> dict[str, str]:
    matches = list(DISCLOSURE_MARKER_RE.finditer(description))
    fragments: dict[str, str] = {}
    for index, match in enumerate(matches):
        marker = match.group(1)
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(description)
        fragments[marker] = description[start:end].strip(" /\n\t")
    return fragments


def relationship_hint(text: str) -> dict[str, Any]:
    lowered = text.lower()
    phrases = []
    candidates = []
    for phrase, relationship in RELATIONSHIP_HINTS:
        if phrase in lowered:
            phrases.append(phrase)
            candidates.append(relationship)
    return {
        "phrase_hints": phrases,
        "candidate_relationship_hint": candidates[0] if candidates else "",
    }


def build_checkpoint_report(
    *,
    raw_export: Path,
    workbook: Path,
    run_id: str,
    source_layout: list[dict[str, Any]],
    variant_matrix: list[dict[str, Any]],
    raw_rows: list[dict[str, Any]],
    disclosure_links: list[dict[str, Any]],
    invariant_failures: list[dict[str, Any]],
) -> str:
    matrix_sheets = [sheet for sheet in source_layout if sheet["sheet_type"] == "matrix"]
    skipped_sheets = [sheet for sheet in source_layout if sheet["sheet_type"] != "matrix"]
    unresolved_variants = [row for row in variant_matrix if row["resolution_status"] != "matched"]
    unresolved_rows = [row for row in raw_rows if row["row_flags"]]
    lines = [
        "# Order-guide ingest Pass 0 checkpoint",
        "",
        f"Run ID: `{run_id}`",
        f"Raw export: `{raw_export}`",
        f"Workbook reference: `{workbook}`",
        "",
        "## Summary",
        "",
        f"- Source sheets: {len(source_layout)}",
        f"- Parsed matrix sheets: {len(matrix_sheets)}",
        f"- Skipped/non-matrix sheets: {len(skipped_sheets)}",
        f"- Variant columns: {len(variant_matrix)}",
        f"- Raw evidence rows: {len(raw_rows)}",
        f"- Disclosure links: {len(disclosure_links)}",
        f"- Unresolved variant headers: {len(unresolved_variants)}",
        f"- Rows with parse flags: {len(unresolved_rows)}",
        f"- Invariant status: {'FAIL' if invariant_failures else 'PASS'}",
        "",
        "No workbook, generated runtime, registry, or browser runtime files were written by this profiler.",
        "",
        "## Sheets",
        "",
    ]
    for sheet in source_layout:
        lines.extend(
            [
                f"### {sheet['source_sheet']}",
                "",
                f"- Type: {sheet['sheet_type']}",
                f"- Header row: {sheet['header_row']}",
                f"- Data rows: {sheet['data_row_count']}",
                f"- Variant columns: {len(sheet['variant_columns'])}",
                f"- Status vocabulary: {sheet['status_vocabulary']}",
            ]
        )
        if sheet["skipped_reason"]:
            lines.append(f"- Skipped reason: {sheet['skipped_reason']}")
        if sheet["invariant_warnings"]:
            lines.append(f"- Warnings: {sheet['invariant_warnings']}")
        lines.append("")
    if invariant_failures:
        lines.extend(["## Invariant failures", ""])
        for failure in invariant_failures:
            lines.append(f"- `{failure['check_id']}` on `{failure['source_sheet']}`: {failure['message']}")
    return "\n".join(lines).rstrip()
