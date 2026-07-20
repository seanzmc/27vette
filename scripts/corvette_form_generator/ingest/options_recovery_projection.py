#!/usr/bin/env python3
"""Build the read-only Deliverable 4.1 options recovery projection.

This module opens every workbook with ``read_only=True`` and only writes the
requested Markdown/JSON reports.  It never saves or mutates a workbook.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import subprocess
import tempfile
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from corvette_form_generator.ingest.wizard.copy_split import propose_copy_split
from corvette_form_generator.ingest.wizard.decisions import (
    candidate_fingerprint,
    model_scoped_statuses,
    scope_candidates,
)
from corvette_form_generator.runtime_metadata import truthy
from corvette_form_generator.workbook import clean, intish


SCHEMA_VERSION = "options-recovery-projection-1"
CHECKPOINT_1_APPROVAL_SCHEMA_VERSION = "options-recovery-checkpoint-1-approval-1"
CHECKPOINT_1_EXCEPTION_SCHEMA_VERSION = "options-recovery-checkpoint-1-exceptions-1"
CHECKPOINT_1_DECISIONS_SCHEMA_VERSION = "options-recovery-checkpoint-1-decisions-1"
CHECKPOINT_1_PENDING_SCHEMA_VERSION = "options-recovery-checkpoint-1-pending-1"
TARGET_MODELS = ("grand_sport_x", "zr1", "zr1x")
FRESH_GSX_RPOS = frozenset({"N26", "PRB", "R6P", "R9L", "R9V", "R9W", "R9Y", "TU7"})
MANDATORY_CHARGE_RPOS = frozenset({"R8E"})
HASH_OPTION_ID_RE = re.compile(r"^opt_std_[0-9a-f]{12,}$")
NUMERIC_OPTION_ID_RE = re.compile(r"^opt_(\d{3})$")
TOKEN_RE = re.compile(r"[a-z0-9]+")
COPY_COMPARISON_STOPWORDS = frozenset(
    {"a", "an", "and", "for", "in", "include", "includes", "lpo", "new", "of", "on", "or", "the", "to", "with"}
)

OPTION_VIEW_FIELDS = (
    "option_id",
    "option_name",
    "description",
    "detail_raw",
    "section_id",
    "price",
    "active",
    "selectable",
    "display_order",
)


@dataclass(frozen=True)
class ProjectionInputs:
    workbook_path: Path
    pre_integration_workbook_path: Path
    reviewed_plan_path: Path
    reviewed_candidates_path: Path
    reconciliation_candidates_path: Path
    reviewed_decisions_path: Path
    exception_queue_path: Path
    exception_resolutions_path: Path
    canonical_manifest_path: Path
    comparator_evidence_path: Path


def _json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha(payload: Any) -> str:
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _rows(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    """Return non-empty rows with row numbers, ignoring stale dimensions."""

    if sheet_name not in wb.sheetnames:
        return []
    values = wb[sheet_name].iter_rows(values_only=True)
    try:
        raw_headers = next(values)
    except StopIteration:
        return []
    headers = [clean(value) for value in raw_headers]
    rows: list[dict[str, Any]] = []
    for row_number, raw_row in enumerate(values, start=2):
        if not any(value is not None for value in raw_row):
            continue
        record = {
            header: value
            for header, value in zip(headers, raw_row)
            if header
        }
        if not any(value is not None and clean(value) for value in record.values()):
            continue
        record["_row_number"] = row_number
        rows.append(record)
    return rows


def _option_row(row: Mapping[str, Any]) -> dict[str, Any]:
    price_text = clean(row.get("price"))
    order_text = clean(row.get("display_order"))
    return {
        "option_id": clean(row.get("option_id")),
        "rpo": clean(row.get("rpo")).upper(),
        "price": None if not price_text else int(float(price_text)),
        "option_name": clean(row.get("option_name")),
        "description": clean(row.get("description")),
        "detail_raw": clean(row.get("detail_raw")),
        "section_id": clean(row.get("section_id")),
        "selectable": truthy(row.get("selectable"), default=True),
        "display_order": None if not order_text else intish(order_text),
        "active": truthy(row.get("active"), default=True),
        "display_behavior": clean(row.get("display_behavior")),
        "row_number": int(row.get("_row_number") or 0),
    }


def _option_view(row: Mapping[str, Any]) -> dict[str, Any]:
    return {field: row.get(field) for field in OPTION_VIEW_FIELDS}


def _source_option_sheets(wb: Any) -> dict[str, str]:
    """Return configured option sheets, including inactive/unpromoted models."""

    sources: dict[str, str] = {}
    for row in _rows(wb, "model_workbook_sources"):
        if clean(row.get("source_role")) != "source_option_sheet":
            continue
        sources[clean(row.get("model_key")).lower()] = clean(row.get("sheet_name"))
    return sources


def _section_modes(wb: Any) -> dict[str, str]:
    return {
        clean(row.get("section_id")): clean(row.get("selection_mode"))
        for row in _rows(wb, "section_master")
        if clean(row.get("section_id"))
    }


def _index_options(rows: Iterable[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    by_id: dict[str, dict[str, Any]] = {}
    by_rpo: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row["option_id"]:
            by_id[row["option_id"]] = row
        if row["rpo"]:
            by_rpo[row["rpo"]] = row
    return by_id, by_rpo


def _target_all_unavailable_by_rpo(
    candidates: Mapping[str, Any],
    model: str,
) -> dict[str, dict[str, Any]]:
    unavailable: dict[str, list[dict[str, Any]]] = {}
    applicable_rpos: set[str] = set()
    for candidate in scope_candidates(list(candidates.get("candidates") or []), model):
        statuses = model_scoped_statuses(candidate, model)
        if not statuses:
            continue
        rpo = clean(candidate.get("rpo") or candidate.get("refOnlyRpo")).upper()
        if not rpo:
            continue
        if all(status.get("status") == "unavailable" for status in statuses):
            unavailable.setdefault(rpo, []).append({**candidate, "statuses": statuses})
        else:
            applicable_rpos.add(rpo)
    return {
        rpo: {
            "sourceCandidateIds": sorted(clean(candidate.get("candidateId")) for candidate in rows),
            "targetStatuses": [
                clean(status.get("raw"))
                for candidate in rows
                for status in candidate.get("statuses") or []
            ],
        }
        for rpo, rows in sorted(unavailable.items())
        if rpo not in applicable_rpos
    }


def _tokens(value: str) -> set[str]:
    return set(TOKEN_RE.findall(value.lower().replace("new!", " ")))


def _semantic_text(row: Mapping[str, Any]) -> str:
    return " ".join(clean(row.get(field)) for field in ("option_name", "description", "detail_raw"))


def _semantic_score(left: Mapping[str, Any], right: Mapping[str, Any]) -> float:
    left_tokens = _tokens(_semantic_text(left))
    right_tokens = _tokens(_semantic_text(right))
    if not left_tokens or not right_tokens:
        return 0.0
    overlap = len(left_tokens & right_tokens)
    overlap_coefficient = overlap / min(len(left_tokens), len(right_tokens))
    jaccard = overlap / len(left_tokens | right_tokens)
    left_name = clean(left.get("option_name")).lower()
    right_name = clean(right.get("option_name")).lower()
    name_bonus = 1.0 if left_name and right_name and (left_name in right_name or right_name in left_name) else 0.0
    return (0.60 * overlap_coefficient) + (0.30 * jaccard) + (0.10 * name_bonus)


def _comparator_copy_comparison(target: Mapping[str, Any], comparator: Mapping[str, Any]) -> dict[str, Any]:
    target_text = " ".join(
        value
        for value in (
            clean(target.get("option_name")),
            clean(target.get("description")),
            clean(target.get("detail_raw")),
        )
        if value
    )
    target_tokens = _tokens(target_text) - COPY_COMPARISON_STOPWORDS
    comparator_name_tokens = _tokens(clean(comparator.get("option_name"))) - COPY_COMPARISON_STOPWORDS
    matched = target_tokens & comparator_name_tokens
    coverage = len(matched) / len(comparator_name_tokens) if comparator_name_tokens else 1.0
    return {
        "materialDisagreement": bool(comparator_name_tokens) and coverage < 0.60,
        "comparatorNameTokenCoverage": round(coverage, 6),
        "comparatorNameTokens": sorted(comparator_name_tokens),
        "matchedNameTokens": sorted(matched),
        "targetRawText": target_text,
    }


def _copy_title(value: str) -> str:
    """Title-case customer copy while preserving known product abbreviations."""

    titled = value.title()
    for old, new in (
        ("Zr1", "ZR1"),
        ("Ztk", "ZTK"),
        ("Lpo", "LPO"),
    ):
        titled = titled.replace(old, new)
    return titled


def _identifying_copy_proposal(
    current: Mapping[str, Any],
) -> dict[str, Any]:
    """Propose concise identifying copy without reducing equipment to a generic noun."""

    raw = next(
        (
            clean(current.get(field))
            for field in ("option_name", "description", "detail_raw")
            if clean(current.get(field))
        ),
        "",
    )
    wheels = re.match(r'^Wheels,\s*(?P<size>.+?\brear)\s+(?P<design>\d+-spoke,\s*.+)$', raw, re.IGNORECASE)
    if wheels:
        design = _copy_title(wheels.group("design").replace(",", "").strip())
        return {
            "option_name": f"{design} Wheels",
            "description": wheels.group("size").strip(),
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    calipers = re.match(r"^Calipers,\s*(?P<finish>.+)$", raw, re.IGNORECASE)
    if calipers:
        return {
            "option_name": f"{_copy_title(calipers.group('finish').strip())} Calipers",
            "description": "",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    lower = raw.lower()
    transmission = re.match(
        r"^Transmission,\s*(?P<speeds>\d+)-speed dual clutch,\s*includes manual and auto modes$",
        raw,
        re.IGNORECASE,
    )
    if transmission:
        return {
            "option_name": f"{transmission.group('speeds')}-Speed Dual-Clutch Transmission",
            "description": "Includes manual and automatic modes",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    stitching = re.match(
        r"^(?P<color>.+?) custom leather stitch,\s*includes seats, instrument panel, doors and console$",
        raw,
        re.IGNORECASE,
    )
    if stitching:
        return {
            "option_name": f"{_copy_title(stitching.group('color'))} Custom Leather Stitching",
            "description": "Includes seats, instrument panel, doors, and console",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    stripes = re.match(
        r"^(?:NEW!\s*)?(?P<color>.+?) Full Length Dual Racing Stripes(?:\n1\.\s*(?P<restriction>.+))?$",
        raw,
        re.IGNORECASE,
    )
    if stripes:
        return {
            "option_name": f"{_copy_title(stripes.group('color'))} Full-Length Dual Racing Stripes",
            "description": clean(stripes.group("restriction")),
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("exterior trim, carbon fiber split window trim, painted body-color"):
        return {
            "option_name": "Body-Color Carbon Fiber Split-Window Trim",
            "description": "",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("steering wheel, carbon fiber and sueded microfiber-wrapped"):
        return {
            "option_name": "Carbon Fiber and Sueded Microfiber Steering Wheel",
            "description": "",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if "carbon fiber aero package" in lower and "includes" in lower:
        _, ancillary = raw.split(", includes", 1)
        return {
            "option_name": "Visible Carbon Fiber Aero Package",
            "description": f"Includes{ancillary}",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("ztk track performance package, includes"):
        _, ancillary = raw.split(", includes", 1)
        return {
            "option_name": "ZTK Track Performance Package",
            "description": f"Includes{ancillary}",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("convertible top,") and "visible carbon fiber" in lower:
        return {
            "option_name": "Visible Carbon Fiber Retractable Hardtop",
            "description": (
                "Power-folding with remote control (down only) and power glass rear window with integral "
                "defogger; nacelles, A-pillars, and header are Carbon Flash-painted (body-color with GBA "
                "Black exterior)"
            ),
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("front axle, electrified propulsion"):
        axle = re.match(
            r"^Front axle, electrified propulsion \((?P<axle>.+)\)\.\s*"
            r"(?P<hp>\d+) \((?P<kw>[^)]+)\) total combined hp with LT7 engine$",
            raw,
            re.IGNORECASE,
        )
        return {
            "option_name": "Electrified Front Axle",
            "description": (
                f"{axle.group('axle')}; {axle.group('hp')} hp [{axle.group('kw')}] total combined with LT7 engine"
                if axle
                else re.sub(r"^Front axle, electrified propulsion\s*", "", raw, flags=re.IGNORECASE)
            ),
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if "hood and roof decal package" in lower and "genuine corvette accessory" in lower:
        return {
            "option_name": "Hood and Roof Decal Package",
            "description": "LPO. Genuine Corvette Accessory",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("brakes,"):
        if "zr1-specific" in lower:
            name = "ZR1-Specific Carbon Ceramic Brakes"
        else:
            piston = re.search(
                r"(?P<front>\d+)-piston front(?: and|,)\s*(?P<rear>\d+)-piston rear",
                raw,
                re.IGNORECASE,
            )
            name = (
                f"{piston.group('front')}-Piston Front / {piston.group('rear')}-Piston Rear "
                "Carbon Ceramic Brakes"
                if piston
                else "Carbon Ceramic Brakes"
            )
        return {
            "option_name": name,
            "description": (
                "4-wheel antilock disc"
                if "4-wheel antilock disc" in lower
                or ("4-wheel antilock" in lower and "4-wheel disc" in lower)
                else ""
            ),
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("engine,") or ("5.5l" in lower and "v8" in lower):
        description = re.sub(r"^Engine,\s*", "", raw, flags=re.IGNORECASE)
        description = re.sub(r"^5\.5L V8 Twin Turbo,?\s*", "", description, flags=re.IGNORECASE)
        return {
            "option_name": "5.5L Twin-Turbo V8 Engine",
            "description": description.strip(" ,"),
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    if lower.startswith("suspension,"):
        is_ztk = "ztk" in lower
        return {
            "option_name": "ZTK Performance Suspension" if is_ztk else "Performance Suspension with Magnetic Ride Control",
            "description": "Magnetic Selective Ride Control" if is_ztk else "",
            "sourceType": "identifying_copy_derivation",
            "flags": [],
        }

    return {
        "option_name": clean(current.get("option_name")),
        "description": clean(current.get("description")),
        "sourceType": "identifying_copy_derivation",
        "flags": ["needs_curated_copy"],
    }


def _comparator_ref(model: str, sheet: str, row: Mapping[str, Any], score: float | None = None) -> dict[str, Any]:
    ref = {
        "model": model,
        "sheet": sheet,
        "rowNumber": row.get("row_number"),
        "optionId": row.get("option_id"),
        "rpo": row.get("rpo"),
    }
    if score is not None:
        ref["semanticScore"] = round(score, 6)
    return ref


def _best_no_rpo_comparator(
    target: Mapping[str, Any],
    comparator_sets: Iterable[tuple[str, str, list[dict[str, Any]]]],
) -> tuple[str, str, dict[str, Any], float] | None:
    candidates: list[tuple[float, int, str, str, dict[str, Any]]] = []
    target_section = clean(target.get("section_id"))
    for priority, (model, sheet, rows) in enumerate(comparator_sets):
        for row in rows:
            if row["rpo"]:
                continue
            score = _semantic_score(target, row)
            if target_section and row["section_id"] == target_section:
                score += 0.05
            candidates.append((score, -priority, model, sheet, row))
    if not candidates:
        return None
    score, _, model, sheet, row = max(candidates, key=lambda item: (item[0], item[1], item[4]["option_id"]))
    if score < 0.45:
        return None
    return model, sheet, row, score


def _plan_rows(
    plan: Mapping[str, Any],
    candidates: Mapping[str, Any],
    decisions: Mapping[str, Any],
    current_by_rpo: Mapping[str, dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    candidate_by_id = {
        clean(candidate.get("candidateId")): candidate
        for candidate in candidates.get("candidates", [])
    }
    decision_by_id = {
        clean(decision.get("decisionId")): decision
        for decision in decisions.get("decisions", [])
    }
    valid: dict[str, dict[str, Any]] = {}
    invalid: dict[str, list[str]] = {}
    for item in (plan.get("stage2") or {}).get("items", []):
        row = item.get("row") or {}
        rpo = clean(row.get("rpo")).upper()
        if not rpo or rpo not in current_by_rpo:
            continue
        reasons: list[str] = []
        decision_ids = [clean(value) for value in item.get("_decisions", [])]
        if not decision_ids:
            reasons.append("missing_plan_decision_binding")
        for decision_id in decision_ids:
            decision = decision_by_id.get(decision_id)
            if decision is None:
                reasons.append(f"missing_decision:{decision_id}")
                continue
            candidate = candidate_by_id.get(clean(decision.get("candidateId")))
            if candidate is None:
                reasons.append(f"missing_candidate:{decision_id}")
                continue
            if candidate_fingerprint(candidate) != clean(decision.get("candidateFingerprint")):
                reasons.append(f"candidate_fingerprint_mismatch:{decision_id}")
        current_raw = clean(current_by_rpo[rpo].get("detail_raw"))
        planned_raw = clean(row.get("detail_raw"))
        if current_raw != planned_raw:
            reasons.append("current_detail_raw_differs_from_reviewed_candidate")
        if reasons:
            invalid[rpo] = sorted(set(reasons))
        else:
            valid[rpo] = dict(row)
    return valid, invalid


def _next_option_ids(rows: list[dict[str, Any]], count: int) -> list[str]:
    reserved = {row["option_id"] for row in rows}
    result: list[str] = []
    number = 1
    while len(result) < count:
        candidate = f"opt_{number:03d}"
        if candidate not in reserved:
            reserved.add(candidate)
            result.append(candidate)
        number += 1
    return result


def _cascade_references(wb: Any, old_option_id: str, target_sheet: str) -> list[dict[str, Any]]:
    refs: list[dict[str, Any]] = []
    token = re.compile(rf"(?<![A-Za-z0-9_]){re.escape(old_option_id)}(?![A-Za-z0-9_])")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration:
            continue
        for row_number, raw_row in enumerate(rows, start=2):
            if not any(value is not None for value in raw_row):
                continue
            for header, value in zip(headers, raw_row):
                if not header or value is None:
                    continue
                text = clean(value)
                if not token.search(text):
                    continue
                if sheet_name == target_sheet and header == "option_id" and text == old_option_id:
                    continue
                refs.append(
                    {
                        "sheet": sheet_name,
                        "rowNumber": row_number,
                        "column": header,
                        "value": text,
                        "matchType": "exact" if text == old_option_id else "embedded",
                    }
                )
    return sorted(refs, key=lambda ref: (ref["sheet"], ref["rowNumber"], ref["column"]))


def _deletion_projection(
    wb: Any,
    model: str,
    target_sheet: str,
    current: Mapping[str, Any],
    reason: str,
    **evidence: Any,
) -> dict[str, Any]:
    references = _cascade_references(wb, clean(current.get("option_id")), target_sheet)
    owned_references = [
        reference
        for reference in references
        if reference["sheet"].startswith(f"{model}_")
    ]
    return {
        "model": model,
        "targetSheet": target_sheet,
        "rowNumber": current["row_number"],
        "optionId": current["option_id"],
        "rpo": current["rpo"],
        "reason": reason,
        **evidence,
        "ownedReferenceRows": owned_references,
        "externalSameIdReferencesPreserved": [
            reference
            for reference in references
            if reference not in owned_references
        ],
    }


def _manifest_identity_by_evidence(manifest: Mapping[str, Any]) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for row in manifest.get("rows", []):
        if row.get("family") != "options":
            continue
        values = row.get("values") or {}
        identity = {
            "model": clean(row.get("model")).lower(),
            "option_id": clean(values.get("option_id")),
            "rpo": clean(values.get("rpo")).upper(),
        }
        for dependency in row.get("evidenceDependencies", []):
            evidence_id = clean(dependency.get("evidenceId"))
            if ":candidate:" in evidence_id:
                result[evidence_id] = identity
    return result


def _section_reconciliation(
    queue: Mapping[str, Any],
    resolutions: Mapping[str, Any],
    manifest: Mapping[str, Any],
    options_by_model: Mapping[str, tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]],
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, dict[str, int]]]:
    resolution_by_subject = {
        clean(entry.get("subjectId")): entry
        for entry in resolutions.get("entries", [])
        if entry.get("action") == "choose_section" and entry.get("disposition") == "resolved"
    }
    identity_by_evidence = _manifest_identity_by_evidence(manifest)
    result = {model: [] for model in TARGET_MODELS}
    checks = {
        model: {"decisionCount": 0, "matchCount": 0, "mismatchCount": 0, "missingLandedRowCount": 0}
        for model in TARGET_MODELS
    }
    for subject in queue.get("subjects", []):
        subject_id = clean(subject.get("subjectId"))
        resolution = resolution_by_subject.get(subject_id)
        if resolution is None:
            continue
        if clean(subject.get("subjectVersion")) != clean(resolution.get("subjectVersion")):
            continue
        model = clean(subject.get("model")).lower()
        if model not in result:
            continue
        identity = None
        evidence_id = ""
        for dependency in subject.get("evidenceDependencies", []):
            evidence_id = clean(dependency.get("evidenceId"))
            identity = identity_by_evidence.get(evidence_id)
            if identity:
                break
        if not identity:
            continue
        checks[model]["decisionCount"] += 1
        by_id, by_rpo = options_by_model[model]
        landed = by_rpo.get(identity["rpo"]) if identity["rpo"] else by_id.get(identity["option_id"])
        decided = clean((resolution.get("payload") or {}).get("sectionId"))
        landed_section = clean((landed or {}).get("section_id"))
        if landed and landed_section == decided:
            checks[model]["matchCount"] += 1
            continue
        if landed:
            checks[model]["mismatchCount"] += 1
        else:
            checks[model]["missingLandedRowCount"] += 1
        result[model].append(
            {
                "subjectId": subject_id,
                "subjectVersion": clean(subject.get("subjectVersion")),
                "reviewer": clean(resolution.get("reviewer")),
                "evidenceId": evidence_id,
                "optionId": identity["option_id"],
                "rpo": identity["rpo"],
                "landedSectionId": landed_section,
                "decidedSectionId": decided,
                "landedRowNumber": (landed or {}).get("row_number"),
                "status": "mismatch" if landed else "landed_row_missing",
            }
        )
    for model in result:
        result[model].sort(key=lambda item: (item["rpo"], item["optionId"], item["subjectId"]))
    return result, checks


def _review_item(
    model: str,
    row: Mapping[str, Any],
    lane: str,
    proposed: Any,
    provenance: Mapping[str, Any],
    *,
    note: str = "",
) -> dict[str, Any]:
    seed = {
        "model": model,
        "optionId": row.get("option_id"),
        "rpo": row.get("rpo"),
        "lane": lane,
        "proposed": proposed,
        "provenance": provenance,
    }
    return {
        "reviewId": f"review:{model}:{lane}:{_canonical_sha(seed)[:16]}",
        "lane": lane,
        "proposed": proposed,
        "provenance": dict(provenance),
        "note": note,
        "decision": {
            "status": "pending",
            "action": "accept_or_override",
            "override": None,
            "reviewer": None,
            "reviewedAt": None,
        },
    }


def _field_change(
    after: dict[str, Any],
    provenance: dict[str, Any],
    field: str,
    value: Any,
    source: Mapping[str, Any],
) -> None:
    if after.get(field) != value:
        after[field] = value
        provenance[field] = dict(source)


def _allocate_order(used: set[int], comparator_order: int | None) -> tuple[int, str]:
    if comparator_order is not None and comparator_order not in used:
        used.add(comparator_order)
        return comparator_order, "comparator_proposal"
    highest = max(used, default=0)
    proposed = int(math.ceil((highest + 1) / 10.0) * 10)
    while proposed in used:
        proposed += 10
    used.add(proposed)
    return proposed, "deterministic_section_local"


def _markdown(report: Mapping[str, Any]) -> str:
    def cell(value: Any) -> str:
        if value is None:
            return "blank"
        if isinstance(value, bool):
            return "true" if value else "false"
        text = str(value).replace("\n", " ↵ ").replace("|", "\\|")
        return text if len(text) <= 120 else text[:117] + "..."

    lines = [
        f"# {report['model']} Options Recovery Projection",
        "",
        "> Read-only Deliverable 4.1 output. PENDING CHECKPOINT 1. This report does not authorize or perform a workbook write.",
        "",
        "## Summary",
        "",
        "| Metric | Count |",
        "|---|---:|",
    ]
    for key, value in report["summary"].items():
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## Source bindings", ""])
    for key, value in report["sources"].items():
        lines.append(f"- {key}: `{cell(value)}`")

    check = report["sectionReconciliationCheck"]
    lines.extend(
        [
            "",
            "## Section-decision reconciliation check",
            "",
            "| Check | Count |",
            "|---|---:|",
            f"| Decisions checked | {check['decisionCount']} |",
            f"| Landed matches | {check['matchCount']} |",
            f"| Landed mismatches | {check['mismatchCount']} |",
            f"| Missing landed rows | {check['missingLandedRowCount']} |",
            "",
            "### Decisions that do not match the landed workbook",
            "",
        ]
    )
    if report["sectionReconciliation"]:
        lines.extend(["| RPO / option | Landed | Decided | Reviewer | Subject |", "|---|---|---|---|---|"])
        for item in report["sectionReconciliation"]:
            identity = item["rpo"] or item["optionId"]
            lines.append(
                f"| {cell(identity)} | {cell(item['landedSectionId'])} | {cell(item['decidedSectionId'])} | "
                f"{cell(item['reviewer'])} | `{cell(item['subjectId'])}` |"
            )
    else:
        lines.append("None.")

    bulk = report["bulkDecisions"]
    lines.extend(
        [
            "",
            "## Bulk decision sets",
            "",
            "| Set | Review items |",
            "|---|---:|",
            f"| Comparator copy: bulk-safe | {len(bulk['comparatorCopy']['reviewIds'])} |",
            "| Comparator copy: material-disagreement exclusions | "
            f"{len(bulk['comparatorCopy']['excludedMaterialDisagreementReviewIds'])} |",
            f"| Copy-split derivation | {len(bulk['copySplit']['reviewIds'])} |",
            f"| Comparator display order | {len(bulk['comparatorDisplayOrder']['reviewIds'])} |",
        ]
    )

    lines.extend(["", "## ID repair preview", ""])
    if report["idRepairs"]:
        lines.extend(["| Old ID | Proposed ID | Cascade references |", "|---|---|---:|"])
        for item in report["idRepairs"]:
            lines.append(f"| `{item['oldOptionId']}` | `{item['proposedOptionId']}` | {len(item['cascade'])} |")
            for ref in item["cascade"]:
                lines.append(
                    f"  - `{ref['sheet']}` row {ref['rowNumber']} `{ref['column']}` ({ref['matchType']})"
                )
    else:
        lines.append("None.")

    lines.extend(["", "## Target-applicability deletions", ""])
    if report["targetApplicabilityDeletions"]:
        lines.extend(
            [
                "| RPO | Option ID | Target raw statuses | Owned reference rows |",
                "|---|---|---|---:|",
            ]
        )
        for item in report["targetApplicabilityDeletions"]:
            lines.append(
                f"| {cell(item['rpo'])} | `{cell(item['optionId'])}` | "
                f"{cell(', '.join(item['targetStatuses']))} | {len(item['ownedReferenceRows'])} |"
            )
    else:
        lines.append("None.")

    lines.extend(["", "## Recorded cross-target deletions", ""])
    if report["recordedInstructionDeletions"]:
        lines.extend(
            [
                "| RPO | Option ID | Reason | Owned reference rows |",
                "|---|---|---|---:|",
            ]
        )
        for item in report["recordedInstructionDeletions"]:
            lines.append(
                f"| {cell(item['rpo'])} | `{cell(item['optionId'])}` | "
                f"{cell(item['reason'])} | {len(item['ownedReferenceRows'])} |"
            )
    else:
        lines.append("None.")

    lines.extend(["", "## Pending Checkpoint 1 decisions", ""])
    lines.extend(["| Review ID | Lane | RPO / option | Proposed | Provenance |", "|---|---|---|---|---|"])
    for row in report["residualRows"]:
        identity = row["identity"]["rpo"] or row["identity"]["optionId"]
        for item in row["reviewItems"]:
            lines.append(
                f"| `{item['reviewId']}` | {cell(item['lane'])} | {cell(identity)} | "
                f"{cell(json.dumps(item['proposed'], sort_keys=True, ensure_ascii=False))} | "
                f"{cell(json.dumps(item['provenance'], sort_keys=True, ensure_ascii=False))} |"
            )

    lines.extend(["", "## Residual before / after", ""])
    lines.extend(["| Partition | RPO / option | Field | Before | Proposed after |", "|---|---|---|---|---|"])
    for row in report["residualRows"]:
        identity = row["identity"]["rpo"] or row["identity"]["optionId"]
        changed_fields = [field for field in OPTION_VIEW_FIELDS if row["before"].get(field) != row["after"].get(field)]
        if not changed_fields:
            lines.append(f"| {cell(row['partition'])} | {cell(identity)} | review only | — | — |")
        for field in changed_fields:
            lines.append(
                f"| {cell(row['partition'])} | {cell(identity)} | `{field}` | "
                f"{cell(row['before'].get(field))} | {cell(row['after'].get(field))} |"
            )
    lines.extend(["", "Full provenance, cascade cell values, and decision placeholders are in the paired JSON report.", ""])
    return "\n".join(lines)


def _validate_checkpoint_report(report: Mapping[str, Any], model: str, path: Path) -> None:
    if report.get("schemaVersion") != SCHEMA_VERSION:
        raise ValueError(f"Unsupported recovery report schema for {model}: {path}")
    if report.get("model") != model:
        raise ValueError(f"Recovery report model mismatch for {model}: {path}")
    stored_fingerprint = clean(report.get("reportFingerprint"))
    unsigned = {key: value for key, value in report.items() if key != "reportFingerprint"}
    if not stored_fingerprint or _canonical_sha(unsigned) != stored_fingerprint:
        raise ValueError(f"Recovery report fingerprint mismatch for {model}: {path}")


def _checkpoint_record(model: str, row: Mapping[str, Any], item: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "model": model,
        "reviewId": item["reviewId"],
        "lane": item["lane"],
        "identity": deepcopy(row["identity"]),
        "partition": row["partition"],
        "before": deepcopy(row["before"]),
        "after": deepcopy(row["after"]),
        "proposed": deepcopy(item["proposed"]),
        "provenance": deepcopy(item["provenance"]),
        "note": item.get("note", ""),
    }


def _bulk_approval_category(record: Mapping[str, Any]) -> str | None:
    lane = record["lane"]
    provenance = record["provenance"]
    if lane == "reviewed_plan_recovery":
        return "reviewed_plan_recovery"
    if lane == "comparator_copy" and provenance.get("bulkEligible") is True:
        return "exact_rpo_comparator_copy"
    if lane == "copy_split" and not provenance.get("flags"):
        return "unflagged_copy_split"
    if lane == "price":
        return "approved_standard_price_rule"
    if lane == "display_order" and provenance.get("sourceType") == "comparator_proposal":
        return "comparator_display_order"
    if lane == "id_repair":
        return "sequential_id_repair"
    return None


def _exception_group_signature(record: Mapping[str, Any]) -> dict[str, Any]:
    identity = record["identity"]
    model = record["model"]
    before = record["before"]
    lane = record["lane"]
    shared_zr = model in {"zr1", "zr1x"}
    identity_key = identity.get("rpo") or before.get("option_name") or identity.get("optionId")
    signature: dict[str, Any] = {
        "scope": "zr_shared" if shared_zr else model,
        "lane": lane,
        "identityKey": identity_key,
        "proposed": record["proposed"],
        "flags": sorted(record["provenance"].get("flags") or []),
    }
    if lane in {"copy_split", "identifying_copy_review", "comparator_copy_material_disagreement"}:
        signature["currentCopy"] = {
            field: before.get(field) for field in ("option_name", "description", "detail_raw")
        }
    elif lane == "active":
        signature["currentActive"] = before.get("active")
    elif lane == "display_order":
        signature["currentOrder"] = before.get("display_order")
        signature["proposedSection"] = record["after"].get("section_id")
    return signature


def _required_decision(lane: str) -> str:
    return {
        "copy_split": "Accept or override option_name and description.",
        "identifying_copy_review": "Accept or override the identifying option_name and non-repeating description.",
        "comparator_copy_material_disagreement": "Review target text; accept or override comparator copy.",
        "no_rpo_mapping": "Accept or override the comparator mapping, copy, and sequential ID.",
        "full_review": "Confirm copy, section, active, selectable, and order; override any field that differs.",
        "display_order": "Accept or override the proposed display order after confirming section placement.",
        "active": "Accept or override the proposed active flag.",
    }.get(lane, "Accept or override the proposed value.")


def _exception_groups(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    signatures: dict[str, dict[str, Any]] = {}
    for record in records:
        signature = _exception_group_signature(record)
        key = _canonical_sha(signature)
        signatures[key] = signature
        grouped.setdefault(key, []).append(record)

    result: list[dict[str, Any]] = []
    for key, members in grouped.items():
        members.sort(key=lambda item: (TARGET_MODELS.index(item["model"]), item["reviewId"]))
        models = sorted({member["model"] for member in members}, key=TARGET_MODELS.index)
        flags = sorted(
            {
                flag
                for member in members
                for flag in (member["provenance"].get("flags") or [])
            }
        )
        proposed_state: Any = deepcopy(members[0]["proposed"])
        if members[0]["lane"] == "full_review":
            proposed_state = {
                field: members[0]["after"].get(field)
                for field in (
                    "option_name",
                    "description",
                    "section_id",
                    "active",
                    "selectable",
                    "display_order",
                )
            }
        result.append(
            {
                "decisionGroupId": f"checkpoint-1:{key[:16]}",
                "lane": members[0]["lane"],
                "models": models,
                "reviewIds": [member["reviewId"] for member in members],
                "identityKey": signatures[key]["identityKey"],
                "proposed": deepcopy(members[0]["proposed"]),
                "proposedState": proposed_state,
                "flags": flags,
                "requiredDecision": _required_decision(members[0]["lane"]),
                "decision": {
                    "status": "pending",
                    "action": "accept_or_override",
                    "override": None,
                    "reviewer": None,
                    "reviewedAt": None,
                },
                "members": members,
            }
        )
    return sorted(
        result,
        key=lambda group: (
            group["lane"],
            group["identityKey"],
            group["decisionGroupId"],
        ),
    )


def _checkpoint_markdown(approval: Mapping[str, Any], packet: Mapping[str, Any]) -> str:
    def cell(value: Any, limit: int = 100) -> str:
        if value is None:
            return "blank"
        text = str(value).replace("\n", " ↵ ").replace("|", "\\|")
        return text if len(text) <= limit else text[: limit - 3] + "..."

    lines = [
        "# Checkpoint 1 Exception-Only Review",
        "",
        "> The safe bulk items are approved. Only the decisions in this packet remain pending. This is read-only and does not authorize a workbook write.",
        "",
        "## Approval summary",
        "",
        f"- Reviewer: `{approval['reviewer']}`",
        f"- Reviewed at: `{approval['reviewedAt']}`",
        f"- Workbook SHA-256: `{approval['workbookSha256']}`",
        f"- Bulk-approved review records: **{approval['approvedReviewCount']}**",
        f"- Pending exception records: **{packet['pendingReviewCount']}**",
        f"- Condensed decision groups: **{packet['exceptionGroupCount']}**",
        "",
        "### Bulk-approved categories",
        "",
        "| Category | Approved records |",
        "|---|---:|",
    ]
    for category, count in approval["approvedCategoryCounts"].items():
        lines.append(f"| {category} | {count} |")

    lane_titles = {
        "copy_split": "Flagged copy-split choices",
        "identifying_copy_review": "Identifying copy choices",
        "comparator_copy_material_disagreement": "Comparator-copy disagreements",
        "full_review": "Fresh GSX full review",
        "no_rpo_mapping": "GSX no-RPO mappings",
        "display_order": "Deterministic display order",
        "active": "Active-status proposals",
    }
    lanes = sorted({group["lane"] for group in packet["exceptionGroups"]})
    for lane in lanes:
        lines.extend(
            [
                "",
                f"## {lane_titles.get(lane, lane)}",
                "",
                "| Decision group | Models | RPO / option | Current | Proposed | Flags |",
                "|---|---|---|---|---|---|",
            ]
        )
        for group in packet["exceptionGroups"]:
            if group["lane"] != lane:
                continue
            member = group["members"][0]
            identity = member["identity"].get("rpo") or member["identity"].get("optionId")
            if lane in {"copy_split", "identifying_copy_review", "comparator_copy_material_disagreement", "no_rpo_mapping", "full_review"}:
                current = {
                    field: member["before"].get(field)
                    for field in ("option_name", "description", "section_id", "active", "selectable")
                }
            elif lane == "display_order":
                current = member["before"].get("display_order")
            elif lane == "active":
                current = member["before"].get("active")
            else:
                current = member["before"]
            lines.append(
                f"| `{group['decisionGroupId']}` | {', '.join(group['models'])} | {cell(identity)} | "
                f"{cell(json.dumps(current, sort_keys=True, ensure_ascii=False))} | "
                f"{cell(json.dumps(group['proposedState'], sort_keys=True, ensure_ascii=False))} | "
                f"{cell(', '.join(group['flags']) or 'none')} |"
            )
        lines.extend(
            [
                "",
                f"Decision required: {next(group['requiredDecision'] for group in packet['exceptionGroups'] if group['lane'] == lane)}",
            ]
        )

    lines.extend(
        [
            "",
            "Reply with each decision-group ID and either `accept` or the field override. Shared ZR groups apply to every model listed unless you explicitly split them.",
            "",
            "The paired JSON contains complete before/after values, provenance, member review IDs, and typed decision placeholders.",
            "",
        ]
    )
    return "\n".join(lines)


def generate_checkpoint_1_packet(
    report_dir: Path,
    current_workbook_path: Path,
    *,
    reviewer: str,
    reviewed_at: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Record the approved safe bulk set and emit the remaining exception review."""

    report_dir = Path(report_dir)
    current_workbook_path = Path(current_workbook_path)
    reviewer = clean(reviewer)
    reviewed_at = clean(reviewed_at)
    if not reviewer or not reviewed_at:
        raise ValueError("Checkpoint 1 approval requires reviewer and reviewed_at")

    reports: dict[str, dict[str, Any]] = {}
    report_bindings: dict[str, dict[str, str]] = {}
    workbook_sha = _sha256(current_workbook_path)
    for model in TARGET_MODELS:
        path = report_dir / f"{model}-recovery-projection.json"
        report = _json(path)
        _validate_checkpoint_report(report, model, path)
        bound_sha = clean((report.get("sources") or {}).get("workbookSha256"))
        if bound_sha != workbook_sha:
            raise ValueError(
                f"Current workbook SHA does not match recovery report for {model}: {path}"
            )
        reports[model] = report
        report_bindings[model] = {
            "path": str(path),
            "reportFingerprint": report["reportFingerprint"],
            "workbookSha256": bound_sha,
        }

    records: list[dict[str, Any]] = []
    seen_review_ids: set[str] = set()
    for model in TARGET_MODELS:
        for row in reports[model]["residualRows"]:
            for item in row["reviewItems"]:
                review_id = clean(item.get("reviewId"))
                if not review_id or review_id in seen_review_ids:
                    raise ValueError(f"Duplicate or blank Checkpoint 1 review ID: {review_id}")
                seen_review_ids.add(review_id)
                records.append(_checkpoint_record(model, row, item))

    approved_records: list[tuple[str, dict[str, Any]]] = []
    exception_records: list[dict[str, Any]] = []
    for record in records:
        category = _bulk_approval_category(record)
        if category:
            approved_records.append((category, record))
        else:
            exception_records.append(record)

    approved_category_counts: dict[str, int] = {}
    for category, _ in approved_records:
        approved_category_counts[category] = approved_category_counts.get(category, 0) + 1
    approved_ids = sorted(record["reviewId"] for _, record in approved_records)
    exception_ids = sorted(record["reviewId"] for record in exception_records)
    if set(approved_ids) & set(exception_ids) or set(approved_ids) | set(exception_ids) != seen_review_ids:
        raise ValueError("Checkpoint 1 review partition is incomplete or overlapping")

    approval: dict[str, Any] = {
        "schemaVersion": CHECKPOINT_1_APPROVAL_SCHEMA_VERSION,
        "status": "bulk_approved_exceptions_pending" if exception_ids else "approved",
        "reviewer": reviewer,
        "reviewedAt": reviewed_at,
        "workbookPath": str(current_workbook_path),
        "workbookSha256": workbook_sha,
        "sourceReports": report_bindings,
        "approvalPolicy": {
            "reviewedPlanRecovery": "approve_all",
            "exactRpoComparatorCopy": "approve_only_bulk_eligible",
            "copySplit": "approve_only_without_flags",
            "price": "approve_under_authoritative_standard_price_rule",
            "displayOrder": "approve_only_comparator_proposals",
            "idRepair": "approve_sequential_zr_repairs",
        },
        "approvedCategoryCounts": dict(sorted(approved_category_counts.items())),
        "approvedReviewCount": len(approved_ids),
        "approvedReviewIds": approved_ids,
        "exceptionReviewCount": len(exception_ids),
        "exceptionReviewIds": exception_ids,
    }
    approval["approvalFingerprint"] = _canonical_sha(approval)

    groups = _exception_groups(exception_records)
    packet: dict[str, Any] = {
        "schemaVersion": CHECKPOINT_1_EXCEPTION_SCHEMA_VERSION,
        "status": "pending_exception_review" if groups else "complete",
        "bulkApprovalFingerprint": approval["approvalFingerprint"],
        "workbookPath": str(current_workbook_path),
        "workbookSha256": workbook_sha,
        "sourceReports": report_bindings,
        "bulkApprovedReviewCount": len(approved_ids),
        "pendingReviewCount": len(exception_ids),
        "exceptionGroupCount": len(groups),
        "exceptionGroups": groups,
    }
    packet["packetFingerprint"] = _canonical_sha(packet)

    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / "checkpoint-1-bulk-approval.json").write_text(
        json.dumps(approval, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    (report_dir / "checkpoint-1-exception-review.json").write_text(
        json.dumps(packet, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    (report_dir / "checkpoint-1-exception-review.md").write_text(
        _checkpoint_markdown(approval, packet),
        encoding="utf-8",
    )
    return approval, packet


def _validate_fingerprinted_artifact(
    payload: Mapping[str, Any],
    fingerprint_field: str,
    label: str,
) -> str:
    stored = clean(payload.get(fingerprint_field))
    unsigned = {key: value for key, value in payload.items() if key != fingerprint_field}
    if not stored or _canonical_sha(unsigned) != stored:
        raise ValueError(f"{label} fingerprint mismatch")
    return stored


def _delete_scope(
    workbook: Any,
    source_sheets: Mapping[str, str],
    group: Mapping[str, Any],
) -> dict[str, Any]:
    if group["lane"] not in {"full_review", "identifying_copy_review"}:
        raise ValueError("Delete decisions require an option-row review group")

    option_rows: list[dict[str, Any]] = []
    owned_by_row: dict[tuple[str, int], dict[str, Any]] = {}
    member_targets: list[tuple[str, str, str, int]] = []
    for member in group["members"]:
        model = member["model"]
        identity = member["identity"]
        option_id = clean(identity.get("optionId"))
        target_sheet = source_sheets[model]
        row_number = int(identity.get("rowNumber"))
        member_targets.append((model, target_sheet, option_id, row_number))
        option_rows.append(
            {
                "sheet": target_sheet,
                "rowNumber": row_number,
                "optionId": option_id,
                "rpo": identity.get("rpo"),
            }
        )

    owned_prefixes = tuple(f"{model}_" for model, _, _, _ in member_targets)
    option_cells = {
        (sheet, row_number, "option_id")
        for _, sheet, _, row_number in member_targets
    }
    seen_refs: set[tuple[str, int, str, str, str]] = set()
    external_refs: list[dict[str, Any]] = []
    for _, target_sheet, option_id, _ in member_targets:
        for ref in _cascade_references(workbook, option_id, target_sheet):
            ref_key = (
                ref["sheet"],
                ref["rowNumber"],
                ref["column"],
                ref["value"],
                ref["matchType"],
            )
            if ref_key in seen_refs or (ref["sheet"], ref["rowNumber"], ref["column"]) in option_cells:
                continue
            seen_refs.add(ref_key)
            if ref["sheet"].startswith(owned_prefixes):
                key = (ref["sheet"], ref["rowNumber"])
                row = owned_by_row.setdefault(
                    key,
                    {
                        "sheet": ref["sheet"],
                        "rowNumber": ref["rowNumber"],
                        "matches": [],
                    },
                )
                row["matches"].append(
                    {
                        "column": ref["column"],
                        "value": ref["value"],
                        "matchType": ref["matchType"],
                    }
                )
            else:
                external_refs.append(ref)
    owned_rows = sorted(owned_by_row.values(), key=lambda row: (row["sheet"], row["rowNumber"]))
    for row in owned_rows:
        row["matches"].sort(key=lambda match: (match["column"], match["value"]))
    return {
        "optionRows": sorted(option_rows, key=lambda row: (row["sheet"], row["rowNumber"])),
        "ownedReferenceRows": owned_rows,
        "ovsRowCount": sum(1 for row in owned_rows if row["sheet"] == "grand_sport_x_ovs"),
        "additionalOwnedReferenceRowCount": sum(
            1 for row in owned_rows if row["sheet"] != "grand_sport_x_ovs"
        ),
        "externalSameIdReferencesPreserved": sorted(
            external_refs,
            key=lambda ref: (ref["sheet"], ref["rowNumber"], ref["column"]),
        ),
    }


def _full_pending_markdown(
    decision_artifact: Mapping[str, Any],
    pending: Mapping[str, Any],
) -> str:
    lane_titles = {
        "copy_split": "Flagged copy choices",
        "identifying_copy_review": "Identifying copy choices",
        "comparator_copy_material_disagreement": "Comparator-copy disagreements",
        "full_review": "Fresh GSX full review",
        "no_rpo_mapping": "GSX no-RPO mappings",
        "display_order": "Deterministic display order",
        "active": "Active-status proposals",
    }
    lines = [
        "# Checkpoint 1 Pending Exception Review",
        "",
        "> Complete, unabridged current and proposed values. This remains read-only and does not authorize a workbook write.",
        "",
        "## Decision status",
        "",
        f"- Recorded decision groups: **{decision_artifact['resolvedGroupCount']}**",
        f"- Recorded review records: **{decision_artifact['resolvedReviewCount']}**",
        f"- Recorded bulk-approval overrides: **{decision_artifact.get('bulkApprovalOverrideCount', 0)}**",
        f"- Pending decision groups: **{pending['pendingGroupCount']}**",
        f"- Pending review records: **{pending['pendingReviewCount']}**",
        f"- Decision artifact fingerprint: `{decision_artifact['decisionArtifactFingerprint']}`",
    ]
    for lane in sorted({group["lane"] for group in pending["exceptionGroups"]}):
        lines.extend(["", f"## {lane_titles.get(lane, lane)}", ""])
        for group in pending["exceptionGroups"]:
            if group["lane"] != lane:
                continue
            lines.extend(
                [
                    f"### `{group['decisionGroupId']}` — {', '.join(group['models'])} — {group['identityKey']}",
                    "",
                    group["requiredDecision"],
                    "",
                    f"Flags: `{', '.join(group['flags']) or 'none'}`",
                    "",
                ]
            )
            for member in group["members"]:
                identity = member["identity"]
                lines.extend(
                    [
                        f"#### {member['model']} — RPO `{identity.get('rpo') or 'none'}` — option `{identity.get('optionId')}`",
                        "",
                        "Current state:",
                        "",
                        "```json",
                        json.dumps(member["before"], indent=2, sort_keys=True, ensure_ascii=False),
                        "```",
                        "",
                        "Proposed state:",
                        "",
                        "```json",
                        json.dumps(group["proposedState"], indent=2, sort_keys=True, ensure_ascii=False),
                        "```",
                        "",
                        "Provenance:",
                        "",
                        "```json",
                        json.dumps(member["provenance"], indent=2, sort_keys=True, ensure_ascii=False),
                        "```",
                        "",
                    ]
                )
    lines.extend(
        [
            "Reply with a decision-group ID and `accept`, or provide the exact field override. Shared ZR groups apply to all listed models unless explicitly split.",
            "",
        ]
    )
    return "\n".join(lines)


def record_checkpoint_1_decisions(
    report_dir: Path,
    current_workbook_path: Path,
    *,
    decisions: Iterable[Mapping[str, Any]],
    bulk_overrides: Iterable[Mapping[str, Any]] = (),
    reviewer: str,
    reviewed_at: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Record typed exception decisions without mutating their source packet."""

    report_dir = Path(report_dir)
    current_workbook_path = Path(current_workbook_path)
    reviewer = clean(reviewer)
    reviewed_at = clean(reviewed_at)
    if not reviewer or not reviewed_at:
        raise ValueError("Checkpoint 1 decisions require reviewer and reviewed_at")

    approval = _json(report_dir / "checkpoint-1-bulk-approval.json")
    packet = _json(report_dir / "checkpoint-1-exception-review.json")
    approval_fingerprint = _validate_fingerprinted_artifact(
        approval,
        "approvalFingerprint",
        "Checkpoint 1 bulk approval",
    )
    packet_fingerprint = _validate_fingerprinted_artifact(
        packet,
        "packetFingerprint",
        "Checkpoint 1 exception packet",
    )
    if packet.get("bulkApprovalFingerprint") != approval_fingerprint:
        raise ValueError("Checkpoint 1 exception packet is not bound to the bulk approval")
    workbook_sha = _sha256(current_workbook_path)
    if packet.get("workbookSha256") != workbook_sha or approval.get("workbookSha256") != workbook_sha:
        raise ValueError("Current workbook SHA does not match Checkpoint 1 artifacts")

    groups_by_id = {group["decisionGroupId"]: group for group in packet["exceptionGroups"]}
    raw_decisions = [dict(decision) for decision in decisions]
    raw_bulk_overrides = [dict(override) for override in bulk_overrides]
    supplied_ids = [clean(decision.get("decisionGroupId")) for decision in raw_decisions]
    if len(set(supplied_ids)) != len(supplied_ids) or any(not value for value in supplied_ids):
        raise ValueError("Checkpoint 1 decision-group IDs must be nonblank and unique")
    missing = sorted(set(supplied_ids) - set(groups_by_id))
    if missing:
        raise ValueError(f"Unknown Checkpoint 1 decision groups: {', '.join(missing)}")

    approved_ids = set(approval.get("approvedReviewIds") or [])
    approved_records: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for model, binding in (approval.get("sourceReports") or {}).items():
        report_path = Path(binding["path"])
        report = _json(report_path)
        _validate_checkpoint_report(report, model, report_path)
        if report["reportFingerprint"] != binding.get("reportFingerprint"):
            raise ValueError(f"Checkpoint 1 source report binding mismatch for {model}")
        for row in report["residualRows"]:
            for item in row["reviewItems"]:
                if item["reviewId"] not in approved_ids:
                    continue
                record = _checkpoint_record(model, row, item)
                key = (model, clean(row["identity"].get("rpo")).upper(), item["lane"])
                approved_records.setdefault(key, []).append(record)

    recorded_bulk_overrides: list[dict[str, Any]] = []
    seen_bulk_keys: set[tuple[str, str, str]] = set()
    for instruction in raw_bulk_overrides:
        key = (
            clean(instruction.get("model")).lower(),
            clean(instruction.get("rpo")).upper(),
            clean(instruction.get("lane")),
        )
        if key in seen_bulk_keys:
            raise ValueError(f"Duplicate Checkpoint 1 bulk override: {'/'.join(key)}")
        seen_bulk_keys.add(key)
        matches = approved_records.get(key) or []
        if len(matches) != 1:
            raise ValueError(
                f"Checkpoint 1 bulk override must match exactly one approved review: {'/'.join(key)}"
            )
        override = instruction.get("override")
        if not isinstance(override, Mapping) or not override:
            raise ValueError(f"Checkpoint 1 bulk override is empty: {'/'.join(key)}")
        unknown_fields = set(override) - set(OPTION_VIEW_FIELDS)
        if unknown_fields:
            raise ValueError(
                f"Unsupported Checkpoint 1 bulk override fields: {', '.join(sorted(unknown_fields))}"
            )
        record = matches[0]
        typed_override = {
            "reviewId": record["reviewId"],
            "model": record["model"],
            "identity": deepcopy(record["identity"]),
            "lane": record["lane"],
            "action": "override_bulk_approval",
            "override": deepcopy(dict(override)),
            "reviewer": reviewer,
            "reviewedAt": reviewed_at,
            "sourceApprovalFingerprint": approval_fingerprint,
        }
        typed_override["decisionFingerprint"] = _canonical_sha(typed_override)
        recorded_bulk_overrides.append(typed_override)
    recorded_bulk_overrides.sort(key=lambda item: (item["model"], item["identity"]["rpo"], item["reviewId"]))

    allowed_actions = {
        "accept",
        "override",
        "override_by_model",
        "delete_option_and_owned_references",
        "not_applicable_due_to_delete",
    }
    delete_identity_keys = {
        groups_by_id[decision["decisionGroupId"]]["identityKey"]
        for decision in raw_decisions
        if decision.get("action") == "delete_option_and_owned_references"
    }
    workbook = load_workbook(current_workbook_path, read_only=True, data_only=True)
    try:
        source_sheets = _source_option_sheets(workbook)
        recorded: list[dict[str, Any]] = []
        for instruction in raw_decisions:
            group_id = instruction["decisionGroupId"]
            group = groups_by_id[group_id]
            action = clean(instruction.get("action"))
            if action not in allowed_actions:
                raise ValueError(f"Unsupported Checkpoint 1 decision action: {action}")
            if action == "override" and "override" not in instruction:
                raise ValueError(f"Override decision is missing its value: {group_id}")
            if action == "override_by_model":
                overrides = instruction.get("overrideByModel")
                expected_models = set(group["models"])
                if not isinstance(overrides, Mapping) or set(overrides) != expected_models:
                    raise ValueError(
                        f"Model-scoped override must cover exactly {', '.join(sorted(expected_models))}: {group_id}"
                    )
                for model, override in overrides.items():
                    if not isinstance(override, Mapping) or not override:
                        raise ValueError(f"Model-scoped override is empty for {model}: {group_id}")
                    unknown_fields = set(override) - set(OPTION_VIEW_FIELDS)
                    if unknown_fields:
                        raise ValueError(
                            f"Unsupported model-scoped override fields for {model}: {', '.join(sorted(unknown_fields))}"
                        )
            if action == "not_applicable_due_to_delete" and group["identityKey"] not in delete_identity_keys:
                raise ValueError(f"Not-applicable decision has no matching delete: {group_id}")

            decision: dict[str, Any] = {
                "decisionGroupId": group_id,
                "reviewIds": list(group["reviewIds"]),
                "models": list(group["models"]),
                "identityKey": group["identityKey"],
                "lane": group["lane"],
                "action": action,
                "reviewer": reviewer,
                "reviewedAt": reviewed_at,
            }
            if action == "accept":
                decision["acceptedState"] = deepcopy(group["proposedState"])
            elif action == "override":
                decision["override"] = deepcopy(instruction["override"])
            elif action == "override_by_model":
                decision["overrideByModel"] = deepcopy(instruction["overrideByModel"])
            elif action == "delete_option_and_owned_references":
                decision["deleteScope"] = _delete_scope(workbook, source_sheets, group)
            else:
                decision["reason"] = "option_deleted_by_checkpoint_1_decision"
            decision["decisionFingerprint"] = _canonical_sha(decision)
            recorded.append(decision)
    finally:
        workbook.close()

    recorded.sort(key=lambda decision: decision["decisionGroupId"])
    resolved_ids = set(supplied_ids)
    pending_groups = [
        deepcopy(group)
        for group in packet["exceptionGroups"]
        if group["decisionGroupId"] not in resolved_ids
    ]
    resolved_review_ids = sorted(
        review_id for decision in recorded for review_id in decision["reviewIds"]
    )
    pending_review_ids = sorted(
        review_id for group in pending_groups for review_id in group["reviewIds"]
    )
    if set(resolved_review_ids) & set(pending_review_ids):
        raise ValueError("Resolved and pending Checkpoint 1 review IDs overlap")

    decision_artifact: dict[str, Any] = {
        "schemaVersion": CHECKPOINT_1_DECISIONS_SCHEMA_VERSION,
        "status": "exceptions_pending" if pending_groups else "complete",
        "reviewer": reviewer,
        "reviewedAt": reviewed_at,
        "workbookPath": str(current_workbook_path),
        "workbookSha256": workbook_sha,
        "bulkApprovalFingerprint": approval_fingerprint,
        "sourcePacketFingerprint": packet_fingerprint,
        "resolvedGroupCount": len(recorded),
        "resolvedReviewCount": len(resolved_review_ids),
        "resolvedReviewIds": resolved_review_ids,
        "pendingGroupCount": len(pending_groups),
        "pendingReviewCount": len(pending_review_ids),
        "decisions": recorded,
        "bulkApprovalOverrideCount": len(recorded_bulk_overrides),
        "bulkApprovalOverrides": recorded_bulk_overrides,
    }
    decision_artifact["decisionArtifactFingerprint"] = _canonical_sha(decision_artifact)

    pending: dict[str, Any] = {
        "schemaVersion": CHECKPOINT_1_PENDING_SCHEMA_VERSION,
        "status": "pending_exception_review" if pending_groups else "complete",
        "bulkApprovalFingerprint": approval_fingerprint,
        "sourcePacketFingerprint": packet_fingerprint,
        "decisionArtifactFingerprint": decision_artifact["decisionArtifactFingerprint"],
        "workbookPath": str(current_workbook_path),
        "workbookSha256": workbook_sha,
        "resolvedGroupCount": len(recorded),
        "resolvedReviewCount": len(resolved_review_ids),
        "pendingGroupCount": len(pending_groups),
        "pendingReviewCount": len(pending_review_ids),
        "bulkApprovalOverrideCount": len(recorded_bulk_overrides),
        "exceptionGroups": pending_groups,
    }
    pending["pendingPacketFingerprint"] = _canonical_sha(pending)

    (report_dir / "checkpoint-1-exception-decisions.json").write_text(
        json.dumps(decision_artifact, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    (report_dir / "checkpoint-1-pending-review.json").write_text(
        json.dumps(pending, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    (report_dir / "checkpoint-1-pending-review.md").write_text(
        _full_pending_markdown(decision_artifact, pending),
        encoding="utf-8",
    )
    return decision_artifact, pending


def generate_recovery_reports(inputs: ProjectionInputs, output_dir: Path) -> dict[str, dict[str, Any]]:
    """Generate one Markdown/JSON residual report per target model."""

    inputs = ProjectionInputs(**{field: Path(getattr(inputs, field)) for field in inputs.__dataclass_fields__})
    reviewed_plan = _json(inputs.reviewed_plan_path)
    reviewed_candidates = _json(inputs.reviewed_candidates_path)
    reconciliation_candidates = _json(inputs.reconciliation_candidates_path)
    reviewed_decisions = _json(inputs.reviewed_decisions_path)
    exception_queue = _json(inputs.exception_queue_path)
    exception_resolutions = _json(inputs.exception_resolutions_path)
    canonical_manifest = _json(inputs.canonical_manifest_path)
    comparator_evidence = _json(inputs.comparator_evidence_path)

    workbook_sha = _sha256(inputs.workbook_path)
    pre_sha = _sha256(inputs.pre_integration_workbook_path)
    workbook = load_workbook(inputs.workbook_path, read_only=True, data_only=True)
    pre_workbook = load_workbook(inputs.pre_integration_workbook_path, read_only=True, data_only=True)
    try:
        source_sheets = _source_option_sheets(workbook)
        section_modes = _section_modes(workbook)
        options: dict[str, list[dict[str, Any]]] = {}
        options_index: dict[str, tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]] = {}
        for model, sheet in source_sheets.items():
            model_rows = [_option_row(row) for row in _rows(workbook, sheet)]
            options[model] = model_rows
            options_index[model] = _index_options(model_rows)
        for model in TARGET_MODELS:
            if model not in options:
                raise ValueError(f"Missing source_option_sheet for {model}")

        comparator_models = {
            model: clean(((comparator_evidence.get("targets") or {}).get(model) or {}).get("comparator")).lower()
            for model in TARGET_MODELS
        }
        for model, comparator in comparator_models.items():
            if comparator not in options:
                raise ValueError(f"Missing comparator option sheet for {model}: {comparator}")

        pre_options: dict[str, list[dict[str, Any]]] = {}
        for model in ("zr1", "zr1x"):
            sheet = source_sheets[model]
            pre_options[model] = [_option_row(row) for row in _rows(pre_workbook, sheet)]

        reconciliation, reconciliation_checks = _section_reconciliation(
            exception_queue,
            exception_resolutions,
            canonical_manifest,
            options_index,
        )
        reconciliation_by_identity: dict[str, dict[tuple[str, str], dict[str, Any]]] = {
            model: {
                (item["rpo"], item["optionId"]): item
                for item in items
            }
            for model, items in reconciliation.items()
        }
        target_all_unavailable = {
            model: _target_all_unavailable_by_rpo(reconciliation_candidates, model)
            for model in ("zr1", "zr1x")
        }

        gsx_by_id, gsx_by_rpo = options_index["grand_sport_x"]
        valid_plan, invalid_plan = _plan_rows(
            reviewed_plan,
            reviewed_candidates,
            reviewed_decisions,
            gsx_by_rpo,
        )
        plan_rpos = {
            clean((item.get("row") or {}).get("rpo")).upper()
            for item in (reviewed_plan.get("stage2") or {}).get("items", [])
            if clean((item.get("row") or {}).get("rpo"))
        }

        reports: dict[str, dict[str, Any]] = {}
        for model in TARGET_MODELS:
            sheet = source_sheets[model]
            current_rows = options[model]
            current_by_id, current_by_rpo = options_index[model]
            comparator_model = comparator_models[model]
            comparator_sheet = source_sheets[comparator_model]
            comparator_rows = options[comparator_model]
            comparator_by_id, comparator_by_rpo = options_index[comparator_model]

            fallback_sets: list[tuple[str, str, list[dict[str, Any]]]] = [
                (comparator_model, comparator_sheet, comparator_rows)
            ]
            for fallback_model in ("z06", "grand_sport", "stingray"):
                if fallback_model != comparator_model and fallback_model in options:
                    fallback_sets.append((fallback_model, source_sheets[fallback_model], options[fallback_model]))

            hex_rows = [row for row in current_rows if HASH_OPTION_ID_RE.match(row["option_id"])]
            proposed_ids = _next_option_ids(current_rows, len(hex_rows))
            id_by_old = {row["option_id"]: proposed for row, proposed in zip(hex_rows, proposed_ids)}
            z06_no_rpo_set = [("z06", source_sheets["z06"], options["z06"])]
            no_rpo_matches: dict[str, tuple[str, str, dict[str, Any], float] | None] = {
                row["option_id"]: _best_no_rpo_comparator(row, z06_no_rpo_set)
                for row in current_rows
                if not row["rpo"]
            }
            no_rpo_order_matches: dict[str, tuple[str, str, dict[str, Any], float] | None] = {
                row["option_id"]: _best_no_rpo_comparator(row, fallback_sets)
                for row in current_rows
                if not row["rpo"]
            }

            pre_by_id: dict[str, dict[str, Any]] = {}
            pre_by_rpo: dict[str, dict[str, Any]] = {}
            if model in pre_options:
                pre_by_id, pre_by_rpo = _index_options(pre_options[model])

            row_states: list[dict[str, Any]] = []
            rows_needing_order: list[tuple[dict[str, Any], dict[str, Any] | None, Mapping[str, Any]]] = []
            target_applicability_deletions: list[dict[str, Any]] = []
            recorded_instruction_deletions: list[dict[str, Any]] = []
            for current in current_rows:
                if model in {"zr1", "zr1x"} and current["rpo"] in FRESH_GSX_RPOS:
                    recorded_instruction_deletions.append(
                        _deletion_projection(
                            workbook,
                            model,
                            sheet,
                            current,
                            "checkpoint_1_delete_rpo_across_targets",
                        )
                    )
                    continue
                unavailable_evidence = target_all_unavailable.get(model, {}).get(current["rpo"])
                if unavailable_evidence is not None:
                    target_applicability_deletions.append(
                        _deletion_projection(
                            workbook,
                            model,
                            sheet,
                            current,
                            "all_target_variant_statuses_unavailable",
                            **unavailable_evidence,
                        )
                    )
                    continue
                before = _option_view(current)
                after = deepcopy(before)
                provenance: dict[str, Any] = {}
                review_items: list[dict[str, Any]] = []
                comparator = comparator_by_rpo.get(current["rpo"]) if current["rpo"] else None
                comparator_ref = _comparator_ref(comparator_model, comparator_sheet, comparator) if comparator else None

                if model == "grand_sport_x":
                    if not current["rpo"]:
                        partition = "z06_no_rpo_copy_recovery"
                        match = no_rpo_matches.get(current["option_id"])
                        if match is None:
                            raise ValueError(
                                f"No Z06 no-RPO copy match for {model} {current['option_id']}"
                            )
                        source = {
                            "sourceType": "exact_z06_no_rpo_copy",
                            "reference": _comparator_ref(match[0], match[1], match[2], match[3]),
                            "bulkEligible": True,
                        }
                        _field_change(after, provenance, "option_id", id_by_old[current["option_id"]], source)
                        _field_change(after, provenance, "option_name", match[2]["option_name"], source)
                        _field_change(after, provenance, "description", match[2]["description"], source)
                        order_match = no_rpo_order_matches.get(current["option_id"])
                        rows_needing_order.append(
                            (
                                after,
                                order_match[2] if order_match else None,
                                {
                                    "sourceType": "deterministic_section_local",
                                    "reference": (
                                        _comparator_ref(
                                            order_match[0],
                                            order_match[1],
                                            order_match[2],
                                            order_match[3],
                                        )
                                        if order_match
                                        else None
                                    ),
                                },
                            )
                        )
                    elif current["rpo"] in FRESH_GSX_RPOS:
                        partition = "fresh_rpo_full_review"
                        split = propose_copy_split({"description": current["detail_raw"], "statuses": []})
                        split_source = {"sourceType": "copy_split", "flags": split["flags"]}
                        _field_change(after, provenance, "option_name", split["name"], split_source)
                        _field_change(after, provenance, "description", split["description"], split_source)
                        rows_needing_order.append((after, comparator, comparator_ref or {"sourceType": "deterministic_section_local"}))
                        if comparator:
                            _field_change(after, provenance, "selectable", comparator["selectable"], {"sourceType": "comparator_proposal", "reference": comparator_ref})
                        review_items.append(
                            _review_item(
                                model,
                                current,
                                "full_review",
                                {"placement": after["section_id"], "active": after["active"], "selectable": after["selectable"]},
                                {"sourceType": "fresh_rpo", "reference": comparator_ref},
                                note="Fresh GSX RPO absent from the July 9 reviewed plan.",
                            )
                        )
                    elif current["rpo"] in plan_rpos:
                        if current["rpo"] in valid_plan:
                            partition = "reviewed_plan_recovery"
                            planned = valid_plan[current["rpo"]]
                            source = {
                                "sourceType": "reviewed_plan",
                                "path": str(inputs.reviewed_plan_path),
                                "candidateFingerprintStatus": "matched",
                            }
                            for field in ("option_name", "description", "section_id", "display_order"):
                                value = planned.get(field)
                                if field in {"option_name", "description", "section_id"}:
                                    value = clean(value)
                                elif value is not None:
                                    value = intish(value)
                                _field_change(after, provenance, field, value, source)
                            if any(before.get(field) != after.get(field) for field in ("option_name", "description", "section_id", "display_order")):
                                review_items.append(
                                    _review_item(
                                        model,
                                        current,
                                        "reviewed_plan_recovery",
                                        {field: after[field] for field in ("option_name", "description", "section_id", "display_order")},
                                        source,
                                    )
                                )
                        else:
                            partition = "reviewed_plan_fingerprint_mismatch"
                            review_items.append(
                                _review_item(
                                    model,
                                    current,
                                    "full_review",
                                    _option_view(current),
                                    {"sourceType": "reviewed_plan_rejected", "reasons": invalid_plan.get(current["rpo"], ["missing_plan_row"])},
                                )
                            )
                    else:
                        partition = "paint_keep_current"
                else:
                    prior = pre_by_id.get(current["option_id"]) or (pre_by_rpo.get(current["rpo"]) if current["rpo"] else None)
                    partition = "forward_copy_repair" if prior else "new_option_review"
                    if not current["rpo"]:
                        match = no_rpo_matches.get(current["option_id"])
                        if match is None:
                            raise ValueError(
                                f"No Z06 no-RPO copy match for {model} {current['option_id']}"
                            )
                        partition = "z06_no_rpo_copy_recovery"
                        source = {
                            "sourceType": "exact_z06_no_rpo_copy",
                            "reference": _comparator_ref(match[0], match[1], match[2], match[3]),
                            "bulkEligible": True,
                        }
                        _field_change(after, provenance, "option_name", match[2]["option_name"], source)
                        _field_change(after, provenance, "description", match[2]["description"], source)
                    elif comparator:
                        comparison = _comparator_copy_comparison(current, comparator)
                        bulk_eligible = not comparison["materialDisagreement"]
                        source = {
                            "sourceType": "exact_rpo_comparator_copy",
                            "reference": comparator_ref,
                            "comparison": comparison,
                            "bulkEligible": bulk_eligible,
                        }
                        _field_change(after, provenance, "option_name", comparator["option_name"], source)
                        _field_change(after, provenance, "description", comparator["description"], source)
                        review_items.append(
                            _review_item(
                                model,
                                current,
                                "comparator_copy" if bulk_eligible else "comparator_copy_material_disagreement",
                                {"option_name": after["option_name"], "description": after["description"]},
                                source,
                            )
                        )
                    else:
                        proposal = _identifying_copy_proposal(current)
                        source = {
                            "sourceType": proposal["sourceType"],
                            "flags": proposal["flags"],
                            "reference": proposal.get("reference"),
                            "bulkEligible": False,
                        }
                        _field_change(after, provenance, "option_name", proposal["option_name"], source)
                        _field_change(after, provenance, "description", proposal["description"], source)
                        review_items.append(
                            _review_item(
                                model,
                                current,
                                "identifying_copy_review",
                                {"option_name": after["option_name"], "description": after["description"]},
                                source,
                            )
                        )
                    if prior is None:
                        rows_needing_order.append((after, comparator, comparator_ref or {"sourceType": "deterministic_section_local"}))
                    if current["option_id"] in id_by_old:
                        id_source = {"sourceType": "sequential_id_allocation", "reservedAgainst": sheet}
                        _field_change(after, provenance, "option_id", id_by_old[current["option_id"]], id_source)
                        review_items.append(
                            _review_item(model, current, "id_repair", after["option_id"], id_source)
                        )

                reconciliation_item = reconciliation_by_identity[model].get((current["rpo"], current["option_id"]))
                if reconciliation_item:
                    section_source = {
                        "sourceType": "resolved_choose_section",
                        "subjectId": reconciliation_item["subjectId"],
                        "subjectVersion": reconciliation_item["subjectVersion"],
                        "reviewer": reconciliation_item["reviewer"],
                    }
                    _field_change(after, provenance, "section_id", reconciliation_item["decidedSectionId"], section_source)
                    review_items.append(
                        _review_item(model, current, "section_reconciliation", after["section_id"], section_source)
                    )

                row_states.append(
                    {
                        "current": current,
                        "before": before,
                        "after": after,
                        "partition": partition,
                        "fieldProvenance": provenance,
                        "reviewItems": review_items,
                        "comparator": comparator,
                        "comparatorRef": comparator_ref,
                    }
                )

            used_orders: dict[str, set[int]] = {}
            pending_ids = {id(after) for after, _, _ in rows_needing_order}
            for state in row_states:
                after = state["after"]
                if id(after) in pending_ids or after["display_order"] is None:
                    continue
                used_orders.setdefault(after["section_id"], set()).add(after["display_order"])
            for after, comparator, source in rows_needing_order:
                used = used_orders.setdefault(after["section_id"], set())
                comparator_order = comparator["display_order"] if comparator and comparator["section_id"] == after["section_id"] else None
                proposed_order, tier = _allocate_order(used, comparator_order)
                order_source = {
                    "sourceType": tier,
                    "reference": source.get("reference") if isinstance(source, Mapping) else None,
                }
                state = next(state for state in row_states if state["after"] is after)
                _field_change(after, state["fieldProvenance"], "display_order", proposed_order, order_source)
                state["reviewItems"].append(
                    _review_item(model, state["current"], "display_order", proposed_order, order_source)
                )

            for state in row_states:
                current = state["current"]
                after = state["after"]
                if state["partition"] != "paint_keep_current":
                    mode = section_modes.get(after["section_id"], "")
                    if not after["selectable"]:
                        rule_price = None if mode == "display_only" else 0
                        if current["rpo"] in MANDATORY_CHARGE_RPOS and current["price"] is not None:
                            proposed_price = current["price"]
                            classification = "mandatory_charge_exception"
                        else:
                            proposed_price = rule_price
                            classification = "standard_price_rule"
                        if current["price"] != proposed_price or classification == "mandatory_charge_exception":
                            price_source = {
                                "sourceType": classification,
                                "selectionMode": mode,
                                "comparatorReference": state["comparatorRef"],
                            }
                            _field_change(after, state["fieldProvenance"], "price", proposed_price, price_source)
                            state["reviewItems"].append(
                                _review_item(model, current, "price", proposed_price, price_source)
                            )

                    comparator = state["comparator"]
                    if not current["rpo"]:
                        match = no_rpo_matches.get(current["option_id"])
                        comparator = match[2] if match else None
                        comparator_reference = _comparator_ref(match[0], match[1], match[2], match[3]) if match else None
                    else:
                        comparator_reference = state["comparatorRef"]
                    if comparator and (after["active"] != comparator["active"] or state["partition"] == "fresh_rpo_full_review"):
                        active_source = {"sourceType": "comparator_proposal", "reference": comparator_reference}
                        _field_change(after, state["fieldProvenance"], "active", comparator["active"], active_source)
                        state["reviewItems"].append(
                            _review_item(model, current, "active", comparator["active"], active_source)
                        )

            residual_rows: list[dict[str, Any]] = []
            for state in row_states:
                if state["before"] == state["after"] and not state["reviewItems"]:
                    continue
                current = state["current"]
                residual_rows.append(
                    {
                        "identity": {
                            "optionId": current["option_id"],
                            "rpo": current["rpo"],
                            "rowNumber": current["row_number"],
                        },
                        "partition": state["partition"],
                        "before": state["before"],
                        "after": state["after"],
                        "fieldProvenance": state["fieldProvenance"],
                        "reviewItems": state["reviewItems"],
                    }
                )

            id_repairs = [
                {
                    "oldOptionId": row["option_id"],
                    "proposedOptionId": id_by_old[row["option_id"]],
                    "targetSheet": sheet,
                    "targetRowNumber": row["row_number"],
                    "cascade": _cascade_references(workbook, row["option_id"], sheet),
                }
                for row in hex_rows
            ]

            partition_counts: dict[str, int] = {}
            for state in row_states:
                partition_counts[state["partition"]] = partition_counts.get(state["partition"], 0) + 1
            pending_count = sum(len(row["reviewItems"]) for row in residual_rows)
            sources = {
                "workbookPath": str(inputs.workbook_path),
                "workbookSha256": workbook_sha,
                "reviewedPlanPath": str(inputs.reviewed_plan_path),
                "targetApplicabilityCandidatesPath": str(inputs.reconciliation_candidates_path),
                "targetApplicabilityCandidatesSha256": _sha256(inputs.reconciliation_candidates_path),
                "sectionResolutionPath": str(inputs.exception_resolutions_path),
                "targetSheet": sheet,
                "comparatorModel": comparator_model,
                "comparatorSheet": comparator_sheet,
            }
            if model in pre_options:
                sources.update(
                    {
                        "preIntegrationUsage": "row_existence_only",
                        "preIntegrationWorkbookSha256": pre_sha,
                    }
                )
            report: dict[str, Any] = {
                "schemaVersion": SCHEMA_VERSION,
                "model": model,
                "status": "pending_checkpoint_1",
                "readOnly": True,
                "sources": sources,
                "summary": {
                    "sourceRowCount": len(current_rows),
                    "targetApplicabilityDeletionCount": len(target_applicability_deletions),
                    "recordedInstructionDeletionCount": len(recorded_instruction_deletions),
                    "residualRowCount": len(residual_rows),
                    "pendingReviewCount": pending_count,
                    "sectionMismatchCount": len(reconciliation[model]),
                    "idRepairCount": len(id_repairs),
                    **{f"partition.{key}": value for key, value in sorted(partition_counts.items())},
                },
                "sectionReconciliation": reconciliation[model],
                "sectionReconciliationCheck": reconciliation_checks[model],
                "idRepairs": id_repairs,
                "targetApplicabilityDeletions": target_applicability_deletions,
                "recordedInstructionDeletions": recorded_instruction_deletions,
                "residualRows": residual_rows,
                "bulkDecisions": {
                    "comparatorDisplayOrder": {
                        "status": "pending",
                        "action": "accept_or_override",
                        "reviewIds": [
                            item["reviewId"]
                            for row in residual_rows
                            for item in row["reviewItems"]
                            if item["lane"] == "display_order"
                            and item["provenance"].get("sourceType") == "comparator_proposal"
                        ],
                    },
                    "comparatorCopy": {
                        "status": "pending",
                        "action": "accept_or_override",
                        "reviewIds": [
                            item["reviewId"]
                            for row in residual_rows
                            for item in row["reviewItems"]
                            if item["lane"] == "comparator_copy"
                        ],
                        "excludedMaterialDisagreementReviewIds": [
                            item["reviewId"]
                            for row in residual_rows
                            for item in row["reviewItems"]
                            if item["lane"] == "comparator_copy_material_disagreement"
                        ],
                    },
                    "copySplit": {
                        "status": "pending",
                        "action": "accept_or_override",
                        "reviewIds": [
                            item["reviewId"]
                            for row in residual_rows
                            for item in row["reviewItems"]
                            if item["lane"] == "copy_split"
                        ],
                    },
                },
            }
            report["reportFingerprint"] = _canonical_sha(report)
            reports[model] = report

        output_dir.mkdir(parents=True, exist_ok=True)
        for model, report in reports.items():
            json_path = output_dir / f"{model}-recovery-projection.json"
            markdown_path = output_dir / f"{model}-recovery-projection.md"
            json_path.write_text(json.dumps(report, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")
            markdown_path.write_text(_markdown(report), encoding="utf-8")
        return reports
    finally:
        pre_workbook.close()
        workbook.close()


def _default_inputs(root: Path, pre_integration_workbook_path: Path) -> ProjectionInputs:
    reviewed_run = root / "form-output/ingest-wizard/20260709-184223-960eb1"
    reconciliation_run = root / "form-output/ingest-wizard/20260717-091317-470292"
    return ProjectionInputs(
        workbook_path=root / "stingray_master.xlsx",
        pre_integration_workbook_path=pre_integration_workbook_path,
        reviewed_plan_path=reviewed_run / "apply-plan.json",
        reviewed_candidates_path=reviewed_run / "option-candidates.json",
        reconciliation_candidates_path=reconciliation_run / "option-candidates.json",
        reviewed_decisions_path=reviewed_run / "decisions.json",
        exception_queue_path=reconciliation_run / "exception-queue.json",
        exception_resolutions_path=reconciliation_run / "exception-resolutions.json",
        canonical_manifest_path=reconciliation_run / "canonical-row-manifest.json",
        comparator_evidence_path=reconciliation_run / "comparator-evidence.json",
    )


def _extract_git_workbook(root: Path, revision: str, destination: Path) -> None:
    with destination.open("wb") as handle:
        result = subprocess.run(
            ["git", "show", revision],
            cwd=root,
            stdout=handle,
            stderr=subprocess.PIPE,
            check=False,
        )
    if result.returncode:
        raise RuntimeError(result.stderr.decode("utf-8", errors="replace").strip())


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--pre-integration-workbook", type=Path)
    parser.add_argument("--pre-integration-ref", default="281eb14^:stingray_master.xlsx")
    parser.add_argument("--approve-checkpoint-1", action="store_true")
    parser.add_argument("--reviewer")
    parser.add_argument("--reviewed-at")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("form-output/ingest-wizard/20260720-options-recovery-projection"),
    )
    args = parser.parse_args(argv)
    root = args.root.resolve()
    output_dir = args.output_dir if args.output_dir.is_absolute() else root / args.output_dir

    if args.approve_checkpoint_1:
        if not args.reviewer or not args.reviewed_at:
            parser.error("--approve-checkpoint-1 requires --reviewer and --reviewed-at")
        workbook_path = args.workbook.resolve() if args.workbook else root / "stingray_master.xlsx"
        approval, packet = generate_checkpoint_1_packet(
            output_dir,
            workbook_path,
            reviewer=args.reviewer,
            reviewed_at=args.reviewed_at,
        )
        print(f"bulk-approved review records: {approval['approvedReviewCount']}")
        print(f"pending exception records: {packet['pendingReviewCount']}")
        print(f"condensed decision groups: {packet['exceptionGroupCount']}")
        print(f"packet: {output_dir / 'checkpoint-1-exception-review.md'}")
        return 0

    if args.pre_integration_workbook:
        pre_path = args.pre_integration_workbook.resolve()
        inputs = _default_inputs(root, pre_path)
        if args.workbook:
            inputs = ProjectionInputs(**{**inputs.__dict__, "workbook_path": args.workbook.resolve()})
        reports = generate_recovery_reports(inputs, output_dir)
    else:
        with tempfile.TemporaryDirectory(prefix="27vette-options-recovery-") as tmp:
            pre_path = Path(tmp) / "pre-integration.xlsx"
            _extract_git_workbook(root, args.pre_integration_ref, pre_path)
            inputs = _default_inputs(root, pre_path)
            if args.workbook:
                inputs = ProjectionInputs(**{**inputs.__dict__, "workbook_path": args.workbook.resolve()})
            reports = generate_recovery_reports(inputs, output_dir)

    for model in TARGET_MODELS:
        summary = reports[model]["summary"]
        print(
            f"{model}: {summary['residualRowCount']} residual rows, "
            f"{summary['pendingReviewCount']} pending decisions"
        )
    print(f"reports: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
