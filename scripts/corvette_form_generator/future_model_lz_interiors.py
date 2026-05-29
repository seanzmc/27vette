"""Z06/ZR1/ZR1X LZ interior preview and workbook write helpers.

This module reads workbook-owned LZ_Interiors rows and projects the rows that
future Z-family models need in model_interior_scope and interior_components.
Dry-run preview remains the default caller behavior. Explicit write-mode callers
can apply the projected rows to workbook source sheets without touching generated
form_* sheets or runtime app data.
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook

from corvette_form_generator.future_model_ingest import FUTURE_MODEL_SPECS
from corvette_form_generator.inspection import (
    interior_component_metadata,
    price_ref_component_prices,
    rows_from_sheet,
)
from corvette_form_generator.workbook import clean

LZ_INTERIOR_SOURCE_SHEET = "LZ_Interiors"
PRICE_REF_SHEET = "PriceRef"
SECTION_MASTER_SHEET = "section_master"
MODEL_INTERIOR_SCOPE_SHEET = "model_interior_scope"
INTERIOR_COMPONENTS_SHEET = "interior_components"

FUTURE_MODEL_KEYS = tuple(FUTURE_MODEL_SPECS)
LZ_TRIM_SECTION_MAP: dict[str, str] = {
    "1LZ": "sec_lzint_001",
    "2LZ": "sec_lzint_002",
    "3LZ": "sec_lzint_003",
    "3LZ_R6X": "sec_lzint_003",
}
LZ_SECTION_ROWS: tuple[dict[str, Any], ...] = (
    {
        "section_id": "sec_lzint_001",
        "section_name": "1LZ Interior",
        "selection_mode": "single_select_req",
        "is_required": False,
        "display_order": 15,
        "standard_behavior": "locked_included",
        "help_text": "",
        "step_key": "base_interior",
    },
    {
        "section_id": "sec_lzint_002",
        "section_name": "2LZ Interior",
        "selection_mode": "single_select_req",
        "is_required": False,
        "display_order": 16,
        "standard_behavior": "locked_included",
        "help_text": "",
        "step_key": "base_interior",
    },
    {
        "section_id": "sec_lzint_003",
        "section_name": "3LZ Interior",
        "selection_mode": "single_select_req",
        "is_required": False,
        "display_order": 17,
        "standard_behavior": "locked_included",
        "help_text": "",
        "step_key": "base_interior",
    },
)
MODEL_TRIM_SCOPE: dict[str, tuple[str, ...]] = {
    "z06": ("1LZ", "2LZ", "3LZ", "3LZ_R6X"),
    "zr1": ("1LZ", "3LZ", "3LZ_R6X"),
    "zr1x": ("1LZ", "3LZ", "3LZ_R6X"),
}
SECTION_HEADERS = ("section_id", "section_name", "selection_mode", "is_required", "display_order", "standard_behavior", "help_text", "step_key")
MODEL_INTERIOR_SCOPE_HEADERS = ("model_key", "interior_id", "trim_level", "active", "requires_option_id", "notes")
INTERIOR_COMPONENT_HEADERS = (
    "model_key",
    "interior_id",
    "rpo",
    "component_type",
    "label",
    "price_ref_type",
    "price_ref_code",
    "price_trim_scope",
    "display_order",
    "active",
    "notes",
)
_BOOL_FIELDS = {"active", "is_required", "review_flag", "selectable", "active_for_stingray", "requires_r6x"}


def selected_future_model_keys(model_keys: Iterable[str]) -> list[str]:
    requested = [clean(model_key) for model_key in model_keys if clean(model_key)]
    if not requested or requested == ["all"] or "all" in requested:
        return list(FUTURE_MODEL_KEYS)
    unknown = [model_key for model_key in requested if model_key not in FUTURE_MODEL_KEYS]
    if unknown:
        raise ValueError(f"Unknown future model key(s): {', '.join(unknown)}")
    return requested


def lz_rows_from_workbook(wb: Any) -> list[dict[str, Any]]:
    if LZ_INTERIOR_SOURCE_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing source sheet {LZ_INTERIOR_SOURCE_SHEET}")
    rows = []
    for row in rows_from_sheet(wb, LZ_INTERIOR_SOURCE_SHEET):
        if clean(row.get("interior_id")):
            rows.append(row)
    return rows


def proposed_section_id(row: dict[str, Any]) -> str:
    return LZ_TRIM_SECTION_MAP.get(clean(row.get("Trim")), "")


def scoped_lz_rows(source_rows: list[dict[str, Any]], model_key: str) -> list[dict[str, Any]]:
    trim_scope = set(MODEL_TRIM_SCOPE[model_key])
    return [row for row in source_rows if clean(row.get("Trim")) in trim_scope]


def _scope_preview_row(model_key: str, row: dict[str, Any]) -> dict[str, Any]:
    trim = clean(row.get("Trim"))
    interior_id = clean(row.get("interior_id"))
    return {
        "model_key": model_key,
        "interior_id": interior_id,
        "trim_level": trim.replace("_R6X", ""),
        "active": True,
        "requires_option_id": "opt_r6x_001" if trim.endswith("_R6X") or interior_id.endswith("_R6X") else "",
        "notes": "Workbook write-mode row projected from LZ_Interiors for future Z-family model prep.",
    }


def _component_preview_rows(
    *,
    model_key: str,
    source_rows: list[dict[str, Any]],
    component_price_ref: dict[tuple[str, str, str], int],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source_row in source_rows:
        interior_id = clean(source_row.get("interior_id"))
        for display_order, component in enumerate(interior_component_metadata(source_row, component_price_ref), start=1):
            rpo = clean(component.get("rpo"))
            component_type = clean(component.get("component_type"))
            rows.append(
                {
                    "model_key": model_key,
                    "interior_id": interior_id,
                    "rpo": rpo,
                    "component_type": component_type,
                    "label": clean(component.get("label")),
                    "price": component.get("price", 0),
                    "price_ref_type": component_type,
                    "price_ref_code": rpo,
                    "price_trim_scope": clean(source_row.get("Trim")),
                    "display_order": str(display_order * 10),
                    "active": True,
                    "notes": "Workbook write-mode row projected from legacy LZ component semantics for future Z-family model prep.",
                }
            )
    return rows


def build_lz_interiors_preview(wb: Any, model_keys: Iterable[str] = ("all",), *, include_details: bool = False) -> dict[str, Any]:
    selected = selected_future_model_keys(model_keys)
    source_rows = lz_rows_from_workbook(wb)
    price_ref_rows = rows_from_sheet(wb, PRICE_REF_SHEET) if PRICE_REF_SHEET in wb.sheetnames else []
    component_price_ref = price_ref_component_prices(price_ref_rows)

    source_trim_counts = Counter(clean(row.get("Trim")) for row in source_rows)
    current_section_counts = Counter(clean(row.get("section_id")) for row in source_rows)
    proposed_section_counts = Counter(proposed_section_id(row) for row in source_rows)

    preview: dict[str, Any] = {
        "status": "dry_run",
        "source_sheet": LZ_INTERIOR_SOURCE_SHEET,
        "source_row_count": len(source_rows),
        "source_trim_counts": dict(sorted(source_trim_counts.items())),
        "current_section_counts": dict(sorted(current_section_counts.items())),
        "proposed_section_mapping": dict(LZ_TRIM_SECTION_MAP),
        "proposed_section_counts": dict(sorted(proposed_section_counts.items())),
        "selected_model_keys": selected,
        "models": {},
        "would_write_workbook": False,
        "would_mutate_generated_runtime_data": False,
    }

    for model_key in selected:
        model_source_rows = scoped_lz_rows(source_rows, model_key)
        trim_counts = Counter(clean(row.get("Trim")) for row in model_source_rows)
        section_counts = Counter(proposed_section_id(row) for row in model_source_rows)
        scope_rows = [_scope_preview_row(model_key, row) for row in model_source_rows]
        component_rows = _component_preview_rows(
            model_key=model_key,
            source_rows=model_source_rows,
            component_price_ref=component_price_ref,
        )
        component_type_counts = Counter(row["component_type"] for row in component_rows)
        model_preview: dict[str, Any] = {
            "model_key": model_key,
            "trim_scope": list(MODEL_TRIM_SCOPE[model_key]),
            "interior_scope_row_count": len(scope_rows),
            "trim_counts": dict(sorted(trim_counts.items())),
            "proposed_section_counts": dict(sorted(section_counts.items())),
            "component_row_count": len(component_rows),
            "component_type_counts": dict(sorted(component_type_counts.items())),
        }
        if include_details:
            model_preview["interior_scope_rows"] = scope_rows
            model_preview["component_rows"] = component_rows
            model_preview["interior_rows"] = [
                {
                    "interior_id": clean(row.get("interior_id")),
                    "trim": clean(row.get("Trim")),
                    "current_section_id": clean(row.get("section_id")),
                    "proposed_section_id": proposed_section_id(row),
                    "requires_r6x": "True" if clean(row.get("Trim")).endswith("_R6X") else "False",
                }
                for row in model_source_rows
            ]
        preview["models"][model_key] = model_preview

    return preview


def _sheet_headers(ws: Any) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def _header_index(wb: Any, sheet_name: str, required_headers: Iterable[str]) -> dict[str, int]:
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Missing required sheet {sheet_name}")
    ws = wb[sheet_name]
    headers = _sheet_headers(ws)
    index = {header: offset + 1 for offset, header in enumerate(headers) if header}
    missing = [header for header in required_headers if header not in index]
    if missing:
        raise RuntimeError(f"{sheet_name} missing required header(s): {', '.join(missing)}")
    return index


def _row_dict(ws: Any, row_number: int, headers: Iterable[str]) -> dict[str, Any]:
    index = _header_index(ws.parent, ws.title, headers)
    return {header: ws.cell(row_number, index[header]).value for header in headers}


def _excel_value(header: str, value: Any) -> Any:
    text = clean(value)
    if header in _BOOL_FIELDS:
        if text.casefold() in {"true", "1", "yes", "y", "active"}:
            return True
        if text.casefold() in {"false", "0", "no", "n", "inactive"}:
            return False
    return value


def _append_dict_rows(ws: Any, headers: Iterable[str], rows: Iterable[dict[str, Any]]) -> int:
    count = 0
    for row in rows:
        ws.append([_excel_value(header, row.get(header, "")) for header in headers])
        count += 1
    return count


def _delete_rows_matching(ws: Any, headers: Iterable[str], predicate: Callable[[dict[str, Any]], bool]) -> int:
    deleted = 0
    for row_number in range(ws.max_row, 1, -1):
        row = _row_dict(ws, row_number, headers)
        if any(clean(value) for value in row.values()) and predicate(row):
            ws.delete_rows(row_number, 1)
            deleted += 1
    return deleted


def _replace_rows_for_models(wb: Any, sheet_name: str, headers: tuple[str, ...], model_keys: Iterable[str], rows: list[dict[str, Any]]) -> dict[str, int]:
    _header_index(wb, sheet_name, headers)
    ws = wb[sheet_name]
    selected = set(model_keys)
    deleted = _delete_rows_matching(ws, headers, lambda row: clean(row.get("model_key")) in selected)
    inserted = _append_dict_rows(ws, headers, rows)
    return {"deleted": deleted, "inserted": inserted}


def _ensure_lz_sections(wb: Any) -> dict[str, int]:
    _header_index(wb, SECTION_MASTER_SHEET, SECTION_HEADERS)
    ws = wb[SECTION_MASTER_SHEET]
    index = _header_index(wb, SECTION_MASTER_SHEET, SECTION_HEADERS)
    existing: dict[str, int] = {}
    for row_number in range(2, ws.max_row + 1):
        section_id = clean(ws.cell(row_number, index["section_id"]).value)
        if section_id:
            existing[section_id] = row_number

    inserted = 0
    for expected in LZ_SECTION_ROWS:
        section_id = clean(expected["section_id"])
        row_number = existing.get(section_id)
        if row_number is None:
            ws.append([_excel_value(header, expected.get(header, "")) for header in SECTION_HEADERS])
            inserted += 1
            continue
        for header in SECTION_HEADERS:
            actual = clean(ws.cell(row_number, index[header]).value)
            wanted = clean(expected.get(header, ""))
            if actual != wanted:
                raise RuntimeError(f"{SECTION_MASTER_SHEET} existing {section_id} has {header}={actual!r}; expected {wanted!r}")
    return {"inserted": inserted, "existing": len(LZ_SECTION_ROWS) - inserted}


def _apply_lz_source_section_ids(wb: Any) -> dict[str, Any]:
    required = ("interior_id", "Trim", "section_id")
    index = _header_index(wb, LZ_INTERIOR_SOURCE_SHEET, required)
    ws = wb[LZ_INTERIOR_SOURCE_SHEET]
    updated = 0
    section_counts: Counter[str] = Counter()
    for row_number in range(2, ws.max_row + 1):
        interior_id = clean(ws.cell(row_number, index["interior_id"]).value)
        if not interior_id:
            continue
        trim = clean(ws.cell(row_number, index["Trim"]).value)
        desired = LZ_TRIM_SECTION_MAP.get(trim, "")
        if not desired:
            raise RuntimeError(f"{LZ_INTERIOR_SOURCE_SHEET} row {row_number} has unsupported Trim={trim!r}")
        current = clean(ws.cell(row_number, index["section_id"]).value)
        if current != desired:
            ws.cell(row_number, index["section_id"]).value = desired
            updated += 1
        section_counts[desired] += 1
    return {"updated": updated, "section_counts": dict(sorted(section_counts.items()))}


def _model_detail_rows(preview: dict[str, Any], key: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    missing: list[str] = []
    for model_key in preview["selected_model_keys"]:
        model_preview = preview["models"][model_key]
        model_rows = model_preview.get(key)
        if model_rows is None:
            missing.append(f"{model_key}.{key}")
            continue
        rows.extend(model_rows)
    if missing:
        raise RuntimeError(f"Preview missing write-mode detail rows: {', '.join(missing)}. Rebuild preview with include_details=True.")
    return rows


def apply_lz_interiors_to_workbook(wb: Any, preview: dict[str, Any]) -> dict[str, Any]:
    """Apply an include_details LZ interior preview to workbook source sheets."""

    selected = selected_future_model_keys(preview.get("selected_model_keys", ("all",)))
    section_report = _ensure_lz_sections(wb)
    source_report = _apply_lz_source_section_ids(wb)
    scope_rows = _model_detail_rows(preview, "interior_scope_rows")
    component_rows = _model_detail_rows(preview, "component_rows")
    scope_report = _replace_rows_for_models(wb, MODEL_INTERIOR_SCOPE_SHEET, MODEL_INTERIOR_SCOPE_HEADERS, selected, scope_rows)
    component_report = _replace_rows_for_models(wb, INTERIOR_COMPONENTS_SHEET, INTERIOR_COMPONENT_HEADERS, selected, component_rows)
    return {
        "status": "applied",
        "selected_model_keys": selected,
        "section_master": section_report,
        LZ_INTERIOR_SOURCE_SHEET: source_report,
        MODEL_INTERIOR_SCOPE_SHEET: scope_report,
        INTERIOR_COMPONENTS_SHEET: component_report,
        "would_mutate_generated_runtime_data": False,
    }


def assert_future_metadata_inactive(wb: Any) -> None:
    future_models = set(FUTURE_MODEL_KEYS)
    checks = {
        "model_master": ("model_key", ["active"]),
        "model_workbook_sources": ("model_key", ["active"]),
        "model_registry_promotion": ("model_key", ["promoted_to_runtime", "active"]),
    }
    for sheet_name, (model_field, inactive_fields) in checks.items():
        if sheet_name not in wb.sheetnames:
            continue
        index = _header_index(wb, sheet_name, (model_field, *inactive_fields))
        ws = wb[sheet_name]
        for row_number in range(2, ws.max_row + 1):
            model_key = clean(ws.cell(row_number, index[model_field]).value)
            if model_key not in future_models:
                continue
            for field in inactive_fields:
                value = clean(ws.cell(row_number, index[field]).value).casefold()
                if value in {"true", "1", "yes", "y", "active"}:
                    raise RuntimeError(f"{sheet_name} row {row_number} unexpectedly has active {field}={value}")


def verify_saved_lz_interiors_workbook(path: Path, preview: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        selected = selected_future_model_keys(preview.get("selected_model_keys", ("all",)))
        source_rows = lz_rows_from_workbook(wb)
        section_counts = Counter(clean(row.get("section_id")) for row in source_rows)
        expected_sections = {clean(row["section_id"]): clean(row["section_name"]) for row in LZ_SECTION_ROWS}
        section_rows = {clean(row.get("section_id")): row for row in rows_from_sheet(wb, SECTION_MASTER_SHEET)}
        missing_sections = [section_id for section_id in expected_sections if section_id not in section_rows]
        if missing_sections:
            raise RuntimeError(f"Saved workbook is missing LZ section rows: {', '.join(missing_sections)}")
        expected_section_counts = dict(sorted(Counter(proposed_section_id(row) for row in source_rows).items()))
        actual_section_counts = dict(sorted(section_counts.items()))
        if actual_section_counts != expected_section_counts:
            raise RuntimeError(f"Saved LZ section counts differ: expected {expected_section_counts}, found {actual_section_counts}")

        scope_rows = rows_from_sheet(wb, MODEL_INTERIOR_SCOPE_SHEET)
        component_rows = rows_from_sheet(wb, INTERIOR_COMPONENTS_SHEET)
        verification: dict[str, Any] = {
            "source_row_count": len(source_rows),
            "section_counts": actual_section_counts,
            "models": {},
        }
        for model_key in selected:
            expected_scope = preview["models"][model_key]["interior_scope_row_count"]
            expected_components = preview["models"][model_key]["component_row_count"]
            actual_scope = sum(1 for row in scope_rows if clean(row.get("model_key")) == model_key)
            actual_components = sum(1 for row in component_rows if clean(row.get("model_key")) == model_key)
            if actual_scope != expected_scope:
                raise RuntimeError(f"{MODEL_INTERIOR_SCOPE_SHEET} {model_key} expected {expected_scope}, found {actual_scope}")
            if actual_components != expected_components:
                raise RuntimeError(f"{INTERIOR_COMPONENTS_SHEET} {model_key} expected {expected_components}, found {actual_components}")
            verification["models"][model_key] = {
                "interior_scope_rows": actual_scope,
                "component_rows": actual_components,
            }
        assert_future_metadata_inactive(wb)
        verification["future_metadata_inactive"] = True
        return verification
    finally:
        wb.close()


def strip_lz_preview_details(preview: dict[str, Any]) -> dict[str, Any]:
    stripped = dict(preview)
    stripped["models"] = {}
    for model_key, model_preview in preview["models"].items():
        stripped["models"][model_key] = {
            key: value
            for key, value in model_preview.items()
            if key not in {"interior_scope_rows", "component_rows", "interior_rows"}
        }
    return stripped
