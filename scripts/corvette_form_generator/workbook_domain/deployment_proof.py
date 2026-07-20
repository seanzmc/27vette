#!/usr/bin/env python3
"""Temporary deployment-proof helpers, separated from ingest sessions.

The shared workbook service owns application authority. The legacy mixin remains
available only for historical diagnostics; :func:`prove_changeset_deployment`
is the public ``workbook-changeset-1`` proof entry point and never writes the
canonical workbook.
"""

from __future__ import annotations

import copy
import hashlib
import json
import re
import shutil
import subprocess
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from corvette_form_generator import editor_ops
from corvette_form_generator.ingest.wizard.canonical_rows import (
    semantic_hash,
    validate_artifact_graph,
)
from corvette_form_generator.workbook_domain.changeset import (
    ChangeSetError,
    canonical_json,
    changeset_fingerprint,
    changeset_to_editor_batch,
    parse_changeset,
)

ROOT = Path(__file__).resolve().parents[3]
WRITABLE_PLAN_SCHEMA = "pass-c-3"


class TemporaryDeploymentProofMixin:
    """Mixin for a non-ingest diagnostic host supplying workbook helpers."""

    def _combined_plan_batch(self, plan: dict[str, Any], workbook: Path) -> dict[str, Any]:
        return {
            "workbookMtimeNs": str(workbook.stat().st_mtime_ns),
            "forceTypedBools": plan.get("schemaVersion") == WRITABLE_PLAN_SCHEMA,
            "items": [*plan["stage1"]["items"], *plan["stage2"]["items"]],
        }

    def _per_sheet_action_counts(self, items: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
        counts: dict[str, dict[str, int]] = {}
        for item in items:
            sheet = item.get("sheet")
            action = item.get("action")
            if not sheet or not action:
                continue
            sheet_counts = counts.setdefault(str(sheet), {})
            sheet_counts[str(action)] = sheet_counts.get(str(action), 0) + 1
        return {sheet: dict(sorted(actions.items())) for sheet, actions in sorted(counts.items())}

    def _deployment_continuity_from_plan(self, plan: dict[str, Any]) -> dict[str, Any]:
        """Action-aware source-coverage diagnostic for the dry-run report.

        The temp generation probe added by Pass D.1 can enrich these entries;
        this source-op layer is intentionally conservative and never treats
        create/delete-only sheet activity as runtime coverage.
        """

        report = plan.get("report") or {}
        plan_continuity = report.get("runtimeContinuity") or {}
        output: dict[str, Any] = {}
        for model in plan.get("targets") or []:
            source_ops = (plan_continuity.get(model) or {}).get("sourceOps") or {}
            source_actions = ("add", "update", "noop")
            price_adds = sum((source_ops.get("priceRules") or {}).get(action, 0) for action in source_actions)
            rule_group_adds = sum((source_ops.get("ruleGroups") or {}).get(action, 0) for action in source_actions)
            color_adds = sum((source_ops.get("colorOverrides") or {}).get(action, 0) for action in source_actions)
            component_adds = sum((source_ops.get("interiorComponents") or {}).get(action, 0) for action in source_actions)
            asset_adds = sum((source_ops.get("assetMap") or {}).get(action, 0) for action in source_actions)
            blockers: list[dict[str, str]] = []
            deferrals: list[dict[str, str]] = []
            if model in {"zr1", "zr1x"} and price_adds == 0:
                blockers.append({"kind": "price_rules_required_for_runtime", "detail": "no price-rule add/update ops"})
            if model in {"zr1", "zr1x"} and rule_group_adds == 0:
                blockers.append({"kind": "rule_groups_required_for_runtime", "detail": "no rule-group add/update ops"})
            output[model] = {
                "status": "not_deployment_ready" if blockers else "source_ops_diagnostic",
                "registryLoadable": None,
                "registryLoadableNote": "temp generation probe not run in source-op diagnostic layer",
                "sourceOps": source_ops,
                "sourceCoverage": {
                    "priceRuleAddOrUpdateCount": price_adds,
                    "ruleGroupAddOrUpdateCount": rule_group_adds,
                    "colorOverrideAddOrUpdateCount": color_adds,
                    "interiorComponentAddOrUpdateCount": component_adds,
                    "assetMapAddOrUpdateCount": asset_adds,
                },
                "deploymentBlockers": blockers,
                "deploymentDeferrals": deferrals,
            }
        return output

    def _activate_probe_models(self, workbook: Path, models: list[str]) -> None:
        from openpyxl import load_workbook

        wb = load_workbook(workbook)
        try:
            targets = set(models)
            variant_ids: set[str] = set()
            if "model_variants" in wb.sheetnames:
                ws = wb["model_variants"]
                headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value is not None}
                model_col = headers.get("model_key")
                variant_col = headers.get("variant_id")
                if model_col and variant_col:
                    for row in range(2, ws.max_row + 1):
                        model_key = str(ws.cell(row=row, column=model_col).value or "").strip().lower()
                        if model_key in targets:
                            variant_id = str(ws.cell(row=row, column=variant_col).value or "").strip()
                            if variant_id:
                                variant_ids.add(variant_id)
            for sheet_name, fields in {
                "model_master": {"active": True},
                "model_variants": {"active": True},
                "model_workbook_sources": {"active": True},
                "model_registry_promotion": {"active": True, "promoted_to_runtime": True},
            }.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value is not None}
                model_col = headers.get("model_key")
                if not model_col:
                    continue
                for row in range(2, ws.max_row + 1):
                    model_key = str(ws.cell(row=row, column=model_col).value or "").strip().lower()
                    if model_key not in targets:
                        if "active" in headers:
                            ws.cell(row=row, column=headers["active"], value=False)
                        if sheet_name == "model_registry_promotion" and "promoted_to_runtime" in headers:
                            ws.cell(
                                row=row,
                                column=headers["promoted_to_runtime"],
                                value=False,
                            )
                        if sheet_name == "model_registry_promotion" and "default_model" in headers:
                            ws.cell(
                                row=row,
                                column=headers["default_model"],
                                value=False,
                            )
                        continue
                    for column, value in fields.items():
                        if column in headers:
                            ws.cell(row=row, column=headers[column], value=value)
                    if sheet_name == "model_registry_promotion":
                        if "default_model" in headers:
                            ws.cell(
                                row=row,
                                column=headers["default_model"],
                                value=model_key == models[0],
                            )
                        if "artifact_type" in headers:
                            ws.cell(row=row, column=headers["artifact_type"], value="runtime_contract")
                        if "artifact_path" in headers:
                            ws.cell(
                                row=row,
                                column=headers["artifact_path"],
                                value=f"form-output/runtime/{model_key.replace('_', '-')}-runtime-contract.json",
                            )
            if variant_ids and "variant_master" in wb.sheetnames:
                ws = wb["variant_master"]
                headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value is not None}
                variant_col = headers.get("variant_id")
                active_col = headers.get("active")
                if variant_col and active_col:
                    for row in range(2, ws.max_row + 1):
                        variant_id = str(ws.cell(row=row, column=variant_col).value or "").strip()
                        if variant_id in variant_ids:
                            ws.cell(row=row, column=active_col, value=True)
            wb.save(workbook)
        finally:
            wb.close()

    def _validate_probe_activation(
        self,
        before_workbook: Path,
        activated_workbook: Path,
        models: list[str],
    ) -> dict[str, Any]:
        """Revalidate and exactly read back a temp-only activation save."""

        from openpyxl import load_workbook

        from corvette_form_generator.schema_validation import (
            result_payload as schema_result_payload,
            validate_workbook_schema,
        )
        from corvette_form_generator.workbook_bool_hygiene import (
            compare_bool_like_workbooks,
            result_payload as bool_hygiene_result_payload,
        )
        from corvette_form_generator.workbook_package import (
            assert_valid_workbook_package,
        )

        errors: list[str] = []
        try:
            assert_valid_workbook_package(activated_workbook)
        except Exception as exc:
            errors.append(str(exc))
            package_result = {"status": "invalid", "error": str(exc)}
        else:
            package_result = {"status": "valid", "issues": []}

        schema_issues = validate_workbook_schema(
            activated_workbook,
            check_live_contract=False,
        )
        schema_result = schema_result_payload(str(activated_workbook), schema_issues)
        if schema_result.get("error_count"):
            errors.append(
                f"post-activation schema validation returned "
                f"{schema_result['error_count']} error(s)"
            )

        bool_issues = compare_bool_like_workbooks(
            before_workbook,
            activated_workbook,
        )
        bool_result = bool_hygiene_result_payload(
            before_workbook,
            activated_workbook,
            bool_issues,
        )
        bool_result["issues"] = bool_result.get("issues", [])[:20]
        if bool_result.get("error_count"):
            errors.append(
                f"post-activation Boolean hygiene returned "
                f"{bool_result['error_count']} error(s)"
            )

        targets = set(models)
        checked = 0

        def is_true(value: Any) -> bool:
            return value is True or str(value or "").strip().lower() in {
                "1",
                "true",
                "yes",
            }

        wb = load_workbook(activated_workbook, read_only=True, data_only=True)
        try:
            target_variants: set[str] = set()
            for sheet_name in (
                "model_master",
                "model_variants",
                "model_workbook_sources",
                "model_registry_promotion",
            ):
                if sheet_name not in wb.sheetnames:
                    errors.append(f"post-activation readback missing sheet {sheet_name!r}")
                    continue
                ws = wb[sheet_name]
                headers = {
                    str(cell.value): index
                    for index, cell in enumerate(ws[1])
                    if cell.value is not None
                }
                if "model_key" not in headers or "active" not in headers:
                    errors.append(
                        f"post-activation readback missing model_key/active in {sheet_name!r}"
                    )
                    continue
                seen_targets: set[str] = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    model_key = str(row[headers["model_key"]] or "").strip().lower()
                    if not model_key:
                        continue
                    expected_active = model_key in targets
                    if is_true(row[headers["active"]]) != expected_active:
                        errors.append(
                            f"post-activation readback {sheet_name}:{model_key} active mismatch"
                        )
                    if expected_active:
                        seen_targets.add(model_key)
                    if sheet_name == "model_variants" and expected_active:
                        variant_id = str(row[headers.get("variant_id", -1)] or "").strip()
                        if variant_id:
                            target_variants.add(variant_id)
                    if sheet_name == "model_registry_promotion":
                        if is_true(row[headers.get("promoted_to_runtime", -1)]) != expected_active:
                            errors.append(
                                f"post-activation readback {model_key} promotion mismatch"
                            )
                        expected_default = expected_active and model_key == models[0]
                        if is_true(row[headers.get("default_model", -1)]) != expected_default:
                            errors.append(
                                f"post-activation readback {model_key} default mismatch"
                            )
                    checked += 1
                missing_targets = targets - seen_targets
                if missing_targets:
                    errors.append(
                        f"post-activation readback {sheet_name} missing targets "
                        f"{sorted(missing_targets)}"
                    )

            if "variant_master" not in wb.sheetnames:
                errors.append("post-activation readback missing sheet 'variant_master'")
            else:
                ws = wb["variant_master"]
                headers = {
                    str(cell.value): index
                    for index, cell in enumerate(ws[1])
                    if cell.value is not None
                }
                active_variants = {
                    str(row[headers["variant_id"]] or "").strip()
                    for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[headers.get("variant_id", -1)]
                    and is_true(row[headers.get("active", -1)])
                }
                inactive_targets = target_variants - active_variants
                if inactive_targets:
                    errors.append(
                        f"post-activation readback has inactive target variants "
                        f"{sorted(inactive_targets)}"
                    )
                checked += len(target_variants)
        finally:
            wb.close()

        return {
            "ok": not errors,
            "errors": errors,
            "package": package_result,
            "schema": schema_result,
            "boolHygiene": bool_result,
            "exactReadback": {"checked": checked, "errors": list(errors)},
        }

    def _source_count(self, payload: dict[str, Any], key: str) -> int:
        value = payload.get(key)
        return len(value) if isinstance(value, list) else 0

    def _source_feature_coverage_agreement(
        self,
        plan: dict[str, Any],
        model: str,
        signature_mismatches: list[dict[str, Any]],
    ) -> dict[str, Any]:
        coverage = plan.get("sourceFeatureCoverage") or {}
        model_coverage = (coverage.get("byModel") or {}).get(model) or {}
        blocking_features = list(model_coverage.get("blockingFeatures") or [])
        feature_count = model_coverage.get("featureCount")
        errors: list[str] = []
        if (
            not coverage.get("semanticSha")
            or not isinstance(feature_count, int)
            or feature_count <= 0
        ):
            errors.append("bound compiler source-feature coverage is absent or empty")
        if blocking_features:
            errors.append(
                f"compiler source-feature coverage contains {len(blocking_features)} blocking feature(s)"
            )
        if signature_mismatches:
            errors.append(
                f"runtime parity contains {len(signature_mismatches)} semantic mismatch(es)"
            )
        return {
            "ok": not errors,
            "semanticSha": coverage.get("semanticSha"),
            "featureCount": feature_count,
            "dispositionCounts": dict(model_coverage.get("dispositionCounts") or {}),
            "blockingFeatures": blocking_features,
            "runtimeMismatchCount": len(signature_mismatches),
            "errors": errors,
        }

    def _manifest_runtime_signature_mismatches(
        self,
        plan: dict[str, Any],
        model: str,
        contract: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Compare target-authored manifest semantics to generated runtime rows."""

        def scope(value: Any) -> str:
            return str(value or "*").strip().lower() or "*"

        choices = {
            str(row.get("id") or row.get("option_id") or ""): str(row.get("rpo") or "").upper()
            for row in contract.get("choices") or []
            if isinstance(row, dict)
        }
        interior_ids = {
            str(row.get("interior_id") or "")
            for row in contract.get("interiors") or []
            if isinstance(row, dict) and row.get("interior_id")
        }

        def rpo(option_id: Any) -> str:
            endpoint_id = str(option_id or "")
            if endpoint_id in choices:
                return choices[endpoint_id]
            if endpoint_id in interior_ids:
                return endpoint_id.upper()
            return ""

        runtime_rows = {
            "rule_mapping": {
                str(row.get("rule_id") or ""): row for row in contract.get("rules") or []
            },
            "rule_groups": {
                str(row.get("group_id") or ""): row for row in contract.get("ruleGroups") or []
            },
            "exclusive_groups": {
                str(row.get("group_id") or ""): row
                for row in contract.get("exclusiveGroups") or []
            },
            "default_selection_rules": {
                str(row.get("rule_id") or ""): row
                for row in contract.get("defaultSelectionRules") or []
            },
            "price_rules": {
                str(row.get("price_rule_id") or ""): row
                for row in contract.get("priceRules") or []
            },
        }
        mismatches: list[dict[str, Any]] = []
        material = [
            row
            for row in (plan.get("coverage") or {}).get("manifestRows") or []
            if row.get("model") in {model, "*"}
            and row.get("action") in {"add", "update", "noop"}
            and row.get("projectionScopeDisposition")
            != "retained_existing_noop_outside_target_domain"
        ]
        expected_runtime_ids: dict[str, set[str]] = {
            family: set() for family in runtime_rows
        }
        endpoint_runtime_visibility: dict[str, bool | None] = {}
        for endpoint_row in material:
            endpoint_family = str(endpoint_row.get("family") or "")
            endpoint_values = endpoint_row.get("canonicalValues") or {}
            endpoint_key = endpoint_row.get("key") or {}
            if endpoint_family == "options":
                endpoint_id = str(
                    endpoint_values.get("option_id")
                    or endpoint_key.get("option_id")
                    or ""
                )
                if (
                    endpoint_values.get("active") is False
                    or endpoint_values.get("selectable") is False
                ):
                    visible: bool | None = False
                elif (
                    endpoint_values.get("active") is True
                    and endpoint_values.get("selectable") is True
                ):
                    visible = True
                else:
                    visible = None
            elif endpoint_family == "interiors":
                endpoint_id = str(
                    endpoint_values.get("interior_id")
                    or endpoint_key.get("interior_id")
                    or ""
                )
                if endpoint_values.get("active") is True:
                    visible = True
                elif endpoint_values.get("active") is False:
                    visible = False
                else:
                    visible = None
            else:
                continue
            if endpoint_id:
                previous = endpoint_runtime_visibility.get(endpoint_id, False)
                if previous is True or visible is True:
                    endpoint_runtime_visibility[endpoint_id] = True
                elif previous is None or visible is None:
                    endpoint_runtime_visibility[endpoint_id] = None
                else:
                    endpoint_runtime_visibility[endpoint_id] = False
        for row in material:
            family = str(row.get("family") or "")
            if family not in {
                "rule_mapping",
                "rule_groups",
                "rule_group_members",
                "exclusive_groups",
                "exclusive_members",
                "default_selection_rules",
                "price_rules",
            }:
                continue
            key = row.get("key") or {}
            values = row.get("canonicalValues") or {}
            if family == "rule_mapping" and values:
                missing_endpoints = [
                    str(endpoint_id or "")
                    for endpoint_id in (
                        values.get("source_id"),
                        values.get("target_id"),
                    )
                    if not rpo(endpoint_id)
                ]
                if missing_endpoints and all(
                    endpoint_id in endpoint_runtime_visibility
                    and endpoint_runtime_visibility[endpoint_id] is False
                    for endpoint_id in missing_endpoints
                ):
                    # The runtime generator intentionally omits relationships
                    # whose missing endpoints are explicitly non-runtime in
                    # the canonical manifest. Endpoint absence by itself is
                    # not authority because it can also indicate generator
                    # regression.
                    continue
            expected = row.get("semanticSignature")
            if values:
                if family == "rule_mapping":
                    expected = {
                        "bodyStyleScope": scope(values.get("body_style_scope")),
                        "ruleType": str(values.get("rule_type") or ""),
                        "sourceRpo": rpo(values.get("source_id")),
                        "targetRpo": rpo(values.get("target_id")),
                        "trimLevelScope": scope(values.get("trim_level_scope")),
                        "variantScope": scope(values.get("variant_scope")),
                    }
                elif family == "rule_groups":
                    group_id = str(key.get("group_id") or values.get("group_id") or "")
                    expected = {
                        "bodyStyleScope": scope(values.get("body_style_scope")),
                        "groupType": str(values.get("group_type") or ""),
                        "memberRpos": sorted(
                            rpo((member.get("canonicalValues") or {}).get("target_id"))
                            for member in material
                            if member.get("family") == "rule_group_members"
                            and str((member.get("key") or {}).get("group_id") or "")
                            == group_id
                        ),
                        "sourceRpo": rpo(values.get("source_id")),
                        "trimLevelScope": scope(values.get("trim_level_scope")),
                        "variantScope": scope(values.get("variant_scope")),
                    }
                elif family == "exclusive_groups":
                    group_id = str(key.get("group_id") or values.get("group_id") or "")
                    expected = {
                        "memberRpos": sorted(
                            rpo((member.get("canonicalValues") or {}).get("option_id"))
                            for member in material
                            if member.get("family") == "exclusive_members"
                            and str((member.get("key") or {}).get("group_id") or "")
                            == group_id
                        ),
                        "selectionMode": str(values.get("selection_mode") or ""),
                    }
                elif family == "default_selection_rules":
                    expected = {
                        "bodyStyleScope": scope(values.get("body_style_scope")),
                        "conditionId": str(values.get("condition_id") or ""),
                        "conditionType": str(values.get("condition_type") or ""),
                        "modelKey": str(values.get("model_key") or model),
                        "targetOptionId": str(values.get("target_option_id") or ""),
                        "trimLevelScope": scope(values.get("trim_level_scope")),
                        "variantScope": scope(values.get("variant_scope")),
                    }
                elif family == "price_rules":
                    expected = {
                        "bodyStyleScope": scope(values.get("body_style_scope")),
                        "conditionOptionId": str(values.get("condition_option_id") or ""),
                        "priceRuleType": str(values.get("price_rule_type") or ""),
                        "priceValue": values.get("price_value"),
                        "targetOptionId": str(values.get("target_option_id") or ""),
                        "trimLevelScope": scope(values.get("trim_level_scope")),
                        "variantScope": scope(values.get("variant_scope")),
                    }
            if not isinstance(expected, dict):
                mismatches.append(
                    {"family": family, "key": key, "kind": "missing_manifest_semantic_signature"}
                )
                continue
            if family in {"rule_group_members", "exclusive_members"}:
                group_family = "rule_groups" if family == "rule_group_members" else "exclusive_groups"
                group_id = str(key.get("group_id") or "")
                generated = runtime_rows[group_family].get(group_id)
                target_field = "target_ids" if family == "rule_group_members" else "option_ids"
                option_field = "target_id" if family == "rule_group_members" else "option_id"
                generated_rpos = sorted(
                    rpo(value) for value in ((generated or {}).get(target_field) or [])
                )
                expected_rpo = str(
                    expected.get("memberRpo") or rpo(key.get(option_field))
                ).upper()
                if generated is None or expected_rpo not in generated_rpos:
                    mismatches.append(
                        {
                            "family": family,
                            "key": key,
                            "kind": "generated_member_missing",
                            "expectedRpo": expected_rpo,
                            "generatedRpos": generated_rpos,
                            "optionId": key.get(option_field),
                        }
                    )
                continue
            key_field = {
                "rule_mapping": "rule_id",
                "rule_groups": "group_id",
                "exclusive_groups": "group_id",
                "default_selection_rules": "rule_id",
                "price_rules": "price_rule_id",
            }[family]
            runtime_id = str(key.get(key_field) or "")
            expected_runtime_ids[family].add(runtime_id)
            generated = runtime_rows.get(family, {}).get(runtime_id)
            if generated is None:
                mismatches.append(
                    {"family": family, "key": key, "kind": "generated_row_missing"}
                )
                continue
            if expected.get("retainedKey") and not values:
                continue
            if family == "rule_mapping":
                actual = {
                    "bodyStyleScope": scope(generated.get("body_style_scope")),
                    "ruleType": str(generated.get("rule_type") or ""),
                    "sourceRpo": rpo(generated.get("source_id")),
                    "targetRpo": rpo(generated.get("target_id")),
                    "trimLevelScope": scope(generated.get("trim_level_scope")),
                    "variantScope": scope(generated.get("variant_scope")),
                }
            elif family == "rule_groups":
                actual = {
                    "bodyStyleScope": scope(generated.get("body_style_scope")),
                    "groupType": str(generated.get("group_type") or ""),
                    "memberRpos": sorted(rpo(value) for value in generated.get("target_ids") or []),
                    "sourceRpo": rpo(generated.get("source_id")),
                    "trimLevelScope": scope(generated.get("trim_level_scope")),
                    "variantScope": scope(generated.get("variant_scope")),
                }
            elif family == "exclusive_groups":
                actual = {
                    "memberRpos": sorted(rpo(value) for value in generated.get("option_ids") or []),
                    "selectionMode": str(generated.get("selection_mode") or ""),
                }
            elif family == "default_selection_rules":
                actual = {
                    "bodyStyleScope": scope(generated.get("body_style_scope")),
                    "conditionId": str(generated.get("condition_id") or ""),
                    "conditionType": str(generated.get("condition_type") or ""),
                    "modelKey": model,
                    "targetOptionId": str(generated.get("target_option_id") or ""),
                    "trimLevelScope": scope(generated.get("trim_level_scope")),
                    "variantScope": scope(generated.get("variant_scope")),
                }
            elif family == "price_rules":
                actual = {
                    "bodyStyleScope": scope(generated.get("body_style_scope")),
                    "priceRuleType": str(generated.get("price_rule_type") or ""),
                    "priceValue": generated.get("price_value"),
                    "trimLevelScope": scope(generated.get("trim_level_scope")),
                    "variantScope": scope(generated.get("variant_scope")),
                }
                if "conditionOptionId" in expected or "targetOptionId" in expected:
                    actual.update(
                        {
                            "conditionOptionId": str(
                                generated.get("condition_option_id") or ""
                            ),
                            "targetOptionId": str(
                                generated.get("target_option_id") or ""
                            ),
                        }
                    )
                else:
                    actual.update(
                        {
                            "conditionRpo": rpo(generated.get("condition_option_id")),
                            "targetRpo": rpo(generated.get("target_option_id")),
                        }
                    )
            else:
                continue
            comparable_expected = dict(expected)
            for scope_field in (
                "bodyStyleScope",
                "trimLevelScope",
                "variantScope",
            ):
                if scope_field in comparable_expected:
                    comparable_expected[scope_field] = scope(
                        comparable_expected[scope_field]
                    )
            comparable_actual = {
                key: actual.get(key) for key in comparable_expected
            }
            if semantic_hash(comparable_actual) != semantic_hash(comparable_expected):
                mismatches.append(
                    {
                        "family": family,
                        "key": key,
                        "kind": "semantic_signature_mismatch",
                        "expected": expected,
                        "generated": actual,
                    }
                )

        for row in (plan.get("coverage") or {}).get("manifestRows") or []:
            if row.get("model") not in {model, "*"} or row.get("action") != "delete":
                continue
            family = str(row.get("family") or "")
            key_field = {
                "rule_mapping": "rule_id",
                "rule_groups": "group_id",
                "exclusive_groups": "group_id",
                "default_selection_rules": "rule_id",
                "price_rules": "price_rule_id",
            }.get(family)
            if key_field is None:
                continue
            runtime_id = str((row.get("key") or {}).get(key_field) or "")
            if runtime_id in runtime_rows[family]:
                mismatches.append(
                    {
                        "family": family,
                        "key": row.get("key") or {},
                        "kind": "deleted_row_still_generated",
                    }
                )

        for family, rows_by_id in runtime_rows.items():
            for runtime_id in sorted(set(rows_by_id) - expected_runtime_ids[family]):
                mismatches.append(
                    {
                        "family": family,
                        "key": {"runtimeId": runtime_id},
                        "kind": "generated_row_unexpected",
                    }
                )

        expected_variants = {
            str((row.get("key") or {}).get("variant_id") or "")
            for row in (plan.get("coverage") or {}).get("manifestRows") or []
            if row.get("model") == model
            and row.get("family") == "model_variants"
            and row.get("action") in {"add", "update", "noop"}
        }
        generated_variants = {
            str(row.get("variant_id") or "")
            for row in contract.get("variants") or []
            if isinstance(row, dict)
        }
        if expected_variants != generated_variants:
            mismatches.append(
                {
                    "family": "model_variants",
                    "kind": "variant_set_mismatch",
                    "missing": sorted(expected_variants - generated_variants),
                    "unexpected": sorted(generated_variants - expected_variants),
                }
            )
        return mismatches

    def _deployment_phase_plan(
        self,
        plan: dict[str, Any],
        models: list[str],
    ) -> dict[str, Any]:
        """Return a mechanical target subset for temp-only phased proof."""

        selected = set(models)
        coverage = plan.get("coverage") or {}
        manifest_rows = [
            dict(row)
            for row in coverage.get("manifestRows") or []
            if row.get("model") == "*" or row.get("model") in selected
        ]
        manifest_refs = {str(row.get("manifestRef") or "") for row in manifest_rows}

        def include(item: dict[str, Any]) -> bool:
            if (
                item.get("_scaffoldRule")
                == "pass_c3_greenfield_registry_promotion"
            ):
                return str((item.get("key") or {}).get("model_key") or "") in selected
            if item.get("action") == "create_sheet":
                return bool(manifest_refs & set(item.get("_manifestRefs") or []))
            return str(item.get("_manifestRef") or "") in manifest_refs

        stage1 = [dict(item) for item in plan.get("stage1", {}).get("items") or [] if include(item)]
        stage2 = [dict(item) for item in plan.get("stage2", {}).get("items") or [] if include(item)]
        noops = [
            dict(row)
            for row in coverage.get("noops") or []
            if row.get("model") == "*" or row.get("model") in selected
        ]
        return {
            **plan,
            "targets": list(models),
            "targetModes": {
                model: (plan.get("targetModes") or {}).get(model) for model in models
            },
            "stage1": {"items": stage1},
            "stage2": {"items": stage2},
            "coverage": {
                **coverage,
                "manifestRows": manifest_rows,
                "noops": noops,
            },
            "planReadiness": {
                "planReady": True,
                "manifestRowCount": len(manifest_rows),
                "mutationCount": len(stage1) + len(stage2),
                "noopCount": len(noops),
                "uncoveredCount": 0,
            },
        }

    def _deployment_proof_phases(
        self,
        workbook: Path,
        plan: dict[str, Any],
        *,
        schema_validation: bool,
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        """Run the required GSX/ZR1, ZR1X, then all-target temp proofs."""

        targets = [str(model) for model in plan.get("targets") or []]
        phase_models: list[tuple[str, list[str]]] = []
        first = [model for model in ("grand_sport_x", "zr1") if model in targets]
        if first:
            phase_models.append(("grand_sport_x_plus_zr1", first))
        if "zr1x" in targets:
            phase_models.append(("zr1x_repeatability", ["zr1x"]))
        phase_models.append(("all_targets_atomic", targets))

        phases: list[dict[str, Any]] = []
        final: dict[str, Any] = {}
        for phase_id, models in phase_models:
            phase_plan = (
                plan if models == targets else self._deployment_phase_plan(plan, models)
            )
            continuity = self._deployment_continuity_probe(
                workbook,
                self._combined_plan_batch(phase_plan, workbook),
                phase_plan,
                schema_validation=schema_validation,
            )
            passed = bool(models) and all(
                (continuity.get(model) or {}).get("status") == "deployment_probe_passed"
                for model in models
            )
            phases.append(
                {
                    "phaseId": phase_id,
                    "models": models,
                    "passed": passed,
                    "continuity": continuity,
                }
            )
            if phase_id == "all_targets_atomic":
                final = continuity
        return phases, final

    def _expected_deployment_phase_specs(
        self,
        plan: dict[str, Any],
    ) -> list[tuple[str, list[str]]]:
        targets = [str(model) for model in plan.get("targets") or []]
        specs: list[tuple[str, list[str]]] = []
        first = [model for model in ("grand_sport_x", "zr1") if model in targets]
        if first:
            specs.append(("grand_sport_x_plus_zr1", first))
        if "zr1x" in targets:
            specs.append(("zr1x_repeatability", ["zr1x"]))
        specs.append(("all_targets_atomic", targets))
        return specs

    def _deployment_phase_blockers(
        self,
        plan: dict[str, Any],
        phases: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        expected = self._expected_deployment_phase_specs(plan)
        actual = [
            (
                str(phase.get("phaseId") or ""),
                [str(model) for model in phase.get("models") or []],
            )
            for phase in phases
        ]
        blockers: list[dict[str, Any]] = []
        if actual != expected:
            blockers.append(
                {
                    "kind": "deployment_proof_phases_invalid",
                    "detail": (
                        "required deployment proof phases are absent, reordered, or scoped "
                        f"incorrectly; expected {expected}, got {actual}"
                    ),
                }
            )
            return blockers
        for phase, (phase_id, models) in zip(phases, expected, strict=True):
            continuity = phase.get("continuity") or {}
            failed_models = [
                model
                for model in models
                if (continuity.get(model) or {}).get("status")
                != "deployment_probe_passed"
            ]
            if phase.get("passed") is not True or failed_models:
                blockers.append(
                    {
                        "kind": "deployment_proof_phase_failed",
                        "detail": (
                            f"deployment proof phase {phase_id!r} did not pass for "
                            f"{failed_models or models}"
                        ),
                    }
                )
        return blockers

    def _deployment_continuity_probe(
        self,
        workbook: Path,
        batch: dict[str, Any],
        plan: dict[str, Any],
        *,
        schema_validation: bool,
    ) -> dict[str, Any]:
        """Apply the plan to a temp workbook and run temp-only generation probes."""

        import shutil
        import tempfile

        from openpyxl import load_workbook

        from corvette_form_generator.contract import ASSET_IMAGE_FIELDS, load_model_asset_map
        from corvette_form_generator.editor_ops import apply_batch
        from corvette_form_generator.model_configs import discover_generation_model_configs
        from corvette_form_generator.registry_promotion import (
            assert_runtime_contract,
            build_registry_from_artifacts,
            registry_model_key,
            runtime_contract_artifact_path,
        )
        from corvette_form_generator.source_assembly import assemble_model_source

        targets = [str(model) for model in plan.get("targets") or []]
        continuity = self._deployment_continuity_from_plan(plan)
        if not targets:
            return continuity
        with tempfile.TemporaryDirectory(prefix="ingest-d1-deployment-") as tmp_dir:
            tmp_root = Path(tmp_dir)
            tmp_workbook = tmp_root / workbook.name
            shutil.copy2(workbook, tmp_workbook)
            temp_batch = dict(batch, workbookMtimeNs=str(tmp_workbook.stat().st_mtime_ns))
            preview = apply_batch(tmp_workbook, temp_batch, write=False, run_schema_validation=schema_validation)
            confirmed = [warning["id"] for warning in preview.get("warnings", [])]
            applied = apply_batch(
                tmp_workbook,
                temp_batch,
                write=True,
                confirmed_warnings=confirmed,
                source="ingest_wizard_deployment_probe",
                log_path=tmp_root / "probe-edit-log.jsonl",
                # This is a saved scratch workbook, so the writer boundary
                # always keeps schema validation enabled. Compact tests patch
                # the issue scan rather than disabling the write gate.
                run_schema_validation=True,
            )
            if not applied.get("ok"):
                error = "; ".join(applied.get("errors", [])) or str(applied.get("status"))
                for model in targets:
                    entry = continuity.setdefault(model, {})
                    blockers = list(entry.get("deploymentBlockers") or [])
                    blockers.append({"kind": "deployment_probe_apply_failed", "detail": error})
                    entry.update(
                        {
                            "status": "probe_apply_failed",
                            "registryLoadable": False,
                            "registryError": error,
                            "deploymentBlockers": blockers,
                        }
                    )
                return continuity

            pre_activation_workbook = tmp_root / "pre-activation.xlsx"
            shutil.copy2(tmp_workbook, pre_activation_workbook)
            self._activate_probe_models(tmp_workbook, targets)
            activation_validation = self._validate_probe_activation(
                pre_activation_workbook,
                tmp_workbook,
                targets,
            )
            for model in targets:
                continuity.setdefault(model, {})["postActivationValidation"] = (
                    activation_validation
                )
            if not activation_validation.get("ok"):
                detail = "; ".join(activation_validation.get("errors") or [])
                for model in targets:
                    entry = continuity.setdefault(model, {})
                    blockers = list(entry.get("deploymentBlockers") or [])
                    blockers.append(
                        {
                            "kind": "post_activation_validation_failed",
                            "detail": detail or "temporary activation validation failed",
                        }
                    )
                    entry.update(
                        {
                            "status": "not_deployment_ready",
                            "registryLoadable": False,
                            "registryError": detail,
                            "deploymentBlockers": blockers,
                        }
                    )
                return continuity
            try:
                configs = discover_generation_model_configs(tmp_workbook)
            except Exception as exc:
                configs = {}
                discovery_error = str(exc)
            else:
                discovery_error = ""

            for model in targets:
                entry = continuity.setdefault(model, {})
                blockers = list(entry.get("deploymentBlockers") or [])
                deferrals = list(entry.get("deploymentDeferrals") or [])
                config = configs.get(model)
                if config is None:
                    error = discovery_error or f"model {model!r} was not discoverable after temp activation"
                    blockers.append({"kind": "registry_load_failed", "detail": error})
                    entry.update(
                        {
                            "status": "not_deployment_ready",
                            "registryLoadable": False,
                            "registryError": error,
                            "deploymentBlockers": blockers,
                            "deploymentDeferrals": deferrals,
                        }
                    )
                    continue

                config = config.with_overrides(
                    root=tmp_root,
                    workbook_path=tmp_workbook,
                    output_dir=tmp_root / "form-output",
                    app_dir=tmp_root / "form-app",
                )
                entry["probePaths"] = {
                    "root": str(tmp_root),
                    "workbook": str(tmp_workbook),
                    "outputDir": str(tmp_root / "form-output"),
                    "appDir": str(tmp_root / "form-app"),
                    "allTemporary": all(
                        path.is_relative_to(tmp_root)
                        for path in (
                            tmp_workbook,
                            tmp_root / "form-output",
                            tmp_root / "form-app",
                        )
                    ),
                }
                try:
                    assembly = assemble_model_source(config)
                except Exception as exc:
                    blockers.append({"kind": "registry_load_failed", "detail": str(exc)})
                    entry.update(
                        {
                            "status": "not_deployment_ready",
                            "registryLoadable": False,
                            "registryError": str(exc),
                            "deploymentBlockers": blockers,
                            "deploymentDeferrals": deferrals,
                        }
                    )
                    continue

                source = assembly.source_data
                contract = getattr(assembly, "runtime_contract", source)
                try:
                    assert_runtime_contract(
                        contract, source=f"temporary deployment proof:{model}"
                    )
                    runtime_path = runtime_contract_artifact_path(tmp_root, model)
                    runtime_path.parent.mkdir(parents=True, exist_ok=True)
                    runtime_path.write_text(
                        json.dumps(contract, indent=2, ensure_ascii=False) + "\n",
                        encoding="utf-8",
                    )

                except Exception as exc:
                    blockers.append(
                        {"kind": "runtime_contract_invalid", "detail": str(exc)}
                    )
                raw_validation = source.get("validation")
                validation: list[dict[str, Any]] = raw_validation if isinstance(raw_validation, list) else []
                pricing_deferred = any(row.get("check_id") == "pricing_deferred" for row in validation if isinstance(row, dict))
                raw_choices = contract.get("choices")
                choices: list[dict[str, Any]] = raw_choices if isinstance(raw_choices, list) else []
                raw_interiors = contract.get("interiors")
                interiors: list[dict[str, Any]] = raw_interiors if isinstance(raw_interiors, list) else []
                media_count = sum(
                    1
                    for choice in choices
                    if isinstance(choice, dict) and any(choice.get(field) for field in ASSET_IMAGE_FIELDS)
                )
                component_count = sum(
                    len(interior.get("interior_components") or [])
                    for interior in interiors
                    if isinstance(interior, dict) and isinstance(interior.get("interior_components"), list)
                )
                counts = {
                    "choices": len(choices),
                    "customerSelectableChoices": sum(
                        1
                        for choice in choices
                        if isinstance(choice, dict)
                        and str(choice.get("selectable") or "").strip().lower()
                        in {"true", "1", "yes"}
                        and str(choice.get("active") or "true").strip().lower()
                        not in {"false", "0", "no"}
                    ),
                    "directRules": self._source_count(contract, "rules"),
                    "ruleGroups": self._source_count(contract, "ruleGroups"),
                    "exclusiveGroups": self._source_count(contract, "exclusiveGroups"),
                    "priceRules": self._source_count(contract, "priceRules"),
                    "pricingDeferred": pricing_deferred,
                    "colorOverrides": self._source_count(contract, "colorOverrides"),
                    "interiors": len(interiors),
                    "interiorComponentLineItems": component_count,
                    "optionMediaCoveredChoices": media_count,
                    "optionMediaTotalChoices": len(choices),
                    "validationWarnings": sum(1 for row in validation if isinstance(row, dict) and row.get("severity") == "warning"),
                    "validationErrors": sum(1 for row in validation if isinstance(row, dict) and row.get("severity") == "error"),
                }
                if counts["validationErrors"]:
                    blockers.append(
                        {
                            "kind": "generated_validation_errors",
                            "detail": (
                                f"generated contract has {counts['validationErrors']} validation error(s)"
                            ),
                        }
                    )
                if counts["customerSelectableChoices"] == 0:
                    blockers.append(
                        {
                            "kind": "customer_selectable_choices_missing",
                            "detail": "generated contract has zero active customer-selectable choices",
                        }
                    )
                signature_mismatches = self._manifest_runtime_signature_mismatches(
                    plan, model, contract
                )
                source_feature_agreement = self._source_feature_coverage_agreement(
                    plan,
                    model,
                    signature_mismatches,
                )
                if signature_mismatches:
                    blockers.append(
                        {
                            "kind": "manifest_runtime_signature_mismatch",
                            "detail": (
                                f"{len(signature_mismatches)} manifest/runtime semantic "
                                "mismatch(es)"
                            ),
                        }
                    )
                if not source_feature_agreement["ok"]:
                    blockers.append(
                        {
                            "kind": "source_feature_coverage_disagreement",
                            "detail": "; ".join(source_feature_agreement["errors"]),
                        }
                    )
                if model in {"zr1", "zr1x"} and counts["priceRules"] == 0 and pricing_deferred:
                    blockers.append({"kind": "price_rules_required_for_runtime", "detail": "generated contract has pricing_deferred and zero priceRules"})
                if model in {"zr1", "zr1x"} and counts["ruleGroups"] == 0:
                    blockers.append({"kind": "rule_groups_required_for_runtime", "detail": "generated contract has zero ruleGroups"})
                if counts["colorOverrides"] == 0:
                    blockers.append(
                        {
                            "kind": "color_overrides_missing_or_unproven",
                            "detail": "generated contract has zero colorOverrides without not-applicable proof",
                        }
                    )
                if counts["interiors"] == 0:
                    blockers.append(
                        {
                            "kind": "interiors_missing_or_unproven",
                            "detail": "generated contract has zero interiors without not-applicable proof",
                        }
                    )
                if counts["interiorComponentLineItems"] == 0:
                    blockers.append(
                        {
                            "kind": "interior_components_missing_or_unproven",
                            "detail": (
                                "generated contract has zero interior component line items "
                                "without not-applicable proof"
                            ),
                        }
                    )
                if counts["optionMediaCoveredChoices"] == 0:
                    blockers.append(
                        {
                            "kind": "asset_map_media_missing",
                            "detail": "generated choices have zero option/card asset fields",
                        }
                    )
                entry.update(
                    {
                        "status": "not_deployment_ready" if blockers else "deployment_probe_passed",
                        "registryLoadable": None,
                        "registryError": "registry proof pending",
                        "counts": counts,
                        "signatureMismatches": signature_mismatches,
                        "sourceFeatureCoverageAgreement": source_feature_agreement,
                        "validationErrors": [
                            row
                            for row in validation
                            if isinstance(row, dict) and row.get("severity") == "error"
                        ][:50],
                        "validationWarnings": [
                            row
                            for row in validation
                            if isinstance(row, dict) and row.get("severity") == "warning"
                        ][:50],
                        "deploymentBlockers": blockers,
                        "deploymentDeferrals": deferrals,
                    }
                )
            registry_error = ""
            registry_models: set[str] = set()
            registry_promotion_rows: list[dict[str, Any]] = []
            try:
                wb = load_workbook(tmp_workbook, read_only=True, data_only=True)
                try:
                    promotion_sheet = wb["model_registry_promotion"]
                    promotion_headers = [cell.value for cell in promotion_sheet[1]]
                    registry_promotion_rows = [
                        {
                            str(header): value
                            for header, value in zip(promotion_headers, values)
                            if header is not None
                        }
                        for values in promotion_sheet.iter_rows(
                            min_row=2, values_only=True
                        )
                        if str(values[0] or "") in targets
                    ]
                    registry = build_registry_from_artifacts(
                        wb,
                        model_assets=load_model_asset_map(wb, registry_model_key),
                        root=tmp_root,
                    )
                finally:
                    wb.close()
                registry_models = set((registry.get("models") or {}).keys())
            except Exception as exc:
                registry_error = str(exc)
            expected_registry_models = {registry_model_key(model) for model in targets}
            unexpected_registry_models = registry_models - expected_registry_models
            for model in targets:
                entry = continuity.setdefault(model, {})
                blockers = list(entry.get("deploymentBlockers") or [])
                registry_key = registry_model_key(model)
                if registry_error or registry_key not in registry_models or unexpected_registry_models:
                    detail = registry_error or (
                        f"registry models were {sorted(registry_models)}, expected exactly "
                        f"{sorted(expected_registry_models)}"
                    )
                    blockers.append({"kind": "registry_load_failed", "detail": detail})
                    entry["registryLoadable"] = False
                    entry["registryError"] = detail
                else:
                    entry["registryLoadable"] = True
                    entry["registryError"] = ""
                entry["registryModels"] = sorted(registry_models)
                entry["registryPromotionRows"] = registry_promotion_rows
                entry["deploymentBlockers"] = blockers
                entry["status"] = (
                    "not_deployment_ready" if blockers else "deployment_probe_passed"
                )
            return continuity


DEPLOYMENT_PROOF_SCHEMA = "workbook-changeset-deployment-proof-1"
_TEMP_PATH_RE = re.compile(r"/[^\s]*/ingest-d1-deployment-[^/\s]+")


class _ChangeSetDeploymentProofEngine(TemporaryDeploymentProofMixin):
    """Run temp generator/registry checks without legacy plan policy."""

    def _deployment_continuity_from_plan(self, plan: dict[str, Any]) -> dict[str, Any]:
        return {
            str(model): {
                "status": "deployment_probe_pending",
                "registryLoadable": None,
                "deploymentBlockers": [],
                "deploymentDeferrals": [],
            }
            for model in plan.get("targets") or []
        }


def _proof_fingerprint(payload: dict[str, Any]) -> str:
    body = {key: value for key, value in payload.items() if key != "proofFingerprint"}
    return hashlib.sha256(canonical_json(body).encode("utf-8")).hexdigest()


def _normalize_temporary_paths(value: Any) -> Any:
    """Remove random temp-root names from the durable proof receipt."""

    if isinstance(value, dict):
        return {
            str(key): _normalize_temporary_paths(item)
            for key, item in value.items()
        }
    if isinstance(value, list):
        return [_normalize_temporary_paths(item) for item in value]
    if isinstance(value, str):
        return _TEMP_PATH_RE.sub("<temporary-root>", value)
    return value


def _manifest_rows_with_refs(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {**dict(row), "manifestRef": f"manifest-{index:05d}"}
        for index, row in enumerate(manifest.get("rows") or [])
    ]


def _change_models(
    change: dict[str, Any],
    manifest_models: dict[str, str],
) -> set[str]:
    provenance = list(change.get("provenance") or [])
    if len(provenance) != 1:
        return set()
    entry = provenance[0]
    kind = str(entry.get("kind") or "")
    if kind == "manifest":
        manifest_ref = str(entry.get("manifestRef") or entry.get("id") or "")
        if manifest_ref not in manifest_models:
            return set()
        return {manifest_models[manifest_ref]}
    if (
        kind == "scaffold"
        and str(entry.get("id") or "") == "pass_c3_greenfield_registry_promotion"
        and str(change.get("family") or "") == "model_registry_promotion"
    ):
        models: set[str] = set()
        key_model = str((change.get("key") or {}).get("model_key") or "")
        if key_model:
            models.add(key_model)
        field_model = str(
            ((change.get("fields") or {}).get("model_key") or {}).get("after") or ""
        )
        if field_model:
            models.add(field_model)
        return models if len(models) == 1 else set()
    return set()


def _phase_changeset(
    changeset: dict[str, Any],
    manifest_rows: list[dict[str, Any]],
    targets: list[str],
) -> dict[str, Any]:
    """Build a mechanically filtered, valid ChangeSet for one proof phase."""

    selected = set(targets)
    manifest_models = {
        str(row["manifestRef"]): str(row.get("model") or "")
        for row in manifest_rows
    }

    def include(change: dict[str, Any]) -> bool:
        models = _change_models(change, manifest_models)
        return "*" in models or bool(models & selected)

    row_changes = [dict(change) for change in changeset["rowChanges"] if include(change)]
    noops = [dict(change) for change in changeset["noops"] if include(change)]
    referenced_sheets = {
        str(change.get("sheet") or "") for change in [*row_changes, *noops]
    }
    referenced_sheets.update(
        str(row.get("sheet") or "")
        for row in manifest_rows
        if str(row.get("model") or "") in selected | {"*"}
    )
    for change in row_changes:
        for pair in (change.get("fields") or {}).values():
            after = pair.get("after") if isinstance(pair, dict) else None
            if isinstance(after, str):
                referenced_sheets.add(after)
    sheet_creates = [
        dict(create)
        for create in changeset["sheetCreates"]
        if str(create.get("sheet") or "") in referenced_sheets
    ]
    phase = {
        **changeset,
        "targets": sorted(targets),
        "sheetCreates": sheet_creates,
        "rowChanges": row_changes,
        "noops": noops,
    }
    phase["semanticFingerprint"] = changeset_fingerprint(phase)
    phase["changeSetId"] = phase["semanticFingerprint"][:24]
    return parse_changeset(phase)


def _proof_context(
    manifest: dict[str, Any],
    compile_report: dict[str, Any],
    manifest_rows: list[dict[str, Any]],
    targets: list[str],
) -> dict[str, Any]:
    selected = set(targets)
    selected_rows = []
    for row in manifest_rows:
        if str(row.get("model") or "") not in selected | {"*"}:
            continue
        item = copy.deepcopy(row)
        item["canonicalValues"] = copy.deepcopy(item.get("values") or {})
        selected_rows.append(item)
    raw_feature_coverage = compile_report.get("sourceFeatureCoverage") or []
    if isinstance(raw_feature_coverage, list):
        by_model: dict[str, dict[str, Any]] = {}
        for model in targets:
            features = [
                copy.deepcopy(feature)
                for feature in raw_feature_coverage
                if str(feature.get("model") or "") in {model, "*"}
            ]
            dispositions = Counter(
                str(feature.get("disposition") or "") for feature in features
            )
            by_model[model] = {
                "featureCount": len(features),
                "dispositionCounts": dict(sorted(dispositions.items())),
                "blockingFeatures": [
                    feature
                    for feature in features
                    if str(feature.get("disposition") or "")
                    in {"exception_open", "unsupported_blocker"}
                ],
            }
        source_feature_coverage = {
            "semanticSha": semantic_hash(raw_feature_coverage),
            "byModel": by_model,
        }
    else:
        source_feature_coverage = copy.deepcopy(raw_feature_coverage)
    return {
        "targets": list(targets),
        "targetModes": {
            model: (manifest.get("modelModes") or {}).get(model)
            for model in targets
        },
        "coverage": {
            "manifestRows": selected_rows,
        },
        "sourceFeatureCoverage": source_feature_coverage,
    }


def _proof_error(status: str, message: str) -> dict[str, Any]:
    return {"ok": False, "status": status, "errors": [message]}


def prove_changeset_deployment(
    workbook_path: Path | str,
    changeset: dict[str, Any],
    *,
    canonical_manifest_path: Path | str,
    compile_report_path: Path | str,
) -> dict[str, Any]:
    """Prove an exact ChangeSet through ordered, isolated temp-workbook phases."""

    workbook = Path(workbook_path)
    try:
        parsed = parse_changeset(changeset)
    except ChangeSetError as exc:
        return _proof_error("invalid_changeset", str(exc))
    manifest_path = Path(canonical_manifest_path)
    report_path = Path(compile_report_path)
    try:
        manifest_bytes = manifest_path.read_bytes()
        report_bytes = report_path.read_bytes()
        manifest = json.loads(manifest_bytes)
        compile_report = json.loads(report_bytes)
    except (OSError, json.JSONDecodeError) as exc:
        return _proof_error("invalid_proof_input", str(exc))
    if not isinstance(manifest, dict) or not isinstance(compile_report, dict):
        return _proof_error(
            "invalid_proof_input",
            "canonical manifest and compile report must be JSON objects",
        )
    manifest_sha256 = hashlib.sha256(manifest_bytes).hexdigest()
    compile_report_sha256 = hashlib.sha256(report_bytes).hexdigest()
    try:
        validate_artifact_graph(manifest, compile_report)
    except ValueError as exc:
        return _proof_error("binding_mismatch", str(exc))
    live_workbook = {
        "sha256": hashlib.sha256(workbook.read_bytes()).hexdigest(),
        "mtimeNs": str(workbook.stat().st_mtime_ns),
    }
    if live_workbook != parsed["workbook"]:
        return _proof_error(
            "stale",
            "live workbook no longer matches the ChangeSet workbook binding",
        )
    manifest_semantic = str(manifest.get("manifestSemanticSha") or "")
    if manifest_semantic != str(
        (parsed.get("bindings") or {}).get("canonicalManifestSemanticSha") or ""
    ):
        return _proof_error(
            "binding_mismatch",
            "canonical manifest semantic SHA does not match the ChangeSet binding",
        )
    bindings = parsed.get("bindings") or {}
    compiler_bindings = bindings.get("compilerBindings") or {}
    if (
        manifest_sha256 != str(bindings.get("canonicalManifestSha") or "")
        or compile_report_sha256
        != str(compiler_bindings.get("compileReportSha") or "")
    ):
        return _proof_error(
            "binding_mismatch",
            "manifest or compile-report file SHA does not match the ChangeSet binding",
        )
    targets = [str(target) for target in parsed["targets"]]
    if targets != ["grand_sport_x", "zr1", "zr1x"]:
        return _proof_error(
            "binding_mismatch",
            "deployment proof requires exact Task 8 targets: grand_sport_x, zr1, zr1x",
        )
    manifest_modes = {
        str(model): str(mode)
        for model, mode in (manifest.get("modelModes") or {}).items()
    }
    report_models = {
        str(model): dict(result or {})
        for model, result in (compile_report.get("models") or {}).items()
    }
    if set(targets) != set(manifest_modes) or set(targets) != set(report_models):
        return _proof_error(
            "binding_mismatch",
            "ChangeSet, manifest, and compile report target sets do not match",
        )
    not_ready = [
        model
        for model in targets
        if report_models[model].get("compileReady") is not True
        or report_models[model].get("blockers")
        or str(report_models[model].get("mode") or "") != manifest_modes[model]
    ]
    if not_ready or compile_report.get("deferrals"):
        detail = not_ready or ["compile report contains deferrals"]
        return _proof_error(
            "compiler_not_ready",
            f"compiler is not ready for exact targets: {detail}",
        )

    manifest_rows = _manifest_rows_with_refs(manifest)
    manifest_models = {
        str(row["manifestRef"]): str(row.get("model") or "")
        for row in manifest_rows
    }
    manifest_by_ref = {
        str(row["manifestRef"]): row
        for row in manifest_rows
    }
    unbound_changes = [
        change
        for change in [*parsed["rowChanges"], *parsed["noops"]]
        if not _change_models(change, manifest_models)
    ]
    if unbound_changes:
        return _proof_error(
            "phase_projection_invalid",
            "ChangeSet row lacks exact manifest or scaffold model authority",
        )
    projection_mismatches = []
    for change in [*parsed["rowChanges"], *parsed["noops"]]:
        provenance = list(change.get("provenance") or [])
        if str((provenance[0] if provenance else {}).get("kind") or "") != "manifest":
            continue
        manifest_ref = str(
            provenance[0].get("manifestRef") or provenance[0].get("id") or ""
        )
        row = manifest_by_ref[manifest_ref]
        structural_match = all(
            change.get(field) == row.get(field)
            for field in ("family", "sheet", "key")
        )
        if change.get("action") == "noop":
            value_match = change.get("canonicalValues") == row.get("values")
        else:
            fields = change.get("fields") or {}
            value_match = all(
                isinstance(pair, dict)
                and pair.get("after") == (row.get("values") or {}).get(column)
                for column, pair in fields.items()
            )
            if change.get("action") == "add":
                missing_fields = set(row.get("values") or {}) - set(fields)
                value_match = value_match and all(
                    (row.get("values") or {}).get(column) in (None, "")
                    for column in missing_fields
                )
        if not structural_match or not value_match:
            projection_mismatches.append(manifest_ref)
    if projection_mismatches:
        return _proof_error(
            "phase_projection_invalid",
            "ChangeSet row does not match its manifest row: "
            + ", ".join(projection_mismatches[:5]),
        )
    manifest_ref_counts = Counter(
        str(entry.get("manifestRef") or entry.get("id") or "")
        for change in [*parsed["rowChanges"], *parsed["noops"]]
        for entry in change.get("provenance") or []
        if str(entry.get("kind") or "") == "manifest"
    )
    expected_manifest_refs = set(manifest_models)
    if (
        set(manifest_ref_counts) != expected_manifest_refs
        or any(count != 1 for count in manifest_ref_counts.values())
    ):
        return _proof_error(
            "phase_projection_invalid",
            "ChangeSet manifest coverage is incomplete or duplicated",
        )
    phase_specs = [
        (
            "grand_sport_x_plus_zr1",
            [model for model in ("grand_sport_x", "zr1") if model in targets],
        ),
        ("zr1x_repeatability", ["zr1x"] if "zr1x" in targets else []),
        ("all_targets_atomic", targets),
    ]
    engine = _ChangeSetDeploymentProofEngine()
    extract = editor_ops.extract_workbook(workbook)
    phases: list[dict[str, Any]] = []
    all_blockers: list[dict[str, Any]] = []
    all_deferrals: list[dict[str, Any]] = []
    for phase_id, phase_targets in phase_specs:
        if not phase_targets:
            continue
        phase_changeset = _phase_changeset(parsed, manifest_rows, phase_targets)
        batch = changeset_to_editor_batch(phase_changeset, extract)
        context = _proof_context(
            manifest,
            compile_report,
            manifest_rows,
            phase_targets,
        )
        continuity = engine._deployment_continuity_probe(
            workbook,
            batch,
            context,
            schema_validation=True,
        )
        blockers = [
            {"phaseId": phase_id, "model": model, **dict(blocker)}
            for model in phase_targets
            for blocker in (continuity.get(model) or {}).get("deploymentBlockers") or []
        ]
        deferrals = [
            {"phaseId": phase_id, "model": model, **dict(deferral)}
            for model in phase_targets
            for deferral in (continuity.get(model) or {}).get("deploymentDeferrals") or []
        ]
        passed = not blockers and not deferrals and all(
            (continuity.get(model) or {}).get("status") == "deployment_probe_passed"
            for model in phase_targets
        )
        phases.append(
            {
                "phaseId": phase_id,
                "targets": phase_targets,
                "passed": passed,
                "changeSetId": phase_changeset["changeSetId"],
                "operationCounts": {
                    "sheetCreates": len(phase_changeset["sheetCreates"]),
                    "rowChanges": len(phase_changeset["rowChanges"]),
                    "noops": len(phase_changeset["noops"]),
                },
                "continuity": continuity,
            }
        )
        all_blockers.extend(blockers)
        all_deferrals.extend(deferrals)

    passed = bool(phases) and not all_blockers and not all_deferrals and all(
        phase["passed"] for phase in phases
    )
    receipt = _normalize_temporary_paths(
        {
            "ok": passed,
            "schemaVersion": DEPLOYMENT_PROOF_SCHEMA,
            "status": (
                "deployment_proof_passed" if passed else "deployment_proof_blocked"
            ),
            "changeSetId": parsed["changeSetId"],
            "semanticFingerprint": parsed["semanticFingerprint"],
            "workbook": dict(parsed["workbook"]),
            "bindings": dict(parsed.get("bindings") or {}),
            "targets": targets,
            "phases": phases,
            "blockers": all_blockers,
            "deferrals": all_deferrals,
            "errors": [],
        }
    )
    receipt["proofFingerprint"] = _proof_fingerprint(receipt)
    return receipt
