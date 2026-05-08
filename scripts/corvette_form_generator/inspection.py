"""Read-only model source inspection helpers."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.mapping import best_status, normalize_mode, selection_mode_label, status_to_label, step_for_section
from corvette_form_generator.model_config import ModelConfig
from corvette_form_generator.workbook import clean, intish, money, rows_from_sheet


ALLOWED_STATUSES = {"available", "standard", "unavailable"}
STATUS_ALIASES = {
    "available": "available",
    "standard": "standard",
    "not available": "unavailable",
    "unavailable": "unavailable",
}
RULE_HOT_SPOT_PATTERNS = {
    "requires": re.compile(r"\brequires?\b", re.IGNORECASE),
    "not_available": re.compile(r"\bnot available\b", re.IGNORECASE),
    "included_with": re.compile(r"\bincluded with\b", re.IGNORECASE),
    "includes": re.compile(r"\bincludes?\b", re.IGNORECASE),
    "only": re.compile(r"\bonly\b", re.IGNORECASE),
    "replaces": re.compile(r"\breplaces?\b", re.IGNORECASE),
    "not_recommended": re.compile(r"\bnot recommended\b", re.IGNORECASE),
    "except": re.compile(r"\bexcept\b", re.IGNORECASE),
}
SPECIAL_REVIEW_RPOS = {"EL9", "Z25", "FEY", "Z15"}
GRAND_SPORT_ONLY_INTERIOR_IDS = {"3LT_AE4_EL9", "3LT_AH2_EL9"}
INTERIOR_COMPONENT_LABELS = {
    "36S": "Yellow Stitching",
    "37S": "Blue Stitching",
    "38S": "Red Stitching",
    "N26": "Sueded Microfiber",
    "N2Z": "Sueded Microfiber",
    "TU7": "Two-Tone",
    "R6X": "Custom Interior Trim and Seat Combination",
}


def normalize_status(value: Any) -> str:
    text = clean(value).lower()
    return STATUS_ALIASES.get(text, text)


def normalize_selectable(value: Any) -> str:
    text = clean(value).lower()
    if text in {"yes", "true", "1", "y"}:
        return "True"
    if text in {"no", "false", "0", "n"}:
        return "False"
    return clean(value)


def normalized_option_row(row: dict[str, str], config: ModelConfig) -> dict[str, Any]:
    option_id = clean(row.get("option_id", ""))
    original_section_id = clean(row.get("section_id", ""))
    resolved_section_id = original_section_id or config.blank_section_overrides.get(option_id, "")
    source_category_id = (
        clean(row.get("category_id", ""))
        or config.source_category_overrides.get(option_id, "")
        or config.option_category_overrides.get(option_id, "")
        or config.section_category_overrides.get(resolved_section_id, "")
    )
    return {
        "option_id": option_id,
        "rpo": clean(row.get("rpo", "")),
        "price": money(row.get("price")),
        "option_name": clean(row.get("option_name", "")),
        "description": clean(row.get("description", "")),
        "detail_raw": clean(row.get("detail_raw", "")),
        "category_id": source_category_id,
        "selectable": normalize_selectable(row.get("selectable", "")),
        "active": normalize_selectable(row.get("active", "True")),
        "display_behavior": clean(row.get("display_behavior", "")),
        "original_section_id": original_section_id,
        "section_id": resolved_section_id,
        "section_override_applied": bool(not original_section_id and resolved_section_id),
        "statuses": {variant_id: normalize_status(row.get(variant_id, "")) for variant_id in config.variant_ids},
    }


def status_lookup_from_sheet(wb, config: ModelConfig) -> dict[tuple[str, str], str]:
    if config.status_sheet not in wb.sheetnames:
        return {}
    statuses: dict[tuple[str, str], str] = {}
    for row in rows_from_sheet(wb, config.status_sheet):
        option_id = clean(row.get("option_id", ""))
        variant_id = clean(row.get("variant_id", ""))
        if option_id and variant_id:
            statuses[(option_id, variant_id)] = normalize_status(row.get("status", ""))
    return statuses


def apply_status_lookup(rows: list[dict[str, Any]], status_lookup: dict[tuple[str, str], str], config: ModelConfig) -> None:
    if not status_lookup:
        return
    for row in rows:
        row["statuses"] = {
            variant_id: status_lookup.get((row["option_id"], variant_id), "")
            for variant_id in config.variant_ids
        }


def cleanup_display_text(value: str, config: ModelConfig) -> tuple[str, list[str]]:
    notes: list[str] = []
    original = clean(value)
    text = original
    if not text or not config.text_cleanup.get("enabled"):
        return text, notes

    collapsed = re.sub(r"\s+", " ", text).strip()
    if collapsed != text:
        notes.append("collapsed_whitespace")
        text = collapsed

    punctuation = re.sub(r"([!?.,])\1+", r"\1", text)
    punctuation = re.sub(r"\s+([,.;:!?])", r"\1", punctuation)
    if punctuation != text:
        notes.append("collapsed_repeated_punctuation")
        text = punctuation

    if config.text_cleanup.get("normalize_new_prefix"):
        normalized_new = re.sub(r"^NEW!\s*", "New ", text, flags=re.IGNORECASE)
        if normalized_new != text:
            notes.append("normalized_new_prefix")
            text = normalized_new

    exact_replacements = {
        "New Ground effects": "New Ground Effects",
    }
    replacement = exact_replacements.get(text)
    if replacement:
        notes.append("normalized_capitalization")
        text = replacement

    if config.text_cleanup.get("remove_adjacent_duplicate_phrases"):
        deduped = re.sub(r"\b(\w+)(\s+\1\b)+", r"\1", text, flags=re.IGNORECASE)
        if deduped != text:
            notes.append("removed_adjacent_duplicate_word")
            text = deduped

    return text, notes


def price_ref_key(trim: str, code: str) -> tuple[str, str]:
    return (clean(trim).replace("_", " "), clean(code))


def price_ref_prices(rows: list[dict[str, str]]) -> dict[tuple[str, str], int]:
    prices: dict[tuple[str, str], int] = {}
    for row in rows:
        if clean(row.get("OptionType", "")).lower() != "seat":
            continue
        trim = clean(row.get("Trim", ""))
        code = clean(row.get("Code", ""))
        if trim and code:
            prices[price_ref_key(trim, code)] = money(row.get("Price"))
    return prices


def price_ref_component_prices(rows: list[dict[str, str]]) -> dict[tuple[str, str, str], int]:
    prices: dict[tuple[str, str, str], int] = {}
    for row in rows:
        option_type = clean(row.get("OptionType", "")).lower()
        code = clean(row.get("Code", ""))
        if not option_type or not code:
            continue
        prices[(option_type, clean(row.get("Trim", "")).replace("_", " "), code)] = money(row.get("Price"))
    return prices


def price_ref_component_price(
    price_ref: dict[tuple[str, str, str], int],
    option_type: str,
    code: str,
    trim: str = "",
) -> int:
    normalized_type = clean(option_type).lower()
    normalized_trim = clean(trim).replace("_", " ")
    normalized_code = clean(code)
    if (normalized_type, normalized_trim, normalized_code) in price_ref:
        return price_ref[(normalized_type, normalized_trim, normalized_code)]
    return price_ref.get((normalized_type, "", normalized_code), 0)


def r6x_price_component(row: dict[str, str], price_ref: dict[tuple[str, str], int]) -> int:
    trim = clean(row.get("Trim", ""))
    interior_id = clean(row.get("interior_id", "") or row.get("ID", ""))
    if "R6X" not in trim and "R6X" not in interior_id:
        return 0

    seat = clean(row.get("Seat", ""))
    r6x_trim = trim if "R6X" in trim else f"{trim}_R6X"
    base_trim = r6x_trim.replace("_R6X", "")
    r6x_price = price_ref.get(price_ref_key(r6x_trim, seat))
    if r6x_price is None:
        return 0
    return max(0, r6x_price - price_ref.get(price_ref_key(base_trim, seat), 0))


def generated_interior_price(row: dict[str, str], price_ref: dict[tuple[str, str], int]) -> int:
    return money(row.get("Price") or row.get("Cost")) + r6x_price_component(row, price_ref)


def interior_component_metadata(
    row: dict[str, str],
    price_ref: dict[tuple[str, str, str], int],
) -> list[dict[str, Any]]:
    trim = clean(row.get("Trim", ""))
    interior_id = clean(row.get("interior_id", "") or row.get("ID", ""))
    seat = clean(row.get("Seat", ""))
    tokens = set(interior_id.split("_"))
    components: list[dict[str, Any]] = []

    if "R6X" in trim or "R6X" in tokens:
        r6x_trim = trim if "R6X" in trim else f"{trim}_R6X"
        components.append(
            {
                "rpo": "R6X",
                "label": INTERIOR_COMPONENT_LABELS["R6X"],
                "price": price_ref_component_price(price_ref, "seat", seat, r6x_trim),
                "component_type": "r6x",
            }
        )
    else:
        seat_price = price_ref_component_price(price_ref, "seat", seat, trim)
        if seat_price:
            components.append(
                {
                    "rpo": seat,
                    "label": f"{seat} Seat Upgrade",
                    "price": seat_price,
                    "component_type": "seat",
                }
            )

    for rpo in ("36S", "37S", "38S"):
        if rpo in tokens:
            components.append(
                {
                    "rpo": rpo,
                    "label": INTERIOR_COMPONENT_LABELS[rpo],
                    "price": price_ref_component_price(price_ref, "stitching", rpo),
                    "component_type": "stitching",
                }
            )

    for rpo in ("N26", "N2Z"):
        if rpo in tokens:
            components.append(
                {
                    "rpo": rpo,
                    "label": INTERIOR_COMPONENT_LABELS[rpo],
                    "price": price_ref_component_price(price_ref, "suede", rpo),
                    "component_type": "suede",
                }
            )

    if "TU7" in tokens:
        components.append(
            {
                "rpo": "TU7",
                "label": INTERIOR_COMPONENT_LABELS["TU7"],
                "price": price_ref_component_price(price_ref, "twotone", "TU7"),
                "component_type": "two_tone",
            }
        )

    return [component for component in components if component["price"] or component["rpo"] == "R6X"]


def clean_reference_label(value: str) -> str:
    label = clean(value)
    if " - " in label:
        head, tail = label.split(" - ", 1)
        if re.search(r"\b(option|expandable|choice|card)\b", tail, re.IGNORECASE):
            label = head
    label = re.sub(r"\s*\([^)]*(?:expandable|only one option|no need)[^)]*\)\s*$", "", label, flags=re.IGNORECASE)
    return label.strip()


def read_interior_reference(config: ModelConfig) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    reference_by_id: dict[str, dict[str, Any]] = {}
    reference_rows: list[dict[str, Any]] = []
    current_levels = [""] * 6
    if not config.interior_reference_path.exists():
        return reference_by_id, reference_rows
    with config.interior_reference_path.open(newline="", encoding="utf-8") as handle:
        import csv

        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            for index in range(6):
                key = f"level{index}"
                value = clean_reference_label(row.get(key, ""))
                if value:
                    current_levels[index] = value
                    for deeper in range(index + 1, 6):
                        current_levels[deeper] = ""
            interior_id = clean(row.get("interior_id", ""))
            levels = [level for level in current_levels if level]
            record = {
                "row_number": row_number,
                "interior_id": interior_id,
                "levels": levels,
            }
            reference_rows.append(record)
            if interior_id:
                reference_by_id[interior_id] = record
    return reference_by_id, reference_rows


def seat_code_from_label(label: str) -> str:
    return clean(label).split(" ", 1)[0]


def grouping_fields_for_interior(
    interior: dict[str, Any],
    reference: dict[str, Any] | None,
    reference_order: int,
    fallback: bool = False,
) -> dict[str, Any]:
    seat_label = reference["levels"][1] if reference and len(reference["levels"]) > 1 else f"{interior['seat_code']} Seats"
    levels = reference["levels"] if reference else [
        interior["trim_level"],
        seat_label,
        interior["interior_name"] or "Other Interior Choices",
        interior["material"] or "Standard interior",
        interior["interior_name"] or interior["interior_id"],
    ]
    leaf_label = levels[-1] if levels else interior["interior_name"] or interior["interior_id"]
    color_family = levels[2] if len(levels) > 2 else leaf_label
    material_family = interior.get("material") or "Standard interior"
    if len(levels) > 3 and levels[-2] != color_family:
        material_family = levels[-2]
    parent_group = levels[-2] if len(levels) > 1 else color_family
    return {
        "interior_trim_level": levels[0] if levels else interior["trim_level"],
        "interior_seat_code": seat_code_from_label(seat_label) or interior["seat_code"],
        "interior_seat_label": seat_label,
        "interior_color_family": "Other Interior Choices" if fallback else color_family,
        "interior_material_family": material_family,
        "interior_variant_label": leaf_label,
        "interior_group_display_order": reference_order,
        "interior_material_display_order": reference_order,
        "interior_choice_display_order": reference_order,
        "interior_hierarchy_levels": json.dumps(levels, ensure_ascii=False),
        "interior_hierarchy_path": " > ".join(levels),
        "interior_parent_group_label": parent_group,
        "interior_leaf_label": leaf_label,
        "interior_reference_order": reference_order,
    }


def build_grand_sport_interiors(config: ModelConfig) -> list[dict[str, Any]]:
    wb = load_workbook(config.workbook_path, data_only=True, read_only=True)
    lt_interiors_raw = rows_from_sheet(wb, "lt_interiors")
    price_ref_rows = rows_from_sheet(wb, "PriceRef")
    interior_price_ref = price_ref_prices(price_ref_rows)
    interior_component_price_ref = price_ref_component_prices(price_ref_rows)
    reference_by_id, reference_rows = read_interior_reference(config)
    reference_order_by_id = {
        row["interior_id"]: index
        for index, row in enumerate((row for row in reference_rows if row["interior_id"]), start=1)
    }
    fallback_order = len(reference_order_by_id) + 1
    interiors: list[dict[str, Any]] = []

    for row in lt_interiors_raw:
        trim = clean(row.get("Trim", ""))
        if trim not in {"1LT", "2LT", "3LT", "3LT_R6X"}:
            continue
        interior_id = clean(row.get("interior_id", ""))
        components = interior_component_metadata(row, interior_component_price_ref)
        interior = {
            "interior_id": interior_id,
            "source_sheet": "lt_interiors",
            "active_for_stingray": False,
            "active_for_grand_sport": True,
            "requires_z25": "True" if interior_id in GRAND_SPORT_ONLY_INTERIOR_IDS else "False",
            "trim_level": trim.replace("_R6X", ""),
            "requires_r6x": "True" if "_R6X" in trim or interior_id.endswith("_R6X") else "False",
            "seat_code": clean(row.get("Seat", "")),
            "interior_code": clean(row.get("Interior Code", "")),
            "interior_name": clean(row.get("Interior Name", "")),
            "material": clean(row.get("Material", "")),
            "price": generated_interior_price(row, interior_price_ref),
            "suede": clean(row.get("Suede", "")),
            "stitch": clean(row.get("Stitch", "")),
            "two_tone": clean(row.get("Two Tone", "")),
            "section_id": clean(row.get("section_id", "")),
            "color_overrides_raw": clean(row.get("Color Overrides", "")),
            "source_note": clean(row.get("Detail from Disclosure", "")),
            "interior_components": components,
            "interior_components_json": json.dumps(components, separators=(",", ":")),
        }
        reference = reference_by_id.get(interior_id)
        if reference:
            interior.update(grouping_fields_for_interior(interior, reference, reference_order_by_id[interior_id]))
        else:
            interior.update(grouping_fields_for_interior(interior, None, fallback_order, fallback=False))
            fallback_order += 1
        interiors.append(interior)

    return interiors


def resolve_category(
    row: dict[str, Any],
    sections: dict[str, dict[str, str]],
    config: ModelConfig,
) -> tuple[str, str]:
    option_id = row["option_id"]
    section_id = row["section_id"]
    source_category_id = row["category_id"]
    section_category_id = sections.get(section_id, {}).get("category_id", "")

    if option_id in config.option_category_overrides:
        return config.option_category_overrides[option_id], "option_override"
    if section_id in config.section_category_overrides:
        return config.section_category_overrides[section_id], "section_override"
    if source_category_id:
        return source_category_id, "source"
    return section_category_id, "section_master"


def resolved_step_key(
    section_id: str,
    sections: dict[str, dict[str, str]],
    resolved_category_id: str,
    config: ModelConfig,
) -> str:
    section = sections.get(section_id, {})
    return step_for_section(
        section_id,
        section.get("section_name", ""),
        section.get("category_id", resolved_category_id),
        standard_sections=config.standard_sections,
        section_step_overrides=config.section_step_overrides,
    )


def classify_rule_hot_spots(
    rows: list[dict[str, Any]],
    config: ModelConfig,
    include_special_bucket: bool = False,
) -> dict[str, Any]:
    rule_hot_spots = []
    hot_spot_counts: Counter[str] = Counter()
    special_mentions: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    special_rpos = set(config.special_rule_review_rpos or tuple(SPECIAL_REVIEW_RPOS))
    for row in rows:
        text = "\n".join(part for part in (row["detail_raw"], row["description"], row["option_name"]) if part)
        matched_terms = [name for name, pattern in RULE_HOT_SPOT_PATTERNS.items() if pattern.search(text)]
        mentioned_specials = [rpo for rpo in sorted(special_rpos) if re.search(rf"\b{re.escape(rpo)}\b", text)]
        if include_special_bucket and (row["rpo"] in special_rpos or mentioned_specials):
            matched_terms.append("special_package_review")
        matched_terms = sorted(set(matched_terms), key=matched_terms.index)
        for term in matched_terms:
            hot_spot_counts[term] += 1
        if matched_terms or row["rpo"] in special_rpos or mentioned_specials:
            record = {
                "option_id": row["option_id"],
                "rpo": row["rpo"],
                "option_name": row["option_name"],
                "section_id": row["section_id"],
                "matched_terms": matched_terms,
                "special_mentions": mentioned_specials,
                "detail_raw": row["detail_raw"],
            }
            rule_hot_spots.append(record)
            for rpo in mentioned_specials:
                special_mentions[rpo].append(
                    {
                        "option_id": row["option_id"],
                        "rpo": row["rpo"],
                        "option_name": row["option_name"],
                    }
                )
    return {
        "counts": dict(sorted(hot_spot_counts.items())),
        "rows": rule_hot_spots,
        "special_mentions": {rpo: rows for rpo, rows in sorted(special_mentions.items())},
    }


def rows_from_optional_sheet(wb, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    return rows_from_sheet(wb, sheet_name)


def active_source_row(row: dict[str, str]) -> bool:
    return clean(row.get("active", "True")) == "True"


def display_behavior_status(
    status: str,
    selectable: str,
    active: str,
    display_behavior: str,
) -> tuple[str, str, str]:
    if display_behavior == "auto_only":
        return "unavailable", "False", "False"
    if display_behavior == "display_only":
        return "available", "False", "True"
    return status, selectable, active


def build_color_overrides(
    wb,
    config: ModelConfig,
    interiors: list[dict[str, Any]],
    option_rows: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    interior_ids = {row["interior_id"] for row in interiors if row.get("interior_id")}
    valid_option_ids = set(option_rows)
    rows: list[dict[str, Any]] = []
    for source in rows_from_optional_sheet(wb, config.color_overrides_sheet):
        interior_id = source.get("interior_id", "")
        option_id = source.get("option_id", "")
        adds_rpo = source.get("adds_rpo", "")
        if interior_id not in interior_ids or option_id not in valid_option_ids or adds_rpo not in valid_option_ids:
            continue
        rows.append(
            {
                "override_id": f"co_{len(rows) + 1:03d}",
                "interior_id": interior_id,
                "option_id": option_id,
                "rule_type": source.get("rule_type", "").lower(),
                "adds_rpo": adds_rpo,
                "notes": "Exterior/interior pairing requires the listed override RPO.",
            }
        )
    return rows


def load_rule_groups(wb, config: ModelConfig) -> list[dict[str, Any]]:
    members_by_group: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows_from_optional_sheet(wb, config.rule_group_members_sheet):
        if active_source_row(row):
            members_by_group[row.get("group_id", "")].append(row)

    rule_groups: list[dict[str, Any]] = []
    for row in rows_from_optional_sheet(wb, config.rule_groups_sheet):
        if not active_source_row(row):
            continue
        group_id = row.get("group_id", "")
        members = sorted(members_by_group.get(group_id, []), key=lambda member: intish(member.get("display_order")))
        rule_groups.append(
            {
                "group_id": group_id,
                "group_type": row.get("group_type", ""),
                "source_id": row.get("source_id", ""),
                "target_ids": [member.get("target_id", "") for member in members if member.get("target_id", "")],
                "body_style_scope": row.get("body_style_scope", ""),
                "trim_level_scope": row.get("trim_level_scope", ""),
                "variant_scope": row.get("variant_scope", ""),
                "disabled_reason": row.get("disabled_reason", ""),
                "active": row.get("active", ""),
                "notes": row.get("notes", ""),
            }
        )
    return rule_groups


def load_exclusive_groups(wb, config: ModelConfig) -> list[dict[str, Any]]:
    members_by_group: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows_from_optional_sheet(wb, config.exclusive_group_members_sheet):
        if active_source_row(row):
            members_by_group[row.get("group_id", "")].append(row)

    exclusive_groups: list[dict[str, Any]] = []
    for row in rows_from_optional_sheet(wb, config.exclusive_groups_sheet):
        if not active_source_row(row):
            continue
        group_id = row.get("group_id", "")
        members = sorted(members_by_group.get(group_id, []), key=lambda member: intish(member.get("display_order")))
        exclusive_groups.append(
            {
                "group_id": group_id,
                "option_ids": [member.get("option_id", "") for member in members if member.get("option_id", "")],
                "selection_mode": row.get("selection_mode", ""),
                "active": row.get("active", ""),
                "notes": row.get("notes", ""),
            }
        )
    return exclusive_groups


def grouped_requirement_pairs(rule_groups: list[dict[str, Any]]) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for group in rule_groups:
        if group.get("active") != "True" or group.get("group_type") != "requires_any":
            continue
        source_id = group.get("source_id", "")
        for target_id in group.get("target_ids", []):
            pairs.add((source_id, target_id))
    return pairs


def exclusive_group_pairs(exclusive_groups: list[dict[str, Any]]) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for group in exclusive_groups:
        if group.get("active") != "True":
            continue
        option_ids = [option_id for option_id in group.get("option_ids", []) if option_id]
        for source_id in option_ids:
            for target_id in option_ids:
                if source_id != target_id:
                    pairs.add((source_id, target_id))
    return pairs


def truncate_reason(text: str, limit: int = 180) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "..."


def draft_label_for(entity_id: str, option_rows: dict[str, dict[str, Any]], interiors_by_id: dict[str, dict[str, Any]]) -> str:
    if entity_id in option_rows:
        option = option_rows[entity_id]
        return f"{option.get('rpo')} {option.get('label')}".strip()
    if entity_id in interiors_by_id:
        interior = interiors_by_id[entity_id]
        return f"{interior.get('interior_id')} {interior.get('interior_name')}".strip()
    return entity_id


def build_draft_rules(
    wb,
    config: ModelConfig,
    option_rows: dict[str, dict[str, Any]],
    sections_by_id: dict[str, dict[str, Any]],
    interiors: list[dict[str, Any]],
    grouped_requires: set[tuple[str, str]],
    grouped_excludes: set[tuple[str, str]],
) -> list[dict[str, Any]]:
    interiors_by_id = {row["interior_id"]: row for row in interiors if row.get("interior_id")}
    valid_ids = set(option_rows) | set(interiors_by_id)
    raw_rules: list[dict[str, Any]] = []
    for rule in rows_from_optional_sheet(wb, config.rule_mapping_sheet):
        rule_type = rule.get("rule_type", "").lower()
        source_id = rule.get("source_id", "")
        target_id = rule.get("target_id", "")
        if not rule_type or source_id not in valid_ids or target_id not in valid_ids:
            continue
        if rule.get("generation_action", "") == "omit_grouped_requirement":
            continue
        if rule_type == "requires" and (source_id, target_id) in grouped_requires:
            continue
        if rule_type == "excludes" and (source_id, target_id) in grouped_excludes:
            continue
        source_section = rule.get("source_section", "")
        target_section = rule.get("target_section", "")
        source_mode = sections_by_id.get(source_section, {}).get("selection_mode") or rule.get("source_selection_mode", "")
        target_mode = sections_by_id.get(target_section, {}).get("selection_mode") or rule.get("target_selection_mode", "")
        replaces_default = rule.get("runtime_action", "") == "replace"
        redundant = (
            rule_type == "excludes"
            and source_section
            and source_section == target_section
            and source_mode.startswith("single")
            and target_mode.startswith("single")
            and not replaces_default
        )
        source_label = draft_label_for(source_id, option_rows, interiors_by_id)
        target_label = draft_label_for(target_id, option_rows, interiors_by_id)
        disabled_reason = rule.get("disabled_reason", "")
        auto_add = "False"
        if not disabled_reason and replaces_default:
            disabled_reason = f"{source_label} removes this default."
        elif not disabled_reason and rule_type == "excludes":
            disabled_reason = f"Blocked by {source_label}."
        elif not disabled_reason and rule_type == "requires":
            disabled_reason = f"Requires {target_label}."
        elif not disabled_reason and rule_type == "includes":
            disabled_reason = f"Included with {source_label}."
            auto_add = "True"
        elif rule_type == "includes":
            auto_add = "True"
        raw_rules.append(
            {
                "rule_id": rule.get("rule_id", ""),
                "source_id": source_id,
                "rule_type": rule_type,
                "target_id": target_id,
                "target_type": rule.get("target_type", ""),
                "source_type": rule.get("source_type", ""),
                "source_section": source_section,
                "target_section": target_section,
                "source_selection_mode": source_mode,
                "target_selection_mode": target_mode,
                "body_style_scope": rule.get("body_style_scope", ""),
                "disabled_reason": disabled_reason,
                "auto_add": auto_add,
                "active": "False" if redundant else "True",
                "runtime_action": "replace" if replaces_default else "omit_redundant_same_section_exclude" if redundant else "active",
                "source_note": truncate_reason(rule.get("original_detail_raw", ""), 500),
                "review_flag": rule.get("review_flag", ""),
            }
        )
    return raw_rules


def inspect_model_sources(config: ModelConfig) -> dict[str, Any]:
    wb = load_workbook(config.workbook_path, data_only=True, read_only=True)
    raw_rows = rows_from_sheet(wb, config.source_option_sheet)
    variants_raw = rows_from_sheet(wb, "variant_master")
    categories = {row["category_id"]: row for row in rows_from_sheet(wb, "category_master")}
    sections = {row["section_id"]: row for row in rows_from_sheet(wb, "section_master")}
    rows = [normalized_option_row(row, config) for row in raw_rows]
    status_lookup = status_lookup_from_sheet(wb, config)
    apply_status_lookup(rows, status_lookup, config)

    variant_rows = [row for row in variants_raw if row.get("variant_id", "") in config.variant_ids]
    active_variant_ids = {row["variant_id"] for row in variant_rows if row.get("active") == "True"}
    configured_variant_ids = set(config.variant_ids)

    status_counts: Counter[str] = Counter()
    status_counts_by_variant: dict[str, Counter[str]] = {variant_id: Counter() for variant_id in config.variant_ids}
    unknown_status_cells: list[dict[str, Any]] = []
    missing_status_cells: list[dict[str, Any]] = []
    matrix_rows: list[dict[str, Any]] = []
    candidate_choices: list[dict[str, Any]] = []
    candidate_standard_equipment: list[dict[str, Any]] = []

    for row in rows:
        row_active = row["active"] == "True"
        normalized_statuses = [status for status in row["statuses"].values() if status]
        effective_status = best_status(*normalized_statuses)
        matrix_rows.append(
            {
                "option_id": row["option_id"],
                "rpo": row["rpo"],
                "effective_status": effective_status,
                "effective_status_label": status_to_label(effective_status),
                "statuses": row["statuses"],
            }
        )
        for variant_id, status in row["statuses"].items():
            if not status:
                missing_status_cells.append({"option_id": row["option_id"], "rpo": row["rpo"], "variant_id": variant_id})
                continue
            status_counts[status] += 1
            status_counts_by_variant[variant_id][status] += 1
            if status not in ALLOWED_STATUSES:
                unknown_status_cells.append(
                    {
                        "option_id": row["option_id"],
                        "rpo": row["rpo"],
                        "variant_id": variant_id,
                        "status": status,
                    }
                )
            candidate_status, candidate_selectable, candidate_active = display_behavior_status(
                status,
                row["selectable"],
                row["active"],
                row["display_behavior"],
            )
            if row_active and candidate_status in {"available", "standard", "unavailable"}:
                candidate_choices.append(
                    {
                        "choice_id": f"{variant_id}__{row['option_id']}",
                        "option_id": row["option_id"],
                        "rpo": row["rpo"],
                        "variant_id": variant_id,
                        "status": candidate_status,
                        "selectable": candidate_selectable,
                        "active": candidate_active,
                        "section_id": row["section_id"],
                    }
                )
            if row_active and candidate_status == "standard":
                candidate_standard_equipment.append(
                    {
                        "option_id": row["option_id"],
                        "rpo": row["rpo"],
                        "variant_id": variant_id,
                        "section_id": row["section_id"],
                        "option_name": row["option_name"],
                    }
                )

    selectable_counts = Counter(row["selectable"] or "<blank>" for row in rows)
    active_rows = [
        row
        for row in rows
        if row["active"] == "True"
        and any(display_behavior_status(status, row["selectable"], row["active"], row["display_behavior"])[0] in {"available", "standard"} for status in row["statuses"].values())
    ]
    standard_rows = [
        row
        for row in rows
        if row["active"] == "True"
        and any(display_behavior_status(status, row["selectable"], row["active"], row["display_behavior"])[0] == "standard" for status in row["statuses"].values())
    ]
    unique_rpos = sorted({row["rpo"] for row in rows if row["rpo"]})

    section_mapping_rows: list[dict[str, Any]] = []
    missing_sections: Counter[str] = Counter()
    unknown_sections: Counter[str] = Counter()
    unknown_categories: Counter[str] = Counter()
    section_category_mismatches: list[dict[str, Any]] = []
    for row in rows:
        if row["active"] != "True":
            continue
        section_id = row["section_id"]
        category_id = row["category_id"]
        if not section_id:
            missing_sections[row["option_id"]] += 1
        elif section_id not in sections:
            unknown_sections[section_id] += 1
        if category_id and category_id not in categories:
            unknown_categories[category_id] += 1
        section = sections.get(section_id, {})
        if section and category_id and section.get("category_id") != category_id:
            section_category_mismatches.append(
                {
                    "option_id": row["option_id"],
                    "rpo": row["rpo"],
                    "section_id": section_id,
                    "row_category_id": category_id,
                    "section_category_id": section.get("category_id", ""),
                }
            )
        step_key = step_for_section(
            section_id,
            section.get("section_name", ""),
            section.get("category_id", category_id),
            standard_sections=config.standard_sections,
            section_step_overrides=config.section_step_overrides,
        )
        section_mapping_rows.append(
            {
                "option_id": row["option_id"],
                "rpo": row["rpo"],
                "category_id": category_id,
                "section_id": section_id,
                "original_section_id": row["original_section_id"],
                "section_override_applied": row["section_override_applied"],
                "step_key": step_key,
            }
        )

    blank_section_overrides = []
    for option_id, configured_section_id in config.blank_section_overrides.items():
        row = next((candidate for candidate in rows if candidate["option_id"] == option_id), None)
        blank_section_overrides.append(
            {
                "option_id": option_id,
                "rpo": row["rpo"] if row else "",
                "source_section_blank": bool(row and not row["original_section_id"]),
                "configured_section_id": configured_section_id,
                "resolved_section_id": row["section_id"] if row else "",
                "handled_by_explicit_config": bool(row and row["section_override_applied"] and row["section_id"] == configured_section_id),
            }
        )

    rule_hot_spots = classify_rule_hot_spots(rows, config)

    warnings = []
    if len(variant_rows) != config.expected_variant_count:
        warnings.append(f"Expected {config.expected_variant_count} configured variants, found {len(variant_rows)} in variant_master.")
    if configured_variant_ids - active_variant_ids:
        warnings.append(
            "Configured Grand Sport variants are present but inactive in variant_master, preserving the live Stingray-only generator path: "
            f"{', '.join(sorted(configured_variant_ids - active_variant_ids))}."
        )
    if missing_status_cells:
        warnings.append(f"Missing status cells: {len(missing_status_cells)}.")
    if unknown_status_cells:
        warnings.append(f"Unknown status cells: {len(unknown_status_cells)}.")
    if missing_sections:
        warnings.append(f"Rows still missing resolved sections: {sum(missing_sections.values())}.")
    if unknown_sections:
        warnings.append(f"Unknown resolved section ids: {', '.join(sorted(unknown_sections))}.")
    if unknown_categories:
        warnings.append(f"Unknown category ids: {', '.join(sorted(unknown_categories))}.")
    if section_category_mismatches:
        warnings.append(f"Section/category mismatches: {len(section_category_mismatches)}.")
    unresolved_blank_overrides = [row["option_id"] for row in blank_section_overrides if not row["handled_by_explicit_config"]]
    if unresolved_blank_overrides:
        warnings.append(f"Configured blank-section overrides not resolved: {', '.join(unresolved_blank_overrides)}.")

    return {
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "status": "inspection_generated",
        "config": {
            "model_key": config.model_key,
            "model_label": config.model_label,
            "model_year": config.model_year,
            "source_option_sheet": config.source_option_sheet,
            "status_sheet": config.status_sheet,
            "variant_ids": list(config.variant_ids),
            "expected_variant_count": config.expected_variant_count,
            "blank_section_overrides": dict(config.blank_section_overrides),
            "notes": list(config.notes),
        },
        "variants": {
            "configured_count": len(config.variant_ids),
            "expected_count": config.expected_variant_count,
            "found_in_variant_master": variant_rows,
            "missing_from_variant_master": sorted(configured_variant_ids - {row["variant_id"] for row in variant_rows}),
            "inactive_configured": sorted(configured_variant_ids - active_variant_ids),
        },
        "counts": {
            "option_rows": len(rows),
            "unique_rpos": len(unique_rpos),
            "variant_status_cells": len(rows) * len(config.variant_ids),
            "status_source_rows": len(status_lookup),
            "candidate_choice_rows_available_or_standard": len(candidate_choices),
            "candidate_standard_equipment_cells": len(candidate_standard_equipment),
            "candidate_standard_option_rows": len(standard_rows),
            "selectable_counts": dict(sorted(selectable_counts.items())),
            "active_option_rows_available_or_standard": len(active_rows),
            "inactive_option_rows_all_unavailable_or_blank": len(rows) - len(active_rows),
            "status_counts": dict(sorted(status_counts.items())),
            "status_counts_by_variant": {
                variant_id: dict(sorted(counts.items())) for variant_id, counts in status_counts_by_variant.items()
            },
            "missing_status_cells": len(missing_status_cells),
            "unknown_status_cells": len(unknown_status_cells),
        },
        "unique_rpos": unique_rpos,
        "status_matrix": matrix_rows,
        "section_mappings": {
            "rows": section_mapping_rows,
            "missing_sections": dict(sorted(missing_sections.items())),
            "unknown_sections": dict(sorted(unknown_sections.items())),
            "unknown_categories": dict(sorted(unknown_categories.items())),
            "section_category_mismatches": section_category_mismatches,
        },
        "blank_section_overrides": blank_section_overrides,
        "candidate_standard_equipment": candidate_standard_equipment,
        "candidate_choices_sample": candidate_choices[:100],
        "rule_detail_hot_spots": {
            "counts": rule_hot_spots["counts"],
            "rows": rule_hot_spots["rows"],
            "special_mentions": rule_hot_spots["special_mentions"],
        },
        "missing_status_cells": missing_status_cells,
        "unknown_status_cells": unknown_status_cells,
        "warnings": warnings,
    }


def build_contract_preview(config: ModelConfig) -> dict[str, Any]:
    wb = load_workbook(config.workbook_path, data_only=True, read_only=True)
    raw_rows = rows_from_sheet(wb, config.source_option_sheet)
    variants_raw = rows_from_sheet(wb, "variant_master")
    categories = {row["category_id"]: row for row in rows_from_sheet(wb, "category_master")}
    sections = {row["section_id"]: row for row in rows_from_sheet(wb, "section_master")}
    rows = [normalized_option_row(row, config) for row in raw_rows]
    apply_status_lookup(rows, status_lookup_from_sheet(wb, config), config)

    variant_source_rows = {row["variant_id"]: row for row in variants_raw if row.get("variant_id", "") in config.variant_ids}
    variants: list[dict[str, Any]] = []
    for variant_id in config.variant_ids:
        source = variant_source_rows.get(variant_id, {})
        variants.append(
            {
                "variant_id": variant_id,
                "model_year": intish(source.get("model_year"), intish(config.model_year)),
                "model": config.model_label,
                "trim_level": clean(source.get("trim_level", variant_id.split("_", 1)[0])).upper(),
                "body_style": clean(source.get("body_style", "convertible" if variant_id.endswith("e67") else "coupe")).lower(),
                "display_name": clean(source.get("display_name", variant_id)),
                "base_price": money(source.get("base_price")),
                "display_order": intish(source.get("display_order")),
                "source_active": clean(source.get("active", "")),
                "preview_included": True,
            }
        )

    body_context_choices = []
    body_styles = sorted(
        {row["body_style"] for row in variants},
        key=lambda body_style: config.body_style_display_order.get(body_style, 99),
    )
    for body_style in body_styles:
        body_variants = [row for row in variants if row["body_style"] == body_style]
        body_context_choices.append(
            {
                "context_choice_id": f"body_style__{body_style}",
                "context_type": "body_style",
                "value": body_style,
                "label": body_style.title(),
                "description": f"{len(body_variants)} trims available",
                "section_id": "sec_context_body_style",
                "step_key": "body_style",
                "body_style": body_style,
                "trim_level": "",
                "variant_id": "",
                "base_price": "",
                "display_order": config.body_style_display_order.get(body_style, 99),
            }
        )
    trim_context_choices = [
        {
            "context_choice_id": f"trim_level__{variant['body_style']}__{variant['trim_level'].lower()}",
            "context_type": "trim_level",
            "value": variant["trim_level"],
            "label": variant["trim_level"],
            "description": variant["display_name"],
            "section_id": "sec_context_trim_level",
            "step_key": "trim_level",
            "body_style": variant["body_style"],
            "trim_level": variant["trim_level"],
            "variant_id": variant["variant_id"],
            "base_price": variant["base_price"],
            "display_order": variant["display_order"],
        }
        for variant in variants
    ]
    context_choices = body_context_choices + trim_context_choices
    variants_by_id = {row["variant_id"]: row for row in variants}

    choices: list[dict[str, Any]] = []
    candidate_standard_equipment: list[dict[str, Any]] = []
    section_category_resolutions: list[dict[str, Any]] = []
    unresolved_issues: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []
    text_cleanup_counter: Counter[str] = Counter()
    section_source_categories: defaultdict[str, set[str]] = defaultdict(set)
    section_ids_with_choices: set[str] = set()

    for row in rows:
        if row["active"] != "True":
            continue
        section_id = row["section_id"]
        source_category_id = row["category_id"]
        section_category_id = sections.get(section_id, {}).get("category_id", "")
        resolved_category_id, category_resolution_source = resolve_category(row, sections, config)
        step_key = resolved_step_key(section_id, sections, resolved_category_id, config)
        section_source_categories[section_id].add(source_category_id or "")
        if not section_id:
            issue = {
                "issue_type": "unresolved_section",
                "option_id": row["option_id"],
                "rpo": row["rpo"],
                "message": "Grand Sport row has no resolved section.",
            }
            unresolved_issues.append(issue)
            validation_rows.append({**issue, "severity": "error"})
            continue
        if section_id not in sections:
            issue = {
                "issue_type": "unknown_section",
                "option_id": row["option_id"],
                "rpo": row["rpo"],
                "section_id": section_id,
                "message": "Grand Sport row resolves to a section id missing from section_master.",
            }
            unresolved_issues.append(issue)
            validation_rows.append({**issue, "severity": "error"})
            continue
        if not resolved_category_id or resolved_category_id not in categories:
            issue = {
                "issue_type": "unknown_category",
                "option_id": row["option_id"],
                "rpo": row["rpo"],
                "resolved_category_id": resolved_category_id,
                "message": "Grand Sport row resolves to a category id missing from category_master.",
            }
            unresolved_issues.append(issue)
            validation_rows.append({**issue, "severity": "error"})
            continue
        if source_category_id and section_category_id and source_category_id != section_category_id:
            resolution = {
                "option_id": row["option_id"],
                "rpo": row["rpo"],
                "source_category_id": source_category_id,
                "section_category_id": section_category_id,
                "resolved_category_id": resolved_category_id,
                "resolved_section_id": section_id,
                "category_resolution_source": category_resolution_source,
            }
            section_category_resolutions.append(resolution)
            if category_resolution_source not in {"section_override", "option_override"}:
                issue = {
                    "issue_type": "unresolved_section_category_mismatch",
                    "option_id": row["option_id"],
                    "rpo": row["rpo"],
                    "source_category_id": source_category_id,
                    "section_category_id": section_category_id,
                    "message": "Section/category mismatch does not have an explicit config resolution.",
                }
                unresolved_issues.append(issue)
                validation_rows.append({**issue, "severity": "warning"})

        label, label_notes = cleanup_display_text(row["option_name"], config)
        description, description_notes = cleanup_display_text(row["description"], config)
        text_cleanup_notes = [f"label:{note}" for note in label_notes] + [f"description:{note}" for note in description_notes]
        for note in text_cleanup_notes:
            text_cleanup_counter[note] += 1

        category = categories.get(resolved_category_id, {})
        section = sections.get(section_id, {})
        section_name = config.section_label_overrides.get(section_id, section.get("section_name", ""))
        option_base = {
            "option_id": row["option_id"],
            "rpo": row["rpo"],
            "label": label,
            "description": description,
            "source_option_name": row["option_name"],
            "source_description": row["description"],
            "source_detail_raw": row["detail_raw"],
            "source_category_id": source_category_id,
            "source_section_id": row["original_section_id"],
            "section_id": section_id,
            "resolved_section_id": section_id,
            "section_category_id": section_category_id,
            "resolved_category_id": resolved_category_id,
            "category_resolution_source": category_resolution_source,
            "section_name": section_name,
            "category_name": category.get("category_name", ""),
            "step_key": step_key,
            "selectable": row["selectable"],
            "active": row["active"],
            "display_behavior": row["display_behavior"],
            "base_price": row["price"],
            "text_cleanup_notes": text_cleanup_notes,
        }
        section_ids_with_choices.add(section_id)

        for variant_id, status in row["statuses"].items():
            status, selectable, active = display_behavior_status(
                status,
                row["selectable"],
                row["active"],
                row["display_behavior"],
            )
            if status not in {"available", "standard"} and row["display_behavior"] != "auto_only":
                continue
            variant = variants_by_id[variant_id]
            choice = {
                **option_base,
                "choice_id": f"{variant_id}__{row['option_id']}",
                "variant_id": variant_id,
                "body_style": variant["body_style"],
                "trim_level": variant["trim_level"],
                "status": status,
                "status_label": status_to_label(status),
                "selectable": selectable,
                "active": active,
            }
            choices.append(choice)
            if status == "standard":
                candidate_standard_equipment.append(choice)

    step_order_index = {step_key: index for index, step_key in enumerate(config.step_order)}

    def section_sort_key(section_id: str) -> tuple[int, int, str]:
        section = sections.get(section_id, {})
        category_id = section.get("category_id", "")
        step_key = resolved_step_key(section_id, sections, category_id, config)
        return (
            intish(section.get("display_order"), 9999),
            step_order_index.get(step_key, 9999),
            section_id,
        )

    section_rows: list[dict[str, Any]] = [dict(section) for section in config.context_sections]
    for section_id in sorted(section_ids_with_choices, key=section_sort_key):
        section = sections.get(section_id, {})
        resolved_categories = sorted(
            {
                choice["resolved_category_id"]
                for choice in choices
                if choice["resolved_section_id"] == section_id
            }
        )
        resolved_category_id = resolved_categories[0] if resolved_categories else section.get("category_id", "")
        category = categories.get(resolved_category_id, {})
        step_key = resolved_step_key(section_id, sections, resolved_category_id, config)
        selection_mode = section.get("selection_mode", "")
        section_rows.append(
            {
                "section_id": section_id,
                "section_name": config.section_label_overrides.get(section_id, section.get("section_name", "")),
                "source_section_name": section.get("section_name", ""),
                "category_id": resolved_category_id,
                "category_name": category.get("category_name", ""),
                "source_category_ids": sorted(value for value in section_source_categories.get(section_id, set()) if value),
                "section_category_id": section.get("category_id", ""),
                "selection_mode": selection_mode,
                "selection_mode_label": selection_mode_label(selection_mode, config.selection_mode_labels),
                "choice_mode": normalize_mode(selection_mode),
                "is_required": section.get("is_required", ""),
                "standard_behavior": section.get("standard_behavior", ""),
                "section_display_order": intish(section.get("display_order")),
                "step_key": step_key,
                "step_label": config.step_labels.get(step_key, step_key.replace("_", " ").title()),
            }
        )

    section_ids_by_step: dict[str, list[str]] = defaultdict(list)
    for row in section_rows:
        section_ids_by_step[row["step_key"]].append(row["section_id"])
    step_rows = [
        {
            "step_key": step_key,
            "step_label": config.step_labels[step_key],
            "runtime_order": index + 1,
            "source": "runtime",
            "section_ids": "|".join(sorted(section_ids_by_step.get(step_key, []))),
        }
        for index, step_key in enumerate(config.step_order)
    ]

    blank_overrides = []
    for option_id, configured_section_id in config.blank_section_overrides.items():
        row = next((candidate for candidate in rows if candidate["option_id"] == option_id), None)
        blank_overrides.append(
            {
                "option_id": option_id,
                "rpo": row["rpo"] if row else "",
                "source_section_blank": bool(row and not row["original_section_id"]),
                "configured_section_id": configured_section_id,
                "resolved_section_id": row["section_id"] if row else "",
                "handled_by_explicit_config": bool(row and row["section_override_applied"] and row["section_id"] == configured_section_id),
            }
        )

    rule_hot_spots = classify_rule_hot_spots(rows, config, include_special_bucket=True)
    text_cleanup_summary = {
        "changed_fields": sum(text_cleanup_counter.values()),
        "notes": dict(sorted(text_cleanup_counter.items())),
    }

    return {
        "dataset": {
            "name": "2027 Corvette Grand Sport contract preview",
            "model": config.model_label,
            "model_year": config.model_year,
            "source_workbook": config.workbook_path.name,
            "source_sheet": config.source_option_sheet,
            "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "status": "read_only_preview",
        },
        "variants": variants,
        "steps": step_rows,
        "sections": section_rows,
        "contextChoices": context_choices,
        "choices": choices,
        "candidateStandardEquipment": candidate_standard_equipment,
        "ruleDetailHotSpots": rule_hot_spots,
        "normalization": {
            "blankSectionOverrides": blank_overrides,
            "sectionCategoryResolutions": section_category_resolutions,
            "textCleanupSummary": text_cleanup_summary,
            "unresolvedIssues": unresolved_issues,
        },
        "validation": validation_rows,
    }


def write_contract_preview_artifacts(preview: dict[str, Any], output_dir: Path, artifact_prefix: str) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / f"{artifact_prefix}.json"
    md_path = output_dir / f"{artifact_prefix}.md"
    json_path.write_text(json.dumps(preview, indent=2), encoding="utf-8")
    md_path.write_text(render_contract_preview_markdown(preview), encoding="utf-8")
    return {"json": str(json_path), "markdown": str(md_path)}


def build_form_data_draft(config: ModelConfig) -> dict[str, Any]:
    preview = build_contract_preview(config)
    variants_by_id = {row["variant_id"]: row for row in preview["variants"]}
    sections_by_id = {row["section_id"]: row for row in preview["sections"]}
    interiors = build_grand_sport_interiors(config)
    option_rows: dict[str, dict[str, Any]] = {}
    statuses_by_option: defaultdict[str, dict[str, str]] = defaultdict(dict)
    order_by_option: dict[str, int] = {}

    for index, choice in enumerate(preview["choices"], start=1):
        option_id = choice["option_id"]
        statuses_by_option[option_id][choice["variant_id"]] = choice["status"]
        if option_id not in order_by_option:
            order_by_option[option_id] = index
        if option_id not in option_rows:
            option_rows[option_id] = {
                key: choice.get(key, "")
                for key in (
                    "option_id",
                    "rpo",
                    "label",
                    "description",
                    "source_detail_raw",
                    "source_option_name",
                    "source_description",
                    "section_id",
                    "section_name",
                    "resolved_section_id",
                    "resolved_category_id",
                    "category_name",
                    "step_key",
                    "selectable",
                    "active",
                    "display_behavior",
                    "base_price",
                    "text_cleanup_notes",
                )
            }

    draft_choices: list[dict[str, Any]] = []
    for option_id, option in sorted(option_rows.items(), key=lambda item: order_by_option[item[0]]):
        section = sections_by_id.get(option["section_id"], {})
        for variant_id in config.variant_ids:
            variant = variants_by_id[variant_id]
            status = statuses_by_option[option_id].get(variant_id, "unavailable")
            status, selectable, active = display_behavior_status(
                status,
                option["selectable"],
                option.get("active", "True"),
                option.get("display_behavior", ""),
            )
            draft_choices.append(
                {
                    "choice_id": f"{variant_id}__{option_id}",
                    "option_id": option_id,
                    "rpo": option["rpo"],
                    "label": option["label"],
                    "description": option["description"],
                    "section_id": option["section_id"],
                    "section_name": option["section_name"],
                    "category_id": option["resolved_category_id"],
                    "category_name": option["category_name"],
                    "step_key": option["step_key"],
                    "variant_id": variant_id,
                    "body_style": variant["body_style"],
                    "trim_level": variant["trim_level"],
                    "status": status,
                    "status_label": status_to_label(status),
                    "selectable": selectable,
                    "active": active,
                    "choice_mode": section.get("choice_mode", ""),
                    "selection_mode": section.get("selection_mode", ""),
                    "selection_mode_label": section.get("selection_mode_label", ""),
                    "base_price": option["base_price"],
                    "display_order": order_by_option[option_id],
                    "source_detail_raw": option["source_detail_raw"],
                    "source_option_name": option["source_option_name"],
                    "source_description": option["source_description"],
                    "text_cleanup_notes": option["text_cleanup_notes"],
                }
            )

    standard_equipment = [
        {
            "equipment_id": f"std_{choice['choice_id']}",
            "variant_id": choice["variant_id"],
            "body_style": choice["body_style"],
            "trim_level": choice["trim_level"],
            "option_id": choice["option_id"],
            "rpo": choice["rpo"],
            "label": choice["label"],
            "description": choice["description"],
            "section_id": choice["section_id"],
            "section_name": choice["section_name"],
            "category_name": choice["category_name"],
            "display_order": choice["display_order"],
            "source_detail_raw": choice["source_detail_raw"],
        }
        for choice in draft_choices
        if choice["status"] == "standard"
    ]

    wb = load_workbook(config.workbook_path, data_only=True, read_only=True)
    rule_groups = load_rule_groups(wb, config)
    exclusive_groups = load_exclusive_groups(wb, config)
    color_overrides = build_color_overrides(wb, config, interiors, option_rows)
    rules = build_draft_rules(
        wb,
        config,
        option_rows,
        sections_by_id,
        interiors,
        grouped_requirement_pairs(rule_groups),
        exclusive_group_pairs(exclusive_groups),
    )

    validation = [
        {
            "check_id": "grand_sport_draft_status",
            "severity": "warning",
            "entity_type": "dataset",
            "entity_id": "",
            "message": "Grand Sport form data is a draft inspection artifact and is not runtime active.",
        },
        {
            "check_id": "active_variants",
            "severity": "pass",
            "entity_type": "variant",
            "entity_id": "",
            "message": f"{len(preview['variants'])} configured Grand Sport variants included by model config; workbook active flags are unchanged.",
        },
        {
            "check_id": "availability_rows",
            "severity": "pass",
            "entity_type": "availability",
            "entity_id": "",
            "message": f"{len(draft_choices)} draft choice rows exported from the Grand Sport variant matrix.",
        },
        {
            "check_id": "rules",
            "severity": "pass",
            "entity_type": "rule",
            "entity_id": "",
            "message": f"{len(rules)} active compatibility rules exported from {config.rule_mapping_sheet}.",
        },
        {
            "check_id": "interior_contract",
            "severity": "pass",
            "entity_type": "interior",
            "entity_id": "",
            "message": f"{len(interiors)} model-scoped Grand Sport LT interiors exported.",
        },
        {
            "check_id": "pricing_deferred",
            "severity": "warning",
            "entity_type": "price_rule",
            "entity_id": "",
            "message": "Final Grand Sport price rules are deferred unless directly represented in normalized option prices.",
        },
    ]
    if color_overrides:
        validation.append(
            {
                "check_id": "color_overrides",
                "severity": "pass",
                "entity_type": "color_override",
                "entity_id": "",
                "message": f"{len(color_overrides)} color override rows exported from {config.color_overrides_sheet}.",
            }
        )

    return {
        "dataset": {
            "name": "2027 Corvette Grand Sport form data draft",
            "model": config.model_label,
            "model_year": config.model_year,
            "source_workbook": config.workbook_path.name,
            "source_sheet": config.source_option_sheet,
            "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "status": "draft_not_runtime_active",
        },
        "variants": preview["variants"],
        "steps": preview["steps"],
        "sections": preview["sections"],
        "contextChoices": preview["contextChoices"],
        "choices": draft_choices,
        "standardEquipment": standard_equipment,
        "ruleGroups": rule_groups,
        "exclusiveGroups": exclusive_groups,
        "rules": rules,
        "priceRules": [],
        "interiors": interiors,
        "colorOverrides": color_overrides,
        "validation": validation,
        "draftMetadata": {
            "sourcePreviewStatus": preview["dataset"]["status"],
            "candidateAvailableOrStandardChoices": len(preview["choices"]),
            "fullVariantMatrixChoices": len(draft_choices),
            "ruleDetailHotSpots": preview["ruleDetailHotSpots"],
            "normalization": preview["normalization"],
            "deferredSurfaces": ["priceRules"],
        },
    }


def write_form_data_draft_artifacts(draft: dict[str, Any], output_dir: Path, artifact_prefix: str) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / f"{artifact_prefix}.json"
    md_path = output_dir / f"{artifact_prefix}.md"
    json_path.write_text(json.dumps(draft, indent=2), encoding="utf-8")
    md_path.write_text(render_form_data_draft_markdown(draft), encoding="utf-8")
    return {"json": str(json_path), "markdown": str(md_path)}


def write_inspection_artifacts(report: dict[str, Any], output_dir: Path) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "grand-sport-inspection.json"
    md_path = output_dir / "grand-sport-inspection.md"
    json_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    md_path.write_text(render_markdown_report(report), encoding="utf-8")
    return {"json": str(json_path), "markdown": str(md_path)}


def render_markdown_report(report: dict[str, Any]) -> str:
    counts = report["counts"]
    warnings = report["warnings"]
    section_mappings = report["section_mappings"]
    hot_spots = report["rule_detail_hot_spots"]
    lines = [
        "# Grand Sport Inspection",
        "",
        f"Generated: `{report['generated_at']}`",
        f"Source sheet: `{report['config']['source_option_sheet']}`",
        f"Status: `{report['status']}`",
        "",
        "## Variant Validation",
        "",
        f"- Configured variants: {', '.join(report['config']['variant_ids'])}",
        f"- Found in `variant_master`: {len(report['variants']['found_in_variant_master'])}",
        f"- Missing from `variant_master`: {', '.join(report['variants']['missing_from_variant_master']) or 'none'}",
        f"- Inactive configured variants: {', '.join(report['variants']['inactive_configured']) or 'none'}",
        "",
        "## Counts",
        "",
        f"- Option rows: {counts['option_rows']}",
        f"- Unique RPOs: {counts['unique_rpos']}",
        f"- Variant status cells: {counts['variant_status_cells']}",
        f"- Candidate choice rows with available/standard status: {counts['candidate_choice_rows_available_or_standard']}",
        f"- Candidate standard equipment cells: {counts['candidate_standard_equipment_cells']}",
        f"- Candidate standard option rows: {counts['candidate_standard_option_rows']}",
        f"- Active option rows with available/standard status: {counts['active_option_rows_available_or_standard']}",
        f"- Inactive option rows all unavailable/blank: {counts['inactive_option_rows_all_unavailable_or_blank']}",
        f"- Selectable counts: `{json.dumps(counts['selectable_counts'], sort_keys=True)}`",
        f"- Status counts: `{json.dumps(counts['status_counts'], sort_keys=True)}`",
        f"- Missing status cells: {counts['missing_status_cells']}",
        f"- Unknown status cells: {counts['unknown_status_cells']}",
        "",
        "## Section Mapping",
        "",
        f"- Rows still missing resolved sections: {sum(section_mappings['missing_sections'].values())}",
        f"- Unknown section ids: {', '.join(section_mappings['unknown_sections']) or 'none'}",
        f"- Unknown category ids: {', '.join(section_mappings['unknown_categories']) or 'none'}",
        f"- Section/category mismatches: {len(section_mappings['section_category_mismatches'])}",
        "",
        "## Blank-Section Overrides",
        "",
    ]
    for row in report["blank_section_overrides"]:
        handled = "yes" if row["handled_by_explicit_config"] else "no"
        lines.append(
            f"- `{row['rpo']}` / `{row['option_id']}`: blank source section -> `{row['resolved_section_id']}` "
            f"(configured `{row['configured_section_id']}`, handled by config: {handled})"
        )
    lines.extend(
        [
            "",
            "## Rule/Detail Hot Spots",
            "",
            f"- Hot spot counts: `{json.dumps(hot_spots['counts'], sort_keys=True)}`",
            f"- Rows requiring later rule review: {len(hot_spots['rows'])}",
            "",
            "| RPO | Option | Section | Matched Terms | Special Mentions |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for row in hot_spots["rows"][:40]:
        lines.append(
            "| "
            + " | ".join(
                [
                    f"`{row['rpo']}`",
                    row["option_name"].replace("|", "\\|"),
                    f"`{row['section_id']}`",
                    ", ".join(row["matched_terms"]) or "",
                    ", ".join(row["special_mentions"]) or "",
                ]
            )
            + " |"
        )
    if len(hot_spots["rows"]) > 40:
        lines.append(f"| ... | {len(hot_spots['rows']) - 40} additional rows in JSON artifact |  |  |  |")
    lines.extend(["", "## Warnings", ""])
    if warnings:
        lines.extend(f"- {warning}" for warning in warnings)
    else:
        lines.append("- none")
    lines.append("")
    return "\n".join(lines)


def render_contract_preview_markdown(preview: dict[str, Any]) -> str:
    normalization = preview["normalization"]
    hot_spots = preview["ruleDetailHotSpots"]
    unresolved = normalization["unresolvedIssues"]
    text_cleanup = normalization["textCleanupSummary"]
    lines = [
        "# Grand Sport Contract Preview",
        "",
        f"Generated: `{preview['dataset']['generated_at']}`",
        f"Status: `{preview['dataset']['status']}`",
        f"Source sheet: `{preview['dataset']['source_sheet']}`",
        "",
        "## Summary",
        "",
        f"- Variants: {len(preview['variants'])}",
        f"- Context choices: {len(preview['contextChoices'])}",
        f"- Steps: {len(preview['steps'])}",
        f"- Sections: {len(preview['sections'])}",
        f"- Choices: {len(preview['choices'])}",
        f"- Candidate standard equipment cells: {len(preview['candidateStandardEquipment'])}",
        f"- Rule/detail hot spot rows: {len(hot_spots['rows'])}",
        f"- Unresolved normalization issues: {len(unresolved)}",
        "",
        "## Variants",
        "",
        "| Variant | Display Name | Source Active | Preview Included |",
        "| --- | --- | --- | --- |",
    ]
    for variant in preview["variants"]:
        lines.append(
            f"| `{variant['variant_id']}` | {variant['display_name']} | `{variant['source_active']}` | `{variant['preview_included']}` |"
        )
    lines.extend(["", "## Blank-Section Overrides", ""])
    for row in normalization["blankSectionOverrides"]:
        handled = "yes" if row["handled_by_explicit_config"] else "no"
        lines.append(
            f"- `{row['rpo']}` / `{row['option_id']}`: `{row['configured_section_id']}` "
            f"(handled by explicit config: {handled})"
        )
    lines.extend(
        [
            "",
            "## Section/Category Resolution",
            "",
            f"- Resolved mismatch rows: {len(normalization['sectionCategoryResolutions'])}",
            f"- Unresolved issues: {len(unresolved)}",
            "",
            "## Text Cleanup",
            "",
            f"- Changed display fields: {text_cleanup['changed_fields']}",
            f"- Notes: `{json.dumps(text_cleanup['notes'], sort_keys=True)}`",
            "",
            "## Rule/Detail Hot Spots",
            "",
            f"- Counts: `{json.dumps(hot_spots['counts'], sort_keys=True)}`",
            f"- Rows: {len(hot_spots['rows'])}",
            "",
            "## Unresolved Normalization Issues",
            "",
        ]
    )
    if unresolved:
        for issue in unresolved:
            lines.append(f"- `{issue['issue_type']}`: {issue.get('option_id', '')} {issue.get('message', '')}")
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## Live Output Safety",
            "",
            "- Grand Sport preview generation writes only inspection artifacts under `form-output/inspection/`.",
            "- It does not write `form-app/data.js` or final Grand Sport app data.",
            "",
        ]
    )
    return "\n".join(lines)


def render_form_data_draft_markdown(draft: dict[str, Any]) -> str:
    metadata = draft["draftMetadata"]
    warning_rows = [row for row in draft["validation"] if row["severity"] == "warning"]
    lines = [
        "# Grand Sport Form Data Draft",
        "",
        f"Generated: `{draft['dataset']['generated_at']}`",
        f"Status: `{draft['dataset']['status']}`",
        f"Source sheet: `{draft['dataset']['source_sheet']}`",
        "",
        "## Contract Surface",
        "",
        f"- Variants: {len(draft['variants'])}",
        f"- Context choices: {len(draft['contextChoices'])}",
        f"- Steps: {len(draft['steps'])}",
        f"- Sections: {len(draft['sections'])}",
        f"- Choices: {len(draft['choices'])}",
        f"- Standard equipment rows: {len(draft['standardEquipment'])}",
        f"- Rule groups: {len(draft['ruleGroups'])} (workbook-backed)",
        f"- Exclusive groups: {len(draft['exclusiveGroups'])} (workbook-backed)",
        f"- Rules: {len(draft['rules'])} (workbook-backed)",
        f"- Price rules: {len(draft['priceRules'])} (deferred)",
        f"- Interiors: {len(draft['interiors'])} (model-scoped)",
        f"- Color overrides: {len(draft['colorOverrides'])}",
        "",
        "## Draft Notes",
        "",
        f"- Candidate available/standard choices from preview: {metadata['candidateAvailableOrStandardChoices']}",
        f"- Full variant-matrix draft choices: {metadata['fullVariantMatrixChoices']}",
        f"- Rule/detail hot spot rows preserved: {len(metadata['ruleDetailHotSpots']['rows'])}",
        f"- Unresolved normalization issues: {len(metadata['normalization']['unresolvedIssues'])}",
        "",
        "## Validation Warnings",
        "",
    ]
    if warning_rows:
        lines.extend(f"- `{row['check_id']}`: {row['message']}" for row in warning_rows)
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## Live Output Safety",
            "",
            "- This draft writes only inspection artifacts under `form-output/inspection/`.",
            "- It does not write `form-app/data.js` or activate Grand Sport in the app.",
            "",
        ]
    )
    return "\n".join(lines)
