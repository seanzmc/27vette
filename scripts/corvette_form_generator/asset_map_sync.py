"""Safe asset_map maintenance helpers and CLI implementation.

This module keeps WordPress media discovery separate from workbook-authored
runtime image metadata. Dry-run/report mode is the default; apply mode saves
through the project safe-save helper.
"""

from __future__ import annotations

import argparse
import base64
import csv
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field, replace
from datetime import datetime, timedelta, timezone
import hashlib
import json
import os
from pathlib import Path
import re
import subprocess
import sys
import time
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qsl, urlencode, urlparse, unquote, urlunparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from corvette_form_generator.contract import WILDCARD_MODEL_KEY
from corvette_form_generator.output import write_text_atomic
from corvette_form_generator.workbook_domain.registry import WRITABLE_COLUMNS
from corvette_form_generator.workbook import (
    clean,
    restore_workbook_backup,
    rows_from_sheet,
    save_workbook_safely,
    workbook_truthy,
)

SITE = "stingraychevroletcorvette.com"
MEDIA_ENDPOINT = f"https://{SITE}/wp-json/wp/v2/media"
PATH_FILTER = "/wp-content/uploads/pictures/27vette/"
WORDPRESS_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 Chrome/125 Safari/537.36 27vette-asset-map-sync/1.0"
)
ASSET_SHEET = "asset_map"
ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = ROOT / "stingray_master.xlsx"
DEFAULT_REPORT_DIR = ROOT / ".asset-map-sync"
TARGET_TYPE_OPTION = "option"
TARGET_TYPE_MODEL = "model"
TARGET_TYPE_CONTEXT_CHOICE = "context_choice"
SUPPORTED_TARGET_TYPES = {TARGET_TYPE_OPTION, TARGET_TYPE_MODEL, TARGET_TYPE_CONTEXT_CHOICE}
NEW_ROW_FIT = "cover"
NEW_ROW_POSITION = "center"
NEW_ROW_NOTE = "auto-seeded"
MISSING_IMAGE_ACTIONS = {"flag_missing", "flag_ambiguous", "flag_dead_no_match"}

COVERAGE_EXPECTED = "expected"
COVERAGE_NOT_EXPECTED = "not_expected"
COVERAGE_INTENTS = (COVERAGE_EXPECTED, COVERAGE_NOT_EXPECTED)
ACTIONABLE_COVERAGE_INTENTS = {COVERAGE_EXPECTED, ""}
COVERAGE_RULESET_VERSION = "phase4b-v1"
COVERAGE_RULESET = (
    "target_type:model|context_choice -> expected",
    "section selection_mode=display_only -> not_expected",
    "section_presentation standard_equipment_bucket -> not_expected",
    "all other active+selectable option targets -> expected (universal policy)",
)

IMGI_RE = re.compile(r"^imgi_\d+_(.+)$")
PREFIX_RE = re.compile(r"^([cehrsg])-(.+)$")
SHARED_PREFIX_RE = re.compile(r"^([cehrsg](?:-[cehrsg])+)-(.+)$")
MODEL_BODY_STYLE_RE = re.compile(r"^([cehrsg])(07|67)-([12])$")
SPLIT_RE = re.compile(r"[-_]")
RPO_RE = re.compile(r"^[0-9a-z]{3}$")

MODEL_PREFIX = {
    "c": "stingray",
    "e": "grand_sport",
    "h": "z06",
    "r": "zr1",
    "s": "zr1x",
    "g": "grand_sport_x",
}
# Ordered option-media inheritance after an exact model prefix and before a
# bare shared filename. Chains are flattened deliberately so resolution is
# deterministic and cannot cycle.
OPTION_MODEL_FALLBACKS = {
    "grand_sport": ("stingray",),
    "grand_sport_x": ("grand_sport", "stingray"),
    "zr1": ("z06",),
    "zr1x": ("z06",),
}
MODEL_TARGET_STEMS = {
    "stingray": ("stingray", "stingray"),
    "grand-sport": ("grand_sport", "grandSport"),
    "grand_sport": ("grand_sport", "grandSport"),
    "grandsport": ("grand_sport", "grandSport"),
    "z06": ("z06", "z06"),
    "zr1": ("zr1", "zr1"),
    "zr1x": ("zr1x", "zr1x"),
    "grand-sport-x": ("grand_sport_x", "grandSportX"),
    "grand_sport_x": ("grand_sport_x", "grandSportX"),
    "grandsportx": ("grand_sport_x", "grandSportX"),
}
BODY_STYLE_CODE = {"07": "coupe", "67": "convertible"}
BODY_STYLE_IMAGE_FIELD = {"1": "image_url", "2": "hover_image_url"}


@dataclass(frozen=True)
class SyncResult:
    report_path: Path
    missing_path: Path
    unmatched_path: Path
    manifest_path: Path
    url_write_count: int
    insert_count: int
    action_counts: dict[str, int]
    unmatched_count: int
    unparseable_count: int
    backup_path: Path | None = None


class WordPressMediaFetchError(RuntimeError):
    """Raised when live WordPress media fetch is blocked with actionable guidance."""


@dataclass(frozen=True)
class SyncPlan:
    report: list[dict[str, str]]
    url_writes: dict[tuple[int, str], str]
    inserts: list[dict[str, Any]]
    status: dict[int, str]
    used: set[str]
    section_coverage: dict[str, Any] = field(default_factory=dict)
    resolutions: dict[tuple[str, str, str], "CandidateResolution"] = field(default_factory=dict)
    desired_targets: dict[tuple[str, str, str], dict[str, str]] = field(default_factory=dict)
    current_rows: dict[tuple[str, str, str], dict[str, Any]] = field(default_factory=dict)

    def __iter__(self):
        """Preserve the older tuple-unpack test/helper contract."""

        yield self.report
        yield self.url_writes
        yield self.inserts
        yield self.status
        yield self.used


@dataclass(frozen=True)
class MediaInventory:
    option_exact: dict[tuple[str, str], list[str]]
    option_bare: dict[str, list[str]]
    model: dict[tuple[str, str], list[str]]
    bodystyle: dict[tuple[str, str, str], list[str]]
    unparseable: list[str]
    option_shared: dict[tuple[str, str], list[str]] = field(default_factory=dict)


@dataclass(frozen=True)
class MediaSnapshot:
    urls: list[str]
    modified_by_url: dict[str, str]


@dataclass(frozen=True)
class CandidateAlternative:
    """One equal-priority media candidate surfaced for human inspection."""

    field: str
    url: str
    source: str
    priority: int


@dataclass(frozen=True)
class CandidateResolution:
    """Typed output of the shared candidate resolver used by sync and Manager."""

    fields: dict[str, str]
    source: str
    reason: str
    priority: int | None
    alternatives: tuple[CandidateAlternative, ...] = ()


@dataclass(frozen=True)
class AssetManagerSnapshot:
    """Immutable shared reconciliation view consumed by API filtering and tests."""

    fingerprints: dict[str, str]
    media: dict[str, Any]
    coverage_ruleset: dict[str, Any]
    items: tuple[dict[str, Any], ...]
    action_counts: dict[str, int]
    media_urls: tuple[str, ...]


def filename_stem(url: str) -> str:
    base = unquote(os.path.basename(urlparse(url).path))
    return os.path.splitext(base)[0].lower()


def parse_media(url: str) -> tuple[str | None, str, bool]:
    """Return ``(model_key_or_none, rpo, is_valid)`` parsed from a media URL."""

    stem = filename_stem(url)
    match = IMGI_RE.match(stem)
    if match:
        stem = match.group(1)
    model = None
    match = PREFIX_RE.match(stem)
    if match:
        model = MODEL_PREFIX[match.group(1)]
        stem = match.group(2)
    rpo = SPLIT_RE.split(stem)[0]
    return model, rpo, bool(RPO_RE.match(rpo))


def parse_shared_option_media(url: str) -> tuple[str | None, str, bool]:
    """Return ``(prefix_group, rpo, is_valid)`` for multi-model option media."""

    stem = filename_stem(url)
    match = IMGI_RE.match(stem)
    if match:
        stem = match.group(1)
    match = SHARED_PREFIX_RE.match(stem)
    if not match:
        return None, "", False
    prefix_group = match.group(1)
    prefix_codes = prefix_group.split("-")
    rpo = SPLIT_RE.split(match.group(2))[0]
    valid = len(set(prefix_codes)) == len(prefix_codes) and bool(RPO_RE.match(rpo))
    return prefix_group, rpo, valid


def parse_model_media(url: str) -> tuple[str | None, str | None]:
    """Return ``(model_key, target_id)`` for model-card media filenames."""

    return MODEL_TARGET_STEMS.get(filename_stem(url), (None, None))


def parse_bodystyle_media(url: str) -> tuple[str | None, str | None, str | None]:
    """Return ``(model_key, body_style_target_id, image_field)`` for body style media."""

    match = MODEL_BODY_STYLE_RE.match(filename_stem(url))
    if not match:
        return None, None, None
    model_key = MODEL_PREFIX[match.group(1)]
    body_style = BODY_STYLE_CODE[match.group(2)]
    image_field = BODY_STYLE_IMAGE_FIELD[match.group(3)]
    return model_key, f"body_style__{body_style}", image_field


def build_media_index(media_urls: Iterable[str]) -> tuple[dict[tuple[str, str], list[str]], dict[str, list[str]], list[str]]:
    exact: dict[tuple[str, str], list[str]] = defaultdict(list)
    bare: dict[str, list[str]] = defaultdict(list)
    unparseable: list[str] = []
    for url in dict.fromkeys(media_urls):
        model, rpo, ok = parse_media(url)
        if not ok:
            unparseable.append(url)
        elif model:
            exact[(model, rpo)].append(url)
        else:
            bare[rpo].append(url)
    return exact, bare, unparseable


def build_media_inventory(media_urls: Iterable[str]) -> MediaInventory:
    option_exact: dict[tuple[str, str], list[str]] = defaultdict(list)
    option_shared: dict[tuple[str, str], list[str]] = defaultdict(list)
    option_bare: dict[str, list[str]] = defaultdict(list)
    model_media: dict[tuple[str, str], list[str]] = defaultdict(list)
    bodystyle_media: dict[tuple[str, str, str], list[str]] = defaultdict(list)
    unparseable: list[str] = []
    # WordPress can expose multiple attachment records for the same physical
    # source URL. Matching is URL-based, so duplicate records are one candidate,
    # not an ambiguity.
    for url in dict.fromkeys(media_urls):
        parsed_any = False
        model_key, target_id = parse_model_media(url)
        if model_key and target_id:
            model_media[(model_key, target_id)].append(url)
            parsed_any = True
        body_model, body_target, image_field = parse_bodystyle_media(url)
        if body_model and body_target and image_field:
            bodystyle_media[(body_model, body_target, image_field)].append(url)
            parsed_any = True
        shared_prefix, rpo, shared_ok = parse_shared_option_media(url)
        if shared_ok and not parsed_any:
            parsed_any = True
            option_shared[(shared_prefix, rpo)].append(url)
        if not shared_ok:
            option_model, rpo, ok = parse_media(url)
            if ok and not parsed_any:
                parsed_any = True
                if option_model:
                    option_exact[(option_model, rpo)].append(url)
                else:
                    option_bare[rpo].append(url)
        if not parsed_any:
            unparseable.append(url)
    return MediaInventory(
        option_exact=dict(option_exact),
        option_bare=dict(option_bare),
        model=dict(model_media),
        bodystyle=dict(bodystyle_media),
        unparseable=unparseable,
        option_shared=dict(option_shared),
    )


def _auth_header_from_env() -> str | None:
    user = os.environ.get("WP_USER")
    password = os.environ.get("WP_APP_PASSWORD")
    if not user or not password:
        return None
    token = f"{user}:{password.replace(' ', '')}".encode("utf-8")
    return "Basic " + base64.b64encode(token).decode("ascii")


def _open_json(url: str, *, auth_header: str | None, timeout: float) -> tuple[list[dict[str, Any]], dict[str, str]]:
    headers = {
        "Accept": "application/json",
        "Cache-Control": "no-cache, no-store, max-age=0",
        "Pragma": "no-cache",
        "User-Agent": WORDPRESS_USER_AGENT,
    }
    if auth_header:
        headers["Authorization"] = auth_header
    request = Request(url, headers=headers)
    with urlopen(request, timeout=timeout) as response:  # noqa: S310 - project-controlled endpoint/CLI URL
        payload = json.loads(response.read().decode("utf-8"))
        response_headers = {key.lower(): value for key, value in response.headers.items()}
    return payload, response_headers


def _media_fetch_error(exc: HTTPError, api_url: str) -> WordPressMediaFetchError:
    return WordPressMediaFetchError(
        f"WordPress media fetch failed with HTTP {exc.code} from {api_url}. "
        "If media is private, set WP_USER/WP_APP_PASSWORD; otherwise use "
        "--media-url-list <path> for deterministic report/apply review."
    )


def fetch_media_snapshot(
    timeout: float,
    modified_after: str | None = None,
    *,
    cache_token: str | None = None,
) -> MediaSnapshot:
    """Fetch one uncached WordPress media inventory snapshot."""

    auth_header = _auth_header_from_env()
    urls: list[str] = []
    modified_by_url: dict[str, str] = {}
    page = 1
    while True:
        params: dict[str, Any] = {
            "per_page": 100,
            "page": page,
            "_fields": "source_url,modified",
            "media_type": "image",
        }
        if cache_token:
            params["asset_sync_cache_bust"] = cache_token
        if modified_after:
            params.update({"modified_after": modified_after, "orderby": "modified", "order": "desc"})
        api_url = f"{MEDIA_ENDPOINT}?{urlencode(params)}"
        try:
            batch, headers = _open_json(api_url, auth_header=auth_header, timeout=timeout)
        except HTTPError as exc:
            if exc.code == 400 and page > 1:
                break
            if exc.code in {401, 403}:
                raise _media_fetch_error(exc, api_url) from exc
            raise
        if not batch:
            break
        for item in batch:
            url = clean(item.get("source_url"))
            if PATH_FILTER in url:
                urls.append(url)
                modified = clean(item.get("modified"))
                if modified > modified_by_url.get(url, ""):
                    modified_by_url[url] = modified
        total_pages = int(headers.get("x-wp-totalpages", page))
        if page >= total_pages:
            break
        page += 1
    # Multiple WordPress attachment rows may resolve to the same physical URL.
    # Preserve first-seen API order while collapsing those duplicate records.
    return MediaSnapshot(urls=list(dict.fromkeys(urls)), modified_by_url=modified_by_url)


def fetch_media(timeout: float, modified_after: str | None = None) -> list[str]:
    """Compatibility wrapper returning URLs from one media snapshot."""

    return fetch_media_snapshot(timeout, modified_after).urls


def fetch_media_stable(
    timeout: float,
    modified_after: str | None = None,
    *,
    attempts: int = 4,
) -> MediaSnapshot:
    """Require two identical uncached snapshots before a complete apply."""

    previous: MediaSnapshot | None = None
    for attempt in range(max(2, attempts)):
        snapshot = fetch_media_snapshot(
            timeout,
            modified_after,
            cache_token=f"{time.time_ns()}-{attempt}",
        )
        if previous is not None and (
            sorted(previous.urls) == sorted(snapshot.urls)
            and previous.modified_by_url == snapshot.modified_by_url
        ):
            return snapshot
        previous = snapshot
        time.sleep(0.25)
    raise RuntimeError(
        "WordPress media inventory did not stabilize after "
        f"{max(2, attempts)} uncached snapshots; no workbook write was attempted."
    )


def read_media_url_list(path: Path | str) -> list[str]:
    media_path = Path(path)
    urls: list[str] = []
    for line in media_path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if stripped and not stripped.startswith("#"):
            urls.append(stripped)
    return urls


def state_path(report_dir: Path) -> Path:
    return report_dir / ".asset_map_sync_state.json"


def read_state(report_dir: Path) -> tuple[dict[str, Any], bool]:
    path = state_path(report_dir)
    if not path.exists():
        return {}, False
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return (payload if isinstance(payload, dict) else {}), True
    except (OSError, TypeError, json.JSONDecodeError):
        return {}, True


def read_since_auto(report_dir: Path, cushion_hours: int = 6) -> tuple[str | None, bool]:
    payload, exists = read_state(report_dir)
    if not exists:
        return None, False
    try:
        timestamp = payload.get("last_run_utc")
        if not timestamp:
            return None, True
        return (datetime.fromisoformat(timestamp) - timedelta(hours=cushion_hours)).strftime("%Y-%m-%dT%H:%M:%S"), True
    except (ValueError, TypeError):
        return None, True


def _with_asset_revision(url: str, token: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["asset_rev"] = token
    return urlunparse(parsed._replace(query=urlencode(query)))


def prepare_revisioned_media_urls(
    snapshot: MediaSnapshot,
    prior_state: dict[str, Any],
) -> tuple[list[str], dict[str, str]]:
    """Version only URLs whose attachment modified time changed after a known baseline."""

    prior_modified = prior_state.get("media_modified", {})
    if not isinstance(prior_modified, dict):
        prior_modified = {}
    prior_tokens = prior_state.get("revision_tokens", {})
    if not isinstance(prior_tokens, dict):
        prior_tokens = {}
    tokens: dict[str, str] = {}
    urls: list[str] = []
    for url in snapshot.urls:
        modified = snapshot.modified_by_url.get(url, "")
        token = clean(prior_tokens.get(url))
        if url in prior_modified and modified and clean(prior_modified.get(url)) != modified:
            token = re.sub(r"[^0-9A-Za-z]", "", modified)
        if token:
            tokens[url] = token
            urls.append(_with_asset_revision(url, token))
        else:
            urls.append(url)
    return urls, tokens


def write_state(
    report_dir: Path,
    *,
    media_modified: dict[str, str] | None = None,
    revision_tokens: dict[str, str] | None = None,
) -> None:
    payload, _exists = read_state(report_dir)
    payload["last_run_utc"] = datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec="seconds")
    if media_modified is not None:
        payload["media_modified"] = media_modified
    if revision_tokens is not None:
        payload["revision_tokens"] = revision_tokens
    report_dir.mkdir(parents=True, exist_ok=True)
    state_path(report_dir).write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def url_alive(url: str, timeout: float) -> bool:
    try:
        request = Request(url, method="HEAD")
        with urlopen(request, timeout=timeout) as response:  # noqa: S310 - user/workbook-provided URL liveness check
            return response.status < 400
    except HTTPError as exc:
        if exc.code != 405:
            return False
        try:
            request = Request(url, method="GET")
            with urlopen(request, timeout=timeout) as response:  # noqa: S310 - user/workbook-provided URL liveness check
                return response.status < 400
        except (HTTPError, URLError, TimeoutError):
            return False
    except (URLError, TimeoutError):
        return False


def check_existing(urls: Iterable[str], timeout: float, workers: int) -> dict[str, bool]:
    unique_urls = sorted(set(urls))
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        return dict(executor.map(lambda url: (url, url_alive(url, timeout)), unique_urls))


def discover_promoted_option_sources(wb) -> dict[str, str]:
    """Return promoted runtime model -> source option sheet mapping."""

    if "model_registry_promotion" not in wb.sheetnames:
        raise ValueError("Missing required workbook sheet 'model_registry_promotion'.")
    if "model_workbook_sources" not in wb.sheetnames:
        raise ValueError("Missing required workbook sheet 'model_workbook_sources'.")

    promoted_rows = [
        row
        for row in rows_from_sheet(wb, "model_registry_promotion")
        if workbook_truthy(row.get("active")) and workbook_truthy(row.get("promoted_to_runtime"))
    ]
    promoted_rows.sort(key=lambda row: (int(clean(row.get("display_order")) or 0), clean(row.get("model_key"))))
    promoted_models = [clean(row.get("model_key")).lower() for row in promoted_rows if clean(row.get("model_key"))]

    source_rows = [
        row
        for row in rows_from_sheet(wb, "model_workbook_sources")
        if workbook_truthy(row.get("active")) and clean(row.get("source_role")) == "source_option_sheet"
    ]
    source_by_model = {clean(row.get("model_key")).lower(): clean(row.get("sheet_name")) for row in source_rows}

    sources: dict[str, str] = {}
    for model_key in promoted_models:
        sheet_name = source_by_model.get(model_key)
        if not sheet_name:
            raise ValueError(f"Promoted model {model_key!r} is missing an active source_option_sheet row.")
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Promoted model {model_key!r} source option sheet {sheet_name!r} is missing.")
        sources[model_key] = sheet_name
    return sources


def _header_index(ws) -> dict[str, int]:
    headers = [clean(cell.value).lower() for cell in ws[1]]
    return {header: index for index, header in enumerate(headers) if header}


def read_option_sheets(wb, sources: dict[str, str]) -> dict[tuple[str, str, str], dict[str, str]]:
    desired: dict[tuple[str, str, str], dict[str, str]] = {}
    for model_key, sheet_name in sources.items():
        ws = wb[sheet_name]
        index = _header_index(ws)
        missing = {"option_id", "rpo", "selectable", "active"} - set(index)
        if missing:
            raise ValueError(f"{sheet_name} missing required columns: {', '.join(sorted(missing))}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(clean(value) for value in row):
                continue
            option_id = clean(row[index["option_id"]]).lower()
            if not option_id:
                continue
            if not (workbook_truthy(row[index["active"]]) and workbook_truthy(row[index["selectable"]])):
                continue
            rpo = clean(row[index["rpo"]]).lower()
            name = clean(row[index["option_name"]]) if "option_name" in index else ""
            section_id = clean(row[index["section_id"]]).lower() if "section_id" in index else ""
            desired[(model_key, TARGET_TYPE_OPTION, option_id)] = {
                "target_type": TARGET_TYPE_OPTION,
                "rpo": rpo,
                "name": name,
                "section_id": section_id,
                "source_sheet": sheet_name,
            }
    return desired


def read_model_targets(wb) -> dict[tuple[str, str, str], dict[str, str]]:
    desired: dict[tuple[str, str, str], dict[str, str]] = {}
    if "model_registry_promotion" not in wb.sheetnames:
        return desired
    for row in rows_from_sheet(wb, "model_registry_promotion"):
        if not (workbook_truthy(row.get("active")) and workbook_truthy(row.get("promoted_to_runtime"))):
            continue
        model_key = clean(row.get("model_key")).lower()
        target_id = clean(row.get("registry_key")) or model_key
        if not model_key or not target_id:
            continue
        desired[(model_key, TARGET_TYPE_MODEL, target_id)] = {
            "target_type": TARGET_TYPE_MODEL,
            "rpo": "",
            "name": clean(row.get("model_label")) or target_id,
            "section_id": "",
            "source_sheet": "model_registry_promotion",
        }
    return desired


def read_bodystyle_targets(sources: dict[str, str]) -> dict[tuple[str, str, str], dict[str, str]]:
    desired: dict[tuple[str, str, str], dict[str, str]] = {}
    for model_key in sources:
        for body_style, display_name in (("coupe", "Coupe"), ("convertible", "Convertible")):
            target_id = f"body_style__{body_style}"
            desired[(model_key, TARGET_TYPE_CONTEXT_CHOICE, target_id)] = {
                "target_type": TARGET_TYPE_CONTEXT_CHOICE,
                "rpo": "",
                "name": display_name,
                "section_id": "sec_context_body_style",
                "source_sheet": "generated_body_style_context",
            }
    return desired


@dataclass(frozen=True)
class SectionCoverageMetadata:
    """Workbook-derived section metadata used by the coverage-intent classifier."""

    selection_mode: dict[str, str]
    standard_equipment_buckets: set[tuple[str, str]]


def read_section_coverage_metadata(wb) -> SectionCoverageMetadata:
    """Read section metadata for coverage classification; tolerates missing sheets."""

    selection_mode: dict[str, str] = {}
    buckets: set[tuple[str, str]] = set()
    if "section_master" in wb.sheetnames:
        for row in rows_from_sheet(wb, "section_master"):
            section_id = clean(row.get("section_id")).lower()
            if not section_id:
                continue
            selection_mode[section_id] = clean(row.get("selection_mode")).lower()
    if "section_presentation" in wb.sheetnames:
        for row in rows_from_sheet(wb, "section_presentation"):
            if not (workbook_truthy(row.get("active")) and workbook_truthy(row.get("standard_equipment_bucket"))):
                continue
            model_key = clean(row.get("model_key")).lower()
            section_id = clean(row.get("section_id")).lower()
            if model_key and section_id:
                buckets.add((model_key, section_id))
    return SectionCoverageMetadata(
        selection_mode=selection_mode,
        standard_equipment_buckets=buckets,
    )


def build_coverage_classifier(
    desired: dict[tuple[str, str, str], dict[str, str]],
    section_metadata: SectionCoverageMetadata,
    existing_rows: dict[tuple[str, str, str], dict[str, Any]],
) -> Callable[[str, str, str], tuple[str, str]]:
    """Return a pure ``(model_key, target_type, target_id) -> (intent, reason)`` classifier.

    Universal-expected policy (phase4b-v1): every active+selectable option card
    is expected to carry a visual element. ``not_expected`` derives only from
    structural presentation metadata (display-only sections, standard-equipment
    buckets) — never from media inventory or asset_map coverage state, so
    classifications self-correct when workbook presentation metadata changes.
    ``existing_rows`` is accepted for signature stability but unused by policy.
    """

    del existing_rows  # policy must not depend on coverage state (no circularity)

    def classify(model_key: str, target_type: str, target_id: str) -> tuple[str, str]:
        if target_type in (TARGET_TYPE_MODEL, TARGET_TYPE_CONTEXT_CHOICE):
            return COVERAGE_EXPECTED, f"target_type:{target_type}"
        info = desired.get((model_key, target_type, target_id), {})
        section_id = info.get("section_id", "")
        if section_id and section_metadata.selection_mode.get(section_id) == "display_only":
            return COVERAGE_NOT_EXPECTED, f"section-display-only:{section_id}"
        if section_id and (model_key, section_id) in section_metadata.standard_equipment_buckets:
            return COVERAGE_NOT_EXPECTED, f"standard-equipment-bucket:{section_id}"
        if not section_id or section_id not in section_metadata.selection_mode:
            return COVERAGE_EXPECTED, "unmatched-section"
        return COVERAGE_EXPECTED, "universal-expected"

    return classify


def existing_asset_rows(ws, header_index: dict[str, int]) -> dict[tuple[str, str, str], dict[str, Any]]:
    rows: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        target_type = clean(row[header_index["target_type"]]).lower()
        if target_type not in SUPPORTED_TARGET_TYPES:
            continue
        model_key = clean(row[header_index["model_key"]]).lower()
        target_id = clean(row[header_index["target_id"]])
        target_id_key = target_id.lower() if target_type == TARGET_TYPE_OPTION else target_id
        if not model_key or not target_id_key:
            continue
        values: dict[str, Any] = {
            "row": row_number,
            "source_sheet": ws.title,
            "source_row": row_number,
            "target_type": target_type,
            "url": clean(row[header_index["image_url"]]),
        }
        for column in WRITABLE_COLUMNS["asset_map"]:
            if column in header_index:
                values[column] = clean(row[header_index[column]])
        rows[(model_key, target_type, target_id_key)] = values
    return rows


def resolve_candidate(
    media: MediaInventory,
    model_key: str,
    target_id: str,
    info: dict[str, str],
) -> CandidateResolution:
    """Resolve one target using the sync owner's exact fail-closed precedence."""

    target_type = info.get("target_type", TARGET_TYPE_OPTION)
    if target_type == TARGET_TYPE_OPTION:
        rpo = info.get("rpo", "")
        if not rpo:
            return CandidateResolution({}, "no-rpo", "no rpo in option sheet", None)

        candidates = media.option_exact.get((model_key, rpo), [])
        if len(candidates) == 1:
            return CandidateResolution({"image_url": candidates[0]}, "prefixed", "", 1)
        if len(candidates) > 1:
            source = "prefixed-ambiguous"
            return CandidateResolution(
                {}, source,
                f"multiple {model_key}-prefixed files for '{rpo}'; keep one file at this priority",
                1,
                tuple(CandidateAlternative("image_url", url, source, 1) for url in candidates),
            )

        eligible_groups: list[tuple[int, str, list[str]]] = []
        for (prefix_group, candidate_rpo), group_candidates in media.option_shared.items():
            if candidate_rpo != rpo:
                continue
            group_models = tuple(MODEL_PREFIX[code] for code in prefix_group.split("-"))
            if model_key in group_models:
                eligible_groups.append((len(group_models), prefix_group, group_candidates))
        if eligible_groups:
            winning_size = min(size for size, _prefix, _candidates in eligible_groups)
            winning_groups = [
                (prefix, group_candidates)
                for size, prefix, group_candidates in eligible_groups
                if size == winning_size
            ]
            winning_candidates = [
                (prefix, url)
                for prefix, group_candidates in winning_groups
                for url in group_candidates
            ]
            if len(winning_candidates) == 1:
                prefix_group, url = winning_candidates[0]
                source = f"shared-prefix:{prefix_group}"
                return CandidateResolution(
                    {"image_url": url}, source,
                    f"using {prefix_group}-prefixed media shared with {model_key}", 2,
                )
            prefix_labels = ", ".join(sorted(prefix for prefix, _candidates in winning_groups))
            source_label = winning_groups[0][0] if len(winning_groups) == 1 else prefix_labels
            source = f"shared-prefix:{source_label}:ambiguous"
            return CandidateResolution(
                {}, source,
                f"multiple {prefix_labels} shared-prefix files for '{rpo}' at the winning specificity",
                2,
                tuple(
                    CandidateAlternative("image_url", url, f"shared-prefix:{prefix}", 2)
                    for prefix, url in winning_candidates
                ),
            )

        for candidate_model in OPTION_MODEL_FALLBACKS.get(model_key, ()):
            source = f"model-fallback:{candidate_model}"
            candidates = media.option_exact.get((candidate_model, rpo), [])
            if len(candidates) == 1:
                return CandidateResolution(
                    {"image_url": candidates[0]}, source,
                    f"using {candidate_model}-prefixed media as configured fallback for {model_key}", 3,
                )
            if len(candidates) > 1:
                ambiguous_source = f"{source}:ambiguous"
                return CandidateResolution(
                    {}, ambiguous_source,
                    f"multiple {candidate_model}-prefixed files for '{rpo}'; keep one file at this priority",
                    3,
                    tuple(
                        CandidateAlternative("image_url", url, source, 3)
                        for url in candidates
                    ),
                )
        candidates = media.option_bare.get(rpo, [])
        if len(candidates) == 1:
            return CandidateResolution({"image_url": candidates[0]}, "bare-shared", "", 4)
        if len(candidates) > 1:
            return CandidateResolution(
                {}, "bare-ambiguous",
                f"multiple bare files for '{rpo}'; keep one shared file or add c/e/h/r/s/g prefixes",
                4,
                tuple(CandidateAlternative("image_url", url, "bare-shared", 4) for url in candidates),
            )
        return CandidateResolution({}, "none", "", None)

    if target_type == TARGET_TYPE_MODEL:
        candidates = media.model.get((model_key, target_id), [])
        if len(candidates) == 1:
            return CandidateResolution({"image_url": candidates[0]}, "model-filename", "", 1)
        if len(candidates) > 1:
            return CandidateResolution(
                {}, "model-ambiguous", f"multiple model files for '{target_id}'", 1,
                tuple(CandidateAlternative("image_url", url, "model-filename", 1) for url in candidates),
            )
        return CandidateResolution({}, "none", "", None)

    if target_type == TARGET_TYPE_CONTEXT_CHOICE:
        fields: dict[str, str] = {}
        alternatives: list[CandidateAlternative] = []
        ambiguous: list[str] = []
        for candidate_field in ("image_url", "hover_image_url"):
            candidates = media.bodystyle.get((model_key, target_id, candidate_field), [])
            if len(candidates) == 1:
                fields[candidate_field] = candidates[0]
            elif len(candidates) > 1:
                ambiguous.append(candidate_field)
                alternatives.extend(
                    CandidateAlternative(candidate_field, url, "bodystyle-filename", 1)
                    for url in candidates
                )
        if ambiguous:
            return CandidateResolution(
                {}, "bodystyle-ambiguous",
                f"multiple body-style files for {', '.join(ambiguous)}", 1,
                tuple(alternatives),
            )
        if fields:
            return CandidateResolution(fields, "bodystyle-filename", "", 1)
        return CandidateResolution({}, "none", "", None)

    return CandidateResolution(
        {}, "unsupported-target-type", f"unsupported target_type '{target_type}'", None
    )


def reconcile(
    desired: dict[tuple[str, str, str], dict[str, str]],
    media: MediaInventory | dict[tuple[str, str], list[str]],
    bare: dict[str, list[str]] | None = None,
    existing_rows: dict[tuple[str, str, str], dict[str, Any]] | None = None,
    alive: dict[str, bool] | None = None,
    incremental: bool = False,
    classify_coverage: Callable[[str, str, str], tuple[str, str]] | None = None,
    update_safe_wildcards: bool = False,
) -> SyncPlan:
    """Pure reconciliation of desired asset targets vs current hosted media inventory."""

    if not isinstance(media, MediaInventory):
        media = MediaInventory(
            option_exact=media,
            option_bare=bare or {},
            model={},
            bodystyle={},
            unparseable=[],
        )
    existing_rows = existing_rows or {}
    alive = alive or {}
    # Wildcard rows (model_key "*", option targets) cover every promoted model.
    # Sync never writes/edits/inserts wildcard rows; divergence between a
    # wildcard row and canonical media is a human decision (wildcard_conflict).
    wildcard_rows = {
        (target_type, target_id): existing
        for (model_key, target_type, target_id), existing in existing_rows.items()
        if model_key == WILDCARD_MODEL_KEY
    }

    safe_wildcard_candidates: dict[tuple[str, str], str] = {}
    if update_safe_wildcards:
        for target_type, target_id in wildcard_rows:
            if target_type != TARGET_TYPE_OPTION:
                continue
            resolutions = [
                resolve_candidate(media, model_key, target_id, info)
                for (model_key, desired_type, desired_id), info in desired.items()
                if desired_type == target_type and desired_id == target_id
            ]
            candidate_urls = {
                resolution.fields.get("image_url", "")
                for resolution in resolutions
                if resolution.fields and resolution.source == "bare-shared"
            }
            if resolutions and len(candidate_urls) == 1 and all(
                resolution.fields and resolution.source == "bare-shared"
                for resolution in resolutions
            ):
                safe_wildcard_candidates[(target_type, target_id)] = candidate_urls.pop()

    report: list[dict[str, str]] = []
    url_writes: dict[tuple[int, str], str] = {}
    inserts: list[dict[str, Any]] = []
    status: dict[int, str] = {}
    used: set[str] = set()
    resolutions: dict[tuple[str, str, str], CandidateResolution] = {}

    def add_report(
        scope: str,
        model_key: str,
        source_sheet: str,
        target_type: str,
        target_id: str,
        rpo: str,
        action: str,
        source: str,
        existing_url: str,
        new_url: str,
        image_status: str,
        note: str = "",
    ) -> None:
        info = desired.get((model_key, target_type, target_id), {})
        if classify_coverage is None:
            coverage_intent, coverage_reason = "", ""
        else:
            coverage_intent, coverage_reason = classify_coverage(model_key, target_type, target_id)
        report.append(
            {
                "scope": scope,
                "model_key": model_key,
                "source_sheet": source_sheet,
                "section_id": info.get("section_id", ""),
                "target_type": target_type or info.get("target_type", ""),
                "target_id": target_id,
                "rpo": rpo,
                "option_name": info.get("name", ""),
                "action": action,
                "candidate_source": source,
                "existing_url": existing_url,
                "new_url": new_url,
                "image_status": image_status,
                "coverage_intent": coverage_intent,
                "coverage_intent_reason": coverage_reason,
                "note": note,
            }
        )

    for (model_key, target_type, target_id), info in desired.items():
        rpo = info.get("rpo", "")
        source_sheet = info.get("source_sheet", "")
        resolution = resolve_candidate(media, model_key, target_id, info)
        resolutions[(model_key, target_type, target_id)] = resolution
        fields, source, note = resolution.fields, resolution.source, resolution.reason
        for url in fields.values():
            used.add(url)
        for alternative in resolution.alternatives:
            used.add(alternative.url)

        existing = existing_rows.get((model_key, target_type, target_id))
        if existing:
            row_number = int(existing["row"])
            if fields:
                changed_fields = []
                for field, candidate in fields.items():
                    existing_value = clean(existing.get("url" if field == "image_url" else field))
                    if existing_value != candidate:
                        url_writes[(row_number, field)] = candidate
                        changed_fields.append(field)
                if changed_fields:
                    status[row_number] = "ok"
                    existing_url = clean(existing.get("url"))
                    action = "fill" if not existing_url else "replace_canonical"
                    add_report(
                        "existing",
                        model_key,
                        source_sheet,
                        target_type,
                        target_id,
                        rpo,
                        action,
                        source,
                        existing_url,
                        fields.get("image_url", existing_url),
                        "ok",
                        f"canonical media inventory differs in: {', '.join(changed_fields)}",
                    )
                else:
                    status[row_number] = "ok"
                    existing_url = clean(existing.get("url"))
                    add_report("existing", model_key, source_sheet, target_type, target_id, rpo, "keep", source, existing_url, existing_url, "ok")
            elif source.endswith("ambiguous"):
                status[row_number] = "ambiguous"
                add_report("existing", model_key, source_sheet, target_type, target_id, rpo, "flag_ambiguous", source, clean(existing.get("url")), "", "ambiguous", note)
            elif incremental:
                add_report("existing", model_key, source_sheet, target_type, target_id, rpo, "skip_no_candidate_incremental", source, clean(existing.get("url")), "", "missing", note)
            elif alive.get(clean(existing.get("url")), True) is False:
                action = "dead_no_match_incremental" if incremental else "flag_dead_no_match"
                status[row_number] = "url_dead"
                add_report("existing", model_key, source_sheet, target_type, target_id, rpo, action, source, clean(existing.get("url")), "", "url_dead", note)
            else:
                status[row_number] = "missing"
                add_report("existing", model_key, source_sheet, target_type, target_id, rpo, "flag_missing", source, clean(existing.get("url")), "", "missing", note)
            continue

        wildcard = wildcard_rows.get((target_type, target_id)) if target_type == TARGET_TYPE_OPTION else None
        if wildcard is not None:
            wildcard_url = clean(wildcard.get("url"))
            candidate_url = fields.get("image_url", "") if fields else ""
            safe_candidate = safe_wildcard_candidates.get((target_type, target_id), "")
            if safe_candidate and safe_candidate != wildcard_url:
                row_number = int(wildcard["row"])
                url_writes[(row_number, "image_url")] = safe_candidate
                status[row_number] = "ok"
                add_report(
                    "existing",
                    model_key,
                    source_sheet,
                    target_type,
                    target_id,
                    rpo,
                    "replace_shared_canonical",
                    "bare-shared",
                    wildcard_url,
                    safe_candidate,
                    "ok",
                    "unique bare media safely replaces the shared wildcard row",
                )
            elif candidate_url and candidate_url != wildcard_url:
                add_report(
                    "existing",
                    model_key,
                    source_sheet,
                    target_type,
                    target_id,
                    rpo,
                    "wildcard_conflict",
                    source,
                    wildcard_url,
                    candidate_url,
                    "ok",
                    "canonical media differs from shared wildcard row; routine sync never edits wildcard rows, "
                    "and complete sync edits only a unique bare shared candidate",
                )
            else:
                add_report(
                    "existing",
                    model_key,
                    source_sheet,
                    target_type,
                    target_id,
                    rpo,
                    "keep",
                    source,
                    wildcard_url,
                    wildcard_url,
                    "ok",
                    "covered by shared wildcard row",
                )
            continue

        if fields:
            insert = {
                "model": model_key,
                "target_type": info.get("target_type", TARGET_TYPE_OPTION),
                "tid": target_id,
                "rpo": rpo,
                "name": info.get("name", ""),
                "fields": fields,
                "url": fields.get("image_url", ""),
                "status": "ok",
            }
            if source_sheet:
                insert["source_sheet"] = source_sheet
            inserts.append(insert)
            add_report("new", model_key, source_sheet, target_type, target_id, rpo, "insert_filled", source, "", fields.get("image_url", ""), "ok", note)
        elif source.endswith("ambiguous"):
            add_report("new", model_key, source_sheet, target_type, target_id, rpo, "flag_ambiguous", source, "", "", "ambiguous", note)
        elif incremental:
            add_report("new", model_key, source_sheet, target_type, target_id, rpo, "skip_no_candidate_incremental", source, "", "", "missing", note)
        else:
            add_report("new", model_key, source_sheet, target_type, target_id, rpo, "flag_missing", source, "", "", "missing", note)

    for (model_key, target_type, target_id), existing in existing_rows.items():
        if model_key == WILDCARD_MODEL_KEY:
            # A wildcard row is stale only if NO promoted model desires the target.
            if any(
                key[1] == target_type and key[2] == target_id
                for key in desired
            ):
                continue
        elif (model_key, target_type, target_id) in desired:
            continue
        row_number = int(existing["row"])
        status[row_number] = "stale_target"
        add_report(
            "stale",
            model_key,
            "",
            target_type,
            target_id,
            "",
            "stale_target",
            "",
            clean(existing.get("url")),
            clean(existing.get("url")),
            "stale_target",
            "asset target no longer desired by current promoted model inventory",
        )

    return SyncPlan(
        report=report,
        url_writes=url_writes,
        inserts=inserts,
        status=status,
        used=used,
        resolutions=resolutions,
    )


def _ensure_asset_headers(headers: list[str]) -> dict[str, int]:
    index = {header: offset for offset, header in enumerate(headers)}
    missing = {"model_key", "target_type", "target_id", "image_url"} - set(index)
    if missing:
        raise ValueError(f"asset_map missing required columns: {', '.join(sorted(missing))}")
    return index


def _write_reports(
    report_dir: Path,
    report: list[dict[str, str]],
    unmatched: list[str],
    unparseable: list[str],
    incremental: bool,
) -> tuple[Path, Path, Path, int, int]:
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / "asset_map_sync_report.csv"
    report_fieldnames = [
        "scope",
        "model_key",
        "source_sheet",
        "section_id",
        "target_type",
        "target_id",
        "rpo",
        "option_name",
        "action",
        "candidate_source",
        "existing_url",
        "new_url",
        "image_status",
        "coverage_intent",
        "coverage_intent_reason",
        "note",
    ]
    with report_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=report_fieldnames)
        writer.writeheader()
        writer.writerows(report)

    missing_path = report_dir / "asset_map_missing_images.csv"
    missing_fieldnames = [
        "model_key",
        "source_sheet",
        "section_id",
        "target_type",
        "target_id",
        "rpo",
        "option_name",
        "action",
        "candidate_source",
        "image_status",
        "coverage_intent",
        "coverage_intent_reason",
        "note",
    ]
    broad_missing_rows = [row for row in report if row["action"] in MISSING_IMAGE_ACTIONS]
    # Actionable review queue: expected only (universal policy). Structural
    # not_expected missing rows stay visible in the broad report CSV with
    # intent columns populated. Sorted model -> section -> target for triage.
    missing_rows = sorted(
        (row for row in broad_missing_rows if row.get("coverage_intent", "") in ACTIONABLE_COVERAGE_INTENTS),
        key=lambda row: (row.get("model_key", ""), row.get("section_id", ""), row.get("target_id", "")),
    )
    with missing_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=missing_fieldnames)
        writer.writeheader()
        for row in missing_rows:
            writer.writerow({field: row.get(field, "") for field in missing_fieldnames})

    unmatched_path = report_dir / "asset_map_unmatched_media.csv"
    reason = "new media in window, no row yet" if incremental else "no desired (model, rpo) for this file"
    with unmatched_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["source_url", "parsed_model", "parsed_rpo", "reason"])
        for url in unmatched:
            shared_prefix, rpo, shared_ok = parse_shared_option_media(url)
            if shared_ok:
                writer.writerow([url, shared_prefix, rpo, reason])
            else:
                model_key, rpo, _ = parse_media(url)
                writer.writerow([url, model_key or "", rpo, reason])
        for url in unparseable:
            writer.writerow([url, "", "", "filename did not yield a 3-char RPO"])
    return report_path, missing_path, unmatched_path, len(missing_rows), len(broad_missing_rows)


def build_section_coverage_stats(
    desired: dict[tuple[str, str, str], dict[str, str]],
    existing_rows: dict[tuple[str, str, str], dict[str, Any]],
) -> dict[str, Any]:
    """Per-model/per-section coverage stats over ALL desired option targets.

    A target counts as covered when an existing asset_map row carries an
    image_url — either an exact-model row or a shared wildcard row
    (model_key "*"). Computed over the full desired set (not the missing
    subset) so percentages converge to 100 as images land.
    """

    per_model: dict[str, dict[str, dict[str, int]]] = {}
    for (model_key, target_type, target_id), info in desired.items():
        if target_type != TARGET_TYPE_OPTION:
            continue
        section_id = info.get("section_id", "") or "(none)"
        bucket = per_model.setdefault(model_key, {}).setdefault(
            section_id, {"total_targets": 0, "covered": 0, "missing": 0}
        )
        bucket["total_targets"] += 1
        existing = existing_rows.get((model_key, target_type, target_id)) or existing_rows.get(
            (WILDCARD_MODEL_KEY, target_type, target_id)
        )
        if existing and clean(existing.get("url")):
            bucket["covered"] += 1
        else:
            bucket["missing"] += 1

    stats: dict[str, Any] = {}
    for model_key in sorted(per_model):
        sections: dict[str, Any] = {}
        model_total = model_covered = 0
        for section_id in sorted(per_model[model_key]):
            bucket = per_model[model_key][section_id]
            model_total += bucket["total_targets"]
            model_covered += bucket["covered"]
            sections[section_id] = {
                **bucket,
                "coverage_pct": round(100.0 * bucket["covered"] / bucket["total_targets"], 1),
            }
        stats[model_key] = {
            "sections": sections,
            "total_targets": model_total,
            "covered": model_covered,
            "missing": model_total - model_covered,
            "coverage_pct": round(100.0 * model_covered / model_total, 1) if model_total else 0.0,
        }
    return stats


def build_coverage_summary(report: list[dict[str, str]]) -> dict[str, Any]:
    """Coverage-intent metrics over broad missing rows for the manifest."""

    broad_missing = [row for row in report if row["action"] in MISSING_IMAGE_ACTIONS]
    intent_counts = Counter(row.get("coverage_intent", "") or "unclassified" for row in broad_missing)
    section_counts: dict[str, dict[str, dict[str, int]]] = {}
    for row in broad_missing:
        model_key = row.get("model_key", "")
        section_id = row.get("section_id", "") or "(none)"
        intent = row.get("coverage_intent", "") or "unclassified"
        section_counts.setdefault(model_key, {}).setdefault(section_id, {}).setdefault(intent, 0)
        section_counts[model_key][section_id][intent] += 1
    return {
        "ruleset_version": COVERAGE_RULESET_VERSION,
        "ruleset": list(COVERAGE_RULESET),
        "broad_missing_count": len(broad_missing),
        "intent_counts": dict(intent_counts),
        "actionable_missing_count": sum(
            count for intent, count in intent_counts.items() if intent != COVERAGE_NOT_EXPECTED
        ),
        "missing_by_model_section_intent": section_counts,
    }


def _write_manifest(
    *,
    report_dir: Path,
    workbook_path: Path,
    asset_sheet: str,
    apply: bool,
    media_source: str,
    since_argument: str | None,
    resolved_modified_after: str | None,
    incremental: bool,
    state_read: bool,
    state_written: bool,
    media_url_count: int,
    verify_existing: bool,
    included_sources: dict[str, str],
    action_counts: dict[str, int],
    url_write_count: int,
    insert_count: int,
    unmatched_count: int,
    unparseable_count: int,
    report_path: Path,
    missing_path: Path,
    missing_count: int,
    broad_missing_count: int,
    coverage_summary: dict[str, Any],
    unmatched_path: Path,
) -> Path:
    manifest_path = report_dir / "asset_map_sync_manifest.json"
    payload = {
        "version": 1,
        "generated_at_utc": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec="seconds") + "Z",
        "workbook_path": str(workbook_path),
        "asset_sheet": asset_sheet,
        "apply": apply,
        "media_source": media_source,
        "since_argument": since_argument,
        "resolved_modified_after": resolved_modified_after,
        "incremental": incremental,
        "state_path": str(state_path(report_dir)),
        "state_read": state_read,
        "state_written": state_written,
        "media_url_count": media_url_count,
        "verify_existing": verify_existing,
        "included_sources": [
            {"model_key": model_key, "option_sheet": option_sheet}
            for model_key, option_sheet in included_sources.items()
        ],
        "action_counts": action_counts,
        "url_write_count": url_write_count,
        "insert_count": insert_count,
        "unmatched_count": unmatched_count,
        "unparseable_count": unparseable_count,
        "report_path": str(report_path),
        "missing_images_path": str(missing_path),
        "missing_images_count": missing_count,
        "broad_missing_images_count": broad_missing_count,
        "coverage": coverage_summary,
        "unmatched_path": str(unmatched_path),
    }
    manifest_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return manifest_path


def build_sync_plan(
    wb,
    *,
    asset_sheet: str,
    media_urls: list[str],
    verify_existing: bool,
    timeout: float,
    workers: int,
    incremental: bool,
    update_safe_wildcards: bool = False,
) -> tuple[SyncPlan, dict[str, str], list[str], list[str]]:
    if asset_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {asset_sheet!r} not found.")
    ws = wb[asset_sheet]
    headers = [clean(cell.value) for cell in ws[1]]
    header_index = _ensure_asset_headers(headers)

    sources = discover_promoted_option_sources(wb)
    desired = read_option_sheets(wb, sources)
    desired.update(read_model_targets(wb))
    desired.update(read_bodystyle_targets(sources))
    existing_rows = existing_asset_rows(ws, header_index)
    media = build_media_inventory(media_urls)
    classify_coverage = build_coverage_classifier(desired, read_section_coverage_metadata(wb), existing_rows)

    if verify_existing:
        alive = check_existing([row["url"] for row in existing_rows.values() if row.get("url")], timeout, workers)
    else:
        alive = {}

    plan = reconcile(
        desired,
        media,
        existing_rows=existing_rows,
        alive=alive,
        incremental=incremental,
        classify_coverage=classify_coverage,
        update_safe_wildcards=update_safe_wildcards,
    )
    plan = replace(
        plan,
        section_coverage=build_section_coverage_stats(desired, existing_rows),
        desired_targets=desired,
        current_rows=existing_rows,
    )
    unmatched = sorted(set(media_urls) - plan.used - set(media.unparseable))
    return plan, sources, unmatched, media.unparseable


ASSET_MANAGER_STATUSES = (
    "safe_proposal",
    "covered",
    "missing",
    "ambiguous",
    "unmatched",
    "unparseable",
    "dead_url",
    "stale_target",
    "wildcard_conflict",
    "ignored",
)
SAFE_PROPOSAL_ACTIONS = {
    "fill", "replace_canonical", "insert_filled", "replace_shared_canonical",
}


def _manager_status(action: str) -> str:
    if action in SAFE_PROPOSAL_ACTIONS:
        return "safe_proposal"
    if action == "keep":
        return "covered"
    if action in {"flag_missing", "skip_no_candidate_incremental"}:
        return "missing"
    if action == "flag_ambiguous":
        return "ambiguous"
    if action in {"flag_dead_no_match", "dead_no_match_incremental"}:
        return "dead_url"
    if action == "stale_target":
        return "stale_target"
    if action == "wildcard_conflict":
        return "wildcard_conflict"
    return action or "missing"


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _stable_item_id(*parts: str) -> str:
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()[:20]


def _current_asset_for_target(
    current_rows: dict[tuple[str, str, str], dict[str, Any]],
    model_key: str,
    target_type: str,
    target_id: str,
) -> tuple[dict[str, Any], str]:
    exact = current_rows.get((model_key, target_type, target_id))
    if exact is not None:
        return exact, "exact"
    if target_type == TARGET_TYPE_OPTION:
        shared = current_rows.get((WILDCARD_MODEL_KEY, target_type, target_id))
        if shared is not None:
            return shared, "shared"
    return {}, "missing"


def add_wildcard_ownership_resolution(items: list[dict[str, Any]]) -> None:
    """Annotate wildcard conflicts with truthful ordinary-operation choices."""

    for item in items:
        if item.get("status") != "wildcard_conflict":
            continue
        target_type = item.get("target_type", "")
        target_id = item.get("target_id", "")
        shared_targets = [
            candidate for candidate in items
            if candidate.get("kind") == "target"
            and candidate.get("target_type") == target_type
            and candidate.get("target_id") == target_id
            and candidate.get("current_values", {}).get("model_key") == WILDCARD_MODEL_KEY
        ]
        candidate_urls = {
            str(candidate.get("proposed_values", {}).get("image_url") or "")
            for candidate in shared_targets
        }
        candidate_urls.discard("")
        shared_candidate = next(iter(candidate_urls)) if len(candidate_urls) == 1 else ""
        current = item.get("current_values", {})
        exact_owner = next((
            candidate.get("current_values")
            for candidate in items
            if candidate.get("kind") == "target"
            and candidate.get("model_key") == item.get("model_key")
            and candidate.get("target_type") == target_type
            and candidate.get("target_id") == target_id
            and candidate.get("current_values", {}).get("model_key") == item.get("model_key")
        ), None)
        item["ownership_resolution"] = {
            "shared_owner": {
                "model_key": WILDCARD_MODEL_KEY,
                "target_type": target_type,
                "target_id": target_id,
                "image_url": current.get("image_url", ""),
            },
            "current_exact_owner": exact_owner,
            "affected_models": sorted({
                str(candidate.get("model_key") or "") for candidate in shared_targets
                if candidate.get("model_key")
            }),
            "exact_operation": {
                "allowed": bool(item.get("model_key") and item.get("proposed_values", {}).get("image_url")),
                "candidate_url": item.get("proposed_values", {}).get("image_url", ""),
                "blocked_reason": "" if item.get("proposed_values", {}).get("image_url") else "No candidate URL is available for this exact model target.",
            },
            "shared_operation": {
                "allowed": bool(shared_targets and len(candidate_urls) == 1),
                "candidate_url": shared_candidate,
                "blocked_reason": "" if shared_targets and len(candidate_urls) == 1 else (
                    "Affected models have different candidate URLs; shared wildcard ownership cannot be changed unambiguously."
                ),
            },
        }


def build_asset_manager_snapshot(
    workbook_path: Path | str,
    media_urls: Iterable[str],
    *,
    media_source: str = "live",
    verify_existing: bool = False,
    timeout: float = 10.0,
    workers: int = 16,
) -> AssetManagerSnapshot:
    """Build the read-only Manager view from the exact CLI reconciliation owner."""

    workbook_path = Path(workbook_path)
    normalized_urls = sorted(set(media_urls))
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        plan, sources, unmatched, unparseable = build_sync_plan(
            wb,
            asset_sheet=ASSET_SHEET,
            media_urls=normalized_urls,
            verify_existing=verify_existing,
            timeout=timeout,
            workers=workers,
            incremental=False,
        )
    finally:
        wb.close()

    items: list[dict[str, Any]] = []
    for row in plan.report:
        model_key = row.get("model_key", "")
        target_type = row.get("target_type", "")
        target_id = row.get("target_id", "")
        resolution = plan.resolutions.get(
            (model_key, target_type, target_id),
            CandidateResolution({}, row.get("candidate_source", ""), row.get("note", ""), None),
        )
        current, coverage_kind = _current_asset_for_target(
            plan.current_rows, model_key, target_type, target_id
        )
        current_values = {
            column: clean(current.get(column, ""))
            for column in WRITABLE_COLUMNS["asset_map"]
        }
        proposed_values = dict(current_values)
        proposed_values.update(resolution.fields)
        if resolution.fields and not current:
            proposed_values.update({
                "model_key": model_key,
                "target_type": target_type,
                "target_id": target_id,
                "image_alt": row.get("option_name", ""),
                "image_fit": NEW_ROW_FIT,
                "image_position": NEW_ROW_POSITION,
                "active": "True",
                "notes": NEW_ROW_NOTE,
            })
            if resolution.fields.get("hover_image_url"):
                proposed_values["hover_image_alt"] = row.get("option_name", "")
                proposed_values["hover_image_position"] = NEW_ROW_POSITION
        status = _manager_status(row.get("action", ""))
        items.append({
            "id": _stable_item_id("target", model_key, target_type, target_id),
            "kind": "target",
            "status": status,
            "action": row.get("action", ""),
            "model_key": model_key,
            "section_id": row.get("section_id", "") or "(none)",
            "target_type": target_type,
            "target_id": target_id,
            "rpo": row.get("rpo", ""),
            "label": row.get("option_name", "") or target_id,
            "expected": row.get("coverage_intent", "") != COVERAGE_NOT_EXPECTED,
            "coverage_intent": row.get("coverage_intent", "") or "unclassified",
            "coverage_intent_reason": row.get("coverage_intent_reason", ""),
            "coverage": {
                "kind": coverage_kind,
                "covered": bool(current_values.get("image_url")),
                "asset_model_key": current_values.get("model_key", ""),
            },
            "workbook_target": {
                "model_key": model_key,
                "target_type": target_type,
                "target_id": target_id,
            },
            "lineage": {
                "target_source_sheet": row.get("source_sheet", ""),
                "asset_source_sheet": current.get("source_sheet", ""),
                "asset_source_row": current.get("source_row"),
            },
            "current_values": current_values,
            "proposed_values": proposed_values,
            "candidate": {
                "selected": [
                    {"field": candidate_field, "url": url}
                    for candidate_field, url in sorted(resolution.fields.items())
                ],
                "source": resolution.source,
                "priority": resolution.priority,
                "reason": resolution.reason or row.get("note", ""),
                "alternatives": [
                    {
                        "field": alternative.field,
                        "url": alternative.url,
                        "source": alternative.source,
                        "priority": alternative.priority,
                    }
                    for alternative in resolution.alternatives
                ],
            },
            "supports_hover": (
                target_type == TARGET_TYPE_CONTEXT_CHOICE
                and target_id.startswith("body_style__")
            ),
        })

    for url in unmatched:
        shared_prefix, rpo, shared_ok = parse_shared_option_media(url)
        if shared_ok:
            parsed_model = shared_prefix or ""
        else:
            parsed_model, rpo, _ = parse_media(url)
        items.append({
            "id": _stable_item_id("unmatched", url),
            "kind": "media",
            "status": "unmatched",
            "action": "unmatched_media",
            "model_key": parsed_model or "",
            "section_id": "(media inventory)",
            "target_type": "media",
            "target_id": "",
            "rpo": rpo,
            "label": filename_stem(url),
            "expected": False,
            "coverage_intent": "unclassified",
            "coverage_intent_reason": "media has no desired workbook target",
            "coverage": {"kind": "unmatched", "covered": False, "asset_model_key": ""},
            "workbook_target": {},
            "lineage": {"target_source_sheet": "", "asset_source_sheet": "", "asset_source_row": None},
            "current_values": {},
            "proposed_values": {"image_url": url},
            "candidate": {
                "selected": [], "source": "unmatched-media", "priority": None,
                "reason": "media filename parsed but did not match a desired workbook target",
                "alternatives": [{"field": "image_url", "url": url, "source": "inventory", "priority": None}],
            },
            "supports_hover": False,
        })
    for url in unparseable:
        items.append({
            "id": _stable_item_id("unparseable", url),
            "kind": "media",
            "status": "unparseable",
            "action": "unparseable_media",
            "model_key": "",
            "section_id": "(media inventory)",
            "target_type": "media",
            "target_id": "",
            "rpo": "",
            "label": filename_stem(url),
            "expected": False,
            "coverage_intent": "unclassified",
            "coverage_intent_reason": "filename did not yield a supported target identity",
            "coverage": {"kind": "unparseable", "covered": False, "asset_model_key": ""},
            "workbook_target": {},
            "lineage": {"target_source_sheet": "", "asset_source_sheet": "", "asset_source_row": None},
            "current_values": {},
            "proposed_values": {"image_url": url},
            "candidate": {
                "selected": [], "source": "unparseable-media", "priority": None,
                "reason": "filename did not yield a supported model, RPO, model card, or body-style identity",
                "alternatives": [{"field": "image_url", "url": url, "source": "inventory", "priority": None}],
            },
            "supports_hover": False,
        })

    add_wildcard_ownership_resolution(items)
    items.sort(key=lambda item: (
        ASSET_MANAGER_STATUSES.index(item["status"])
        if item["status"] in ASSET_MANAGER_STATUSES else len(ASSET_MANAGER_STATUSES),
        item.get("model_key", ""), item.get("section_id", ""),
        item.get("rpo", ""), item.get("target_id", ""), item["id"],
    ))
    workbook_sha256 = _sha256_path(workbook_path)
    media_sha256 = hashlib.sha256(
        json.dumps(normalized_urls, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    ).hexdigest()
    reconciliation_sha256 = hashlib.sha256(
        f"{workbook_sha256}:{media_sha256}:{COVERAGE_RULESET_VERSION}".encode("utf-8")
    ).hexdigest()
    return AssetManagerSnapshot(
        fingerprints={
            "workbook_sha256": workbook_sha256,
            "workbook_mtime_ns": str(workbook_path.stat().st_mtime_ns),
            "media_inventory_sha256": media_sha256,
            "reconciliation_sha256": reconciliation_sha256,
        },
        media={
            "source": media_source,
            "url_count": len(normalized_urls),
            "unmatched_count": len(unmatched),
            "unparseable_count": len(unparseable),
            "existing_url_verification": "enabled" if verify_existing else "not_run",
        },
        coverage_ruleset={
            "version": COVERAGE_RULESET_VERSION,
            "rules": list(COVERAGE_RULESET),
        },
        items=tuple(items),
        action_counts=dict(Counter(item["action"] for item in items)),
        media_urls=tuple(normalized_urls),
    )


def filter_asset_manager_snapshot(
    snapshot: AssetManagerSnapshot,
    *,
    model_key: str = "",
    section_id: str = "",
    target_type: str = "",
    coverage_intent: str = "",
    status: str = "",
    offset: int = 0,
    limit: int = 24,
    ignored_item_ids: set[str] | None = None,
) -> dict[str, Any]:
    """Filter/paginate while keeping every count and percentage domain-owned."""

    def matches(item: dict[str, Any], *, include_status: bool = True) -> bool:
        return (
            (not model_key or item.get("model_key") == model_key)
            and (not section_id or item.get("section_id") == section_id)
            and (not target_type or item.get("target_type") == target_type)
            and (not coverage_intent or item.get("coverage_intent") == coverage_intent)
            and (not include_status or not status or item.get("status") == status)
        )

    ignored_item_ids = ignored_item_ids or set()
    effective_items = [
        {**item, "status": "ignored"} if item["id"] in ignored_item_ids else item
        for item in snapshot.items
    ]
    scoped = [item for item in effective_items if matches(item, include_status=False)]
    filtered = [item for item in scoped if not status or item.get("status") == status]
    target_items = [item for item in scoped if item.get("kind") == "target"]

    def coverage_bucket(rows: list[dict[str, Any]]) -> dict[str, Any]:
        total = len(rows)
        covered = sum(1 for item in rows if item.get("coverage", {}).get("covered"))
        return {
            "total_targets": total,
            "covered": covered,
            "missing": total - covered,
            "coverage_pct": round(100.0 * covered / total, 1) if total else 0.0,
        }

    models: list[dict[str, Any]] = []
    for current_model in sorted({item.get("model_key", "") for item in target_items if item.get("model_key")}):
        model_items = [item for item in target_items if item.get("model_key") == current_model]
        sections = []
        for current_section in sorted({item.get("section_id", "") for item in model_items}):
            section_items = [item for item in model_items if item.get("section_id") == current_section]
            sections.append({"section_id": current_section, **coverage_bucket(section_items)})
        models.append({"model_key": current_model, **coverage_bucket(model_items), "sections": sections})

    status_counter = Counter(item.get("status", "") for item in scoped)
    status_counts = {name: status_counter.get(name, 0) for name in ASSET_MANAGER_STATUSES}
    facets = {
        "models": sorted({item.get("model_key", "") for item in effective_items if item.get("model_key")}),
        "sections": sorted({item.get("section_id", "") for item in scoped if item.get("section_id")}),
        "target_types": sorted({item.get("target_type", "") for item in scoped if item.get("target_type")}),
        "coverage_intents": sorted({item.get("coverage_intent", "") for item in scoped if item.get("coverage_intent")}),
        "statuses": list(ASSET_MANAGER_STATUSES),
    }
    return {
        "mode": "read_only_asset_resolution",
        "fingerprints": snapshot.fingerprints,
        "media": snapshot.media,
        "coverage_ruleset": snapshot.coverage_ruleset,
        "coverage": {"overall": coverage_bucket(target_items), "models": models},
        "status_counts": status_counts,
        "action_counts": snapshot.action_counts,
        "facets": facets,
        "filters": {
            "model_key": model_key,
            "section_id": section_id,
            "target_type": target_type,
            "coverage_intent": coverage_intent,
            "status": status,
        },
        "queue": {
            "total": len(filtered),
            "offset": offset,
            "limit": limit,
            "items": filtered[offset:offset + limit],
        },
        "assignment_targets": [
            {
                "item_id": item["id"],
                "model_key": item.get("model_key", ""),
                "section_id": item.get("section_id", ""),
                "target_type": item.get("target_type", ""),
                "target_id": item.get("target_id", ""),
                "rpo": item.get("rpo", ""),
                "label": item.get("label", ""),
            }
            for item in snapshot.items
            if item.get("kind") == "target" and item.get("model_key") != WILDCARD_MODEL_KEY
        ],
    }


def search_asset_manager_media(
    snapshot: AssetManagerSnapshot,
    query: str = "",
    *,
    offset: int = 0,
    limit: int = 50,
) -> dict[str, Any]:
    """Return a bounded inventory selector bound to the snapshot fingerprints."""

    needle = query.strip().lower()
    matches = [
        url for url in snapshot.media_urls
        if not needle or needle in url.lower() or needle in filename_stem(url)
    ]
    bounded_limit = max(1, min(limit, 100))
    bounded_offset = max(0, offset)
    return {
        "fingerprints": snapshot.fingerprints,
        "query": query,
        "total": len(matches),
        "offset": bounded_offset,
        "limit": bounded_limit,
        "items": [
            {"url": url, "label": filename_stem(url)}
            for url in matches[bounded_offset:bounded_offset + bounded_limit]
        ],
    }


def search_asset_manager_targets(
    snapshot: AssetManagerSnapshot,
    query: str = "",
    *,
    model_key: str = "",
    offset: int = 0,
    limit: int = 25,
) -> dict[str, Any]:
    """Return bounded human-labeled workbook targets from one bound snapshot."""

    needle = query.strip().lower()
    matches = []
    for item in snapshot.items:
        if item.get("kind") != "target" or item.get("model_key") == WILDCARD_MODEL_KEY:
            continue
        if model_key and item.get("model_key") != model_key:
            continue
        searchable = " ".join(str(item.get(field, "")) for field in (
            "label", "rpo", "target_id", "target_type", "section_id", "model_key"
        )).lower()
        if needle and needle not in searchable:
            continue
        matches.append({
            "item_id": item["id"],
            "label": item.get("label", "") or item.get("target_id", ""),
            "canonical_id": item.get("target_id", ""),
            "model_key": item.get("model_key", ""),
            "section_id": item.get("section_id", ""),
            "target_type": item.get("target_type", ""),
            "target_id": item.get("target_id", ""),
            "rpo": item.get("rpo", ""),
        })
    matches.sort(key=lambda item: (
        item["label"].lower(), item["canonical_id"], item["model_key"], item["item_id"]
    ))
    bounded_limit = max(1, min(limit, 100))
    bounded_offset = max(0, offset)
    return {
        "fingerprints": snapshot.fingerprints,
        "query": query,
        "model_key": model_key,
        "total": len(matches),
        "offset": bounded_offset,
        "limit": bounded_limit,
        "items": matches[bounded_offset:bounded_offset + bounded_limit],
    }


def apply_sync_plan(wb, *, asset_sheet: str, plan: SyncPlan) -> None:
    if asset_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {asset_sheet!r} not found.")
    ws = wb[asset_sheet]
    headers = [clean(cell.value) for cell in ws[1]]
    header_index = _ensure_asset_headers(headers)

    for (row_number, field), url in plan.url_writes.items():
        if field in header_index:
            ws.cell(row_number, header_index[field] + 1).value = url

    for insert in plan.inserts:
        row_values: list[Any] = [""] * len(headers)
        row_values[header_index["model_key"]] = insert["model"]
        row_values[header_index["target_type"]] = insert.get("target_type", TARGET_TYPE_OPTION)
        row_values[header_index["target_id"]] = insert["tid"]
        for field, value in insert.get("fields", {"image_url": insert.get("url", "")}).items():
            if field in header_index:
                row_values[header_index[field]] = value
        if "image_alt" in header_index:
            row_values[header_index["image_alt"]] = insert["name"]
        if "hover_image_alt" in header_index and insert.get("fields", {}).get("hover_image_url"):
            row_values[header_index["hover_image_alt"]] = insert["name"]
        if "image_fit" in header_index:
            row_values[header_index["image_fit"]] = NEW_ROW_FIT
        if "image_position" in header_index:
            row_values[header_index["image_position"]] = NEW_ROW_POSITION
        if "hover_image_position" in header_index and insert.get("fields", {}).get("hover_image_url"):
            row_values[header_index["hover_image_position"]] = NEW_ROW_POSITION
        if "active" in header_index:
            row_values[header_index["active"]] = True
        if "notes" in header_index:
            row_values[header_index["notes"]] = NEW_ROW_NOTE
        ws.append(row_values)


def run_sync(
    *,
    workbook_path: Path | str,
    report_dir: Path | str,
    media_urls: Iterable[str],
    apply: bool = False,
    asset_sheet: str = ASSET_SHEET,
    verify_existing: bool = False,
    timeout: float = 10.0,
    workers: int = 16,
    incremental: bool = False,
    media_source: str = "live",
    since_argument: str | None = None,
    resolved_modified_after: str | None = None,
    state_read: bool = False,
    state_media_modified: dict[str, str] | None = None,
    state_revision_tokens: dict[str, str] | None = None,
    update_safe_wildcards: bool = False,
    write_apply_state: bool = True,
    save_fn: Callable[..., Path] = save_workbook_safely,
) -> SyncResult:
    workbook_path = Path(workbook_path)
    report_dir = Path(report_dir)
    media_url_list = list(media_urls)
    loaded_mtime_ns = workbook_path.stat().st_mtime_ns if workbook_path.exists() else None
    backup_path: Path | None = None
    state_written = False

    plan_wb = load_workbook(workbook_path, read_only=not apply, data_only=not apply)
    try:
        plan, sources, unmatched, unparseable = build_sync_plan(
            plan_wb,
            asset_sheet=asset_sheet,
            media_urls=media_url_list,
            verify_existing=verify_existing,
            timeout=timeout,
            workers=workers,
            incremental=incremental,
            update_safe_wildcards=update_safe_wildcards,
        )
    finally:
        plan_wb.close()

    report_path, missing_path, unmatched_path, missing_count, broad_missing_count = _write_reports(
        report_dir,
        plan.report,
        unmatched,
        unparseable,
        incremental,
    )
    action_counts = dict(Counter(row["action"] for row in plan.report))
    coverage_summary = build_coverage_summary(plan.report)
    coverage_summary["section_coverage"] = plan.section_coverage

    has_workbook_changes = bool(plan.url_writes or plan.inserts)
    if apply and has_workbook_changes:
        apply_wb = load_workbook(workbook_path)
        try:
            apply_sync_plan(apply_wb, asset_sheet=asset_sheet, plan=plan)
            backup_path = save_fn(apply_wb, workbook_path, loaded_mtime_ns=loaded_mtime_ns)
        finally:
            apply_wb.close()
        check_wb = load_workbook(workbook_path, read_only=True, data_only=True)
        check_wb.close()
    if apply and since_argument == "auto" and write_apply_state:
        write_state(
            report_dir,
            media_modified=state_media_modified,
            revision_tokens=state_revision_tokens,
        )
        state_written = True

    manifest_path = _write_manifest(
        report_dir=report_dir,
        workbook_path=workbook_path,
        asset_sheet=asset_sheet,
        apply=apply,
        media_source=media_source,
        since_argument=since_argument,
        resolved_modified_after=resolved_modified_after,
        incremental=incremental,
        state_read=state_read,
        state_written=state_written,
        media_url_count=len(media_url_list),
        verify_existing=verify_existing,
        included_sources=sources,
        action_counts=action_counts,
        url_write_count=len(plan.url_writes),
        insert_count=len(plan.inserts),
        unmatched_count=len(unmatched),
        unparseable_count=len(unparseable),
        report_path=report_path,
        missing_path=missing_path,
        missing_count=missing_count,
        broad_missing_count=broad_missing_count,
        coverage_summary=coverage_summary,
        unmatched_path=unmatched_path,
    )

    return SyncResult(
        report_path=report_path,
        missing_path=missing_path,
        unmatched_path=unmatched_path,
        manifest_path=manifest_path,
        url_write_count=len(plan.url_writes),
        insert_count=len(plan.inserts),
        action_counts=action_counts,
        unmatched_count=len(unmatched),
        unparseable_count=len(unparseable),
        backup_path=backup_path,
    )


def _affected_models_from_report(report_path: Path) -> list[str]:
    write_actions = {"fill", "insert_filled", "replace_canonical", "replace_shared_canonical"}
    with report_path.open(encoding="utf-8") as handle:
        return sorted(
            {
                clean(row.get("model_key")).lower()
                for row in csv.DictReader(handle)
                if row.get("action") in write_actions and clean(row.get("model_key")) not in {"", WILDCARD_MODEL_KEY}
            }
        )


def _snapshot_files(paths: Iterable[Path]) -> dict[Path, bytes | None]:
    return {path: path.read_bytes() if path.exists() else None for path in paths}


def _restore_files(snapshot: dict[Path, bytes | None]) -> None:
    for path, content in snapshot.items():
        if content is None:
            path.unlink(missing_ok=True)
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(content)


def _bump_data_cache_version(index_path: Path) -> None:
    text = index_path.read_text(encoding="utf-8")
    pattern = re.compile(r'(src="\./data\.js\?v=)(\d+)(")')
    match = pattern.search(text)
    if not match:
        raise RuntimeError(f"Could not find data.js cache version in {index_path}")
    replacement = f"{match.group(1)}{int(match.group(2)) + 1}{match.group(3)}"
    write_text_atomic(index_path, text[: match.start()] + replacement + text[match.end() :])


def _mark_manifest_state_written(manifest_path: Path) -> None:
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    payload["state_written"] = True
    manifest_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def run_complete_pipeline(result: SyncResult, *, workbook_path: Path) -> list[str]:
    """Validate, regenerate, publish, and cache-bust one complete sync atomically enough to roll back."""

    affected_models = _affected_models_from_report(result.report_path)
    if not affected_models or not (result.url_write_count or result.insert_count):
        return []
    tracked_paths = [
        *(ROOT / "form-output" / "runtime" / f"{model.replace('_', '-')}-runtime-contract.json" for model in affected_models),
        *(ROOT / "form-output" / "inspection" / f"{model.replace('_', '-')}-derived-swap-manifest.json" for model in affected_models),
        ROOT / "form-app" / "data.js",
        ROOT / "form-app" / "index.html",
    ]
    snapshot = _snapshot_files(tracked_paths)
    commands = [
        ("workbook package", [sys.executable, "scripts/validate_workbook_package.py", str(workbook_path)]),
        ("workbook schema", [sys.executable, "scripts/validate_workbook_schema.py", str(workbook_path)]),
        *(
            (
                f"generate {model}",
                [sys.executable, "scripts/generate_form.py", "--model", model, "--workbook", str(workbook_path)],
            )
            for model in affected_models
        ),
        ("publish registry", [sys.executable, "scripts/generate_registry.py", "--workbook", str(workbook_path)]),
    ]
    try:
        for label, command in commands:
            print(f"  {label} ...", end="", flush=True)
            completed = subprocess.run(  # noqa: S603 - fixed project commands
                command,
                cwd=ROOT,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                check=False,
            )
            if completed.returncode:
                print(" failed", flush=True)
                if completed.stdout:
                    print(completed.stdout, file=sys.stderr, flush=True)
                raise RuntimeError(f"Complete asset sync failed during {label}")
            print(" ok", flush=True)
        _bump_data_cache_version(ROOT / "form-app" / "index.html")
    except Exception:
        if result.backup_path is not None:
            restore_workbook_backup(workbook_path, result.backup_path)
        _restore_files(snapshot)
        raise
    return affected_models


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Report or safely apply asset_map image URL sync from current hosted media inventory.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, type=Path)
    parser.add_argument("--asset-sheet", default=ASSET_SHEET)
    parser.add_argument("--report-dir", default=DEFAULT_REPORT_DIR, type=Path)
    parser.add_argument("--media-url-list", type=Path, help="Deterministic newline-delimited media URL fixture/list")
    parser.add_argument("--apply", action="store_true", help="Write workbook changes through save_workbook_safely()")
    parser.add_argument(
        "--complete",
        action="store_true",
        help="Stable full live pull, apply every unambiguous change, validate, regenerate, and publish",
    )
    parser.add_argument("--timeout", type=float, default=10.0)
    parser.add_argument("--workers", type=int, default=16)
    parser.add_argument("--verify-existing-network", dest="verify_existing", action="store_true", help="Optionally probe existing workbook URLs for dead-link reporting")
    parser.add_argument("--no-verify-existing", dest="verify_existing", action="store_false", help="Deprecated no-op; network verification is off by default")
    parser.add_argument("--since", default=None, metavar="DATE", help="Media modified-after date, or 'auto' for saved cursor")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.complete and args.media_url_list:
        parser.error("--complete uses the live stable inventory and cannot be combined with --media-url-list")
    if args.complete and args.since:
        parser.error("--complete always uses a full stable inventory and cannot be combined with --since")
    if args.complete and args.workbook.resolve() != DEFAULT_WORKBOOK.resolve():
        parser.error("--complete is restricted to the canonical stingray_master.xlsx workflow")

    prior_state, prior_state_exists = read_state(args.report_dir)
    state_read = False
    media_modified: dict[str, str] | None = None
    revision_tokens: dict[str, str] | None = None
    if args.complete:
        modified_after = None
        state_read = prior_state_exists
    elif args.since == "auto":
        modified_after, state_read = read_since_auto(args.report_dir)
    else:
        modified_after = args.since
    if args.media_url_list:
        media_urls = read_media_url_list(args.media_url_list)
        media_source = "media-url-list"
        print(f"Loaded {len(media_urls)} media URLs from {args.media_url_list}")
    else:
        media_source = "live"
        label = "stable full" if args.complete else (f"incremental after {modified_after}" if modified_after else "full")
        print(f"Pulling media [{label}] ...", flush=True)
        try:
            if args.complete:
                snapshot = fetch_media_stable(args.timeout)
                media_urls, revision_tokens = prepare_revisioned_media_urls(snapshot, prior_state)
                media_modified = snapshot.modified_by_url
            else:
                media_urls = fetch_media(args.timeout, modified_after)
        except WordPressMediaFetchError as exc:
            print(str(exc), file=sys.stderr)
            return 1
        print(f"  {len(media_urls)} images under {PATH_FILTER}", flush=True)

    result = run_sync(
        workbook_path=args.workbook,
        report_dir=args.report_dir,
        media_urls=media_urls,
        apply=args.apply or args.complete,
        asset_sheet=args.asset_sheet,
        verify_existing=args.verify_existing,
        timeout=args.timeout,
        workers=args.workers,
        incremental=modified_after is not None,
        media_source=media_source,
        since_argument="auto" if args.complete else args.since,
        resolved_modified_after=modified_after,
        state_read=state_read,
        state_media_modified=media_modified,
        state_revision_tokens=revision_tokens,
        update_safe_wildcards=args.complete,
        write_apply_state=not args.complete,
    )

    print("\n=== Summary ===")
    for action, count in sorted(result.action_counts.items()):
        print(f"  {action:<30} {count}")
    print(f"  {'unmatched media':<30} {result.unmatched_count}")
    print(f"  {'unparseable files':<30} {result.unparseable_count}")
    print(
        f"\nReports: {result.report_path}\n"
        f"         {result.missing_path}\n"
        f"         {result.unmatched_path}\n"
        f"Manifest: {result.manifest_path}"
    )

    if args.apply or args.complete:
        print(
            f"\nAPPLIED: {result.url_write_count} url change(s), {result.insert_count} row insert(s).",
            flush=True,
        )
        if result.backup_path:
            print(f"Backup -> {result.backup_path}")
        if args.complete:
            affected_models = run_complete_pipeline(result, workbook_path=args.workbook)
            write_state(
                args.report_dir,
                media_modified=media_modified,
                revision_tokens=revision_tokens,
            )
            _mark_manifest_state_written(result.manifest_path)
            if affected_models:
                print("COMPLETE: validated, regenerated, and published " + ", ".join(affected_models))
            else:
                print("COMPLETE: no unambiguous workbook changes were needed.")
    else:
        print(f"\nDRY RUN -- would write {result.url_write_count} url change(s) and {result.insert_count} new row(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
