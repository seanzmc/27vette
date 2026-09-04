#!/usr/bin/env python3
"""Workbook-editor engine: sheet metadata, op model, validation, apply.

The workbook owns its data; this module owns only what the workbook cannot
express about itself: which columns form each sheet family's primary key,
intended cell types, enum domains, and which columns reference other
workbook entities — plus the Phase 2 op pipeline (flatten -> coalesce ->
coerce/validate -> dry-run -> apply through ``save_workbook_safely``).
See workbook-editor-phase2-spec.md.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from corvette_form_generator.schema_validation import result_payload, validate_workbook_schema
from corvette_form_generator.workbook import (
    excel_lock_path,
    remove_table_sheet_auto_filters,
    restore_workbook_backup,
    save_workbook_safely,
    workbook_truthy,
)
from corvette_form_generator.workbook_bool_hygiene import (
    BOOL_TEXT_VALUES,
    compare_bool_like_workbooks,
    result_payload as bool_hygiene_result_payload,
)
# Compatibility aliases re-exported from workbook_domain.registry; the
# canonical registry literals now live in that module.
from corvette_form_generator.workbook_domain.registry import (
    ACTIVE_ROW_REQUIRED_COLUMNS,
    EDITOR_SHEET_META,
    GLOBAL_SHEET_FAMILIES,
    SOURCE_ROLE_FAMILIES,
    optional_key_columns,
)
from corvette_form_generator.workbook_package import assert_valid_workbook_package

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_LOG_PATH = ROOT / "form-output" / "workbook-edit-log.jsonl"


# ─────────────────────────────────────────────────────────────
# Workbook extraction (shared with the server)
# ─────────────────────────────────────────────────────────────

def jsonable(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def extract_workbook(path: Path) -> dict:
    """Load the whole workbook into plain dicts and close the file."""
    path = Path(path)
    mtime_ns = path.stat().st_mtime_ns
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, dict] = {}
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None) or ()
        cols = [(i, str(v)) for i, v in enumerate(header_row) if v is not None]
        rows = []
        for raw in rows_iter:
            row = {
                name: jsonable(raw[i]) if i < len(raw) else None
                for i, name in cols
            }
            if all(v in (None, "") for v in row.values()):
                continue
            rows.append(row)
        sheets[ws.title] = {"headers": [name for _, name in cols], "rows": rows}
    wb.close()
    return {"path": str(path), "mtime_ns": mtime_ns, "sheets": sheets}


def rows_of(extract: dict, name: str) -> list[dict]:
    sheet = extract["sheets"].get(name)
    return sheet["rows"] if sheet else []


def model_sheet_registry(extract: dict) -> tuple[dict, dict]:
    """Per-model sheet registry plus a sheet-name -> family reverse map."""
    registry: dict[str, list[dict]] = {}
    sheet_family: dict[str, str] = {}
    for row in rows_of(extract, "model_workbook_sources"):
        if not workbook_truthy(row.get("active")):
            continue
        family = SOURCE_ROLE_FAMILIES.get(row.get("source_role"))
        sheet_name = row.get("sheet_name")
        model_key = row.get("model_key")
        if not (family and sheet_name and model_key):
            continue
        registry.setdefault(model_key, []).append({
            "sheet": sheet_name,
            "role": row.get("source_role"),
            "family": family,
        })
        sheet_family.setdefault(sheet_name, family)
    return registry, sheet_family


# ─────────────────────────────────────────────────────────────
# Op primitives: flatten, coalesce, coerce
# ─────────────────────────────────────────────────────────────

def _raw_effect(operation: dict, raw_index: int) -> dict:
    return {
        "index": raw_index,
        "action": operation.get("action"),
        "fields": sorted(
            str(column)
            for column in (operation.get("row") or {})
            if not str(column).startswith("_")
        ),
    }


def flatten_items(items) -> list[dict]:
    ops: list[dict] = []
    raw_index = 0
    for item in items or []:
        if isinstance(item, dict) and item.get("kind") == "composite":
            label = item.get("label") or item.get("compositeType") or "composite"
            for member in item.get("ops", []):
                member = dict(member)
                member["_composite"] = label
                member["_raw_indices"] = [raw_index]
                member["_raw_effects"] = [_raw_effect(member, raw_index)]
                ops.append(member)
                raw_index += 1
        else:
            operation = dict(item)
            operation["_raw_indices"] = [raw_index]
            operation["_raw_effects"] = [_raw_effect(operation, raw_index)]
            ops.append(operation)
            raw_index += 1
    return ops


def _key_id(op: dict) -> tuple:
    return (op.get("sheet"), tuple(sorted(
        (str(k), str(v).strip()) for k, v in (op.get("key") or {}).items())))


def coalesce_ops(ops: list[dict]) -> list[dict]:
    result: list[dict | None] = []
    last_live: dict[tuple, int] = {}
    for raw_index, original in enumerate(ops):
        op = dict(original)
        op["_raw_indices"] = list(op.get("_raw_indices") or [raw_index])
        op["_raw_effects"] = list(
            op.get("_raw_effects") or [_raw_effect(op, op["_raw_indices"][0])]
        )
        op["_coverage_errors"] = list(op.get("_coverage_errors") or [])
        kid = _key_id(op)
        pos = last_live.get(kid)
        prev = result[pos] if pos is not None else None
        action = op.get("action")
        if prev is None:
            last_live[kid] = len(result)
            result.append(dict(op))
            continue
        if prev.get("action") == "delete":
            if action == "add":
                last_live[kid] = len(result)
                result.append(dict(op))
                continue
            prev["_raw_indices"].extend(op["_raw_indices"])
            prev["_raw_effects"].extend(op["_raw_effects"])
            prev["_coverage_errors"].extend(op["_coverage_errors"])
            prev["_coverage_errors"].append(
                f"dropped raw effect for {_key_id(op)!r}: delete followed by {action}"
            )
            continue
        prev_action = prev.get("action")
        if action == "update" and prev_action in ("add", "update"):
            overlapping = set(prev.get("row") or {}) & set(op.get("row") or {})
            contradictions = sorted(
                column
                for column in overlapping
                if (prev.get("row") or {}).get(column) != (op.get("row") or {}).get(column)
            )
            if contradictions:
                prev["_coverage_errors"].append(
                    f"contradictory raw effects for {_key_id(op)!r}: "
                    f"later {action} replaces {contradictions}"
                )
            prev["row"] = {**(prev.get("row") or {}), **(op.get("row") or {})}
            prev["_raw_indices"].extend(op["_raw_indices"])
            prev["_raw_effects"].extend(op["_raw_effects"])
            prev["_coverage_errors"].extend(op["_coverage_errors"])
        elif action == "delete" and prev_action == "add":
            prev["_raw_indices"].extend(op["_raw_indices"])
            prev["_raw_effects"].extend(op["_raw_effects"])
            prev["_coverage_errors"].extend(op["_coverage_errors"])
            prev["_coverage_errors"].append(
                f"dropped raw effects for {_key_id(op)!r}: add followed by delete"
            )
            prev["_coverage_only"] = True
            last_live.pop(kid, None)
        elif action == "delete" and prev_action == "update":
            replacement = dict(op)
            replacement["_raw_indices"] = prev["_raw_indices"] + op["_raw_indices"]
            replacement["_raw_effects"] = prev["_raw_effects"] + op["_raw_effects"]
            replacement["_coverage_errors"] = (
                list(prev.get("_coverage_errors") or [])
                + list(op.get("_coverage_errors") or [])
                + [f"dropped raw effect for {_key_id(op)!r}: update followed by delete"]
            )
            result[pos] = replacement
        else:  # e.g. add after add/update — leave both; validation flags it
            last_live[kid] = len(result)
            result.append(op)
    return [op for op in result if op is not None]


def _bool_storage_for_sheet(extract: dict, sheet: str, family: str) -> dict[str, dict]:
    meta = EDITOR_SHEET_META[family]
    bool_columns = {column for column, kind in meta.get("types", {}).items() if kind == "bool"}
    result: dict[str, dict] = {}
    if not bool_columns:
        return result
    rows = extract["sheets"].get(sheet, {}).get("rows", [])
    for column in bool_columns:
        family_counts: dict[str, int] = {}
        true_text_counts: dict[str, int] = {}
        false_text_counts: dict[str, int] = {}
        for row in rows:
            value = row.get(column)
            if isinstance(value, bool):
                family_counts["excel_boolean"] = family_counts.get("excel_boolean", 0) + 1
                continue
            if isinstance(value, str) and value in BOOL_TEXT_VALUES:
                family_counts["text"] = family_counts.get("text", 0) + 1
                if value.lower() == "true":
                    true_text_counts[value] = true_text_counts.get(value, 0) + 1
                else:
                    false_text_counts[value] = false_text_counts.get(value, 0) + 1
        if len(family_counts) != 1:
            continue
        storage_family = next(iter(family_counts))
        entry = {"storageFamily": storage_family}
        if storage_family == "text":
            entry["trueText"] = max(true_text_counts, key=lambda key: true_text_counts[key]) if true_text_counts else "True"
            entry["falseText"] = max(false_text_counts, key=lambda key: false_text_counts[key]) if false_text_counts else "False"
        result[column] = entry
    return result


def _bool_storage_conventions(extract: dict, sheet_family: dict[str, str], created_templates: dict[str, str]) -> dict[tuple[str, str], dict]:
    conventions: dict[tuple[str, str], dict] = {}
    for sheet, family in sheet_family.items():
        source_sheet = created_templates.get(sheet, sheet)
        for column, convention in _bool_storage_for_sheet(extract, source_sheet, family).items():
            conventions[(sheet, column)] = convention
    return conventions


def _parse_bool_value(column: str, value) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, str) and value.strip().lower() in ("true", "false"):
        return value.strip().lower() == "true"
    raise ValueError(f"{column}: expected True/False, got {value!r}")


def coerce_value(family: str, column: str, value, *, bool_storage: dict | None = None):
    meta = EDITOR_SHEET_META[family]
    if value == "":
        value = None
    enums = meta.get("enums", {}).get(column)
    if enums is not None:
        text = "" if value is None else str(value).strip()
        if text not in enums:
            raise ValueError(f"{column}: {text!r} not in enum {sorted(enums)}")
        return text or None
    kind = meta.get("types", {}).get(column)
    if kind == "int":
        if value is None:
            return None
        if isinstance(value, bool):
            raise ValueError(f"{column}: expected integer, got boolean")
        if isinstance(value, int):
            return value
        if isinstance(value, float) and value.is_integer():
            return int(value)
        text = str(value).strip()
        if text.lstrip("-").isdigit():
            return int(text)
        raise ValueError(f"{column}: expected integer, got {value!r}")
    if kind == "bool":
        logical = _parse_bool_value(column, value)
        if logical is None:
            return None
        if bool_storage and bool_storage.get("storageFamily") == "text":
            return bool_storage.get("trueText", "True") if logical else bool_storage.get("falseText", "False")
        return logical
    if value is None:
        return None
    return str(value).strip() or None


# ─────────────────────────────────────────────────────────────
# Batch validation (non-breaking writes only)
# ─────────────────────────────────────────────────────────────

_REF_FAMILY = {"options": ("options", "option_id"), "rule_groups": ("rule_groups", "group_id"),
               "exclusive_groups": ("exclusive_groups", "group_id"),
               "interiors": ("interiors", "interior_id")}

_DORDER_GROUP_COL = {"options": "section_id", "rule_group_members": "group_id",
                     "exclusive_members": "group_id"}

# Derived reference domains whose consumer compares case-insensitively
# (contract.py:126-140 lowercases both sides for context copy lookups).
_CASE_INSENSITIVE_REF_KINDS = frozenset(("variant_trim_levels", "variant_body_styles"))


def price_ref_normalized_key(row: dict) -> tuple[str, str, str]:
    """The key PriceRef readers actually look up (pricing.py:17,33,43).

    OptionType drops non-alphanumerics and lowercases; Trim maps ``_`` to a
    space; Code is stripped. Two physical rows with one normalized key are a
    silent last-row-wins collision for the generator, so the writer refuses it.
    """
    from corvette_form_generator.pricing import price_ref_component_type_key

    return (
        price_ref_component_type_key(str(row.get("OptionType") or "")),
        str(row.get("Trim") or "").strip().replace("_", " "),
        str(row.get("Code") or "").strip(),
    )


def _registry_maps(extract):
    registry, sheet_family = model_sheet_registry(extract)
    models_by_sheet: dict[str, set] = {}
    by_model_family: dict[tuple, str] = {}
    for model_key, entries in registry.items():
        for entry in entries:
            models_by_sheet.setdefault(entry["sheet"], set()).add(model_key)
            by_model_family[(model_key, entry["family"])] = entry["sheet"]
    return registry, sheet_family, models_by_sheet, by_model_family


def _key_tuple(key, keycols):
    return tuple(str(key.get(k) or "").strip() for k in keycols)


def _sheet_key_index(extract, sheet, keycols, blank_keys=frozenset()):
    index = {}
    for row in rows_of(extract, sheet):
        kt = tuple(str(row.get(k) or "").strip() for k in keycols)
        if all(value or column in blank_keys for value, column in zip(kt, keycols)):
            index[kt] = row
    return index


def _ref_kinds(refkind) -> tuple[str, ...]:
    if isinstance(refkind, (tuple, list)):
        return tuple(str(kind) for kind in refkind)
    return (str(refkind),)


def _reference_models(maps, sheet: str, row: dict | None = None) -> set[str]:
    _registry, _sheet_family, models_by_sheet, _by_model_family = maps
    model_key = str((row or {}).get("model_key") or "").strip()
    if model_key == "*":
        return set(_registry)
    if model_key:
        return {model_key}
    return set(models_by_sheet.get(sheet, set()))


def _single_ref_domain(extract, maps, batch_adds, sheet, refkind, *, models=None):
    _registry, _sheet_family, models_by_sheet, by_model_family = maps
    models = set(models) if models is not None else set(models_by_sheet.get(sheet, set()))
    if refkind == "sections":
        return {str(r.get("section_id")).strip()
                for r in rows_of(extract, "section_master") if r.get("section_id")}
    if refkind == "variants":
        ids = {str(r.get("variant_id")).strip()
               for r in rows_of(extract, "model_variants")
               if r.get("model_key") in models and workbook_truthy(r.get("active"))}
        return ids or {str(r.get("variant_id")).strip()
                       for r in rows_of(extract, "variant_master") if r.get("variant_id")}
    if refkind in ("variant_trim_levels", "variant_body_styles"):
        # Derived domains over variant_master, compared case-insensitively the
        # way contract.py:126-135 compares context copy against variants. A
        # concrete context_choice_copy row resolves against only the trims and
        # body styles its affected models actually generate (active
        # model_variants memberships): the generator builds each model's
        # choices from config.variant_ids, so copy for a trim another model
        # supplies would never be emitted. Only a wildcard row (affected
        # models = every registry model) validates against the union. The
        # fallback to every variant_master value keeps the domain usable when
        # the affected models have no active membership to read.
        column = "trim_level" if refkind == "variant_trim_levels" else "body_style"
        master_rows = [
            r for r in rows_of(extract, "variant_master")
            if str(r.get(column) or "").strip()
        ]
        by_variant_id = {
            str(r.get("variant_id")).strip(): r for r in master_rows
            if str(r.get("variant_id") or "").strip()
        }
        memberships = [
            r for r in rows_of(extract, "model_variants")
            if str(r.get("model_key") or "").strip() in models
            and workbook_truthy(r.get("active"))
        ]
        scoped = set()
        for r in memberships:
            master = by_variant_id.get(str(r.get("variant_id") or "").strip())
            if master is not None:
                scoped.add(str(master.get(column)).strip().lower())
        return scoped or {
            str(r.get(column)).strip().lower() for r in master_rows
        }
    if refkind == "option_rpos":
        rpos = set()
        for model in models:
            src = by_model_family.get((model, "options"))
            if not src:
                continue
            rpos |= {str(r.get("rpo")).strip() for r in rows_of(extract, src) if r.get("rpo")}
            rpos |= {
                str((o.get("row") or {}).get("rpo")).strip()
                for o in batch_adds.get(src, [])
                if (o.get("row") or {}).get("rpo")
            }
        return rpos
    family, id_col = _REF_FAMILY[refkind]
    ids = set()
    for model in models:
        src = by_model_family.get((model, family))
        if not src:
            continue
        ids |= {str(r.get(id_col)).strip() for r in rows_of(extract, src) if r.get(id_col)}
        ids |= {str((o.get("row") or {}).get(id_col)).strip()
                for o in batch_adds.get(src, []) if (o.get("row") or {}).get(id_col)}
    return ids


def _ref_domain(extract, maps, batch_adds, sheet, refkind, *, row=None):
    models = _reference_models(maps, sheet, row)
    domain: set[str] = set()
    for kind in _ref_kinds(refkind):
        domain |= _single_ref_domain(
            extract,
            maps,
            batch_adds,
            sheet,
            kind,
            models=models,
        )
    return domain


def _ref_label(refkind) -> str:
    return " or ".join(_ref_kinds(refkind))


def _meta_ref_items(meta: dict):
    refs = dict(meta.get("refs", {}))
    refs.update(meta.get("ref_unions", {}))
    return refs.items()


def _conditional_ref_spec(meta: dict, row: dict) -> tuple[str, str | None] | None:
    spec = meta.get("conditional_ref")
    if not spec:
        return None
    discriminator = str(row.get(spec["discriminator"]) or "").strip().lower()
    domains = meta.get("conditional_refs", {})
    if discriminator not in domains:
        return None
    return spec["column"], domains[discriminator]


def _final_rows_by_sheet(extract: dict, prepared: list[dict]) -> dict[str, list[dict]]:
    final_rows = {
        sheet: [dict(row) for row in data.get("rows", [])]
        for sheet, data in extract["sheets"].items()
    }
    row_indexes: dict[str, dict[tuple[str, ...], dict]] = {}
    deleted_row_ids: dict[str, set[int]] = {}
    for operation in prepared:
        if operation["action"] == "create_sheet":
            final_rows.setdefault(operation["sheet"], [])
            continue
        sheet = operation["sheet"]
        family = operation["_family"]
        keycols = EDITOR_SHEET_META[family]["key"]
        key = operation["_kt"]
        rows = final_rows.setdefault(sheet, [])
        if sheet not in row_indexes:
            row_indexes[sheet] = {
                tuple(str(row.get(column) or "").strip() for column in keycols): row
                for row in rows
            }
        row_index = row_indexes[sheet]
        match = row_index.get(key)
        if operation["action"] == "delete":
            if match is not None:
                deleted_row_ids.setdefault(sheet, set()).add(id(match))
            continue
        if operation["action"] == "update":
            if match is not None:
                match.update(operation["_coerced_row"])
            continue
        row = dict(operation.get("key") or {})
        row.update(operation["_coerced_row"])
        rows.append(row)
        row_index[key] = row
    for sheet, deleted_ids in deleted_row_ids.items():
        final_rows[sheet] = [row for row in final_rows[sheet] if id(row) not in deleted_ids]
    return final_rows


def _reference_row_models(row: dict, sheet: str, maps) -> set[str]:
    _registry, _sheet_family, models_by_sheet, _by_model_family = maps
    if sheet in GLOBAL_SHEET_FAMILIES:
        model_key = str(row.get("model_key") or "").strip()
        return {model_key} if model_key else set()
    return set(models_by_sheet.get(sheet, set()))


def _build_reverse_reference_index(
    final_rows: dict[str, list[dict]],
    maps,
    sheet_family: dict[str, str],
) -> tuple[dict[tuple[str, str, str], set[str]], dict[str, set[str]]]:
    reverse_index: dict[tuple[str, str, str], set[str]] = {}
    for sheet, family in sheet_family.items():
        meta = EDITOR_SHEET_META.get(family)
        if not meta:
            continue
        for row in final_rows.get(sheet, []):
            models = _reference_row_models(row, sheet, maps)
            if not models:
                continue
            for column, refkind in _meta_ref_items(meta):
                value = str(row.get(column) or "").strip()
                if not value:
                    continue
                for model in models:
                    for kind in _ref_kinds(refkind):
                        reverse_index.setdefault((model, kind, value), set()).add(f"{sheet}.{column}")
            conditional = _conditional_ref_spec(meta, row)
            if conditional is None:
                continue
            column, refkind = conditional
            value = str(row.get(column) or "").strip()
            if refkind is None or not value:
                continue
            for model in models:
                reverse_index.setdefault((model, refkind, value), set()).add(f"{sheet}.{column}")

    _registry, _sheet_family, _models_by_sheet, by_model_family = maps
    option_rpos: dict[str, set[str]] = {}
    option_models = {model for model, family in by_model_family if family == "options"}
    for model in option_models:
        option_sheet = by_model_family.get((model, "options"))
        if not option_sheet:
            continue
        option_rpos[model] = {
            str(row.get("rpo") or "").strip()
            for row in final_rows.get(option_sheet, [])
            if str(row.get("rpo") or "").strip()
        }
    return reverse_index, option_rpos


def _incoming_references(
    reverse_index: dict[tuple[str, str, str], set[str]],
    final_option_rpos: dict[str, set[str]],
    maps,
    *,
    deleted_sheet: str,
    deleted_family: str,
    target_id: str,
    target_rpo: str = "",
) -> list[str]:
    _registry, _registered_families, models_by_sheet, _by_model_family = maps
    target_models = set(models_by_sheet.get(deleted_sheet, set()))
    references: set[str] = set()
    for model in target_models:
        references.update(reverse_index.get((model, deleted_family, target_id), set()))
        references.update(reverse_index.get(("*", deleted_family, target_id), set()))
        if deleted_family == "options" and target_rpo and target_rpo not in final_option_rpos.get(model, set()):
            references.update(reverse_index.get((model, "option_rpos", target_rpo), set()))
            references.update(reverse_index.get(("*", "option_rpos", target_rpo), set()))
    return sorted(references)


def reference_graph_summary(extract: dict) -> dict:
    """Build the canonical reverse-reference graph without preparing writes."""

    maps = _registry_maps(extract)
    _registry, sheet_family, _models_by_sheet, _by_model_family = maps
    sheet_family = {**GLOBAL_SHEET_FAMILIES, **sheet_family}
    final_rows = {
        sheet: [dict(row) for row in (payload.get("rows") or [])]
        for sheet, payload in (extract.get("sheets") or {}).items()
    }
    reverse_index, option_rpos = _build_reverse_reference_index(
        final_rows,
        maps,
        sheet_family,
    )
    return {
        "graphBuilt": True,
        "referenceKeys": len(reverse_index),
        "referenceEdges": sum(len(locations) for locations in reverse_index.values()),
        "models": sorted(option_rpos),
    }


def _price_ref_lints(extract, prepared, errors):
    """Fail closed on PriceRef rows the readers would silently mis-price.

    Seat rows need a Trim (pricing.py:27 drops blank-trim seat rows), and the
    final sheet may not hold two rows with one normalized key (dict lookups keep
    the last row). Both are checked over the batch-final row set so an update
    cannot sneak a collision past the literal-key duplicate check.
    """
    price_ops = [o for o in prepared if o.get("_family") == "price_ref"]
    if not price_ops:
        return
    sheet = price_ops[0]["sheet"]
    final_rows = _final_rows_by_sheet(extract, prepared).get(sheet, [])
    seen: dict[tuple[str, str, str], int] = {}
    for row in final_rows:
        normalized = price_ref_normalized_key(row)
        if normalized[0] == "seat" and not normalized[1]:
            errors.append(
                f"{sheet}: Seat rows require Trim (blank Trim is only a component fallback key): "
                f"{row.get('Code')!r}"
            )
        seen[normalized] = seen.get(normalized, 0) + 1
    for normalized, count in sorted(seen.items()):
        if count > 1:
            errors.append(
                f"{sheet}: {count} rows share normalized PriceRef key {normalized!r}; "
                "readers keep only the last row"
            )


def _prepare_batch(extract, batch):
    errors: list[str] = []
    warnings: list[dict] = []
    flat = flatten_items(batch.get("items") or [])
    creates = [o for o in flat if o.get("action") == "create_sheet"]
    ops = coalesce_ops([o for o in flat if o.get("action") != "create_sheet"])
    coalescing_errors = [
        error
        for operation in ops
        for error in operation.get("_coverage_errors", [])
    ]
    if coalescing_errors:
        return coalescing_errors, warnings, []
    maps = _registry_maps(extract)
    _registry, sheet_family, models_by_sheet, by_model_family = maps
    sheet_family = {**GLOBAL_SHEET_FAMILIES, **sheet_family}

    # Editing may target an existing scaffold sheet that is deliberately
    # inactive for runtime discovery. Keep this edit-time registration local
    # so model_sheet_registry() remains the active-only discovery boundary.
    for row in rows_of(extract, "model_workbook_sources"):
        model_key = str(row.get("model_key") or "").strip()
        sheet_name = str(row.get("sheet_name") or "").strip()
        family = SOURCE_ROLE_FAMILIES.get(row.get("source_role"))
        if not (model_key and sheet_name in extract["sheets"] and family):
            continue
        sheet_family.setdefault(sheet_name, family)
        models_by_sheet.setdefault(sheet_name, set()).add(model_key)
        by_model_family.setdefault((model_key, family), sheet_name)

    prepared_creates: list[dict] = []
    created_templates: dict[str, str] = {}
    for i, o in enumerate(creates):
        sheet = str(o.get("sheet") or "").strip()
        family = str(o.get("family") or "").strip()
        template = str(o.get("headersFrom") or "").strip()
        ctx = f"create_sheet[{i}] {sheet}"
        if not sheet or sheet.startswith("form_"):
            errors.append(f"{ctx}: invalid sheet name")
            continue
        if sheet in extract["sheets"]:
            errors.append(f"{ctx}: sheet already exists")
            continue
        if family not in EDITOR_SHEET_META:
            errors.append(f"{ctx}: unknown family {family!r}")
            continue
        template_data = extract["sheets"].get(template)
        if template_data is None:
            errors.append(f"{ctx}: headersFrom sheet not found: {template!r}")
            continue
        headers = list(template_data["headers"])
        missing_keys = [k for k in EDITOR_SHEET_META[family]["key"] if k not in headers]
        if missing_keys:
            errors.append(f"{ctx}: template headers lack key column(s) {missing_keys}")
            continue
        # Register the new sheet so later ops in this batch validate against it.
        extract["sheets"][sheet] = {"headers": headers, "rows": []}
        sheet_family[sheet] = family
        created_templates[sheet] = template
        prepared_creates.append(
            {
                "action": "create_sheet",
                "sheet": sheet,
                "_family": family,
                "_headers": headers,
                "_raw_indices": list(o.get("_raw_indices") or []),
                "_raw_effects": list(o.get("_raw_effects") or []),
            }
        )

    # A scaffold plan can register model_workbook_sources and write to those
    # sheets in the same combined batch. Reflect every pending registration in
    # the edit-time maps even when it remains inactive for runtime discovery;
    # the source-row operation itself is the explicit batch authority.
    source_keycols = EDITOR_SHEET_META["model_workbook_sources"]["key"]
    source_index = _sheet_key_index(extract, "model_workbook_sources", source_keycols)
    for o in ops:
        if o.get("sheet") != "model_workbook_sources" or o.get("action") not in ("add", "update"):
            continue
        key = o.get("key") or {}
        row = dict(source_index.get(_key_tuple(key, source_keycols), {}))
        row.update(key)
        row.update(o.get("row") or {})
        model_key = str(row.get("model_key") or "").strip()
        source_role = str(row.get("source_role") or "").strip()
        sheet_name = str(row.get("sheet_name") or "").strip()
        family = SOURCE_ROLE_FAMILIES.get(source_role)
        if not (model_key and sheet_name and family):
            continue
        sheet_family.setdefault(sheet_name, family)
        models_by_sheet.setdefault(sheet_name, set()).add(model_key)
        by_model_family[(model_key, family)] = sheet_name
    bool_storage = (
        {}
        if batch.get("forceTypedBools") is True
        else _bool_storage_conventions(extract, sheet_family, created_templates)
    )
    promoted = {r.get("model_key"): workbook_truthy(r.get("promoted_to_runtime"))
                for r in rows_of(extract, "model_registry_promotion")}

    batch_adds: dict[str, list] = {}
    for o in ops:
        if o.get("action") == "add":
            batch_adds.setdefault(o.get("sheet"), []).append(o)

    key_indexes: dict[str, dict] = {}
    seen_adds: set = set()
    prepared: list[dict] = []
    scaffold_warned: set[str] = set()

    for i, o in enumerate(ops):
        action, sheet = o.get("action"), o.get("sheet")
        ctx = f"op[{i}] {action} {sheet} {o.get('key')}"
        if action not in ("add", "update", "delete"):
            errors.append(f"{ctx}: unknown action")
            continue
        family = sheet_family.get(sheet)
        if not family or str(sheet).startswith("form_"):
            errors.append(f"{ctx}: sheet is not editable")
            continue
        data = extract["sheets"].get(sheet)
        if data is None:
            errors.append(f"{ctx}: sheet not found in workbook")
            continue
        headers = set(data["headers"])
        meta = EDITOR_SHEET_META[family]
        keycols = list(meta["key"])
        key = o.get("key") or {}
        if sorted(key) != sorted(keycols):
            errors.append(f"{ctx}: key must be exactly {keycols}")
            continue
        blank_keys = optional_key_columns(family)
        if any(not str(v or "").strip() and k not in blank_keys for k, v in key.items()):
            errors.append(f"{ctx}: blank key value")
            continue
        row = {k: v for k, v in (o.get("row") or {}).items() if not str(k).startswith("_")}
        unknown = sorted(c for c in row if c not in headers)
        if unknown:
            errors.append(f"{ctx}: unknown column(s) {unknown}")
            continue
        # A physical Excel column is not writable authority. The shared registry
        # owns the writable set, so a rogue column that exists only in the sheet
        # is rejected even though it passes the header check above.
        unregistered = sorted(c for c in row if c not in EDITOR_SHEET_META[family]["columns"])
        if unregistered:
            errors.append(
                f"{ctx}: column(s) {unregistered} are not owned by family {family!r} "
                "in the shared workbook registry"
            )
            continue
        if action == "update" and any(c in row for c in keycols):
            errors.append(f"{ctx}: key columns are immutable on update")
            continue
        coerced = {}
        bad = False
        for col, val in row.items():
            try:
                coerced[col] = coerce_value(family, col, val, bool_storage=bool_storage.get((str(sheet), str(col))))
            except ValueError as exc:
                errors.append(f"{ctx}: {exc}")
                bad = True
        if bad:
            continue
        if sheet not in key_indexes:
            key_indexes[sheet] = _sheet_key_index(extract, sheet, keycols, blank_keys)
        kt = _key_tuple(key, keycols)
        effective_row = dict(key_indexes[sheet].get(kt, {}))
        effective_row.update(key)
        effective_row.update(coerced)
        required_columns = set(meta.get("required_on_add", ())) if action == "add" else set()
        effective_is_active = "active" not in meta.get("columns", ()) or workbook_truthy(
            effective_row.get("active")
        )
        if effective_is_active:
            required_columns.update(meta.get("required_on_effective_active_row", ()))
        # A requirement introduced by a completed migration applies only once
        # the sheet actually carries the column, and never to a delete: the
        # pre-migration case is distinguished by the column being absent, and
        # removing a row must not require filling in its copy first (§7.1).
        required_columns.difference_update(
            column
            for column in ACTIVE_ROW_REQUIRED_COLUMNS.get(family, ())
            if action == "delete" or column not in headers
        )
        for required_column in sorted(required_columns):
            value = effective_row.get(required_column)
            if value is None or not str(value).strip():
                errors.append(f"{ctx}: required field {required_column} is blank")
                bad = True
        if bad:
            continue
        for col, refkind in _meta_ref_items(meta):
            if action == "delete":
                continue
            if action != "add" and col not in coerced and col not in key:
                continue
            value = effective_row.get(col)
            if value is not None and str(value).strip():
                domain = _ref_domain(extract, maps, batch_adds, sheet, refkind, row=effective_row)
                if str(value) not in domain:
                    errors.append(f"{ctx}: {col}={value!r} not found in {_ref_label(refkind)}")
                    bad = True
        conditional_meta = meta.get("conditional_ref") or {}
        conditional_column = conditional_meta.get("column")
        discriminator_column = conditional_meta.get("discriminator")
        if action != "delete" and conditional_column and (
            action == "add"
            or conditional_column in coerced
            or discriminator_column in coerced
            or conditional_column in key
            or discriminator_column in key
        ):
            conditional = _conditional_ref_spec(meta, effective_row)
            if conditional is not None:
                column, refkind = conditional
                value = str(effective_row.get(column) or "").strip()
                if refkind is None:
                    if value:
                        errors.append(
                            f"{ctx}: {column} must be blank for "
                            f"{discriminator_column}={effective_row.get(discriminator_column)!r}"
                        )
                        bad = True
                elif not value:
                    errors.append(
                        f"{ctx}: {column} is required for "
                        f"{discriminator_column}={effective_row.get(discriminator_column)!r}"
                    )
                    bad = True
                else:
                    domain = _ref_domain(extract, maps, batch_adds, sheet, refkind, row=effective_row)
                    probe = value.lower() if refkind in _CASE_INSENSITIVE_REF_KINDS else value
                    if probe not in domain:
                        errors.append(f"{ctx}: {column}={value!r} not found in {_ref_label(refkind)}")
                        bad = True
        if bad:
            continue
        if action == "add":
            if any(str(coerced.get(k) or "").strip() != kt[idx] for idx, k in enumerate(keycols)):
                errors.append(f"{ctx}: add row must include key columns matching the key")
                continue
            if kt in key_indexes[sheet] or (sheet, kt) in seen_adds:
                errors.append(f"{ctx}: duplicate key")
                continue
            seen_adds.add((sheet, kt))
        else:
            if kt not in key_indexes[sheet]:
                errors.append(f"{ctx}: row not found for key")
                continue
        models = models_by_sheet.get(sheet, set())
        if models and not any(promoted.get(m) for m in models) and sheet not in scaffold_warned:
            scaffold_warned.add(sheet)
            warnings.append({"id": f"scaffold:{sheet}",
                             "message": f"{sheet}: model is not promoted to runtime (scaffold)"})
        o = dict(o)
        o["_family"] = family
        o["_coerced_row"] = coerced
        o["_kt"] = kt
        prepared.append(o)

    prepared = prepared_creates + prepared
    if errors:
        return errors, warnings, prepared

    # composite-level integrity
    _price_ref_lints(extract, prepared, errors)
    for o in prepared:
        if o["action"] != "add":
            continue
        family, sheet = o["_family"], o["sheet"]
        models = models_by_sheet.get(sheet, set())
        if family == "options":
            oid = o["_kt"][0]
            for model in models:
                ovs_sheet = by_model_family.get((model, "ovs"))
                if not ovs_sheet:
                    continue
                for vrow in rows_of(extract, "model_variants"):
                    if vrow.get("model_key") != model or not workbook_truthy(vrow.get("active")):
                        continue
                    vid = str(vrow.get("variant_id")).strip()
                    covered = any(p["action"] == "add" and p["sheet"] == ovs_sheet
                                  and p["_kt"] == (oid, vid) for p in prepared)
                    if not covered:
                        errors.append(f"add option {oid}: missing OVS coverage for variant {vid} in {ovs_sheet}")
        if family in ("rule_groups", "exclusive_groups"):
            member_family = "rule_group_members" if family == "rule_groups" else "exclusive_members"
            minimum = 1 if family == "rule_groups" else 2
            gid = o["_kt"][0]
            count = 0
            for model in models:
                member_sheet = by_model_family.get((model, member_family))
                count += sum(1 for p in prepared if p["action"] == "add"
                             and p["sheet"] == member_sheet and p["_kt"][0] == gid)
            if count < minimum:
                errors.append(f"add {family[:-1]} {gid}: requires at least {minimum} member row(s) in the same batch")

    # display-order collision warnings
    for o in prepared:
        if o["action"] not in ("add", "update"):
            continue
        group_col = _DORDER_GROUP_COL.get(o["_family"])
        dorder = o["_coerced_row"].get("display_order")
        if group_col is None or dorder is None:
            continue
        sheet = o["sheet"]
        existing = key_indexes[sheet].get(o["_kt"], {})
        group_val = o["_coerced_row"].get(group_col) or str(existing.get(group_col) or "").strip()
        clash = False
        for kt2, row2 in key_indexes[sheet].items():
            if kt2 == o["_kt"]:
                continue
            if str(row2.get(group_col) or "").strip() == str(group_val) and \
                    str(row2.get("display_order") or "").strip() == str(dorder):
                clash = True
        for p in prepared:
            if p is o or p["sheet"] != sheet or p["action"] in ("create_sheet", "delete"):
                continue
            p_group = p["_coerced_row"].get(group_col)
            if p_group is None:
                p_existing = key_indexes[sheet].get(p["_kt"], {})
                p_group = str(p_existing.get(group_col) or "").strip()
            if str(p_group) == str(group_val) and p["_coerced_row"].get("display_order") == dorder:
                clash = True
        if clash:
            warnings.append({"id": f"dorder:{sheet}:{'+'.join(o['_kt'])}",
                             "message": f"{sheet}: display_order {dorder} duplicates another row "
                                        f"in {group_col}={group_val}"})

    # referenced-delete warnings
    delete_ops = [operation for operation in prepared if operation["action"] == "delete"]
    if delete_ops:
        final_rows = _final_rows_by_sheet(extract, prepared)
        reverse_references, final_option_rpos = _build_reverse_reference_index(
            final_rows,
            maps,
            sheet_family,
        )
    for o in delete_ops:
        family, sheet = o["_family"], o["sheet"]
        target_id = o["_kt"][0]
        source_row = key_indexes.get(sheet, {}).get(o["_kt"], {})
        referencing = _incoming_references(
            reverse_references,
            final_option_rpos,
            maps,
            deleted_sheet=sheet,
            deleted_family=family,
            target_id=target_id,
            target_rpo=str(source_row.get("rpo") or "").strip(),
        )
        if referencing:
            warnings.append({"id": f"refdel:{sheet}:{'+'.join(o['_kt'])}",
                             "message": f"delete {target_id}: still referenced by {referencing}"})
    return errors, warnings, prepared


def validate_batch(extract, batch):
    errors, warnings, _prepared = _prepare_batch(extract, batch)
    return {"errors": errors, "warnings": warnings}


def _operation_coverage(batch: dict, prepared: list[dict]) -> tuple[dict, list[str]]:
    raw_count = len(flatten_items(batch.get("items") or []))
    expected = set(range(raw_count))
    occurrences: dict[int, int] = {}
    invalid_indices: set[int] = set()
    unknown: list[str] = []
    errors: list[str] = []
    for prepared_index, operation in enumerate(prepared):
        raw_indices = list(operation.get("_raw_indices") or [])
        effects = list(operation.get("_raw_effects") or [])
        effect_indices = [effect.get("index") for effect in effects]
        if raw_indices != effect_indices:
            errors.append(
                f"prepared operation {prepared_index} raw provenance differs: "
                f"indices {raw_indices!r}, effects {effect_indices!r}"
            )
            invalid_indices.update(index for index in raw_indices if isinstance(index, int))
            invalid_indices.update(index for index in effect_indices if isinstance(index, int))
        for effect in effects:
            raw_index = effect.get("index")
            if not isinstance(raw_index, int) or raw_index not in expected:
                unknown.append(repr(raw_index))
                continue
            occurrences[raw_index] = occurrences.get(raw_index, 0) + 1
            raw_action = effect.get("action")
            prepared_action = operation.get("action")
            action_matches = (
                raw_action == prepared_action
                or (raw_action == "update" and prepared_action == "add")
            )
            if not action_matches:
                errors.append(
                    f"raw operation {raw_index} action {raw_action!r} does not map to "
                    f"prepared action {prepared_action!r}"
                )
                invalid_indices.add(raw_index)
            missing_fields = sorted(
                set(effect.get("fields") or {})
                - set(operation.get("_coerced_row") or {})
            )
            if missing_fields:
                errors.append(
                    f"raw operation {raw_index} fields are absent from its prepared effect: "
                    f"{missing_fields}"
                )
                invalid_indices.add(raw_index)
    duplicated = sorted(raw_index for raw_index, count in occurrences.items() if count > 1)
    invalid_indices.update(duplicated)
    covered = {
        raw_index
        for raw_index, count in occurrences.items()
        if count == 1 and raw_index not in invalid_indices
    }
    missing = sorted(expected - covered)
    if missing:
        errors.append(f"raw operation coverage missing indices {missing}")
    if duplicated:
        errors.append(f"raw operation coverage duplicated indices {duplicated}")
    if unknown:
        errors.append(f"raw operation coverage contains unknown indices {sorted(unknown)}")
    coverage = {
        "rawCount": raw_count,
        "rawCovered": len(covered),
        "preparedCount": len(prepared),
    }
    if coverage["rawCovered"] != coverage["rawCount"]:
        errors.append(
            "raw operation coverage gate failed: "
            f"{coverage['rawCovered']} != {coverage['rawCount']}"
        )
    return coverage, errors


# ─────────────────────────────────────────────────────────────
# Apply pipeline
# ─────────────────────────────────────────────────────────────

CONFIRMABLE_WARNING_KINDS = {"scaffold"}


def warning_kind(warning_id: str) -> str:
    return str(warning_id or "").partition(":")[0].strip()


def warning_fingerprint(warnings) -> str:
    warning_ids = sorted({str(warning.get("id") or "") for warning in warnings})
    payload = json.dumps(warning_ids, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def classify_warnings(warnings) -> dict:
    warning_list = list(warnings)
    confirmable_ids: list[str] = []
    blocking_ids: list[str] = []
    unknown_ids: list[str] = []
    known_kinds = CONFIRMABLE_WARNING_KINDS | {"dorder", "refdel"}
    for warning in warning_list:
        warning_id = str(warning.get("id") or "")
        kind = warning_kind(warning_id)
        if kind in CONFIRMABLE_WARNING_KINDS:
            confirmable_ids.append(warning_id)
        else:
            blocking_ids.append(warning_id)
            if kind not in known_kinds:
                unknown_ids.append(warning_id)
    return {
        "confirmableIds": sorted(set(confirmable_ids)),
        "blockingIds": sorted(set(blocking_ids)),
        "unknownIds": sorted(set(unknown_ids)),
        "fingerprint": warning_fingerprint(warning_list),
    }


def gate_reminders(models: set[str]) -> list[str]:
    """Return the current post-write route; README owns the full gate matrix.

    Preview/draft tests are optional diagnostics, not readiness authority. The
    composed candidate lane is the one current command that performs complete
    fresh generation, strict validation, registry construction, and browser
    proof without writing tracked artifacts.
    """

    model_keys = sorted(models)
    changed_models = " ".join(f"--changed-model {model}" for model in model_keys)
    commands = [
        ".venv/bin/python scripts/validate_workbook_package.py stingray_master.xlsx",
        ".venv/bin/python scripts/validate_workbook_schema.py stingray_master.xlsx",
    ]
    if changed_models:
        commands.append(
            ".venv/bin/python scripts/verify_workbook_candidate.py --workbook stingray_master.xlsx "
            f"{changed_models} --report /tmp/27vette-workbook-candidate.json"
        )
    commands.extend(f".venv/bin/python scripts/generate_form.py --model {model}" for model in model_keys)
    commands.append(".venv/bin/python scripts/generate_registry.py")
    return commands


def apply_ops_to_workbook(wb, prepared_ops, sheet_family) -> set[str]:
    touched: set[str] = set()
    for o in (x for x in prepared_ops if x["action"] == "create_sheet"):
        ws = wb.create_sheet(title=o["sheet"])
        ws.append(o["_headers"])
        touched.add(o["sheet"])
    by_sheet: dict[str, list] = {}
    for o in prepared_ops:
        if o["action"] == "create_sheet":
            continue
        by_sheet.setdefault(o["sheet"], []).append(o)
    for sheet, sheet_ops in by_sheet.items():
        ws = wb[sheet]
        col_of = {str(c.value): i + 1 for i, c in enumerate(ws[1]) if c.value is not None}
        # Ops carry their resolved family (global sheets and batch-created
        # sheets aren't in the model registry's sheet_family map).
        family = sheet_ops[0].get("_family") or sheet_family[sheet]
        keycols = list(EDITOR_SHEET_META[family]["key"])

        def key_at(r):
            return tuple(str(ws.cell(row=r, column=col_of[k]).value or "").strip() for k in keycols)

        kmap = {key_at(r): r for r in range(2, ws.max_row + 1)}
        for o in (x for x in sheet_ops if x["action"] == "update"):
            r = kmap[o["_kt"]]
            for col, val in o["_coerced_row"].items():
                ws.cell(row=r, column=col_of[col]).value = val
        for r in sorted((kmap[o["_kt"]] for o in sheet_ops if o["action"] == "delete"), reverse=True):
            ws.delete_rows(r)
        for o in (x for x in sheet_ops if x["action"] == "add"):
            values = [None] * max(col_of.values())
            for col, val in o["_coerced_row"].items():
                values[col_of[col] - 1] = val
            ws.append(values)
        touched.add(sheet)
    return touched


def resize_sheet_tables(ws) -> None:
    last = 1
    for r in range(1, ws.max_row + 1):
        if any(c.value is not None for c in ws[r]):
            last = r
    last = max(last, 2)
    for name in list(ws.tables):
        table = ws.tables[name]
        ref = str(table.ref)
        if not ref.startswith("A1:"):
            continue
        end = ref.split(":", 1)[1]
        letters = "".join(ch for ch in end if ch.isalpha())
        table.ref = f"A1:{letters}{last}"


def _save_scratch_workbook(workbook, path: Path) -> None:
    """Single scratch-save seam so tests can corrupt the saved copy before readback."""
    workbook.save(path)


def _canonical_readback_key(values, columns: dict[str, int], key_columns) -> tuple[str, ...]:
    return tuple(
        str(values[columns[column]] or "").strip()
        for column in key_columns
    )


def verify_prepared_workbook(path: Path, prepared: list[dict]) -> dict:
    """Reopen ``path`` and verify every prepared effect exactly as saved."""
    errors: list[str] = []
    prepared_checked = 0
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        grouped: dict[tuple[str, str], list[dict]] = {}
        for operation in prepared:
            if operation.get("action") == "create_sheet":
                continue
            grouped.setdefault(
                (str(operation.get("sheet") or ""), str(operation.get("_family") or "")),
                [],
            ).append(operation)

        readback: dict[tuple[str, str], dict] = {}
        for (sheet, family), operations in grouped.items():
            if sheet not in workbook.sheetnames:
                readback[(sheet, family)] = {"error": "target sheet is absent"}
                continue
            worksheet = workbook[sheet]
            headers = [cell.value for cell in worksheet[1]]
            columns = {
                str(header): column_index
                for column_index, header in enumerate(headers)
                if header is not None
            }
            key_columns = tuple(EDITOR_SHEET_META.get(family, {}).get("key") or ())
            missing_key_columns = [column for column in key_columns if column not in columns]
            if missing_key_columns:
                readback[(sheet, family)] = {
                    "error": f"key columns are absent: {missing_key_columns}"
                }
                continue
            requested_keys = {operation.get("_kt") for operation in operations}
            matches_by_key = {key: [] for key in requested_keys}
            for values in worksheet.iter_rows(min_row=2, values_only=True):
                row_key = _canonical_readback_key(values, columns, key_columns)
                if row_key in matches_by_key:
                    matches_by_key[row_key].append(values)
            readback[(sheet, family)] = {
                "columns": columns,
                "matches": matches_by_key,
            }

        for index, operation in enumerate(prepared):
            action = operation.get("action")
            sheet = str(operation.get("sheet") or "")
            context = f"prepared[{index}] {action} {sheet}"
            if action == "create_sheet":
                if sheet not in workbook.sheetnames:
                    errors.append(f"{context}: created sheet is absent")
                    continue
                actual_headers = [cell.value for cell in workbook[sheet][1]]
                expected_headers = list(operation.get("_headers") or [])
                if actual_headers != expected_headers:
                    errors.append(
                        f"{context}: exact headers differ; "
                        f"expected {expected_headers!r}, got {actual_headers!r}"
                    )
                    continue
                prepared_checked += 1
                continue
            family = operation.get("_family")
            cached = readback[(sheet, str(family or ""))]
            if cached.get("error"):
                errors.append(f"{context}: {cached['error']}")
                continue
            columns = cached["columns"]
            matches = cached["matches"].get(operation.get("_kt"), [])
            if action == "delete":
                if matches:
                    errors.append(
                        f"{context}: delete key {operation.get('_kt')!r} still exists "
                        f"in {len(matches)} row(s)"
                    )
                    continue
                prepared_checked += 1
                continue
            if action not in ("add", "update"):
                errors.append(f"{context}: unsupported prepared action")
                continue
            if len(matches) != 1:
                errors.append(
                    f"{context}: key {operation.get('_kt')!r} matched {len(matches)} row(s), expected 1"
                )
                continue
            actual_row = matches[0]
            field_errors = []
            for column, expected in (operation.get("_coerced_row") or {}).items():
                if column not in columns:
                    field_errors.append(f"field {column} is absent")
                    continue
                actual = actual_row[columns[column]]
                if actual != expected:
                    field_errors.append(
                        f"field {column} expected {expected!r}, got {actual!r}"
                    )
            if field_errors:
                errors.extend(f"{context}: {error}" for error in field_errors)
                continue
            prepared_checked += 1
    finally:
        workbook.close()
    prepared_count = len(prepared)
    if prepared_checked != prepared_count:
        errors.append(
            "prepared operation verification gate failed: "
            f"{prepared_checked} != {prepared_count}"
        )
    return {
        "ok": not errors,
        "preparedChecked": prepared_checked,
        "preparedCount": prepared_count,
        "errors": errors,
    }


def _workbook_identity_matches(path: Path, expected_mtime_ns, expected_sha256) -> bool:
    """True when the live file still matches the reviewed mtime/SHA identity."""
    if str(path.stat().st_mtime_ns) != str(expected_mtime_ns):
        return False
    if expected_sha256 is not None:
        actual = hashlib.sha256(path.read_bytes()).hexdigest()
        if actual != expected_sha256:
            return False
    return True


def _append_edit_log(log_file: Path, entry: dict) -> None:
    """Append one completed write entry through a fault-injectable seam."""
    log_file.parent.mkdir(parents=True, exist_ok=True)
    with log_file.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry) + "\n")


def _post_save_failure_result(
    path: Path,
    backup_path: Path,
    *,
    failure: dict,
    errors: list[str],
    restored_status: str,
    base: dict,
    extra: dict | None = None,
) -> dict:
    """Restore and hash-verify a backup while retaining original failure evidence."""
    restoration = {
        "attempted": False,
        "verified": False,
        "backupSha256": None,
        "workbookSha256": None,
        "error": None,
    }
    restore_error = None
    try:
        restoration["backupSha256"] = hashlib.sha256(backup_path.read_bytes()).hexdigest()
        restoration["attempted"] = True
        restore_workbook_backup(path, backup_path)
        restoration["workbookSha256"] = hashlib.sha256(path.read_bytes()).hexdigest()
        restoration["verified"] = (
            restoration["workbookSha256"] == restoration["backupSha256"]
        )
        if not restoration["verified"]:
            restore_error = "restored workbook SHA-256 does not match backup"
    except Exception as exc:  # never claim the workbook is safe
        restore_error = f"{type(exc).__name__}: {exc}"
    restoration["error"] = restore_error

    result = dict(base)
    result.update(extra or {})
    result.update({
        "ok": False,
        "status": restored_status if restoration["verified"] else "workbook_restore_failed",
        "workbookState": "restored" if restoration["verified"] else "unknown",
        "backupPath": str(backup_path),
        "workbookPath": str(path),
        "failure": failure,
        "restoration": restoration,
    })
    if restoration["verified"]:
        result["errors"] = list(errors)
    else:
        original_detail = "; ".join(errors) or str(failure.get("detail") or "unknown")
        result["errors"] = [
            f"{failure['phase']} failed ({original_detail}) and backup restoration "
            f"could not be verified; workbook path: {path}; backup path: {backup_path}"
            + (f"; restore error: {restore_error}" if restore_error else "")
        ]
    return result


def apply_batch(path, batch, *, write=False, confirmed_warnings=(), source="cli",
                log_path=None, allow_stale=False, run_schema_validation=True) -> dict:
    path = Path(path)
    if write and not run_schema_validation:
        return {
            "ok": False,
            "status": "schema_validation_required",
            "errors": ["live workbook writes require schema validation"],
            "warnings": [],
        }
    lock = excel_lock_path(path)
    if lock.exists():
        return {"ok": False, "status": "locked",
                "errors": [f"Excel lock file present: {lock}. Close Excel first."], "warnings": []}
    # mtime compares as strings: st_mtime_ns exceeds JS Number.MAX_SAFE_INTEGER,
    # so the browser must round-trip it as a string to keep full precision.
    if not allow_stale and str(batch.get("workbookMtimeNs")) != str(path.stat().st_mtime_ns):
        return {"ok": False, "status": "stale",
                "errors": ["workbook changed since this batch was prepared; reload and re-verify"],
                "warnings": []}
    # Reviewed workbook identity: mtime always, SHA-256 when the batch carries
    # one (ChangeSet-derived batches). Rechecked before live mutation and
    # again before the safe save so mid-flight drift fails closed.
    expected_mtime = batch.get("workbookMtimeNs")
    expected_sha = batch.get("workbookSha256")
    extract = extract_workbook(path)
    errors, warnings, prepared = _prepare_batch(extract, batch)
    operation_coverage, coverage_errors = _operation_coverage(batch, prepared)
    if errors or coverage_errors:
        return {
            "ok": False,
            "status": "invalid",
            "errors": errors + coverage_errors,
            "warnings": warnings,
            "operationCoverage": operation_coverage,
        }
    if not prepared:
        return {
            "ok": False,
            "status": "empty",
            "errors": ["batch contains no operations"],
            "warnings": [],
            "operationCoverage": operation_coverage,
        }
    confirmed = set(confirmed_warnings or ())
    warning_policy = classify_warnings(warnings)
    emitted_ids = {str(warning.get("id") or "") for warning in warnings}
    stale_confirmations = sorted(confirmed - emitted_ids)
    if write and stale_confirmations:
        return {
            "ok": False,
            "status": "warning_confirmation_mismatch",
            "errors": [f"confirmed warning IDs were not emitted: {stale_confirmations}"],
            "warnings": warnings,
            "warningPolicy": warning_policy,
            "operationCoverage": operation_coverage,
        }
    if write and warning_policy["blockingIds"]:
        return {
            "ok": False,
            "status": "warning_blocked",
            "errors": ["batch emitted unconfirmable warning IDs"],
            "warnings": warnings,
            "warningPolicy": warning_policy,
            "operationCoverage": operation_coverage,
        }
    unconfirmed_ids = set(warning_policy["confirmableIds"]) - confirmed
    if write and unconfirmed_ids:
        unconfirmed = [warning for warning in warnings if warning["id"] in unconfirmed_ids]
        return {
            "ok": False,
            "status": "needs_confirmation",
            "errors": [],
            "warnings": unconfirmed,
            "warningPolicy": warning_policy,
            "operationCoverage": operation_coverage,
        }

    _registry, sheet_family, models_by_sheet, _bmf = _registry_maps(extract)
    schema_result = None
    bool_hygiene_result = None
    verification = None
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp = Path(tmp_dir) / path.name
        shutil.copy2(path, tmp)
        wb_tmp = load_workbook(tmp)
        touched = apply_ops_to_workbook(wb_tmp, prepared, sheet_family)
        for name in touched:
            resize_sheet_tables(wb_tmp[name])
        remove_table_sheet_auto_filters(wb_tmp)
        _save_scratch_workbook(wb_tmp, tmp)
        wb_tmp.close()
        verification = verify_prepared_workbook(tmp, prepared)
        if not verification["ok"]:
            return {
                "ok": False,
                "status": "readback_failed",
                "errors": verification["errors"],
                "warnings": warnings,
                "warningPolicy": warning_policy,
                "operationCoverage": operation_coverage,
                "verification": verification,
            }
        assert_valid_workbook_package(tmp)
        approved_bool_type_migrations = None
        if batch.get("forceTypedBools") is True:
            typed_sheet_family = dict(sheet_family)
            for operation in prepared:
                operation_sheet = str(operation.get("sheet") or "")
                operation_family = str(operation.get("_family") or "")
                if operation_sheet and operation_family:
                    typed_sheet_family[operation_sheet] = operation_family
            approved_bool_type_migrations = [
                (sheet, column)
                for sheet, family in typed_sheet_family.items()
                for column, kind in EDITOR_SHEET_META.get(family, {}).get(
                    "types", {}
                ).items()
                if kind == "bool"
            ]
        bool_issues = compare_bool_like_workbooks(
            path,
            tmp,
            approved_bool_type_migrations=approved_bool_type_migrations,
        )
        bool_hygiene_result = bool_hygiene_result_payload(path, tmp, bool_issues)
        bool_hygiene_result["issues"] = bool_hygiene_result["issues"][:20]
        if bool_hygiene_result["error_count"]:
            return {"ok": False, "status": "bool_hygiene_failed",
                    "errors": [f"dry-run bool hygiene failed with "
                               f"{bool_hygiene_result['error_count']} error(s)"],
                    "warnings": warnings, "boolHygieneResult": bool_hygiene_result,
                    "operationCoverage": operation_coverage, "verification": verification}
        if run_schema_validation:
            issues = validate_workbook_schema(str(tmp), check_live_contract=False)
            schema_result = result_payload(str(tmp), issues)
            if schema_result["error_count"]:
                return {"ok": False, "status": "schema_failed",
                        "errors": [f"dry-run schema validation failed with "
                                   f"{schema_result['error_count']} error(s)"],
                        "warnings": warnings, "schemaResult": schema_result,
                        "operationCoverage": operation_coverage, "verification": verification}

    models_touched = {m for s in touched for m in models_by_sheet.get(s, set())}
    base = {"opCount": len(prepared), "sheets": sorted(touched), "warnings": warnings,
            "warningPolicy": warning_policy,
            "operationCoverage": operation_coverage, "verification": verification,
            "schemaResult": schema_result, "boolHygieneResult": bool_hygiene_result,
            "gateReminders": gate_reminders(models_touched)}
    if not write:
        return {"ok": True, "status": "validated", "errors": [], **base}

    if not allow_stale and not _workbook_identity_matches(path, expected_mtime, expected_sha):
        return {
            "ok": False,
            "status": "stale_before_save",
            "workbookState": "untouched",
            "errors": ["workbook changed after review but before live mutation; "
                       "re-preview and re-approve"],
            "warnings": warnings,
            "warningPolicy": warning_policy,
            "operationCoverage": operation_coverage,
            "verification": verification,
        }
    wb = load_workbook(path)
    loaded_mtime = path.stat().st_mtime_ns
    touched = apply_ops_to_workbook(wb, prepared, sheet_family)
    for name in touched:
        resize_sheet_tables(wb[name])
    if not allow_stale and not _workbook_identity_matches(path, expected_mtime, expected_sha):
        return {
            "ok": False,
            "status": "stale_before_save",
            "workbookState": "untouched",
            "errors": ["workbook changed after live mutation but before save; "
                       "re-preview and re-approve"],
            "warnings": warnings,
            "warningPolicy": warning_policy,
            "operationCoverage": operation_coverage,
            "verification": verification,
        }
    backup_path = save_workbook_safely(
        wb,
        path,
        loaded_mtime_ns=int(expected_mtime) if not allow_stale else loaded_mtime,
        approved_bool_type_migrations=approved_bool_type_migrations,
    )
    phase = "live_readback"
    try:
        live_verification = verify_prepared_workbook(path, prepared)
        if not live_verification["ok"]:
            failure = {
                "phase": phase,
                "kind": "returned_failure",
                "detail": "; ".join(live_verification["errors"]),
            }
            return _post_save_failure_result(
                path,
                backup_path,
                failure=failure,
                errors=live_verification["errors"],
                restored_status="apply_verification_failed_rolled_back",
                base=base,
                extra={"verification": live_verification},
            )

        phase = "live_package"
        assert_valid_workbook_package(path)

        phase = "live_schema"
        live_schema_result = schema_result
        if run_schema_validation:
            live_schema_issues = validate_workbook_schema(str(path), check_live_contract=False)
            live_schema_result = result_payload(str(path), live_schema_issues)
            if live_schema_result["error_count"]:
                failure = {
                    "phase": phase,
                    "kind": "returned_failure",
                    "detail": (
                        "live schema validation returned "
                        f"{live_schema_result['error_count']} error(s)"
                    ),
                }
                return _post_save_failure_result(
                    path,
                    backup_path,
                    failure=failure,
                    errors=[failure["detail"]],
                    restored_status="apply_verification_failed_rolled_back",
                    base=base,
                    extra={
                        "verification": live_verification,
                        "schemaResult": live_schema_result,
                    },
                )

        log_file = Path(log_path) if log_path else DEFAULT_LOG_PATH
        entry = {"ts": datetime.now().isoformat(timespec="seconds"), "source": source,
                 "workbook": str(path), "opCount": len(prepared),
                 "composites": sorted({o["_composite"] for o in prepared if o.get("_composite")}),
                 "sheets": sorted(touched), "backupPath": str(backup_path),
                 "schemaErrors": None if live_schema_result is None else live_schema_result["error_count"],
                 "warningsConfirmed": sorted(confirmed)}
        phase = "success_result"
        success_result = {
            "ok": True,
            "status": "applied",
            "errors": [],
            "applied": len(prepared),
            "backupPath": str(backup_path),
            "logPath": str(log_file),
            **base,
            "schemaResult": live_schema_result,
            "verification": live_verification,
        }

        phase = "write_log"
        _append_edit_log(log_file, entry)
        return success_result
    except Exception as exc:
        detail = f"{type(exc).__name__}: {exc}"
        failure = {"phase": phase, "kind": "exception", "detail": detail}
        return _post_save_failure_result(
            path,
            backup_path,
            failure=failure,
            errors=[detail],
            restored_status="apply_verification_failed_rolled_back",
            base=base,
            extra={
                "verification": locals().get("live_verification"),
                "schemaResult": locals().get("live_schema_result", schema_result),
            },
        )
