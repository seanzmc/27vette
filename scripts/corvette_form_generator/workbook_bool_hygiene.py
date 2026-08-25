"""Boolean-looking workbook cell type hygiene guards.

The generator normalizes bool-like values for emitted JSON, but workbook writes must
not silently convert source cells between text booleans ("TRUE"/"False"/etc.) and
real Excel booleans unless that sheet/column migration is explicitly approved.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BOOL_TEXT_VALUES = {"True", "False", "TRUE", "FALSE", "true", "false"}

ROW_KEY_CANDIDATES: tuple[tuple[str, ...], ...] = (
    ("option_id",),
    ("choice_id",),
    ("rule_id",),
    ("group_id", "option_id"),
    ("group_id", "target_id"),
    ("group_id",),
    ("model_key", "source_role"),
    ("model_key", "variant_id"),
    ("model_key", "step_key"),
    ("model_key", "context_type"),
    ("model_key", "section_id"),
    ("model_key", "section_key"),
    ("model_key", "interior_id", "trim_level"),
    ("model_key", "interior_id", "rpo", "component_type"),
    ("model_key", "interior_id", "rpo"),
    ("interior_id", "rpo", "component_type"),
    ("interior_id", "rpo"),
    ("interior_id",),
    ("section_id",),
    ("variant_id",),
    ("rpo",),
    ("key",),
    ("id",),
)


@dataclass(frozen=True)
class BoolLikeCell:
    sheet: str
    coordinate: str
    row: int
    column: str
    value: Any
    python_type: str
    logical_value: bool
    storage_family: str
    row_key_kind: str
    row_key_columns: tuple[str, ...]
    row_key_values: tuple[str, ...]

    @property
    def compare_key(self) -> tuple[Any, ...]:
        if self.row_key_kind == "stable":
            return (self.sheet, self.column, "stable", self.row_key_columns, self.row_key_values)
        return (self.sheet, self.column, "coordinate", self.coordinate)

    @property
    def sheet_column(self) -> tuple[str, str]:
        return (self.sheet, self.column)

    def payload(self) -> dict[str, Any]:
        data = asdict(self)
        data["row_key_columns"] = list(self.row_key_columns)
        data["row_key_values"] = list(self.row_key_values)
        return data


@dataclass(frozen=True)
class BoolHygieneIssue:
    check_id: str
    severity: str
    sheet: str
    column: str
    message: str
    before: dict[str, Any] | None = None
    after: dict[str, Any] | None = None
    convention: str = ""

    def payload(self) -> dict[str, Any]:
        return asdict(self)


def bool_like(value: Any) -> tuple[bool, str] | None:
    if isinstance(value, bool):
        return value, "excel_boolean"
    if isinstance(value, str) and value in BOOL_TEXT_VALUES:
        return value.lower() == "true", "text"
    return None


def clean_key_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def header_row(ws) -> list[str]:
    return [clean_key_value(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]


def row_key_counts(
    value_rows: list[tuple[Any, ...]],
    headers: list[str],
) -> dict[tuple[str, ...], Counter[tuple[str, ...]]]:
    """Count each candidate row key over rows already read into memory.

    This used to take the worksheet and call ``ws.iter_rows`` once per
    candidate. The workbook is opened ``read_only=True``, where every
    ``iter_rows`` re-parses the sheet XML from the start, so a sheet matching
    four candidates was parsed four times, and then a fifth time by the
    snapshot loop below. Profiled during a real apply that was 19.9s of
    read-only row parsing here plus 8.7s in the snapshot loop.
    """

    counts: dict[tuple[str, ...], Counter[tuple[str, ...]]] = {}
    header_set = set(headers)
    for candidate in ROW_KEY_CANDIDATES:
        if not set(candidate).issubset(header_set):
            continue
        indexes = [headers.index(column) for column in candidate]
        counter: Counter[tuple[str, ...]] = Counter()
        for row in value_rows:
            values = tuple(clean_key_value(row[index]) if index < len(row) else "" for index in indexes)
            if all(values):
                counter[values] += 1
        counts[candidate] = counter
    return counts


def stable_row_key(row_values: tuple[Any, ...], headers: list[str], counts: dict[tuple[str, ...], Counter[tuple[str, ...]]]) -> tuple[tuple[str, ...], tuple[str, ...]] | None:
    for candidate in ROW_KEY_CANDIDATES:
        counter = counts.get(candidate)
        if not counter:
            continue
        indexes = [headers.index(column) for column in candidate]
        values = tuple(clean_key_value(row_values[index]) if index < len(row_values) else "" for index in indexes)
        if all(values) and counter.get(values) == 1:
            return candidate, values
    return None


def snapshot_bool_like_cells(workbook_path: str | Path) -> list[BoolLikeCell]:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    cells: list[BoolLikeCell] = []
    try:
        for ws in wb.worksheets:
            headers = header_row(ws)
            # One read-only parse of the sheet feeds both the candidate-key
            # counters and the scan below.
            rows = list(ws.iter_rows(min_row=2))
            value_rows = [tuple(cell.value for cell in row) for row in rows]
            counts = row_key_counts(value_rows, headers)
            for row_number, (row, row_values) in enumerate(zip(rows, value_rows), start=2):
                stable_key = stable_row_key(row_values, headers, counts)
                if stable_key:
                    row_key_kind = "stable"
                    row_key_columns, row_key_values = stable_key
                else:
                    row_key_kind = "coordinate"
                    row_key_columns = ()
                    row_key_values = ()
                for cell in row:
                    detected = bool_like(cell.value)
                    if not detected:
                        continue
                    logical_value, storage_family = detected
                    col_index = cell.column if isinstance(cell.column, int) else 0
                    column = headers[col_index - 1] if 0 < col_index <= len(headers) else get_column_letter(col_index or 1)
                    cells.append(
                        BoolLikeCell(
                            sheet=ws.title,
                            coordinate=cell.coordinate,
                            row=row_number,
                            column=column,
                            value=cell.value,
                            python_type=type(cell.value).__name__,
                            logical_value=logical_value,
                            storage_family=storage_family,
                            row_key_kind=row_key_kind,
                            row_key_columns=row_key_columns,
                            row_key_values=row_key_values,
                        )
                    )
    finally:
        wb.close()
    return cells


def convention_by_sheet_column(cells: Iterable[BoolLikeCell]) -> dict[tuple[str, str], str]:
    families: dict[tuple[str, str], set[str]] = {}
    for cell in cells:
        families.setdefault(cell.sheet_column, set()).add(cell.storage_family)
    return {sheet_column: next(iter(values)) for sheet_column, values in families.items() if len(values) == 1}


def normalize_approved_migrations(approved: Iterable[str | tuple[str, str]] | None) -> set[tuple[str, str]]:
    result: set[tuple[str, str]] = set()
    for item in approved or []:
        if isinstance(item, tuple):
            if len(item) != 2:
                raise ValueError(f"Approved bool migration tuple must be (sheet, column), got {item!r}")
            result.add((str(item[0]), str(item[1])))
            continue
        text = str(item)
        if "." not in text:
            raise ValueError(f"Approved bool migration must use sheet.column, got {text!r}")
        sheet, column = text.split(".", 1)
        result.add((sheet, column))
    return result


def compare_bool_like_workbooks(
    before_workbook: str | Path,
    after_workbook: str | Path,
    *,
    approved_bool_type_migrations: Iterable[str | tuple[str, str]] | None = None,
) -> list[BoolHygieneIssue]:
    approved = normalize_approved_migrations(approved_bool_type_migrations)
    before_cells = snapshot_bool_like_cells(before_workbook)
    after_cells = snapshot_bool_like_cells(after_workbook)
    before_by_key = {cell.compare_key: cell for cell in before_cells}
    conventions = convention_by_sheet_column(before_cells)
    issues: list[BoolHygieneIssue] = []

    for after_cell in after_cells:
        if after_cell.sheet_column in approved:
            continue
        before_cell = before_by_key.get(after_cell.compare_key)
        if before_cell:
            if (
                before_cell.logical_value == after_cell.logical_value
                and before_cell.storage_family != after_cell.storage_family
            ):
                issues.append(
                    BoolHygieneIssue(
                        check_id="bool_type_family_changed",
                        severity="error",
                        sheet=after_cell.sheet,
                        column=after_cell.column,
                        before=before_cell.payload(),
                        after=after_cell.payload(),
                        message=(
                            f"{after_cell.sheet}.{after_cell.column} logical {after_cell.logical_value} changed "
                            f"from {before_cell.storage_family} to {after_cell.storage_family}."
                        ),
                    )
                )
            continue

        expected_family = conventions.get(after_cell.sheet_column)
        if expected_family and after_cell.storage_family != expected_family:
            issues.append(
                BoolHygieneIssue(
                    check_id="added_bool_type_convention_mismatch",
                    severity="error",
                    sheet=after_cell.sheet,
                    column=after_cell.column,
                    after=after_cell.payload(),
                    convention=expected_family,
                    message=(
                        f"Added {after_cell.sheet}.{after_cell.column} bool-like cell uses "
                        f"{after_cell.storage_family}; existing unambiguous convention is {expected_family}."
                    ),
                )
            )

    return issues


def result_payload(before_workbook: str | Path, after_workbook: str | Path, issues: list[BoolHygieneIssue]) -> dict[str, Any]:
    before_cells = snapshot_bool_like_cells(before_workbook)
    after_cells = snapshot_bool_like_cells(after_workbook)
    return {
        "before_workbook": str(before_workbook),
        "after_workbook": str(after_workbook),
        "status": "valid" if not any(issue.severity == "error" for issue in issues) else "invalid",
        "before_bool_like_count": len(before_cells),
        "after_bool_like_count": len(after_cells),
        "issue_count": len(issues),
        "error_count": sum(1 for issue in issues if issue.severity == "error"),
        "issues": [issue.payload() for issue in issues],
    }


def format_issues(issues: Iterable[BoolHygieneIssue], *, limit: int = 20) -> str:
    lines: list[str] = []
    for issue in list(issues)[:limit]:
        after = issue.after or {}
        before = issue.before or {}
        location = after.get("coordinate") or before.get("coordinate") or ""
        lines.append(f"{issue.check_id}: {issue.sheet}.{issue.column} {location}: {issue.message}")
    return "\n".join(lines)
