"""Workbook loading and writing helpers shared by model generators."""

from __future__ import annotations

import shutil
import tempfile
from typing import Any
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from corvette_form_generator.workbook_package import assert_valid_workbook_package


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def money(value: Any) -> int:
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        return 0
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


def intish(value: Any, default: int = 0) -> int:
    text = clean(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def rows_from_sheet(wb, sheet_name: str) -> list[dict[str, str]]:
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record: dict[str, str] = {}
        for header, value in zip(headers, row):
            if header:
                record[header] = clean(value)
        if any(record.values()):
            rows.append(record)
    return rows


def write_sheet(wb, name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width


def excel_lock_path(path: Path) -> Path:
    return path.with_name(f"~${path.name}")


def backup_workbook(path: Path) -> Path:
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    backup_path = backup_dir / f"{path.stem}-{datetime.now().strftime('%Y%m%d-%H%M%S')}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def save_workbook_safely(wb, path: Path, *, loaded_mtime_ns: int | None = None) -> Path:
    path = Path(path)
    lock_path = excel_lock_path(path)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to save {path}; Excel lock file is present: {lock_path}. Close Excel first.")
    if loaded_mtime_ns is not None and path.exists() and path.stat().st_mtime_ns != loaded_mtime_ns:
        raise RuntimeError(f"Refusing to save {path}; file changed after it was loaded.")

    with tempfile.NamedTemporaryFile(prefix=f"{path.stem}-", suffix=path.suffix, delete=False, dir=path.parent) as handle:
        tmp_path = Path(handle.name)
    try:
        wb.save(tmp_path)
        assert_valid_workbook_package(tmp_path)
        check_wb = load_workbook(tmp_path, read_only=True, data_only=True)
        check_wb.close()
        backup_path = backup_workbook(path)
        shutil.move(tmp_path, path)
        return backup_path
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise
