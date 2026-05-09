"""Validate and repair workbook package metadata that Excel is strict about."""

from __future__ import annotations

import json
import posixpath
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"main": SPREADSHEET_NS, "rel": REL_NS, "pkg": PACKAGE_REL_NS}
ET.register_namespace("", SPREADSHEET_NS)


@dataclass(frozen=True)
class TableBinding:
    table_path: str
    sheet_name: str
    worksheet_path: str


def _zip_path(base: str, target: str) -> str:
    if target.startswith("/"):
        target = target[1:]
    if target.startswith("xl/"):
        return posixpath.normpath(target)
    base_dir = PurePosixPath(base).parent
    return posixpath.normpath(str(PurePosixPath(base_dir, target)))


def _rels_path(part_path: str) -> str:
    path = PurePosixPath(part_path)
    return str(path.parent / "_rels" / f"{path.name}.rels")


def _relationship_targets(zf: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    if rels_path not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read(rels_path))
    return {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in root.findall("pkg:Relationship", NS)
        if rel.attrib.get("Id") and rel.attrib.get("Target")
    }


def table_bindings(workbook_path: Path) -> dict[str, TableBinding]:
    with zipfile.ZipFile(workbook_path) as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        workbook_rels = _relationship_targets(zf, "xl/_rels/workbook.xml.rels")
        bindings: dict[str, TableBinding] = {}

        for sheet_node in workbook_root.findall("main:sheets/main:sheet", NS):
            sheet_name = sheet_node.attrib["name"]
            rel_id = sheet_node.attrib[f"{{{REL_NS}}}id"]
            worksheet_path = _zip_path("xl/workbook.xml", workbook_rels[rel_id])
            worksheet_rels = _relationship_targets(zf, _rels_path(worksheet_path))
            worksheet_root = ET.fromstring(zf.read(worksheet_path))
            for table_part in worksheet_root.findall("main:tableParts/main:tablePart", NS):
                table_rel_id = table_part.attrib[f"{{{REL_NS}}}id"]
                table_path = _zip_path(worksheet_path, worksheet_rels[table_rel_id])
                bindings[table_path] = TableBinding(table_path, sheet_name, worksheet_path)
        return bindings


def _table_column_nodes(table_root: ET.Element) -> list[ET.Element]:
    table_columns = table_root.find("main:tableColumns", NS)
    if table_columns is None:
        return []
    return list(table_columns.findall("main:tableColumn", NS))


def _header_values(wb, binding: TableBinding, ref: str, count: int) -> list[str]:
    min_col, min_row, max_col, _ = range_boundaries(ref)
    ws = wb[binding.sheet_name]
    return [
        "" if ws.cell(min_row, col).value is None else str(ws.cell(min_row, col).value).strip()
        for col in range(min_col, min(max_col, min_col + count - 1) + 1)
    ]


def _ref_for_bounds(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def validate_workbook_package(workbook_path: Path) -> list[dict[str, object]]:
    workbook_path = Path(workbook_path)
    issues: list[dict[str, object]] = []
    bindings = table_bindings(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    with zipfile.ZipFile(workbook_path) as zf:
        for table_path in sorted(name for name in zf.namelist() if name.startswith("xl/tables/") and name.endswith(".xml")):
            binding = bindings.get(table_path)
            root = ET.fromstring(zf.read(table_path))
            table_name = root.attrib.get("displayName") or root.attrib.get("name") or table_path
            ref = root.attrib.get("ref", "")
            header_row_count = int(root.attrib.get("headerRowCount", "1"))
            columns_parent = root.find("main:tableColumns", NS)
            columns = _table_column_nodes(root)
            context = {
                "table_path": table_path,
                "table_name": table_name,
                "sheet_name": binding.sheet_name if binding else "",
                "ref": ref,
            }

            if binding is None:
                issues.append({**context, "issue": "table_not_referenced_by_worksheet"})
                continue
            if columns_parent is None:
                issues.append({**context, "issue": "missing_tableColumns"})
                continue

            declared_count = int(columns_parent.attrib.get("count", "-1"))
            if declared_count != len(columns):
                issues.append({**context, "issue": "tableColumns_count_mismatch", "declared": declared_count, "actual": len(columns)})

            ids = [column.attrib.get("id", "") for column in columns]
            expected_ids = [str(index) for index in range(1, len(columns) + 1)]
            if ids != expected_ids:
                issues.append({**context, "issue": "non_sequential_table_column_ids", "actual": ids, "expected": expected_ids})

            if ref:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                ws = wb[binding.sheet_name]
                ref_width = max_col - min_col + 1
                if ref_width != len(columns):
                    issues.append({**context, "issue": "table_ref_width_mismatch", "ref_width": ref_width, "column_count": len(columns)})
                if max_col > ws.max_column or max_row > ws.max_row:
                    issues.append(
                        {
                            **context,
                            "issue": "table_ref_exceeds_sheet_dimensions",
                            "sheet_dimensions": {"rows": ws.max_row, "columns": ws.max_column},
                        }
                    )
                if header_row_count:
                    header_values = _header_values(wb, binding, ref, len(columns))
                    table_names = [column.attrib.get("name", "") for column in columns]
                    if table_names != header_values:
                        issues.append(
                            {
                                **context,
                                "issue": "table_column_names_do_not_match_headers",
                                "actual": table_names,
                                "expected": header_values,
                            }
                        )
    wb.close()
    return issues


def assert_valid_workbook_package(workbook_path: Path) -> None:
    issues = validate_workbook_package(workbook_path)
    if issues:
        raise ValueError(f"Workbook package validation failed for {workbook_path}:\n{json.dumps(issues, indent=2)}")


def repair_workbook_tables(workbook_path: Path, *, backup: bool = True) -> dict[str, object]:
    workbook_path = Path(workbook_path)
    before = validate_workbook_package(workbook_path)
    bindings = table_bindings(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    repaired_tables: list[dict[str, object]] = []

    with tempfile.NamedTemporaryFile(prefix=f"{workbook_path.stem}-", suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    with zipfile.ZipFile(workbook_path, "r") as source, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename.startswith("xl/tables/") and item.filename.endswith(".xml") and item.filename in bindings:
                root = ET.fromstring(data)
                columns_parent = root.find("main:tableColumns", NS)
                ref = root.attrib.get("ref", "")
                if columns_parent is not None and ref:
                    columns = _table_column_nodes(root)
                    binding = bindings[item.filename]
                    ws = wb[binding.sheet_name]
                    min_col, min_row, max_col, max_row = range_boundaries(ref)
                    repaired_max_col = min(max_col, ws.max_column)
                    repaired_max_row = min(max_row, ws.max_row)
                    if repaired_max_col - min_col + 1 != len(columns):
                        repaired_max_col = min_col + len(columns) - 1
                    repaired_ref = _ref_for_bounds(min_col, min_row, repaired_max_col, repaired_max_row)
                    headers = _header_values(wb, binding, repaired_ref, len(columns))
                    changed = False
                    if root.attrib.get("ref") != repaired_ref:
                        root.attrib["ref"] = repaired_ref
                        auto_filter = root.find("main:autoFilter", NS)
                        if auto_filter is not None:
                            auto_filter.attrib["ref"] = repaired_ref
                        sort_state = root.find("main:sortState", NS)
                        if sort_state is not None:
                            sort_state.attrib["ref"] = _ref_for_bounds(min_col, min_row + 1, repaired_max_col, repaired_max_row)
                        changed = True
                    for index, column in enumerate(columns, start=1):
                        expected_name = headers[index - 1] if root.attrib.get("headerRowCount", "1") != "0" else ""
                        if column.attrib.get("id") != str(index):
                            column.attrib["id"] = str(index)
                            changed = True
                        if expected_name and column.attrib.get("name") != expected_name:
                            column.attrib["name"] = expected_name
                            changed = True
                    if columns_parent.attrib.get("count") != str(len(columns)):
                        columns_parent.attrib["count"] = str(len(columns))
                        changed = True
                    if changed:
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
                        repaired_tables.append(
                            {
                                "table_path": item.filename,
                                "table_name": root.attrib.get("displayName") or root.attrib.get("name") or item.filename,
                                "sheet_name": bindings[item.filename].sheet_name,
                                "columns": len(columns),
                            }
                        )
            target.writestr(item, data)
    wb.close()

    backup_path = None
    if repaired_tables:
        assert_valid_workbook_package(tmp_path)
        if backup:
            backup_dir = workbook_path.parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            backup_path = backup_dir / f"{workbook_path.stem}-{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook_path.suffix}"
            shutil.copy2(workbook_path, backup_path)
        shutil.move(tmp_path, workbook_path)
    else:
        tmp_path.unlink(missing_ok=True)

    after = validate_workbook_package(workbook_path)
    return {
        "workbook": str(workbook_path),
        "backup": str(backup_path) if backup_path else None,
        "issues_before": before,
        "issues_after": after,
        "repaired_tables": repaired_tables,
    }
