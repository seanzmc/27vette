"""Model configuration: shared constants plus workbook-first base configs.

The workbook owns model-specific metadata through ``model_master``,
``model_workbook_sources``, ``model_variants``, and
``model_registry_promotion``; resolve a base config against it with
``runtime_metadata.load_model_config_overrides``. Python keeps filesystem
paths, unpromoted-model compatibility defaults, and promoted-model
completeness expectations. Promoted runtime metadata such as steps, context
sections, and order-summary grouping must come from workbook rows.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from corvette_form_generator.model_config import ModelConfig
from corvette_form_generator.runtime_metadata import load_model_config_overrides, truthy
from corvette_form_generator.workbook import clean, intish, rows_from_sheet


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "stingray_master.xlsx"
OUTPUT_DIR = ROOT / "form-output"
APP_DIR = ROOT / "form-app"

STEP_ORDER = (
    "body_style",
    "trim_level",
    "paint",
    "exterior_appearance",
    "wheels",
    "packages_performance",
    "aero_exhaust_stripes_accessories",
    "seat",
    "base_interior",
    "seat_belt",
    "interior_trim",
    "accessories",
    "delivery",
    "summary",
)

STEP_LABELS = {
    "body_style": "Body Style",
    "trim_level": "Trim Level",
    "paint": "Exterior Paint",
    "exterior_appearance": "Exterior Appearance",
    "wheels": "Wheels & Brake Calipers",
    "packages_performance": "Performance & Aero",
    "aero_exhaust_stripes_accessories": "Stripes",
    "seat": "Seats",
    "base_interior": "Interior Color",
    "seat_belt": "Seat Belt",
    "interior_trim": "Interior Trim",
    "accessories": "Accessories",
    "delivery": "Custom Delivery",
    "summary": "Summary",
    "standard_equipment": "Standard Equipment",
}

CONTEXT_SECTIONS = (
    {
        "section_id": "sec_context_body_style",
        "section_name": "Body Style",
        "selection_mode": "single_select_req",
        "selection_mode_label": "Required single choice",
        "choice_mode": "single",
        "is_required": "True",
        "standard_behavior": "user_selected",
        "section_display_order": 1,
        "step_key": "body_style",
        "step_label": "Body Style",
    },
    {
        "section_id": "sec_context_trim_level",
        "section_name": "Trim Level",
        "selection_mode": "single_select_req",
        "selection_mode_label": "Required single choice",
        "choice_mode": "single",
        "is_required": "True",
        "standard_behavior": "user_selected",
        "section_display_order": 2,
        "step_key": "trim_level",
        "step_label": "Trim Level",
    },
)

SECTION_STEP_OVERRIDES = {
    "sec_pain_001": "paint",
    "sec_whee_002": "wheels",
    "sec_cali_001": "wheels",
    "sec_roof_001": "exterior_appearance",
    "sec_exte_001": "exterior_appearance",
    "sec_badg_001": "exterior_appearance",
    "sec_engi_001": "exterior_appearance",
    "sec_perf_001": "packages_performance",
    "sec_susp_001": "packages_performance",
    "sec_seat_002": "seat",
    "sec_intc_001": "base_interior",
    "sec_intc_002": "base_interior",
    "sec_intc_003": "base_interior",
    "sec_seat_001": "seat_belt",
    "sec_inte_001": "interior_trim",
    "sec_lpoi_001": "interior_trim",
    "sec_whee_001": "wheels",
    "sec_gsce_001": "exterior_appearance",
    "sec_gsha_001": "exterior_appearance",
    "sec_colo_001": "interior_trim",
    "sec_onst_001": "interior_trim",
    "sec_cust_002": "interior_trim",
    "sec_spec_001": "packages_performance",
    "sec_cust_001": "delivery",
}

BODY_STYLE_DISPLAY_ORDER = {
    "coupe": 1,
    "convertible": 2,
}

SELECTION_MODE_LABELS = {
    "single_select_req": "Required single choice",
    "single_select_opt": "Optional single choice",
    "multi_select_opt": "Optional multiple choice",
    "display_only": "Display only",
}

STANDARD_SECTIONS = frozenset(
    {
        "sec_1lte_001",
        "sec_2lte_001",
        "sec_3lte_001",
        "sec_incl_001",
        "sec_stan_001",
        "sec_stan_002",
        "sec_safe_001",
        "sec_tech_001",
    }
)

DEFAULT_TEXT_CLEANUP = {
    "enabled": True,
    "normalize_new_prefix": True,
    "collapse_whitespace": True,
    "collapse_repeated_punctuation": True,
    "remove_adjacent_duplicate_phrases": True,
}

_MODEL_NOTES = {
    "grand_sport": (
        "Grand Sport is a promoted runtime model generated through the shared source-assembly path.",
        "Grand Sport option rows are read from the normalized grandSport_options sheet.",
    ),
    "z06": (
        "Z06 is eligible for runtime promotion through workbook-owned model_registry_promotion rows.",
        "Z06 option rows are read from normalized z06_options source rows.",
    ),
}

REQUIRED_GENERATION_SOURCE_ROLES = (
    "source_option_sheet",
    "status_sheet",
    "rule_mapping_sheet",
    "price_rules_sheet",
    "rule_groups_sheet",
    "rule_group_members_sheet",
    "exclusive_groups_sheet",
    "exclusive_group_members_sheet",
    "color_overrides_sheet",
    "interior_source_sheet",
)
OPTIONAL_GENERATION_SOURCE_ROLES = ("variant_option_overrides_sheet",)
_GENERATION_SOURCE_ROLES = frozenset(REQUIRED_GENERATION_SOURCE_ROLES + OPTIONAL_GENERATION_SOURCE_ROLES)


def base_model_config(model_key: str) -> ModelConfig:
    """Build the Python-side base config for any model key.

    Every workbook-expressible field (label, year, dataset name, sheet roles,
    variants) carries only a conventional default here; the workbook metadata
    sheets are authoritative once the config is resolved through
    ``load_model_config_overrides``.
    """

    label = model_key.replace("_", " ").title()
    slug = model_key.replace("_", "-")
    return ModelConfig(
        model_key=model_key,
        model_label=label,
        model_year="2027",
        dataset_name=f"2027 Corvette {label} operational form",
        source_option_sheet=f"{model_key}_options",
        status_sheet=f"{model_key}_ovs",
        variant_ids=(),
        expected_variant_count=0,
        root=ROOT,
        workbook_path=WORKBOOK_PATH,
        output_dir=OUTPUT_DIR,
        app_dir=APP_DIR,
        step_order=STEP_ORDER,
        step_labels=STEP_LABELS,
        context_sections=CONTEXT_SECTIONS,
        body_style_display_order=BODY_STYLE_DISPLAY_ORDER,
        selection_mode_labels=SELECTION_MODE_LABELS,
        standard_sections=STANDARD_SECTIONS,
        section_step_overrides=SECTION_STEP_OVERRIDES,
        preview_artifact_prefix=f"{slug}-contract-preview",
        draft_artifact_prefix=f"{slug}-form-data-draft",
        text_cleanup=dict(DEFAULT_TEXT_CLEANUP),
        notes=_MODEL_NOTES.get(model_key, ()),
    )


def _metadata_rows(wb, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing required workbook metadata sheet {sheet_name!r}.")
    return rows_from_sheet(wb, sheet_name)


def _active_exact_model_rows(wb, sheet_name: str, model_key: str) -> list[dict[str, str]]:
    model = clean(model_key).lower()
    return [
        row
        for row in _metadata_rows(wb, sheet_name)
        if clean(row.get("model_key")).lower() == model
        and truthy(row.get("active", "True"), default=True)
    ]


def _active_model_master_rows(wb) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in _metadata_rows(wb, "model_master"):
        model_key = clean(row.get("model_key")).lower()
        if not model_key or not truthy(row.get("active", "True"), default=True):
            continue
        if model_key in seen:
            raise ValueError(f"Duplicate active model_master rows for model {model_key}")
        seen.add(model_key)
        rows.append(row)
    return rows


def _duplicate_values(values: list[str]) -> list[str]:
    return sorted({value for value in values if values.count(value) > 1})


def _require_generation_sources(wb, model_key: str) -> None:
    rows = _active_exact_model_rows(wb, "model_workbook_sources", model_key)
    roles = [clean(row.get("source_role")) for row in rows if clean(row.get("source_role"))]
    duplicate_roles = _duplicate_values(roles)
    if duplicate_roles:
        raise ValueError(
            f"Duplicate active model_workbook_sources roles for {model_key}: {', '.join(duplicate_roles)}"
        )

    unknown_roles = sorted(set(roles) - _GENERATION_SOURCE_ROLES)
    if unknown_roles:
        raise ValueError(f"Unknown model_workbook_sources roles for {model_key}: {', '.join(unknown_roles)}")

    missing_roles = [role for role in REQUIRED_GENERATION_SOURCE_ROLES if role not in roles]
    if missing_roles:
        raise ValueError(
            f"Model {model_key} is missing required active model_workbook_sources roles: "
            f"{', '.join(missing_roles)}"
        )


def _require_generation_variants(wb, model_key: str, model_row: dict[str, str]) -> None:
    expected_variant_count = intish(model_row.get("expected_variant_count"), 0)
    if expected_variant_count <= 0:
        raise ValueError(f"Model {model_key} requires model_master.expected_variant_count greater than 0.")

    rows = _active_exact_model_rows(wb, "model_variants", model_key)
    if not rows:
        raise ValueError(f"Model {model_key} requires active model_variants rows.")

    variant_ids = [clean(row.get("variant_id")) for row in rows if clean(row.get("variant_id"))]
    duplicate_variants = _duplicate_values(variant_ids)
    if duplicate_variants:
        raise ValueError(
            f"Duplicate active model_variants rows for {model_key}: {', '.join(duplicate_variants)}"
        )
    if len(variant_ids) != expected_variant_count:
        raise ValueError(
            f"Model {model_key} expected {expected_variant_count} active model_variants rows; "
            f"found {len(variant_ids)}."
        )


def discover_generation_model_configs(workbook_path: Path = WORKBOOK_PATH) -> dict[str, ModelConfig]:
    """Return workbook-discovered configs for active/generatable models.

    Discovery is stricter than runtime compatibility fallback metadata: a model
    is generatable only when it has an active model_master row, complete
    exact-match active source roles, and exact-match active variant rows whose
    count equals model_master.expected_variant_count.
    """

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        configs: dict[str, ModelConfig] = {}
        for model_row in _active_model_master_rows(wb):
            model_key = clean(model_row.get("model_key")).lower()
            _require_generation_sources(wb, model_key)
            _require_generation_variants(wb, model_key, model_row)
            configs[model_key] = load_model_config_overrides(wb, base_model_config(model_key))
        return configs
    finally:
        wb.close()
