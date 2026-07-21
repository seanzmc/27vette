"""Read-only customer-facing quality lint for workbook option sheets."""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from corvette_form_generator.runtime_metadata import truthy
from corvette_form_generator.workbook import clean


ALLOWLIST_SCHEMA_VERSION = "options-sheet-quality-allowlist-1"
HASH_OPTION_ID_RE = re.compile(r"^opt_std_[0-9a-f]{16,}$")
MAX_OPTION_NAME_LENGTH = 60
MAX_REFERENCE_STUB_COUNT = 6


@dataclass(frozen=True)
class QualityIssue:
    check_id: str
    model: str
    sheet: str
    row: int | None
    option_id: str
    rpo: str
    value: Any
    message: str


def _records(ws: Any) -> Iterable[tuple[int, dict[str, Any]]]:
    values = ws.iter_rows(values_only=True)
    try:
        headers = [clean(value) for value in next(values)]
    except StopIteration:
        return
    for row_number, raw_row in enumerate(values, start=2):
        if not any(value is not None for value in raw_row):
            continue
        row = {header: value for header, value in zip(headers, raw_row) if header}
        if any(value is not None for value in row.values()):
            yield row_number, row


def _source_option_sheets(wb: Any) -> list[tuple[str, str]]:
    if "model_workbook_sources" not in wb.sheetnames:
        raise ValueError("Workbook is missing model_workbook_sources")
    result: list[tuple[str, str]] = []
    for _, row in _records(wb["model_workbook_sources"]):
        if clean(row.get("source_role")) != "source_option_sheet":
            continue
        model = clean(row.get("model_key")).lower()
        sheet = clean(row.get("sheet_name"))
        if not model or not sheet:
            continue
        if sheet not in wb.sheetnames:
            raise ValueError(f"Configured option sheet is missing: {model}/{sheet}")
        result.append((model, sheet))
    if not result:
        raise ValueError("Workbook has no configured source_option_sheet rows")
    return result


def _section_modes(wb: Any) -> dict[str, str]:
    if "section_master" not in wb.sheetnames:
        raise ValueError("Workbook is missing section_master")
    return {
        clean(row.get("section_id")): clean(row.get("selection_mode"))
        for _, row in _records(wb["section_master"])
        if clean(row.get("section_id"))
    }


def _load_allowlist(path: str | Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schemaVersion") != ALLOWLIST_SCHEMA_VERSION:
        raise ValueError("Unsupported options-sheet quality allowlist schema")
    entries = payload.get("entries")
    if not isinstance(entries, list):
        raise ValueError("Options-sheet quality allowlist entries must be a list")
    required = {"model", "sheet", "optionId", "checkId", "value", "reason"}
    normalized: list[dict[str, Any]] = []
    for index, entry in enumerate(entries):
        if not isinstance(entry, Mapping) or set(entry) != required:
            raise ValueError(f"Invalid options-sheet quality allowlist entry {index}")
        if not clean(entry.get("reason")):
            raise ValueError(f"Options-sheet quality allowlist entry {index} needs a reason")
        normalized.append(dict(entry))
    return normalized


def _is_allowed(issue: QualityIssue, entries: Iterable[Mapping[str, Any]]) -> bool:
    return any(
        clean(entry.get("model")).lower() == issue.model
        and clean(entry.get("sheet")) == issue.sheet
        and clean(entry.get("optionId")) == issue.option_id
        and clean(entry.get("checkId")) == issue.check_id
        and entry.get("value") == issue.value
        for entry in entries
    )


def _issue(
    check_id: str,
    model: str,
    sheet: str,
    row_number: int | None,
    row: Mapping[str, Any],
    value: Any,
    message: str,
) -> QualityIssue:
    return QualityIssue(
        check_id=check_id,
        model=model,
        sheet=sheet,
        row=row_number,
        option_id=clean(row.get("option_id")),
        rpo=clean(row.get("rpo")).upper(),
        value=value,
        message=message,
    )


def _row_issues(
    model: str,
    sheet: str,
    row_number: int,
    row: Mapping[str, Any],
    section_modes: Mapping[str, str],
) -> list[QualityIssue]:
    issues: list[QualityIssue] = []
    option_name = clean(row.get("option_name"))
    description = clean(row.get("description"))
    detail_raw = clean(row.get("detail_raw"))
    option_id = clean(row.get("option_id"))
    active = truthy(row.get("active"))
    selectable = truthy(row.get("selectable"))
    display_order = clean(row.get("display_order"))
    raw_price = row.get("price")
    price = None if not clean(raw_price) else float(clean(raw_price))
    section_mode = section_modes.get(clean(row.get("section_id")), "")

    checks = [
        (
            bool(option_name and description and option_name == description),
            "option_name_equals_description",
            option_name,
            "option_name duplicates description",
        ),
        (
            bool(description and detail_raw and description == detail_raw),
            "description_equals_detail_raw",
            description,
            "description duplicates detail_raw",
        ),
        (
            "\n" in option_name or "\r" in option_name,
            "option_name_multiline",
            option_name,
            "option_name contains a line break",
        ),
        (
            len(option_name) > MAX_OPTION_NAME_LENGTH,
            "option_name_too_long",
            option_name,
            f"option_name exceeds {MAX_OPTION_NAME_LENGTH} characters",
        ),
        (
            option_name.upper() == "LPO",
            "bare_lpo_option_name",
            option_name,
            "option_name is the non-identifying label LPO",
        ),
        (
            bool(HASH_OPTION_ID_RE.match(option_id)),
            "hash_derived_option_id",
            option_id,
            "option_id uses the forbidden hash-derived no-RPO format",
        ),
        (
            active and not display_order,
            "active_option_missing_display_order",
            row.get("display_order"),
            "active option has no display_order",
        ),
    ]
    for failed, check_id, value, message in checks:
        if failed:
            issues.append(_issue(check_id, model, sheet, row_number, row, value, message))

    if not selectable:
        if price is not None and price != 0:
            issues.append(
                _issue(
                    "standard_option_nonzero_price",
                    model,
                    sheet,
                    row_number,
                    row,
                    raw_price,
                    "non-selectable standard row carries a nonzero price",
                )
            )
        elif section_mode == "display_only" and price is not None:
            issues.append(
                _issue(
                    "display_only_standard_has_price",
                    model,
                    sheet,
                    row_number,
                    row,
                    raw_price,
                    "display-only standard row must have a blank price",
                )
            )
        elif section_mode and section_mode != "display_only" and price is None:
            issues.append(
                _issue(
                    "selectable_section_standard_missing_zero_price",
                    model,
                    sheet,
                    row_number,
                    row,
                    None,
                    "standard row in a selectable section must use price 0 or an exact reviewed exception",
                )
            )
    return issues


def lint_options_sheet_quality(
    workbook_path: str | Path,
    *,
    allowlist_path: str | Path | None = None,
) -> list[QualityIssue]:
    workbook_path = Path(workbook_path)
    allowlist = _load_allowlist(allowlist_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        section_modes = _section_modes(wb)
        issues: list[QualityIssue] = []
        for model, sheet in _source_option_sheets(wb):
            short_rows: list[tuple[int, Mapping[str, Any]]] = []
            for row_number, row in _records(wb[sheet]):
                issues.extend(_row_issues(model, sheet, row_number, row, section_modes))
                option_name = clean(row.get("option_name"))
                if option_name and len(option_name) <= 12:
                    short_issue = _issue(
                        "short_option_name",
                        model,
                        sheet,
                        row_number,
                        row,
                        option_name,
                        "short option name",
                    )
                    if not _is_allowed(short_issue, allowlist):
                        short_rows.append((row_number, row))
            if len(short_rows) > MAX_REFERENCE_STUB_COUNT:
                issues.append(
                    QualityIssue(
                        check_id="stub_name_count_exceeds_reference_band",
                        model=model,
                        sheet=sheet,
                        row=None,
                        option_id="",
                        rpo="",
                        value=len(short_rows),
                        message=(
                            f"{len(short_rows)} unallowlisted option names are 12 characters or shorter; "
                            f"reference maximum is {MAX_REFERENCE_STUB_COUNT}"
                        ),
                    )
                )
        return [issue for issue in issues if not _is_allowed(issue, allowlist)]
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--allowlist", type=Path)
    parser.add_argument("--json", action="store_true", dest="as_json")
    args = parser.parse_args(argv)

    issues = lint_options_sheet_quality(args.workbook, allowlist_path=args.allowlist)
    if args.as_json:
        print(
            json.dumps(
                {
                    "status": "failed" if issues else "passed",
                    "workbook": str(args.workbook),
                    "allowlist": str(args.allowlist) if args.allowlist else None,
                    "issueCount": len(issues),
                    "issues": [asdict(issue) for issue in issues],
                },
                indent=2,
                sort_keys=True,
                ensure_ascii=False,
            )
        )
    else:
        for issue in issues:
            location = f"{issue.sheet}:{issue.row}" if issue.row else issue.sheet
            identity = issue.rpo or issue.option_id or "sheet"
            print(f"{issue.check_id} {location} {identity}: {issue.message}")
        print(f"options-sheet quality: {'FAILED' if issues else 'PASSED'} ({len(issues)} issues)")
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
