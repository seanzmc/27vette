"""Read-only workbook schema validation for Corvette form source sheets."""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from corvette_form_generator.contract import load_model_asset_map
from corvette_form_generator.model_configs import OPTIONAL_GENERATION_SOURCE_ROLES, REQUIRED_GENERATION_SOURCE_ROLES
from corvette_form_generator.registry_promotion import (
    build_registry_from_artifacts,
    parse_app_data_registry,
    registry_model_key,
)
from corvette_form_generator.workbook import workbook_truthy


BOOLEAN_COLUMNS: dict[str, tuple[str, ...]] = {
    "model_registry_promotion": ("promoted_to_runtime", "default_model", "active"),
    # NOTE: LZ_Interiors / model_interior_scope / interior_components intentionally
    # excluded. Their bool cell-typing is cosmetic (generator coerces; data.js is
    # byte-identical with or without it). Model source sheets are added by source role.
}

PRICE_COLUMNS: dict[str, tuple[str, ...]] = {
    "PriceRef": ("Price",),
}

RPO_COLUMNS: dict[str, tuple[str, ...]] = {
    "interior_components": ("rpo",),
}

MODEL_SOURCE_ROLES = frozenset(REQUIRED_GENERATION_SOURCE_ROLES + OPTIONAL_GENERATION_SOURCE_ROLES)

ROLE_BOOLEAN_COLUMNS: dict[str, tuple[str, ...]] = {
    "source_option_sheet": ("selectable", "active"),
    "rule_groups_sheet": ("active",),
    "rule_group_members_sheet": ("active",),
    "exclusive_groups_sheet": ("active",),
    "exclusive_group_members_sheet": ("active",),
    "variant_option_overrides_sheet": ("active", "selectable"),
    "interior_source_sheet": ("active_for_stingray", "requires_r6x"),
}

ROLE_PRICE_COLUMNS: dict[str, tuple[str, ...]] = {
    "source_option_sheet": ("price",),
    "price_rules_sheet": ("price_value",),
    "interior_source_sheet": ("Price",),
}

ROLE_RPO_COLUMNS: dict[str, tuple[str, ...]] = {
    "source_option_sheet": ("rpo",),
}

HEADER_MATCH_ROLES: tuple[str, ...] = (
    "source_option_sheet",
    "status_sheet",
    "rule_mapping_sheet",
    "price_rules_sheet",
    "rule_groups_sheet",
    "rule_group_members_sheet",
    "exclusive_groups_sheet",
    "exclusive_group_members_sheet",
    "interior_source_sheet",
)

REQUIRED_SHEETS: tuple[str, ...] = (
    "model_master",
    "model_workbook_sources",
    "model_variants",
    "model_registry_promotion",
    "variant_master",
    "section_master",
    "lt_interiors",
    "LZ_Interiors",
    "model_interior_scope",
    "interior_components",
    "PriceRef",
)

DRAFT_ONLY_CHOICE_FIELDS: set[str] = {"source_option_name", "source_description", "text_cleanup_notes"}
DRAFT_ONLY_PROVENANCE_FIELDS: set[str] = {
    "draftMetadata",
    "copy_from_model_key",
    "suggested_copy_from",
    "raw_source_sheet",
    "raw_source_sheets",
    "review_status",
    "review_flags",
}
DRAFT_ONLY_LIVE_CONTRACT_FIELDS: set[str] = DRAFT_ONLY_CHOICE_FIELDS | DRAFT_ONLY_PROVENANCE_FIELDS
RUNTIME_CHOICE_ROW_TRIM_FIELDS: set[str] = {
    "source_detail_raw",
    "choice_mode",
    "selection_mode",
    "selection_mode_label",
}
RUNTIME_STANDARD_EQUIPMENT_ROW_TRIM_FIELDS: set[str] = {"source_detail_raw"}
FORBIDDEN_LIVE_LINEAGE_VALUE_TOKENS: tuple[str, ...] = ("grand_sport:",)
GENERATED_TIMESTAMP_KEYS: frozenset[str] = frozenset(("generated_at", "sourceGeneratedAt", "generatedAt"))

MODEL_MASTER_HEADERS: tuple[str, ...] = (
    "model_key",
    "registry_key",
    "model_label",
    "model_year",
    "dataset_name",
    "export_slug",
    "expected_variant_count",
    "default_model",
    "active",
    "setup_card_subtitle",
    "setup_eyebrow",
    "setup_title",
    "setup_description",
    "setup_fact_1",
    "setup_fact_2",
    "setup_fact_3",
    "notes",
)
MODEL_SETUP_COPY_FIELDS: tuple[str, ...] = MODEL_MASTER_HEADERS[9:16]

MODEL_REGISTRY_PROMOTION_HEADERS: tuple[str, ...] = (
    "model_key",
    "registry_key",
    "promoted_to_runtime",
    "default_model",
    "artifact_path",
    "artifact_type",
    "legacy_alias",
    "active",
    "display_order",
    "notes",
)
VALID_REGISTRY_PROMOTION_ARTIFACT_TYPES: set[str] = {"current_generation", "draft_artifact", "runtime_contract"}
DEFAULT_SELECTION_DISPLAY_BEHAVIORS: frozenset[str] = frozenset(("default_selected",))


@dataclass
class SchemaIssue:
    severity: str
    check_id: str
    sheet: str = ""
    row: int | None = None
    column: str = ""
    value: Any = None
    message: str = ""


def nonblank_headers(ws) -> list[str]:
    return [str(value).strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if value]


def header_index(ws) -> dict[str, int]:
    return {
        str(value).strip(): index
        for index, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if value
    }


def records(ws) -> Iterable[tuple[int, dict[str, Any]]]:
    headers = [str(value).strip() if value else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {header: value for header, value in zip(headers, values) if header}
        if any(value is not None for value in record.values()):
            yield row_number, record


def add_issue(
    issues: list[SchemaIssue],
    severity: str,
    check_id: str,
    *,
    sheet: str = "",
    row: int | None = None,
    column: str = "",
    value: Any = None,
    message: str,
) -> None:
    issues.append(
        SchemaIssue(
            severity=severity,
            check_id=check_id,
            sheet=sheet,
            row=row,
            column=column,
            value=value,
            message=message,
        )
    )


def runtime_payload_trim_fields(path: tuple[str, ...]) -> set[str]:
    if path == ("choices", "[]"):
        return RUNTIME_CHOICE_ROW_TRIM_FIELDS
    if path == ("standardEquipment", "[]"):
        return RUNTIME_STANDARD_EQUIPMENT_ROW_TRIM_FIELDS
    return set()


def json_path_parts(path: str) -> tuple[str, ...]:
    if path == "$":
        return ()
    parts: list[str] = []
    for part in path.removeprefix("$.").split("."):
        if "[" in part and part.endswith("]"):
            name = part.split("[", 1)[0]
            if name:
                parts.append(name)
            parts.append("[]")
        else:
            parts.append(part)
    return tuple(parts)


def live_contract_provenance_leaks(value: Any, path: str = "$") -> Iterable[tuple[str, str, Any]]:
    if isinstance(value, dict):
        runtime_payload_fields = runtime_payload_trim_fields(json_path_parts(path))
        for key, child in value.items():
            child_path = f"{path}.{key}"
            if key in DRAFT_ONLY_LIVE_CONTRACT_FIELDS or key in runtime_payload_fields:
                yield (child_path, "field", child)
                continue
            yield from live_contract_provenance_leaks(child, child_path)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            yield from live_contract_provenance_leaks(child, f"{path}[{index}]")
    elif isinstance(value, str):
        normalized = value.lower()
        for token in FORBIDDEN_LIVE_LINEAGE_VALUE_TOKENS:
            if token in normalized:
                yield (path, "value", value)
                break


def without_generated_timestamps(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            key: without_generated_timestamps(child)
            for key, child in value.items()
            if key not in GENERATED_TIMESTAMP_KEYS
        }
    if isinstance(value, list):
        return [without_generated_timestamps(item) for item in value]
    return value


def validate_app_registry_freshness(wb, workbook: Path, issues: list[SchemaIssue]) -> None:
    app_data_path = workbook.parent / "form-app" / "data.js"
    if not app_data_path.exists():
        return
    try:
        actual_registry = parse_app_data_registry(app_data_path)
        expected_registry = build_registry_from_artifacts(
            wb,
            model_assets=load_model_asset_map(wb, registry_model_key),
            root=workbook.parent,
        )
        expected_registry.pop("legacyAliases", None)
    except (FileNotFoundError, ValueError, RuntimeError, json.JSONDecodeError) as exc:
        add_issue(
            issues,
            "error",
            "app_registry_freshness_check_failed",
            sheet="form-app/data.js",
            message=f"Could not validate app registry freshness: {exc}",
        )
        return
    actual_registry_for_compare = without_generated_timestamps(actual_registry)
    expected_registry_for_compare = without_generated_timestamps(expected_registry)
    if actual_registry_for_compare != expected_registry_for_compare:
        stale_models: list[str] = []
        actual_models = actual_registry_for_compare.get("models", {}) if isinstance(actual_registry_for_compare, dict) else {}
        expected_models = expected_registry_for_compare.get("models", {}) if isinstance(expected_registry_for_compare, dict) else {}
        for model_key in sorted(set(actual_models) | set(expected_models)):
            if actual_models.get(model_key) != expected_models.get(model_key):
                stale_models.append(model_key)
        add_issue(
            issues,
            "error",
            "app_registry_stale",
            sheet="form-app/data.js",
            value={"stale_models": stale_models, "expected_default_model": expected_registry.get("defaultModelKey")},
            message="form-app/data.js is stale relative to promoted runtime artifacts; run scripts/generate_registry.py.",
        )


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def option_ids(wb, sheet: str) -> set[str]:
    if sheet not in wb.sheetnames:
        return set()
    return {str(row.get("option_id")) for _, row in records(wb[sheet]) if row.get("option_id")}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def display_order_key(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if isinstance(value, bool):
        return text
    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else str(value)
    try:
        numeric = float(text)
    except ValueError:
        return text
    return str(int(numeric)) if numeric.is_integer() else text


def intish(value: Any, default: int = 0) -> int:
    text = clean_text(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def truthy(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = clean_text(value).lower()
    if not text:
        return default
    if text in {"1", "true", "t", "yes", "y", "on", "active", "enabled"}:
        return True
    if text in {"0", "false", "f", "no", "n", "off", "inactive", "disabled"}:
        return False
    return default


def active_metadata_rows(wb, sheet: str, model_key: str | None = None) -> list[dict[str, Any]]:
    if sheet not in wb.sheetnames:
        return []
    rows = [row for _, row in records(wb[sheet]) if truthy(row.get("active"), default=True)]
    if model_key is None:
        return rows
    return [row for row in rows if clean_text(row.get("model_key")).lower() in {model_key.lower(), "*", "all", "shared"}]


def metadata_source_graph(wb, issues: list[SchemaIssue]) -> dict[str, dict[str, str]]:
    """Return exact-match active model source-role mapping from workbook metadata."""

    active_models = {clean_text(row.get("model_key")).lower() for row in active_metadata_rows(wb, "model_master")}
    active_models.discard("")
    graph: dict[str, dict[str, str]] = {model_key: {} for model_key in active_models}

    if "model_workbook_sources" not in wb.sheetnames:
        return graph

    seen: set[tuple[str, str]] = set()
    for row_number, row in records(wb["model_workbook_sources"]):
        if not truthy(row.get("active"), default=True):
            continue
        model_key = clean_text(row.get("model_key")).lower()
        source_role = clean_text(row.get("source_role"))
        sheet_name = clean_text(row.get("sheet_name"))
        if not model_key or not source_role:
            continue
        if model_key not in active_models:
            continue
        if source_role not in MODEL_SOURCE_ROLES:
            add_issue(
                issues,
                "error",
                "unknown_model_source_role",
                sheet="model_workbook_sources",
                row=row_number,
                column="source_role",
                value=source_role,
                message=f"Unknown model_workbook_sources source_role {source_role!r}.",
            )
            continue
        if not sheet_name:
            continue
        duplicate_key = (model_key, source_role)
        if duplicate_key in seen:
            add_issue(
                issues,
                "error",
                "duplicate_model_source_role",
                sheet="model_workbook_sources",
                row=row_number,
                column="source_role",
                value={"model_key": model_key, "source_role": source_role},
                message=f"Duplicate active source role {source_role!r} for model {model_key!r}.",
            )
            continue
        seen.add(duplicate_key)
        graph.setdefault(model_key, {})[source_role] = sheet_name

    for model_key in sorted(active_models):
        sources = graph.setdefault(model_key, {})
        for source_role in REQUIRED_GENERATION_SOURCE_ROLES:
            if source_role in sources:
                continue
            add_issue(
                issues,
                "error",
                "missing_model_source_role",
                sheet="model_workbook_sources",
                column="source_role",
                value={"model_key": model_key, "source_role": source_role},
                message=f"Active model {model_key!r} is missing required model_workbook_sources role {source_role!r}.",
            )

    for model_key, sources in graph.items():
        for source_role, sheet_name in sources.items():
            if sheet_name and sheet_name not in wb.sheetnames:
                add_issue(
                    issues,
                    "error",
                    "missing_model_source_sheet",
                    sheet=sheet_name,
                    value={"model_key": model_key, "source_role": source_role},
                    message=f"Active {model_key}.{source_role} sheet {sheet_name!r} does not exist.",
                )
    return graph


def sheets_by_role(source_graph: dict[str, dict[str, str]]) -> dict[str, list[str]]:
    by_role: dict[str, list[str]] = {}
    for sources in source_graph.values():
        for source_role, sheet_name in sources.items():
            if not sheet_name:
                continue
            by_role.setdefault(source_role, [])
            if sheet_name not in by_role[source_role]:
                by_role[source_role].append(sheet_name)
    return by_role


def validate_model_master_metadata(wb, issues: list[SchemaIssue]) -> bool:
    if "model_master" not in wb.sheetnames:
        add_issue(
            issues,
            "error",
            "missing_required_sheet",
            sheet="model_master",
            message="Missing required sheet model_master.",
        )
        return False

    headers = nonblank_headers(wb["model_master"])
    if headers != list(MODEL_MASTER_HEADERS):
        add_issue(
            issues,
            "error",
            "model_master_header_drift",
            sheet="model_master",
            value={"expected": list(MODEL_MASTER_HEADERS), "actual": headers},
            message="model_master headers must match the workbook-owned model metadata contract.",
        )
        return False

    seen_active_model_keys: dict[str, int] = {}
    for row_number, row in records(wb["model_master"]):
        if not truthy(row.get("active"), default=True):
            continue
        model_key = clean_text(row.get("model_key")).lower()
        if not model_key:
            continue
        first_row = seen_active_model_keys.get(model_key)
        if first_row is not None:
            add_issue(
                issues,
                "error",
                "duplicate_active_model_master_row",
                sheet="model_master",
                row=row_number,
                column="model_key",
                value={"model_key": model_key, "first_row": first_row, "duplicate_row": row_number},
                message=f"Duplicate active model_master row for model_key {model_key!r}.",
            )
            continue
        seen_active_model_keys[model_key] = row_number
    return True


def validate_asset_map_uniqueness(wb, issues: list[SchemaIssue]) -> None:
    """Validate active asset_map rows are unique per (model_key, target_type, target_id).

    Mirrors validate_model_master_metadata's active-row uniqueness pattern: the runtime read
    path (contract.py:load_asset_map) keys assets by (target_type, target_id) per model and only
    loads rows where workbook_truthy(active) is true, so a duplicate active row for the same
    identity is silently last-write-wins at load time with no error. This check makes that
    failure loud instead of silent while preserving the runtime's inactive/blank-row semantics.
    """

    if "asset_map" not in wb.sheetnames:
        return

    header_cols = nonblank_headers(wb["asset_map"])
    required = {"model_key", "target_type", "target_id"}
    if not required.issubset(header_cols):
        return

    seen_active_keys: dict[tuple[str, str, str], int] = {}
    for row_number, row in records(wb["asset_map"]):
        if not workbook_truthy(row.get("active")):
            continue
        model_key = clean_text(row.get("model_key")).lower()
        target_type = clean_text(row.get("target_type")).lower()
        target_id = clean_text(row.get("target_id"))
        target_id_key = target_id.lower() if target_type == "option" else target_id
        if not model_key or not target_type or not target_id_key:
            continue
        key = (model_key, target_type, target_id_key)
        first_row = seen_active_keys.get(key)
        if first_row is not None:
            add_issue(
                issues,
                "error",
                "duplicate_active_asset_map_row",
                sheet="asset_map",
                row=row_number,
                column="target_id",
                value={
                    "model_key": model_key,
                    "target_type": target_type,
                    "target_id": target_id,
                    "first_row": first_row,
                    "duplicate_row": row_number,
                },
                message=(
                    f"Duplicate active asset_map row for model_key {model_key!r}, "
                    f"target_type {target_type!r}, target_id {target_id!r}."
                ),
            )
            continue
        seen_active_keys[key] = row_number


def validate_asset_map_wildcard_rows(wb, issues: list[SchemaIssue]) -> None:
    """Validate active wildcard asset_map rows are option targets only.

    Mirrors validate_asset_map_uniqueness's active-row walk: the runtime read
    path (contract.py:load_asset_map) supports model_key="*" wildcard rows for
    option targets only (wildcard-first, exact-model overlay). Wildcard rows
    with target_type model or context_choice would be silently ignored by
    load_model_asset_map or apply model-specific hover media to every model,
    so reject them loudly here instead.
    """

    if "asset_map" not in wb.sheetnames:
        return

    header_cols = nonblank_headers(wb["asset_map"])
    required = {"model_key", "target_type", "target_id"}
    if not required.issubset(header_cols):
        return

    for row_number, row in records(wb["asset_map"]):
        if not workbook_truthy(row.get("active")):
            continue
        if clean_text(row.get("model_key")) != "*":
            continue
        target_type = clean_text(row.get("target_type")).lower()
        if target_type == "option":
            continue
        add_issue(
            issues,
            "error",
            "invalid_wildcard_asset_map_row",
            sheet="asset_map",
            row=row_number,
            column="target_type",
            value={
                "model_key": "*",
                "target_type": target_type,
                "target_id": clean_text(row.get("target_id")),
            },
            message=(
                f"Active wildcard asset_map row (model_key '*') has target_type "
                f"{target_type!r}; wildcard rows are supported for target_type 'option' only."
            ),
        )


def validate_default_selection_display_behavior(wb, issues: list[SchemaIssue]) -> None:
    """Validate workbook-authored default-selection display behavior values.

    ``default_selection_rules.display_behavior`` is a workbook-only authoring
    signal consumed by runtime_metadata.load_default_selection_display_rules().
    The loader intentionally recognizes only ``default_selected``; reject typos
    here so a workbook authoring mistake cannot silently suppress display
    derivation.
    """

    sheet_name = "default_selection_rules"
    if sheet_name not in wb.sheetnames:
        return

    headers = header_index(wb[sheet_name])
    if "display_behavior" not in headers:
        return

    for row_number, row in records(wb[sheet_name]):
        value = clean_text(row.get("display_behavior"))
        if not value:
            continue
        if value in DEFAULT_SELECTION_DISPLAY_BEHAVIORS:
            continue
        add_issue(
            issues,
            "error",
            "invalid_default_selection_display_behavior",
            sheet=sheet_name,
            row=row_number,
            column="display_behavior",
            value=value,
            message=(
                "default_selection_rules.display_behavior must be blank or one of "
                f"{sorted(DEFAULT_SELECTION_DISPLAY_BEHAVIORS)!r}."
            ),
        )


def validate_model_variant_topology(wb, issues: list[SchemaIssue]) -> None:
    """Validate active model membership references active variant fact rows."""

    if not all(sheet in wb.sheetnames for sheet in ("model_master", "model_variants", "variant_master")):
        return

    active_models: dict[str, dict[str, Any]] = {}
    for _, row in records(wb["model_master"]):
        if not truthy(row.get("active"), default=True):
            continue
        model_key = clean_text(row.get("model_key")).lower()
        if model_key:
            active_models[model_key] = row

    variant_master: dict[str, tuple[int, dict[str, Any]]] = {}
    for row_number, row in records(wb["variant_master"]):
        variant_id = clean_text(row.get("variant_id"))
        if variant_id and variant_id not in variant_master:
            variant_master[variant_id] = (row_number, row)

    active_variant_counts: dict[str, int] = {model_key: 0 for model_key in active_models}
    display_orders: dict[tuple[str, str], tuple[int, str]] = {}

    for row_number, row in records(wb["model_variants"]):
        if not truthy(row.get("active"), default=True):
            continue
        model_key = clean_text(row.get("model_key")).lower()
        if model_key not in active_models:
            continue
        variant_id = clean_text(row.get("variant_id"))
        active_variant_counts[model_key] += 1

        if variant_id:
            source = variant_master.get(variant_id)
            if source is None:
                add_issue(
                    issues,
                    "error",
                    "model_variant_unknown_variant_master",
                    sheet="model_variants",
                    row=row_number,
                    column="variant_id",
                    value={"model_key": model_key, "variant_id": variant_id},
                    message=(
                        f"Active model_variants row for active model {model_key!r} references missing "
                        f"variant_master row {variant_id!r}."
                    ),
                )
            elif not truthy(source[1].get("active"), default=True):
                add_issue(
                    issues,
                    "error",
                    "model_variant_inactive_variant_master",
                    sheet="model_variants",
                    row=row_number,
                    column="variant_id",
                    value={"model_key": model_key, "variant_id": variant_id, "variant_master_row": source[0]},
                    message=(
                        f"Active model_variants row for active model {model_key!r} references inactive "
                        f"variant_master row {variant_id!r}."
                    ),
                )

        display_order = display_order_key(row.get("display_order"))
        if not display_order:
            continue
        order_key = (model_key, display_order)
        previous = display_orders.get(order_key)
        if previous:
            previous_row, previous_variant = previous
            add_issue(
                issues,
                "error",
                "duplicate_model_variant_display_order",
                sheet="model_variants",
                row=row_number,
                column="display_order",
                value={
                    "model_key": model_key,
                    "display_order": display_order,
                    "first_row": previous_row,
                    "first_variant_id": previous_variant,
                    "duplicate_row": row_number,
                    "duplicate_variant_id": variant_id,
                },
                message=(
                    f"Active model_variants rows for model {model_key!r} duplicate display_order "
                    f"{display_order!r}."
                ),
            )
        else:
            display_orders[order_key] = (row_number, variant_id)

    for model_key, row in active_models.items():
        expected_variant_count = intish(row.get("expected_variant_count"), 0)
        if expected_variant_count <= 0:
            continue
        actual_count = active_variant_counts.get(model_key, 0)
        if actual_count == expected_variant_count:
            continue
        add_issue(
            issues,
            "error",
            "model_variant_count_mismatch",
            sheet="model_variants",
            column="active",
            value={
                "model_key": model_key,
                "expected_variant_count": expected_variant_count,
                "active_model_variant_count": actual_count,
            },
            message=(
                f"Active model {model_key!r} expected {expected_variant_count} active model_variants rows; "
                f"found {actual_count}."
            ),
        )


def validate_registry_promotion_metadata(wb, issues: list[SchemaIssue]) -> None:
    if "model_registry_promotion" not in wb.sheetnames:
        return
    if "model_master" not in wb.sheetnames:
        return

    headers = nonblank_headers(wb["model_registry_promotion"])
    if headers != list(MODEL_REGISTRY_PROMOTION_HEADERS):
        add_issue(
            issues,
            "error",
            "registry_promotion_header_drift",
            sheet="model_registry_promotion",
            value={"expected": list(MODEL_REGISTRY_PROMOTION_HEADERS), "actual": headers},
            message="model_registry_promotion headers must match the workbook-owned runtime promotion contract.",
        )

    model_rows = {
        clean_text(row.get("model_key")).lower(): row
        for _, row in records(wb["model_master"])
        if clean_text(row.get("model_key"))
    }
    promoted_rows: list[tuple[int, dict[str, Any]]] = []
    seen_registry_keys: set[str] = set()
    for row_number, row in records(wb["model_registry_promotion"]):
        if not truthy(row.get("active"), default=True) or not truthy(row.get("promoted_to_runtime"), default=False):
            continue
        promoted_rows.append((row_number, row))
        model_key = clean_text(row.get("model_key")).lower()
        registry_key = clean_text(row.get("registry_key"))
        artifact_path = clean_text(row.get("artifact_path"))
        artifact_type = clean_text(row.get("artifact_type")) or "draft_artifact"
        if registry_key in seen_registry_keys:
            add_issue(
                issues,
                "error",
                "registry_promotion_duplicate_registry_key",
                sheet="model_registry_promotion",
                row=row_number,
                column="registry_key",
                value=registry_key,
                message=f"Duplicate promoted runtime registry_key {registry_key!r}.",
            )
        seen_registry_keys.add(registry_key)
        model_row = model_rows.get(model_key)
        if not model_row:
            add_issue(
                issues,
                "error",
                "registry_promotion_unknown_model",
                sheet="model_registry_promotion",
                row=row_number,
                column="model_key",
                value=model_key,
                message=f"Promoted runtime model {model_key!r} is not present in model_master.",
            )
            continue
        expected_registry_key = clean_text(model_row.get("registry_key")) or model_key
        if registry_key != expected_registry_key:
            add_issue(
                issues,
                "error",
                "registry_promotion_registry_key_mismatch",
                sheet="model_registry_promotion",
                row=row_number,
                column="registry_key",
                value={"model_key": model_key, "registry_key": registry_key, "expected": expected_registry_key},
                message=f"Promoted runtime registry_key for {model_key!r} must match model_master.registry_key {expected_registry_key!r}.",
            )
        model_is_active = truthy(model_row.get("active"), default=True)
        if not model_is_active:
            add_issue(
                issues,
                "error",
                "registry_promotion_inactive_model",
                sheet="model_registry_promotion",
                row=row_number,
                column="model_key",
                value=model_key,
                message=f"Promoted runtime model {model_key!r} is inactive in model_master.",
            )
        if model_is_active:
            for field in MODEL_SETUP_COPY_FIELDS:
                if clean_text(model_row.get(field)):
                    continue
                add_issue(
                    issues,
                    "error",
                    "promoted_model_setup_copy_incomplete",
                    sheet="model_master",
                    column=field,
                    value={"model_key": model_key, "field": field},
                    message=f"Promoted runtime model {model_key!r} requires nonblank {field}.",
                )
        if artifact_type not in VALID_REGISTRY_PROMOTION_ARTIFACT_TYPES:
            add_issue(
                issues,
                "error",
                "registry_promotion_unknown_artifact_type",
                sheet="model_registry_promotion",
                row=row_number,
                column="artifact_type",
                value=artifact_type,
                message=f"Unsupported runtime promotion artifact_type {artifact_type!r}.",
            )
        if artifact_type != "current_generation" and not artifact_path:
            add_issue(
                issues,
                "error",
                "registry_promotion_missing_artifact_path",
                sheet="model_registry_promotion",
                row=row_number,
                column="artifact_path",
                value={"model_key": model_key, "artifact_type": artifact_type},
                message="Promoted non-current-generation runtime models must provide artifact_path.",
            )

    default_count = sum(1 for _, row in promoted_rows if truthy(row.get("default_model"), default=False))
    if promoted_rows and default_count != 1:
        add_issue(
            issues,
            "error",
            "registry_promotion_default_count",
            sheet="model_registry_promotion",
            value={"promoted_rows": len(promoted_rows), "default_count": default_count},
            message="model_registry_promotion must have exactly one active promoted default model.",
        )


def validate_option_display_order_uniqueness(
    wb,
    option_sheets: Iterable[str],
    issues: list[SchemaIssue],
    *,
    check_id: str = "duplicate_option_display_order",
) -> None:
    for sheet in dict.fromkeys(option_sheets):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = header_index(ws)
        if not all(column in headers for column in ("option_id", "section_id", "display_order")):
            continue
        buckets: dict[tuple[str, str], list[tuple[int, str]]] = {}
        for row_number, row in records(ws):
            if "active" in headers and not truthy(row.get("active"), default=True):
                continue
            section_id = clean_text(row.get("section_id"))
            display_order = display_order_key(row.get("display_order"))
            option_id = clean_text(row.get("option_id"))
            if not section_id or not display_order or not option_id:
                continue
            buckets.setdefault((section_id, display_order), []).append((row_number, option_id))
        for (section_id, display_order), matches in sorted(buckets.items()):
            if len(matches) < 2:
                continue
            add_issue(
                issues,
                "error",
                check_id,
                sheet=sheet,
                row=matches[0][0],
                column="display_order",
                value={
                    "section_id": section_id,
                    "display_order": display_order,
                    "matches": [{"row": row_number, "option_id": option_id} for row_number, option_id in matches],
                },
                message=(
                    f"{sheet} section {section_id!r} has duplicate option display_order {display_order!r}; "
                    "use deterministic unique ordering, including standard/included sections."
                ),
            )


def inactive_source_option_sheets(wb, active_option_sheets: Iterable[str]) -> list[str]:
    if "model_workbook_sources" not in wb.sheetnames:
        return []

    active_sheets = set(active_option_sheets)
    sheets: list[str] = []
    for _, row in records(wb["model_workbook_sources"]):
        if truthy(row.get("active"), default=True):
            continue
        if clean_text(row.get("source_role")) != "source_option_sheet":
            continue
        sheet_name = clean_text(row.get("sheet_name"))
        if not sheet_name or sheet_name in active_sheets or sheet_name not in wb.sheetnames:
            continue
        if sheet_name not in sheets:
            sheets.append(sheet_name)
    return sheets


def merge_sheet_columns(
    base: dict[str, tuple[str, ...]],
    by_role: dict[str, list[str]],
    role_columns: dict[str, tuple[str, ...]],
) -> dict[str, tuple[str, ...]]:
    merged = {sheet: tuple(columns) for sheet, columns in base.items()}
    for source_role, columns in role_columns.items():
        for sheet in by_role.get(source_role, []):
            merged[sheet] = tuple(dict.fromkeys((*merged.get(sheet, ()), *columns)))
    return merged


def validate_workbook_schema(workbook: str | Path, *, check_live_contract: bool = True) -> list[SchemaIssue]:
    workbook = Path(workbook)
    wb = load_workbook(workbook, read_only=True, data_only=True)
    issues: list[SchemaIssue] = []
    try:
        for sheet in REQUIRED_SHEETS:
            if sheet not in wb.sheetnames:
                add_issue(issues, "error", "missing_required_sheet", sheet=sheet, message=f"Missing required sheet {sheet}.")

        if "category_master" in wb.sheetnames:
            add_issue(
                issues,
                "error",
                "category_master_active",
                sheet="category_master",
                message="category_master should not be an active source sheet; historical category evidence lives in archive/stingray_archive.xlsx.",
            )

        model_master_valid = validate_model_master_metadata(wb, issues)
        source_graph = metadata_source_graph(wb, issues)
        source_sheets_by_role = sheets_by_role(source_graph)
        if model_master_valid:
            validate_model_variant_topology(wb, issues)
            validate_registry_promotion_metadata(wb, issues)

        validate_asset_map_uniqueness(wb, issues)
        validate_asset_map_wildcard_rows(wb, issues)
        validate_default_selection_display_behavior(wb, issues)

        for source_role in HEADER_MATCH_ROLES:
            existing_sheets = [sheet for sheet in source_sheets_by_role.get(source_role, []) if sheet in wb.sheetnames]
            if len(existing_sheets) < 2:
                continue
            canonical_sheet = existing_sheets[0]
            canonical_headers = nonblank_headers(wb[canonical_sheet])
            for sheet in existing_sheets[1:]:
                current_headers = nonblank_headers(wb[sheet])
                if current_headers != canonical_headers:
                    add_issue(
                        issues,
                        "error",
                        "source_role_header_drift",
                        sheet=source_role,
                        value={
                            "canonical_sheet": canonical_sheet,
                            "canonical_headers": canonical_headers,
                            "sheet": sheet,
                            "headers": current_headers,
                        },
                        message=f"All active {source_role} sheets must share headers; {sheet} differs from {canonical_sheet}.",
                    )

        if "lt_interiors" in wb.sheetnames and "LZ_Interiors" in wb.sheetnames:
            lt_headers = nonblank_headers(wb["lt_interiors"])
            lz_headers = nonblank_headers(wb["LZ_Interiors"])
            if lt_headers != lz_headers:
                add_issue(
                    issues,
                    "error",
                    "lz_interiors_header_drift",
                    sheet="LZ_Interiors",
                    value={"lt_interiors": lt_headers, "LZ_Interiors": lz_headers},
                    message="LZ_Interiors headers must exactly match lt_interiors headers.",
                )

        boolean_columns = merge_sheet_columns(BOOLEAN_COLUMNS, source_sheets_by_role, ROLE_BOOLEAN_COLUMNS)
        rpo_columns = merge_sheet_columns(RPO_COLUMNS, source_sheets_by_role, ROLE_RPO_COLUMNS)
        price_columns = merge_sheet_columns(PRICE_COLUMNS, source_sheets_by_role, ROLE_PRICE_COLUMNS)
        active_option_sheets = source_sheets_by_role.get("source_option_sheet", [])
        validate_option_display_order_uniqueness(wb, active_option_sheets, issues)
        validate_option_display_order_uniqueness(
            wb,
            inactive_source_option_sheets(wb, active_option_sheets),
            issues,
            check_id="duplicate_future_scaffold_option_display_order",
        )

        for sheet, columns in boolean_columns.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = header_index(ws)
            for column in columns:
                if column not in headers:
                    continue
                for row_number in range(2, ws.max_row + 1):
                    value = ws.cell(row_number, headers[column]).value
                    if value is None:
                        continue
                    if not isinstance(value, bool):
                        add_issue(
                            issues,
                            "error",
                            "boolean_type_drift",
                            sheet=sheet,
                            row=row_number,
                            column=column,
                            value=value,
                            message=f"{sheet}.{column} must be a real Excel boolean, not {type(value).__name__}.",
                        )

        for sheet, columns in rpo_columns.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = header_index(ws)
            for column in columns:
                if column not in headers:
                    continue
                for row_number in range(2, ws.max_row + 1):
                    value = ws.cell(row_number, headers[column]).value
                    if value is None:
                        continue
                    if not isinstance(value, str):
                        add_issue(
                            issues,
                            "error",
                            "rpo_type_drift",
                            sheet=sheet,
                            row=row_number,
                            column=column,
                            value=value,
                            message=f"{sheet}.{column} must be stored as text, including numeric-looking RPOs.",
                        )

        for sheet, columns in price_columns.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = header_index(ws)
            for column in columns:
                if column not in headers:
                    continue
                for row_number in range(2, ws.max_row + 1):
                    value = ws.cell(row_number, headers[column]).value
                    if value is None:
                        continue
                    if not is_number(value):
                        add_issue(
                            issues,
                            "error",
                            "price_type_drift",
                            sheet=sheet,
                            row=row_number,
                            column=column,
                            value=value,
                            message=f"{sheet}.{column} must be numeric or blank; blank means null/not-priced and 0 means explicit zero-price.",
                        )

        validated_ovs_pairs: set[tuple[str, str]] = set()
        for model_key, sources in source_graph.items():
            option_sheet = sources.get("source_option_sheet", "")
            ovs_sheet = sources.get("status_sheet", "")
            if not option_sheet or not ovs_sheet or (option_sheet, ovs_sheet) in validated_ovs_pairs:
                continue
            validated_ovs_pairs.add((option_sheet, ovs_sheet))
            if option_sheet not in wb.sheetnames or ovs_sheet not in wb.sheetnames:
                continue
            valid_options = option_ids(wb, option_sheet)
            for row_number, row in records(wb[ovs_sheet]):
                option_id = row.get("option_id")
                if option_id and option_id not in valid_options:
                    add_issue(
                        issues,
                        "error",
                        "ovs_unknown_option_id",
                        sheet=ovs_sheet,
                        row=row_number,
                        column="option_id",
                        value=option_id,
                        message=f"{ovs_sheet}.option_id does not resolve to {option_sheet} for model {model_key}.",
                    )

        if check_live_contract:
            app_data_path = workbook.parent / "form-app" / "data.js"
            if app_data_path.exists():
                try:
                    registry = parse_app_data_registry(app_data_path)
                    for model_key, entry in registry.get("models", {}).items():
                        data = entry.get("data", {})
                        for path, leak_type, leaked_value in live_contract_provenance_leaks(data):
                            add_issue(
                                issues,
                                "error",
                                "draft_provenance_in_live_contract",
                                sheet="form-app/data.js",
                                value={
                                    "model_key": model_key,
                                    "path": path,
                                    "leak_type": leak_type,
                                    "value": leaked_value,
                                },
                                message="Draft/review provenance and runtime-trim payload fields must not leak into live app data.",
                            )
                except (ValueError, json.JSONDecodeError) as exc:
                    add_issue(
                        issues,
                        "error",
                        "app_data_parse_failed",
                        sheet="form-app/data.js",
                        message=f"Could not parse window.CORVETTE_FORM_DATA: {exc}",
                    )
                validate_app_registry_freshness(wb, workbook, issues)

        return issues
    finally:
        wb.close()


def result_payload(workbook: str | Path, issues: list[SchemaIssue]) -> dict[str, Any]:
    return {
        "workbook": str(workbook),
        "status": "valid" if not any(issue.severity == "error" for issue in issues) else "invalid",
        "issue_count": len(issues),
        "error_count": sum(1 for issue in issues if issue.severity == "error"),
        "warning_count": sum(1 for issue in issues if issue.severity == "warning"),
        "issues": [asdict(issue) for issue in issues],
    }
