#!/usr/bin/env python3
"""Promote a workbook-authored model into the static app runtime registry.

Usage:
    python scripts/promote_model.py --model z06 [--write]
    python scripts/promote_model.py --model grand_sport_x --model zr1 --model zr1x [--write]

All promotion values derive from workbook metadata: ``model_master`` supplies
the label, year, registry key, and export slug; ``model_variants`` supplies
the exact membership rows and variant ids whose membership and
``variant_master`` active flags are activated; the draft artifact path derives
from the export slug. Without ``--write`` the script reports the changes it
would make and leaves the workbook untouched.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.model_configs import WORKBOOK_PATH, discover_generation_model_configs
from corvette_form_generator.registry_promotion import export_slug, registry_model_key, runtime_contract_artifact_path
from corvette_form_generator.workbook import (
    clean,
    excel_lock_path,
    restore_workbook_backup,
    rows_from_sheet,
    save_workbook_safely,
)


def headers_for(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}


def row_by_key(ws, key_column: str, key_value: str) -> int:
    headers = headers_for(ws)
    if key_column not in headers:
        raise ValueError(f"{ws.title} is missing required column {key_column!r}")
    key_col = headers[key_column]
    matches: list[int] = []
    for row_idx in range(2, ws.max_row + 1):
        if clean(ws.cell(row_idx, key_col).value).lower() == key_value.lower():
            matches.append(row_idx)
    if not matches:
        raise ValueError(f"{ws.title} is missing {key_column}={key_value!r}")
    if len(matches) > 1:
        raise ValueError(f"{ws.title} has duplicate {key_column}={key_value!r} rows: {matches}")
    return matches[0]


def row_by_membership(ws, model_key: str, variant_id: str) -> int:
    headers = headers_for(ws)
    required = {"model_key", "variant_id"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{ws.title} is missing required columns: {', '.join(missing)}")
    matches: list[int] = []
    for row_idx in range(2, ws.max_row + 1):
        row_model = clean(ws.cell(row_idx, headers["model_key"]).value).lower()
        row_variant = clean(ws.cell(row_idx, headers["variant_id"]).value)
        if row_model == model_key.lower() and row_variant == variant_id:
            matches.append(row_idx)
    if not matches:
        raise ValueError(f"{ws.title} is missing model_key={model_key!r}, variant_id={variant_id!r}")
    if len(matches) > 1:
        raise ValueError(
            f"{ws.title} has duplicate model_key={model_key!r}, variant_id={variant_id!r} rows: {matches}"
        )
    return matches[0]


def set_cell(ws, row_idx: int, headers: dict[str, int], column: str, value: Any, changes: list[dict[str, Any]]) -> None:
    if column not in headers:
        raise ValueError(f"{ws.title} is missing required column {column!r}")
    cell = ws.cell(row_idx, headers[column])
    old_value = cell.value
    if old_value == value:
        return
    changes.append(
        {
            "sheet": ws.title,
            "row": row_idx,
            "column": column,
            "old": clean(old_value),
            "new": clean(value),
        }
    )
    cell.value = value


def model_promotion_plan(wb, model_key: str) -> dict[str, Any]:
    models = {clean(row.get("model_key")).lower(): row for row in rows_from_sheet(wb, "model_master")}
    model = models.get(model_key)
    if not model:
        raise ValueError(f"model_master has no row for model_key {model_key!r}")
    model_label = clean(model.get("model_label")) or model_key.replace("_", " ").title()
    model_year = clean(model.get("model_year")) or "2027"
    slug = clean(model.get("export_slug")) or export_slug(model_key)
    registry_key = clean(model.get("registry_key")) or registry_model_key(model_key)
    variant_ids = [
        clean(row.get("variant_id"))
        for row in rows_from_sheet(wb, "model_variants")
        if clean(row.get("model_key")).lower() == model_key and clean(row.get("variant_id"))
    ]
    if not variant_ids:
        raise ValueError(f"model_variants has no rows for model_key {model_key!r}")
    return {
        "model_label": model_label,
        "dataset_name": f"{model_year} Corvette {model_label} operational form",
        "registry_key": registry_key,
        "export_slug": slug,
        "artifact_path": str(runtime_contract_artifact_path(Path("."), model_key)),
        "variant_ids": variant_ids,
        "model_notes": f"{model_label} promoted to runtime after source data review.",
        "promotion_notes": f"{model_label} runtime contract promoted to branch runtime registry.",
    }


def promote_model(wb, model_key: str, plan: dict[str, Any]) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []

    model_ws = wb["model_master"]
    model_headers = headers_for(model_ws)
    model_row = row_by_key(model_ws, "model_key", model_key)
    set_cell(model_ws, model_row, model_headers, "active", True, changes)
    set_cell(model_ws, model_row, model_headers, "dataset_name", plan["dataset_name"], changes)
    set_cell(model_ws, model_row, model_headers, "default_model", False, changes)
    set_cell(model_ws, model_row, model_headers, "registry_key", plan["registry_key"], changes)
    set_cell(model_ws, model_row, model_headers, "export_slug", plan["export_slug"], changes)
    set_cell(model_ws, model_row, model_headers, "notes", plan["model_notes"], changes)

    promotion_ws = wb["model_registry_promotion"]
    promotion_headers = headers_for(promotion_ws)
    promotion_row = row_by_key(promotion_ws, "model_key", model_key)
    set_cell(promotion_ws, promotion_row, promotion_headers, "registry_key", plan["registry_key"], changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "promoted_to_runtime", True, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "default_model", False, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "artifact_path", plan["artifact_path"], changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "artifact_type", "runtime_contract", changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "active", True, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "notes", plan["promotion_notes"], changes)

    variant_ws = wb["variant_master"]
    variant_headers = headers_for(variant_ws)
    for variant_id in plan["variant_ids"]:
        variant_row = row_by_key(variant_ws, "variant_id", variant_id)
        set_cell(variant_ws, variant_row, variant_headers, "active", True, changes)

    membership_ws = wb["model_variants"]
    membership_headers = headers_for(membership_ws)
    for variant_id in plan["variant_ids"]:
        membership_row = row_by_membership(membership_ws, model_key, variant_id)
        set_cell(membership_ws, membership_row, membership_headers, "active", True, changes)

    source_ws = wb["model_workbook_sources"]
    source_headers = headers_for(source_ws)
    if "model_key" not in source_headers:
        raise ValueError("model_workbook_sources is missing required column 'model_key'")
    source_rows = [
        row_idx
        for row_idx in range(2, source_ws.max_row + 1)
        if clean(source_ws.cell(row_idx, source_headers["model_key"]).value).lower() == model_key.lower()
    ]
    if not source_rows:
        raise ValueError(f"model_workbook_sources has no rows for model_key {model_key!r}")
    for source_row in source_rows:
        set_cell(source_ws, source_row, source_headers, "active", True, changes)

    return changes


def verify_workbook(path: Path, model_key: str, plan: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        models = {row["model_key"]: row for row in rows_from_sheet(wb, "model_master")}
        promotions = {row["model_key"]: row for row in rows_from_sheet(wb, "model_registry_promotion")}
        variants = {row["variant_id"]: row for row in rows_from_sheet(wb, "variant_master")}
        memberships = {
            clean(row.get("variant_id")): row
            for row in rows_from_sheet(wb, "model_variants")
            if clean(row.get("model_key")).lower() == model_key.lower()
        }
        model = models.get(model_key) or {}
        promotion = promotions.get(model_key) or {}
        variant_status = {variant_id: (variants.get(variant_id) or {}).get("active") for variant_id in plan["variant_ids"]}
        membership_status = {
            variant_id: (memberships.get(variant_id) or {}).get("active") for variant_id in plan["variant_ids"]
        }
        failures: list[str] = []
        if model.get("active") != "True":
            failures.append(f"model_master {model_key} active is not True")
        if model.get("dataset_name") != plan["dataset_name"]:
            failures.append(f"model_master {model_key} dataset_name was not updated")
        expected_promotion = {
            "promoted_to_runtime": "True",
            "active": "True",
            "artifact_path": plan["artifact_path"],
            "artifact_type": "runtime_contract",
            "default_model": "False",
            "registry_key": plan["registry_key"],
        }
        for key, expected in expected_promotion.items():
            if promotion.get(key) != expected:
                failures.append(
                    f"model_registry_promotion {model_key} {key} expected {expected!r}, found {promotion.get(key)!r}"
                )
        for variant_id, active in variant_status.items():
            if active != "True":
                failures.append(f"variant_master {variant_id} active expected 'True', found {active!r}")
        for variant_id, active in membership_status.items():
            if active != "True":
                failures.append(
                    f"model_variants {model_key}/{variant_id} active expected 'True', found {active!r}"
                )
    finally:
        wb.close()

    expected_variant_ids = list(plan["variant_ids"])
    discovery: dict[str, Any] = {
        "ok": False,
        "model_key": model_key,
        "expected_variant_ids": expected_variant_ids,
        "discovered_model_keys": [],
        "variant_ids": [],
        "error": None,
    }
    try:
        configs = discover_generation_model_configs(path)
        discovery["discovered_model_keys"] = list(configs)
        config = configs.get(model_key)
        if config is None:
            failures.append(f"generator discovery did not return target model {model_key!r}")
        else:
            discovered_variant_ids = list(config.variant_ids)
            discovery["variant_ids"] = discovered_variant_ids
            if set(discovered_variant_ids) != set(expected_variant_ids) or len(discovered_variant_ids) != len(
                expected_variant_ids
            ):
                failures.append(
                    f"generator discovery variants for {model_key!r} expected {expected_variant_ids!r}, "
                    f"found {discovered_variant_ids!r}"
                )
            else:
                discovery["ok"] = True
    except Exception as exc:  # noqa: BLE001 - verification must preserve discovery evidence
        discovery["error"] = {"type": type(exc).__name__, "message": str(exc)}
        failures.append(f"generator discovery failed for {model_key!r}: {type(exc).__name__}: {exc}")

    return {
        f"model_master_{model_key}": model,
        f"model_registry_promotion_{model_key}": promotion,
        "variant_active": variant_status,
        "model_variant_active": membership_status,
        "discovery": discovery,
        "failures": failures,
    }


def verify_promotions(path: Path, plans: dict[str, dict[str, Any]]) -> dict[str, Any]:
    models = {
        model_key: verify_workbook(path, model_key, plan)
        for model_key, plan in plans.items()
    }
    failures = [
        f"{model_key}: {failure}"
        for model_key, verification in models.items()
        for failure in verification["failures"]
    ]
    return {"ok": not failures, "models": models, "failures": failures}


def execute_promotion(workbook_path: Path, model_keys: list[str], *, write: bool = False) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    normalized_keys = list(dict.fromkeys(clean(model_key).lower() for model_key in model_keys if clean(model_key)))
    if not normalized_keys:
        raise ValueError("at least one model key is required")
    lock_path = excel_lock_path(workbook_path)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to run; Excel lock file is present: {lock_path}. Close Excel first.")

    loaded_mtime_ns = workbook_path.stat().st_mtime_ns
    wb = load_workbook(workbook_path)
    try:
        plans = {model_key: model_promotion_plan(wb, model_key) for model_key in normalized_keys}
        changes: list[dict[str, Any]] = []
        for model_key in normalized_keys:
            changes.extend(promote_model(wb, model_key, plans[model_key]))
        result: dict[str, Any] = {
            "ok": False,
            "status": "pending",
            "workbook": str(workbook_path),
            "model_keys": normalized_keys,
            "write": write,
            "plans": plans,
            "change_count": len(changes),
            "changes": changes,
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            scratch_path = Path(tmp_dir) / workbook_path.name
            shutil.copy2(workbook_path, scratch_path)
            scratch_mtime_ns = scratch_path.stat().st_mtime_ns
            save_workbook_safely(wb, scratch_path, loaded_mtime_ns=scratch_mtime_ns)
            preflight = verify_promotions(scratch_path, plans)
            result["preflight"] = preflight
            if not preflight["ok"]:
                result["status"] = "preflight_failed"
                return result

        if not write:
            result["ok"] = True
            result["status"] = "validated"
            return result

        backup_path = save_workbook_safely(wb, workbook_path, loaded_mtime_ns=loaded_mtime_ns)
        result["backup_path"] = str(backup_path)
        verification = verify_promotions(workbook_path, plans)
        result["verification"] = verification
        if verification["ok"]:
            result["ok"] = True
            result["status"] = "applied"
            return result

        restore_workbook_backup(workbook_path, backup_path)
        restored = hashlib.sha256(workbook_path.read_bytes()).hexdigest() == hashlib.sha256(backup_path.read_bytes()).hexdigest()
        result["status"] = "post_save_verification_failed_restored" if restored else "restore_failed"
        result["restored"] = restored
        return result
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--model", action="append", required=True, help="model key to promote; repeat for one atomic batch")
    parser.add_argument("--workbook", default=str(WORKBOOK_PATH), help="workbook to inspect or promote")
    parser.add_argument("--write", action="store_true", help="write changes to stingray_master.xlsx")
    args = parser.parse_args()
    result = execute_promotion(Path(args.workbook), args.model, write=args.write)
    print(json.dumps(result, indent=2))
    if not result["ok"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
