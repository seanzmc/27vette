#!/usr/bin/env python3
"""Set workbook-owned asset-card sizing/alignment by RPO.

The command is dry-run by default. Pass ``--write`` only after reviewing the
resolved asset_map rows and the guarded editor-operation validation result.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.asset_map_sync import (  # noqa: E402
    clean,
    discover_promoted_option_sources,
    read_option_sheets,
)
from corvette_form_generator.editor_ops import apply_batch  # noqa: E402
from corvette_form_generator.workbook import workbook_truthy  # noqa: E402

FIT_CHOICES = ("cover", "contain", "swatch")
POSITION_PATTERN = re.compile(r"^[A-Za-z0-9 .%_-]+$")


def _asset_rows(ws) -> dict[tuple[str, str, str], dict[str, str]]:
    headers = [clean(cell.value) for cell in ws[1]]
    index = {header: column for column, header in enumerate(headers) if header}
    required = {"model_key", "target_type", "target_id", "image_url", "active"}
    missing = required - set(index)
    if missing:
        raise ValueError(f"asset_map missing required columns: {', '.join(sorted(missing))}")

    rows: dict[tuple[str, str, str], dict[str, str]] = {}
    for row_number, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not workbook_truthy(raw[index["active"]]):
            continue
        model_key = clean(raw[index["model_key"]]).lower()
        target_type = clean(raw[index["target_type"]]).lower()
        target_id = clean(raw[index["target_id"]]).lower()
        if not (model_key and target_type and target_id):
            continue
        key = (model_key, target_type, target_id)
        if key in rows:
            raise ValueError(f"duplicate active asset_map row for {key}")
        rows[key] = {
            "row_number": str(row_number),
            "image_url": clean(raw[index["image_url"]]),
            "image_fit": clean(raw[index["image_fit"]]).lower() if "image_fit" in index else "",
            "image_position": clean(raw[index["image_position"]]) if "image_position" in index else "",
        }
    return rows


def resolve_updates(
    workbook: Path,
    *,
    rpos: list[str],
    models: list[str],
    image_fit: str | None,
    image_position: str | None,
) -> dict:
    requested_rpos = sorted({rpo.strip().lower() for rpo in rpos if rpo.strip()})
    requested_models = sorted({model.strip().lower() for model in models if model.strip()})
    if not requested_rpos:
        raise ValueError("at least one nonblank --rpo is required")

    wb = load_workbook(workbook, read_only=True, data_only=False)
    try:
        sources = discover_promoted_option_sources(wb)
        unknown_models = sorted(set(requested_models) - set(sources))
        if unknown_models:
            raise ValueError(
                "models are not promoted active option sources: " + ", ".join(unknown_models)
            )
        selected_models = set(requested_models or sources)
        desired = read_option_sheets(wb, sources)
        assets = _asset_rows(wb["asset_map"])
    finally:
        wb.close()

    resolved: dict[tuple[str, str, str], dict] = {}
    missing_assets: list[dict[str, str]] = []
    matched_rpos: set[str] = set()
    for (model_key, target_type, target_id), option in sorted(desired.items()):
        rpo = option.get("rpo", "").lower()
        if model_key not in selected_models or rpo not in requested_rpos:
            continue
        matched_rpos.add(rpo)
        exact_key = (model_key, target_type, target_id)
        wildcard_key = ("*", target_type, target_id)
        asset_key = exact_key if exact_key in assets else wildcard_key if wildcard_key in assets else None
        if asset_key is None or not assets[asset_key]["image_url"]:
            missing_assets.append({"model_key": model_key, "rpo": rpo.upper(), "target_id": target_id})
            continue
        item = resolved.setdefault(
            asset_key,
            {
                "key": {
                    "model_key": asset_key[0],
                    "target_type": asset_key[1],
                    "target_id": asset_key[2],
                },
                "current": {
                    "image_fit": assets[asset_key]["image_fit"] or "cover",
                    "image_position": assets[asset_key]["image_position"] or "center",
                },
                "covers": [],
            },
        )
        item["covers"].append({"model_key": model_key, "rpo": rpo.upper()})

    unmatched_rpos = sorted(set(requested_rpos) - matched_rpos)
    if unmatched_rpos:
        raise ValueError("RPOs were not found in the selected promoted models: " + ", ".join(r.upper() for r in unmatched_rpos))

    operations: list[dict] = []
    rows = []
    for asset_key, item in sorted(resolved.items()):
        updates: dict[str, str] = {}
        if image_fit is not None and item["current"]["image_fit"] != image_fit:
            updates["image_fit"] = image_fit
        if image_position is not None and item["current"]["image_position"] != image_position:
            updates["image_position"] = image_position
        row_result = {**item, "updates": updates}
        rows.append(row_result)
        if updates:
            operations.append(
                {
                    "action": "update",
                    "sheet": "asset_map",
                    "key": item["key"],
                    "row": updates,
                }
            )

    return {
        "workbook": str(workbook),
        "requested_rpos": [rpo.upper() for rpo in requested_rpos],
        "selected_models": sorted(selected_models),
        "resolved_rows": rows,
        "missing_asset_targets": missing_assets,
        "operations": operations,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=str(ROOT / "stingray_master.xlsx"))
    parser.add_argument("--rpo", action="append", required=True, help="RPO to update; repeat as needed")
    parser.add_argument("--model", action="append", default=[], help="promoted model_key scope; default is all")
    parser.add_argument("--fit", choices=FIT_CHOICES, help="card image sizing mode")
    parser.add_argument("--position", help="CSS object-position value such as center, top, or 50%% 30%%")
    parser.add_argument("--write", action="store_true", help="apply after guarded validation; default is dry-run")
    args = parser.parse_args()

    if args.fit is None and args.position is None:
        parser.error("at least one of --fit or --position is required")
    position = args.position.strip() if args.position is not None else None
    if position is not None and (not position or len(position) > 64 or not POSITION_PATTERN.fullmatch(position)):
        parser.error("--position must be a simple CSS object-position value")

    workbook = Path(args.workbook).resolve()
    try:
        resolution = resolve_updates(
            workbook,
            rpos=args.rpo,
            models=args.model,
            image_fit=args.fit,
            image_position=position,
        )
    except (OSError, ValueError) as exc:
        print(json.dumps({"ok": False, "status": "resolution_failed", "errors": [str(exc)]}, indent=2))
        return 1

    operations = resolution.pop("operations")
    if not operations:
        print(json.dumps({"ok": True, "status": "no_changes", **resolution}, indent=2))
        return 0

    batch = {
        "version": 1,
        "workbook": workbook.name,
        "workbookMtimeNs": str(workbook.stat().st_mtime_ns),
        "items": operations,
    }
    result = apply_batch(workbook, batch, write=args.write, source="asset-display-cli")
    payload = {
        "ok": bool(result.get("ok")),
        "mode": "write" if args.write else "dry-run",
        **resolution,
        "operation_count": len(operations),
        "apply_result": result,
    }
    print(json.dumps(payload, indent=2, default=str))
    return 0 if payload["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
