#!/usr/bin/env python3
"""Repair Z option-sheet pricing, sections, selectability, and display order.

Dry-run is the default. Use --write to safely save stingray_master.xlsx.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import FUTURE_MODEL_SPECS, OPTION_SOURCE_HEADERS  # noqa: E402
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely  # noqa: E402

TARGET_SHEETS = {
    "z06_options": "z06",
    "zr1_options": "zr1",
    "zr1x_options": "zr1x",
}

STANDARD_EQUIPMENT_SECTIONS = {
    "sec_stan_001",
    "sec_1lte_001",
    "sec_2lte_001",
    "sec_3lte_001",
    "sec_incl_001",
    "sec_safe_001",
    "sec_stan_002",
    "sec_tech_001",
}

SECTION_MOVES = {
    "WUB": "sec_stan_001",
    "UV6": "sec_1lte_001",
}

ARCHIVE_PATHS = {
    "z06_options": ROOT / "archive-2026-05-29/archived/referenceSheets/stingray_master - z06_options.csv",
    "zr1_options": ROOT / "archive-2026-05-29/archived/referenceSheets/stingray_master - zr1_options.csv",
    "zr1x_options": ROOT / "archive-2026-05-29/archived/referenceSheets/stingray_master - zr1x_options.csv",
}

NO_PRICE_SENTINELS = {"", "no price", "n/a", "na", "none"}


@dataclass(frozen=True)
class CellUpdate:
    sheet: str
    row_number: int
    option_id: str
    rpo: str
    field: str
    current: Any
    desired: Any
    reason: str


def _headers(ws) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def _header_index(ws) -> dict[str, int]:
    headers = _headers(ws)
    if headers != list(OPTION_SOURCE_HEADERS):
        raise RuntimeError(f"{ws.title} header drift: expected {list(OPTION_SOURCE_HEADERS)}, found {headers}")
    return {header: offset + 1 for offset, header in enumerate(headers)}


def _option_identity(ws, index: dict[str, int], row_number: int) -> tuple[str, str]:
    option_id = clean(ws.cell(row_number, index["option_id"]).value)
    rpo = clean(ws.cell(row_number, index["rpo"]).value)
    return option_id, rpo


def _is_active(ws, index: dict[str, int], row_number: int) -> bool:
    return clean(ws.cell(row_number, index["active"]).value).casefold() in {"true", "1", "yes", "y", "active"}


def _parse_price(value: str) -> int | float | str:
    text = clean(value)
    if text.casefold() in NO_PRICE_SENTINELS:
        return ""
    numeric = float(text.replace("$", "").replace(",", ""))
    return int(numeric) if numeric.is_integer() else numeric


def _price_equal(current: Any, desired: Any) -> bool:
    current_clean = clean(current)
    if desired == "":
        return current_clean == ""
    try:
        current_value = _parse_price(current_clean)
    except Exception:
        return False
    return current_value == desired


def _load_archive_orders() -> dict[str, dict[tuple[str, str, str], str]]:
    archive_orders: dict[str, dict[tuple[str, str, str], str]] = {}
    for sheet_name, path in ARCHIVE_PATHS.items():
        if not path.exists():
            raise RuntimeError(f"Missing archive display_order source: {path}")
        rows: dict[tuple[str, str, str], str] = {}
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                key = (clean(row.get("option_id")), clean(row.get("rpo")), clean(row.get("section_id")))
                display_order = clean(row.get("display_order"))
                if key != ("", "", "") and display_order:
                    rows.setdefault(key, display_order)
        archive_orders[sheet_name] = rows
    return archive_orders


def _load_price_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    required = {"option_sheet", "option_row", "option_id", "rpo", "approved_price"}
    headers = set(rows[0].keys()) if rows else set()
    missing = sorted(required - headers)
    if missing:
        raise RuntimeError(f"{path} missing required column(s): {', '.join(missing)}")
    return rows


def _effective_section(rpo: str, current_section: str) -> str:
    return SECTION_MOVES.get(rpo, current_section)


def _collect_section_moves(wb) -> tuple[list[CellUpdate], dict[str, Any]]:
    updates: list[CellUpdate] = []
    by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for sheet_name in TARGET_SHEETS:
        ws = wb[sheet_name]
        index = _header_index(ws)
        section_col = index["section_id"]
        for row_number in range(2, ws.max_row + 1):
            option_id, rpo = _option_identity(ws, index, row_number)
            if rpo not in SECTION_MOVES:
                continue
            current = clean(ws.cell(row_number, section_col).value)
            desired = SECTION_MOVES[rpo]
            if current != desired:
                updates.append(CellUpdate(sheet_name, row_number, option_id, rpo, "section_id", current, desired, "approved section move"))
                by_sheet[sheet_name].append({"row_number": row_number, "rpo": rpo, "current": current, "desired": desired})
    return updates, {sheet: rows for sheet, rows in sorted(by_sheet.items())}


def _collect_default_selectable_updates(wb) -> tuple[list[CellUpdate], dict[str, Any]]:
    updates: list[CellUpdate] = []
    by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for sheet_name in TARGET_SHEETS:
        ws = wb[sheet_name]
        index = _header_index(ws)
        selectable_col = index["selectable"]
        for row_number in range(2, ws.max_row + 1):
            if clean(ws.cell(row_number, index["display_behavior"]).value) != "default_selected":
                continue
            option_id, rpo = _option_identity(ws, index, row_number)
            current = clean(ws.cell(row_number, selectable_col).value)
            if current.casefold() != "true":
                updates.append(CellUpdate(sheet_name, row_number, option_id, rpo, "selectable", current, "True", "default_selected rows must remain selectable"))
                by_sheet[sheet_name].append({"row_number": row_number, "rpo": rpo, "current": current, "desired": "True"})
    return updates, {sheet: rows for sheet, rows in sorted(by_sheet.items())}


def _collect_price_updates(wb, price_rows: list[dict[str, str]]) -> tuple[list[CellUpdate], dict[str, Any]]:
    updates: list[CellUpdate] = []
    summary: dict[str, Any] = {
        "csv_rows": len(price_rows),
        "current_standard_section_rows_ignored": 0,
        "effective_standard_section_rows_ignored": 0,
        "current_standard_by_section": Counter(),
        "effective_standard_by_section": Counter(),
        "applied_by_bucket": Counter(),
        "ignored_standard_with_filled_approved_price": 0,
        "row_identity_errors": [],
        "unsupported_approved_price_errors": [],
        "updates_by_sheet": Counter(),
    }
    sheet_indexes = {sheet: _header_index(wb[sheet]) for sheet in TARGET_SHEETS}
    option_locations: dict[str, dict[str, int]] = {}
    for sheet_name, index in sheet_indexes.items():
        ws = wb[sheet_name]
        locations: dict[str, int] = {}
        for row_number in range(2, ws.max_row + 1):
            option_id = clean(ws.cell(row_number, index["option_id"]).value)
            if not option_id:
                continue
            if option_id in locations:
                raise RuntimeError(f"{sheet_name} has duplicate option_id {option_id!r}; cannot safely join pricing CSV")
            locations[option_id] = row_number
        option_locations[sheet_name] = locations

    for csv_row_number, csv_row in enumerate(price_rows, start=2):
        sheet_name = clean(csv_row.get("option_sheet"))
        if sheet_name not in TARGET_SHEETS:
            summary["row_identity_errors"].append(f"CSV row {csv_row_number}: unsupported option_sheet {sheet_name!r}")
            continue
        csv_option_id = clean(csv_row.get("option_id"))
        if not csv_option_id:
            summary["row_identity_errors"].append(f"CSV row {csv_row_number}: missing option_id")
            continue
        row_number = option_locations[sheet_name].get(csv_option_id)
        if row_number is None:
            summary["row_identity_errors"].append(f"CSV row {csv_row_number}: option_id {csv_option_id!r} was not found in {sheet_name}")
            continue
        ws = wb[sheet_name]
        index = sheet_indexes[sheet_name]
        option_id, rpo = _option_identity(ws, index, row_number)
        csv_rpo = clean(csv_row.get("rpo"))
        if csv_rpo and csv_rpo != rpo:
            summary["row_identity_errors"].append(f"CSV row {csv_row_number}: expected rpo {csv_rpo!r}, found {rpo!r} in {sheet_name} row {row_number} for option_id {option_id!r}")
            continue

        current_section = clean(ws.cell(row_number, index["section_id"]).value)
        effective_section = _effective_section(rpo, current_section)
        approved_text = clean(csv_row.get("approved_price"))
        if current_section in STANDARD_EQUIPMENT_SECTIONS:
            summary["current_standard_section_rows_ignored"] += 1
            summary["current_standard_by_section"][current_section] += 1
        if effective_section in STANDARD_EQUIPMENT_SECTIONS:
            summary["effective_standard_section_rows_ignored"] += 1
            summary["effective_standard_by_section"][effective_section] += 1
            if approved_text:
                summary["ignored_standard_with_filled_approved_price"] += 1
            continue

        try:
            desired = _parse_price(approved_text)
        except Exception:
            summary["unsupported_approved_price_errors"].append(
                f"CSV row {csv_row_number}: unsupported approved_price {approved_text!r} for {sheet_name} row {row_number} {rpo}"
            )
            continue
        bucket = "blank_or_no_price" if desired == "" else "numeric"
        summary["applied_by_bucket"][bucket] += 1
        price_col = index["price"]
        current = ws.cell(row_number, price_col).value
        if not _price_equal(current, desired):
            updates.append(CellUpdate(sheet_name, row_number, option_id, rpo, "price", current, desired, "approved_price CSV non-standard row"))
            summary["updates_by_sheet"][sheet_name] += 1

    if summary["row_identity_errors"] or summary["unsupported_approved_price_errors"]:
        raise RuntimeError(json.dumps({"price_csv_errors": summary}, indent=2, default=str))
    summary["current_standard_by_section"] = dict(sorted(summary["current_standard_by_section"].items()))
    summary["effective_standard_by_section"] = dict(sorted(summary["effective_standard_by_section"].items()))
    summary["applied_by_bucket"] = dict(sorted(summary["applied_by_bucket"].items()))
    summary["updates_by_sheet"] = dict(sorted(summary["updates_by_sheet"].items()))
    return updates, summary


def _numeric_order(value: Any) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _collect_display_order_updates(wb) -> tuple[list[CellUpdate], dict[str, Any]]:
    archive_orders = _load_archive_orders()
    updates: list[CellUpdate] = []
    summary: dict[str, Any] = {"by_sheet": {}}
    for sheet_name in TARGET_SHEETS:
        ws = wb[sheet_name]
        index = _header_index(ws)
        order_col = index["display_order"]
        section_col = index["section_id"]
        section_max: dict[str, int] = defaultdict(int)
        missing_after_archive: list[int] = []
        archive_restored = 0

        for row_number in range(2, ws.max_row + 1):
            current_order = ws.cell(row_number, order_col).value
            section = clean(ws.cell(row_number, section_col).value)
            rpo = clean(ws.cell(row_number, index["rpo"]).value)
            effective_section = _effective_section(rpo, section)
            numeric = _numeric_order(current_order)
            if numeric is not None:
                section_max[effective_section] = max(section_max[effective_section], numeric)

        for row_number in range(2, ws.max_row + 1):
            if not _is_active(ws, index, row_number):
                continue
            current_order = clean(ws.cell(row_number, order_col).value)
            if current_order:
                continue
            option_id, rpo = _option_identity(ws, index, row_number)
            section = clean(ws.cell(row_number, section_col).value)
            effective_section = _effective_section(rpo, section)
            key = (option_id, rpo, section)
            archived = archive_orders[sheet_name].get(key)
            if archived:
                updates.append(CellUpdate(sheet_name, row_number, option_id, rpo, "display_order", current_order, archived, "archive display_order restore"))
                numeric = _numeric_order(archived)
                if numeric is not None:
                    section_max[effective_section] = max(section_max[effective_section], numeric)
                archive_restored += 1
            else:
                missing_after_archive.append(row_number)

        deterministic_assigned = 0
        for row_number in missing_after_archive:
            option_id, rpo = _option_identity(ws, index, row_number)
            section = clean(ws.cell(row_number, section_col).value)
            effective_section = _effective_section(rpo, section)
            base = section_max[effective_section]
            next_order = int(math.ceil(base / 10) * 10 + 10) if base else 10
            section_max[effective_section] = next_order
            updates.append(CellUpdate(sheet_name, row_number, option_id, rpo, "display_order", "", str(next_order), "deterministic section-local display_order assignment"))
            deterministic_assigned += 1

        summary["by_sheet"][sheet_name] = {
            "archive_restored": archive_restored,
            "deterministic_assigned": deterministic_assigned,
            "total_updates": archive_restored + deterministic_assigned,
        }
    return updates, summary


def _dedupe_updates(updates: list[CellUpdate]) -> list[CellUpdate]:
    seen: set[tuple[str, int, str]] = set()
    deduped: list[CellUpdate] = []
    for update in updates:
        key = (update.sheet, update.row_number, update.field)
        if key in seen:
            raise RuntimeError(f"Duplicate update for {key}")
        seen.add(key)
        deduped.append(update)
    return deduped


def build_plan(wb, price_csv_path: Path) -> dict[str, Any]:
    price_rows = _load_price_csv(price_csv_path)
    section_updates, section_summary = _collect_section_moves(wb)
    selectable_updates, selectable_summary = _collect_default_selectable_updates(wb)
    price_updates, price_summary = _collect_price_updates(wb, price_rows)
    order_updates, order_summary = _collect_display_order_updates(wb)
    updates = _dedupe_updates([*section_updates, *selectable_updates, *price_updates, *order_updates])
    by_sheet: Counter[str] = Counter(update.sheet for update in updates)
    by_field: Counter[str] = Counter(update.field for update in updates)
    return {
        "workbook": str(WORKBOOK_PATH),
        "price_csv": str(price_csv_path),
        "updates": updates,
        "summary": {
            "total_updates": len(updates),
            "updates_by_sheet": dict(sorted(by_sheet.items())),
            "updates_by_field": dict(sorted(by_field.items())),
            "section_moves": section_summary,
            "default_selected_selectable_fixes": selectable_summary,
            "price_application": price_summary,
            "display_order": order_summary,
        },
    }


def apply_updates(wb, updates: list[CellUpdate]) -> None:
    indexes = {sheet: _header_index(wb[sheet]) for sheet in TARGET_SHEETS}
    for update in updates:
        ws = wb[update.sheet]
        col = indexes[update.sheet][update.field]
        ws.cell(update.row_number, col).value = update.desired


def _verify_saved_workbook(path: Path, plan: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        indexes = {sheet: _header_index(wb[sheet]) for sheet in TARGET_SHEETS}
        checked = 0
        for update in plan["updates"]:
            ws = wb[update.sheet]
            actual = ws.cell(update.row_number, indexes[update.sheet][update.field]).value
            if update.field == "price":
                if not _price_equal(actual, update.desired):
                    raise RuntimeError(f"{update.sheet} row {update.row_number} {update.field}: expected {update.desired!r}, found {actual!r}")
            elif clean(actual) != clean(update.desired):
                raise RuntimeError(f"{update.sheet} row {update.row_number} {update.field}: expected {update.desired!r}, found {actual!r}")
            checked += 1

        default_selectable_false = []
        section_checks = []
        missing_order = []
        standard_prices = []
        for sheet_name in TARGET_SHEETS:
            ws = wb[sheet_name]
            index = indexes[sheet_name]
            for row_number in range(2, ws.max_row + 1):
                option_id, rpo = _option_identity(ws, index, row_number)
                section = clean(ws.cell(row_number, index["section_id"]).value)
                if rpo in SECTION_MOVES:
                    expected = SECTION_MOVES[rpo]
                    section_checks.append({"sheet": sheet_name, "row_number": row_number, "rpo": rpo, "section_id": section})
                    if section != expected:
                        raise RuntimeError(f"{sheet_name} row {row_number} {rpo}: expected section {expected}, found {section}")
                if clean(ws.cell(row_number, index["display_behavior"]).value) == "default_selected" and clean(ws.cell(row_number, index["selectable"]).value).casefold() != "true":
                    default_selectable_false.append({"sheet": sheet_name, "row_number": row_number, "rpo": rpo})
                if _is_active(ws, index, row_number) and not clean(ws.cell(row_number, index["display_order"]).value):
                    missing_order.append({"sheet": sheet_name, "row_number": row_number, "option_id": option_id, "rpo": rpo, "section_id": section})
                if section in STANDARD_EQUIPMENT_SECTIONS and clean(ws.cell(row_number, index["price"]).value):
                    standard_prices.append({"sheet": sheet_name, "row_number": row_number, "rpo": rpo, "section_id": section, "price": clean(ws.cell(row_number, index["price"]).value)})
        if default_selectable_false:
            raise RuntimeError(f"default_selected rows still non-selectable: {default_selectable_false}")
        if missing_order:
            raise RuntimeError(f"active rows still missing display_order: {missing_order[:20]}")
        if standard_prices:
            raise RuntimeError(f"standard equipment rows have prices after repair: {standard_prices[:20]}")
        return {
            "checked_updates": checked,
            "section_checks": section_checks,
            "default_selected_rows_selectable": True,
            "active_display_orders_complete": True,
            "standard_equipment_prices_blank": True,
        }
    finally:
        wb.close()


def _json_report(status: str, plan: dict[str, Any], *, include_updates: bool, backup_path: Path | None = None, verification: dict[str, Any] | None = None) -> dict[str, Any]:
    report = {"status": status, **plan["summary"], "workbook": plan["workbook"], "price_csv": plan["price_csv"]}
    if backup_path is not None:
        report["backup"] = str(backup_path)
    if verification is not None:
        report["verification"] = verification
    if include_updates:
        report["updates"] = [update.__dict__ for update in plan["updates"]]
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, default=ROOT / "z-option-canonical-pricing-matrix.csv", help="Approved pricing matrix CSV.")
    parser.add_argument("--write", action="store_true", help="Write workbook changes. Default is dry-run.")
    parser.add_argument("--include-updates", action="store_true", help="Include every cell update in the JSON report.")
    args = parser.parse_args(argv)

    price_csv_path = args.csv if args.csv.is_absolute() else ROOT / args.csv
    if not price_csv_path.exists():
        raise RuntimeError(f"Missing price CSV: {price_csv_path}")
    if args.write and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    backup_path = None
    try:
        for sheet in TARGET_SHEETS:
            if sheet not in wb.sheetnames:
                raise RuntimeError(f"Missing target sheet: {sheet}")
        # Prove future models remain inactive; this pass is source canonicalization, not runtime promotion.
        for model_key in FUTURE_MODEL_SPECS:
            if FUTURE_MODEL_SPECS[model_key].target_option_sheet in TARGET_SHEETS and TARGET_SHEETS[FUTURE_MODEL_SPECS[model_key].target_option_sheet] != model_key:
                raise RuntimeError(f"Future model spec mismatch for {model_key}")
        plan = build_plan(wb, price_csv_path)
        if not args.write:
            print(json.dumps(_json_report("dry_run", plan, include_updates=args.include_updates), indent=2, default=str))
            return 0
        apply_updates(wb, plan["updates"])
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = _verify_saved_workbook(WORKBOOK_PATH, plan)
    print(json.dumps(_json_report("written", plan, include_updates=args.include_updates, backup_path=backup_path, verification=verification), indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
