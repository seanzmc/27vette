#!/usr/bin/env python3
"""Create/update the future_model_source_review sheet from raw order-guide sheets."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SOURCE_REVIEW_HEADERS,
    build_source_review_rows,
)
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely, write_sheet  # noqa: E402

REVIEW_SHEET = "future_model_source_review"
PRESERVED_REVIEW_FIELDS = {
    "approved_option_id",
    "approved_rpo",
    "approved_price",
    "approved_option_name",
    "approved_description",
    "approved_detail_raw",
    "approved_section_id",
    "approved_selectable",
    "approved_display_behavior",
    "approved_display_order",
    "copy_from_model_key",
    "copy_from_option_id",
    "duplicate_group_id",
    "review_status",
    "review_reason",
    "active",
    "notes",
}


def row_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        clean(row.get("model_key")),
        clean(row.get("source_primary_rpo")),
        clean(row.get("source_option_description")),
        clean(row.get("raw_source_spans")),
    )


def existing_review_rows(wb) -> dict[tuple[str, str, str, str], dict[str, str]]:
    if REVIEW_SHEET not in wb.sheetnames:
        return {}
    ws = wb[REVIEW_SHEET]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        record = {header: clean(value) for header, value in zip(headers, values) if header}
        if any(record.values()):
            rows[row_key(record)] = record
    return rows


def _has_manual_review_decision(row: dict[str, str]) -> bool:
    status = clean(row.get("review_status"))
    if status in {"approved", "inactive", "deferred"}:
        return True
    if clean(row.get("active")).casefold() in {"true", "1", "yes", "y", "active"}:
        return True
    return False


def merge_preserved_fields(generated_rows: list[dict[str, Any]], existing_rows: dict[tuple[str, str, str, str], dict[str, str]], *, reset_reviewed_fields: bool) -> list[dict[str, Any]]:
    if reset_reviewed_fields:
        return generated_rows
    merged: list[dict[str, Any]] = []
    for generated in generated_rows:
        prior = existing_rows.get(row_key(generated))
        if prior and _has_manual_review_decision(prior):
            for field in PRESERVED_REVIEW_FIELDS:
                if clean(prior.get(field)):
                    generated[field] = prior[field]
        merged.append(generated)
    return merged


def verify_saved_rows(path: Path, expected_rows: int) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if REVIEW_SHEET not in wb.sheetnames:
            raise RuntimeError(f"{REVIEW_SHEET} was not created")
        ws = wb[REVIEW_SHEET]
        headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        if headers != list(FUTURE_MODEL_SOURCE_REVIEW_HEADERS):
            raise RuntimeError(f"{REVIEW_SHEET} header drift: {headers}")
        data_rows = max(ws.max_row - 1, 0)
        if data_rows != expected_rows:
            raise RuntimeError(f"Expected {expected_rows} review rows, found {data_rows}")
        counts: dict[str, int] = {}
        model_col = headers.index("model_key") + 1
        for row_number in range(2, ws.max_row + 1):
            model_key = clean(ws.cell(row_number, model_col).value)
            counts[model_key] = counts.get(model_key, 0) + 1
        return {"rows": data_rows, "counts_by_model": counts}
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reset-reviewed-fields", action="store_true", help="Overwrite any existing manual review/approved fields.")
    args = parser.parse_args(argv)

    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {lock_path}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    try:
        generated_rows = build_source_review_rows(wb)
        existing_rows = existing_review_rows(wb)
        rows = merge_preserved_fields(generated_rows, existing_rows, reset_reviewed_fields=args.reset_reviewed_fields)
        write_sheet(wb, REVIEW_SHEET, list(FUTURE_MODEL_SOURCE_REVIEW_HEADERS), rows)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_rows(WORKBOOK_PATH, len(rows))
    print(json.dumps({"workbook": str(WORKBOOK_PATH), "backup": str(backup_path), "verification": verification}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
