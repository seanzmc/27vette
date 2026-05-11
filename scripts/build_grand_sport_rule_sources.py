#!/usr/bin/env python3
"""Audit workbook-authored Grand Sport rule source sheets."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.model_configs import GRAND_SPORT_MODEL
from corvette_form_generator.workbook import clean, rows_from_sheet


RULE_MAPPING_HEADERS = [
    "rule_id",
    "source_id",
    "rule_type",
    "target_id",
    "target_type",
    "original_detail_raw",
    "review_flag",
    "source_type",
    "target_selection_mode",
    "source_selection_mode",
    "target_section",
    "source_section",
    "generation_action",
    "body_style_scope",
    "runtime_action",
    "disabled_reason",
]
RULE_GROUP_HEADERS = [
    "group_id",
    "group_type",
    "source_id",
    "body_style_scope",
    "trim_level_scope",
    "variant_scope",
    "disabled_reason",
    "active",
    "notes",
]
RULE_GROUP_MEMBER_HEADERS = ["group_id", "target_id", "display_order", "active"]
EXCLUSIVE_GROUP_HEADERS = ["group_id", "selection_mode", "active", "notes"]
EXCLUSIVE_GROUP_MEMBER_HEADERS = ["group_id", "option_id", "display_order", "active"]
RULE_PHRASES = (
    "not available with",
    "requires",
    "includes",
    "included with",
    "only available with",
    "included and only available with",
)
ENGINE_COVER_RPOS = {"BC7", "BCP", "BCS", "BC4", "B6P", "ZZ3", "D3V", "SL9"}


def active_source_row(row: dict[str, str]) -> bool:
    return clean(row.get("active", "True")).lower() == "true"


def option_indexes(rows: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    options_by_id = {row["option_id"]: row for row in rows if row.get("option_id")}
    option_ids_by_rpo: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        option_id = row.get("option_id", "")
        rpo = row.get("rpo", "")
        if option_id and rpo:
            option_ids_by_rpo[rpo].append(option_id)
    return options_by_id, option_ids_by_rpo


def runtime_available_option_ids(options: list[dict[str, str]], status_rows: list[dict[str, str]], variant_ids: tuple[str, ...]) -> set[str]:
    statuses_by_option: dict[str, set[str]] = defaultdict(set)
    for row in status_rows:
        if row.get("variant_id") in variant_ids:
            statuses_by_option[row.get("option_id", "")].add(row.get("status", "").lower())
    return {
        row["option_id"]
        for row in options
        if row.get("option_id")
        and active_source_row(row)
        and (
            row.get("display_behavior", "") == "auto_only"
            or bool(statuses_by_option.get(row["option_id"], set()) & {"available", "standard"})
        )
    }


def runtime_rule_rows(workbook_rules: list[dict[str, str]], runtime_option_ids: set[str], grouped_excludes: set[tuple[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in workbook_rules:
        if not runtime_authored_rule(row):
            continue
        source_id = row.get("source_id", "")
        target_id = row.get("target_id", "")
        source_type = row.get("source_type", "option") or "option"
        target_type = row.get("target_type", "option") or "option"
        if source_type == "option" and source_id not in runtime_option_ids:
            continue
        if target_type == "option" and target_id not in runtime_option_ids:
            continue
        if (
            row.get("rule_type", "").lower() == "excludes"
            and (source_id, target_id) in grouped_excludes
            and row.get("generation_action", "") != "preserve_runtime_exclude"
        ):
            continue
        rows.append(row)
    return rows


def rpo_codes(text: str) -> list[str]:
    codes: list[str] = []
    for group in re.findall(r"\(([A-Z0-9][A-Z0-9/,\s-]*)\)", text):
        for code in re.split(r"[/,\s]+", group.replace("-", " ")):
            if re.fullmatch(r"[A-Z0-9]{3}", code):
                codes.append(code)
    return codes


def referenced_rpo_codes(text: str, option_ids_by_rpo: dict[str, list[str]]) -> list[str]:
    codes = rpo_codes(text)
    for code in re.findall(r"\b[A-Z0-9]{3}\b", text):
        if code in option_ids_by_rpo and code not in codes:
            codes.append(code)
    return codes


def phrase_tail(text: str, phrase: str, stop_phrases: tuple[str, ...] = ()) -> str:
    match = re.search(re.escape(phrase), text, flags=re.IGNORECASE)
    if not match:
        return ""
    tail = text[match.end() :]
    lower_tail = tail.lower()
    stop_indexes = [
        lower_tail.find(stop_phrase.lower())
        for stop_phrase in stop_phrases
        if lower_tail.find(stop_phrase.lower()) >= 0
    ]
    if stop_indexes:
        tail = tail[: min(stop_indexes)]
    return tail


def detail_fragments(detail_raw: str) -> list[str]:
    fragments = []
    for line in re.split(r"(?:^|\n)\s*\d+\.\s*", detail_raw):
        text = re.sub(r"\s+", " ", line).strip()
        if text:
            fragments.append(text)
    return fragments or ([re.sub(r"\s+", " ", detail_raw).strip()] if detail_raw.strip() else [])


def interior_combination_codes(wb) -> set[str]:
    codes: set[str] = set()
    for row in rows_from_sheet(wb, "lt_interiors"):
        for key in ("Interior Code", "Seat", "Stitch", "Suede", "Two Tone"):
            value = clean(row.get(key, "")).upper()
            if re.fullmatch(r"[A-Z0-9]{3}", value):
                codes.add(value)
        for token in clean(row.get("interior_id", "")).upper().split("_"):
            if re.fullmatch(r"[A-Z0-9]{3}", token):
                codes.add(token)
    return codes


def ids_for_codes(codes: list[str], option_ids_by_rpo: dict[str, list[str]]) -> list[str]:
    option_ids: list[str] = []
    for code in codes:
        option_ids.extend(option_ids_by_rpo.get(code, []))
    return option_ids


def ids_for_text(text: str, option_ids_by_rpo: dict[str, list[str]]) -> list[str]:
    return ids_for_codes(referenced_rpo_codes(text, option_ids_by_rpo), option_ids_by_rpo)


def candidate_rule_keys(
    options: list[dict[str, str]],
    option_ids_by_rpo: dict[str, list[str]],
    interior_codes: set[str],
) -> tuple[set[tuple[str, str, str]], list[dict[str, Any]], list[dict[str, Any]]]:
    candidates: set[tuple[str, str, str]] = set()
    review_rows: list[dict[str, Any]] = []
    unresolved_mentions: list[dict[str, Any]] = []

    for option in options:
        source_id = option.get("option_id", "")
        source_rpo = option.get("rpo", "")
        detail_raw = option.get("detail_raw", "")
        if not source_id or not detail_raw:
            continue
        for fragment in detail_fragments(detail_raw):
            lower = fragment.lower()
            mentioned_codes = set(referenced_rpo_codes(fragment, option_ids_by_rpo))
            for code in sorted(mentioned_codes):
                if code in interior_codes:
                    continue
                if code != source_rpo and not option_ids_by_rpo.get(code):
                    unresolved_mentions.append(
                        {
                            "option_id": source_id,
                            "rpo": source_rpo,
                            "mentioned_rpo": code,
                            "fragment": fragment,
                        }
                    )
            before_count = len(candidates)
            if "not available with" in lower:
                for target_id in ids_for_text(phrase_tail(fragment, "not available with"), option_ids_by_rpo):
                    candidates.add((source_id, "excludes", target_id))
            if "requires" in lower:
                tail = phrase_tail(fragment, "requires", (" or included with", " included with"))
                for target_id in ids_for_text(tail, option_ids_by_rpo):
                    candidates.add((source_id, "requires", target_id))
            if "includes" in lower and "included with" not in lower:
                for target_id in ids_for_text(phrase_tail(fragment, "includes", (" requires",)), option_ids_by_rpo):
                    candidates.add((source_id, "includes", target_id))
            if "included and only available with" in lower:
                for including_source_id in ids_for_text(phrase_tail(fragment, "included and only available with"), option_ids_by_rpo):
                    candidates.add((including_source_id, "includes", source_id))
            if "included with" in lower:
                for including_source_id in ids_for_text(phrase_tail(fragment, "included with"), option_ids_by_rpo):
                    candidates.add((including_source_id, "includes", source_id))
            if "only available with" in lower and "included" not in lower:
                for target_id in ids_for_text(phrase_tail(fragment, "only available with"), option_ids_by_rpo):
                    candidates.add((source_id, "requires", target_id))
            if len(candidates) == before_count and any(phrase in lower for phrase in RULE_PHRASES):
                if mentioned_codes and mentioned_codes <= interior_codes:
                    continue
                review_rows.append(
                    {
                        "option_id": source_id,
                        "rpo": source_rpo,
                        "reason": "supported_clause_without_resolved_rule",
                        "fragment": fragment,
                    }
                )
    return candidates, review_rows, unresolved_mentions


def rule_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (row.get("source_id", ""), row.get("rule_type", "").lower(), row.get("target_id", ""))


def full_rule_key(row: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        row.get("source_id", ""),
        row.get("rule_type", "").lower(),
        row.get("target_id", ""),
        row.get("body_style_scope", ""),
        row.get("runtime_action", ""),
    )


def grand_sport_rule_dedupe_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (
        row.get("source_id", ""),
        row.get("rule_type", "").lower(),
        row.get("target_id", ""),
    )


def grand_sport_rule_exact_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        row.get("source_id", ""),
        row.get("rule_type", "").lower(),
        row.get("target_id", ""),
        clean(row.get("body_style_scope", "")),
    )


def runtime_authored_rule(row: dict[str, Any]) -> bool:
    return not clean(row.get("generation_action", "")).lower().startswith("omit")


def canonical_rule_row(rows: list[dict[str, str]]) -> dict[str, str]:
    return sorted(
        rows,
        key=lambda row: (
            row.get("rule_id", "").startswith("gs_copy_"),
            len(row.get("rule_id", "")),
            bool(clean(row.get("body_style_scope", ""))),
            row.get("rule_id", ""),
        ),
    )[0]


def focused_duplicate_rule_rows(workbook_rules: list[dict[str, str]]) -> list[dict[str, Any]]:
    rows_by_key: dict[tuple[str, str, str, str], list[dict[str, str]]] = defaultdict(list)
    for row in workbook_rules:
        if runtime_authored_rule(row):
            rows_by_key[grand_sport_rule_exact_key(row)].append(row)
    duplicates = []
    for key, rows in sorted(rows_by_key.items()):
        if len(rows) <= 1:
            continue
        canonical = canonical_rule_row(rows)
        duplicates.append(
            {
                "source_id": key[0],
                "rule_type": key[1],
                "target_id": key[2],
                "body_style_scope": key[3],
                "rule_ids": [row.get("rule_id", "") for row in rows],
                "copied_rule_ids": [row.get("rule_id", "") for row in rows if row.get("rule_id", "").startswith("gs_copy_")],
                "canonical_rule_id": canonical.get("rule_id", ""),
                "recommended_action": "omit copied duplicate rows and keep canonical rule row",
                "count": len(rows),
            }
        )
    return duplicates


def option_label(option_id: str, options_by_id: dict[str, dict[str, str]]) -> str:
    option = options_by_id.get(option_id, {})
    return " ".join(part for part in (option.get("rpo", ""), option.get("option_name", "")) if part).strip()


def option_body_style_coverage(
    status_rows: list[dict[str, str]],
    variant_rows: list[dict[str, str]],
    variant_ids: tuple[str, ...],
) -> dict[str, list[str]]:
    body_style_by_variant = {
        row.get("variant_id", ""): row.get("body_style", "")
        for row in variant_rows
        if row.get("variant_id", "") in variant_ids
    }
    body_styles_by_option: dict[str, set[str]] = defaultdict(set)
    for row in status_rows:
        if row.get("status", "").lower() not in {"available", "standard"}:
            continue
        body_style = body_style_by_variant.get(row.get("variant_id", ""))
        option_id = row.get("option_id", "")
        if option_id and body_style:
            body_styles_by_option[option_id].add(body_style)
    return {option_id: sorted(body_styles) for option_id, body_styles in body_styles_by_option.items()}


def scoped_rule_dedupe_audit(
    workbook_rules: list[dict[str, str]],
    options_by_id: dict[str, dict[str, str]],
    body_styles_by_option: dict[str, list[str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows_by_key: dict[tuple[str, str, str], list[dict[str, str]]] = defaultdict(list)
    for row in workbook_rules:
        if runtime_authored_rule(row):
            rows_by_key[grand_sport_rule_dedupe_key(row)].append(row)

    overlapping_rows: list[dict[str, Any]] = []
    redundant_rows: list[dict[str, Any]] = []
    for key, rows in sorted(rows_by_key.items()):
        global_rows = [row for row in rows if not clean(row.get("body_style_scope", ""))]
        scoped_rows = [row for row in rows if clean(row.get("body_style_scope", ""))]
        if not global_rows or not scoped_rows:
            continue
        canonical = canonical_rule_row(global_rows)
        source_id, rule_type, target_id = key
        source_option = options_by_id.get(source_id, {})
        target_option = options_by_id.get(target_id, {})
        source_body_styles = body_styles_by_option.get(source_id, [])
        redundant_for_key: list[dict[str, Any]] = []
        for scoped_row in scoped_rows:
            scope = clean(scoped_row.get("body_style_scope", ""))
            rule_id = scoped_row.get("rule_id", "")
            if (
                rule_id.startswith("gs_copy_")
                and len(source_body_styles) == 1
                and source_body_styles[0] == scope
            ):
                redundant = {
                    "redundant_rule_id": rule_id,
                    "canonical_rule_id": canonical.get("rule_id", ""),
                    "source_id": source_id,
                    "source_rpo": source_option.get("rpo", ""),
                    "rule_type": rule_type,
                    "target_id": target_id,
                    "target_rpo": target_option.get("rpo", ""),
                    "body_style_scope": scope,
                    "source_option_body_styles": source_body_styles,
                    "reason": "A global canonical row already covers this relationship, and the source option is only available for the scoped body style.",
                    "recommended_generation_action": "omit_redundant_scoped_duplicate",
                }
                redundant_rows.append(redundant)
                redundant_for_key.append(redundant)
        overlapping_rows.append(
            {
                "source_id": source_id,
                "source_rpo": source_option.get("rpo", ""),
                "rule_type": rule_type,
                "target_id": target_id,
                "target_rpo": target_option.get("rpo", ""),
                "global_rule_ids": [row.get("rule_id", "") for row in global_rows],
                "scoped_rule_ids": [row.get("rule_id", "") for row in scoped_rows],
                "scopes": sorted({clean(row.get("body_style_scope", "")) for row in scoped_rows if clean(row.get("body_style_scope", ""))}),
                "source_option_body_styles": source_body_styles,
                "classification": "has_redundant_scoped_rows" if redundant_for_key else "scope_overlap_requires_review",
                "recommended_action": "omit redundant scoped copied rows only" if redundant_for_key else "review before changing workbook rows",
            }
        )
    return overlapping_rows, redundant_rows


def rule_reference_issues(
    workbook_rules: list[dict[str, str]],
    options_by_id: dict[str, dict[str, str]],
    option_ids_by_rpo: dict[str, list[str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    missing_references: list[dict[str, Any]] = []
    inactive_references: list[dict[str, Any]] = []
    active_ids_by_rpo: dict[str, list[str]] = defaultdict(list)
    for option_id, option in options_by_id.items():
        rpo = option.get("rpo", "")
        if rpo and active_source_row(option):
            active_ids_by_rpo[rpo].append(option_id)

    for row in workbook_rules:
        if not runtime_authored_rule(row):
            continue
        for field in ("source_id", "target_id"):
            type_field = "source_type" if field == "source_id" else "target_type"
            if (row.get(type_field, "option") or "option") != "option":
                continue
            option_id = row.get(field, "")
            if not option_id:
                missing_references.append(
                    {
                        "rule_id": row.get("rule_id", ""),
                        "field": field,
                        "option_id": option_id,
                        "message": "Rule row has a blank option reference.",
                    }
                )
                continue
            option = options_by_id.get(option_id)
            if not option:
                missing_references.append(
                    {
                        "rule_id": row.get("rule_id", ""),
                        "field": field,
                        "option_id": option_id,
                        "message": "Rule row references an option_id missing from grandSport_options.",
                    }
                )
                continue
            if not active_source_row(option):
                rpo = option.get("rpo", "")
                inactive_references.append(
                    {
                        "rule_id": row.get("rule_id", ""),
                        "field": field,
                        "option_id": option_id,
                        "rpo": rpo,
                        "label": option_label(option_id, options_by_id),
                        "canonical_active_option_ids": active_ids_by_rpo.get(rpo, []),
                        "same_rpo_option_ids": option_ids_by_rpo.get(rpo, []),
                        "source_id": row.get("source_id", ""),
                        "rule_type": row.get("rule_type", ""),
                        "target_id": row.get("target_id", ""),
                        "body_style_scope": row.get("body_style_scope", ""),
                    }
                )
    return missing_references, inactive_references


def engine_cover_rule_audit(
    workbook_rules: list[dict[str, str]],
    options_by_id: dict[str, dict[str, str]],
    inactive_references: list[dict[str, Any]],
) -> dict[str, Any]:
    engine_cover_option_ids = {
        option_id
        for option_id, option in options_by_id.items()
        if option.get("rpo", "") in ENGINE_COVER_RPOS
    }
    rows = []
    for row in workbook_rules:
        source_id = row.get("source_id", "")
        target_id = row.get("target_id", "")
        if source_id not in engine_cover_option_ids and target_id not in engine_cover_option_ids:
            continue
        rows.append(
            {
                "rule_id": row.get("rule_id", ""),
                "source_id": source_id,
                "source_rpo": options_by_id.get(source_id, {}).get("rpo", ""),
                "source_active": active_source_row(options_by_id.get(source_id, {})) if source_id in options_by_id else False,
                "rule_type": row.get("rule_type", ""),
                "target_id": target_id,
                "target_rpo": options_by_id.get(target_id, {}).get("rpo", ""),
                "target_active": active_source_row(options_by_id.get(target_id, {})) if target_id in options_by_id else False,
                "body_style_scope": row.get("body_style_scope", ""),
                "runtime_action": row.get("runtime_action", ""),
                "generation_action": row.get("generation_action", ""),
            }
        )
    inactive_engine_cover_refs = [
        row
        for row in inactive_references
        if row.get("rpo", "") in ENGINE_COVER_RPOS
        or row.get("source_id", "") in engine_cover_option_ids
        or row.get("target_id", "") in engine_cover_option_ids
    ]
    return {
        "rpos": sorted(ENGINE_COVER_RPOS),
        "option_ids": sorted(engine_cover_option_ids),
        "rules": rows,
        "inactiveReferences": inactive_engine_cover_refs,
    }


def exclusive_group_pairs(exclusive_groups: list[dict[str, Any]], exclusive_group_members: list[dict[str, Any]]) -> set[tuple[str, str]]:
    option_ids_by_group: dict[str, list[str]] = defaultdict(list)
    for member in exclusive_group_members:
        if active_source_row(member):
            option_ids_by_group[member.get("group_id", "")].append(member.get("option_id", ""))

    pairs: set[tuple[str, str]] = set()
    for group in exclusive_groups:
        if not active_source_row(group):
            continue
        option_ids = [option_id for option_id in option_ids_by_group.get(group.get("group_id", ""), []) if option_id]
        for source_id in option_ids:
            for target_id in option_ids:
                if source_id != target_id:
                    pairs.add((source_id, target_id))
    return pairs


def rule_group_pairs(
    rule_groups: list[dict[str, Any]],
    rule_group_members: list[dict[str, Any]],
    group_type: str,
) -> set[tuple[str, str]]:
    target_ids_by_group: dict[str, list[str]] = defaultdict(list)
    for member in rule_group_members:
        if active_source_row(member):
            target_ids_by_group[member.get("group_id", "")].append(member.get("target_id", ""))

    pairs: set[tuple[str, str]] = set()
    for group in rule_groups:
        if not active_source_row(group) or group.get("group_type", "") != group_type:
            continue
        source_id = group.get("source_id", "")
        for target_id in target_ids_by_group.get(group.get("group_id", ""), []):
            if source_id and target_id:
                pairs.add((source_id, target_id))
    return pairs


def workbook_rule_origin(row: dict[str, str], candidate_keys: set[tuple[str, str, str]]) -> str:
    rule_id = row.get("rule_id", "")
    if rule_id.startswith("gs_copy_"):
        return "workbook_copied_from_stingray"
    if rule_id.startswith("gs_approved_"):
        return "workbook_approved_manual"
    if rule_key(row) in candidate_keys:
        return "workbook_matches_detail_raw"
    return "workbook_manual_or_normalized"


def copied_rule_audit_rows(rules: list[dict[str, str]]) -> list[dict[str, Any]]:
    return [
        {
            "grand_sport_rule_id": row.get("rule_id", ""),
            "source_id": row.get("source_id", ""),
            "rule_type": row.get("rule_type", ""),
            "target_id": row.get("target_id", ""),
            "body_style_scope": row.get("body_style_scope", ""),
            "reason": "Workbook-authored Grand Sport rule id indicates a copied Stingray source.",
        }
        for row in rules
        if row.get("_audit_origin") == "workbook_copied_from_stingray"
    ]


def parsed_rule_audit_rows(rules: list[dict[str, str]]) -> list[dict[str, Any]]:
    return [
        {
            "rule_id": row.get("rule_id", ""),
            "source_id": row.get("source_id", ""),
            "rule_type": row.get("rule_type", ""),
            "target_id": row.get("target_id", ""),
            "body_style_scope": row.get("body_style_scope", ""),
            "matched_phrase": "workbook_matches_detail_raw",
        }
        for row in rules
        if row.get("_audit_origin") == "workbook_matches_detail_raw"
    ]


def review_hot_spots(
    options: list[dict[str, str]],
    option_ids_by_rpo: dict[str, list[str]],
    special_review_rpos: tuple[str, ...],
) -> dict[str, Any]:
    duplicate_rpos = []
    for rpo, option_ids in sorted(option_ids_by_rpo.items()):
        if len(option_ids) <= 1:
            continue
        duplicate_rpos.append(
            {
                "rpo": rpo,
                "option_ids": option_ids,
                "reason": "Duplicate active Grand Sport RPO; verify workbook rules point at the intended option variant.",
            }
        )

    special_mentions = []
    for option in options:
        text = " ".join(part for part in (option.get("detail_raw", ""), option.get("description", ""), option.get("option_name", "")) if part)
        mentioned = [rpo for rpo in special_review_rpos if re.search(rf"\b{re.escape(rpo)}\b", text)]
        if option.get("rpo", "") in special_review_rpos or mentioned:
            special_mentions.append(
                {
                    "option_id": option.get("option_id", ""),
                    "rpo": option.get("rpo", ""),
                    "mentioned_rpos": mentioned,
                    "detail_raw": option.get("detail_raw", ""),
                }
            )

    return {
        "duplicateRpos": duplicate_rpos,
        "specialPackageMentions": special_mentions,
    }


def render_audit_markdown(audit: dict[str, Any]) -> str:
    summary = audit["summary"]
    lines = [
        "# Grand Sport Rule Audit",
        "",
        f"Generated: `{audit['dataset']['generated_at']}`",
        f"Status: `{audit['dataset']['status']}`",
        "",
        "## Summary",
        "",
        f"- Workbook rule rows: {summary['finalWorkbookRuleRows']}",
        f"- Expected draft runtime rules: {summary['expectedDraftRuntimeRules']}",
        f"- Omitted duplicate exclusive-group rules: {summary['omittedDuplicateExclusiveGroup']}",
        f"- Raw detail candidate keys: {summary['rawDetailCandidateKeys']}",
        f"- Workbook rules matching detail raw: {summary['workbookRulesMatchingDetailRaw']}",
        f"- Workbook manual/approved/copied rules: {summary['workbookManualOrApprovedRules']}",
        f"- Skipped/review rows: {summary['skippedRequiresReview']}",
        f"- Unresolved non-interior RPO mentions: {summary['unresolvedRpoMentions']}",
        f"- Duplicate semantic rule keys: {summary['duplicateSemanticRuleKeys']}",
        f"- Exact duplicate rule rows: {summary['exactDuplicateRuleRows']}",
        f"- Overlapping scoped rule rows: {summary['overlappingScopedRuleRows']}",
        f"- Redundant scoped rule rows: {summary['redundantScopedRuleRows']}",
        f"- Missing option references: {summary['missingOptionReferences']}",
        f"- Inactive option references: {summary['inactiveOptionReferences']}",
        f"- Engine-cover inactive references: {summary['engineCoverInactiveReferences']}",
        "",
        "## Workbook-Copied Rules",
        "",
    ]
    for row in audit["copiedFromStingray"][:25]:
        lines.append(f"- `{row['grand_sport_rule_id']}`: {row['source_id']} {row['rule_type']} {row['target_id']}")
    if len(audit["copiedFromStingray"]) > 25:
        lines.append(f"- ... {len(audit['copiedFromStingray']) - 25} more")

    lines.extend(["", "## Workbook Rules Matching Detail Raw", ""])
    for row in audit["parsedFromDetailRaw"][:35]:
        scope = f" [{row['body_style_scope']}]" if row.get("body_style_scope") else ""
        lines.append(f"- `{row['rule_id']}`: {row['source_id']} {row['rule_type']} {row['target_id']}{scope}")
    if len(audit["parsedFromDetailRaw"]) > 35:
        lines.append(f"- ... {len(audit['parsedFromDetailRaw']) - 35} more")

    lines.extend(["", "## Omitted Duplicate Exclusive Group Rules", ""])
    for row in audit["omittedDuplicateExclusiveGroup"][:35]:
        lines.append(f"- `{row['rule_id']}`: {row['source_id']} excludes {row['target_id']}")
    if not audit["omittedDuplicateExclusiveGroup"]:
        lines.append("- none")

    lines.extend(["", "## Skipped Requires Review", ""])
    for row in audit["skippedRequiresReview"][:50]:
        lines.append(f"- `{row['rpo']}` {row['option_id']} [{row['reason']}]: {row['fragment']}")
    if not audit["skippedRequiresReview"]:
        lines.append("- none")

    lines.extend(["", "## Unresolved RPO Mentions", ""])
    for row in audit["unresolvedRpoMentions"][:50]:
        lines.append(f"- `{row['rpo']}` {row['option_id']} mentions `{row['mentioned_rpo']}`: {row['fragment']}")
    if not audit["unresolvedRpoMentions"]:
        lines.append("- none")

    lines.extend(["", "## Review Hot Spots", ""])
    lines.append(f"- Duplicate active RPOs: {len(audit['reviewHotSpots']['duplicateRpos'])}")
    lines.append(f"- Special package mention rows: {len(audit['reviewHotSpots']['specialPackageMentions'])}")

    lines.extend(["", "## Duplicate Semantic Rule Keys", ""])
    if audit["focusedReview"]["duplicateSemanticRuleKeys"]:
        for row in audit["focusedReview"]["duplicateSemanticRuleKeys"][:50]:
            scope = f" [{row['body_style_scope']}]" if row.get("body_style_scope") else ""
            lines.append(f"- {row['source_id']} {row['rule_type']} {row['target_id']}{scope}: {', '.join(row['rule_ids'])}")
    else:
        lines.append("- none")

    lines.extend(["", "## Exact Duplicate Rule Rows", ""])
    if audit["focusedReview"]["exactDuplicateRuleRows"]:
        for row in audit["focusedReview"]["exactDuplicateRuleRows"][:50]:
            scope = f" [{row['body_style_scope']}]" if row.get("body_style_scope") else ""
            lines.append(f"- {row['source_id']} {row['rule_type']} {row['target_id']}{scope}: keep `{row['canonical_rule_id']}`; rows: {', '.join(row['rule_ids'])}")
    else:
        lines.append("- none")

    lines.extend(["", "## Overlapping Scoped Rule Rows", ""])
    if audit["focusedReview"]["overlappingScopedRuleRows"]:
        for row in audit["focusedReview"]["overlappingScopedRuleRows"][:75]:
            scopes = ", ".join(row.get("scopes", [])) or "none"
            body_styles = ", ".join(row.get("source_option_body_styles", [])) or "none"
            lines.append(
                f"- {row['source_id']} {row['rule_type']} {row['target_id']}: scopes {scopes}; source body styles {body_styles}; {row['classification']}"
            )
    else:
        lines.append("- none")

    lines.extend(["", "## Redundant Scoped Rule Rows", ""])
    if audit["focusedReview"]["redundantScopedRuleRows"]:
        for row in audit["focusedReview"]["redundantScopedRuleRows"][:75]:
            body_styles = ", ".join(row.get("source_option_body_styles", [])) or "none"
            lines.append(
                f"- omit `{row['redundant_rule_id']}`; keep `{row['canonical_rule_id']}`; scope `{row['body_style_scope']}`; source body styles {body_styles}"
            )
    else:
        lines.append("- none")

    lines.extend(["", "## Missing Option References", ""])
    if audit["focusedReview"]["missingOptionReferences"]:
        for row in audit["focusedReview"]["missingOptionReferences"][:50]:
            lines.append(f"- `{row['rule_id']}` {row['field']} -> `{row['option_id']}`: {row['message']}")
    else:
        lines.append("- none")

    lines.extend(["", "## Inactive Option References", ""])
    if audit["focusedReview"]["inactiveOptionReferences"]:
        for row in audit["focusedReview"]["inactiveOptionReferences"][:75]:
            active_ids = ", ".join(row.get("canonical_active_option_ids", [])) or "none"
            lines.append(
                f"- `{row['rule_id']}` {row['field']} -> `{row['option_id']}` ({row['rpo']}): active same-RPO ids: {active_ids}"
            )
    else:
        lines.append("- none")

    lines.extend(["", "## Engine Cover Rule Focus", ""])
    engine_cover = audit["focusedReview"]["engineCoverRules"]
    lines.append(f"- Engine-cover rule rows: {len(engine_cover['rules'])}")
    lines.append(f"- Engine-cover inactive references: {len(engine_cover['inactiveReferences'])}")
    for row in engine_cover["inactiveReferences"][:50]:
        active_ids = ", ".join(row.get("canonical_active_option_ids", [])) or "none"
        lines.append(f"- `{row['rule_id']}` references inactive `{row['option_id']}` ({row['rpo']}); active same-RPO ids: {active_ids}")
    return "\n".join(lines) + "\n"


def write_rule_audit(
    config,
    options: list[dict[str, str]],
    all_options: list[dict[str, str]],
    status_rows: list[dict[str, str]],
    variant_rows: list[dict[str, str]],
    option_ids_by_rpo: dict[str, list[str]],
    candidate_keys: set[tuple[str, str, str]],
    review_rows: list[dict[str, Any]],
    unresolved_mentions: list[dict[str, Any]],
    workbook_rules: list[dict[str, str]],
    exclusive_groups: list[dict[str, str]],
    exclusive_group_members: list[dict[str, str]],
    rule_groups: list[dict[str, str]],
    rule_group_members: list[dict[str, str]],
) -> dict[str, str]:
    exclusive_group_excludes = exclusive_group_pairs(exclusive_groups, exclusive_group_members)
    rule_group_excludes = rule_group_pairs(rule_groups, rule_group_members, "excludes_any")
    grouped_excludes = exclusive_group_excludes | rule_group_excludes
    duplicate_keys = [key for key, count in Counter(full_rule_key(row) for row in workbook_rules).items() if count > 1]
    options_by_id = {row.get("option_id", ""): row for row in all_options if row.get("option_id", "")}
    all_option_ids_by_rpo: dict[str, list[str]] = defaultdict(list)
    for row in all_options:
        if row.get("option_id", "") and row.get("rpo", ""):
            all_option_ids_by_rpo[row.get("rpo", "")].append(row.get("option_id", ""))
    duplicate_semantic_rule_keys = focused_duplicate_rule_rows(workbook_rules)
    missing_references, inactive_references = rule_reference_issues(workbook_rules, options_by_id, all_option_ids_by_rpo)
    engine_cover_rules = engine_cover_rule_audit(workbook_rules, options_by_id, inactive_references)
    body_styles_by_option = option_body_style_coverage(status_rows, variant_rows, tuple(config.variant_ids))
    overlapping_scoped_rule_rows, redundant_scoped_rule_rows = scoped_rule_dedupe_audit(
        workbook_rules,
        options_by_id,
        body_styles_by_option,
    )
    annotated_rules = [
        {**row, "_audit_origin": workbook_rule_origin(row, candidate_keys)}
        for row in workbook_rules
    ]
    omitted_exclusive = [
        row
        for row in annotated_rules
        if runtime_authored_rule(row)
        and row.get("rule_type", "").lower() == "excludes"
        and row.get("generation_action", "") != "preserve_runtime_exclude"
        and (row.get("source_id", ""), row.get("target_id", "")) in grouped_excludes
    ]
    runtime_ids = runtime_available_option_ids(options, status_rows, tuple(config.variant_ids))
    runtime_rows = runtime_rule_rows(workbook_rules, runtime_ids, grouped_excludes)
    omitted_inactive_or_unemitted = [
        row
        for row in annotated_rules
        if not runtime_authored_rule(row)
        or ((row.get("source_type", "option") or "option") == "option" and row.get("source_id", "") not in runtime_ids)
        or ((row.get("target_type", "option") or "option") == "option" and row.get("target_id", "") not in runtime_ids)
    ]
    origin_counts = Counter(row["_audit_origin"] for row in annotated_rules)
    audit = {
        "dataset": {
            "name": "2027 Corvette Grand Sport rule audit",
            "model": config.model_label,
            "model_year": config.model_year,
            "source_workbook": config.workbook_path.name,
            "source_option_sheet": config.source_option_sheet,
            "source_rule_sheet": config.rule_mapping_sheet,
            "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "status": "rule_audit_generated",
        },
        "summary": {
            "copiedRuleCandidates": origin_counts["workbook_copied_from_stingray"],
            "rawDetailRuleCandidates": len(candidate_keys),
            "rawDetailCandidateKeys": len(candidate_keys),
            "dedupedOrSelfReferenceCandidates": len(duplicate_keys),
            "finalWorkbookRuleRows": len(workbook_rules),
            "expectedDraftRuntimeRules": len(runtime_rows),
            "omittedDuplicateExclusiveGroup": len(omitted_exclusive),
            "omittedInactiveOrUnemitted": len(omitted_inactive_or_unemitted),
            "skippedRequiresReview": len(review_rows),
            "unresolvedRpoMentions": len(unresolved_mentions),
            "exclusiveGroups": len(exclusive_groups),
            "exclusiveGroupMembers": len(exclusive_group_members),
            "ruleGroups": len(rule_groups),
            "ruleGroupMembers": len(rule_group_members),
            "groupedExclusionPairs": len(rule_group_excludes),
            "duplicateSemanticRuleKeys": len(duplicate_semantic_rule_keys),
            "exactDuplicateRuleRows": len(duplicate_semantic_rule_keys),
            "overlappingScopedRuleRows": len(overlapping_scoped_rule_rows),
            "redundantScopedRuleRows": len(redundant_scoped_rule_rows),
            "missingOptionReferences": len(missing_references),
            "inactiveOptionReferences": len(inactive_references),
            "engineCoverRuleRows": len(engine_cover_rules["rules"]),
            "engineCoverInactiveReferences": len(engine_cover_rules["inactiveReferences"]),
            "workbookRulesMatchingDetailRaw": origin_counts["workbook_matches_detail_raw"],
            "workbookManualOrApprovedRules": (
                origin_counts["workbook_manual_or_normalized"]
                + origin_counts["workbook_approved_manual"]
                + origin_counts["workbook_copied_from_stingray"]
            ),
        },
        "sourceSheets": {
            "optionSheet": config.source_option_sheet,
            "ruleSheet": config.rule_mapping_sheet,
            "ruleGroupsSheet": config.rule_groups_sheet,
            "ruleGroupMembersSheet": config.rule_group_members_sheet,
            "exclusiveGroupsSheet": config.exclusive_groups_sheet,
            "exclusiveMembersSheet": config.exclusive_group_members_sheet,
        },
        "copiedFromStingray": copied_rule_audit_rows(annotated_rules),
        "parsedFromDetailRaw": parsed_rule_audit_rows(annotated_rules),
        "omittedDeduped": [{"source_id": key[0], "rule_type": key[1], "target_id": key[2], "body_style_scope": key[3], "runtime_action": key[4]} for key in duplicate_keys],
        "omittedDuplicateExclusiveGroup": omitted_exclusive,
        "omittedInactiveOrUnemitted": omitted_inactive_or_unemitted,
        "skippedRequiresReview": review_rows,
        "unresolvedRpoMentions": unresolved_mentions,
        "reviewHotSpots": review_hot_spots(options, option_ids_by_rpo, tuple(config.special_rule_review_rpos)),
        "focusedReview": {
            "duplicateSemanticRuleKeys": duplicate_semantic_rule_keys,
            "exactDuplicateRuleRows": duplicate_semantic_rule_keys,
            "overlappingScopedRuleRows": overlapping_scoped_rule_rows,
            "redundantScopedRuleRows": redundant_scoped_rule_rows,
            "missingOptionReferences": missing_references,
            "inactiveOptionReferences": inactive_references,
            "engineCoverRules": engine_cover_rules,
        },
    }
    output_dir = config.output_dir / "inspection"
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "grand-sport-rule-audit.json"
    md_path = output_dir / "grand-sport-rule-audit.md"
    json_path.write_text(json.dumps(audit, indent=2), encoding="utf-8")
    md_path.write_text(render_audit_markdown(audit), encoding="utf-8")
    return {"json": str(json_path), "markdown": str(md_path)}


def main() -> None:
    config = GRAND_SPORT_MODEL
    wb = load_workbook(config.workbook_path, data_only=True, read_only=True)
    all_grand_sport_options = rows_from_sheet(wb, config.source_option_sheet)
    status_rows = rows_from_sheet(wb, config.status_sheet)
    variant_rows = rows_from_sheet(wb, "variant_master")
    grand_sport_options = [row for row in all_grand_sport_options if active_source_row(row)]
    _, active_option_ids_by_rpo = option_indexes(grand_sport_options)
    _, all_option_ids_by_rpo = option_indexes(all_grand_sport_options)
    interior_codes = interior_combination_codes(wb)
    candidate_keys, review_rows, unresolved_mentions = candidate_rule_keys(grand_sport_options, all_option_ids_by_rpo, interior_codes)
    workbook_rules = rows_from_sheet(wb, config.rule_mapping_sheet)
    exclusive_groups = rows_from_sheet(wb, config.exclusive_groups_sheet)
    exclusive_group_members = rows_from_sheet(wb, config.exclusive_group_members_sheet)
    rule_groups = rows_from_sheet(wb, config.rule_groups_sheet) if config.rule_groups_sheet in wb.sheetnames else []
    rule_group_members = rows_from_sheet(wb, config.rule_group_members_sheet) if config.rule_group_members_sheet in wb.sheetnames else []
    audit_paths = write_rule_audit(
        config,
        grand_sport_options,
        all_grand_sport_options,
        status_rows,
        variant_rows,
        active_option_ids_by_rpo,
        candidate_keys,
        review_rows,
        unresolved_mentions,
        workbook_rules,
        exclusive_groups,
        exclusive_group_members,
        rule_groups,
        rule_group_members,
    )

    print(
        json.dumps(
            {
                "rule_mapping_rows": len(workbook_rules),
                "copied_rule_candidates": sum(1 for row in workbook_rules if row.get("rule_id", "").startswith("gs_copy_")),
                "raw_detail_rule_candidates": len(candidate_keys),
                "exclusive_groups": len(exclusive_groups),
                "exclusive_group_members": len(exclusive_group_members),
                "rule_groups": len(rule_groups),
                "rule_group_members": len(rule_group_members),
                "skipped_requires_review": len(review_rows),
                "unresolved_rpo_mentions": len(unresolved_mentions),
                "rule_audit_artifacts": audit_paths,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
