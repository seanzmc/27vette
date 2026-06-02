#!/usr/bin/env python3
"""Apply approved Z06 Pass 2 interior/accessory cleanup rows.

Dry-run is the default. Use --write to safely save stingray_master.xlsx.
This script intentionally does not call apply_future_model_option_review.py.
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely  # noqa: E402

OPTION_SHEET = "z06_options"
VARIANT_OVERRIDES_SHEET = "z06_variant_overrides"
PRICE_RULES_SHEET = "z06_price_rules"
EXCLUSIVE_GROUPS_SHEET = "z06_exclusive_groups"
EXCLUSIVE_MEMBERS_SHEET = "z06_exclusive_members"

UQT_STANDARD_OVERRIDES = {
    "2lz_h07": "sec_2lte_001",
    "2lz_h67": "sec_2lte_001",
    "3lz_h07": "sec_3lte_001",
    "3lz_h67": "sec_3lte_001",
}
SEAT_PRICE_RULES = [
    ("z06_pr_3lz_ah2_seat_001", "AH2", 0, "3LZ standard AH2 GT2 seats should not add a charge.", "3LZ"),
    ("z06_pr_3lz_ae4_seat_001", "AE4", 595, "3LZ AE4 Competition Sport seats should add $595.", "3LZ"),
    ("z06_pr_1lz_uqt_001", "UQT", 1495, "1LZ selectable UQT should retain its approved charge while 2LZ/3LZ standard UQT stays unpriced.", "1LZ"),
]
ACCESSORY_ZERO_RULES = [
    ("z06_pr_pcq_vwe_zero", "PCQ", "VWE", "PCQ includes VWE, so VWE does not add a second charge."),
    ("z06_pr_pcq_vwt_zero", "PCQ", "VWT", "PCQ includes VWT, so VWT does not add a second charge."),
    ("z06_pr_pdy_ryt_zero", "PDY", "RYT", "PDY includes RYT, so RYT does not add a second charge."),
    ("z06_pr_pdy_s08_zero", "PDY", "S08", "PDY includes S08, so S08 does not add a second charge."),
    ("z06_pr_pef_cav_zero", "PEF", "CAV", "PEF includes CAV, so CAV does not add a second charge."),
    ("z06_pr_pef_ria_zero", "PEF", "RIA", "PEF includes RIA, so RIA does not add a second charge."),
]


@dataclass(frozen=True)
class Change:
    sheet: str
    row_number: int | str
    key: str
    field: str
    current: Any
    desired: Any
    reason: str


def headers(ws) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def index(ws) -> dict[str, int]:
    return {header: offset + 1 for offset, header in enumerate(headers(ws)) if header}


def rows(ws):
    idx = index(ws)
    for row_number in range(2, ws.max_row + 1):
        row = {header: ws.cell(row_number, col).value for header, col in idx.items()}
        if any(value not in (None, "") for value in row.values()):
            yield row_number, row


def price_equal(current: Any, desired: Any) -> bool:
    try:
        return float(clean(current) or 0) == float(desired)
    except ValueError:
        return False


def values_equal(field: str, current: Any, desired: Any) -> bool:
    if field in {"price", "price_value"}:
        return price_equal(current, desired)
    return clean(current) == clean(desired)


def set_cell(ws, row_number: int, field: str, desired: Any, changes: list[Change], *, key: str, reason: str) -> None:
    idx = index(ws)
    if field not in idx:
        return
    current = ws.cell(row_number, idx[field]).value
    if not values_equal(field, current, desired):
        changes.append(Change(ws.title, row_number, key, field, current, desired, reason))
        ws.cell(row_number, idx[field]).value = desired


def option_maps(wb) -> tuple[dict[str, dict[str, Any]], dict[str, str]]:
    by_rpo: dict[str, dict[str, Any]] = {}
    id_by_rpo: dict[str, str] = {}
    for row_number, row in rows(wb[OPTION_SHEET]):
        rpo = clean(row.get("rpo"))
        option_id = clean(row.get("option_id"))
        if rpo and option_id:
            by_rpo[rpo] = {**row, "_row_number": row_number}
            id_by_rpo[rpo] = option_id
    return by_rpo, id_by_rpo


def require_rpos(id_by_rpo: dict[str, str], rpos: list[str] | tuple[str, ...]) -> None:
    missing = [rpo for rpo in rpos if rpo not in id_by_rpo]
    if missing:
        raise RuntimeError(f"Missing required Z06 RPO(s): {', '.join(missing)}")


def ensure_row_by_key(ws, key_field: str, key: str, values: dict[str, Any], changes: list[Change], *, reason: str) -> None:
    row_number = None
    for candidate_number, row in rows(ws):
        if clean(row.get(key_field)) == key:
            row_number = candidate_number
            break
    if row_number is None:
        row_number = ws.max_row + 1
        idx = index(ws)
        for field, value in values.items():
            if field in idx:
                ws.cell(row_number, idx[field]).value = value
        changes.append(Change(ws.title, f"new:{row_number}", key, "row", None, values, reason))
        return
    for field, desired in values.items():
        if field == key_field:
            continue
        set_cell(ws, row_number, field, desired, changes, key=key, reason=reason)


def ensure_variant_override(wb, changes: list[Change], option_id: str, variant_id: str, section_id: str) -> None:
    ws = wb[VARIANT_OVERRIDES_SHEET]
    existing = None
    for row_number, row in rows(ws):
        if clean(row.get("option_id")) == option_id and clean(row.get("variant_id")) == variant_id:
            existing = row_number
            break
    values = {
        "option_id": option_id,
        "variant_id": variant_id,
        "selectable": False,
        "display_behavior": "display_only",
        "section_id": section_id,
        "active": True,
        "note": f"{variant_id} includes UQT as display-only standard equipment.",
    }
    key = f"{option_id}:{variant_id}"
    if existing is None:
        row_number = ws.max_row + 1
        idx = index(ws)
        for field, value in values.items():
            if field in idx:
                ws.cell(row_number, idx[field]).value = value
        changes.append(Change(ws.title, f"new:{row_number}", key, "row", None, values, "Z06 UQT trim-scoped standard equipment override"))
    else:
        for field, value in values.items():
            if field in {"option_id", "variant_id"}:
                continue
            set_cell(ws, existing, field, value, changes, key=key, reason="Z06 UQT trim-scoped standard equipment override")


def ensure_price_rule(wb, changes: list[Change], rule_id: str, condition_rpo: str, target_rpo: str, price: int, notes: str, *, trim: str = "*") -> None:
    _, id_by_rpo = option_maps(wb)
    require_rpos(id_by_rpo, [condition_rpo, target_rpo])
    ensure_row_by_key(
        wb[PRICE_RULES_SHEET],
        "price_rule_id",
        rule_id,
        {
            "price_rule_id": rule_id,
            "condition_option_id": id_by_rpo[condition_rpo],
            "price_rule_type": "override",
            "target_option_id": id_by_rpo[target_rpo],
            "price_value": price,
            "body_style_scope": "*",
            "trim_level_scope": trim,
            "review_flag": False,
            "notes": notes,
        },
        changes,
        reason="approved Z06 Pass 2 price rule",
    )


def apply_options(wb, changes: list[Change]) -> None:
    by_rpo, id_by_rpo = option_maps(wb)
    require_rpos(id_by_rpo, ["UQT", "N3W", "FA5", "FA6", "AH2", "AE4", "PCQ", "VWE", "VWT", "PDY", "RYT", "S08", "PEF", "CAV", "RIA"])
    n3w = by_rpo["N3W"]
    set_cell(wb[OPTION_SHEET], int(n3w["_row_number"]), "active", "False", changes, key="N3W", reason="N3W should be represented as interior component/standard equipment, not as a customer option card")

    uqt = by_rpo["UQT"]
    set_cell(wb[OPTION_SHEET], int(uqt["_row_number"]), "price", 0, changes, key="UQT", reason="UQT is standard on 2LZ/3LZ; 1LZ charge is represented by a trim-scoped price rule")

    for variant_id, section_id in UQT_STANDARD_OVERRIDES.items():
        ensure_variant_override(wb, changes, id_by_rpo["UQT"], variant_id, section_id)

    for rule_id, rpo, price, notes, trim in SEAT_PRICE_RULES:
        ensure_price_rule(wb, changes, rule_id, rpo, rpo, price, notes, trim=trim)

    for rule_id, source_rpo, target_rpo, notes in ACCESSORY_ZERO_RULES:
        ensure_price_rule(wb, changes, rule_id, source_rpo, target_rpo, 0, notes)

    ensure_row_by_key(
        wb[EXCLUSIVE_GROUPS_SHEET],
        "group_id",
        "z06_excl_fa5_fa6_interior_trim",
        {
            "group_id": "z06_excl_fa5_fa6_interior_trim",
            "selection_mode": "single_within_group",
            "active": "True",
            "notes": "Z06 FA5 and FA6 carbon-fiber interior trim options are mutually exclusive.",
        },
        changes,
        reason="approved Z06 FA5/FA6 mutual exclusivity",
    )
    members_ws = wb[EXCLUSIVE_MEMBERS_SHEET]
    for display_order, rpo in enumerate(("FA5", "FA6"), start=1):
        option_id = id_by_rpo[rpo]
        existing = None
        for row_number, row in rows(members_ws):
            if clean(row.get("group_id")) == "z06_excl_fa5_fa6_interior_trim" and clean(row.get("option_id")) == option_id:
                existing = row_number
                break
        key = f"z06_excl_fa5_fa6_interior_trim:{option_id}"
        values = {"group_id": "z06_excl_fa5_fa6_interior_trim", "option_id": option_id, "display_order": display_order * 10, "active": "True"}
        if existing is None:
            row_number = members_ws.max_row + 1
            idx = index(members_ws)
            for field, value in values.items():
                members_ws.cell(row_number, idx[field]).value = value
            changes.append(Change(members_ws.title, f"new:{row_number}", key, "row", None, values, "approved Z06 FA5/FA6 mutual exclusivity member"))
        else:
            for field, value in values.items():
                if field in {"group_id", "option_id"}:
                    continue
                set_cell(members_ws, existing, field, value, changes, key=key, reason="approved Z06 FA5/FA6 mutual exclusivity member")


def build_plan(wb) -> list[Change]:
    required = [OPTION_SHEET, VARIANT_OVERRIDES_SHEET, PRICE_RULES_SHEET, EXCLUSIVE_GROUPS_SHEET, EXCLUSIVE_MEMBERS_SHEET]
    missing = [sheet for sheet in required if sheet not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Missing required sheets: {', '.join(missing)}")
    changes: list[Change] = []
    apply_options(wb, changes)
    return changes


def verify_saved(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        by_rpo, id_by_rpo = option_maps(wb)
        checks: dict[str, Any] = {
            "n3w_active": clean(by_rpo["N3W"].get("active")),
            "uqt_overrides": 0,
            "seat_price_rules": {},
            "accessory_zero_rules": {},
            "fa_group_members": [],
        }
        if checks["n3w_active"] != "False":
            raise RuntimeError(f"N3W should be inactive, got {checks['n3w_active']!r}")
        for variant_id, section_id in UQT_STANDARD_OVERRIDES.items():
            matched = [row for _, row in rows(wb[VARIANT_OVERRIDES_SHEET]) if clean(row.get("option_id")) == id_by_rpo["UQT"] and clean(row.get("variant_id")) == variant_id and clean(row.get("section_id")) == section_id and clean(row.get("selectable")) == "False" and clean(row.get("display_behavior")) == "display_only" and clean(row.get("active")) == "True"]
            if not matched:
                raise RuntimeError(f"Missing UQT standard override for {variant_id}")
            checks["uqt_overrides"] += 1
        price_rules = {clean(row.get("price_rule_id")): row for _, row in rows(wb[PRICE_RULES_SHEET])}
        for rule_id, rpo, price, _notes, trim in SEAT_PRICE_RULES:
            row = price_rules.get(rule_id)
            if not row or clean(row.get("target_option_id")) != id_by_rpo[rpo] or clean(row.get("price_value")) != str(price) or clean(row.get("trim_level_scope")) != trim:
                raise RuntimeError(f"Missing/corrupt seat price rule {rule_id}")
            checks["seat_price_rules"][rule_id] = clean(row.get("price_value"))
        for rule_id, source_rpo, target_rpo, _notes in ACCESSORY_ZERO_RULES:
            row = price_rules.get(rule_id)
            if not row or clean(row.get("condition_option_id")) != id_by_rpo[source_rpo] or clean(row.get("target_option_id")) != id_by_rpo[target_rpo] or clean(row.get("price_value")) != "0":
                raise RuntimeError(f"Missing/corrupt accessory zero price rule {rule_id}")
            checks["accessory_zero_rules"][rule_id] = "0"
        member_ids = {clean(row.get("option_id")) for _, row in rows(wb[EXCLUSIVE_MEMBERS_SHEET]) if clean(row.get("group_id")) == "z06_excl_fa5_fa6_interior_trim" and clean(row.get("active")) == "True"}
        if member_ids != {id_by_rpo["FA5"], id_by_rpo["FA6"]}:
            raise RuntimeError(f"Missing FA5/FA6 exclusive members: {member_ids}")
        checks["fa_group_members"] = sorted(member_ids)
        return checks
    finally:
        wb.close()


def summarize(changes: list[Change]) -> dict[str, Any]:
    by_sheet: dict[str, int] = {}
    by_field: dict[str, int] = {}
    for change in changes:
        by_sheet[change.sheet] = by_sheet.get(change.sheet, 0) + 1
        by_field[change.field] = by_field.get(change.field, 0) + 1
    return {"total_changes": len(changes), "changes_by_sheet": dict(sorted(by_sheet.items())), "changes_by_field": dict(sorted(by_field.items()))}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="Write workbook changes. Default is dry-run.")
    parser.add_argument("--include-changes", action="store_true", help="Include per-cell changes in JSON report.")
    args = parser.parse_args(argv)

    if args.write and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")
    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    backup_path = None
    try:
        changes = build_plan(wb)
        if not args.write:
            report: dict[str, Any] = {"status": "dry_run", **summarize(changes), "workbook": str(WORKBOOK_PATH)}
            if args.include_changes:
                report["changes"] = [change.__dict__ for change in changes]
            print(json.dumps(report, indent=2, default=str))
            return 0
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved(WORKBOOK_PATH)
    report = {"status": "written", **summarize(changes), "workbook": str(WORKBOOK_PATH), "backup": str(backup_path), "verification": verification}
    if args.include_changes:
        report["changes"] = [change.__dict__ for change in changes]
    print(json.dumps(report, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
