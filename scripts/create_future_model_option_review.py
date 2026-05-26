#!/usr/bin/env python3
"""Create a simplified all-row future_model_option_review sheet."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import FUTURE_MODEL_SPECS  # noqa: E402
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely, write_sheet  # noqa: E402

SOURCE_REVIEW_SHEET = "future_model_source_review"
OPTION_REVIEW_SHEET = "future_model_option_review"

OPTION_REVIEW_HEADERS = [
    "model_key",
    "raw_source_sheet",
    "raw_source_span",
    "orderable_rpo",
    "ref_only_rpo",
    "source_rpo",
    "source_option_description",
    "source_disclosure_raw",
    "raw_status_summary",
    "normalized_status_summary",
    "suggested_option_id",
    "suggested_section_id",
    "suggested_display_order",
    "suggested_copy_from",
    "final_option_id",
    "final_option_name",
    "final_description",
    "final_detail_raw",
    "final_section_id",
    "final_selectable",
    "final_display_order",
    "final_display_behavior",
    "review_status",
    "active",
    "notes",
]

HUMAN_DECISION_FIELDS = {
    "final_option_id",
    "final_option_name",
    "final_description",
    "final_detail_raw",
    "final_section_id",
    "final_selectable",
    "final_display_order",
    "final_display_behavior",
    "review_status",
    "active",
    "notes",
}


def row_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        clean(row.get("model_key")),
        clean(row.get("raw_source_sheet")),
        clean(row.get("raw_source_span")),
        clean(row.get("source_rpo")),
    )


def rows_from_sheet(wb, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header: clean(value) for header, value in zip(headers, values) if header}
        if any(row.values()):
            rows.append(row)
    return rows


def existing_option_review_rows(wb) -> dict[tuple[str, str, str, str], dict[str, str]]:
    return {row_key(row): row for row in rows_from_sheet(wb, OPTION_REVIEW_SHEET)}


def _variant_ids_for_model(model_key: str) -> list[str]:
    spec = FUTURE_MODEL_SPECS.get(model_key)
    if not spec:
        return []
    return list(spec.variant_columns.values())


def _status_summary(source_row: dict[str, Any], *, prefix: str, variant_ids: list[str]) -> str:
    parts: list[str] = []
    for variant_id in variant_ids:
        value = clean(source_row.get(f"{prefix}_{variant_id}"))
        if value:
            parts.append(f"{variant_id}={value}")
    return "; ".join(parts)


def _suggested_copy_from(source_row: dict[str, Any]) -> str:
    model_key = clean(source_row.get("copy_from_model_key"))
    option_id = clean(source_row.get("copy_from_option_id"))
    if model_key and option_id:
        return f"{model_key}:{option_id}"
    return option_id or model_key


def simplified_row(source_row: dict[str, Any]) -> dict[str, Any]:
    model_key = clean(source_row.get("model_key"))
    orderable_rpo = clean(source_row.get("source_orderable_rpo"))
    ref_only_rpo = clean(source_row.get("source_ref_rpo"))
    source_rpo = orderable_rpo or ref_only_rpo
    variant_ids = _variant_ids_for_model(model_key)
    return {
        "model_key": model_key,
        "raw_source_sheet": clean(source_row.get("raw_source_sheets")),
        "raw_source_span": clean(source_row.get("raw_source_spans")),
        "orderable_rpo": orderable_rpo,
        "ref_only_rpo": ref_only_rpo,
        "source_rpo": source_rpo,
        "source_option_description": clean(source_row.get("source_option_description")),
        "source_disclosure_raw": clean(source_row.get("source_disclosure_raw")),
        "raw_status_summary": _status_summary(source_row, prefix="raw_status", variant_ids=variant_ids),
        "normalized_status_summary": _status_summary(source_row, prefix="status", variant_ids=variant_ids),
        "suggested_option_id": clean(source_row.get("approved_option_id")) or clean(source_row.get("candidate_option_id")),
        "suggested_section_id": clean(source_row.get("approved_section_id")),
        "suggested_display_order": clean(source_row.get("approved_display_order")),
        "suggested_copy_from": _suggested_copy_from(source_row),
        "final_option_id": clean(source_row.get("approved_option_id")),
        "final_option_name": clean(source_row.get("approved_option_name")),
        "final_description": clean(source_row.get("approved_description")),
        "final_detail_raw": clean(source_row.get("approved_detail_raw")),
        "final_section_id": clean(source_row.get("approved_section_id")),
        "final_selectable": clean(source_row.get("approved_selectable")),
        "final_display_order": clean(source_row.get("approved_display_order")),
        "final_display_behavior": clean(source_row.get("approved_display_behavior")),
        "review_status": clean(source_row.get("review_status")),
        "active": clean(source_row.get("active")),
        "notes": clean(source_row.get("notes")),
    }


def build_option_review_rows(wb, *, preserve_existing: bool = True) -> list[dict[str, Any]]:
    source_rows = rows_from_sheet(wb, SOURCE_REVIEW_SHEET)
    existing_rows = existing_option_review_rows(wb) if preserve_existing else {}
    rows: list[dict[str, Any]] = []
    for source_row in source_rows:
        row = simplified_row(source_row)
        prior = existing_rows.get(row_key(row))
        if prior:
            for field in HUMAN_DECISION_FIELDS:
                if clean(prior.get(field)):
                    row[field] = prior[field]
        rows.append(row)
    return rows


def verify_saved_rows(path: Path, expected_rows: int) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if OPTION_REVIEW_SHEET not in wb.sheetnames:
            raise RuntimeError(f"{OPTION_REVIEW_SHEET} was not created")
        ws = wb[OPTION_REVIEW_SHEET]
        headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        if headers != OPTION_REVIEW_HEADERS:
            raise RuntimeError(f"{OPTION_REVIEW_SHEET} header drift: {headers}")
        data_rows = max(ws.max_row - 1, 0)
        if data_rows != expected_rows:
            raise RuntimeError(f"Expected {expected_rows} rows, found {data_rows}")
        model_col = headers.index("model_key") + 1
        status_col = headers.index("review_status") + 1
        counts_by_model: dict[str, int] = {}
        counts_by_status: dict[str, int] = {}
        for row_number in range(2, ws.max_row + 1):
            model_key = clean(ws.cell(row_number, model_col).value)
            status = clean(ws.cell(row_number, status_col).value)
            counts_by_model[model_key] = counts_by_model.get(model_key, 0) + 1
            counts_by_status[status] = counts_by_status.get(status, 0) + 1
        return {
            "rows": data_rows,
            "counts_by_model": counts_by_model,
            "counts_by_status": counts_by_status,
        }
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reset-review-fields", action="store_true", help="Do not preserve existing final/review fields in the simplified sheet.")
    args = parser.parse_args(argv)

    lock_path = excel_lock_path(WORKBOOK_PATH)
    if lock_path.exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {lock_path}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    try:
        rows = build_option_review_rows(wb, preserve_existing=not args.reset_review_fields)
        write_sheet(wb, OPTION_REVIEW_SHEET, OPTION_REVIEW_HEADERS, rows)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_rows(WORKBOOK_PATH, len(rows))
    print(json.dumps({"workbook": str(WORKBOOK_PATH), "backup": str(backup_path), "verification": verification}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
