#!/usr/bin/env python3
"""Generate the Stingray form contract and static-app data from stingray_master.xlsx."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook
from corvette_form_generator.mapping import (
    best_status,
    normalize_mode,
    selection_mode_label as shared_selection_mode_label,
    status_to_label,
    step_for_section as shared_step_for_section,
)
from corvette_form_generator.model_configs import GRAND_SPORT_MODEL, STINGRAY_MODEL
from corvette_form_generator.output import write_app_data_registry, write_json_output
from corvette_form_generator.registry_promotion import build_registry_from_promotions, live_contract_data
from corvette_form_generator.runtime_metadata import (
    load_context_sections,
    load_default_selection_rules,
    load_interior_components,
    load_model_config_overrides,
    load_order_summary_metadata,
    load_runtime_rule_exceptions,
    load_runtime_steps,
    load_section_presentation,
    load_variant_option_overrides,
    presentation_bool,
)
from corvette_form_generator.validation import validation_error_count
from corvette_form_generator.workbook import clean, intish, money, rows_from_sheet, save_workbook_safely, write_sheet


MODEL_CONFIG = STINGRAY_MODEL
ROOT = MODEL_CONFIG.root
WORKBOOK_PATH = MODEL_CONFIG.workbook_path
OUTPUT_DIR = MODEL_CONFIG.output_dir
APP_DIR = MODEL_CONFIG.app_dir
INTERIOR_REFERENCE_PATH = MODEL_CONFIG.interior_reference_path
GENERATED_SHEETS = list(MODEL_CONFIG.generated_sheets)
STEP_ORDER = list(MODEL_CONFIG.step_order)
STEP_LABELS = dict(MODEL_CONFIG.step_labels)
CONTEXT_SECTIONS = [dict(section) for section in MODEL_CONFIG.context_sections]
SECTION_STEP_OVERRIDES = dict(MODEL_CONFIG.section_step_overrides)
BODY_STYLE_DISPLAY_ORDER = dict(MODEL_CONFIG.body_style_display_order)
SELECTION_MODE_LABELS = dict(MODEL_CONFIG.selection_mode_labels)
STANDARD_SECTIONS = set(MODEL_CONFIG.standard_sections)


def export_slug(model_key: str) -> str:
    return model_key.replace("_", "-")


def registry_model_key(model_key: str) -> str:
    return "grandSport" if model_key == "grand_sport" else model_key


def model_registry_entry(
    model_key: str,
    model_label: str,
    data: dict[str, Any],
    asset: dict[str, str] | None = None,
) -> dict[str, Any]:
    entry = {
        "key": registry_model_key(model_key),
        "label": model_label,
        "modelName": f"Corvette {model_label}",
        "exportSlug": export_slug(model_key),
        "data": data,
    }
    if asset and asset.get("image_url"):
        entry.update(asset)
    return entry


def workbook_truthy(value: Any) -> bool:
    return clean(value).lower() in {"true", "yes", "1", "y"}


ASSET_IMAGE_FIELDS = ("image_url", "image_alt", "image_fit", "image_position")


def asset_fields(row: dict[str, Any]) -> dict[str, str]:
    return {
        "image_url": clean(row.get("image_url")),
        "image_alt": clean(row.get("image_alt")),
        "image_fit": clean(row.get("image_fit")),
        "image_position": clean(row.get("image_position")),
    }


def load_asset_map(wb, model_key: str, target_type: str) -> dict[str, dict[str, str]]:
    if "asset_map" not in wb.sheetnames:
        return {}
    assets: dict[str, dict[str, str]] = {}
    for row in rows_from_sheet(wb, "asset_map"):
        if not workbook_truthy(row.get("active")):
            continue
        if clean(row.get("model_key")) != model_key:
            continue
        if clean(row.get("target_type")) != target_type:
            continue
        target_id = clean(row.get("target_id"))
        fields = asset_fields(row)
        if target_id and fields["image_url"]:
            assets[target_id] = fields
    return assets


def load_model_asset_map() -> dict[str, dict[str, str]]:
    if not WORKBOOK_PATH.exists():
        return {}
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        if "asset_map" not in wb.sheetnames:
            return {}
        assets: dict[str, dict[str, str]] = {}
        for row in rows_from_sheet(wb, "asset_map"):
            if not workbook_truthy(row.get("active")):
                continue
            if clean(row.get("target_type")) != "model":
                continue
            model_key = clean(row.get("model_key"))
            target_id = clean(row.get("target_id")) or registry_model_key(model_key)
            fields = asset_fields(row)
            if not target_id or not fields["image_url"]:
                continue
            assets[target_id] = fields
        return assets
    finally:
        wb.close()


def context_choice_copy_rows(wb, model_key: str) -> list[dict[str, str]]:
    if "context_choice_copy" not in wb.sheetnames:
        return []
    rows: list[dict[str, str]] = []
    for row in rows_from_sheet(wb, "context_choice_copy"):
        if not workbook_truthy(row.get("active")):
            continue
        row_model = clean(row.get("model_key")) or "*"
        if row_model not in {"*", model_key}:
            continue
        if clean(row.get("info_tooltip")):
            rows.append(row)
    return rows


def context_choice_info_tooltip(
    copy_rows: list[dict[str, str]],
    *,
    model_key: str,
    context_type: str,
    value: str,
    body_style: str = "",
) -> str:
    context_type_key = clean(context_type).lower()
    value_key = clean(value).lower()
    body_style_key = clean(body_style).lower()
    best: tuple[int, str] = (-1, "")
    for row in copy_rows:
        row_context_type = clean(row.get("context_type")).lower()
        row_value = clean(row.get("value")).lower()
        row_model = clean(row.get("model_key")) or "*"
        row_body_style = (clean(row.get("body_style")) or "*").lower()
        if row_context_type != context_type_key or row_value != value_key:
            continue
        if row_model not in {"*", model_key}:
            continue
        if row_body_style not in {"*", body_style_key}:
            continue
        score = (2 if row_model == model_key else 0) + (1 if row_body_style == body_style_key else 0)
        tooltip = clean(row.get("info_tooltip"))
        if tooltip and score > best[0]:
            best = (score, tooltip)
    return best[1]


def load_grand_sport_registry_data() -> dict[str, Any] | None:
    draft_path = OUTPUT_DIR / "inspection" / f"{GRAND_SPORT_MODEL.draft_artifact_prefix}.json"
    if draft_path.exists():
        return json.loads(draft_path.read_text(encoding="utf-8"))
    app_data_path = APP_DIR / "data.js"
    if app_data_path.exists():
        app_data = app_data_path.read_text(encoding="utf-8")
        try:
            registry_json = app_data.split("window.CORVETTE_FORM_DATA = ", 1)[1].split(
                ";\nwindow.STINGRAY_FORM_DATA",
                1,
            )[0]
            return json.loads(registry_json).get("models", {}).get("grandSport", {}).get("data")
        except (IndexError, json.JSONDecodeError):
            pass
    return None


def refresh_grand_sport_registry_source() -> None:
    return


def build_app_data_registry(stingray_data: dict[str, Any]) -> tuple[dict[str, Any], dict[str, str]]:
    model_assets = load_model_asset_map()
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        promoted_registry = build_registry_from_promotions(
            wb,
            current_model_key=MODEL_CONFIG.model_key,
            current_data=stingray_data,
            model_assets=model_assets,
            root=ROOT,
        )
    finally:
        wb.close()
    if promoted_registry is not None:
        legacy_aliases = promoted_registry.pop("legacyAliases", {})
        return promoted_registry, legacy_aliases

    stingray_key = registry_model_key(MODEL_CONFIG.model_key)
    models = {
        stingray_key: model_registry_entry(
            MODEL_CONFIG.model_key,
            MODEL_CONFIG.model_label,
            stingray_data,
            model_assets.get(stingray_key),
        ),
    }
    grand_sport_data = load_grand_sport_registry_data()
    if grand_sport_data is not None:
        grand_sport_data = live_contract_data(grand_sport_data)
        grand_sport_key = registry_model_key(GRAND_SPORT_MODEL.model_key)
        models[grand_sport_key] = model_registry_entry(
            GRAND_SPORT_MODEL.model_key,
            GRAND_SPORT_MODEL.model_label,
            grand_sport_data,
            model_assets.get(grand_sport_key),
        )
    return {
        "defaultModelKey": "stingray",
        "models": models,
    }, {"STINGRAY_FORM_DATA": "stingray"}

def step_for_section(
    section_id: str,
    section_name: str,
    section_step_key: str = "",
    *,
    standard_sections: set[str] | frozenset[str] | None = None,
) -> str:
    return shared_step_for_section(
        section_id,
        section_name,
        section_step_key=section_step_key,
        standard_sections=standard_sections or STANDARD_SECTIONS,
        section_step_overrides=SECTION_STEP_OVERRIDES,
    )


def selection_mode_label(selection_mode: str) -> str:
    return shared_selection_mode_label(selection_mode, SELECTION_MODE_LABELS)


def workbook_bool(row: dict[str, str], field: str, fallback: bool) -> bool:
    value = clean(row.get(field, ""))
    if value in {"True", "False"}:
        return value == "True"
    return fallback


def active_source_row(row: dict[str, str]) -> bool:
    return clean(row.get("active", "True")) == "True"


def runtime_authored_rule(row: dict[str, str]) -> bool:
    status = clean(row.get("normalization_status", "")).lower()
    if status in {"omitted", "replaced"}:
        return False
    if status == "preserved":
        return True
    return not clean(row.get("generation_action", "")).lower().startswith("omit")


def rows_from_optional_sheet(wb, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    return rows_from_sheet(wb, sheet_name)


def load_rule_groups(wb) -> list[dict[str, Any]]:
    members_by_group: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows_from_optional_sheet(wb, MODEL_CONFIG.rule_group_members_sheet):
        if active_source_row(row):
            members_by_group[row.get("group_id", "")].append(row)

    rule_groups: list[dict[str, Any]] = []
    for row in rows_from_optional_sheet(wb, MODEL_CONFIG.rule_groups_sheet):
        if not active_source_row(row):
            continue
        group_id = row.get("group_id", "")
        members = sorted(members_by_group.get(group_id, []), key=lambda member: intish(member.get("display_order")))
        rule_groups.append(
            {
                "group_id": group_id,
                "group_type": row.get("group_type", ""),
                "source_id": row.get("source_id", ""),
                "target_ids": [member.get("target_id", "") for member in members if member.get("target_id", "")],
                "body_style_scope": row.get("body_style_scope", ""),
                "trim_level_scope": row.get("trim_level_scope", ""),
                "variant_scope": row.get("variant_scope", ""),
                "disabled_reason": row.get("disabled_reason", ""),
                "active": row.get("active", ""),
                "notes": row.get("notes", ""),
            }
        )
    return rule_groups


def load_exclusive_groups(wb) -> list[dict[str, Any]]:
    members_by_group: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows_from_optional_sheet(wb, MODEL_CONFIG.exclusive_group_members_sheet):
        if active_source_row(row):
            members_by_group[row.get("group_id", "")].append(row)

    exclusive_groups: list[dict[str, Any]] = []
    for row in rows_from_optional_sheet(wb, MODEL_CONFIG.exclusive_groups_sheet):
        if not active_source_row(row):
            continue
        group_id = row.get("group_id", "")
        members = sorted(members_by_group.get(group_id, []), key=lambda member: intish(member.get("display_order")))
        exclusive_groups.append(
            {
                "group_id": group_id,
                "option_ids": [member.get("option_id", "") for member in members if member.get("option_id", "")],
                "selection_mode": row.get("selection_mode", ""),
                "active": row.get("active", ""),
                "notes": row.get("notes", ""),
            }
        )
    return exclusive_groups


def grouped_requirement_pairs(rule_groups: list[dict[str, Any]]) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for group in rule_groups:
        if group.get("active") != "True" or group.get("group_type") != "requires_any":
            continue
        source_id = group.get("source_id", "")
        for target_id in group.get("target_ids", []):
            pairs.add((source_id, target_id))
    return pairs


def option_key(option: dict[str, str]) -> str:
    return option["option_id"]


def interior_price(row: dict[str, str]) -> int:
    return money(row.get("Price") or row.get("Cost"))


def price_ref_key(trim: str, code: str) -> tuple[str, str]:
    return (clean(trim).replace("_", " "), clean(code))


def price_ref_prices(rows: list[dict[str, str]]) -> dict[tuple[str, str], int]:
    prices: dict[tuple[str, str], int] = {}
    for row in rows:
        if clean(row.get("OptionType", "")).lower() != "seat":
            continue
        trim = clean(row.get("Trim", ""))
        code = clean(row.get("Code", ""))
        if trim and code:
            prices[price_ref_key(trim, code)] = money(row.get("Price"))
    return prices


def price_ref_component_prices(rows: list[dict[str, str]]) -> dict[tuple[str, str, str], int]:
    prices: dict[tuple[str, str, str], int] = {}
    for row in rows:
        option_type = clean(row.get("OptionType", "")).lower()
        code = clean(row.get("Code", ""))
        if not option_type or not code:
            continue
        prices[(option_type, clean(row.get("Trim", "")).replace("_", " "), code)] = money(row.get("Price"))
    return prices


def price_ref_component_price(
    price_ref: dict[tuple[str, str, str], int],
    option_type: str,
    code: str,
    trim: str = "",
) -> int:
    normalized_type = clean(option_type).lower()
    normalized_trim = clean(trim).replace("_", " ")
    normalized_code = clean(code)
    if (normalized_type, normalized_trim, normalized_code) in price_ref:
        return price_ref[(normalized_type, normalized_trim, normalized_code)]
    return price_ref.get((normalized_type, "", normalized_code), 0)


def r6x_price_component(row: dict[str, str], price_ref: dict[tuple[str, str], int]) -> int:
    trim = clean(row.get("Trim", ""))
    interior_id = clean(row.get("interior_id", "") or row.get("ID", ""))
    if "R6X" not in trim and "R6X" not in interior_id:
        return 0

    seat = clean(row.get("Seat", ""))
    r6x_trim = trim if "R6X" in trim else f"{trim}_R6X"
    base_trim = r6x_trim.replace("_R6X", "")
    r6x_price = price_ref.get(price_ref_key(r6x_trim, seat))
    if r6x_price is None:
        return 0
    return max(0, r6x_price - price_ref.get(price_ref_key(base_trim, seat), 0))


def generated_interior_price(row: dict[str, str], price_ref: dict[tuple[str, str], int]) -> int:
    return interior_price(row) + r6x_price_component(row, price_ref)


INTERIOR_COMPONENT_LABELS = {
    "36S": "Yellow Stitching",
    "37S": "Blue Stitching",
    "38S": "Red Stitching",
    "N26": "Sueded Microfiber",
    "N2Z": "Sueded Microfiber",
    "TU7": "Two-Tone",
    "R6X": "Custom Interior Trim and Seat Combination",
}


def interior_component_metadata(
    row: dict[str, str],
    price_ref: dict[tuple[str, str, str], int],
) -> list[dict[str, Any]]:
    trim = clean(row.get("Trim", ""))
    interior_id = clean(row.get("interior_id", "") or row.get("ID", ""))
    seat = clean(row.get("Seat", ""))
    tokens = set(interior_id.split("_"))
    components: list[dict[str, Any]] = []

    if "R6X" in trim or "R6X" in tokens:
        r6x_trim = trim if "R6X" in trim else f"{trim}_R6X"
        components.append(
            {
                "rpo": "R6X",
                "label": INTERIOR_COMPONENT_LABELS["R6X"],
                "price": price_ref_component_price(price_ref, "seat", seat, r6x_trim),
                "component_type": "r6x",
            }
        )
    else:
        seat_price = price_ref_component_price(price_ref, "seat", seat, trim)
        if seat_price:
            components.append(
                {
                    "rpo": seat,
                    "label": f"{seat} Seat Upgrade",
                    "price": seat_price,
                    "component_type": "seat",
                }
            )

    for rpo in ("36S", "37S", "38S"):
        if rpo in tokens:
            components.append(
                {
                    "rpo": rpo,
                    "label": INTERIOR_COMPONENT_LABELS[rpo],
                    "price": price_ref_component_price(price_ref, "stitching", rpo),
                    "component_type": "stitching",
                }
            )

    for rpo in ("N26", "N2Z"):
        if rpo in tokens:
            components.append(
                {
                    "rpo": rpo,
                    "label": INTERIOR_COMPONENT_LABELS[rpo],
                    "price": price_ref_component_price(price_ref, "suede", rpo),
                    "component_type": "suede",
                }
            )

    if "TU7" in tokens:
        components.append(
            {
                "rpo": "TU7",
                "label": INTERIOR_COMPONENT_LABELS["TU7"],
                "price": price_ref_component_price(price_ref, "twotone", "TU7"),
                "component_type": "two_tone",
            }
        )

    return [component for component in components if component["price"] or component["rpo"] == "R6X"]


def workbook_interior_component_metadata(
    row: dict[str, str],
    workbook_components_by_interior_id: dict[str, list[dict[str, Any]]],
    price_ref: dict[tuple[str, str, str], int],
) -> list[dict[str, Any]]:
    interior_id = clean(row.get("interior_id", "") or row.get("ID", ""))
    component_rows = workbook_components_by_interior_id.get(interior_id, [])
    if not component_rows:
        return []
    trim = clean(row.get("Trim", ""))
    components: list[dict[str, Any]] = []
    for component in component_rows:
        price_ref_type = clean(component.get("price_ref_type"))
        price_ref_code = clean(component.get("price_ref_code")) or clean(component.get("rpo"))
        price_trim_scope = clean(component.get("price_trim_scope")) or trim
        price = price_ref_component_price(price_ref, price_ref_type, price_ref_code, price_trim_scope)
        normalized = {
            "rpo": clean(component.get("rpo")),
            "label": clean(component.get("label")),
            "price": price,
            "component_type": clean(component.get("component_type")),
        }
        if normalized["price"] or normalized["rpo"] == "R6X":
            components.append(normalized)
    return components


def clean_reference_label(value: str) -> str:
    label = clean(value)
    if " - " in label:
        head, tail = label.split(" - ", 1)
        if re.search(r"\b(option|expandable|choice|card)\b", tail, re.IGNORECASE):
            label = head
    label = re.sub(r"\s*\([^)]*(?:expandable|only one option|no need)[^)]*\)\s*$", "", label, flags=re.IGNORECASE)
    return label.strip()


def read_interior_reference() -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    reference_by_id: dict[str, dict[str, Any]] = {}
    reference_rows: list[dict[str, Any]] = []
    current_levels = [""] * 6
    with INTERIOR_REFERENCE_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            for index in range(6):
                key = f"level{index}"
                value = clean_reference_label(row.get(key, ""))
                if value:
                    current_levels[index] = value
                    for deeper in range(index + 1, 6):
                        current_levels[deeper] = ""
            interior_id = clean(row.get("interior_id", ""))
            levels = [level for level in current_levels if level]
            record = {
                "row_number": row_number,
                "interior_id": interior_id,
                "levels": levels,
            }
            reference_rows.append(record)
            if interior_id:
                reference_by_id[interior_id] = record
    return reference_by_id, reference_rows


def seat_code_from_label(label: str) -> str:
    return clean(label).split(" ", 1)[0]


def grouping_fields_for_interior(
    interior: dict[str, Any],
    reference: dict[str, Any] | None,
    reference_order: int,
    fallback: bool = False,
) -> dict[str, Any]:
    seat_label = reference["levels"][1] if reference and len(reference["levels"]) > 1 else f"{interior['seat_code']} Seats"
    levels = reference["levels"] if reference else [
        interior["trim_level"],
        seat_label,
        "Other Interior Choices",
        interior["interior_name"] or interior["interior_id"],
    ]
    leaf_label = levels[-1] if levels else interior["interior_name"] or interior["interior_id"]
    color_family = levels[2] if len(levels) > 2 else leaf_label
    material_family = interior.get("material") or "Standard interior"
    if len(levels) > 3 and levels[-2] != color_family:
        material_family = levels[-2]
    parent_group = levels[-2] if len(levels) > 1 else color_family
    return {
        "interior_trim_level": levels[0] if levels else interior["trim_level"],
        "interior_seat_code": seat_code_from_label(seat_label) or interior["seat_code"],
        "interior_seat_label": seat_label,
        "interior_color_family": "Other Interior Choices" if fallback else color_family,
        "interior_material_family": material_family,
        "interior_variant_label": leaf_label,
        "interior_group_display_order": reference_order,
        "interior_material_display_order": reference_order,
        "interior_choice_display_order": reference_order,
        "interior_hierarchy_levels": json.dumps(levels, ensure_ascii=False),
        "interior_hierarchy_path": " > ".join(levels),
        "interior_parent_group_label": parent_group,
        "interior_leaf_label": leaf_label,
        "interior_reference_order": reference_order,
    }


def label_for(entity_id: str, options: dict[str, dict[str, Any]], interiors: dict[str, dict[str, Any]]) -> str:
    if entity_id in options:
        option = options[entity_id]
        rpo = option.get("rpo") or entity_id
        return f"{rpo} {option.get('label', '')}".strip()
    if entity_id in interiors:
        interior = interiors[entity_id]
        return f"{interior.get('interior_id')} {interior.get('interior_name')}".strip()
    return entity_id


def truncate_reason(text: str, limit: int = 180) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "..."


def standard_equipment_key(choice: dict[str, Any]) -> tuple[str, str]:
    rpo = clean(choice.get("rpo", ""))
    return choice["variant_id"], rpo or choice["option_id"]


def standard_equipment_preference(choice: dict[str, Any], index: int) -> tuple[int, int]:
    option_id = clean(choice.get("option_id", ""))
    section_id = clean(choice.get("section_id", ""))
    is_canonical = option_id.endswith("_001")
    if is_canonical and section_id != "sec_stan_002":
        rank = 0
    elif section_id != "sec_stan_002":
        rank = 1
    else:
        rank = 2
    return rank, index


def standard_equipment_row(choice: dict[str, Any]) -> dict[str, Any]:
    return {
        "equipment_id": f"std_{choice['variant_id']}__{choice['option_id']}",
        "variant_id": choice["variant_id"],
        "body_style": choice["body_style"],
        "trim_level": choice["trim_level"],
        "option_id": choice["option_id"],
        "rpo": choice["rpo"],
        "label": choice["label"],
        "description": choice["description"],
        "section_id": choice["section_id"],
        "section_name": choice["section_name"],
        "standard_equipment_group_type": choice.get("standard_equipment_group_type", ""),
        "display_order": choice["display_order"],
        "source_detail_raw": choice["source_detail_raw"],
    }


def build_standard_equipment(choices: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected: dict[tuple[str, str], tuple[int, dict[str, Any]]] = {}
    key_order: list[tuple[str, str]] = []
    for index, choice in enumerate(choices):
        if choice["status"] != "standard":
            continue
        key = standard_equipment_key(choice)
        if key not in selected:
            selected[key] = (index, choice)
            key_order.append(key)
            continue
        existing_index, existing_choice = selected[key]
        if standard_equipment_preference(choice, index) < standard_equipment_preference(existing_choice, existing_index):
            selected[key] = (index, choice)
    return [standard_equipment_row(selected[key][1]) for key in key_order]


def main() -> None:
    global MODEL_CONFIG

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH)
    MODEL_CONFIG = load_model_config_overrides(wb, STINGRAY_MODEL)

    variants_raw = rows_from_sheet(wb, "variant_master")
    sections = {row["section_id"]: row for row in rows_from_sheet(wb, "section_master")}
    options_raw = rows_from_sheet(wb, MODEL_CONFIG.source_option_sheet)
    statuses_raw = rows_from_sheet(wb, MODEL_CONFIG.status_sheet)
    context_copy_rows = context_choice_copy_rows(wb, MODEL_CONFIG.model_key)
    rules_raw = rows_from_sheet(wb, MODEL_CONFIG.rule_mapping_sheet)
    price_rules_raw = rows_from_sheet(wb, MODEL_CONFIG.price_rules_sheet)
    d30_r6x_price_rules_raw = [row for row in price_rules_raw if row.get("price_rule_id") == "pr_d30_r6x_001"]
    price_rules_raw = [row for row in price_rules_raw if row.get("price_rule_id") != "pr_d30_r6x_001"]
    lt_interiors_raw = rows_from_sheet(wb, "lt_interiors")
    price_ref_rows = rows_from_sheet(wb, "PriceRef")
    price_ref = price_ref_prices(price_ref_rows)
    interior_component_price_ref = price_ref_component_prices(price_ref_rows)
    color_overrides_raw = rows_from_sheet(wb, MODEL_CONFIG.color_overrides_sheet)
    rule_groups = load_rule_groups(wb)
    exclusive_groups = load_exclusive_groups(wb)
    default_selection_rules = load_default_selection_rules(wb, MODEL_CONFIG.model_key)
    runtime_rule_exceptions = load_runtime_rule_exceptions(wb, MODEL_CONFIG.model_key)
    order_summary_metadata = load_order_summary_metadata(wb, MODEL_CONFIG.model_key)
    runtime_steps = load_runtime_steps(wb, MODEL_CONFIG.model_key, MODEL_CONFIG.step_order, MODEL_CONFIG.step_labels)
    context_sections = [
        {**row, "selection_mode_label": selection_mode_label(row.get("selection_mode", ""))}
        for row in load_context_sections(wb, MODEL_CONFIG.model_key, MODEL_CONFIG.context_sections)
    ]
    section_presentation_rows = load_section_presentation(wb, MODEL_CONFIG.model_key)
    section_presentation = {row["section_id"]: row for row in section_presentation_rows}
    standard_section_ids = {
        section_id
        for section_id, presentation in section_presentation.items()
        if presentation_bool(presentation, "standard_equipment_bucket", default=False)
    } or set(STANDARD_SECTIONS)
    variant_option_override_rows = load_variant_option_overrides(wb, MODEL_CONFIG.model_key)
    variant_option_overrides = {
        (row["option_id"], row["variant_id"]): row
        for row in variant_option_override_rows
    }
    workbook_components_by_interior_id = load_interior_components(wb, MODEL_CONFIG.model_key)
    option_asset_map = load_asset_map(wb, MODEL_CONFIG.model_key, "option")
    grouped_requires = grouped_requirement_pairs(rule_groups)
    interior_reference_by_id, interior_reference_rows = read_interior_reference()

    display_behavior_by_option_id = {
        row["option_id"]: clean(row.get("display_behavior", "")).lower()
        for row in options_raw
        if clean(row.get("display_behavior", ""))
    }
    hidden_option_ids = {
        option_id
        for option_id, display_behavior in display_behavior_by_option_id.items()
        if display_behavior == "hidden"
    }
    for option in options_raw:
        option_display_behavior = display_behavior_by_option_id.get(option["option_id"], "")
        section_display_behavior = clean(
            section_presentation.get(option.get("section_id", ""), {}).get("display_behavior", "")
        ).lower()
        display_behavior = option_display_behavior or section_display_behavior
        option["_display_behavior"] = display_behavior
        if display_behavior == "hidden":
            option["active"] = "False"
            hidden_option_ids.add(option["option_id"])

    price_rules_raw.extend(d30_r6x_price_rules_raw)

    active_variants = [
        {
            "variant_id": row["variant_id"],
            "model_year": intish(row.get("model_year")),
            "trim_level": row["trim_level"].upper(),
            "body_style": row["body_style"].lower(),
            "display_name": row["display_name"],
            "base_price": money(row.get("base_price")),
            "display_order": intish(row.get("display_order")),
        }
        for row in variants_raw
        if row.get("active") == "True" and row.get("variant_id", "") in MODEL_CONFIG.variant_ids
    ]
    variant_by_id = {row["variant_id"]: row for row in active_variants}

    section_rows: list[dict[str, Any]] = [dict(row) for row in context_sections]
    for section_id, section in sections.items():
        presentation = section_presentation.get(section_id, {})
        section_name = clean(presentation.get("display_label")) or section.get("section_name", "")
        presentation_step_key = clean(presentation.get("step_key"))
        step_key = step_for_section(
            section_id,
            section_name,
            presentation_step_key or section.get("step_key", ""),
            standard_sections=standard_section_ids,
        )
        selection_mode = section.get("selection_mode", "")
        section_display_order = (
            intish(presentation.get("section_display_order"))
            if clean(presentation.get("section_display_order"))
            else intish(section.get("display_order"))
        )
        section_rows.append(
            {
                "section_id": section_id,
                "section_name": section_name,
                "selection_mode": selection_mode,
                "selection_mode_label": selection_mode_label(selection_mode),
                "choice_mode": normalize_mode(selection_mode),
                "is_required": section.get("is_required", ""),
                "standard_behavior": section.get("standard_behavior", ""),
                "section_display_order": section_display_order,
                "step_key": step_key,
                "step_label": STEP_LABELS.get(step_key, step_key.replace("_", " ").title()),
            }
        )

    step_rows: list[dict[str, Any]] = [dict(row) for row in runtime_steps]
    section_ids_by_step: dict[str, list[str]] = defaultdict(list)
    for row in section_rows:
        section_ids_by_step[row["step_key"]].append(row["section_id"])
    for row in step_rows:
        row["section_ids"] = "|".join(sorted(section_ids_by_step.get(row["step_key"], [])))

    body_context_choices = []
    body_styles = sorted(
        {row["body_style"] for row in active_variants},
        key=lambda body_style: BODY_STYLE_DISPLAY_ORDER.get(body_style, 99),
    )
    for body_style in body_styles:
        body_variants = [row for row in active_variants if row["body_style"] == body_style]
        body_context_choices.append(
            {
                "context_choice_id": f"body_style__{body_style}",
                "context_type": "body_style",
                "value": body_style,
                "label": body_style.title(),
                "description": f"{len(body_variants)} trims available",
                "info_tooltip": context_choice_info_tooltip(
                    context_copy_rows,
                    model_key=MODEL_CONFIG.model_key,
                    context_type="body_style",
                    value=body_style,
                    body_style=body_style,
                ),
                "section_id": "sec_context_body_style",
                "step_key": "body_style",
                "body_style": body_style,
                "trim_level": "",
                "variant_id": "",
                "base_price": "",
                "display_order": BODY_STYLE_DISPLAY_ORDER.get(body_style, 99),
            }
        )
    trim_context_choices = []
    for variant in active_variants:
        trim_context_choices.append(
            {
                "context_choice_id": f"trim_level__{variant['body_style']}__{variant['trim_level'].lower()}",
                "context_type": "trim_level",
                "value": variant["trim_level"],
                "label": variant["trim_level"],
                "description": variant["display_name"],
                "info_tooltip": context_choice_info_tooltip(
                    context_copy_rows,
                    model_key=MODEL_CONFIG.model_key,
                    context_type="trim_level",
                    value=variant["trim_level"],
                    body_style=variant["body_style"],
                ),
                "section_id": "sec_context_trim_level",
                "step_key": "trim_level",
                "body_style": variant["body_style"],
                "trim_level": variant["trim_level"],
                "variant_id": variant["variant_id"],
                "base_price": variant["base_price"],
                "display_order": variant["display_order"],
            }
        )
    context_choices = body_context_choices + trim_context_choices

    options_by_id: dict[str, dict[str, Any]] = {}
    for option in options_raw:
        if option["option_id"] in options_by_id and option.get("active") != "True":
            continue
        section = sections.get(option.get("section_id", ""), {})
        presentation = section_presentation.get(option.get("section_id", ""), {})
        section_name = clean(presentation.get("display_label")) or section.get("section_name", "")
        presentation_step_key = clean(presentation.get("step_key"))
        step_key = step_for_section(
            option.get("section_id", ""),
            section_name,
            presentation_step_key or section.get("step_key", ""),
            standard_sections=standard_section_ids,
        )
        mode = section.get("selection_mode", "")
        options_by_id[option["option_id"]] = {
            "option_id": option["option_id"],
            "rpo": option.get("rpo", ""),
            "label": option.get("option_name", ""),
            "description": option.get("description", ""),
            "source_detail_raw": option.get("detail_raw", ""),
            "section_id": option.get("section_id", ""),
            "section_name": section_name,
            "standard_equipment_group_type": clean(presentation.get("standard_equipment_group_type")),
            "step_key": step_key,
            "selection_mode": mode,
            "selection_mode_label": selection_mode_label(mode),
            "choice_mode": normalize_mode(mode),
            "selectable": option.get("selectable", ""),
            "active": option.get("active", ""),
            "display_behavior": option.get("_display_behavior", ""),
            "base_price": money(option.get("price")),
            "display_order": intish(option.get("display_order")),
        }

    status_by_option_variant: dict[tuple[str, str], str] = {}
    for row in statuses_raw:
        key = (row["option_id"], row["variant_id"])
        status_by_option_variant[key] = best_status(status_by_option_variant.get(key, ""), row["status"])
    choices: list[dict[str, Any]] = []
    for option_id, option in options_by_id.items():
        if option["active"] != "True":
            continue
        for variant in active_variants:
            status = status_by_option_variant.get((option_id, variant["variant_id"]), "unavailable")
            selectable = option["selectable"]
            active = option["active"]
            display_behavior = option.get("display_behavior", "")
            override = variant_option_overrides.get((option_id, variant["variant_id"]), {})
            if clean(override.get("status")):
                status = clean(override["status"]).lower()
            if clean(override.get("selectable")):
                selectable = clean(override["selectable"])
            if clean(override.get("active")):
                active = clean(override["active"])
            if clean(override.get("display_behavior")):
                display_behavior = clean(override["display_behavior"]).lower()
            if display_behavior == "auto_only":
                status = "unavailable"
                selectable = "False"
                active = "False"
            elif display_behavior == "display_only":
                status = "available"
                selectable = "False"
                active = "True"
            choice = {
                "choice_id": f"{variant['variant_id']}__{option_id}",
                "option_id": option_id,
                "rpo": option["rpo"],
                "label": option["label"],
                "description": option["description"],
                "section_id": option["section_id"],
                "section_name": option["section_name"],
                "standard_equipment_group_type": option.get("standard_equipment_group_type", ""),
                "step_key": option["step_key"],
                "variant_id": variant["variant_id"],
                "body_style": variant["body_style"],
                "trim_level": variant["trim_level"],
                "status": status,
                "status_label": status_to_label(status),
                "selectable": selectable,
                "active": active,
                "choice_mode": option["choice_mode"],
                "selection_mode": option["selection_mode"],
                "selection_mode_label": option["selection_mode_label"],
                "base_price": option["base_price"],
                "display_order": option["display_order"],
                "source_detail_raw": option["source_detail_raw"],
            }
            if display_behavior:
                choice["display_behavior"] = display_behavior
            if asset := option_asset_map.get(option_id):
                choice.update(asset)
            choices.append(choice)

    validation_rows: list[dict[str, Any]] = []
    interiors: list[dict[str, Any]] = []
    for row in lt_interiors_raw:
        trim = row.get("Trim", "")
        interior_id = row.get("interior_id", "")
        active_for_stingray = workbook_bool(
            row,
            "active_for_stingray",
            False,
        )
        requires_r6x = workbook_bool(
            row,
            "requires_r6x",
            False,
        )
        included_option_id = clean(row.get("included_option_id", ""))
        if active_for_stingray and requires_r6x and not included_option_id:
            validation_rows.append(
                {
                    "check_id": f"missing_r6x_included_option_{interior_id}",
                    "severity": "error",
                    "entity_type": "interior",
                    "entity_id": interior_id,
                    "message": "R6X interior requires included_option_id in lt_interiors.",
                }
            )
        components = workbook_interior_component_metadata(row, workbook_components_by_interior_id, interior_component_price_ref)
        if not active_for_stingray and not components:
            components = interior_component_metadata(row, interior_component_price_ref)
        if active_for_stingray and not components and interior_component_metadata(row, interior_component_price_ref):
            validation_rows.append(
                {
                    "check_id": f"missing_workbook_components_{interior_id}",
                    "severity": "error",
                    "entity_type": "interior",
                    "entity_id": interior_id,
                    "message": "Active Stingray interior has component-bearing legacy output but no active interior_components workbook rows.",
                }
            )
        interiors.append(
            {
                "interior_id": interior_id,
                "source_sheet": "lt_interiors",
                "active_for_stingray": active_for_stingray,
                "trim_level": trim.replace("_R6X", ""),
                "requires_r6x": "True" if requires_r6x else "False",
                "_included_option_id": included_option_id,
                "seat_code": row.get("Seat", ""),
                "interior_code": row.get("Interior Code", ""),
                "interior_name": row.get("Interior Name", ""),
                "material": row.get("Material", ""),
                "price": generated_interior_price(row, price_ref),
                "suede": row.get("Suede", ""),
                "stitch": row.get("Stitch", ""),
                "two_tone": row.get("Two Tone", ""),
                "section_id": row.get("section_id", ""),
                "color_overrides_raw": row.get("Color Overrides", ""),
                "source_note": row.get("Detail from Disclosure", ""),
                "interior_components": components,
                "interior_components_json": json.dumps(components, separators=(",", ":")),
            }
        )
    reference_order_by_id = {
        row["interior_id"]: index
        for index, row in enumerate((row for row in interior_reference_rows if row["interior_id"]), start=1)
    }
    active_interior_ids = {
        row["interior_id"]
        for row in interiors
        if row["interior_id"] and row["active_for_stingray"]
    }
    all_interior_ids = {row["interior_id"] for row in interiors if row["interior_id"]}
    for interior_id, reference in interior_reference_by_id.items():
        if interior_id not in all_interior_ids:
            validation_rows.append(
                {
                    "check_id": f"missing_reference_interior_{interior_id}",
                    "severity": "error",
                    "entity_type": "interior",
                    "entity_id": interior_id,
                    "message": f"Interior reference row {reference['row_number']} does not resolve to generated interior data.",
                }
            )
        elif interior_id not in active_interior_ids:
            validation_rows.append(
                {
                    "check_id": f"inactive_reference_interior_{interior_id}",
                    "severity": "error",
                    "entity_type": "interior",
                    "entity_id": interior_id,
                    "message": f"Interior reference row {reference['row_number']} resolves to an inactive Stingray interior.",
                }
            )

    fallback_order = len(reference_order_by_id) + 1
    for row in interiors:
        if not row["interior_id"]:
            continue
        reference = interior_reference_by_id.get(row["interior_id"])
        if row["active_for_stingray"] and reference:
            row.update(grouping_fields_for_interior(row, reference, reference_order_by_id[row["interior_id"]]))
        elif row["active_for_stingray"]:
            row.update(grouping_fields_for_interior(row, None, fallback_order, fallback=True))
            fallback_order += 1
            validation_rows.append(
                {
                    "check_id": f"unmapped_active_interior_{row['interior_id']}",
                    "severity": "warning",
                    "entity_type": "interior",
                    "entity_id": row["interior_id"],
                    "message": "Active Stingray interior is not represented in the CSV hierarchy and was placed in Other Interior Choices.",
                }
            )
        else:
            row.update(grouping_fields_for_interior(row, reference, reference_order_by_id.get(row["interior_id"], fallback_order)))

    interiors_by_id = {row["interior_id"]: row for row in interiors if row["interior_id"]}
    interior_include_ids = [
        (row["interior_id"], row.get("_included_option_id", ""))
        for row in interiors
        if row["interior_id"]
        and row["active_for_stingray"]
        and row.get("_included_option_id", "")
        and options_by_id.get(row.get("_included_option_id", ""), {}).get("rpo")
        not in {component.get("rpo") for component in row.get("interior_components", [])}
    ]

    raw_rules: list[dict[str, Any]] = []
    manual_rules = []
    for interior_id, included_option_id in interior_include_ids:
        manual_rules.append(
            {
                "rule_id": f"rule_{interior_id.lower()}_includes_{included_option_id}",
                "source_id": interior_id,
                "rule_type": "includes",
                "target_id": included_option_id,
                "target_type": "option",
                "source_type": "interior",
                "source_section": interiors_by_id[interior_id].get("section_id", ""),
                "target_section": "sec_colo_001",
                "source_selection_mode": "single_select_req",
                "target_selection_mode": "multi_select_opt",
                "original_detail_raw": "R6X is included with this custom interior trim and seat combination."
                if included_option_id == "opt_r6x_001"
                else f"{included_option_id} is included with this custom interior trim and seat combination.",
                "review_flag": "False",
            }
        )
    for rule in rules_raw + manual_rules:
        rule_type = rule.get("rule_type", "").lower()
        source_id = rule.get("source_id", "")
        target_id = rule.get("target_id", "")
        if not runtime_authored_rule(rule):
            continue
        if rule_type == "requires" and (source_id, target_id) in grouped_requires:
            continue
        if source_id in hidden_option_ids or target_id in hidden_option_ids:
            continue
        source_section = rule.get("source_section", "")
        target_section = rule.get("target_section", "")
        source_mode = sections.get(source_section, {}).get("selection_mode") or rule.get("source_selection_mode", "")
        target_mode = sections.get(target_section, {}).get("selection_mode") or rule.get("target_selection_mode", "")
        body_style_scope = rule.get("body_style_scope", "")
        replaces_t0a = rule.get("runtime_action", "") == "replace"
        redundant = (
            rule_type == "excludes"
            and source_section
            and source_section == target_section
            and source_mode.startswith("single")
            and target_mode.startswith("single")
            and not replaces_t0a
        )
        action = "replace" if replaces_t0a else "omit_redundant_same_section_exclude" if redundant else "active"
        if redundant:
            validation_rows.append(
                {
                    "check_id": f"redundant_{rule.get('rule_id', '')}",
                    "severity": "info",
                    "entity_type": "rule",
                    "entity_id": rule.get("rule_id", ""),
                    "message": "Same-section single-select excludes are redundant because the section choice mode already prevents multiple selections.",
                }
            )
        disabled_reason = ""
        auto_add = "False"
        source_label = label_for(source_id, options_by_id, interiors_by_id)
        target_label = label_for(target_id, options_by_id, interiors_by_id)
        if rule.get("disabled_reason", ""):
            disabled_reason = rule.get("disabled_reason", "")
        elif replaces_t0a:
            disabled_reason = f"{source_label} removes this default."
        elif rule_type == "excludes":
            disabled_reason = f"Blocked by {source_label}."
        elif rule_type == "requires":
            disabled_reason = f"Requires {target_label}."
        elif rule_type == "includes":
            disabled_reason = f"Included with {source_label}."
            auto_add = "True"
        raw_rules.append(
            {
                "rule_id": rule.get("rule_id", ""),
                "source_id": source_id,
                "rule_type": rule_type,
                "target_id": target_id,
                "target_type": rule.get("target_type", ""),
                "source_type": rule.get("source_type", ""),
                "source_section": source_section,
                "target_section": target_section,
                "source_selection_mode": source_mode,
                "target_selection_mode": target_mode,
                "body_style_scope": body_style_scope,
                "disabled_reason": disabled_reason,
                "auto_add": auto_add,
                "active": "False" if redundant else "True",
                "runtime_action": action,
                "source_note": truncate_reason(rule.get("original_detail_raw", ""), 500),
                "review_flag": rule.get("review_flag", ""),
            }
        )

    price_rules = [
        {
            "price_rule_id": row.get("price_rule_id", ""),
            "condition_option_id": row.get("condition_option_id", ""),
            "target_option_id": row.get("target_option_id", ""),
            "price_rule_type": row.get("price_rule_type", "").lower(),
            "price_value": money(row.get("price_value")),
            "body_style_scope": row.get("body_style_scope", ""),
            "trim_level_scope": row.get("trim_level_scope", ""),
            "variant_scope": row.get("variant_scope", ""),
            "review_flag": row.get("review_flag", ""),
            "notes": row.get("notes", ""),
        }
        for row in price_rules_raw
        if row.get("condition_option_id", "") not in hidden_option_ids and row.get("target_option_id", "") not in hidden_option_ids
    ]

    color_overrides = [
        {
            "override_id": f"co_{idx:03d}",
            "interior_id": row.get("interior_id", ""),
            "option_id": row.get("option_id", ""),
            "rule_type": row.get("rule_type", "").lower(),
            "adds_rpo": row.get("adds_rpo", ""),
            "notes": "Exterior/interior pairing requires the listed override RPO.",
        }
        for idx, row in enumerate(color_overrides_raw, start=1)
    ]
    for row in interiors:
        row.pop("_included_option_id", None)

    # Validation floor
    if len(active_variants) != 6:
        validation_rows.append(
            {
                "check_id": "active_variant_count",
                "severity": "error",
                "entity_type": "variant",
                "entity_id": "",
                "message": f"Expected 6 active Stingray variants; found {len(active_variants)}.",
            }
        )
    expected_status_rows = len(active_variants) * len(options_by_id)
    canonical_status_rows = len({(row["option_id"], row["variant_id"]) for row in statuses_raw})
    if canonical_status_rows != expected_status_rows:
        validation_rows.append(
            {
                "check_id": "availability_row_count",
                "severity": "error",
                "entity_type": "availability",
                "entity_id": "",
                "message": f"Expected {expected_status_rows} canonical {MODEL_CONFIG.status_sheet} rows; found {canonical_status_rows}.",
            }
        )
    valid_ids = set(options_by_id) | set(interiors_by_id)
    for rule in raw_rules:
        for key in ["source_id", "target_id"]:
            if rule[key] not in valid_ids:
                validation_rows.append(
                    {
                        "check_id": f"missing_{key}_{rule['rule_id']}",
                        "severity": "error",
                        "entity_type": "rule",
                        "entity_id": rule["rule_id"],
                        "message": f"{key} {rule[key]} does not resolve to an option or interior.",
                    }
                )
    for rule in price_rules:
        for key in ["condition_option_id", "target_option_id"]:
            if rule[key] not in valid_ids:
                validation_rows.append(
                    {
                        "check_id": f"missing_{key}_{rule['price_rule_id']}",
                        "severity": "error",
                        "entity_type": "price_rule",
                        "entity_id": rule["price_rule_id"],
                        "message": f"{key} {rule[key]} does not resolve to an option or interior.",
                    }
                )

    status_counts = Counter(row["status"] for row in choices)
    standard_equipment = build_standard_equipment(choices)
    validation_rows.extend(
        [
            {
                "check_id": "active_variants",
                "severity": "pass",
                "entity_type": "variant",
                "entity_id": "",
                "message": f"{len(active_variants)} active Stingray variants exported.",
            },
            {
                "check_id": "availability_rows",
                "severity": "pass",
                "entity_type": "availability",
                "entity_id": "",
                "message": f"{len(choices)} choice rows exported ({dict(status_counts)}).",
            },
            {
                "check_id": "rules",
                "severity": "pass",
                "entity_type": "rule",
                "entity_id": "",
                "message": f"{sum(1 for row in raw_rules if row['active'] == 'True')} active compatibility rules exported from {len(raw_rules)} source rules.",
            },
        ]
    )

    write_sheet(
        wb,
        "form_steps",
        ["step_key", "step_label", "runtime_order", "source", "section_ids"],
        step_rows,
    )
    write_sheet(
        wb,
        "form_context_choices",
        [
            "context_choice_id",
            "context_type",
            "value",
            "label",
            "description",
            "info_tooltip",
            "section_id",
            "step_key",
            "body_style",
            "trim_level",
            "variant_id",
            "base_price",
            "display_order",
        ],
        context_choices,
    )
    write_sheet(
        wb,
        "form_choices",
        [
            "choice_id",
            "option_id",
            "rpo",
            "label",
            "description",
            "section_id",
            "section_name",
            "step_key",
            "variant_id",
            "body_style",
            "trim_level",
            "status",
            "status_label",
            "selectable",
            "active",
            "choice_mode",
            "selection_mode",
            "selection_mode_label",
            "base_price",
            "display_behavior",
            "display_order",
            "source_detail_raw",
            *ASSET_IMAGE_FIELDS,
        ],
        choices,
    )
    write_sheet(
        wb,
        "form_standard_equipment",
        [
            "equipment_id",
            "variant_id",
            "body_style",
            "trim_level",
            "option_id",
            "rpo",
            "label",
            "description",
            "section_id",
            "section_name",
            "standard_equipment_group_type",
            "display_order",
            "source_detail_raw",
        ],
        standard_equipment,
    )
    write_sheet(
        wb,
        "form_rule_groups",
        [
            "group_id",
            "group_type",
            "source_id",
            "target_ids",
            "body_style_scope",
            "trim_level_scope",
            "variant_scope",
            "disabled_reason",
            "active",
            "notes",
        ],
        [{**row, "target_ids": "|".join(row["target_ids"])} for row in rule_groups],
    )
    write_sheet(
        wb,
        "form_exclusive_groups",
        ["group_id", "option_ids", "selection_mode", "active", "notes"],
        [{**row, "option_ids": "|".join(row["option_ids"])} for row in exclusive_groups],
    )
    write_sheet(
        wb,
        "form_rules",
        [
            "rule_id",
            "source_id",
            "rule_type",
            "target_id",
            "target_type",
            "source_type",
            "source_section",
            "target_section",
            "source_selection_mode",
            "target_selection_mode",
            "body_style_scope",
            "disabled_reason",
            "auto_add",
            "active",
            "runtime_action",
            "source_note",
            "review_flag",
        ],
        raw_rules,
    )
    write_sheet(
        wb,
        "form_price_rules",
        [
            "price_rule_id",
            "condition_option_id",
            "target_option_id",
            "price_rule_type",
            "price_value",
            "body_style_scope",
            "trim_level_scope",
            "variant_scope",
            "review_flag",
            "notes",
        ],
        price_rules,
    )
    write_sheet(
        wb,
        "form_interiors",
        [
            "interior_id",
            "source_sheet",
            "active_for_stingray",
            "trim_level",
            "requires_r6x",
            "seat_code",
            "interior_code",
            "interior_name",
            "material",
            "price",
            "suede",
            "stitch",
            "two_tone",
            "section_id",
            "color_overrides_raw",
            "source_note",
            "interior_components_json",
            "interior_trim_level",
            "interior_seat_code",
            "interior_seat_label",
            "interior_color_family",
            "interior_material_family",
            "interior_variant_label",
            "interior_group_display_order",
            "interior_material_display_order",
            "interior_choice_display_order",
            "interior_hierarchy_levels",
            "interior_hierarchy_path",
            "interior_parent_group_label",
            "interior_leaf_label",
            "interior_reference_order",
        ],
        interiors,
    )
    write_sheet(
        wb,
        "form_color_overrides",
        ["override_id", "interior_id", "option_id", "rule_type", "adds_rpo", "notes"],
        color_overrides,
    )
    write_sheet(
        wb,
        "form_validation",
        ["check_id", "severity", "entity_type", "entity_id", "message"],
        validation_rows,
    )
    workbook_backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)

    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    data = {
        "dataset": {
            "name": MODEL_CONFIG.dataset_name,
            "source_workbook": WORKBOOK_PATH.name,
            "generated_at": generated_at,
        },
        "variants": active_variants,
        "steps": step_rows,
        "sections": section_rows,
        "contextChoices": context_choices,
        "choices": choices,
        "standardEquipment": standard_equipment,
        "ruleGroups": rule_groups,
        "exclusiveGroups": exclusive_groups,
        "rules": [row for row in raw_rules if row["active"] == "True"],
        "priceRules": price_rules,
        "defaultSelectionRules": default_selection_rules,
        "runtimeRuleExceptions": runtime_rule_exceptions,
        "orderSummary": order_summary_metadata,
        "interiors": [row for row in interiors if row["active_for_stingray"]],
        "colorOverrides": color_overrides,
        "validation": validation_rows,
    }

    OUTPUT_DIR.mkdir(exist_ok=True)
    APP_DIR.mkdir(exist_ok=True)
    json_path = OUTPUT_DIR / "stingray-form-data.json"
    write_json_output(json_path, data)
    refresh_grand_sport_registry_source()
    app_registry, legacy_aliases = build_app_data_registry(data)
    write_app_data_registry(
        APP_DIR / "data.js",
        app_registry,
        legacy_aliases=legacy_aliases,
    )
    csv_path = OUTPUT_DIR / "stingray-form-data.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            lineterminator="\n",
            fieldnames=[
                "choice_id",
                "option_id",
                "rpo",
                "label",
                "section_id",
                "step_key",
                "variant_id",
                "body_style",
                "trim_level",
                "status",
                "selectable",
                "base_price",
                *ASSET_IMAGE_FIELDS,
            ],
        )
        writer.writeheader()
        for row in choices:
            writer.writerow({key: row.get(key, "") for key in writer.fieldnames})

    print(json.dumps({
        "workbook": str(WORKBOOK_PATH),
        "workbook_backup": str(workbook_backup_path),
        "json": str(json_path),
        "csv": str(csv_path),
        "choices": len(choices),
        "context_choices": len(context_choices),
        "standard_equipment": len(standard_equipment),
        "rules": len(data["rules"]),
        "price_rules": len(price_rules),
        "interiors": len(data["interiors"]),
        "validation_errors": validation_error_count(validation_rows),
    }, indent=2))


if __name__ == "__main__":
    main()
