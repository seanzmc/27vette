#!/usr/bin/env python3
"""Preview or write future Z-family LZ interior workbook source rows."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import FUTURE_MODEL_SPECS  # noqa: E402
from corvette_form_generator.future_model_lz_interiors import (  # noqa: E402
    apply_lz_interiors_to_workbook,
    build_lz_interiors_preview,
    strip_lz_preview_details,
    verify_saved_lz_interiors_workbook,
)
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import excel_lock_path, save_workbook_safely  # noqa: E402

FUTURE_MODEL_KEYS = tuple(FUTURE_MODEL_SPECS)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model-key", choices=("all", *FUTURE_MODEL_KEYS), default="all")
    parser.add_argument("--dry-run", action="store_true", help="Report projected rows without saving the workbook. This is the default unless --write is passed.")
    parser.add_argument("--write", action="store_true", help="Persist projected LZ section, scope, and component rows into the workbook source sheets.")
    parser.add_argument("--include-details", action="store_true", help="Include projected row-level scope/component details in JSON output.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=WORKBOOK_PATH,
        help="Workbook to inspect or write. Default: stingray_master.xlsx in the repository root.",
    )
    args = parser.parse_args(argv)

    workbook_path = args.workbook
    selected = list(FUTURE_MODEL_KEYS) if args.model_key == "all" else [args.model_key]

    if not args.write:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            preview = build_lz_interiors_preview(wb, selected, include_details=args.include_details)
        finally:
            wb.close()
        report = preview if args.include_details else strip_lz_preview_details(preview)
        print(json.dumps(report, indent=2))
        return 0

    if excel_lock_path(workbook_path).exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {excel_lock_path(workbook_path)}")

    loaded_mtime_ns = workbook_path.stat().st_mtime_ns
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    backup_path = None
    try:
        preview = build_lz_interiors_preview(wb, selected, include_details=True)
        apply_report = apply_lz_interiors_to_workbook(wb, preview)
        backup_path = save_workbook_safely(wb, workbook_path, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_lz_interiors_workbook(workbook_path, preview)
    preview_report = preview if args.include_details else strip_lz_preview_details(preview)
    print(
        json.dumps(
            {
                **preview_report,
                "status": "written",
                "workbook": str(workbook_path),
                "backup": str(backup_path),
                "would_write_workbook": True,
                "apply_report": apply_report,
                "verification": verification,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
