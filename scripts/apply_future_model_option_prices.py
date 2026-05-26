#!/usr/bin/env python3
"""Fill first-pass Z06/ZR1/ZR1X option prices from price_sched_raw.

This intentionally handles only direct option-level prices and a small set of
safe model/application disambiguations. Conditional package, wheel, body-style,
and trim-specific prices remain blank for later price-rule work.
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SPECS,
    OPTION_SOURCE_HEADERS,
    build_price_schedule_rows,
)
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely  # noqa: E402

FUTURE_MODEL_KEYS = tuple(FUTURE_MODEL_SPECS)
DISPLAY_ONLY_SELECTION_MODE = "display_only"


@dataclass(frozen=True)
class PriceResolution:
    price: int | float | str
    status: str
    reason: str
    raw_price_row: int | None = None
    match_count: int = 0


def rows_from_sheet(wb, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header: clean(value) for header, value in zip(headers, values) if header}
        if any(row.values()):
            rows.append(row)
    return rows


def _selected_future_model_keys(model_keys: Iterable[str]) -> list[str]:
    requested = [clean(model_key) for model_key in model_keys if clean(model_key)]
    if not requested or requested == ["all"] or "all" in requested:
        return list(FUTURE_MODEL_KEYS)
    unknown = [model_key for model_key in requested if model_key not in FUTURE_MODEL_SPECS]
    if unknown:
        raise ValueError(f"Unknown future model key(s): {', '.join(unknown)}")
    return requested


def _price_value(value: Any) -> int | float | str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = clean(value)
    if not text:
        return ""
    numeric = float(text.replace("$", "").replace(",", ""))
    return int(numeric) if numeric.is_integer() else numeric


def _section_selection_modes(wb) -> dict[str, str]:
    modes: dict[str, str] = {}
    for row in rows_from_sheet(wb, "section_master"):
        section_id = clean(row.get("section_id"))
        if section_id:
            modes[section_id] = clean(row.get("selection_mode"))
    return modes


def _price_rows_by_rpo(wb) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in build_price_schedule_rows(wb).get("option_price_rows", []):
        rpo = clean(row.get("price_rpo"))
        price = _price_value(row.get("price_list"))
        if rpo and price != "":
            grouped[rpo].append(row)
    return dict(grouped)


def _application(row: dict[str, Any]) -> str:
    return clean(row.get("price_application")).casefold()


def _single_match(matches: list[dict[str, Any]], predicate) -> dict[str, Any] | None:
    filtered = [row for row in matches if predicate(row)]
    return filtered[0] if len(filtered) == 1 else None


def _resolution_from_row(row: dict[str, Any], status: str, reason: str, match_count: int) -> PriceResolution:
    return PriceResolution(
        price=_price_value(row.get("price_list")),
        status=status,
        reason=reason,
        raw_price_row=int(row["raw_price_row"]) if row.get("raw_price_row") is not None else None,
        match_count=match_count,
    )


def _resolve_multi_match(model_key: str, rpo: str, matches: list[dict[str, Any]]) -> PriceResolution:
    match_count = len(matches)
    selected: dict[str, Any] | None = None
    reason = ""

    if rpo == "ZTK" and model_key == "zr1":
        selected = _single_match(matches, lambda row: "zr1 only" in _application(row) and "zr1x" not in _application(row))
        reason = "resolved by ZR1-only price application"
    elif rpo == "ZTK" and model_key == "zr1x":
        selected = _single_match(matches, lambda row: "zr1x only" in _application(row))
        reason = "resolved by ZR1X-only price application"
    elif rpo in {"PCQ", "VWE"} and model_key in FUTURE_MODEL_KEYS:
        selected = _single_match(matches, lambda row: "stingray" not in _application(row))
        reason = "resolved by non-Stingray/general price application"
    elif rpo == "5V5" and model_key == "z06":
        selected = _single_match(matches, lambda row: "z06" in _application(row))
        reason = "resolved by Z06 price application"

    if selected:
        return _resolution_from_row(selected, "resolved_multi_match", reason, match_count)
    return PriceResolution("", "ambiguous_price_deferred", "multiple price schedule rows require conditional price-rule review", match_count=match_count)


def resolve_option_price(
    *,
    model_key: str,
    option_row: dict[str, Any],
    section_modes: dict[str, str],
    price_rows_by_rpo: dict[str, list[dict[str, Any]]],
) -> PriceResolution:
    section_id = clean(option_row.get("section_id"))
    if section_modes.get(section_id) == DISPLAY_ONLY_SELECTION_MODE:
        return PriceResolution("", "display_only_blank", "display-only/standard section prices stay blank")

    rpo = clean(option_row.get("rpo"))
    if not rpo:
        return PriceResolution("", "missing_rpo", "blank RPO cannot match price schedule")

    matches = price_rows_by_rpo.get(rpo, [])
    if not matches:
        return PriceResolution("", "no_price_match", "no price_sched_raw option row matched the RPO")
    if len(matches) == 1:
        return _resolution_from_row(matches[0], "unique_price_match", "single price schedule row matched the RPO", 1)
    return _resolve_multi_match(model_key, rpo, matches)


def _sheet_headers(ws) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def build_future_option_price_plan(wb, model_keys: Iterable[str]) -> dict[str, Any]:
    selected_model_keys = _selected_future_model_keys(model_keys)
    section_modes = _section_selection_modes(wb)
    prices_by_rpo = _price_rows_by_rpo(wb)
    plan: dict[str, Any] = {
        "selected_model_keys": selected_model_keys,
        "models": {},
        "price_schedule_rpo_count": len(prices_by_rpo),
        "price_schedule_row_count": sum(len(rows) for rows in prices_by_rpo.values()),
    }

    for model_key in selected_model_keys:
        sheet_name = FUTURE_MODEL_SPECS[model_key].target_option_sheet
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Missing target option sheet {sheet_name}")
        ws = wb[sheet_name]
        headers = _sheet_headers(ws)
        if headers != list(OPTION_SOURCE_HEADERS):
            raise RuntimeError(f"{sheet_name} header drift: expected {list(OPTION_SOURCE_HEADERS)}, found {headers}")
        price_col = headers.index("price") + 1
        rows: list[dict[str, Any]] = []
        desired_prices: dict[str, int | float | str] = {}
        resolutions: dict[str, dict[str, Any]] = {}
        updates: list[dict[str, Any]] = []
        counts: Counter[str] = Counter()

        for row_number in range(2, ws.max_row + 1):
            row = {header: ws.cell(row_number, col).value for col, header in enumerate(headers, start=1)}
            if not any(clean(value) for value in row.values()):
                continue
            option_id = clean(row.get("option_id")) or f"row_{row_number}"
            current_price = _price_value(ws.cell(row_number, price_col).value)
            resolution = resolve_option_price(model_key=model_key, option_row=row, section_modes=section_modes, price_rows_by_rpo=prices_by_rpo)
            desired_price = resolution.price
            counts[resolution.status] += 1
            desired_prices[option_id] = desired_price
            resolutions[option_id] = {
                "row_number": row_number,
                "rpo": clean(row.get("rpo")),
                "section_id": clean(row.get("section_id")),
                "current_price": current_price,
                "desired_price": desired_price,
                "status": resolution.status,
                "reason": resolution.reason,
                "raw_price_row": resolution.raw_price_row,
                "match_count": resolution.match_count,
            }
            if current_price != desired_price:
                updates.append(
                    {
                        "row_number": row_number,
                        "option_id": option_id,
                        "rpo": clean(row.get("rpo")),
                        "current_price": current_price,
                        "desired_price": desired_price,
                        "status": resolution.status,
                        "raw_price_row": resolution.raw_price_row,
                    }
                )
            rows.append({"row_number": row_number, "option_id": option_id, "desired_price": desired_price})

        plan["models"][model_key] = {
            "model_key": model_key,
            "target_option_sheet": sheet_name,
            "option_rows": len(rows),
            "update_count": len(updates),
            "resolution_counts": dict(sorted(counts.items())),
            "desired_prices": desired_prices,
            "resolutions": resolutions,
            "updates": updates,
            "rows": rows,
        }
    return plan


def apply_price_plan_to_workbook(wb, plan: dict[str, Any]) -> None:
    for model_plan in plan["models"].values():
        ws = wb[model_plan["target_option_sheet"]]
        headers = _sheet_headers(ws)
        price_col = headers.index("price") + 1
        for row in model_plan["rows"]:
            ws.cell(int(row["row_number"]), price_col).value = row["desired_price"]


def _strip_details(plan: dict[str, Any]) -> dict[str, Any]:
    report = {
        "selected_model_keys": plan["selected_model_keys"],
        "price_schedule_rpo_count": plan["price_schedule_rpo_count"],
        "price_schedule_row_count": plan["price_schedule_row_count"],
        "models": {},
    }
    for model_key, model_plan in plan["models"].items():
        report["models"][model_key] = {
            key: value
            for key, value in model_plan.items()
            if key not in {"rows", "resolutions", "desired_prices", "updates"}
        }
    return report


def _assert_future_metadata_inactive(wb) -> None:
    future_models = set(FUTURE_MODEL_KEYS)
    checks = {
        "model_master": ("model_key", ["active"]),
        "model_workbook_sources": ("model_key", ["active"]),
        "model_registry_promotion": ("model_key", ["promoted_to_runtime", "active"]),
    }
    for sheet_name, (model_field, inactive_fields) in checks.items():
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Missing metadata sheet {sheet_name}")
        ws = wb[sheet_name]
        headers = _sheet_headers(ws)
        index = {header: offset + 1 for offset, header in enumerate(headers)}
        if model_field not in index:
            raise RuntimeError(f"{sheet_name} missing {model_field}")
        for field in inactive_fields:
            if field not in index:
                raise RuntimeError(f"{sheet_name} missing {field}")
        for row_number in range(2, ws.max_row + 1):
            model_key = clean(ws.cell(row_number, index[model_field]).value)
            if model_key not in future_models:
                continue
            for field in inactive_fields:
                value = clean(ws.cell(row_number, index[field]).value).casefold()
                if value in {"true", "1", "yes", "y", "active"}:
                    raise RuntimeError(f"{sheet_name} row {row_number} unexpectedly has active {field}={value}")


def verify_saved_workbook(path: Path, plan: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        verification: dict[str, Any] = {"models": {}}
        for model_key, model_plan in plan["models"].items():
            sheet_name = model_plan["target_option_sheet"]
            ws = wb[sheet_name]
            headers = _sheet_headers(ws)
            price_col = headers.index("price") + 1
            option_id_col = headers.index("option_id") + 1
            checked = 0
            for row_number in range(2, ws.max_row + 1):
                option_id = clean(ws.cell(row_number, option_id_col).value)
                if not option_id or option_id not in model_plan["desired_prices"]:
                    continue
                actual = _price_value(ws.cell(row_number, price_col).value)
                expected = model_plan["desired_prices"][option_id]
                if actual != expected:
                    raise RuntimeError(f"{sheet_name} {option_id} expected price {expected!r}, found {actual!r}")
                checked += 1
            if checked != model_plan["option_rows"]:
                raise RuntimeError(f"{sheet_name} expected to verify {model_plan['option_rows']} rows, verified {checked}")
            verification["models"][model_key] = {
                "option_sheet": sheet_name,
                "verified_option_rows": checked,
                "update_count": model_plan["update_count"],
                "resolution_counts": model_plan["resolution_counts"],
            }
        _assert_future_metadata_inactive(wb)
        verification["future_metadata_inactive"] = True
        return verification
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model-key", choices=("all", *FUTURE_MODEL_KEYS), default="all")
    parser.add_argument("--dry-run", action="store_true", help="Report option price changes without saving the workbook.")
    parser.add_argument("--include-details", action="store_true", help="Include per-option resolution details in the JSON report.")
    args = parser.parse_args(argv)

    if not args.dry_run and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    selected = list(FUTURE_MODEL_KEYS) if args.model_key == "all" else [args.model_key]
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    backup_path = None
    try:
        plan = build_future_option_price_plan(wb, selected)
        report = _strip_details(plan) if not args.include_details else plan
        if args.dry_run:
            print(json.dumps({"status": "dry_run", **report}, indent=2))
            return 0
        apply_price_plan_to_workbook(wb, plan)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_workbook(WORKBOOK_PATH, plan)
    print(
        json.dumps(
            {
                "status": "written",
                "workbook": str(WORKBOOK_PATH),
                "backup": str(backup_path),
                **(_strip_details(plan) if not args.include_details else plan),
                "verification": verification,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
