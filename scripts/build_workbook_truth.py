#!/usr/bin/env python3
"""The independent workbook-truth snapshot (spec §6.2).

    python scripts/build_workbook_truth.py [--workbook <path>] [--out <path>]

A parity test compares two independent paths (spec §4.2):

    expected: direct, simple read of authoritative workbook rows
    actual:   generator -> runtime contract -> candidate registry/runtime

This module is the expected side, built once per run instead of once per gate.
It reads a workbook with a read-only openpyxl handle and shapes the result using
`workbook_domain.registry` metadata — sheet families, key columns, source roles.
Nothing else is imported. Specifically it does NOT import `model_configs`,
`model_generation`, `source_assembly`, `rule_derivation`, `runtime_contract`,
`inspection`, `contract`, `rules`, or `output`: a shared generator function
cannot be both implementation and oracle, and `tests/test_workbook_truth.py`
asserts that import boundary rather than trusting this comment.

For the same reason the two cell-representation helpers below are implemented
here rather than imported from `corvette_form_generator.workbook`, even though
that module owns the same contract for the generator. They are six lines; a
shared helper would make one class of representation bug invisible to every
parity gate at once. `tests/test_workbook_truth.py` pins them to the generator's
behavior over a value table, so "independent" does not mean "allowed to drift".

What this module may normalize, and nothing beyond it:

- cell representation, matching `workbook.clean`;
- the workbook's truthiness convention, matching `workbook.workbook_truthy`;
- physical row identity, from the registry's declared key columns;
- `asset_map` addressing, which is a workbook lookup rule rather than a
  generation rule — and even there an unadjudicable case is reported as a
  conflict instead of resolved.

It does not filter rows by scope, derive rules, apply fallbacks, or decide what
generation would emit. Callers state those relationships themselves, so that the
rule a gate is checking is visible in the gate.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.workbook_domain import registry as reg

SCHEMA_VERSION = "workbook-truth-1"
DEFAULT_WORKBOOK = SCRIPTS_DIR.parent / "stingray_master.xlsx"

# asset_map addressing, stated here so the snapshot's one derived view says what
# rule it applied. Wildcard rows are shared media for option targets only; an
# exact model row wins over a wildcard for the same target; a blank model_key is
# neither and applies to nothing.
WILDCARD_MODEL_KEY = "*"
WILDCARD_TARGET_TYPES = ("option",)
ASSET_IMAGE_FIELDS = (
    "image_url",
    "image_alt",
    "image_fit",
    "image_position",
    "hover_image_url",
    "hover_image_alt",
    "hover_image_position",
)


def clean(value: Any) -> str:
    """One cell as the string the workbook contract represents it by."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def truthy(value: Any) -> bool:
    """The workbook's own truthiness convention."""

    return clean(value).lower() in {"true", "yes", "1", "y"}


def read_sheet(wb, sheet_name: str) -> dict[str, Any] | None:
    """Headers and cleaned rows for one sheet, or None when it is absent.

    A missing sheet reads back as None rather than as an empty list, so a caller
    can tell "this workbook has no such sheet" apart from "this sheet is empty".
    Blank rows are dropped; every other row is kept exactly as authored.
    """

    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = {header: clean(value) for header, value in zip(headers, raw) if header}
        if any(record.values()):
            rows.append(record)
    return {"headers": [h for h in headers if h], "rows": rows}


def registered_sheets(source_rows: list[dict[str, str]]) -> dict[str, str]:
    """Every sheet the registry knows about, mapped to its family.

    Global and read-only families are addressed by their own fixed sheet names;
    per-model source sheets are discovered from the workbook's own active
    `model_workbook_sources` rows, so registering a new model or renaming a
    source sheet widens the snapshot without editing this file.
    """

    sheets = dict(reg.GLOBAL_SHEET_FAMILIES)
    for family, meta in reg.READONLY_SHEET_META.items():
        sheets[meta["sheet"]] = family
    for row in source_rows:
        if not truthy(row.get("active")):
            continue
        family = reg.SOURCE_ROLE_FAMILIES.get(clean(row.get("source_role")))
        sheet = clean(row.get("sheet_name"))
        if family and sheet:
            sheets[sheet] = family
    return sheets


def family_key_columns() -> dict[str, tuple[str, ...]]:
    keys = {family: tuple(meta.get("key", ())) for family, meta in reg.EDITOR_SHEET_META.items()}
    for family, meta in reg.READONLY_SHEET_META.items():
        keys[family] = tuple(meta.get("key", ()))
    return keys


def row_identity(row: dict[str, str], key_columns: tuple[str, ...]) -> str:
    """A row's physical identity, from the registry's declared key columns."""

    return "::".join(clean(row.get(column)) for column in key_columns)


def index_unique(
    rows: list[dict[str, str]],
    key_column: str,
    sheet_name: str,
    conflicts: list[dict[str, str]],
) -> dict[str, dict[str, str]]:
    """Index rows by a key that must be unique, recording collisions.

    A second row for the same key is not adjudicable by any rule the workbook
    states, so it is reported the same way a duplicate `asset_map` row is:
    surfaced as a conflict rather than silently resolved by row order. Without
    this, a duplicate promotion or variant row would be a last-write-wins
    decision made by whichever row happened to be lower in the sheet.
    """

    indexed: dict[str, dict[str, str]] = {}
    for row in rows:
        key = clean(row.get(key_column))
        if not key:
            continue
        if key in indexed:
            conflicts.append({"sheet": sheet_name, key_column: key})
            continue
        indexed[key] = row
    return indexed


def build_models(sheets: dict[str, Any]) -> tuple[dict[str, dict[str, Any]], list[dict[str, str]]]:
    """Model topology, assembled only from the workbook's own metadata rows."""

    def rows(name: str) -> list[dict[str, str]]:
        entry = sheets.get(name)
        return entry["rows"] if entry else []

    conflicts: list[dict[str, str]] = []
    variant_facts = index_unique(rows("variant_master"), "variant_id", "variant_master", conflicts)
    promotion_by_model = index_unique(
        rows("model_registry_promotion"), "model_key", "model_registry_promotion", conflicts
    )

    models: dict[str, dict[str, Any]] = {}
    for row in rows("model_master"):
        model_key = clean(row.get("model_key"))
        if not model_key:
            continue
        models[model_key] = {
            "model_key": model_key,
            "registry_key": clean(row.get("registry_key")),
            "model_label": clean(row.get("model_label")),
            "model_year": clean(row.get("model_year")),
            "expected_variant_count": clean(row.get("expected_variant_count")),
            "default_model": truthy(row.get("default_model")),
            "active": truthy(row.get("active")),
            "export_slug": clean(row.get("export_slug")),
            "dataset_name": clean(row.get("dataset_name")),
            # The whole authored row as well, so a parity gate can compare a
            # column this dict has not named without the snapshot needing an
            # edit first. Presentation copy in particular lives in a growing
            # set of `setup_*` columns.
            "master_row": dict(row),
            "source_sheets": {},
            "variants": [],
            "inactive_variant_ids": [],
            "interior_scope": [],
            "promotion": None,
        }

    for row in rows("model_workbook_sources"):
        model_key = clean(row.get("model_key"))
        if model_key not in models or not truthy(row.get("active")):
            continue
        role = clean(row.get("source_role"))
        if role:
            models[model_key]["source_sheets"][role] = clean(row.get("sheet_name"))

    for row in rows("model_variants"):
        model_key = clean(row.get("model_key"))
        if model_key not in models:
            continue
        variant_id = clean(row.get("variant_id"))
        if not truthy(row.get("active")):
            models[model_key]["inactive_variant_ids"].append(variant_id)
            continue
        fact = variant_facts.get(variant_id) or {}
        models[model_key]["variants"].append(
            {
                "variant_id": variant_id,
                "membership_display_order": clean(row.get("display_order")),
                # Present and active in variant_master, or not. Generation and
                # this snapshot disagreeing about that is exactly the kind of
                # thing a parity gate exists to surface, so it is recorded here
                # rather than filtered out.
                "declared_in_variant_master": variant_id in variant_facts,
                "active_in_variant_master": truthy(fact.get("active")),
                "variant_label": clean(fact.get("variant_label")),
                "body_style": clean(fact.get("body_style")),
                "trim_level": clean(fact.get("trim_level")),
                "base_price": clean(fact.get("base_price")),
                "display_order": clean(fact.get("display_order")),
                "model_year": clean(fact.get("model_year")),
            }
        )

    for row in rows("model_interior_scope"):
        model_key = clean(row.get("model_key"))
        if model_key not in models or not truthy(row.get("active")):
            continue
        models[model_key]["interior_scope"].append(
            {
                "interior_id": clean(row.get("interior_id")),
                "trim_level": clean(row.get("trim_level")),
            }
        )

    for model_key, model in models.items():
        promotion = promotion_by_model.get(model_key)
        if promotion is None:
            continue
        model["promotion"] = {
            "model_key": model_key,
            "registry_key": clean(promotion.get("registry_key")),
            "artifact_type": clean(promotion.get("artifact_type")),
            "artifact_path": clean(promotion.get("artifact_path")),
            "legacy_alias": clean(promotion.get("legacy_alias")),
            "promoted_to_runtime": truthy(promotion.get("promoted_to_runtime")),
            "default_model": truthy(promotion.get("default_model")),
            "display_order": clean(promotion.get("display_order")),
            "active": truthy(promotion.get("active")),
        }

    return models, conflicts


def build_promotions(models: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """The published-registry membership the workbook declares.

    A promoted row is active AND promoted_to_runtime. Ordering is by the
    declared `display_order`, with the model key breaking ties so the sequence
    is total rather than dependent on row order.
    """

    promoted = [
        model["promotion"]
        for model in models.values()
        if model["promotion"] and model["promotion"]["active"] and model["promotion"]["promoted_to_runtime"]
    ]
    promoted.sort(key=lambda row: (int(row["display_order"] or 0), row["model_key"]))
    defaults = [row["model_key"] for row in promoted if row["default_model"]]
    return {
        "promoted_model_keys": [row["model_key"] for row in promoted],
        "promoted_registry_keys": [models[row["model_key"]]["registry_key"] for row in promoted],
        "rows": promoted,
        # A list, not a value: exactly-one is an assertion the gate makes, not
        # something this snapshot may quietly pick a winner for.
        "default_model_keys": defaults,
    }


def build_assets(
    asset_rows: list[dict[str, str]], model_keys: list[str]
) -> tuple[dict[str, dict[str, dict[str, str]]], list[dict[str, str]]]:
    """Applicable asset per model, keyed by `<target_type>::<target_id>`.

    The addressing rule is the wildcard/exact precedence described at the top of
    this module. Two active rows sharing a model_key and a target cannot be
    adjudicated by any rule, so they are reported as conflicts rather than
    resolved by row order — resolving them silently is how an oracle ends up
    wrong in the same direction as the code it is checking.
    """

    conflicts: list[dict[str, str]] = []
    resolved: dict[str, dict[str, dict[str, str]]] = {key: {} for key in model_keys}

    def usable(row: dict[str, str]) -> tuple[str, str, dict[str, str]] | None:
        target_type = clean(row.get("target_type"))
        target_id = clean(row.get("target_id"))
        fields = {field: clean(row.get(field)) for field in ASSET_IMAGE_FIELDS}
        if not target_type or not target_id or not fields["image_url"]:
            return None
        return target_type, target_id, fields

    for pass_model in [WILDCARD_MODEL_KEY, *model_keys]:
        seen: set[str] = set()
        for row in asset_rows:
            if not truthy(row.get("active")):
                continue
            if clean(row.get("model_key")) != pass_model:
                continue
            usable_row = usable(row)
            if usable_row is None:
                continue
            target_type, target_id, fields = usable_row
            if pass_model == WILDCARD_MODEL_KEY and target_type not in WILDCARD_TARGET_TYPES:
                continue
            target = f"{target_type}::{target_id}"
            if target in seen:
                conflicts.append({"model_key": pass_model, "target": target})
                continue
            seen.add(target)
            if pass_model == WILDCARD_MODEL_KEY:
                for model_key in model_keys:
                    resolved[model_key][target] = fields
            else:
                resolved[pass_model][target] = fields

    return resolved, conflicts


def build_workbook_truth(workbook_path: Path) -> dict[str, Any]:
    """The whole snapshot for one workbook."""

    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        source_entry = read_sheet(wb, "model_workbook_sources")
        source_rows = source_entry["rows"] if source_entry else []
        families = registered_sheets(source_rows)
        keys = family_key_columns()

        sheets: dict[str, Any] = {}
        missing: list[str] = []
        for sheet_name, family in sorted(families.items()):
            entry = read_sheet(wb, sheet_name)
            if entry is None:
                missing.append(sheet_name)
                continue
            key_columns = keys.get(family, ())
            sheets[sheet_name] = {
                "family": family,
                "key_columns": list(key_columns),
                "headers": entry["headers"],
                "rows": entry["rows"],
                "row_identities": [row_identity(row, key_columns) for row in entry["rows"]]
                if key_columns
                else [],
            }
    finally:
        wb.close()

    models, topology_conflicts = build_models(sheets)
    promotions = build_promotions(models)
    asset_entry = sheets.get("asset_map")
    assets, asset_conflicts = build_assets(
        asset_entry["rows"] if asset_entry else [],
        sorted(models),
    )

    return {
        "schemaVersion": SCHEMA_VERSION,
        "workbook": {
            "name": workbook_path.name,
            "sha256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        },
        "registeredSheetFamilies": families,
        "familyKeyColumns": {family: list(columns) for family, columns in sorted(keys.items())},
        "registeredSheetsMissingFromWorkbook": sorted(missing),
        "sheets": sheets,
        "models": models,
        "promotions": promotions,
        "assets": assets,
        "assetConflicts": asset_conflicts,
        # Duplicate rows on the sheets whose keys the topology assumes unique.
        # Reported, never resolved — see `index_unique`.
        "topologyConflicts": topology_conflicts,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", type=Path, help="write the snapshot JSON here instead of stdout")
    args = parser.parse_args(argv)

    snapshot = build_workbook_truth(args.workbook)
    payload = json.dumps(snapshot, sort_keys=True)
    if args.out is None:
        sys.stdout.write(payload + "\n")
    else:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(payload + "\n", encoding="utf-8")
        sys.stdout.write(f"{args.out}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
