#!/usr/bin/env python3
"""Apply approved Z06 package/pricing cascade workbook rows.

Dry-run is the default. Use --write to safely save stingray_master.xlsx.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely  # noqa: E402

SECTION_MASTER_SHEET = "section_master"
OPTION_SHEET = "z06_options"
RULE_MAPPING_SHEET = "z06_rule_mapping"
RULE_GROUPS_SHEET = "z06_rule_groups"
RULE_GROUP_MEMBERS_SHEET = "z06_rule_group_members"
PRICE_RULES_SHEET = "z06_price_rules"
EXCLUSIVE_GROUPS_SHEET = "z06_exclusive_groups"
EXCLUSIVE_MEMBERS_SHEET = "z06_exclusive_members"

PACKAGE_SECTION_ID = "sec_z06_pkg_001"
CARBON_WHEEL_SECTION_ID = "sec_z06_cf_whee_001"
WHEELS_SECTION_ID = "sec_whee_002"

PACKAGE_RPOS = ("PDB", "PDD", "PDF")
CARBON_WHEEL_RPOS = ("ROY", "ROZ", "STZ")
DIRECT_PRICES = {
    "Z07": 9500,
    "T0F": 8995,
    "T0G": 10995,
    "CFZ": 3495,
    "CFV": 4495,
    "ROY": 11995,
    "ROZ": 13995,
    "STZ": 15500,
}
PACKAGE_PRICES = {
    "PDB": {"ROY": 16000, "ROZ": 17000, "STZ": 17500},
    "PDD": {"ROY": 25495, "ROZ": 26495, "STZ": 26995},
    "PDF": {"ROY": 26495, "ROZ": 27495, "STZ": 27995},
}


@dataclass(frozen=True)
class Change:
    sheet: str
    row_number: int | str
    key: str
    field: str
    current: Any
    desired: Any
    reason: str


def headers(ws) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def index(ws) -> dict[str, int]:
    return {header: offset + 1 for offset, header in enumerate(headers(ws))}


def rows(ws):
    idx = index(ws)
    for row_number in range(2, ws.max_row + 1):
        row = {header: ws.cell(row_number, col).value for header, col in idx.items()}
        if any(value not in (None, "") for value in row.values()):
            yield row_number, row


def boolish(value: Any) -> str:
    text = clean(value).casefold()
    if text in {"true", "1", "yes", "y"}:
        return "True"
    if text in {"false", "0", "no", "n"}:
        return "False"
    return clean(value)


def price_equal(current: Any, desired: Any) -> bool:
    if desired in (None, ""):
        return clean(current) == ""
    try:
        return float(clean(current) or 0) == float(desired)
    except ValueError:
        return False


def row_by_key(ws, key_field: str, key: str) -> tuple[int | None, dict[str, Any] | None]:
    for row_number, row in rows(ws):
        if clean(row.get(key_field)) == key:
            return row_number, row
    return None, None


def option_maps(wb) -> tuple[dict[str, dict[str, Any]], dict[str, str]]:
    ws = wb[OPTION_SHEET]
    by_rpo: dict[str, dict[str, Any]] = {}
    id_by_rpo: dict[str, str] = {}
    for row_number, row in rows(ws):
        rpo = clean(row.get("rpo"))
        option_id = clean(row.get("option_id"))
        if rpo and option_id:
            by_rpo[rpo] = {**row, "_row_number": row_number}
            id_by_rpo[rpo] = option_id
    return by_rpo, id_by_rpo


def set_cell(ws, row_number: int, field: str, desired: Any, changes: list[Change], *, key: str, reason: str) -> None:
    idx = index(ws)
    current = ws.cell(row_number, idx[field]).value
    equal = price_equal(current, desired) if field in {"price", "price_value"} else clean(current) == clean(desired)
    if not equal:
        changes.append(Change(ws.title, row_number, key, field, current, desired, reason))
        ws.cell(row_number, idx[field]).value = desired


def ensure_row(ws, key_field: str, key: str, values: dict[str, Any], changes: list[Change], *, reason: str) -> None:
    idx = index(ws)
    row_number, existing = row_by_key(ws, key_field, key)
    if row_number is None:
        row_number = ws.max_row + 1
        for field, value in values.items():
            ws.cell(row_number, idx[field]).value = value
        changes.append(Change(ws.title, f"new:{row_number}", key, "row", None, values, reason))
        return
    for field, desired in values.items():
        if field == key_field:
            continue
        set_cell(ws, row_number, field, desired, changes, key=key, reason=reason)


def ensure_section_rows(wb, changes: list[Change]) -> None:
    ws = wb[SECTION_MASTER_SHEET]
    ensure_row(
        ws,
        "section_id",
        PACKAGE_SECTION_ID,
        {
            "section_id": PACKAGE_SECTION_ID,
            "section_name": "Z06 Carbon Fiber Wheel and Brake Packages",
            "selection_mode": "single_select_opt",
            "is_required": "FALSE",
            "display_order": 12,
            "standard_behavior": "user_selected",
            "help_text": "Choose PDB, PDD, or PDF here; Z07 remains the adjacent standalone performance package.",
            "step_key": "packages_performance",
        },
        changes,
        reason="approved Z06 package section",
    )
    ensure_row(
        ws,
        "section_id",
        CARBON_WHEEL_SECTION_ID,
        {
            "section_id": CARBON_WHEEL_SECTION_ID,
            "section_name": "Z06 Carbon Fiber Wheel Selection",
            "selection_mode": "single_select_opt",
            "is_required": "FALSE",
            "display_order": 15,
            "standard_behavior": "user_selected",
            "help_text": "For Z06 carbon fiber wheel packages, select ROY, ROZ, or STZ in this Performance & Aero step. These replace the wheel choice from Wheels & Brake Calipers.",
            "step_key": "packages_performance",
        },
        changes,
        reason="approved Z06 carbon wheel package section",
    )
    row_number, _ = row_by_key(ws, "section_id", WHEELS_SECTION_ID)
    if row_number is not None:
        set_cell(
            ws,
            row_number,
            "help_text",
            "Z06 carbon fiber wheel choices ROY, ROZ, and STZ are selected in Performance & Aero with the carbon fiber wheel packages.",
            changes,
            key=WHEELS_SECTION_ID,
            reason="approved wheels step guidance message",
        )


def update_options(wb, changes: list[Change]) -> None:
    ws = wb[OPTION_SHEET]
    by_rpo, _ = option_maps(wb)
    for display_order, rpo in enumerate(PACKAGE_RPOS, start=1):
        row = by_rpo[rpo]
        row_number = int(row["_row_number"])
        set_cell(ws, row_number, "section_id", PACKAGE_SECTION_ID, changes, key=rpo, reason="approved package section move")
        set_cell(ws, row_number, "display_order", str(display_order * 10), changes, key=rpo, reason="package section display order")
    for display_order, rpo in enumerate(CARBON_WHEEL_RPOS, start=1):
        row = by_rpo[rpo]
        row_number = int(row["_row_number"])
        set_cell(ws, row_number, "section_id", CARBON_WHEEL_SECTION_ID, changes, key=rpo, reason="approved carbon wheel performance placement")
        set_cell(ws, row_number, "display_order", str(display_order * 10), changes, key=rpo, reason="carbon wheel section display order")
    for rpo, price in DIRECT_PRICES.items():
        row = by_rpo[rpo]
        set_cell(ws, int(row["_row_number"]), "price", price, changes, key=rpo, reason="approved direct Z06 package-related price")


def rule_values(rule_id: str, source_id: str, rule_type: str, target_id: str, original: str, *, source_row: dict[str, Any], target_row: dict[str, Any], runtime_action: str = "", body_style_scope: str = "") -> dict[str, Any]:
    return {
        "rule_id": rule_id,
        "source_id": source_id,
        "rule_type": rule_type,
        "target_id": target_id,
        "target_type": "option",
        "original_detail_raw": original,
        "review_flag": False,
        "source_type": "option",
        "target_selection_mode": "",
        "source_selection_mode": "",
        "target_section": clean(target_row.get("section_id")),
        "source_section": clean(source_row.get("section_id")),
        "generation_action": "",
        "body_style_scope": body_style_scope,
        "runtime_action": runtime_action,
        "disabled_reason": "",
        "normalization_status": "active",
        "normalization_reason": "",
        "replacement_group_id": "",
        "replacement_rule_id": "",
    }


def ensure_rule_mapping(wb, changes: list[Change]) -> None:
    ws = wb[RULE_MAPPING_SHEET]
    by_rpo, id_by_rpo = option_maps(wb)

    include_pairs = [
        ("PDB", "J57", "PDB includes J57 carbon ceramic brakes."),
        ("PDB", "J6D", "PDB includes J6D Dark Gray Metallic-painted calipers."),
        ("PDD", "Z07", "PDD includes Z07 Performance Package."),
        ("PDD", "T0F", "PDD includes Carbon Flash-painted Carbon Fiber Aero Package."),
        ("PDD", "CFZ", "PDD includes Carbon Flash-painted carbon fiber ground effects."),
        ("PDF", "Z07", "PDF includes Z07 Performance Package."),
        ("PDF", "T0G", "PDF includes visible Carbon Fiber Aero Package."),
        ("PDF", "CFV", "PDF includes visible carbon fiber ground effects."),
        ("Z07", "J57", "Z07 includes J57 carbon ceramic brakes."),
        ("Z07", "FE7", "Z07 includes FE7 suspension."),
        ("Z07", "XFS", "Z07 includes XFS Michelin Pilot Sport Cup 2 R tires."),
        ("T0G", "CFV", "T0G includes visible carbon fiber ground effects."),
    ]
    for source_rpo, target_rpo, note in include_pairs:
        source_id = id_by_rpo[source_rpo]
        target_id = id_by_rpo[target_rpo]
        rule_id = f"z06_rule_{source_id}_includes_{target_id}"
        ensure_row(
            ws,
            "rule_id",
            rule_id,
            rule_values(rule_id, source_id, "includes", target_id, note, source_row=by_rpo[source_rpo], target_row=by_rpo[target_rpo]),
            changes,
            reason="approved package include rule",
        )

    exclude_pairs = [
        ("PDD", "T0G", "PDD uses T0F and blocks the visible aero package peer.", ""),
        ("PDF", "T0F", "PDF uses T0G and blocks the Carbon Flash aero package peer.", ""),
    ]
    for source_rpo, target_rpo, note, body_scope in exclude_pairs:
        source_id = id_by_rpo[source_rpo]
        target_id = id_by_rpo[target_rpo]
        rule_id = f"z06_rule_{source_id}_excludes_{target_id}"
        values = rule_values(rule_id, source_id, "excludes", target_id, note, source_row=by_rpo[source_rpo], target_row=by_rpo[target_rpo], body_style_scope=body_scope)
        values["disabled_reason"] = f"Blocked by {source_rpo}."
        ensure_row(ws, "rule_id", rule_id, values, changes, reason="approved package peer exclusion")

    bcw_rule_id = f"z06_rule_{id_by_rpo['BCW']}_requires_{id_by_rpo['ZZ3']}_convertible"
    bcw_values = rule_values(
        bcw_rule_id,
        id_by_rpo["BCW"],
        "requires",
        id_by_rpo["ZZ3"],
        "BCW red engine intake is available on convertible only with ZZ3.",
        source_row=by_rpo["BCW"],
        target_row=by_rpo["ZZ3"],
        body_style_scope="convertible",
    )
    bcw_values["disabled_reason"] = "Convertible BCW requires ZZ3 Convertible Engine Appearance Package."
    ensure_row(ws, "rule_id", bcw_rule_id, bcw_values, changes, reason="approved BCW convertible availability rule")

    # Moving carbon wheels into a separate performance section requires workbook-owned replace rules
    # so either a normal wheel or a carbon wheel can replace the current wheel selection across sections.
    all_wheel_rows = [row for row in by_rpo.values() if clean(row.get("section_id")) in {WHEELS_SECTION_ID, CARBON_WHEEL_SECTION_ID} and boolish(row.get("active")) == "True"]
    carbon_ids = {id_by_rpo[rpo] for rpo in CARBON_WHEEL_RPOS}
    normal_wheel_rows = [row for row in all_wheel_rows if clean(row.get("option_id")) not in carbon_ids]
    carbon_rows = [by_rpo[rpo] for rpo in CARBON_WHEEL_RPOS]
    for source in [*normal_wheel_rows, *carbon_rows]:
        for target in [*normal_wheel_rows, *carbon_rows]:
            source_id = clean(source.get("option_id"))
            target_id = clean(target.get("option_id"))
            if source_id == target_id:
                continue
            # Same-section wheel choices are already handled by the section's single-select behavior.
            if clean(source.get("section_id")) == clean(target.get("section_id")):
                continue
            source_rpo = clean(source.get("rpo"))
            target_rpo = clean(target.get("rpo"))
            rule_id = f"z06_rule_{source_id}_replaces_{target_id}"
            values = rule_values(
                rule_id,
                source_id,
                "excludes",
                target_id,
                f"{source_rpo} replaces wheel choice {target_rpo} across the Z06 wheel/package sections.",
                source_row=source,
                target_row=target,
                runtime_action="replace",
            )
            values["disabled_reason"] = f"Replaced by {source_rpo}."
            ensure_row(ws, "rule_id", rule_id, values, changes, reason="cross-section Z06 wheel replacement")


def ensure_rule_groups(wb, changes: list[Change]) -> None:
    groups_ws = wb[RULE_GROUPS_SHEET]
    members_ws = wb[RULE_GROUP_MEMBERS_SHEET]
    _, id_by_rpo = option_maps(wb)
    z07_id = id_by_rpo["Z07"]
    group_id = "z06_group_z07_requires_aero"
    ensure_row(
        groups_ws,
        "group_id",
        group_id,
        {
            "group_id": group_id,
            "group_type": "requires_any",
            "source_id": z07_id,
            "body_style_scope": "*",
            "trim_level_scope": "*",
            "variant_scope": "*",
            "disabled_reason": "Requires T0F Carbon Flash aero or T0G visible carbon aero.",
            "active": True,
            "notes": "Z07 requires one of T0F or T0G; PDD/PDF include the required aero choice.",
        },
        changes,
        reason="approved Z07 aero requires-any group",
    )
    for order, rpo in enumerate(("T0F", "T0G"), start=1):
        key = f"{group_id}:{id_by_rpo[rpo]}"
        # Composite key sheet has no id field; find/update manually.
        existing_row = None
        for row_number, row in rows(members_ws):
            if clean(row.get("group_id")) == group_id and clean(row.get("target_id")) == id_by_rpo[rpo]:
                existing_row = row_number
                break
        if existing_row is None:
            row_number = members_ws.max_row + 1
            idx = index(members_ws)
            values = {"group_id": group_id, "target_id": id_by_rpo[rpo], "display_order": order * 10, "active": True}
            for field, value in values.items():
                members_ws.cell(row_number, idx[field]).value = value
            changes.append(Change(members_ws.title, f"new:{row_number}", key, "row", None, values, "approved Z07 requires-any member"))
        else:
            set_cell(members_ws, existing_row, "display_order", order * 10, changes, key=key, reason="approved Z07 requires-any member")
            set_cell(members_ws, existing_row, "active", True, changes, key=key, reason="approved Z07 requires-any member")


def ensure_exclusive_groups(wb, changes: list[Change]) -> None:
    groups_ws = wb[EXCLUSIVE_GROUPS_SHEET]
    members_ws = wb[EXCLUSIVE_MEMBERS_SHEET]
    _, id_by_rpo = option_maps(wb)
    ensure_row(
        groups_ws,
        "group_id",
        "z06_excl_carbon_wheel_packages",
        {
            "group_id": "z06_excl_carbon_wheel_packages",
            "selection_mode": "single_within_group",
            "active": True,
            "notes": "PDB, PDD, and PDF are mutually exclusive Z06 package choices. Z07 remains adjacent and can be included by PDD/PDF.",
        },
        changes,
        reason="approved package exclusive group",
    )
    for order, rpo in enumerate(PACKAGE_RPOS, start=1):
        key = f"z06_excl_carbon_wheel_packages:{id_by_rpo[rpo]}"
        existing_row = None
        for row_number, row in rows(members_ws):
            if clean(row.get("group_id")) == "z06_excl_carbon_wheel_packages" and clean(row.get("option_id")) == id_by_rpo[rpo]:
                existing_row = row_number
                break
        if existing_row is None:
            row_number = members_ws.max_row + 1
            idx = index(members_ws)
            values = {"group_id": "z06_excl_carbon_wheel_packages", "option_id": id_by_rpo[rpo], "display_order": order * 10, "active": True}
            for field, value in values.items():
                members_ws.cell(row_number, idx[field]).value = value
            changes.append(Change(members_ws.title, f"new:{row_number}", key, "row", None, values, "approved package exclusive member"))
        else:
            set_cell(members_ws, existing_row, "display_order", order * 10, changes, key=key, reason="approved package exclusive member")
            set_cell(members_ws, existing_row, "active", True, changes, key=key, reason="approved package exclusive member")

    # CFV is the visible-carbon peer of CFZ and should be in the same ground-effects group.
    cfv_id = id_by_rpo["CFV"]
    key = f"z06_excl_ground_effects:{cfv_id}"
    existing_row = None
    for row_number, row in rows(members_ws):
        if clean(row.get("group_id")) == "z06_excl_ground_effects" and clean(row.get("option_id")) == cfv_id:
            existing_row = row_number
            break
    if existing_row is None:
        row_number = members_ws.max_row + 1
        idx = index(members_ws)
        values = {"group_id": "z06_excl_ground_effects", "option_id": cfv_id, "display_order": 30, "active": True}
        for field, value in values.items():
            members_ws.cell(row_number, idx[field]).value = value
        changes.append(Change(members_ws.title, f"new:{row_number}", key, "row", None, values, "approved ground-effects peer member"))
    else:
        set_cell(members_ws, existing_row, "active", True, changes, key=key, reason="approved ground-effects peer member")


def ensure_price_rules(wb, changes: list[Change]) -> None:
    ws = wb[PRICE_RULES_SHEET]
    _, id_by_rpo = option_maps(wb)

    def ensure_price_rule(rule_id: str, condition_rpo: str, target_rpo: str, price: int, note: str, *, body: str = "*", trim: str = "*") -> None:
        ensure_row(
            ws,
            "price_rule_id",
            rule_id,
            {
                "price_rule_id": rule_id,
                "condition_option_id": id_by_rpo[condition_rpo],
                "price_rule_type": "override",
                "target_option_id": id_by_rpo[target_rpo],
                "price_value": price,
                "body_style_scope": body,
                "trim_level_scope": trim,
                "review_flag": False,
                "notes": note,
            },
            changes,
            reason="approved Z06 package/conditional price rule",
        )

    ensure_price_rule("z06_pr_b6p_bcw_895_coupe", "B6P", "BCW", 895, "BCW is $895 on coupe when B6P is selected.", body="coupe")
    ensure_price_rule("z06_pr_zz3_bcw_895_convertible", "ZZ3", "BCW", 895, "BCW is $895 on convertible when ZZ3 is selected.", body="convertible")
    ensure_price_rule("z06_pr_t0f_cfz_zero", "T0F", "CFZ", 0, "T0F includes CFZ, so CFZ does not add a second charge.")
    ensure_price_rule("z06_pr_t0g_cfv_zero", "T0G", "CFV", 0, "T0G includes CFV, so CFV does not add a second charge.")
    ensure_price_rule("z06_pr_z07_j57_zero", "Z07", "J57", 0, "Z07 includes J57, so J57 does not add a second charge.")

    zero_by_package = {
        "PDB": ("J57", "J6D", "ROY", "ROZ", "STZ"),
        "PDD": ("Z07", "J57", "T0F", "CFZ", "ROY", "ROZ", "STZ"),
        "PDF": ("Z07", "J57", "T0G", "CFV", "ROY", "ROZ", "STZ"),
    }
    for package_rpo, targets in zero_by_package.items():
        for target_rpo in targets:
            ensure_price_rule(
                f"z06_pr_{package_rpo.lower()}_{target_rpo.lower()}_zero",
                package_rpo,
                target_rpo,
                0,
                f"{package_rpo} package price includes {target_rpo}, so {target_rpo} does not add a second charge.",
            )

    for package_rpo, wheel_prices in PACKAGE_PRICES.items():
        for wheel_rpo, price in wheel_prices.items():
            ensure_price_rule(
                f"z06_pr_{wheel_rpo.lower()}_{package_rpo.lower()}_{price}",
                wheel_rpo,
                package_rpo,
                price,
                f"{package_rpo} package price with {wheel_rpo} carbon fiber wheels from the raw price schedule.",
            )


def build_plan(wb) -> list[Change]:
    changes: list[Change] = []
    required = [
        SECTION_MASTER_SHEET,
        OPTION_SHEET,
        RULE_MAPPING_SHEET,
        RULE_GROUPS_SHEET,
        RULE_GROUP_MEMBERS_SHEET,
        PRICE_RULES_SHEET,
        EXCLUSIVE_GROUPS_SHEET,
        EXCLUSIVE_MEMBERS_SHEET,
    ]
    missing = [sheet for sheet in required if sheet not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Missing required sheet(s): {', '.join(missing)}")
    ensure_section_rows(wb, changes)
    update_options(wb, changes)
    ensure_rule_mapping(wb, changes)
    ensure_rule_groups(wb, changes)
    ensure_exclusive_groups(wb, changes)
    ensure_price_rules(wb, changes)
    return changes


def verify_saved(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        by_rpo, id_by_rpo = option_maps(wb)
        checks: dict[str, Any] = {}
        checks["package_section"] = {rpo: clean(by_rpo[rpo].get("section_id")) for rpo in PACKAGE_RPOS}
        checks["carbon_wheel_section"] = {rpo: clean(by_rpo[rpo].get("section_id")) for rpo in CARBON_WHEEL_RPOS}
        checks["direct_prices"] = {rpo: by_rpo[rpo].get("price") for rpo in DIRECT_PRICES}
        checks["r8e_section"] = clean(by_rpo["R8E"].get("section_id"))
        if checks["r8e_section"] != "sec_incl_001":
            raise RuntimeError(f"R8E moved unexpectedly: {checks['r8e_section']}")
        for rpo in PACKAGE_RPOS:
            if checks["package_section"][rpo] != PACKAGE_SECTION_ID:
                raise RuntimeError(f"{rpo} not in {PACKAGE_SECTION_ID}")
        for rpo in CARBON_WHEEL_RPOS:
            if checks["carbon_wheel_section"][rpo] != CARBON_WHEEL_SECTION_ID:
                raise RuntimeError(f"{rpo} not in {CARBON_WHEEL_SECTION_ID}")
        for rpo, price in DIRECT_PRICES.items():
            if not price_equal(checks["direct_prices"][rpo], price):
                raise RuntimeError(f"{rpo} price expected {price}, found {checks['direct_prices'][rpo]!r}")

        price_rules = {clean(row.get("price_rule_id")): row for _, row in rows(wb[PRICE_RULES_SHEET])}
        expected_price_rules = [
            "z06_pr_b6p_bcw_895_coupe",
            "z06_pr_zz3_bcw_895_convertible",
            "z06_pr_z07_j57_zero",
            "z06_pr_roy_pdb_16000",
            "z06_pr_stz_pdf_27995",
        ]
        missing_price_rules = [rule_id for rule_id in expected_price_rules if rule_id not in price_rules]
        if missing_price_rules:
            raise RuntimeError(f"Missing expected price rules: {missing_price_rules}")
        rules = {clean(row.get("rule_id")): row for _, row in rows(wb[RULE_MAPPING_SHEET])}
        expected_rules = [
            f"z06_rule_{id_by_rpo['Z07']}_includes_{id_by_rpo['FE7']}",
            f"z06_rule_{id_by_rpo['Z07']}_includes_{id_by_rpo['XFS']}",
            f"z06_rule_{id_by_rpo['PDF']}_includes_{id_by_rpo['T0G']}",
            f"z06_rule_{id_by_rpo['T0G']}_includes_{id_by_rpo['CFV']}",
        ]
        missing_rules = [rule_id for rule_id in expected_rules if rule_id not in rules]
        if missing_rules:
            raise RuntimeError(f"Missing expected rules: {missing_rules}")
        checks["price_rule_count"] = len(price_rules)
        checks["rule_mapping_count"] = len(rules)
        return checks
    finally:
        wb.close()


def summarize(changes: list[Change]) -> dict[str, Any]:
    by_sheet: dict[str, int] = {}
    by_field: dict[str, int] = {}
    for change in changes:
        by_sheet[change.sheet] = by_sheet.get(change.sheet, 0) + 1
        by_field[change.field] = by_field.get(change.field, 0) + 1
    return {
        "total_changes": len(changes),
        "changes_by_sheet": dict(sorted(by_sheet.items())),
        "changes_by_field": dict(sorted(by_field.items())),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="Write workbook changes. Default is dry-run.")
    parser.add_argument("--include-changes", action="store_true", help="Include each cell/row change in the JSON report.")
    args = parser.parse_args(argv)

    if args.write and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")
    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    backup_path = None
    try:
        changes = build_plan(wb)
        report: dict[str, Any] = {"status": "dry_run", **summarize(changes), "workbook": str(WORKBOOK_PATH)}
        if args.include_changes:
            report["changes"] = [change.__dict__ for change in changes]
        if not args.write:
            print(json.dumps(report, indent=2, default=str))
            return 0
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved(WORKBOOK_PATH)
    report = {"status": "written", **summarize(changes), "workbook": str(WORKBOOK_PATH), "backup": str(backup_path), "verification": verification}
    if args.include_changes:
        report["changes"] = [change.__dict__ for change in changes]
    print(json.dumps(report, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
