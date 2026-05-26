#!/usr/bin/env python3
"""Apply simplified future_model_option_review CSV decisions to future source sheets."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from create_future_model_option_review import OPTION_REVIEW_HEADERS, OPTION_REVIEW_SHEET  # noqa: E402
from corvette_form_generator.future_model_ingest import (  # noqa: E402
    FUTURE_MODEL_SPECS,
    OPTION_SOURCE_HEADERS,
    OVS_SOURCE_HEADERS,
    VALID_SOURCE_STATUSES,
)
from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely, write_sheet  # noqa: E402

DEFAULT_CSV_PATH = ROOT / "stingray_master - future_model_option_review.csv"
FUTURE_MODEL_KEYS = tuple(FUTURE_MODEL_SPECS)


def option_review_row_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        clean(row.get("model_key")),
        clean(row.get("raw_source_sheet")),
        clean(row.get("raw_source_span")),
        clean(row.get("source_rpo")),
    )


def rows_from_sheet(wb, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header: clean(value) for header, value in zip(headers, values) if header}
        if any(row.values()):
            rows.append(row)
    return rows


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        expected = list(OPTION_REVIEW_HEADERS)
        if headers != expected:
            raise RuntimeError(f"CSV header drift: expected {expected}, found {headers}")
        return [{header: clean(row.get(header)) for header in expected} for row in reader]


def merge_csv_with_existing_option_review_rows(existing_rows: list[dict[str, Any]], csv_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Overlay CSV decisions but keep workbook-only rows for review provenance."""

    csv_keys = {option_review_row_key(row) for row in csv_rows}
    merged = [dict(row) for row in csv_rows]
    merged.extend(dict(row) for row in existing_rows if option_review_row_key(row) not in csv_keys)
    return [{header: clean(row.get(header)) for header in OPTION_REVIEW_HEADERS} for row in merged]


def _selected_future_model_keys(model_keys: Iterable[str]) -> list[str]:
    requested = [clean(model_key) for model_key in model_keys if clean(model_key)]
    if not requested or requested == ["all"] or "all" in requested:
        return list(FUTURE_MODEL_KEYS)
    unknown = [model_key for model_key in requested if model_key not in FUTURE_MODEL_SPECS]
    if unknown:
        raise ValueError(f"Unknown future model key(s): {', '.join(unknown)}")
    return requested


def _section_ids(wb) -> set[str]:
    return {clean(row.get("section_id")) for row in rows_from_sheet(wb, "section_master") if clean(row.get("section_id"))}


def _current_sheet_row_count(wb, sheet_name: str) -> int:
    return len(rows_from_sheet(wb, sheet_name))


def _parse_status_summary(value: Any) -> dict[str, str]:
    statuses: dict[str, str] = {}
    for part in clean(value).split(";"):
        part = part.strip()
        if not part:
            continue
        if "=" not in part:
            statuses[part] = ""
            continue
        variant_id, status = part.split("=", 1)
        statuses[clean(variant_id)] = clean(status)
    return statuses


def _display_order_value(row: dict[str, Any]) -> str:
    return clean(row.get("final_display_order")) or clean(row.get("suggested_display_order"))


def _intish_string(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    numeric = float(text)
    if numeric.is_integer():
        return str(int(numeric))
    return text


def _sort_key_for_option(row: dict[str, Any]) -> tuple[float, str]:
    display_order = clean(row.get("display_order"))
    if not display_order:
        return (float("inf"), clean(row.get("option_id")))
    return (float(display_order), clean(row.get("option_id")))


def _option_id_for_row(row: dict[str, Any]) -> str:
    return clean(row.get("final_option_id")) or clean(row.get("suggested_option_id"))


def _option_name_for_row(row: dict[str, Any]) -> str:
    return clean(row.get("final_option_name")) or clean(row.get("source_option_description"))


def _detail_raw_for_row(row: dict[str, Any]) -> str:
    return clean(row.get("final_detail_raw")) or clean(row.get("source_disclosure_raw"))


def _validation_errors_for_selected_row(
    row: dict[str, Any],
    *,
    section_ids: set[str],
    variant_ids: Iterable[str],
    duplicate_option_ids: set[str],
) -> list[str]:
    errors: list[str] = []
    option_id = _option_id_for_row(row)
    section_id = clean(row.get("suggested_section_id"))
    statuses = _parse_status_summary(row.get("normalized_status_summary"))
    if not option_id:
        errors.append("option_id is required")
    if option_id and option_id in duplicate_option_ids:
        errors.append(f"duplicate option_id {option_id}")
    if not _option_name_for_row(row):
        errors.append("option_name is required")
    if not section_id:
        errors.append("suggested_section_id is required")
    elif section_id not in section_ids:
        errors.append(f"suggested_section_id {section_id} is not in section_master")
    display_order = _display_order_value(row)
    if display_order:
        try:
            _intish_string(display_order)
        except ValueError:
            errors.append("display_order must be numeric when provided")
    for variant_id in variant_ids:
        status = clean(statuses.get(variant_id))
        if not status:
            errors.append(f"status_{variant_id} is required")
        elif status not in VALID_SOURCE_STATUSES:
            errors.append(f"status_{variant_id} has invalid value {status}")
    return errors


def _option_row_from_review(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "option_id": _option_id_for_row(row),
        "rpo": clean(row.get("source_rpo")),
        "price": "",
        "option_name": _option_name_for_row(row),
        "description": clean(row.get("final_description")),
        "detail_raw": _detail_raw_for_row(row),
        "section_id": clean(row.get("suggested_section_id")),
        "selectable": "True" if clean(row.get("orderable_rpo")) else "False",
        "display_order": _intish_string(_display_order_value(row)),
        "active": "True",
        "display_behavior": clean(row.get("final_display_behavior")),
    }


def _source_label(row: dict[str, Any]) -> str:
    return clean(row.get("raw_source_span")) or _option_id_for_row(row) or clean(row.get("source_option_description")) or "option review row"


def build_future_option_population_plan(wb, option_review_rows: list[dict[str, Any]], model_keys: Iterable[str]) -> dict[str, Any]:
    """Materialize section-filled simplified review rows without mutating a workbook."""

    selected_model_keys = _selected_future_model_keys(model_keys)
    section_ids = _section_ids(wb)
    plan: dict[str, Any] = {
        "selected_model_keys": selected_model_keys,
        "models": {},
        "error_count": 0,
    }

    for model_key in selected_model_keys:
        spec = FUTURE_MODEL_SPECS[model_key]
        variant_ids = list(spec.variant_columns.values())
        model_rows = [row for row in option_review_rows if clean(row.get("model_key")) == model_key]
        selected_rows = [row for row in model_rows if clean(row.get("suggested_section_id"))]
        option_id_counts = Counter(_option_id_for_row(row) for row in selected_rows if _option_id_for_row(row))
        duplicate_option_ids = {option_id for option_id, count in option_id_counts.items() if count > 1}

        option_rows: list[dict[str, Any]] = []
        ovs_rows: list[dict[str, Any]] = []
        blocked_counts: Counter[str] = Counter()
        errors: list[str] = []

        for row in model_rows:
            if not clean(row.get("suggested_section_id")):
                blocked_counts["blank_suggested_section_id"] += 1
                continue
            row_errors = _validation_errors_for_selected_row(row, section_ids=section_ids, variant_ids=variant_ids, duplicate_option_ids=duplicate_option_ids)
            if row_errors:
                for error in row_errors:
                    blocked_counts[error] += 1
                errors.append(f"{_source_label(row)}: {'; '.join(row_errors)}")
                continue
            option_row = _option_row_from_review(row)
            option_rows.append(option_row)
            statuses = _parse_status_summary(row.get("normalized_status_summary"))
            for variant_id in variant_ids:
                ovs_rows.append({"option_id": option_row["option_id"], "variant_id": variant_id, "status": clean(statuses.get(variant_id))})

        display_order = {row["option_id"]: _sort_key_for_option(row) for row in option_rows}
        option_rows.sort(key=lambda row: display_order[row["option_id"]])
        variant_order = {variant_id: index for index, variant_id in enumerate(variant_ids)}
        ovs_rows.sort(key=lambda row: (display_order[row["option_id"]], variant_order[row["variant_id"]]))

        model_plan = {
            "model_key": model_key,
            "target_option_sheet": spec.target_option_sheet,
            "target_ovs_sheet": spec.target_ovs_sheet,
            "current_option_rows": _current_sheet_row_count(wb, spec.target_option_sheet),
            "current_ovs_rows": _current_sheet_row_count(wb, spec.target_ovs_sheet),
            "would_write_option_rows": len(option_rows),
            "would_write_ovs_rows": len(ovs_rows),
            "eligible_option_count": len(option_rows),
            "emitted_ovs_count": len(ovs_rows),
            "blocked_counts": dict(sorted(blocked_counts.items())),
            "errors": errors,
            "option_rows": option_rows,
            "ovs_rows": ovs_rows,
        }
        plan["models"][model_key] = model_plan
        plan["error_count"] += len(errors)

    return plan


def apply_plan_to_workbook(wb, plan: dict[str, Any]) -> None:
    for model_plan in plan["models"].values():
        write_sheet(wb, model_plan["target_option_sheet"], list(OPTION_SOURCE_HEADERS), model_plan["option_rows"])
        write_sheet(wb, model_plan["target_ovs_sheet"], list(OVS_SOURCE_HEADERS), model_plan["ovs_rows"])


def _strip_rows_from_plan(plan: dict[str, Any]) -> dict[str, Any]:
    report = {
        "selected_model_keys": plan["selected_model_keys"],
        "error_count": plan["error_count"],
        "models": {},
    }
    for model_key, model_plan in plan["models"].items():
        report["models"][model_key] = {key: value for key, value in model_plan.items() if key not in {"option_rows", "ovs_rows"}}
    return report


def _sheet_headers(ws) -> list[str]:
    return [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]


def _assert_future_metadata_inactive(wb) -> None:
    future_models = set(FUTURE_MODEL_KEYS)
    checks = {
        "model_master": ("model_key", ["active"]),
        "model_workbook_sources": ("model_key", ["active"]),
        "model_registry_promotion": ("model_key", ["promoted_to_runtime", "active"]),
    }
    for sheet_name, (model_field, inactive_fields) in checks.items():
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Missing metadata sheet {sheet_name}")
        ws = wb[sheet_name]
        headers = _sheet_headers(ws)
        index = {header: offset + 1 for offset, header in enumerate(headers)}
        if model_field not in index:
            raise RuntimeError(f"{sheet_name} missing {model_field}")
        for field in inactive_fields:
            if field not in index:
                raise RuntimeError(f"{sheet_name} missing {field}")
        for row_number in range(2, ws.max_row + 1):
            model_key = clean(ws.cell(row_number, index[model_field]).value)
            if model_key not in future_models:
                continue
            for field in inactive_fields:
                value = clean(ws.cell(row_number, index[field]).value).casefold()
                if value in {"true", "1", "yes", "y", "active"}:
                    raise RuntimeError(f"{sheet_name} row {row_number} unexpectedly has active {field}={value}")


def verify_saved_workbook(path: Path, plan: dict[str, Any], expected_option_review_rows: int) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        verification: dict[str, Any] = {"models": {}}
        if OPTION_REVIEW_SHEET not in wb.sheetnames:
            raise RuntimeError(f"Missing {OPTION_REVIEW_SHEET}")
        review_ws = wb[OPTION_REVIEW_SHEET]
        review_headers = _sheet_headers(review_ws)
        if review_headers != list(OPTION_REVIEW_HEADERS):
            raise RuntimeError(f"{OPTION_REVIEW_SHEET} header drift: {review_headers}")
        review_rows = max(review_ws.max_row - 1, 0)
        if review_rows != expected_option_review_rows:
            raise RuntimeError(f"{OPTION_REVIEW_SHEET} expected {expected_option_review_rows} rows, found {review_rows}")
        verification[OPTION_REVIEW_SHEET] = {"rows": review_rows}
        for model_key, model_plan in plan["models"].items():
            option_sheet = model_plan["target_option_sheet"]
            ovs_sheet = model_plan["target_ovs_sheet"]
            if option_sheet not in wb.sheetnames:
                raise RuntimeError(f"Missing target option sheet {option_sheet}")
            if ovs_sheet not in wb.sheetnames:
                raise RuntimeError(f"Missing target OVS sheet {ovs_sheet}")
            option_ws = wb[option_sheet]
            ovs_ws = wb[ovs_sheet]
            option_headers = _sheet_headers(option_ws)
            ovs_headers = _sheet_headers(ovs_ws)
            if option_headers != list(OPTION_SOURCE_HEADERS):
                raise RuntimeError(f"{option_sheet} header drift: {option_headers}")
            if ovs_headers != list(OVS_SOURCE_HEADERS):
                raise RuntimeError(f"{ovs_sheet} header drift: {ovs_headers}")
            option_rows = max(option_ws.max_row - 1, 0)
            ovs_rows = max(ovs_ws.max_row - 1, 0)
            expected_options = model_plan["eligible_option_count"]
            expected_ovs = model_plan["emitted_ovs_count"]
            if option_rows != expected_options:
                raise RuntimeError(f"{option_sheet} expected {expected_options} rows, found {option_rows}")
            if ovs_rows != expected_ovs:
                raise RuntimeError(f"{ovs_sheet} expected {expected_ovs} rows, found {ovs_rows}")
            verification["models"][model_key] = {
                "option_sheet": option_sheet,
                "option_rows": option_rows,
                "ovs_sheet": ovs_sheet,
                "ovs_rows": ovs_rows,
            }
        _assert_future_metadata_inactive(wb)
        verification["future_metadata_inactive"] = True
        return verification
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv-path", type=Path, default=DEFAULT_CSV_PATH)
    parser.add_argument("--model-key", choices=("all", *FUTURE_MODEL_KEYS), default="all")
    parser.add_argument("--dry-run", action="store_true", help="Report rows that would be written without saving the workbook.")
    args = parser.parse_args(argv)

    if not args.dry_run and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write workbook while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")

    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    csv_rows = load_csv_rows(args.csv_path)
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    backup_path = None
    try:
        existing_rows = rows_from_sheet(wb, OPTION_REVIEW_SHEET)
        merged_rows = merge_csv_with_existing_option_review_rows(existing_rows, csv_rows)
        selected = list(FUTURE_MODEL_KEYS) if args.model_key == "all" else [args.model_key]
        plan = build_future_option_population_plan(wb, csv_rows, selected)
        report = {
            "csv_path": str(args.csv_path),
            "csv_rows": len(csv_rows),
            "existing_option_review_rows": len(existing_rows),
            "merged_option_review_rows": len(merged_rows),
            **_strip_rows_from_plan(plan),
        }
        if plan["error_count"]:
            print(json.dumps({"status": "blocked", **report}, indent=2))
            return 1
        if args.dry_run:
            print(json.dumps({"status": "dry_run", **report}, indent=2))
            return 0
        write_sheet(wb, OPTION_REVIEW_SHEET, list(OPTION_REVIEW_HEADERS), merged_rows)
        apply_plan_to_workbook(wb, plan)
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved_workbook(WORKBOOK_PATH, plan, len(merged_rows))
    print(
        json.dumps(
            {
                "status": "written",
                "workbook": str(WORKBOOK_PATH),
                "backup": str(backup_path),
                "csv_path": str(args.csv_path),
                "csv_rows": len(csv_rows),
                "merged_option_review_rows": len(merged_rows),
                **_strip_rows_from_plan(plan),
                "verification": verification,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
