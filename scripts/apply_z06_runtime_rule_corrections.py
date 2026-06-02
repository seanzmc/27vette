#!/usr/bin/env python3
"""Apply approved Pass 1 Z06 runtime rule corrections.

Dry-run is the default. Use --write to safely save stingray_master.xlsx.
This script intentionally does not call apply_future_model_option_review.py.
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from corvette_form_generator.model_configs import WORKBOOK_PATH  # noqa: E402
from corvette_form_generator.workbook import clean, excel_lock_path, save_workbook_safely  # noqa: E402

OPTION_SHEET = "z06_options"
OVS_SHEET = "z06_ovs"
RULE_MAPPING_SHEET = "z06_rule_mapping"
RULE_GROUPS_SHEET = "z06_rule_groups"
RULE_GROUP_MEMBERS_SHEET = "z06_rule_group_members"
PRICE_RULES_SHEET = "z06_price_rules"
EXCLUSIVE_GROUPS_SHEET = "z06_exclusive_groups"
EXCLUSIVE_MEMBERS_SHEET = "z06_exclusive_members"

PACKAGE_RPOS = ("PDB", "PDD", "PDF")
CARBON_WHEEL_RPOS = ("ROY", "ROZ", "STZ")
GBA_INCOMPATIBLE_RPOS = ("EFY", "ZYC", "D84", "D86")


@dataclass(frozen=True)
class Change:
    sheet: str
    row_number: int | str
    key: str
    field: str
    current: Any
    desired: Any
    reason: str


def headers(ws) -> list[str]:
    return [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def index(ws) -> dict[str, int]:
    return {header: offset + 1 for offset, header in enumerate(headers(ws)) if header}


def rows(ws):
    idx = index(ws)
    for row_number in range(2, ws.max_row + 1):
        row = {header: ws.cell(row_number, col).value for header, col in idx.items()}
        if any(value not in (None, "") for value in row.values()):
            yield row_number, row


def price_equal(current: Any, desired: Any) -> bool:
    if desired in (None, ""):
        return clean(current) == ""
    try:
        return float(clean(current) or 0) == float(desired)
    except ValueError:
        return False


def values_equal(field: str, current: Any, desired: Any) -> bool:
    if field in {"price", "price_value"}:
        return price_equal(current, desired)
    return clean(current) == clean(desired)


def set_cell(ws, row_number: int, field: str, desired: Any, changes: list[Change], *, key: str, reason: str) -> None:
    idx = index(ws)
    if field not in idx:
        return
    current = ws.cell(row_number, idx[field]).value
    if not values_equal(field, current, desired):
        changes.append(Change(ws.title, row_number, key, field, current, desired, reason))
        ws.cell(row_number, idx[field]).value = desired


def row_by_key(ws, key_field: str, key: str) -> tuple[int | None, dict[str, Any] | None]:
    for row_number, row in rows(ws):
        if clean(row.get(key_field)) == key:
            return row_number, row
    return None, None


def ensure_row(ws, key_field: str, key: str, values: dict[str, Any], changes: list[Change], *, reason: str) -> None:
    idx = index(ws)
    row_number, _ = row_by_key(ws, key_field, key)
    if row_number is None:
        row_number = ws.max_row + 1
        for field, value in values.items():
            if field in idx:
                ws.cell(row_number, idx[field]).value = value
        changes.append(Change(ws.title, f"new:{row_number}", key, "row", None, values, reason))
        return
    for field, desired in values.items():
        if field == key_field:
            continue
        set_cell(ws, row_number, field, desired, changes, key=key, reason=reason)


def option_maps(wb) -> tuple[dict[str, dict[str, Any]], dict[str, str], dict[str, str]]:
    by_rpo: dict[str, dict[str, Any]] = {}
    id_by_rpo: dict[str, str] = {}
    rpo_by_id: dict[str, str] = {}
    for row_number, row in rows(wb[OPTION_SHEET]):
        rpo = clean(row.get("rpo"))
        option_id = clean(row.get("option_id"))
        if rpo and option_id:
            by_rpo[rpo] = {**row, "_row_number": row_number}
            id_by_rpo[rpo] = option_id
            rpo_by_id[option_id] = rpo
    return by_rpo, id_by_rpo, rpo_by_id


def required_option_ids(id_by_rpo: dict[str, str], rpos: tuple[str, ...] | list[str]) -> list[str]:
    missing = [rpo for rpo in rpos if rpo not in id_by_rpo]
    if missing:
        raise RuntimeError(f"Missing required Z06 RPO(s): {', '.join(missing)}")
    return [id_by_rpo[rpo] for rpo in rpos]


def rule_values(rule_id: str, source_id: str, rule_type: str, target_id: str, note: str, *, by_id: dict[str, dict[str, Any]], body: str = "", runtime_action: str = "", disabled_reason: str = "", status: str = "active") -> dict[str, Any]:
    source = by_id[source_id]
    target = by_id[target_id]
    return {
        "rule_id": rule_id,
        "source_id": source_id,
        "rule_type": rule_type,
        "target_id": target_id,
        "target_type": "option",
        "original_detail_raw": note,
        "review_flag": False,
        "source_type": "option",
        "target_selection_mode": "",
        "source_selection_mode": "",
        "target_section": clean(target.get("section_id")),
        "source_section": clean(source.get("section_id")),
        "generation_action": "",
        "body_style_scope": body,
        "runtime_action": runtime_action,
        "disabled_reason": disabled_reason,
        "normalization_status": status,
        "normalization_reason": "",
        "replacement_group_id": "",
        "replacement_rule_id": "",
    }


def ensure_rule(wb, changes: list[Change], source_rpo: str, rule_type: str, target_rpo: str, note: str, *, body: str = "", runtime_action: str = "", disabled_reason: str = "", rule_id: str | None = None) -> None:
    by_rpo, id_by_rpo, _ = option_maps(wb)
    required_option_ids(id_by_rpo, [source_rpo, target_rpo])
    by_id = {clean(row.get("option_id")): row for row in by_rpo.values()}
    source_id = id_by_rpo[source_rpo]
    target_id = id_by_rpo[target_rpo]
    rule_id = rule_id or f"z06_rule_{source_id}_{rule_type}_{target_id}"
    ensure_row(
        wb[RULE_MAPPING_SHEET],
        "rule_id",
        rule_id,
        rule_values(rule_id, source_id, rule_type, target_id, note, by_id=by_id, body=body, runtime_action=runtime_action, disabled_reason=disabled_reason),
        changes,
        reason="approved Z06 runtime rule correction",
    )


def omit_rule(wb, changes: list[Change], rule_id: str, reason: str) -> None:
    ws = wb[RULE_MAPPING_SHEET]
    row_number, row = row_by_key(ws, "rule_id", rule_id)
    if row_number is None:
        return
    set_cell(ws, row_number, "normalization_status", "omitted", changes, key=rule_id, reason=reason)
    set_cell(ws, row_number, "normalization_reason", reason, changes, key=rule_id, reason=reason)


def ensure_price_rule(wb, changes: list[Change], rule_id: str, condition_rpo: str, target_rpo: str, price: int, note: str, *, body: str = "*", trim: str = "*") -> None:
    _, id_by_rpo, _ = option_maps(wb)
    required_option_ids(id_by_rpo, [condition_rpo, target_rpo])
    ensure_row(
        wb[PRICE_RULES_SHEET],
        "price_rule_id",
        rule_id,
        {
            "price_rule_id": rule_id,
            "condition_option_id": id_by_rpo[condition_rpo],
            "price_rule_type": "override",
            "target_option_id": id_by_rpo[target_rpo],
            "price_value": price,
            "body_style_scope": body,
            "trim_level_scope": trim,
            "review_flag": False,
            "notes": note,
        },
        changes,
        reason="approved Z06 runtime price correction",
    )


def ensure_rule_group(wb, changes: list[Change], source_rpo: str, group_id: str, target_rpos: tuple[str, ...], disabled_reason: str, notes: str) -> None:
    _, id_by_rpo, _ = option_maps(wb)
    required_option_ids(id_by_rpo, [source_rpo, *target_rpos])
    ensure_row(
        wb[RULE_GROUPS_SHEET],
        "group_id",
        group_id,
        {
            "group_id": group_id,
            "group_type": "requires_any",
            "source_id": id_by_rpo[source_rpo],
            "body_style_scope": "*",
            "trim_level_scope": "*",
            "variant_scope": "*",
            "disabled_reason": disabled_reason,
            "active": "True",
            "notes": notes,
        },
        changes,
        reason="approved Z06 grouped requirement",
    )
    members_ws = wb[RULE_GROUP_MEMBERS_SHEET]
    for order, target_rpo in enumerate(target_rpos, start=1):
        target_id = id_by_rpo[target_rpo]
        existing_row = None
        for row_number, row in rows(members_ws):
            if clean(row.get("group_id")) == group_id and clean(row.get("target_id")) == target_id:
                existing_row = row_number
                break
        key = f"{group_id}:{target_id}"
        values = {"group_id": group_id, "target_id": target_id, "display_order": order * 10, "active": "True"}
        if existing_row is None:
            row_number = members_ws.max_row + 1
            idx = index(members_ws)
            for field, value in values.items():
                members_ws.cell(row_number, idx[field]).value = value
            changes.append(Change(members_ws.title, f"new:{row_number}", key, "row", None, values, "approved Z06 grouped requirement member"))
        else:
            for field, value in values.items():
                if field == "group_id" or field == "target_id":
                    continue
                set_cell(members_ws, existing_row, field, value, changes, key=key, reason="approved Z06 grouped requirement member")


def deactivate_package_exclusive_group(wb, changes: list[Change]) -> None:
    ws = wb[EXCLUSIVE_GROUPS_SHEET]
    row_number, _ = row_by_key(ws, "group_id", "z06_excl_carbon_wheel_packages")
    if row_number is not None:
        set_cell(ws, row_number, "active", "False", changes, key="z06_excl_carbon_wheel_packages", reason="package peers now use explicit deactivate/blocking rules")
        set_cell(ws, row_number, "notes", "Inactive: PDB, PDD, and PDF now use explicit workbook excludes so selected packages deactivate peers instead of radio-switching them.", changes, key="z06_excl_carbon_wheel_packages", reason="package peers now use explicit deactivate/blocking rules")


def active_normal_wheel_rpos(wb) -> list[str]:
    by_rpo, _, _ = option_maps(wb)
    wheel_rpos: list[str] = []
    for rpo, row in by_rpo.items():
        if clean(row.get("section_id")) == "sec_whee_002" and clean(row.get("active")) == "True":
            wheel_rpos.append(rpo)
    return sorted(wheel_rpos)


def update_options(wb, changes: list[Change]) -> None:
    by_rpo, _, _ = option_maps(wb)
    ws = wb[OPTION_SHEET]
    for rpo in ("V8X", "RYQ"):
        row = by_rpo.get(rpo)
        if row:
            set_cell(ws, int(row["_row_number"]), "active", "False", changes, key=rpo, reason="unreleased Z06 option should not appear on front end")
    # Keep EFY in the exterior accent required group while preserving existing EDU/EFR rows.
    _, id_by_rpo, _ = option_maps(wb)
    required_option_ids(id_by_rpo, ["EFY"])
    members_ws = wb[EXCLUSIVE_MEMBERS_SHEET]
    group_id = "z06_excl_exterior_accents"
    efy_id = id_by_rpo["EFY"]
    existing_row = None
    for row_number, row in rows(members_ws):
        if clean(row.get("group_id")) == group_id and clean(row.get("option_id")) == efy_id:
            existing_row = row_number
            break
    values = {"group_id": group_id, "option_id": efy_id, "display_order": 30, "active": "True"}
    key = f"{group_id}:{efy_id}"
    if existing_row is None:
        row_number = members_ws.max_row + 1
        idx = index(members_ws)
        for field, value in values.items():
            members_ws.cell(row_number, idx[field]).value = value
        changes.append(Change(members_ws.title, f"new:{row_number}", key, "row", None, values, "EFY belongs in exterior accent group"))
    else:
        set_cell(members_ws, existing_row, "active", "True", changes, key=key, reason="EFY belongs in exterior accent group")


def build_plan(wb) -> list[Change]:
    required = [OPTION_SHEET, OVS_SHEET, RULE_MAPPING_SHEET, RULE_GROUPS_SHEET, RULE_GROUP_MEMBERS_SHEET, PRICE_RULES_SHEET, EXCLUSIVE_GROUPS_SHEET, EXCLUSIVE_MEMBERS_SHEET]
    missing = [sheet for sheet in required if sheet not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Missing required sheets: {', '.join(missing)}")
    changes: list[Change] = []
    by_rpo, id_by_rpo, _ = option_maps(wb)
    required_option_ids(id_by_rpo, [
        "Z07", "J57", "J6A", "T0F", "T0G", "PDB", "PDD", "PDF", "ROY", "ROZ", "STZ", "B6P", "ZZ3", "BCW", "D3V", "SL9", "NWI", "WUB", "GBA", "PBC", "EFY", "ZYC", "D84", "D86",
    ])

    update_options(wb, changes)

    omit_rule(wb, changes, "z06_rule_opt_t0f_001_requires_opt_j57_001", "Z06 T0F should not require J57 before selection.")
    omit_rule(wb, changes, "z06_copy_rule_opt_nwi_001_requires_opt_wub_001_opt_nwi_001_requires_opt_wub_001", "Z06 WUB is standard equipment; NWI should not require WUB.")

    ensure_rule(wb, changes, "J57", "excludes", "J6A", "J57 carbon ceramic brakes make J6A black calipers unavailable.", disabled_reason="Blocked by J57 carbon ceramic brakes.")
    ensure_rule(wb, changes, "D3V", "includes", "BCW", "D3V engine cover selection adds BCW red engine intake.")
    ensure_rule(wb, changes, "PBC", "requires", "ZZ3", "PBC requires ZZ3 on convertible.", body="convertible", disabled_reason="Convertible PBC requires ZZ3 Convertible Engine Appearance Package.", rule_id="z06_rule_opt_pbc_001_requires_opt_zz3_001_convertible")

    for target_rpo in GBA_INCOMPATIBLE_RPOS:
        ensure_rule(wb, changes, "GBA", "excludes", target_rpo, f"{target_rpo} is not available with GBA Black exterior paint.", disabled_reason=f"Not available with GBA Black exterior paint.")

    deactivate_package_exclusive_group(wb, changes)
    for package_rpo in PACKAGE_RPOS:
        ensure_rule_group(
            wb,
            changes,
            package_rpo,
            f"z06_group_{package_rpo.lower()}_requires_carbon_wheel",
            CARBON_WHEEL_RPOS,
            f"{package_rpo} requires one Z06 carbon fiber wheel choice: ROY, ROZ, or STZ.",
            f"{package_rpo} wheel-and-brake package requires a carbon fiber wheel selection.",
        )
        for peer_rpo in PACKAGE_RPOS:
            if peer_rpo != package_rpo:
                ensure_rule(wb, changes, package_rpo, "excludes", peer_rpo, f"{package_rpo} deactivates peer wheel-and-brake package {peer_rpo}.", disabled_reason=f"Blocked by selected {package_rpo} package.")
        for wheel_rpo in active_normal_wheel_rpos(wb):
            ensure_rule(wb, changes, package_rpo, "excludes", wheel_rpo, f"{package_rpo} requires carbon fiber wheels and deactivates aluminum wheel {wheel_rpo}.", disabled_reason=f"{package_rpo} requires a carbon fiber wheel selection.")

    ensure_price_rule(wb, changes, "z06_pr_b6p_d3v_zero", "B6P", "D3V", 0, "B6P includes D3V, so D3V does not add a second charge.")
    ensure_price_rule(wb, changes, "z06_pr_b6p_sl9_zero", "B6P", "SL9", 0, "B6P includes SL9, so SL9 does not add a second charge.")
    ensure_price_rule(wb, changes, "z06_pr_zz3_sl9_zero", "ZZ3", "SL9", 0, "ZZ3 includes SL9, so SL9 does not add a second charge.")

    return changes


def verify_saved(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        by_rpo, id_by_rpo, _ = option_maps(wb)
        checks = {
            "v8x_active": clean(by_rpo["V8X"].get("active")),
            "ryq_active": clean(by_rpo["RYQ"].get("active")),
            "rule_group_count": 0,
            "price_rules": {},
        }
        for package_rpo in PACKAGE_RPOS:
            gid = f"z06_group_{package_rpo.lower()}_requires_carbon_wheel"
            groups = [row for _, row in rows(wb[RULE_GROUPS_SHEET]) if clean(row.get("group_id")) == gid and clean(row.get("active")) == "True"]
            members = [row for _, row in rows(wb[RULE_GROUP_MEMBERS_SHEET]) if clean(row.get("group_id")) == gid and clean(row.get("active")) == "True"]
            if not groups or {clean(row.get("target_id")) for row in members} != {id_by_rpo[rpo] for rpo in CARBON_WHEEL_RPOS}:
                raise RuntimeError(f"Missing {package_rpo} carbon wheel requires_any group")
            checks["rule_group_count"] += 1
        if checks["v8x_active"] != "False" or checks["ryq_active"] != "False":
            raise RuntimeError(f"V8X/RYQ should be inactive: {checks}")
        price_rules = {clean(row.get("price_rule_id")): row for _, row in rows(wb[PRICE_RULES_SHEET])}
        for rid in ["z06_pr_b6p_d3v_zero", "z06_pr_b6p_sl9_zero", "z06_pr_zz3_sl9_zero"]:
            if rid not in price_rules or clean(price_rules[rid].get("price_value")) != "0":
                raise RuntimeError(f"Missing zero price rule {rid}")
            checks["price_rules"][rid] = clean(price_rules[rid].get("price_value"))
        return checks
    finally:
        wb.close()


def summarize(changes: list[Change]) -> dict[str, Any]:
    by_sheet: dict[str, int] = {}
    by_field: dict[str, int] = {}
    for change in changes:
        by_sheet[change.sheet] = by_sheet.get(change.sheet, 0) + 1
        by_field[change.field] = by_field.get(change.field, 0) + 1
    return {"total_changes": len(changes), "changes_by_sheet": dict(sorted(by_sheet.items())), "changes_by_field": dict(sorted(by_field.items()))}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="Write workbook changes. Default is dry-run.")
    parser.add_argument("--include-changes", action="store_true", help="Include per-cell changes in the JSON report.")
    args = parser.parse_args(argv)

    if args.write and excel_lock_path(WORKBOOK_PATH).exists():
        raise RuntimeError(f"Refusing to write while Excel lock file exists: {excel_lock_path(WORKBOOK_PATH)}")
    loaded_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    backup_path = None
    try:
        changes = build_plan(wb)
        if not args.write:
            report: dict[str, Any] = {"status": "dry_run", **summarize(changes), "workbook": str(WORKBOOK_PATH)}
            if args.include_changes:
                report["changes"] = [change.__dict__ for change in changes]
            print(json.dumps(report, indent=2, default=str))
            return 0
        backup_path = save_workbook_safely(wb, WORKBOOK_PATH, loaded_mtime_ns=loaded_mtime_ns)
    finally:
        wb.close()

    verification = verify_saved(WORKBOOK_PATH)
    report = {"status": "written", **summarize(changes), "workbook": str(WORKBOOK_PATH), "backup": str(backup_path), "verification": verification}
    if args.include_changes:
        report["changes"] = [change.__dict__ for change in changes]
    print(json.dumps(report, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
