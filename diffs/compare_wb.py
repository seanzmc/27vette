#!/usr/bin/env python
"""Read-only comparison of two stingray_master workbooks. Never saves."""
import sys
from openpyxl import load_workbook

LIVE = "diffs/stingray_master.xlsx"   # production main snapshot
LOCAL = "stingray_master.xlsx"        # local divergent

def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v

def sheet_data(ws):
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(tuple(norm(c) for c in r))
    # trim trailing fully-empty rows
    while rows and all(c is None for c in rows[-1]):
        rows.pop()
    return rows

wb_live = load_workbook(LIVE, read_only=True, data_only=False)
wb_local = load_workbook(LOCAL, read_only=True, data_only=False)

live_sheets = wb_live.sheetnames
local_sheets = wb_local.sheetnames

print("=== SHEET LISTS ===")
print("LIVE  (%d): %s" % (len(live_sheets), ", ".join(live_sheets)))
print()
print("LOCAL (%d): %s" % (len(local_sheets), ", ".join(local_sheets)))
print()
added = [s for s in local_sheets if s not in live_sheets]
removed = [s for s in live_sheets if s not in local_sheets]
common = [s for s in live_sheets if s in local_sheets]
print("Added in LOCAL:", added)
print("Removed from LIVE:", removed)
order_live = [s for s in live_sheets if s in common]
order_local = [s for s in local_sheets if s in common]
print("Order of common sheets differs:", order_live != order_local)
if order_live != order_local:
    print("  live order :", order_live)
    print("  local order:", order_local)
print()

print("=== PER-SHEET STRUCTURE (common sheets) ===")
data_cache = {}
for s in common:
    dl = sheet_data(wb_live[s])
    dc = sheet_data(wb_local[s])
    data_cache[s] = (dl, dc)
    hl = list(dl[0]) if dl else []
    hc = list(dc[0]) if dc else []
    line = "%-28s rows live=%d local=%d delta=%+d" % (s, len(dl), len(dc), len(dc)-len(dl))
    if hl == hc:
        print(line, "| headers: same (%d cols)" % len(hl))
    else:
        print(line, "| HEADERS DIFFER")
        ha = [h for h in hc if h not in hl]
        hr = [h for h in hl if h not in hc]
        if ha: print("    cols added in local:", ha)
        if hr: print("    cols removed in local:", hr)
        if not ha and not hr:
            print("    same cols, order changed: live=%s local=%s" % (hl, hc))
print()

import re
KEY_SHEETS = [s for s in common if (
    s in ("model_master","variant_master","stingray_options","rule_mapping",
          "price_rules","lt_interiors","LZ_Interiors","model_registry_promotion")
    or s.startswith("z06_") or s.startswith("grandSport_"))]

print("=== DATA DIFFS (key sheets, keyed on first column) ===")
for s in KEY_SHEETS:
    dl, dc = data_cache[s]
    if not dl or not dc:
        print("[%s] empty on one side" % s); continue
    hl, hc = list(dl[0]), list(dc[0])
    if hl != hc:
        print("[%s] headers differ -- skipping row diff (see structure section)" % s)
        continue
    hdr = hl
    def keyed(rows):
        d = {}
        for i, r in enumerate(rows[1:], start=2):
            k = r[0]
            if k is None:
                k = "__row%d" % i
            # handle dup keys
            kk = k
            n = 1
            while kk in d:
                n += 1
                kk = "%s#%d" % (k, n)
            d[kk] = r
        return d
    kl, kc = keyed(dl), keyed(dc)
    only_live = [k for k in kl if k not in kc]
    only_local = [k for k in kc if k not in kl]
    changed = []
    for k in kl:
        if k in kc and kl[k] != kc[k]:
            diffs = []
            rl, rc = kl[k], kc[k]
            mx = max(len(rl), len(rc))
            for i in range(mx):
                a = rl[i] if i < len(rl) else None
                b = rc[i] if i < len(rc) else None
                if a != b:
                    col = hdr[i] if i < len(hdr) else "col%d" % (i+1)
                    diffs.append((col, a, b))
            changed.append((k, diffs))
    print("[%s] rows-only-in-live=%d rows-only-in-local=%d changed-rows=%d" %
          (s, len(only_live), len(only_local), len(changed)))
    if only_live:
        print("   only in LIVE keys (up to 10):", only_live[:10])
    if only_local:
        print("   only in LOCAL keys (up to 10):", only_local[:10])
    if changed:
        # categorize by column
        from collections import Counter
        colcount = Counter()
        for k, diffs in changed:
            for col, a, b in diffs:
                colcount[col] += 1
        print("   changed cells by column:", dict(colcount))
        for k, diffs in changed[:6]:
            print("   key=%r:" % (k,))
            for col, a, b in diffs[:6]:
                print("      %s: live=%r -> local=%r" % (col, a, b))
        if len(changed) > 6:
            print("   ... %d more changed rows" % (len(changed)-6))
    print()

print("=== form_* SHEETS (summary only) ===")
form_live = [s for s in live_sheets if s.startswith("form_")]
form_local = [s for s in local_sheets if s.startswith("form_")]
print("form_* live=%d local=%d" % (len(form_live), len(form_local)))
print("only live:", [s for s in form_live if s not in form_local])
print("only local:", [s for s in form_local if s not in form_live])
for s in form_live:
    if s in form_local:
        dl, dc = data_cache.get(s, (None, None))
        if dl is None:
            dl = sheet_data(wb_live[s]); dc = sheet_data(wb_local[s])
        same = dl == dc
        print("  %-30s rows live=%d local=%d identical=%s" % (s, len(dl), len(dc), same))

wb_live.close()
wb_local.close()
print("\nDone (no writes performed).")
