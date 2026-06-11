#!/usr/bin/env python
"""Pass 2: type-normalized diffs + common-column diffs for header-changed sheets. Read-only."""
from openpyxl import load_workbook
from collections import Counter

LIVE = "diffs/stingray_master.xlsx"
LOCAL = "stingray_master.xlsx"

def norm(v):
    if v is None: return ""
    if isinstance(v, bool): return "True" if v else "False"
    if isinstance(v, float) and v == int(v): v = int(v)
    return str(v).strip()

def sheet_data(ws):
    rows = [tuple(norm(c) for c in r) for r in ws.iter_rows(values_only=True)]
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows

wb_live = load_workbook(LIVE, read_only=True, data_only=False)
wb_local = load_workbook(LOCAL, read_only=True, data_only=False)

SHEETS = ["model_master","variant_master","stingray_options","rule_mapping","price_rules",
          "grandSport_price_rules","grandSport_rule_mapping","grandSport_rule_groups",
          "grandSport_rule_group_members",
          "z06_rule_mapping","z06_price_rules","z06_options","z06_rule_groups",
          "z06_rule_group_members","z06_exclusive_groups","z06_exclusive_members",
          "lt_interiors","LZ_Interiors","model_registry_promotion",
          "form_rules","form_price_rules","form_interiors",
          "section_master","section_presentation","asset_map","context_choice_copy",
          "zr1_rule_mapping","zr1x_rule_mapping","zr1_price_rules","zr1x_price_rules"]

for s in SHEETS:
    if s not in wb_live.sheetnames or s not in wb_local.sheetnames:
        continue
    dl = sheet_data(wb_live[s]); dc = sheet_data(wb_local[s])
    hl, hc = list(dl[0]), list(dc[0])
    common_cols = [h for h in hl if h in hc]
    il = [hl.index(h) for h in common_cols]
    ic = [hc.index(h) for h in common_cols]
    def keyed(rows, idxs):
        d = {}
        for r in rows[1:]:
            base = "|".join(r[i] if i < len(r) else "" for i in idxs[:2])
            k = base; n = 1
            while k in d:
                n += 1; k = "%s#%d" % (base, n)
            d[k] = tuple(r[i] if i < len(r) else "" for i in idxs)
        return d
    kl = keyed(dl, il); kc = keyed(dc, ic)
    only_l = [k for k in kl if k not in kc]
    only_c = [k for k in kc if k not in kl]
    changed = []
    for k in kl:
        if k in kc and kl[k] != kc[k]:
            ds = [(common_cols[i], kl[k][i], kc[k][i]) for i in range(len(common_cols)) if kl[k][i] != kc[k][i]]
            changed.append((k, ds))
    if not (only_l or only_c or changed):
        print("[%s] IDENTICAL on common columns (type-normalized)" % s)
        continue
    print("[%s] only-live=%d only-local=%d changed=%d" % (s, len(only_l), len(only_c), len(changed)))
    if only_l: print("   only LIVE:", only_l[:8])
    if only_c: print("   only LOCAL:", only_c[:8])
    if changed:
        cc = Counter()
        for k, ds in changed:
            for col, a, b in ds: cc[col] += 1
        print("   changed cells by column:", dict(cc))
        for k, ds in changed[:5]:
            print("   key=%s" % k)
            for col, a, b in ds[:5]:
                print("      %s: live=%r -> local=%r" % (col, a, b))
        if len(changed) > 5:
            print("   ... %d more" % (len(changed)-5))

wb_live.close(); wb_local.close()
print("\nDone (no writes).")
