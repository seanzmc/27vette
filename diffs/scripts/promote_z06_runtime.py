#!/usr/bin/env python3
"""Promote the workbook-authored Z06 data into the static app runtime registry."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from corvette_form_generator.model_configs import WORKBOOK_PATH
from corvette_form_generator.workbook import clean, excel_lock_path, rows_from_sheet, save_workbook_safely

Z06_VARIANT_IDS = ("1lz_h07", "2lz_h07", "3lz_h07", "1lz_h67", "2lz_h67", "3lz_h67")
Z06_DATASET_NAME = "2027 Corvette Z06 operational form"
Z06_MODEL_NOTES = "Z06 promoted to runtime after source data review."
Z06_PROMOTION_NOTES = "Z06 draft artifact promoted to branch runtime registry."
Z06_DRAFT_ARTIFACT_PATH = "form-output/inspection/z06-form-data-draft.json"


def headers_for(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}


def row_by_key(ws, key_column: str, key_value: str) -> int:
    headers = headers_for(ws)
    if key_column not in headers:
        raise ValueError(f"{ws.title} is missing required column {key_column!r}")
    key_col = headers[key_column]
    matches: list[int] = []
    for row_idx in range(2, ws.max_row + 1):
        if clean(ws.cell(row_idx, key_col).value).lower() == key_value.lower():
            matches.append(row_idx)
    if not matches:
        raise ValueError(f"{ws.title} is missing {key_column}={key_value!r}")
    if len(matches) > 1:
        raise ValueError(f"{ws.title} has duplicate {key_column}={key_value!r} rows: {matches}")
    return matches[0]


def set_cell(ws, row_idx: int, headers: dict[str, int], column: str, value: Any, changes: list[dict[str, Any]]) -> None:
    if column not in headers:
        raise ValueError(f"{ws.title} is missing required column {column!r}")
    cell = ws.cell(row_idx, headers[column])
    old_value = cell.value
    if old_value == value:
        return
    changes.append(
        {
            "sheet": ws.title,
            "row": row_idx,
            "column": column,
            "old": clean(old_value),
            "new": clean(value),
        }
    )
    cell.value = value


def promote_z06(wb) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []

    model_ws = wb["model_master"]
    model_headers = headers_for(model_ws)
    model_row = row_by_key(model_ws, "model_key", "z06")
    set_cell(model_ws, model_row, model_headers, "active", True, changes)
    set_cell(model_ws, model_row, model_headers, "dataset_name", Z06_DATASET_NAME, changes)
    set_cell(model_ws, model_row, model_headers, "default_model", False, changes)
    set_cell(model_ws, model_row, model_headers, "registry_key", "z06", changes)
    set_cell(model_ws, model_row, model_headers, "export_slug", "z06", changes)
    set_cell(model_ws, model_row, model_headers, "notes", Z06_MODEL_NOTES, changes)

    promotion_ws = wb["model_registry_promotion"]
    promotion_headers = headers_for(promotion_ws)
    promotion_row = row_by_key(promotion_ws, "model_key", "z06")
    set_cell(promotion_ws, promotion_row, promotion_headers, "registry_key", "z06", changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "promoted_to_runtime", True, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "default_model", False, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "artifact_path", Z06_DRAFT_ARTIFACT_PATH, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "artifact_type", "draft_artifact", changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "active", True, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "display_order", 3, changes)
    set_cell(promotion_ws, promotion_row, promotion_headers, "notes", Z06_PROMOTION_NOTES, changes)

    variant_ws = wb["variant_master"]
    variant_headers = headers_for(variant_ws)
    for variant_id in Z06_VARIANT_IDS:
        variant_row = row_by_key(variant_ws, "variant_id", variant_id)
        set_cell(variant_ws, variant_row, variant_headers, "active", True, changes)

    return changes


def verify_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        models = {row["model_key"]: row for row in rows_from_sheet(wb, "model_master")}
        promotions = {row["model_key"]: row for row in rows_from_sheet(wb, "model_registry_promotion")}
        variants = {row["variant_id"]: row for row in rows_from_sheet(wb, "variant_master")}
        model = models.get("z06") or {}
        promotion = promotions.get("z06") or {}
        variant_status = {variant_id: (variants.get(variant_id) or {}).get("active") for variant_id in Z06_VARIANT_IDS}
        failures: list[str] = []
        if model.get("active") != "True":
            failures.append("model_master z06 active is not True")
        if model.get("dataset_name") != Z06_DATASET_NAME:
            failures.append("model_master z06 dataset_name was not updated")
        expected_promotion = {
            "promoted_to_runtime": "True",
            "active": "True",
            "artifact_path": Z06_DRAFT_ARTIFACT_PATH,
            "artifact_type": "draft_artifact",
            "default_model": "False",
            "registry_key": "z06",
        }
        for key, expected in expected_promotion.items():
            if promotion.get(key) != expected:
                failures.append(f"model_registry_promotion z06 {key} expected {expected!r}, found {promotion.get(key)!r}")
        for variant_id, active in variant_status.items():
            if active != "True":
                failures.append(f"variant_master {variant_id} active expected 'True', found {active!r}")
        return {
            "model_master_z06": model,
            "model_registry_promotion_z06": promotion,
            "variant_active": variant_status,
            "failures": failures,
        }
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="write changes to stingray_master.xlsx")
    args = parser.parse_args()

    workbook_path = Path(WORKBOOK_PATH)
    lock_path = excel_lock_path(workbook_path)
    if lock_path.exists():
        raise SystemExit(f"Refusing to run; Excel lock file is present: {lock_path}. Close Excel first.")

    loaded_mtime_ns = workbook_path.stat().st_mtime_ns
    wb = load_workbook(workbook_path)
    try:
        changes = promote_z06(wb)
        result: dict[str, Any] = {
            "workbook": str(workbook_path),
            "write": args.write,
            "change_count": len(changes),
            "changes": changes,
        }
        if args.write:
            backup_path = save_workbook_safely(wb, workbook_path, loaded_mtime_ns=loaded_mtime_ns)
            verification = verify_workbook(workbook_path)
            result["backup_path"] = str(backup_path)
            result["verification"] = verification
            if verification["failures"]:
                raise SystemExit(json.dumps(result, indent=2))
        print(json.dumps(result, indent=2))
    finally:
        wb.close()


if __name__ == "__main__":
    main()
